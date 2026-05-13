from __future__ import annotations

import hashlib
import io
import json
import os
import re
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable
from urllib.parse import parse_qs, unquote_plus, urlparse

from bs4 import BeautifulSoup

from bpmis_jira_tool.config import Settings
from bpmis_jira_tool.codex_model_router import CODEX_ROUTE_DEEP, resolve_codex_model, resolve_codex_reasoning_effort
from bpmis_jira_tool.errors import ToolError
from bpmis_jira_tool.source_code_qa import CodexCliBridgeSourceCodeQALLMProvider

from .confluence import ConfluenceConnector, IngestedConfluencePage
from .storage import BriefingStore


PRD_REVIEW_PROMPT_VERSION = "v11_prd_review_table_only_no_image_evidence"
PRD_SUMMARY_PROMPT_VERSION = "v2_prd_summary_hybrid_sections"
PRD_REVIEW_MAX_SOURCE_CHARS = 90_000
PRD_HYBRID_BATCH_SOURCE_CHARS = 45_000
PRD_SECTION_LONG_CHAR_THRESHOLD = 8_000
PRD_LINKED_SPREADSHEET_MAX_SHEETS = 10
PRD_LINKED_SPREADSHEET_MAX_ROWS_PER_SHEET = 80
PRD_LINKED_SPREADSHEET_MAX_COLS = 20
PRD_LINKED_SPREADSHEET_MAX_TEXT_CHARS = 60_000
PRD_REVIEW_TABLE_MEDIA_MAX_ROWS = 120
PRD_REVIEW_TABLE_MEDIA_MAX_CHARS = 20_000
PRD_REPORT_SECTION_KEYWORDS = ("report", "layout", "format", "template", "register", "field name", "字段", "报表", "报告", "模板", "格式")
PRD_GOOGLE_SHEET_ARTIFACT_CACHE_VERSION = "v1"
PRD_GOOGLE_SHEET_SCREENSHOT_EVIDENCE_CACHE_VERSION = "v1"
PRD_ANTI_FRAUD_KEYWORDS = (
    "anti-fraud",
    "antifraud",
    "anti fraud",
    "risk decision",
    "fraud",
    "blacklist",
    "black list",
    "whitelist",
    "white list",
)
PRD_BRIEFING_REVIEW_CACHE_KEY = "__prd_briefing_url_review__"
PRD_URL_SUMMARY_CACHE_KEY = "__prd_url_summary__"
PRDProgressCallback = Callable[[str, str, int, int], None]


@dataclass(frozen=True)
class PRDReviewRequest:
    owner_key: str
    jira_id: str
    jira_link: str
    prd_url: str
    force_refresh: bool = False
    google_credentials: dict[str, Any] | None = None


@dataclass(frozen=True)
class PRDBriefingReviewRequest:
    owner_key: str
    prd_url: str
    language: str = "zh"
    force_refresh: bool = False
    selected_section_indexes: list[int] | None = None
    include_linked_spreadsheets: bool = True
    google_credentials: dict[str, Any] | None = None


class PRDReviewService:
    def __init__(
        self,
        *,
        store: BriefingStore,
        confluence: ConfluenceConnector,
        settings: Settings,
        workspace_root: Path,
    ) -> None:
        self.store = store
        self.confluence = confluence
        self.settings = settings
        self.workspace_root = Path(workspace_root)

    def review(self, request: PRDReviewRequest, *, progress_callback: PRDProgressCallback | None = None) -> dict[str, Any]:
        normalized = self._normalize_request(request)
        _emit_prd_progress(progress_callback, "reading_prd", "Reading PRD.", 0, 4)
        page = self.confluence.ingest_page(normalized.prd_url, "prd-review")
        if not page.sections:
            raise ToolError("PRD page did not contain readable sections.")
        linked_spreadsheet_evidence = _resolve_linked_spreadsheet_evidence(
            confluence=self.confluence,
            page=page,
            selected_section_indexes=list(range(1, len(page.sections) + 1)),
            google_credentials=normalized.google_credentials,
            include_linked_spreadsheets=True,
            google_sheet_cache_dir=_google_sheet_artifact_cache_dir(self.settings),
            progress_callback=progress_callback,
        )
        google_sheet_screenshot_evidence = _resolve_google_sheet_screenshot_evidence(
            confluence=self.confluence,
            page=page,
            selected_section_indexes=list(range(1, len(page.sections) + 1)),
            settings=self.settings,
            workspace_root=self.workspace_root,
            progress_callback=progress_callback,
        )
        coverage = _merge_google_sheet_screenshot_coverage(_merge_linked_spreadsheet_coverage(
            _build_generation_coverage(page, mode=_generation_mode_for_page(page)),
            linked_spreadsheet_evidence,
        ), google_sheet_screenshot_evidence)
        cache_jira_id = _cache_key_for_section_selection(
            normalized.jira_id,
            page=page,
            selected_section_indexes=None,
            linked_artifact_fingerprint=linked_spreadsheet_evidence["cache_fingerprint"],
            google_sheet_screenshot_fingerprint=google_sheet_screenshot_evidence["cache_fingerprint"],
        )
        cached = None if normalized.force_refresh else self.store.get_prd_review_result(
            owner_key=normalized.owner_key,
            jira_id=cache_jira_id,
            prd_url=page.source_url,
            prd_updated_at=page.updated_at,
            prompt_version=PRD_REVIEW_PROMPT_VERSION,
        )
        if cached and cached.get("status") == "completed" and cached.get("result_markdown"):
            return {"status": "ok", "cached": True, "review": cached, "prd": self._page_metadata(page), "coverage": coverage}
        try:
            _emit_prd_progress(progress_callback, "generating_review", "Generating review.", 3, 4)
            generated = self._generate_prd_review(
                jira_id=normalized.jira_id,
                jira_link=normalized.jira_link,
                prd_url=page.source_url,
                page=page,
                language="zh",
                prompt_version=PRD_REVIEW_PROMPT_VERSION,
                linked_spreadsheet_evidence=linked_spreadsheet_evidence,
                google_sheet_screenshot_evidence=google_sheet_screenshot_evidence,
            )
        except Exception as error:  # noqa: BLE001 - persist the failure for visible retry context.
            self.store.save_prd_review_result(
                owner_key=normalized.owner_key,
                jira_id=cache_jira_id,
                jira_link=normalized.jira_link,
                prd_url=page.source_url,
                prd_updated_at=page.updated_at,
                prompt_version=PRD_REVIEW_PROMPT_VERSION,
                status="failed",
                error=str(error),
            )
            raise ToolError(str(error)) from error

        review = self.store.save_prd_review_result(
            owner_key=normalized.owner_key,
            jira_id=cache_jira_id,
            jira_link=normalized.jira_link,
            prd_url=page.source_url,
            prd_updated_at=page.updated_at,
            prompt_version=PRD_REVIEW_PROMPT_VERSION,
            status="completed",
            result_markdown=generated["result_markdown"],
            model_id=generated["model_id"],
            trace=generated["trace"],
        )
        return {"status": "ok", "cached": False, "review": review, "prd": self._page_metadata(page), "coverage": coverage}

    def review_url(self, request: PRDBriefingReviewRequest, *, progress_callback: PRDProgressCallback | None = None) -> dict[str, Any]:
        normalized = self._normalize_briefing_review_request(request)
        _emit_prd_progress(progress_callback, "reading_prd", "Reading PRD.", 0, 4)
        page = self.confluence.ingest_page(normalized.prd_url, "prd-briefing-review")
        if not page.sections:
            raise ToolError("PRD page did not contain readable sections.")
        selected_page, selection = self._select_sections(page, normalized.selected_section_indexes)
        linked_spreadsheet_evidence = _resolve_linked_spreadsheet_evidence(
            confluence=self.confluence,
            page=selected_page,
            selected_section_indexes=selection["coverage"]["selected_section_indexes"],
            google_credentials=normalized.google_credentials,
            include_linked_spreadsheets=normalized.include_linked_spreadsheets,
            google_sheet_cache_dir=_google_sheet_artifact_cache_dir(self.settings),
            progress_callback=progress_callback,
        )
        google_sheet_screenshot_evidence = _resolve_google_sheet_screenshot_evidence(
            confluence=self.confluence,
            page=selected_page,
            selected_section_indexes=selection["coverage"]["selected_section_indexes"],
            settings=self.settings,
            workspace_root=self.workspace_root,
            progress_callback=progress_callback,
        )
        generation_coverage = _build_generation_coverage(
            selected_page,
            mode=_generation_mode_for_page(selected_page),
            total_sections=int(selection["coverage"].get("sections_total") or len(selected_page.sections)),
            section_indexes=selection["coverage"]["selected_section_indexes"],
        )
        coverage = _merge_google_sheet_screenshot_coverage(_merge_linked_spreadsheet_coverage(
            {**selection["coverage"], **generation_coverage},
            linked_spreadsheet_evidence,
        ), google_sheet_screenshot_evidence)
        prompt_version = prd_briefing_review_prompt_version(normalized.language)
        cache_jira_id = _cache_key_for_section_selection(
            PRD_BRIEFING_REVIEW_CACHE_KEY,
            page=page,
            selected_section_indexes=selection["cache_section_indexes"],
            linked_artifact_fingerprint=linked_spreadsheet_evidence["cache_fingerprint"],
            google_sheet_screenshot_fingerprint=google_sheet_screenshot_evidence["cache_fingerprint"],
        )
        cached = None if normalized.force_refresh else self.store.get_prd_review_result(
            owner_key=normalized.owner_key,
            jira_id=cache_jira_id,
            prd_url=page.source_url,
            prd_updated_at=page.updated_at,
            prompt_version=prompt_version,
        )
        if cached and cached.get("status") == "completed" and cached.get("result_markdown"):
            return {
                "status": "ok",
                "cached": True,
                "review": cached,
                "prd": self._page_metadata(page),
                "language": normalized.language,
                "coverage": coverage,
            }

        try:
            _emit_prd_progress(progress_callback, "generating_review", "Generating review.", 3, 4)
            generated = self._generate_prd_review(
                jira_id="",
                jira_link="",
                prd_url=page.source_url,
                page=selected_page,
                language=normalized.language,
                prompt_version=prompt_version,
                linked_spreadsheet_evidence=linked_spreadsheet_evidence,
                google_sheet_screenshot_evidence=google_sheet_screenshot_evidence,
            )
        except Exception as error:  # noqa: BLE001 - persist the failure for visible retry context.
            self.store.save_prd_review_result(
                owner_key=normalized.owner_key,
                jira_id=cache_jira_id,
                jira_link="",
                prd_url=page.source_url,
                prd_updated_at=page.updated_at,
                prompt_version=prompt_version,
                status="failed",
                error=str(error),
            )
            raise ToolError(str(error)) from error

        review = self.store.save_prd_review_result(
            owner_key=normalized.owner_key,
            jira_id=cache_jira_id,
            jira_link="",
            prd_url=page.source_url,
            prd_updated_at=page.updated_at,
            prompt_version=prompt_version,
            status="completed",
            result_markdown=generated["result_markdown"],
            model_id=generated["model_id"],
            trace=generated["trace"],
        )
        return {
            "status": "ok",
            "cached": False,
            "review": review,
            "prd": self._page_metadata(page),
            "language": normalized.language,
            "coverage": coverage,
        }

    def summarize(self, request: PRDReviewRequest, *, progress_callback: PRDProgressCallback | None = None) -> dict[str, Any]:
        normalized = self._normalize_request(request)
        _emit_prd_progress(progress_callback, "reading_prd", "Reading PRD.", 0, 2)
        page = self.confluence.ingest_page(normalized.prd_url, "prd-summary")
        if not page.sections:
            raise ToolError("PRD page did not contain readable sections.")
        cached = None if normalized.force_refresh else self.store.get_prd_review_result(
            owner_key=normalized.owner_key,
            jira_id=normalized.jira_id,
            prd_url=page.source_url,
            prd_updated_at=page.updated_at,
            prompt_version=PRD_SUMMARY_PROMPT_VERSION,
        )
        if cached and cached.get("status") == "completed" and cached.get("result_markdown"):
            return {
                "status": "ok",
                "cached": True,
                "summary": cached,
                "prd": self._page_metadata(page),
                "coverage": _build_generation_coverage(page, mode=_generation_mode_for_page(page)),
            }
        try:
            _emit_prd_progress(progress_callback, "generating_summary", "Generating summary.", 1, 2)
            generated = self._generate_prd_summary(
                jira_id=normalized.jira_id,
                jira_link=normalized.jira_link,
                prd_url=page.source_url,
                page=page,
                language="zh",
                prompt_version=PRD_SUMMARY_PROMPT_VERSION,
            )
        except Exception as error:  # noqa: BLE001 - persist the failure for visible retry context.
            self.store.save_prd_review_result(
                owner_key=normalized.owner_key,
                jira_id=normalized.jira_id,
                jira_link=normalized.jira_link,
                prd_url=page.source_url,
                prd_updated_at=page.updated_at,
                prompt_version=PRD_SUMMARY_PROMPT_VERSION,
                status="failed",
                error=str(error),
            )
            raise ToolError(str(error)) from error

        summary = self.store.save_prd_review_result(
            owner_key=normalized.owner_key,
            jira_id=normalized.jira_id,
            jira_link=normalized.jira_link,
            prd_url=page.source_url,
            prd_updated_at=page.updated_at,
            prompt_version=PRD_SUMMARY_PROMPT_VERSION,
            status="completed",
            result_markdown=generated["result_markdown"],
            model_id=generated["model_id"],
            trace=generated["trace"],
        )
        return {
            "status": "ok",
            "cached": False,
            "summary": summary,
            "prd": self._page_metadata(page),
            "coverage": _build_generation_coverage(page, mode=_generation_mode_for_page(page)),
        }

    def summarize_url(self, request: PRDBriefingReviewRequest, *, progress_callback: PRDProgressCallback | None = None) -> dict[str, Any]:
        normalized = self._normalize_briefing_review_request(request)
        _emit_prd_progress(progress_callback, "reading_prd", "Reading PRD.", 0, 2)
        page = self.confluence.ingest_page(normalized.prd_url, "prd-self-assessment-summary")
        if not page.sections:
            raise ToolError("PRD page did not contain readable sections.")
        prompt_version = prd_summary_prompt_version(normalized.language)
        cached = None if normalized.force_refresh else self.store.get_prd_review_result(
            owner_key=normalized.owner_key,
            jira_id=PRD_URL_SUMMARY_CACHE_KEY,
            prd_url=page.source_url,
            prd_updated_at=page.updated_at,
            prompt_version=prompt_version,
        )
        if cached and cached.get("status") == "completed" and cached.get("result_markdown"):
            return {
                "status": "ok",
                "cached": True,
                "summary": cached,
                "prd": self._page_metadata(page),
                "language": normalized.language,
                "coverage": _build_generation_coverage(page, mode=_generation_mode_for_page(page)),
            }
        try:
            _emit_prd_progress(progress_callback, "generating_summary", "Generating summary.", 1, 2)
            generated = self._generate_prd_summary(
                jira_id="",
                jira_link="",
                prd_url=page.source_url,
                page=page,
                language=normalized.language,
                prompt_version=prompt_version,
            )
        except Exception as error:  # noqa: BLE001 - persist the failure for visible retry context.
            self.store.save_prd_review_result(
                owner_key=normalized.owner_key,
                jira_id=PRD_URL_SUMMARY_CACHE_KEY,
                jira_link="",
                prd_url=page.source_url,
                prd_updated_at=page.updated_at,
                prompt_version=prompt_version,
                status="failed",
                error=str(error),
            )
            raise ToolError(str(error)) from error

        summary = self.store.save_prd_review_result(
            owner_key=normalized.owner_key,
            jira_id=PRD_URL_SUMMARY_CACHE_KEY,
            jira_link="",
            prd_url=page.source_url,
            prd_updated_at=page.updated_at,
            prompt_version=prompt_version,
            status="completed",
            result_markdown=generated["result_markdown"],
            model_id=generated["model_id"],
            trace=generated["trace"],
        )
        return {
            "status": "ok",
            "cached": False,
            "summary": summary,
            "prd": self._page_metadata(page),
            "language": normalized.language,
            "coverage": _build_generation_coverage(page, mode=_generation_mode_for_page(page)),
        }

    def _generate_prd_review(
        self,
        *,
        jira_id: str,
        jira_link: str,
        prd_url: str,
        page: IngestedConfluencePage,
        language: str,
        prompt_version: str,
        linked_spreadsheet_evidence: dict[str, Any] | None,
        google_sheet_screenshot_evidence: dict[str, Any] | None,
    ) -> dict[str, Any]:
        if _generation_mode_for_page(page) == "single":
            prompt = build_prd_review_prompt(
                jira_id=jira_id,
                jira_link=jira_link,
                prd_url=prd_url,
                page=page,
                language=language,
                linked_spreadsheet_evidence=linked_spreadsheet_evidence,
                google_sheet_screenshot_evidence=google_sheet_screenshot_evidence,
            )
            generated = generate_prd_review_with_codex(
                prompt=prompt,
                settings=self.settings,
                workspace_root=self.workspace_root,
                language=language,
                prompt_version=prompt_version,
            )
            generated["result_markdown"] = _postprocess_prd_review_markdown(
                generated.get("result_markdown") or "",
                page=page,
            )
            return generated

        batch_outputs: list[dict[str, Any]] = []
        traces: list[Any] = []
        batches = _split_prd_page_batches(page)
        for batch_index, batch_page in enumerate(batches, start=1):
            prompt = build_prd_review_batch_prompt(
                jira_id=jira_id,
                jira_link=jira_link,
                prd_url=prd_url,
                page=batch_page,
                language=language,
                batch_index=batch_index,
                batch_total=len(batches),
            )
            generated = generate_prd_review_with_codex(
                prompt=prompt,
                settings=self.settings,
                workspace_root=self.workspace_root,
                language=language,
                prompt_version=f"{prompt_version}_batch",
            )
            batch_outputs.append(
                {
                    "batch_index": batch_index,
                    "section_titles": [section.section_path or section.title for section in batch_page.sections],
                    "result_markdown": generated["result_markdown"],
                }
            )
            traces.append(generated.get("trace") or {})
        synthesis_prompt = build_prd_review_synthesis_prompt(
            jira_id=jira_id,
            jira_link=jira_link,
            prd_url=prd_url,
            page=page,
            language=language,
            batch_outputs=batch_outputs,
            linked_spreadsheet_evidence=linked_spreadsheet_evidence,
            google_sheet_screenshot_evidence=google_sheet_screenshot_evidence,
        )
        final = generate_prd_review_with_codex(
            prompt=synthesis_prompt,
            settings=self.settings,
            workspace_root=self.workspace_root,
            language=language,
            prompt_version=f"{prompt_version}_synthesis",
        )
        final["result_markdown"] = _postprocess_prd_review_markdown(
            final.get("result_markdown") or "",
            page=page,
        )
        final["trace"] = {"hybrid": True, "batch_count": len(batches), "batch_traces": traces, "final_trace": final.get("trace") or {}}
        return final

    def _generate_prd_summary(
        self,
        *,
        jira_id: str,
        jira_link: str,
        prd_url: str,
        page: IngestedConfluencePage,
        language: str,
        prompt_version: str,
    ) -> dict[str, Any]:
        if _generation_mode_for_page(page) == "single":
            prompt = build_prd_summary_prompt(
                jira_id=jira_id,
                jira_link=jira_link,
                prd_url=prd_url,
                page=page,
                language=language,
            )
            return generate_prd_summary_with_codex(
                prompt=prompt,
                settings=self.settings,
                workspace_root=self.workspace_root,
                language=language,
                prompt_version=prompt_version,
            )

        batch_outputs: list[dict[str, Any]] = []
        traces: list[Any] = []
        batches = _split_prd_page_batches(page)
        for batch_index, batch_page in enumerate(batches, start=1):
            prompt = build_prd_summary_batch_prompt(
                jira_id=jira_id,
                jira_link=jira_link,
                prd_url=prd_url,
                page=batch_page,
                language=language,
                batch_index=batch_index,
                batch_total=len(batches),
            )
            generated = generate_prd_summary_with_codex(
                prompt=prompt,
                settings=self.settings,
                workspace_root=self.workspace_root,
                language=language,
                prompt_version=f"{prompt_version}_batch",
            )
            batch_outputs.append(
                {
                    "batch_index": batch_index,
                    "section_titles": [section.section_path or section.title for section in batch_page.sections],
                    "result_markdown": generated["result_markdown"],
                }
            )
            traces.append(generated.get("trace") or {})
        synthesis_prompt = build_prd_summary_synthesis_prompt(
            jira_id=jira_id,
            jira_link=jira_link,
            prd_url=prd_url,
            page=page,
            language=language,
            batch_outputs=batch_outputs,
        )
        final = generate_prd_summary_with_codex(
            prompt=synthesis_prompt,
            settings=self.settings,
            workspace_root=self.workspace_root,
            language=language,
            prompt_version=f"{prompt_version}_synthesis",
        )
        final["trace"] = {"hybrid": True, "batch_count": len(batches), "batch_traces": traces, "final_trace": final.get("trace") or {}}
        return final

    def list_url_sections(self, request: PRDBriefingReviewRequest) -> dict[str, Any]:
        normalized = self._normalize_briefing_review_request(request)
        page = self.confluence.ingest_page(normalized.prd_url, "prd-self-assessment-sections")
        if not page.sections:
            raise ToolError("PRD page did not contain readable sections.")
        return {
            "status": "ok",
            "prd": self._page_metadata(page),
            "sections": [
                {
                    "index": index,
                    "title": section.section_path or section.title or f"Section {index}",
                    "char_count": len(str(section.content or "")),
                    "long": len(str(section.content or "")) >= PRD_SECTION_LONG_CHAR_THRESHOLD,
                    "linked_spreadsheet_count": len(getattr(section, "spreadsheet_links", []) or []),
                }
                for index, section in enumerate(page.sections, start=1)
            ],
        }

    @staticmethod
    def _normalize_request(request: PRDReviewRequest) -> PRDReviewRequest:
        owner_key = str(request.owner_key or "").strip()
        jira_id = str(request.jira_id or "").strip()
        prd_url = str(request.prd_url or "").strip()
        jira_link = str(request.jira_link or "").strip()
        if not owner_key:
            raise ToolError("Owner identity is required.")
        if not jira_id:
            raise ToolError("Jira ID is required.")
        if not prd_url:
            raise ToolError("PRD link is required.")
        if not prd_url.lower().startswith(("http://", "https://")):
            raise ToolError("PRD link must be an HTTP or HTTPS URL.")
        return PRDReviewRequest(
            owner_key=owner_key,
            jira_id=jira_id,
            jira_link=jira_link,
            prd_url=prd_url,
            force_refresh=bool(request.force_refresh),
            google_credentials=dict(request.google_credentials or {}) or None,
        )

    @staticmethod
    def _normalize_briefing_review_request(request: PRDBriefingReviewRequest) -> PRDBriefingReviewRequest:
        owner_key = str(request.owner_key or "").strip()
        prd_url = str(request.prd_url or "").strip()
        language = normalize_prd_review_language(request.language)
        selected_section_indexes = _normalize_selected_section_indexes(request.selected_section_indexes)
        if not owner_key:
            raise ToolError("Owner identity is required.")
        if not prd_url:
            raise ToolError("PRD link is required.")
        if not prd_url.lower().startswith(("http://", "https://")):
            raise ToolError("PRD link must be an HTTP or HTTPS URL.")
        return PRDBriefingReviewRequest(
            owner_key=owner_key,
            prd_url=prd_url,
            language=language,
            force_refresh=bool(request.force_refresh),
            selected_section_indexes=selected_section_indexes,
            include_linked_spreadsheets=bool(request.include_linked_spreadsheets),
            google_credentials=dict(request.google_credentials or {}) or None,
        )

    @staticmethod
    def _select_sections(page: IngestedConfluencePage, selected_section_indexes: list[int] | None) -> tuple[IngestedConfluencePage, dict[str, Any]]:
        total_sections = len(page.sections)
        all_indexes = list(range(1, total_sections + 1))
        if selected_section_indexes is None:
            effective_indexes = all_indexes
            cache_section_indexes = None
        else:
            if not selected_section_indexes:
                raise ToolError("Select at least one PRD section to review.")
            invalid = [index for index in selected_section_indexes if index < 1 or index > total_sections]
            if invalid:
                raise ToolError(f"Selected PRD section index is out of range: {invalid[0]}.")
            effective_indexes = selected_section_indexes
            cache_section_indexes = None if effective_indexes == all_indexes else effective_indexes

        selected_sections = [page.sections[index - 1] for index in effective_indexes]
        selected_page = IngestedConfluencePage(
            page_id=page.page_id,
            title=page.title,
            source_url=page.source_url,
            updated_at=page.updated_at,
            language=page.language,
            sections=selected_sections,
            version_number=page.version_number,
            ancestor_titles=list(getattr(page, "ancestor_titles", []) or []),
            media_dict=page.media_dict,
            presentation_source_text=page.presentation_source_text,
        )
        coverage = _build_section_coverage(
            total_sections=total_sections,
            selected_indexes=effective_indexes,
            selected_sections=selected_sections,
            selection_hash=_section_selection_hash(effective_indexes) if cache_section_indexes else "",
        )
        return selected_page, {"coverage": coverage, "cache_section_indexes": cache_section_indexes}

    @staticmethod
    def _page_metadata(page: IngestedConfluencePage) -> dict[str, Any]:
        return {
            "title": page.title,
            "source_url": page.source_url,
            "updated_at": page.updated_at,
            "page_id": page.page_id,
            "ancestor_titles": list(getattr(page, "ancestor_titles", []) or []),
        }


def _build_prd_source(page: IngestedConfluencePage, *, max_chars: int = PRD_REVIEW_MAX_SOURCE_CHARS) -> str:
    sections = []
    used_chars = 0
    for index, section in enumerate(page.sections, start=1):
        content = _build_review_section_content(page=page, section=section)
        if not content:
            continue
        block = f"## Section {index}: {section.section_path}\n{content}"
        remaining = max_chars - used_chars
        if remaining <= 0:
            break
        if len(block) > remaining:
            block = block[:remaining].rstrip() + "\n[Truncated because the PRD is long.]"
        sections.append(block)
        used_chars += len(block)
    source = "\n\n".join(sections).strip()
    if not source:
        raise ToolError("PRD page did not contain readable text.")
    return source


def _build_review_section_content(*, page: IngestedConfluencePage, section: Any) -> str:
    content = str(getattr(section, "content", "") or "").strip()
    media_blocks: list[str] = []
    media_dict = getattr(page, "media_dict", {}) or {}
    for media_ref in getattr(section, "media_refs", []) or []:
        media_id = str(media_ref or "").strip()
        if not media_id:
            continue
        media = media_dict.get(media_id) or {}
        if str(media.get("type") or "") != "table":
            continue
        table_text = _format_table_media_for_review(media_id, media)
        if not table_text:
            continue
        placeholder = f"[{media_id}]"
        if placeholder in content:
            content = content.replace(placeholder, table_text)
        else:
            media_blocks.append(table_text)
    return "\n".join(part for part in [content, *media_blocks] if str(part or "").strip()).strip()


def _format_table_media_for_review(media_id: str, media: dict[str, Any]) -> str:
    html = str(media.get("content") or "")
    if not html.strip():
        return ""
    soup = BeautifulSoup(html, "html.parser")
    rows: list[str] = []
    for row in soup.find_all("tr"):
        cells = [
            re.sub(r"\s+", " ", cell.get_text(" ", strip=True)).strip()
            for cell in row.find_all(["th", "td"], recursive=False)
        ]
        cells = [cell for cell in cells if cell]
        if not cells:
            continue
        rows.append(" | ".join(cells))
        if len(rows) >= PRD_REVIEW_TABLE_MEDIA_MAX_ROWS:
            rows.append("[Table truncated after 120 rows.]")
            break
    if not rows:
        fallback = re.sub(r"\n{3,}", "\n\n", soup.get_text("\n", strip=True)).strip()
        if not fallback:
            return ""
        rows = [fallback]
    table_text = "\n".join(rows).strip()
    if len(table_text) > PRD_REVIEW_TABLE_MEDIA_MAX_CHARS:
        table_text = table_text[:PRD_REVIEW_TABLE_MEDIA_MAX_CHARS].rstrip() + "\n[Table truncated because it is long.]"
    return f"[{media_id} table content]\n{table_text}"


def _postprocess_prd_review_markdown(markdown: str, *, page: IngestedConfluencePage) -> str:
    text = str(markdown or "")
    if not text.strip() or "### Secondary Clarifications" not in text:
        return text
    source_text = _page_plain_text_for_postprocessing(page)
    if not source_text.strip():
        return text
    source_identifier_text = _normalize_identifier_for_comparison(source_text)

    lines = text.splitlines()
    output: list[str] = []
    index = 0
    while index < len(lines):
        line = lines[index]
        if not line.startswith("- **Priority") and not line.startswith("- **优先级"):
            output.append(line)
            index += 1
            continue

        block = [line]
        index += 1
        while index < len(lines) and not lines[index].startswith("- **Priority") and not lines[index].startswith("- **优先级"):
            if lines[index].startswith("### "):
                break
            block.append(lines[index])
            index += 1

        block_text = "\n".join(block)
        if not _is_spurious_identifier_typo_finding(block_text, source_text=source_text, source_identifier_text=source_identifier_text):
            output.extend(block)
        if index < len(lines) and lines[index].startswith("### "):
            continue
    return "\n".join(output).strip() + ("\n" if text.endswith("\n") else "")


def _page_plain_text_for_postprocessing(page: IngestedConfluencePage) -> str:
    parts = [str(page.title or "")]
    for section in page.sections:
        parts.append(str(section.section_path or section.title or ""))
        parts.append(str(section.content or ""))
    return "\n".join(parts)


def _is_spurious_identifier_typo_finding(block_text: str, *, source_text: str, source_identifier_text: str) -> bool:
    lowered = block_text.casefold()
    typo_keywords = (
        "fix typo",
        "fix the section title typo",
        "typo",
        "spelling",
        "misspell",
        "修正拼写",
        "拼写",
        "错别字",
        "文案清理",
    )
    if not any(keyword in lowered for keyword in typo_keywords):
        return False
    if not ("p2" in lowered or "priority" in lowered or "优先级" in lowered):
        return False

    source_exact = source_text.casefold()
    for quoted in re.findall(r"`([^`]+)`", block_text):
        if quoted.strip() and quoted.casefold() in source_exact:
            return False

    candidates = _identifier_candidates_from_text(block_text)
    for display_value, normalized in candidates:
        if not normalized or len(normalized) < 8:
            continue
        if normalized not in source_identifier_text:
            continue
        if display_value.casefold() in source_exact:
            continue
        return True
    return False


def _identifier_candidates_from_text(value: str) -> list[tuple[str, str]]:
    tokens = re.findall(r"[A-Za-z][A-Za-z0-9_]*", value)
    candidates: list[tuple[str, str]] = []
    for window_size in range(1, 5):
        for start in range(0, max(len(tokens) - window_size + 1, 0)):
            window = tokens[start:start + window_size]
            joined = "".join(window)
            if not _looks_like_identifier(joined):
                continue
            display_value = " ".join(window)
            candidates.append((display_value, _normalize_identifier_for_comparison(joined)))
    return candidates


def _looks_like_identifier(value: str) -> bool:
    if len(value) < 8:
        return False
    if "_" in value or any(ch.isdigit() for ch in value):
        return True
    return sum(1 for ch in value if ch.isupper()) >= 2 and any(ch.islower() for ch in value)


def _normalize_identifier_for_comparison(value: str) -> str:
    return re.sub(r"[^a-z0-9_]+", "", str(value or "").casefold())


def _build_prd_source_block(section: Any, index: int) -> str:
    content = str(getattr(section, "content", "") or "").strip()
    title = str(getattr(section, "section_path", "") or getattr(section, "title", "") or f"Section {index}")
    return f"## Section {index}: {title}\n{content}" if content else ""


def _prd_source_char_count(page: IngestedConfluencePage) -> int:
    return sum(len(_build_prd_source_block(section, index)) for index, section in enumerate(page.sections, start=1))


def _generation_mode_for_page(page: IngestedConfluencePage) -> str:
    return "hybrid" if _prd_source_char_count(page) > PRD_REVIEW_MAX_SOURCE_CHARS else "single"


def _build_generation_coverage(
    page: IngestedConfluencePage,
    *,
    mode: str,
    total_sections: int | None = None,
    section_indexes: list[int] | None = None,
) -> dict[str, Any]:
    covered_sections = sum(1 for section in page.sections if str(section.content or "").strip())
    table_media = _collect_table_media_coverage(page, section_indexes=section_indexes)
    return {
        "mode": "hybrid" if mode == "hybrid" else "single",
        "sections_total": int(total_sections if total_sections is not None else len(page.sections)),
        "sections_covered": covered_sections,
        "truncated": False if mode == "hybrid" else _prd_source_char_count(page) > PRD_REVIEW_MAX_SOURCE_CHARS,
        "confluence_tables_total": len(table_media),
        "confluence_tables_reviewed": len(table_media),
        "confluence_tables": table_media,
    }


def _collect_table_media_coverage(page: IngestedConfluencePage, *, section_indexes: list[int] | None = None) -> list[dict[str, Any]]:
    media_dict = getattr(page, "media_dict", {}) or {}
    seen: set[str] = set()
    tables: list[dict[str, Any]] = []
    effective_indexes = section_indexes if section_indexes and len(section_indexes) == len(page.sections) else list(range(1, len(page.sections) + 1))
    for section_index, section in zip(effective_indexes, page.sections):
        for media_ref in getattr(section, "media_refs", []) or []:
            media_id = str(media_ref or "").strip()
            if not media_id or media_id in seen:
                continue
            media = media_dict.get(media_id) or {}
            if str(media.get("type") or "") != "table":
                continue
            table_text = _format_table_media_for_review(media_id, media)
            if not table_text:
                continue
            seen.add(media_id)
            row_count = max(len(table_text.splitlines()) - 1, 0)
            tables.append(
                {
                    "media_id": media_id,
                    "source_section_index": section_index,
                    "source_section_title": str(getattr(section, "section_path", "") or getattr(section, "title", "") or f"Section {section_index}"),
                    "row_count": row_count,
                    "char_count": len(table_text),
                }
            )
    return tables


def _page_with_sections(page: IngestedConfluencePage, sections: list[Any]) -> IngestedConfluencePage:
    return IngestedConfluencePage(
        page_id=page.page_id,
        title=page.title,
        source_url=page.source_url,
        updated_at=page.updated_at,
        language=page.language,
        sections=sections,
        version_number=page.version_number,
        ancestor_titles=list(getattr(page, "ancestor_titles", []) or []),
        media_dict=page.media_dict,
        presentation_source_text=page.presentation_source_text,
    )


def _split_prd_page_batches(page: IngestedConfluencePage, *, max_chars: int = PRD_HYBRID_BATCH_SOURCE_CHARS) -> list[IngestedConfluencePage]:
    batches: list[list[Any]] = []
    current: list[Any] = []
    current_chars = 0
    for index, section in enumerate(page.sections, start=1):
        block_chars = len(_build_prd_source_block(section, index))
        if current and current_chars + block_chars > max_chars:
            batches.append(current)
            current = []
            current_chars = 0
        current.append(section)
        current_chars += block_chars
    if current:
        batches.append(current)
    if not batches:
        raise ToolError("PRD page did not contain readable text.")
    return [_page_with_sections(page, sections) for sections in batches]


def _emit_prd_progress(
    progress_callback: PRDProgressCallback | None,
    stage: str,
    message: str,
    current: int,
    total: int,
) -> None:
    if progress_callback is None:
        return
    progress_callback(stage, message, current, total)


def _resolve_linked_spreadsheet_evidence(
    *,
    confluence: ConfluenceConnector,
    page: IngestedConfluencePage,
    selected_section_indexes: list[int],
    google_credentials: dict[str, Any] | None,
    include_linked_spreadsheets: bool,
    google_sheet_cache_dir: Path | None = None,
    progress_callback: PRDProgressCallback | None = None,
) -> dict[str, Any]:
    if not include_linked_spreadsheets:
        return {"artifacts": [], "cache_fingerprint": ""}

    artifacts: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    for section_index, section in zip(selected_section_indexes, page.sections):
        if not _is_report_related_section(section):
            continue
        for raw_link in getattr(section, "spreadsheet_links", []) or []:
            url = str(getattr(raw_link, "url", "") or "").strip()
            if not url or url in seen_urls:
                continue
            seen_urls.add(url)
            artifacts.append(
                {
                    "title": str(getattr(raw_link, "title", "") or getattr(raw_link, "filename", "") or url),
                    "url": url,
                    "filename": str(getattr(raw_link, "filename", "") or ""),
                    "source_section_index": section_index,
                    "source_section_title": section.section_path or section.title or f"Section {section_index}",
                    "status": "pending",
                    "reason": "",
                    "text": "",
                    "content_hash": "",
                    "metadata_hash": "",
                    "sheet_count": 0,
                    "sheets_extracted": [],
                    "skipped_sheet_count": 0,
                    "template_metadata": {},
                }
            )

    remaining_text_chars = PRD_LINKED_SPREADSHEET_MAX_TEXT_CHARS
    total = len(artifacts)
    if total:
        _emit_prd_progress(progress_callback, "reading_report_templates", f"Reading {total} report templates.", 1, 4)
    for index, artifact in enumerate(artifacts, start=1):
        _emit_prd_progress(
            progress_callback,
            "reading_report_templates",
            f"Reading {index}/{total} report templates.",
            min(index, total),
            max(total, 1),
        )
        _resolve_one_linked_spreadsheet(
            confluence=confluence,
            artifact=artifact,
            google_credentials=google_credentials,
            remaining_text_chars=max(remaining_text_chars, 0),
            google_sheet_cache_dir=google_sheet_cache_dir,
            progress_callback=progress_callback,
            artifact_index=index,
            artifact_total=total,
        )
        if artifact.get("status") == "ok":
            remaining_text_chars -= len(str(artifact.get("text") or ""))
    if total:
        _emit_prd_progress(progress_callback, "analyzing_template_metadata", "Analyzing template metadata.", 2, 4)

    fingerprint = "|".join(
        f"{item.get('source_section_index')}:{item.get('url')}:{item.get('status')}:{item.get('reason')}:{item.get('content_hash')}:{item.get('metadata_hash')}:{item.get('sheet_count')}:{item.get('skipped_sheet_count')}"
        for item in artifacts
    )
    return {"artifacts": artifacts, "cache_fingerprint": fingerprint}


def _is_report_related_section(section: Any) -> bool:
    haystack = f"{getattr(section, 'section_path', '')} {getattr(section, 'title', '')}".casefold()
    return any(keyword.casefold() in haystack for keyword in PRD_REPORT_SECTION_KEYWORDS)


def _is_report_or_template_prd(page: IngestedConfluencePage, linked_spreadsheet_evidence: dict[str, Any] | None = None) -> bool:
    if any(_is_report_related_section(section) for section in page.sections):
        return True
    return bool((linked_spreadsheet_evidence or {}).get("artifacts"))


def _is_anti_fraud_prd(page: IngestedConfluencePage) -> bool:
    ancestor_titles = [str(title or "") for title in (getattr(page, "ancestor_titles", []) or [])]
    ancestor_haystack = " > ".join(ancestor_titles).casefold()
    if _contains_anti_fraud_marker(ancestor_haystack):
        return True
    return _contains_anti_fraud_marker(str(page.title or "").casefold())


def _contains_anti_fraud_marker(value: str) -> bool:
    haystack = str(value or "").casefold()
    if any(keyword in haystack for keyword in PRD_ANTI_FRAUD_KEYWORDS):
        return True
    return re.search(r"(?<![a-z0-9])af(?![a-z0-9])", haystack) is not None


def _resolve_google_sheet_screenshot_evidence(
    *,
    confluence: ConfluenceConnector,
    page: IngestedConfluencePage,
    selected_section_indexes: list[int],
    settings: Settings,
    workspace_root: Path,
    progress_callback: PRDProgressCallback | None = None,
) -> dict[str, Any]:
    # Image OCR/vision evidence is intentionally disabled. Confluence table media is
    # expanded as text in the PRD source, which covers copied spreadsheet-like tables
    # without scanning every embedded image.
    return {"enabled": False, "artifacts": [], "cache_fingerprint": ""}

    if not _is_anti_fraud_prd(page):
        return {"enabled": False, "artifacts": [], "cache_fingerprint": ""}

    artifacts = _collect_section_image_artifacts(page=page, selected_section_indexes=selected_section_indexes)
    if artifacts:
        _emit_prd_progress(
            progress_callback,
            "reading_google_sheet_screenshots",
            f"Reading {len(artifacts)} PRD images for Google Sheet screenshots.",
            0,
            max(len(artifacts), 1),
        )

    cache_dir = _google_sheet_screenshot_cache_dir(settings)
    for index, artifact in enumerate(artifacts, start=1):
        _emit_prd_progress(
            progress_callback,
            "reading_google_sheet_screenshots",
            f"Reading image {index}/{len(artifacts)} for Google Sheet screenshot evidence.",
            index,
            max(len(artifacts), 1),
        )
        _resolve_one_google_sheet_screenshot(
            confluence=confluence,
            artifact=artifact,
            cache_dir=cache_dir,
            workspace_root=workspace_root,
            settings=settings,
        )

    fingerprint = "|".join(
        f"{item.get('source_section_index')}:{item.get('image_id')}:{item.get('url')}:{item.get('status')}:{item.get('reason')}:{item.get('content_hash')}:{item.get('evidence_hash')}:{item.get('is_google_sheet_screenshot')}"
        for item in artifacts
    )
    return {"enabled": True, "artifacts": artifacts, "cache_fingerprint": fingerprint}


def _collect_section_image_artifacts(*, page: IngestedConfluencePage, selected_section_indexes: list[int]) -> list[dict[str, Any]]:
    artifacts: list[dict[str, Any]] = []
    seen_urls: set[str] = set()
    media_dict = page.media_dict or {}
    for section_index, section in zip(selected_section_indexes, page.sections):
        title = str(section.section_path or section.title or f"Section {section_index}")
        media_refs = [str(ref or "").strip() for ref in (getattr(section, "media_refs", []) or []) if str(ref or "").strip()]
        for media_ref in media_refs:
            media = media_dict.get(media_ref) or {}
            if str(media.get("type") or "") != "image":
                continue
            url = str(media.get("source_url") or "").strip()
            if not url or url in seen_urls:
                continue
            seen_urls.add(url)
            artifacts.append(
                {
                    "image_id": media_ref,
                    "url": url,
                    "source_section_index": section_index,
                    "source_section_title": title,
                    "status": "pending",
                    "reason": "",
                    "is_google_sheet_screenshot": False,
                    "evidence_text": "",
                    "content_hash": "",
                    "evidence_hash": "",
                    "used_in_findings": False,
                }
            )
        for image_index, url in enumerate(getattr(section, "image_refs", []) or [], start=1):
            url = str(url or "").strip()
            if not url or url in seen_urls:
                continue
            seen_urls.add(url)
            artifacts.append(
                {
                    "image_id": f"IMAGE_{section_index}_{image_index}",
                    "url": url,
                    "source_section_index": section_index,
                    "source_section_title": title,
                    "status": "pending",
                    "reason": "",
                    "is_google_sheet_screenshot": False,
                    "evidence_text": "",
                    "content_hash": "",
                    "evidence_hash": "",
                    "used_in_findings": False,
                }
            )
    return artifacts


def _resolve_one_google_sheet_screenshot(
    *,
    confluence: ConfluenceConnector,
    artifact: dict[str, Any],
    cache_dir: Path | None,
    workspace_root: Path,
    settings: Settings,
) -> None:
    url = str(artifact.get("url") or "")
    try:
        response = confluence._request(  # noqa: SLF001 - local PRD connector owns authenticated Confluence fetches.
            url,
            accept="image/png,image/jpeg,image/webp,image/gif,image/*,*/*;q=0.8",
        )
        content = bytes(getattr(response, "content", b"") or b"")
        if not content:
            artifact.update({"status": "failed", "reason": "empty_download"})
            return
        content_hash = hashlib.sha256(content).hexdigest()[:16]
        artifact["content_hash"] = content_hash
        cached = _load_cached_google_sheet_screenshot_evidence(
            url=url,
            content_hash=content_hash,
            cache_dir=cache_dir,
        )
        if cached is not None:
            artifact.update(cached, {"cache_hit": True})
            return
        extracted = _extract_google_sheet_screenshot_evidence_from_image(
            image_bytes=content,
            image_url=url,
            image_id=str(artifact.get("image_id") or ""),
            source_section=str(artifact.get("source_section_title") or ""),
            settings=settings,
            workspace_root=workspace_root,
        )
        evidence_text = str(extracted.get("evidence_text") or "").strip()
        status = "ok" if bool(extracted.get("is_google_sheet_screenshot")) else "skipped"
        reason = "" if status == "ok" else str(extracted.get("reason") or "not_google_sheet_screenshot")
        evidence_hash = hashlib.sha256(
            json.dumps(
                {
                    "is_google_sheet_screenshot": bool(extracted.get("is_google_sheet_screenshot")),
                    "reason": reason,
                    "evidence_text": evidence_text,
                },
                ensure_ascii=False,
                sort_keys=True,
            ).encode("utf-8")
        ).hexdigest()[:16]
        payload = {
            "status": status,
            "reason": reason,
            "is_google_sheet_screenshot": status == "ok",
            "evidence_text": evidence_text if status == "ok" else "",
            "evidence_hash": evidence_hash,
            "classification": str(extracted.get("classification") or ""),
        }
        artifact.update(payload)
        _store_cached_google_sheet_screenshot_evidence(
            url=url,
            content_hash=content_hash,
            cache_dir=cache_dir,
            extracted=payload,
        )
    except ToolError as error:
        artifact.update({"status": "failed", "reason": str(error)})
    except Exception as error:  # noqa: BLE001 - image evidence gaps should not block the PRD review.
        artifact.update({"status": "failed", "reason": f"ocr_failed: {error}"})


def _google_sheet_screenshot_cache_key(*, url: str, content_hash: str) -> str:
    digest = hashlib.sha256(
        json.dumps(
            {
                "version": PRD_GOOGLE_SHEET_SCREENSHOT_EVIDENCE_CACHE_VERSION,
                "url": str(url or ""),
                "content_hash": str(content_hash or ""),
            },
            ensure_ascii=False,
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()
    return digest


def _load_cached_google_sheet_screenshot_evidence(
    *,
    url: str,
    content_hash: str,
    cache_dir: Path | None,
) -> dict[str, Any] | None:
    if cache_dir is None:
        return None
    cache_path = cache_dir / f"{_google_sheet_screenshot_cache_key(url=url, content_hash=content_hash)}.json"
    if not cache_path.exists():
        return None
    try:
        payload = json.loads(cache_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if payload.get("version") != PRD_GOOGLE_SHEET_SCREENSHOT_EVIDENCE_CACHE_VERSION:
        return None
    if payload.get("url") != url or payload.get("content_hash") != content_hash:
        return None
    extracted = payload.get("extracted")
    return extracted if isinstance(extracted, dict) else None


def _store_cached_google_sheet_screenshot_evidence(
    *,
    url: str,
    content_hash: str,
    cache_dir: Path | None,
    extracted: dict[str, Any],
) -> None:
    if cache_dir is None:
        return
    payload = {
        "version": PRD_GOOGLE_SHEET_SCREENSHOT_EVIDENCE_CACHE_VERSION,
        "url": url,
        "content_hash": content_hash,
        "extracted": extracted,
    }
    try:
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache_path = cache_dir / f"{_google_sheet_screenshot_cache_key(url=url, content_hash=content_hash)}.json"
        cache_path.write_text(json.dumps(payload, ensure_ascii=False, sort_keys=True), encoding="utf-8")
    except OSError:
        return


def _extract_google_sheet_screenshot_evidence_from_image(
    *,
    image_bytes: bytes,
    image_url: str,
    image_id: str,
    source_section: str,
    settings: Settings,
    workspace_root: Path,
) -> dict[str, Any]:
    suffix = _image_suffix_from_url(image_url)
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=True) as image_file:
        image_file.write(image_bytes)
        image_file.flush()
        prompt = f"""# Task
Classify this PRD image and extract evidence only if it is a screenshot of Google Sheets or a spreadsheet-like Google Sheet template.

Treat the image as a Google Sheet screenshot only when visible evidence shows spreadsheet grid cells, row/column headers, sheet tabs, formula bar/toolbar, Google Sheets UI labels, or a clear table-template layout copied from Google Sheets.

If it is a normal product screenshot, flow diagram, icon, avatar, decorative image, or non-spreadsheet UI, return is_google_sheet_screenshot=false and reason="not_google_sheet_screenshot".

If true, extract compact evidence for PRD review: visible sheet/tab names, field/header names, column structure, sample values, empty/duplicate headers, multi-level header or merged-cell signs, formula/format risks, unclear text with uncertainty markers, and any QA/template mapping concerns. Do not invent unreadable text.

Return strict JSON only:
{{
  "is_google_sheet_screenshot": true,
  "reason": "",
  "classification": "why this is or is not a Google Sheet screenshot",
  "evidence_text": "Markdown evidence for PRD review"
}}

# Image Context
- Image ID: {image_id or "-"}
- Source section: {source_section or "-"}
- Source URL: {image_url or "-"}
"""
        generated = _generate_with_codex(
            prompt=prompt,
            settings=settings,
            workspace_root=workspace_root,
            system_text=(
                "You extract visible evidence from PRD images. Return strict JSON only. "
                "Do not guess unreadable text. Classify non-Google-Sheet images as not_google_sheet_screenshot."
            ),
            prompt_mode=f"{PRD_REVIEW_PROMPT_VERSION}_google_sheet_screenshot_extraction",
            image_paths=[image_file.name],
        )
    payload = _parse_json_object(generated.get("result_markdown") or "")
    if not payload:
        raise ToolError("ocr_failed")
    return payload


def _image_suffix_from_url(url: str) -> str:
    path = unquote_plus(str(urlparse(str(url or "")).path or "")).casefold()
    for suffix in (".png", ".jpg", ".jpeg", ".webp", ".gif"):
        if path.endswith(suffix):
            return suffix
    return ".png"


def _parse_json_object(value: str) -> dict[str, Any] | None:
    text = str(value or "").strip()
    if not text:
        return None
    fenced = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, flags=re.DOTALL)
    if fenced:
        text = fenced.group(1).strip()
    elif not text.startswith("{"):
        start = text.find("{")
        end = text.rfind("}")
        if start >= 0 and end > start:
            text = text[start : end + 1]
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        return None
    return parsed if isinstance(parsed, dict) else None


def _resolve_one_linked_spreadsheet(
    *,
    confluence: ConfluenceConnector,
    artifact: dict[str, Any],
    google_credentials: dict[str, Any] | None,
    remaining_text_chars: int,
    google_sheet_cache_dir: Path | None = None,
    progress_callback: PRDProgressCallback | None = None,
    artifact_index: int = 0,
    artifact_total: int = 0,
) -> None:
    url = str(artifact.get("url") or "")
    lowered = f"{url} {artifact.get('filename') or ''} {artifact.get('title') or ''}".casefold()
    if ".xls" in lowered and not (".xlsx" in lowered or ".xlsm" in lowered):
        artifact.update({"status": "failed", "reason": "unsupported_xls_format"})
        return
    try:
        if _is_google_spreadsheet_url(url):
            cached = _load_cached_google_spreadsheet_artifact(
                url=url,
                google_credentials=google_credentials,
                cache_dir=google_sheet_cache_dir,
                max_chars=remaining_text_chars,
            )
            if cached is not None:
                artifact.update({"status": "ok", "reason": "", **cached, "cache_hit": True})
                return
            content, drive_metadata = _download_google_spreadsheet(url=url, google_credentials=google_credentials)
        elif _is_confluence_spreadsheet_url(url):
            response = confluence._request(  # noqa: SLF001 - local PRD connector owns authenticated Confluence fetches.
                url,
                accept="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel.sheet.macroEnabled.12,*/*;q=0.8",
            )
            content = bytes(getattr(response, "content", b"") or b"")
            drive_metadata = {}
        else:
            artifact.update({"status": "failed", "reason": "unsupported_link"})
            return
        if not content:
            artifact.update({"status": "failed", "reason": "empty_download"})
            return
        _emit_prd_progress(
            progress_callback,
            "analyzing_template_metadata",
            f"Analyzing template metadata {artifact_index}/{artifact_total}." if artifact_total else "Analyzing template metadata.",
            min(max(artifact_index, 1), max(artifact_total, 1)),
            max(artifact_total, 1),
        )
        content_hash = hashlib.sha256(content).hexdigest()[:16]
        artifact["content_hash"] = content_hash
        is_google = _is_google_spreadsheet_url(url)
        extract_max_chars = PRD_LINKED_SPREADSHEET_MAX_TEXT_CHARS if is_google else remaining_text_chars
        extracted_full = _extract_workbook_text(content, max_chars=extract_max_chars)
        extracted = dict(extracted_full)
        if is_google:
            extracted = _limit_extracted_workbook_text(extracted, max_chars=remaining_text_chars)
        artifact.update({"status": "ok", "reason": "", **extracted})
        if is_google:
            _store_cached_google_spreadsheet_artifact(
                url=url,
                drive_metadata=drive_metadata,
                cache_dir=google_sheet_cache_dir,
                extracted={**extracted_full, "content_hash": content_hash},
            )
    except ToolError as error:
        artifact.update({"status": "failed", "reason": str(error)})
    except Exception as error:  # noqa: BLE001 - artifact coverage should not block the PRD review.
        artifact.update({"status": "failed", "reason": f"download_or_parse_failed: {error}"})


def _is_google_spreadsheet_url(url: str) -> bool:
    parsed = urlparse(str(url or ""))
    host = str(parsed.netloc or "").casefold()
    return host in {"docs.google.com", "drive.google.com"} and (
        "/spreadsheets/" in parsed.path or parsed.path.startswith("/file/") or parsed.path.startswith("/open")
    )


def _is_confluence_spreadsheet_url(url: str) -> bool:
    parsed = urlparse(str(url or ""))
    path = unquote_plus(str(parsed.path or "")).casefold()
    return "/download/attachments/" in path and (path.endswith(".xlsx") or path.endswith(".xlsm") or ".xlsx" in path or ".xlsm" in path)


def _download_google_spreadsheet(*, url: str, google_credentials: dict[str, Any] | None) -> tuple[bytes, dict[str, Any]]:
    if not google_credentials:
        raise ToolError("missing_google_credentials")
    try:
        from google.oauth2.credentials import Credentials
        from googleapiclient.errors import HttpError

        from bpmis_jira_tool.gmail_dashboard import build_drive_api_service
    except ImportError as error:
        raise ToolError("google_drive_reader_unavailable") from error
    file_id = _google_drive_file_id_from_url(url)
    if not file_id:
        raise ToolError("google_file_id_not_found")
    service = build_drive_api_service(Credentials(**google_credentials))
    try:
        metadata = service.files().get(fileId=file_id, fields="id,name,mimeType,modifiedTime,md5Checksum,size").execute()
        mime_type = str(metadata.get("mimeType") or "")
        if mime_type == "application/vnd.google-apps.spreadsheet":
            content = (
                service.files()
                .export_media(
                    fileId=file_id,
                    mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                .execute()
            )
            return _coerce_download_bytes(content), _normalize_drive_metadata(metadata)
        if mime_type in {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel.sheet.macroEnabled.12",
        }:
            content = service.files().get_media(fileId=file_id).execute()
            return _coerce_download_bytes(content), _normalize_drive_metadata(metadata)
        raise ToolError(f"unsupported_google_drive_mime_type:{mime_type or 'unknown'}")
    except HttpError as error:
        status = int(getattr(getattr(error, "resp", None), "status", 0) or 0)
        if status in {401, 403, 404}:
            raise ToolError("permission_denied") from error
        raise ToolError(f"google_drive_download_failed:{status or 'unknown'}") from error


def _google_drive_file_id_from_url(url: str) -> str:
    parsed = urlparse(str(url or ""))
    path_parts = [part for part in parsed.path.split("/") if part]
    if "d" in path_parts:
        index = path_parts.index("d")
        if index + 1 < len(path_parts):
            return path_parts[index + 1]
    query = parse_qs(parsed.query)
    if query.get("id"):
        return str(query["id"][0] or "").strip()
    return ""


def _google_sheet_artifact_cache_dir(settings: Settings) -> Path | None:
    data_dir = getattr(settings, "team_portal_data_dir", None)
    if not data_dir:
        return None
    return Path(data_dir) / "prd_briefing" / "cache" / "google_sheet_artifacts"


def _google_sheet_screenshot_cache_dir(settings: Settings) -> Path | None:
    data_dir = getattr(settings, "team_portal_data_dir", None)
    if not data_dir:
        return None
    return Path(data_dir) / "prd_briefing" / "cache" / "google_sheet_screenshot_evidence"


def _normalize_drive_metadata(metadata: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": str(metadata.get("id") or ""),
        "name": str(metadata.get("name") or ""),
        "mimeType": str(metadata.get("mimeType") or ""),
        "modifiedTime": str(metadata.get("modifiedTime") or ""),
        "md5Checksum": str(metadata.get("md5Checksum") or ""),
        "size": str(metadata.get("size") or ""),
    }


def _google_sheet_artifact_cache_key(*, url: str, drive_metadata: dict[str, Any]) -> str:
    file_id = str(drive_metadata.get("id") or _google_drive_file_id_from_url(url) or "")
    modified_time = str(drive_metadata.get("modifiedTime") or "")
    mime_type = str(drive_metadata.get("mimeType") or "")
    digest = hashlib.sha256(
        json.dumps(
            {
                "version": PRD_GOOGLE_SHEET_ARTIFACT_CACHE_VERSION,
                "file_id": file_id,
                "modified_time": modified_time,
                "mime_type": mime_type,
            },
            ensure_ascii=False,
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()
    return digest


def _load_cached_google_spreadsheet_artifact(
    *,
    url: str,
    google_credentials: dict[str, Any] | None,
    cache_dir: Path | None,
    max_chars: int,
) -> dict[str, Any] | None:
    if not google_credentials or cache_dir is None:
        return None
    try:
        from google.oauth2.credentials import Credentials
        from googleapiclient.errors import HttpError

        from bpmis_jira_tool.gmail_dashboard import build_drive_api_service
    except ImportError as error:
        raise ToolError("google_drive_reader_unavailable") from error
    file_id = _google_drive_file_id_from_url(url)
    if not file_id:
        raise ToolError("google_file_id_not_found")
    service = build_drive_api_service(Credentials(**google_credentials))
    try:
        metadata = _normalize_drive_metadata(
            service.files().get(fileId=file_id, fields="id,name,mimeType,modifiedTime,md5Checksum,size").execute()
        )
    except HttpError as error:
        status = int(getattr(getattr(error, "resp", None), "status", 0) or 0)
        if status in {401, 403, 404}:
            raise ToolError("permission_denied") from error
        raise ToolError(f"google_drive_metadata_failed:{status or 'unknown'}") from error
    cache_path = cache_dir / f"{_google_sheet_artifact_cache_key(url=url, drive_metadata=metadata)}.json"
    if not cache_path.exists():
        return None
    try:
        payload = json.loads(cache_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if payload.get("version") != PRD_GOOGLE_SHEET_ARTIFACT_CACHE_VERSION:
        return None
    if payload.get("drive_metadata") != metadata:
        return None
    extracted = payload.get("extracted")
    if not isinstance(extracted, dict):
        return None
    return _limit_extracted_workbook_text({**extracted, "drive_metadata": metadata}, max_chars=max_chars)


def _store_cached_google_spreadsheet_artifact(
    *,
    url: str,
    drive_metadata: dict[str, Any],
    cache_dir: Path | None,
    extracted: dict[str, Any],
) -> None:
    if cache_dir is None:
        return
    metadata = _normalize_drive_metadata(drive_metadata)
    if not metadata.get("id") or not metadata.get("modifiedTime"):
        return
    payload = {
        "version": PRD_GOOGLE_SHEET_ARTIFACT_CACHE_VERSION,
        "url": url,
        "drive_metadata": metadata,
        "extracted": extracted,
    }
    try:
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache_path = cache_dir / f"{_google_sheet_artifact_cache_key(url=url, drive_metadata=metadata)}.json"
        cache_path.write_text(json.dumps(payload, ensure_ascii=False, sort_keys=True), encoding="utf-8")
    except OSError:
        return


def _limit_extracted_workbook_text(extracted: dict[str, Any], *, max_chars: int) -> dict[str, Any]:
    payload = dict(extracted)
    text = str(payload.get("text") or "")
    if max_chars <= 0:
        payload["text"] = "[Linked spreadsheet text omitted because the linked artifact text limit was reached.]"
        return payload
    if len(text) > max_chars:
        payload["text"] = text[:max_chars]
    return payload


def _coerce_download_bytes(content: Any) -> bytes:
    if isinstance(content, bytes):
        return content
    if isinstance(content, str):
        return content.encode("utf-8")
    return bytes(content)


def _extract_workbook_text(content: bytes, *, max_chars: int) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ToolError("openpyxl_unavailable") from error
    workbook = load_workbook(io.BytesIO(content), read_only=False, data_only=False)
    lines: list[str] = []
    sheets = workbook.worksheets[:PRD_LINKED_SPREADSHEET_MAX_SHEETS]
    template_metadata = _extract_workbook_template_metadata(workbook, sheets=sheets)
    metadata_hash = hashlib.sha256(json.dumps(template_metadata, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()[:16]
    if max_chars <= 0:
        return {
            "text": "[Linked spreadsheet text omitted because the linked artifact text limit was reached.]",
            "sheet_count": len(workbook.worksheets),
            "sheets_extracted": [sheet.title for sheet in sheets],
            "skipped_sheet_count": max(len(workbook.worksheets) - len(sheets), 0),
            "template_metadata": template_metadata,
            "metadata_hash": metadata_hash,
        }
    for worksheet in sheets:
        lines.append(f"[Sheet: {worksheet.title}]")
        for row in worksheet.iter_rows(
            max_row=PRD_LINKED_SPREADSHEET_MAX_ROWS_PER_SHEET,
            max_col=PRD_LINKED_SPREADSHEET_MAX_COLS,
            values_only=True,
        ):
            values = ["" if value is None else str(value).strip().replace("\r\n", "\n").replace("\r", "\n") for value in row]
            if any(values):
                lines.append("\t".join(values).rstrip())
            if sum(len(line) for line in lines) >= max_chars:
                lines.append("...[linked spreadsheet text truncated]")
                text = "\n".join(lines).strip()[:max_chars]
                return {
                    "text": text,
                    "sheet_count": len(workbook.worksheets),
                    "sheets_extracted": [sheet.title for sheet in sheets],
                    "skipped_sheet_count": max(len(workbook.worksheets) - len(sheets), 0),
                    "template_metadata": template_metadata,
                    "metadata_hash": metadata_hash,
                }
    text = "\n".join(lines).strip()
    if not text:
        raise ToolError("empty_workbook")
    return {
        "text": text[:max_chars],
        "sheet_count": len(workbook.worksheets),
        "sheets_extracted": [sheet.title for sheet in sheets],
        "skipped_sheet_count": max(len(workbook.worksheets) - len(sheets), 0),
        "template_metadata": template_metadata,
        "metadata_hash": metadata_hash,
    }


def _extract_workbook_template_metadata(workbook: Any, *, sheets: list[Any]) -> dict[str, Any]:
    sheet_payloads: list[dict[str, Any]] = []
    hidden_sheets = [sheet.title for sheet in workbook.worksheets if str(getattr(sheet, "sheet_state", "visible") or "visible") != "visible"]
    for worksheet in sheets:
        non_empty_rows: list[list[str]] = []
        formula_cells: list[str] = []
        for row in worksheet.iter_rows(max_row=PRD_LINKED_SPREADSHEET_MAX_ROWS_PER_SHEET, max_col=PRD_LINKED_SPREADSHEET_MAX_COLS):
            values: list[str] = []
            for cell in row:
                value = cell.value
                text = "" if value is None else str(value).strip()
                values.append(text)
                if text.startswith("=") and len(formula_cells) < 8:
                    formula_cells.append(cell.coordinate)
            if any(values):
                non_empty_rows.append(values)
        header = _likely_header_row(non_empty_rows)
        duplicate_headers = sorted({value for value in header if value and header.count(value) > 1})
        empty_header_count = sum(1 for value in header if not value)
        merged_ranges = [str(range_ref) for range_ref in getattr(worksheet, "merged_cells", []) .ranges] if getattr(worksheet, "merged_cells", None) else []
        hidden_rows = [index for index, dimension in worksheet.row_dimensions.items() if bool(getattr(dimension, "hidden", False))]
        hidden_cols = [key for key, dimension in worksheet.column_dimensions.items() if bool(getattr(dimension, "hidden", False))]
        sheet_payloads.append(
            {
                "name": worksheet.title,
                "state": str(getattr(worksheet, "sheet_state", "visible") or "visible"),
                "max_row": int(worksheet.max_row or 0),
                "max_column": int(worksheet.max_column or 0),
                "sampled_rows": len(non_empty_rows),
                "header_row": header[:PRD_LINKED_SPREADSHEET_MAX_COLS],
                "empty_header_count": empty_header_count,
                "duplicate_headers": duplicate_headers,
                "merged_ranges": merged_ranges[:20],
                "merged_range_count": len(merged_ranges),
                "formula_cells": formula_cells,
                "formula_cell_count": _count_formula_cells(worksheet),
                "hidden_rows_count": len(hidden_rows),
                "hidden_columns": hidden_cols[:20],
                "hidden_columns_count": len(hidden_cols),
                "multi_level_header_risk": _has_multi_level_header_risk(non_empty_rows, merged_ranges),
                "wide_template_risk": int(worksheet.max_column or 0) > PRD_LINKED_SPREADSHEET_MAX_COLS,
            }
        )
    return {
        "workbook_sheet_count": len(workbook.worksheets),
        "hidden_sheets": hidden_sheets,
        "sheets": sheet_payloads,
    }


def _likely_header_row(rows: list[list[str]]) -> list[str]:
    candidates = [row for row in rows[:5] if any(row)]
    if not candidates:
        return []
    return max(candidates, key=lambda row: sum(1 for value in row if value))


def _count_formula_cells(worksheet: Any) -> int:
    count = 0
    for row in worksheet.iter_rows(max_row=PRD_LINKED_SPREADSHEET_MAX_ROWS_PER_SHEET, max_col=PRD_LINKED_SPREADSHEET_MAX_COLS):
        for cell in row:
            if str(cell.value or "").strip().startswith("="):
                count += 1
    return count


def _has_multi_level_header_risk(rows: list[list[str]], merged_ranges: list[str]) -> bool:
    non_empty = [row for row in rows[:3] if sum(1 for value in row if value) >= 2]
    if len(non_empty) >= 2:
        return True
    return any(any(ch.isdigit() for ch in range_ref) for range_ref in merged_ranges)


def _merge_linked_spreadsheet_coverage(coverage: dict[str, Any], linked_spreadsheet_evidence: dict[str, Any]) -> dict[str, Any]:
    artifacts = list(linked_spreadsheet_evidence.get("artifacts") or [])
    reviewed = [item for item in artifacts if item.get("status") == "ok"]
    failed = [item for item in artifacts if item.get("status") != "ok"]
    merged = dict(coverage)
    merged.update(
        {
            "linked_artifacts_total": len(artifacts),
            "linked_artifacts_reviewed": len(reviewed),
            "linked_artifacts_failed": len(failed),
            "linked_artifacts": [_linked_artifact_public_payload(item) for item in artifacts],
            "report_templates_total": len(artifacts),
            "report_templates_reviewed": len(reviewed),
            "report_templates_failed": len(failed),
            "report_templates": [_linked_artifact_public_payload(item) for item in artifacts],
        }
    )
    return merged


def _merge_google_sheet_screenshot_coverage(
    coverage: dict[str, Any],
    google_sheet_screenshot_evidence: dict[str, Any],
) -> dict[str, Any]:
    artifacts = list(google_sheet_screenshot_evidence.get("artifacts") or [])
    candidates = [
        item
        for item in artifacts
        if item.get("status") != "skipped" or item.get("reason") != "not_google_sheet_screenshot"
    ]
    reviewed = [item for item in candidates if item.get("status") == "ok"]
    failed = [item for item in candidates if item.get("status") != "ok"]
    merged = dict(coverage)
    public_artifacts = [_google_sheet_screenshot_public_payload(item) for item in artifacts]
    merged.update(
        {
            "google_sheet_screenshot_evidence_enabled": bool(google_sheet_screenshot_evidence.get("enabled")),
            "google_sheet_screenshots_total": len(candidates),
            "google_sheet_screenshots_reviewed": len(reviewed),
            "google_sheet_screenshots_failed": len(failed),
            "google_sheet_screenshot_images": public_artifacts,
        }
    )
    return merged


def _google_sheet_screenshot_public_payload(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "image_id": str(item.get("image_id") or ""),
        "url": str(item.get("url") or ""),
        "source_section_index": item.get("source_section_index"),
        "source_section_title": str(item.get("source_section_title") or ""),
        "status": str(item.get("status") or ""),
        "reason": str(item.get("reason") or ""),
        "is_google_sheet_screenshot": bool(item.get("is_google_sheet_screenshot")),
        "char_count": len(str(item.get("evidence_text") or "")),
        "evidence_hash": str(item.get("evidence_hash") or ""),
        "cache_hit": bool(item.get("cache_hit")),
        "used_in_findings": bool(item.get("used_in_findings")),
    }


def _linked_artifact_public_payload(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "title": str(item.get("title") or ""),
        "url": str(item.get("url") or ""),
        "source_section_index": item.get("source_section_index"),
        "source_section_title": str(item.get("source_section_title") or ""),
        "status": str(item.get("status") or ""),
        "reason": str(item.get("reason") or ""),
        "sheet_count": int(item.get("sheet_count") or 0),
        "sheets_extracted": list(item.get("sheets_extracted") or []),
        "skipped_sheet_count": int(item.get("skipped_sheet_count") or 0),
        "char_count": len(str(item.get("text") or "")),
        "metadata_hash": str(item.get("metadata_hash") or ""),
        "cache_hit": bool(item.get("cache_hit")),
        "template_metadata": item.get("template_metadata") if isinstance(item.get("template_metadata"), dict) else {},
    }


def _build_linked_spreadsheet_prompt_section(linked_spreadsheet_evidence: dict[str, Any]) -> str:
    artifacts = list(linked_spreadsheet_evidence.get("artifacts") or [])
    if not artifacts:
        return (
            "# Linked Spreadsheet Evidence\n"
            "No report-template spreadsheet links were found in report-related PRD sections."
        )
    blocks = ["# Linked Spreadsheet Evidence", "Use this section specifically for report generation feasibility and PRD-to-template mapping assessment."]
    for index, artifact in enumerate(artifacts, start=1):
        title = str(artifact.get("title") or artifact.get("url") or f"Linked spreadsheet {index}")
        section = str(artifact.get("source_section_title") or f"Section {artifact.get('source_section_index') or '-'}")
        status = str(artifact.get("status") or "failed")
        if status == "ok":
            sheets = ", ".join(str(sheet) for sheet in (artifact.get("sheets_extracted") or []))
            skipped = int(artifact.get("skipped_sheet_count") or 0)
            skipped_line = f"\nSkipped sheets: {skipped}" if skipped else ""
            metadata = _format_template_metadata_for_prompt(artifact.get("template_metadata") if isinstance(artifact.get("template_metadata"), dict) else {})
            blocks.append(
                f"## Linked artifact reviewed: {title}\n"
                f"Source section: {section}\n"
                f"Sheets extracted: {sheets or '-'}{skipped_line}\n"
                f"Use this artifact as evidence only for the linked source section.\n"
                f"Report template metadata:\n{metadata}\n"
                f"{str(artifact.get('text') or '').strip()}"
            )
        else:
            blocks.append(
                f"## Linked artifact not reviewed: {title}\n"
                f"Source section: {section}\n"
                f"Reason: {artifact.get('reason') or 'unavailable'}\n"
                f"Do not infer this artifact's contents."
            )
    return "\n\n".join(blocks).strip()


def _build_google_sheet_screenshot_prompt_section(google_sheet_screenshot_evidence: dict[str, Any]) -> str:
    artifacts = list(google_sheet_screenshot_evidence.get("artifacts") or [])
    reviewed = [item for item in artifacts if item.get("status") == "ok" and item.get("is_google_sheet_screenshot")]
    failed = [
        item
        for item in artifacts
        if item.get("status") == "failed" or (item.get("status") == "skipped" and item.get("reason") != "not_google_sheet_screenshot")
    ]
    if not reviewed and not failed:
        return ""
    blocks = [
        "# Google Sheet Screenshot Evidence",
        "Use this evidence only for the source PRD section shown on each image. These images were classified as Google Sheet screenshots or unreadable candidate screenshots. Do not treat `[MEDIA_ID_x]` placeholders themselves as PRD defects.",
    ]
    for index, artifact in enumerate(reviewed, start=1):
        blocks.append(
            f"## Google Sheet screenshot reviewed: {artifact.get('image_id') or f'Image {index}'}\n"
            f"Source section: {artifact.get('source_section_title') or '-'}\n"
            f"Source URL: {artifact.get('url') or '-'}\n"
            f"Use this screenshot evidence for template/configuration feasibility, field mapping, QA acceptance, and PRD patch suggestions.\n"
            f"{str(artifact.get('evidence_text') or '').strip()}"
        )
    for artifact in failed:
        blocks.append(
            f"## Google Sheet screenshot not reviewed: {artifact.get('image_id') or 'Image'}\n"
            f"Source section: {artifact.get('source_section_title') or '-'}\n"
            f"Source URL: {artifact.get('url') or '-'}\n"
            f"Reason: {artifact.get('reason') or 'unavailable'}\n"
            "Do not infer this screenshot's contents."
        )
    return "\n\n".join(blocks).strip()


def _format_template_metadata_for_prompt(metadata: dict[str, Any]) -> str:
    if not metadata:
        return "- Metadata unavailable."
    lines = [
        f"- Workbook sheets: {metadata.get('workbook_sheet_count') or 0}",
        f"- Hidden sheets: {', '.join(str(item) for item in (metadata.get('hidden_sheets') or [])) or '-'}",
    ]
    for sheet in list(metadata.get("sheets") or [])[:PRD_LINKED_SPREADSHEET_MAX_SHEETS]:
        if not isinstance(sheet, dict):
            continue
        header = ", ".join(str(value) for value in (sheet.get("header_row") or []) if str(value or "").strip())
        lines.append(
            "- Sheet: {name}; rows={rows}; columns={cols}; header={header}; merged_ranges={merged}; "
            "formula_cells={formula}; hidden_rows={hidden_rows}; hidden_columns={hidden_cols}; "
            "empty_headers={empty_headers}; duplicate_headers={duplicates}; multi_level_header_risk={multi}; wide_template_risk={wide}".format(
                name=sheet.get("name") or "-",
                rows=sheet.get("max_row") or 0,
                cols=sheet.get("max_column") or 0,
                header=header or "-",
                merged=sheet.get("merged_range_count") or 0,
                formula=sheet.get("formula_cell_count") or 0,
                hidden_rows=sheet.get("hidden_rows_count") or 0,
                hidden_cols=sheet.get("hidden_columns_count") or 0,
                empty_headers=sheet.get("empty_header_count") or 0,
                duplicates=", ".join(str(value) for value in (sheet.get("duplicate_headers") or [])) or "-",
                multi=bool(sheet.get("multi_level_header_risk")),
                wide=bool(sheet.get("wide_template_risk")),
            )
        )
    return "\n".join(lines)


def _build_section_coverage(
    *,
    total_sections: int,
    selected_indexes: list[int],
    selected_sections: list[Any],
    selection_hash: str,
) -> dict[str, Any]:
    used_chars = 0
    assessed_indexes: list[int] = []
    omitted_titles: list[str] = []
    truncated = False
    for original_index, section in zip(selected_indexes, selected_sections):
        content = str(section.content or "").strip()
        if not content:
            continue
        title = section.section_path or section.title or f"Section {original_index}"
        block = f"## Section {original_index}: {title}\n{content}"
        remaining = PRD_REVIEW_MAX_SOURCE_CHARS - used_chars
        if remaining <= 0:
            truncated = True
            omitted_titles.append(title)
            continue
        assessed_indexes.append(original_index)
        if len(block) > remaining:
            truncated = True
        used_chars += min(len(block), max(remaining, 0))
    selected_titles = [
        section.section_path or section.title or f"Section {index}"
        for index, section in zip(selected_indexes, selected_sections)
    ]
    return {
        "status": "truncated" if truncated else "full",
        "sections_total": total_sections,
        "selected_sections_total": len(selected_indexes),
        "sections_assessed": len(assessed_indexes),
        "selected_section_indexes": selected_indexes,
        "assessed_section_indexes": assessed_indexes,
        "selected_section_titles": selected_titles,
        "selection_hash": selection_hash,
        "truncated": truncated,
        "omitted_sections": omitted_titles,
    }


def _normalize_selected_section_indexes(value: list[int] | None) -> list[int] | None:
    if value is None:
        return None
    if not isinstance(value, list):
        raise ToolError("selected_section_indexes must be a list of section indexes.")
    indexes: list[int] = []
    seen: set[int] = set()
    for raw_index in value:
        if isinstance(raw_index, bool):
            raise ToolError("selected_section_indexes must contain integer section indexes.")
        try:
            index = int(raw_index)
        except (TypeError, ValueError) as error:
            raise ToolError("selected_section_indexes must contain integer section indexes.") from error
        if index not in seen:
            indexes.append(index)
            seen.add(index)
    return indexes


def _section_selection_hash(selected_section_indexes: list[int]) -> str:
    payload = ",".join(str(index) for index in selected_section_indexes)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]


def _cache_key_for_section_selection(
    prefix: str,
    *,
    page: IngestedConfluencePage,
    selected_section_indexes: list[int] | None,
    linked_artifact_fingerprint: str = "",
    google_sheet_screenshot_fingerprint: str = "",
) -> str:
    cache_key = prefix
    if selected_section_indexes:
        section_fingerprint = "|".join(
            f"{index}:{page.sections[index - 1].section_path}:{hashlib.sha256(str(page.sections[index - 1].content or '').encode('utf-8')).hexdigest()[:12]}"
            for index in selected_section_indexes
        )
        selection_hash = hashlib.sha256(section_fingerprint.encode("utf-8")).hexdigest()[:16]
        cache_key = f"{cache_key}:sections:{selection_hash}"
    if linked_artifact_fingerprint:
        linked_hash = hashlib.sha256(linked_artifact_fingerprint.encode("utf-8")).hexdigest()[:16]
        cache_key = f"{cache_key}:linked:{linked_hash}"
    if google_sheet_screenshot_fingerprint:
        screenshot_hash = hashlib.sha256(google_sheet_screenshot_fingerprint.encode("utf-8")).hexdigest()[:16]
        cache_key = f"{cache_key}:sheetshots:{screenshot_hash}"
    return cache_key


def normalize_prd_review_language(language: str | None) -> str:
    normalized = str(language or "zh").strip().lower()
    return "en" if normalized in {"en", "english"} else "zh"


def prd_briefing_review_prompt_version(language: str | None) -> str:
    return f"{PRD_REVIEW_PROMPT_VERSION}_briefing_{normalize_prd_review_language(language)}"


def prd_summary_prompt_version(language: str | None) -> str:
    return f"{PRD_SUMMARY_PROMPT_VERSION}_{normalize_prd_review_language(language)}"


def _report_output_structure_en() -> str:
    return """
### Report Generation Feasibility
1. **Template / Sheet:** [template title / sheet name, or Cannot assess report template feasibility because linked artifact was not reviewed.]
   - **Generation feasibility:** [Ready / Risky / Not Ready]
   - **Format issue:** [Header, merged-cell, formula, hidden row/column, duplicate field, dynamic row, or layout issue.]
   - **Technical generation risk:** [Why system-generated XLSX may be unstable or ambiguous.]
   - **PRD mapping gap:** [Missing PRD field definition, source rule, row duplication rule, snapshot rule, or fallback.]
   - **Suggested PRD patch:** [Concrete PRD wording/table/rule the PM should add.]
   - **QA acceptance check:** [How QA can verify generated report against the template.]
   - **Evidence basis:** [Spreadsheet evidence reviewed / Linked artifact not reviewed]
   - **Used in findings:** [Yes/No]
   - **Related finding:** [P0/P1/P2 item id or "-"]
   - **Reason:** [What changed in the assessment because of this artifact.]

### Report Template Risks
- [Prioritized risks from reviewed templates, or coverage gap if no report template was readable.]

### PRD-to-Template Mapping Gaps
- [Fields, rows, statuses, dates, source systems, snapshot timing, or fallback rules that the PRD must define for template generation.]
"""


def _report_output_structure_zh() -> str:
    return """
### Report Generation Feasibility / 报表生成可行性
1. **Template / Sheet：** [template title / sheet name，或 Cannot assess report template feasibility because linked artifact was not reviewed.]
   - **Generation feasibility：** [Ready / Risky / Not Ready]
   - **Format issue：** [表头、合并单元格、公式、隐藏行列、重复字段、动态行或版式问题。]
   - **Technical generation risk：** [为什么系统生成 XLSX 会不稳定或有歧义。]
   - **PRD mapping gap：** [缺失的 PRD 字段定义、来源规则、行复制规则、快照规则或 fallback。]
   - **Suggested PRD patch：** [PM 应补进 PRD 的具体文案/表格/规则。]
   - **QA acceptance check：** [QA 如何按模板验收生成结果。]
   - **Evidence basis：** [Spreadsheet evidence reviewed / Linked artifact not reviewed]
   - **Used in findings：** [Yes/No]
   - **Related finding：** [P0/P1/P2 item id 或 "-"]
   - **Reason：** [这个模板证据如何改变了评审结论。]

### Report Template Risks / 报表模板风险
- [来自已读模板的优先级风险；如果没有可读模板，说明 coverage gap。]

### PRD-to-Template Mapping Gaps / PRD 到模板映射缺口
- [PRD 为支持模板生成必须定义的字段、行规则、状态、日期、数据源、快照时间或兜底规则。]
"""


def _google_sheet_screenshot_output_structure_en() -> str:
    return """
### Google Sheet Screenshot Template Assessment
1. **Template / Screenshot:** [image id or screenshot title]
   - **Generation / configuration feasibility:** [Ready / Risky / Not Ready]
   - **Format issue:** [Visible field/header/layout issue from the screenshot.]
   - **PRD mapping gap:** [Missing PRD rule or mapping needed to use this screenshot as delivery evidence.]
   - **Suggested PRD patch:** [Concrete wording/table/rule the PM should add.]
   - **QA acceptance check:** [How QA can validate the template/configuration against the screenshot.]
   - **Evidence basis:** [Google Sheet screenshot evidence reviewed / Google Sheet screenshot not reviewed]
"""


def _google_sheet_screenshot_output_structure_zh() -> str:
    return """
### Google Sheet Screenshot Template Assessment / Google Sheet 截图模板评估
1. **Template / Screenshot：** [image id 或截图标题]
   - **Generation / configuration feasibility：** [Ready / Risky / Not Ready]
   - **Format issue：** [截图中可见的字段/表头/版式问题。]
   - **PRD mapping gap：** [要把截图作为交付依据时，PRD 缺失的规则或映射。]
   - **Suggested PRD patch：** [PM 应补进 PRD 的具体文案/表格/规则。]
   - **QA acceptance check：** [QA 如何按截图验收模板/配置。]
   - **Evidence basis：** [Google Sheet screenshot evidence reviewed / Google Sheet screenshot not reviewed]
"""


def build_prd_review_prompt(
    *,
    jira_id: str,
    jira_link: str,
    prd_url: str,
    page: IngestedConfluencePage,
    language: str = "zh",
    linked_spreadsheet_evidence: dict[str, Any] | None = None,
    google_sheet_screenshot_evidence: dict[str, Any] | None = None,
) -> str:
    source = _build_prd_source(page)
    linked_source = _build_linked_spreadsheet_prompt_section(linked_spreadsheet_evidence or {})
    screenshot_source = _build_google_sheet_screenshot_prompt_section(google_sheet_screenshot_evidence or {})
    include_report_sections = _is_report_or_template_prd(page, linked_spreadsheet_evidence)
    include_screenshot_sections = bool(screenshot_source)
    report_required_check_en = (
        "6. **Report generation feasibility:** For report/template/register sections, do reviewed Google Sheet or Excel templates have clear headers, stable columns, avoid risky merged/multi-level headers or formula dependencies, and map cleanly back to PRD field definitions?"
        if include_report_sections
        else ""
    )
    report_required_check_zh = (
        "6. **Report 生成可行性：** 对 report/template/register 类 section，已读取的 Google Sheet 或 Excel 模板是否有清晰表头、稳定列结构，是否避免高风险合并单元格/多层表头/公式依赖，并能和 PRD 字段定义清楚映射？"
        if include_report_sections
        else ""
    )
    report_output_en = _report_output_structure_en() if include_report_sections else ""
    report_output_zh = _report_output_structure_zh() if include_report_sections else ""
    screenshot_boundary_en = (
        "- Treat Google Sheet screenshot evidence as part of the source PRD section when it is listed under `# Google Sheet Screenshot Evidence`. Use it for template/configuration feasibility, field mapping, QA acceptance, and PRD patch suggestions.\n"
        if include_screenshot_sections
        else ""
    )
    screenshot_boundary_zh = (
        "- `# Google Sheet Screenshot Evidence` 中列出的截图 evidence 属于对应 PRD Section 的评审依据；用于评估模板/配置可行性、字段映射、QA 验收和 PRD 补写建议。\n"
        if include_screenshot_sections
        else ""
    )
    screenshot_output_en = _google_sheet_screenshot_output_structure_en() if include_screenshot_sections else ""
    screenshot_output_zh = _google_sheet_screenshot_output_structure_zh() if include_screenshot_sections else ""
    report_coverage_en = "- **Report templates:** [Reviewed count and unreadable artifacts with reasons]" if include_report_sections else ""
    report_coverage_zh = "- **Report templates：** [已读数量和未读模板及原因]" if include_report_sections else ""
    screenshot_coverage_en = (
        "- **Google Sheet screenshots:** [Reviewed count and unreadable screenshot reasons]"
        if include_screenshot_sections
        else ""
    )
    screenshot_coverage_zh = (
        "- **Google Sheet screenshots：** [已读数量和未读截图原因]"
        if include_screenshot_sections
        else ""
    )
    if normalize_prd_review_language(language) == "en":
        return f"""# Role
You are a senior PRD delivery-readiness reviewer using the `prd-review` skill. Your job is to help the PM identify the few delivery blockers that matter most and turn them into concrete PRD patches.

# Review Boundary
- Review only delivery logic: business-flow closure, role actions, ownership, approvals, handoffs, status transitions, rule precedence, exception paths, reverse paths, operational fallback, and QA/dev acceptance clarity.
- Do not evaluate business value, ROI, KPI choice, API design, database design, code architecture, implementation approach, or engineering effort.
- Do not invent rules. If the selected PRD sections do not contain evidence for a claim, write `Source not found in selected sections`.
- Preserve scenario names, API names, field names, enum values, and other identifiers exactly as written in the PRD, preferably in code format. Do not create spelling/typo findings for identifiers unless you can quote the exact typo from the selected PRD text.
- Treat linked spreadsheet evidence as part of the selected PRD section when it is listed under `# Linked Spreadsheet Evidence`. If an expected linked spreadsheet is listed as not reviewed, call out the coverage gap instead of guessing its content.
{screenshot_boundary_en}- Never list `[MEDIA_ID_x]` placeholders themselves as PRD defects. If an image or screenshot cannot be reviewed, treat that only as an evidence coverage gap.
{"- When report-template spreadsheet evidence is reviewed, assess whether the report format is reasonable, technically generatable, and testable by QA. This is a template feasibility review, not backend architecture or effort estimation." if include_report_sections else ""}

# Task
Assess the selected PRD sections for delivery readiness. Prioritize only the issues that would materially block development, QA, UAT, release, operations, or support. Every important blocker must name the PRD section that should be changed and include text the PM can add back to the PRD.

# Scoring Rubric
- **9 - 10 (Ready):** Main flow, exception paths, reverse paths, rule precedence, role actions, and fallback are clear enough for dev and QA.
- **7 - 8 (Needs Minor Clarification):** Happy path is clear, with only minor non-blocking clarifications.
- **4 - 6 (Needs Major Clarification):** Development, QA, release, operations, or support will likely be blocked by missing logic or unclear rules.
- **1 - 3 (Not Ready):** Core flow is broken, contradictory, or impossible to validate from the PRD.

# Prioritization Rules
- Use `P0` only when dev/QA/UAT/release/operations cannot proceed or would likely implement inconsistent behavior.
- Use `P1` for important rule, exception, access, status, or evidence gaps that can cause rework or failed acceptance.
- Use `P2` only for useful clarifications. Put P2 items under `Secondary Clarifications`, not in the blocker list.
- `Top Must-Fix Delivery Blockers` must contain at most 5 items total. Merge duplicate symptoms into one blocker.
- Avoid duplication between `Top Must-Fix Delivery Blockers` and `Section Patch Suggestions`: blockers explain the issue and priority; patch suggestions provide copy-ready PRD changes without restating the full problem.
- Do not create more than 2 main findings for the same selected section unless there is a true P0 contradiction.
- Keep the whole answer concise, roughly 800-1200 words unless the selected section is extremely complex.

# Required Review Checks
1. **Flow closure:** Can each user/system role move from trigger to terminal state without an undefined next step?
2. **Rule rigor:** Are preconditions, status changes, priorities, validation rules, and overlapping rules unambiguous?
3. **Role and ownership clarity:** Is it clear who can act, approve, reject, withdraw, edit, retry, or override at each step?
4. **Exception and reverse paths:** Are timeout, interruption, retry, rejection, rollback, cancellation, blacklist, permission failure, and manual fallback handled?
5. **Acceptance clarity:** Can QA/dev verify the expected behavior after the PRD is updated?
{report_required_check_en}

# Output Format
Return only concise Markdown in English. Every blocker or gap must include a section name and evidence basis.
---
### Executive Verdict
- **Final Score:** [X] / 10
- **Dev-Readiness Verdict:** [Ready / Needs Clarification / Not Ready]
- **Bottom line:** [One concise reason grounded in the PRD.]
- **Top 3 attention points:** [At most 3 short bullets.]

### Top Must-Fix Delivery Blockers
1. **Priority:** [P0/P1]
   - **Section:** [section title, or Source not found in selected sections]
   - **Problem:** [What is missing, contradictory, or undefined.]
   - **Why this is important:** [Why this blocks dev, QA, UAT, release, ops, or support.]
   - **Suggested PRD patch:** [Specific rule/state/fallback/wording the PM can add.]
   - **Acceptance check:** [What QA/dev should be able to verify.]
   - **Evidence basis:** [Current PRD section / Spreadsheet evidence reviewed / Linked artifact not reviewed / Source not found in selected sections]

### Section Patch Suggestions
1. **Section: [section title]**
   - **Priority:** [P0/P1/P2]
   - **Suggested PRD patch:** [Copy-ready sentence/table/rule to add or rewrite.]
   - **Acceptance check:** [Expected observable validation.]
   - **Evidence basis:** [One of the allowed evidence labels.]

### Secondary Clarifications
- **Priority:** P2
  - **Section:** [section title]
  - **Clarification:** [Non-blocking question or cleanup.]
{screenshot_output_en}{report_output_en}

### Evidence Coverage
- **Current PRD sections reviewed:** [brief coverage]
{report_coverage_en}
{screenshot_coverage_en}

### PM Decision Checklist
- [Concrete yes/no or rule-choice question the PM should answer, ordered by priority.]
---

# Review Context
- Jira ID: {jira_id or "-"}
- Jira Link: {jira_link or "-"}
- PRD Title: {page.title}
- PRD Link: {prd_url}
- PRD Updated At: {page.updated_at or "-"}

# PRD Content
{source}

{linked_source}
{screenshot_source}
"""
    return f"""# Role
你是一位资深 PRD 交付就绪度评审专家，使用 `prd-review` Skill 的方法论工作。你的目标不是评价产品想法，而是帮 PM 找出真正最重要的交付阻塞点，并转成可直接写回 PRD 的修改建议。

# 评审边界
- 只评审交付逻辑：业务流程闭环、角色动作、权限/归属、审批/交接、状态流转、规则优先级、异常路径、逆向路径、运营兜底、QA/研发验收清晰度。
- 不评价商业价值、ROI、指标选择、API 设计、数据库设计、代码架构、实现方案或研发工期。
- 不臆测 PRD 没写的规则。如果选中的 PRD Section 中没有依据，必须写 `Source not found in selected sections`。
- 场景名、API 名、字段名、枚举值等 identifier 必须按 PRD 原文保留，优先用 code format；除非能引用选中 PRD 原文里的 exact typo，否则不要输出 identifier 拼写/typo 类 finding。
- `# Linked Spreadsheet Evidence` 中列出的 spreadsheet evidence 属于对应 PRD Section 的评审依据；如果某个关键 spreadsheet 标记为 not reviewed，必须把它当作覆盖缺口说明，不能臆测其内容。
{screenshot_boundary_zh}- 不能把 `[MEDIA_ID_x]` 占位符本身当作业务缺口；如果图片或截图未能读取，只能作为 evidence coverage gap 说明。
{"- 当读取到 report-template spreadsheet evidence 时，必须评估报表格式是否合理、系统是否能稳定生成、QA 是否能按模板验收。这是模板可生成性和 PRD 映射评估，不是后端架构或工期评估。" if include_report_sections else ""}

# Task
请评估选中的 PRD Section 是否具备交付就绪度。优先输出真正会阻塞研发、QA、UAT、上线、运营或客服承接的问题。每个关键 blocker 都必须指出应该修改哪个 PRD Section，并给出 PM 可直接写回 PRD 的补写文本。

# 评分标准
- **9 - 10 分 (Ready)：** 主流程、异常、逆向、规则优先级、角色动作、运营兜底都足够清晰，研发和 QA 可以直接接手。
- **7 - 8 分 (Needs Minor Clarification)：** 主流程清楚，仅有少量不阻断交付的澄清项。
- **4 - 6 分 (Needs Major Clarification)：** 存在会阻塞研发、QA、上线、运营或客服承接的流程/规则缺口。
- **1 - 3 分 (Not Ready)：** 核心流程断裂、规则自相矛盾，或无法从 PRD 判断用户/系统下一步。

# 优先级规则
- `P0` 只用于研发/QA/UAT/上线/运营无法继续，或极可能导致实现不一致的问题。
- `P1` 用于重要规则、异常、权限、状态或证据缺口，会导致返工或验收失败。
- `P2` 只用于有价值但不阻塞的问题，必须放到 `Secondary Clarifications / 次要澄清项`，不要混进 blocker。
- `Top Must-Fix Delivery Blockers` 最多 5 条。相同类型问题必须合并，不能平铺刷屏。
- 避免 `Top Must-Fix Delivery Blockers` 和 `Section Patch Suggestions` 重复：blocker 说明问题和优先级，patch suggestions 只给可写回 PRD 的补写文本，不重复完整问题描述。
- 同一个 selected section 最多输出 2 条主建议，除非存在真正 P0 矛盾。
- 输出要压缩，默认约 800-1200 words 的中文等价长度。

# 必查维度
1. **流程闭环：** 每个用户/系统角色是否能从触发条件走到终态，没有未定义的下一步？
2. **规则严密性：** 前置条件、状态变化、优先级、校验规则、多规则叠加是否清楚？
3. **角色与权限：** 谁能提交、审批、拒绝、撤回、编辑、重试、覆盖或人工介入是否明确？
4. **异常与逆向：** 超时、中断、重试、拒绝、回滚、取消、拉黑、权限失败、人工兜底是否有处理方式？
5. **验收清晰度：** PRD 补完后，QA/研发是否能验证行为正确？
{report_required_check_zh}

# Output Format
请严格按以下结构输出精炼 Markdown。每个 blocker 或 gap 都必须包含 Section 名称和证据依据。
---
### Executive Verdict
- **最终得分：** [X] / 10
- **Dev-Readiness Verdict：** [Ready / Needs Clarification / Not Ready]
- **Bottom line：** [一句话说明，必须基于 PRD 内容。]
- **Top 3 attention points：** [最多 3 个短 bullet。]

### Top Must-Fix Delivery Blockers
1. **优先级：** [P0/P1]
   - **Section：** [section title，或 Source not found in selected sections]
   - **Problem：** [缺什么、哪里矛盾、哪里未定义。]
   - **Why this is important：** [为什么会阻塞研发、QA、UAT、上线、运营或客服。]
   - **Suggested PRD patch：** [PM 可直接补进 PRD 的规则/状态/fallback/文案。]
   - **Acceptance check：** [QA/研发应该能验证什么。]
   - **Evidence basis：** [Current PRD section / Spreadsheet evidence reviewed / Linked artifact not reviewed / Source not found in selected sections]

### Section Patch Suggestions
1. **Section：[section title]**
   - **优先级：** [P0/P1/P2]
   - **建议补写：** [可直接复制的句子/表格/规则。]
   - **验收检查：** [可观察的验收结果。]
   - **证据依据：** [上述证据标签之一。]

### Secondary Clarifications / 次要澄清项
- **优先级：** P2
  - **Section：** [section title]
  - **Clarification：** [非阻塞澄清问题或文案清理。]
{screenshot_output_zh}{report_output_zh}

### Evidence Coverage / 证据覆盖
- **Current PRD sections reviewed：** [简要覆盖范围]
{report_coverage_zh}
{screenshot_coverage_zh}

### PM Decision Checklist
- [PM 需要回答的具体 yes/no 或规则选择问题，按优先级排序。]
---

# Review Context
- Jira ID: {jira_id or "-"}
- Jira Link: {jira_link or "-"}
- PRD Title: {page.title}
- PRD Link: {prd_url}
- PRD Updated At: {page.updated_at or "-"}

# PRD Content
{source}

{linked_source}
{screenshot_source}
"""


def build_prd_summary_prompt(
    *,
    jira_id: str,
    jira_link: str,
    prd_url: str,
    page: IngestedConfluencePage,
    language: str = "zh",
) -> str:
    source = _build_prd_source(page)
    if normalize_prd_review_language(language) == "en":
        return f"""# Role
You are a senior product manager who turns PRDs into concise summaries that business, product, and engineering readers can use immediately.

# Task
Read the PRD content and produce a concise but complete English summary. Focus on:
1. Background and objective.
2. User/business flow.
3. Main functional scope.
4. Key rules, data definitions, and status transitions.
5. Launch dependencies and open questions.

# Output Format
Return only Markdown:
---
### PRD Summary
[3-5 sentence overview]

### Scope
- ...

### Key Logic
- ...

### Dependencies / Open Questions
- ...
---

# Context
- Jira ID: {jira_id or "-"}
- Jira Link: {jira_link or "-"}
- PRD Title: {page.title}
- PRD Link: {prd_url}
- PRD Updated At: {page.updated_at or "-"}

# PRD Content
{source}
"""
    return f"""# Role
你是一位资深产品经理，擅长把 PRD 快速整理成业务、产品、研发都能直接阅读的摘要。

# Task
请阅读 PRD 内容，输出一份简洁但完整的中文摘要。重点覆盖：
1. 背景和目标。
2. 用户/业务流程。
3. 主要功能范围。
4. 关键规则、数据口径、状态流转。
5. 上线/依赖/待确认事项。

# Output Format
请严格按 Markdown 输出：
---
### PRD Summary
[3-5 句话概括]

### Scope
- ...

### Key Logic
- ...

### Dependencies / Open Questions
- ...
---

# Context
- Jira ID: {jira_id or "-"}
- Jira Link: {jira_link or "-"}
- PRD Title: {page.title}
- PRD Link: {prd_url}
- PRD Updated At: {page.updated_at or "-"}

# PRD Content
{source}
"""


def _batch_outputs_markdown(batch_outputs: list[dict[str, Any]]) -> str:
    blocks = []
    for item in batch_outputs:
        titles = ", ".join(str(title or "-") for title in item.get("section_titles") or [])
        blocks.append(
            f"## Batch {item.get('batch_index')}\n"
            f"Sections: {titles or '-'}\n\n"
            f"{str(item.get('result_markdown') or '').strip()}"
        )
    return "\n\n".join(blocks).strip()


def build_prd_summary_batch_prompt(
    *,
    jira_id: str,
    jira_link: str,
    prd_url: str,
    page: IngestedConfluencePage,
    language: str,
    batch_index: int,
    batch_total: int,
) -> str:
    source = _build_prd_source(page, max_chars=PRD_REVIEW_MAX_SOURCE_CHARS)
    if normalize_prd_review_language(language) == "en":
        return f"""# Task
Summarize this PRD section batch for a later whole-PRD synthesis. Preserve important scope, flow, rules, data definitions, status transitions, dependencies, and open questions. Do not omit details just because they appear late in the document.

# Context
- Jira ID: {jira_id or "-"}
- Jira Link: {jira_link or "-"}
- PRD Title: {page.title}
- PRD Link: {prd_url}
- Batch: {batch_index}/{batch_total}

# PRD Batch Content
{source}
"""
    return f"""# Task
请总结这个 PRD section batch，供后续整份 PRD synthesis 使用。保留重要范围、流程、规则、数据口径、状态流转、依赖和待确认事项。不要因为内容在文档后半段就省略。

# Context
- Jira ID: {jira_id or "-"}
- Jira Link: {jira_link or "-"}
- PRD Title: {page.title}
- PRD Link: {prd_url}
- Batch: {batch_index}/{batch_total}

# PRD Batch Content
{source}
"""


def build_prd_summary_synthesis_prompt(
    *,
    jira_id: str,
    jira_link: str,
    prd_url: str,
    page: IngestedConfluencePage,
    language: str,
    batch_outputs: list[dict[str, Any]],
) -> str:
    batches = _batch_outputs_markdown(batch_outputs)
    if normalize_prd_review_language(language) == "en":
        return f"""# Role
You are a senior product manager synthesizing a full PRD summary from complete section-batch notes.

# Task
Produce one concise but complete English summary of the whole PRD. Cover all batches, including late sections. Do not mention the batching process.

# Output Format
Return only Markdown:
---
### PRD Summary
[3-5 sentence overview]

### Scope
- ...

### Key Logic
- ...

### Dependencies / Open Questions
- ...
---

# Context
- Jira ID: {jira_id or "-"}
- Jira Link: {jira_link or "-"}
- PRD Title: {page.title}
- PRD Link: {prd_url}
- PRD Updated At: {page.updated_at or "-"}

# Batch Summaries
{batches}
"""
    return f"""# Role
你是一位资深产品经理，正在基于完整的 section batch notes 合成整份 PRD 摘要。

# Task
请输出一份简洁但完整的中文整份 PRD 摘要。必须覆盖所有 batch，包括文档后半段 section。不要提及 batch 处理过程。

# Output Format
请严格按 Markdown 输出：
---
### PRD Summary
[3-5 句话概括]

### Scope
- ...

### Key Logic
- ...

### Dependencies / Open Questions
- ...
---

# Context
- Jira ID: {jira_id or "-"}
- Jira Link: {jira_link or "-"}
- PRD Title: {page.title}
- PRD Link: {prd_url}
- PRD Updated At: {page.updated_at or "-"}

# Batch Summaries
{batches}
"""


def build_prd_review_batch_prompt(
    *,
    jira_id: str,
    jira_link: str,
    prd_url: str,
    page: IngestedConfluencePage,
    language: str,
    batch_index: int,
    batch_total: int,
) -> str:
    source = _build_prd_source(page, max_chars=PRD_REVIEW_MAX_SOURCE_CHARS)
    if normalize_prd_review_language(language) == "en":
        return f"""# Task
Review this PRD section batch for delivery-readiness evidence. Identify only material P0/P1/P2 delivery gaps from these sections, with section titles and evidence basis. Keep it concise; this will be synthesized with other batches. Preserve scenario/API/field identifiers exactly as written in the PRD and do not create spelling/typo findings for identifiers unless the exact typo is quoted from this batch. Do not treat `[MEDIA_ID_x]` placeholders themselves as PRD defects.

# Context
- Jira ID: {jira_id or "-"}
- Jira Link: {jira_link or "-"}
- PRD Title: {page.title}
- PRD Link: {prd_url}
- Batch: {batch_index}/{batch_total}

# PRD Batch Content
{source}
"""
    return f"""# Task
请评审这个 PRD section batch 的交付就绪度证据。只识别这些 section 内真正重要的 P0/P1/P2 交付缺口，必须带 section title 和 evidence basis。保持精炼，后续会和其他 batch 合成最终评审。场景名、API 名、字段名等 identifier 必须按 PRD 原文保留；除非能引用本 batch 原文里的 exact typo，否则不要输出 identifier 拼写/typo 类 finding。不能把 `[MEDIA_ID_x]` 占位符本身当作 PRD 缺口。

# Context
- Jira ID: {jira_id or "-"}
- Jira Link: {jira_link or "-"}
- PRD Title: {page.title}
- PRD Link: {prd_url}
- Batch: {batch_index}/{batch_total}

# PRD Batch Content
{source}
"""


def build_prd_review_synthesis_prompt(
    *,
    jira_id: str,
    jira_link: str,
    prd_url: str,
    page: IngestedConfluencePage,
    language: str,
    batch_outputs: list[dict[str, Any]],
    linked_spreadsheet_evidence: dict[str, Any] | None,
    google_sheet_screenshot_evidence: dict[str, Any] | None = None,
) -> str:
    batch_reviews = _batch_outputs_markdown(batch_outputs)
    linked_source = _build_linked_spreadsheet_prompt_section(linked_spreadsheet_evidence or {})
    screenshot_source = _build_google_sheet_screenshot_prompt_section(google_sheet_screenshot_evidence or {})
    include_report_sections = _is_report_or_template_prd(page, linked_spreadsheet_evidence)
    include_screenshot_sections = bool(screenshot_source)
    report_headings = (
        "### Report Generation Feasibility\n### Report Template Risks\n### PRD-to-Template Mapping Gaps"
        if include_report_sections
        else ""
    )
    screenshot_headings = "### Google Sheet Screenshot Template Assessment" if include_screenshot_sections else ""
    report_instruction = (
        "Report-template findings must include Template / Sheet, Generation feasibility, Format issue, Technical generation risk, PRD mapping gap, QA acceptance check, Used in findings, Related finding, and Reason. If a linked report template was not readable, write `Cannot assess report template feasibility because linked artifact was not reviewed.`"
        if include_report_sections
        else ""
    )
    screenshot_instruction = (
        "Google Sheet screenshot findings must include Template / Screenshot, Generation / configuration feasibility, Format issue, PRD mapping gap, Suggested PRD patch, QA acceptance check, and Evidence basis. Never treat `[MEDIA_ID_x]` placeholders themselves as PRD defects."
        if include_screenshot_sections
        else "Never treat `[MEDIA_ID_x]` placeholders themselves as PRD defects."
    )
    identifier_instruction_en = "Preserve scenario/API/field identifiers exactly as written in the PRD. Do not create spelling or typo findings for identifiers unless the exact typo is quoted from the reviewed PRD text."
    identifier_instruction_zh = "场景名、API 名、字段名、枚举值等 identifier 必须按 PRD 原文保留；除非能引用 reviewed PRD 原文里的 exact typo，否则不要输出 identifier 拼写/typo 类 finding。"
    if normalize_prd_review_language(language) == "en":
        return f"""# Role
You are a senior PRD delivery-readiness reviewer using the `prd-review` skill.

# Task
Synthesize the batch reviews into one final prioritized PRD assessment. Keep only the most important blockers. P0/P1 must be at most 5 items total. Merge duplicates across batches. Use linked spreadsheet evidence if provided. Do not mention the batching process.

# Required Output Structure
### Executive Verdict
### Top Must-Fix Delivery Blockers
### Section Patch Suggestions
### Secondary Clarifications
{screenshot_headings}
{report_headings}
### Evidence Coverage
### PM Decision Checklist

Every finding must include Priority, Section, Problem, Why this is important, Suggested PRD patch, Acceptance check, and Evidence basis. {report_instruction} {screenshot_instruction} {identifier_instruction_en} If evidence is not in the reviewed sections, write `Source not found in selected sections`.

# Context
- Jira ID: {jira_id or "-"}
- Jira Link: {jira_link or "-"}
- PRD Title: {page.title}
- PRD Link: {prd_url}
- PRD Updated At: {page.updated_at or "-"}

# Batch Reviews
{batch_reviews}

{linked_source}
{screenshot_source}
"""
    return f"""# Role
你是一位资深 PRD 交付就绪度评审专家，使用 `prd-review` Skill 的方法论工作。

# Task
请把 batch reviews 合成为一份最终的高优先级 PRD Assessment。只保留最重要的阻塞点，P0/P1 总数最多 5 条，跨 batch 重复问题必须合并。如有 linked spreadsheet evidence，必须纳入证据依据。不要提及 batch 处理过程。

# Required Output Structure
### Executive Verdict
### Top Must-Fix Delivery Blockers
### Section Patch Suggestions
### Secondary Clarifications
{screenshot_headings}
{report_headings}
### Evidence Coverage
### PM Decision Checklist

每条 finding 必须包含优先级、Section、Problem、Why this is important、Suggested PRD patch、Acceptance check、Evidence basis。{report_instruction} {screenshot_instruction} {identifier_instruction_zh} 如果 reviewed sections 没有依据，写 `Source not found in selected sections`。

# Context
- Jira ID: {jira_id or "-"}
- Jira Link: {jira_link or "-"}
- PRD Title: {page.title}
- PRD Link: {prd_url}
- PRD Updated At: {page.updated_at or "-"}

# Batch Reviews
{batch_reviews}

{linked_source}
{screenshot_source}
"""


def generate_prd_review_with_codex(
    *,
    prompt: str,
    settings: Settings,
    workspace_root: Path,
    language: str = "zh",
    prompt_version: str = PRD_REVIEW_PROMPT_VERSION,
) -> dict[str, Any]:
    return _generate_with_codex(
        prompt=prompt,
        settings=settings,
        workspace_root=workspace_root,
        system_text=build_prd_review_system_text(language),
        prompt_mode=prompt_version,
    )


def build_prd_review_system_text(language: str | None = "zh") -> str:
    output_language = "English" if normalize_prd_review_language(language) == "en" else "Chinese"
    return (
        "You are using the prd-review skill to assess PRD delivery readiness. "
        "Review only delivery logic: flow closure, role actions, ownership, status transitions, rule rigor, "
        "exception and reverse paths, operational fallback, and QA/dev acceptance clarity. "
        "Do not evaluate business value, metrics, architecture, APIs, databases, code, implementation effort, or ROI. "
        "Prioritize the few issues that materially block delivery; do not flatten all findings into equal severity. "
        "Every important blocker or gap must name the affected PRD section and include Priority, Problem, Why this is important, "
        "Suggested PRD patch, Acceptance check, and Evidence basis. If evidence is not in the selected sections, "
        "say `Source not found in selected sections`. "
        "Preserve scenario names, API names, field names, enum values, and other identifiers exactly as written in the PRD; "
        "do not create spelling or typo findings for identifiers unless the exact typo is quoted from the selected PRD text. "
        "Never treat `[MEDIA_ID_x]` placeholders themselves as PRD defects; only use extracted image evidence when it is provided. "
        "For report-generation PRDs, assess reviewed Google Sheet or Excel templates for report format reasonableness, "
        "generation feasibility, PRD-to-template mapping gaps, and QA acceptance checks. State whether spreadsheet evidence "
        "was used in findings and which finding it changed. "
        f"Return only the requested Markdown review in {output_language}. Do not include tool logs."
    )


def generate_prd_summary_with_codex(
    *,
    prompt: str,
    settings: Settings,
    workspace_root: Path,
    language: str = "zh",
    prompt_version: str = PRD_SUMMARY_PROMPT_VERSION,
) -> dict[str, Any]:
    output_language = "English" if normalize_prd_review_language(language) == "en" else "Chinese"
    return _generate_with_codex(
        prompt=prompt,
        settings=settings,
        workspace_root=workspace_root,
        system_text=(
            "You are a senior product manager. "
            f"Return only the requested Markdown PRD summary in {output_language}. Do not include tool logs."
        ),
        prompt_mode=prompt_version,
    )


def _generate_with_codex(
    *,
    prompt: str,
    settings: Settings,
    workspace_root: Path,
    system_text: str,
    prompt_mode: str,
    image_paths: list[str] | None = None,
) -> dict[str, Any]:
    provider = CodexCliBridgeSourceCodeQALLMProvider(
        workspace_root=workspace_root,
        timeout_seconds=settings.source_code_qa_codex_timeout_seconds,
        concurrency_limit=settings.source_code_qa_codex_concurrency,
        session_mode="ephemeral",
        codex_binary=os.getenv("SOURCE_CODE_QA_CODEX_BINARY") or None,
    )
    codex_model = resolve_codex_model(
        CODEX_ROUTE_DEEP,
        legacy_env_names=("PRD_REVIEWER_CODEX_MODEL", "SOURCE_CODE_QA_CODEX_MODEL"),
    )
    result = provider.generate(
        payload={
            "systemInstruction": {"parts": [{"text": system_text}]},
            "contents": [{"parts": [{"text": prompt}]}],
            "codex_prompt_mode": prompt_mode,
            "_codex_reasoning_effort": resolve_codex_reasoning_effort(CODEX_ROUTE_DEEP),
            **({"_codex_image_paths": list(image_paths)} if image_paths else {}),
        },
        primary_model=codex_model,
        fallback_model=codex_model,
    )
    return {
        "result_markdown": provider.extract_text(result.payload),
        "model_id": result.model,
        "trace": result.payload.get("codex_cli_trace") if isinstance(result.payload, dict) else {},
    }
