import json
import io
import os
import subprocess
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
import sqlite3
import tempfile
import threading
import time
import unittest
import zipfile
from unittest.mock import patch
from types import SimpleNamespace

from openpyxl import Workbook

from bpmis_jira_tool import web as portal_web
from bpmis_jira_tool import source_code_qa_llm_providers as llm_providers
from bpmis_jira_tool.errors import ToolError
from bpmis_jira_tool.source_code_qa import (
    CODE_INDEX_VERSION,
    CodexCliBridgeSourceCodeQALLMProvider,
    RepositoryEntry,
    SourceCodeQAService,
)
from bpmis_jira_tool.source_code_qa_stores import (
    SourceCodeQAAttachmentStore,
    SourceCodeQAGeneratedArtifactStore,
    SourceCodeQARuntimeEvidenceStore,
)
from bpmis_jira_tool.source_code_qa_types import LLMGenerateResult
from bpmis_jira_tool.user_config import TEAM_PROFILE_DEFAULTS
from bpmis_jira_tool.web import (
    JobStore,
    SourceCodeQAQueryScheduler,
    SourceCodeQASessionStore,
    _build_source_code_qa_effort_assessment_prompt,
    _build_source_code_qa_effort_business_plan,
    _build_source_code_qa_effort_evidence_matrix,
    _build_source_code_qa_effort_estimation_rubric,
    _build_source_code_qa_effort_structured_assessment,
    _build_source_code_qa_effort_technical_candidates,
    _build_source_code_qa_session_context,
    _load_source_code_qa_effort_dictionaries,
    _prepare_source_code_qa_auto_sync,
    _source_code_qa_effort_scope_guard,
    create_app,
)
from scripts.promote_source_code_qa_eval_candidates import promote_candidates
from scripts.run_source_code_qa_evals import _build_fixture_repositories, _evaluate_case, _guard_fixture_data_root, _summarize_results
from scripts.source_code_qa_scheduled_sync import run_scheduled_sync
from scripts.source_code_qa_auto_eval_candidates import build_auto_eval_candidates
from scripts.source_code_qa_feedback_to_eval import build_eval_candidates


# These tests build the Flask app via create_app(), which loads the host's .env.
# On the live host that .env enables local-agent passthrough (LOCAL_AGENT_MODE,
# LOCAL_AGENT_SOURCE_CODE_QA_ENABLED, a real LOCAL_AGENT_BASE_URL + secret). Without
# this guard, a config-save test would forward through the live local-agent server
# and clobber the real Source Code Q&A repository mapping, ignoring the per-test
# temp data dir. Force direct mode for the whole module so every test stays isolated
# to its TemporaryDirectory. Tests that need local-agent behavior re-enable it via
# their own inner patch.dict (and mock the client / use a fake host).
_LOCAL_AGENT_DISABLE_ENV = {
    "BPMIS_CALL_MODE": "direct",
    "LOCAL_AGENT_MODE": "disabled",
    "LOCAL_AGENT_SOURCE_CODE_QA_ENABLED": "false",
    "LOCAL_AGENT_BASE_URL": "",
    "LOCAL_AGENT_HMAC_SECRET": "",
}
_module_local_agent_env_patch = None


def setUpModule():
    global _module_local_agent_env_patch
    _module_local_agent_env_patch = patch.dict(os.environ, _LOCAL_AGENT_DISABLE_ENV, clear=False)
    _module_local_agent_env_patch.start()


def tearDownModule():
    if _module_local_agent_env_patch is not None:
        _module_local_agent_env_patch.stop()


class SourceCodeQARouteTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        from bpmis_jira_tool.source_code_qa import CODE_INDEX_VERSION

        with patch.dict(
            os.environ,
            {
                "FLASK_SECRET_KEY": "test-secret",
                "TEAM_PORTAL_DATA_DIR": self.temp_dir.name,
                "TEAM_ALLOWED_EMAIL_DOMAINS": "npt.sg",
                "SOURCE_CODE_QA_GITLAB_TOKEN": "secret-token",
            },
            clear=False,
        ):
            self.app = create_app()
            self.app.testing = True

    def tearDown(self):
        self.temp_dir.cleanup()

    @staticmethod
    def _login(client, email="xiaodong.zheng@npt.sg"):
        with client.session_transaction() as session:
            session["google_profile"] = {"email": email, "name": "Portal User"}
            session["google_credentials"] = {"token": "x", "scopes": []}

    def _seed_repo_download_scope(self, *, pm_team="AF", country="All", display_name="Portal Repo", url="https://git.example.com/team/portal.git"):
        service = self.app.config["SOURCE_CODE_QA_SERVICE"]
        service.save_mapping(pm_team=pm_team, country=country, repositories=[{"display_name": display_name, "url": url}])
        key = service.mapping_key(pm_team, country)
        entry = service._load_entries_for_key(key)[0]
        repo_path = service._repo_path(key, entry)
        repo_path.mkdir(parents=True, exist_ok=True)
        (repo_path / "README.md").write_text("# portal repo\n", encoding="utf-8")
        (repo_path / ".gitignore").write_text(".pytest_cache/\n", encoding="utf-8")
        subprocess.run(["git", "init"], cwd=repo_path, check=True, capture_output=True, text=True)
        subprocess.run(["git", "add", "."], cwd=repo_path, check=True, capture_output=True, text=True)
        subprocess.run(
            ["git", "-c", "user.name=Test User", "-c", "user.email=test@example.com", "commit", "-m", "init"],
            cwd=repo_path,
            check=True,
            capture_output=True,
            text=True,
        )
        return repo_path

    def test_npt_user_gets_public_repo_download_page_but_blocked_apis(self):
        with self.app.test_client() as client:
            self._login(client, "teammate@npt.sg")
            default_response = client.get("/", follow_redirects=False)
            self._login(client, "teammate@npt.sg")
            page_response = client.get("/source-code-qa", follow_redirects=False)
            api_response = client.get("/api/source-code-qa/config")
            self._login(client, "teammate@npt.sg")
            sessions_response = client.get("/api/source-code-qa/sessions")

        # Non-admin NPT users are blocked from the signed-in home but still get
        # the public Repo Download page; the config/session APIs are admin-only.
        self.assertEqual(default_response.status_code, 302)
        self.assertEqual(default_response.headers["Location"], "/access-denied")
        self.assertEqual(page_response.status_code, 200)
        self.assertEqual(api_response.status_code, 403)
        self.assertEqual(sessions_response.status_code, 403)
        page_html = page_response.get_data(as_text=True)
        self.assertIn("Source Code Repo Download", page_html)
        self.assertNotIn('data-source-view-tab="chat"', page_html)

    def test_admin_user_can_access_source_code_qa_directly(self):
        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            response = client.get("/source-code-qa", follow_redirects=True)

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("Source Code Q&amp;A", html)
        self.assertIn('href="/source-code-qa">Source Code</a>', html)
        self.assertIn(">Projects<", html)
        self.assertIn('data-source-question rows="2"', html)
        self.assertLess(html.index("data-source-attachments"), html.index("data-source-question"))
        self.assertLess(html.index("data-source-question"), html.index("data-source-attachment-upload"))

    def test_source_code_qa_view_panels_are_siblings_not_nested(self):
        # Regression: the chat <section> was missing a closing tag, nesting the
        # download/effort/admin panels inside it. They then inherited the chat
        # panel's display:none and never showed when their tab was selected.
        from html.parser import HTMLParser

        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            response = client.get("/source-code-qa", follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)

        class _PanelNesting(HTMLParser):
            def __init__(self):
                super().__init__()
                self._section_stack = []  # one bool per open <section>: is it a view panel?
                self.nested = []
                self.seen = []

            def handle_starttag(self, tag, attrs):
                if tag != "section":
                    return
                panel = dict(attrs).get("data-source-view-panel")
                if panel is not None:
                    self.seen.append(panel)
                    if any(self._section_stack):
                        self.nested.append(panel)
                self._section_stack.append(panel is not None)

            def handle_endtag(self, tag):
                if tag == "section" and self._section_stack:
                    self._section_stack.pop()

        parser = _PanelNesting()
        parser.feed(html)
        self.assertEqual(
            sorted(parser.seen), ["admin", "chat", "download", "effort"]
        )
        self.assertEqual(
            parser.nested, [], f"view panels must not be nested inside each other: {parser.nested}"
        )
        self.assertLess(html.index("data-source-attachment-upload"), html.index("data-source-query"))

    def test_non_npt_user_gets_public_download_page_but_blocked_api(self):
        with self.app.test_client() as client:
            with client.session_transaction() as session:
                session["google_profile"] = {"email": "teammate@example.com", "name": "External User"}
                session["google_credentials"] = {"token": "x", "scopes": []}
            page_response = client.get("/source-code-qa", follow_redirects=False)
            api_response = client.post("/api/source-code-qa/query", json={"pm_team": "AF", "country": "All", "question": "test"})

        # The page is public (Repo Download view only); the chat API stays blocked.
        self.assertEqual(page_response.status_code, 200)
        self.assertIn(b"Source Code Repo Download", page_response.data)
        self.assertNotIn(b'data-source-view-tab="chat"', page_response.data)
        self.assertEqual(api_response.status_code, 403)

    def test_owner_sees_admin_controls_but_teammate_does_not(self):
        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            owner_response = client.get("/source-code-qa")
            self._login(client, "teammate@npt.sg")
            teammate_response = client.get("/source-code-qa", follow_redirects=False)

        self.assertIn(b"Repository Mapping", owner_response.data)
        self.assertIn(b"Repo Admin", owner_response.data)
        self.assertIn(b"Effort Assessment", owner_response.data)
        self.assertIn(b"Repo Download", owner_response.data)
        self.assertIn(b">Source Code<", owner_response.data)
        self.assertIn(b">PRDs<", owner_response.data)
        self.assertLess(owner_response.data.index(b"Chat"), owner_response.data.index(b"Effort Assessment"))
        self.assertLess(owner_response.data.index(b"Effort Assessment"), owner_response.data.index(b"Repo Admin"))
        self.assertLess(owner_response.data.index(b"Chat"), owner_response.data.index(b"Repo Download"))
        self.assertIn(b'data-tab-trigger="admin"', owner_response.data)
        self.assertIn(b'data-tab-trigger="effort"', owner_response.data)
        self.assertIn(b'data-source-view-tab="admin"', owner_response.data)
        self.assertIn(b'data-source-view-tab="effort"', owner_response.data)
        self.assertIn(b'data-tab-panel="admin"', owner_response.data)
        self.assertIn(b'data-tab-panel="effort"', owner_response.data)
        self.assertIn(b'data-source-view-panel="admin"', owner_response.data)
        self.assertIn(b'data-source-view-panel="effort"', owner_response.data)
        self.assertIn(b"data-source-effort-requirement", owner_response.data)
        self.assertIn(b"data-source-effort-run", owner_response.data)
        self.assertIn(b"Save Config", owner_response.data)
        self.assertIn(b"Sync / Refresh", owner_response.data)
        self.assertEqual(teammate_response.status_code, 200)
        self.assertIn(b"Repo Download", teammate_response.data)
        self.assertNotIn(b"Chat", teammate_response.data)
        self.assertNotIn(b"Repository Mapping", teammate_response.data)
        self.assertNotIn(b"Repo Admin", teammate_response.data)
        self.assertNotIn(b"Effort Assessment", teammate_response.data)
        self.assertNotIn(b"Save Config", teammate_response.data)
        self.assertNotIn(b"Sync / Refresh", teammate_response.data)
        self.assertNotIn(b'data-source-view-tab="chat"', teammate_response.data)
        self.assertIn(b'data-source-view-tab="download"', teammate_response.data)
        self.assertNotIn(b"Model Availability", owner_response.data)

    def test_repo_download_endpoint_is_public_after_password_unlock(self):
        self._seed_repo_download_scope()
        with self.app.test_client() as client, patch.dict(
            os.environ, {"BUSINESS_INSIGHTS_DOWNLOAD_PASSWORD": "test-bi-pass"}, clear=False
        ):
            self._login(client, "teammate@npt.sg")
            locked = client.get("/api/source-code-qa/repo-downloads/AF:All")
            unlock = client.post("/api/business-insights/download-unlock", json={"password": "test-bi-pass"})
            download = client.get("/api/source-code-qa/repo-downloads/AF:All")

        # The endpoint is public but password-gated: 401 until the shared
        # download password unlocks the session.
        self.assertEqual(locked.status_code, 401)
        self.assertEqual(unlock.status_code, 200)
        self.assertEqual(download.status_code, 200)
        self.assertEqual(download.headers["Content-Type"], "application/zip")
        with zipfile.ZipFile(io.BytesIO(download.data)) as archive:
            names = archive.namelist()
            self.assertIn("manifest.json", names)
            self.assertIn("Portal-Repo/README.md", names)
            self.assertNotIn("Portal-Repo/.git/config", names)

    def test_sync_refresh_uses_sync_job_status_endpoint(self):
        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            response = client.get("/source-code-qa")

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('data-sync-jobs-url="/api/source-code-qa/sync-jobs/__JOB_ID__"', html)
        script = Path("static/source_code_qa.js").read_text(encoding="utf-8")
        self.assertIn("const syncJobsUrlTemplate", script)
        self.assertIn("readJobStatus(jobId, syncJobsUrlTemplate)", script)

    def test_sync_job_status_endpoint_accepts_sync_jobs_only(self):
        with self.app.app_context():
            sync_job = self.app.config["JOB_STORE"].create("source-code-qa-sync", title="Sync Source Code Repositories")
            query_job = self.app.config["JOB_STORE"].create("source-code-qa-query", title="Source Code Q&A Query")

        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            sync_response = client.get(f"/api/source-code-qa/sync-jobs/{sync_job.job_id}")
            query_response = client.get(f"/api/source-code-qa/sync-jobs/{query_job.job_id}")

        self.assertEqual(sync_response.status_code, 200)
        self.assertEqual(sync_response.get_json()["action"], "source-code-qa-sync")
        self.assertEqual(query_response.status_code, 404)
        self.assertEqual(query_response.get_json()["error_category"], "job_not_found")

    def test_frontend_pasted_images_use_attachment_upload_flow(self):
        script = Path("static/source_code_qa.js").read_text(encoding="utf-8")

        self.assertIn("const handleAttachmentPaste", script)
        self.assertIn("imageFilesFromClipboard(event.clipboardData)", script)
        self.assertIn("await addAttachmentFiles(pastedImages)", script)
        self.assertIn("addEventListener('paste', handleAttachmentPaste)", script)
        self.assertIn("image.${extension}", script)
        self.assertIn("Uploading...", script)
        self.assertIn("data-source-preview-attachment", script)
        self.assertIn("Please wait for image upload to finish.", script)

    def test_frontend_effort_assessment_clears_stale_result_on_new_run(self):
        script = Path("static/source_code_qa.js").read_text(encoding="utf-8")

        self.assertIn("lastEffortPayload = null;", script)
        self.assertIn("effortMeta.textContent = 'running';", script)
        self.assertIn("result: null", script)
        self.assertIn("effortMeta.textContent = ['failed'", script)

    def test_frontend_css_is_loaded_from_page_specific_asset(self):
        template = Path("templates/source_code_qa.html").read_text(encoding="utf-8")
        global_stylesheet = Path("static/style.css").read_text(encoding="utf-8")
        page_stylesheet = Path("static/source_code_qa.css").read_text(encoding="utf-8")
        api_script = Path("static/source_code_qa_api.js").read_text(encoding="utf-8")

        self.assertIn("source_code_qa.css", template)
        self.assertIn("source_code_qa_api.js", template)
        self.assertIn(".source-qa-chat-shell", page_stylesheet)
        self.assertIn(".source-qa-session-messages", page_stylesheet)
        self.assertIn(".source-qa-status", page_stylesheet)
        self.assertIn("window.SourceCodeQAApi", api_script)
        self.assertNotIn(".source-qa-chat-shell", global_stylesheet)
        self.assertNotIn(".source-qa-session-messages", global_stylesheet)
        self.assertNotIn(".source-qa-status", global_stylesheet)

    def test_frontend_caches_effort_assessment_draft_and_result(self):
        script = Path("static/source_code_qa.js").read_text(encoding="utf-8")
        template = Path("templates/source_code_qa.html").read_text(encoding="utf-8")

        self.assertIn("source-code-qa:effort-assessment:last:v1", script)
        self.assertIn("persistEffortAssessmentDraft", script)
        self.assertIn("restoreEffortAssessmentCache", script)
        self.assertIn("persistEffortAssessmentCache({ result: cachedEffortPayload(payload)", script)
        self.assertIn("effortRequirement?.addEventListener('input', persistEffortAssessmentDraft)", script)
        self.assertIn("renderEffortAssessment(cached.result, { persist: false })", script)
        self.assertIn("renderEffortHybridSummary", script)
        self.assertIn("renderEffortStructuredAnswer", script)
        self.assertIn("effortRenderChangePoints", script)
        self.assertIn("effortParsedAnswer", script)
        self.assertIn("Assessment Summary", script)
        self.assertIn("Code Change Points", script)
        self.assertIn("assessment.business_plan", script)
        self.assertIn("assessment.technical_candidates", script)
        self.assertIn("structured_assessment", script)
        self.assertNotIn("Evidence Used", script)
        self.assertNotIn("renderEffortEvidence", script)
        self.assertNotIn("data-source-effort-evidence", template)
        self.assertNotIn("effort_evidence_query", script)
        self.assertNotIn("effort_evidence_result", script)
        self.assertNotIn("Hybrid Assessment Context", script)
        self.assertIn("data-source-effort-copy", template)
        self.assertIn("data-effort-latest-url", template)
        self.assertIn("Business Requirement", template)
        self.assertIn("Code Impact &amp; Man-day Assessment", template)
        self.assertNotIn("Code Impact &amp; Person-days", template)
        self.assertNotIn("业务需求", template)
        self.assertIn("loadLatestEffortAssessment", script)
        self.assertIn("loadLatestEffortAssessment();", script)
        self.assertIn("copyEffortPmDevSummary", script)
        self.assertIn("effortDisplayedAssessmentText", script)
        self.assertIn("navigator.clipboard.writeText(text)", script)
        self.assertNotIn("data-source-effort-summary", template)
        self.assertNotIn("Effort Assessment PM/Dev Summary", script)

    def test_frontend_uses_deep_mode_and_raw_codex_answer(self):
        template = Path("templates/source_code_qa.html").read_text(encoding="utf-8")
        script = Path("static/source_code_qa.js").read_text(encoding="utf-8")

        self.assertNotIn("Answer Speed", template)
        self.assertNotIn("data-source-query-mode", template)
        self.assertNotIn("Fast Mode", script)
        self.assertIn("query_mode: 'deep'", script)
        self.assertIn("llm_budget_mode: 'auto'", script)
        self.assertIn("source-qa-raw-codex-answer", script)
        self.assertIn("scrollLatestAssistantAnswerToStart", script)
        self.assertIn("scrollIntoView({ behavior: 'smooth', block: 'start'", script)
        self.assertIn("const renderSessionMessages = (session, options = {}) =>", script)
        self.assertIn("if (options.autoScroll) {\n      scrollLatestAssistantAnswerToStart();\n    }", script)
        self.assertIn("applyActiveSession(payload.session || null, { autoScroll: Boolean(options.autoScroll) });", script)
        self.assertIn("applyActiveSession(payload.session, { autoScroll: true });", script)
        self.assertIn("await loadSession(activeSessionId, { preserveLive: true, autoScroll: true });", script)
        self.assertIn("const formatSingaporeTimestamp = (value) =>", script)
        self.assertIn("const formatSessionTime = (value) => {\n    return formatSingaporeTimestamp(value);\n  };", script)
        self.assertIn("timeZone: 'Asia/Singapore'", script)
        self.assertIn("${parts.year}-${parts.month}-${parts.day} ${parts.hour}:${parts.minute}:${parts.second} SGT", script)
        self.assertIn("formatSingaporeTimestamp(item.created_at || '')", script)
        self.assertNotIn("(item.created_at || '').replace('T', ' ').replace('Z', '')", script)
        self.assertIn("data-source-runtime-delete-team", script)
        self.assertIn("data-source-runtime-delete-country", script)
        self.assertIn("const deleteRuntimeEvidence = async (evidenceId, itemScope = {}) =>", script)
        self.assertIn("pm_team: itemScope.pm_team || scope.pm_team", script)
        self.assertIn("await loadRuntimeEvidence();", script)
        self.assertIn("data-source-message-role", script)
        self.assertIn("data-source-message-live", script)
        self.assertNotIn("Continue with Deep Mode", script)

    def test_codex_answer_path_is_split_from_main_service(self):
        service_source = Path("bpmis_jira_tool/source_code_qa.py").read_text(encoding="utf-8")
        component_source = Path("bpmis_jira_tool/source_code_qa_components.py").read_text(encoding="utf-8")
        codex_answer_source = Path("bpmis_jira_tool/source_code_qa_codex_answer.py").read_text(encoding="utf-8")

        self.assertNotIn("def _build_codex_llm_answer_impl", service_source)
        self.assertIn("from bpmis_jira_tool.source_code_qa_codex_answer import build_codex_llm_answer", component_source)
        self.assertIn("def build_codex_llm_answer(", codex_answer_source)
        self.assertIn("service._codex_initial_answer_result(", codex_answer_source)
        self.assertIn("service._codex_repair_answer_context(", codex_answer_source)

    def test_answer_generation_prelude_is_split_from_main_service(self):
        service_source = Path("bpmis_jira_tool/source_code_qa.py").read_text(encoding="utf-8")
        component_source = Path("bpmis_jira_tool/source_code_qa_components.py").read_text(encoding="utf-8")
        answer_generation_source = Path("bpmis_jira_tool/source_code_qa_answer_generation.py").read_text(encoding="utf-8")

        self.assertNotIn("def _build_llm_answer_impl", service_source)
        self.assertIn("from bpmis_jira_tool.source_code_qa_answer_generation import build_llm_answer", component_source)
        self.assertIn("def build_llm_answer(", answer_generation_source)
        self.assertIn("service._resolve_llm_budget(", answer_generation_source)
        self.assertIn("service._llm_answer_evidence_context(", answer_generation_source)
        self.assertIn("COMPACT_DEEP_BUDGET_MODE", answer_generation_source)
        self.assertIn("service._build_codex_llm_answer(", answer_generation_source)

    def test_removed_fast_query_mode_normalizes_to_deep(self):
        self.assertEqual(SourceCodeQAService.normalize_query_mode("fast"), "deep")
        self.assertEqual(SourceCodeQAService.normalize_query_mode("deep"), "deep")
        self.assertEqual(SourceCodeQAService.normalize_query_mode(None), "deep")

    def test_source_code_qa_admin_allowlist_does_not_grant_portal_admin(self):
        with patch.dict(
            os.environ,
            {
                "FLASK_SECRET_KEY": "test-secret",
                "TEAM_PORTAL_DATA_DIR": self.temp_dir.name,
                "TEAM_ALLOWED_EMAIL_DOMAINS": "npt.sg",
                "SOURCE_CODE_QA_ADMIN_EMAILS": "other-owner@npt.sg,xiaodong.zheng1991@gmail.com",
            },
            clear=False,
        ):
            app = create_app()
            app.testing = True

        with app.test_client() as client:
            self._login(client, "xiaodong.zheng1991@gmail.com")
            page_response = client.get("/source-code-qa")
            config_response = client.get("/api/source-code-qa/config")

        # The allowlisted gmail user is just another blocked non-admin now: the
        # public page renders without admin controls and the config API is 403.
        self.assertEqual(page_response.status_code, 200)
        self.assertNotIn(b"Repository Mapping", page_response.data)
        self.assertNotIn(b"Repo Admin", page_response.data)
        self.assertEqual(config_response.status_code, 403)
        self.assertEqual(config_response.get_json()["status"], "error")

    def test_source_code_qa_builtin_owner_can_manage_when_env_owner_changes(self):
        with patch.dict(
            os.environ,
            {
                "FLASK_SECRET_KEY": "test-secret",
                "TEAM_PORTAL_DATA_DIR": self.temp_dir.name,
                "TEAM_ALLOWED_EMAIL_DOMAINS": "npt.sg",
                "SOURCE_CODE_QA_OWNER_EMAIL": "temporary-owner@npt.sg",
                "SOURCE_CODE_QA_ADMIN_EMAILS": "",
            },
            clear=False,
        ):
            app = create_app()
            app.testing = True

        with app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            page_response = client.get("/source-code-qa")
            config_response = client.get("/api/source-code-qa/config")

        config_payload = config_response.get_json()
        self.assertIn(b"Repository Mapping", page_response.data)
        self.assertTrue(config_payload["can_manage"])
        self.assertEqual(config_payload["auth"]["admin_match_source"], "portal_admin")

    def test_source_code_qa_test_user_is_not_admin_even_with_default_allowlist(self):
        with patch.dict(
            os.environ,
            {
                "FLASK_SECRET_KEY": "test-secret",
                "TEAM_PORTAL_DATA_DIR": self.temp_dir.name,
                "TEAM_ALLOWED_EMAIL_DOMAINS": "npt.sg",
                "SOURCE_CODE_QA_GITLAB_TOKEN": "secret-token",
            },
            clear=True,
        ):
            app = create_app()
            app.testing = True

        sync_result = {
            "status": "ok",
            "key": "AF:All",
            "results": [{"state": "ok", "display_name": "Repo One"}],
            "repo_status": [{"display_name": "Repo One", "state": "synced", "message": "ready"}],
        }
        with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.sync", return_value=sync_result):
            with app.test_client() as client:
                self._login(client, "xiaodong.zheng1991@gmail.com")
                page_response = client.get("/source-code-qa")
                config_response = client.get("/api/source-code-qa/config")
                self._login(client, "xiaodong.zheng1991@gmail.com")
                save_response = client.post(
                    "/api/source-code-qa/config",
                    json={"pm_team": "AF", "country": "All", "repositories": [{"url": "https://git.example.com/team/repo.git"}]},
                )
                self._login(client, "xiaodong.zheng1991@gmail.com")
                sync_response = client.post("/api/source-code-qa/sync", json={"pm_team": "AF", "country": "All"})

        # The gmail test user is blocked like any non-admin: public page only,
        # all config/sync APIs rejected.
        self.assertNotIn(b"Repository Mapping", page_response.data)
        self.assertEqual(config_response.status_code, 403)
        self.assertEqual(config_response.get_json()["status"], "error")
        self.assertEqual(save_response.status_code, 403)
        self.assertEqual(sync_response.status_code, 403)

    def test_source_code_qa_manage_blocked_for_non_admin_with_error_payload(self):
        with self.app.test_client() as client:
            self._login(client, "teammate@npt.sg")
            response = client.post(
                "/api/source-code-qa/config",
                json={"pm_team": "AF", "country": "All", "repositories": [{"url": "https://git.example.com/team/repo.git"}]},
            )

        # Non-admins are blocked before the manage gate: generic blocked payload.
        payload = response.get_json()
        self.assertEqual(response.status_code, 403)
        self.assertEqual(payload["status"], "error")
        self.assertIn("not authorized", payload["message"])

    def test_config_save_is_owner_only_and_validates_https(self):
        with self.app.test_client() as client:
            self._login(client, "teammate@npt.sg")
            forbidden = client.post(
                "/api/source-code-qa/config",
                json={"pm_team": "AF", "country": "All", "repositories": [{"url": "https://git.example.com/team/repo.git"}]},
            )
            self._login(client, "xiaodong.zheng@npt.sg")
            invalid = client.post(
                "/api/source-code-qa/config",
                json={"pm_team": "AF", "country": "All", "repositories": [{"url": "git@git.example.com:team/repo.git"}]},
            )
            placeholder = client.post(
                "/api/source-code-qa/config",
                json={
                    "pm_team": "AF",
                    "country": "All",
                    "repositories": [{"display_name": "Repo One", "url": "https://git.example.com/team/repo.git"}],
                },
            )
            valid = client.post(
                "/api/source-code-qa/config",
                json={
                    "pm_team": "AF",
                    "country": "SG",
                    "repositories": [{"display_name": "Repo One", "url": "https://gitlab.npt.seabank.io/team/repo.git"}],
                },
            )

        self.assertEqual(forbidden.status_code, 403)
        self.assertEqual(invalid.status_code, 400)
        self.assertEqual(placeholder.status_code, 400)
        self.assertIn("placeholder host", placeholder.get_json()["message"])
        self.assertEqual(valid.status_code, 200)
        payload = valid.get_json()
        self.assertEqual(payload["key"], "AF:All")
        self.assertEqual(payload["repositories"][0]["display_name"], "Repo One")

    def test_effort_assessment_is_admin_only_and_validates_input(self):
        with self.app.test_client() as client:
            self._login(client, "teammate@npt.sg")
            forbidden = client.post(
                "/api/source-code-qa/effort-assessment",
                json={"pm_team": "AF", "country": "All", "requirement": "new flow", "llm_provider": "codex_cli_bridge"},
            )
            forbidden_status = client.get("/api/source-code-qa/effort-assessment-jobs/missing")
            forbidden_events = client.get("/api/source-code-qa/effort-assessment-jobs/missing/events")
            self._login(client, "xiaodong.zheng@npt.sg")
            empty = client.post(
                "/api/source-code-qa/effort-assessment",
                json={"pm_team": "AF", "country": "All", "requirement": "   ", "llm_provider": "codex_cli_bridge"},
            )

        self.assertEqual(forbidden.status_code, 403)
        self.assertEqual(forbidden_status.status_code, 403)
        self.assertEqual(forbidden_events.status_code, 403)
        self.assertEqual(empty.status_code, 400)
        self.assertEqual(empty.get_json()["message"], "Business requirement is empty.")

    def test_source_code_qa_route_access_gate_and_json_error_boundaries(self):
        with self.app.test_client() as client:
            # The page itself is public now; every Q&A API stays blocked for a
            # non-admin (each blocked request clears the session, so re-login).
            self._login(client, "external@example.com")
            public_page = client.get("/source-code-qa", follow_redirects=False)
            blocked_calls = [
                ("get", "/api/source-code-qa/config", None),
                ("get", "/api/source-code-qa/sync-jobs/missing", None),
                ("get", "/api/source-code-qa/sessions", None),
                ("get", "/api/source-code-qa/sessions/missing", None),
                ("post", "/api/source-code-qa/sessions/missing/archive", None),
                ("post", "/api/source-code-qa/attachments", None),
                ("get", "/api/source-code-qa/attachments/missing", None),
                ("get", "/api/source-code-qa/generated-artifacts/missing", None),
                ("post", "/api/source-code-qa/query", {"question": "x"}),
                ("post", "/api/source-code-qa/feedback", {"rating": "ok"}),
                ("get", "/api/source-code-qa/query-jobs/missing", None),
                ("get", "/api/source-code-qa/query-jobs/missing/events", None),
            ]
            blocked = []
            for method, path, payload in blocked_calls:
                self._login(client, "external@example.com")
                caller = getattr(client, method)
                blocked.append(caller(path, json=payload) if payload is not None else caller(path))
            self._login(client, "xiaodong.zheng@npt.sg")
            with patch("bpmis_jira_tool.web._build_source_code_qa_service", side_effect=RuntimeError("config exploded")):
                config_error = client.get("/api/source-code-qa/config")
            save_empty_payload = client.post("/api/source-code-qa/config", data="not-json", content_type="text/plain")
            invalid_limit = client.get("/api/source-code-qa/sessions?limit=not-a-number")

        self.assertEqual(public_page.status_code, 200)
        self.assertTrue(all(response.status_code == 403 for response in blocked))
        self.assertEqual(config_error.status_code, 500)
        self.assertEqual(config_error.get_json()["error_category"], "source_code_qa_internal")
        self.assertIn(save_empty_payload.status_code, {200, 400})
        self.assertEqual(invalid_limit.status_code, 200)

    def test_source_code_qa_route_attachment_artifact_runtime_and_feedback_boundaries(self):
        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            created = client.post(
                "/api/source-code-qa/sessions",
                json={"pm_team": "AF", "country": "All", "llm_provider": "codex_cli_bridge"},
            )
            session_id = created.get_json()["session"]["id"]

            missing_file = client.post(
                "/api/source-code-qa/attachments",
                data={"session_id": session_id},
                content_type="multipart/form-data",
            )
            missing_attachment = client.get(f"/api/source-code-qa/attachments/0bad?session_id={session_id}")
            missing_artifact_session = client.get("/api/source-code-qa/generated-artifacts/0bad")
            missing_artifact = client.get(f"/api/source-code-qa/generated-artifacts/0bad?session_id={session_id}")

            runtime_store = self.app.config["SOURCE_CODE_QA_RUNTIME_EVIDENCE_STORE"]
            with patch.object(runtime_store, "list", side_effect=ToolError("bad runtime scope")):
                runtime_list_error = client.get("/api/source-code-qa/runtime-evidence?pm_team=AF&country=All")
            runtime_missing_file = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={"pm_team": "AF", "country": "All"},
                content_type="multipart/form-data",
            )
            saved_runtime = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={
                    "pm_team": "AF",
                    "country": "All",
                    "source_type": "other",
                    "file": (io.BytesIO(b"runtime note"), "runtime.txt"),
                },
                content_type="multipart/form-data",
            )
            evidence_id = saved_runtime.get_json()["evidence"]["id"]
            runtime_deleted = client.delete(f"/api/source-code-qa/runtime-evidence/{evidence_id}?pm_team=AF&country=All")
            runtime_delete_error = client.delete("/api/source-code-qa/runtime-evidence/not-a-valid-id?pm_team=AF&country=All")

            with patch("bpmis_jira_tool.web._build_source_code_qa_service") as build_service:
                build_service.return_value.save_feedback.side_effect = ToolError("feedback rejected")
                feedback_error = client.post("/api/source-code-qa/feedback", json={"rating": "bad"})

        self.assertEqual(missing_file.status_code, 400)
        self.assertEqual(missing_attachment.status_code, 404)
        self.assertEqual(missing_artifact_session.status_code, 400)
        self.assertEqual(missing_artifact.status_code, 404)
        self.assertEqual(runtime_list_error.status_code, 400)
        self.assertEqual(runtime_missing_file.status_code, 400)
        self.assertEqual(saved_runtime.status_code, 200)
        self.assertEqual(runtime_deleted.status_code, 200)
        self.assertTrue(runtime_deleted.get_json()["deleted"])
        self.assertEqual(runtime_delete_error.status_code, 400)
        self.assertEqual(feedback_error.status_code, 400)

    def test_source_code_qa_route_async_queue_and_event_stream_boundaries(self):
        class CapturingScheduler:
            def __init__(self):
                self.submissions = []

            def submit(self, **kwargs):
                self.submissions.append(kwargs)

        scheduler = CapturingScheduler()
        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            with patch("bpmis_jira_tool.web._source_code_qa_provider_available", return_value=False):
                async_unavailable = client.post(
                    "/api/source-code-qa/query",
                    json={"pm_team": "AF", "country": "All", "question": "x", "llm_provider": "missing", "async": True},
                )
            created = client.post(
                "/api/source-code-qa/sessions",
                json={"pm_team": "AF", "country": "All", "llm_provider": "codex_cli_bridge"},
            )
            session_id = created.get_json()["session"]["id"]
            original_scheduler = self.app.config["SOURCE_CODE_QA_QUERY_SCHEDULER"]
            self.app.config["SOURCE_CODE_QA_QUERY_SCHEDULER"] = scheduler
            try:
                with patch("bpmis_jira_tool.web._source_code_qa_provider_available", return_value=True):
                    queued = client.post(
                        "/api/source-code-qa/query",
                        json={
                            "pm_team": "AF",
                            "country": "All",
                            "question": "x",
                            "llm_provider": "codex_cli_bridge",
                            "session_id": session_id,
                            "attachment_ids": ["bad-attachment-id"],
                            "async": True,
                        },
                    )
            finally:
                self.app.config["SOURCE_CODE_QA_QUERY_SCHEDULER"] = original_scheduler

            job_store: JobStore = self.app.config["JOB_STORE"]
            failed_query_job = job_store.create("source-code-qa-query", title="Source Code Q&A Query", owner_email="teammate@npt.sg")
            job_store.fail(failed_query_job.job_id, "query failed", error_category="test", error_code="failed", error_retryable=False)
            failed_query_events = client.get(f"/api/source-code-qa/query-jobs/{failed_query_job.job_id}/events")
            failed_query_events_data = failed_query_events.get_data()
            failed_query_events.close()

            running_query_job = job_store.create("source-code-qa-query", title="Source Code Q&A Query", owner_email="teammate@npt.sg")
            job_store.update(running_query_job.job_id, state="running", stage="retrieval", message="running", current=1, total=2)
            with patch("time.sleep", return_value=None):
                keepalive_response = client.get(f"/api/source-code-qa/query-jobs/{running_query_job.job_id}/events", buffered=False)
                keepalive_iter = iter(keepalive_response.response)
                first_event = next(keepalive_iter)
                keepalive_event = next(keepalive_iter)
                keepalive_response.close()

            transient_snapshot = {
                "job_id": "transient",
                "action": "source-code-qa-query",
                "state": "running",
                "owner_email": "teammate@npt.sg",
            }
            with patch("bpmis_jira_tool.web._source_code_qa_job_snapshot_for_current_user", side_effect=[transient_snapshot, None]):
                missing_after_open = client.get("/api/source-code-qa/query-jobs/transient/events")
                missing_after_open_data = missing_after_open.get_data()
                missing_after_open.close()

            self._login(client, "xiaodong.zheng@npt.sg")
            missing_effort_status = client.get("/api/source-code-qa/effort-assessment-jobs/missing")
            missing_effort_events = client.get("/api/source-code-qa/effort-assessment-jobs/missing/events")
            missing_effort_events_data = missing_effort_events.get_data()
            missing_effort_events.close()
            failed_effort_job = job_store.create("source-code-qa-effort-assessment", title="Source Code Q&A Effort Assessment")
            job_store.fail(failed_effort_job.job_id, "effort failed", error_category="test", error_code="failed", error_retryable=False)
            failed_effort_events = client.get(f"/api/source-code-qa/effort-assessment-jobs/{failed_effort_job.job_id}/events")
            failed_effort_events_data = failed_effort_events.get_data()
            failed_effort_events.close()

            running_effort_job = job_store.create("source-code-qa-effort-assessment", title="Source Code Q&A Effort Assessment")
            job_store.update(running_effort_job.job_id, state="running", stage="retrieval", message="running", current=1, total=2)
            with patch("time.sleep", return_value=None):
                effort_keepalive_response = client.get(f"/api/source-code-qa/effort-assessment-jobs/{running_effort_job.job_id}/events", buffered=False)
                effort_keepalive_iter = iter(effort_keepalive_response.response)
                effort_first_event = next(effort_keepalive_iter)
                effort_keepalive_event = next(effort_keepalive_iter)
                effort_keepalive_response.close()

        self.assertEqual(async_unavailable.status_code, 400)
        self.assertEqual(queued.status_code, 200)
        self.assertEqual(queued.get_json()["status"], "queued")
        self.assertEqual(len(scheduler.submissions), 1)
        self.assertEqual(failed_query_events.status_code, 200)
        self.assertIn(b"event: failed", failed_query_events_data)
        self.assertIn(b"event: message", first_event)
        self.assertEqual(keepalive_event, b": keepalive\n\n")
        self.assertIn(b"event: failed", missing_after_open_data)
        self.assertEqual(missing_effort_status.status_code, 404)
        self.assertIn(b"event: failed", missing_effort_events_data)
        self.assertIn(b"event: failed", failed_effort_events_data)
        self.assertIn(b"event: message", effort_first_event)
        self.assertEqual(effort_keepalive_event, b": keepalive\n\n")

    def test_source_code_qa_split_route_access_gate_early_returns_are_bound(self):
        from flask import Flask
        from bpmis_jira_tool import web_source_code_qa_routes as route_module
        from bpmis_jira_tool.web_source_code_qa_routes import register_source_code_qa_routes

        app = Flask("source-code-qa-access-gate-smoke")
        app.secret_key = "secret"
        blocked_response = ("blocked", 499)
        override_globals = {
            "_require_source_code_qa_access": lambda *_args, **_kwargs: blocked_response,
            "_require_source_code_qa_manage_access": lambda *_args, **_kwargs: blocked_response,
            "_require_source_code_qa_chat_access": lambda *_args, **_kwargs: blocked_response,
        }
        missing = object()
        original_globals = {key: route_module.__dict__.get(key, missing) for key in override_globals}
        try:
            register_source_code_qa_routes(app, object(), override_globals)
            direct_calls = [
                # NOTE: the /source-code-qa page itself is public now and does
                # not consult the access gate, so it is not probed here.
                ("source_code_qa_config_api", "/api/source-code-qa/config"),
                ("source_code_qa_sync_job_api", "/api/source-code-qa/sync-jobs/missing", "missing"),
                ("source_code_qa_sessions_api", "/api/source-code-qa/sessions"),
                ("source_code_qa_session_api", "/api/source-code-qa/sessions/missing", "missing"),
                ("source_code_qa_session_archive_api", "/api/source-code-qa/sessions/missing/archive", "missing"),
                ("source_code_qa_attachments_api", "/api/source-code-qa/attachments"),
                ("source_code_qa_attachment_api", "/api/source-code-qa/attachments/missing", "missing"),
                ("source_code_qa_generated_artifact_api", "/api/source-code-qa/generated-artifacts/missing", "missing"),
                ("source_code_qa_runtime_evidence_delete_api", "/api/source-code-qa/runtime-evidence/missing", "missing"),
                ("source_code_qa_query_api", "/api/source-code-qa/query"),
                ("source_code_qa_feedback_api", "/api/source-code-qa/feedback"),
                ("source_code_qa_query_job_api", "/api/source-code-qa/query-jobs/missing", "missing"),
                ("source_code_qa_query_job_events_api", "/api/source-code-qa/query-jobs/missing/events", "missing"),
            ]

            for item in direct_calls:
                endpoint, path, *args = item
                with app.test_request_context(path):
                    self.assertEqual(app.view_functions[endpoint](*args), blocked_response)
        finally:
            for key, value in original_globals.items():
                if value is missing:
                    route_module.__dict__.pop(key, None)
                else:
                    route_module.__dict__[key] = value

    def test_source_code_qa_query_api_accepts_pre_resolved_attachments(self):
        captured = {}

        def fake_query(**kwargs):
            captured.update(kwargs)
            return {"status": "ok", "summary": "done", "matches": []}

        with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today", return_value={"attempted": False, "status": "fresh"}), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
            side_effect=fake_query,
        ):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/query",
                    json={
                        "pm_team": "AF",
                        "country": "All",
                        "question": "x",
                        "llm_provider": "codex_cli_bridge",
                        "_resolved_attachments": [{"id": "pre-resolved", "filename": "context.txt"}],
                    },
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(captured["attachments"], [{"id": "pre-resolved", "filename": "context.txt"}])

    def test_effort_assessment_builds_crms_business_plan_and_candidates(self):
        requirement = (
            "高收入人群(annual income >120K)的现金分期额度被信用卡现在给的额度低限制住。"
            "方案1: 做信用卡现金分期专项额度 + 信用卡日常消费上限，报送还是报15w信用卡额度，通过103、104子产品额度控制。"
            "方案2: 如果信用卡10w，现金贷能拿40w，按需申请一个独立cashline。"
        )

        dictionaries = _load_source_code_qa_effort_dictionaries()
        crms_entries = dictionaries["domains"]["CRMS"]["entries"]
        self.assertGreaterEqual(len(crms_entries), 10)
        self.assertTrue(all(entry.get("business_aliases") for entry in crms_entries))
        self.assertTrue(all(entry.get("technical_terms") for entry in crms_entries))
        self.assertTrue(all(entry.get("surfaces") for entry in crms_entries))
        self.assertTrue(all(entry.get("complexity_hint") for entry in crms_entries))
        business_plan = _build_source_code_qa_effort_business_plan(
            pm_team="CRMS",
            country="SG",
            language="zh",
            requirement=requirement,
        )
        candidates = _build_source_code_qa_effort_technical_candidates(
            pm_team="CRMS",
            country="SG",
            business_plan=business_plan,
            requirement=requirement,
        )
        rubric = _build_source_code_qa_effort_estimation_rubric(
            business_plan=business_plan,
            technical_candidates=candidates,
        )

        self.assertIn("high income customers", business_plan["user_segments"])
        self.assertIn("cash installment", business_plan["products"])
        self.assertIn("cashline", business_plan["products"])
        self.assertEqual(len(business_plan["options"]), 2)
        self.assertIn("cash installment dedicated limit", business_plan["limit_types"])
        self.assertIn("sub-product limit 103/104", business_plan["limit_types"])
        search_terms = set(candidates["search_terms"])
        self.assertIn("annualIncome", search_terms)
        self.assertIn("cashline", search_terms)
        self.assertIn("borrowerLimit", search_terms)
        self.assertIn("productLimit", search_terms)
        self.assertIn("subProductLimitInfos", search_terms)
        self.assertIn("subProductCode=103", candidates["configs_or_tables"])
        self.assertIn("subProductCode=104", candidates["configs_or_tables"])
        self.assertIn("cash_installment_limit", candidates["matched_dictionary_entries"])
        self.assertIn("backend_service", candidates["typed_candidates"])
        self.assertIn("frontend_surface", candidates["typed_candidates"])
        self.assertIn("downstream_reporting", candidates["typed_candidates"])
        self.assertEqual(rubric["complexity_drivers"]["backend"], "high")
        self.assertEqual(len(rubric["option_estimates"]), 2)
        matrix = _build_source_code_qa_effort_evidence_matrix(
            evidence_result={
                "matches": [
                    {
                        "repo": "CRMS",
                        "path": "src/services/LimitStrategyService.java",
                        "line_start": 12,
                        "line_end": 18,
                        "reason": "cash installment limit strategy service",
                        "snippet": "class LimitStrategyService { void calculateAnnualIncomeLimit() {} }",
                    },
                    {
                        "repo": "CRMS",
                        "path": "src/config/limit_mapper.xml",
                        "line_start": 21,
                        "line_end": 28,
                        "reason": "subProductCode=103 config table mapping",
                        "snippet": "subProductCode=103 subProductCode=104",
                    },
                ],
            },
            business_plan=business_plan,
            technical_candidates=candidates,
        )
        self.assertEqual(matrix["version"], 1)
        self.assertIn("quality", matrix)
        self.assertGreaterEqual(matrix["quality"]["confirmed_group_count"], 1)
        self.assertTrue(any(group["key"] == "config_table" and group["status"] == "confirmed" for group in matrix["groups"]))
        structured = _build_source_code_qa_effort_structured_assessment(
            result={"matches": [], "effort_evidence_matrix": matrix},
            language="zh",
            business_plan=business_plan,
            technical_candidates=candidates,
            estimation_rubric=rubric,
            missing_evidence=["missing code evidence"],
            confidence="low",
        )
        self.assertIn("business_understanding", structured)
        self.assertTrue(structured["be_estimate"])
        self.assertTrue(structured["fe_estimate"])
        self.assertTrue(structured["inferred_impact"])
        self.assertTrue(structured["code_change_points"])
        self.assertIn("change", structured["code_change_points"][0])
        self.assertIn("evidence_status", structured["code_change_points"][0])
        self.assertIn("evidence_matrix_quality", structured)

    def test_effort_assessment_scope_guard_blocks_wrong_repo_selection(self):
        requirement = (
            "CRMS SG BTI suspended case review: income review, Manual Review, APC approval, "
            "CBS purchase, Suspension Appeal and Monthly Review need final decision workflow."
        )

        guard = _source_code_qa_effort_scope_guard(pm_team="GRC", country="SG", requirement=requirement)

        self.assertEqual(guard["status"], "mismatch")
        self.assertEqual(guard["selected_pm_team"], "GRC")
        self.assertEqual(guard["suggested_pm_team"], "CRMS")
        self.assertEqual(guard["suggested_country"], "SG")
        self.assertGreater(guard["scores"]["CRMS"], guard["scores"]["GRC"])
        self.assertIn("crms", [item.lower() for item in guard["matched_terms"]["CRMS"]])

    def test_effort_assessment_single_requirement_does_not_invent_options(self):
        requirement = (
            "Scope Retrieve and validate SME Keyman information from CIF using the SME CIF number from the DWH file. "
            "Pass Keyman indicator and guarantor-level CBS account status into the DWH-driven review code category."
        )
        business_plan = _build_source_code_qa_effort_business_plan(
            pm_team="CRMS",
            country="SG",
            language="en",
            requirement=requirement,
        )
        candidates = _build_source_code_qa_effort_technical_candidates(
            pm_team="CRMS",
            country="SG",
            business_plan=business_plan,
            requirement=requirement,
        )
        rubric = _build_source_code_qa_effort_estimation_rubric(
            business_plan=business_plan,
            technical_candidates=candidates,
        )
        prompt = _build_source_code_qa_effort_assessment_prompt(
            pm_team="CRMS",
            country="SG",
            language="en",
            requirement=requirement,
            llm_provider="codex_cli_bridge",
            runtime_evidence=[],
            business_plan=business_plan,
            technical_candidates=candidates,
            estimation_rubric=rubric,
        )

        self.assertFalse(business_plan["has_explicit_options"])
        self.assertEqual(len(business_plan["options"]), 1)
        self.assertEqual(business_plan["options"][0]["label"], "single proposed change")
        self.assertEqual(rubric["option_estimates"][0]["label"], "single proposed change")
        self.assertIn("do not invent option labels", prompt)
        self.assertIn("2. 代码改动点 / Code Change Points", prompt)
        self.assertNotIn("2. 方案 1/2 代码改动点", prompt)
        self.assertIn("Do not include visible sections named Evidence", prompt)
        self.assertNotIn("5. Confirmed / Inferred / Missing Evidence", prompt)

    def test_effort_assessment_builds_optimized_prompt_and_passes_runtime_evidence(self):
        captured_calls = []

        def fake_query(**kwargs):
            captured_calls.append(kwargs)
            kwargs["progress_callback"]("direct_search", "Searching direct matches in Repo One.", 1, 1)
            if kwargs.get("answer_mode") == "retrieval_only":
                return {
                    "status": "ok",
                    "answer_mode": "retrieval_only",
                    "summary": "evidence done",
                    "trace_id": "trace-effort-evidence",
                    "matches": [{"repo": "Repo One", "path": "src/api.py", "line_start": 10, "line_end": 12}],
                    "citations": [{"id": "S1", "repo": "Repo One", "path": "src/api.py", "line_start": 10, "line_end": 12}],
                    "index_freshness": {"status": "fresh", "git_revisions": [{"repo": "Repo One", "git_revision": "abc123"}]},
                    "answer_quality": {"status": "sufficient", "confidence": "medium", "missing": []},
                }
            return {
                "status": "ok",
                "answer_mode": "auto",
                "summary": "effort done",
                "llm_answer": "业务理解\n- ok\n\n代码改动点\n- 调整审批接口和状态流转，让高风险订单走新的审批路径 [S1]，不要暴露 src/api.py:10。\n\nBE 人天\n- 1-2 PD\n\nFE 人天\n- 0 PD\n\nConfirmed / Inferred / Missing Evidence\n- Confirmed S1\n\nQA / Integration Impact\n- 覆盖审批通过、拒绝和回退场景。",
                "llm_provider": "codex_cli_bridge",
                "llm_model": "codex-cli",
                "llm_route": {"task": "effort_assessment", "codex_repair_attempted": False},
                "trace_id": "trace-effort",
                "matches": [{"repo": "Repo One", "path": "src/api.py", "line_start": 10, "line_end": 12}],
            }

        with patch("bpmis_jira_tool.web._source_code_qa_provider_available", return_value=True), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today",
            return_value={"attempted": False, "status": "fresh"},
        ) as ensure_synced, patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
            side_effect=fake_query,
        ):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                upload = client.post(
                    "/api/source-code-qa/runtime-evidence",
                    data={
                        "pm_team": "AF",
                        "country": "SG",
                        "source_type": "apollo",
                        "file": (io.BytesIO(b"new.feature.enabled=true"), "apollo.properties"),
                    },
                    content_type="multipart/form-data",
                )
                response = client.post(
                    "/api/source-code-qa/effort-assessment",
                    json={
                        "pm_team": "AF",
                        "country": "SG",
                        "language": "zh",
                        "requirement": "Need to add a new approval flow for high-risk orders",
                        "llm_provider": "codex_cli_bridge",
                    },
                )
                self.assertEqual(response.status_code, 200)
                payload = response.get_json()
                snapshot = {}
                for _ in range(20):
                    status_response = client.get(f"/api/source-code-qa/effort-assessment-jobs/{payload['job_id']}")
                    snapshot = status_response.get_json()
                    if snapshot.get("state") == "completed":
                        break
                    time.sleep(0.05)
                events_response = client.get(f"/api/source-code-qa/effort-assessment-jobs/{payload['job_id']}/events")

        self.assertEqual(upload.status_code, 200)
        self.assertEqual(snapshot.get("state"), "completed")
        result = snapshot["results"][0]
        self.assertEqual(result["assessment"]["type"], "effort_assessment")
        self.assertEqual(result["assessment"]["language"], "zh")
        self.assertEqual(result["assessment"]["confidence"], "medium")
        self.assertIn("business_plan", result["assessment"])
        self.assertIn("technical_candidates", result["assessment"])
        self.assertIn("estimation_rubric", result["assessment"])
        self.assertEqual(result["runtime_evidence"][0]["filename"], "apollo.properties")
        self.assertGreaterEqual(len(captured_calls), 2)
        evidence_call = captured_calls[0]
        synthesis_call = captured_calls[1]
        self.assertEqual(evidence_call["answer_mode"], "retrieval_only")
        self.assertEqual(evidence_call["llm_budget_mode"], "cheap")
        self.assertLess(len(evidence_call["question"]), 2500)
        self.assertIn("Technical search terms:", evidence_call["question"])
        self.assertIn("Need to add a new approval flow for high-risk orders", evidence_call["question"])
        self.assertEqual(synthesis_call["answer_mode"], "auto")
        self.assertEqual(synthesis_call["llm_budget_mode"], "auto")
        self.assertTrue(synthesis_call["effort_assessment"])
        self.assertIn("Compact source-code evidence pack", synthesis_call["question"])
        self.assertIn("Internal evidence matrix for planning quality", synthesis_call["question"])
        self.assertIn("Every visible code change point must be grounded", synthesis_call["question"])
        self.assertIn("rule/workflow, backend/API, config/data, frontend, and integration/QA", synthesis_call["question"])
        self.assertIn("Required output sections:", synthesis_call["question"])
        self.assertIn("代码改动点 / Code Change Points", synthesis_call["question"])
        self.assertIn("BE 人天", synthesis_call["question"])
        self.assertIn("FE 人天", synthesis_call["question"])
        self.assertIn("Do not include visible sections named Evidence", synthesis_call["question"])
        self.assertNotIn("8. Source / Runtime Evidence", synthesis_call["question"])
        self.assertEqual(synthesis_call["pm_team"], "AF")
        self.assertEqual(synthesis_call["country"], "SG")
        self.assertEqual(synthesis_call["query_mode"], "deep")
        self.assertIn("new.feature.enabled=true", synthesis_call["runtime_evidence"][0]["text"])
        self.assertIn("effort_evidence_query", result)
        self.assertIn("effort_evidence_matrix", result)
        self.assertIn("effort_evidence_matrix_quality", result)
        self.assertIn("effort_generic_output_guard", result)
        self.assertEqual(result["effort_evidence_matrix"]["version"], 1)
        self.assertIn("effort_timing", result)
        self.assertIn("evidence_matrix_quality", result["effort_timing"])
        self.assertIn("generic_output_guard", result["effort_timing"])
        self.assertIn("代码改动点", result["llm_answer"])
        self.assertNotIn("[S1]", result["llm_answer"])
        self.assertNotIn("src/api.py", result["llm_answer"])
        self.assertNotIn("Confirmed / Inferred / Missing Evidence", result["llm_answer"])
        self.assertNotIn("Evidence Used", result["llm_answer"])
        ensure_synced.assert_not_called()
        self.assertIn(b"event: completed", events_response.data)

    def test_effort_assessment_wrong_repo_scope_returns_mismatch_without_query(self):
        requirement = (
            "CRMS SG BTI suspended case review: income review, Manual Review, APC approval, "
            "CBS purchase, Suspension Appeal and Monthly Review need final decision workflow."
        )

        with patch("bpmis_jira_tool.web._source_code_qa_provider_available", return_value=True), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query"
        ) as query:
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/effort-assessment",
                    json={
                        "pm_team": "GRC",
                        "country": "SG",
                        "language": "zh",
                        "requirement": requirement,
                        "llm_provider": "codex_cli_bridge",
                    },
                )
                self.assertEqual(response.status_code, 200)
                payload = response.get_json()
                snapshot = {}
                for _ in range(20):
                    status_response = client.get(f"/api/source-code-qa/effort-assessment-jobs/{payload['job_id']}")
                    snapshot = status_response.get_json()
                    if snapshot.get("state") == "completed":
                        break
                    time.sleep(0.05)

        query.assert_not_called()
        result = snapshot["results"][0]
        self.assertEqual(result["status"], "scope_mismatch")
        self.assertEqual(result["effort_evidence_status"], "scope_mismatch")
        self.assertEqual(result["assessment"]["confidence"], "scope_mismatch")
        self.assertEqual(result["assessment"]["scope_guard"]["suggested_pm_team"], "CRMS")
        self.assertEqual(result["assessment"]["scope_guard"]["suggested_country"], "SG")
        self.assertIn("请切换 PM Team 到 CRMS、Country 到 SG", result["llm_answer"])
        self.assertIn("不会基于当前 repo 生成 BE/FE 人天估算", result["llm_answer"])
        self.assertEqual(result["structured_assessment"]["code_change_points"], [])
        self.assertEqual(result["structured_assessment"]["be_estimate"], [])
        self.assertEqual(result["structured_assessment"]["fe_estimate"], [])

    def test_effort_assessment_exact_miss_still_returns_low_confidence_assessment(self):
        def fake_query(**kwargs):
            kwargs["progress_callback"]("direct_search", "Searching focused candidates.", 1, 1)
            return {
                "status": "no_match",
                "summary": "No exact table/path references were found in the indexed repositories.",
                "trace_id": "trace-effort-no-match",
                "matches": [],
                "exact_lookup": {"terms": ["missing.limit.table"], "matched_terms": []},
            }

        requirement = (
            "高收入人群 annual income >120K 的现金分期额度被信用卡额度限制。"
            "方案1: 用现金分期专项额度和103/104子产品额度控制。"
            "方案2: 引导申请独立cashline。"
        )
        with patch("bpmis_jira_tool.web._source_code_qa_provider_available", return_value=True), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
            side_effect=fake_query,
        ):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/effort-assessment",
                    json={
                        "pm_team": "CRMS",
                        "country": "SG",
                        "language": "zh",
                        "requirement": requirement,
                        "llm_provider": "codex_cli_bridge",
                    },
                )
                self.assertEqual(response.status_code, 200)
                payload = response.get_json()
                snapshot = {}
                for _ in range(20):
                    status_response = client.get(f"/api/source-code-qa/effort-assessment-jobs/{payload['job_id']}")
                    snapshot = status_response.get_json()
                    if snapshot.get("state") == "completed":
                        break
                    time.sleep(0.05)

        result = snapshot["results"][0]
        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["effort_evidence_status"], "warning")
        self.assertEqual(result["assessment"]["confidence"], "low")
        self.assertIn("business_plan", result["assessment"])
        self.assertIn("technical_candidates", result["assessment"])
        self.assertIn("estimation_rubric", result["assessment"])
        self.assertIn("missing_evidence", result["assessment"])
        self.assertIn("effort_evidence_matrix", result)
        self.assertEqual(result["effort_evidence_matrix_quality"]["status"], "planning_assumption")
        self.assertIn("code_change_points", result["structured_assessment"])
        self.assertTrue(result["structured_assessment"]["code_change_points"])
        self.assertIn("BE", result["llm_answer"])
        self.assertIn("FE", result["llm_answer"])
        self.assertIn("代码改动点", result["llm_answer"])
        self.assertIn("低置信度", result["llm_answer"])
        self.assertNotIn("Confirmed / Inferred / Missing Evidence", result["llm_answer"])
        self.assertNotIn("Source / Runtime Evidence", result["llm_answer"])
        self.assertNotIn("证据", result["llm_answer"])
        self.assertIn("cashline", result["assessment"]["technical_candidates"]["search_terms"])
        self.assertIn("subProductLimitInfos", result["assessment"]["technical_candidates"]["search_terms"])

    def test_effort_assessment_latest_result_restores_from_job_store(self):
        with patch("bpmis_jira_tool.web._source_code_qa_provider_available", return_value=True), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
            return_value={
                "status": "ok",
                "summary": "effort done",
                "llm_answer": "业务理解\n- ok\n\nBE 人天\n- 1-2 PD\n\nFE 人天\n- 0-1 PD",
                "trace_id": "trace-latest-effort",
                "matches": [{"repo": "CRMS", "path": "src/LimitService.java", "line_start": 3}],
            },
        ):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                initial_latest = client.get("/api/source-code-qa/effort-assessment/latest")
                response = client.post(
                    "/api/source-code-qa/effort-assessment",
                    json={
                        "pm_team": "CRMS",
                        "country": "SG",
                        "language": "zh",
                        "requirement": "现金分期额度需要和信用卡额度区分展示。",
                        "llm_provider": "codex_cli_bridge",
                    },
                )
                self.assertEqual(response.status_code, 200)
                payload = response.get_json()
                snapshot = {}
                for _ in range(20):
                    status_response = client.get(f"/api/source-code-qa/effort-assessment-jobs/{payload['job_id']}")
                    snapshot = status_response.get_json()
                    if snapshot.get("state") == "completed":
                        break
                    time.sleep(0.05)
                latest = client.get("/api/source-code-qa/effort-assessment/latest")

        self.assertEqual(initial_latest.status_code, 200)
        self.assertEqual(snapshot.get("state"), "completed")
        self.assertEqual(latest.status_code, 200)
        latest_payload = latest.get_json()
        self.assertEqual(latest_payload["status"], "ok")
        self.assertEqual(latest_payload["result"]["job_id"], payload["job_id"])
        self.assertIn("structured_assessment", latest_payload["result"]["assessment"])
        self.assertIn("business_plan", latest_payload["result"]["assessment"])

    def test_effort_assessment_rejects_unavailable_model_provider(self):
        with patch("bpmis_jira_tool.web._source_code_qa_provider_available", return_value=False):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/effort-assessment",
                    json={
                        "pm_team": "AF",
                        "country": "All",
                        "language": "en",
                        "requirement": "Add a new workflow",
                        "llm_provider": "not-real",
                    },
                )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.get_json()["message"], "Selected Source Code Q&A model is unavailable.")

    def test_source_code_qa_session_api_creates_lists_and_loads_session(self):
        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            created = client.post(
                "/api/source-code-qa/sessions",
                json={"pm_team": "AF", "country": "All", "llm_provider": "codex_cli_bridge"},
            )
            self.assertEqual(created.status_code, 200)
            session_payload = created.get_json()["session"]
            listed = client.get("/api/source-code-qa/sessions")
            loaded = client.get(f"/api/source-code-qa/sessions/{session_payload['id']}")

        self.assertEqual(listed.status_code, 200)
        self.assertEqual(loaded.status_code, 200)
        self.assertEqual(listed.get_json()["sessions"][0]["id"], session_payload["id"])
        self.assertEqual(loaded.get_json()["session"]["llm_provider"], "codex_cli_bridge")

    def test_source_code_qa_session_archive_hides_but_preserves_session(self):
        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            created = client.post(
                "/api/source-code-qa/sessions",
                json={"pm_team": "AF", "country": "All", "llm_provider": "codex_cli_bridge"},
            )
            session_id = created.get_json()["session"]["id"]
            archived = client.post(f"/api/source-code-qa/sessions/{session_id}/archive")
            listed = client.get("/api/source-code-qa/sessions")
            loaded = client.get(f"/api/source-code-qa/sessions/{session_id}")

        self.assertEqual(archived.status_code, 200)
        self.assertEqual(archived.get_json()["status"], "ok")
        self.assertEqual(listed.status_code, 200)
        self.assertEqual(listed.get_json()["sessions"], [])
        self.assertEqual(loaded.status_code, 200)
        self.assertEqual(loaded.get_json()["session"]["archived_at"], archived.get_json()["archived_at"])

    def test_query_api_uses_session_context_and_appends_exchange(self):
        captured = {}

        def fake_query(**kwargs):
            captured.update(kwargs)
            return {
                "status": "ok",
                "answer_mode": "auto",
                "summary": "answer summary",
                "llm_answer": "direct answer",
                "llm_provider": "codex_cli_bridge",
                "llm_model": "codex-cli",
                "trace_id": "trace-123",
                "structured_answer": {"direct_answer": "direct answer", "claims": [], "citations": [], "missing_evidence": [], "confidence": "high"},
                "matches": [{"repo": "Repo", "path": "src/App.java", "line_start": 1, "line_end": 3, "score": 9, "reason": "hit"}],
            }

        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            created = client.post("/api/source-code-qa/sessions", json={"pm_team": "AF", "country": "All", "llm_provider": "codex_cli_bridge"})
            session_id = created.get_json()["session"]["id"]
            store: SourceCodeQASessionStore = self.app.config["SOURCE_CODE_QA_SESSION_STORE"]
            store.append_exchange(
                session_id,
                owner_email="xiaodong.zheng@npt.sg",
                pm_team="AF",
                country="All",
                llm_provider="codex_cli_bridge",
                question="first question",
                result={"status": "ok", "summary": "first", "llm_answer": "first answer", "trace_id": "trace-1", "matches": []},
                context={"question": "first question", "trace_id": "trace-1", "matches_snapshot": [{"path": "src/First.java"}]},
            )
            with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today", return_value={"attempted": False, "status": "fresh"}), patch(
                "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
                side_effect=fake_query,
            ):
                response = client.post(
                    "/api/source-code-qa/query",
                    json={
                        "session_id": session_id,
                        "pm_team": "AF",
                        "country": "All",
                        "question": "follow up",
                        "answer_mode": "auto",
                        "llm_provider": "codex_cli_bridge",
                    },
                )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(captured["conversation_context"]["trace_id"], "trace-1")
        self.assertEqual(payload["session_id"], session_id)
        self.assertEqual(payload["session"]["message_count"], 4)
        self.assertEqual(payload["session"]["last_context"]["question"], "follow up")
        self.assertEqual(payload["session"]["last_context"]["recent_turns"][0]["question"], "first question")
        self.assertEqual(payload["session"]["last_context"]["recent_turns"][0]["trace_id"], "trace-1")

    def test_session_context_preserves_raw_codex_answer(self):
        context = _build_source_code_qa_session_context(
            {
                "status": "ok",
                "trace_id": "trace-deep",
                "summary": "Found evidence",
                "llm_answer": "Raw Codex answer",
                "llm_provider": "codex_cli_bridge",
                "llm_model": "gpt-5.5",
                "query_mode": "deep",
                "structured_answer": {
                    "direct_answer": "Structured answer should not replace raw text",
                    "claims": [],
                    "missing_evidence": [],
                    "confidence": "medium",
                },
                "matches": [],
            },
            {"pm_team": "AF", "country": "All", "question": "Can you tell me how blacklist is used?"},
        )

        self.assertEqual(context["query_mode"], "deep")
        self.assertEqual(context["answer"], "Raw Codex answer")

    def test_query_passes_attachment_metadata_and_text_to_service(self):
        captured = {}

        def fake_query(**kwargs):
            captured.update(kwargs)
            return {
                "status": "ok",
                "answer_mode": "auto",
                "summary": "answer summary",
                "llm_answer": "direct answer",
                "llm_provider": "codex_cli_bridge",
                "llm_model": "codex-cli",
                "trace_id": "trace-attachment",
                "matches": [],
            }

        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            created = client.post("/api/source-code-qa/sessions", json={"pm_team": "AF", "country": "All", "llm_provider": "codex_cli_bridge"})
            session_id = created.get_json()["session"]["id"]
            uploaded = client.post(
                "/api/source-code-qa/attachments",
                data={
                    "session_id": session_id,
                    "file": (io.BytesIO(b"uploaded source question context"), "context.md"),
                },
                content_type="multipart/form-data",
            ).get_json()["attachment"]
            with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today", return_value={"attempted": False, "status": "fresh"}), patch(
                "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
                side_effect=fake_query,
            ):
                response = client.post(
                    "/api/source-code-qa/query",
                    json={
                        "session_id": session_id,
                        "pm_team": "AF",
                        "country": "All",
                        "question": "use attachment",
                        "llm_provider": "codex_cli_bridge",
                        "attachment_ids": [uploaded["id"]],
                    },
                )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(captured["attachments"][0]["filename"], "context.md")
        self.assertIn("uploaded source question context", captured["attachments"][0]["text"])
        self.assertEqual(payload["attachments"][0]["id"], uploaded["id"])
        self.assertEqual(payload["session"]["messages"][-2]["attachments"][0]["filename"], "context.md")

    def test_runtime_evidence_upload_is_scoped_and_passed_to_query(self):
        captured = {}

        def fake_query(**kwargs):
            captured.update(kwargs)
            return {
                "status": "ok",
                "answer_mode": "auto",
                "summary": "answer summary",
                "llm_answer": "direct answer",
                "llm_provider": "codex_cli_bridge",
                "llm_model": "codex-cli",
                "trace_id": "trace-runtime",
                "matches": [],
            }

        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            sg_upload = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={
                    "pm_team": "AF",
                    "country": "SG",
                    "source_type": "apollo",
                    "file": (io.BytesIO(b"apollo.sg.rule.enabled=true"), "apollo.properties"),
                },
                content_type="multipart/form-data",
            )
            ph_upload = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={
                    "pm_team": "AF",
                    "country": "PH",
                    "source_type": "db",
                    "file": (io.BytesIO(b"rule_id,status\nC0204v2,online"), "rules.csv"),
                },
                content_type="multipart/form-data",
            )
            listed = client.get("/api/source-code-qa/runtime-evidence?pm_team=AF&country=SG")
            with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today", return_value={"attempted": False, "status": "fresh"}), patch(
                "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
                side_effect=fake_query,
            ):
                response = client.post(
                    "/api/source-code-qa/query",
                    json={
                        "pm_team": "AF",
                        "country": "SG",
                        "question": "compare runtime config",
                        "llm_provider": "codex_cli_bridge",
                    },
                )

        self.assertEqual(sg_upload.status_code, 200)
        self.assertEqual(ph_upload.status_code, 200)
        self.assertEqual(listed.status_code, 200)
        self.assertEqual(listed.get_json()["evidence"][0]["country"], "SG")
        self.assertEqual(response.status_code, 200)
        countries = {item["country"] for item in captured["runtime_evidence"]}
        self.assertEqual(countries, {"SG"})
        self.assertTrue(any("apollo.sg.rule.enabled" in item.get("text", "") for item in captured["runtime_evidence"]))
        self.assertEqual(response.get_json()["runtime_evidence"][0]["pm_team"], "AF")

    def test_runtime_evidence_all_scope_does_not_merge_country_evidence(self):
        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            for country, filename in [("SG", "sg.txt"), ("PH", "ph.txt"), ("ID", "id.txt"), ("All", "all.txt")]:
                response = client.post(
                    "/api/source-code-qa/runtime-evidence",
                    data={
                        "pm_team": "AF",
                        "country": country,
                        "source_type": "db",
                        "file": (io.BytesIO(f"{country} runtime evidence".encode("utf-8")), filename),
                    },
                    content_type="multipart/form-data",
                )
                self.assertEqual(response.status_code, 200)
                time.sleep(0.01)
            listed = client.get("/api/source-code-qa/runtime-evidence?pm_team=AF&country=All")

        self.assertEqual(listed.status_code, 200)
        payload = listed.get_json()
        self.assertEqual({(item["pm_team"], item["country"]) for item in payload["evidence"]}, {("AF", "All")})

    def test_all_country_query_ignores_runtime_evidence_and_uses_code_only(self):
        captured = {}

        def fake_query(**kwargs):
            captured.update(kwargs)
            return {
                "status": "ok",
                "answer_mode": "auto",
                "summary": "code-only answer",
                "llm_answer": "code-only answer",
                "llm_provider": "codex_cli_bridge",
                "llm_model": "codex-cli",
                "trace_id": "trace-all-code-only",
                "matches": [],
            }

        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            for country, filename, content in [
                ("SG", "apollo-sg.properties", b"apollo.sg.rule.enabled=true"),
                ("All", "apollo-all.properties", b"apollo.all.rule.enabled=true"),
            ]:
                upload = client.post(
                    "/api/source-code-qa/runtime-evidence",
                    data={
                        "pm_team": "AF",
                        "country": country,
                        "source_type": "apollo",
                        "file": (io.BytesIO(content), filename),
                    },
                    content_type="multipart/form-data",
                )
                self.assertEqual(upload.status_code, 200)
            with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today", return_value={"attempted": False, "status": "fresh"}), patch(
                "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
                side_effect=fake_query,
            ):
                response = client.post(
                    "/api/source-code-qa/query",
                    json={"pm_team": "AF", "country": "All", "question": "check all-country code path", "llm_provider": "codex_cli_bridge"},
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(captured["runtime_evidence"], [])
        self.assertEqual(response.get_json()["runtime_evidence"], [])

    def test_grc_and_af_country_query_uses_all_country_data_dictionary(self):
        captured = {}

        def fake_query(**kwargs):
            captured.update(kwargs)
            return {
                "status": "ok",
                "answer_mode": "auto",
                "summary": "answer summary",
                "llm_answer": "direct answer",
                "llm_provider": "codex_cli_bridge",
                "llm_model": "codex-cli",
                "trace_id": "trace-dictionary",
                "matches": [],
            }

        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            grc_dictionary = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={
                    "pm_team": "GRC",
                    "country": "All",
                    "source_type": "data_dictionary",
                    "file": (io.BytesIO(b"table,column,meaning\nbcf_global_lock,lock_key,Global lock key"), "grc-dictionary.csv"),
                },
                content_type="multipart/form-data",
            )
            grc_db = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={
                    "pm_team": "GRC",
                    "country": "SG",
                    "source_type": "db",
                    "file": (io.BytesIO(b"bcf_global_lock rows from SG"), "grc-sg-db.txt"),
                },
                content_type="multipart/form-data",
            )
            af_dictionary = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={
                    "pm_team": "AF",
                    "country": "All",
                    "source_type": "data_dictionary",
                    "file": (io.BytesIO(b"table,column,meaning\naf_rule,rule_id,Risk rule id"), "af-dictionary.csv"),
                },
                content_type="multipart/form-data",
            )
            with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today", return_value={"attempted": False, "status": "fresh"}), patch(
                "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
                side_effect=fake_query,
            ):
                grc_response = client.post(
                    "/api/source-code-qa/query",
                    json={"pm_team": "GRC", "country": "SG", "question": "write SQL for global lock", "llm_provider": "codex_cli_bridge"},
                )
                grc_runtime = list(captured["runtime_evidence"])
                af_response = client.post(
                    "/api/source-code-qa/query",
                    json={"pm_team": "AF", "country": "PH", "question": "write SQL for rule", "llm_provider": "codex_cli_bridge"},
                )
                af_runtime = list(captured["runtime_evidence"])

        self.assertEqual(grc_dictionary.status_code, 200)
        self.assertEqual(grc_db.status_code, 200)
        self.assertEqual(af_dictionary.status_code, 200)
        self.assertEqual(grc_response.status_code, 200)
        self.assertEqual(af_response.status_code, 200)
        self.assertEqual({(item["pm_team"], item["country"], item["source_type"]) for item in grc_runtime}, {("GRC", "All", "data_dictionary"), ("GRC", "SG", "db")})
        self.assertTrue(any("bcf_global_lock" in item.get("text", "") for item in grc_runtime))
        self.assertEqual({(item["pm_team"], item["country"], item["source_type"]) for item in af_runtime}, {("AF", "All", "data_dictionary")})
        self.assertTrue(any("af_rule" in item.get("text", "") for item in af_runtime))

    def test_data_dictionary_xlsx_extracts_many_table_sheets(self):
        workbook = Workbook()
        overview = workbook.active
        overview.title = "Full Dictionary"
        overview.append(["table catalog"])
        for index in range(1, 8):
            sheet = workbook.create_sheet(f"table_{index}_info")
            sheet.append(["字段名", "数据类型", "是否必填", "业务含义"])
            sheet.append([f"`field_{index}`", "varchar(64)", "是", f"Meaning for table {index}"])
        buffer = io.BytesIO()
        workbook.save(buffer)

        store = self.app.config["SOURCE_CODE_QA_RUNTIME_EVIDENCE_STORE"]
        saved = store.save_bytes(
            pm_team="GRC",
            country="All",
            source_type="data_dictionary",
            uploaded_by="xiaodong.zheng@npt.sg",
            filename="dictionary.xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=buffer.getvalue(),
        )
        resolved = store.resolve_scope(pm_team="GRC", country="SG")
        dictionary_text = next(item["text"] for item in resolved if item["id"] == saved["id"])

        self.assertIn("[Data dictionary sheet: table_7_info]", dictionary_text)
        self.assertIn("field_7", dictionary_text)
        self.assertGreater(saved["text_char_count"], 700)

    def test_data_dictionary_prompt_allows_larger_context(self):
        section = SourceCodeQAService._runtime_evidence_prompt_section(
            [
                {
                    "id": "dict-1",
                    "filename": "dictionary.xlsx",
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "kind": "document",
                    "source_type": "data_dictionary",
                    "pm_team": "GRC",
                    "country": "All",
                    "size": 100,
                    "sha256": "a" * 64,
                    "created_at": "2026-05-05T00:00:00Z",
                    "uploaded_by": "owner@npt.sg",
                    "text": "start\n" + ("x" * 7000) + "\nlate_table_name",
                }
            ]
        )

        self.assertIn("late_table_name", section)
        self.assertIn("Data dictionary handling", section)
        self.assertIn("AF and GRC, data_dictionary uploads apply to all country selections", section)
        self.assertIn("SG, ID, and PH share the same table and data-field definitions", section)
        self.assertIn("RC and Compliance are business aliases", section)
        self.assertIn("first explain the actual SQL/table logic", section)
        self.assertIn("chosen table names, key filters, joins, reviewer/status rows, and timestamp assumptions", section)
        self.assertIn("selected country points to that country's separate runtime DB instance", section)
        self.assertIn("not the first sentence or main conclusion", section)

    def test_crms_cannot_use_all_country_runtime_evidence(self):
        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            response = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={
                    "pm_team": "CRMS",
                    "country": "All",
                    "source_type": "data_dictionary",
                    "file": (io.BytesIO(b"table,column,meaning\ncr_customer,id,Customer id"), "crms-dictionary.csv"),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn("supported only for AF and GRC", response.get_json()["message"])

    def test_source_code_qa_sql_answer_creates_downloadable_package(self):
        def fake_query(**kwargs):
            return {
                "status": "ok",
                "answer_mode": "auto",
                "summary": "SQL generated",
                "llm_answer": "Use this SQL:\n```sql\nselect lock_key from bcf_global_lock where status = 'ACTIVE';\n```\n",
                "llm_provider": "codex_cli_bridge",
                "llm_model": "codex-cli",
                "trace_id": "trace-sql-package",
                "matches": [{"repo": "GRC Portal", "path": "mapper/GlobalLockMapper.xml", "line_start": 12}],
            }

        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            created = client.post("/api/source-code-qa/sessions", json={"pm_team": "GRC", "country": "SG", "llm_provider": "codex_cli_bridge"})
            session_id = created.get_json()["session"]["id"]
            upload = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={
                    "pm_team": "GRC",
                    "country": "All",
                    "source_type": "data_dictionary",
                    "file": (io.BytesIO(b"table,column,meaning\nbcf_global_lock,lock_key,Global lock key"), "grc-dictionary.csv"),
                },
                content_type="multipart/form-data",
            )
            with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today", return_value={"attempted": False, "status": "fresh"}), patch(
                "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
                side_effect=fake_query,
            ):
                response = client.post(
                    "/api/source-code-qa/query",
                    json={
                        "session_id": session_id,
                        "pm_team": "GRC",
                        "country": "SG",
                        "question": "write SQL for global lock",
                        "llm_provider": "codex_cli_bridge",
                    },
                )

            payload = response.get_json()
            artifact = payload["generated_artifacts"][0]
            download = client.get(f"/api/source-code-qa/generated-artifacts/{artifact['id']}?session_id={session_id}")

        self.assertEqual(upload.status_code, 200)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(download.status_code, 200)
        self.assertEqual(download.headers["Content-Type"], "application/zip")
        with zipfile.ZipFile(io.BytesIO(download.data)) as archive:
            self.assertIn("query.sql", archive.namelist())
            self.assertIn("README.md", archive.namelist())
            sql = archive.read("query.sql").decode("utf-8")
            self.assertIn("SELECT lock_key", sql)
            self.assertRegex(sql, r"\n\s+FROM bcf_global_lock")
            readme = archive.read("README.md").decode("utf-8")
            self.assertIn("GRC:SG", readme)
            self.assertIn("grc-dictionary.csv", readme)
            self.assertIn("mapper/GlobalLockMapper.xml", readme)
            self.assertIn("## SQL Rough Logic", readme)
            self.assertIn("Uses table(s): bcf_global_lock", readme)
            self.assertIn("## Tables Used", readme)
            self.assertIn("`bcf_global_lock`", readme)

    def test_source_code_qa_json_direct_answer_sql_creates_downloadable_package(self):
        def fake_query(**kwargs):
            return {
                "status": "ok",
                "answer_mode": "auto",
                "summary": "SQL generated",
                "answer": json.dumps(
                    {
                        "direct_answer": (
                            "Use reviewer tables. SQL: WITH latest_ticket AS "
                            "(SELECT * FROM ticket_info WHERE authorization_type = 'Incident - Authorize New Incident') "
                            "SELECT event_id FROM latest_ticket;"
                        ),
                        "source_code_evidence": ["Mapper evidence with other SQL: SELECT should_not_leak FROM evidence;"],
                    }
                ),
                "llm_provider": "codex_cli_bridge",
                "llm_model": "codex-cli",
                "trace_id": "trace-json-sql-package",
                "matches": [{"repo": "GRC Portal", "path": "mapper/TicketMapper.xml", "line_start": 8}],
            }

        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            created = client.post("/api/source-code-qa/sessions", json={"pm_team": "GRC", "country": "All", "llm_provider": "codex_cli_bridge"})
            session_id = created.get_json()["session"]["id"]
            with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today", return_value={"attempted": False, "status": "fresh"}), patch(
                "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
                side_effect=fake_query,
            ):
                response = client.post(
                    "/api/source-code-qa/query",
                    json={
                        "session_id": session_id,
                        "pm_team": "GRC",
                        "country": "All",
                        "question": "write SQL for reviewer status",
                        "llm_provider": "codex_cli_bridge",
                    },
                )
            payload = response.get_json()
            artifact = payload["generated_artifacts"][0]
            download = client.get(f"/api/source-code-qa/generated-artifacts/{artifact['id']}?session_id={session_id}")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(download.status_code, 200)
        with zipfile.ZipFile(io.BytesIO(download.data)) as archive:
            sql = archive.read("query.sql").decode("utf-8")
            readme = archive.read("README.md").decode("utf-8")
        self.assertIn("WITH latest_ticket", sql)
        self.assertIn("SELECT event_id", sql)
        self.assertNotIn("should_not_leak", sql)
        self.assertIn("Builds intermediate CTE(s): latest_ticket", readme)
        self.assertIn("`ticket_info`", readme)

    def test_source_code_qa_config_includes_runtime_capabilities(self):
        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            apollo_upload = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={
                    "pm_team": "AF",
                    "country": "SG",
                    "source_type": "apollo",
                    "file": (io.BytesIO(b"apollo.sg.rule.enabled=true"), "apollo.properties"),
                },
                content_type="multipart/form-data",
            )
            db_upload = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={
                    "pm_team": "AF",
                    "country": "SG",
                    "source_type": "db",
                    "file": (io.BytesIO(b"rule_id,status\nC0204v2,online"), "rules.csv"),
                },
                content_type="multipart/form-data",
            )
            response = client.get("/api/source-code-qa/config")

        self.assertEqual(apollo_upload.status_code, 200)
        self.assertEqual(db_upload.status_code, 200)
        self.assertEqual(response.status_code, 200)
        capabilities = response.get_json()["options"]["runtime_capabilities"]
        self.assertTrue(capabilities["AF"]["SG"]["hasConfig"])
        self.assertTrue(capabilities["AF"]["SG"]["hasDB"])
        self.assertFalse(capabilities["AF"]["PH"]["hasConfig"])
        self.assertFalse(capabilities["CRMS"]["ID"]["hasDB"])

    def test_runtime_evidence_prompt_marks_apollo_as_uat_reference_only(self):
        section = SourceCodeQAService._runtime_evidence_prompt_section(
            [
                {
                    "id": "e1",
                    "filename": "apollo.properties",
                    "source_type": "apollo",
                    "pm_team": "AF",
                    "country": "SG",
                    "kind": "text",
                    "text": "feature.enabled=true",
                }
            ]
        )

        self.assertIn("UAT/non-Live", section)
        self.assertIn("never use them as confirmed Live/production configuration facts", section)
        self.assertIn("app/env/namespace", section)
        self.assertIn("Do not say an Apollo export is missing when it is present in uploaded runtime evidence", section)
        self.assertIn("choose the matching app namespace", section)
        self.assertNotIn("production DB/Apollo/config snapshots", section)

    def test_runtime_evidence_apollo_zip_extracts_nested_text_configs(self):
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as archive:
            archive.writestr("sg/application.properties", "apollo.sg.rule.enabled=true")
            archive.writestr("sg/nested/rules.yaml", "challenge: C0204v2")
            archive.writestr("sg/ignore.bin", b"\x00\x01binary")
        zip_buffer.seek(0)

        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            upload = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={
                    "pm_team": "AF",
                    "country": "SG",
                    "source_type": "apollo",
                    "file": (zip_buffer, "apollo-config.zip"),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(upload.status_code, 200)
        evidence = upload.get_json()["evidence"]
        self.assertEqual(evidence["filename"], "apollo-config.zip")
        self.assertEqual(evidence["kind"], "archive")
        self.assertIn("sg/application.properties", evidence["summary"])
        self.assertIn("apollo.sg.rule.enabled=true", evidence["summary"])
        self.assertIn("sg/nested/rules.yaml", evidence["summary"])

    def test_runtime_evidence_apollo_zip_allows_larger_config_trees(self):
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as archive:
            for index in range(120):
                archive.writestr(f"apollo/apps/app-{index}/application.properties", f"feature.{index}=true")
        zip_buffer.seek(0)

        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            upload = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={
                    "pm_team": "AF",
                    "country": "PH",
                    "source_type": "apollo",
                    "file": (zip_buffer, "apollo-large.zip"),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(upload.status_code, 200)
        evidence = upload.get_json()["evidence"]
        self.assertEqual(evidence["kind"], "archive")
        self.assertIn("apollo/apps/app-0/application.properties", evidence["summary"])

    def test_regular_source_code_qa_attachment_still_rejects_zip(self):
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as archive:
            archive.writestr("notes.txt", "hello")
        zip_buffer.seek(0)

        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            created = client.post("/api/source-code-qa/sessions", json={"pm_team": "AF", "country": "All", "llm_provider": "codex_cli_bridge"})
            session_id = created.get_json()["session"]["id"]
            upload = client.post(
                "/api/source-code-qa/attachments",
                data={
                    "session_id": session_id,
                    "file": (zip_buffer, "notes.zip"),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(upload.status_code, 400)
        self.assertIn("archive", upload.get_json()["message"].lower())

    def test_query_empty_config_is_controlled(self):
        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            response = client.post(
                "/api/source-code-qa/query",
                json={"pm_team": "CRMS", "country": "SG", "question": "where is approval logic"},
            )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["status"], "empty_config")
        self.assertEqual(payload["answer_mode"], "auto")

    def test_query_api_defaults_llm_budget_to_auto(self):
        captured = {}

        def fake_query(**kwargs):
            captured.update(kwargs)
            return {"status": "ok", "answer_mode": "retrieval_only", "matches": []}

        with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.query", side_effect=fake_query), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today",
            return_value={"attempted": False, "status": "fresh"},
        ) as ensure_synced:
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/query",
                    json={"pm_team": "AF", "country": "All", "question": "where is createIssue", "answer_mode": "auto"},
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(captured["llm_budget_mode"], "auto")
        ensure_synced.assert_not_called()
        self.assertEqual(response.get_json()["auto_sync"]["status"], "skipped")

    def test_af_country_query_uses_country_runtime_evidence_with_shared_repo_scope(self):
        captured = {}

        def fake_query(**kwargs):
            captured.update(kwargs)
            return {
                "status": "ok",
                "answer_mode": "auto",
                "matches": [],
                "trace_id": "trace-af-ph",
                "summary": "answer",
            }

        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            upload = client.post(
                "/api/source-code-qa/runtime-evidence",
                data={
                    "pm_team": "AF",
                    "country": "PH",
                    "source_type": "apollo",
                    "file": (io.BytesIO(b"apollo.ph.only=true"), "apollo-ph.properties"),
                },
                content_type="multipart/form-data",
            )
            self._login(client, "xiaodong.zheng@npt.sg")
            with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.query", side_effect=fake_query), patch(
                "bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today",
                return_value={"attempted": False, "status": "fresh", "key": "AF:All"},
            ) as ensure_synced:
                response = client.post(
                    "/api/source-code-qa/query",
                    json={"pm_team": "AF", "country": "PH", "question": "check PH config", "answer_mode": "auto"},
                )

        self.assertEqual(upload.status_code, 200)
        self.assertEqual(response.status_code, 200)
        ensure_synced.assert_not_called()
        self.assertEqual(captured["pm_team"], "AF")
        self.assertEqual(captured["country"], "PH")
        self.assertEqual(captured["runtime_evidence"][0]["country"], "PH")
        self.assertIn("apollo.ph.only=true", captured["runtime_evidence"][0]["text"])
        self.assertEqual(response.get_json()["runtime_evidence"][0]["country"], "PH")

    def test_query_api_coerces_legacy_retrieval_only_to_auto(self):
        captured = {}

        def fake_query(**kwargs):
            captured.update(kwargs)
            return {"status": "ok", "answer_mode": kwargs["answer_mode"], "matches": []}

        with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.query", side_effect=fake_query), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today",
            return_value={"attempted": False, "status": "fresh"},
        ):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/query",
                    json={"pm_team": "AF", "country": "All", "question": "where is createIssue", "answer_mode": "retrieval_only"},
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(captured["answer_mode"], "auto")
        self.assertEqual(response.get_json()["answer_mode"], "auto")

    def test_query_api_uses_requested_llm_provider(self):
        selected = []

        def fake_query(**kwargs):
            return {"status": "ok", "answer_mode": "auto", "matches": [], "llm_provider": "codex_cli_bridge"}

        original = SourceCodeQAService.with_llm_provider

        def fake_with_provider(service, provider):
            selected.append(provider)
            return original(service, provider)

        with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.with_llm_provider", new=fake_with_provider), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today",
            return_value={"attempted": False, "status": "fresh"},
        ), patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.query", side_effect=fake_query):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/query",
                    json={
                        "pm_team": "AF",
                        "country": "All",
                        "question": "where is createIssue",
                        "answer_mode": "auto",
                        "llm_provider": "codex_cli_bridge",
                    },
                )

        self.assertEqual(response.status_code, 200)
        self.assertIn("codex_cli_bridge", selected)

    def test_query_api_invalid_llm_provider_is_rejected(self):
        selected = []

        def fake_query(**kwargs):
            return {"status": "ok", "answer_mode": "auto", "matches": [], "llm_provider": "codex_cli_bridge"}

        original = SourceCodeQAService.with_llm_provider

        def fake_with_provider(service, provider):
            selected.append(provider)
            return original(service, provider)

        with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.with_llm_provider", new=fake_with_provider), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today",
            return_value={"attempted": False, "status": "fresh"},
        ), patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.query", side_effect=fake_query):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/query",
                    json={
                        "pm_team": "AF",
                        "country": "All",
                        "question": "where is createIssue",
                        "answer_mode": "auto",
                        "llm_provider": "not-real",
                    },
                )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(selected, [])

    def test_query_api_ignores_client_llm_budget_selection(self):
        captured = {}

        def fake_query(**kwargs):
            captured.update(kwargs)
            return {"status": "ok", "answer_mode": "auto", "matches": []}

        with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.query", side_effect=fake_query), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today",
            return_value={"attempted": False, "status": "fresh"},
        ):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/query",
                    json={
                        "pm_team": "AF",
                        "country": "All",
                        "question": "where is createIssue",
                        "answer_mode": "auto",
                        "llm_budget_mode": "deep",
                    },
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(captured["llm_budget_mode"], "auto")

    def test_query_api_normalizes_removed_fast_query_mode_to_deep(self):
        captured = {}

        def fake_query(**kwargs):
            captured.update(kwargs)
            return {"status": "ok", "answer_mode": "auto", "query_mode": kwargs["query_mode"], "matches": []}

        with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.query", side_effect=fake_query), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today",
            return_value={"attempted": False, "status": "fresh"},
        ):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/query",
                    json={
                        "pm_team": "AF",
                        "country": "All",
                        "question": "where is createIssue",
                        "answer_mode": "auto",
                        "query_mode": "fast",
                    },
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(captured["query_mode"], "deep")
        self.assertEqual(captured["llm_budget_mode"], "auto")
        self.assertEqual(response.get_json()["query_mode"], "deep")

    def test_query_api_returns_json_for_unexpected_failures(self):
        with patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today",
            return_value={"attempted": False, "status": "fresh"},
        ), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
            side_effect=RuntimeError("boom"),
        ):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/query",
                    json={"pm_team": "CRMS", "country": "SG", "question": "where is income logic"},
                )

        self.assertEqual(response.status_code, 500)
        self.assertEqual(response.content_type, "application/json")
        payload = response.get_json()
        self.assertEqual(payload["status"], "error")
        self.assertEqual(payload["error_category"], "source_code_qa_internal")
        self.assertTrue(payload["error_retryable"])

    def test_query_api_does_not_sync_before_answer_by_default(self):
        with patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today",
            return_value={"attempted": True, "status": "ok", "reason": "synced before query"},
        ) as ensure_synced, patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
            return_value={"status": "ok", "answer_mode": "retrieval_only", "matches": []},
        ):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/query",
                    json={"pm_team": "AF", "country": "All", "question": "where is createIssue"},
                )

        self.assertEqual(response.status_code, 200)
        ensure_synced.assert_not_called()
        self.assertEqual(response.get_json()["auto_sync"]["status"], "skipped")

    def test_background_query_sync_blocks_when_scope_has_no_queryable_index(self):
        calls = []
        progress_events = []

        class ColdIndexService:
            def mapping_key(self, pm_team, country):
                return f"{pm_team}:{country}"

            def index_health_payload(self):
                calls.append("health")
                return {
                    "keys": {
                        "GRC:All": {
                            "repos": [
                                {"index": {"state": "missing", "queryable": False}},
                                {"index": {"state": "missing", "queryable": False}},
                            ]
                        }
                    }
                }

            def ensure_synced_today(self, *, pm_team, country):
                calls.append(("sync", pm_team, country))
                return {"attempted": True, "status": "ok", "key": f"{pm_team}:{country}"}

            def ensure_synced_today_background(self, *, pm_team, country):  # pragma: no cover - must not be called.
                calls.append(("background", pm_team, country))
                return {"attempted": False, "status": "background_queued"}

        with patch.dict(os.environ, {"SOURCE_CODE_QA_QUERY_SYNC_MODE": "background"}, clear=False):
            with self.app.app_context():
                result = _prepare_source_code_qa_auto_sync(
                    ColdIndexService(),
                    pm_team="GRC",
                    country="All",
                    progress_callback=lambda event, message, current, total: progress_events.append(event),
                )

        self.assertEqual(result["status"], "ok")
        self.assertEqual(calls, ["health", ("sync", "GRC", "All")])
        self.assertEqual(progress_events, ["auto_sync_check", "auto_sync_completed"])

    def test_cloud_run_local_agent_query_queues_auto_sync_in_background(self):
        calls = []

        class FakeLocalAgentClient:
            def source_code_qa_ensure_synced_today(self, *, pm_team, country, background=False):
                calls.append(("ensure", pm_team, country, background))
                return {"status": "background_queued", "attempted": False, "key": f"{pm_team}:{country}"}

            def source_code_qa_query(self, payload, *, progress_callback=None):
                calls.append(("query", payload["pm_team"], payload["country"], payload.get("question")))
                if progress_callback:
                    progress_callback("codex_stream", "Reading repo files.", 0, 0)
                return {"status": "ok", "answer_mode": "auto", "summary": "agent answer", "matches": []}

            def source_code_qa_runtime_evidence_resolve(self, *, pm_team, country):
                return []

        with patch.dict(
            os.environ,
            {
                "FLASK_SECRET_KEY": "test-secret",
                "TEAM_PORTAL_DATA_DIR": self.temp_dir.name,
                "TEAM_ALLOWED_EMAIL_DOMAINS": "npt.sg",
                "LOCAL_AGENT_MODE": "sync",
                "LOCAL_AGENT_BASE_URL": "https://agent.example",
                "LOCAL_AGENT_HMAC_SECRET": "shared-secret",
                "LOCAL_AGENT_SOURCE_CODE_QA_ENABLED": "true",
                "SOURCE_CODE_QA_QUERY_SYNC_MODE": "background",
            },
            clear=False,
        ):
            app = create_app()
            app.testing = True

        with patch.dict(os.environ, {"SOURCE_CODE_QA_QUERY_SYNC_MODE": "background"}, clear=False), patch(
            "bpmis_jira_tool.web._source_code_qa_provider_available", return_value=True
        ), patch(
            "bpmis_jira_tool.web._build_local_agent_client",
            return_value=FakeLocalAgentClient(),
        ):
            with app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/query",
                    json={"pm_team": "AF", "country": "All", "question": "where is createIssue", "llm_provider": "codex_cli_bridge"},
                )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(calls[0], ("ensure", "AF", "All", True))
        self.assertEqual(calls[1], ("query", "AF", "All", "where is createIssue"))
        self.assertEqual(response.get_json()["auto_sync"]["status"], "background_queued")

    def test_config_api_reports_llm_not_ready_by_default(self):
        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            response = client.get("/api/source-code-qa/config")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertFalse(payload["llm_ready"])
        self.assertEqual(payload["llm_provider"], "codex_cli_bridge")
        self.assertEqual(payload["llm_policy"]["provider"]["provider"], "codex_cli_bridge")
        self.assertEqual(payload["llm_policy"]["router"]["version"], 12)
        self.assertEqual(payload["llm_policy"]["versions"]["cache"], 21)
        self.assertEqual(payload["llm_policy"]["versions"]["runtime"], 2)
        self.assertNotIn("max_retries", payload["llm_policy"]["runtime"])
        self.assertEqual(payload["llm_policy"]["model_policy"]["answer"]["model"], payload["llm_model"])
        self.assertIn("codex_model_routes", payload["llm_policy"])
        self.assertEqual(payload["llm_policy"]["codex_model_routes"]["routes"]["cheap"]["reasoning_effort"], "low")
        self.assertEqual(payload["llm_policy"]["codex_model_routes"]["reasoning_control"], "codex_cli_model_reasoning_effort")
        self.assertTrue(payload["llm_policy"]["judge"]["enabled"])
        self.assertEqual(payload["llm_policy"]["judge"]["mode"], "deterministic_evidence_judge")
        self.assertNotIn("cache", payload["llm_policy"]["judge"])
        self.assertEqual(payload["llm_policy"]["planner_tools"]["version"], 1)
        self.assertEqual(payload["llm_policy"]["semantic_retrieval"]["model"], "local-token-hybrid-v1")
        self.assertEqual(payload["llm_policy"]["semantic_retrieval"]["embedding_provider"]["provider"], "local_token_hybrid")
        self.assertEqual(payload["index_health"]["status"], "not_configured")
        self.assertEqual(payload["release_gate"]["status"], "missing")
        self.assertEqual(payload["domain_knowledge"]["domains"]["CRMS"]["label"], "Credit Risk")
        self.assertIn("GRC", payload["domain_knowledge"]["domains"])
        self.assertEqual(payload["options"]["answer_modes"][0]["value"], "auto")
        self.assertEqual([item["value"] for item in payload["options"]["llm_providers"]], ["codex_cli_bridge"])
        self.assertEqual(set(payload["llm_providers"]), {"codex_cli_bridge"})
        self.assertNotIn("llm_budget_modes", payload["options"])

    def test_scheduled_sync_dry_run_catches_up_latest_due_date(self):
        class FakeService:
            def load_config(self):
                return {"mappings": {"AF:All": [{}], "GRC:All": [{}]}}

        with tempfile.TemporaryDirectory() as temp_dir, patch(
            "scripts.source_code_qa_scheduled_sync.Settings.from_env",
            return_value=SimpleNamespace(),
        ), patch(
            "scripts.source_code_qa_scheduled_sync.source_code_qa_data_root",
            return_value=Path(temp_dir),
        ), patch(
            "scripts.source_code_qa_scheduled_sync.build_source_code_qa_service_from_settings",
            return_value=FakeService(),
        ):
            payload = run_scheduled_sync(
                dry_run=True,
                now=datetime.fromisoformat("2026-05-23T10:00:00+08:00"),
            )

        self.assertEqual(payload["status"], "would_run")
        self.assertEqual(payload["scheduled_date"], "2026-05-22")
        self.assertEqual(payload["mappings"], ["AF:All", "GRC:All"])

    def test_scheduled_sync_runs_all_configured_mappings_once_per_scheduled_date(self):
        calls = []

        class FakeService:
            def load_config(self):
                return {"mappings": {"AF:All": [{}], "GRC:All": [{}]}}

            def sync(self, *, pm_team, country):
                calls.append((pm_team, country))
                return {"status": "ok", "results": [{"display_name": "Repo"}], "job": {"job_id": f"{pm_team}-{country}"}}

        with tempfile.TemporaryDirectory() as temp_dir, patch(
            "scripts.source_code_qa_scheduled_sync.Settings.from_env",
            return_value=SimpleNamespace(),
        ), patch(
            "scripts.source_code_qa_scheduled_sync.source_code_qa_data_root",
            return_value=Path(temp_dir),
        ), patch(
            "scripts.source_code_qa_scheduled_sync.build_source_code_qa_service_from_settings",
            return_value=FakeService(),
        ):
            first = run_scheduled_sync(now=datetime.fromisoformat("2026-05-23T10:00:00+08:00"))
            second = run_scheduled_sync(now=datetime.fromisoformat("2026-05-23T11:00:00+08:00"))

        self.assertEqual(first["status"], "ok")
        self.assertEqual(calls, [("AF", "All"), ("GRC", "All")])
        self.assertEqual(second["status"], "skipped")
        self.assertIn("already completed", second["reason"])

    def test_sync_api_runs_as_background_job(self):
        sync_result = {
            "status": "ok",
            "key": "AF:All",
            "results": [{"state": "ok", "display_name": "Repo One"}],
            "repo_status": [{"display_name": "Repo One", "state": "synced", "message": "ready"}],
        }
        with patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.sync", return_value=sync_result):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post("/api/source-code-qa/sync", json={"pm_team": "AF", "country": "All"})
                self.assertEqual(response.status_code, 200)
                payload = response.get_json()
                self.assertEqual(payload["status"], "queued")
                job_id = payload["job_id"]
                snapshot = {}
                for _ in range(20):
                    job_response = client.get(f"/api/jobs/{job_id}")
                    snapshot = job_response.get_json()
                    if snapshot.get("state") == "completed":
                        break
                    time.sleep(0.05)

        self.assertEqual(snapshot.get("state"), "completed")
        self.assertEqual(snapshot["results"][0]["repo_status"][0]["state"], "synced")

    def test_query_api_async_reports_real_backend_progress(self):
        captured = {}
        selected = []

        def fake_query(**kwargs):
            captured.update(kwargs)
            kwargs["progress_callback"]("direct_search", "Searching direct matches in Repo One.", 1, 2)
            return {"status": "ok", "answer_mode": "retrieval_only", "summary": "done", "matches": []}

        original = SourceCodeQAService.with_llm_provider

        def fake_with_provider(service, provider):
            selected.append(provider)
            return original(service, provider)

        with patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.with_llm_provider",
            new=fake_with_provider,
        ), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today",
            return_value={"attempted": False, "status": "fresh"},
        ) as ensure_synced, patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
            side_effect=fake_query,
        ):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                response = client.post(
                    "/api/source-code-qa/query",
                    json={
                        "pm_team": "AF",
                        "country": "All",
                        "question": "where is createIssue",
                        "answer_mode": "auto",
                        "llm_provider": "codex_cli_bridge",
                        "async": True,
                    },
                )
                self.assertEqual(response.status_code, 200)
                payload = response.get_json()
                self.assertEqual(payload["status"], "queued")
                snapshot = {}
                for _ in range(20):
                    job_response = client.get(f"/api/jobs/{payload['job_id']}")
                    snapshot = job_response.get_json()
                    if snapshot.get("state") == "completed":
                        break
                    time.sleep(0.05)
                source_job_response = client.get(f"/api/source-code-qa/query-jobs/{payload['job_id']}")
                source_job_payload = source_job_response.get_json()
                events_response = client.get(f"/api/source-code-qa/query-jobs/{payload['job_id']}/events")

        self.assertEqual(snapshot.get("state"), "completed")
        self.assertEqual(snapshot["results"][0]["summary"], "done")
        self.assertEqual(source_job_response.status_code, 200)
        self.assertEqual(source_job_payload["state"], "completed")
        self.assertEqual(source_job_payload["progress"]["stage"], "completed")
        self.assertIn(b"event: completed", events_response.data)
        self.assertIn(b'"state": "completed"', events_response.data)
        self.assertIn("progress_callback", captured)
        self.assertIn("codex_cli_bridge", selected)
        ensure_synced.assert_not_called()

    def test_query_api_async_returns_before_slow_backend_step_and_exposes_timing(self):
        backend_entered = threading.Event()
        release_backend = threading.Event()

        def slow_query(**kwargs):
            kwargs["progress_callback"]("retrieval", "Slow fake retrieval started.", 1, 3)
            backend_entered.set()
            release_backend.wait(timeout=2)
            return {
                "status": "ok",
                "answer_mode": "retrieval_only",
                "summary": "slow retrieval finished",
                "matches": [],
                "trace_id": "trace-slow-step",
                "retrieval_latency_ms": 1500,
                "llm_latency_ms": 25,
                "timing": {"slow_component": "retrieval", "retrieval_ms": 1500, "llm_ms": 25},
            }

        with patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.ensure_synced_today",
            return_value={"attempted": False, "status": "fresh"},
        ), patch(
            "bpmis_jira_tool.source_code_qa.SourceCodeQAService.query",
            side_effect=slow_query,
        ):
            with self.app.test_client() as client:
                self._login(client, "xiaodong.zheng@npt.sg")
                created = client.post(
                    "/api/source-code-qa/sessions",
                    json={"pm_team": "AF", "country": "All", "llm_provider": "codex_cli_bridge"},
                )
                session_id = created.get_json()["session"]["id"]
                started = time.monotonic()
                response = client.post(
                    "/api/source-code-qa/query",
                    json={
                        "session_id": session_id,
                        "pm_team": "AF",
                        "country": "All",
                        "question": "which step is slow",
                        "answer_mode": "auto",
                        "llm_provider": "codex_cli_bridge",
                        "async": True,
                    },
                )
                elapsed = time.monotonic() - started
                payload = response.get_json()
                self.assertEqual(response.status_code, 200)
                self.assertEqual(payload["status"], "queued")
                self.assertLess(elapsed, 0.25)
                self.assertTrue(backend_entered.wait(timeout=1))
                running_snapshot = client.get(f"/api/source-code-qa/query-jobs/{payload['job_id']}").get_json()
                release_backend.set()
                completed_snapshot = {}
                for _ in range(40):
                    completed_snapshot = client.get(f"/api/source-code-qa/query-jobs/{payload['job_id']}").get_json()
                    if completed_snapshot.get("state") == "completed":
                        break
                    time.sleep(0.05)

        self.assertEqual(running_snapshot["progress"]["stage"], "retrieval")
        self.assertEqual(completed_snapshot["state"], "completed")
        result = completed_snapshot["results"][0]
        self.assertEqual(result["timing"]["slow_component"], "retrieval")
        self.assertEqual(result["retrieval_latency_ms"], 1500)
        self.assertEqual(result["llm_latency_ms"], 25)

    def test_job_store_persists_background_job_snapshots(self):
        path = Path(self.temp_dir.name) / "run" / "jobs.json"
        first_store = JobStore(path)
        job = first_store.create("source-code-qa-sync", "Sync Source Code Repositories")
        first_store.update(job.job_id, state="running", message="Indexing.")
        first_store.complete(job.job_id, results=[{"status": "ok"}], notice={"summary": "done"})

        second_store = JobStore(path)
        snapshot = second_store.snapshot(job.job_id)

        self.assertIsNotNone(snapshot)
        self.assertEqual(snapshot["state"], "completed")
        self.assertEqual(snapshot["results"][0]["status"], "ok")

    def test_job_store_refreshes_snapshots_written_by_another_worker(self):
        path = Path(self.temp_dir.name) / "run" / "jobs.json"
        first_store = JobStore(path)
        second_store = JobStore(path)
        job = first_store.create("source-code-qa-sync", "Sync Source Code Repositories")
        first_store.update(job.job_id, state="running", message="Indexing.")
        first_store.complete(job.job_id, results=[{"status": "ok"}], notice={"summary": "done"})

        snapshot = second_store.snapshot(job.job_id)

        self.assertIsNotNone(snapshot)
        self.assertEqual(snapshot["state"], "completed")
        self.assertEqual(snapshot["results"][0]["status"], "ok")

    def test_job_store_marks_loaded_running_jobs_interrupted(self):
        path = Path(self.temp_dir.name) / "run" / "jobs.json"
        first_store = JobStore(path)
        job = first_store.create("source-code-qa-query", "Answer Source Code Question")
        first_store.update(job.job_id, state="running", stage="codex_stream", message="Calling Codex.")

        second_store = JobStore(path)
        snapshot = second_store.snapshot(job.job_id)

        self.assertIsNotNone(snapshot)
        self.assertEqual(snapshot["state"], "failed")
        self.assertEqual(snapshot["stage"], "failed")
        self.assertIn("interrupted by a server restart", snapshot["error"])
        self.assertEqual(snapshot["error_category"], "server_restart")
        self.assertEqual(snapshot["error_code"], "job_interrupted")
        self.assertTrue(snapshot["error_retryable"])
        self.assertGreater(snapshot["completed_at"], 0)

    def test_job_store_completed_job_clears_stale_error_fields(self):
        path = Path(self.temp_dir.name) / "run" / "jobs.json"
        store = JobStore(path)
        job = store.create("source-code-qa-query", "Answer Source Code Question")
        store.fail(
            job.job_id,
            "Temporary failure.",
            error_category="codex_timeout_or_rate_limit",
            error_code="llm_timeout",
            error_retryable=True,
        )

        store.update(job.job_id, state="running", stage="retry", message="Retrying.")
        store.complete(job.job_id, results=[{"status": "ok"}], notice={"summary": "done"})
        snapshot = store.snapshot(job.job_id)

        self.assertIsNotNone(snapshot)
        self.assertEqual(snapshot["state"], "completed")
        self.assertIsNone(snapshot["error"])
        self.assertEqual(snapshot["error_category"], "")
        self.assertEqual(snapshot["error_code"], "")
        self.assertFalse(snapshot["error_retryable"])

    def test_source_code_qa_scheduler_fairly_rotates_between_users(self):
        path = Path(self.temp_dir.name) / "run" / "jobs.json"
        store = JobStore(path)
        jobs = [store.create("source-code-qa-query", f"job-{index}") for index in range(4)]
        release_events = {job.job_id: threading.Event() for job in jobs}
        started: list[str] = []

        def fake_run(_app, job_id, _payload):
            started.append(job_id)
            store.update(job_id, state="running", stage="test", message="running")
            release_events[job_id].wait(timeout=2)
            store.complete(job_id, results=[{"status": "ok"}], notice={"summary": "done"})

        scheduler = SourceCodeQAQueryScheduler(job_store=store, max_running=2, default_runner=fake_run)
        scheduler.submit(app=self.app, job_id=jobs[0].job_id, payload={}, owner_email="a@npt.sg")
        scheduler.submit(app=self.app, job_id=jobs[1].job_id, payload={}, owner_email="a@npt.sg")
        scheduler.submit(app=self.app, job_id=jobs[2].job_id, payload={}, owner_email="a@npt.sg")
        scheduler.submit(app=self.app, job_id=jobs[3].job_id, payload={}, owner_email="b@npt.sg")
        self.assertEqual(started[:2], [jobs[0].job_id, jobs[1].job_id])
        release_events[jobs[0].job_id].set()
        for _ in range(20):
            if jobs[3].job_id in started:
                break
            time.sleep(0.05)
        for event in release_events.values():
            event.set()
        for _ in range(20):
            snapshots = [store.snapshot(job.job_id) for job in jobs]
            if all(snapshot and snapshot.get("state") == "completed" for snapshot in snapshots):
                break
            time.sleep(0.05)

        self.assertIn(jobs[3].job_id, started[:3])
        self.assertNotEqual(started[:3], [jobs[0].job_id, jobs[1].job_id, jobs[2].job_id])

    def test_source_code_qa_scheduler_persists_unhandled_worker_exception(self):
        path = Path(self.temp_dir.name) / "run" / "jobs.json"
        store = JobStore(path)
        job = store.create("source-code-qa-query", "job")
        scheduler = SourceCodeQAQueryScheduler(job_store=store, max_running=1)
        scheduler._running.add(job.job_id)
        scheduler._running_users["owner@npt.sg"] = 1

        def broken_runner(_app, _job_id, _payload):
            store.update(job.job_id, state="running", stage="test", message="running")
            raise RuntimeError("boom")

        scheduler._run_job(self.app, job.job_id, {}, "owner@npt.sg", broken_runner)

        snapshot = store.snapshot(job.job_id)
        self.assertIsNotNone(snapshot)
        self.assertEqual(snapshot["state"], "failed")
        self.assertEqual(snapshot["error_category"], "unexpected_internal")
        self.assertEqual(snapshot["error_code"], "background_worker_unhandled_exception")
        self.assertIn("boom", snapshot["error"])
        self.assertEqual(scheduler._running, set())
        self.assertEqual(scheduler._running_users, {})

    def test_feedback_api_saves_user_signal(self):
        with self.app.test_client() as client:
            self._login(client, "xiaodong.zheng@npt.sg")
            response = client.post(
                "/api/source-code-qa/feedback",
                json={
                    "rating": "incorrect",
                    "reason": "opposite_logic",
                    "pm_team": "AF",
                    "country": "All",
                    "question": "where is createIssue",
                    "top_paths": ["controller/IssueController.java"],
                    "trace_id": "trace-123",
                    "replay_context": {
                        "trace_id": "trace-123",
                        "answer_mode": "auto",
                        "llm_model": "codex-cli",
                        "rendered_answer": "createIssue is handled by IssueController.",
                        "answer_contract": {"status": "satisfied"},
                        "evidence_pack": {"items": [{"type": "entry_point", "claim": "IssueController"}]},
                        "matches_snapshot": [
                            {
                                "repo": "Portal Repo",
                                "path": "controller/IssueController.java",
                                "line_start": 10,
                                "line_end": 20,
                                "snippet": "public Issue createIssue() {}",
                            }
                        ],
                    },
                },
            )

        self.assertEqual(response.status_code, 200)
        feedback_path = Path(self.temp_dir.name) / "source_code_qa" / "feedback.jsonl"
        self.assertTrue(feedback_path.exists())
        feedback = json.loads(feedback_path.read_text(encoding="utf-8").strip())
        self.assertEqual(feedback["rating"], "incorrect")
        self.assertEqual(feedback["reason"], "opposite_logic")
        self.assertEqual(feedback["trace_id"], "trace-123")
        self.assertEqual(feedback["replay_context"]["answer_contract"]["status"], "satisfied")
        self.assertEqual(feedback["replay_context"]["matches_snapshot"][0]["path"], "controller/IssueController.java")

    def test_feedback_records_can_be_promoted_to_eval_candidates(self):
        candidates = build_eval_candidates(
            [
                {
                    "rating": "needs_deeper_trace",
                    "pm_team": "AF",
                    "country": "All",
                    "question_preview": "where does createIssue load data from",
                    "question_sha1": "abc123456789",
                    "trace_id": "trace-abc",
                    "top_paths": ["controller/IssueController.java", "repository/IssueRepository.java"],
                    "comment": "missed repository",
                    "replay_context": {
                        "trace_id": "trace-abc",
                        "answer_mode": "auto",
                        "llm_model": "codex-cli",
                        "answer_contract": {"status": "satisfied"},
                        "evidence_pack": {"items": [{"type": "entry_point", "claim": "IssueController"}]},
                        "matches_snapshot": [
                            {"path": "controller/IssueController.java"},
                            {"path": "repository/IssueRepository.java"},
                        ],
                    },
                },
                {
                    "rating": "useful",
                    "pm_team": "AF",
                    "country": "All",
                    "question_preview": "where is createIssue",
                },
            ]
        )

        self.assertEqual(len(candidates), 1)
        self.assertEqual(candidates[0]["id"], "feedback-needs_deeper_trace-abc1234567")
        self.assertEqual(candidates[0]["expected_paths"], [])
        self.assertEqual(candidates[0]["observed_paths"], ["controller/IssueController.java", "repository/IssueRepository.java"])
        self.assertEqual(candidates[0]["draft_status"], "needs_human_expected_evidence")
        self.assertEqual(candidates[0]["review_context"]["trace_id"], "trace-abc")

    def test_useful_feedback_can_become_positive_smoke_eval(self):
        candidates = build_eval_candidates(
            [
                {
                    "rating": "useful",
                    "pm_team": "AF",
                    "country": "All",
                    "question_preview": "where is createIssue",
                    "question_sha1": "def123456789",
                    "replay_context": {
                        "matches_snapshot": [{"path": "controller/IssueController.java"}],
                    },
                }
            ],
            include_useful=True,
        )

        self.assertEqual(len(candidates), 1)
        self.assertEqual(candidates[0]["expected_paths"], ["controller/IssueController.java"])
        self.assertEqual(candidates[0]["draft_status"], "ready_positive_smoke")

    def test_reviewed_feedback_candidates_promote_to_real_eval_cases(self):
        merged, summary = promote_candidates(
            [
                {
                    "id": "feedback-needs_deeper_trace-abc1234567",
                    "pm_team": "AF",
                    "country": "All",
                    "question": "where does createIssue load data from",
                    "answer_mode": "retrieval_only",
                    "draft_status": "approved",
                    "expected_paths": ["repository/IssueRepository.java"],
                    "observed_paths": ["controller/IssueController.java"],
                    "review_context": {"trace_id": "trace-abc"},
                },
                {
                    "id": "feedback-too_vague-def1234567",
                    "pm_team": "AF",
                    "country": "All",
                    "question": "where is createIssue",
                    "draft_status": "needs_human_expected_evidence",
                    "expected_paths": [],
                },
            ],
            [],
        )

        self.assertEqual(summary["promoted"], 1)
        self.assertEqual(summary["rejected"], 1)
        self.assertEqual(merged[0]["expected_paths"], ["repository/IssueRepository.java"])
        self.assertNotIn("review_context", merged[0])
        self.assertNotIn("observed_paths", merged[0])

    def test_positive_smoke_candidates_can_be_promoted_when_allowed(self):
        merged, summary = promote_candidates(
            [
                {
                    "id": "feedback-useful-def1234567",
                    "pm_team": "AF",
                    "country": "All",
                    "question": "where is createIssue",
                    "draft_status": "ready_positive_smoke",
                    "expected_paths": ["controller/IssueController.java"],
                }
            ],
            [],
            allow_positive_smoke=True,
        )

        self.assertEqual(summary["promoted"], 1)
        self.assertEqual(merged[0]["id"], "feedback-useful-def1234567")

    def test_auto_eval_candidates_collect_telemetry_feedback_and_dedupe(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            data_root = Path(temp_dir)
            source_root = data_root / "source_code_qa"
            source_root.mkdir(parents=True)
            (source_root / "feedback.jsonl").write_text(
                "\n".join(
                    [
                        json.dumps(
                            {
                                "rating": "useful",
                                "pm_team": "AF",
                                "country": "All",
                                "question_preview": "where is createIssue",
                                "question_sha1": "useful123456",
                                "replay_context": {"matches_snapshot": [{"path": "controller/IssueController.java"}]},
                            }
                        ),
                        json.dumps(
                            {
                                "rating": "wrong_file",
                                "pm_team": "AF",
                                "country": "All",
                                "question_preview": "where is broken answer",
                                "question_sha1": "wrong123456",
                                "replay_context": {"matches_snapshot": [{"path": "wrong/File.java"}]},
                            }
                        ),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            (source_root / "telemetry.jsonl").write_text(
                "\n".join(
                    [
                        json.dumps(
                            {
                                "key": "GRC:All",
                                "question_preview": "where is globallock config",
                                "question_sha1": "slow111",
                                "status": "ok",
                                "answer_mode": "retrieval_only",
                                "latency_ms": 45000,
                                "top_paths": ["config/GlobalLockConfig.java"],
                                "slow_query_attribution": {"status": "slow"},
                            }
                        ),
                        json.dumps(
                            {
                                "key": "CRMS:SG",
                                "question_preview": "which SG API runs precheck",
                                "question_sha1": "fallback222",
                                "status": "ok",
                                "answer_mode": "auto",
                                "deadline_hit": True,
                                "fallback_used": True,
                                "top_paths": ["controller/SgCreditApplicationController.java"],
                            }
                        ),
                        json.dumps(
                            {
                                "key": "CRMS:PH",
                                "question_preview": "where is missing PH flow",
                                "question_sha1": "nomatch333",
                                "status": "no_match",
                                "top_paths": [],
                            }
                        ),
                        json.dumps(
                            {
                                "key": "GRC:All",
                                "question_preview": "where is globallock config",
                                "question_sha1": "slow111",
                                "status": "ok",
                                "answer_mode": "retrieval_only",
                                "latency_ms": 46000,
                                "top_paths": ["config/GlobalLockConfig.java"],
                                "slow_query_attribution": {"status": "slow"},
                            }
                        ),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            candidates, summary = build_auto_eval_candidates(data_root, limit=20)
            runnable_candidates, runnable_summary = build_auto_eval_candidates(data_root, limit=20, runnable_only=True)

        categories = {candidate["category"] for candidate in candidates}
        self.assertIn("feedback_useful", categories)
        self.assertIn("feedback_wrong_file", categories)
        self.assertIn("telemetry_slow_query", categories)
        self.assertIn("telemetry_deadline_fallback", categories)
        self.assertIn("telemetry_no_match", categories)
        self.assertEqual(summary["review_only_candidates"], 1)
        self.assertGreaterEqual(runnable_summary["runnable_candidates"], 4)
        self.assertNotIn("feedback_wrong_file", {candidate["category"] for candidate in runnable_candidates})
        self.assertEqual(
            len({(candidate["pm_team"], candidate["country"], candidate["question"], candidate["answer_mode"]) for candidate in runnable_candidates}),
            len(runnable_candidates),
        )


class SourceCodeQAServiceTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self._env_patch = patch.dict(
            os.environ,
            {
                "TEAM_PORTAL_DATA_DIR": self.temp_dir.name,
                "LLM_CALL_LEDGER_PATH": str(Path(self.temp_dir.name) / "llm_call_ledger.jsonl"),
            },
            clear=False,
        )
        self._env_patch.start()
        self.service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )

    def tearDown(self):
        self._env_patch.stop()
        self.temp_dir.cleanup()

    def _ensure_runtime_helper_app(self):
        if hasattr(self, "app"):
            return self.app
        with patch.dict(
            os.environ,
            {
                "FLASK_SECRET_KEY": "test-secret",
                "TEAM_PORTAL_DATA_DIR": self.temp_dir.name,
                "TEAM_ALLOWED_EMAIL_DOMAINS": "npt.sg",
                "SOURCE_CODE_QA_GITLAB_TOKEN": "secret-token",
            },
            clear=False,
        ):
            self.app = create_app()
        self.app.testing = True
        return self.app

    def test_service_uses_internal_responsibility_components(self):
        from bpmis_jira_tool.source_code_qa_components import (
            SourceCodeQAAnswerGenerationComponent,
            SourceCodeQAIndexingSyncComponent,
            SourceCodeQAQualityJudgeComponent,
            SourceCodeQARetrievalComponent,
        )

        self.assertIsInstance(self.service._retrieval, SourceCodeQARetrievalComponent)
        self.assertIsInstance(self.service._answer_generation, SourceCodeQAAnswerGenerationComponent)
        self.assertIsInstance(self.service._quality_judge, SourceCodeQAQualityJudgeComponent)
        self.assertIsInstance(self.service._indexing_sync, SourceCodeQAIndexingSyncComponent)
        self.assertIs(self.service._retrieval._service, self.service)
        self.assertIs(self.service._answer_generation._service, self.service)
        self.assertIs(self.service._quality_judge._service, self.service)
        self.assertIs(self.service._indexing_sync._service, self.service)
        self.assertTrue(callable(self.service._rank_and_expand_query_matches))
        self.assertTrue(callable(self.service._build_llm_answer))
        self.assertTrue(callable(self.service._run_answer_judge))
        self.assertTrue(callable(self.service.sync))
        self.assertTrue(callable(self.service.ensure_synced_today))
        self.assertTrue(callable(self.service.index_health_payload))

    def test_with_codex_timeout_seconds_keeps_original_service_unchanged(self):
        tuned = self.service.with_codex_timeout_seconds(600)

        self.assertIsNot(tuned, self.service)
        self.assertEqual(self.service.codex_timeout_seconds, 360)
        self.assertEqual(tuned.codex_timeout_seconds, 600)
        self.assertEqual(tuned.base_data_root, self.service.base_data_root)

    def test_service_misc_boundary_helpers_cover_source_code_qa_edges(self):
        from bpmis_jira_tool import source_code_qa

        source_code_qa._log_source_code_qa_timing(
            "unit",
            elapsed_ms=-1,
            skipped=None,
            tags=("a", "b"),
            details={"ok": True, "nested": []},
            marker=object(),
        )
        self.assertIs(self.service.with_llm_provider(self.service.llm_provider_name), self.service)
        with patch.object(source_code_qa, "LLM_PROVIDER_ALLOWED_QUERY_CHOICES", {self.service.llm_provider_name, "other"}):
            self.assertIsNot(self.service.with_llm_provider("other"), self.service)
        self.assertIn("codex login", self.service.llm_unavailable_message())
        self.assertIsInstance(self.service._build_llm_provider(), CodexCliBridgeSourceCodeQALLMProvider)

        self.service.model_policy = {
            "answer": {"model": "gpt-answer", "override": False},
            "repair": {"model": "gpt-repair", "override": True},
            "empty": {"model": "   ", "override": True},
        }
        self.assertEqual(self.service._model_for_role("answer"), "gpt-answer")
        self.assertTrue(self.service._model_for_role("missing", fallback="fallback-model"))
        self.assertEqual(self.service._model_for_role_or_budget("repair", {"model": "budget-model"}), "gpt-repair")
        self.assertEqual(self.service._model_for_role_or_budget("empty", {"model": "budget-model"}), "budget-model")
        self.assertEqual(
            self.service._normalize_llm_usage({"promptTokenCount": "2", "candidatesTokenCount": "3"})["total_tokens"],
            5,
        )
        self.assertIsNone(
            self.service._normalize_llm_usage({"prompt_tokens": "bad", "completion_tokens": 3}).get("total_tokens")
        )
        merged_usage = self.service._merge_llm_usage({"prompt_tokens": 2, "label": "a"}, {"prompt_tokens": "bad", "label": "b"})
        self.assertEqual(merged_usage["prompt_tokens"], "bad")
        self.assertEqual(merged_usage["label"], "a")
        self.assertEqual(self.service._llm_finish_reason({"candidates": [{"finishReason": "MAX_TOKENS"}]}), "MAX_TOKENS")
        self.assertEqual(self.service._llm_finish_reason({"choices": [{"finish_reason": "length"}]}), "length")
        self.assertEqual(self.service._llm_finish_reason({"finish_reason": "stop"}), "stop")
        self.assertEqual(self.service._llm_finish_reason({}), "")

        self.service.config_path.parent.mkdir(parents=True, exist_ok=True)
        self.service.config_path.write_text("{bad", encoding="utf-8")
        with self.assertRaises(ToolError):
            self.service.load_config()
        self.service.config_path.write_text(json.dumps({"mappings": []}), encoding="utf-8")
        self.assertEqual(self.service.load_config()["mappings"], {})

        self.service.domain_profile_path = Path(self.temp_dir.name) / "profiles.json"
        self.service.domain_knowledge_pack_path = Path(self.temp_dir.name) / "knowledge.json"
        self.service.domain_profile_path.write_text("{bad", encoding="utf-8")
        self.assertEqual(self.service.load_domain_profiles(), {"default": {}})
        self.service.domain_profile_path.write_text(json.dumps({"default": {"api_terms": ["Api"]}, "AF": {"api_terms": ["RiskApi"]}}), encoding="utf-8")
        self.service.domain_knowledge_pack_path.write_text("[]", encoding="utf-8")
        self.assertEqual(self.service.load_domain_knowledge_packs(), {"version": 1, "domains": {}})

        pack = {
            "label": "Anti Fraud",
            "summary": "Risk decisions",
            "module_map": [
                "bad",
                {"name": "Decision Engine", "aliases": ["risk decision"], "code_hints": ["DecisionService"], "business_flows": ["review"]},
            ],
            "terminology": ["bad", {"term": "Rule", "aliases": ["policy"], "code_terms": ["RuleService"]}],
            "key_artifacts": {
                "apis": ["bad", {"name": "Decision API", "path": "/api/decision", "purpose": "entry"}],
                "configs": [{"path": "application.yml"}],
                "tables": [{"name": "", "path": "", "purpose": "skip"}],
            },
            "question_seeds": ["plain question", {"question": "How decision works?", "expected_terms": ["DecisionService"]}],
            "evidence_rules": ["Prefer runtime evidence"],
            "retrieval_terms": {"api_terms": ["DecisionClient"]},
        }
        self.service.domain_knowledge_pack_path.write_text(json.dumps({"domains": {"AF": pack}}), encoding="utf-8")
        self.assertIn("DecisionService", self.service._domain_knowledge_terms(pack))
        self.assertIn("DecisionClient", self.service._domain_profile("AF", "SG")["api_terms"])
        domain_payload = self.service.domain_knowledge_payload()
        self.assertEqual(domain_payload["domains"]["AF"]["module_count"], 2)
        self.assertEqual(self.service._domain_knowledge_pack("AF")["label"], "Anti Fraud")
        context_text = self.service._llm_domain_context(
            pm_team="AF",
            country="SG",
            question="risk decision api config",
            evidence_summary={"intent": {"api": True, "config": True}},
        )
        self.assertIn("Domain guidance", context_text)
        self.assertEqual(self.service._matched_domain_modules({"module_map": ["bad"]}, "", {}), [])
        self.assertTrue(self.service._domain_artifact_lines(pack, {"data_source": True, "module_dependency": True}))
        self.assertTrue(self.service._domain_artifact_lines(pack, {"rule_logic": True, "impact_analysis": True}))
        self.assertTrue(self.service._llm_answer_blueprint({"api": True}))
        self.assertTrue(self.service._llm_answer_blueprint({"config": True}))
        self.assertTrue(self.service._llm_answer_blueprint({"static_qa": True}))
        self.assertTrue(self.service._llm_answer_blueprint({"impact_analysis": True}))
        self.assertTrue(self.service._llm_answer_blueprint({"test_coverage": True}))
        self.assertIn("CreditReviewCbsService", self.service._question_specific_retrieval_terms("credit review cbs report"))
        self.assertEqual(self.service._question_specific_retrieval_terms("payslip only"), [])
        self.assertIn("ExtractRecord", self.service._question_specific_retrieval_terms("payslip extracted fields stored"))

        previous_repo = RepositoryEntry("Risk Engine", "https://git.example.com/af/risk-engine.git")
        other_repo = RepositoryEntry("Other Service", "https://git.example.com/af/other-service.git")
        conversation_context = {
            "key": self.service.mapping_key("AF", "SG"),
            "question": "How does DecisionService work?",
            "answer": "It uses DecisionService and risk_table.",
            "repo_scope": ["Risk Engine"],
            "matches": [{"repo": "Risk Engine", "path": "src/DecisionService.java", "reason": "risk decision"}],
            "recent_turns": [
                "bad",
                {
                    "question": "previous DecisionService",
                    "answer": "answer RiskPolicy",
                    "matches_snapshot": [{"path": "src/RiskPolicy.java", "reason": "policy reason"}],
                    "evidence_pack": {"confirmed_facts": ["risk_table"]},
                    "codex_candidate_paths": [{"path": "src/CodexPath.java", "reason": "candidate"}],
                },
            ],
            "codex_candidate_paths": [{"path": "src/DecisionClient.java", "reason": "client"}],
            "codex_citation_validation": {"direct_file_refs": [{"path": "src/DirectRef.java"}]},
            "evidence_pack": {"confirmed_facts": ["decision_table"]},
            "answer_contract": {"confirmed_sources": ["DecisionSource"]},
        }
        augmented, followup = self.service._apply_conversation_context(
            "What about this method?",
            conversation_context,
            current_key=self.service.mapping_key("AF", "SG"),
            current_repositories=[previous_repo],
        )
        self.assertIn("Previous Source Code Q&A context terms", augmented)
        self.assertTrue(followup["used"])
        self.assertEqual(self.service._apply_conversation_context("q", None)[1], {"used": False})
        self.assertEqual(
            self.service._apply_conversation_context("q", {"key": "CRMS:SG"}, current_key=self.service.mapping_key("AF", "SG"))[1]["reason"],
            "scope_mismatch",
        )
        self.assertEqual(
            self.service._conversation_repo_scope("ask other service", conversation_context, [previous_repo, other_repo])["mismatch"],
            True,
        )
        self.assertFalse(self.service._conversation_repo_scope("", {}, [previous_repo])["mismatch"])
        self.assertFalse(self.service._conversation_repo_scope("risk engine", {"repo_scope": ["Risk Engine"]}, [previous_repo])["mismatch"])
        self.assertEqual(self.service._mentioned_repository_aliases("", [previous_repo]), set())
        selected, scope = self.service._filter_entries_for_question_repository_scope("risk engine details", [previous_repo, other_repo])
        self.assertEqual([entry.display_name for entry in selected], ["Risk Engine"])
        self.assertTrue(scope["active"])
        all_selected, all_scope = self.service._filter_entries_for_question_repository_scope("", [previous_repo])
        self.assertFalse(all_scope["active"])
        self.assertEqual([entry.display_name for entry in all_selected], ["Risk Engine"])
        self.assertFalse(self.service._repo_scope_alias_is_useful(""))
        self.assertFalse(self.service._repo_scope_alias_is_useful("repo"))
        self.assertTrue(self.service._repo_scope_alias_is_specific("risk engine"))
        self.assertFalse(self.service._repo_scope_alias_is_specific(""))
        self.assertFalse(self.service._repo_alias_in_text("", "risk engine"))
        self.assertIn("Decision", self.service._conversation_title_terms("S1 Decision Decision"))

    def test_service_sync_feedback_and_progress_boundaries(self):
        from bpmis_jira_tool import source_code_qa

        key = self.service.mapping_key("AF", "SG")
        entry_payload = {"display_name": "Repo", "url": "https://git.example.com/team/repo.git"}
        entry = RepositoryEntry("Repo", "https://git.example.com/team/repo.git")

        self.assertIs(self.service.with_codex_timeout_seconds(self.service.codex_timeout_seconds), self.service)
        self.assertTrue(self.service._llm_default_model())
        self.service.domain_profile_path = Path(self.temp_dir.name) / "missing_profiles.json"
        self.service.domain_knowledge_pack_path = Path(self.temp_dir.name) / "missing_knowledge.json"
        self.assertEqual(self.service.load_domain_profiles(), {"default": {}})
        self.assertEqual(self.service.load_domain_knowledge_packs(), {"version": 1, "domains": {}})
        self.service.domain_knowledge_pack_path.write_text("{bad", encoding="utf-8")
        self.assertEqual(self.service.load_domain_knowledge_packs(), {"version": 1, "domains": {}})
        self.assertTrue(
            self.service._apply_conversation_context(
                "q",
                {"pm_team": "UNKNOWN", "country": "SG"},
                current_key=key,
            )[1]["used"]
        )
        self.assertEqual(self.service._apply_conversation_context("this method", {}, current_key=None)[1], {"used": False})
        self.assertEqual(self.service._conversation_title_terms("S1234 Decision Decision"), ["Decision"])
        self.assertTrue(
            self.service._conversation_repo_scope(
                "risk engine",
                {},
                [RepositoryEntry("Risk Engine", "https://git.example.com/af/risk-engine.git")],
            )["mismatch"]
        )

        with self.assertRaises(ToolError):
            self.service.save_mapping(pm_team="AF", country="SG", repositories="bad")
        no_token_service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name) / "no-token",
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        with self.assertRaises(ToolError):
            no_token_service._sync_impl(pm_team="AF", country="SG")
        self.assertEqual(self.service._sync_impl(pm_team="AF", country="SG")["status"], "empty_config")

        self.service.save_mapping(pm_team="AF", country="SG", repositories=[entry_payload])
        finished_jobs: list[tuple[str, str, list[dict[str, object]]]] = []
        with patch.object(self.service, "_start_sync_job", return_value={"job_id": "job-1"}), patch.object(
            self.service,
            "_sync_entry",
            side_effect=RuntimeError("boom"),
        ), patch.object(
            self.service,
            "_finish_sync_job",
            side_effect=lambda _key, _job_id, *, status, results: finished_jobs.append((_key, status, results)),
        ):
            with self.assertRaises(RuntimeError):
                self.service._sync_impl(pm_team="AF", country="SG")
        self.assertEqual(finished_jobs, [(key, "failed", [])])

        with patch.dict(os.environ, {"SOURCE_CODE_QA_AUTO_SYNC_START_DATE": "bad", "SOURCE_CODE_QA_AUTO_SYNC_INTERVAL_DAYS": "bad"}):
            self.assertEqual(self.service._auto_sync_start_date(), source_code_qa.DEFAULT_AUTO_SYNC_START_DATE)
            self.assertEqual(self.service._auto_sync_interval_days(), source_code_qa.DEFAULT_AUTO_SYNC_INTERVAL_DAYS)
        with patch.dict(os.environ, {"SOURCE_CODE_QA_AUTO_SYNC_START_DATE": "2099-01-01", "SOURCE_CODE_QA_AUTO_SYNC_INTERVAL_DAYS": "3"}):
            self.assertEqual(self.service._auto_sync_start_date(), date(2099, 1, 1))
            self.assertEqual(self.service._auto_sync_interval_days(), 3)
            self.assertEqual(self.service._latest_completed_scheduled_sync_date(date(2098, 12, 31)), date(2099, 1, 1))

        empty_service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name) / "empty-auto",
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        self.assertEqual(empty_service._ensure_synced_today_impl(pm_team="AF", country="SG")["status"], "empty_config")
        skipped_service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name) / "skipped-auto",
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        skipped_service.save_mapping(pm_team="AF", country="SG", repositories=[entry_payload])
        with patch.object(skipped_service, "repo_status", return_value=[]), patch.object(
            skipped_service,
            "_index_freshness_payload",
            return_value={"status": "stale_or_missing"},
        ):
            self.assertEqual(skipped_service._ensure_synced_today_impl(pm_team="AF", country="SG")["status"], "skipped")
        with patch.object(self.service, "repo_status", return_value=[]), patch.object(
            self.service,
            "_index_freshness_payload",
            return_value={"status": "stale_or_missing"},
        ), patch.object(self.service, "sync", side_effect=ToolError("sync failed")):
            failed_auto_sync = self.service._ensure_synced_today_impl(pm_team="AF", country="SG")
        self.assertEqual(failed_auto_sync["status"], "failed")
        with patch.object(self.service, "repo_status", return_value=[]), patch.object(
            self.service,
            "_index_freshness_payload",
            return_value={"status": "fresh", "newest_indexed_at": datetime.now().astimezone().isoformat()},
        ):
            self.assertEqual(self.service._ensure_synced_today_impl(pm_team="AF", country="SG")["status"], "fresh")
        with patch.object(self.service, "repo_status", return_value=[]), patch.object(
            self.service,
            "_index_freshness_payload",
            return_value={"status": "fresh", "newest_indexed_at": "bad-date"},
        ), patch.object(self.service, "sync", return_value={"status": "ok"}):
            self.assertEqual(self.service._ensure_synced_today_impl(pm_team="AF", country="SG")["status"], "ok")

        self.assertEqual(self.service._normalize_answer_mode("unexpected"), "auto")
        self.assertFalse(self.service._query_deadline_hit(time.time(), reserve_seconds=0))
        self.service.query_deadline_seconds = 1
        self.assertTrue(self.service._query_deadline_hit(time.time() - 5, reserve_seconds=0))
        self.service.query_deadline_seconds = 0
        self.assertFalse(self.service._query_deadline_hit(time.time() - 5, reserve_seconds=0))
        self.service._report_query_progress(None, "stage", "message")
        self.service._report_query_progress(lambda *_args: (_ for _ in ()).throw(RuntimeError("callback")), "stage", "message")
        phase_cache: dict[str, object] = {}
        self.service._mark_query_phase(phase_cache, "forced", time.perf_counter(), force_log=True)
        self.assertIn("forced", phase_cache["timing"])

        slow = self.service._build_slow_query_attribution(
            {
                "query_timing": {"components": {"rank": "bad", "search": 10}},
                "llm_timing": {"answer": "bad"},
                "retrieval_latency_ms": 20,
                "llm_cached": True,
            },
            latency_ms=1,
        )
        self.assertEqual(slow["reason"], "cache_hit")
        deadline_slow = self.service._build_slow_query_attribution({"deadline_hit": True}, latency_ms=30_000)
        self.assertEqual(deadline_slow["reason"], "deadline_hit")

        with self.assertRaises(ToolError):
            self.service.query(pm_team="AF", country="SG", question="")
        with self.assertRaises(ToolError):
            self.service.normalize_country("CRMS", "TH")
        freshness = self.service._index_freshness_payload_impl(
            [{"display_name": "Repo", "index": {"state": "ready", "git_revision": "abc", "updated_at": "2026-01-01T00:00:00+00:00"}}]
        )
        self.assertEqual(freshness["git_revisions"], [{"repo": "Repo", "git_revision": "abc"}])
        with self.assertRaises(ToolError):
            self.service.save_feedback(user_email="u@example.com", payload={"rating": "bad", "question": "q"})
        with self.assertRaises(ToolError):
            self.service.save_feedback(user_email="u@example.com", payload={"rating": "useful", "reason": "bad", "question": "q"})
        feedback_payload = {
            "rating": "useful",
            "reason": "deprecated_class",
            "question": "How?",
            "matches": ["bad", {"repo": "Repo", "path": "src/App.java", "retrieval": "direct"}],
            "answer_quality": [],
            "extra": {"nested": {"value": object()}},
        }
        self.assertEqual(self.service.save_feedback(user_email="USER@EXAMPLE.COM", payload=feedback_payload)["status"], "ok")
        self.assertIsNone(self.service._trim_feedback_value("deep", depth=6))
        self.assertEqual(len(self.service._trim_feedback_value({str(i): i for i in range(90)}, dict_limit=80)), 80)
        self.assertTrue(str(self.service._trim_feedback_value(object())).startswith("<object object at"))

        with patch("bpmis_jira_tool.source_code_qa.subprocess.run", side_effect=source_code_qa.subprocess.TimeoutExpired(cmd=["git"], timeout=1)):
            self.assertIn("timed out", self.service._sync_entry_impl(key, entry)["message"])
        with patch("bpmis_jira_tool.source_code_qa.subprocess.run", side_effect=OSError("missing git")):
            self.assertIn("could not start", self.service._sync_entry_impl(key, entry)["message"])
        with patch(
            "bpmis_jira_tool.source_code_qa.subprocess.run",
            return_value=SimpleNamespace(returncode=1, stderr="https://alice:secret@git.example.com/repo.git", stdout=""),
        ):
            self.assertIn("Git clone failed", self.service._sync_entry_impl(key, entry)["message"])
        with patch(
            "bpmis_jira_tool.source_code_qa.subprocess.run",
            return_value=SimpleNamespace(returncode=0, stderr="", stdout=""),
        ), patch.object(self.service, "_build_repo_index", side_effect=ValueError("bad index")):
            self.assertIn("code index failed", self.service._sync_entry_impl(key, entry)["message"])
        pull_repo = self.service._repo_path(key, entry)
        (pull_repo / ".git").mkdir(parents=True, exist_ok=True)
        with patch(
            "bpmis_jira_tool.source_code_qa.subprocess.run",
            return_value=SimpleNamespace(returncode=0, stderr="", stdout=""),
        ), patch.object(self.service, "_build_repo_index", return_value={"state": "ready", "files": 2, "lines": 10, "definitions": 1, "references": 1}):
            self.assertEqual(self.service._sync_entry_impl(key, entry)["state"], "ok")
        clone_entry = RepositoryEntry("Clone Repo", "https://git.example.com/team/clone.git")
        clone_repo = self.service._repo_path(key, clone_entry)
        clone_repo.mkdir(parents=True)
        with patch.object(self.service, "_remove_incomplete_repo_dir") as remove_dir, patch(
            "bpmis_jira_tool.source_code_qa.subprocess.run",
            return_value=SimpleNamespace(returncode=0, stderr="", stdout=""),
        ), patch.object(self.service, "_build_repo_index", return_value={"state": "ready", "files": 1, "lines": 5, "definitions": 0, "references": 0}):
            self.assertEqual(self.service._sync_entry_impl(key, clone_entry)["state"], "ok")
        remove_dir.assert_called_once_with(clone_repo)

        self.service.sync_jobs_path.parent.mkdir(parents=True, exist_ok=True)
        self.service.sync_jobs_path.write_text("{bad", encoding="utf-8")
        self.assertEqual(self.service._sync_job_status_impl(key), {"status": "idle", "key": key})
        with patch.object(type(self.service.sync_jobs_path), "exists", side_effect=OSError("stat failed")):
            self.assertEqual(self.service._sync_job_status_impl(key), {"status": "idle", "key": key})
        self.service.sync_jobs_path.write_text("{}", encoding="utf-8")
        job = self.service._start_sync_job_impl(key, [entry])
        self.service._finish_sync_job_impl(key, "wrong-job", status="ok", results=[])
        self.service._finish_sync_job_impl(key, job["job_id"], status="ok", results=[{"state": "ok"}])
        self.assertEqual(self.service._sync_job_status_impl(key)["status"], "ok")
        self.service.sync_jobs_path.write_text("[]", encoding="utf-8")
        self.service._write_sync_job_impl(key, {"job_id": "list-payload"})
        self.assertEqual(self.service._sync_job_status_impl(key)["job_id"], "list-payload")
        lock_path = self.service.sync_jobs_path.with_suffix(".lock")
        lock_path.write_text("locked", encoding="utf-8")
        with patch("bpmis_jira_tool.source_code_qa.SYNC_JOB_LOCK_TIMEOUT_SECONDS", 0), patch(
            "bpmis_jira_tool.source_code_qa.time.monotonic",
            side_effect=[0, 1],
        ), patch("bpmis_jira_tool.source_code_qa.time.sleep", return_value=None):
            self.service._write_sync_job_impl(key, {"job_id": "locked"})
        lock_path.unlink(missing_ok=True)
        with patch("bpmis_jira_tool.source_code_qa.os.open", side_effect=[FileExistsError(), OSError("denied")]), patch(
            "bpmis_jira_tool.source_code_qa.time.monotonic",
            return_value=0,
        ), patch("bpmis_jira_tool.source_code_qa.time.sleep", return_value=None) as sleep:
            self.service._write_sync_job_impl(key, {"job_id": "sleep-then-denied"})
        sleep.assert_called()
        with patch("bpmis_jira_tool.source_code_qa.os.open", side_effect=OSError("denied")):
            self.service._write_sync_job_impl(key, {"job_id": "denied"})

    def test_service_ranking_evidence_and_agent_helper_boundaries(self):
        entry = RepositoryEntry("Repo", "https://git.example.com/team/repo.git")
        repo_path = self.service._repo_path("AF:All", entry)
        report_events: list[tuple[str, str, int, int]] = []

        def report(*args):
            report_events.append(args)

        self.assertEqual(self.service._dedupe_agent_plan_steps([{"name": ""}, {"name": "x", "terms": []}]), [])
        deduped_steps = self.service._dedupe_agent_plan_steps(
            [
                {"name": "x", "terms": ["Term", "Term", ""], "tools": ["search_code", "bad"]},
                {"name": "x", "terms": ["Other"], "tools": ["search_code"]},
            ]
        )
        self.assertEqual(deduped_steps[0]["tools"], ["search_code"])
        self.assertEqual(self.service._planner_suffix_terms(["abc", "Risk"], ("Service",)), ["RiskService"])

        base_match = {
            "repo": "Repo",
            "path": "src/main/java/RiskRepository.java",
            "line_start": 1,
            "line_end": 3,
            "score": 100,
            "snippet": "class RiskRepository { select * from risk_table; WebClient client; validate(permission); }",
            "reason": "bm25 content match",
            "retrieval": "direct",
            "trace_stage": "direct",
        }
        self.assertEqual(self.service._rank_matches("q", []), [])
        ranked_without_cache = self.service._rank_matches("risk table", [base_match])
        self.assertEqual(ranked_without_cache[0]["path"], base_match["path"])
        request_cache = self.service._new_retrieval_request_cache()
        score_one = self.service._rerank_score_cached("risk table", base_match, request_cache=request_cache)
        score_two = self.service._rerank_score_cached("risk table", base_match, request_cache=request_cache)
        self.assertEqual(score_one, score_two)
        self.assertGreaterEqual(request_cache["stats"]["rerank_hits"], 1)

        exact_match = {
            **base_match,
            "path": "src/main/java/RiskMapper.java",
            "snippet": "select * from risk_table",
            "retrieval": "exact_table_path_lookup",
            "trace_stage": "exact_lookup",
            "exact_lookup": {"term": "risk_table", "lookup_value": "risk_table"},
        }
        module_match = {
            **base_match,
            "path": "build.gradle",
            "snippet": "dependencies { implementation project(':risk') }",
            "retrieval": "direct",
            "trace_stage": "query_decomposition",
        }
        message_match = {
            **base_match,
            "path": "src/EventProducer.java",
            "snippet": "kafkaTemplate.send(topic, event)",
            "reason": "message_publish risk.topic",
        }
        test_match = {
            **base_match,
            "path": "src/test/java/RiskServiceTest.java",
            "retrieval": "test_coverage",
            "test_coverage": {"has_assertion": True},
        }
        boundary_match = {
            **base_match,
            "path": "src/RiskService.java",
            "retrieval": "operational_boundary",
            "operational_boundary": {"kind": "transactional"},
        }
        static_match = {
            **base_match,
            "retrieval": "static_qa",
            "static_qa": {"kind": "hardcoded_secret", "severity": "high", "reason": "secret"},
            "snippet": "String password = \"secret\";",
        }
        question_features = {
            "intent": {
                "data_source": True,
                "api": True,
                "config": True,
                "module_dependency": True,
                "message_flow": True,
                "rule_logic": True,
                "static_qa": True,
            },
            "tokens": {"riskmapper"},
            "specific_terms": {"risk_table", "risk"},
        }
        self.assertGreater(self.service._rerank_score("credit review cbs", exact_match, question_features=question_features), 0)
        self.assertGreater(self.service._rerank_score("gradle multi-module", module_match, question_features=question_features), 0)
        self.assertGreater(
            self.service._rerank_score(
                "focused",
                {**base_match, "trace_stage": "focused_search"},
                question_features={"intent": {}, "tokens": set(), "specific_terms": set()},
            ),
            base_match["score"],
        )
        self.assertLess(
            self.service._rerank_score(
                "credit review cbs",
                {**base_match, "path": "src/bulk/BulkJob.java", "snippet": "cbs"},
                question_features={"intent": {}, "tokens": set(), "specific_terms": set()},
            ),
            base_match["score"],
        )
        self.assertGreater(
            self.service._rerank_score(
                "credit review cbs",
                {**base_match, "path": "src/CreditReviewCbsService.java", "snippet": "cbs"},
                question_features={"intent": {}, "tokens": set(), "specific_terms": set()},
            ),
            base_match["score"],
        )
        self.assertGreater(
            self.service._rerank_score(
                "maven pom",
                {**module_match, "path": "pom.xml", "snippet": "<artifactId>risk</artifactId>"},
                question_features={"intent": {"module_dependency": True}, "tokens": set(), "specific_terms": set()},
            ),
            module_match["score"],
        )
        self.assertGreater(
            self.service._rerank_score(
                "rule",
                {**base_match, "snippet": "validate permission"},
                question_features={"intent": {"rule_logic": True}, "tokens": set(), "specific_terms": set()},
            ),
            base_match["score"],
        )
        self.assertGreater(self.service._rerank_score("message flow", message_match, question_features=question_features), 0)
        self.assertGreater(self.service._rerank_score("static qa", static_match, question_features=question_features), 0)
        self.assertLess(
            self.service._rerank_score("normal question", {**test_match, "retrieval": "direct"}, question_features={"intent": {}, "tokens": set(), "specific_terms": set()}),
            test_match["score"],
        )

        rich_matches = [exact_match, module_match, message_match, test_match, boundary_match, static_match]
        evidence_cache: dict[str, object] = {}
        evidence_question = "risk data source api config maven dependency message test coverage transactional static qa"
        evidence_summary = self.service._compress_evidence_cached(evidence_question, rich_matches, request_cache=evidence_cache)
        cached_summary = self.service._compress_evidence_cached(evidence_question, rich_matches, request_cache=evidence_cache)
        self.assertEqual(evidence_summary, cached_summary)
        self.assertTrue(evidence_summary["entry_points"])
        self.assertTrue(evidence_summary["data_sources"])
        self.assertTrue(evidence_summary["module_dependencies"])
        self.assertTrue(evidence_summary["message_flows"])
        self.assertTrue(evidence_summary["test_coverage"])
        self.assertTrue(evidence_summary["operational_boundaries"])
        self.assertTrue(evidence_summary["static_findings"])

        pack = self.service._build_evidence_pack(
            question="risk data source api config",
            evidence_summary={
                **evidence_summary,
                "entry_points": [""],
                "source_conflicts": ["conflict"],
                "impact_surfaces": ["upstream caller"],
                "field_population": ["field set"],
                "downstream_components": ["RiskClient"],
                "source_tiers": ["test:1"],
            },
            matches=rich_matches,
            trace_paths=[{}, {"edges": [{"to_name": "RiskController"}, {"edge_kind": "call"}]}],
            quality_gate={"missing": ["missing hop"]},
        )
        self.assertTrue(pack["items"])
        self.assertIn("missing hop", pack["missing_hops"])
        empty_pack = self.service._new_evidence_pack("q", {"intent": {"data_source": True}})
        empty_pack["items"] = ["bad", {"type": "table", "claim": ""}]
        self.service._classify_evidence_pack_items(empty_pack)
        self.assertTrue(empty_pack["evidence_limits"])
        self.assertEqual(self.service._fact_source_label("Repo:file.java:1-2: table"), "Repo:file.java:1-2")
        adder_target: list[str] = []
        adder = self.service._limited_fact_adder(adder_target, 1)
        adder("")
        adder("first")
        adder("second")
        self.assertEqual(adder_target, ["first"])

        self.assertEqual(self.service._evidence_source_tier("src/test/AppTest.java"), "test")
        self.assertEqual(self.service._evidence_source_tier("docs/readme.md"), "docs")
        self.assertEqual(self.service._evidence_source_tier("build/generated/App.java"), "generated")
        self.assertEqual(self.service._evidence_source_tier("src/main/resources/application.yml"), "config")
        self.assertEqual(self.service._interesting_lines("\nselect * from risk_table\nplain", ("missing",)), ["select * from risk_table"])
        self.assertEqual(self.service._static_qa_findings_for_line(""), [])
        self.assertFalse(self.service._is_concrete_source_line(""))
        self.assertFalse(self.service._is_concrete_source_line("import java.util.List;"))
        self.assertFalse(self.service._is_concrete_source_line("private RiskRepository riskRepository;"))
        self.assertTrue(self.service._is_concrete_source_line("select * from risk_table"))
        self.assertTrue(self.service._has_production_source_tier({"data_source_tiers": ["production:1"]}))
        self.assertTrue(self.service._has_production_source_tier({"data_sources": ["source"]}))
        self.assertEqual(self.service._best_snippet_window([], 1), (1, 1))
        long_lines = ["x"] * 60
        long_lines[14] = "class A {"
        self.assertEqual(self.service._best_snippet_window(long_lines, 30), (15, 39))
        self.assertIn("static QA", self.service._build_summary([static_match]))
        self.assertIn("test coverage", self.service._build_summary([test_match]))
        self.assertIn("operational boundary", self.service._build_summary([boundary_match]))
        self.assertEqual(len(self.service._build_citations([base_match, dict(base_match)])), 1)
        self.assertGreater(self.service._result_match_priority_sort_key(exact_match, intent={}, question="q")[0], 0)
        self.assertGreater(self.service._result_match_priority_sort_key(module_match, intent={"module_dependency": True}, question="npm package gradle maven pom")[0], 0)
        self.assertGreater(self.service._result_match_priority_sort_key({**module_match, "path": "pom.xml"}, intent={"module_dependency": True}, question="maven pom")[0], 0)
        self.assertGreater(self.service._result_match_priority_sort_key(message_match, intent={"message_flow": True}, question="message")[0], 0)
        self.assertGreater(self.service._result_match_priority_sort_key(test_match, intent={"test_coverage": True}, question="test")[0], 0)
        self.assertGreater(self.service._result_match_priority_sort_key(boundary_match, intent={"operational_boundary": True}, question="boundary")[0], 0)
        self.assertGreater(self.service._result_match_priority_sort_key({**base_match, "retrieval": "planner_caller"}, intent={"impact_analysis": True}, question="impact")[0], 0)

        with patch.object(self.service, "_compress_evidence_cached", return_value={"intent": {}}), patch.object(
            self.service,
            "_quality_gate_cached",
            return_value={"status": "sufficient", "confidence": "medium"},
        ):
            self.assertFalse(
                self.service._should_expand_query_matches(
                    question="q",
                    top_matches=[base_match],
                    exact_lookup_sufficient=False,
                    simple_quality_trace=True,
                    synced_entries=[(entry, repo_path)] * 5,
                    started_at=time.time(),
                    latency_guarded_query_expansion=False,
                    request_cache=self.service._new_retrieval_request_cache(),
                    report=report,
                )
            )
        self.assertFalse(
            self.service._should_expand_query_matches(
                question="q",
                top_matches=[base_match],
                exact_lookup_sufficient=False,
                simple_quality_trace=False,
                synced_entries=[(entry, repo_path)],
                started_at=time.time(),
                latency_guarded_query_expansion=True,
                request_cache=self.service._new_retrieval_request_cache(),
                report=report,
            )
        )
        self.assertFalse(
            self.service._should_expand_query_matches(
                question="q",
                top_matches=[base_match],
                exact_lookup_sufficient=False,
                simple_quality_trace=False,
                synced_entries=[(entry, repo_path)],
                started_at=time.time() - 9,
                latency_guarded_query_expansion=False,
                request_cache=self.service._new_retrieval_request_cache(),
                report=report,
            )
        )

        with patch.object(self.service, "_agent_step_terms", return_value=[]):
            self.assertEqual(
                self.service._run_agent_plan(
                    entries=[entry],
                    key="AF:All",
                    question="q",
                    matches=[base_match],
                    evidence_summary={},
                    quality_gate={},
                    agent_plan={"steps": [{"name": "empty", "tools": ["search_code"]}]},
                    limit=3,
                    tool_trace=[],
                    request_cache=self.service._new_retrieval_request_cache(),
                ),
                [base_match],
            )
        repo_path.mkdir(parents=True, exist_ok=True)
        (repo_path / ".git").mkdir(exist_ok=True)
        new_match = {**base_match, "path": "src/New.java", "line_start": 2, "line_end": 4, "score": 200}
        with patch.object(self.service, "_agent_step_terms", return_value=["RiskService"]), patch.object(
            self.service,
            "_execute_tool_loop_step",
            return_value=[dict(base_match), new_match],
        ), patch.object(self.service, "_search_repo", return_value=[]), patch.object(
            self.service,
            "_compress_evidence_cached",
            return_value={"intent": {}},
        ), patch.object(self.service, "_quality_gate_cached", return_value={"status": "sufficient"}):
            collected = self.service._run_agent_plan(
                entries=[entry],
                key="AF:All",
                question="q",
                matches=[base_match],
                evidence_summary={},
                quality_gate={},
                agent_plan={"steps": [{"name": "step", "terms": ["RiskService"], "tools": ["find_definition"]}]},
                limit=3,
                tool_trace=[],
                request_cache=self.service._new_retrieval_request_cache(),
            )
        self.assertTrue(any(match["path"] == "src/New.java" for match in collected))
        missing_repo = RepositoryEntry("Missing Repo", "https://git.example.com/team/missing.git")
        with patch.object(self.service, "_agent_step_terms", return_value=["RiskService"]), patch.object(
            self.service,
            "_search_repo",
            return_value=[],
        ), patch.object(self.service, "_execute_tool_loop_step", return_value=[]) as execute_step:
            self.assertEqual(
                self.service._run_agent_plan(
                    entries=[missing_repo],
                    key="AF:All",
                    question="q",
                    matches=[base_match],
                    evidence_summary={},
                    quality_gate={},
                    agent_plan={"steps": [{"name": "missing", "terms": ["RiskService"], "tools": ["find_definition"]}]},
                    limit=3,
                    tool_trace=[],
                    request_cache=self.service._new_retrieval_request_cache(),
                ),
                [base_match],
            )
            execute_step.assert_called_once()
        two_step_trace: list[dict[str, object]] = []
        with patch.object(self.service, "_agent_step_terms", return_value=["RiskService"]), patch.object(
            self.service,
            "_execute_tool_loop_step",
            return_value=[new_match],
        ), patch.object(self.service, "_search_repo", return_value=[]), patch.object(
            self.service,
            "_compress_evidence_cached",
            return_value={"intent": {}, "entry_points": ["src/New.java"]},
        ), patch.object(self.service, "_quality_gate_cached", return_value={"status": "sufficient", "confidence": "high"}):
            stopped = self.service._run_agent_plan(
                entries=[entry],
                key="AF:All",
                question="q",
                matches=[base_match],
                evidence_summary={},
                quality_gate={},
                agent_plan={
                    "steps": [
                        {"name": "one", "terms": ["RiskService"], "tools": ["find_definition"]},
                        {"name": "two", "terms": ["RiskService"], "tools": ["find_definition"]},
                        {"name": "three", "terms": ["RiskService"], "tools": ["find_definition"]},
                    ]
                },
                limit=3,
                tool_trace=two_step_trace,
                request_cache=self.service._new_retrieval_request_cache(),
            )
        self.assertTrue(any(match["path"] == "src/New.java" for match in stopped))
        self.assertFalse(any(item.get("step") == "three" for item in two_step_trace))

    def test_codex_context_runtime_and_answer_helper_boundaries(self):
        key = self.service.mapping_key("AF", "SG")
        entry = RepositoryEntry("Repo", "https://git.example.com/team/repo.git")
        repo_path = self.service._repo_path(key, entry)
        repo_path.mkdir(parents=True, exist_ok=True)
        (repo_path / ".git").mkdir(exist_ok=True)
        (repo_path / "src").mkdir(exist_ok=True)
        (repo_path / "src" / "App.java").write_text("class App {}", encoding="utf-8")
        index_path = self.service._index_path(repo_path)
        index_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(index_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute("insert into files values ('src/App.java', 'src/app.java', '[]')")
            connection.execute("insert into files values ('other/App.java', 'other/app.java', '[]')")
            connection.commit()

        base_match = {
            "repo": "Repo",
            "path": "src/App.java",
            "line_start": 1,
            "line_end": 1,
            "score": 100,
            "snippet": "select * from risk_table",
            "reason": "direct",
            "retrieval": "exact_table_path_lookup",
            "trace_stage": "exact_lookup",
        }
        self.assertEqual(self.service._codex_cli_session_id({}), "")
        self.service.codex_session_mode = "resume"
        self.assertEqual(self.service._codex_cli_session_id({"codex_cli_session": {}}), "")
        self.assertEqual(self.service._codex_cli_session_id({"codex_cli_session": "bad"}), "")
        self.assertEqual(self.service._codex_cli_session_id({"codex_cli_session": {"session_id": " sess "}}), "sess")

        evidence_pack = {
            "tables": ["risk_table"],
            "read_write_points": ["RiskRepository select"],
            "data_sources": [{"claim": "risk_table", "source_id": "S1"}],
            "typed_items": [{"claim": "Risk API", "source_id": "S2"}],
        }
        self.assertEqual(self.service._codex_sql_relevance_terms("", {"tables": [""]}), [])
        self.assertIn("risk_table", [term.lower() for term in self.service._codex_sql_relevance_terms("select risk_table", evidence_pack)])
        long_text = "\n".join(["header", "risk_table column", "nearby", *[f"line {i}" for i in range(400)]])
        self.assertIn("relevance-filtered", self.service._compact_sql_runtime_evidence_text(long_text, question="risk_table", evidence_pack=evidence_pack, limit=500))
        self.assertIn(
            "runtime evidence text truncated",
            self.service._compact_sql_runtime_evidence_text("x" * 2000, question="zzzz", evidence_pack={}, limit=100),
        )
        runtime_evidence = [
            {
                "id": "r1",
                "filename": "dict.txt",
                "mime_type": "text/plain",
                "kind": "text",
                "source_type": "data_dictionary",
                "pm_team": "AF",
                "country": "SG",
                "size": 10,
                "sha256": "a" * 64,
                "text": "risk_table\n" + "x" * 70000,
            },
            {
                "id": "r2",
                "filename": "apollo.zip",
                "mime_type": "application/zip",
                "kind": "text",
                "source_type": "apollo",
                "pm_team": "AF",
                "country": "SG",
                "size": 10,
                "sha256": "b" * 64,
                "text": "config=value",
            },
        ]
        self.assertEqual(self.service._runtime_evidence_sql_prompt_section([], question="q", evidence_pack={}), "")
        self.assertIn("Uploaded runtime evidence for SQL generation", self.service._runtime_evidence_sql_prompt_section(runtime_evidence, question="risk_table", evidence_pack=evidence_pack, text_limit=500))
        attachment_section = self.service._attachment_prompt_section(
            [
                {"filename": "long.txt", "mime_type": "text/plain", "kind": "text", "size": 1, "sha256": "c" * 64, "text": "x" * 7000},
                {"filename": "shot.png", "mime_type": "image/png", "kind": "image", "size": 1, "sha256": "d" * 64},
            ]
        )
        self.assertIn("attachment text truncated", attachment_section)
        self.assertIn("Image content is attached", attachment_section)
        self.assertIn("Data dictionary handling", self.service._runtime_evidence_prompt_section(runtime_evidence))
        self.assertNotIn("text", self.service._runtime_evidence_for_budget(runtime_evidence, "cheap")[0])
        self.assertLessEqual(len(self.service._runtime_evidence_for_budget([{"source_type": "other", "summary": "x" * 2000}], "cheap")[0]["summary"]), 1000)
        self.assertIn("compacted", self.service._runtime_evidence_for_budget(runtime_evidence, "balanced")[0]["text"])
        self.assertEqual(len(self.service._repair_candidate_paths_for_runtime_evidence([{"path": str(i)} for i in range(20)], [{}] * 4)), 8)
        self.assertEqual(self.service._attachment_image_paths([{"kind": "text"}, {"kind": "image", "path": "/tmp/a.png"}]), ["/tmp/a.png"])
        self.assertIn(
            "No explicit allowlist",
            self.service._codex_investigation_brief(
                pm_team="AF",
                country="SG",
                question="q",
                candidate_paths=[],
                evidence_pack={},
                quality_gate={},
                followup_context=None,
                scope_roots=[],
            ),
        )

        candidate_paths = self.service._codex_candidate_paths(
            entries=[entry],
            key=key,
            matches=[{}, base_match, dict(base_match), {**base_match, "repo": "", "path": "missing/App.java"}],
        )
        self.assertEqual(candidate_paths[0]["path_status"], "exact")
        self.assertEqual(self.service._codex_candidate_path_status(repo_path, "/abs.java")["status"], "invalid")
        self.assertEqual(self.service._codex_candidate_path_status(repo_path, "../App.java")["status"], "invalid")
        self.assertEqual(self.service._codex_candidate_path_status(repo_path, "")["status"], "invalid")
        self.assertEqual(self.service._codex_candidate_path_status(repo_path, ".")["status"], "missing")
        self.assertEqual(self.service._codex_candidate_path_status(repo_path, "Nope.java")["status"], "missing")
        self.assertEqual(self.service._codex_candidate_path_status(repo_path, "App.java")["status"], "ambiguous_filename")
        outside_dir = Path(self.temp_dir.name) / "outside"
        outside_dir.mkdir(exist_ok=True)
        (outside_dir / "Leak.java").write_text("class Leak {}", encoding="utf-8")
        (repo_path / "outlink").symlink_to(outside_dir, target_is_directory=True)
        self.assertEqual(self.service._codex_candidate_path_status(repo_path, "outlink/Leak.java")["status"], "invalid")
        with sqlite3.connect(index_path) as connection:
            connection.execute("delete from files where path = 'other/App.java'")
            connection.commit()
        self.assertEqual(self.service._codex_candidate_path_status(repo_path, "App.java")["status"], "resolved_by_filename")
        index_path.write_text("not sqlite", encoding="utf-8")
        self.assertEqual(self.service._codex_candidate_path_status(repo_path, "App.java")["status"], "missing")
        roots = self.service._codex_scope_roots(entries=[entry, entry], key=key)
        self.assertEqual(len(roots), 1)

        valid_followup_file = repo_path / "src" / "Followup.java"
        valid_followup_file.write_text("class Followup {}", encoding="utf-8")
        merged_paths = self.service._merge_codex_followup_candidate_paths(
            candidate_paths[:1],
            {
                "codex_inspected_paths": [{"repo": "Repo", "repo_root": str(repo_path), "path": "src/Followup.java"}],
                "codex_candidate_paths": [{"repo": "Repo", "path": "../bad.java"}, {"repo": "Repo", "path": "src/App.java"}],
                "recent_turns": [{"codex_candidate_paths": [{"repo": "Repo", "repo_root": str(repo_path), "path": "missing.java"}]}],
            },
        )
        self.assertTrue(any(item["path"] == "src/Followup.java" for item in merged_paths))
        self.service.codex_top_path_limit = 1
        self.assertEqual(
            len(
                self.service._merge_codex_followup_candidate_paths(
                    candidate_paths[:1],
                    {"codex_candidate_paths": [{"repo": "Repo", "repo_root": str(repo_path), "path": "src/Followup.java"}]},
                )
            ),
            1,
        )
        self.service.codex_top_path_limit = 24
        escaped_paths = self.service._merge_codex_followup_candidate_paths(
            [],
            {"codex_candidate_paths": [{"repo": "Repo", "repo_root": str(repo_path), "path": "outlink/Leak.java"}]},
        )
        self.assertEqual(escaped_paths, [])

        context = self.service._codex_initial_candidate_context(
            entries=[entry],
            key=key,
            question="generate SQL for risk_table",
            matches=[],
            selected_matches=[base_match],
            followup_context={"codex_candidate_paths": merged_paths},
        )
        self.assertIn("sql_generation", context["prompt_mode"])
        followup_match = {"repo": "Repo", "path": "src/App.java", "line_start": 1, "line_end": 1}
        rich_brief = self.service._codex_investigation_brief(
            pm_team="AF",
            country="SG",
            question="follow up",
            candidate_paths=candidate_paths,
            evidence_pack={"entry_points": ["src/App.java"], "tables": ["risk_table"], "missing_hops": ["missing upstream"]},
            quality_gate={"status": "missing", "confidence": "low", "missing": ["caller"]},
            followup_context={
                "question": "previous q",
                "answer": "previous a",
                "rendered_answer": "previous rendered",
                "summary": "previous summary",
                "trace_id": "trace-1",
                "matches_snapshot": [followup_match],
                "codex_candidate_paths": [{**followup_match, "repo_root": str(repo_path)}],
                "codex_citation_validation": {"status": "ok", "cited_path_count": 1, "direct_file_refs": [followup_match]},
                "codex_cli_summary": {"prompt_mode": "answer", "candidate_path_count": 1, "repair_attempted": False},
                "codex_inspected_paths": [{**followup_match, "repo_root": str(repo_path), "source": "manual"}],
                "evidence_pack": {"entry_points": ["src/App.java"], "tables": ["risk_table"], "missing_hops": ["missing upstream"]},
                "recent_turns": [
                    {
                        "question": "turn q",
                        "answer": "turn a",
                        "trace_id": "trace-0",
                        "matches_snapshot": [followup_match],
                        "codex_candidate_paths": [followup_match],
                    }
                ],
            },
            attachments=[{"filename": "shot.png", "mime_type": "image/png", "kind": "image", "size": 1, "sha256": "d" * 64}],
            runtime_evidence=runtime_evidence,
            scope_roots=[{"repo": "Repo", "repo_root": str(repo_path), "repo_relative_root": "Repo"}],
            repair_issues=["fix citations"],
        )
        self.assertIn("User attachments", rich_brief)
        self.assertIn("Previous Codex inspected paths", rich_brief)
        self.assertIn("Prior evidence summary", rich_brief)
        self.assertIn("Repair required", rich_brief)

        answer = json.dumps({"claims": [{"text": "RiskRepository selects risk_table src/App.java:1", "citations": ["S99", "missing/File.java"]}]})
        validation = self.service._validate_codex_citations(
            answer,
            candidate_paths,
            [base_match],
            scope_roots=[{"repo_root": str(repo_path)}],
        )
        self.assertEqual(validation["status"], "warn")
        self.assertTrue(validation["warnings"])
        self.assertEqual(self.service._append_fact_citation("", [base_match]), "")
        self.assertEqual(self.service._append_fact_citation("already [S1]", [base_match]), "already [S1]")
        self.assertIn("[S1]", self.service._append_fact_citation("Repo:src/App.java:1-1 source", [base_match]))
        self.assertEqual(self.service._split_answer_claims("\n- First sentence. Second?"), ["First sentence.", "Second?"])
        parsed = self.service._parse_structured_answer(
            '```json\n{"claims":[{"text":"A","citations":["1"]},"B"],"investigation_steps":{"candidate_evidence":["x"]},"confidence":"high"}\n```'
        )
        self.assertEqual(parsed["claims"][0]["citations"], ["S1"])
        self.assertFalse(self.service._codex_high_risk_question("What is RiskStatus?"))
        self.assertTrue(self.service._codex_high_risk_question("why production fail data source"))
        self.assertTrue(self.service._is_hard_codex_repair_reason("finish_reason_length"))
        self.assertEqual(self.service._filter_codex_repair_reasons_for_tier(question="What is status?", repair_reasons=[], routed_budget_mode="cheap"), ([], [], ""))
        allowed, suppressed, reason = self.service._filter_codex_repair_reasons_for_tier(
            question="What is status?",
            repair_reasons=["empty_codex_answer", "high_risk_claims_missing_scoped_file_evidence"],
            routed_budget_mode="cheap",
        )
        self.assertEqual(allowed, ["empty_codex_answer"])
        self.assertTrue(suppressed)
        self.assertEqual(reason, "cheap_simple_first_pass")
        self.assertEqual(
            self.service._filter_codex_repair_reasons_for_tier(
                question="why data source fail",
                repair_reasons=["empty_codex_answer", "high_risk_claims_missing_scoped_file_evidence"],
                routed_budget_mode="cheap",
            )[2],
            "cheap_hard_failures",
        )
        self.assertEqual(self.service._filter_codex_repair_reasons_for_tier(question="why fail", repair_reasons=["x"], routed_budget_mode="deep")[2], "severe_only")
        self.assertEqual(self.service._codex_repair_remaining_timeout_seconds(time.time()), (None, ""))
        self.assertTrue(
            self.service._codex_deep_investigation_needed(
                question="why data source failed",
                answer="confirmed",
                structured_answer={"confidence": "high"},
                quality_gate={"status": "sufficient", "confidence": "high"},
                answer_judge={"status": "repair"},
                codex_validation={"status": "ok"},
            )
        )
        with patch.object(self.service, "_codex_deep_investigation_terms", return_value=[]):
            self.assertEqual(
                self.service._codex_deep_investigation_matches(
                    entries=[entry],
                    key=key,
                    question="why",
                    matches=[base_match],
                    selected_matches=[base_match],
                    evidence_summary={},
                    quality_gate={},
                    structured_answer={},
                    answer_judge={},
                    codex_validation={},
                    limit=3,
                ),
                [base_match],
            )
        self.service._codex_answer_timeout_seconds = 1
        remaining, skip_reason = self.service._codex_repair_remaining_timeout_seconds(time.time() - 10)
        self.assertEqual(remaining, 0)
        self.assertIn("insufficient", skip_reason)
        self.service._codex_answer_timeout_seconds = 100
        self.assertGreaterEqual(self.service._codex_repair_remaining_timeout_seconds(time.time())[0], 10)
        delattr(self.service, "_codex_answer_timeout_seconds")

        severe = self.service._codex_severe_repair_reasons(
            question="why production fail data source",
            answer='{"detail":"Bad Request"}',
            structured_answer={"format": "prose_fallback"},
            quality_gate={"missing": ["source"]},
            evidence_pack={"items": [{"type": "table", "claim": "risk_table", "support_level": "confirmed"}]},
            answer_judge={"status": "repair"},
            codex_validation={"status": "needs_citation", "out_of_scope_refs": [{}], "unsupported_claims": ["claim"]},
            finish_reason="MAX_TOKENS",
        )
        self.assertIn("bad_request_answer", severe)
        final = self.service._finalize_trusted_model_answer(
            question="data source",
            answer="RiskRepository selects risk_table",
            structured_answer={"confidence": "high", "confirmed_from_code": ["confirmed"], "not_found": ["missing"]},
            evidence_summary={"intent": {"data_source": True}, "data_sources": ["Repo:src/App.java:1-1: select * from risk_table"], "data_carriers": ["Carrier"], "field_population": ["field set"]},
            quality_gate={"missing": ["missing"], "confidence": "medium"},
            claim_check=self.service._trusted_provider_check(),
            selected_matches=[base_match],
            answer_judge=self.service._skipped_codex_answer_check(),
            finish_reason="stop",
        )
        self.assertEqual(final["answer_contract"]["confidence"], "high")
        with patch.object(self.service, "_build_agent_plan", return_value={"steps": []}), patch.object(
            self.service,
            "_quality_gate_trace_terms",
            return_value=["RiskService"],
        ), patch.object(self.service, "_run_agent_plan", return_value=[base_match]) as run_agent_plan:
            self.assertEqual(
                self.service._expand_answer_retry_matches(
                    entries=[entry],
                    key=key,
                    question="q",
                    matches=[base_match],
                    evidence_summary={},
                    quality_gate={},
                    limit=3,
                ),
                [base_match],
            )
            self.assertEqual(run_agent_plan.call_args.kwargs["agent_plan"]["steps"][0]["name"], "answer_retry_deeper_trace")

        mixed_matches = [
            base_match,
            {**base_match, "path": "src/SourceRepository.java", "snippet": "select * from risk_table", "retrieval": "direct", "line_start": 2},
            {**base_match, "path": "src/CarrierDTO.java", "snippet": "class RiskDTO {}", "retrieval": "direct", "line_start": 3},
            {**base_match, "path": "src/FieldSetter.java", "snippet": "target.setRisk(value)", "retrieval": "direct", "line_start": 4},
            {**base_match, "path": "src/Static.java", "retrieval": "static_qa", "static_qa": {"kind": "todo"}, "line_start": 5},
            {**base_match, "path": "src/Impact.java", "retrieval": "planner_caller", "line_start": 6},
            {**base_match, "path": "src/Test.java", "retrieval": "test_coverage", "line_start": 7},
            {**base_match, "path": "src/Boundary.java", "retrieval": "operational_boundary", "line_start": 8},
        ]
        self.assertEqual(self.service._select_llm_matches([], 3), [])
        selected = self.service._select_llm_matches(
            mixed_matches,
            20,
            question="data source static qa impact test coverage transactional boundary",
        )
        self.assertTrue(any(match["retrieval"] == "static_qa" for match in selected))
        self.assertTrue(any(match["retrieval"] == "planner_caller" for match in selected))
        self.assertTrue(any(match["retrieval"] == "test_coverage" for match in selected))
        self.assertTrue(any(match["retrieval"] == "operational_boundary" for match in selected))
        self.assertEqual(
            len(self.service._select_llm_matches([{**base_match, "path": "src/Static.java", "retrieval": "static_qa", "trace_stage": "other", "static_qa": {"kind": "todo"}}], 1, question="static qa")),
            1,
        )
        self.assertEqual(
            len(self.service._select_llm_matches([{**base_match, "path": "src/Test.java", "retrieval": "test_coverage", "trace_stage": "test_coverage"}], 1, question="test coverage")),
            1,
        )
        self.assertEqual(
            len(self.service._select_llm_matches([{**base_match, "path": "src/Boundary.java", "retrieval": "operational_boundary", "trace_stage": "operational_boundary"}], 1, question="transactional boundary")),
            1,
        )
        deep_mode = self.service._resolve_llm_budget(
            "auto",
            "why data source root cause",
            [{**base_match, "trace_stage": "agent_plan_1", "retrieval": "flow_graph", "line_start": index} for index in range(8)],
        )
        self.assertEqual(deep_mode[0], "deep")
        balanced_mode = self.service._resolve_llm_budget(
            "auto",
            "tell me implementation details",
            [{**base_match, "retrieval": "direct", "line_start": index} for index in range(5)],
        )
        self.assertEqual(balanced_mode[0], "balanced")
        self.assertFalse(self.service._is_short_definition_or_status_question(""))
        self.assertFalse(self.service._is_short_definition_or_status_question("x" * 121 + " status"))
        judge = self.service._judge_answer_impl(
            "data source api static qa impact test coverage operational boundary",
            "It likely works.",
            evidence_pack={
                "intent": {
                    "data_source": True,
                    "api": True,
                    "static_qa": True,
                    "impact_analysis": True,
                    "test_coverage": True,
                    "operational_boundary": True,
                },
                "items": [{"type": "table", "claim": "risk_table", "supports_answer": True}],
                "missing_hops": ["missing upstream"],
                "apis": ["POST /risk"],
                "static_findings": ["high hardcoded_secret"],
                "impact_surfaces": ["upstream caller"],
                "test_coverage": ["RiskServiceTest"],
                "operational_boundaries": ["@Transactional"],
            },
            claim_check={"status": "ok", "issues": []},
        )
        self.assertEqual(judge["status"], "repair")
        self.assertTrue(judge["issues"])
        cached_payload = self.service._cached_codex_answer_payload(
            cached={"answer": "RiskRepository selects risk_table", "usage": {"totalTokenCount": 1}, "finish_reason": "stop"},
            question="data source",
            structured_answer={"claims": [], "confidence": "high"},
            evidence_summary={"intent": {"data_source": True}, "data_sources": ["risk_table"]},
            quality_gate={"status": "sufficient", "confidence": "high"},
            evidence_pack={"intent": {"data_source": True}, "items": [{"type": "table", "claim": "risk_table"}]},
            candidate_matches=[base_match],
            candidate_paths=candidate_paths,
            scope_roots=[{"repo_root": str(repo_path)}],
            prompt_mode="codex_investigation_brief_v5",
            llm_route={"prompt_mode": "codex_investigation_brief_v5"},
            llm_budget_mode="auto",
            routed_budget_mode="cheap",
            cache_key="cache-key",
        )
        self.assertTrue(cached_payload["llm_cached"])
        self.service.codex_repair_prompt_token_limit = 1
        repair_context = self.service._codex_repair_answer_context(
            pm_team="AF",
            country="SG",
            question="why data source failed",
            answer="initial answer",
            structured_answer={"claims": []},
            scope_roots=[{"repo_root": str(repo_path)}],
            candidate_paths=candidate_paths,
            runtime_evidence=[],
            repair_issues=["missing citation"],
            deep_needed=False,
            repair_issue_count=1,
            repair_reason="severe_only",
            deep_investigation_added=0,
            selected_model="codex",
            query_mode="deep",
            trace_id="trace",
            progress_callback=None,
            codex_cli_session_id="",
            attachments=[],
            timing={},
            evidence_pack={},
            codex_validation={"status": "needs_citation"},
            claim_check={"status": "needs_citation"},
            answer_judge={"status": "repair"},
            usage={},
            effective_model="codex",
            attempts=1,
            llm_latency_ms=1,
            llm_attempt_log=[],
            finish_reason="stop",
            codex_cli_trace={},
            repair_attempted=True,
            repair_skipped_reason="",
        )
        self.assertFalse(repair_context["repair_attempted"])
        self.assertIn("repair_prompt_too_large", repair_context["repair_skipped_reason"])
        self.service.codex_repair_prompt_token_limit = 60_000
        with patch.object(self.service.llm_provider, "generate", side_effect=ToolError("initial failed")):
            with self.assertRaises(ToolError):
                self.service._codex_initial_answer_result(
                    prompt_context="context",
                    prompt_mode="codex_investigation_brief_v5",
                    progress_callback=None,
                    codex_cli_session_id="",
                    attachments=[],
                    trace_id="trace",
                    initial_prompt_stats={"estimated_prompt_tokens": 1},
                    candidate_paths=candidate_paths,
                    candidate_repo_count=1,
                    selected_model="codex",
                    reasoning_effort="low",
                    routed_budget_mode="cheap",
                    query_mode="deep",
                    question="q",
                    evidence_pack={},
                    timing={},
                    scope_roots=[{"repo_root": str(repo_path)}],
                )
        with patch.object(self.service, "_codex_deep_investigation_matches", return_value=[{**base_match, "path": "src/Deep.java", "line_start": 9}]), patch.object(
            self.service,
            "_build_trace_paths",
            return_value=[{"edges": [{"to_name": "Deep"}]}],
        ), patch.object(self.service, "_quality_gate_cached", return_value={"status": "sufficient", "confidence": "high"}):
            deep_context = self.service._codex_deep_investigation_context(
                entries=[entry],
                key=key,
                question="why data source failed",
                matches=[base_match],
                candidate_matches=[base_match],
                candidate_paths=candidate_paths,
                candidate_path_layers={},
                llm_route={},
                evidence_summary={"intent": {"data_source": True}},
                quality_gate={"status": "missing"},
                evidence_pack={},
                answer="missing",
                structured_answer={"direct_answer": "missing"},
                answer_judge={"issues": ["missing"]},
                codex_validation={"unsupported_claims": ["claim"]},
                budget={"match_limit": 3},
                request_cache=self.service._new_retrieval_request_cache(),
                followup_context=None,
                progress_callback=None,
                trace_id="trace",
                selected_model="codex",
                query_mode="deep",
            )
        self.assertIn("trace_paths", deep_context["evidence_summary"])

    def test_query_orchestration_deadline_and_direct_search_edges(self):
        key = self.service.mapping_key("AF", "SG")
        entry = RepositoryEntry("Repo", "https://git.example.com/team/repo.git")
        entry_payload = {"display_name": entry.display_name, "url": entry.url}
        repo_path = self.service._repo_path(key, entry)
        repo_path.mkdir(parents=True, exist_ok=True)
        (repo_path / ".git").mkdir(exist_ok=True)
        self.service.save_mapping(pm_team="AF", country="SG", repositories=[entry_payload])

        weak = self.service.query(pm_team="AF", country="SG", question="???")
        self.assertEqual(weak["status"], "weak_question")

        common_patches = [
            patch.object(self.service, "repo_status", return_value=[]),
            patch.object(self.service, "_index_freshness_payload", return_value={"status": "fresh"}),
            patch.object(self.service, "_synced_query_entries", return_value=[(entry, repo_path)]),
            patch.object(self.service, "_queryable_index_entries", return_value=[(entry, repo_path)]),
        ]
        with common_patches[0], common_patches[1], common_patches[2], common_patches[3], patch.object(
            self.service,
            "_query_exact_lookup_terms",
            return_value=(["src/Missing.java"], []),
        ), patch.object(self.service, "_exact_table_path_lookup_repo", return_value=[]):
            strict_miss = self.service.query(pm_team="AF", country="SG", question="src/Missing.java")
        self.assertEqual(strict_miss["status"], "no_match")

        focused_match = {
            "repo": entry.display_name,
            "path": "src/Focused.java",
            "line_start": 1,
            "line_end": 1,
            "score": 100,
            "snippet": "class Focused {}",
            "reason": "focused",
            "retrieval": "direct",
            "trace_stage": "focused_search",
        }
        with patch.object(self.service, "repo_status", return_value=[]), patch.object(
            self.service,
            "_index_freshness_payload",
            return_value={"status": "stale_or_missing"},
        ), patch.object(self.service, "_synced_query_entries", return_value=[(entry, repo_path)]), patch.object(
            self.service,
            "_queryable_index_entries",
            return_value=[(entry, repo_path)],
        ), patch.object(self.service, "_query_exact_lookup_terms", return_value=([], ["FocusedTerm"])), patch.object(
            self.service,
            "_search_repo",
            return_value=[focused_match],
        ), patch.object(
            self.service,
            "_query_direct_and_decomposed_matches",
            return_value={"matches": [focused_match], "latency_guarded_query_expansion": False},
        ), patch.object(
            self.service,
            "_rank_and_expand_query_matches",
            return_value=([], False),
        ):
            no_top = self.service.query(pm_team="AF", country="SG", question="FocusedTerm")
        self.assertEqual(no_top["status"], "no_match")

        report_events: list[tuple[object, ...]] = []

        def report(*args):
            report_events.append(args)

        many_matches = [
            {
                "repo": entry.display_name,
                "path": f"src/Hit{index}.java",
                "line_start": 1,
                "line_end": 1,
                "score": 100 + index,
                "snippet": "class Hit {}",
                "reason": "hit",
                "retrieval": "direct",
                "trace_stage": "direct",
            }
            for index in range(30)
        ]
        with patch.object(self.service, "_search_repo", return_value=many_matches), patch.object(
            self.service,
            "_rank_matches",
            side_effect=lambda _question, matches, request_cache=None: matches,
        ), patch.object(
            self.service,
            "_select_result_matches",
            side_effect=lambda matches, _limit, question="": matches,
        ), patch.object(
            self.service,
            "_compress_evidence_cached",
            return_value={"intent": {}},
        ), patch.object(self.service, "_quality_gate_cached", return_value={"status": "sufficient", "confidence": "high"}), patch(
            "bpmis_jira_tool.source_code_qa.time.time",
            return_value=100.0,
        ):
            direct_context = self.service._query_direct_and_decomposed_matches_impl(
                question="api config",
                matches=[],
                tokens=["api"],
                synced_entries=[(entry, repo_path)] * 5,
                simple_quality_trace=True,
                started_at=92.0,
                result_limit=12,
                limit=12,
                query_plan={"intent": {"api": True}, "components": [{"terms": [""]}]},
                request_cache=self.service._new_retrieval_request_cache(),
                report=report,
            )
        self.assertFalse(direct_context["latency_guarded_query_expansion"])

        with patch.object(self.service, "_search_repo", return_value=many_matches), patch.object(
            self.service,
            "_rank_matches",
            side_effect=lambda _question, matches, request_cache=None: matches,
        ), patch.object(
            self.service,
            "_select_result_matches",
            side_effect=lambda matches, _limit, question="": matches,
        ), patch.object(
            self.service,
            "_compress_evidence_cached",
            return_value={"intent": {}},
        ), patch.object(self.service, "_quality_gate_cached", return_value={"status": "missing", "confidence": "low"}), patch(
            "bpmis_jira_tool.source_code_qa.time.time",
            return_value=100.0,
        ):
            latency_context = self.service._query_direct_and_decomposed_matches_impl(
                question="api config",
                matches=[],
                tokens=["api"],
                synced_entries=[(entry, repo_path)] * 3,
                simple_quality_trace=True,
                started_at=92.0,
                result_limit=12,
                limit=12,
                query_plan={"intent": {"api": True}, "components": [{"terms": [""]}]},
                request_cache=self.service._new_retrieval_request_cache(),
                report=report,
            )
        self.assertTrue(latency_context["matches"])

        payload: dict[str, object] = {}
        non_codex_provider = SimpleNamespace(name="other")
        with patch.object(self.service, "llm_provider", non_codex_provider):
            self.service._augment_query_payload_with_llm_answer(
                payload=payload,
                entries=[entry],
                key=key,
                pm_team="AF",
                country="SG",
                question="q",
                matches=[focused_match],
                llm_budget_mode="cheap",
                query_mode="deep",
                trace_id="trace",
                followup_context={},
                normalized_answer_mode="auto",
                request_cache=self.service._new_retrieval_request_cache(),
                progress_callback=None,
                attachments=[],
                runtime_evidence=[],
                effort_assessment=False,
                retrieval_latency_ms=1,
                evidence_pack={},
                report=report,
                query_started_at=time.time() - 200,
            )
        self.assertTrue(payload["deadline_hit"])

        payload = {}
        with patch.object(self.service, "_build_llm_answer", side_effect=ToolError("timed out")):
            self.service._augment_query_payload_with_llm_answer(
                payload=payload,
                entries=[entry],
                key=key,
                pm_team="AF",
                country="SG",
                question="q",
                matches=[focused_match],
                llm_budget_mode="cheap",
                query_mode="deep",
                trace_id="trace",
                followup_context={},
                normalized_answer_mode="auto",
                request_cache=self.service._new_retrieval_request_cache(),
                progress_callback=None,
                attachments=[],
                runtime_evidence=[],
                effort_assessment=False,
                retrieval_latency_ms=1,
                evidence_pack={},
                report=report,
                query_started_at=time.time(),
            )
        self.assertTrue(payload["deadline_hit"])
        payload = {}
        self.service._codex_answer_timeout_seconds = 123
        with patch.object(self.service, "_build_llm_answer", side_effect=ToolError("bad request")):
            with self.assertRaises(ToolError):
                self.service._augment_query_payload_with_llm_answer(
                    payload=payload,
                    entries=[entry],
                    key=key,
                    pm_team="AF",
                    country="SG",
                    question="q",
                    matches=[focused_match],
                    llm_budget_mode="cheap",
                    query_mode="deep",
                    trace_id="trace",
                    followup_context={},
                    normalized_answer_mode="auto",
                    request_cache=self.service._new_retrieval_request_cache(),
                    progress_callback=None,
                    attachments=[],
                    runtime_evidence=[],
                    effort_assessment=False,
                    retrieval_latency_ms=1,
                    evidence_pack={},
                    report=report,
                    query_started_at=time.time(),
                )
        self.assertEqual(self.service._codex_answer_timeout_seconds, 123)
        delattr(self.service, "_codex_answer_timeout_seconds")

        with patch.object(self.service, "_search_repo", return_value=many_matches), patch(
            "bpmis_jira_tool.source_code_qa.time.time",
            return_value=100.0,
        ):
            guarded = self.service._query_direct_and_decomposed_matches_impl(
                question="broad",
                matches=list(many_matches) * 2,
                tokens=["broad"],
                synced_entries=[(entry, repo_path)],
                simple_quality_trace=False,
                started_at=92.0,
                result_limit=12,
                limit=12,
                query_plan={"intent": {}, "components": [{"terms": ["risk"]}]},
                request_cache=self.service._new_retrieval_request_cache(),
                report=report,
            )
        self.assertTrue(guarded["latency_guarded_query_expansion"])
        with patch.object(self.service, "_search_repo", return_value=[]):
            empty_expansion = self.service._query_direct_and_decomposed_matches_impl(
                question="broad",
                matches=[],
                tokens=["broad"],
                synced_entries=[(entry, repo_path)],
                simple_quality_trace=False,
                started_at=time.time(),
                result_limit=12,
                limit=12,
                query_plan={"intent": {}, "components": [{"terms": ["of"]}]},
                request_cache=self.service._new_retrieval_request_cache(),
                report=report,
            )
        self.assertEqual(empty_expansion["matches"], [])

    def _build_index_for_entry(self, key, entry):
        repo_entry = type("Entry", (), entry)()
        repo_path = self.service._repo_path(key, repo_entry)
        self.service._build_repo_index(key, repo_entry, repo_path)
        return repo_path

    def _build_all_indexes(self, service=None):
        target_service = service or self.service
        for key, entries in (target_service.load_config().get("mappings") or {}).items():
            for entry in entries:
                repo_entry = type("Entry", (), entry)()
                repo_path = target_service._repo_path(key, repo_entry)
                if (repo_path / ".git").exists():
                    target_service._build_repo_index(key, repo_entry, repo_path)

    def _llm_result(self, text, *, usage=None, model="codex-cli", finish_reason="STOP"):
        payload = {
            "text": text,
            "candidates": [
                {
                    "content": {"parts": [{"text": text}]},
                    "finishReason": finish_reason,
                }
            ],
            "usageMetadata": usage or {"promptTokenCount": 123, "candidatesTokenCount": 45, "totalTokenCount": 168},
        }
        return LLMGenerateResult(
            payload=payload,
            usage=payload["usageMetadata"],
            model=model,
            attempts=1,
            latency_ms=10,
            attempt_log=({"model": model, "status": "ok"},),
        )

    def test_merge_expanded_matches_dedupes_and_preserves_duplicate_signal(self):
        current_matches = [
            {
                "repo": "Repo",
                "path": "service/IssueService.java",
                "line_start": 10,
                "line_end": 12,
                "retrieval": "file_scan",
                "reason": "direct match",
                "score": 4,
            }
        ]
        expanded_matches = [
            {
                "repo": "Repo",
                "path": "service/IssueService.java",
                "line_start": 10,
                "line_end": 12,
                "retrieval": "flow_graph",
                "reason": "flow edge",
                "score": 9,
            },
            {
                "repo": "Repo",
                "path": "controller/IssueController.java",
                "line_start": 3,
                "line_end": 5,
                "retrieval": "planner_caller",
                "reason": "caller",
                "score": 7,
            },
        ]

        with patch.object(self.service, "_rank_matches", side_effect=lambda _question, matches, request_cache=None: matches), patch.object(
            self.service,
            "_select_result_matches",
            side_effect=lambda matches, _limit, question="": matches,
        ):
            merged = self.service._merge_expanded_matches(
                question="what is impacted",
                current_matches=current_matches,
                expanded_matches=expanded_matches,
                limit=12,
            )

        self.assertEqual(len(merged), 2)
        self.assertEqual(merged[0]["retrieval"], "flow_graph")
        self.assertEqual(merged[0]["retrieval_chain"], ["file_scan", "flow_graph"])
        self.assertIn("corroborated by flow edge", merged[0]["reason"])
        self.assertEqual(merged[0]["score"], 9)
        self.assertEqual(merged[1]["path"], "controller/IssueController.java")

    def test_attachment_prompt_and_codex_image_paths_keep_attachment_evidence_separate(self):
        image_path = Path(self.temp_dir.name) / "screen.png"
        image_path.write_bytes(b"fake-image")
        attachment = {
            "id": "att-1",
            "filename": "screen.png",
            "mime_type": "image/png",
            "kind": "image",
            "size": 10,
            "sha256": "a" * 64,
            "path": str(image_path),
        }

        section = self.service._attachment_prompt_section([attachment])
        image_paths = self.service._attachment_image_paths([attachment])

        self.assertIn("User attachments", section)
        self.assertIn("not repository facts", section)
        self.assertIn("extract visible facts exactly", section)
        self.assertIn("missing_production_evidence", section)
        self.assertEqual(image_paths, [str(image_path)])

    def test_runtime_evidence_prompt_dry_run_dedupes_identical_uploads(self):
        duplicate = {
            "id": "runtime-1",
            "filename": "apollo.json",
            "mime_type": "application/json",
            "kind": "text",
            "source_type": "apollo",
            "pm_team": "AF",
            "country": "SG",
            "sha256": "a" * 64,
            "text": "AF rollout flag enabled",
        }
        unique = {
            "id": "runtime-2",
            "filename": "db.txt",
            "mime_type": "text/plain",
            "kind": "text",
            "source_type": "db",
            "pm_team": "AF",
            "country": "SG",
            "sha256": "b" * 64,
            "text": "Ticket status is APPROVED",
        }

        section = self.service._runtime_evidence_prompt_section([duplicate, dict(duplicate), unique])

        self.assertEqual(section.count("AF rollout flag enabled"), 1)
        self.assertIn("Ticket status is APPROVED", section)
        legacy_duplicate_section = section + "\n  Extracted text/summary:\nAF rollout flag enabled"
        self.assertLess(len(section), len(legacy_duplicate_section))

    def test_non_crms_country_normalizes_to_shared_repository_mapping(self):
        self.assertEqual(self.service.mapping_key("AF", "PH"), "AF:All")
        self.assertEqual(self.service.mapping_key("GRC", "SG"), "GRC:All")
        self.assertEqual(self.service.mapping_key("CRMS", "PH"), "CRMS:PH")

    def test_sync_subprocess_timeout_is_controlled(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Repo One", "url": "https://git.example.com/team/repo.git"}],
        )
        import subprocess

        with patch("subprocess.run", side_effect=subprocess.TimeoutExpired(["git"], timeout=5)):
            payload = self.service.sync(pm_team="AF", country="All")

        self.assertEqual(payload["status"], "partial")
        self.assertIn("timed out", payload["results"][0]["message"])
        self.assertEqual(payload["job"]["status"], "partial")

    def test_ensure_synced_today_skips_before_scheduled_start_date(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Repo One", "url": "https://git.example.com/team/repo.git"}],
        )

        with patch.object(self.service, "_today", return_value=date(2026, 4, 25)), patch.object(
            self.service,
            "sync",
            return_value={"status": "ok", "results": []},
        ) as sync:
            payload = self.service.ensure_synced_today(pm_team="AF", country="All")

        self.assertFalse(payload["attempted"])
        self.assertEqual(payload["status"], "scheduled")
        self.assertEqual(payload["next_sync_date"], "2026-05-08")
        sync.assert_not_called()

    def test_ensure_synced_today_runs_sync_on_scheduled_start_date(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Repo One", "url": "https://git.example.com/team/repo.git"}],
        )

        with patch.object(self.service, "_today", return_value=date(2026, 5, 8)), patch.object(
            self.service,
            "sync",
            return_value={"status": "ok", "results": []},
        ) as sync:
            payload = self.service.ensure_synced_today(pm_team="AF", country="All")

        self.assertTrue(payload["attempted"])
        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["next_sync_date"], "2026-05-22")
        sync.assert_called_once_with(pm_team="AF", country="All")

    def test_ensure_synced_today_skips_fresh_current_schedule_window(self):
        entry = RepositoryEntry(display_name="Repo One", url="https://git.example.com/team/repo.git")
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": entry.display_name, "url": entry.url}],
        )
        repo_path = self.service._repo_path("AF:All", entry)
        (repo_path / ".git").mkdir(parents=True)
        (repo_path / "IssueService.java").write_text("class IssueService {}\n", encoding="utf-8")
        self.service._build_repo_index("AF:All", entry, repo_path)
        scheduled_time = date(2026, 5, 8)
        index_path = self.service._index_path(repo_path)
        with sqlite3.connect(index_path) as connection:
            connection.execute(
                "update metadata set value = ? where key = 'updated_at'",
                [f"{scheduled_time.isoformat()}T01:00:00+00:00"],
            )

        with patch.object(self.service, "_today", return_value=scheduled_time + timedelta(days=3)), patch.object(
            self.service,
            "sync",
            return_value={"status": "ok"},
        ) as sync:
            payload = self.service.ensure_synced_today(pm_team="AF", country="All")

        self.assertFalse(payload["attempted"])
        self.assertEqual(payload["status"], "fresh")
        self.assertEqual(payload["next_sync_date"], "2026-05-22")
        sync.assert_not_called()

    def test_search_repo_request_cache_reuses_identical_search(self):
        entry = RepositoryEntry(display_name="Portal Repo", url="https://git.example.com/team/portal.git")
        repo_path = Path(self.temp_dir.name) / "portal"
        (repo_path / ".git").mkdir(parents=True)
        request_cache = self.service._new_retrieval_request_cache()
        match = {
            "repo": "Portal Repo",
            "path": "src/App.java",
            "line_start": 1,
            "line_end": 2,
            "score": 99,
            "snippet": "class App {}",
            "reason": "test match",
            "trace_stage": "direct",
            "retrieval": "persistent_index",
        }

        with patch.object(self.service, "_require_ready_repo_index", return_value={"state": "ready"}) as mocked_ensure, patch.object(
            self.service,
            "_search_repo_index",
            return_value=[dict(match)],
        ) as mocked_index:
            first = self.service._search_repo(
                entry,
                repo_path,
                ["app"],
                question="where is app",
                request_cache=request_cache,
            )
            first[0]["reason"] = "mutated by caller"
            second = self.service._search_repo(
                entry,
                repo_path,
                ["app"],
                question="where is app",
                request_cache=request_cache,
            )

        self.assertEqual(mocked_ensure.call_count, 1)
        self.assertEqual(mocked_index.call_count, 1)
        self.assertEqual(second[0]["reason"], "test match")
        self.assertEqual(request_cache["stats"]["search_hits"], 1)
        self.assertEqual(request_cache["stats"]["search_misses"], 1)
        self.assertEqual(request_cache["stats"]["index_ensure_hits"], 1)
        self.assertEqual(request_cache["stats"]["index_ensure_misses"], 1)

    def test_rank_matches_reuses_rerank_scores_within_request_cache(self):
        request_cache = self.service._new_retrieval_request_cache()
        matches = [
            {
                "repo": "Portal Repo",
                "path": "src/main/java/com/example/IssueService.java",
                "line_start": 10,
                "line_end": 12,
                "score": 100,
                "snippet": "class IssueService { void createIssue() {} }",
                "reason": "test match",
                "trace_stage": "direct",
                "retrieval": "persistent_index",
            }
        ]

        with patch.object(self.service, "_rerank_score", wraps=self.service._rerank_score) as mocked_rerank:
            first = self.service._rank_matches("where is IssueService createIssue", matches, request_cache=request_cache)
            first[0]["rerank_score"] = -1
            second = self.service._rank_matches("where is IssueService createIssue", matches, request_cache=request_cache)

        self.assertEqual(mocked_rerank.call_count, 1)
        self.assertGreater(second[0]["rerank_score"], 0)
        self.assertNotEqual(second[0]["rerank_score"], -1)
        self.assertEqual(request_cache["stats"]["rerank_hits"], 1)
        self.assertEqual(request_cache["stats"]["rerank_misses"], 1)
        self.assertEqual(request_cache["stats"]["question_feature_hits"], 1)
        self.assertEqual(request_cache["stats"]["question_feature_misses"], 1)

    def test_question_tokens_cache_returns_caller_safe_lists(self):
        first = self.service._question_tokens("where is IssueService createIssue")
        first.append("mutated")
        second = self.service._question_tokens("where is IssueService createIssue")

        self.assertIn("issueservice", second)
        self.assertNotIn("mutated", second)

    def test_index_rows_are_reused_within_request_cache(self):
        index_path = Path(self.temp_dir.name) / "index.sqlite3"
        with sqlite3.connect(index_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute(
                "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            connection.execute(
                "create table semantic_chunks (file_path text, start_line integer, end_line integer, chunk_text text, lower_text text, tokens text, symbols text)"
            )
            connection.execute("insert into files values ('src/App.java', 'src/app.java', '[]')")
            connection.execute("insert into lines values ('src/App.java', 1, 'class App {}', 'class app {}', '[]', 1, 0)")
            connection.execute("insert into semantic_chunks values ('src/App.java', 1, 1, 'class App {}', 'class app {}', '[]', '[]')")
            connection.commit()
        request_cache = self.service._new_retrieval_request_cache()
        with sqlite3.connect(index_path) as connection:
            connection.row_factory = sqlite3.Row
            first = self.service._cached_index_rows(connection, index_path, request_cache=request_cache)
            second = self.service._cached_index_rows(connection, index_path, request_cache=request_cache)

        self.assertIs(first, second)
        self.assertEqual(first["lines_by_path"]["src/App.java"], ["class App {}"])
        self.assertEqual(request_cache["stats"]["index_rows_hits"], 1)
        self.assertEqual(request_cache["stats"]["index_rows_misses"], 1)

    def test_targeted_index_rows_keep_late_matches_without_loading_broad_snapshot(self):
        index_path = Path(self.temp_dir.name) / "targeted.sqlite3"
        with sqlite3.connect(index_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute(
                "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            connection.execute("create table file_tokens (token text, file_path text)")
            connection.execute("create table line_tokens (token text, file_path text, line_no integer)")
            connection.execute(
                "create table semantic_chunks (chunk_id text, file_path text, start_line integer, end_line integer, chunk_text text, lower_text text, tokens text, symbols text)"
            )
            connection.execute("create table semantic_chunk_tokens (token text, chunk_id text, file_path text)")
            for index in range(300):
                path = f"src/noise/Noise{index}.java"
                connection.execute("insert into files values (?, ?, ?)", (path, path.lower(), "[]"))
                connection.execute(
                    "insert into lines values (?, ?, ?, ?, ?, ?, ?)",
                    (path, 1, "class Noise {}", "class noise {}", "[]", 1, 0),
                )
            connection.execute("insert into files values ('src/service/Late.java', 'src/service/late.java', '[]')")
            connection.execute("insert into file_tokens values ('needlevalue', 'src/service/Late.java')")
            connection.execute(
                "insert into lines values ('src/service/Late.java', 9001, 'return needleValue;', 'return needlevalue;', '[\"needlevalue\"]', 0, 0)"
            )
            connection.execute("insert into line_tokens values ('needlevalue', 'src/service/Late.java', 9001)")
            connection.execute(
                "insert into semantic_chunks values ('late-1', 'src/service/Late.java', 9001, 9001, 'return needleValue;', 'return needlevalue;', '[\"needlevalue\"]', '[\"needlevalue\"]')"
            )
            connection.execute("insert into semantic_chunk_tokens values ('needlevalue', 'late-1', 'src/service/Late.java')")
            connection.commit()
        request_cache = self.service._new_retrieval_request_cache()
        with sqlite3.connect(index_path) as connection:
            connection.row_factory = sqlite3.Row
            rows = self.service._targeted_index_rows(
                connection,
                index_path,
                tokens=["needlevalue"],
                focus_terms=[],
                intent={},
                request_cache=request_cache,
            )

        self.assertIn("src/service/Late.java", rows["files_by_path"])
        self.assertTrue(any(row["file_path"] == "src/service/Late.java" for row in rows["lines"]))
        self.assertLessEqual(len(rows["files"]), 220)
        self.assertLessEqual(len(rows["lines"]), 1200)
        self.assertEqual(request_cache["stats"]["targeted_index_rows_misses"], 1)

    def test_retrieval_tool_cache_and_fallback_edges(self):
        from bpmis_jira_tool import source_code_qa_retrieval_tools as retrieval_tools

        index_path = Path(self.temp_dir.name) / "retrieval_edges.sqlite3"
        with sqlite3.connect(index_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute(
                "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            connection.execute("create table definitions (name text, lower_name text, kind text, file_path text, line_no integer, signature text)")
            connection.execute("create table references_index (target text, lower_target text, kind text, file_path text, line_no integer, context text)")
            connection.execute(
                "create table semantic_chunks (chunk_id text, file_path text, start_line integer, end_line integer, chunk_text text, lower_text text, tokens text, symbols text)"
            )
            connection.execute("create virtual table files_fts using fts5(path)")
            connection.execute("create virtual table lines_fts using fts5(file_path, line_no unindexed, line_text)")
            connection.execute("create virtual table semantic_chunks_fts using fts5(chunk_id unindexed, file_path, chunk_text)")
            connection.execute("insert into files values ('src/App.java', 'src/app.java', '[\"appsymbol\"]')")
            connection.execute("insert into lines values ('src/App.java', 1, 'class App { jdbcTemplate.query(); }', 'class app { jdbctemplate.query(); }', '[\"appsymbol\"]', 1, 1)")
            connection.execute("insert into definitions values ('AppService', 'appservice', 'class', 'src/App.java', 1, 'class AppService')")
            connection.execute("insert into references_index values ('/api/app', '/api/app', 'route', 'src/App.java', 1, 'RequestMapping')")
            connection.execute("insert into semantic_chunks values ('chunk-1', 'src/App.java', 1, 1, 'jdbcTemplate selects account_table', 'jdbctemplate selects account_table', '[\"account_table\"]', '[\"AppService\"]')")
            connection.execute("insert into files_fts(rowid, path) values (1, 'src/App.java')")
            connection.execute("insert into lines_fts(rowid, file_path, line_no, line_text) values (1, 'src/App.java', 1, 'class App jdbcTemplate')")
            connection.execute("insert into semantic_chunks_fts(rowid, chunk_id, file_path, chunk_text) values (1, 'chunk-1', 'src/App.java', 'jdbcTemplate account_table')")
            connection.commit()

        request_cache = self.service._new_retrieval_request_cache()
        with sqlite3.connect(index_path) as connection:
            connection.row_factory = sqlite3.Row
            self.assertEqual(self.service._cached_file_lines(connection, index_path, ""), [])
            self.assertEqual(self.service._file_fts_search_rows(connection, ["the"], [], index_path=index_path), [])
            self.assertEqual(self.service._fts_search_rows(connection, ["the"], [], index_path=index_path), [])
            self.assertEqual(self.service._semantic_fts_search_rows(connection, ["the"], [], index_path=index_path), [])
            self.assertEqual(
                self.service._cached_structure_like_rows(
                    connection,
                    index_path,
                    table="bad",
                    lower_column="lower_name",
                    term="app",
                    limit=5,
                    request_cache=request_cache,
                ),
                [],
            )
            self.assertEqual(
                self.service._cached_structure_like_rows(
                    connection,
                    index_path,
                    table="definitions",
                    lower_column="lower_name",
                    term="ap",
                    limit=5,
                    request_cache=request_cache,
                ),
                [],
            )
            first_file_fts = self.service._file_fts_search_rows(connection, ["app"], [], index_path=index_path, request_cache=request_cache)
            second_file_fts = self.service._file_fts_search_rows(connection, ["app"], [], index_path=index_path, request_cache=request_cache)
            first_line_fts = self.service._fts_search_rows(connection, ["jdbctemplate"], [], index_path=index_path, request_cache=request_cache)
            second_line_fts = self.service._fts_search_rows(connection, ["jdbctemplate"], [], index_path=index_path, request_cache=request_cache)
            first_semantic_fts = self.service._semantic_fts_search_rows(connection, ["account_table"], [], index_path=index_path, request_cache=request_cache)
            second_semantic_fts = self.service._semantic_fts_search_rows(connection, ["account_table"], [], index_path=index_path, request_cache=request_cache)
            first_structure = self.service._cached_structure_like_rows(
                connection,
                index_path,
                table="definitions",
                lower_column="lower_name",
                term="appservice",
                limit=5,
                request_cache=request_cache,
            )
            second_structure = self.service._cached_structure_like_rows(
                connection,
                index_path,
                table="definitions",
                lower_column="lower_name",
                term="appservice",
                limit=5,
                request_cache=request_cache,
            )
            semantic_matches = self.service._semantic_chunk_matches(
                connection,
                entry=RepositoryEntry("Repo", "https://git.example.com/repo.git"),
                tokens=["account_table"],
                question="Which data source does App use?",
                focus_terms=["AppService"],
                trace_stage="dependency",
                repo_score=3,
                trace_stage_bonus=5,
                rows=None,
                intent={"data_source": True},
            )

        self.assertEqual(first_file_fts, second_file_fts)
        self.assertEqual(first_line_fts, second_line_fts)
        self.assertEqual(first_semantic_fts, second_semantic_fts)
        self.assertEqual(first_structure, second_structure)
        self.assertTrue(semantic_matches)
        self.assertIn("dependency trace", semantic_matches[0]["reason"])
        self.assertEqual(request_cache["stats"]["file_fts_hits"], 1)
        self.assertEqual(request_cache["stats"]["fts_hits"], 1)
        self.assertEqual(request_cache["stats"]["semantic_fts_hits"], 1)
        self.assertEqual(request_cache["stats"]["structure_like_hits"], 1)
        self.assertFalse(retrieval_tools._is_large_index_file(Path(self.temp_dir.name) / "missing.sqlite3"))
        self.assertEqual(
            self.service._structure_lookup_query_terms(
                ["the", "simple", "veryveryveryverylongplainterm", "path/name", "module.name", "low"],
                ["focused_term"],
                large_index=True,
                limit=4,
            ),
            ["focused_term", "simple", "path/name", "module.name"],
        )

    def test_targeted_index_rows_use_degraded_sqlite_fallbacks(self):
        index_path = Path(self.temp_dir.name) / "targeted_fallback.sqlite3"
        with sqlite3.connect(index_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute(
                "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            connection.execute(
                "create table semantic_chunks (chunk_id text, file_path text, start_line integer, end_line integer, chunk_text text, lower_text text, tokens text, symbols text)"
            )
            connection.execute("insert into files values ('src/NeedleConfig.yaml', 'src/needleconfig.yaml', '[]')")
            connection.execute(
                "insert into lines values ('src/NeedleConfig.yaml', 7, 'needleValue: true', 'needlevalue: true', '[]', 0, 1)"
            )
            connection.execute(
                "insert into semantic_chunks values ('needle-chunk', 'src/NeedleConfig.yaml', 7, 7, 'needleValue config', 'needlevalue config', '[\"needlevalue\"]', '[]')"
            )
            connection.commit()

        with sqlite3.connect(index_path) as connection:
            connection.row_factory = sqlite3.Row
            rows = self.service._targeted_index_rows(
                connection,
                index_path,
                tokens=["needlevalue"],
                focus_terms=["needleconfig"],
                intent={"config": True},
                request_cache=self.service._new_retrieval_request_cache(),
            )

        self.assertIn("src/NeedleConfig.yaml", rows["files_by_path"])
        self.assertTrue(any(row["file_path"] == "src/NeedleConfig.yaml" for row in rows["lines"]))
        self.assertTrue(any(row["file_path"] == "src/NeedleConfig.yaml" for row in rows["semantic_chunks"]))

    def test_exact_lookup_and_search_repo_edge_paths(self):
        entry = RepositoryEntry(display_name="Repo", url="https://git.example.com/team/repo.git")
        repo_path = self.service._repo_path("AF:All", entry)
        (repo_path / ".git").mkdir(parents=True)
        index_path = self.service._index_path(repo_path)
        index_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(index_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute(
                "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            connection.execute("create table line_tokens (token text, file_path text, line_no integer)")
            connection.execute("create table references_index (target text, lower_target text, kind text, file_path text, line_no integer, context text)")
            connection.execute("create virtual table lines_fts using fts5(file_path, line_no unindexed, line_text)")
            connection.execute("insert into files values ('src/AccountMapper.java', 'src/accountmapper.java', '[\"AccountMapper\"]')")
            connection.execute(
                "insert into lines values ('src/AccountMapper.java', 4, 'select * from ads_account_snapshot_table', 'select * from ads_account_snapshot_table', '[\"ads_account_snapshot_table\"]', 0, 0)"
            )
            connection.execute(
                "insert into references_index values ('ads_account_snapshot_table', 'ads_account_snapshot_table', 'sql_table', 'src/AccountMapper.java', 4, 'select')"
            )
            connection.execute("insert into line_tokens values ('fallback_only_table', 'src/AccountMapper.java', 4)")
            connection.execute("insert into lines_fts(rowid, file_path, line_no, line_text) values (1, 'src/AccountMapper.java', 4, 'text_only_lookup_table')")
            connection.commit()

        terms = self.service._extract_exact_lookup_terms(
            "Open https://example.com then inspect src/AccountMapper.java and db.ads_account_snapshot_table plus very_long_identifier_with_many_parts"
        )
        self.assertIn("src/accountmapper.java", terms)
        self.assertIn("db.ads_account_snapshot_table", terms)
        self.assertIn("very_long_identifier_with_many_parts", terms)
        self.assertNotIn("https://example.com", terms)
        self.assertTrue(self.service._is_strict_exact_lookup_term("src/AccountMapper.java"))
        self.assertFalse(self.service._is_strict_exact_lookup_term("api/account"))
        self.assertFalse(self.service._exact_lookup_is_sufficient([], []))
        with patch.object(self.service, "_require_ready_repo_index", return_value={"state": "ready"}):
            matches = self.service._exact_table_path_lookup_repo(
                entry,
                repo_path,
                [
                    "db.ads_account_snapshot_table",
                    "src/accountmapper.java",
                    "fallback_only_table",
                    "text_only_lookup_table",
                ],
                question="where is account snapshot table",
                request_cache=self.service._new_retrieval_request_cache(),
            )
        lookup_sources = {match["exact_lookup"]["source"] for match in matches}
        self.assertIn("references_index", lookup_sources)
        self.assertIn("files", lookup_sources)
        self.assertIn("line_tokens", lookup_sources)
        self.assertIn("lines_fts", lookup_sources)
        self.assertTrue(self.service._exact_lookup_is_sufficient(["src/accountmapper.java"], matches))
        with patch.object(self.service, "_require_ready_repo_index", side_effect=sqlite3.Error("broken")):
            self.assertEqual(self.service._search_repo(entry, repo_path, ["account"], question="account"), [])

    def test_planner_tools_cover_index_edges_and_fallbacks(self):
        entry = RepositoryEntry(display_name="Repo", url="https://git.example.com/team/repo.git")
        repo_path = self.service._repo_path("AF:All", entry)
        (repo_path / ".git").mkdir(parents=True)
        index_path = self.service._index_path(repo_path)
        index_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(index_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute(
                "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            connection.execute("create table definitions (name text, lower_name text, kind text, file_path text, line_no integer, signature text)")
            connection.execute("create table code_entities (name text, lower_name text, kind text, file_path text, line_no integer)")
            connection.execute(
                "create table references_index (target text, lower_target text, kind text, file_path text, line_no integer, context text)"
            )
            connection.execute(
                "create table flow_edges (edge_kind text, from_name text, to_name text, evidence text, from_kind text, to_kind text, from_file text, from_line integer, to_file text, to_line integer)"
            )
            connection.execute(
                "create table graph_edges (edge_kind text, symbol text, from_file text, from_line integer, to_file text, to_line integer)"
            )
            connection.execute(
                "create table entity_edges (edge_kind text, from_name text, to_name text, from_file text, from_line integer, to_file text, to_line integer)"
            )
            rows = {
                "src/service/IssueService.java": [
                    "class IssueService {",
                    "  String password = \"secret-token\";",
                    "  void createIssue() { mapper.save(); }",
                    "}",
                ],
                "src/controller/IssueController.java": [
                    "@PostMapping(\"/api/issues\")",
                    "class IssueController { IssueService service; }",
                ],
                "src/repository/IssueMapper.java": [
                    "class IssueMapper {",
                    "  void save() { select * from issue_table; }",
                    "}",
                ],
                "src/client/ExternalClient.java": ["class ExternalClient {}"],
                "tests/IssueServiceTest.java": [
                    "@Test",
                    "void createIssueTest() { assertThat(service.createIssue()).isNotNull(); }",
                ],
                "node_modules/dist/app.min.js": ["const password = \"ignored-secret\";"],
            }
            for path, lines in rows.items():
                connection.execute("insert into files values (?, ?, ?)", (path, path.lower(), "[]"))
                for line_no, text in enumerate(lines, start=1):
                    connection.execute(
                        "insert into lines values (?, ?, ?, ?, ?, ?, ?)",
                        (path, line_no, text, text.lower(), "[]", 1 if "class " in text else 0, 0),
                    )
            connection.execute("insert into definitions values ('IssueService', 'issueservice', 'class', 'src/service/IssueService.java', 1, '')")
            connection.execute("insert into definitions values ('IssueMapper', 'issuemapper', 'class', 'src/repository/IssueMapper.java', 1, '')")
            connection.execute("insert into code_entities values ('ExternalClient', 'externalclient', 'class', 'src/client/ExternalClient.java', 1)")
            connection.execute(
                "insert into references_index values ('issue_table', 'issue_table', 'sql_table', 'src/repository/IssueMapper.java', 2, 'select from issue_table')"
            )
            connection.execute(
                "insert into references_index values ('unused_table', 'unused_table', 'sql_table', 'src/repository/IssueMapper.java', 2, 'plain context')"
            )
            connection.execute(
                "insert into references_index values ('/api/issues', '/api/issues', 'route', 'src/controller/IssueController.java', 1, 'PostMapping')"
            )
            connection.execute(
                "insert into references_index values ('IssueService', 'issueservice', 'test_subject', 'tests/IssueServiceTest.java', 2, 'assertThat')"
            )
            connection.execute(
                "insert into references_index values ('Transactional', 'transactional', 'operational_boundary', 'src/service/IssueService.java', 3, '@Transactional createIssue')"
            )
            connection.execute(
                "insert into references_index values ('Transactional', 'transactional', 'operational_boundary', 'src/service/IssueService.java', 3, '@Transactional duplicate')"
            )
            connection.execute(
                "insert into flow_edges values ('service', 'IssueController', 'IssueService', 'controller calls service', 'controller', 'service', 'src/controller/IssueController.java', 2, 'src/service/IssueService.java', 3)"
            )
            connection.execute(
                "insert into flow_edges values ('repository', 'IssueService', 'IssueMapper', 'service calls mapper', 'service', 'repository', 'src/service/IssueService.java', 3, 'src/repository/IssueMapper.java', 2)"
            )
            connection.execute(
                "insert into flow_edges values ('repository', 'IssueService', 'IssueMapper', 'service calls mapper duplicate', 'service', 'repository', 'src/service/IssueService.java', 3, 'src/repository/IssueMapper.java', 2)"
            )
            connection.execute(
                "insert into flow_edges values ('client', 'IssueService', 'ExternalClient', 'client without file', 'service', 'client', 'src/service/IssueService.java', 3, '', 0)"
            )
            connection.execute(
                "insert into flow_edges values ('client', 'IssueService', 'IssueMapper', 'definition fallback without file', 'service', 'client', 'src/service/IssueService.java', 3, '', 0)"
            )
            connection.execute(
                "insert into graph_edges values ('call', 'createIssue', 'src/service/IssueService.java', 3, 'src/repository/IssueMapper.java', 2)"
            )
            connection.execute(
                "insert into entity_edges values ('injects', 'IssueController', 'IssueService', 'src/controller/IssueController.java', 2, 'src/service/IssueService.java', 1)"
            )
            connection.commit()

        request_cache = self.service._new_retrieval_request_cache()
        base_matches = [
            {
                "repo": entry.display_name,
                "path": "src/service/IssueService.java",
                "line_start": 3,
                "line_end": 3,
                "score": 200,
                "snippet": "createIssue mapper.save",
                "reason": "seed",
                "retrieval": "persistent_index",
            }
        ]
        with patch.object(self.service, "_require_ready_repo_index", return_value={"state": "ready"}):
            static_matches = self.service._tool_find_static_findings(
                entry,
                repo_path,
                ["secret"],
                "find static secret risks",
                1,
                request_cache=request_cache,
            )
            test_matches = self.service._tool_find_test_coverage(
                entry,
                repo_path,
                ["IssueService"],
                "does IssueService have test coverage?",
                2,
                request_cache=request_cache,
            )
            boundary_matches = self.service._tool_find_operational_boundaries(
                entry,
                repo_path,
                ["Transactional"],
                "what transactional boundary exists?",
                3,
                request_cache=request_cache,
            )
            caller_matches = self.service._tool_find_callers(
                entry,
                repo_path,
                base_matches,
                ["IssueService"],
                "who calls IssueService?",
                4,
                request_cache=request_cache,
            )
            callee_matches = self.service._tool_find_callees(
                entry,
                repo_path,
                base_matches,
                ["IssueMapper"],
                "what does IssueService call?",
                5,
                request_cache=request_cache,
            )
            open_matches = self.service._tool_open_file_window(
                entry,
                repo_path,
                base_matches,
                "open service",
                6,
                request_cache=request_cache,
            )
            structure_matches = self.service._tool_lookup_structure(
                entry,
                repo_path,
                ["IssueMapper"],
                question="find mapper definition",
                table="definitions",
                name_column="name",
                lower_column="lower_name",
                line_column="line_no",
                kind_column="kind",
                trace_stage="tool_loop_7",
                retrieval="planner_definition",
                request_cache=request_cache,
            )
            graph_matches = self.service._tool_trace_graph(
                entry,
                repo_path,
                base_matches,
                "trace graph",
                8,
                request_cache=request_cache,
            )
            flow_matches = self.service._tool_trace_flow(
                entry,
                repo_path,
                base_matches,
                "trace flow",
                9,
                request_cache=request_cache,
            )
            entity_matches = self.service._tool_trace_entity(
                entry,
                repo_path,
                [{"repo": entry.display_name, "path": "src/controller/IssueController.java"}],
                "trace entity",
                10,
                request_cache=request_cache,
            )
            impact_matches = self.service._expand_impact_matches(
                entries=[entry],
                key="AF:All",
                question="IssueService impact upstream downstream IssueMapper",
                base_matches=base_matches,
                limit=8,
                request_cache=request_cache,
            )
            skipped_table_matches = self.service._tool_lookup_references_by_kind(
                entry,
                repo_path,
                ["notpresent"],
                kinds={"sql_table"},
                question="unmatched table",
                trace_stage="tool_loop_11",
                retrieval="planner_table",
                score=100,
                request_cache=request_cache,
            )

        self.assertTrue(static_matches)
        self.assertEqual({match["path"] for match in static_matches}, {"src/service/IssueService.java"})
        self.assertTrue(
            any(
                "issueservice" in match.get("test_coverage", {}).get("terms", [])
                or match.get("test_coverage", {}).get("target") == "IssueService"
                for match in test_matches
            )
        )
        self.assertEqual(boundary_matches[0]["operational_boundary"]["kind"], "operational_boundary")
        self.assertTrue(any(match["path"] == "src/controller/IssueController.java" for match in caller_matches))
        self.assertTrue(any(match["path"] == "src/repository/IssueMapper.java" for match in callee_matches))
        self.assertTrue(open_matches)
        self.assertEqual(structure_matches[0]["path"], "src/repository/IssueMapper.java")
        self.assertTrue(graph_matches)
        self.assertTrue(flow_matches)
        self.assertTrue(entity_matches)
        self.assertTrue(any(match["retrieval"] == "planner_caller" for match in impact_matches))
        self.assertTrue(any(match["retrieval"] == "planner_callee" for match in impact_matches))
        self.assertTrue(all(match["reason"].endswith("issue_table") for match in skipped_table_matches))

        with patch.object(self.service, "_require_ready_repo_index", side_effect=sqlite3.Error("broken")):
            self.assertEqual(self.service._tool_find_static_findings(entry, repo_path, ["secret"], "q", 1), [])
            self.assertEqual(self.service._tool_find_test_coverage(entry, repo_path, ["IssueService"], "q", 1), [])
            self.assertEqual(self.service._tool_find_operational_boundaries(entry, repo_path, ["Transactional"], "q", 1), [])
            self.assertEqual(
                self.service._tool_lookup_references_by_kind(
                    entry,
                    repo_path,
                    ["issue_table"],
                    kinds={"sql_table"},
                    question="q",
                    trace_stage="tool_loop_1",
                    retrieval="planner_table",
                    score=100,
                ),
                [],
            )
        with patch.object(self.service, "_require_ready_repo_index", return_value={"state": "ready"}):
            self.assertEqual(
                self.service._expand_impact_matches(
                    entries=[entry],
                    key="AF:All",
                    question="Fallback",
                    base_matches=[{"repo": entry.display_name, "path": "src/Fallback.java"}],
                    limit=3,
                ),
                [],
            )
            self.assertEqual(
                self.service._tool_lookup_structure(
                    entry,
                    repo_path,
                    ["IssueMapper"],
                    question="q",
                    table="missing_definitions",
                    name_column="name",
                    lower_column="lower_name",
                    line_column="line_no",
                    kind_column="kind",
                    trace_stage="tool_loop_1",
                    retrieval="planner_definition",
                ),
                [],
            )
            with patch(
                "bpmis_jira_tool.source_code_qa_retrieval_tools.sqlite3.connect",
                side_effect=sqlite3.Error("broken"),
            ):
                self.assertEqual(self.service._tool_trace_graph(entry, repo_path, base_matches, "q", 1), [])

    def test_retrieval_edge_helpers_cover_remaining_boundaries(self):
        from bpmis_jira_tool import source_code_qa_retrieval_tools as retrieval_tools

        self.assertEqual(self.service._structure_lookup_query_terms(["", "of", "valid", "valid"], [], large_index=False, limit=3), ["valid"])

        indexed_text_root = Path(self.temp_dir.name) / "text-files"
        (indexed_text_root / ".git").mkdir(parents=True)
        (indexed_text_root / "src").mkdir()
        (indexed_text_root / "node_modules").mkdir()
        (indexed_text_root / "src" / "App.py").write_text("print('ok')\n", encoding="utf-8")
        (indexed_text_root / "node_modules" / "lib.py").write_text("skip\n", encoding="utf-8")
        (indexed_text_root / "src" / "image.bin").write_bytes(b"skip")
        (indexed_text_root / "src" / "large.py").write_text("x" * (self.service.max_file_bytes + 1), encoding="utf-8")
        iterated_paths = {path.name for path in self.service._iter_text_files(indexed_text_root)}
        self.assertEqual(iterated_paths, {"App.py"})

        with self.assertRaises(ToolError):
            self.service._normalize_entry("bad")
        with self.assertRaises(ToolError):
            self.service._normalize_entry({"url": ""})
        self.assertEqual(self.service._entry_to_dict({"display_name": "", "url": "https://git.example.com/team/demo.git"})["display_name"], "demo")
        self.assertIn("https://***:***@", self.service._sanitize_error_detail("https://alice:secret-token@git.example.com/repo.git"))

        request_cache = self.service._new_retrieval_request_cache()
        index_path = Path(self.temp_dir.name) / "large_targeted.sqlite3"
        with sqlite3.connect(index_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute(
                "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            connection.execute(
                "create table semantic_chunks (chunk_id text, file_path text, start_line integer, end_line integer, chunk_text text, lower_text text, tokens text, symbols text)"
            )
            connection.execute("create virtual table lines_fts using fts5(file_path, line_no unindexed, line_text)")
            connection.execute("create virtual table files_fts using fts5(path)")
            connection.execute("create virtual table semantic_chunks_fts using fts5(chunk_id unindexed, file_path, chunk_text)")
            for index in range(2001):
                path = f"src/File{index}.java"
                connection.execute("insert into files values (?, ?, ?)", (path, path.lower(), "[]"))
                connection.execute("insert into lines values (?, 1, 'class File {}', 'class file {}', '[]', 1, 0)", (path,))
            connection.execute("insert into semantic_chunks values ('chunk-1', 'src/File1.java', 1, 1, 'api config', 'api config', '[\"config\"]', '[]')")
            connection.commit()
        with sqlite3.connect(index_path) as connection:
            connection.row_factory = sqlite3.Row
            targeted = self.service._targeted_index_rows(
                connection,
                index_path,
                tokens=["loan", "averyveryveryveryveryveryverylongterm", "mediumlengthtermxxx", "src/file1.java"],
                focus_terms=["focusedplainterm"],
                intent={"api": True},
                request_cache=request_cache,
            )
            targeted_again = self.service._targeted_index_rows(
                connection,
                index_path,
                tokens=["loan", "averyveryveryveryveryveryverylongterm", "mediumlengthtermxxx", "src/file1.java"],
                focus_terms=["focusedplainterm"],
                intent={"api": True},
                request_cache=request_cache,
            )
            self.assertIs(targeted, targeted_again)
            self.assertLessEqual(len(targeted["files"]), 48)

            old_semantic_enabled = self.service.semantic_index_enabled
            try:
                self.service.semantic_index_enabled = False
                self.assertEqual(
                    self.service._targeted_semantic_rows_by_id(
                        connection,
                        index_path,
                        tokens=["config"],
                        focus_terms=[],
                        query_terms=["config"],
                        file_paths=["src/File1.java"],
                        simple_intent=True,
                        max_target_semantic_chunks=2,
                    ),
                    {},
                )
                self.assertEqual(
                    self.service._semantic_chunk_matches(
                        connection,
                        entry=RepositoryEntry("Repo", "https://git.example.com/repo.git"),
                        tokens=["config"],
                        question="config",
                        focus_terms=[],
                        trace_stage="direct",
                        repo_score=0,
                        trace_stage_bonus=0,
                    ),
                    [],
                )
            finally:
                self.service.semantic_index_enabled = old_semantic_enabled
            self.assertEqual(
                self.service._semantic_chunk_matches(
                    connection,
                    entry=RepositoryEntry("Repo", "https://git.example.com/repo.git"),
                    tokens=["the"],
                    question="the",
                    focus_terms=[],
                    trace_stage="direct",
                    repo_score=0,
                    trace_stage_bonus=0,
                    rows=[],
                    intent={},
                ),
                [],
            )

        with sqlite3.connect(Path(self.temp_dir.name) / "no_semantic_table.sqlite3") as connection:
            connection.row_factory = sqlite3.Row
            self.assertEqual(
                self.service._semantic_chunk_matches(
                    connection,
                    entry=RepositoryEntry("Repo", "https://git.example.com/repo.git"),
                    tokens=["config"],
                    question="config",
                    focus_terms=[],
                    trace_stage="direct",
                    repo_score=0,
                    trace_stage_bonus=0,
                    rows=None,
                    intent={},
                ),
                [],
            )

        missing_entry = RepositoryEntry("Missing", "https://git.example.com/missing.git")
        self.assertEqual(
            self.service._expand_dependency_matches(entries=[missing_entry], key="AF:All", question="", base_matches=[], limit=3),
            [],
        )
        with patch.object(self.service, "_dependency_focus_terms", return_value=["IssueService"]):
            self.assertEqual(
                self.service._expand_dependency_matches(entries=[missing_entry], key="AF:All", question="dependency", base_matches=[], limit=3),
                [],
            )
        with patch.object(self.service, "_two_hop_trace_terms", return_value=["IssueService"]):
            self.assertEqual(
                self.service._expand_two_hop_matches(entries=[missing_entry], key="AF:All", question="two hop", base_matches=[], limit=3),
                [],
            )
        with patch.object(self.service, "_agent_trace_terms", return_value=["IssueService"]):
            self.assertEqual(
                self.service._expand_agent_trace_matches(entries=[missing_entry], key="AF:All", question="agent", base_matches=[], limit=3),
                [],
            )

        with patch.object(self.service, "_tool_loop_terms", return_value=[]), patch.object(
            self.service,
            "_build_tool_loop_plan",
            return_value=[{"tool": "unknown", "terms": ["x"]}, {"tool": "find_definition", "terms": []}],
        ):
            self.assertIsNone(
                self.service._choose_next_tool_step(
                    question="",
                    matches=[],
                    evidence_summary={"intent": {}},
                    quality_gate={},
                    executed_steps=set(),
                )
            )
        self.assertEqual(
            self.service._execute_tool_loop_step(
                entries=[missing_entry],
                key="AF:All",
                question="q",
                matches=[],
                step={"tool": "trace_graph", "terms": ["IssueService"]},
                step_index=1,
            ),
            [],
        )
        self.assertFalse(self.service._is_strict_exact_lookup_term(""))
        self.assertTrue(self.service._is_strict_exact_lookup_term("./src/App.java"))
        self.assertTrue(self.service._exact_lookup_miss_should_stop(["./src/App.java"]))
        self.assertFalse(self.service._exact_lookup_miss_should_stop(["api/account"]))
        self.assertFalse(retrieval_tools._is_large_index_file(Path(self.temp_dir.name) / "does-not-exist.sqlite3"))
        self.assertNotIn("https://example.com/path/to/app.java", retrieval_tools._extract_exact_lookup_terms("https://example.com/path/to/app.java"))
        self.assertNotIn("abc.def", retrieval_tools._extract_exact_lookup_terms("abc.def"))
        self.assertEqual(
            retrieval_tools._extract_exact_lookup_terms("pkg.alpha_beta_gamma_delta_value alpha_beta_gamma_delta_value"),
            ["pkg.alpha_beta_gamma_delta_value"],
        )
        class FakeExactUrlPattern:
            def finditer(self, _text):
                class Match:
                    def group(self, _index):
                        return "https://example.com/path.java"

                return [Match()]

        with patch.object(retrieval_tools, "EXACT_LOOKUP_TERM_PATTERN", FakeExactUrlPattern()):
            self.assertEqual(retrieval_tools._extract_exact_lookup_terms(""), [])
        self.assertNotIn("http://example.com/very/long/path.java", self.service._extract_exact_lookup_terms("http://example.com/very/long/path.java"))
        self.assertIn("foo/bar/app.java", self.service._extract_exact_lookup_terms("foo/bar/app.java"))
        self.assertEqual(
            self.service._extract_exact_lookup_terms("db.ads_account_snapshot_table ads_account_snapshot_table"),
            ["db.ads_account_snapshot_table"],
        )
        self.assertTrue(self.service._is_strict_exact_lookup_term("foo/bar/App.java"))
        self.assertTrue(retrieval_tools._is_strict_exact_lookup_term("src/foo"))
        temp_repo = Path(self.temp_dir.name) / "incomplete"
        temp_repo.mkdir()
        self.service._remove_incomplete_repo_dir(temp_repo)
        self.assertFalse(temp_repo.exists())
        self.assertIn("quality-gate trace", self.service._match_reason([], "", "", file_symbols=set(), question="", trace_stage="quality_gate"))

    def test_repo_dependency_graph_matches_routes_messages_tables_and_artifacts(self):
        source = RepositoryEntry(display_name="source-service", url="https://git.example.com/team/source-service.git")
        target = RepositoryEntry(display_name="target-service", url="https://git.example.com/team/target-service.git")
        entries = [source, target]
        source_config = {
            "target.base-url": ["https://target.internal/api"],
            "target.topic": ["risk.events"],
            "inventory.service": ["inventory-service"],
        }
        route_index = {
            target.display_name: [
                {"route": "/api/issues/create", "file": "src/TargetController.java", "line": 8},
                {"route": "/create", "file": "src/TooGenericController.java", "line": 2},
            ]
        }
        message_index = {target.display_name: [{"message": "risk.events", "file": "src/Consumer.java", "line": 12}]}
        artifact_index = {target.display_name: [{"artifact": "target-client", "file": "pom.xml", "line": 4}]}
        table_index = {target.display_name: [{"table": "issue_table", "edge_kind": "db_read", "file": "src/Reader.java", "line": 7}]}

        route_candidate = self.service._match_repo_dependency_candidate(
            row={
                "edge_kind": "client",
                "to_name": "${target.base-url}/issues/create",
                "evidence": "WebClient ${target.base-url}/issues/create",
                "from_file": "src/TargetClient.java",
            },
            entries=entries,
            source_name=source.display_name,
            route_index=route_index,
            source_config=source_config,
        )
        artifact_candidate = self.service._match_repo_dependency_candidate(
            row={"edge_kind": "module_dependency", "to_name": "com.example:target-client", "evidence": "com.example:target-client"},
            entries=entries,
            source_name=source.display_name,
            route_index={},
            artifact_index=artifact_index,
        )
        table_candidate = self.service._match_repo_dependency_candidate(
            row={"edge_kind": "db_write", "to_name": "db.issue_table", "evidence": "insert issue_table"},
            entries=entries,
            source_name=source.display_name,
            route_index={},
            table_index=table_index,
        )
        message_candidate = self.service._match_repo_dependency_candidate(
            row={"edge_kind": "event_publish", "to_name": "${target.topic}", "evidence": "publish"},
            entries=entries,
            source_name=source.display_name,
            route_index={},
            message_index=message_index,
            source_config=source_config,
        )
        skipped_table_candidate = self.service._match_repo_dependency_candidate(
            row={"edge_kind": "db_write", "to_name": "db.issue_table", "evidence": "insert issue_table"},
            entries=entries,
            source_name=source.display_name,
            route_index={},
            table_index={target.display_name: [{"table": "issue_table", "edge_kind": "db_write"}]},
        )
        skipped_message_candidate = self.service._match_repo_dependency_candidate(
            row={"edge_kind": "message_publish", "to_name": "!!!", "evidence": ""},
            entries=entries,
            source_name=source.display_name,
            route_index={},
            message_index=message_index,
            source_config={},
        )
        alias_candidate = self.service._match_repo_dependency_candidate(
            row={"edge_kind": "import", "to_name": "import com.example.targetservice.Client", "evidence": "", "from_file": "src/Client.java"},
            entries=entries,
            source_name=source.display_name,
            route_index={},
        )

        self.assertEqual(route_candidate["edge_kind"], "http_path")
        self.assertEqual(route_candidate["to_file"], "src/TargetController.java")
        self.assertEqual(artifact_candidate["edge_kind"], "module_dependency")
        self.assertEqual(table_candidate["edge_kind"], "shared_table")
        self.assertEqual(message_candidate["edge_kind"], "event_flow")
        self.assertIsNone(skipped_table_candidate)
        self.assertIsNone(skipped_message_candidate)
        self.assertEqual(alias_candidate["edge_kind"], "import")
        self.assertEqual(self.service._normalize_message_name("${target.topic:default}"), "target.topic")
        self.assertIn("target-client", self.service._artifact_values_from_text("com.example:target-client"))
        self.assertEqual(self.service._normalize_table_name("db.issue_table"), "issue_table")
        self.assertEqual(
            self.service._prefer_specific_routes([{"route": "/issues"}, {"route": "/api/issues"}]),
            [{"route": "/api/issues"}],
        )
        self.assertEqual(self.service._join_config_routes("/create", ["https://target.internal/api"]), ["/api/create"])
        self.assertEqual(self.service._resolve_config_placeholders("${target.base-url}/x", source_config), ["https://target.internal/api"])
        self.assertIn("https://target.internal/api", self.service._candidate_dependency_config_values(source_config))
        self.assertEqual(self.service._route_overlap_score("/api/issues/create", "https://target.internal/api/issues/create"), 0.96)
        self.assertEqual(self.service._route_overlap_score("/", "/"), 0.0)
        self.assertEqual(self.service._route_overlap_score("/{id}", "/{name}"), 0.0)
        self.assertEqual(self.service._join_routes("", "/health"), "/health")
        self.assertEqual(self.service._join_routes("/health", ""), "/health")
        self.assertEqual(self.service._join_routes("https://host/base", "/v1"), "https://host/base/v1")
        self.assertGreater(self.service._repo_alias_match_score("targetservice client", target), 0)
        self.assertGreater(self.service._repo_alias_match_score("target", target), 0)
        self.assertEqual(self.service._repo_alias_match_score("", target), 0)

    def test_repo_dependency_indexes_build_from_persistent_indexes(self):
        source = RepositoryEntry(display_name="source-service", url="https://git.example.com/team/source-service.git")
        target = RepositoryEntry(display_name="target-service", url="https://git.example.com/team/target-service.git")
        missing = RepositoryEntry(display_name="missing-service", url="https://git.example.com/team/missing-service.git")
        entries = [source, target, missing]

        def create_repo_index(entry, *, lines, flow_rows):
            repo_path = self.service._repo_path("AF:All", entry)
            (repo_path / ".git").mkdir(parents=True)
            index_path = self.service._index_path(repo_path)
            index_path.parent.mkdir(parents=True, exist_ok=True)
            with sqlite3.connect(index_path) as connection:
                connection.execute(
                    "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
                )
                connection.execute(
                    "create table flow_edges (edge_kind text, to_name text, evidence text, from_file text, from_line integer, to_file text, to_line integer)"
                )
                for file_path, line_no, line_text in lines:
                    connection.execute(
                        "insert into lines values (?, ?, ?, ?, '[]', 0, 0)",
                        (file_path, line_no, line_text, line_text.lower()),
                    )
                for row in flow_rows:
                    connection.execute("insert into flow_edges values (?, ?, ?, ?, ?, ?, ?)", row)
                connection.commit()

        create_repo_index(
            source,
            lines=[
                ("src/main/resources/application.properties", 1, "target.base-url=https://target.internal"),
                ("src/main/resources/application.yml", 1, "messaging:"),
                ("src/main/resources/application.yml", 2, "  topic: risk.events"),
            ],
            flow_rows=[
                ("client", "${target.base-url}/api/issues/create", "WebClient ${target.base-url}/api/issues/create", "src/TargetClient.java", 10, "", 0),
                ("module_dependency", "com.example:target-client", "com.example:target-client", "pom.xml", 4, "", 0),
                ("event_publish", "${messaging.topic}", "publish risk event", "src/Publisher.java", 12, "", 0),
                ("db_write", "db.issue_table", "insert issue_table", "src/Writer.java", 20, "", 0),
            ],
        )
        create_repo_index(
            target,
            lines=[
                ("src/main/resources/application.properties", 1, "target.client.enabled=true"),
                ("src/main/resources/application.properties", 2, "messaging.topic=risk.events"),
            ],
            flow_rows=[
                ("route", "/api/issues/create", "@PostMapping", "src/TargetController.java", 8, "", 0),
                ("module_dependency", "target-client", "target-client", "pom.xml", 4, "", 0),
                ("message_consume", "${messaging.topic}", "consume topic", "src/Consumer.java", 12, "", 0),
                ("db_read", "issue_table", "select issue_table", "src/Reader.java", 7, "", 0),
            ],
        )
        request_cache = self.service._new_retrieval_request_cache()
        with patch.object(self.service, "_require_ready_repo_index", return_value={"state": "ready"}):
            route_index = self.service._repo_route_index(key="AF:All", entries=entries, request_cache=request_cache)
            config_index = self.service._repo_config_index(key="AF:All", entries=entries, request_cache=request_cache)
            message_index = self.service._repo_message_index(
                key="AF:All",
                entries=entries,
                config_index=config_index,
                request_cache=request_cache,
            )
            artifact_index = self.service._repo_artifact_index(key="AF:All", entries=entries, request_cache=request_cache)
            table_index = self.service._repo_table_index(key="AF:All", entries=entries, request_cache=request_cache)
            graph = self.service._build_repo_dependency_graph(key="AF:All", entries=entries, request_cache=request_cache)

        self.assertEqual(route_index[target.display_name][0]["route"], "/api/issues/create")
        self.assertEqual(config_index[source.display_name]["target.base-url"], ["https://target.internal"])
        self.assertIn("risk.events", message_index[target.display_name][0]["message"])
        self.assertEqual(artifact_index[target.display_name][0]["artifact"], "target-client")
        self.assertEqual(table_index[target.display_name][0]["table"], "issue_table")
        self.assertTrue(any(edge["edge_kind"] in {"http_path", "module_dependency", "event_flow", "shared_table"} for edge in graph["edges"]))

        broken = RepositoryEntry(display_name="broken-service", url="https://git.example.com/team/broken-service.git")
        broken_path = self.service._repo_path("AF:All", broken)
        (broken_path / ".git").mkdir(parents=True)
        self.service._index_path(broken_path).parent.mkdir(parents=True, exist_ok=True)
        self.service._index_path(broken_path).write_text("not sqlite", encoding="utf-8")
        with patch.object(self.service, "_require_ready_repo_index", return_value={"state": "ready"}):
            self.assertEqual(self.service._repo_message_index(key="AF:All", entries=[broken]), {broken.display_name: []})
            self.assertEqual(self.service._repo_artifact_index(key="AF:All", entries=[broken]), {broken.display_name: []})
            self.assertEqual(self.service._repo_table_index(key="AF:All", entries=[broken]), {broken.display_name: []})
            self.assertEqual(self.service._repo_route_index(key="AF:All", entries=[broken]), {broken.display_name: []})
            self.assertEqual(self.service._repo_config_index(key="AF:All", entries=[broken]), {broken.display_name: {}})
            broken_graph = self.service._build_repo_dependency_graph(key="AF:All", entries=[broken])
        self.assertEqual(broken_graph["edges"], [])

    def test_retrieval_fallbacks_and_empty_guards(self):
        entry = RepositoryEntry(display_name="Repo", url="https://git.example.com/team/repo.git")
        repo_path = self.service._repo_path("AF:All", entry)
        (repo_path / ".git").mkdir(parents=True)
        index_path = self.service._index_path(repo_path)
        index_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(index_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute(
                "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            connection.execute("create table references_index (target text, lower_target text, kind text, file_path text, line_no integer, context text)")
            connection.execute(
                "create table flow_edges (edge_kind text, from_name text, to_name text, evidence text, from_kind text, to_kind text, from_file text, from_line integer, to_file text, to_line integer)"
            )
            connection.execute("create table definitions (name text, lower_name text, kind text, file_path text, line_no integer, signature text)")
            connection.execute("create table code_entities (name text, lower_name text, kind text, file_path text, line_no integer)")
            connection.execute("insert into files values ('src/Fallback.java', 'src/fallback.java', '[\"Fallback\"]')")
            connection.execute(
                "insert into lines values ('src/Fallback.java', 1, 'class Fallback { void act() {} }', 'class fallback { void act() {} }', '[\"act\"]', 1, 0)"
            )
            connection.execute("insert into files values ('tests/NoHitTest.java', 'tests/nohittest.java', '[]')")
            connection.execute("insert into lines values ('tests/NoHitTest.java', 1, 'void testOther() {}', 'void testother() {}', '[]', 0, 0)")
            connection.execute("insert into references_index values ('plain_table', 'plain_table', 'sql_table', 'src/Fallback.java', 1, 'plain context')")
            connection.execute("insert into flow_edges values ('client', 'Unknown', '', 'empty target', 'client', 'service', 'src/Fallback.java', 1, '', 0)")
            connection.execute("insert into flow_edges values ('client', 'Unknown', 'Fallback', 'empty upstream path', 'client', 'service', '', 1, 'src/Fallback.java', 1)")
            connection.commit()

        request_cache = self.service._new_retrieval_request_cache()
        with sqlite3.connect(index_path) as connection:
            connection.row_factory = sqlite3.Row
            many_hits = {
                f"src/File{index}.java": {
                    "best_score": index,
                    "best_line": 1,
                    "path_text": f"src/file{index}.java",
                    "file_symbols": set(),
                    "structure_hits": [],
                }
                for index in range(65)
            }
            many_hits["src/Fallback.java"] = {
                "best_score": 100,
                "best_line": 1,
                "path_text": "src/fallback.java",
                "file_symbols": {"fallback"},
                "structure_hits": ["class definition Fallback"],
            }
            many_hits["src/Zero.java"] = {"best_score": 0}
            many_hits["src/Missing.java"] = {"best_score": 99, "best_line": 1, "path_text": "src/missing.java", "file_symbols": set()}
            pruned_matches = self.service._persistent_index_matches_from_hits(
                connection,
                index_path,
                entry=entry,
                tokens=["fallback"],
                question="fallback",
                focus_terms=[],
                trace_stage="direct",
                simple_intent=True,
                file_hits=many_hits,
                request_cache=request_cache,
            )
            self.assertEqual(request_cache["stats"]["simple_file_hit_prunes"], 1)
            self.assertTrue(any(match["path"] == "src/Fallback.java" for match in pruned_matches))
            self.assertNotIn("src/Zero.java", {match["path"] for match in pruned_matches})
            self.assertEqual(
                self.service._persistent_index_matches_from_hits(
                    connection,
                    index_path,
                    entry=entry,
                    tokens=["zero"],
                    question="zero",
                    focus_terms=[],
                    trace_stage="direct",
                    simple_intent=False,
                    file_hits={"src/Zero.java": {"best_score": 0}},
                    request_cache=request_cache,
                ),
                [],
            )

            self.assertIsNone(
                self.service._match_from_index_location(
                    entry,
                    connection,
                    "src/DoesNotExist.java",
                    1,
                    score=1,
                    reason="missing",
                    question="q",
                    trace_stage="direct",
                    retrieval="test",
                )
            )

        empty_index = Path(self.temp_dir.name) / "empty_targeted.sqlite3"
        with sqlite3.connect(empty_index) as connection:
            connection.row_factory = sqlite3.Row
            self.assertEqual(
                self.service._targeted_index_rows(
                    connection,
                    empty_index,
                    tokens=[],
                    focus_terms=[],
                    intent={},
                    request_cache=self.service._new_retrieval_request_cache(),
                )["files"],
                [],
            )

        with patch.object(self.service, "_choose_next_tool_step", return_value=None), patch.object(
            self.service,
            "_compress_evidence_cached",
            return_value={},
        ), patch.object(self.service, "_quality_gate_cached", return_value={}):
            trace = []
            self.assertEqual(
                self.service._run_planner_tool_loop(
                    entries=[entry],
                    key="AF:All",
                    question="q",
                    base_matches=[],
                    limit=3,
                    tool_trace=trace,
                    request_cache=self.service._new_retrieval_request_cache(),
                ),
                [],
            )
            self.assertEqual(trace[0]["reason"], "no_next_tool")

        with patch.object(self.service, "_require_ready_repo_index", return_value={"state": "ready"}), patch.object(
            self.service,
            "_static_qa_findings_for_line",
            return_value=[
                {"kind": "duplicate", "severity": "medium", "reason": "first", "score": 10},
                {"kind": "duplicate", "severity": "medium", "reason": "second", "score": 10},
            ],
        ):
            static_matches = self.service._tool_find_static_findings(entry, repo_path, ["duplicate"], "q", 0)
            self.assertEqual(len([match for match in static_matches if match["path"] == "src/Fallback.java"]), 1)

        with patch.object(self.service, "_require_ready_repo_index", return_value={"state": "ready"}):
            self.assertEqual(self.service._tool_find_test_coverage(entry, repo_path, ["coverage"], "", 1), [])
            self.assertEqual(self.service._tool_find_test_coverage(entry, repo_path, ["MissingSubject"], "MissingSubject", 1), [])

            class DuplicateItems(dict):
                def __bool__(self):
                    return True

                def items(self):
                    return [
                        ("tests/DuplicateTest.java", ["void testSubject() { assertTrue(subject); }"]),
                        ("tests/DuplicateTest.java", ["void testSubject() { assertTrue(subject); }"]),
                    ]

            with patch.object(self.service, "_cached_index_rows", return_value={"lines_by_path": DuplicateItems()}), patch.object(
                self.service,
                "_match_from_index_location",
                return_value={"path": "tests/DuplicateTest.java", "score": 1},
            ):
                duplicate_test_matches = self.service._tool_find_test_coverage(entry, repo_path, ["subject"], "subject", 1)
            self.assertEqual(len(duplicate_test_matches), 1)
            self.assertEqual(duplicate_test_matches[0]["path"], "tests/DuplicateTest.java")
            self.assertEqual(
                self.service._tool_lookup_structure(
                    entry,
                    repo_path,
                    ["ab"],
                    question="q",
                    table="references_index",
                    name_column="target",
                    lower_column="lower_target",
                    line_column="line_no",
                    kind_column="kind",
                    trace_stage="tool_loop_1",
                    retrieval="planner_reference",
                ),
                [],
            )
            self.assertEqual(
                self.service._execute_tool_loop_step(
                    entries=[entry],
                    key="AF:All",
                    question="q",
                    matches=[{"repo": entry.display_name, "path": "src/Fallback.java"}],
                    step={"tool": "trace_graph", "terms": ["Fallback"]},
                    step_index=1,
                    request_cache=self.service._new_retrieval_request_cache(),
                ),
                [],
            )
        with patch.object(self.service, "_require_ready_repo_index", side_effect=sqlite3.Error("broken")):
            self.assertEqual(self.service._tool_lookup_flow_edges(entry, repo_path, terms=["x"], seed_paths=["src/Fallback.java"], direction="callers", question="q", step_index=1), [])
            self.assertEqual(self.service._tool_open_file_window(entry, repo_path, [{"repo": entry.display_name, "path": "src/Fallback.java"}], "q", 1), [])
            self.assertEqual(self.service._tool_trace_flow(entry, repo_path, [{"repo": entry.display_name, "path": "src/Fallback.java"}], "q", 1), [])
            self.assertEqual(self.service._tool_trace_entity(entry, repo_path, [{"repo": entry.display_name, "path": "src/Fallback.java"}], "q", 1), [])
            self.assertEqual(
                self.service._expand_impact_matches(
                    entries=[entry],
                    key="AF:All",
                    question="Fallback",
                    base_matches=[{"repo": entry.display_name, "path": "src/Fallback.java"}],
                    limit=3,
                ),
                [],
            )
        with patch.object(self.service, "_require_ready_repo_index", return_value={"state": "ready"}):
            impact_matches = self.service._expand_impact_matches(
                entries=[entry],
                key="AF:All",
                question="Fallback",
                base_matches=[{"repo": entry.display_name, "path": "src/Fallback.java"}],
                limit=3,
                request_cache=self.service._new_retrieval_request_cache(),
            )
        self.assertTrue(any(match["path"] == "src/Fallback.java" for match in impact_matches))

        no_git = RepositoryEntry("No Git", "https://git.example.com/no-git.git")
        self.assertEqual(
            self.service._expand_impact_matches(
                entries=[no_git],
                key="AF:All",
                question="q",
                base_matches=[{"repo": no_git.display_name, "path": "src/App.java"}],
                limit=3,
            ),
            [],
        )
        self.assertEqual(
            self.service._expand_impact_matches(entries=[entry], key="AF:All", question="q", base_matches=[], limit=3),
            [],
        )
        self.assertEqual(
            self.service._build_trace_paths(entries=[entry], key="AF:All", matches=[], question="q", request_cache=request_cache),
            [],
        )

    def test_sqlite_exception_fallbacks_for_targeted_and_exact_lookup(self):
        entry = RepositoryEntry(display_name="Repo", url="https://git.example.com/team/repo.git")
        repo_path = self.service._repo_path("AF:All", entry)
        (repo_path / ".git").mkdir(parents=True)
        index_path = self.service._index_path(repo_path)
        index_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(index_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute(
                "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            connection.execute("create table references_index (target text, lower_target text, kind text, file_path text, line_no integer, context text)")
            connection.execute("create virtual table lines_fts using fts5(file_path, line_no unindexed, line_text)")
            connection.execute("insert into files values ('src/Exact.java', 'src/exact.java', '[]')")
            connection.execute("insert into lines values ('src/Exact.java', 1, 'select exact_table', 'select exact_table', '[]', 0, 0)")
            connection.execute("insert into references_index values ('missing_file_table', 'missing_file_table', 'sql_table', 'src/Missing.java', 1, 'select')")
            connection.commit()

        with patch.object(self.service, "_require_ready_repo_index", return_value={"state": "ready"}):
            misses = self.service._exact_table_path_lookup_repo(
                entry,
                repo_path,
                ["missing_file_table", "not_found_table"],
                question="q",
                request_cache=self.service._new_retrieval_request_cache(),
            )
        self.assertFalse(any(match.get("path") == "src/Missing.java" for match in misses))
        with patch.object(self.service, "_require_ready_repo_index", side_effect=sqlite3.Error("broken")):
            self.assertEqual(self.service._exact_table_path_lookup_repo(entry, repo_path, ["exact_table"], question="q"), [])
        with patch.object(self.service, "_require_ready_repo_index", return_value={"state": "ready"}):
            empty_term_misses = self.service._exact_table_path_lookup_repo(
                entry,
                repo_path,
                ["", "not_found_table"],
                question="q",
                request_cache=self.service._new_retrieval_request_cache(),
            )
        self.assertEqual(empty_term_misses, [])

        no_fts_entry = RepositoryEntry(display_name="No FTS", url="https://git.example.com/team/no-fts.git")
        no_fts_repo_path = self.service._repo_path("AF:All", no_fts_entry)
        (no_fts_repo_path / ".git").mkdir(parents=True)
        no_fts_index_path = self.service._index_path(no_fts_repo_path)
        no_fts_index_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(no_fts_index_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute(
                "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            connection.execute("create table references_index (target text, lower_target text, kind text, file_path text, line_no integer, context text)")
            connection.commit()
        with patch.object(self.service, "_require_ready_repo_index", return_value={"state": "ready"}):
            self.assertEqual(
                self.service._exact_table_path_lookup_repo(
                    no_fts_entry,
                    no_fts_repo_path,
                    ["not_found_table"],
                    question="q",
                    request_cache=self.service._new_retrieval_request_cache(),
                ),
                [],
            )

        no_tables_path = Path(self.temp_dir.name) / "targeted_no_tables.sqlite3"
        with sqlite3.connect(no_tables_path) as connection:
            connection.row_factory = sqlite3.Row
            with patch.object(self.service, "_file_fts_search_rows", return_value=[{"path": "src/Missing.java"}]), patch.object(
                self.service,
                "_fts_search_rows",
                return_value=[{"file_path": "src/Missing.java", "line_no": 1}],
            ), patch.object(
                self.service,
                "_cached_structure_like_rows",
                return_value=[{"file_path": "src/Missing.java", "line_no": 1}],
            ):
                payload = self.service._targeted_index_rows(
                    connection,
                    no_tables_path,
                    tokens=["needle"],
                    focus_terms=[],
                    intent={"config": True},
                    request_cache=self.service._new_retrieval_request_cache(),
                )
        self.assertEqual(payload["files"], [])
        self.assertEqual(payload["lines"], [])

        line_token_only_path = Path(self.temp_dir.name) / "targeted_line_token_only.sqlite3"
        with sqlite3.connect(line_token_only_path) as connection:
            connection.execute(
                "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            connection.execute("create table line_tokens (token text, file_path text, line_no integer)")
            connection.execute("insert into lines values ('src/OnlyLine.java', 1, 'needle', 'needle', '[]', 0, 0)")
            connection.execute("insert into line_tokens values ('needle', 'src/OnlyLine.java', 1)")
            connection.commit()
            connection.row_factory = sqlite3.Row
            line_only_payload = self.service._targeted_index_rows(
                connection,
                line_token_only_path,
                tokens=["needle"],
                focus_terms=[],
                intent={},
                request_cache=self.service._new_retrieval_request_cache(),
            )
        self.assertEqual(line_only_payload["files"], [])

        files_only_path = Path(self.temp_dir.name) / "targeted_files_only.sqlite3"
        with sqlite3.connect(files_only_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute("insert into files values ('src/OnlyFile.java', 'src/onlyfile.java', '[]')")
            connection.commit()
            connection.row_factory = sqlite3.Row
            with patch.object(self.service, "_file_fts_search_rows", return_value=[{"path": "src/OnlyFile.java"}]):
                files_only_payload = self.service._targeted_index_rows(
                    connection,
                    files_only_path,
                    tokens=["onlyfile"],
                    focus_terms=[],
                    intent={},
                    request_cache=self.service._new_retrieval_request_cache(),
                )
        self.assertEqual(len(files_only_payload["files"]), 1)

        semantic_path = Path(self.temp_dir.name) / "semantic_no_tables.sqlite3"
        with sqlite3.connect(semantic_path) as connection:
            connection.row_factory = sqlite3.Row
            with patch.object(self.service, "_semantic_fts_search_rows", return_value=[{"chunk_id": "missing"}]):
                semantic_rows = self.service._targeted_semantic_rows_by_id(
                    connection,
                    semantic_path,
                    tokens=["needle"],
                    focus_terms=[],
                    query_terms=["needle"],
                    file_paths=["src/Missing.java"],
                    simple_intent=False,
                    max_target_semantic_chunks=1,
                    request_cache=self.service._new_retrieval_request_cache(),
                )
        self.assertEqual(semantic_rows, {})

        semantic_limited_path = Path(self.temp_dir.name) / "semantic_limited.sqlite3"
        with sqlite3.connect(semantic_limited_path) as connection:
            connection.execute(
                "create table semantic_chunks (chunk_id text, file_path text, start_line integer, end_line integer, chunk_text text, lower_text text, tokens text, symbols text)"
            )
            connection.execute("insert into semantic_chunks values ('c1', 'a.java', 1, 1, 'needle', 'needle', '[]', '[]')")
            connection.execute("insert into semantic_chunks values ('c2', 'b.java', 1, 1, 'needle', 'needle', '[]', '[]')")
            connection.execute("create table semantic_chunk_tokens (token text, chunk_id text, file_path text)")
            connection.execute("insert into semantic_chunk_tokens values ('needle', 'c1', 'a.java')")
            connection.execute("insert into semantic_chunk_tokens values ('needle', 'c2', 'b.java')")
            connection.commit()
            connection.row_factory = sqlite3.Row
            with patch.object(self.service, "_semantic_fts_search_rows", return_value=[{"chunk_id": "c1"}, {"chunk_id": "c2"}]):
                semantic_limited = self.service._targeted_semantic_rows_by_id(
                    connection,
                    semantic_limited_path,
                    tokens=["needle"],
                    focus_terms=[],
                    query_terms=["needle"],
                    file_paths=["a.java", "b.java"],
                    simple_intent=False,
                    max_target_semantic_chunks=1,
                    request_cache=self.service._new_retrieval_request_cache(),
                )
            semantic_from_tokens = self.service._targeted_semantic_rows_by_id(
                connection,
                semantic_limited_path,
                tokens=[],
                focus_terms=[],
                query_terms=["needle"],
                file_paths=[],
                simple_intent=False,
                max_target_semantic_chunks=1,
                request_cache=self.service._new_retrieval_request_cache(),
            )
        semantic_fallback_path = Path(self.temp_dir.name) / "semantic_fallback_limited.sqlite3"
        with sqlite3.connect(semantic_fallback_path) as connection:
            connection.execute(
                "create table semantic_chunks (chunk_id text, file_path text, start_line integer, end_line integer, chunk_text text, lower_text text, tokens text, symbols text)"
            )
            connection.execute("insert into semantic_chunks values ('f1', 'a.java', 1, 1, 'needle', 'needle', '[]', '[]')")
            connection.execute("insert into semantic_chunks values ('f2', 'b.java', 1, 1, 'other', 'other', '[]', '[]')")
            connection.commit()
            connection.row_factory = sqlite3.Row
            semantic_from_fallback = self.service._targeted_semantic_rows_by_id(
                connection,
                semantic_fallback_path,
                tokens=[],
                focus_terms=[],
                query_terms=["needle", "other"],
                file_paths=[],
                simple_intent=False,
                max_target_semantic_chunks=1,
                request_cache=self.service._new_retrieval_request_cache(),
            )
        self.assertEqual(len(semantic_limited), 1)
        self.assertEqual(len(semantic_from_tokens), 1)
        self.assertEqual(len(semantic_from_fallback), 1)

    def test_search_repo_index_fallback_rows_and_targeted_limits(self):
        entry = RepositoryEntry(display_name="Repo", url="https://git.example.com/team/repo.git")
        repo_path = self.service._repo_path("AF:All", entry)
        (repo_path / ".git").mkdir(parents=True)
        index_path = self.service._index_path(repo_path)
        index_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(index_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute(
                "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            for index in range(200):
                path = f"src/Limit{index}.java"
                connection.execute("insert into files values (?, ?, ?)", (path, path.lower(), "[]"))
                connection.execute("insert into lines values (?, 1, 'class Limit {}', 'class limit {}', '[]', 1, 0)", (path,))
            connection.execute("insert into files values ('src/SymbolOnly.java', 'src/symbolonly.java', '[\"act\"]')")
            connection.execute("insert into lines values ('src/SymbolOnly.java', 1, 'class SymbolOnly {}', 'class symbolonly {}', '[]', 1, 0)")
            connection.execute("insert into files values ('src/LineOnly.java', 'src/lineonly.java', '[]')")
            connection.execute("insert into lines values ('src/LineOnly.java', 1, 'void run() { act(); }', 'void run() { act(); }', '[\"act\"]', 0, 0)")
            connection.commit()

        request_cache = self.service._new_retrieval_request_cache()
        with sqlite3.connect(index_path) as connection:
            connection.row_factory = sqlite3.Row
            all_files = connection.execute("select * from files order by path").fetchall()
            all_lines = connection.execute("select * from lines order by file_path, line_no").fetchall()
            files_by_path = {str(row["path"]): row for row in all_files}

        with patch.object(self.service, "_question_retrieval_features", return_value={"intent": {}}), patch.object(
            self.service,
            "_structure_lookup_query_terms",
            return_value=["ab", "act"],
        ), patch.object(
            self.service,
            "_targeted_index_rows",
            return_value={
                "files": all_files,
                "files_by_path": files_by_path,
                "lines": all_lines,
                "semantic_chunks": [],
            },
        ), patch.object(self.service, "_cached_structure_like_rows", return_value=[]), patch.object(
            self.service,
            "_fts_search_rows",
            return_value=[],
        ), patch.object(self.service, "_semantic_chunk_matches", return_value=[]):
            matches = self.service._search_repo_index(
                entry,
                repo_path,
                ["act"],
                question="where is act",
                focus_terms=[],
                request_cache=request_cache,
            )

        self.assertTrue(any(match["path"] == "src/SymbolOnly.java" for match in matches))
        self.assertTrue(any(match["path"] == "src/LineOnly.java" for match in matches))

        orphan_line = dict(all_lines[0])
        orphan_line["file_path"] = "src/Orphan.java"
        with patch.object(self.service, "_question_retrieval_features", return_value={"intent": {}}), patch.object(
            self.service,
            "_structure_lookup_query_terms",
            return_value=[],
        ), patch.object(
            self.service,
            "_targeted_index_rows",
            return_value={
                "files": [],
                "files_by_path": {},
                "lines": [orphan_line],
                "semantic_chunks": [],
            },
        ), patch.object(self.service, "_fts_search_rows", return_value=[]), patch.object(
            self.service,
            "_semantic_chunk_matches",
            return_value=[],
        ):
            self.assertEqual(
                self.service._search_repo_index(entry, repo_path, ["act"], question="orphan", request_cache=self.service._new_retrieval_request_cache()),
                [],
            )

        with sqlite3.connect(index_path) as connection:
            connection.row_factory = sqlite3.Row
            with patch.object(
                self.service,
                "_file_fts_search_rows",
                return_value=[{"path": f"src/Limit{index}.java"} for index in range(200)],
            ), patch.object(
                self.service,
                "_fts_search_rows",
                return_value=[{"file_path": f"src/Limit{index}.java", "line_no": 1} for index in range(200)],
            ), patch.object(self.service, "_targeted_semantic_rows_by_id", return_value={}):
                limited = self.service._targeted_index_rows(
                    connection,
                    index_path,
                    tokens=["limit"],
                    focus_terms=[],
                    intent={"api": True},
                    request_cache=self.service._new_retrieval_request_cache(),
                )

        self.assertLessEqual(len(limited["files"]), 48)
        self.assertLessEqual(len(limited["lines"]), 160)

    def test_evidence_outline_summarizes_support_and_primary_sources(self):
        outline = self.service._build_evidence_outline(
            {
                "items": [
                    {"type": "table", "source_id": "S1", "support_level": "confirmed", "claim": "issue_table"},
                    {"type": "call_chain", "source_id": "S2", "support_level": "inferred", "claim": "IssueService"},
                ],
                "confirmed_facts": ["issue_table"],
                "inferred_facts": ["IssueService"],
                "evidence_limits": ["No runtime trace."],
            },
            [
                {"repo": "Portal Repo", "path": "repository/IssueRepository.java", "line_start": 1, "line_end": 3, "retrieval": "persistent_index"},
                {"repo": "Portal Repo", "path": "service/IssueService.java", "line_start": 5, "line_end": 8, "retrieval": "code_graph"},
            ],
        )

        self.assertEqual(outline["type_counts"]["table"], 1)
        self.assertEqual(outline["support_counts"]["confirmed"], 1)
        self.assertEqual(outline["primary_sources"][0]["evidence_items"], 1)
        self.assertEqual(outline["evidence_limits"], ["No runtime trace."])

    def test_match_from_index_location_uses_cached_lines_without_changing_snippet(self):
        entry = RepositoryEntry(display_name="Portal Repo", url="https://git.example.com/team/portal.git")
        index_path = Path(self.temp_dir.name) / "index.sqlite3"
        with sqlite3.connect(index_path) as connection:
            connection.execute("create table files (path text, lower_path text, symbols text)")
            connection.execute(
                "create table lines (file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            connection.execute(
                "create table semantic_chunks (file_path text, start_line integer, end_line integer, chunk_text text, lower_text text, tokens text, symbols text)"
            )
            connection.execute("insert into files values ('src/App.java', 'src/app.java', '[]')")
            connection.execute("insert into lines values ('src/App.java', 1, 'class App {', 'class app {', '[]', 1, 0)")
            connection.execute("insert into lines values ('src/App.java', 2, '  void run() {}', '  void run() {}', '[]', 1, 0)")
            connection.execute("insert into lines values ('src/App.java', 3, '}', '}', '[]', 0, 0)")
            connection.execute("insert into semantic_chunks values ('src/App.java', 1, 3, 'class App {}', 'class app {}', '[]', '[]')")
            connection.commit()
        request_cache = self.service._new_retrieval_request_cache()
        with sqlite3.connect(index_path) as connection:
            connection.row_factory = sqlite3.Row
            uncached = self.service._match_from_index_location(
                entry,
                connection,
                "src/App.java",
                2,
                score=100,
                reason="uncached",
                question="where is run",
                trace_stage="test",
                retrieval="test",
            )
            cached = self.service._match_from_index_location(
                entry,
                connection,
                "src/App.java",
                2,
                score=100,
                reason="cached",
                question="where is run",
                trace_stage="test",
                retrieval="test",
                index_path=index_path,
                request_cache=request_cache,
            )

        self.assertIsNotNone(uncached)
        self.assertIsNotNone(cached)
        self.assertEqual(cached["snippet"], uncached["snippet"])
        self.assertEqual(request_cache["stats"]["file_lines_misses"], 1)

    def test_trace_paths_are_reused_within_request_cache(self):
        entry = RepositoryEntry(display_name="Portal Repo", url="https://git.example.com/team/portal.git")
        key = "AF::All"
        repo_path = self.service._repo_path(key, entry)
        (repo_path / ".git").mkdir(parents=True)
        index_path = self.service._index_path(repo_path)
        index_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(index_path) as connection:
            connection.execute(
                """
                create table flow_edges (
                    edge_kind text,
                    from_file text,
                    from_line integer,
                    from_name text,
                    to_file text,
                    to_line integer,
                    to_name text,
                    evidence text
                )
                """
            )
            connection.execute(
                "insert into flow_edges values ('service', 'src/Controller.java', 10, 'Controller.handle', 'src/Service.java', 20, 'Service.run', 'controller calls service')"
            )
            connection.commit()
        matches = [
            {
                "repo": "Portal Repo",
                "path": "src/Controller.java",
                "line_start": 10,
                "line_end": 12,
                "score": 100,
                "snippet": "class Controller {}",
                "reason": "seed",
                "trace_stage": "direct",
                "retrieval": "persistent_index",
            }
        ]
        request_cache = self.service._new_retrieval_request_cache()
        with patch.object(self.service, "_require_ready_repo_index", return_value={"state": "ready"}) as mocked_ensure:
            first = self.service._build_trace_paths(
                entries=[entry],
                key=key,
                matches=matches,
                question="trace controller",
                request_cache=request_cache,
            )
            first[0]["confidence"] = -1
            second = self.service._build_trace_paths(
                entries=[entry],
                key=key,
                matches=matches,
                question="trace controller",
                request_cache=request_cache,
            )

        self.assertEqual(mocked_ensure.call_count, 1)
        self.assertGreater(second[0]["confidence"], 0)
        self.assertNotEqual(second[0]["confidence"], -1)
        self.assertEqual(request_cache["stats"]["trace_paths_hits"], 1)
        self.assertEqual(request_cache["stats"]["trace_paths_misses"], 1)

    def test_index_lock_reports_concurrent_indexing(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "bpmis" / "jira_client.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text("class BPMISClient:\n    pass\n", encoding="utf-8")
        lock_path = self.service._index_lock_path(repo_path)
        lock_path.parent.mkdir(parents=True, exist_ok=True)
        lock_path.write_text("busy", encoding="utf-8")

        with self.assertRaisesRegex(Exception, "already being indexed"):
            self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        lock_path.unlink(missing_ok=True)

    def test_index_lock_recovers_stale_lock(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "bpmis" / "jira_client.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text("class BPMISClient:\n    pass\n", encoding="utf-8")
        lock_path = self.service._index_lock_path(repo_path)
        lock_path.parent.mkdir(parents=True, exist_ok=True)
        lock_path.write_text("2000-01-01T00:00:00+00:00", encoding="utf-8")

        info = self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        self.assertEqual(info["state"], "ready")
        self.assertFalse(lock_path.exists())

    def test_indexing_low_level_helpers_cover_lock_git_and_sqlite_edges(self):
        repo_path = Path(self.temp_dir.name) / "edge-repo"
        repo_path.mkdir()

        lock_path = repo_path / "missing.lock"
        with patch.dict(os.environ, {"SOURCE_CODE_QA_INDEX_LOCK_STALE_SECONDS": "0"}, clear=False):
            self.assertFalse(self.service._index_lock_is_stale(lock_path))
        self.assertTrue(self.service._index_lock_is_stale(lock_path))

        fake_file = SimpleNamespace(stat=lambda: (_ for _ in ()).throw(OSError("gone")))
        with patch.object(self.service, "_iter_text_files", return_value=[fake_file]):
            self.assertEqual(self.service._repo_fingerprint(repo_path), {"file_count": 0, "latest_mtime_ns": 0, "total_size": 0})

        with patch("subprocess.run", side_effect=OSError("git missing")):
            self.assertFalse(self.service._repo_worktree_clean(repo_path))
            self.assertEqual(self.service._repo_git_revision(repo_path), "")
        with patch("subprocess.run", return_value=SimpleNamespace(returncode=0, stdout="")):
            self.assertTrue(self.service._repo_worktree_clean(repo_path))
        with patch("subprocess.run", return_value=SimpleNamespace(returncode=1, stdout="abc\n")):
            self.assertEqual(self.service._repo_git_revision(repo_path), "")
        with patch("subprocess.run", return_value=SimpleNamespace(returncode=0, stdout="abc123\n")):
            self.assertEqual(self.service._repo_git_revision(repo_path), "abc123")

        missing_entry = RepositoryEntry(display_name="Missing", url="https://git.example.com/missing.git")
        self.assertEqual(self.service._repo_index_info("AF:All", missing_entry, repo_path)["state"], "missing")
        bad_index = self.service._index_path(repo_path)
        bad_index.parent.mkdir(parents=True, exist_ok=True)
        bad_index.write_text("not sqlite", encoding="utf-8")
        self.assertEqual(self.service._repo_index_info("AF:All", missing_entry, repo_path)["state"], "stale")
        bad_index.unlink()
        with sqlite3.connect(bad_index) as connection:
            connection.execute("drop table if exists metadata")
            connection.execute("create table metadata (key text primary key, value text not null)")
            for key, value in {
                "version": str(CODE_INDEX_VERSION),
                "git_revision": "abc123",
                "indexed_files": "0",
                "indexed_lines": "0",
            }.items():
                connection.execute("insert into metadata values (?, ?)", (key, value))
            connection.commit()
        with patch.object(self.service, "_repo_git_revision", return_value="abc123"), patch.object(
            self.service,
            "_repo_worktree_clean",
            return_value=True,
        ):
            self.assertEqual(self.service._repo_index_info("AF:All", missing_entry, repo_path)["state"], "ready")

        old_index = repo_path / "old.sqlite3"
        with sqlite3.connect(old_index) as connection:
            connection.execute("create table metadata (key text primary key, value text not null)")
            connection.execute("insert into metadata values ('version', ?)", (str(CODE_INDEX_VERSION - 1),))
            connection.commit()
        self.assertIsNone(self.service._open_reusable_index(old_index))
        old_index.write_text("broken", encoding="utf-8")
        self.assertIsNone(self.service._open_reusable_index(old_index))
        self.assertIsNone(self.service._open_reusable_index(repo_path / "absent.sqlite3"))

        class StatFailingPath:
            def relative_to(self, _root):
                return Path("stat-failed.py")

            def stat(self):
                raise OSError("stat failed")

        class ReadFailingPath:
            def relative_to(self, _root):
                return Path("read-failed.py")

            def stat(self):
                return SimpleNamespace(st_size=10, st_mtime_ns=20)

            def read_text(self, **_kwargs):
                raise OSError("read failed")

        build_repo = Path(self.temp_dir.name) / "build-edge-repo"
        build_repo.mkdir()
        with patch.object(self.service, "_iter_text_files", return_value=[StatFailingPath(), ReadFailingPath()]):
            info = self.service._build_repo_index("AF:All", missing_entry, build_repo)
        self.assertEqual(info["files"], 0)

    def test_indexing_reuse_and_token_helpers_cover_error_fallbacks(self):
        old_db = Path(self.temp_dir.name) / "old-index.sqlite3"
        new_db = Path(self.temp_dir.name) / "new-index.sqlite3"
        with sqlite3.connect(old_db) as old_connection, sqlite3.connect(new_db) as new_connection:
            old_connection.row_factory = sqlite3.Row
            new_connection.row_factory = sqlite3.Row
            old_connection.execute(
                "create table files(path text primary key, lower_path text, size integer, mtime_ns integer, line_count integer, symbols text)"
            )
            old_connection.execute(
                "create table lines(file_path text, line_no integer, line_text text, lower_text text, symbols text, is_declaration integer, has_pathish integer)"
            )
            old_connection.execute("create table file_tokens(token text, file_path text)")
            old_connection.execute("create table line_tokens(token text, file_path text, line_no integer)")
            old_connection.execute("create table definitions(name text, lower_name text, kind text, file_path text, line_no integer, signature text)")
            old_connection.execute(
                "create table references_index(target text, lower_target text, kind text, file_path text, line_no integer, context text)"
            )
            old_connection.execute(
                "create table code_entities(entity_id text, name text, lower_name text, kind text, language text, file_path text, line_no integer, parent text, signature text)"
            )
            old_connection.execute(
                "create table entity_edges(from_entity_id text, from_file text, from_line integer, edge_kind text, to_name text, lower_to_name text, to_entity_id text, to_file text, to_line integer, evidence text)"
            )
            old_connection.execute(
                "create table semantic_chunks(chunk_id text, file_path text, start_line integer, end_line integer, chunk_text text, lower_text text, tokens text, symbols text)"
            )
            old_connection.execute("create table semantic_chunk_tokens(token text, chunk_id text, file_path text)")
            old_connection.execute(
                "insert into files values ('src/App.java', 'src/app.java', 12, 34, 1, 'not-json')"
            )
            old_connection.execute(
                "insert into lines values ('src/App.java', 1, 'class App {}', 'class app {}', '[]', 1, 0)"
            )
            old_connection.execute(
                "insert into semantic_chunks values ('chunk-1', 'src/App.java', 1, 1, 'class App {}', 'class app {}', 'bad-json', 'also-bad')"
            )
            old_connection.commit()

            self.assertEqual(
                self.service._copy_reused_index_rows(
                    None,
                    new_connection,
                    "src/App.java",
                    "select 1 where ?",
                    "insert into nowhere values (?)",
                ),
                [],
            )
            self.service._create_repo_index_schema(new_connection)
            copied = self.service._copy_unchanged_index_file(
                old_connection,
                new_connection,
                "src/App.java",
                SimpleNamespace(st_size=12, st_mtime_ns=34),
                file_fts_enabled=True,
                fts_enabled=True,
                semantic_fts_enabled=True,
            )
            self.assertEqual(copied["lines"], 1)
            self.assertEqual(copied["semantic_chunks"], 1)
            self.assertIsNone(
                self.service._copy_unchanged_index_file(
                    old_connection,
                    new_connection,
                    "missing.java",
                    SimpleNamespace(st_size=1, st_mtime_ns=1),
                    file_fts_enabled=False,
                    fts_enabled=False,
                    semantic_fts_enabled=False,
                )
            )
            self.assertIsNone(
                self.service._copy_unchanged_index_file(
                    old_connection,
                    new_connection,
                    "src/App.java",
                    SimpleNamespace(st_size=99, st_mtime_ns=34),
                    file_fts_enabled=False,
                    fts_enabled=False,
                    semantic_fts_enabled=False,
                )
            )
            row = old_connection.execute("select * from files where path = 'src/App.java'").fetchone()
            with sqlite3.connect(":memory:") as no_fts_connection:
                self.service._create_repo_index_schema(no_fts_connection)
                self.service._insert_reused_file_row(
                    no_fts_connection,
                    "src/App.java",
                    file_fts_enabled=False,
                    file_row=row,
                )
                self.assertEqual(no_fts_connection.execute("select count(*) from files_fts").fetchone()[0], 0)

        class AlwaysBrokenConnection:
            def execute(self, *_args, **_kwargs):
                raise sqlite3.Error("broken")

        with sqlite3.connect(":memory:") as broken_new, sqlite3.connect(old_db) as old_connection:
            old_connection.row_factory = sqlite3.Row
            self.assertIsNone(
                self.service._copy_unchanged_index_file(
                    old_connection,
                    broken_new,
                    "src/App.java",
                    SimpleNamespace(st_size=12, st_mtime_ns=34),
                    file_fts_enabled=False,
                    fts_enabled=False,
                    semantic_fts_enabled=False,
                )
            )
            self.assertIsNone(
                self.service._copy_unchanged_index_file(
                    old_connection,
                    AlwaysBrokenConnection(),
                    "src/App.java",
                    SimpleNamespace(st_size=12, st_mtime_ns=34),
                    file_fts_enabled=False,
                    fts_enabled=False,
                    semantic_fts_enabled=False,
                )
            )

        with sqlite3.connect(":memory:") as connection:
            connection.execute("create table file_tokens(token text, file_path text)")
            connection.execute("create table semantic_chunk_tokens(token text, chunk_id text, file_path text)")
            self.service._insert_file_tokens(connection, "src/App.java", "client gateway helper", {"SymbolName"})
            self.service._insert_semantic_chunk_tokens(
                connection,
                [("chunk-2", "src/App.java", 1, 1, "gateway helper", "", "bad-json", "also-bad")],
            )
            self.assertGreater(connection.execute("select count(*) from file_tokens").fetchone()[0], 0)
            self.assertGreater(connection.execute("select count(*) from semantic_chunk_tokens").fetchone()[0], 0)

        self.assertEqual(self.service._build_semantic_chunks("src/Empty.java", []), [])
        self.assertEqual(self.service._build_semantic_chunks("src/Blank.java", ["", "   "]), [])
        build_chunks_globals = getattr(self.service._build_semantic_chunks, "__globals__", {})
        with patch.dict(build_chunks_globals, {"range": lambda *_args: [100]}, clear=False):
            self.assertEqual(self.service._build_semantic_chunks("src/OutOfRange.java", ["line"]), [])
        self.assertIn("identifier", self.service._index_tokens_for_text("identifier client", {"client"}))
        self.assertIn("gateway", self.service._semantic_tokens("api/gateway:GatewayClient.run()"))

    def test_indexing_schema_spring_aop_and_flow_helpers_cover_fallbacks(self):
        with sqlite3.connect(":memory:") as connection:
            self.service._create_repo_index_schema(connection)
            self.assertFalse(self.service._try_create_file_fts(connection))
            self.assertFalse(self.service._try_create_fts(connection))
            self.assertFalse(self.service._try_create_semantic_fts(connection))
        with sqlite3.connect(":memory:") as connection:
            self.assertEqual(self.service._spring_config_rows(connection), [])

        with sqlite3.connect(":memory:") as connection:
            connection.execute("create table lines(file_path text, line_no integer, line_text text)")
            connection.executemany(
                "insert into lines values (?, ?, ?)",
                [
                    ("src/main/resources/application.yml", 1, "spring:"),
                    ("src/main/resources/application.yml", 2, "  profiles:"),
                    ("src/main/resources/application.yml", 3, "    active: prod,qa"),
                    ("src/main/resources/application.yml", 4, "---"),
                    ("src/main/resources/application.yml", 5, "spring:"),
                    ("src/main/resources/application.yml", 6, "  config:"),
                    ("src/main/resources/application.yml", 7, "    activate:"),
                    ("src/main/resources/application.yml", 8, "      on-profile: prod"),
                    ("src/main/resources/application.yml", 9, "issue:"),
                    ("src/main/resources/application.yml", 10, "  enabled: true"),
                    ("src/main/resources/application.yml", 11, "# ignored"),
                    ("src/main/resources/application.yml", 12, "empty.value="),
                    ("src/main/resources/application-dev.properties", 1, "issue.enabled=false"),
                ],
            )
            rows = self.service._spring_config_rows(connection)
            values = self.service._spring_config_values(connection)
            self.assertIn("prod", self.service._active_spring_profiles_from_rows(rows))
            self.assertIn("true", values["issue.enabled"])
            self.assertNotIn("false", values["issue.enabled"])
        with sqlite3.connect(":memory:") as connection, patch.object(
            SourceCodeQAService,
            "_spring_config_rows",
            return_value=[("application.yml", "issue.empty", "", "")],
        ):
            self.assertEqual(self.service._spring_config_values(connection), {})

        self.assertEqual(self.service._spring_profile_from_config_path("bootstrap-prod.yaml"), "prod")
        self.assertFalse(self.service._spring_profile_matches("", {"prod"}))
        self.assertTrue(self.service._spring_profile_matches("!dev | prod", {"prod"}))
        self.assertFalse(self.service._spring_condition_matches("missing", {}))
        self.assertTrue(self.service._spring_condition_matches("issue.enabled=<present>", {"issue.enabled": {"true"}}))
        self.assertTrue(self.service._spring_condition_matches("issue.enabled=<missing:true>", {}))
        self.assertEqual(self.service._symbol_lookup_keys("com.example.IssueService"), ["com.example.issueservice", "issueservice"])
        self.assertEqual(self.service._symbol_lookup_keys(""), [])
        self.assertEqual(self.service._aop_pointcut_expression("ref:myPointcut()", {"mypointcut": "execution(* IssueService.run(..))"}), "execution(* IssueService.run(..))")
        self.assertEqual(
            self.service._aop_execution_target_names(
                "execution(* com.example.IssueService.run(..)) || call(* *.*.save(..)) || * *..IssueRepository.find(..)",
                {"issueservice": {"PrimaryIssueService"}, "issuerepository": {"JpaIssueRepository"}},
            ),
            [
                "IssueService.run",
                "PrimaryIssueService.run",
                "save",
                "IssueRepository.find",
                "JpaIssueRepository.find",
            ],
        )
        definitions = {
            "issueservice.run": [("service/IssueService.java", 10, "IssueService.run")],
            "com.example.issueservice.run": [("service/FqIssueService.java", 11, "com.example.IssueService.run")],
            "run": [("service/Run.java", 12, "run")],
        }
        self.assertEqual(self.service._definition_matches_for_aop_target(definitions, ""), [])
        self.assertEqual(len(self.service._definition_matches_for_aop_target(definitions, "IssueService.run")), 2)
        self.assertEqual(len(self.service._definition_matches_for_aop_target(definitions, "run")), 1)
        self.assertEqual(self.service._member_call_variable("return issueService.create();", "create"), "issueService")
        self.assertEqual(self.service._qualified_variable_targets('@Qualifier("primary") final IssueService issueService;'), {"issueService": ["primary"]})
        self.assertEqual(self.service._service_like_types_from_generic("List<IssueService> Map<AuditGateway, Repo>"), ["IssueService", "AuditGateway"])
        self.assertEqual(self.service._flow_role_for_path("src/web/IssueController.java"), "controller")
        self.assertEqual(self.service._flow_role_for_path("src/IssueDao.java"), "dao")
        self.assertEqual(self.service._flow_role_for_path("src/GatewayClient.java"), "client")
        self.assertEqual(self.service._flow_role_for_path("src/application.yml"), "config")
        self.assertEqual(self.service._classify_flow_edge("implementation_call", "", "src/IssueService.java", "run"), "service")
        self.assertEqual(self.service._classify_flow_edge("implementation_call", "", "src/App.java", "run"), "implementation")
        self.assertEqual(self.service._classify_flow_edge("aop_applies_to", "", "src/IssueRepository.java", "run"), "repository")
        self.assertEqual(self.service._classify_flow_edge("aop_applies_to", "", "src/App.java", "run"), "framework")
        self.assertEqual(self.service._classify_flow_edge("unknown", "", "src/IssueController.java", "run"), "controller")
        self.assertEqual(self.service._classify_flow_edge("unknown", "", "src/App.java", "IssueGateway"), "client")
        self.assertEqual(self.service._classify_flow_edge("unknown", "", "src/App.java", "run"), "call")
        self.assertEqual(len(self.service._entity_id("a.py", "class", "A", 1)), 24)

        with sqlite3.connect(":memory:") as connection:
            self.service._create_repo_index_schema(connection)
            connection.execute(
                "insert into code_entities values ('impl-id', 'IssueServiceImpl', 'issueserviceimpl', 'class', 'java', 'IssueServiceImpl.java', 1, '', '')"
            )
            connection.execute(
                "insert into code_entities values ('caller-id', 'IssueController', 'issuecontroller', 'class', 'java', 'IssueController.java', 1, '', '')"
            )
            connection.execute(
                "insert into entity_edges values ('impl-id', 'IssueServiceImpl.java', 1, 'implements', 'IssueService', 'issueservice', '', '', 0, '')"
            )
            connection.execute(
                "insert into entity_edges values ('caller-id', 'IssueController.java', 2, 'call', 'IssueService.create', 'issueservice.create', '', '', 0, 'issueService.create()')"
            )
            connection.commit()
            self.assertEqual(self.service._build_implementation_edges(connection), 0)

    def test_structure_node_parser_and_runtime_trace_helpers_cover_edges(self):
        self.assertIsNone(self.service._tree_sitter_parser_for_language(""))
        self.assertIsNone(self.service._tree_sitter_parser_for_language("ruby"))
        self.assertEqual(self.service._tree_sitter_language_for_suffix(".jsx"), "javascript")
        self.assertEqual(self.service._tree_sitter_language_for_suffix(".unknown"), "")
        self.assertEqual(self.service._language_for_suffix(".sql"), "sql")
        self.assertEqual(self.service._language_for_suffix(".unknown"), "text")

        bad_node = SimpleNamespace(start_byte="bad", end_byte=1, start_point=("bad",))
        self.assertEqual(self.service._node_text(b"abc", bad_node), "")
        self.assertEqual(self.service._node_line(["one"], bad_node), "one")
        self.assertEqual(self.service._node_line([], SimpleNamespace(start_point=(99, 0))), "")
        self.assertEqual(self.service._node_start_line(bad_node), 1)

        child = SimpleNamespace(type="identifier", start_byte=0, end_byte=4)
        self.assertEqual(
            self.service._first_named_child_text(b"Name()", SimpleNamespace(named_children=[child]), {"identifier"}),
            "Name",
        )
        self.assertEqual(self.service._first_named_child_text(b"Name()", SimpleNamespace(named_children=[]), {"identifier"}), "")

        class FieldBrokenNode:
            type = "call"
            named_children = [child]
            start_point = (0, 0)
            start_byte = 0
            end_byte = 10

            def child_by_field_name(self, _name):
                raise RuntimeError("field unavailable")

        self.assertEqual(self.service._tree_sitter_name_for_node(b"Fallback()", FieldBrokenNode()), "Fall")
        self.assertEqual(self.service._tree_sitter_call_target(b"this.issueService.create()", FieldBrokenNode()), "issue")
        self.assertEqual(self.service._tree_sitter_type_text_for_node(b"Type field", FieldBrokenNode()), "")
        self.assertEqual(self.service._tree_sitter_string_values(b'@GetMapping("/api/a")', SimpleNamespace(start_byte=0, end_byte=21)), ["/api/a"])

        callbacks = {"definitions": [], "references": [], "entities": [], "edges": []}

        def add_definition(*args):
            callbacks["definitions"].append(args)

        def add_reference(*args):
            callbacks["references"].append(args)

        def add_entity(name, kind, line_no, signature, parent=""):
            entity_id = f"{kind}:{name}:{line_no}:{parent}"
            callbacks["entities"].append((entity_id, name, kind, line_no, signature, parent))
            return entity_id

        def add_edge(*args):
            callbacks["edges"].append(args)

        class RaisingParser:
            def parse(self, _source):
                raise RuntimeError("parse failed")

        self.service._tree_sitter_parsers["edge"] = RaisingParser()
        self.assertEqual(
            self.service._extract_tree_sitter_structure(
                relative_path="src/App.edge",
                lines=["class App {}"],
                language="edge",
                add_definition=add_definition,
                add_reference=add_reference,
                add_entity=add_entity,
                add_entity_edge=add_edge,
                file_entity_id="file-id",
            ),
            (False, "parse failed"),
        )

        class ErrorParser:
            def parse(self, _source):
                return SimpleNamespace(root_node=SimpleNamespace(has_error=True))

        self.service._tree_sitter_parsers["bad"] = ErrorParser()
        self.assertEqual(
            self.service._extract_tree_sitter_structure(
                relative_path="src/App.bad",
                lines=["class App {"],
                language="bad",
                add_definition=add_definition,
                add_reference=add_reference,
                add_entity=add_entity,
                add_entity_edge=add_edge,
                file_entity_id="file-id",
            ),
            (False, "parse error"),
        )

        self.assertFalse(self.service._is_runtime_trace_file("notes.txt"))
        self.assertTrue(self.service._is_runtime_trace_file("runtime-traces/trace.jsonl"))
        self.assertEqual(self.service._runtime_trace_edge({"path": "/api/issues"}), ("runtime_route", "/api/issues"))
        self.assertEqual(self.service._runtime_trace_edge({"sql": "select * from issue_table"}), ("runtime_sql", "issue_table"))
        self.assertEqual(self.service._runtime_trace_edge({"topic": "issue.created"}), ("runtime_message", "issue.created"))
        self.assertEqual(self.service._runtime_trace_edge({"key": "issue.enabled"}), ("runtime_config", "issue.enabled"))
        self.assertEqual(self.service._runtime_trace_edge({"operation": "IssueService.create"}), ("runtime_call", "IssueService.create"))
        self.assertEqual(self.service._runtime_trace_string({"payload": {"a": 1}}, ("payload",)), '{"a": 1}')
        self.assertEqual(self.service._runtime_trace_string({"empty": "   ", "none": None}, ("empty", "none")), "")
        self.assertIn('"kind": "unknown"', self.service._runtime_trace_evidence({"kind": "unknown"}))

        refs: list[tuple[str, str, int, str]] = []
        edges: list[tuple[str, str, str, int, str]] = []
        self.service._extract_runtime_trace_structure(
            relative_path="source-code-qa-traces/events.jsonl",
            lines=[
                "",
                "not json",
                "[]",
                json.dumps({"kind": "http_request", "path": "/api/issues", "from": "Controller", "trace_id": "t1"}),
            ],
            add_reference=lambda target, kind, line_no, context: refs.append((target, kind, line_no, context)),
            add_entity_edge=lambda entity_id, kind, target, line_no, context: edges.append((entity_id, kind, target, line_no, context)),
            file_entity_id="file-id",
        )
        self.assertEqual(refs, [("/api/issues", "runtime_route", 4, "from=Controller | to=/api/issues | t1")])
        self.assertEqual(edges[0][1:4], ("runtime_route", "/api/issues", 4))

    def test_structure_parser_loader_and_callback_guards_cover_edges(self):
        class FakeLanguage:
            def __init__(self, grammar):
                self.grammar = grammar

        class FakeParser:
            def __init__(self, language):
                self.language = language

        fake_tree_sitter = SimpleNamespace(Language=FakeLanguage, Parser=FakeParser)
        fake_javascript = SimpleNamespace(language=lambda: "javascript-grammar")
        fake_typescript = SimpleNamespace(
            language_typescript=lambda: "typescript-grammar",
            language_tsx=lambda: "tsx-grammar",
        )
        with patch.dict(
            sys.modules,
            {
                "tree_sitter": fake_tree_sitter,
                "tree_sitter_javascript": fake_javascript,
                "tree_sitter_typescript": fake_typescript,
            },
        ):
            self.service._tree_sitter_parsers.pop("javascript", None)
            self.service._tree_sitter_parsers.pop("tsx", None)
            javascript_parser = self.service._tree_sitter_parser_for_language("javascript")
            tsx_parser = self.service._tree_sitter_parser_for_language("tsx")

        self.assertEqual(javascript_parser.language.grammar, "javascript-grammar")
        self.assertEqual(tsx_parser.language.grammar, "tsx-grammar")

        broken_javascript = SimpleNamespace(language=lambda: (_ for _ in ()).throw(RuntimeError("missing grammar")))
        with patch.dict(
            sys.modules,
            {"tree_sitter": fake_tree_sitter, "tree_sitter_javascript": broken_javascript},
        ):
            self.service._tree_sitter_parsers.pop("javascript", None)
            self.assertIsNone(self.service._tree_sitter_parser_for_language("javascript"))

        self.assertIn("missing grammar", self.service._tree_sitter_load_errors["javascript"])

        def exercise_callbacks(*, add_definition, add_reference, add_entity, add_entity_edge, **_kwargs):
            add_definition("", "empty", 1, "")
            add_reference("x", "too_short", 2, "")
            add_entity("", "empty_entity", 3, "")
            add_entity_edge("", "too_short_edge", "x", 4, "")

        with patch.object(self.service, "_extract_python_ast_structure", side_effect=exercise_callbacks):
            rows = self.service._extract_structure_rows("guard.py", ["print('guard')"])

        self.assertEqual(rows["definitions"], [])
        self.assertFalse(any(row[4] == "x" for row in rows["entity_edges"]))

        structure_globals = self.service._extract_structure_rows.__func__.__globals__
        fake_mybatis_attr_pattern = SimpleNamespace(findall=lambda _line: [("type", "")])
        fake_generic_pattern = SimpleNamespace(findall=lambda _line: [("List<IssueService>", "issueService")])
        with patch.dict(
            structure_globals,
            {
                "MYBATIS_ATTR_REFERENCE_PATTERN": fake_mybatis_attr_pattern,
                "GENERIC_FIELD_VAR_TYPE_PATTERN": fake_generic_pattern,
            },
        ):
            guarded_rows = self.service._extract_structure_rows(
                "mapper/IssueMapper.xml",
                ["<select id=\"find\" resultType=\"Issue\">", "IssueService issueService"],
            )

        self.assertFalse(any(row[2] == "mybatis_type_ref" and row[0] == "" for row in guarded_rows["references"]))
        self.assertIn(
            ("IssueService", "issueservice", "field_type", "mapper/IssueMapper.xml", 2, "IssueService issueService"),
            guarded_rows["references"],
        )

    def test_structure_static_extractors_and_build_file_edges(self):
        self.assertIsNone(self.service._extract_config_assignment("# comment"))
        self.assertIsNone(self.service._extract_config_assignment("not assignment"))
        self.assertIsNone(self.service._extract_config_assignment("empty.value="))
        self.assertIsNone(self.service._extract_config_assignment('empty.value=""'))
        self.assertEqual(self.service._extract_config_assignment("issue.enabled=true"), ("issue.enabled", "true"))
        yaml_stack: list[tuple[int, str]] = []
        self.assertIsNone(self.service._extract_yaml_config_assignment("# ignored", yaml_stack))
        self.assertEqual(self.service._extract_yaml_config_assignment("spring:", yaml_stack), ("spring", ""))
        self.assertEqual(self.service._extract_yaml_config_assignment("  profiles:", yaml_stack), ("spring.profiles", ""))
        self.assertEqual(self.service._extract_yaml_config_assignment("    active: prod # comment", yaml_stack), ("spring.profiles.active", "prod"))

        class BlankYamlKeyMatch:
            def group(self, index):
                return {1: "", 2: "", 3: "value"}[index]

        with patch("bpmis_jira_tool.source_code_qa_structure.re.match", return_value=BlankYamlKeyMatch()):
            self.assertIsNone(self.service._extract_yaml_config_assignment("blank: value", []))

        self.assertEqual(self.service._spring_annotation_arg_values('name={"a","b"}, matchIfMissing=true', "name"), ["a", "b"])
        self.assertEqual(
            self.service._spring_conditional_on_property_entries('prefix="issue", name={"enabled","ready"}, havingValue="true", matchIfMissing=true'),
            ["issue.enabled=true", "issue.enabled=<missing:true>", "issue.ready=true", "issue.ready=<missing:true>"],
        )
        self.assertEqual(self.service._annotation_target_text('pointcut = "execution(* run())"'), "execution(* run())")
        self.assertEqual(self.service._annotation_target_text("value = issuePointcut()"), "issuePointcut()")
        self.assertEqual(self.service._scheduled_target_text(""), "scheduled")
        self.assertEqual(self.service._scheduled_target_text('fixedRate=5000, initialDelayString="PT1S"'), "initialDelayString=PT1S;fixedRate=5000")
        self.assertEqual(self.service._extract_message_names('"issue.created", "${topic.name:default}", "plain"'), ["issue.created", "${topic.name:default}", "topic.name"])
        self.assertEqual(
            self.service._extract_event_names('publisher.publishEvent(new IssueCreatedEvent()); @EventListener(IssueCommand.class); "custom-message"'),
            ["IssueCreatedEvent", "IssueCommand", "custom-message"],
        )
        self.assertEqual(self.service._runtime_trace_sql_target({"sql": "nonsense sql"}), "nonsense sql")
        self.assertEqual(self.service._spring_conditional_on_property_entries('prefix="issue", name=., havingValue=true'), [])
        self.assertEqual(self.service._normalize_gradle_module_name(":risk:api"), "risk-api")
        self.assertEqual(self.service._first_line_number_containing(["abc", "needle"], "needle"), 2)
        self.assertEqual(self.service._first_line_number_containing(["abc"], ""), 1)

        refs: list[tuple[str, str, int, str]] = []
        defs: list[tuple[str, str, int, str]] = []
        edges: list[tuple[str, str, str, int, str]] = []

        def add_definition(target, kind, line_no, context):
            defs.append((target, kind, line_no, context))

        def add_reference(target, kind, line_no, context):
            refs.append((target, kind, line_no, context))

        def add_edge(entity_id, kind, target, line_no, context):
            edges.append((entity_id, kind, target, line_no, context))

        self.service._extract_build_file_structure(
            relative_path="package.json",
            lines=['{"name":"portal","dependencies":{"axios":"1.0.0"},"devDependencies":"","peerDependencies":{"": "x"}}'],
            add_definition=add_definition,
            add_reference=add_reference,
            add_entity_edge=add_edge,
            file_entity_id="file-id",
        )
        self.service._extract_build_file_structure(
            relative_path="package.json",
            lines=["{bad json"],
            add_definition=add_definition,
            add_reference=add_reference,
            add_entity_edge=add_edge,
            file_entity_id="file-id",
        )
        self.service._extract_build_file_structure(
            relative_path="pom.xml",
            lines=[
                "<project><groupId>com.example</groupId><artifactId>portal</artifactId><dependencies>",
                "<dependency><groupId>org.demo</groupId><artifactId>demo-client</artifactId></dependency>",
                "<dependency><groupId>${skip}</groupId><artifactId>${skip}</artifactId></dependency>",
                "</dependencies></project>",
            ],
            add_definition=add_definition,
            add_reference=add_reference,
            add_entity_edge=add_edge,
            file_entity_id="file-id",
        )
        self.service._extract_build_file_structure(
            relative_path="settings.gradle",
            lines=[
                "// ignored",
                "include ':risk:api'",
                "implementation 'org.demo:demo-client:1.0.0'",
                "implementation project(':risk:worker')",
            ],
            add_definition=add_definition,
            add_reference=add_reference,
            add_entity_edge=add_edge,
            file_entity_id="file-id",
        )
        self.assertIn(("portal", "npm_package", 1, "portal"), defs)
        self.assertTrue(any(item[0] == "axios" and item[1] == "module_dependency" for item in refs))
        self.assertTrue(any(item[0] == "com.example:portal" and item[1] == "maven_coordinate" for item in defs))
        self.assertTrue(any(item[0] == "org.demo:demo-client" and item[1] == "module_dependency" for item in refs))
        self.assertTrue(any(item[0] == "risk-api" and item[1] == "gradle_module" for item in defs))
        self.assertTrue(any(item[0] == "risk-worker" and item[1] == "gradle_project_dependency" for item in refs))

    def test_structure_rows_cover_inline_spring_runtime_config_and_events(self):
        java_source = (
            "public class InlineAnnotations {\n"
            "    private IssueService issueService;\n"
            "    public void run() { @Transactional @Before(\"execution(* com.example.IssueService.create(..))\") @Scheduled(fixedRate=100) issueService.create(); }\n"
            "    private @Component(\"inlineBean\") @Primary @Aspect @Profile(\"uat\") @ConditionalOnProperty(prefix=\"feature\", name=\"enabled\") String marker;\n"
            "    public void events(ApplicationEventPublisher publisher) { publisher.publishEvent(new IssueCreatedEvent()); }\n"
            "    @EventListener(IssueCreatedEvent.class)\n"
            "    public void consume(IssueCreatedEvent event) { }\n"
            "}\n"
        )
        java_rows = self.service._extract_structure_rows("src/InlineAnnotations.java", java_source.splitlines())
        java_refs = {(row[2], row[0]) for row in java_rows["references"]}
        java_edges = {(row[3], row[4]) for row in java_rows["entity_edges"]}
        self.assertIn(("bean_name", "inlineBean"), java_refs)
        self.assertIn(("bean_primary", "InlineAnnotations"), java_refs)
        self.assertIn(("framework_binding", "Aspect"), java_refs)
        self.assertIn(("spring_profile", "uat"), java_refs)
        self.assertIn(("bean_condition", "feature.enabled=<present>"), java_refs)
        self.assertIn(("operational_boundary", "Transactional"), java_edges)
        self.assertTrue(any(kind == "aop_advice" for kind, _target in java_edges))
        self.assertIn(("scheduled_job", "fixedRate=100"), java_edges)
        self.assertIn(("event_publish", "IssueCreatedEvent"), java_refs)
        self.assertIn(("event_consume", "IssueCreatedEvent"), java_refs)

        generic_skip_rows = self.service._extract_structure_rows(
            "src/GenericSkip.java",
            ["public class GenericSkip {", "    private List<IssueService> services;", "}"],
        )
        self.assertEqual(
            [row for row in generic_skip_rows["references"] if row[2] == "field_type" and row[0] == "IssueService"],
            [("IssueService", "issueservice", "field_type", "src/GenericSkip.java", 2, "private List<IssueService> services;")],
        )

        config_rows = self.service._extract_structure_rows("application.properties", ["server:", "feature.enabled="])
        config_defs = {(row[2], row[0]) for row in config_rows["definitions"]}
        self.assertIn(("config_key", "feature.enabled"), config_defs)

        runtime_rows = self.service._extract_structure_rows(
            "source-code-qa-traces/runtime.jsonl",
            [json.dumps({"kind": "http_request"})],
        )
        self.assertEqual([row for row in runtime_rows["references"] if row[2].startswith("runtime_")], [])

        python_rows = self.service._extract_structure_rows(
            "worker/calls.py",
            [
                "class Client:",
                "    def run(self):",
                "        pkg.client.send()",
                "        (factory())()",
            ],
        )
        python_refs = {(row[2], row[0]) for row in python_rows["references"]}
        self.assertIn(("call", "pkg.client.send"), python_refs)

    def test_structure_rows_cover_pending_spring_and_python_ast_edges(self):
        java_source = (
            "package com.example;\n"
            "import com.example.repo.IssueRepository;\n"
            "@Component(\"issueBean\")\n"
            "@Primary\n"
            "@Aspect\n"
            "@Profile(\"prod\")\n"
            "@ConditionalOnProperty(prefix=\"issue\", name=\"enabled\", matchIfMissing=true)\n"
            "@RequestMapping(\"/api/issues\")\n"
            "public class IssueController implements IssueApi {\n"
            "    @Qualifier(\"primaryRepo\")\n"
            "    private IssueRepository issueRepository;\n"
            "    @Transactional\n"
            "    @Scheduled(fixedDelay=1000)\n"
            "    @Before(\"execution(* com.example.IssueService.create(..))\")\n"
            "    @PostMapping(\"/create\")\n"
            "    public Issue create() {\n"
            "        issueRepository.findIssue();\n"
            "        this.primaryRepository = issueRepository;\n"
            "        return new Issue();\n"
            "    }\n"
            "}\n"
        )
        rows = self.service._extract_structure_rows("controller/IssueController.java", java_source.splitlines())
        references = {(row[2], row[0]) for row in rows["references"]}
        edges = {(row[3], row[4]) for row in rows["entity_edges"]}
        self.assertIn(("bean_name", "issueBean"), references)
        self.assertIn(("bean_primary", "IssueController"), references)
        self.assertIn(("framework_binding", "Aspect"), references)
        self.assertIn(("spring_profile", "prod"), references)
        self.assertIn(("bean_condition", "issue.enabled=<missing:true>"), references)
        self.assertIn(("route", "/api/issues/create"), references)
        self.assertIn(("bean_qualifier_target", "issueRepository=primaryRepo"), references)
        self.assertIn(("bean_qualifier_target", "primaryRepository=primaryRepo"), references)
        self.assertIn(("operational_boundary", "Transactional"), references)
        self.assertIn(("scheduled_job", "fixedDelay=1000"), references)
        self.assertTrue(any(kind == "aop_advice" for kind, _target in edges))

        generic_source = (
            "import com.example.gateway.AuditGateway;\n"
            "public class IssueService {\n"
            "    @Qualifier(\"audit\")\n"
            "    private List<AuditGateway> gateways;\n"
            "    public void run() {\n"
            "        gateways.stream().map(gateway -> gateway.sendAudit()).collect(toList());\n"
            "        gateways.get().sendAudit();\n"
            "    }\n"
            "}\n"
        )
        generic_rows = self.service._extract_structure_rows("service/IssueService.java", generic_source.splitlines())
        generic_refs = {(row[2], row[0]) for row in generic_rows["references"]}
        self.assertIn(("field_type", "AuditGateway"), generic_refs)
        self.assertIn(("field_type", "com.example.gateway.AuditGateway"), generic_refs)
        self.assertIn(("call", "AuditGateway.sendAudit"), generic_refs)
        self.assertIn(("call", "com.example.gateway.AuditGateway.sendAudit"), generic_refs)

        test_rows = self.service._extract_structure_rows(
            "src/test/java/IssueServiceTest.java",
            ["@Test", "void shouldCreateIssue() {", "  assertEquals(1, issueService.create());", "  IssueService subject = new IssueService();", "}"],
        )
        test_refs = {(row[2], row[0]) for row in test_rows["references"]}
        self.assertIn(("test_case", "test_case"), test_refs)
        self.assertIn(("test_assertion", "assertion"), test_refs)
        self.assertIn(("test_subject", "IssueService"), test_refs)

        syntax_rows = self.service._extract_structure_rows("bad.py", ["def broken(:"])
        self.assertNotIn("python_function", {row[2] for row in syntax_rows["definitions"]})
        python_rows = self.service._extract_structure_rows(
            "worker/tasks.py",
            [
                "import os",
                "from service.issue import create_issue",
                "class Worker:",
                "    async def run(self):",
                "        await create_issue()",
            ],
        )
        python_defs = {(row[2], row[0]) for row in python_rows["definitions"]}
        python_refs = {(row[2], row[0]) for row in python_rows["references"]}
        self.assertIn(("python_class", "Worker"), python_defs)
        self.assertIn(("python_async_function", "run"), python_defs)
        self.assertIn(("import", "os"), python_refs)
        self.assertIn(("import", "service.issue.create_issue"), python_refs)
        self.assertIn(("call", "create_issue"), python_refs)

    def test_source_code_qa_session_store_covers_persistence_and_turn_edges(self):
        broken_path = Path(self.temp_dir.name) / "broken-sessions.json"
        broken_path.write_text("{bad", encoding="utf-8")
        self.assertEqual(SourceCodeQASessionStore(broken_path).list(owner_email="owner@npt.sg"), [])
        valid_path = Path(self.temp_dir.name) / "valid-sessions.json"
        valid_path.write_text(json.dumps({"sessions": {"s1": {"id": "s1"}, "bad": "ignored"}}), encoding="utf-8")
        self.assertEqual(SourceCodeQASessionStore(valid_path).list(owner_email="owner@npt.sg"), [])
        with patch("bpmis_jira_tool.source_code_qa_stores.os.replace", side_effect=OSError("replace failed")):
            SourceCodeQASessionStore(valid_path).create(owner_email="owner@npt.sg", pm_team="AF", country="All", llm_provider="", title="")

        memory_store = SourceCodeQASessionStore()
        self.assertEqual(memory_store._title_from_question(""), "New Source Code Chat")
        self.assertIsNone(memory_store._recent_turn_from_context({"question": "", "answer": ""}))
        extended = memory_store._extend_recent_turns(
            {"question": "new", "answer": "new", "codex_session_max_turns": 1},
            {"question": "old", "answer": "old", "trace_id": "t", "recent_turns": [{"question": "old", "trace_id": "t"}]},
        )
        self.assertEqual(len(extended["recent_turns"]), 1)
        self.assertIsNone(memory_store.get_context("missing", owner_email="owner@npt.sg"))
        self.assertIsNone(memory_store.append_pending_question("missing", owner_email="owner@npt.sg", pm_team="AF", country="All", llm_provider="", question="", job_id=""))

        session = memory_store.create(owner_email="owner@npt.sg", pm_team="", country="", llm_provider="", title="")
        self.assertIsNone(memory_store.get(session["id"], owner_email="other@npt.sg"))
        self.assertIsNone(memory_store.archive("missing", owner_email="owner@npt.sg"))
        self.assertIsNone(memory_store.archive(session["id"], owner_email="other@npt.sg"))
        self.assertIsNone(
            memory_store.append_exchange(
                "missing",
                owner_email="owner@npt.sg",
                pm_team="AF",
                country="All",
                llm_provider="codex_cli_bridge",
                question="missing",
                result={},
                context={},
            )
        )
        self.assertIsNone(
            memory_store.append_exchange(
                session["id"],
                owner_email="other@npt.sg",
                pm_team="AF",
                country="All",
                llm_provider="codex_cli_bridge",
                question="wrong owner",
                result={},
                context={},
            )
        )
        self.assertIsNone(
            memory_store.append_pending_question(
                "missing",
                owner_email="owner@npt.sg",
                pm_team="AF",
                country="All",
                llm_provider="codex_cli_bridge",
                question="pending",
                job_id="job-1",
            )
        )
        self.assertIsNone(
            memory_store.append_pending_question(
                session["id"],
                owner_email="other@npt.sg",
                pm_team="AF",
                country="All",
                llm_provider="codex_cli_bridge",
                question="pending",
                job_id="job-1",
            )
        )
        pending = memory_store.append_pending_question(
            session["id"],
            owner_email="owner@npt.sg",
            pm_team="AF",
            country="SG",
            llm_provider="codex_cli_bridge",
            question=" pending question ",
            job_id="job-1",
            attachments=[{"id": "a1", "filename": "notes.txt", "mime_type": "text/plain", "kind": "text", "size": 1}],
        )
        self.assertEqual(pending["messages"][-1]["pending_job_id"], "job-1")
        updated = memory_store.append_exchange(
            session["id"],
            owner_email="owner@npt.sg",
            pm_team="GRC",
            country="All",
            llm_provider="codex_cli_bridge",
            question="pending question",
            result={
                "status": "ok",
                "summary": "done",
                "llm_answer": "answer",
                "trace_id": "trace-2",
                "llm_route": {"candidate_paths": [{"path": "src/App.java"}]},
                "structured_answer": {"claims": [{"claim": "c"}]},
                "attachments": [{"id": "a1", "filename": "notes.txt"}],
                "runtime_evidence": [{"id": "e1", "filename": "apollo.properties"}],
                "generated_artifacts": [{"id": "g1", "filename": "source-code-qa-sql-package.zip"}],
                "matches": [{"repo": "Repo", "path": "src/App.java", "score": 9}],
            },
            context={
                "question": "pending question",
                "rendered_answer": "rendered",
                "trace_id": "trace-2",
                "codex_cli_session": {"id": "old"},
                "codex_session_max_turns": "bad",
            },
            attachments=[{"id": "a1", "filename": "notes.txt"}],
        )

        self.assertEqual(updated["message_count"], 2)
        self.assertNotIn("codex_cli_session", updated["last_context"])
        self.assertEqual(memory_store.get_context(session["id"], owner_email="owner@npt.sg")["session_title"], "pending question")
        self.assertIsNotNone(memory_store.archive(session["id"], owner_email="owner@npt.sg"))
        self.assertEqual(memory_store.list(owner_email="owner@npt.sg"), [])

    def test_source_code_qa_attachment_store_covers_file_boundaries(self):
        store = SourceCodeQAAttachmentStore(Path(self.temp_dir.name) / "attachments")
        self.assertEqual(store._safe_filename("../bad\x00:name?.txt"), "bad_name_.txt")
        with self.assertRaisesRegex(ToolError, "valid Source Code"):
            store._safe_session_id("bad/session")
        with self.assertRaisesRegex(ToolError, "not configured"):
            SourceCodeQAAttachmentStore()._session_dir(owner_email="owner@npt.sg", session_id="session123")
        with self.assertRaisesRegex(ToolError, "empty"):
            store.save_bytes(owner_email="owner@npt.sg", session_id="session123", filename="empty.txt", content=b"")
        with self.assertRaisesRegex(ToolError, "too large"):
            store.save_bytes(owner_email="owner@npt.sg", session_id="session123", filename="large.txt", content=b"x" * (store.MAX_FILE_BYTES + 1))
        with self.assertRaisesRegex(ToolError, "Executable"):
            store.save_bytes(owner_email="owner@npt.sg", session_id="session123", filename="tool.exe", content=b"x")
        with self.assertRaisesRegex(ToolError, "Unknown binary"):
            store.save_bytes(owner_email="owner@npt.sg", session_id="session123", filename="blob", content=b"\x00binary")
        with self.assertRaisesRegex(ToolError, "Unsupported attachment"):
            store.save_bytes(owner_email="owner@npt.sg", session_id="session123", filename="blob", content=b"text")

        metadata_path = store._metadata_path(owner_email="owner@npt.sg", session_id="session123")
        metadata_path.parent.mkdir(parents=True, exist_ok=True)
        metadata_path.write_text("{bad", encoding="utf-8")
        self.assertEqual(store._load_metadata_locked(owner_email="owner@npt.sg", session_id="session123"), {})

        text_item = store.save_bytes(
            owner_email="owner@npt.sg",
            session_id="session123",
            filename="notes.txt",
            content="hello\r\nworld".encode("utf-8"),
            mime_type="text/plain",
        )
        image_items = [
            store.save_bytes(
                owner_email="owner@npt.sg",
                session_id="session123",
                filename=f"img{index}.png",
                content=b"png-bytes",
                mime_type="image/png",
            )
            for index in range(store.MAX_IMAGES + 1)
        ]
        with self.assertRaisesRegex(ToolError, "At most"):
            store.resolve_many(owner_email="owner@npt.sg", session_id="session123", attachment_ids=[item["id"] for item in image_items])
        with self.assertRaisesRegex(ToolError, "At most"):
            store.resolve_many(owner_email="owner@npt.sg", session_id="session123", attachment_ids=[str(index) for index in range(store.MAX_ATTACHMENTS + 1)])
        with self.assertRaisesRegex(ToolError, "not found"):
            store.resolve_many(owner_email="owner@npt.sg", session_id="session123", attachment_ids=["0" * 32])

        odd = store.save_bytes(
            owner_email="owner@npt.sg",
            session_id="session123",
            filename="notes.unknown",
            content=b"fallback text",
        )
        self.assertEqual(odd["kind"], "text")
        self.assertEqual(store._extract_attachment_text("raw.txt", "text/plain", b"\xfflatin")[0], "ÿ")
        with self.assertRaisesRegex(ToolError, "Unable to parse"):
            store._extract_attachment_text("raw.txt", "text/plain", b"")
        with patch.object(SourceCodeQAAttachmentStore, "_extract_pdf_text", return_value="pdf text") as pdf_extract:
            self.assertEqual(store._extract_attachment_text("file.pdf", "application/pdf", b"pdf"), "pdf text")
            pdf_extract.assert_called_once()
        with patch.object(SourceCodeQAAttachmentStore, "_extract_docx_text", return_value="docx text") as docx_extract:
            self.assertEqual(store._extract_attachment_text("file.docx", "", b"docx"), "docx text")
            docx_extract.assert_called_once()
        with patch.object(SourceCodeQAAttachmentStore, "_extract_xlsx_text", return_value="xlsx text") as xlsx_extract:
            self.assertEqual(store._extract_attachment_text("file.xlsx", "", b"xlsx"), "xlsx text")
            xlsx_extract.assert_called_once()

        metadata = store._load_metadata_locked(owner_email="owner@npt.sg", session_id="session123")
        (store._session_dir(owner_email="owner@npt.sg", session_id="session123") / metadata[text_item["id"]]["stored_name"]).unlink()
        with self.assertRaisesRegex(ToolError, "missing"):
            store.resolve_many(owner_email="owner@npt.sg", session_id="session123", attachment_ids=[text_item["id"]])
        metadata[text_item["id"]]["stored_name"] = metadata[image_items[0]["id"]]["stored_name"]
        store._persist_metadata_locked(owner_email="owner@npt.sg", session_id="session123", metadata=metadata)
        with patch("pathlib.Path.read_bytes", side_effect=OSError("unreadable")):
            with self.assertRaisesRegex(ToolError, "unreadable"):
                store.resolve_many(owner_email="owner@npt.sg", session_id="session123", attachment_ids=[text_item["id"]])

        with patch.object(SourceCodeQAAttachmentStore, "resolve_many", return_value=[]):
            with self.assertRaisesRegex(ToolError, "was not found"):
                store.get_bytes(owner_email="owner@npt.sg", session_id="session123", attachment_id=text_item["id"])
        with patch.object(SourceCodeQAAttachmentStore, "resolve_many", return_value=[{"path": str(Path(self.temp_dir.name) / "missing.txt")}]):
            with self.assertRaisesRegex(ToolError, "unreadable"):
                store.get_bytes(owner_email="owner@npt.sg", session_id="session123", attachment_id=text_item["id"])
        with patch("builtins.__import__", side_effect=ImportError("no module")):
            with self.assertRaisesRegex(ToolError, "PDF attachments"):
                store._extract_pdf_text(b"%PDF")
            with self.assertRaisesRegex(ToolError, "DOCX attachments"):
                store._extract_docx_text(b"docx")
            with self.assertRaisesRegex(ToolError, "XLSX attachments"):
                store._extract_xlsx_text(b"xlsx")
        fake_pdf_module = SimpleNamespace(PdfReader=lambda _buffer: SimpleNamespace(pages=[SimpleNamespace(extract_text=lambda: "page text")]))
        with patch.dict(sys.modules, {"pypdf": fake_pdf_module}):
            self.assertEqual(store._extract_pdf_text(b"pdf"), "page text")
        fake_empty_pdf_module = SimpleNamespace(PdfReader=lambda _buffer: SimpleNamespace(pages=[SimpleNamespace(extract_text=lambda: "")]))
        with patch.dict(sys.modules, {"pypdf": fake_empty_pdf_module}):
            with self.assertRaisesRegex(ToolError, "readable text"):
                store._extract_pdf_text(b"pdf")
        fake_docx_module = SimpleNamespace(Document=lambda _buffer: SimpleNamespace(paragraphs=[SimpleNamespace(text="doc text")]))
        with patch.dict(sys.modules, {"docx": fake_docx_module}):
            self.assertEqual(store._extract_docx_text(b"docx"), "doc text")
        fake_empty_docx_module = SimpleNamespace(Document=lambda _buffer: SimpleNamespace(paragraphs=[SimpleNamespace(text="")]))
        with patch.dict(sys.modules, {"docx": fake_empty_docx_module}):
            with self.assertRaisesRegex(ToolError, "readable text"):
                store._extract_docx_text(b"docx")
        real_import = __import__
        fake_empty_openpyxl = SimpleNamespace(load_workbook=lambda *_args, **_kwargs: SimpleNamespace(worksheets=[]))
        with patch("builtins.__import__", side_effect=lambda name, *args, **kwargs: fake_empty_openpyxl if name == "openpyxl" else real_import(name, *args, **kwargs)):
            with self.assertRaisesRegex(ToolError, "readable text"):
                store._extract_xlsx_text(b"xlsx")
        xlsx_workbook = Workbook()
        xlsx_workbook.active.append(["field", "meaning"])
        xlsx_buffer = io.BytesIO()
        xlsx_workbook.save(xlsx_buffer)
        self.assertIn("field\tmeaning", store._extract_xlsx_text(xlsx_buffer.getvalue()))

    def test_source_code_qa_runtime_evidence_store_covers_scope_zip_and_data_dictionary_edges(self):
        root = Path(self.temp_dir.name) / "runtime-evidence"
        store = SourceCodeQARuntimeEvidenceStore(root)
        with self.assertRaisesRegex(ToolError, "PM Team"):
            store._safe_scope(pm_team="OPS", country="SG")
        with self.assertRaisesRegex(ToolError, "Shared All-country"):
            store._safe_scope(pm_team="CRMS", country="All")
        with self.assertRaisesRegex(ToolError, "country"):
            store._safe_scope(pm_team="AF", country="MY")
        with self.assertRaisesRegex(ToolError, "source type"):
            store._safe_source_type("secret")
        with self.assertRaisesRegex(ToolError, "not configured"):
            SourceCodeQARuntimeEvidenceStore()._scope_dir(pm_team="AF", country="SG")
        with self.assertRaisesRegex(ToolError, "empty"):
            store.save_bytes(pm_team="AF", country="SG", source_type="apollo", uploaded_by="owner", filename="empty.txt", content=b"")
        with self.assertRaisesRegex(ToolError, "too large"):
            store.save_bytes(pm_team="AF", country="SG", source_type="apollo", uploaded_by="owner", filename="large.txt", content=b"x" * (store.MAX_FILE_BYTES + 1))
        with self.assertRaisesRegex(ToolError, "not an image"):
            store.save_bytes(pm_team="AF", country="SG", source_type="apollo", uploaded_by="owner", filename="img.png", content=b"png", mime_type="image/png")
        with self.assertRaisesRegex(ToolError, "Executable"):
            store.save_bytes(pm_team="AF", country="SG", source_type="apollo", uploaded_by="owner", filename="archive.jar", content=b"jar")

        scope_dir = store._scope_dir(pm_team="AF", country="SG")
        scope_dir.mkdir(parents=True, exist_ok=True)
        (scope_dir / "metadata.json").write_text("{bad", encoding="utf-8")
        self.assertEqual(store.list(pm_team="AF", country="SG"), [])

        with patch.object(SourceCodeQARuntimeEvidenceStore, "MAX_FILES_PER_SCOPE", 1), patch("pathlib.Path.unlink", side_effect=OSError("unlink failed")):
            stale = store.save_bytes(
                pm_team="AF",
                country="SG",
                source_type="db",
                uploaded_by="owner@npt.sg",
                filename="stale.txt",
                content=b"stale",
                mime_type="text/plain",
            )
            latest = store.save_bytes(
                pm_team="AF",
                country="SG",
                source_type="db",
                uploaded_by="owner@npt.sg",
                filename="latest.txt",
                content=b"latest",
                mime_type="text/plain",
            )
        self.assertEqual([item["id"] for item in store.list(pm_team="AF", country="SG")], [latest["id"]])
        self.assertNotEqual(stale["id"], latest["id"])
        self.assertTrue(store.delete(pm_team="AF", country="SG", evidence_id=latest["id"]))

        first = store.save_bytes(
            pm_team="AF",
            country="All",
            source_type="apollo",
            uploaded_by="owner@npt.sg",
            filename="apollo.properties",
            content=b"feature.enabled=true",
            mime_type="text/plain",
        )
        second = store.save_bytes(
            pm_team="AF",
            country="SG",
            source_type="db",
            uploaded_by="owner@npt.sg",
            filename="rules.csv",
            content=b"rule,status\nA,on",
            mime_type="text/csv",
        )
        resolved = store.resolve_scope(pm_team="AF", country="SG")
        self.assertEqual({item["id"] for item in resolved}, {first["id"], second["id"]})
        self.assertEqual(store.resolve_scope(pm_team="CRMS", country=""), [])

        missing_metadata = store._load_metadata_locked(pm_team="AF", country="SG")
        missing_metadata[first["id"]] = dict(store._load_metadata_locked(pm_team="AF", country="All")[first["id"]])
        missing_metadata[second["id"]]["stored_name"] = "missing.txt"
        store._persist_metadata_locked(pm_team="AF", country="SG", metadata=missing_metadata)
        with patch("pathlib.Path.read_bytes", side_effect=OSError("read failed")):
            self.assertEqual(store.resolve_scope(pm_team="AF", country="All"), [])
        self.assertEqual([item["id"] for item in store.resolve_scope(pm_team="AF", country="SG")], [first["id"]])
        self.assertFalse(store.delete(pm_team="AF", country="SG", evidence_id="f" * 32))
        with self.assertRaisesRegex(ToolError, "invalid"):
            store.delete(pm_team="AF", country="SG", evidence_id="bad")
        deletable = store.save_bytes(
            pm_team="AF",
            country="PH",
            source_type="db",
            uploaded_by="owner@npt.sg",
            filename="delete-me.txt",
            content=b"delete me",
            mime_type="text/plain",
        )
        with patch("pathlib.Path.unlink", side_effect=OSError("delete failed")):
            self.assertTrue(store.delete(pm_team="AF", country="PH", evidence_id=deletable["id"]))

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Tables"
        worksheet.append([None, None, None])
        worksheet.append(["table", "column", "meaning"])
        worksheet.append(["risk_rule", "rule_id", "Risk rule id"])
        empty_sheet = workbook.create_sheet("Empty")
        skipped_sheet = workbook.create_sheet("Skipped")
        workbook_buffer = io.BytesIO()
        workbook.save(workbook_buffer)
        with patch.object(SourceCodeQARuntimeEvidenceStore, "MAX_DATA_DICTIONARY_XLSX_SHEETS", 2):
            dictionary_text = store._extract_data_dictionary_xlsx_text(workbook_buffer.getvalue())
        self.assertIn("[Data dictionary sheet: Tables]", dictionary_text)
        self.assertIn("risk_rule", dictionary_text)
        self.assertIn("(empty sheet)", dictionary_text)
        self.assertIn("[Data dictionary skipped sheets: 1]", dictionary_text)
        with patch("builtins.__import__", side_effect=ImportError("no openpyxl")):
            with self.assertRaisesRegex(ToolError, "openpyxl"):
                store._extract_data_dictionary_xlsx_text(b"xlsx")
        real_import = __import__
        fake_empty_openpyxl = SimpleNamespace(load_workbook=lambda *_args, **_kwargs: SimpleNamespace(worksheets=[]))
        with patch("builtins.__import__", side_effect=lambda name, *args, **kwargs: fake_empty_openpyxl if name == "openpyxl" else real_import(name, *args, **kwargs)):
            with self.assertRaisesRegex(ToolError, "readable text"):
                store._extract_data_dictionary_xlsx_text(b"xlsx")
        with patch.object(SourceCodeQARuntimeEvidenceStore, "MAX_DATA_DICTIONARY_XLSX_TEXT_CHARS", 30):
            truncated_dictionary_text = store._extract_data_dictionary_xlsx_text(workbook_buffer.getvalue())
        self.assertLessEqual(len(truncated_dictionary_text), 30)

        with self.assertRaisesRegex(ToolError, "Unable to read"):
            store._extract_zip_text(b"not a zip")
        too_many_zip = io.BytesIO()
        with zipfile.ZipFile(too_many_zip, "w") as archive:
            archive.writestr("a.txt", "a")
            archive.writestr("b.txt", "b")
        with patch.object(SourceCodeQARuntimeEvidenceStore, "MAX_ZIP_MEMBERS", 1):
            with self.assertRaisesRegex(ToolError, "too many"):
                store._extract_zip_text(too_many_zip.getvalue())
        too_large_zip = io.BytesIO()
        with zipfile.ZipFile(too_large_zip, "w") as archive:
            archive.writestr("large.txt", "large text")
        with patch.object(SourceCodeQARuntimeEvidenceStore, "MAX_ZIP_UNCOMPRESSED_BYTES", 1):
            with self.assertRaisesRegex(ToolError, "too large"):
                store._extract_zip_text(too_large_zip.getvalue())
        class BrokenArchive:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def infolist(self):
                return [SimpleNamespace(is_dir=lambda: False, filename="broken.txt", file_size=1)]

            def read(self, _member):
                raise RuntimeError("read failed")

        with patch("bpmis_jira_tool.source_code_qa_stores.zipfile.ZipFile", return_value=BrokenArchive()):
            with self.assertRaisesRegex(ToolError, "Unable to read broken.txt"):
                store._extract_zip_text(b"fake zip")
        empty_zip = io.BytesIO()
        with zipfile.ZipFile(empty_zip, "w") as archive:
            archive.writestr("../unsafe.txt", "skip")
            archive.writestr("binary.txt", b"\x00skip")
            archive.writestr("empty.txt", "")
        with self.assertRaisesRegex(ToolError, "did not contain"):
            store._extract_zip_text(empty_zip.getvalue())

        readable_zip = io.BytesIO()
        with zipfile.ZipFile(readable_zip, "w") as archive:
            archive.writestr("__MACOSX/ignored.txt", "skip")
            archive.writestr("config/app.properties", "app.enabled=true")
            archive.writestr("latin.txt", b"\xfflatin")
            archive.writestr("image.png", "skip")
        zip_text = store._extract_zip_text(readable_zip.getvalue())
        self.assertIn("[ZIP file: config/app.properties]", zip_text)
        self.assertIn("[ZIP skipped files:", zip_text)
        with patch.object(SourceCodeQARuntimeEvidenceStore, "MAX_ZIP_TEXT_CHARS", 20):
            truncated_zip = store._extract_zip_text(readable_zip.getvalue())
        self.assertLessEqual(len(truncated_zip), 20)

    def test_source_code_qa_generated_artifact_store_covers_errors_and_zip_roundtrip(self):
        store = SourceCodeQAGeneratedArtifactStore(Path(self.temp_dir.name) / "artifacts")
        with self.assertRaisesRegex(ToolError, "not configured"):
            SourceCodeQAGeneratedArtifactStore()._session_dir(owner_email="owner@npt.sg", session_id="session123")
        with self.assertRaisesRegex(ToolError, "empty"):
            store.save_sql_package(owner_email="owner@npt.sg", session_id="session123", pm_team="AF", country="SG", question="", sql="", readme="")
        with patch.object(SourceCodeQAGeneratedArtifactStore, "MAX_SQL_BYTES", 4):
            with self.assertRaisesRegex(ToolError, "too large"):
                store.save_sql_package(owner_email="owner@npt.sg", session_id="session123", pm_team="AF", country="SG", question="", sql="select * from table", readme="")

        metadata_path = store._metadata_path(owner_email="owner@npt.sg", session_id="session123")
        metadata_path.parent.mkdir(parents=True, exist_ok=True)
        metadata_path.write_text("{bad", encoding="utf-8")
        self.assertEqual(store._load_metadata_locked(owner_email="owner@npt.sg", session_id="session123"), {})

        artifact = store.save_sql_package(
            owner_email="owner@npt.sg",
            session_id="session123",
            pm_team="af",
            country="sg",
            question="how to query",
            sql="select 1",
            readme="read me",
        )
        public_metadata, content = store.get_bytes(owner_email="owner@npt.sg", session_id="session123", artifact_id=artifact["id"])
        self.assertEqual(public_metadata["pm_team"], "AF")
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            self.assertIn("query.sql", archive.namelist())
        with self.assertRaisesRegex(ToolError, "invalid"):
            store.get_bytes(owner_email="owner@npt.sg", session_id="session123", artifact_id="bad")
        with self.assertRaisesRegex(ToolError, "not found"):
            store.get_bytes(owner_email="owner@npt.sg", session_id="session123", artifact_id="f" * 32)

        metadata = store._load_metadata_locked(owner_email="owner@npt.sg", session_id="session123")
        (store._session_dir(owner_email="owner@npt.sg", session_id="session123") / metadata[artifact["id"]]["stored_name"]).unlink()
        with self.assertRaisesRegex(ToolError, "unreadable"):
            store.get_bytes(owner_email="owner@npt.sg", session_id="session123", artifact_id=artifact["id"])

    def test_retrieval_query_returns_path_line_and_snippet(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "bpmis" / "jira_client.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "class BPMISClient:\n"
            "    def batchCreateJiraIssue(self):\n"
            "        return self.post('/api/v1/issues/batchCreateJiraIssue')\n",
            encoding="utf-8",
        )
        self._build_index_for_entry("AF:All", entry)
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="batchCreateJiraIssue BPMIS API")

        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["answer_mode"], "retrieval_only")
        self.assertEqual(payload["matches"][0]["path"], "bpmis/jira_client.py")
        self.assertIn("batchCreateJiraIssue", payload["matches"][0]["snippet"])
        self.assertIn("query_timing", payload)
        self.assertIn("rank_and_expand", payload["query_timing"]["components"])
        self.assertIn("slow_query_attribution", payload)
        self.assertIn("cache_preload", payload)

    def test_deadline_fallback_skips_llm_when_retrieval_uses_deadline_budget(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "bpmis" / "jira_client.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "class BPMISClient:\n"
            "    def batchCreateJiraIssue(self):\n"
            "        return self.post('/api/v1/issues/batchCreateJiraIssue')\n",
            encoding="utf-8",
        )
        self._build_index_for_entry("AF:All", entry)

        self.service.query_deadline_seconds = 1
        with patch.object(self.service, "_query_deadline_hit", return_value=True), patch.object(
            self.service,
            "_build_llm_answer",
            side_effect=AssertionError("deadline fallback should skip LLM generation"),
        ):
            payload = self.service.query(
                pm_team="AF",
                country="All",
                question="batchCreateJiraIssue BPMIS API",
                answer_mode="auto",
            )

        self.assertEqual(payload["status"], "ok")
        self.assertTrue(payload["deadline_hit"])
        self.assertTrue(payload["fallback_used"])
        self.assertEqual(payload["answer_contract"]["status"], "deadline_fallback")
        self.assertEqual(payload["llm_finish_reason"], "deadline_fallback")
        self.assertEqual(payload["slow_query_attribution"]["reason"], "retrieval_exceeded_deadline")

    def test_answer_generation_timeout_uses_deadline_fallback(self):
        payload = {
            "summary": "Relevant code references were found.",
            "query_mode": "deep",
        }
        matches = [
            {
                "repo": "Portal Repo",
                "path": "bpmis/jira_client.py",
                "line_start": 1,
                "line_end": 3,
            }
        ]
        reports = []
        self.service.query_deadline_seconds = 30
        with patch.object(
            SourceCodeQAService,
            "_build_llm_answer",
            side_effect=ToolError("Codex unavailable; used code search fallback. Codex CLI timed out after 24s."),
        ) as build_answer:
            self.service._augment_query_payload_with_llm_answer(
                payload=payload,
                entries=[],
                key="AF:All",
                pm_team="AF",
                country="All",
                question="batchCreateJiraIssue BPMIS API",
                matches=matches,
                llm_budget_mode="auto",
                query_mode="deep",
                trace_id="trace-deadline",
                followup_context={},
                normalized_answer_mode="auto",
                request_cache={},
                progress_callback=None,
                attachments=[],
                runtime_evidence=[],
                effort_assessment=False,
                retrieval_latency_ms=1200,
                evidence_pack={},
                report=lambda *args: reports.append(args),
                query_started_at=time.time(),
            )

        self.assertTrue(build_answer.called)
        self.assertTrue(payload["deadline_hit"])
        self.assertTrue(payload["fallback_used"])
        self.assertEqual(payload["deadline_fallback_reason"], "answer_generation_exceeded_deadline")
        self.assertEqual(payload["llm_finish_reason"], "deadline_fallback")
        self.assertIn("Top evidence", payload["llm_answer"])
        self.assertTrue(any(item[0] == "deadline_fallback" for item in reports))

    def test_chinese_business_intent_detection(self):
        intent = self.service._question_intent("IssueService 改了会影响哪些下游，测试有没有覆盖，事务缓存边界是什么")

        self.assertTrue(intent["impact_analysis"])
        self.assertTrue(intent["test_coverage"])
        self.assertTrue(intent["operational_boundary"])
        self.assertTrue(any("影响" in token for token in self.service._question_tokens("IssueService 改了会影响哪些下游")))

    def test_domain_labels_do_not_trigger_unrelated_quality_intents(self):
        crms_intent = self.service._question_intent("which SG Credit Risk API runs credit application precheck")
        self.assertTrue(crms_intent["api"])
        self.assertFalse(crms_intent["static_qa"])

        grc_config_intent = self.service._question_intent("GRC globallock config is where and which table backs it")
        self.assertTrue(grc_config_intent["config"])
        self.assertTrue(grc_config_intent["data_source"])
        self.assertFalse(grc_config_intent["operational_boundary"])

        grc_table_intent = self.service._question_intent("where does GRC read and write bcf_global_lock table")
        self.assertTrue(grc_table_intent["data_source"])
        self.assertFalse(grc_table_intent["operational_boundary"])

    def test_sql_generation_intent_is_explicit_and_not_static_qa(self):
        sql_intent = self.service._question_intent("Based on GRC repo and data dictionary, generate a SQL query for reviewer status")
        self.assertTrue(sql_intent["sql_generation"])
        self.assertTrue(sql_intent["data_source"])

        chinese_sql_intent = self.service._question_intent("帮我写SQL查询incident reviewer状态")
        self.assertTrue(chinese_sql_intent["sql_generation"])

        static_intent = self.service._question_intent("check SQL injection risk in TicketMapper")
        self.assertTrue(static_intent["static_qa"])
        self.assertFalse(static_intent["sql_generation"])

        data_source_intent = self.service._question_intent("which table stores reviewer status")
        self.assertTrue(data_source_intent["data_source"])
        self.assertFalse(data_source_intent["sql_generation"])

    def test_chinese_data_source_query_uses_source_trace(self):
        _build_fixture_repositories(self.service)
        self._build_all_indexes()

        payload = self.service.query(
            pm_team="CRMS",
            country="ID",
            question="Term Loan Pre Check 1 的数据源是哪张表",
        )

        self.assertEqual(payload["status"], "ok")
        self.assertIn("repository/CustomerRepository.java", {match["path"] for match in payload["matches"]})
        self.assertEqual(payload["answer_quality"]["policies"][0]["name"], "data_source")
        self.assertEqual(payload["answer_quality"]["policies"][0]["status"], "satisfied")
        self.assertIn("cr_customer_profile", json.dumps(payload["evidence_pack"], ensure_ascii=False))

    def test_followup_context_uses_evidence_boundaries(self):
        augmented, followup = self.service._apply_conversation_context(
            "继续看下游影响",
            {
                "key": "AF:All",
                "trace_id": "trace-prev",
                "question": "what is impacted if IssueService createIssue changes",
                "answer": "Previous answer mentions IssueRepository.",
                "codex_candidate_paths": [
                    {
                        "repo": "Portal Repo",
                        "repo_root": "/tmp/portal",
                        "path": "repository/IssueRepository.java",
                        "line_start": 3,
                        "line_end": 5,
                    }
                ],
                "matches": [],
                "trace_paths": [],
                "structured_answer": {},
                "answer_contract": {
                    "confirmed_sources": ["Portal Repo:repository/IssueRepository.java:3-5: issue_table [S1]"],
                    "missing_links": ["consumer caller evidence"],
                },
                "evidence_pack": {
                    "confirmed_facts": ["IssueRepository writes issue_table"],
                    "impact_surfaces": ["downstream callee: IssueRepository.findIssue"],
                },
            },
            current_key="AF:All",
        )

        self.assertTrue(followup["used"])
        self.assertIn("issuerepository", augmented)
        self.assertIn("issue_table", augmented)
        self.assertEqual(followup["trace_id"], "trace-prev")
        self.assertIn("Previous answer", followup["answer"])
        self.assertEqual(followup["codex_candidate_paths"][0]["path"], "repository/IssueRepository.java")

    def test_followup_terms_use_codex_context(self):
        augmented, followup = self.service._apply_conversation_context(
            "continue with this",
            {
                "key": "AF:All",
                "question": "previous Codex question",
                "matches": [],
                "trace_paths": [],
                "structured_answer": {},
                "answer_contract": {},
                "evidence_pack": {},
                "codex_candidate_paths": [
                    {
                        "repo": "Portal Repo",
                        "repo_root": "/tmp/portal",
                        "path": "repository/CodexOnlyRepository.java",
                        "reason": "previous Codex-only path",
                    }
                ],
                "codex_citation_validation": {
                    "direct_file_refs": [{"path": "repository/CodexOnlyRepository.java"}],
                },
            },
            current_key="AF:All",
        )

        self.assertTrue(followup["used"])
        self.assertIn("codexonlyrepository", followup["terms"])
        self.assertIn("codexonlyrepository", augmented.lower())

    def test_codex_followup_terms_are_used_for_codex_provider(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            llm_provider="codex_cli_bridge",
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        augmented, followup = service._apply_conversation_context(
            "continue with this",
            {
                "key": "AF:All",
                "question": "previous Codex question",
                "matches": [],
                "trace_paths": [],
                "structured_answer": {},
                "answer_contract": {},
                "evidence_pack": {},
                "codex_candidate_paths": [
                    {
                        "repo": "Portal Repo",
                        "repo_root": "/tmp/portal",
                        "path": "repository/CodexOnlyRepository.java",
                        "reason": "previous Codex-only path",
                    }
                ],
            },
            current_key="AF:All",
        )

        self.assertTrue(followup["used"])
        self.assertIn("codexonlyrepository", augmented)

    def test_followup_context_does_not_cross_repository_scope(self):
        augmented, followup = self.service._apply_conversation_context(
            "继续看这个表哪里写入",
            {
                "key": "AF:All",
                "pm_team": "AF",
                "country": "All",
                "question": "what table does issue creation use",
                "matches": [{"repo": "Anti Fraud", "path": "repository/IssueRepository.java", "snippet": "select * from issue_table"}],
                "trace_paths": [],
                "structured_answer": {},
                "answer_contract": {"confirmed_sources": ["Anti Fraud:repository/IssueRepository.java:1-3: issue_table [S1]"]},
                "evidence_pack": {"confirmed_facts": ["IssueRepository reads issue_table"]},
            },
            current_key="CRMS:SG",
        )

        self.assertFalse(followup["used"])
        self.assertEqual(followup["reason"], "scope_mismatch")
        self.assertEqual(augmented, "继续看这个表哪里写入")

    def test_followup_context_rejects_legacy_scope_fields_when_changed(self):
        augmented, followup = self.service._apply_conversation_context(
            "continue checking this method",
            {
                "pm_team": "AF",
                "country": "All",
                "question": "where is createIssue",
                "matches": [{"repo": "Anti Fraud", "path": "service/IssueService.java", "snippet": "createIssue()"}],
                "trace_paths": [],
                "structured_answer": {},
            },
            current_key="CRMS:ID",
        )

        self.assertFalse(followup["used"])
        self.assertEqual(followup["reason"], "scope_mismatch")
        self.assertEqual(augmented, "continue checking this method")

    def test_followup_context_does_not_match_substrings_inside_normal_questions(self):
        augmented, followup = self.service._apply_conversation_context(
            "When will system query CBS report when performing monthly credit review?",
            {
                "question": "Which table stores payslip extracted fields?",
                "matches": [{"path": "CardIncomeScreeningFlowStatusDAO-ext.xml", "snippet": "process_info"}],
                "trace_paths": [],
                "structured_answer": {"claims": [{"text": "Payslip fields are stored in process_info."}]},
                "answer_contract": {"confirmed_sources": ["card_income_screening_flow_status_tab"]},
                "evidence_pack": {"confirmed_facts": ["extract_record_tab response_body"]},
            },
        )

        self.assertFalse(followup["used"])
        self.assertEqual(augmented, "When will system query CBS report when performing monthly credit review?")

    def test_followup_context_same_scope_session_is_carried_without_query_pollution(self):
        augmented, followup = self.service._apply_conversation_context(
            "Which table stores payslip extracted fields?",
            {
                "key": "CRMS:PH",
                "question": "When is payslip parsed?",
                "answer": "Previous answer mentioned ExtractRecord.",
                "matches": [{"path": "repository/ExtractRecordDAO.xml", "snippet": "response_body"}],
                "trace_paths": [],
                "structured_answer": {"claims": [{"text": "ExtractRecord stores response_body."}]},
                "answer_contract": {"confirmed_sources": ["extract_record_tab.response_body"]},
                "evidence_pack": {"confirmed_facts": ["raw response is stored in extract_record_tab"]},
            },
            current_key="CRMS:PH",
        )

        self.assertTrue(followup["used"])
        self.assertTrue(followup["implicit"])
        self.assertEqual(followup["reason"], "same_scope_session")
        self.assertEqual(augmented, "Which table stores payslip extracted fields?")
        self.assertIn("Previous answer", followup["answer"])

    def test_followup_context_chinese_clarification_uses_previous_terms(self):
        augmented, followup = self.service._apply_conversation_context(
            "我问的是是否要approve，要的话哪些role需要approve？不需要名字",
            {
                "key": "GRC:All",
                "question": "When can an incident be withdrawn?",
                "answer": "Previous answer mentioned AuthorizationContent.",
                "matches": [
                    {
                        "path": "src/pages/AuthorizationManagement/Components/DetailPage/AuthorizationContent.tsx",
                        "snippet": "IncidentApprove IncidentReview pendingWithdraw Approver Reviewer",
                    }
                ],
                "trace_paths": [],
                "structured_answer": {"claims": [{"text": "pendingWithdraw is mapped to IncidentApprove."}]},
                "answer_contract": {"confirmed_sources": ["AuthorizationContent.tsx:241-262"]},
                "evidence_pack": {"confirmed_facts": ["Approver handles IncidentApprove"]},
            },
            current_key="GRC:All",
        )

        self.assertTrue(followup["used"])
        self.assertFalse(followup["implicit"])
        self.assertEqual(followup["reason"], "followup_marker")
        self.assertIn("Previous Source Code Q&A context terms", augmented)
        self.assertIn("authorizationcontent", augmented)
        self.assertIn("incidentapprove", augmented)

    def test_followup_relationship_question_inherits_previous_entities(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            llm_provider="codex_cli_bridge",
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        augmented, followup = service._apply_conversation_context(
            "Understand deeper if there is any relationship between the two fields",
            {
                "key": "AF:All",
                "question": "merchantUId and shopeeUid difference",
                "answer": "Previous answer mentioned merchantUId and shopeeUid.",
                "matches": [],
                "trace_paths": [],
                "structured_answer": {},
                "answer_contract": {},
                "evidence_pack": {},
                "codex_candidate_paths": [
                    {
                        "repo": "AF FE",
                        "repo_root": "/tmp/af-fe",
                        "path": "src/components/ShopeeUid.tsx",
                        "reason": "matched shopeeUid and merchantUId",
                    }
                ],
            },
            current_key="AF:All",
        )

        self.assertTrue(followup["used"])
        self.assertFalse(followup["implicit"])
        self.assertIn("Previous Source Code Q&A context terms", augmented)
        self.assertIn("merchantuid", augmented)
        self.assertIn("shopeeuid", augmented)
        self.assertNotIn("between fields there", augmented)

    def test_followup_relationship_question_prefers_session_title_fields_over_stale_context(self):
        augmented, followup = self.service._apply_conversation_context(
            "Understand deeper if there is any relationship between the two fields",
            {
                "key": "AF:All",
                "session_title": "merchantUId and shopeeUid difference",
                "question": "operationPosition and gpsHex relationship",
                "answer": "operationPosition is used to derive gpsHex fields.",
                "matches": [{"path": "LocationService.java", "snippet": "operationPosition gpsHex action_location"}],
                "trace_paths": [],
                "structured_answer": {},
                "answer_contract": {},
                "evidence_pack": {},
            },
            current_key="AF:All",
        )

        self.assertTrue(followup["used"])
        self.assertIn("merchantuid", augmented)
        self.assertIn("shopeeuid", augmented)
        self.assertNotIn("operationposition", augmented)
        self.assertNotIn("gpshex", augmented)

    def test_followup_context_does_not_pollute_specific_short_lookup(self):
        augmented, followup = self.service._apply_conversation_context(
            "Is fdMaturityDate used in any function?",
            {
                "question": "Where is F44 configured?",
                "matches": [{"path": "apollo.properties", "snippet": "dbp.antifraud.function.F44"}],
                "trace_paths": [],
                "structured_answer": {},
                "answer_contract": {"confirmed_sources": ["dbp.antifraud.function.F44"]},
                "evidence_pack": {},
            },
        )

        self.assertFalse(followup["used"])
        self.assertEqual(augmented, "Is fdMaturityDate used in any function?")

    def test_followup_context_does_not_cross_explicit_repo_scope(self):
        repositories = [
            RepositoryEntry(display_name="Anti Fraud API", url="https://git.example.com/team/anti-fraud-api.git"),
            RepositoryEntry(display_name="GRC Portal", url="https://git.example.com/team/grc-portal.git"),
        ]

        augmented, followup = self.service._apply_conversation_context(
            "继续看 GRC Portal 这个方法",
            {
                "key": "AF:All",
                "repo_scope": ["Anti Fraud API"],
                "question": "where is createIssue",
                "matches": [{"repo": "Anti Fraud API", "path": "service/IssueService.java", "snippet": "createIssue()"}],
                "trace_paths": [],
                "structured_answer": {},
            },
            current_key="AF:All",
            current_repositories=repositories,
        )

        self.assertFalse(followup["used"])
        self.assertEqual(followup["reason"], "repo_scope_mismatch")
        self.assertEqual(augmented, "继续看 GRC Portal 这个方法")

    def test_followup_context_keeps_same_explicit_repo_scope(self):
        repositories = [
            RepositoryEntry(display_name="Anti Fraud API", url="https://git.example.com/team/anti-fraud-api.git"),
            RepositoryEntry(display_name="GRC Portal", url="https://git.example.com/team/grc-portal.git"),
        ]

        augmented, followup = self.service._apply_conversation_context(
            "继续看 Anti Fraud 这个方法下游",
            {
                "key": "AF:All",
                "repo_scope": ["Anti Fraud API"],
                "question": "where is createIssue",
                "matches": [{"repo": "Anti Fraud API", "path": "service/IssueService.java", "snippet": "createIssue()"}],
                "trace_paths": [],
                "structured_answer": {"claims": [{"text": "IssueService handles createIssue."}]},
                "answer_contract": {"confirmed_sources": ["Anti Fraud API:service/IssueService.java:1-3 [S1]"]},
            },
            current_key="AF:All",
            current_repositories=repositories,
        )

        self.assertTrue(followup["used"])
        self.assertIn("issueservice", augmented)

    def test_query_limits_retrieval_to_explicit_repository_scope(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[
                {"display_name": "Anti Fraud API", "url": "https://git.example.com/team/anti-fraud-api.git"},
                {"display_name": "GRC Portal", "url": "https://git.example.com/team/grc-portal.git"},
            ],
        )
        entries = self.service.load_config()["mappings"]["AF:All"]
        anti_entry = entries[0]
        grc_entry = entries[1]
        anti_path = self.service._repo_path("AF:All", type("Entry", (), anti_entry)())
        grc_path = self.service._repo_path("AF:All", type("Entry", (), grc_entry)())
        (anti_path / ".git").mkdir(parents=True)
        (grc_path / ".git").mkdir(parents=True)
        anti_file = anti_path / "service" / "IssueService.java"
        grc_file = grc_path / "service" / "IssueService.java"
        anti_file.parent.mkdir(parents=True)
        grc_file.parent.mkdir(parents=True)
        anti_file.write_text(
            "class IssueService {\n"
            "    void createIssue() { antiFraudOnly(); }\n"
            "}\n",
            encoding="utf-8",
        )
        grc_file.write_text(
            "class IssueService {\n"
            "    void createIssue() { grcPortalOnly(); }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_index_for_entry("AF:All", anti_entry)
        self._build_index_for_entry("AF:All", grc_entry)

        payload = self.service.query(pm_team="AF", country="All", question="where is createIssue in GRC Portal")

        self.assertEqual(payload["status"], "ok")
        self.assertTrue(payload["repository_scope"]["active"])
        self.assertEqual(payload["repository_scope"]["selected_repositories"], ["GRC Portal"])
        self.assertGreaterEqual(payload["retrieval_runtime"].get("repository_scope_filters", 0), 1)
        self.assertEqual({match["repo"] for match in payload["matches"]}, {"GRC Portal"})
        self.assertIn("grcPortalOnly", "\n".join(match["snippet"] for match in payload["matches"]))

    def test_repository_scope_prefers_specific_alias_over_generic_token(self):
        repositories = [
            RepositoryEntry(display_name="Anti Fraud API", url="https://git.example.com/team/anti-fraud-api.git"),
            RepositoryEntry(display_name="GRC Portal", url="https://git.example.com/team/grc-portal.git"),
            RepositoryEntry(display_name="Audit Admin", url="https://git.example.com/team/audit-admin.git"),
        ]

        selected, scope = self.service._filter_entries_for_question_repository_scope(
            "where is fraud rule handled in GRC Portal",
            repositories,
        )

        self.assertTrue(scope["active"])
        self.assertEqual([entry.display_name for entry in selected], ["GRC Portal"])
        self.assertEqual(scope["selected_repositories"], ["GRC Portal"])
        self.assertNotIn("Anti Fraud API", scope["selected_repositories"])

    def test_question_specific_terms_cover_cbs_credit_review_questions(self):
        terms = self.service._question_specific_retrieval_terms(
            "When will system query CBS report when performing monthly credit review?"
        )

        self.assertIn("CreditReviewCbsServiceImpl", terms)
        self.assertIn("CreditReviewCbsBureauReportProvider", terms)
        self.assertIn("CR_CBS_REPORT", terms)

    def test_query_uses_ready_persistent_index_and_citations(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "bpmis" / "jira_client.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "class BPMISClient:\n"
            "    def batchCreateJiraIssue(self):\n"
            "        return self.post('/api/v1/issues/batchCreateJiraIssue')\n",
            encoding="utf-8",
        )
        self._build_index_for_entry("AF:All", entry)
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="where is batchCreateJiraIssue implemented")

        self.assertEqual(payload["status"], "ok")
        self.assertTrue(self.service._index_path(repo_path).exists())
        self.assertIn("persistent_index", {match.get("retrieval") for match in payload["matches"]})
        self.assertEqual(payload["citations"][0]["id"], "S1")
        self.assertEqual(payload["citations"][0]["path"], "bpmis/jira_client.py")
        self.assertEqual(payload["index_freshness"]["status"], "fresh")
        self.assertIn("git_revisions", payload["index_freshness"])
        self.assertEqual(self.service.repo_status("AF:All")[0]["index"]["state"], "ready")

    def test_soft_exact_lookup_miss_falls_back_to_broad_search(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "src" / "RiskEngineTimeoutConfig.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "package com.example.risk.engine;\n"
            "public class RiskEngineTimeoutConfig {\n"
            "    public int timeoutSeconds() { return 30; }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_index_for_entry("AF:All", entry)

        payload = self.service.query(pm_team="AF", country="All", question="where is risk.engine.timeout configured")

        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["exact_lookup"]["terms"], ["risk.engine.timeout"])
        self.assertEqual(payload["exact_lookup"]["matched_terms"], [])
        self.assertGreaterEqual(payload["retrieval_runtime"].get("exact_lookup_soft_misses", 0), 1)
        self.assertIn("src/RiskEngineTimeoutConfig.java", {match["path"] for match in payload["matches"]})

    def test_exact_lookup_miss_still_stops_for_strict_table_terms(self):
        self.assertTrue(self.service._exact_lookup_miss_should_stop(["missing_schema.dwd_missing_table_di"]))
        self.assertTrue(self.service._exact_lookup_miss_should_stop(["mapper/MissingMapper.xml"]))
        self.assertFalse(self.service._exact_lookup_miss_should_stop(["risk.engine.timeout"]))
        self.assertFalse(self.service._exact_lookup_miss_should_stop(["model/provider:", "screens/components", "db/apollo/log"]))
        self.assertFalse(self.service._is_strict_exact_lookup_term("screens/components"))
        self.assertFalse(self.service._is_strict_exact_lookup_term("db/apollo/log"))
        self.assertTrue(self.service._is_strict_exact_lookup_term("spec/usecase/uc007_query_retail_underwriting_result.md"))
        self.assertTrue(self.service._is_strict_exact_lookup_term("src/main/java/com/example/RiskEngine.java"))

    def test_query_does_not_build_missing_index_synchronously(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "bpmis" / "jira_client.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text("class BPMISClient: pass\n", encoding="utf-8")

        payload = self.service.query(pm_team="AF", country="All", question="where is BPMISClient")

        self.assertEqual(payload["status"], "index_not_ready")
        self.assertFalse(self.service._index_path(repo_path).exists())
        self.assertEqual(payload["index_freshness"]["status"], "stale_or_missing")
        self.assertIn("ready queryable index", payload["summary"])
        self.assertGreaterEqual(payload["retrieval_runtime"].get("index_not_ready_scopes", 0), 1)

    def test_query_reports_not_synced_when_no_local_clone_is_available(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )

        payload = self.service.query(pm_team="AF", country="All", question="where is BPMISClient")

        self.assertEqual(payload["status"], "not_synced")
        self.assertEqual(payload["matches"], [])
        self.assertEqual(payload["index_freshness"]["status"], "stale_or_missing")
        self.assertIn("Sync / Refresh", payload["summary"])

    def test_query_reports_not_synced_for_unsynced_explicit_repository_scope(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[
                {"display_name": "Anti Fraud API", "url": "https://git.example.com/team/anti-fraud-api.git"},
                {"display_name": "GRC Portal", "url": "https://git.example.com/team/grc-portal.git"},
            ],
        )
        entries = self.service.load_config()["mappings"]["AF:All"]
        anti_entry = entries[0]
        anti_path = self.service._repo_path("AF:All", type("Entry", (), anti_entry)())
        (anti_path / ".git").mkdir(parents=True)
        source_file = anti_path / "service" / "IssueService.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text("class IssueService { void createIssue() {} }\n", encoding="utf-8")
        self._build_index_for_entry("AF:All", anti_entry)

        payload = self.service.query(pm_team="AF", country="All", question="where is createIssue in GRC Portal")

        self.assertEqual(payload["status"], "not_synced")
        self.assertTrue(payload["repository_scope"]["active"])
        self.assertEqual(payload["repository_scope"]["selected_repositories"], ["GRC Portal"])
        self.assertEqual(payload["matches"], [])
        self.assertGreaterEqual(payload["retrieval_runtime"].get("repository_scope_filters", 0), 1)

    def test_query_reports_index_not_ready_for_explicit_scope_with_clone_but_no_index(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[
                {"display_name": "Anti Fraud API", "url": "https://git.example.com/team/anti-fraud-api.git"},
                {"display_name": "GRC Portal", "url": "https://git.example.com/team/grc-portal.git"},
            ],
        )
        entries = self.service.load_config()["mappings"]["AF:All"]
        anti_entry = entries[0]
        grc_entry = entries[1]
        anti_path = self.service._repo_path("AF:All", type("Entry", (), anti_entry)())
        grc_path = self.service._repo_path("AF:All", type("Entry", (), grc_entry)())
        (anti_path / ".git").mkdir(parents=True)
        (grc_path / ".git").mkdir(parents=True)
        anti_file = anti_path / "service" / "IssueService.java"
        grc_file = grc_path / "service" / "IssueService.java"
        anti_file.parent.mkdir(parents=True)
        grc_file.parent.mkdir(parents=True)
        anti_file.write_text("class IssueService { void createIssue() {} }\n", encoding="utf-8")
        grc_file.write_text("class IssueService { void createIssue() {} }\n", encoding="utf-8")
        self._build_index_for_entry("AF:All", anti_entry)

        payload = self.service.query(pm_team="AF", country="All", question="where is createIssue in GRC Portal")

        self.assertEqual(payload["status"], "index_not_ready")
        self.assertTrue(payload["repository_scope"]["active"])
        self.assertEqual(payload["repository_scope"]["selected_repositories"], ["GRC Portal"])
        self.assertEqual(payload["matches"], [])
        self.assertGreaterEqual(payload["retrieval_runtime"].get("index_not_ready_scopes", 0), 1)

    def test_query_can_use_stale_but_schema_compatible_index(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_entry = type("Entry", (), entry)()
        repo_path = self.service._repo_path("AF:All", repo_entry)
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "bpmis" / "jira_client.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "class BPMISClient:\n"
            "    def batchCreateJiraIssue(self):\n"
            "        return self.post('/api/v1/issues/batchCreateJiraIssue')\n",
            encoding="utf-8",
        )
        self.service._build_repo_index("AF:All", repo_entry, repo_path)
        index_path = self.service._index_path(repo_path)
        indexed_mtime = index_path.stat().st_mtime_ns

        time.sleep(0.01)
        (repo_path / "new_module.py").write_text("class NewModule: pass\n", encoding="utf-8")

        payload = self.service.query(pm_team="AF", country="All", question="where is batchCreateJiraIssue implemented")

        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["index_freshness"]["status"], "stale_or_missing")
        self.assertIn("Portal Repo", payload["index_freshness"]["stale_repos"])
        self.assertIn("persistent_index", {match.get("retrieval") for match in payload["matches"]})
        self.assertEqual(index_path.stat().st_mtime_ns, indexed_mtime)
        index_info = self.service.repo_status("AF:All")[0]["index"]
        self.assertEqual(index_info["state"], "stale")
        self.assertTrue(index_info["schema_compatible"])
        self.assertTrue(index_info["queryable"])

    def test_query_can_use_previous_queryable_index_without_token_tables(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_entry = type("Entry", (), entry)()
        repo_path = self.service._repo_path("AF:All", repo_entry)
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "bpmis" / "jira_client.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "class BPMISClient:\n"
            "    def batchCreateJiraIssue(self):\n"
            "        return self.post('/api/v1/issues/batchCreateJiraIssue')\n",
            encoding="utf-8",
        )
        self.service._build_repo_index("AF:All", repo_entry, repo_path)
        index_path = self.service._index_path(repo_path)
        with sqlite3.connect(index_path) as connection:
            connection.execute("update metadata set value = '28' where key = 'version'")
            connection.execute("drop table file_tokens")
            connection.execute("drop table line_tokens")
            connection.execute("drop table semantic_chunk_tokens")
            connection.commit()

        payload = self.service.query(pm_team="AF", country="All", question="where is batchCreateJiraIssue implemented")

        self.assertEqual(payload["status"], "ok")
        index_info = self.service.repo_status("AF:All")[0]["index"]
        self.assertEqual(index_info["state"], "stale")
        self.assertFalse(index_info["schema_compatible"])
        self.assertTrue(index_info["queryable"])

    def test_structure_index_extracts_definitions_references_and_telemetry(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "controller" / "IssueController.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "@RequestMapping(\"/api/v1/issues\")\n"
            "public class IssueController {\n"
            "    private IssueRepository issueRepository;\n"
            "    @PostMapping(\"/create\")\n"
            "    public Issue createIssue() {\n"
            "        return issueRepository.findIssue();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        repository_file = repo_path / "repository" / "IssueRepository.java"
        repository_file.parent.mkdir(parents=True)
        repository_file.write_text(
            "public class IssueRepository {\n"
            "    public Issue findIssue() {\n"
            "        return jdbcTemplate.queryForObject(\"select * from bpmis_issue\", mapper);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_index_for_entry("AF:All", entry)
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="where is IssueController createIssue API")

        self.assertEqual(payload["status"], "ok")
        self.assertTrue(any("structure matched" in match["reason"] for match in payload["matches"]))
        index_path = self.service._index_path(repo_path)
        with sqlite3.connect(index_path) as connection:
            metadata = dict(connection.execute("select key, value from metadata").fetchall())
            definition_count = connection.execute("select count(*) from definitions").fetchone()[0]
            reference_count = connection.execute("select count(*) from references_index").fetchone()[0]
            entity_count = connection.execute("select count(*) from code_entities").fetchone()[0]
            entity_edge_count = connection.execute("select count(*) from entity_edges").fetchone()[0]
            edge_count = connection.execute("select count(*) from graph_edges").fetchone()[0]
            flow_edge_count = connection.execute("select count(*) from flow_edges").fetchone()[0]
            semantic_chunk_count = connection.execute("select count(*) from semantic_chunks").fetchone()[0]
            file_fts_count = connection.execute("select count(*) from files_fts").fetchone()[0]
            fts_count = connection.execute("select count(*) from lines_fts").fetchone()[0]
            semantic_fts_count = connection.execute("select count(*) from semantic_chunks_fts").fetchone()[0]
            semantic_columns = {row[1] for row in connection.execute("pragma table_info(semantic_chunks)").fetchall()}
        self.assertEqual(metadata["version"], str(CODE_INDEX_VERSION))
        self.assertEqual(CODE_INDEX_VERSION, 30)
        self.assertNotIn("embedding", semantic_columns)
        self.assertGreaterEqual(definition_count, 3)
        self.assertGreaterEqual(reference_count, 3)
        self.assertGreaterEqual(entity_count, 3)
        self.assertGreaterEqual(entity_edge_count, 1)
        self.assertGreaterEqual(edge_count, 1)
        self.assertGreaterEqual(flow_edge_count, 1)
        self.assertGreaterEqual(semantic_chunk_count, 2)
        self.assertGreaterEqual(file_fts_count, 2)
        self.assertGreaterEqual(fts_count, 1)
        self.assertGreaterEqual(semantic_fts_count, 1)
        self.assertEqual(metadata["file_fts_enabled"], "1")
        self.assertEqual(metadata["semantic_fts_enabled"], "1")
        self.assertEqual(metadata["parser_backend"], "tree_sitter+regex")
        self.assertIn("java", metadata["parser_languages"])
        self.assertGreaterEqual(int(metadata["tree_sitter_files"]), 2)
        self.assertIn("semantic_chunk", {match.get("retrieval") for match in payload["matches"]})
        telemetry = self.service.telemetry_path.read_text(encoding="utf-8").strip().splitlines()
        self.assertTrue(telemetry)
        self.assertIn("IssueController", telemetry[-1])

    def test_query_prioritizes_exact_table_lookup_before_broad_search(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Anti Fraud Repo", "url": "https://git.example.com/team/anti-fraud.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        mapper_file = repo_path / "mapper" / "IncomingTransferMapper.xml"
        mapper_file.parent.mkdir(parents=True)
        mapper_file.write_text(
            "<mapper namespace=\"IncomingTransferMapper\">\n"
            "  <select id=\"readDetail\">\n"
            "    select * from bmart_antifraud.dwd_antifraud_incoming_transfer_detailed_di\n"
            "    join tmp_dwd_antifraud_incoming_transfer_detailed_df on id = transfer_id\n"
            "  </select>\n"
            "</mapper>\n",
            encoding="utf-8",
        )
        self._build_index_for_entry("AF:All", entry)

        with patch.object(self.service, "_search_repo", wraps=self.service._search_repo) as broad_search:
            payload = self.service.query(
                pm_team="AF",
                country="All",
                question=(
                    "relation between bmart_antifraud.dwd_antifraud_incoming_transfer_detailed_di "
                    "and tmp_dwd_antifraud_incoming_transfer_detailed_df"
                ),
            )

        self.assertEqual(payload["status"], "ok")
        self.assertTrue(payload["exact_lookup"]["sufficient"])
        self.assertEqual(
            set(payload["exact_lookup"]["matched_terms"]),
            {
                "bmart_antifraud.dwd_antifraud_incoming_transfer_detailed_di",
                "tmp_dwd_antifraud_incoming_transfer_detailed_df",
            },
        )
        self.assertEqual(broad_search.call_count, 0)
        self.assertIn("exact_table_path_lookup", {match.get("retrieval") for match in payload["matches"]})
        self.assertGreaterEqual(payload["retrieval_runtime"].get("exact_lookup_hits", 0), 2)

    def test_exact_table_lookup_keeps_schema_qualified_table_hits_in_evidence(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "anti-fraud-admin", "url": "https://git.example.com/team/anti-fraud-admin.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        transfer_job = repo_path / "anti-fraud-admin-app/src/main/java/com/shopee/banking/af/admin/app/job/xxl/BaseSyncIncomingDataToIvlog.java"
        account_job = repo_path / "anti-fraud-admin-app/src/main/java/com/shopee/banking/af/admin/app/job/xxl/OneTimesSyncDwhIncomingAccountInfoJob.java"
        config_doc = repo_path / "spec/arch/config.md"
        transfer_job.parent.mkdir(parents=True)
        account_job.parent.mkdir(parents=True, exist_ok=True)
        config_doc.parent.mkdir(parents=True)
        transfer_job.write_text(
            "public class BaseSyncIncomingDataToIvlog {\n"
            "    /** dwd_antifraud_incoming_transfer_detailed_di */\n"
            "    @Value(\"${mkt-opt.dwh.incoming.transfer.tab.name:dwd_antifraud_incoming_transfer_detailed_di}\")\n"
            "    protected String incomingTransferTabName;\n"
            "}\n",
            encoding="utf-8",
        )
        account_job.write_text(
            "public class OneTimesSyncDwhIncomingAccountInfoJob {\n"
            "    /** tmp_dwd_antifraud_incoming_transfer_detailed_df */\n"
            "    @Value(\"${mkt-opt.dwh.incoming.account.tab.name:tmp_dwd_antifraud_incoming_transfer_detailed_df}\")\n"
            "    protected String incomingAccountTabName;\n"
            "}\n",
            encoding="utf-8",
        )
        config_doc.write_text(
            "| mkt-opt.dwh.incoming.account.tab.name | tmp_dwd_antifraud_incoming_transfer_detailed_df |\n"
            "| mkt-opt.dwh.incoming.transfer.tab.name | dwd_antifraud_incoming_transfer_detailed_di |\n",
            encoding="utf-8",
        )
        self._build_index_for_entry("AF:All", entry)

        with patch.object(self.service, "_search_repo", wraps=self.service._search_repo) as broad_search:
            payload = self.service.query(
                pm_team="AF",
                country="All",
                question=(
                    "Can you check the relation between "
                    "bmart_antifraud.dwd_antifraud_incoming_transfer_detailed_di and "
                    "tmp_dwd_antifraud_incoming_transfer_detailed_df?"
                ),
            )

        self.assertEqual(payload["status"], "ok")
        self.assertTrue(payload["exact_lookup"]["sufficient"])
        self.assertEqual(
            set(payload["exact_lookup"]["matched_terms"]),
            {
                "bmart_antifraud.dwd_antifraud_incoming_transfer_detailed_di",
                "tmp_dwd_antifraud_incoming_transfer_detailed_df",
            },
        )
        self.assertEqual(broad_search.call_count, 0)
        paths = [match["path"] for match in payload["matches"]]
        self.assertIn(
            "anti-fraud-admin-app/src/main/java/com/shopee/banking/af/admin/app/job/xxl/BaseSyncIncomingDataToIvlog.java",
            paths,
        )
        self.assertIn(
            "anti-fraud-admin-app/src/main/java/com/shopee/banking/af/admin/app/job/xxl/OneTimesSyncDwhIncomingAccountInfoJob.java",
            paths,
        )
        evidence_text = json.dumps(payload["evidence_pack"], ensure_ascii=False)
        self.assertIn("dwd_antifraud_incoming_transfer_detailed_di", evidence_text)
        self.assertIn("tmp_dwd_antifraud_incoming_transfer_detailed_df", evidence_text)

    def test_myinfo_income_logic_does_not_trigger_data_source_from_preposition(self):
        intent = self.service._question_intent("What is the logic to calculate income from MyInfo CPF and NOA?")

        self.assertTrue(intent["rule_logic"])
        self.assertFalse(intent["data_source"])

    def test_sufficient_myinfo_income_logic_skips_agent_plan(self):
        self.service.save_mapping(
            pm_team="CRMS",
            country="SG",
            repositories=[{"display_name": "credit-risk-model", "url": "https://git.example.com/team/credit-risk-model.git"}],
        )
        entry = self.service.load_config()["mappings"]["CRMS:SG"][0]
        repo_path = self.service._repo_path("CRMS:SG", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        feature_file = repo_path / "feature" / "retail_feature_engineering.py"
        feature_file.parent.mkdir(parents=True)
        feature_file.write_text(
            "class RetailFeatureEngineering:\n"
            "    @classmethod\n"
            "    def get_myinfo_monthly_income(self, myInfoData, **kwargs):\n"
            "        myInfoData_dict = myInfoData.copy()\n"
            "        myinfo_monthly_income = self.convert_to_float(myInfoData_dict.get('monthlyIncome'))\n"
            "        if np.isnan(myinfo_monthly_income):\n"
            "            myinfo_monthly_income = 0\n"
            "        return myinfo_monthly_income\n",
            encoding="utf-8",
        )
        self._build_index_for_entry("CRMS:SG", entry)

        with patch.object(self.service, "_run_agent_plan", wraps=self.service._run_agent_plan) as agent_plan:
            payload = self.service.query(
                pm_team="CRMS",
                country="SG",
                question="What is the logic to calculate income from MyInfo CPF and NOA?",
            )

        self.assertEqual(payload["status"], "ok")
        self.assertEqual(agent_plan.call_count, 0)
        self.assertTrue(any(match["path"] == "feature/retail_feature_engineering.py" for match in payload["matches"]))

    def test_index_rebuild_reuses_unchanged_file_rows(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_entry = type("Entry", (), entry)()
        repo_path = self.service._repo_path("AF:All", repo_entry)
        (repo_path / ".git").mkdir(parents=True)
        alpha_file = repo_path / "service" / "AlphaService.java"
        beta_file = repo_path / "service" / "BetaService.java"
        alpha_file.parent.mkdir(parents=True)
        alpha_file.write_text(
            "public class AlphaService {\n"
            "    public String alpha() { return \"alpha\"; }\n"
            "}\n",
            encoding="utf-8",
        )
        beta_file.write_text(
            "public class BetaService {\n"
            "    public String beta() { return \"beta\"; }\n"
            "}\n",
            encoding="utf-8",
        )

        first_result = self.service._build_repo_index("AF:All", repo_entry, repo_path)
        self.assertEqual(first_result["reused_files"], 0)
        self.assertEqual(first_result["reparsed_files"], 2)

        time.sleep(0.01)
        beta_file.write_text(
            "public class BetaServiceV2 {\n"
            "    public String beta() { return \"beta-v2\"; }\n"
            "    public String extra() { return \"extra\"; }\n"
            "}\n",
            encoding="utf-8",
        )
        os.utime(beta_file, None)

        original_extract = self.service._extract_structure_rows
        with patch.object(self.service, "_extract_structure_rows", wraps=original_extract) as mocked_extract:
            second_result = self.service._build_repo_index("AF:All", repo_entry, repo_path)

        self.assertEqual(mocked_extract.call_count, 1)
        self.assertEqual(second_result["reused_files"], 1)
        self.assertEqual(second_result["reparsed_files"], 1)
        index_info = self.service._repo_index_info("AF:All", repo_entry, repo_path)
        self.assertEqual(index_info["reused_files"], 1)
        self.assertEqual(index_info["reparsed_files"], 1)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            names = {
                row[0]
                for row in connection.execute(
                    "select name from definitions where name in ('AlphaService', 'BetaServiceV2')"
                ).fetchall()
            }
            line_count = connection.execute("select count(*) from lines").fetchone()[0]
        self.assertIn("AlphaService", names)
        self.assertIn("BetaServiceV2", names)
        self.assertGreaterEqual(line_count, 7)

    def test_ops_summary_reports_recent_source_qa_signals(self):
        from scripts.source_code_qa_ops_summary import build_summary

        data_root = Path(self.temp_dir.name)
        source_root = data_root / "source_code_qa"
        run_root = data_root / "run"
        source_root.mkdir(parents=True)
        run_root.mkdir(parents=True)
        (source_root / "telemetry.jsonl").write_text(
            json.dumps(
                {
                    "created_at": "2026-04-25T10:00:00+08:00",
                    "status": "ok",
                    "latency_ms": 120,
                    "llm_route": {"mode": "hybrid"},
                    "llm_provider": "codex_cli_bridge",
                    "llm_budget_mode": "cheap",
                    "llm_model": "gpt-5.4-mini",
                    "llm_usage": {"prompt_tokens": 3000},
                    "llm_cached": True,
                    "answer_contract": {"status": "satisfied"},
                }
            )
            + "\n"
            + json.dumps(
                {
                    "created_at": "2026-04-25T10:01:00+08:00",
                    "status": "no_match",
                    "latency_ms": 360,
                    "llm_route": {"mode": "retrieval"},
                    "llm_provider": "codex_cli_bridge",
                    "llm_budget_mode": "deep",
                    "llm_model": "gpt-5.5",
                    "llm_usage": {"prompt_tokens": 18000},
                    "llm_cached": False,
                    "answer_contract": {"status": "insufficient_evidence"},
                    "index_freshness": {"status": "stale_or_missing"},
                }
            )
            + "\n",
            encoding="utf-8",
        )
        (source_root / "feedback.jsonl").write_text(
            json.dumps({"created_at": "2026-04-25T10:02:00+08:00", "rating": "needs_deeper_trace"}) + "\n",
            encoding="utf-8",
        )
        summary = "\n".join(build_summary(data_root, limit=20))

        self.assertIn("telemetry_window=2", summary)
        self.assertIn("query_status=ok=1, no_match=1", summary)
        self.assertIn("llm_budgets=cheap=1, deep=1", summary)
        self.assertIn("llm_models=gpt-5.4-mini=1, gpt-5.5=1", summary)
        self.assertIn("prompt_token_bands=lt6k=1, 12k_24k=1", summary)
        self.assertIn("llm_cache_hits=1/2", summary)
        self.assertIn("no_match_rate=1/2", summary)
        self.assertIn("feedback_window=1", summary)
        self.assertIn("review_queue=0", summary)
        self.assertNotIn("latest_eval_state", summary)

    def test_source_code_qa_warm_answer_cache_selects_recent_slow_uncached_questions(self):
        from scripts.source_code_qa_warm_answer_cache import _recent_slow_questions

        data_root = Path(self.temp_dir.name)
        source_root = data_root / "source_code_qa"
        source_root.mkdir(parents=True)
        (source_root / "telemetry.jsonl").write_text(
            "\n".join(
                [
                    json.dumps({"key": "AF:All", "question_preview": "fast cached", "latency_ms": 1000, "llm_cached": False}),
                    json.dumps({"key": "AF:All", "question_preview": "slow cached", "latency_ms": 90000, "llm_cached": True}),
                    json.dumps({"key": "GRC:All", "question_preview": "slow uncached", "latency_ms": 91000, "llm_cached": False}),
                    json.dumps(
                        {
                            "key": "CRMS:SG",
                            "question_preview": "attributed slow uncached",
                            "latency_ms": 1000,
                            "llm_cached": False,
                            "slow_query_attribution": {"status": "slow", "cache_preload_recommended": True},
                        }
                    ),
                ]
            )
            + "\n",
            encoding="utf-8",
        )

        questions = _recent_slow_questions(data_root, limit=5, min_latency_ms=30000)

        self.assertEqual(
            questions,
            [
                {"pm_team": "GRC", "country": "All", "question": "slow uncached"},
                {"pm_team": "CRMS", "country": "SG", "question": "attributed slow uncached"},
            ],
        )

    def test_ops_summary_flags_fixture_repo_config_in_strict_mode(self):
        from scripts.source_code_qa_ops_summary import build_summary

        data_root = Path(self.temp_dir.name)
        source_root = data_root / "source_code_qa"
        source_root.mkdir(parents=True)
        (source_root / "config.json").write_text(
            json.dumps(
                {
                    "version": 1,
                    "mappings": {
                        "CRMS:SG": [
                            {"display_name": "Credit Risk SG", "url": "https://git.example.com/team/credit-risk-sg.git"}
                        ]
                    },
                }
            ),
            encoding="utf-8",
        )

        summary = "\n".join(build_summary(data_root, limit=20, strict=True))

        self.assertIn("active_config_demo_repos=1", summary)
        self.assertIn("ops_summary_status=fail", summary)
        self.assertIn("fixture/demo repositories", summary)

    def test_ops_summary_uses_local_agent_source_code_qa_data_root_when_enabled(self):
        from scripts.source_code_qa_ops_summary import build_summary

        requested_root = Path(self.temp_dir.name) / "portal-data"
        requested_source_root = requested_root / "source_code_qa"
        requested_source_root.mkdir(parents=True)
        (requested_source_root / "config.json").write_text(
            json.dumps(
                {
                    "version": 1,
                    "mappings": {
                        "CRMS:SG": [
                            {"display_name": "Fixture", "url": "https://git.example.com/team/fixture.git"}
                        ]
                    },
                }
            ),
            encoding="utf-8",
        )

        agent_root = Path(self.temp_dir.name) / "agent-data"
        agent_source_root = agent_root / "source_code_qa"
        agent_source_root.mkdir(parents=True)
        (agent_source_root / "config.json").write_text(
            json.dumps(
                {
                    "version": 1,
                    "mappings": {
                        "CRMS:SG": [
                            {"display_name": "credit-risk", "url": "https://gitlab.npt.seabank.io/team/credit-risk.git"}
                        ]
                    },
                }
            ),
            encoding="utf-8",
        )

        with patch.dict(
            os.environ,
            {
                "LOCAL_AGENT_SOURCE_CODE_QA_ENABLED": "true",
                "LOCAL_AGENT_TEAM_PORTAL_DATA_DIR": str(agent_root),
            },
        ), patch("bpmis_jira_tool.source_code_qa.SourceCodeQAService.index_health_payload") as health:
            health.return_value = {
                "status": "ready",
                "totals": {"repos": 1, "ready": 1, "stale_or_missing": 0},
                "keys": {
                    "CRMS:SG": {
                        "repos": [
                            {
                                "display_name": "credit-risk",
                                "url": "https://gitlab.npt.seabank.io/team/credit-risk.git",
                                "index": {"state": "ready", "index_version": CODE_INDEX_VERSION},
                            }
                        ]
                    }
                },
            }
            summary = "\n".join(build_summary(requested_root, limit=20, strict=True, prefer_local_agent=True))

        self.assertIn(f"data_root={agent_root.resolve()}", summary)
        self.assertIn(f"requested_data_root={requested_root.resolve()}", summary)
        self.assertIn("data_root_resolution=local_agent_source_code_qa", summary)
        self.assertIn("active_config_demo_repos=0", summary)
        self.assertIn("ops_summary_status=pass", summary)
        self.assertNotIn("fixture/demo repositories", summary)

    def test_rebuild_indexes_helper_flags_non_current_ready_indexes(self):
        from scripts.source_code_qa_rebuild_indexes import _verify_health

        class FakeService:
            def index_health_payload(self):
                return {
                    "status": "ready",
                    "keys": {
                        "AF:All": {
                            "repos": [
                                {
                                    "display_name": "AF repo",
                                    "index": {"state": "ready", "index_version": CODE_INDEX_VERSION - 1},
                                }
                            ]
                        }
                    },
                }

        issues, health = _verify_health(FakeService())

        self.assertEqual(health["status"], "ready")
        self.assertEqual(issues, [f"AF:All:AF repo: index_version={CODE_INDEX_VERSION - 1} expected={CODE_INDEX_VERSION}"])

    def test_rebuild_indexes_helper_backs_up_existing_index_root(self):
        from scripts.source_code_qa_rebuild_indexes import _backup_indexes

        source_root = Path(self.temp_dir.name) / "source_code_qa"
        index_root = source_root / "indexes"
        index_root.mkdir(parents=True)
        (index_root / "repo.sqlite").write_text("index", encoding="utf-8")

        backup_path = _backup_indexes(source_root)

        self.assertIsNotNone(backup_path)
        self.assertTrue(str(backup_path.name).startswith(f"indexes.backup.v{CODE_INDEX_VERSION}."))
        self.assertEqual((backup_path / "repo.sqlite").read_text(encoding="utf-8"), "index")

    def test_rebuild_indexes_helper_dry_run_lists_orphan_indexes_without_deleting(self):
        from scripts.source_code_qa_rebuild_indexes import _scan_orphan_index_files

        index_root = Path(self.temp_dir.name) / "source_code_qa" / "indexes"
        index_root.mkdir(parents=True)
        active = index_root / "active.sqlite3"
        orphan = index_root / "orphan.sqlite3"
        non_index = index_root / "notes.txt"
        active.write_bytes(b"active")
        orphan.write_bytes(b"orphan-index")
        non_index.write_text("not an index", encoding="utf-8")

        result = _scan_orphan_index_files(index_root, {active})

        self.assertEqual(result["active_count"], 1)
        self.assertEqual(result["orphan_count"], 1)
        self.assertEqual(result["active_bytes"], len(b"active"))
        self.assertEqual(result["orphan_bytes"], len(b"orphan-index"))
        self.assertEqual(result["orphan_files"][0]["name"], "orphan.sqlite3")
        self.assertTrue(active.exists())
        self.assertTrue(orphan.exists())
        self.assertTrue(non_index.exists())

    def test_cleanup_orphan_indexes_deletes_only_non_active_index_files(self):
        from scripts.source_code_qa_rebuild_indexes import cleanup_orphan_indexes

        index_root = Path(self.temp_dir.name) / "source_code_qa" / "indexes"
        index_root.mkdir(parents=True)
        active = index_root / "active.sqlite3"
        orphan = index_root / "orphan.sqlite3"
        orphan_db = index_root / "orphan.db"
        wal = index_root / "orphan.sqlite3-wal"
        active.write_bytes(b"active")
        orphan.write_bytes(b"orphan")
        orphan_db.write_bytes(b"db")
        wal.write_bytes(b"wal")

        class FakeService:
            def __init__(self):
                self.index_root = index_root

            def load_config(self):
                return {"mappings": {"CRMS:SG": [{"url": "https://git.example.com/team/active.git"}]}}

            def _normalize_entry(self, raw):
                return raw

            def _repo_path(self, key, entry):
                return Path(self.load_config()["mappings"][key][0]["url"])

            def _index_path(self, repo_path):
                return active

        with patch("scripts.source_code_qa_rebuild_indexes.source_code_qa_data_root", return_value=Path(self.temp_dir.name)), patch(
            "scripts.source_code_qa_rebuild_indexes.build_source_code_qa_service_from_settings",
            return_value=FakeService(),
        ):
            result = cleanup_orphan_indexes(SimpleNamespace(), delete=True)

        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["active_count"], 1)
        self.assertEqual(result["orphan_count"], 2)
        self.assertEqual(result["deleted_count"], 2)
        self.assertTrue(active.exists())
        self.assertFalse(orphan.exists())
        self.assertFalse(orphan_db.exists())
        self.assertTrue(wal.exists())

    def test_domain_profiles_include_domain_knowledge_pack_terms(self):
        crms_profile = self.service._domain_profile("CRMS", "SG")
        af_profile = self.service._domain_profile("AF", "All")
        grc_profile = self.service._domain_profile("GRC", "All")
        knowledge = self.service.domain_knowledge_payload()

        self.assertIn("TermLoanPreCheckEngine", crms_profile["source_terms"])
        self.assertIn("Term Loan PreCheck", crms_profile["knowledge_terms"])
        self.assertIn("BlackWhiteList", af_profile["data_carriers"])
        self.assertIn("CaseReviewController", af_profile["api_terms"])
        self.assertIn("mysql-isolate.globallock", grc_profile["config_terms"])
        self.assertIn("RC", grc_profile["knowledge_terms"])
        self.assertIn("Compliance", grc_profile["knowledge_terms"])
        self.assertIn("ApprovalController", grc_profile["api_terms"])
        self.assertEqual(knowledge["domains"]["CRMS"]["label"], "Credit Risk")
        self.assertGreaterEqual(knowledge["domains"]["AF"]["module_count"], 5)
        self.assertGreaterEqual(knowledge["domains"]["GRC"]["question_count"], 4)

    def test_llm_domain_context_includes_team_rules_and_answer_blueprint(self):
        crms_context = self.service._llm_domain_context(
            pm_team="CRMS",
            country="SG",
            question="What is the logic to calculate income from MyInfo CPF and NOA?",
            evidence_summary={"intent": self.service._question_intent("What data source is used for MyInfo income?")},
        )
        af_context = self.service._llm_domain_context(
            pm_team="AF",
            country="All",
            question="Can you check the relation between dwd_antifraud_incoming_transfer_detailed_di and tmp_dwd_antifraud_incoming_transfer_detailed_df?",
            evidence_summary={"intent": self.service._question_intent("what data source relation exists between two tables?")},
        )
        grc_context = self.service._llm_domain_context(
            pm_team="GRC",
            country="All",
            question="GRC globallock 配置在哪里？",
            evidence_summary={"intent": self.service._question_intent("where is globallock config?")},
        )

        self.assertIn("Domain guidance: Credit Risk", crms_context)
        self.assertIn("SG MyInfo Income", crms_context)
        self.assertIn("final source/table/API", crms_context)
        self.assertIn("Domain guidance: Anti-Fraud", af_context)
        self.assertIn("co-occurrence in the same flow", af_context)
        self.assertIn("Domain guidance: Ops Risk", grc_context)
        self.assertIn("mysql-isolate.globallock", grc_context)
        self.assertIn("RC and Compliance as business aliases", grc_context)
        self.assertIn("Evidence priority", grc_context)

    def test_index_health_payload_summarizes_ready_indexes(self):
        entry = RepositoryEntry(display_name="Portal Repo", url="https://git.example.com/team/portal.git")
        self.service.save_mapping(pm_team="AF", country="All", repositories=[{"display_name": entry.display_name, "url": entry.url}])
        repo_path = self.service._repo_path("AF:All", entry)
        (repo_path / ".git").mkdir(parents=True)
        (repo_path / "IssueService.java").write_text(
            "class IssueService { void createIssue() { repository.save(); } }\n",
            encoding="utf-8",
        )
        self.service._build_repo_index("AF:All", entry, repo_path)

        health = self.service.index_health_payload()

        self.assertEqual(health["status"], "ready")
        self.assertEqual(health["totals"]["repos"], 1)
        self.assertEqual(health["totals"]["ready"], 1)
        self.assertIn("AF:All", health["keys"])

    def test_release_gate_evaluator_enforces_case_floor(self):
        from scripts.run_source_code_qa_release_gate import evaluate_release_gate

        gate = evaluate_release_gate(
            {
                "status": "pass",
                "eval": {
                    "status": "pass",
                    "total": 4,
                    "failed": 0,
                    "team_buckets": {"AF": {"total": 4, "failed": 0}},
                    "segment_buckets": {"AF:ALL": {"total": 4, "failed": 0}},
                },
                "llm_smoke": {"status": "pass", "failed": 0},
                "review_queue": {"returncode": 0},
            },
            thresholds={"min_eval_cases": 10, "required_eval_teams": ["AF"], "required_eval_segments": ["AF:ALL"]},
        )

        self.assertEqual(gate["status"], "fail")
        self.assertIn("min_eval_cases", gate["failed_checks"])

    def test_release_gate_evaluator_requires_team_balanced_coverage(self):
        from scripts.run_source_code_qa_release_gate import evaluate_release_gate

        gate = evaluate_release_gate(
            {
                "status": "pass",
                "eval": {
                    "status": "pass",
                    "total": 8,
                    "failed": 0,
                    "team_buckets": {
                        "AF": {"total": 4, "failed": 0},
                        "CRMS": {"total": 4, "failed": 0},
                    },
                    "segment_buckets": {
                        "AF:ALL": {"total": 4, "failed": 0},
                        "CRMS:ID": {"total": 4, "failed": 0},
                    },
                },
                "llm_smoke": {"status": "pass", "failed": 0},
                "review_queue": {"returncode": 0},
            },
            thresholds={
                "min_eval_cases": 8,
                "min_eval_cases_per_team": 4,
                "required_eval_segments": ["AF:ALL", "CRMS:ID"],
                "required_eval_teams": ["AF", "CRMS", "GRC"],
            },
        )

        self.assertEqual(gate["status"], "fail")
        self.assertIn("eval_team_coverage", gate["failed_checks"])
        self.assertEqual(gate["missing_or_thin_teams"], ["GRC"])

    def test_release_gate_evaluator_requires_country_segment_coverage(self):
        from scripts.run_source_code_qa_release_gate import evaluate_release_gate

        gate = evaluate_release_gate(
            {
                "status": "pass",
                "eval": {
                    "status": "pass",
                    "total": 12,
                    "failed": 0,
                    "team_buckets": {
                        "AF": {"total": 4, "failed": 0},
                        "CRMS": {"total": 4, "failed": 0},
                        "GRC": {"total": 4, "failed": 0},
                    },
                    "segment_buckets": {
                        "AF:ALL": {"total": 4, "failed": 0},
                        "CRMS:ID": {"total": 4, "failed": 0},
                        "GRC:ALL": {"total": 4, "failed": 0},
                    },
                },
                "llm_smoke": {"status": "pass", "failed": 0},
                "review_queue": {"returncode": 0},
            },
            thresholds={
                "min_eval_cases": 12,
                "min_eval_cases_per_segment": 2,
                "required_eval_segments": ["AF:ALL", "CRMS:ID", "CRMS:SG", "CRMS:PH", "GRC:ALL"],
            },
        )

        self.assertEqual(gate["status"], "fail")
        self.assertIn("eval_segment_coverage", gate["failed_checks"])
        self.assertEqual(gate["missing_or_thin_segments"], ["CRMS:SG", "CRMS:PH"])

    def test_release_gate_evaluator_accepts_all_required_teams(self):
        from scripts.run_source_code_qa_release_gate import evaluate_release_gate

        gate = evaluate_release_gate(
            {
                "status": "pass",
                "eval": {
                    "status": "pass",
                    "total": 12,
                    "failed": 0,
                    "team_buckets": {
                        "AF": {"total": 4, "failed": 0},
                        "CRMS": {"total": 6, "failed": 0},
                        "GRC": {"total": 4, "failed": 0},
                    },
                    "segment_buckets": {
                        "AF:ALL": {"total": 4, "failed": 0},
                        "CRMS:ID": {"total": 2, "failed": 0},
                        "CRMS:SG": {"total": 2, "failed": 0},
                        "CRMS:PH": {"total": 2, "failed": 0},
                        "GRC:ALL": {"total": 4, "failed": 0},
                    },
                },
                "llm_smoke": {"status": "pass", "failed": 0},
                "review_queue": {"returncode": 0},
            },
            thresholds={"min_eval_cases": 12, "min_eval_cases_per_team": 4},
        )

        self.assertEqual(gate["status"], "pass")
        self.assertEqual(gate["missing_or_thin_teams"], [])
        self.assertEqual(gate["missing_or_thin_segments"], [])

    def test_release_eval_report_uses_isolated_data_root(self):
        from scripts.run_source_code_qa_release_gate import run_release_eval_report

        calls = []

        def fake_run(args):
            calls.append(args)
            if any(str(arg).endswith("source_code_qa_evals.py") for arg in args):
                return {
                    "status": "pass",
                    "total": 1,
                    "failed": 0,
                    "team_buckets": {"AF": {"total": 1, "failed": 0}},
                    "segment_buckets": {"AF:ALL": {"total": 1, "failed": 0}},
                }, "{}", "", 0
            return {"status": "ok", "review_items": 0}, "{}", "", 0

        output_dir = Path(self.temp_dir.name) / "eval_runs"
        with patch("scripts.run_source_code_qa_release_gate._run_json_command", side_effect=fake_run):
            report = run_release_eval_report(output_dir=output_dir, cases=["evals/source_code_qa/golden.jsonl"], fixture=True, include_useful_feedback=False)

        self.assertEqual(report["status"], "pass")
        self.assertEqual(report["eval"]["team_buckets"]["AF"]["total"], 1)
        self.assertEqual(report["eval"]["segment_buckets"]["AF:ALL"]["total"], 1)
        eval_call = calls[0]
        self.assertIn("--data-root", eval_call)
        self.assertEqual(Path(eval_call[eval_call.index("--data-root") + 1]), output_dir / "fixture_data")

    def test_release_gate_uses_mock_llm_by_default(self):
        from scripts.run_source_code_qa_release_gate import run_release_gate

        captured = {}

        def fake_run_release_eval_report(**kwargs):
            captured.update(kwargs)
            return {
                "status": "pass",
                "eval": {
                    "status": "pass",
                    "total": 1,
                    "failed": 0,
                    "team_buckets": {"AF": {"total": 1, "failed": 0}},
                    "segment_buckets": {"AF:ALL": {"total": 1, "failed": 0}},
                },
                "llm_smoke": {"status": "pass", "failed": 0},
                "review_queue": {"returncode": 0},
            }

        with patch("scripts.run_source_code_qa_release_gate.run_release_eval_report", side_effect=fake_run_release_eval_report):
            gate = run_release_gate(
                data_root=Path(self.temp_dir.name),
                cases=["evals/source_code_qa/llm_smoke.jsonl"],
                fixture=True,
                include_useful_feedback=False,
                thresholds={
                    "min_eval_cases": 1,
                    "min_eval_cases_per_team": 1,
                    "min_eval_cases_per_segment": 1,
                    "required_eval_teams": ["AF"],
                    "required_eval_segments": ["AF:ALL"],
                },
            )

        self.assertEqual(gate["status"], "pass")
        self.assertTrue(captured["mock_llm"])

    def test_release_eval_report_can_pass_mock_llm_to_main_eval(self):
        from scripts.run_source_code_qa_release_gate import run_release_eval_report

        calls = []

        def fake_run(args):
            calls.append(args)
            if any(str(arg).endswith("source_code_qa_evals.py") for arg in args):
                return {
                    "status": "pass",
                    "total": 1,
                    "failed": 0,
                    "team_buckets": {"AF": {"total": 1, "failed": 0}},
                    "segment_buckets": {"AF:ALL": {"total": 1, "failed": 0}},
                }, "{}", "", 0
            return {"status": "ok", "review_items": 0}, "{}", "", 0

        output_dir = Path(self.temp_dir.name) / "eval_runs"
        with patch("scripts.run_source_code_qa_release_gate._run_json_command", side_effect=fake_run):
            report = run_release_eval_report(
                output_dir=output_dir,
                cases=["evals/source_code_qa/golden.jsonl"],
                fixture=True,
                include_useful_feedback=False,
                mock_llm=True,
            )

        self.assertTrue(report["mock_llm"])
        self.assertIn("--mock-llm", calls[0])

    def test_eval_summary_groups_latency_cache_fallback_and_answer_mode_buckets(self):
        summary = _summarize_results(
            [
                {
                    "status": "pass",
                    "category": "telemetry_slow_query",
                    "pm_team": "GRC",
                    "country": "All",
                    "answer_mode": "retrieval_only",
                    "fallback_used": False,
                    "llm_cached": False,
                    "slow_query_attribution": {"status": "slow"},
                    "failure_buckets": {},
                    "llm_route": {},
                },
                {
                    "status": "fail",
                    "category": "telemetry_deadline_fallback",
                    "pm_team": "CRMS",
                    "country": "SG",
                    "answer_mode": "auto",
                    "fallback_used": True,
                    "llm_cached": True,
                    "slow_query_attribution": {"status": "ok"},
                    "failure_buckets": {"answer_policy": 1},
                    "llm_route": {"selected": "auto"},
                },
            ]
        )

        self.assertEqual(summary["answer_mode_buckets"]["retrieval_only"]["total"], 1)
        self.assertEqual(summary["fallback_buckets"]["fallback"]["failed"], 1)
        self.assertEqual(summary["cache_buckets"]["cache_hit"]["total"], 1)
        self.assertEqual(summary["slow_query_buckets"]["slow"]["total"], 1)
        self.assertEqual(summary["segment_buckets"]["CRMS:SG"]["failed"], 1)

    def test_broad_eval_records_auto_candidate_failures_without_blocking_stable_layer(self):
        from scripts.run_source_code_qa_broad_eval import run_broad_eval

        data_root = Path(self.temp_dir.name)
        source_root = data_root / "source_code_qa"
        source_root.mkdir(parents=True, exist_ok=True)
        (source_root / "telemetry.jsonl").write_text(
            json.dumps(
                {
                    "key": "AF:All",
                    "question_preview": "where is slow flow",
                    "question_sha1": "slow-flow",
                    "status": "ok",
                    "latency_ms": 40000,
                    "top_paths": ["service/IssueService.java"],
                    "slow_query_attribution": {"status": "slow"},
                }
            )
            + "\n",
            encoding="utf-8",
        )

        calls = []

        def fake_run(args):
            calls.append(args)
            if any("auto_eval_candidates" in str(arg) for arg in args):
                return {"status": "fail", "total": 1, "failed": 1, "failure_buckets": {"retrieval": 1}}, "{}", "", 1
            return {
                "status": "pass",
                "total": 2,
                "failed": 0,
                "team_buckets": {"AF": {"total": 2, "failed": 0}},
                "segment_buckets": {"AF:ALL": {"total": 2, "failed": 0}},
                "answer_mode_buckets": {"retrieval_only": {"total": 2, "failed": 0}},
            }, "{}", "", 0

        with patch("scripts.run_source_code_qa_broad_eval._run_json_command", side_effect=fake_run):
            report = run_broad_eval(data_root=data_root, output_dir=data_root / "source_code_qa" / "eval_runs")

        self.assertEqual(report["status"], "ok")
        self.assertEqual(report["broad_quality_status"], "warn")
        self.assertEqual(report["auto_candidates"]["runnable_candidates"], 1)
        self.assertEqual(report["auto_eval"]["status"], "fail")
        self.assertTrue((data_root / "run" / "source_code_qa_broad_eval.json").exists())

    def test_fixture_eval_requires_isolated_data_root(self):
        data_root = Path(self.temp_dir.name)

        with self.assertRaises(SystemExit):
            _guard_fixture_data_root(data_root=data_root, main_data_root=data_root, data_root_explicit=False)

        with self.assertRaises(SystemExit):
            _guard_fixture_data_root(data_root=data_root, main_data_root=data_root, data_root_explicit=True)

        _guard_fixture_data_root(data_root=data_root / "fixture_data", main_data_root=data_root, data_root_explicit=True)

    def test_review_queue_collects_feedback_and_telemetry_risks(self):
        from scripts.source_code_qa_review_queue import build_review_queue

        data_root = Path(self.temp_dir.name)
        source_root = data_root / "source_code_qa"
        source_root.mkdir(parents=True)
        (source_root / "feedback.jsonl").write_text(
            json.dumps(
                {
                    "timestamp": "2026-04-25T10:05:00+08:00",
                    "rating": "wrong_file",
                    "pm_team": "AF",
                    "country": "All",
                    "question_preview": "where is batch create",
                    "question_sha1": "abc123",
                    "trace_id": "trace-feedback",
                    "replay_context": {
                        "answer_contract": {"status": "grounded"},
                        "matches_snapshot": [{"path": "wrong/File.java"}],
                    },
                }
            )
            + "\n",
            encoding="utf-8",
        )
        (source_root / "telemetry.jsonl").write_text(
            json.dumps(
                {
                    "timestamp": "2026-04-25T10:06:00+08:00",
                    "status": "no_match",
                    "key": "AF:All",
                    "question_preview": "where is missing flow",
                    "question_sha1": "def456",
                    "trace_id": "trace-telemetry",
                    "top_paths": [],
                    "index_freshness": {"status": "fresh"},
                }
            )
            + "\n",
            encoding="utf-8",
        )

        queue = build_review_queue(data_root)

        self.assertEqual(len(queue), 2)
        self.assertEqual(queue[0]["priority"], "high")
        self.assertEqual({item["source"] for item in queue}, {"feedback", "telemetry"})
        feedback_item = next(item for item in queue if item["source"] == "feedback")
        self.assertEqual(feedback_item["observed_paths"], ["wrong/File.java"])
        self.assertEqual(feedback_item["draft_eval"]["draft_status"], "needs_human_expected_evidence")

    def test_planner_tool_loop_adds_code_graph_matches(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "public class IssueController {\n"
            "    private IssueService issueService;\n"
            "    public Issue createIssue() {\n"
            "        return issueService.createIssue();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        service_file = repo_path / "service" / "IssueService.java"
        service_file.parent.mkdir(parents=True)
        service_file.write_text(
            "public class IssueService {\n"
            "    private IssueRepository issueRepository;\n"
            "    public Issue createIssue() {\n"
            "        return issueRepository.findIssue();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        repository_file = repo_path / "repository" / "IssueRepository.java"
        repository_file.parent.mkdir(parents=True)
        repository_file.write_text(
            "public class IssueRepository {\n"
            "    public Issue findIssue() {\n"
            "        return jdbcTemplate.queryForObject(\"select * from issue_table\", mapper);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="trace issue create API data source")

        retrievals = {match.get("retrieval") for match in payload["matches"]}
        paths = {match["path"] for match in payload["matches"]}
        trace_stages = {match.get("trace_stage") for match in payload["matches"]}
        self.assertTrue({"planner_definition", "planner_reference", "code_graph", "flow_graph", "entity_graph"} & retrievals)
        self.assertIn("repository/IssueRepository.java", paths)
        self.assertTrue(any(str(stage).startswith("tool_loop_") for stage in trace_stages) or "semantic_chunk" in retrievals)
        self.assertTrue(payload["tool_trace"])
        self.assertTrue(any(step.get("phase") in {"tool_loop", "agent_plan"} for step in payload["tool_trace"]))
        telemetry = json.loads(self.service.telemetry_path.read_text(encoding="utf-8").strip().splitlines()[-1])
        self.assertGreaterEqual(telemetry["tool_trace_summary"]["steps"], 1)

    def test_python_ast_index_extracts_entities_and_calls(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "services" / "issue_service.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "from bpmis.client import BPMISClient\n"
            "class IssueService:\n"
            "    def create_issue(self):\n"
            "        client = BPMISClient()\n"
            "        return client.batchCreateJiraIssue()\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="where does create_issue call BPMIS")
        index_path = self.service._index_path(repo_path)

        self.assertEqual(payload["status"], "ok")
        with sqlite3.connect(index_path) as connection:
            entity_names = {row[0] for row in connection.execute("select name from code_entities")}
            edge_targets = {row[0] for row in connection.execute("select to_name from entity_edges")}
        self.assertIn("IssueService", entity_names)
        self.assertIn("create_issue", entity_names)
        self.assertTrue(any("batchCreateJiraIssue" in target for target in edge_targets))

    def test_flow_graph_can_trace_controller_service_repository_table_chain(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "@PostMapping(\"/issue/create\")\n"
            "public class IssueController {\n"
            "    private IssueService issueService;\n"
            "    public Issue createIssue() {\n"
            "        return issueService.createIssue();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        service_file = repo_path / "service" / "IssueService.java"
        service_file.parent.mkdir(parents=True)
        service_file.write_text(
            "public class IssueService {\n"
            "    private IssueRepository issueRepository;\n"
            "    public Issue createIssue() {\n"
            "        return issueRepository.findIssue();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        repository_file = repo_path / "repository" / "IssueRepository.java"
        repository_file.parent.mkdir(parents=True)
        repository_file.write_text(
            "public class IssueRepository {\n"
            "    public Issue findIssue() {\n"
            "        return jdbcTemplate.queryForObject(\"select * from issue_table\", mapper);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="which API flow reaches issue data source table")

        self.assertEqual(payload["status"], "ok")
        self.assertIn("repository/IssueRepository.java", {match["path"] for match in payload["matches"]})
        self.assertTrue({"flow_graph", "entity_graph"} & {match.get("retrieval") for match in payload["matches"]})
        self.assertTrue(any("issue_table" in match["snippet"] for match in payload["matches"]))
        self.assertTrue(payload["trace_paths"])
        self.assertIn("issue_table", str(payload["trace_paths"]))

    def test_framework_adapters_extract_spring_mybatis_and_feign_edges(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "public class IssueController {\n"
            "    @PostMapping(\"/issue/create\")\n"
            "    public Issue createIssue() {\n"
            "        return issueClient.createIssue();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        client_file = repo_path / "client" / "IssueClient.java"
        client_file.parent.mkdir(parents=True)
        client_file.write_text(
            "@FeignClient(name = \"issue-service\", url = \"https://issue.example.com\")\n"
            "public interface IssueClient {\n"
            "    @PostMapping(\"/remote/issue\")\n"
            "    Issue createIssue();\n"
            "}\n",
            encoding="utf-8",
        )
        mapper_file = repo_path / "mapper" / "IssueMapper.xml"
        mapper_file.parent.mkdir(parents=True)
        mapper_file.write_text(
            "<mapper namespace=\"com.example.IssueMapper\">\n"
            "  <resultMap id=\"IssueResult\" type=\"com.example.IssueEntity\">\n"
            "    <result property=\"id\" column=\"id\" javaType=\"java.lang.Long\" />\n"
            "  </resultMap>\n"
            "  <sql id=\"BaseColumns\">id, name</sql>\n"
            "  <select id=\"findIssue\" parameterType=\"com.example.IssueQuery\" resultMap=\"IssueResult\">\n"
            "    <include refid=\"BaseColumns\" />\n"
            "    select * from issue_table\n"
            "  </select>\n"
            "</mapper>\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="trace create issue API to downstream service and table")
        index_path = self.service._index_path(repo_path)

        self.assertEqual(payload["status"], "ok")
        self.assertTrue(payload["trace_paths"])
        with sqlite3.connect(index_path) as connection:
            edge_rows = list(connection.execute("select edge_kind, to_name from entity_edges"))
        self.assertIn(("downstream_api", "issue-service"), edge_rows)
        self.assertIn(("http_endpoint", "https://issue.example.com"), edge_rows)
        self.assertTrue(any(kind == "mapper_statement" and "findIssue" in target for kind, target in edge_rows))
        self.assertTrue(any(kind == "result_map" and "IssueResult" in target for kind, target in edge_rows))
        self.assertTrue(any(kind == "mybatis_result_map_ref" and target == "IssueResult" for kind, target in edge_rows))
        self.assertTrue(any(kind == "mybatis_include_refid" and target == "BaseColumns" for kind, target in edge_rows))
        self.assertTrue(any(kind == "mybatis_type_ref" and target == "com.example.IssueEntity" for kind, target in edge_rows))
        self.assertTrue(any(kind == "sql_table" and target == "issue_table" for kind, target in edge_rows))

    def test_cross_repo_graph_links_client_to_service_repo(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[
                {"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"},
                {"display_name": "Issue Service", "url": "https://git.example.com/team/issue-service.git"},
            ],
        )
        entries = self.service.load_config()["mappings"]["AF:All"]
        portal = type("Entry", (), entries[0])()
        service_entry = type("Entry", (), entries[1])()
        portal_path = self.service._repo_path("AF:All", portal)
        service_path = self.service._repo_path("AF:All", service_entry)
        (portal_path / ".git").mkdir(parents=True)
        (service_path / ".git").mkdir(parents=True)
        client_file = portal_path / "client" / "IssueClient.java"
        client_file.parent.mkdir(parents=True)
        client_file.write_text(
            "@FeignClient(name = \"issue-service\")\n"
            "public interface IssueClient { Issue createIssue(); }\n",
            encoding="utf-8",
        )
        service_file = service_path / "controller" / "IssueController.java"
        service_file.parent.mkdir(parents=True)
        service_file.write_text(
            "@PostMapping(\"/issue/create\")\n"
            "public class IssueController { public Issue createIssue() { return new Issue(); } }\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="which service does issue client call")

        self.assertEqual(payload["repo_graph"]["version"], 2)
        edges = payload["repo_graph"]["edges"]
        graph_edge = next(
            edge for edge in edges if edge["from_repo"] == "Portal Repo" and edge["to_repo"] == "Issue Service"
        )
        self.assertIn("confidence", graph_edge)
        self.assertIn("match_reason", graph_edge)
        self.assertGreaterEqual(graph_edge["confidence"], 0.7)

    def test_tree_sitter_java_extracts_injection_routes_and_calls(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "controller" / "IssueController.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "@RestController\n"
            "public class IssueController {\n"
            "    private IssueService issueService;\n"
            "    @PostMapping(\"/issue/create\")\n"
            "    public Issue createIssue() {\n"
            "        return issueService.createIssue();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )

        index_info = self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        self.assertEqual(index_info["parser_backend"], "tree_sitter+regex")
        self.assertIn("java", index_info["parser_languages"])
        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            rows = list(connection.execute("select edge_kind, to_name from entity_edges"))
        self.assertIn(("injects", "IssueService"), rows)
        self.assertIn(("route", "/issue/create"), rows)
        self.assertTrue(any(kind == "call" and "createIssue" in target for kind, target in rows))

    def test_spring_route_prefix_and_constructor_injection_are_indexed(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "controller" / "IssueController.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "@RestController\n"
            "@RequestMapping(\"/api/issues\")\n"
            "public class IssueController {\n"
            "    private final IssueService issueService;\n"
            "    public IssueController(IssueService issueService) {\n"
            "        this.issueService = issueService;\n"
            "    }\n"
            "    @PostMapping(\"/create\")\n"
            "    public Issue createIssue() {\n"
            "        return issueService.createIssue();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            rows = list(connection.execute("select edge_kind, to_name from entity_edges"))
        self.assertIn(("route", "/api/issues/create"), rows)
        self.assertIn(("injects", "IssueService"), rows)

    def test_extract_structure_rows_characterizes_spring_route_config_and_sql_edges(self):
        source = (
            "@RestController\n"
            "@RequestMapping(\"/api/issues\")\n"
            "@ConditionalOnProperty(prefix = \"issue\", name = \"enabled\", havingValue = \"true\")\n"
            "public class IssueController {\n"
            "    @Qualifier(\"primaryIssueService\")\n"
            "    private IssueService issueService;\n"
            "    @Transactional(rollbackFor = Exception.class)\n"
            "    @PostMapping(\"/create\")\n"
            "    public Issue createIssue() {\n"
            "        return jdbcTemplate.queryForObject(\"select * from issue_table\", mapper);\n"
            "    }\n"
            "}\n"
        )

        rows = self.service._extract_structure_rows("controller/IssueController.java", source.splitlines())
        definitions = {(row[2], row[0]) for row in rows["definitions"]}
        references = {(row[2], row[0]) for row in rows["references"]}
        entity_edges = {(row[3], row[4]) for row in rows["entity_edges"]}

        self.assertIn(("class", "IssueController"), definitions)
        self.assertIn(("java_method", "IssueController.createIssue"), definitions)
        self.assertIn(("route", "/api/issues/create"), references)
        self.assertIn(("bean_condition", "issue.enabled=true"), references)
        self.assertIn(("bean_qualifier_target", "issueService=primaryIssueService"), references)
        self.assertIn(("operational_boundary", "Transactional:rollbackFor = Exception.class"), references)
        self.assertIn(("sql_table", "issue_table"), references)
        self.assertIn(("route", "/api/issues/create"), entity_edges)
        self.assertIn(("db_read", "issue_table"), entity_edges)

    def test_java_member_calls_resolve_to_qualified_repository_methods(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        service_file = repo_path / "service" / "IssueService.java"
        service_file.parent.mkdir(parents=True)
        service_file.write_text(
            "public class IssueService {\n"
            "    private IssueRepository issueRepository;\n"
            "    public Issue createIssue() {\n"
            "        return issueRepository.findIssue();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        repository_file = repo_path / "repository" / "IssueRepository.java"
        repository_file.parent.mkdir(parents=True)
        repository_file.write_text(
            "public class IssueRepository {\n"
            "    public Issue findIssue() {\n"
            "        return jdbcTemplate.queryForObject(\"select * from issue_table\", mapper);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            edge_rows = list(
                connection.execute(
                    """
                    select edge_kind, to_name, to_file
                    from entity_edges
                    where from_file = 'service/IssueService.java'
                    """
                )
            )
            flow_rows = list(
                connection.execute(
                    """
                    select edge_kind, to_name, to_file
                    from flow_edges
                    where from_file = 'service/IssueService.java'
                    """
                )
            )
        self.assertIn(("call", "IssueRepository.findIssue", "repository/IssueRepository.java"), edge_rows)
        self.assertTrue(
            any(kind == "repository" and target == "IssueRepository.findIssue" for kind, target, _to_file in flow_rows)
        )

    def test_java_interface_and_mybatis_namespace_connect_to_implementations(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text(
            "public interface IssueService {\n"
            "    Issue createIssue();\n"
            "}\n",
            encoding="utf-8",
        )
        impl_file = repo_path / "service" / "IssueServiceImpl.java"
        impl_file.write_text(
            "public class IssueServiceImpl implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        mapper_interface = repo_path / "mapper" / "IssueMapper.java"
        mapper_interface.parent.mkdir(parents=True)
        mapper_interface.write_text(
            "public interface IssueMapper {\n"
            "    Issue findIssue();\n"
            "}\n",
            encoding="utf-8",
        )
        mapper_xml = repo_path / "mapper" / "IssueMapper.xml"
        mapper_xml.write_text(
            "<mapper namespace=\"com.example.IssueMapper\">\n"
            "  <select id=\"findIssue\">select * from issue_table</select>\n"
            "</mapper>\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            edge_rows = list(connection.execute("select edge_kind, to_name, to_file from entity_edges"))
            flow_rows = list(connection.execute("select edge_kind, to_name, to_file from flow_edges"))
        self.assertIn(("implements", "IssueService", "service/IssueService.java"), edge_rows)
        self.assertIn(("mapper_interface", "IssueMapper", "mapper/IssueMapper.java"), edge_rows)
        self.assertTrue(any(kind == "mapper" and target == "IssueMapper" for kind, target, _to_file in flow_rows))

    def test_interface_method_calls_resolve_to_implementation_methods(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "public class IssueController {\n"
            "    private IssueService issueService;\n"
            "    public Issue createIssue() {\n"
            "        return issueService.createIssue();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text(
            "public interface IssueService {\n"
            "    Issue createIssue();\n"
            "}\n",
            encoding="utf-8",
        )
        impl_file = repo_path / "service" / "IssueServiceImpl.java"
        impl_file.write_text(
            "@Service(\"primaryIssueService\")\n"
            "public class IssueServiceImpl implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            entity_edges = list(connection.execute("select edge_kind, to_name, to_file from entity_edges"))
            flow_edges = list(connection.execute("select edge_kind, to_name, to_file from flow_edges"))

        self.assertIn(("bean_name", "primaryIssueService", ""), entity_edges)
        self.assertIn(("implementation_call", "IssueServiceImpl.createIssue", "service/IssueServiceImpl.java"), entity_edges)
        self.assertTrue(
            any(kind == "service" and target == "IssueServiceImpl.createIssue" for kind, target, _to_file in flow_edges)
        )

    def test_qualifier_filters_interface_implementation_resolution(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "public class IssueController {\n"
            "    @Qualifier(\"fastIssueService\")\n"
            "    private IssueService issueService;\n"
            "    public Issue createIssue() {\n"
            "        return issueService.createIssue();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text(
            "public interface IssueService {\n"
            "    Issue createIssue();\n"
            "}\n",
            encoding="utf-8",
        )
        fast_impl = repo_path / "service" / "FastIssueService.java"
        fast_impl.write_text(
            "@Service(\"fastIssueService\")\n"
            "public class FastIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        slow_impl = repo_path / "service" / "SlowIssueService.java"
        slow_impl.write_text(
            "@Service(\"slowIssueService\")\n"
            "public class SlowIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select edge_kind, to_name, to_file, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    """
                )
            )

        self.assertIn(
            ("implementation_call", "FastIssueService.createIssue", "service/FastIssueService.java"),
            [(kind, target, to_file) for kind, target, to_file, _evidence in implementation_edges],
        )
        self.assertFalse(any(target == "SlowIssueService.createIssue" for _kind, target, _to_file, _evidence in implementation_edges))
        self.assertTrue(any("qualifier=fastIssueService" in evidence for _kind, _target, _to_file, evidence in implementation_edges))

    def test_variable_qualifier_filters_each_interface_call_independently(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "public class IssueController {\n"
            "    @Qualifier(\"fastIssueService\")\n"
            "    private IssueService fastIssueService;\n"
            "    @Qualifier(\"slowIssueService\")\n"
            "    private IssueService slowIssueService;\n"
            "    public Issue createFastIssue() { return fastIssueService.createIssue(); }\n"
            "    public Issue createSlowIssue() { return slowIssueService.createIssue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        fast_impl = repo_path / "service" / "FastIssueService.java"
        fast_impl.write_text(
            "@Service(\"fastIssueService\")\n"
            "public class FastIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        slow_impl = repo_path / "service" / "SlowIssueService.java"
        slow_impl.write_text(
            "@Service(\"slowIssueService\")\n"
            "public class SlowIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select from_line, to_name, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    order by from_line, to_name
                    """
                )
            )

        fast_edges = [(target, evidence) for line_no, target, evidence in implementation_edges if line_no == 6]
        slow_edges = [(target, evidence) for line_no, target, evidence in implementation_edges if line_no == 7]
        self.assertEqual(["FastIssueService.createIssue"], [target for target, _evidence in fast_edges])
        self.assertEqual(["SlowIssueService.createIssue"], [target for target, _evidence in slow_edges])
        self.assertTrue(any("qualifier=fastIssueService" in evidence for _target, evidence in fast_edges))
        self.assertTrue(any("qualifier=slowIssueService" in evidence for _target, evidence in slow_edges))

    def test_constructor_parameter_qualifiers_resolve_assigned_fields_independently(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "public class IssueController {\n"
            "    private IssueService fastIssueService;\n"
            "    private IssueService slowIssueService;\n"
            "    public IssueController(@Qualifier(\"fastIssueService\") IssueService fastDelegate, @Qualifier(\"slowIssueService\") IssueService slowDelegate) {\n"
            "        this.fastIssueService = fastDelegate;\n"
            "        this.slowIssueService = slowDelegate;\n"
            "    }\n"
            "    public Issue createFastIssue() { return fastIssueService.createIssue(); }\n"
            "    public Issue createSlowIssue() { return slowIssueService.createIssue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        fast_impl = repo_path / "service" / "FastIssueService.java"
        fast_impl.write_text(
            "@Service(\"fastIssueService\")\n"
            "public class FastIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        slow_impl = repo_path / "service" / "SlowIssueService.java"
        slow_impl.write_text(
            "@Service(\"slowIssueService\")\n"
            "public class SlowIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select to_name, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    order by from_line, to_name
                    """
                )
            )

        fast_edges = [(target, evidence) for target, evidence in implementation_edges if "fastIssueService.createIssue" in evidence]
        slow_edges = [(target, evidence) for target, evidence in implementation_edges if "slowIssueService.createIssue" in evidence]
        self.assertEqual(["FastIssueService.createIssue"], [target for target, _evidence in fast_edges])
        self.assertEqual(["SlowIssueService.createIssue"], [target for target, _evidence in slow_edges])
        self.assertTrue(any("qualifier=fastIssueService" in evidence for _target, evidence in fast_edges))
        self.assertTrue(any("qualifier=slowIssueService" in evidence for _target, evidence in slow_edges))

    def test_generic_collection_injection_resolves_lambda_interface_calls(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "import java.util.List;\n"
            "public class IssueController {\n"
            "    private List<IssueService> issueServices;\n"
            "    public void createAllIssues() {\n"
            "        issueServices.forEach(issueService -> issueService.createIssue());\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        fast_impl = repo_path / "service" / "FastIssueService.java"
        fast_impl.write_text(
            "@Service(\"fastIssueService\")\n"
            "public class FastIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        slow_impl = repo_path / "service" / "SlowIssueService.java"
        slow_impl.write_text(
            "@Service(\"slowIssueService\")\n"
            "public class SlowIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select to_name, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    """
                )
            )

        targets = {target for target, _evidence in implementation_edges}
        self.assertIn("FastIssueService.createIssue", targets)
        self.assertIn("SlowIssueService.createIssue", targets)
        self.assertTrue(any("issueService.createIssue" in evidence for _target, evidence in implementation_edges))

    def test_qualified_generic_collection_injection_filters_lambda_interface_calls(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "import java.util.List;\n"
            "public class IssueController {\n"
            "    @Qualifier(\"fastIssueService\")\n"
            "    private List<IssueService> issueServices;\n"
            "    public void createFastIssues() {\n"
            "        issueServices.forEach(issueService -> issueService.createIssue());\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        fast_impl = repo_path / "service" / "FastIssueService.java"
        fast_impl.write_text(
            "@Service(\"fastIssueService\")\n"
            "public class FastIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        slow_impl = repo_path / "service" / "SlowIssueService.java"
        slow_impl.write_text(
            "@Service(\"slowIssueService\")\n"
            "public class SlowIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select to_name, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    """
                )
            )

        self.assertIn("FastIssueService.createIssue", [target for target, _evidence in implementation_edges])
        self.assertFalse(any(target == "SlowIssueService.createIssue" for target, _evidence in implementation_edges))
        self.assertTrue(any("qualifier=fastIssueService" in evidence for _target, evidence in implementation_edges))

    def test_generic_map_values_lambda_resolves_interface_calls(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "import java.util.Map;\n"
            "public class IssueController {\n"
            "    private Map<String, IssueService> issueServiceMap;\n"
            "    public void createAllIssues() {\n"
            "        issueServiceMap.values().forEach(issueService -> issueService.createIssue());\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        fast_impl = repo_path / "service" / "FastIssueService.java"
        fast_impl.write_text(
            "@Service(\"fastIssueService\")\n"
            "public class FastIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        slow_impl = repo_path / "service" / "SlowIssueService.java"
        slow_impl.write_text(
            "@Service(\"slowIssueService\")\n"
            "public class SlowIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select to_name, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    """
                )
            )

        targets = {target for target, _evidence in implementation_edges}
        self.assertIn("FastIssueService.createIssue", targets)
        self.assertIn("SlowIssueService.createIssue", targets)
        self.assertTrue(any("issueService.createIssue" in evidence for _target, evidence in implementation_edges))

    def test_generic_stream_map_lambda_resolves_interface_calls(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "import java.util.List;\n"
            "public class IssueController {\n"
            "    private List<IssueService> issueServices;\n"
            "    public void createAllIssues() {\n"
            "        issueServices.stream().map(issueService -> issueService.createIssue()).toList();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        impl_file = repo_path / "service" / "IssueServiceImpl.java"
        impl_file.write_text(
            "@Service(\"issueServiceImpl\")\n"
            "public class IssueServiceImpl implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select to_name, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    """
                )
            )

        self.assertIn("IssueServiceImpl.createIssue", [target for target, _evidence in implementation_edges])
        self.assertTrue(any("issueService.createIssue" in evidence for _target, evidence in implementation_edges))

    def test_object_provider_chain_resolves_interface_call(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "import org.springframework.beans.factory.ObjectProvider;\n"
            "public class IssueController {\n"
            "    private ObjectProvider<IssueService> issueServiceProvider;\n"
            "    public Issue createIssue() {\n"
            "        return issueServiceProvider.getIfAvailable().createIssue();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        impl_file = repo_path / "service" / "IssueServiceImpl.java"
        impl_file.write_text(
            "@Service(\"issueServiceImpl\")\n"
            "public class IssueServiceImpl implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select to_name, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    """
                )
            )

        self.assertIn("IssueServiceImpl.createIssue", [target for target, _evidence in implementation_edges])
        self.assertTrue(any("issueServiceProvider.getIfAvailable().createIssue" in evidence for _target, evidence in implementation_edges))

    def test_spring_aop_edges_are_indexed_as_framework_flow(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        aspect_file = repo_path / "aspect" / "IssueAuditAspect.java"
        aspect_file.parent.mkdir(parents=True)
        aspect_file.write_text(
            "@Aspect\n"
            "@Component\n"
            "public class IssueAuditAspect {\n"
            "    @Pointcut(\"execution(* *..IssueService.createIssue(..))\")\n"
            "    public void issueCreatePointcut() {}\n"
            "    @Around(\"issueCreatePointcut()\")\n"
            "    public Object auditIssueCreate(ProceedingJoinPoint joinPoint) { return joinPoint.proceed(); }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        impl_file = repo_path / "service" / "IssueServiceImpl.java"
        impl_file.write_text(
            "@Service(\"issueServiceImpl\")\n"
            "public class IssueServiceImpl implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            entity_edges = list(
                connection.execute(
                    """
                    select edge_kind, to_name
                    from entity_edges
                    where from_file = 'aspect/IssueAuditAspect.java'
                    """
                )
            )
            flow_edges = list(
                connection.execute(
                    """
                    select edge_kind, to_name
                    from flow_edges
                    where from_file = 'aspect/IssueAuditAspect.java'
                    """
                )
            )

        self.assertIn(("framework_binding", "Aspect"), entity_edges)
        self.assertTrue(any(kind == "aop_pointcut" and "IssueService.createIssue" in target for kind, target in entity_edges))
        self.assertTrue(any(kind == "aop_advice" and "issueCreatePointcut" in target for kind, target in entity_edges))
        self.assertTrue(any(kind == "aop_applies_to" and target == "IssueService.createIssue" for kind, target in entity_edges))
        self.assertTrue(any(kind == "aop_applies_to" and target == "IssueServiceImpl.createIssue" for kind, target in entity_edges))
        self.assertTrue(any(kind == "framework" and "issueCreatePointcut" in target for kind, target in flow_edges))
        self.assertTrue(any(kind == "service" and target == "IssueServiceImpl.createIssue" for kind, target in flow_edges))

    def test_scheduled_job_edges_are_indexed_as_framework_flow(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        job_file = repo_path / "job" / "IssueRetryJob.java"
        job_file.parent.mkdir(parents=True)
        job_file.write_text(
            "@Component\n"
            "public class IssueRetryJob {\n"
            "    private IssueService issueService;\n"
            "    @Scheduled(cron = \"0 0/5 * * * ?\")\n"
            "    public void retryIssues() { issueService.createIssue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        impl_file = repo_path / "service" / "IssueServiceImpl.java"
        impl_file.write_text(
            "@Service(\"issueServiceImpl\")\n"
            "public class IssueServiceImpl implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            entity_edges = list(
                connection.execute(
                    """
                    select edge_kind, to_name
                    from entity_edges
                    where from_file = 'job/IssueRetryJob.java'
                    """
                )
            )
            flow_edges = list(
                connection.execute(
                    """
                    select edge_kind, to_name
                    from flow_edges
                    where from_file = 'job/IssueRetryJob.java'
                    """
                )
            )

        self.assertTrue(any(kind == "scheduled_job" and "cron=0 0/5 * * * ?" in target for kind, target in entity_edges))
        self.assertTrue(any(kind == "implementation_call" and target == "IssueServiceImpl.createIssue" for kind, target in entity_edges))
        self.assertTrue(any(kind == "framework" and "cron=0 0/5 * * * ?" in target for kind, target in flow_edges))

    def test_handler_interceptor_edges_are_indexed_as_framework_flow(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        interceptor_file = repo_path / "web" / "IssueAuthInterceptor.java"
        interceptor_file.parent.mkdir(parents=True)
        interceptor_file.write_text(
            "public class IssueAuthInterceptor implements HandlerInterceptor {\n"
            "    public boolean preHandle(HttpServletRequest request, HttpServletResponse response, Object handler) {\n"
            "        return true;\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            entity_edges = list(
                connection.execute(
                    """
                    select edge_kind, to_name
                    from entity_edges
                    where from_file = 'web/IssueAuthInterceptor.java'
                    """
                )
            )
            flow_edges = list(
                connection.execute(
                    """
                    select edge_kind, to_name
                    from flow_edges
                    where from_file = 'web/IssueAuthInterceptor.java'
                    """
                )
            )

        self.assertTrue(any(kind == "web_interceptor" and target in {"IssueAuthInterceptor", "HandlerInterceptor"} for kind, target in entity_edges))
        self.assertTrue(any(kind == "framework" and target in {"IssueAuthInterceptor", "HandlerInterceptor"} for kind, target in flow_edges))

    def test_primary_filters_interface_implementation_resolution(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "public class IssueController {\n"
            "    private IssueService issueService;\n"
            "    public Issue createIssue() { return issueService.createIssue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        primary_impl = repo_path / "service" / "PrimaryIssueService.java"
        primary_impl.write_text(
            "@Primary\n"
            "@Service(\"primaryIssueService\")\n"
            "public class PrimaryIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        fallback_impl = repo_path / "service" / "FallbackIssueService.java"
        fallback_impl.write_text(
            "@Service(\"fallbackIssueService\")\n"
            "public class FallbackIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select to_name, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    """
                )
            )

        self.assertIn("PrimaryIssueService.createIssue", [target for target, _evidence in implementation_edges])
        self.assertFalse(any(target == "FallbackIssueService.createIssue" for target, _evidence in implementation_edges))
        self.assertTrue(any("primary=true" in evidence for _target, evidence in implementation_edges))

    def test_active_profile_filters_interface_implementation_resolution(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        (repo_path / "application.properties").write_text("spring.profiles.active=prod\n", encoding="utf-8")
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "public class IssueController {\n"
            "    private IssueService issueService;\n"
            "    public Issue createIssue() { return issueService.createIssue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        prod_impl = repo_path / "service" / "ProdIssueService.java"
        prod_impl.write_text(
            "@Profile(\"prod\")\n"
            "@Service(\"prodIssueService\")\n"
            "public class ProdIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        dev_impl = repo_path / "service" / "DevIssueService.java"
        dev_impl.write_text(
            "@Profile(\"dev\")\n"
            "@Service(\"devIssueService\")\n"
            "public class DevIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select to_name, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    """
                )
            )

        self.assertIn("ProdIssueService.createIssue", [target for target, _evidence in implementation_edges])
        self.assertFalse(any(target == "DevIssueService.createIssue" for target, _evidence in implementation_edges))
        self.assertTrue(any("profile=prod" in evidence for _target, evidence in implementation_edges))

    def test_conditional_on_property_filters_interface_implementation_resolution(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        (repo_path / "application.properties").write_text("issue.fast.enabled=true\n", encoding="utf-8")
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "public class IssueController {\n"
            "    private IssueService issueService;\n"
            "    public Issue createIssue() { return issueService.createIssue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        fast_impl = repo_path / "service" / "FastIssueService.java"
        fast_impl.write_text(
            "@ConditionalOnProperty(name = \"issue.fast.enabled\", havingValue = \"true\")\n"
            "@Service(\"fastIssueService\")\n"
            "public class FastIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        slow_impl = repo_path / "service" / "SlowIssueService.java"
        slow_impl.write_text(
            "@ConditionalOnProperty(name = \"issue.fast.enabled\", havingValue = \"false\")\n"
            "@Service(\"slowIssueService\")\n"
            "public class SlowIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select to_name, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    """
                )
            )

        self.assertIn("FastIssueService.createIssue", [target for target, _evidence in implementation_edges])
        self.assertFalse(any(target == "SlowIssueService.createIssue" for target, _evidence in implementation_edges))
        self.assertTrue(any("condition=issue.fast.enabled=true" in evidence for _target, evidence in implementation_edges))

    def test_profile_specific_config_overrides_conditional_on_property_resolution(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        (repo_path / "application.properties").write_text(
            "spring.profiles.active=prod\n"
            "issue.fast.enabled=false\n",
            encoding="utf-8",
        )
        (repo_path / "application-prod.properties").write_text("issue.fast.enabled=true\n", encoding="utf-8")
        (repo_path / "application-dev.properties").write_text("issue.fast.enabled=false\n", encoding="utf-8")
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "public class IssueController {\n"
            "    private IssueService issueService;\n"
            "    public Issue createIssue() { return issueService.createIssue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        fast_impl = repo_path / "service" / "FastIssueService.java"
        fast_impl.write_text(
            "@ConditionalOnProperty(name = \"issue.fast.enabled\", havingValue = \"true\")\n"
            "@Service(\"fastIssueService\")\n"
            "public class FastIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        slow_impl = repo_path / "service" / "SlowIssueService.java"
        slow_impl.write_text(
            "@ConditionalOnProperty(name = \"issue.fast.enabled\", havingValue = \"false\")\n"
            "@Service(\"slowIssueService\")\n"
            "public class SlowIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select to_name, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    """
                )
            )

        self.assertIn("FastIssueService.createIssue", [target for target, _evidence in implementation_edges])
        self.assertFalse(any(target == "SlowIssueService.createIssue" for target, _evidence in implementation_edges))
        self.assertTrue(any("condition=issue.fast.enabled=true" in evidence for _target, evidence in implementation_edges))

    def test_yaml_on_profile_document_overrides_conditional_on_property_resolution(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        (repo_path / "application.yml").write_text(
            "spring:\n"
            "  profiles:\n"
            "    active: prod\n"
            "issue:\n"
            "  fast:\n"
            "    enabled: false\n"
            "---\n"
            "spring:\n"
            "  config:\n"
            "    activate:\n"
            "      on-profile: prod\n"
            "issue:\n"
            "  fast:\n"
            "    enabled: true\n"
            "---\n"
            "spring:\n"
            "  config:\n"
            "    activate:\n"
            "      on-profile: dev\n"
            "issue:\n"
            "  fast:\n"
            "    enabled: false\n",
            encoding="utf-8",
        )
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "public class IssueController {\n"
            "    private IssueService issueService;\n"
            "    public Issue createIssue() { return issueService.createIssue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        fast_impl = repo_path / "service" / "FastIssueService.java"
        fast_impl.write_text(
            "@ConditionalOnProperty(name = \"issue.fast.enabled\", havingValue = \"true\")\n"
            "@Service(\"fastIssueService\")\n"
            "public class FastIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        slow_impl = repo_path / "service" / "SlowIssueService.java"
        slow_impl.write_text(
            "@ConditionalOnProperty(name = \"issue.fast.enabled\", havingValue = \"false\")\n"
            "@Service(\"slowIssueService\")\n"
            "public class SlowIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select to_name, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    """
                )
            )

        self.assertIn("FastIssueService.createIssue", [target for target, _evidence in implementation_edges])
        self.assertFalse(any(target == "SlowIssueService.createIssue" for target, _evidence in implementation_edges))
        self.assertTrue(any("condition=issue.fast.enabled=true" in evidence for _target, evidence in implementation_edges))

    def test_conditional_on_property_match_if_missing_selects_default_implementation(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        controller_file = repo_path / "controller" / "IssueController.java"
        controller_file.parent.mkdir(parents=True)
        controller_file.write_text(
            "public class IssueController {\n"
            "    private IssueService issueService;\n"
            "    public Issue createIssue() { return issueService.createIssue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        interface_file = repo_path / "service" / "IssueService.java"
        interface_file.parent.mkdir(parents=True)
        interface_file.write_text("public interface IssueService { Issue createIssue(); }\n", encoding="utf-8")
        default_impl = repo_path / "service" / "DefaultIssueService.java"
        default_impl.write_text(
            "@ConditionalOnProperty(name = \"issue.optional.enabled\", matchIfMissing = true)\n"
            "@Service(\"defaultIssueService\")\n"
            "public class DefaultIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        disabled_impl = repo_path / "service" / "DisabledIssueService.java"
        disabled_impl.write_text(
            "@ConditionalOnProperty(name = \"issue.optional.enabled\", havingValue = \"false\")\n"
            "@Service(\"disabledIssueService\")\n"
            "public class DisabledIssueService implements IssueService {\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            implementation_edges = list(
                connection.execute(
                    """
                    select to_name, evidence
                    from entity_edges
                    where edge_kind = 'implementation_call'
                      and from_file = 'controller/IssueController.java'
                    """
                )
            )

        self.assertIn("DefaultIssueService.createIssue", [target for target, _evidence in implementation_edges])
        self.assertFalse(any(target == "DisabledIssueService.createIssue" for target, _evidence in implementation_edges))
        self.assertTrue(any("condition=issue.optional.enabled=<missing:true>" in evidence for _target, evidence in implementation_edges))

    def test_java_package_imports_create_fully_qualified_symbol_edges(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        service_file = repo_path / "src" / "main" / "java" / "com" / "example" / "service" / "IssueService.java"
        service_file.parent.mkdir(parents=True)
        service_file.write_text(
            "package com.example.service;\n"
            "import com.example.repository.IssueRepository;\n"
            "public class IssueService {\n"
            "    private IssueRepository issueRepository;\n"
            "    public Issue createIssue() {\n"
            "        return issueRepository.findIssue();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        repository_file = repo_path / "src" / "main" / "java" / "com" / "example" / "repository" / "IssueRepository.java"
        repository_file.parent.mkdir(parents=True)
        repository_file.write_text(
            "package com.example.repository;\n"
            "public class IssueRepository {\n"
            "    public Issue findIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            definitions = list(connection.execute("select name, kind from definitions"))
            edges = list(connection.execute("select edge_kind, to_name, to_file from entity_edges"))
        self.assertIn(("com.example.service.IssueService", "class"), definitions)
        self.assertIn(("com.example.repository.IssueRepository.findIssue", "call", str(repository_file.relative_to(repo_path))), [(to_name, kind, to_file) for kind, to_name, to_file in edges])

    def test_tree_sitter_python_extracts_import_route_and_call(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "api" / "routes.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "from flask import Blueprint\n"
            "from service.issue_service import create_issue\n\n"
            "bp = Blueprint('issue', __name__)\n\n"
            "@bp.route('/api/issues/create', methods=['POST'])\n"
            "def create_issue_route():\n"
            "    return create_issue()\n",
            encoding="utf-8",
        )

        index_info = self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        self.assertEqual(index_info["parser_backend"], "tree_sitter+regex")
        self.assertIn("python", index_info["parser_languages"])
        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            rows = list(connection.execute("select edge_kind, to_name from entity_edges"))
        self.assertTrue(any(kind == "import" and "service.issue_service" in target for kind, target in rows))
        self.assertIn(("route", "/api/issues/create"), rows)
        self.assertTrue(any(kind == "call" and target == "create_issue" for kind, target in rows))

    def test_tree_sitter_typescript_extracts_import_export_and_api_call(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "web" / "issue_client.ts"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "import axios from 'axios';\n"
            "export class IssueClient {\n"
            "  createIssue(payload: unknown) {\n"
            "    return axios.post('/api/issues/create', payload);\n"
            "  }\n"
            "}\n",
            encoding="utf-8",
        )

        index_info = self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        self.assertEqual(index_info["parser_backend"], "tree_sitter+regex")
        self.assertIn("typescript", index_info["parser_languages"])
        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            definitions = list(connection.execute("select name, kind from definitions"))
            rows = list(connection.execute("select edge_kind, to_name from entity_edges"))
        self.assertTrue(any(name == "IssueClient" and "class" in kind for name, kind in definitions))
        self.assertTrue(any(kind == "import" and "axios" in target for kind, target in rows))
        self.assertIn(("http_endpoint", "/api/issues/create"), rows)

    def test_tree_sitter_unavailable_falls_back_to_regex_index(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "controller" / "IssueController.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "public class IssueController { public Issue createIssue() { return new Issue(); } }\n",
            encoding="utf-8",
        )

        with patch.object(self.service, "_tree_sitter_parser_for_language", return_value=None):
            index_info = self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        self.assertEqual(index_info["parser_backend"], "regex")
        self.assertEqual(index_info["tree_sitter_files"], 0)
        self.assertGreaterEqual(index_info["definitions"], 1)

    def test_cross_repo_graph_matches_http_path_to_controller_route(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[
                {"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"},
                {"display_name": "Issue Service", "url": "https://git.example.com/team/issue-service.git"},
            ],
        )
        entries = self.service.load_config()["mappings"]["AF:All"]
        portal = type("Entry", (), entries[0])()
        service_entry = type("Entry", (), entries[1])()
        portal_path = self.service._repo_path("AF:All", portal)
        service_path = self.service._repo_path("AF:All", service_entry)
        (portal_path / ".git").mkdir(parents=True)
        (service_path / ".git").mkdir(parents=True)
        client_file = portal_path / "web" / "issue_client.ts"
        client_file.parent.mkdir(parents=True)
        client_file.write_text(
            "export function createIssue(payload: unknown) {\n"
            "  return fetch('/issue/create', { method: 'POST', body: JSON.stringify(payload) });\n"
            "}\n",
            encoding="utf-8",
        )
        service_file = service_path / "controller" / "IssueController.java"
        service_file.parent.mkdir(parents=True)
        service_file.write_text(
            "@RestController\n"
            "@RequestMapping(\"/issue\")\n"
            "public class IssueController {\n"
            "    @PostMapping(\"/create\")\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="which service handles issue create API")
        edge = next(
            edge
            for edge in payload["repo_graph"]["edges"]
            if edge["edge_kind"] == "http_path" and "/issue/create" in edge["match_reason"]
        )

        self.assertEqual(edge["from_repo"], "Portal Repo")
        self.assertEqual(edge["to_repo"], "Issue Service")
        self.assertGreaterEqual(edge["confidence"], 0.95)
        self.assertEqual(edge["to_file"], "controller/IssueController.java")
        self.assertGreater(edge["to_line"], 0)
        self.assertFalse(
            any(
                graph_edge["from_repo"] == "Issue Service"
                and graph_edge["to_repo"] == "Portal Repo"
                and graph_edge["edge_kind"] == "http_path"
                for graph_edge in payload["repo_graph"]["edges"]
            )
        )

    def test_cross_repo_graph_resolves_feign_config_placeholders(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[
                {"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"},
                {"display_name": "Issue Service", "url": "https://git.example.com/team/issue-service.git"},
            ],
        )
        entries = self.service.load_config()["mappings"]["AF:All"]
        portal = type("Entry", (), entries[0])()
        service_entry = type("Entry", (), entries[1])()
        portal_path = self.service._repo_path("AF:All", portal)
        service_path = self.service._repo_path("AF:All", service_entry)
        (portal_path / ".git").mkdir(parents=True)
        (service_path / ".git").mkdir(parents=True)
        client_file = portal_path / "client" / "IssueClient.java"
        client_file.parent.mkdir(parents=True)
        client_file.write_text(
            "@FeignClient(name = \"${issue.service.name}\", url = \"${issue.service.url}\")\n"
            "public interface IssueClient {\n"
            "    @PostMapping(\"/create\")\n"
            "    Issue createIssue();\n"
            "}\n",
            encoding="utf-8",
        )
        config_file = portal_path / "application.properties"
        config_file.write_text(
            "issue.service.name=issue-service\n"
            "issue.service.url=http://issue-service/issue\n",
            encoding="utf-8",
        )
        service_file = service_path / "controller" / "IssueController.java"
        service_file.parent.mkdir(parents=True)
        service_file.write_text(
            "@RestController\n"
            "@RequestMapping(\"/issue\")\n"
            "public class IssueController {\n"
            "    @PostMapping(\"/create\")\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="which service does feign issue client call")
        edge = next(
            edge
            for edge in payload["repo_graph"]["edges"]
            if edge["edge_kind"] == "http_path" and "/issue/create" in edge["match_reason"]
        )

        self.assertEqual(edge["from_repo"], "Portal Repo")
        self.assertEqual(edge["to_repo"], "Issue Service")
        self.assertGreaterEqual(edge["confidence"], 0.95)
        self.assertIn("/issue/create", edge["match_reason"])

    def test_cross_repo_graph_resolves_feign_yaml_config_placeholders(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[
                {"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"},
                {"display_name": "Issue Service", "url": "https://git.example.com/team/issue-service.git"},
            ],
        )
        entries = self.service.load_config()["mappings"]["AF:All"]
        portal = type("Entry", (), entries[0])()
        service_entry = type("Entry", (), entries[1])()
        portal_path = self.service._repo_path("AF:All", portal)
        service_path = self.service._repo_path("AF:All", service_entry)
        (portal_path / ".git").mkdir(parents=True)
        (service_path / ".git").mkdir(parents=True)
        client_file = portal_path / "client" / "IssueClient.java"
        client_file.parent.mkdir(parents=True)
        client_file.write_text(
            "@FeignClient(name = \"${issue.service.name}\", url = \"${issue.service.url}\")\n"
            "public interface IssueClient {\n"
            "    @PostMapping(\"/create\")\n"
            "    Issue createIssue();\n"
            "}\n",
            encoding="utf-8",
        )
        config_file = portal_path / "application.yml"
        config_file.write_text(
            "issue:\n"
            "  service:\n"
            "    name: issue-service\n"
            "    url: http://issue-service/issue\n",
            encoding="utf-8",
        )
        service_file = service_path / "controller" / "IssueController.java"
        service_file.parent.mkdir(parents=True)
        service_file.write_text(
            "@RestController\n"
            "@RequestMapping(\"/issue\")\n"
            "public class IssueController {\n"
            "    @PostMapping(\"/create\")\n"
            "    public Issue createIssue() { return new Issue(); }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="which service does feign issue client call")
        edge = next(
            edge
            for edge in payload["repo_graph"]["edges"]
            if edge["edge_kind"] == "http_path" and "/issue/create" in edge["match_reason"]
        )
        with sqlite3.connect(self.service._index_path(portal_path)) as connection:
            config_edges = list(connection.execute("select edge_kind, to_name from entity_edges"))

        self.assertEqual(edge["from_repo"], "Portal Repo")
        self.assertEqual(edge["to_repo"], "Issue Service")
        self.assertGreaterEqual(edge["confidence"], 0.95)
        self.assertIn(("config", "issue.service.url"), config_edges)
        self.assertIn(("config_value", "http://issue-service/issue"), config_edges)

    def test_build_files_create_module_dependency_edges_and_repo_graph(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[
                {"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"},
                {"display_name": "Issue Service", "url": "https://git.example.com/team/issue-service.git"},
            ],
        )
        entries = self.service.load_config()["mappings"]["AF:All"]
        portal = type("Entry", (), entries[0])()
        service_entry = type("Entry", (), entries[1])()
        portal_path = self.service._repo_path("AF:All", portal)
        service_path = self.service._repo_path("AF:All", service_entry)
        (portal_path / ".git").mkdir(parents=True)
        (service_path / ".git").mkdir(parents=True)
        (portal_path / "pom.xml").write_text(
            "<project>\n"
            "  <groupId>com.example</groupId>\n"
            "  <artifactId>portal-web</artifactId>\n"
            "  <dependencies>\n"
            "    <dependency>\n"
            "      <groupId>com.example</groupId>\n"
            "      <artifactId>issue-service-api</artifactId>\n"
            "      <version>1.0.0</version>\n"
            "    </dependency>\n"
            "  </dependencies>\n"
            "</project>\n",
            encoding="utf-8",
        )
        (service_path / "pom.xml").write_text(
            "<project>\n"
            "  <groupId>com.example</groupId>\n"
            "  <artifactId>issue-service-api</artifactId>\n"
            "</project>\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="which repo does portal-web depend on")
        policy_statuses = {policy["name"]: policy["status"] for policy in payload["answer_quality"]["policies"]}
        self.assertEqual(policy_statuses["module_dependency"], "satisfied")
        self.assertIn(payload["matches"][0]["path"], {"pom.xml", "package.json", "settings.gradle", "build.gradle"})
        edge = next(
            edge
            for edge in payload["repo_graph"]["edges"]
            if edge["from_repo"] == "Portal Repo"
            and edge["to_repo"] == "Issue Service"
            and edge["edge_kind"] == "module_dependency"
        )
        with sqlite3.connect(self.service._index_path(portal_path)) as connection:
            flow_edges = list(connection.execute("select edge_kind, to_name, to_file from flow_edges"))

        self.assertGreaterEqual(edge["confidence"], 0.84)
        self.assertIn("exact build artifact match", edge["match_reason"])
        self.assertTrue(
            any(kind == "module_dependency" and target == "com.example:issue-service-api" for kind, target, _to_file in flow_edges)
        )

    def test_gradle_multi_module_edges_are_indexed(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        (repo_path / "settings.gradle").write_text(
            "rootProject.name = 'portal'\n"
            "include ':issue-api', ':issue-service'\n",
            encoding="utf-8",
        )
        (repo_path / "issue-service" / "build.gradle").parent.mkdir(parents=True)
        (repo_path / "issue-service" / "build.gradle").write_text(
            "dependencies {\n"
            "    implementation project(':issue-api')\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            entity_edges = list(
                connection.execute(
                    """
                    select edge_kind, to_name, from_file
                    from entity_edges
                    where edge_kind in ('gradle_module', 'gradle_project_dependency', 'module_dependency', 'module_artifact')
                    """
                )
            )
            flow_edges = list(
                connection.execute(
                    """
                    select edge_kind, to_name, from_file
                    from flow_edges
                    where edge_kind = 'module_dependency'
                    """
                )
            )

        self.assertIn(("gradle_module", "issue-api", "settings.gradle"), entity_edges)
        self.assertIn(("gradle_module", ":issue-api", "settings.gradle"), entity_edges)
        self.assertIn(("gradle_project_dependency", "issue-api", "issue-service/build.gradle"), entity_edges)
        self.assertIn(("gradle_project_dependency", ":issue-api", "issue-service/build.gradle"), entity_edges)
        self.assertIn(("module_dependency", "issue-api", "issue-service/build.gradle"), entity_edges)
        self.assertTrue(any(kind == "module_dependency" and target == "issue-api" for kind, target, _from_file in flow_edges))

    def test_runtime_trace_jsonl_creates_dynamic_flow_edges(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        trace_path = repo_path / "runtime-traces" / "source-code-qa.jsonl"
        trace_path.parent.mkdir(parents=True)
        trace_path.write_text(
            "\n".join(
                json.dumps(row)
                for row in [
                    {
                        "kind": "call",
                        "from": "IssueController.createIssue",
                        "to": "IssueServiceImpl.createIssue",
                        "evidence": "observed trace id t1",
                    },
                    {"kind": "sql", "from": "IssueServiceImpl.createIssue", "table": "issue_table"},
                    {"kind": "message", "from": "IssueEventPublisher.publish", "topic": "issue.created"},
                    {"kind": "config", "key": "feature.issue.fast.enabled", "value": True},
                    {"kind": "route", "path": "/api/issues/create", "handler": "IssueController.createIssue"},
                ]
            )
            + "\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            entity_edges = list(
                connection.execute(
                    """
                    select edge_kind, to_name, from_file
                    from entity_edges
                    where edge_kind like 'runtime_%'
                    """
                )
            )
            flow_edges = list(
                connection.execute(
                    """
                    select edge_kind, to_name, evidence
                    from flow_edges
                    where evidence like '%observed trace id t1%'
                       or to_name in ('issue_table', 'issue.created', 'feature.issue.fast.enabled', '/api/issues/create')
                    """
                )
            )

        self.assertIn(("runtime_call", "IssueServiceImpl.createIssue", "runtime-traces/source-code-qa.jsonl"), entity_edges)
        self.assertIn(("runtime_sql", "issue_table", "runtime-traces/source-code-qa.jsonl"), entity_edges)
        self.assertIn(("runtime_message", "issue.created", "runtime-traces/source-code-qa.jsonl"), entity_edges)
        self.assertIn(("runtime_config", "feature.issue.fast.enabled", "runtime-traces/source-code-qa.jsonl"), entity_edges)
        self.assertIn(("runtime_route", "/api/issues/create", "runtime-traces/source-code-qa.jsonl"), entity_edges)
        self.assertTrue(any(kind == "runtime" and target == "IssueServiceImpl.createIssue" for kind, target, _ in flow_edges))
        self.assertTrue(any(kind == "db_runtime" and target == "issue_table" for kind, target, _ in flow_edges))
        self.assertTrue(any(kind == "message_runtime" and target == "issue.created" for kind, target, _ in flow_edges))
        self.assertTrue(any(kind == "config" and target == "feature.issue.fast.enabled" for kind, target, _ in flow_edges))
        self.assertTrue(any(kind == "route" and target == "/api/issues/create" for kind, target, _ in flow_edges))

    def test_package_json_dependencies_create_module_dependency_repo_graph_edges(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[
                {"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"},
                {"display_name": "Issue SDK", "url": "https://git.example.com/team/issue-sdk.git"},
            ],
        )
        entries = self.service.load_config()["mappings"]["AF:All"]
        portal = type("Entry", (), entries[0])()
        sdk_entry = type("Entry", (), entries[1])()
        portal_path = self.service._repo_path("AF:All", portal)
        sdk_path = self.service._repo_path("AF:All", sdk_entry)
        (portal_path / ".git").mkdir(parents=True)
        (sdk_path / ".git").mkdir(parents=True)
        (portal_path / "package.json").write_text(
            json.dumps(
                {
                    "name": "@example/portal-web",
                    "dependencies": {"@example/issue-sdk": "^1.2.3"},
                },
                indent=2,
            ),
            encoding="utf-8",
        )
        (portal_path / "pom.xml").write_text(
            "<project><artifactId>portal-web</artifactId><dependencies>"
            "<dependency><groupId>com.example</groupId><artifactId>unrelated-service-api</artifactId></dependency>"
            "</dependencies></project>\n",
            encoding="utf-8",
        )
        (sdk_path / "package.json").write_text(
            json.dumps({"name": "@example/issue-sdk"}, indent=2),
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="which npm package does portal-web depend on")
        policy_statuses = {policy["name"]: policy["status"] for policy in payload["answer_quality"]["policies"]}
        self.assertEqual(policy_statuses["module_dependency"], "satisfied")
        self.assertEqual(payload["matches"][0]["path"], "package.json")
        edge = next(
            edge
            for edge in payload["repo_graph"]["edges"]
            if edge["from_repo"] == "Portal Repo"
            and edge["to_repo"] == "Issue SDK"
            and edge["edge_kind"] == "module_dependency"
        )
        with sqlite3.connect(self.service._index_path(portal_path)) as connection:
            flow_edges = list(connection.execute("select edge_kind, to_name from flow_edges"))

        self.assertGreaterEqual(edge["confidence"], 0.84)
        self.assertIn(("module_dependency", "@example/issue-sdk"), flow_edges)

    def test_messaging_publish_consume_edges_create_cross_repo_graph(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[
                {"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"},
                {"display_name": "Issue Service", "url": "https://git.example.com/team/issue-service.git"},
            ],
        )
        entries = self.service.load_config()["mappings"]["AF:All"]
        portal = type("Entry", (), entries[0])()
        service_entry = type("Entry", (), entries[1])()
        portal_path = self.service._repo_path("AF:All", portal)
        service_path = self.service._repo_path("AF:All", service_entry)
        (portal_path / ".git").mkdir(parents=True)
        (service_path / ".git").mkdir(parents=True)
        publisher_file = portal_path / "events" / "IssueEventPublisher.java"
        publisher_file.parent.mkdir(parents=True)
        publisher_file.write_text(
            "public class IssueEventPublisher {\n"
            "    public void publish(IssueCreatedEvent event) {\n"
            "        kafkaTemplate.send(\"issue.created\", event);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        listener_file = service_path / "events" / "IssueEventListener.java"
        listener_file.parent.mkdir(parents=True)
        listener_file.write_text(
            "public class IssueEventListener {\n"
            "    @KafkaListener(topics = \"issue.created\")\n"
            "    public void consume(IssueCreatedEvent event) { }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="which service consumes issue created event")
        policy_statuses = {policy["name"]: policy["status"] for policy in payload["answer_quality"]["policies"]}
        self.assertEqual(policy_statuses["message_flow"], "satisfied")
        self.assertIn(payload["matches"][0]["path"], {"events/IssueEventPublisher.java", "events/IssueEventListener.java"})
        edge = next(
            edge
            for edge in payload["repo_graph"]["edges"]
            if edge["from_repo"] == "Portal Repo"
            and edge["to_repo"] == "Issue Service"
            and edge["edge_kind"] == "message_topic"
        )
        with sqlite3.connect(self.service._index_path(portal_path)) as connection:
            publish_edges = list(connection.execute("select edge_kind, to_name from flow_edges"))
        with sqlite3.connect(self.service._index_path(service_path)) as connection:
            consume_edges = list(connection.execute("select edge_kind, to_name from flow_edges"))

        self.assertGreaterEqual(edge["confidence"], 0.93)
        self.assertIn(("message_publish", "issue.created"), publish_edges)
        self.assertIn(("message_consume", "issue.created"), consume_edges)

    def test_configured_message_topic_placeholders_create_cross_repo_graph(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[
                {"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"},
                {"display_name": "Issue Service", "url": "https://git.example.com/team/issue-service.git"},
            ],
        )
        entries = self.service.load_config()["mappings"]["AF:All"]
        portal = type("Entry", (), entries[0])()
        service_entry = type("Entry", (), entries[1])()
        portal_path = self.service._repo_path("AF:All", portal)
        service_path = self.service._repo_path("AF:All", service_entry)
        (portal_path / ".git").mkdir(parents=True)
        (service_path / ".git").mkdir(parents=True)
        publisher_file = portal_path / "events" / "IssueEventPublisher.java"
        publisher_file.parent.mkdir(parents=True)
        publisher_file.write_text(
            "public class IssueEventPublisher {\n"
            "    public void publish(IssueCreatedEvent event) {\n"
            "        kafkaTemplate.send(\"${issue.topic.name}\", event);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        (portal_path / "application.properties").write_text("issue.topic.name=issue.created\n", encoding="utf-8")
        listener_file = service_path / "events" / "IssueEventListener.java"
        listener_file.parent.mkdir(parents=True)
        listener_file.write_text(
            "public class IssueEventListener {\n"
            "    @KafkaListener(topics = \"${issue.topic.name}\")\n"
            "    public void consume(IssueCreatedEvent event) { }\n"
            "}\n",
            encoding="utf-8",
        )
        (service_path / "application.properties").write_text("issue.topic.name=issue.created\n", encoding="utf-8")
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="which service consumes configured issue topic")
        policy_statuses = {policy["name"]: policy["status"] for policy in payload["answer_quality"]["policies"]}
        self.assertEqual(policy_statuses["message_flow"], "satisfied")
        self.assertIn(payload["matches"][0]["path"], {"events/IssueEventPublisher.java", "events/IssueEventListener.java", "application.properties"})
        edge = next(
            edge
            for edge in payload["repo_graph"]["edges"]
            if edge["from_repo"] == "Portal Repo"
            and edge["to_repo"] == "Issue Service"
            and edge["edge_kind"] == "message_topic"
        )

        self.assertGreaterEqual(edge["confidence"], 0.93)
        self.assertIn("issue.created", edge["match_reason"])

    def test_exact_build_artifact_matching_links_repo_without_display_alias(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[
                {"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"},
                {"display_name": "Payments Core", "url": "https://git.example.com/team/payments-core.git"},
            ],
        )
        entries = self.service.load_config()["mappings"]["AF:All"]
        portal = type("Entry", (), entries[0])()
        target = type("Entry", (), entries[1])()
        portal_path = self.service._repo_path("AF:All", portal)
        target_path = self.service._repo_path("AF:All", target)
        (portal_path / ".git").mkdir(parents=True)
        (target_path / ".git").mkdir(parents=True)
        (portal_path / "pom.xml").write_text(
            "<project><dependencies><dependency>"
            "<groupId>com.example</groupId><artifactId>fraud-ledger-api</artifactId><version>1.0.0</version>"
            "</dependency></dependencies></project>",
            encoding="utf-8",
        )
        (target_path / "pom.xml").write_text(
            "<project><groupId>com.example</groupId><artifactId>fraud-ledger-api</artifactId></project>",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="which repo provides fraud-ledger-api dependency")
        edge = next(
            edge
            for edge in payload["repo_graph"]["edges"]
            if edge["from_repo"] == "Portal Repo"
            and edge["to_repo"] == "Payments Core"
            and edge["edge_kind"] == "module_dependency"
        )

        self.assertGreaterEqual(edge["confidence"], 0.97)
        self.assertIn("exact build artifact match", edge["match_reason"])

    def test_shared_table_write_read_creates_cross_repo_lineage_edge(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[
                {"display_name": "Issue Writer", "url": "https://git.example.com/team/issue-writer.git"},
                {"display_name": "Issue Reporting", "url": "https://git.example.com/team/issue-reporting.git"},
            ],
        )
        entries = self.service.load_config()["mappings"]["AF:All"]
        writer = type("Entry", (), entries[0])()
        reader = type("Entry", (), entries[1])()
        writer_path = self.service._repo_path("AF:All", writer)
        reader_path = self.service._repo_path("AF:All", reader)
        (writer_path / ".git").mkdir(parents=True)
        (reader_path / ".git").mkdir(parents=True)
        writer_file = writer_path / "repository" / "IssueWriterRepository.java"
        writer_file.parent.mkdir(parents=True)
        writer_file.write_text(
            "public class IssueWriterRepository {\n"
            "    public void saveIssue() { jdbcTemplate.update(\"insert into shared_issue_table(id) values (?)\", 1); }\n"
            "}\n",
            encoding="utf-8",
        )
        reader_file = reader_path / "repository" / "IssueReportRepository.java"
        reader_file.parent.mkdir(parents=True)
        reader_file.write_text(
            "public class IssueReportRepository {\n"
            "    public Issue loadIssue() { return jdbcTemplate.queryForObject(\"select * from shared_issue_table\", mapper); }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="which repo reads shared_issue_table after it is written")
        edge = next(
            edge
            for edge in payload["repo_graph"]["edges"]
            if edge["from_repo"] == "Issue Writer"
            and edge["to_repo"] == "Issue Reporting"
            and edge["edge_kind"] == "shared_table"
        )

        self.assertGreaterEqual(edge["confidence"], 0.86)
        self.assertIn("db write/read table overlap", edge["match_reason"])

    def test_sql_read_write_and_spring_condition_edges_are_indexed(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        repository_file = repo_path / "repository" / "IssueRepository.java"
        repository_file.parent.mkdir(parents=True)
        repository_file.write_text(
            "@Profile(\"prod\")\n"
            "public class IssueRepository {\n"
            "    @Value(\"${issue.topic.name:issue.created}\")\n"
            "    private String topicName;\n"
            "    @Qualifier(\"primaryJdbcTemplate\")\n"
            "    private JdbcTemplate jdbcTemplate;\n"
            "    public Issue loadIssue() {\n"
            "        return jdbcTemplate.queryForObject(\"select * from issue_table\", mapper);\n"
            "    }\n"
            "    public void saveIssue() {\n"
            "        jdbcTemplate.update(\"insert into issue_table(id) values (?)\", 1);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)

        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            entity_edges = list(connection.execute("select edge_kind, to_name from entity_edges"))
            flow_edges = list(connection.execute("select edge_kind, to_name from flow_edges"))

        self.assertIn(("spring_profile", "prod"), entity_edges)
        self.assertIn(("config", "issue.topic.name"), entity_edges)
        self.assertIn(("bean_qualifier", "primaryJdbcTemplate"), entity_edges)
        self.assertIn(("db_read", "issue_table"), flow_edges)
        self.assertIn(("db_write", "issue_table"), flow_edges)

    def test_followup_context_augments_short_question(self):
        question, context = self.service._apply_conversation_context(
            "继续找这个表哪里写入",
            {
                "question": "what table does issue creation use",
                "matches": [{"path": "repository/IssueRepository.java", "snippet": "select * from issue_table"}],
                "trace_paths": [{"edges": [{"to_name": "issue_table", "to_file": ""}]}],
                "structured_answer": {"claims": [{"text": "IssueRepository reads issue_table"}]},
            },
        )

        self.assertTrue(context["used"])
        self.assertIn("issue_table", question)

    def test_domain_profile_terms_can_boost_configured_data_source_names(self):
        profile_path = Path(self.temp_dir.name) / "profiles.json"
        profile_path.write_text(
            '{'
            '"default": {"source_terms": ["LedgerGateway"], "data_carriers": ["LedgerDTO"]},'
            '"AF": {"source_terms": ["FraudLedgerRepository"]}'
            '}',
            encoding="utf-8",
        )
        self.service.domain_profile_path = profile_path
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        repository_file = repo_path / "repository" / "FraudLedgerRepository.java"
        repository_file.parent.mkdir(parents=True)
        repository_file.write_text(
            "public class FraudLedgerRepository {\n"
            "    public LedgerDTO loadLedger() {\n"
            "        return jdbcTemplate.queryForObject(\"select * from fraud_ledger\", mapper);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="what data source is used for ledger")

        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["matches"][0]["path"], "repository/FraudLedgerRepository.java")

    def test_retrieval_prefers_symbol_definition_over_noisy_readme(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)

        readme = repo_path / "README.md"
        readme.write_text(
            "batchCreateJiraIssue batchCreateJiraIssue batchCreateJiraIssue\n"
            "This repo mentions batchCreateJiraIssue many times in docs.\n",
            encoding="utf-8",
        )

        source_file = repo_path / "bpmis" / "jira_client.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "class BPMISClient:\n"
            "    def batchCreateJiraIssue(self):\n"
            "        payload = {'route': '/api/v1/issues/batchCreateJiraIssue'}\n"
            "        return self.post('/api/v1/issues/batchCreateJiraIssue', json=payload)\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="where is batchCreateJiraIssue implemented")

        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["matches"][0]["path"], "bpmis/jira_client.py")
        self.assertTrue(
            "symbol matched" in payload["matches"][0]["reason"]
            or "planner_definition" in payload["matches"][0]["reason"]
            or "structure matched" in payload["matches"][0]["reason"]
        )

    def test_dependency_question_expands_to_service_and_integration_matches(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)

        engine_file = repo_path / "decision" / "TermLoanPreCheckEngine.java"
        engine_file.parent.mkdir(parents=True)
        engine_file.write_text(
            "public class TermLoanPreCheckEngine {\n"
            "    private ScreeningService screeningService;\n"
            "    private EcInfoIntegration ecInfoIntegration;\n"
            "    public void runPreCheck1() {\n"
            "        screeningService.screenApplicant();\n"
            "        ecInfoIntegration.fetchEcInfo();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )

        service_file = repo_path / "service" / "ScreeningService.java"
        service_file.parent.mkdir(parents=True)
        service_file.write_text(
            "public class ScreeningService {\n"
            "    public void screenApplicant() {\n"
            "        // data screening source logic\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )

        integration_file = repo_path / "integration" / "EcInfoIntegration.java"
        integration_file.parent.mkdir(parents=True)
        integration_file.write_text(
            "public class EcInfoIntegration {\n"
            "    public void fetchEcInfo() {\n"
            "        // upstream ecinfo source\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(
            pm_team="AF",
            country="All",
            question="What data sources and integrations are checked for Term Loan Pre Check 1",
        )

        self.assertEqual(payload["status"], "ok")
        paths = [match["path"] for match in payload["matches"]]
        self.assertIn("decision/TermLoanPreCheckEngine.java", paths)
        self.assertIn("service/ScreeningService.java", paths)
        self.assertIn("integration/EcInfoIntegration.java", paths)
        dependency_reasons = [match["reason"] for match in payload["matches"] if match.get("trace_stage") == "dependency"]
        self.assertTrue(any("dependency trace" in reason for reason in dependency_reasons))

    def test_two_hop_trace_follows_downstream_service_from_entry_point(self):
        self.service.save_mapping(
            pm_team="CRMS",
            country="ID",
            repositories=[{"display_name": "Credit Risk", "url": "https://git.example.com/team/credit-risk.git"}],
        )
        entry = self.service.load_config()["mappings"]["CRMS:ID"][0]
        repo_path = self.service._repo_path("CRMS:ID", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)

        engine_file = repo_path / "underwriting" / "TermLoanPreCheckEngine.java"
        engine_file.parent.mkdir(parents=True)
        engine_file.write_text(
            "public class TermLoanPreCheckEngine {\n"
            "    private EcInfoService ecInfoService;\n"
            "    public void runPreCheck1() {\n"
            "        ecInfoService.loadEcInfo();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )

        service_file = repo_path / "service" / "EcInfoService.java"
        service_file.parent.mkdir(parents=True)
        service_file.write_text(
            "public class EcInfoService {\n"
            "    public CustomerEcInfo loadEcInfo() {\n"
            "        return jdbcTemplate.queryForObject(\"select * from cr_ec_info\", mapper);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(
            pm_team="CRMS",
            country="ID",
            question="What data sources are checked for Term Loan Pre Check 1",
        )

        paths = [match["path"] for match in payload["matches"]]
        self.assertIn("underwriting/TermLoanPreCheckEngine.java", paths)
        self.assertIn("service/EcInfoService.java", paths)
        two_hop_reasons = [match["reason"] for match in payload["matches"] if match.get("trace_stage") == "two_hop"]
        self.assertTrue(any("two-hop trace" in reason for reason in two_hop_reasons))

    def test_agent_trace_can_follow_service_to_repository(self):
        self.service.save_mapping(
            pm_team="CRMS",
            country="ID",
            repositories=[{"display_name": "Credit Risk", "url": "https://git.example.com/team/credit-risk.git"}],
        )
        entry = self.service.load_config()["mappings"]["CRMS:ID"][0]
        repo_path = self.service._repo_path("CRMS:ID", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)

        engine_file = repo_path / "engine" / "TermLoanPreCheckEngine.java"
        engine_file.parent.mkdir(parents=True)
        engine_file.write_text(
            "public class TermLoanPreCheckEngine {\n"
            "    private EligibilityService eligibilityService;\n"
            "    public void runPreCheck1() {\n"
            "        eligibilityService.evaluateEligibility();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )

        service_file = repo_path / "service" / "EligibilityService.java"
        service_file.parent.mkdir(parents=True)
        service_file.write_text(
            "public class EligibilityService {\n"
            "    private CustomerRepository customerRepository;\n"
            "    public CustomerProfile evaluateEligibility() {\n"
            "        return customerRepository.findCustomerProfile();\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )

        repository_file = repo_path / "repository" / "CustomerRepository.java"
        repository_file.parent.mkdir(parents=True)
        repository_file.write_text(
            "public class CustomerRepository {\n"
            "    public CustomerProfile findCustomerProfile() {\n"
            "        return jdbcTemplate.queryForObject(\"select * from cr_customer_profile\", mapper);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        base_matches = self.service._search_repo(
            type("Entry", (), entry)(),
            repo_path,
            self.service._question_tokens("Term Loan Pre Check 1"),
            question="What data source does Term Loan Pre Check 1 check?",
        )
        agent_matches = self.service._expand_agent_trace_matches(
            entries=[type("Entry", (), entry)()],
            key="CRMS:ID",
            question="What data source does Term Loan Pre Check 1 check?",
            base_matches=base_matches,
            limit=12,
        )

        paths = [match["path"] for match in agent_matches]
        self.assertIn("repository/CustomerRepository.java", paths)
        agent_reasons = [match["reason"] for match in agent_matches if str(match.get("trace_stage") or "").startswith("agent_trace")]
        self.assertTrue(any("agent trace" in reason for reason in agent_reasons))

    def test_evidence_compressor_extracts_concrete_data_sources(self):
        matches = [
            {
                "repo": "Credit Risk",
                "path": "repository/CustomerRepository.java",
                "line_start": 10,
                "line_end": 14,
                "score": 100,
                "trace_stage": "agent_trace_2",
                "reason": "agent trace",
                "snippet": (
                    "public CustomerProfile loadProfile() {\n"
                    "    return jdbcTemplate.queryForObject(\"select * from cr_customer_profile\", mapper);\n"
                    "}\n"
                ),
            }
        ]

        compressed = self.service._compress_evidence("What data source does Term Loan Pre Check 1 check?", matches)
        quality = self.service._quality_gate("What data source does Term Loan Pre Check 1 check?", compressed)

        self.assertTrue(compressed["data_sources"])
        self.assertIn("cr_customer_profile", " ".join(compressed["data_sources"]))
        self.assertEqual(quality["status"], "sufficient")
        self.assertIn("production:2", compressed["data_source_tiers"])

    def test_test_only_data_source_requires_production_evidence(self):
        matches = [
            {
                "repo": "Credit Risk",
                "path": "src/test/java/repository/CustomerRepositoryTest.java",
                "line_start": 10,
                "line_end": 14,
                "score": 100,
                "trace_stage": "direct",
                "reason": "test fixture matched",
                "snippet": (
                    "public void loadsProfile() {\n"
                    "    jdbcTemplate.queryForObject(\"select * from cr_customer_profile\", mapper);\n"
                    "}\n"
                ),
            }
        ]

        compressed = self.service._compress_evidence("What data source does Term Loan Pre Check 1 check?", matches)
        quality = self.service._quality_gate("What data source does Term Loan Pre Check 1 check?", compressed)
        pack = self.service._build_evidence_pack(
            question="What data source does Term Loan Pre Check 1 check?",
            evidence_summary=compressed,
            matches=matches,
            trace_paths=[],
            quality_gate=quality,
        )

        self.assertIn("test:2", compressed["data_source_tiers"])
        self.assertTrue(compressed["source_conflicts"])
        self.assertEqual(quality["status"], "needs_more_trace")
        self.assertIn("production repository/mapper/client/table evidence beyond test/docs/generated files", quality["missing"])
        self.assertTrue(pack["source_conflicts"])
        self.assertIn("Production source evidence was not found", " ".join(pack["missing_hops"]))

    def test_evidence_pack_structures_tables_and_missing_hops(self):
        matches = [
            {
                "repo": "Credit Risk",
                "path": "repository/CustomerRepository.java",
                "line_start": 10,
                "line_end": 14,
                "score": 100,
                "trace_stage": "agent_trace_2",
                "reason": "agent trace",
                "snippet": (
                    "public CustomerProfile loadProfile() {\n"
                    "    return jdbcTemplate.queryForObject(\"select * from cr_customer_profile\", mapper);\n"
                    "}\n"
                ),
            }
        ]
        compressed = self.service._compress_evidence("What data source does Term Loan Pre Check 1 check?", matches)
        quality = self.service._quality_gate("What data source does Term Loan Pre Check 1 check?", compressed)

        pack = self.service._build_evidence_pack(
            question="What data source does Term Loan Pre Check 1 check?",
            evidence_summary=compressed,
            matches=matches,
            trace_paths=[],
            quality_gate=quality,
        )

        self.assertEqual(pack["version"], 2)
        self.assertTrue(any(item["type"] == "table" for item in pack["items"]))
        self.assertIn("cr_customer_profile", " ".join(pack["tables"]))
        self.assertTrue(pack["read_write_points"])
        self.assertFalse(pack["missing_hops"])

    def test_planner_tool_dsl_finds_tables_routes_and_callees(self):
        _build_fixture_repositories(self.service)
        af_entry = self.service.load_config()["mappings"]["AF:All"][0]
        entry = type("Entry", (), af_entry)()
        repo_path = self.service._repo_path("AF:All", entry)
        self.service._build_repo_index("AF:All", entry, repo_path)

        table_matches = self.service._tool_find_tables(
            entry,
            repo_path,
            ["issue"],
            "what table does issue creation use",
            1,
        )
        route_matches = self.service._tool_find_api_routes(
            entry,
            repo_path,
            ["issue"],
            "which API handles issue create",
            1,
        )
        base_matches = self.service._search_repo(
            entry,
            repo_path,
            self.service._question_tokens("IssueController createIssue"),
            question="which service or repository does issue creation call",
        )
        callee_matches = self.service._tool_find_callees(
            entry,
            repo_path,
            base_matches,
            ["Issue"],
            "which service or repository does issue creation call",
            1,
        )

        self.assertIn("issue_table", " ".join(match["snippet"] for match in table_matches))
        self.assertTrue(any(match.get("retrieval") == "planner_api_route" for match in route_matches))
        self.assertTrue(any(match.get("retrieval") == "planner_callee" for match in callee_matches))

    def test_quality_gate_requests_deeper_trace_when_data_source_is_missing(self):
        matches = [
            {
                "repo": "Credit Risk",
                "path": "engine/TermLoanPreCheckEngine.java",
                "line_start": 1,
                "line_end": 5,
                "score": 80,
                "trace_stage": "direct",
                "reason": "content matched",
                "snippet": (
                    "public class TermLoanPreCheckEngine {\n"
                    "    public void run(EngineTermLoanPreCheckLayer1Input input) {}\n"
                    "}\n"
                ),
            }
        ]

        compressed = self.service._compress_evidence("What data source does Term Loan Pre Check 1 check?", matches)
        quality = self.service._quality_gate("What data source does Term Loan Pre Check 1 check?", compressed)
        trace_terms = self.service._quality_gate_trace_terms(
            "What data source does Term Loan Pre Check 1 check?",
            compressed,
            quality,
            matches,
        )

        self.assertEqual(quality["status"], "needs_more_trace")
        self.assertIn("concrete upstream source/table/API/repository evidence beyond DTO fields", quality["missing"])
        self.assertIn("repository", trace_terms)

    def test_quality_gate_trace_can_find_repository_after_weak_entry_match(self):
        self.service.save_mapping(
            pm_team="CRMS",
            country="ID",
            repositories=[{"display_name": "Credit Risk", "url": "https://git.example.com/team/credit-risk.git"}],
        )
        entry = self.service.load_config()["mappings"]["CRMS:ID"][0]
        repo_path = self.service._repo_path("CRMS:ID", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)

        engine_file = repo_path / "engine" / "TermLoanPreCheckEngine.java"
        engine_file.parent.mkdir(parents=True)
        engine_file.write_text(
            "public class TermLoanPreCheckEngine {\n"
            "    public void run(EngineTermLoanPreCheckLayer1Input input) {}\n"
            "}\n",
            encoding="utf-8",
        )

        repository_file = repo_path / "repository" / "DataSourceResultRepository.java"
        repository_file.parent.mkdir(parents=True)
        repository_file.write_text(
            "public class DataSourceResultRepository {\n"
            "    public DataSourceResult loadDataSourceResult() {\n"
            "        return jdbcTemplate.queryForObject(\"select * from cr_data_source_result\", mapper);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        base_matches = self.service._search_repo(
            type("Entry", (), entry)(),
            repo_path,
            self.service._question_tokens("Term Loan Pre Check 1"),
            question="What data source does Term Loan Pre Check 1 check?",
        )
        compressed = self.service._compress_evidence("What data source does Term Loan Pre Check 1 check?", base_matches[:1])
        quality = self.service._quality_gate("What data source does Term Loan Pre Check 1 check?", compressed)
        agent_plan = self.service._build_agent_plan(
            "What data source does Term Loan Pre Check 1 check?",
            compressed,
            quality,
        )
        quality_matches = self.service._run_agent_plan(
            entries=[type("Entry", (), entry)()],
            key="CRMS:ID",
            question="What data source does Term Loan Pre Check 1 check?",
            matches=base_matches[:1],
            evidence_summary=compressed,
            quality_gate=quality,
            agent_plan=agent_plan,
            limit=12,
        )

        self.assertIn("repository/DataSourceResultRepository.java", [match["path"] for match in quality_matches])
        self.assertTrue(any(str(match.get("trace_stage") or "").startswith("agent_plan") for match in quality_matches))

    def test_agent_plan_includes_data_source_trace_steps(self):
        compressed = {
            "intent": self.service._question_intent("What data source does Term Loan Pre Check 1 check?"),
            "entry_points": ["Credit Risk:engine/TermLoanPreCheckEngine.java:1-5"],
            "data_carriers": ["EngineTermLoanPreCheckLayer1Input"],
            "field_population": [],
            "downstream_components": [],
            "data_sources": [],
            "api_or_config": [],
            "rule_or_error_logic": [],
            "source_count": 1,
        }
        quality = self.service._quality_gate("What data source does Term Loan Pre Check 1 check?", compressed)

        plan = self.service._build_agent_plan("What data source does Term Loan Pre Check 1 check?", compressed, quality)

        step_names = [step["name"] for step in plan["steps"]]
        self.assertIn("trace_entry_to_source_lineage", step_names)
        self.assertIn("trace_data_carriers", step_names)
        self.assertIn("trace_field_population", step_names)
        self.assertIn("trace_downstream_sources", step_names)
        self.assertIn("trace_dao_mapper_methods", step_names)
        self.assertIn("fill_quality_gap", step_names)
        flattened_terms = " ".join(term for step in plan["steps"] for term in step["terms"])
        self.assertIn("EngineTermLoanPreCheckLayer1Input", flattened_terms)
        self.assertNotIn("UnderwritingInitiationDTO", flattened_terms)

    def test_answer_judge_requests_repair_for_uncited_thin_answer(self):
        evidence_pack = {
            "version": 2,
            "intent": self.service._question_intent("What data source does issue creation use?"),
            "items": [
                {
                    "type": "table",
                    "claim": "issue_table is referenced in Portal Repo:repository/IssueRepository.java:1-5",
                    "source_id": "S1",
                    "confidence": "high",
                    "hop": "source",
                    "supports_answer": True,
                }
            ],
            "tables": ["issue_table (Portal Repo:repository/IssueRepository.java:1-5)"],
            "apis": [],
            "missing_hops": [],
        }

        judge = self.service._judge_answer(
            "What data source does issue creation use?",
            "It uses a repository.",
            evidence_pack,
            {"status": "needs_citation", "issues": ["concrete answer claims need citation-backed evidence"]},
        )

        self.assertEqual(judge["status"], "repair")
        self.assertTrue(judge["repair_targets"])

    def test_deterministic_answer_judge_can_request_repair(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        evidence_pack = {
            "version": 2,
            "intent": service._question_intent("What data source does issue creation use?"),
            "items": [
                {
                    "type": "table",
                    "claim": "issue_table is referenced in Portal Repo:repository/IssueRepository.java:1-5",
                    "source_id": "S1",
                    "confidence": "high",
                    "supports_answer": True,
                }
            ],
            "tables": ["issue_table (Portal Repo:repository/IssueRepository.java:1-5)"],
            "apis": [],
            "missing_hops": [],
        }
        result = LLMGenerateResult(
            payload={
                "candidates": [
                    {
                        "content": {
                            "parts": [
                                {
                                    "text": json.dumps(
                                        {
                                            "status": "repair",
                                            "confidence": "high",
                                            "issues": ["answer omits issue_table"],
                                            "repair_targets": ["include table source"],
                                        }
                                    )
                                }
                            ]
                        }
                    }
                ],
                "usageMetadata": {"promptTokenCount": 10, "candidatesTokenCount": 8},
            },
            usage={"promptTokenCount": 10, "candidatesTokenCount": 8},
            model="judge-lite",
            attempts=1,
        )

        with patch.object(service.llm_provider, "generate", return_value=result) as mocked_generate:
            judge = service._run_answer_judge(
                "What data source does issue creation use?",
                "It uses a repository.",
                evidence_pack,
                {"status": "ok", "issues": []},
            )

        self.assertEqual(judge["mode"], "deterministic_evidence_judge")
        self.assertEqual(judge["status"], "repair")
        self.assertIn("answer omits typed table/API/client source evidence", judge["issues"])
        mocked_generate.assert_not_called()

    def test_deterministic_answer_judge_does_not_call_provider(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        evidence_pack = {
            "version": 2,
            "intent": service._question_intent("What data source does issue creation use?"),
            "items": [
                {
                    "type": "table",
                    "claim": "issue_table is referenced in Portal Repo:repository/IssueRepository.java:1-5",
                    "source_id": "S1",
                    "confidence": "high",
                    "supports_answer": True,
                }
            ],
            "tables": ["issue_table (Portal Repo:repository/IssueRepository.java:1-5)"],
            "apis": [],
            "missing_hops": [],
        }
        result = LLMGenerateResult(
            payload={
                "candidates": [
                    {
                        "content": {
                            "parts": [
                                {
                                    "text": json.dumps(
                                        {
                                            "status": "ok",
                                            "confidence": "high",
                                            "issues": [],
                                            "repair_targets": [],
                                        }
                                    )
                                }
                            ]
                        }
                    }
                ],
                "usageMetadata": {"promptTokenCount": 10, "candidatesTokenCount": 8},
            },
            usage={"promptTokenCount": 10, "candidatesTokenCount": 8},
            model="codex-judge",
            attempts=1,
        )

        with patch.object(service.llm_provider, "generate", return_value=result) as mocked_generate:
            first = service._run_answer_judge(
                "What data source does issue creation use?",
                "IssueRepository reads issue_table [S1].",
                evidence_pack,
                {"status": "ok", "issues": []},
            )
            second = service._run_answer_judge(
                "What data source does issue creation use?",
                "IssueRepository reads issue_table [S1].",
                evidence_pack,
                {"status": "ok", "issues": []},
            )

        self.assertEqual(first["mode"], "deterministic_evidence_judge")
        self.assertEqual(second["mode"], "deterministic_evidence_judge")
        self.assertFalse(second.get("cached", False))
        mocked_generate.assert_not_called()

    def test_imported_dao_is_not_treated_as_concrete_data_source(self):
        matches = [
            {
                "repo": "Credit Risk",
                "path": "service/UnderwritingInitiationService.java",
                "line_start": 1,
                "line_end": 6,
                "score": 100,
                "trace_stage": "agent_plan_3",
                "reason": "agent plan trace",
                "snippet": (
                    "import com.shopee.banking.crm.dao.CustomerInfoDAO;\n"
                    "public class UnderwritingInitiationService {\n"
                    "    private CustomerInfoDAO customerInfoDAO;\n"
                    "}\n"
                ),
            }
        ]

        compressed = self.service._compress_evidence("What data source does Term Loan Pre Check 1 check?", matches)
        quality = self.service._quality_gate("What data source does Term Loan Pre Check 1 check?", compressed)

        self.assertFalse(compressed["data_sources"])
        self.assertEqual(quality["status"], "needs_more_trace")
        self.assertEqual(quality["policies"][0]["name"], "data_source")
        self.assertEqual(quality["policies"][0]["status"], "missing")

    def test_field_level_data_flow_edges_are_indexed(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        service_file = repo_path / "service" / "IssueService.java"
        service_file.parent.mkdir(parents=True)
        service_file.write_text(
            "public class IssueService {\n"
            "    public EngineInput buildInput(DataSourceResult result) {\n"
            "        EngineInput input = new EngineInput();\n"
            "        input.setDataSourceResult(result);\n"
            "        return input;\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )

        self.service._build_repo_index("AF:All", type("Entry", (), entry)(), repo_path)
        with sqlite3.connect(self.service._index_path(repo_path)) as connection:
            entity_edges = {row[0] for row in connection.execute("select edge_kind from entity_edges")}
            flow_edges = {row[0] for row in connection.execute("select edge_kind from flow_edges")}

        self.assertIn("data_flow", entity_edges)
        self.assertIn("field_population", flow_edges)

    def test_static_qa_query_finds_security_and_quality_findings(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        risky_file = repo_path / "service" / "RiskyIssueService.java"
        risky_file.parent.mkdir(parents=True)
        risky_file.write_text(
            "public class RiskyIssueService {\n"
            "    private static final String PASSWORD = \"plain-secret\";\n"
            "    public void load(String issueId) {\n"
            "        String sql = \"select * from issue_table where id = \" + issueId;\n"
            "        try { callRemote(); } catch (Exception e) { e.printStackTrace(); }\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        payload = self.service.query(pm_team="AF", country="All", question="what static QA security risks exist")

        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["matches"][0]["path"], "service/RiskyIssueService.java")
        self.assertTrue(any(match.get("retrieval") == "static_qa" for match in payload["matches"]))
        evidence_text = json.dumps(payload["evidence_pack"], ensure_ascii=False)
        self.assertIn("hardcoded_secret", evidence_text)
        self.assertIn("sql_string_concatenation", evidence_text)
        self.assertEqual(payload["answer_quality"]["policies"][0]["name"], "static_qa")
        self.assertEqual(payload["answer_quality"]["policies"][0]["status"], "satisfied")
        self.assertIn("static QA findings", payload["summary"])

    def test_impact_analysis_query_finds_upstream_and_downstream_surfaces(self):
        _build_fixture_repositories(self.service)
        self._build_all_indexes()

        payload = self.service.query(
            pm_team="AF",
            country="All",
            question="what is impacted if IssueService createIssue changes",
        )

        paths = {match["path"] for match in payload["matches"]}
        self.assertIn("controller/IssueController.java", paths)
        self.assertIn("repository/IssueRepository.java", paths)
        self.assertTrue(any(match.get("retrieval") in {"planner_caller", "planner_callee", "flow_graph", "entity_graph"} for match in payload["matches"]))
        evidence_text = json.dumps(payload["evidence_pack"], ensure_ascii=False)
        self.assertIn("impact_surfaces", payload["evidence_pack"])
        self.assertIn("IssueRepository", evidence_text)
        policy_statuses = {policy["name"]: policy["status"] for policy in payload["answer_quality"]["policies"]}
        self.assertEqual(policy_statuses["impact_analysis"], "satisfied")

    def test_test_coverage_query_finds_tests_and_assertions(self):
        _build_fixture_repositories(self.service)
        self._build_all_indexes()

        payload = self.service.query(
            pm_team="AF",
            country="All",
            question="is IssueService createIssue covered by tests",
        )

        self.assertEqual(payload["matches"][0]["path"], "src/test/java/IssueServiceTest.java")
        paths = {match["path"] for match in payload["matches"]}
        self.assertIn("src/test/java/IssueServiceTest.java", paths)
        self.assertTrue(any(match.get("retrieval") == "test_coverage" for match in payload["matches"]))
        evidence_text = json.dumps(payload["evidence_pack"], ensure_ascii=False)
        self.assertIn("test_coverage", payload["evidence_pack"])
        self.assertIn("IssueServiceTest", evidence_text)
        self.assertTrue("verify" in evidence_text or "assert" in evidence_text)
        policy_statuses = {policy["name"]: policy["status"] for policy in payload["answer_quality"]["policies"]}
        self.assertEqual(policy_statuses["test_coverage"], "satisfied")

    def test_operational_boundary_query_finds_transaction_and_cache_annotations(self):
        _build_fixture_repositories(self.service)
        self._build_all_indexes()

        payload = self.service.query(
            pm_team="AF",
            country="All",
            question="is IssueService createIssue transactional or cached",
        )

        self.assertEqual(payload["matches"][0]["path"], "service/IssueService.java")
        paths = {match["path"] for match in payload["matches"]}
        self.assertIn("service/IssueService.java", paths)
        self.assertTrue(any(match.get("retrieval") == "operational_boundary" for match in payload["matches"]))
        evidence_text = json.dumps(payload["evidence_pack"], ensure_ascii=False)
        self.assertIn("operational_boundaries", payload["evidence_pack"])
        self.assertIn("Transactional", evidence_text)
        self.assertIn("Cacheable", evidence_text)
        policy_statuses = {policy["name"]: policy["status"] for policy in payload["answer_quality"]["policies"]}
        self.assertEqual(policy_statuses["operational_boundary"], "satisfied")

    def test_structured_answer_parser_accepts_json_and_fallback_prose(self):
        parsed = self.service._parse_structured_answer(
            '{"direct_answer":"Uses issue_table","confirmed_from_code":["IssueRepository reads issue_table [S1]"],'
            '"investigation_steps":{"candidate_evidence":["Opened IssueRepository"],"gap_verification":["Checked mapper"],"certainty_split":["Confirmed table only"]},'
            '"inferred_from_code":["IssueService calls the repository"],"not_found":["No caller evidence"],'
            '"claims":[{"text":"IssueRepository reads issue_table","citations":["S1"]}],"missing_evidence":[],"confidence":"high"}'
        )
        fallback = self.service._parse_structured_answer("IssueRepository reads issue_table [S1].")

        self.assertEqual(parsed["format"], "json")
        self.assertEqual(parsed["claims"][0]["citations"], ["S1"])
        self.assertEqual(parsed["confirmed_from_code"], ["IssueRepository reads issue_table [S1]"])
        self.assertEqual(parsed["investigation_steps"]["gap_verification"], ["Checked mapper"])
        self.assertEqual(parsed["inferred_from_code"], ["IssueService calls the repository"])
        self.assertEqual(parsed["not_found"], ["No caller evidence"])
        self.assertEqual(fallback["format"], "prose_fallback")
        self.assertEqual(fallback["claims"][0]["citations"], ["S1"])
        self.assertEqual(fallback["confirmed_from_code"], [])

    def test_trusted_model_finalizer_returns_raw_codex_answer(self):
        structured = {
            "direct_answer": "Only the status migration is confirmed; the full rule expression is not present.",
            "investigation_steps": {
                "candidate_evidence": ["Opened migration SQL"],
                "gap_verification": ["Searched for rule_config_tab inserts"],
                "certainty_split": ["Confirmed status update; full expression missing"],
            },
            "confirmed_from_code": ["Migration updates C0204v2 status [S1]"],
            "inferred_from_code": ["Rule likely maps to the v2 challenge rule family"],
            "not_found": ["No full rule_config_tab row or feature_expr for C0204v2 was found"],
            "claims": [{"text": "Migration updates C0204v2 status", "citations": ["S1"]}],
            "missing_evidence": ["Live rule_config_tab row is required to confirm feature_expr"],
            "confidence": "medium",
            "format": "json",
        }
        raw_answer = json.dumps(structured)

        final = self.service._finalize_trusted_model_answer(
            question="why are C0204v2 and C0205v2 both needed",
            answer=raw_answer,
            structured_answer=structured,
            evidence_summary={"intent": self.service._question_intent("why are C0204v2 and C0205v2 both needed")},
            quality_gate={"status": "sufficient", "confidence": "medium", "missing": []},
            claim_check={"status": "skipped", "issues": []},
            selected_matches=[{"repo": "AF", "path": "dml.sql", "line_start": 1, "line_end": 2, "snippet": "update rule_config_tab"}],
        )

        self.assertEqual(final["answer"], raw_answer)
        self.assertIn("Live rule_config_tab row is required to confirm feature_expr", final["answer_contract"]["missing_links"])
        self.assertEqual(final["answer_contract"]["investigation_steps"]["gap_verification"], ["Searched for rule_config_tab inserts"])
        self.assertTrue(final["answer_contract"]["passthrough"])

    def test_eval_runner_checks_trace_paths_and_structured_claims(self):
        self.service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = self.service.load_config()["mappings"]["AF:All"][0]
        repo_path = self.service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        repository_file = repo_path / "repository" / "IssueRepository.java"
        repository_file.parent.mkdir(parents=True)
        repository_file.write_text(
            "public class IssueRepository {\n"
            "    public Issue findIssue() {\n"
            "        return jdbcTemplate.queryForObject(\"select * from issue_table\", mapper);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes()

        result = _evaluate_case(
            self.service,
            {
                "id": "trace-path-eval",
                "pm_team": "AF",
                "country": "All",
                "question": "what table does IssueRepository use",
                "expected_paths": ["repository/IssueRepository.java"],
                "min_trace_paths": 1,
                "expected_trace_path_terms": ["issue_table"],
                "expected_answer_policy_statuses": {"data_source": "satisfied"},
                "expected_evidence_pack_terms": ["issue_table"],
                "category": "data_source",
            },
        )

        self.assertEqual(result["status"], "pass")
        self.assertEqual(result["category"], "data_source")
        self.assertEqual(result["answer_policies"]["data_source"], "satisfied")

    def test_codex_query_returns_answer_and_usage(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = service.load_config()["mappings"]["AF:All"][0]
        repo_path = service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "bpmis" / "jira_client.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "class BPMISClient:\n"
            "    def batchCreateJiraIssue(self):\n"
            "        return self.post('/api/v1/issues/batchCreateJiraIssue')\n",
            encoding="utf-8",
        )

        expected_model = service.llm_budgets["balanced"]["model"]
        result = self._llm_result(
            '{"direct_answer":"Short answer: batchCreateJiraIssue is in BPMISClient.",'
            '"claims":[{"text":"BPMISClient defines batchCreateJiraIssue","citations":["S1"]}],'
            '"missing_evidence":[],"confidence":"high"}',
            usage={"promptTokenCount": 123, "candidatesTokenCount": 45, "totalTokenCount": 168},
            model=expected_model,
        )
        with patch.object(service.llm_provider, "ready", return_value=True), patch.object(
            service.llm_provider,
            "generate",
            return_value=result,
        ) as mocked_generate:
            self._build_all_indexes(service)
            payload = service.query(
                pm_team="AF",
                country="All",
                question="where is batchCreateJiraIssue",
                answer_mode="auto",
                llm_budget_mode="balanced",
            )

        self.assertEqual(payload["answer_mode"], "auto")
        self.assertIn("Short answer", payload["llm_answer"])
        self.assertEqual(payload["llm_usage"]["promptTokenCount"], 123)
        self.assertEqual(payload["llm_provider"], "codex_cli_bridge")
        self.assertEqual(payload["llm_model"], expected_model)
        self.assertFalse(payload["llm_cached"])
        self.assertEqual(mocked_generate.call_args_list[0].kwargs["primary_model"], expected_model)
        request_payload = mocked_generate.call_args_list[0].kwargs["payload"]
        self.assertEqual(request_payload["generationConfig"]["responseMimeType"], "application/json")
        self.assertNotIn("responseSchema", request_payload["generationConfig"])
        self.assertNotIn("thinkingConfig", request_payload["generationConfig"])
        self.assertEqual(payload["llm_thinking_budget"], 0)
        self.assertEqual(payload["llm_route"]["mode"], "manual")
        telemetry = json.loads(service.telemetry_path.read_text(encoding="utf-8").strip().splitlines()[-1])
        self.assertEqual(telemetry["llm_provider"], "codex_cli_bridge")
        self.assertEqual(telemetry["llm_model"], expected_model)
        self.assertEqual(telemetry["llm_thinking_budget"], 0)
        self.assertEqual(telemetry["llm_route"]["mode"], "manual")
        self.assertEqual(telemetry["versions"]["cache"], 21)
        self.assertIn("llm_latency_ms", telemetry)
        self.assertIn("llm_attempt_log", telemetry)
        self.assertIn("answer_contract", telemetry)
        self.assertIn("evidence_pack_summary", telemetry)
        self.assertIn("answer_judge", telemetry)
        self.assertEqual(payload["evidence_pack"]["version"], 2)
        self.assertIn("answer_judge", payload)

    def test_codex_max_tokens_uses_compact_repair_retry(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = service.load_config()["mappings"]["AF:All"][0]
        repo_path = service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "bpmis" / "jira_client.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "class BPMISClient:\n"
            "    def batchCreateJiraIssue(self):\n"
            "        return self.post('/api/v1/issues/batchCreateJiraIssue')\n",
            encoding="utf-8",
        )
        capped_result = self._llm_result(
            "{",
            usage={"promptTokenCount": 20000, "candidatesTokenCount": 312, "totalTokenCount": 20312},
            finish_reason="MAX_TOKENS",
        )
        repaired_result = self._llm_result(
            '{"direct_answer":"BPMISClient defines batchCreateJiraIssue.",'
            '"claims":[{"text":"BPMISClient defines batchCreateJiraIssue","citations":["S1"]}],'
            '"missing_evidence":[],"confidence":"high"}',
            usage={"promptTokenCount": 4000, "candidatesTokenCount": 140, "totalTokenCount": 4140},
        )

        with patch.object(service.llm_provider, "ready", return_value=True), patch.object(
            service.llm_provider,
            "generate",
            side_effect=[capped_result, repaired_result],
        ) as mocked_generate:
            self._build_all_indexes(service)
            payload = service.query(
                pm_team="AF",
                country="All",
                question="where is batchCreateJiraIssue",
                answer_mode="auto",
                llm_budget_mode="balanced",
            )

        self.assertEqual(mocked_generate.call_count, 2)
        first_request = mocked_generate.call_args_list[0].kwargs["payload"]
        retry_request = mocked_generate.call_args_list[1].kwargs["payload"]
        self.assertNotIn("thinkingConfig", first_request["generationConfig"])
        self.assertNotIn("thinkingConfig", retry_request["generationConfig"])
        self.assertNotIn("maxOutputTokens", retry_request["generationConfig"])
        self.assertLess(
            len(retry_request["contents"][0]["parts"][0]["text"]),
            len(first_request["contents"][0]["parts"][0]["text"]) + 1,
        )
        self.assertIn("BPMISClient defines batchCreateJiraIssue", payload["llm_answer"])
        self.assertEqual(payload["llm_finish_reason"], "STOP")
        self.assertNotEqual(payload["answer_contract"]["status"], "unreliable_llm_answer")
        self.assertEqual(payload["llm_thinking_budget"], 0)

    def test_simple_function_usage_no_hit_still_calls_llm(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Risk Config", "url": "https://git.example.com/team/risk-config.git"}],
        )
        entry = service.load_config()["mappings"]["AF:All"][0]
        repo_path = service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        config_file = repo_path / "anti-fraud-service" / "apollo.properties"
        config_file.parent.mkdir(parents=True)
        config_file.write_text(
            "dbp.antifraud.function.F44={\"functionName\":\"Compare\",\"field\":\"fdMaturityDate\"}\n",
            encoding="utf-8",
        )
        spec_file = repo_path / "spec" / "fields.md"
        spec_file.parent.mkdir(parents=True)
        spec_file.write_text("fdMaturityDate is exposed as a rule form field.\n", encoding="utf-8")
        self._build_all_indexes(service)

        result = self._llm_result(
            '{"direct_answer":"fdMaturityDate only appears in config/spec candidates.",'
            '"claims":[{"text":"fdMaturityDate appears in config/spec evidence","citations":["S1"]}],'
            '"missing_evidence":[],"confidence":"medium"}'
        )

        with patch.object(service.llm_provider, "ready", return_value=True), patch.object(
            service.llm_provider,
            "generate",
            return_value=result,
        ) as mocked_generate:
            payload = service.query(
                pm_team="AF",
                country="All",
                question="Is fdMaturityDate used in any function?",
                answer_mode="auto",
                llm_budget_mode="balanced",
            )

        self.assertEqual(payload["answer_mode"], "auto")
        self.assertEqual(payload["llm_provider"], "codex_cli_bridge")
        self.assertIn("fdMaturityDate only appears", payload["llm_answer"])
        self.assertEqual(mocked_generate.call_count, 1)

    def test_simple_function_usage_with_method_body_still_calls_llm(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Risk Service", "url": "https://git.example.com/team/risk-service.git"}],
        )
        entry = service.load_config()["mappings"]["AF:All"][0]
        repo_path = service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "src" / "main" / "java" / "LoanChecker.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "public class LoanChecker {\n"
            "    public boolean validate() {\n"
            "        return fdMaturityDate != null;\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        self._build_all_indexes(service)
        result = self._llm_result(
            '{"direct_answer":"fdMaturityDate is used in LoanChecker.validate.",'
            '"claims":[{"text":"LoanChecker.validate checks fdMaturityDate","citations":["S1"]}],'
            '"missing_evidence":[],"confidence":"high"}',
            usage={"promptTokenCount": 90, "candidatesTokenCount": 20, "totalTokenCount": 110},
        )

        with patch.object(service.llm_provider, "ready", return_value=True), patch.object(
            service.llm_provider,
            "generate",
            return_value=result,
        ) as mocked_generate:
            payload = service.query(
                pm_team="AF",
                country="All",
                question="Is fdMaturityDate used in any function?",
                answer_mode="auto",
                llm_budget_mode="balanced",
            )

        self.assertEqual(payload["answer_mode"], "auto")
        self.assertNotIn("llm_cost_skip", payload)
        self.assertIn("LoanChecker.validate", payload["llm_answer"])
        mocked_generate.assert_called()

    def test_crms_ph_card_income_extraction_question_prioritizes_storage_tables(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        service.save_mapping(
            pm_team="CRMS",
            country="PH",
            repositories=[{"display_name": "credit-risk", "url": "https://git.example.com/team/credit-risk.git"}],
        )
        entry = service.load_config()["mappings"]["CRMS:PH"][0]
        repo_path = service._repo_path("CRMS:PH", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        extract_mapper = repo_path / "credit-risk-infra" / "src" / "main" / "resources" / "mapper" / "ExtractRecordDAO-ext.xml"
        extract_mapper.parent.mkdir(parents=True)
        extract_mapper.write_text(
            "<mapper>\n"
            "  <update id=\"updateByBizIdAndSceneSelectiveCas\">\n"
            "    update extract_record_tab set response_body = #{updateDO.responseBody} where biz_id = #{bizId}\n"
            "  </update>\n"
            "</mapper>\n",
            encoding="utf-8",
        )
        status_mapper = repo_path / "credit-risk-infra" / "src" / "main" / "resources" / "mapper" / "CardIncomeScreeningFlowStatusDAO-ext.xml"
        status_mapper.parent.mkdir(parents=True, exist_ok=True)
        status_mapper.write_text(
            "<mapper>\n"
            "  <update id=\"updateAllByVersionAndStatusCas\">\n"
            "    update card_income_screening_flow_status_tab set process_info = #{updateDO.processInfo} where underwriting_id = #{underwritingId}\n"
            "  </update>\n"
            "</mapper>\n",
            encoding="utf-8",
        )
        strategy_file = repo_path / "credit-risk-infra" / "src" / "main" / "java" / "PayslipProcessStrategy.java"
        strategy_file.parent.mkdir(parents=True)
        strategy_file.write_text(
            "class PayslipProcessStrategy {\n"
            "  void process(IesResultDTO result, CardIncomeScreenProcessInfo info) {\n"
            "    info.getPayslip().setGrossPay(result.getResult().getGrossPay());\n"
            "  }\n"
            "}\n",
            encoding="utf-8",
        )
        noise_file = repo_path / "credit-risk-experian" / "src" / "main" / "java" / "ExperianEvaluator.java"
        noise_file.parent.mkdir(parents=True)
        noise_file.write_text("class ExperianEvaluator { void query() {} }\n", encoding="utf-8")
        self._build_all_indexes(service)

        payload = service.query(
            pm_team="CRMS",
            country="PH",
            question="Which table stores the Credit Card LLM extracted payslip fields? And which table stores the Ops extracted fields?",
            answer_mode="retrieval_only",
        )

        paths = [match["path"] for match in payload["matches"][:8]]
        self.assertTrue(any("ExtractRecordDAO-ext.xml" in path for path in paths), paths)
        self.assertTrue(any("CardIncomeScreeningFlowStatusDAO-ext.xml" in path for path in paths), paths)
        self.assertIn("extract_record_tab", payload["exact_lookup"]["matched_terms"])
        self.assertIn("card_income_screening_flow_status_tab", payload["exact_lookup"]["matched_terms"])

    def test_llm_answer_cache_requires_current_versions(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        cache_key = service._answer_cache_key(
            provider="codex_cli_bridge",
            model="codex-cli",
            question="where is createIssue",
            answer_mode="auto",
            llm_budget_mode="balanced",
            context="S1 createIssue",
        )
        service._store_cached_answer(
            cache_key,
            answer="createIssue is in IssueController [S1]",
            usage={"totalTokenCount": 12},
            answer_quality={"status": "sufficient"},
            provider="codex_cli_bridge",
            model="codex-cli",
        )
        cached = service._load_cached_answer(cache_key)
        self.assertIsNotNone(cached)
        self.assertEqual(cached["versions"]["cache"], 21)
        cache_path = service.answer_cache_root / f"{cache_key}.json"
        stale_payload = json.loads(cache_path.read_text(encoding="utf-8"))
        stale_payload["versions"]["router"] = -1
        cache_path.write_text(json.dumps(stale_payload), encoding="utf-8")
        self.assertIsNone(service._load_cached_answer(cache_key))

    def test_codex_provider_uses_local_embedding(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            llm_provider="codex_cli_bridge",
            semantic_index_model="local-token-hybrid-v1",
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        self.assertEqual(service.llm_provider_name, "codex_cli_bridge")
        self.assertEqual(service.llm_provider.name, "codex_cli_bridge")
        self.assertEqual(service.llm_policy_payload()["semantic_retrieval"]["embedding_provider"]["provider"], "local_token_hybrid")
        self.assertNotIn("judge", service.model_policy)
        self.assertEqual(service._llm_fallback_model(), service._codex_model_for_route("repair"))
        routed_mode, routed_budget, route = service._resolve_llm_budget("cheap", "where is createIssue", [])
        self.assertEqual(routed_mode, "cheap")
        self.assertEqual(routed_budget["model"], service._codex_model_for_route("cheap"))
        self.assertEqual(route["mode"], "manual")

    def test_codex_cli_bridge_provider_returns_answer(self):
        provider = CodexCliBridgeSourceCodeQALLMProvider(
            workspace_root=Path(self.temp_dir.name),
            timeout_seconds=20,
            codex_binary="codex",
        )
        calls = []

        def fake_run(command, **kwargs):
            calls.append((command, kwargs))
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            Path(output_path).write_text(
                '{"direct_answer":"Codex found the answer.","claims":[{"text":"Codex found the answer","citations":["S1"]}],"missing_evidence":[],"confidence":"high"}',
                encoding="utf-8",
            )
            return SimpleNamespace(returncode=0, stdout='{"type":"done"}\n', stderr="")

        payload = {
            "systemInstruction": {"parts": [{"text": "system"}]},
            "contents": [{"parts": [{"text": "Question: where is createIssue"}]}],
        }
        image_path = Path(self.temp_dir.name) / "screenshot.png"
        image_path.write_bytes(b"fake-png")
        payload["_codex_image_paths"] = [str(image_path)]
        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ):
            result = provider.generate(payload=payload, primary_model="codex-cli", fallback_model="codex-cli")

        self.assertIn("Codex found the answer", provider.extract_text(result.payload))
        exec_command = calls[-1][0]
        self.assertIn("--sandbox", exec_command)
        self.assertIn("read-only", exec_command)
        self.assertIn("--ephemeral", exec_command)
        self.assertIn("--json", exec_command)
        self.assertIn("--skip-git-repo-check", exec_command)
        self.assertIn("--image", exec_command)
        self.assertIn(str(image_path), exec_command)
        self.assertEqual(result.model, "codex-cli")
        self.assertEqual(result.attempt_log[0]["exit_code"], 0)
        self.assertEqual(result.attempt_log[0]["workspace_root"], self.temp_dir.name)
        self.assertEqual(result.attempt_log[0]["concurrency_limit"], 1)
        self.assertIn("queue_wait_ms", result.attempt_log[0])
        self.assertIn("--sandbox", result.attempt_log[0]["command"])

    def test_codex_cli_bridge_provider_passes_selected_model_to_cli(self):
        provider = CodexCliBridgeSourceCodeQALLMProvider(
            workspace_root=Path(self.temp_dir.name),
            timeout_seconds=20,
            codex_binary="codex",
        )
        calls = []

        def fake_run(command, **kwargs):
            calls.append((command, kwargs))
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            Path(output_path).write_text(
                '{"direct_answer":"ok","claims":[],"missing_evidence":[],"confidence":"high"}',
                encoding="utf-8",
            )
            return SimpleNamespace(returncode=0, stdout='{"type":"done"}\n', stderr="")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ):
            result = provider.generate(
                payload={
                    "contents": [{"parts": [{"text": "hi"}]}],
                    "_codex_reasoning_effort": "low",
                },
                primary_model="gpt-5.4-mini",
                fallback_model="gpt-5.5",
            )

        exec_command = calls[-1][0]
        self.assertIn("--model", exec_command)
        self.assertIn("gpt-5.4-mini", exec_command)
        self.assertIn("-c", exec_command)
        self.assertIn('model_reasoning_effort="low"', exec_command)
        self.assertEqual(result.model, "gpt-5.4-mini")

    def test_codex_cli_bridge_provider_rejects_successful_bad_request_payload(self):
        provider = CodexCliBridgeSourceCodeQALLMProvider(
            workspace_root=Path(self.temp_dir.name),
            timeout_seconds=20,
            codex_binary="codex",
        )

        def fake_run(command, **kwargs):
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            Path(output_path).write_text('{"detail":"Bad Request"}', encoding="utf-8")
            return SimpleNamespace(returncode=0, stdout='{"type":"done"}\n', stderr="")

        payload = {
            "contents": [{"parts": [{"text": "hi"}]}],
            "_codex_trace_id": "trace-bad",
            "_codex_phase": "initial",
            "_codex_estimated_prompt_tokens": 123,
            "_codex_candidate_path_count": 7,
            "_codex_candidate_repo_count": 2,
        }

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ), patch("bpmis_jira_tool.source_code_qa_llm_providers.LOGGER.warning") as warning_log:
            with self.assertRaisesRegex(ToolError, "Codex CLI returned API error: Bad Request"):
                provider.generate(payload=payload, primary_model="codex-cli", fallback_model="codex-cli")

        failure_logs = [
            call.args
            for call in warning_log.call_args_list
            if call.args and call.args[0] == "source_code_qa_codex_failure %s"
        ]
        self.assertTrue(failure_logs)
        failure_payload = json.loads(failure_logs[0][1])
        self.assertEqual(failure_payload["reason"], "api_error_payload")
        self.assertEqual(failure_payload["error"], "Bad Request")
        self.assertEqual(failure_payload["exit_code"], 0)
        self.assertEqual(failure_payload["trace_id"], "trace-bad")
        self.assertEqual(failure_payload["phase"], "initial")
        self.assertEqual(failure_payload["estimated_prompt_tokens"], 123)
        self.assertEqual(failure_payload["candidate_path_count"], 7)
        self.assertEqual(failure_payload["candidate_repo_count"], 2)
        self.assertEqual(failure_payload["command_mode"], "ephemeral")
        self.assertIn("Bad Request", failure_payload["answer_tail"])
        self.assertIn("done", failure_payload["stdout_tail"])

    def test_codex_cli_bridge_adds_rg_directory_to_exec_path(self):
        provider = CodexCliBridgeSourceCodeQALLMProvider(
            workspace_root=Path(self.temp_dir.name),
            timeout_seconds=20,
            codex_binary="codex",
        )
        exec_envs = []

        def fake_run(command, **kwargs):
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            exec_envs.append(kwargs.get("env") or {})
            output_path = command[command.index("--output-last-message") + 1]
            Path(output_path).write_text(
                '{"direct_answer":"ok","claims":[],"missing_evidence":[],"confidence":"high"}',
                encoding="utf-8",
            )
            return SimpleNamespace(returncode=0, stdout='{"type":"done"}\n', stderr="")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/opt/tools/rg"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ):
            provider.generate(payload={"contents": [{"parts": [{"text": "hi"}]}]}, primary_model="codex-cli", fallback_model="codex-cli")

        self.assertTrue(exec_envs)
        self.assertTrue(str(exec_envs[-1].get("PATH") or "").startswith("/opt/tools"))

    def test_codex_cli_bridge_resume_mode_uses_resume_command(self):
        provider = CodexCliBridgeSourceCodeQALLMProvider(
            workspace_root=Path(self.temp_dir.name),
            timeout_seconds=20,
            codex_binary="codex",
            session_mode="resume",
        )
        calls = []

        def fake_run(command, **kwargs):
            calls.append((command, kwargs))
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            Path(output_path).write_text(
                '{"direct_answer":"ok","claims":[],"missing_evidence":[],"confidence":"high"}',
                encoding="utf-8",
            )
            return SimpleNamespace(returncode=0, stdout='{"session_id":"session-123","type":"done"}\n', stderr="")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ):
            result = provider.generate(
                payload={"contents": [{"parts": [{"text": "hi"}]}], "codex_cli_session_id": "session-123"},
                primary_model="codex-cli",
                fallback_model="codex-cli",
            )

        exec_command = calls[-1][0]
        self.assertIn("resume", exec_command)
        self.assertIn("session-123", exec_command)
        self.assertNotIn("--ephemeral", exec_command)
        self.assertEqual(result.attempt_log[0]["session_mode"], "resume")
        self.assertEqual(result.payload["codex_cli_trace"]["session_id"], "session-123")

    def test_codex_cli_bridge_streams_json_progress_events(self):
        provider = CodexCliBridgeSourceCodeQALLMProvider(
            workspace_root=Path(self.temp_dir.name),
            timeout_seconds=20,
            codex_binary="codex",
        )
        streamed = []

        class FakePipe:
            def __init__(self, lines):
                self.lines = list(lines)
                self.exhausted = False

            def readline(self):
                if self.lines:
                    return self.lines.pop(0)
                self.exhausted = True
                return ""

            def close(self):
                self.exhausted = True

        class FakeStdin:
            def write(self, _value):
                return None

            def close(self):
                return None

        class FakeProcess:
            def __init__(self, *_args, **_kwargs):
                self.stdout = FakePipe(['{"message":"Reading src/App.java"}\n', '{"type":"done"}\n'])
                self.stderr = FakePipe([])
                self.stdin = FakeStdin()
                self.returncode = 0

            def poll(self):
                if self.stdout.exhausted and self.stderr.exhausted:
                    return self.returncode
                return None

            def kill(self):
                self.returncode = -9
                self.stdout.exhausted = True
                self.stderr.exhausted = True

        def fake_run(command, **_kwargs):
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            raise AssertionError("streaming path should use Popen for codex exec")

        def progress(stage, message, _current, _total):
            streamed.append((stage, message))

        with tempfile.NamedTemporaryFile("w", encoding="utf-8", delete=False) as output_file:
            output_file.write('{"direct_answer":"Final answer","claims":[],"missing_evidence":[],"confidence":"high"}')
            output_path = output_file.name

        def fake_named_temp(*_args, **_kwargs):
            return open(output_path, "r+", encoding="utf-8")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ), patch("bpmis_jira_tool.source_code_qa_llm_providers.subprocess.Popen", side_effect=FakeProcess), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.tempfile.NamedTemporaryFile",
            side_effect=fake_named_temp,
        ):
            try:
                result = provider.generate(
                    payload={"contents": [{"parts": [{"text": "hi"}]}], "_progress_callback": progress},
                    primary_model="codex-cli",
                    fallback_model="codex-cli",
                )
            finally:
                Path(output_path).unlink(missing_ok=True)

        self.assertIn(("codex_stream", "Reading src/App.java"), streamed)
        self.assertIn("Final answer", provider.extract_text(result.payload))

    def test_codex_cli_provider_uses_configured_model(self):
        with patch.dict(os.environ, {"SOURCE_CODE_QA_CODEX_MODEL": "gpt-5.5"}, clear=True):
            service = SourceCodeQAService(
                data_root=Path(self.temp_dir.name),
                team_profiles=TEAM_PROFILE_DEFAULTS,
                llm_provider="codex_cli_bridge",
                gitlab_token="secret-token",
                git_timeout_seconds=5,
                max_file_bytes=200_000,
            )

        self.assertEqual(service.llm_budgets["cheap"]["model"], "gpt-5.5")
        self.assertEqual(service.llm_budgets["balanced"]["model"], "gpt-5.5")
        self.assertEqual(service.llm_budgets["deep"]["model"], "gpt-5.5")
        self.assertEqual(service._llm_fallback_model(), "gpt-5.5")

    def test_codex_provider_uses_route_specific_models(self):
        with patch.dict(
            os.environ,
            {
                "SOURCE_CODE_QA_CODEX_MODEL": "legacy-all-model",
                "SOURCE_CODE_QA_CODEX_MODEL_CHEAP": "cheap-model",
                "SOURCE_CODE_QA_CODEX_MODEL_BALANCED": "balanced-model",
                "SOURCE_CODE_QA_CODEX_MODEL_DEEP": "deep-model",
                "SOURCE_CODE_QA_CODEX_MODEL_COMPACT_DEEP": "compact-model",
                "SOURCE_CODE_QA_CODEX_MODEL_REPAIR": "repair-model",
            },
            clear=True,
        ):
            service = SourceCodeQAService(
                data_root=Path(self.temp_dir.name),
                team_profiles=TEAM_PROFILE_DEFAULTS,
                llm_provider="codex_cli_bridge",
                gitlab_token="secret-token",
                git_timeout_seconds=5,
                max_file_bytes=200_000,
            )

        self.assertEqual(service.llm_budgets["cheap"]["model"], "cheap-model")
        self.assertEqual(service.llm_budgets["balanced"]["model"], "balanced-model")
        self.assertEqual(service.llm_budgets["deep"]["model"], "deep-model")
        self.assertEqual(service.llm_budgets["compact_deep"]["model"], "compact-model")
        self.assertEqual(service.model_policy["repair"]["model"], "repair-model")

    def test_short_definition_status_question_routes_to_cheap(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )

        routed_mode, routed_budget, route = service._resolve_llm_budget(
            "auto",
            "Rule outcome = pass 是什么意思?",
            [{"path": "rules.xml", "retrieval": "persistent_index"} for _ in range(8)],
        )

        self.assertEqual(routed_mode, "cheap")
        self.assertEqual(routed_budget["model"], service.llm_budgets["cheap"]["model"])
        self.assertEqual(route["reason"], "short_definition_or_status")

    def test_cheap_status_question_skips_soft_codex_repair(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )

        decision = service._codex_repair_decision(
            question="Rule outcome = pass 是什么意思?",
            answer="Not found in the repository.",
            structured_answer={"confidence": "low"},
            quality_gate={"status": "sufficient", "confidence": "high"},
            evidence_pack={"items": [{"kind": "config", "target": "Rule outcome = pass"}]},
            answer_judge={"status": "repair", "issues": ["answer missed retrieval hints"]},
            codex_validation={"status": "needs_citation", "unsupported_claims": ["Rule outcome pass was not found"]},
            finish_reason="stop",
            effort_assessment=False,
            trace_id="trace-cheap-repair",
            selected_model=service.llm_budgets["cheap"]["model"],
            query_mode="deep",
            routed_budget_mode="cheap",
        )

        self.assertFalse(service._codex_high_risk_question("Rule outcome = pass 是什么意思?"))
        self.assertFalse(decision["repair_will_run"])
        self.assertEqual(decision["repair_policy"], "cheap_simple_first_pass")
        self.assertIn("not_found_answer_conflicts_with_retrieval_hints", decision["suppressed_repair_reasons"])
        self.assertTrue(decision["repair_skipped_reason"].startswith("codex_repair_tier_skipped:cheap_simple_first_pass"))
        route_fields = service._codex_repair_route_fields(
            codex_validation={"status": "needs_citation", "scoped_file_refs": [], "out_of_scope_refs": []},
            repair_attempted=False,
            repair_policy=decision["repair_policy"],
            repair_reason="",
            repair_skipped_reason=decision["repair_skipped_reason"],
            repair_decision_ms=1,
            deep_investigation_rounds=0,
            deep_investigation_terms=[],
            deep_investigation_added=0,
        )
        self.assertEqual(route_fields["codex_repair_policy"], "cheap_simple_first_pass")

    def test_cheap_budget_still_repairs_hard_codex_failures(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )

        decision = service._codex_repair_decision(
            question="where is batchCreateJiraIssue",
            answer="",
            structured_answer={},
            quality_gate={"status": "sufficient", "confidence": "high"},
            evidence_pack={},
            answer_judge={"status": "ok", "issues": []},
            codex_validation={"status": "skipped"},
            finish_reason="stop",
            effort_assessment=False,
            trace_id="trace-cheap-hard-repair",
            selected_model=service.llm_budgets["cheap"]["model"],
            query_mode="deep",
            routed_budget_mode="cheap",
        )

        self.assertTrue(decision["repair_will_run"])
        self.assertEqual(decision["severe_repair_reasons"], ["empty_codex_answer"])
        self.assertEqual(decision["suppressed_repair_reasons"], [])

    def test_codex_repair_skips_when_answer_generation_budget_is_low(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        service._codex_answer_timeout_seconds = 120

        with patch("bpmis_jira_tool.source_code_qa.time.time", return_value=1040):
            remaining_seconds, skip_reason = service._codex_repair_remaining_timeout_seconds(1000)

        self.assertEqual(remaining_seconds, 75)
        self.assertEqual(skip_reason, "codex_repair_insufficient_query_budget:75<90")

    def test_codex_repair_reserves_deep_investigation_budget(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        service._codex_answer_timeout_seconds = 220

        with patch("bpmis_jira_tool.source_code_qa.time.time", return_value=1100):
            remaining_seconds, skip_reason = service._codex_repair_remaining_timeout_seconds(1000, reserve_seconds=40)

        self.assertEqual(remaining_seconds, 75)
        self.assertEqual(skip_reason, "codex_repair_insufficient_query_budget:75<90")

    def test_sufficient_bounded_negative_answer_skips_soft_codex_repair(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )

        decision = service._codex_repair_decision(
            question="why is SG notification template mapping missing from source evidence?",
            answer=(
                "I cannot confirm those SG template IDs from the scoped source repository. "
                "The code references notification rule wiring, but the concrete template registry appears to be runtime or production DB evidence."
            ),
            structured_answer={
                "direct_answer": "I cannot confirm those SG template IDs from scoped source code.",
                "claims": [
                    {"text": "Rule notification code is present", "citations": ["S1"]},
                    {"text": "Template ID mapping requires production DB evidence", "citations": []},
                ],
                "missing_evidence": ["Production msg-center template registry for SG"],
                "confidence": "medium",
            },
            quality_gate={"status": "sufficient", "confidence": "medium"},
            evidence_pack={"items": [{"kind": "config", "target": "template ID"}]},
            answer_judge={"status": "repair", "issues": ["unsupported runtime ID mapping"]},
            codex_validation={"status": "needs_citation", "unsupported_claims": ["runtime template mapping"]},
            finish_reason="stop",
            effort_assessment=False,
            trace_id="trace-soft-skip",
            selected_model=service.llm_budgets["deep"]["model"],
            query_mode="deep",
            routed_budget_mode="deep",
        )

        self.assertFalse(decision["repair_will_run"])
        self.assertEqual(decision["repair_policy"], "first_pass_quality_gate")
        self.assertEqual(decision["repair_skipped_reason"], "codex_repair_skipped:first_pass_sufficient_for_deep_gap")

    def test_catalog_questions_do_not_skip_soft_codex_repair_on_generic_negative_answer(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )

        decision = service._codex_repair_decision(
            question="What are the available functions I can use for my rules?",
            answer=(
                "This information is not present in the provided Source Code or Runtime Evidence. "
                "Please verify if the feature is implemented or escalate to the engineering lead."
            ),
            structured_answer={"direct_answer": "This information is not present.", "claims": [], "confidence": "medium"},
            quality_gate={"status": "sufficient", "confidence": "medium"},
            evidence_pack={"items": [{"kind": "code", "target": "function catalog"}]},
            answer_judge={"status": "repair", "issues": ["generic negative answer"]},
            codex_validation={"status": "needs_citation", "unsupported_claims": ["no function catalog"]},
            finish_reason="stop",
            effort_assessment=False,
            trace_id="trace-catalog-repair",
            selected_model=service.llm_budgets["balanced"]["model"],
            query_mode="deep",
            routed_budget_mode="balanced",
        )

        self.assertTrue(decision["repair_will_run"])
        self.assertNotEqual(decision["repair_policy"], "first_pass_quality_gate")

    def test_ambiguous_business_flow_still_allows_soft_codex_repair(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )

        decision = service._codex_repair_decision(
            question="why can a user authorize action plan before authorize new incident?",
            answer=(
                "Yes, but only if the action plan task is already created. "
                "The exact state ordering needs source confirmation across incident authorization."
            ),
            structured_answer={
                "direct_answer": "Yes, but only if the action plan task is already created.",
                "claims": [{"text": "Action plan authorization has state checks", "citations": ["S1"]}],
                "missing_evidence": ["Incident authorization ordering guard"],
                "confidence": "medium",
            },
            quality_gate={"status": "sufficient", "confidence": "medium"},
            evidence_pack={"items": [{"kind": "code", "target": "authorize"}]},
            answer_judge={"status": "repair", "issues": ["business state ordering ambiguity"]},
            codex_validation={"status": "needs_citation", "unsupported_claims": ["state ordering"]},
            finish_reason="stop",
            effort_assessment=False,
            trace_id="trace-business-repair",
            selected_model=service.llm_budgets["deep"]["model"],
            query_mode="deep",
            routed_budget_mode="deep",
        )

        self.assertTrue(decision["repair_will_run"])
        self.assertIn("high_risk_claims_missing_scoped_file_evidence", decision["severe_repair_reasons"])

    def test_nested_json_direct_answer_is_normalized_before_final_output(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        inner = json.dumps(
            {
                "direct_answer": "Authorization email links use the ticket id when the template receives one.",
                "claims": [{"text": "Ticket id is passed into the email template", "citations": ["S1"]}],
                "missing_evidence": [],
                "confidence": "high",
            }
        )

        answer, structured = service._normalize_codex_answer_text(
            json.dumps({"direct_answer": inner, "claims": [], "confidence": "medium"}),
            {"direct_answer": inner, "claims": [], "confidence": "medium", "format": "json"},
        )

        self.assertEqual(answer, "Authorization email links use the ticket id when the template receives one.")
        self.assertEqual(structured["confidence"], "high")
        self.assertEqual(structured["claims"][0]["citations"], ["S1"])

    def test_root_cause_definition_like_question_does_not_route_to_cheap(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )

        routed_mode, _routed_budget, route = service._resolve_llm_budget(
            "auto",
            "why is rule outcome pass in the data source chain?",
            [],
        )

        self.assertEqual(routed_mode, "deep")
        self.assertIn("root_cause_or_error", route["reason"])

    def test_api_config_rule_question_routes_to_balanced(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )

        routed_mode, routed_budget, route = service._resolve_llm_budget(
            "auto",
            "how does API config select rule logic?",
            [{"path": "RuleService.java", "retrieval": "persistent_index"}],
        )

        self.assertEqual(routed_mode, "balanced")
        self.assertEqual(routed_budget["model"], service.llm_budgets["balanced"]["model"])
        self.assertEqual(route["reason"], "api_config_rule_logic")

    def test_codex_cli_provider_uses_codex_specific_timeout(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            llm_provider="codex_cli_bridge",
            llm_timeout_seconds=90,
            codex_timeout_seconds=260,
            codex_concurrency=2,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )

        self.assertEqual(service.llm_provider.timeout_seconds, 260)
        self.assertEqual(service.llm_provider.concurrency_limit, 2)

    def test_codex_investigation_brief_uses_paths_not_long_snippet_context(self):
        entry = RepositoryEntry("Portal Repo", "https://git.example.com/team/portal.git")
        key = "AF:All"
        repo_path = self.service._repo_path(key, entry)
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "repository" / "IssueRepository.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text("class IssueRepository {\n  void find() {}\n}\n", encoding="utf-8")
        matches = [
            {
                "repo": "Portal Repo",
                "path": "repository/IssueRepository.java",
                "line_start": 1,
                "line_end": 3,
                "retrieval": "persistent_index",
                "trace_stage": "direct",
                "reason": "IssueRepository matched query",
                "snippet": "x" * 5000,
            }
        ]

        candidate_paths = self.service._codex_candidate_paths(entries=[entry], key=key, matches=matches)
        scope_roots = self.service._codex_scope_roots(entries=[entry], key=key)
        brief = self.service._codex_investigation_brief(
            pm_team="AF",
            country="All",
            question="where is IssueRepository",
            candidate_paths=candidate_paths,
            evidence_pack={"entry_points": ["IssueRepository [S1]"]},
            quality_gate={"status": "sufficient", "confidence": "high", "missing": []},
            scope_roots=scope_roots,
            followup_context={
                "question": "previous question",
                "answer": "previous answer",
                "codex_candidate_paths": [
                    {"repo": "Portal Repo", "repo_root": str(repo_path), "path": "service/PriorService.java", "line_start": 5, "line_end": 9}
                ],
                "codex_citation_validation": {
                    "status": "ok",
                    "cited_path_count": 1,
                    "direct_file_refs": [{"path": "service/PriorService.java", "line_start": 5, "line_end": 9}],
                },
                "codex_cli_summary": {"prompt_mode": "codex_investigation_brief_v1", "candidate_path_count": 1, "repair_attempted": False},
                "matches_snapshot": [{"repo": "Portal Repo", "path": "repository/IssueRepository.java", "line_start": 1, "line_end": 3}],
                "recent_turns": [
                    {
                        "question": "earlier session question",
                        "answer": "earlier session answer",
                        "trace_id": "trace-earlier",
                        "matches_snapshot": [{"repo": "Portal Repo", "path": "service/EarlierService.java", "line_start": 2, "line_end": 4}],
                        "codex_candidate_paths": [{"repo": "Portal Repo", "repo_root": str(repo_path), "path": "service/EarlierService.java"}],
                    }
                ],
            },
        )

        self.assertIn("Prompt mode: codex_investigation_brief_v5", brief)
        self.assertIn("You are the AF PM Team's Source Code & Runtime Evidence Assistant.", brief)
        self.assertIn("central point of truth for developers, QA engineers, and Product Managers", brief)
        self.assertIn("implemented business logic, data flow, and workflows", brief)
        self.assertIn("Code & log analysis", brief)
        self.assertIn("AF domain mapping", brief)
        self.assertIn("Always begin with a direct answer", brief)
        self.assertIn("do not begin the direct answer with country database routing", brief)
        self.assertIn("This information is not present in the provided Source Code or Runtime Evidence", brief)
        self.assertIn("Do not rewrite code, suggest architectural refactoring", brief)
        self.assertIn("Three-stage investigation required:", brief)
        self.assertIn("Stage 2 gap verification", brief)
        self.assertIn("search the uploaded evidence text/summary by app/env/namespace", brief)
        self.assertIn("AMR/FV/authentication flows should inspect authentication-center Apollo keys", brief)
        self.assertIn("uploaded runtime config key/value", brief)
        self.assertIn("screenshot-driven incident questions", brief)
        self.assertIn("missing_production_evidence", brief)
        self.assertIn("source_code_evidence must name concrete files/functions/classes/fields/tables or APIs", brief)
        self.assertIn("lead with the data model and SQL strategy", brief)
        self.assertIn("full rule/config definitions from status-only migration updates", brief)
        self.assertIn("Strict scope boundary: search only the allowed scope roots", brief)
        self.assertIn("Scoped Codex search allowlist:", brief)
        self.assertIn("Starting path hints from local retrieval:", brief)
        self.assertIn("Candidate paths are starting hints from local retrieval", brief)
        self.assertIn("current_high_confidence_paths", brief)
        self.assertIn(str(repo_path), brief)
        self.assertIn("relative_root=AF-All", brief)
        self.assertIn("path=repository/IssueRepository.java", brief)
        self.assertIn("file_exists=True", brief)
        self.assertIn("path_status=exact", brief)
        self.assertIn("Follow-up context:", brief)
        self.assertIn("previous question", brief)
        self.assertIn("Previous Codex candidate paths:", brief)
        self.assertIn("service/PriorService.java", brief)
        self.assertIn("Previous citation validation: status=ok", brief)
        self.assertIn("Previous Codex run: prompt_mode=codex_investigation_brief_v1", brief)
        self.assertIn("Earlier session turns:", brief)
        self.assertIn("earlier session question", brief)
        self.assertIn("service/EarlierService.java", brief)
        self.assertNotIn("x" * 100, brief)

    def test_codex_sql_generation_brief_is_compact_but_requires_code_checks(self):
        entry = RepositoryEntry("GRC Portal", "https://git.example.com/team/grc-portal.git")
        key = "GRC:All"
        repo_path = self.service._repo_path(key, entry)
        (repo_path / ".git").mkdir(parents=True)
        mapper_file = repo_path / "mapper" / "TicketMapper.xml"
        mapper_file.parent.mkdir(parents=True)
        mapper_file.write_text(
            "<select id=\"findTicket\">select * from ticket_info where authorization_type = #{authorizationType}</select>",
            encoding="utf-8",
        )
        matches = [
            {
                "repo": "GRC Portal",
                "path": "mapper/TicketMapper.xml",
                "line_start": 1,
                "line_end": 1,
                "retrieval": "persistent_index",
                "trace_stage": "direct",
                "reason": "ticket_info SQL mapper matched",
            }
        ]
        candidate_paths = self.service._codex_candidate_paths(entries=[entry], key=key, matches=matches)
        scope_roots = self.service._codex_scope_roots(entries=[entry], key=key)
        huge_dictionary = (
            "[Data dictionary sheet: filler]\n"
            + ("filler_field,unused meaning\n" * 2500)
            + "[Data dictionary sheet: reviewer]\n"
            + "ticket_info,event_id,Incident id\n"
            + "ticket_info,authorization_type,Ticket authorization type\n"
            + "ticket_node_info,target_group,Reviewer group such as CRO/ORM/CISO/Compliance\n"
        )

        brief = self.service._codex_sql_generation_brief(
            pm_team="GRC",
            country="SG",
            question="generate a SQL query for incident reviewer status",
            candidate_paths=candidate_paths,
            evidence_pack={"tables": ["ticket_info [S1]", "ticket_node_info [S1]"], "items": []},
            quality_gate={"status": "sufficient", "confidence": "high", "missing": []},
            followup_context={},
            runtime_evidence=[
                {
                    "filename": "grc-dictionary.xlsx",
                    "source_type": "data_dictionary",
                    "pm_team": "GRC",
                    "country": "All",
                    "kind": "document",
                    "sha256": "a" * 64,
                    "text": huge_dictionary,
                }
            ],
            scope_roots=scope_roots,
        )
        stats = self.service._codex_prompt_stats(brief)

        self.assertIn("Prompt mode: codex_sql_generation_brief_v1", brief)
        self.assertIn("Still inspect source code before answering", brief)
        self.assertIn("mapper/XML/DAO/repository/source SQL", brief)
        self.assertIn("data_dictionary uploads apply to SG, ID, PH, and All", brief)
        self.assertIn("separate DB instance", brief)
        self.assertIn("RC and Compliance are business aliases", brief)
        self.assertIn("ticket_info", brief)
        self.assertIn("ticket_node_info", brief)
        self.assertIn("path=mapper/TicketMapper.xml", brief)
        self.assertNotIn("Three-stage investigation required:", brief)
        self.assertLess(stats["estimated_prompt_tokens"], 15000)

    def test_codex_prompt_logging_marks_role_prompt_present(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            llm_provider="codex_cli_bridge",
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        entry = RepositoryEntry("Portal Repo", "https://git.example.com/team/portal.git")
        key = "AF:All"
        repo_path = service._repo_path(key, entry)
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "repository" / "IssueRepository.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text("class IssueRepository {\n  void find(){ jdbc.query(\"select * from issue_table\"); }\n}\n", encoding="utf-8")
        matches = [
            {
                "repo": "Portal Repo",
                "path": "repository/IssueRepository.java",
                "line_start": 1,
                "line_end": 3,
                "retrieval": "persistent_index",
                "trace_stage": "direct",
                "reason": "repository and table matched",
                "score": 10,
                "snippet": "class IssueRepository { void find(){ jdbc.query(\"select * from issue_table\"); } }",
            }
        ]
        timing_logs = []

        def fake_run(command, **kwargs):
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            Path(output_path).write_text(
                '{"direct_answer":"IssueRepository reads issue_table.",'
                '"claims":[{"text":"IssueRepository reads issue_table","citations":["S1"]}],'
                '"missing_evidence":[],"confidence":"high"}',
                encoding="utf-8",
            )
            return SimpleNamespace(returncode=0, stdout='{"type":"done"}\n', stderr="")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ), patch("bpmis_jira_tool.source_code_qa._log_source_code_qa_timing", side_effect=lambda component, **fields: timing_logs.append((component, fields))):
            service._build_llm_answer(
                entries=[entry],
                key=key,
                pm_team="AF",
                country="All",
                question="what table does IssueRepository read",
                matches=matches,
                llm_budget_mode="auto",
                followup_context={},
                requested_answer_mode="auto",
                trace_id="trace-prompt",
            )

        prompt_logs = [fields for component, fields in timing_logs if component == "codex_prompt"]
        self.assertTrue(prompt_logs)
        self.assertTrue(prompt_logs[0]["role_prompt_present"])
        self.assertEqual(prompt_logs[0]["prompt_mode"], "codex_investigation_brief_v5")
        self.assertEqual(prompt_logs[0]["pm_team"], "AF")
        self.assertEqual(prompt_logs[0]["country"], "All")
        self.assertRegex(prompt_logs[0]["prompt_sha256"], r"^[0-9a-f]{16}$")
        self.assertEqual(prompt_logs[0]["candidate_path_count"], 1)
        self.assertEqual(prompt_logs[0]["candidate_repo_count"], 1)
        self.assertEqual(prompt_logs[0]["scope_repo_count"], 1)
        self.assertEqual(prompt_logs[0]["retrieval_role"], "hints")
        self.assertEqual(prompt_logs[0]["repair_policy"], "severe_only")
        self.assertEqual(prompt_logs[0]["requested_budget"], "auto")
        self.assertIn(prompt_logs[0]["routed_budget"], {"cheap", "balanced", "deep", "compact_deep"})
        self.assertIn(prompt_logs[0]["reasoning_effort"], {"low", "medium", "high", "xhigh"})
        self.assertGreater(prompt_logs[0]["prompt_chars"], 0)
        self.assertGreater(prompt_logs[0]["prompt_bytes"], 0)
        self.assertGreater(prompt_logs[0]["estimated_prompt_tokens"], 0)
        components = [component for component, _fields in timing_logs]
        self.assertIn("codex_repair_prepare", components)
        repair_prepare_logs = [fields for component, fields in timing_logs if component == "codex_repair_prepare"]
        self.assertTrue(repair_prepare_logs)
        self.assertIn("repair_will_run", repair_prepare_logs[0])
        self.assertIn("deep_investigation_needed", repair_prepare_logs[0])
        self.assertIn("repair_issue_count", repair_prepare_logs[0])

    def test_sql_generation_question_uses_codex_sql_prompt_mode(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            llm_provider="codex_cli_bridge",
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        entry = RepositoryEntry("GRC Portal", "https://git.example.com/team/grc-portal.git")
        key = "GRC:All"
        repo_path = service._repo_path(key, entry)
        (repo_path / ".git").mkdir(parents=True)
        mapper_file = repo_path / "mapper" / "TicketMapper.xml"
        mapper_file.parent.mkdir(parents=True)
        mapper_file.write_text(
            "<select id=\"findTicket\">select * from ticket_info where authorization_type = #{authorizationType}</select>",
            encoding="utf-8",
        )
        matches = [
            {
                "repo": "GRC Portal",
                "path": "mapper/TicketMapper.xml",
                "line_start": 1,
                "line_end": 1,
                "retrieval": "persistent_index",
                "trace_stage": "direct",
                "reason": "ticket_info mapper matched",
                "score": 10,
                "snippet": "select * from ticket_info",
            }
        ]
        timing_logs = []

        def fake_run(command, **kwargs):
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            Path(output_path).write_text(
                '{"direct_answer":"Use latest ticket_info rows and ticket_node_info reviewer rows.",'
                '"sql":"select * from ticket_info;",'
                '"source_code_evidence":["mapper/TicketMapper.xml uses ticket_info"],'
                '"claims":[{"text":"TicketMapper uses ticket_info","citations":["S1"]}],'
                '"missing_evidence":[],"confidence":"high"}',
                encoding="utf-8",
            )
            return SimpleNamespace(returncode=0, stdout='{"type":"done"}\n', stderr="")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ), patch("bpmis_jira_tool.source_code_qa._log_source_code_qa_timing", side_effect=lambda component, **fields: timing_logs.append((component, fields))):
            payload = service._build_llm_answer(
                entries=[entry],
                key=key,
                pm_team="GRC",
                country="SG",
                question="generate a SQL query for incident reviewer status",
                matches=matches,
                llm_budget_mode="auto",
                followup_context={},
                requested_answer_mode="auto",
                trace_id="trace-sql-prompt",
                runtime_evidence=[
                    {
                        "filename": "grc-dictionary.xlsx",
                        "source_type": "data_dictionary",
                        "pm_team": "GRC",
                        "country": "All",
                        "kind": "document",
                        "sha256": "a" * 64,
                        "text": "ticket_info,event_id,Incident id\n" + ("x" * 40000),
                    }
                ],
            )

        prompt_logs = [fields for component, fields in timing_logs if component == "codex_prompt" and fields.get("phase") == "initial"]
        self.assertTrue(prompt_logs)
        self.assertEqual(prompt_logs[0]["prompt_mode"], "codex_sql_generation_brief_v1")
        self.assertLess(prompt_logs[0]["estimated_prompt_tokens"], 15000)
        self.assertEqual(payload["llm_route"]["prompt_mode"], "codex_sql_generation_brief_v1")

    def test_codex_candidate_paths_resolve_stale_path_by_filename(self):
        entry = RepositoryEntry("Portal Repo", "https://git.example.com/team/portal.git")
        key = "AF:All"
        repo_path = self.service._repo_path(key, entry)
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "src" / "main" / "java" / "IssueRepository.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text("class IssueRepository {}\n", encoding="utf-8")
        self.service._build_repo_index(key, entry, repo_path)

        candidate_paths = self.service._codex_candidate_paths(
            entries=[entry],
            key=key,
            matches=[
                {
                    "repo": "Portal Repo",
                    "path": "repository/IssueRepository.java",
                    "line_start": 1,
                    "line_end": 1,
                }
            ],
        )

        self.assertEqual(candidate_paths[0]["path"], "src/main/java/IssueRepository.java")
        self.assertEqual(candidate_paths[0]["original_path"], "repository/IssueRepository.java")
        self.assertTrue(candidate_paths[0]["repo_relative_root"].startswith("AF-All/"))
        self.assertTrue(candidate_paths[0]["file_exists"])
        self.assertEqual(candidate_paths[0]["path_status"], "resolved_by_filename")

    def test_codex_followup_candidate_paths_are_carried_forward(self):
        entry = RepositoryEntry("Portal Repo", "https://git.example.com/team/portal.git")
        key = "AF:All"
        repo_path = self.service._repo_path(key, entry)
        current_file = repo_path / "controller" / "IssueController.java"
        prior_file = repo_path / "repository" / "IssueRepository.java"
        earlier_file = repo_path / "service" / "EarlierService.java"
        current_file.parent.mkdir(parents=True)
        prior_file.parent.mkdir(parents=True)
        earlier_file.parent.mkdir(parents=True)
        current_file.write_text("class IssueController {}\n", encoding="utf-8")
        prior_file.write_text("class IssueRepository {}\n", encoding="utf-8")
        earlier_file.write_text("class EarlierService {}\n", encoding="utf-8")
        current = self.service._codex_candidate_paths(
            entries=[entry],
            key=key,
            matches=[{"repo": "Portal Repo", "path": "controller/IssueController.java", "line_start": 1, "line_end": 1}],
        )

        merged = self.service._merge_codex_followup_candidate_paths(
            current,
            {
                "codex_candidate_paths": [
                    {
                        "repo": "Portal Repo",
                        "repo_root": str(repo_path),
                        "path": "repository/IssueRepository.java",
                        "line_start": 1,
                        "line_end": 1,
                    }
                ],
                "recent_turns": [
                    {
                        "codex_candidate_paths": [
                            {
                                "repo": "Portal Repo",
                                "repo_root": str(repo_path),
                                "path": "service/EarlierService.java",
                                "line_start": 1,
                                "line_end": 1,
                            }
                        ]
                    }
                ],
            },
        )

        self.assertEqual(
            [item["path"] for item in merged],
            ["controller/IssueController.java", "repository/IssueRepository.java", "service/EarlierService.java"],
        )
        self.assertEqual(merged[1]["retrieval"], "previous_codex_context")

    def test_codex_citation_validation_accepts_direct_file_reference(self):
        entry = RepositoryEntry("Portal Repo", "https://git.example.com/team/portal.git")
        key = "AF:All"
        repo_path = self.service._repo_path(key, entry)
        source_file = repo_path / "repository" / "IssueRepository.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text("line1\nline2\nline3\n", encoding="utf-8")
        candidate_paths = self.service._codex_candidate_paths(
            entries=[entry],
            key=key,
            matches=[{"repo": "Portal Repo", "path": "repository/IssueRepository.java", "line_start": 1, "line_end": 3}],
        )
        answer = (
            '{"direct_answer":"IssueRepository reads issue_table.",'
            '"claims":[{"text":"IssueRepository reads issue_table",'
            '"citations":["repository/IssueRepository.java:1-3"]}],'
            '"missing_evidence":[],"confidence":"high"}'
        )

        validation = self.service._validate_codex_citations(
            answer,
            candidate_paths,
            [{"repo": "Portal Repo", "path": "repository/IssueRepository.java"}],
            scope_roots=self.service._codex_scope_roots(entries=[entry], key=key),
        )

        self.assertEqual(validation["status"], "ok")
        self.assertEqual(validation["cited_path_count"], 1)
        self.assertEqual(validation["scoped_file_refs"][0]["path"], "repository/IssueRepository.java")

    def test_codex_citation_validation_accepts_scoped_file_outside_candidate_paths(self):
        entry = RepositoryEntry("Portal Repo", "https://git.example.com/team/portal.git")
        key = "AF:All"
        repo_path = self.service._repo_path(key, entry)
        candidate_file = repo_path / "repository" / "IssueRepository.java"
        service_file = repo_path / "service" / "IssueService.java"
        candidate_file.parent.mkdir(parents=True)
        service_file.parent.mkdir(parents=True)
        candidate_file.write_text("class IssueRepository {}\n", encoding="utf-8")
        service_file.write_text("class IssueService {\n  void route() {}\n}\n", encoding="utf-8")
        candidate_paths = self.service._codex_candidate_paths(
            entries=[entry],
            key=key,
            matches=[{"repo": "Portal Repo", "path": "repository/IssueRepository.java", "line_start": 1, "line_end": 1}],
        )
        answer = (
            '{"direct_answer":"IssueService routes the request.",'
            '"claims":[{"text":"IssueService routes the request",'
            '"citations":["service/IssueService.java:1-2"]}],'
            '"missing_evidence":[],"confidence":"high"}'
        )

        validation = self.service._validate_codex_citations(
            answer,
            candidate_paths,
            [{"repo": "Portal Repo", "path": "repository/IssueRepository.java"}],
            scope_roots=self.service._codex_scope_roots(entries=[entry], key=key),
        )

        self.assertEqual(validation["status"], "ok")
        self.assertEqual(validation["scoped_file_refs"][0]["path"], "service/IssueService.java")

    def test_codex_citation_validation_rejects_out_of_scope_reference(self):
        allowed_entry = RepositoryEntry("Portal Repo", "https://git.example.com/team/portal.git")
        other_entry = RepositoryEntry("Other Repo", "https://git.example.com/team/other.git")
        allowed_key = "AF:All"
        other_key = "GRC:All"
        allowed_root = self.service._repo_path(allowed_key, allowed_entry)
        other_root = self.service._repo_path(other_key, other_entry)
        allowed_file = allowed_root / "repository" / "IssueRepository.java"
        other_file = other_root / "service" / "OtherService.java"
        allowed_file.parent.mkdir(parents=True)
        other_file.parent.mkdir(parents=True)
        allowed_file.write_text("class IssueRepository {}\n", encoding="utf-8")
        other_file.write_text("class OtherService {\n  void leak() {}\n}\n", encoding="utf-8")
        other_ref = other_file.relative_to(self.service.repo_root)
        candidate_paths = self.service._codex_candidate_paths(
            entries=[allowed_entry],
            key=allowed_key,
            matches=[{"repo": "Portal Repo", "path": "repository/IssueRepository.java", "line_start": 1, "line_end": 1}],
        )
        answer = (
            '{"direct_answer":"OtherService reads issue_table.",'
            f'"claims":[{{"text":"OtherService reads issue_table from table issue_table","citations":["{other_ref}:1-2"]}}],'
            '"missing_evidence":[],"confidence":"high"}'
        )

        validation = self.service._validate_codex_citations(
            answer,
            candidate_paths,
            [{"repo": "Portal Repo", "path": "repository/IssueRepository.java"}],
            scope_roots=self.service._codex_scope_roots(entries=[allowed_entry], key=allowed_key),
        )

        self.assertEqual(validation["status"], "needs_citation")
        self.assertEqual(validation["out_of_scope_refs"][0]["status"], "out_of_scope")

    def test_codex_deep_investigation_triggers_on_missing_business_chain(self):
        structured = {
            "direct_answer": "可能是 merchantUid 被当成 shopeeUid 后查 UC。",
            "confirmed_from_code": ["UserInfoSPIAdapter.queryMerchantInfo calls user-proxy [S1]"],
            "inferred_from_code": ["merchantUid may be copied into shopeeUid"],
            "not_found": ["No caller chain from report ingestion to queryMerchantInfo was found."],
            "missing_evidence": [],
            "claims": [{"text": "queryMerchantInfo uses shopeeUid", "citations": ["S1"]}],
            "confidence": "medium",
        }

        needed = self.service._codex_deep_investigation_needed(
            question="开发说 v2 上报拿 merchantUid 查 UC 填充 merchantInfo 是什么意思，为什么失败？",
            answer=structured["direct_answer"],
            structured_answer=structured,
            quality_gate={"status": "sufficient", "confidence": "medium"},
            answer_judge={"status": "ok", "issues": []},
            codex_validation={"status": "ok", "issues": []},
        )
        terms = self.service._codex_deep_investigation_terms(
            question="v2 report merchantUid queryMerchantInfo",
            answer=structured["direct_answer"],
            structured_answer=structured,
            answer_judge={"issues": []},
            codex_validation={"unsupported_claims": []},
        )

        self.assertTrue(needed)
        self.assertIn("querymerchantinfo", terms)
        self.assertIn("merchantuid", terms)

    def test_codex_raw_passthrough_keeps_mild_citation_warning_without_repair(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            llm_provider="codex_cli_bridge",
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        entry = RepositoryEntry("Portal Repo", "https://git.example.com/team/portal.git")
        key = "AF:All"
        repo_path = service._repo_path(key, entry)
        source_file = repo_path / "repository" / "IssueRepository.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text("class IssueRepository {\n  void find(){ jdbc.query(\"select * from issue_table\"); }\n}\n", encoding="utf-8")
        matches = [
            {
                "repo": "Portal Repo",
                "path": "repository/IssueRepository.java",
                "line_start": 1,
                "line_end": 3,
                "retrieval": "persistent_index",
                "trace_stage": "direct",
                "reason": "repository and table matched",
                "score": 10,
                "snippet": "class IssueRepository { void find(){ jdbc.query(\"select * from issue_table\"); } }",
            }
        ]
        calls = []

        def fake_run(command, **kwargs):
            calls.append((command, kwargs))
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            exec_count = len([item for item in calls if "exec" in item[0]])
            if exec_count == 1:
                Path(output_path).write_text(
                    '{"direct_answer":"IssueRepository is in repository/IssueRepository.java.",'
                    '"claims":[{"text":"IssueRepository is in repository/IssueRepository.java","citations":[]}],'
                    '"missing_evidence":[],"confidence":"high"}',
                    encoding="utf-8",
                )
            else:
                Path(output_path).write_text(
                    '{"direct_answer":"IssueRepository is in repository/IssueRepository.java.",'
                    '"claims":[{"text":"IssueRepository is in repository/IssueRepository.java","citations":["S1"]}],'
                    '"missing_evidence":[],"confidence":"high"}',
                    encoding="utf-8",
                )
            return SimpleNamespace(returncode=0, stdout='{"type":"done"}\n', stderr="")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ):
            payload = service._build_llm_answer(
                entries=[entry],
                key=key,
                pm_team="AF",
                country="All",
                question="where is IssueRepository",
                matches=matches,
                llm_budget_mode="auto",
                followup_context={"question": "previous", "answer": "previous answer"},
                requested_answer_mode="auto",
            )

        exec_calls = [item for item in calls if "exec" in item[0]]
        self.assertEqual(len(exec_calls), 1)
        self.assertEqual(payload["llm_route"]["prompt_mode"], "codex_investigation_brief_v5")
        self.assertIn("candidate_path_layers", payload["llm_route"])
        self.assertFalse(payload["llm_route"]["codex_repair_attempted"])
        self.assertTrue(payload["llm_route"]["codex_repair_allowed"])
        self.assertEqual(payload["llm_route"]["codex_repair_policy"], "severe_only")
        self.assertEqual(payload["llm_route"]["retrieval_role"], "hints")
        self.assertEqual(payload["llm_route"]["codex_repair_skipped_reason"], "")
        self.assertGreater(payload["llm_route"]["codex_validation_warning_count"], 0)
        self.assertEqual(payload["answer_claim_check"]["status"], "ok")
        self.assertEqual(payload["codex_cli_summary"]["citation_validation_status"], "warn")
        self.assertEqual(payload["codex_cli_summary"]["repair_policy"], "severe_only")
        self.assertEqual(payload["codex_cli_summary"]["retrieval_role"], "hints")
        self.assertGreater(payload["codex_cli_summary"]["warning_count"], 0)
        self.assertEqual(payload["codex_cli_summary"]["repair_skipped_reason"], "")
        self.assertEqual(payload["answer_contract"]["status"], "model_answer")
        self.assertIn("IssueRepository is in repository/IssueRepository.java", payload["llm_answer"])

    def test_effort_assessment_missing_evidence_does_not_trigger_codex_repair(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            llm_provider="codex_cli_bridge",
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        entry = RepositoryEntry("Portal Repo", "https://git.example.com/team/portal.git")
        key = "AF:All"
        repo_path = service._repo_path(key, entry)
        source_file = repo_path / "repository" / "IssueRepository.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text("class IssueRepository {\n  void find(){ jdbc.query(\"select * from issue_table\"); }\n}\n", encoding="utf-8")
        matches = [
            {
                "repo": "Portal Repo",
                "path": "repository/IssueRepository.java",
                "line_start": 1,
                "line_end": 3,
                "retrieval": "persistent_index",
                "trace_stage": "direct",
                "reason": "repository and table matched",
                "score": 10,
                "snippet": "class IssueRepository { void find(){ jdbc.query(\"select * from issue_table\"); } }",
            }
        ]
        calls = []

        def fake_run(command, **kwargs):
            calls.append((command, kwargs))
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            answer = "\n".join(
                    [
                        "业务理解 / Business Understanding",
                        "- Need to estimate approval impact for issue_table changes.",
                        "代码改动点 / Code Change Points",
                        "- Update the approval data flow so issue_table changes are handled by the right review process.",
                        "BE 人天 / BE Person-days",
                        "- 2-3 PD, driven by backend approval-flow changes.",
                        "FE 人天 / FE Person-days",
                        "- 0-1 PD, depending on whether the screen needs new copy.",
                        "QA / Integration Impact",
                        "- Cover approval, rejection, and rollback regression scenarios.",
                        "Assumptions / Risks",
                        "- Runtime configuration may change the final scope.",
                        "Confirmation Questions",
                        "- Confirm the affected suspended-user groups.",
                    ]
                )
            Path(output_path).write_text(
                json.dumps(
                    {
                        "direct_answer": answer,
                        "claims": [{"text": "IssueRepository reads issue_table", "citations": ["S1"]}],
                        "missing_evidence": ["Webform template screen evidence"],
                        "confidence": "medium",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            return SimpleNamespace(returncode=0, stdout='{"type":"done"}\n', stderr="")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ):
            payload = service._build_llm_answer(
                entries=[entry],
                key=key,
                pm_team="AF",
                country="All",
                question="Effort assessment impact analysis: why approval changes affect issue_table",
                matches=matches,
                llm_budget_mode="auto",
                followup_context={},
                requested_answer_mode="auto",
                effort_assessment=True,
            )

        exec_calls = [item for item in calls if "exec" in item[0]]
        self.assertEqual(len(exec_calls), 1)
        self.assertEqual(payload["llm_route"]["task"], "effort_assessment")
        self.assertFalse(payload["llm_route"]["codex_repair_attempted"])
        self.assertEqual(payload["llm_route"]["codex_repair_reason"], "")
        self.assertEqual(payload["llm_route"]["codex_deep_investigation_rounds"], 0)
        self.assertIn("right review process", payload["llm_answer"])
        self.assertNotIn("Missing: Webform template screen evidence", payload["llm_answer"])

    def test_oversized_repair_prompt_skips_before_deep_expansion(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            llm_provider="codex_cli_bridge",
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
            codex_repair_deadline_seconds=300,
        )
        service.codex_repair_prompt_token_limit = 1
        entry = RepositoryEntry("Portal Repo", "https://git.example.com/team/portal.git")
        key = "AF:All"
        repo_path = service._repo_path(key, entry)
        source_file = repo_path / "repository" / "IssueRepository.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text("class IssueRepository { void find(){} }\n", encoding="utf-8")
        matches = [
            {
                "repo": "Portal Repo",
                "path": "repository/IssueRepository.java",
                "line_start": 1,
                "line_end": 1,
                "retrieval": "persistent_index",
                "trace_stage": "direct",
                "reason": "IssueRepository matched",
                "score": 10,
                "snippet": "class IssueRepository { void find(){} }",
            }
        ]
        calls = []

        def fake_run(command, **kwargs):
            calls.append((command, kwargs))
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            Path(output_path).write_text(
                '{"direct_answer":"IssueRepository is relevant.",'
                '"claims":[{"text":"IssueRepository is relevant","citations":["S1"]}],'
                '"missing_evidence":[],"confidence":"medium"}',
                encoding="utf-8",
            )
            return SimpleNamespace(returncode=0, stdout='{"type":"done"}\n', stderr="")

        repair_decision = {
            "severe_repair_reasons": ["deep_investigation_needed_for_high_risk_question"],
            "repair_issues": ["deep_investigation_needed_for_high_risk_question"],
            "deep_needed": True,
            "repair_issue_count": 2,
            "repair_will_run": True,
            "repair_decision_ms": 0,
        }

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ), patch.object(service, "_codex_repair_decision", return_value=repair_decision), patch.object(
            service,
            "_codex_deep_investigation_context",
            wraps=service._codex_deep_investigation_context,
        ) as deep_context:
            payload = service._build_llm_answer(
                entries=[entry],
                key=key,
                pm_team="AF",
                country="All",
                question="why is IssueRepository relevant?",
                matches=matches,
                llm_budget_mode="auto",
                followup_context={},
                requested_answer_mode="auto",
            )

        exec_calls = [item for item in calls if "exec" in item[0]]
        self.assertEqual(len(exec_calls), 1)
        deep_context.assert_not_called()
        self.assertFalse(payload["llm_route"]["codex_repair_attempted"])
        self.assertIn("repair_preflight_prompt_too_large", payload["llm_route"]["codex_repair_skipped_reason"])

    def test_codex_out_of_scope_citation_triggers_severe_repair(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            llm_provider="codex_cli_bridge",
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        entry = RepositoryEntry("Portal Repo", "https://git.example.com/team/portal.git")
        key = "AF:All"
        repo_path = service._repo_path(key, entry)
        source_file = repo_path / "repository" / "IssueRepository.java"
        outside_file = Path(self.temp_dir.name) / "outside" / "Outside.java"
        source_file.parent.mkdir(parents=True)
        outside_file.parent.mkdir(parents=True)
        source_file.write_text("class IssueRepository {\n  void find(){ jdbc.query(\"select * from issue_table\"); }\n}\n", encoding="utf-8")
        outside_file.write_text("class Outside {\n  void leak() {}\n}\n", encoding="utf-8")
        matches = [
            {
                "repo": "Portal Repo",
                "path": "repository/IssueRepository.java",
                "line_start": 1,
                "line_end": 3,
                "retrieval": "persistent_index",
                "trace_stage": "direct",
                "reason": "repository and table matched",
                "score": 10,
                "snippet": "class IssueRepository { void find(){ jdbc.query(\"select * from issue_table\"); } }",
            }
        ]
        calls = []

        def fake_run(command, **kwargs):
            calls.append((command, kwargs))
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            exec_count = len([item for item in calls if "exec" in item[0]])
            if exec_count == 1:
                Path(output_path).write_text(
                    '{"direct_answer":"Outside handles issue_table.",'
                    f'"claims":[{{"text":"Outside handles issue_table","citations":["{outside_file}:1-2"]}}],'
                    '"missing_evidence":[],"confidence":"high"}',
                    encoding="utf-8",
                )
            else:
                Path(output_path).write_text(
                    '{"direct_answer":"IssueRepository reads issue_table.",'
                    '"claims":[{"text":"IssueRepository reads issue_table","citations":["S1"]}],'
                    '"missing_evidence":[],"confidence":"high"}',
                    encoding="utf-8",
                )
            return SimpleNamespace(returncode=0, stdout='{"type":"done"}\n', stderr="")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ):
            payload = service._build_llm_answer(
                entries=[entry],
                key=key,
                pm_team="AF",
                country="All",
                question="why does IssueRepository read issue_table",
                matches=matches,
                llm_budget_mode="auto",
                followup_context={},
                requested_answer_mode="auto",
                effort_assessment=True,
            )

        exec_calls = [item for item in calls if "exec" in item[0]]
        self.assertEqual(len(exec_calls), 2)
        self.assertIn(service._model_for_role("repair"), exec_calls[1][0])
        self.assertEqual(payload["llm_route"]["task"], "effort_assessment")
        self.assertTrue(payload["llm_route"]["codex_repair_attempted"])
        self.assertIn("out_of_scope_citations", payload["llm_route"]["codex_repair_reason"])
        self.assertEqual(payload["codex_cli_summary"]["citation_validation_status"], "ok")
        self.assertEqual(payload["answer_contract"]["status"], "model_answer")

    def test_effort_assessment_unsupported_claim_still_triggers_codex_repair(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            llm_provider="codex_cli_bridge",
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        entry = RepositoryEntry("Portal Repo", "https://git.example.com/team/portal.git")
        key = "AF:All"
        repo_path = service._repo_path(key, entry)
        source_file = repo_path / "repository" / "IssueRepository.java"
        source_file.parent.mkdir(parents=True)
        source_file.write_text("class IssueRepository {\n  void find(){ jdbc.query(\"select * from issue_table\"); }\n}\n", encoding="utf-8")
        matches = [
            {
                "repo": "Portal Repo",
                "path": "repository/IssueRepository.java",
                "line_start": 1,
                "line_end": 3,
                "retrieval": "persistent_index",
                "trace_stage": "direct",
                "reason": "repository and table matched",
                "score": 10,
                "snippet": "class IssueRepository { void find(){ jdbc.query(\"select * from issue_table\"); } }",
            }
        ]
        calls = []

        def fake_run(command, **kwargs):
            calls.append((command, kwargs))
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            exec_count = len([item for item in calls if "exec" in item[0]])
            if exec_count == 1:
                answer = "\n".join(
                    [
                        "业务理解 / Business Understanding",
                        "- Need to estimate approval impact.",
                        "代码改动点 / Code Change Points",
                        "- IssueRepository owns all suspended-case routing.",
                        "BE 人天 / BE Person-days",
                        "- 3-5 PD.",
                        "FE 人天 / FE Person-days",
                        "- 1-2 PD.",
                        "QA / Integration Impact",
                        "- Cover suspended-case regression.",
                        "Assumptions / Risks",
                        "- Scope depends on template rules.",
                        "Confirmation Questions",
                        "- Confirm the affected groups.",
                    ]
                )
                Path(output_path).write_text(
                    json.dumps(
                        {
                            "direct_answer": answer,
                            "claims": [{"text": "IssueRepository owns all suspended-case routing", "citations": []}],
                            "missing_evidence": [],
                            "confidence": "high",
                        },
                        ensure_ascii=False,
                    ),
                    encoding="utf-8",
                )
            else:
                Path(output_path).write_text(
                    '{"direct_answer":"IssueRepository reads issue_table [S1].",'
                    '"claims":[{"text":"IssueRepository reads issue_table","citations":["S1"]}],'
                    '"missing_evidence":[],"confidence":"high"}',
                    encoding="utf-8",
                )
            return SimpleNamespace(returncode=0, stdout='{"type":"done"}\n', stderr="")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ):
            payload = service._build_llm_answer(
                entries=[entry],
                key=key,
                pm_team="AF",
                country="All",
                question="Effort assessment impact analysis: why approval changes affect suspended cases",
                matches=matches,
                llm_budget_mode="auto",
                followup_context={},
                requested_answer_mode="auto",
                effort_assessment=True,
            )

        exec_calls = [item for item in calls if "exec" in item[0]]
        self.assertEqual(len(exec_calls), 2)
        self.assertEqual(payload["llm_route"]["task"], "effort_assessment")
        self.assertTrue(payload["llm_route"]["codex_repair_attempted"])
        self.assertIn("high_risk_claims_missing_scoped_file_evidence", payload["llm_route"]["codex_repair_reason"])

    def test_codex_repair_bad_request_keeps_initial_answer(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            llm_provider="codex_cli_bridge",
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        entry = RepositoryEntry("Portal Repo", "https://git.example.com/team/portal.git")
        key = "AF:All"
        repo_path = service._repo_path(key, entry)
        source_file = repo_path / "repository" / "IssueRepository.java"
        outside_file = Path(self.temp_dir.name) / "outside-repair" / "Outside.java"
        source_file.parent.mkdir(parents=True)
        outside_file.parent.mkdir(parents=True)
        source_file.write_text("class IssueRepository {\n  void find(){ jdbc.query(\"select * from issue_table\"); }\n}\n", encoding="utf-8")
        outside_file.write_text("class Outside {\n  void leak() {}\n}\n", encoding="utf-8")
        matches = [
            {
                "repo": "Portal Repo",
                "path": "repository/IssueRepository.java",
                "line_start": 1,
                "line_end": 3,
                "retrieval": "persistent_index",
                "trace_stage": "direct",
                "reason": "repository and table matched",
                "score": 10,
                "snippet": "class IssueRepository { void find(){ jdbc.query(\"select * from issue_table\"); } }",
            }
        ]
        calls = []

        def fake_run(command, **kwargs):
            calls.append((command, kwargs))
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            exec_count = len([item for item in calls if "exec" in item[0]])
            if exec_count == 1:
                Path(output_path).write_text(
                    '{"direct_answer":"Initial answer survived repair failure.",'
                    f'"claims":[{{"text":"IssueRepository reads issue_table","citations":["{outside_file}:1-2"]}}],'
                    '"missing_evidence":[],"confidence":"medium"}',
                    encoding="utf-8",
                )
            else:
                Path(output_path).write_text('{"detail":"Bad Request"}', encoding="utf-8")
            return SimpleNamespace(returncode=0, stdout='{"type":"done"}\n', stderr="")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ):
            payload = service._build_llm_answer(
                entries=[entry],
                key=key,
                pm_team="AF",
                country="All",
                question="what table does IssueRepository read",
                matches=matches,
                llm_budget_mode="auto",
                followup_context={"question": "previous", "answer": "previous answer"},
                requested_answer_mode="auto",
            )

        exec_calls = [item for item in calls if "exec" in item[0]]
        self.assertEqual(len(exec_calls), 2)
        self.assertEqual(payload["llm_route"]["codex_repair_skipped_reason"], "repair_failed_kept_initial_answer")
        self.assertEqual(payload["codex_cli_summary"]["repair_skipped_reason"], "repair_failed_kept_initial_answer")
        self.assertIn("Initial answer survived repair failure", payload["llm_answer"])
        self.assertNotIn('{"detail":"Bad Request"}', payload["llm_answer"])

    def test_codex_cli_bridge_requires_chatgpt_login(self):
        provider = CodexCliBridgeSourceCodeQALLMProvider(
            workspace_root=Path(self.temp_dir.name),
            timeout_seconds=20,
            codex_binary="codex",
        )
        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            return_value=SimpleNamespace(returncode=0, stdout="Logged in using API key\n", stderr=""),
        ):
            self.assertFalse(provider.ready())

    def test_codex_cli_bridge_nonzero_exit_falls_back(self):
        provider = CodexCliBridgeSourceCodeQALLMProvider(
            workspace_root=Path(self.temp_dir.name),
            timeout_seconds=20,
            codex_binary="codex",
        )

        def fake_run(command, **kwargs):
            if command[:3] == ["codex", "login", "status"]:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            return SimpleNamespace(returncode=2, stdout="", stderr="quota exhausted")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ):
            with self.assertRaisesRegex(ToolError, "Codex unavailable"):
                provider.generate(payload={"contents": [{"parts": [{"text": "hi"}]}]}, primary_model="codex-cli", fallback_model="codex-cli")

    def test_auto_simple_lookup_uses_codex_cheap_budget_with_zero_thinking(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = service.load_config()["mappings"]["AF:All"][0]
        repo_path = service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        source_file = repo_path / "bpmis" / "jira_client.py"
        source_file.parent.mkdir(parents=True)
        source_file.write_text(
            "class BPMISClient:\n"
            "    def batchCreateJiraIssue(self):\n"
            "        return self.post('/api/v1/issues/batchCreateJiraIssue')\n",
            encoding="utf-8",
        )
        expected_model = service.llm_budgets["cheap"]["model"]
        result = self._llm_result(
            '{"direct_answer":"batchCreateJiraIssue is in BPMISClient.",'
            '"claims":[{"text":"BPMISClient defines batchCreateJiraIssue","citations":["S1"]}],'
            '"missing_evidence":[],"confidence":"high"}',
            usage={"promptTokenCount": 20, "candidatesTokenCount": 8, "totalTokenCount": 28},
            model=expected_model,
        )

        with patch.object(service.llm_provider, "ready", return_value=True), patch.object(
            service.llm_provider,
            "generate",
            return_value=result,
        ) as mocked_generate:
            self._build_all_indexes(service)
            payload = service.query(
                pm_team="AF",
                country="All",
                question="where is batchCreateJiraIssue",
                answer_mode="auto",
                llm_budget_mode="auto",
            )

        self.assertEqual(payload["llm_budget_mode"], "cheap")
        self.assertEqual(payload["llm_model"], expected_model)
        self.assertEqual(mocked_generate.call_args.kwargs["primary_model"], expected_model)
        self.assertEqual(mocked_generate.call_args.kwargs["payload"]["_codex_reasoning_effort"], "low")
        self.assertEqual(payload["llm_thinking_budget"], 0)
        self.assertNotIn("thinkingConfig", mocked_generate.call_args.kwargs["payload"]["generationConfig"])

    def test_token_heavy_auto_route_uses_compact_deep_budget(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = service.load_config()["mappings"]["AF:All"][0]
        repo_path = service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        repository_file = repo_path / "repository" / "IssueRepository.java"
        repository_file.parent.mkdir(parents=True)
        repository_file.write_text(
            "public class IssueRepository {\n"
            "    public Issue findIssue() {\n"
            "        return jdbcTemplate.queryForObject(\"select * from issue_table\", mapper);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        expected_model = service.llm_budgets["compact_deep"]["model"]
        result = self._llm_result(
            '{"direct_answer":"IssueRepository reads issue_table.",'
            '"claims":[{"text":"IssueRepository reads issue_table","citations":["S1"]}],'
            '"missing_evidence":[],"confidence":"high"}',
            usage={"promptTokenCount": 12000, "candidatesTokenCount": 50, "totalTokenCount": 12050},
            model=expected_model,
        )

        with patch.object(service.llm_provider, "ready", return_value=True), patch.object(
            service.llm_provider,
            "generate",
            return_value=result,
        ) as mocked_generate, patch.object(
            service,
            "_estimate_llm_tokens",
            side_effect=[25000, 12000, 12000],
        ):
            self._build_all_indexes(service)
            payload = service.query(
                pm_team="AF",
                country="All",
                question="what data source does issue creation use",
                answer_mode="auto",
                llm_budget_mode="auto",
            )

        self.assertEqual(payload["llm_budget_mode"], "compact_deep")
        self.assertTrue(payload["llm_route"]["token_pressure"])
        self.assertEqual(payload["llm_route"]["original_selected"], "deep")
        self.assertEqual(payload["llm_model"], expected_model)
        self.assertEqual(mocked_generate.call_args_list[0].kwargs["primary_model"], expected_model)
        request_payload = mocked_generate.call_args_list[0].kwargs["payload"]
        self.assertNotIn("thinkingConfig", request_payload["generationConfig"])
        self.assertEqual(payload["llm_thinking_budget"], 0)

    def test_auto_answer_mode_routes_data_source_questions_to_deep_budget(self):
        service = SourceCodeQAService(
            data_root=Path(self.temp_dir.name),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            gitlab_token="secret-token",
            git_timeout_seconds=5,
            max_file_bytes=200_000,
        )
        service.save_mapping(
            pm_team="AF",
            country="All",
            repositories=[{"display_name": "Portal Repo", "url": "https://git.example.com/team/portal.git"}],
        )
        entry = service.load_config()["mappings"]["AF:All"][0]
        repo_path = service._repo_path("AF:All", type("Entry", (), entry)())
        (repo_path / ".git").mkdir(parents=True)
        repository_file = repo_path / "repository" / "IssueRepository.java"
        repository_file.parent.mkdir(parents=True)
        repository_file.write_text(
            "public class IssueRepository {\n"
            "    public Issue findIssue() {\n"
            "        return jdbcTemplate.queryForObject(\"select * from issue_table\", mapper);\n"
            "    }\n"
            "}\n",
            encoding="utf-8",
        )
        expected_model = service.llm_budgets["deep"]["model"]
        result = self._llm_result(
            '{"direct_answer":"IssueRepository reads issue_table.",'
            '"claims":[{"text":"IssueRepository reads issue_table","citations":["S1"]}],'
            '"missing_evidence":[],"confidence":"high"}',
            model=expected_model,
        )

        with patch.object(service.llm_provider, "ready", return_value=True), patch.object(
            service.llm_provider,
            "generate",
            return_value=result,
        ) as mocked_generate:
            self._build_all_indexes(service)
            payload = service.query(
                pm_team="AF",
                country="All",
                question="what data source does issue creation use",
                answer_mode="auto",
                llm_budget_mode="auto",
            )

        self.assertEqual(payload["answer_mode"], "auto")
        self.assertEqual(payload["llm_budget_mode"], "deep")
        self.assertEqual(payload["llm_model"], expected_model)
        self.assertEqual(mocked_generate.call_args_list[0].kwargs["primary_model"], expected_model)
        self.assertEqual(payload["llm_route"]["mode"], "auto")
        self.assertIn("data_source_trace", payload["llm_route"]["reason"])
        self.assertIn("IssueRepository reads issue_table", payload["llm_answer"])
        request_payload = mocked_generate.call_args_list[0].kwargs["payload"]
        self.assertNotIn("thinkingConfig", request_payload["generationConfig"])
        request_text = request_payload["contents"][0]["parts"][0]["text"]
        system_text = request_payload["systemInstruction"]["parts"][0]["text"]
        self.assertIn("Source Code & Runtime Evidence Assistant", request_text)
        self.assertIn("Final answer contract", request_text)
        self.assertIn("upstream source tables/APIs/repos", request_text)
        self.assertIn("Production code, mapper, client, SQL, and config evidence beats tests", request_text)
        self.assertIn("read-only code investigator", system_text)
        self.assertEqual(payload["llm_thinking_budget"], 0)

    def test_source_code_qa_runtime_helpers_cover_local_agent_and_auth_boundaries(self):
        self._ensure_runtime_helper_app()
        session_lock = portal_web._source_code_qa_codex_session_lock("")
        self.assertIs(session_lock, portal_web._source_code_qa_codex_session_lock(""))

        class FakeRemoteStore:
            def __init__(self, client):
                self.client = client

        class FakeLocalService:
            llm_provider_name = "codex_cli_bridge"

            def with_llm_provider(self, provider):
                self.llm_provider_name = provider
                return self

        fake_client = object()
        fake_service = FakeLocalService()
        self.app.config["SOURCE_CODE_QA_SERVICE"] = fake_service
        globals_dict = portal_web._get_source_code_qa_session_store.__globals__
        with self.app.app_context(), patch.dict(
            globals_dict,
            {
                "_local_agent_source_code_qa_enabled": lambda settings: True,
                "_build_local_agent_client": lambda settings: fake_client,
                "RemoteSourceCodeQASessionStore": FakeRemoteStore,
                "RemoteSourceCodeQAAttachmentStore": FakeRemoteStore,
                "RemoteSourceCodeQAGeneratedArtifactStore": FakeRemoteStore,
                "RemoteSourceCodeQARuntimeEvidenceStore": FakeRemoteStore,
                "RemoteSourceCodeQAService": lambda client, service, llm_provider="": SimpleNamespace(
                    client=client,
                    service=service,
                    llm_provider_name=llm_provider,
                ),
            },
        ):
            self.assertIs(portal_web._get_source_code_qa_session_store().client, fake_client)
            self.assertIs(portal_web._get_source_code_qa_attachment_store().client, fake_client)
            self.assertIs(portal_web._get_source_code_qa_generated_artifact_store().client, fake_client)
            self.assertIs(portal_web._get_source_code_qa_runtime_evidence_store().client, fake_client)
            self.assertIs(portal_web._build_source_code_qa_service("codex_cli_bridge").client, fake_client)

        self.assertFalse(portal_web._source_code_qa_git_auth_ready(SimpleNamespace(git_auth_ready=lambda: False), self.app.config["SETTINGS"]))
        self.assertTrue(portal_web._source_code_qa_git_auth_ready(SimpleNamespace(), SimpleNamespace(source_code_qa_gitlab_token="token")))
        self.assertTrue(portal_web._source_code_qa_provider_available(" codex_cli_bridge "))
        self.assertFalse(portal_web._source_code_qa_provider_available("gemini"))
        self.assertEqual(portal_web._source_code_qa_public_answer_mode("exact"), "auto")
        self.assertEqual(portal_web._source_code_qa_query_mode("cheap"), "deep")

    def test_source_code_qa_runtime_helpers_cover_auto_sync_edges(self):
        self._ensure_runtime_helper_app()
        progress_events = []

        class BrokenHealthService:
            def index_health_payload(self):
                raise RuntimeError("health failed")

        class NonListRepoHealthService:
            def index_health_payload(self):
                return {"keys": {"AF:All": {"repos": "invalid"}}}

        with self.app.app_context():
            self.assertTrue(portal_web._source_code_qa_scope_has_queryable_index(BrokenHealthService(), "AF:All"))
            self.assertTrue(portal_web._source_code_qa_scope_has_queryable_index(NonListRepoHealthService(), "AF:All"))

        class BackgroundService:
            def __init__(self, *, fail=False):
                self.fail = fail
                self.calls = []

            def mapping_key(self, pm_team, country):
                return f"{pm_team}:{country}"

            def index_health_payload(self):
                return {"keys": {"AF:SG": {"repos": [{"index": {"queryable": True, "state": "ready"}}]}}}

            def ensure_synced_today(self, *, pm_team, country):
                self.calls.append(("sync", pm_team, country))
                if self.fail:
                    raise RuntimeError("background failed")
                return {"attempted": True, "status": "ok", "key": f"{pm_team}:{country}"}

        class FakeThread:
            def __init__(self, target, daemon=False):
                self.target = target
                self.daemon = daemon

            def start(self):
                self.target()

        with patch.dict(os.environ, {"SOURCE_CODE_QA_QUERY_SYNC_MODE": "background"}, clear=False), patch.dict(
            portal_web._prepare_source_code_qa_auto_sync.__globals__,
            {"threading": SimpleNamespace(Thread=FakeThread)},
        ):
            with self.app.app_context():
                result = portal_web._prepare_source_code_qa_auto_sync(
                    BackgroundService(),
                    pm_team="AF",
                    country="SG",
                    progress_callback=lambda event, message, current, total: progress_events.append(event),
                )
                failed_background = portal_web._prepare_source_code_qa_auto_sync(
                    BackgroundService(fail=True),
                    pm_team="AF",
                    country="SG",
                )

        self.assertEqual(result["status"], "background_queued")
        self.assertEqual(failed_background["status"], "background_queued")
        self.assertIn("auto_sync_queued", progress_events)

        class BlockingService:
            def __init__(self, attempted=False):
                self.attempted = attempted

            def ensure_synced_today(self, *, pm_team, country):
                return {"attempted": self.attempted, "status": "fresh", "key": f"{pm_team}:{country}"}

        with patch.dict(os.environ, {"SOURCE_CODE_QA_QUERY_SYNC_MODE": "blocking"}, clear=False):
            with self.app.app_context():
                attempted_blocking_result = portal_web._prepare_source_code_qa_auto_sync(
                    BlockingService(attempted=True),
                    pm_team="AF",
                    country="SG",
                    progress_callback=lambda event, message, current, total: progress_events.append(event),
                )
                blocking_result = portal_web._prepare_source_code_qa_auto_sync(
                    BlockingService(),
                    pm_team="AF",
                    country="SG",
                    progress_callback=lambda event, message, current, total: progress_events.append(event),
                )

        self.assertEqual(attempted_blocking_result["status"], "fresh")
        self.assertEqual(blocking_result["status"], "fresh")
        self.assertEqual(progress_events[-4:], ["auto_sync_check", "auto_sync_completed", "auto_sync_check", "auto_sync_completed"])

    def test_source_code_qa_runtime_helpers_cover_capability_dictionary_and_store_failure(self):
        self._ensure_runtime_helper_app()

        class CapabilityStore:
            def __init__(self, fail=False):
                self.fail = fail

            def list(self, *, pm_team, country):
                if self.fail:
                    raise RuntimeError("store unavailable")
                if pm_team == "AF" and country == "SG":
                    return [{"source_type": "data_dictionary"}]
                return []

        with patch.dict(
            portal_web._source_code_qa_runtime_capabilities_payload.__globals__,
            {"_get_source_code_qa_runtime_evidence_store": lambda: CapabilityStore()},
        ):
            capabilities = portal_web._source_code_qa_runtime_capabilities_payload()
        self.assertTrue(capabilities["AF"]["SG"]["hasDictionary"])

        with patch.dict(
            portal_web._source_code_qa_runtime_capabilities_payload.__globals__,
            {"_get_source_code_qa_runtime_evidence_store": lambda: CapabilityStore(fail=True)},
        ):
            failed_capabilities = portal_web._source_code_qa_runtime_capabilities_payload()
        self.assertFalse(failed_capabilities["AF"]["SG"]["hasDictionary"])

    def test_source_code_qa_runtime_helpers_cover_attachment_evidence_and_artifact_edges(self):
        self._ensure_runtime_helper_app()
        max_attachments = portal_web.SourceCodeQAAttachmentStore.MAX_ATTACHMENTS
        with self.assertRaisesRegex(ToolError, "attachment_ids must be a list"):
            portal_web._source_code_qa_attachment_ids({"attachment_ids": "not-list"})
        with self.assertRaisesRegex(ToolError, "At most"):
            portal_web._source_code_qa_attachment_ids({"attachment_ids": [str(index) for index in range(max_attachments + 1)]})
        self.assertEqual(portal_web._source_code_qa_attachment_ids({"attachment_ids": None}), [])

        with self.assertRaisesRegex(ToolError, "session is required"):
            portal_web._resolve_source_code_qa_query_attachments(
                {"attachment_ids": ["a1"]},
                owner_email="teammate@npt.sg",
                session_id="",
            )

        class MissingSessionStore:
            def get(self, session_id, *, owner_email):
                return None

        class FoundSessionStore:
            def get(self, session_id, *, owner_email):
                return {"id": session_id, "owner_email": owner_email}

        class AttachmentStore:
            def resolve_many(self, *, owner_email, session_id, attachment_ids):
                return [{"id": attachment_ids[0], "filename": "evidence.txt", "size": 5}]

        with patch.dict(
            portal_web._resolve_source_code_qa_query_attachments.__globals__,
            {
                "_get_source_code_qa_session_store": lambda: MissingSessionStore(),
                "_get_source_code_qa_attachment_store": lambda: AttachmentStore(),
            },
        ):
            with self.assertRaisesRegex(ToolError, "session was not found"):
                portal_web._resolve_source_code_qa_query_attachments(
                    {"attachment_ids": ["a1"]},
                    owner_email="teammate@npt.sg",
                    session_id="s1",
                )

        with patch.dict(
            portal_web._resolve_source_code_qa_query_attachments.__globals__,
            {
                "_get_source_code_qa_session_store": lambda: FoundSessionStore(),
                "_get_source_code_qa_attachment_store": lambda: AttachmentStore(),
            },
        ):
            resolved = portal_web._resolve_source_code_qa_query_attachments(
                {"attachment_ids": ["a1"]},
                owner_email="teammate@npt.sg",
                session_id="s1",
            )
        self.assertEqual(resolved[0]["filename"], "evidence.txt")
        self.assertEqual(portal_web._source_code_qa_public_attachments([resolved[0], "bad"])[0]["id"], "a1")
        self.assertEqual(
            portal_web._source_code_qa_public_runtime_evidence(
                [{"id": "r1", "filename": "apollo.properties", "source_type": "apollo", "pm_team": "AF", "country": "SG"}, None]
            )[0]["source_type"],
            "apollo",
        )
        self.assertEqual(portal_web._source_code_qa_public_generated_artifacts([{"id": "g1", "question": "q"}, None])[0]["id"], "g1")

        class RuntimeEvidenceStore:
            def __init__(self, result=None, error=None):
                self.result = result
                self.error = error

            def resolve_scope(self, *, pm_team, country):
                if self.error:
                    raise self.error
                return self.result

        self.assertEqual(portal_web._resolve_source_code_qa_runtime_evidence(pm_team="AF", country="All"), [])
        with patch.dict(
            portal_web._resolve_source_code_qa_runtime_evidence.__globals__,
            {"_get_source_code_qa_runtime_evidence_store": lambda: RuntimeEvidenceStore(error=ToolError("bad scope"))},
        ):
            with self.assertRaisesRegex(ToolError, "bad scope"):
                portal_web._resolve_source_code_qa_runtime_evidence(pm_team="AF", country="SG")
        with self.app.app_context(), patch.dict(
            portal_web._resolve_source_code_qa_runtime_evidence.__globals__,
            {"_get_source_code_qa_runtime_evidence_store": lambda: RuntimeEvidenceStore(error=RuntimeError("store down"))},
        ):
            self.assertEqual(portal_web._resolve_source_code_qa_runtime_evidence(pm_team="AF", country="SG"), [])

        class GeneratedArtifactStore:
            def __init__(self, fail=False):
                self.fail = fail

            def save_sql_package(self, **kwargs):
                if self.fail:
                    raise RuntimeError("disk full")
                return {"id": "artifact-1", "sql": kwargs["sql"], "readme": kwargs["readme"]}

        artifact_globals = portal_web._build_source_code_qa_generated_artifacts.__globals__
        with patch.dict(
            artifact_globals,
            {
                "_extract_source_code_qa_sql_blocks": lambda text: ["select 1"] if "sql" in text else [],
                "_build_source_code_qa_sql_readme": lambda **kwargs: "readme",
                "_get_source_code_qa_generated_artifact_store": lambda: GeneratedArtifactStore(),
            },
        ):
            self.assertEqual(
                portal_web._build_source_code_qa_generated_artifacts(
                    owner_email="teammate@npt.sg",
                    session_id="s1",
                    pm_team="AF",
                    country="SG",
                    question="q",
                    result={"llm_answer": "no block"},
                    runtime_evidence=[],
                ),
                [],
            )
            generated = portal_web._build_source_code_qa_generated_artifacts(
                owner_email="teammate@npt.sg",
                session_id="s1",
                pm_team="AF",
                country="SG",
                question="q",
                result={"llm_answer": "sql"},
                runtime_evidence=[],
            )
        self.assertEqual(generated[0]["id"], "artifact-1")

        with self.app.app_context(), patch.dict(
            artifact_globals,
            {
                "_extract_source_code_qa_sql_blocks": lambda text: ["select 1"],
                "_build_source_code_qa_sql_readme": lambda **kwargs: "readme",
                "_get_source_code_qa_generated_artifact_store": lambda: GeneratedArtifactStore(fail=True),
            },
        ):
            self.assertEqual(
                portal_web._build_source_code_qa_generated_artifacts(
                    owner_email="teammate@npt.sg",
                    session_id="s1",
                    pm_team="AF",
                    country="SG",
                    question="q",
                    result={"llm_answer": "sql"},
                    runtime_evidence=[],
                ),
                [],
            )

    def test_source_code_qa_prompt_evidence_dedupes_empty_and_non_dict_items(self):
        deduped = SourceCodeQAService._dedupe_prompt_evidence_items(  # noqa: SLF001
            [
                "not-a-dict",
                {"filename": "first.txt"},
                {"filename": "first.txt"},
            ]
        )

        self.assertEqual(deduped, [{"filename": "first.txt"}])

    def test_source_code_qa_runtime_helpers_cover_context_release_gate_access_and_jobs(self):
        self._ensure_runtime_helper_app()
        result = {
            "summary": "summary",
            "llm_answer": "answer",
            "matches": [{"repo": "repo-a"}, {"repo": "repo-a"}, {"repo": "repo-b"}],
            "llm_route": {
                "candidate_paths": [
                    {"path": "repo/src/Main.java", "repo": "repo-a"},
                    {"path": "repo/src/Followup.java", "repo": "repo-b", "trace_stage": "followup_memory"},
                ],
                "codex_session_max_turns": 12,
            },
            "codex_cli_trace": {
                "session_id": "session-1",
                "session_mode": "resume",
                "probable_inspected_files": ["opened repo/src/Main.java"],
            },
            "trace_paths": ["trace-a", "trace-b", "trace-c", "trace-d", "trace-e", "trace-f"],
        }
        context = portal_web._build_source_code_qa_session_context(result, {"pm_team": "AF", "country": "SG", "question": "q"})
        self.assertEqual(context["codex_cli_session"]["session_id"], "session-1")
        self.assertEqual(context["codex_inspected_paths"][0]["path"], "repo/src/Main.java")
        self.assertEqual(context["repo_scope"], ["repo-a", "repo-b"])
        fallback_context = portal_web._build_source_code_qa_session_context(
            {"llm_route": {"candidate_paths": [{"path": "repo/src/Followup.java", "trace_stage": "followup_memory"}]}},
            {"pm_team": "AF"},
        )
        self.assertEqual(fallback_context["codex_inspected_paths"][0]["trace_stage"], "followup_memory")

        project_root = Path(self.temp_dir.name) / "project-root"
        data_root = project_root / "relative-data-root"
        run_dir = data_root / "run"
        eval_dir = data_root / "source_code_qa" / "eval_runs"
        run_dir.mkdir(parents=True)
        eval_dir.mkdir(parents=True)
        (run_dir / "source_code_qa_release_gate.json").write_text(
            json.dumps({"status": "pass", "timestamp": "2026-01-01T00:00:00Z", "summary": "green", "thresholds": {"failures": 0}}),
            encoding="utf-8",
        )
        (eval_dir / "latest.json").write_text(
            json.dumps({"status": "eval-pass", "eval": {"failures": 0}, "llm_smoke": {"failures": 0}, "report_path": "report.json"}),
            encoding="utf-8",
        )
        settings = SimpleNamespace(team_portal_data_dir=Path("relative-data-root"))
        with patch.dict(portal_web._source_code_qa_release_gate_payload.__globals__, {"PROJECT_ROOT": project_root}):
            gate = portal_web._source_code_qa_release_gate_payload(settings)
        self.assertEqual(gate["status"], "pass")
        self.assertEqual(gate["latest_eval"]["report_path"], "report.json")

        access_globals = portal_web._require_source_code_qa_access.__globals__
        with self.app.test_request_context("/api/source-code-qa/config"), patch.dict(
            access_globals,
            {"_require_google_login": lambda settings, api=False: ("login-required", 401)},
        ):
            self.assertEqual(portal_web._require_source_code_qa_access(settings, api=True), ("login-required", 401))

        with self.app.test_request_context("/api/source-code-qa/config"), patch.dict(
            access_globals,
            {"_require_google_login": lambda settings, api=False: None, "_can_access_source_code_qa": lambda settings: False},
        ):
            response, status = portal_web._require_source_code_qa_access(settings, api=True)
            self.assertEqual(status, 403)
            self.assertEqual(response.get_json()["status"], "error")

        with self.app.test_request_context("/source-code-qa"), patch.dict(
            access_globals,
            {"_require_google_login": lambda settings, api=False: None, "_can_access_source_code_qa": lambda settings: False},
        ):
            response = portal_web._require_source_code_qa_access(settings, api=False)
            self.assertEqual(response.status_code, 302)

        manage_globals = portal_web._require_source_code_qa_manage_access.__globals__
        with self.app.test_request_context("/source-code-qa/admin"), patch.dict(
            manage_globals,
            {
                "_require_source_code_qa_access": lambda settings, api=False: "access-gate",
            },
        ):
            self.assertEqual(portal_web._require_source_code_qa_manage_access(settings, api=False), "access-gate")

        with self.app.test_request_context("/source-code-qa/admin"), patch.dict(
            manage_globals,
            {
                "_require_source_code_qa_access": lambda settings, api=False: None,
                "_source_code_qa_auth_payload": lambda settings: {"signed_in_email": "teammate@npt.sg"},
                "_can_manage_source_code_qa": lambda settings: False,
            },
        ):
            response = portal_web._require_source_code_qa_manage_access(settings, api=False)
            self.assertEqual(response.status_code, 302)

        classifications = [
            portal_web._classify_source_code_qa_job_error("local-agent connection refused"),
            portal_web._classify_source_code_qa_job_error("ERR_NGROK_3200 html error"),
            portal_web._classify_source_code_qa_job_error("rate limit quota"),
            portal_web._classify_source_code_qa_job_error("timed out"),
            portal_web._classify_source_code_qa_job_error("other"),
        ]
        self.assertEqual(classifications[0]["error_code"], "local_agent_unavailable")
        self.assertEqual(classifications[1]["error_code"], "gateway_disconnected")
        self.assertEqual(classifications[2]["error_code"], "llm_rate_limited")
        self.assertEqual(classifications[3]["error_code"], "llm_timeout")
        self.assertEqual(classifications[4]["error_code"], "source_code_qa_job_failed")

        stalled = portal_web._public_source_code_qa_job_snapshot(
            {
                "owner_email": "teammate@npt.sg",
                "state": "queued",
                "eta_seconds_range": [-1, "5", 12],
                "stalled_retryable": True,
            }
        )
        self.assertNotIn("owner_email", stalled)
        self.assertEqual(stalled["error_code"], "job_stalled_retryable")
        self.assertEqual(stalled["eta_seconds_range"], [0, 5])
        self.assertEqual(portal_web._public_source_code_qa_job_snapshot({"state": "running"})["error_category"], "job_running")
        self.assertEqual(portal_web._public_source_code_qa_job_snapshot({"state": "failed", "error": "quota"})["error_code"], "llm_rate_limited")

        with self.app.app_context():
            self.app.config["JOB_STORE"] = SimpleNamespace(snapshot=lambda job_id: None)
            self.assertIsNone(portal_web._source_code_qa_job_snapshot_for_current_user("missing"))
            self.app.config["JOB_STORE"] = SimpleNamespace(snapshot=lambda job_id: {"action": "other"})
            self.assertIsNone(portal_web._source_code_qa_job_snapshot_for_current_user("wrong-action"))

    def test_codex_llm_provider_base_and_static_edges(self):
        base = llm_providers.SourceCodeQALLMProvider()
        self.assertFalse(base.ready())
        with self.assertRaisesRegex(ToolError, "not supported"):
            base.generate(payload={}, primary_model="p", fallback_model="f")
        with self.assertRaisesRegex(ToolError, "unreadable"):
            base.extract_text({})
        self.assertEqual(base.public_config(), {"provider": "unknown", "ready": False})

        unsupported = llm_providers.UnsupportedSourceCodeQALLMProvider("")
        self.assertEqual(unsupported.name, "unknown")
        with self.assertRaisesRegex(ToolError, "unknown"):
            unsupported.generate(payload={}, primary_model="p", fallback_model="f")

        with patch.dict(os.environ, {"SOURCE_CODE_QA_CODEX_BINARY": ""}, clear=False):
            provider = CodexCliBridgeSourceCodeQALLMProvider(
                workspace_root=Path(self.temp_dir.name),
                timeout_seconds=1,
                concurrency_limit=99,
                session_mode="bad-mode",
                codex_binary="",
            )
        self.assertEqual(provider.timeout_seconds, 10)
        self.assertEqual(provider.concurrency_limit, 4)
        self.assertEqual(provider.session_mode, "ephemeral")
        self.assertEqual(provider.codex_binary, "codex")
        self.assertIs(provider._semaphore_for_limit(2), provider._semaphore_for_limit(2))
        self.assertEqual(provider._semaphore_limit, 2)
        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value=None):
            self.assertFalse(provider.ready())

        self.assertEqual(provider._reasoning_effort_from_payload({"_codex_reasoning_effort": "xhigh"}), "xhigh")
        self.assertEqual(provider._reasoning_effort_from_payload({"_codex_reasoning_effort": "verbose"}), "")
        self.assertEqual(provider._reasoning_config_args("medium"), ["-c", 'model_reasoning_effort="medium"'])
        self.assertEqual(provider._reasoning_config_args("bad"), [])
        self.assertEqual(provider._sanitize_cli_output(" a\n\t b "), "a b")
        self.assertEqual(provider._command_summary(["codex", "--output-last-message", "/tmp/out", "-"]), ["codex", "--output-last-message", "<output-file>", "-"])
        self.assertEqual(provider._tail_for_log("x" * 200, limit=10), "x" * 100)
        with self.assertRaisesRegex(ToolError, "no readable answer"):
            provider.extract_text({"text": "   "})

        self.assertEqual(provider._codex_error_answer_detail(""), "")
        self.assertEqual(provider._codex_error_answer_detail("{bad"), "")
        self.assertEqual(provider._codex_error_answer_detail("[]"), "")
        self.assertEqual(provider._codex_error_answer_detail("{}"), "")
        self.assertEqual(provider._codex_error_answer_detail('{"message":"ordinary"}'), "")
        self.assertIn("Invalid request: bad image", provider._codex_error_answer_detail('{"error":{"message":"Invalid request: bad image"}}'))
        self.assertEqual(
            provider._codex_failure_context_from_payload(
                {
                    "_codex_trace_id": "trace",
                    "_codex_phase": "repair",
                    "_codex_prompt_chars": 12,
                    "_codex_prompt_bytes": 13,
                    "_codex_estimated_prompt_tokens": 4,
                    "_codex_candidate_path_count": 2,
                    "_codex_candidate_repo_count": 1,
                    "_codex_repair_issue_count": 3,
                    "_codex_ignored": [],
                }
            ),
            {
                "trace_id": "trace",
                "phase": "repair",
                "prompt_chars": 12,
                "prompt_bytes": 13,
                "estimated_prompt_tokens": 4,
                "candidate_path_count": 2,
                "candidate_repo_count": 1,
                "repair_issue_count": 3,
            },
        )

    def test_codex_llm_provider_command_prompt_and_event_parsing_edges(self):
        provider = CodexCliBridgeSourceCodeQALLMProvider(
            workspace_root=Path(self.temp_dir.name),
            timeout_seconds=20,
            session_mode="resume",
            codex_binary="codex",
        )
        command, mode = provider._build_codex_command(
            output_file="/tmp/out",
            model="gpt-5.4",
            reasoning_effort="high",
            image_paths=["/tmp/a.png"],
        )
        self.assertEqual(mode, "new_persistent")
        self.assertIn("--model", command)
        self.assertIn("-c", command)
        self.assertIn("--image", command)
        self.assertIn("new_persistent", mode)
        resume_command, resume_mode = provider._build_codex_command(
            output_file="/tmp/out",
            model="codex-cli",
            reasoning_effort="",
            session_id="session-1",
        )
        self.assertEqual(resume_mode, "resume")
        self.assertIn("session-1", resume_command)

        ephemeral = CodexCliBridgeSourceCodeQALLMProvider(
            workspace_root=Path(self.temp_dir.name),
            timeout_seconds=20,
            codex_binary="codex",
        )
        ep_command, ep_mode = ephemeral._build_codex_command(
            output_file="/tmp/out",
            model="gpt-5.4",
            reasoning_effort="medium",
            image_paths=["/tmp/a.png"],
        )
        self.assertEqual(ep_mode, "ephemeral")
        self.assertLess(ep_command.index("--model"), ep_command.index("--image"))
        self.assertIn('model_reasoning_effort="medium"', ep_command)

        prompt = ephemeral._prompt_from_llm_payload(
            {
                "systemInstruction": {"parts": [{"text": " system "}, {"text": ""}]},
                "contents": [{"parts": [{"text": " user 1 "}, {"text": ""}]}, {"parts": [{"text": "user 2"}]}],
            }
        )
        self.assertIn("system", prompt)
        self.assertIn("user 1", prompt)
        self.assertIn("user 2", prompt)

        output = "\n".join(
            [
                "not-json",
                "[]",
                '{"message":"first"}',
                '{"text":"second"}',
                '{"output_text":"third"}',
                '{"item":{"content":"final"}}',
            ]
        )
        self.assertEqual(ephemeral._extract_last_json_event_message(output), "final")
        progress = ephemeral._extract_progress_json_event_message(
            json.dumps(
                {
                    "message": "top",
                    "item": {
                        "text": "item text",
                        "content": [{"text": "content text"}, {"output_text": "content output"}],
                    },
                }
            )
        )
        self.assertEqual(progress, "content output")
        self.assertEqual(ephemeral._extract_progress_json_event_message("not-json"), "")
        self.assertEqual(ephemeral._extract_progress_json_event_message("[]"), "")

        stdout = "\n".join(
            [
                '{"session_id":"session-top","command":"rg foo src/Main.java"}',
                '{"item":{"type":"session","id":"session-item","cmd":"grep bar app/Service.java"}}',
                "sed -n 1,20p repo/src/main/java/app/Service.java",
            ]
        )
        trace = ephemeral._extract_codex_trace(stdout, "cat repo/tests/ServiceTest.java")
        self.assertEqual(trace["session_id"], "session-item")
        self.assertTrue(any("rg foo" in command for command in trace["command_summaries"]))
        self.assertTrue(any("Service.java" in path for path in trace["probable_inspected_files"]))

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value=None), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.Path.exists",
            return_value=False,
        ):
            self.assertEqual(ephemeral._codex_rg_hint(), "")

    def test_codex_llm_provider_generate_failure_and_stdout_fallback_edges(self):
        provider = CodexCliBridgeSourceCodeQALLMProvider(
            workspace_root=Path(self.temp_dir.name),
            timeout_seconds=20,
            codex_binary="codex",
        )

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            return_value=SimpleNamespace(returncode=0, stdout="Logged in using API key\n", stderr=""),
        ):
            with self.assertRaisesRegex(ToolError, "Codex is unavailable"):
                provider.generate(payload={"contents": [{"parts": [{"text": "hi"}]}]}, primary_model="codex-cli", fallback_model="codex-cli")

        missing_image = Path(self.temp_dir.name) / "missing.png"
        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            return_value=SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr=""),
        ):
            with self.assertRaisesRegex(ToolError, "image attachment is missing"):
                provider.generate(
                    payload={"contents": [{"parts": [{"text": "hi"}]}], "_codex_image_paths": [str(missing_image)]},
                    primary_model="codex-cli",
                    fallback_model="codex-cli",
                )

        def fake_timeout(command, **kwargs):
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            raise llm_providers.subprocess.TimeoutExpired(command, 10)

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_timeout,
        ):
            with self.assertRaisesRegex(ToolError, "timed out"):
                provider.generate(payload={"contents": [{"parts": [{"text": "hi"}]}]}, primary_model="codex-cli", fallback_model="codex-cli")

        def fake_os_error(command, **kwargs):
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            raise OSError("spawn failed")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_os_error,
        ):
            with self.assertRaisesRegex(ToolError, "spawn failed"):
                provider.generate(payload={"contents": [{"parts": [{"text": "hi"}]}]}, primary_model="codex-cli", fallback_model="codex-cli")

        def fake_stdout_answer(command, **kwargs):
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            Path(output_path).write_text("", encoding="utf-8")
            return SimpleNamespace(returncode=0, stdout='{"message":"answer from stdout","session_id":"s1"}\n', stderr="")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_stdout_answer,
        ):
            result = provider.generate(payload={"contents": [{"parts": [{"text": "hi"}]}]}, primary_model="codex-cli", fallback_model="codex-cli")
        self.assertEqual(result.payload["text"], "answer from stdout")

        def fake_empty_answer(command, **kwargs):
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            output_path = command[command.index("--output-last-message") + 1]
            Path(output_path).write_text("", encoding="utf-8")
            return SimpleNamespace(returncode=0, stdout='{"type":"done"}\n', stderr="")

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_empty_answer,
        ):
            with self.assertRaisesRegex(ToolError, "no readable answer"):
                provider.generate(payload={"contents": [{"parts": [{"text": "hi"}]}]}, primary_model="codex-cli", fallback_model="codex-cli")

    def test_codex_llm_provider_streaming_and_queue_wait_edges(self):
        provider = CodexCliBridgeSourceCodeQALLMProvider(
            workspace_root=Path(self.temp_dir.name),
            timeout_seconds=20,
            codex_binary="codex",
        )
        progress_events = []

        def fake_run(command, **kwargs):
            if "login" in command and "status" in command:
                return SimpleNamespace(returncode=0, stdout="Logged in using ChatGPT\n", stderr="")
            raise AssertionError("streaming path should not call subprocess.run")

        def fake_streaming(**kwargs):
            output_path = kwargs["command"][kwargs["command"].index("--output-last-message") + 1]
            Path(output_path).write_text("streamed answer", encoding="utf-8")
            return SimpleNamespace(returncode=0, stdout='{"session_id":"stream-session"}\n', stderr="")

        time_values = iter([100.0, 100.0, 100.0, 100.4, 101.0, 101.0, 101.0])

        def fake_time():
            return next(time_values, 101.0)

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.shutil.which", return_value="/usr/local/bin/codex"), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.subprocess.run",
            side_effect=fake_run,
        ), patch.object(provider, "_run_codex_streaming", side_effect=fake_streaming), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.time.time",
            side_effect=fake_time,
        ):
            result = provider.generate(
                payload={
                    "contents": [{"parts": [{"text": "hi"}]}],
                    "_progress_callback": lambda event, message, current, total: progress_events.append((event, message)),
                },
                primary_model="codex-cli",
                fallback_model="codex-cli",
            )

        self.assertEqual(result.payload["text"], "streamed answer")
        self.assertTrue(any(event == "codex_queue" and "acquired" in message for event, message in progress_events))

        class FakePipe:
            def __init__(self, lines):
                self.lines = list(lines)

            def readline(self):
                if self.lines:
                    return self.lines.pop(0)
                return ""

            def close(self):
                raise RuntimeError("close failed")

        class FakeStdin:
            def write(self, value):
                self.value = value

            def close(self):
                return None

        class FakeProcess:
            def __init__(self):
                self.stdout = FakePipe(['{"message":"stream progress"}\n'])
                self.stderr = FakePipe(["stderr line\n"])
                self.stdin = FakeStdin()
                self.returncode = 0
                self.poll_count = 0

            def poll(self):
                self.poll_count += 1
                return None if self.poll_count < 4 else 0

            def kill(self):
                self.killed = True

        with patch("bpmis_jira_tool.source_code_qa_llm_providers.subprocess.Popen", return_value=FakeProcess()):
            completed = provider._run_codex_streaming(
                command=["codex", "exec"],
                prompt="prompt",
                progress_callback=lambda event, message, current, total: (_ for _ in ()).throw(RuntimeError("callback failed")),
                timeout_seconds=20,
            )
        self.assertEqual(completed.returncode, 0)
        self.assertIn("stream progress", completed.stdout)
        self.assertIn("stderr line", completed.stderr)

        class TimeoutProcess(FakeProcess):
            def __init__(self):
                super().__init__()
                self.stdout = FakePipe([])
                self.stderr = FakePipe([])
                self.killed = False

            def poll(self):
                return None

        timeout_process = TimeoutProcess()
        with patch("bpmis_jira_tool.source_code_qa_llm_providers.subprocess.Popen", return_value=timeout_process), patch(
            "bpmis_jira_tool.source_code_qa_llm_providers.time.time",
            side_effect=[0.0, 11.0],
        ):
            with self.assertRaises(llm_providers.subprocess.TimeoutExpired):
                provider._run_codex_streaming(
                    command=["codex", "exec"],
                    prompt="prompt",
                    progress_callback=lambda *args: None,
                    timeout_seconds=10,
                )
        self.assertTrue(timeout_process.killed)

if __name__ == "__main__":
    unittest.main()
