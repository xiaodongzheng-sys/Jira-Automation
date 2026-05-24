import tempfile
import unittest
import asyncio
from dataclasses import dataclass, replace
import hashlib
import io
import sys
import types
from pathlib import Path
from unittest.mock import patch
import json

from bpmis_jira_tool.config import Settings
from bpmis_jira_tool.errors import ToolError

import prd_briefing.reviewer as reviewer
import prd_briefing.service as briefing_service
from prd_briefing.confluence import ConfluenceConnector, IngestedConfluencePage, ParsedSection, SpreadsheetLink
from prd_briefing.reviewer import (
    PRD_BRIEFING_REVIEW_CACHE_KEY,
    PRD_REVIEW_PROMPT_VERSION,
    PRD_SUMMARY_PROMPT_VERSION,
    PRD_URL_SUMMARY_CACHE_KEY,
    PRDBriefingReviewRequest,
    PRDReviewRequest,
    PRDReviewService,
    _build_generation_coverage,
    _build_google_sheet_screenshot_prompt_section,
    _build_linked_spreadsheet_prompt_section,
    _build_prd_source_payload,
    _build_review_section_content,
    _build_section_coverage,
    _cache_key_for_section_selection,
    _coerce_download_bytes,
    _collect_section_image_artifacts,
    _collect_table_media_coverage,
    _emit_prd_progress,
    _extract_google_sheet_screenshot_evidence_from_image,
    _format_template_metadata_for_prompt,
    _generation_mode_for_page,
    _google_sheet_screenshot_cache_key,
    _google_drive_file_id_from_url,
    _google_sheet_screenshot_public_payload,
    _image_suffix_from_url,
    _is_anti_fraud_prd,
    _is_confluence_spreadsheet_url,
    _is_google_spreadsheet_url,
    _is_report_or_template_prd,
    _limit_extracted_workbook_text,
    _linked_artifact_public_payload,
    _normalize_selected_section_indexes,
    _parse_json_object,
    _prd_token_risk,
    _resolve_one_google_sheet_screenshot,
    _resolve_linked_spreadsheet_evidence,
    _section_selection_hash,
    _store_cached_google_sheet_screenshot_evidence,
    _load_cached_google_sheet_screenshot_evidence,
    _use_compact_review_prompt,
    build_prd_review_batch_prompt,
    build_prd_review_prompt,
    build_prd_review_synthesis_prompt,
    build_prd_review_system_text,
    build_prd_summary_batch_prompt,
    build_prd_summary_prompt,
    build_prd_summary_synthesis_prompt,
    prd_briefing_review_prompt_version,
    prd_summary_prompt_version,
)
from prd_briefing.service import (
    PRDBriefingService,
    RetrievalService,
    WALKTHROUGH_SCRIPT_PROMPT_VERSION,
    VoiceService,
    attach_presentation_media,
    build_briefing_summary,
    build_block_summary,
    build_detail_grounded_overview,
    build_developer_zh_fallback_script,
    build_presenter_notes,
    build_presentation_source_text,
    build_presentation_system_prompt,
    build_presentation_user_prompt,
    build_pm_briefing_blocks,
    build_scope_items,
    build_sections_from_text,
    build_sentence_timestamps,
    build_walkthrough_block_user_prompt,
    build_walkthrough_section_system_prompt,
    build_walkthrough_section_user_prompt,
    classify_briefing_category,
    describe_section_topic_zh,
    duration_from_edge_boundaries,
    estimate_tts_duration_seconds,
    extract_detail_points,
    extract_text,
    extract_json_array_text,
    format_source_text,
    build_heuristic_session_overview,
    infer_engineering_focus_zh,
    infer_impacted_modules,
    infer_impacted_modules_from_sections,
    localize_detail_point_zh,
    normalize_image_urls,
    normalize_overview_list,
    normalize_detail_sentence,
    normalize_presentation_chunks,
    normalize_presentation_media,
    parse_developer_overview_payload,
    proxy_confluence_image_url,
    overview_is_low_signal,
    optimize_tts_text,
    parse_presentation_chunks,
    parse_session_overview,
    safe_filename,
    select_sections_for_overview,
    split_presentation_sentences,
    summarize_chunk_for_fallback,
    truncate_for_prompt,
)
from prd_briefing.storage import BriefingStore


class FakeTextClient:
    def __init__(self):
        self.last_system_prompt = None
        self.last_user_prompt = None
        self.answer_calls = 0
        self.model_id = "codex:test"
        self.answer_response = "LLM answer"

    def is_configured(self):
        return False

    def create_answer(self, system_prompt, user_prompt):
        self.last_system_prompt = system_prompt
        self.last_user_prompt = user_prompt
        self.answer_calls += 1
        return self.answer_response


class FakeVoiceService:
    def __init__(self):
        self.cached_texts = set()
        self.synthesize_calls = []

    def synthesize(self, **kwargs):
        self.synthesize_calls.append(kwargs)
        return None

    def get_cached_audio_for_text(self, *, owner_key, text, language_code):
        return "audio/cached.mp3" if text in self.cached_texts else None

    def synthesize_presentation_chunk(self, **kwargs):
        chunk = kwargs["chunk"]
        self.synthesize_calls.append(kwargs)
        return {
            "id": chunk["id"],
            "title": chunk["title"],
            "content": chunk["content"],
            "audioUrl": "/prd-briefing/assets/audio/test.mp3",
            "duration": 1.5,
            "timestamps": [{"sentence": chunk["content"], "start": 0, "end": 1.5}],
            "imageUrls": chunk.get("imageUrls") or [],
            "media": chunk.get("media") or {"type": "none", "content": ""},
            "cacheKey": kwargs.get("presentation_cache_key") or "",
        }


@dataclass
class FakeConnector:
    page: IngestedConfluencePage
    calls: int = 0

    def ingest_page(self, page_ref, session_id):
        self.calls += 1
        return self.page


class FakeConfluenceResponse:
    def __init__(self, content: bytes):
        self.content = content


class FakeAttachmentConnector(FakeConnector):
    def __init__(self, page: IngestedConfluencePage, attachment_content: bytes):
        super().__init__(page)
        self.attachment_content = attachment_content
        self.requested_urls = []

    def _request(self, url, accept=None, **_kwargs):
        self.requested_urls.append((url, accept))
        return FakeConfluenceResponse(self.attachment_content)


class FakeDriveExecute:
    def __init__(self, value):
        self.value = value

    def execute(self):
        return self.value


class FakeDriveFiles:
    def __init__(self, workbook_content: bytes, *, modified_time: str = "2026-05-12T10:00:00.000Z"):
        self.workbook_content = workbook_content
        self.modified_time = modified_time
        self.get_calls = 0
        self.export_calls = 0

    def get(self, **_kwargs):
        self.get_calls += 1
        return FakeDriveExecute(
            {
                "id": "sheet123",
                "name": "MAS Outsourcing Register",
                "mimeType": "application/vnd.google-apps.spreadsheet",
                "modifiedTime": self.modified_time,
            }
        )

    def export_media(self, **_kwargs):
        self.export_calls += 1
        return FakeDriveExecute(self.workbook_content)

    def get_media(self, **_kwargs):
        self.export_calls += 1
        return FakeDriveExecute(self.workbook_content)


class FakeDriveService:
    def __init__(self, workbook_content: bytes, *, modified_time: str = "2026-05-12T10:00:00.000Z"):
        self.files_resource = FakeDriveFiles(workbook_content, modified_time=modified_time)

    def files(self):
        return self.files_resource


def _xlsx_bytes(*, sheet_count: int = 1, marker: str = "MAS format") -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    first = workbook.active
    first.title = "Register Format"
    first.append(["Field", "Required", "Format"])
    first.append(["Outsourcing Arrangement ID", "Yes", marker])
    for index in range(2, sheet_count + 1):
        sheet = workbook.create_sheet(f"Sheet {index}")
        sheet.append(["Column", "Rule"])
        sheet.append([f"field_{index}", f"rule_{index}"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _xlsx_template_risk_bytes() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Register Format"
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "Merged Header"
    sheet.append(["Field", "Field", ""])
    sheet.append(["Amount", "=SUM(D4:D5)", ""])
    sheet.column_dimensions["C"].hidden = True
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class PRDBriefingServiceTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.store = BriefingStore(Path(self.temp_dir.name))
        self.text_client = FakeTextClient()
        self.voice_service = FakeVoiceService()
        page = IngestedConfluencePage(
            page_id="123",
            title="Payments PRD",
            source_url="https://example.atlassian.net/wiki/pages/123",
            updated_at="2026-04-15T10:00:00Z",
            language="en",
            sections=[
                ParsedSection(
                    title="Overview",
                    section_path="Overview",
                    content="This PRD introduces approval workflow and reviewer assignment.",
                    html_content="<p>This PRD introduces approval workflow and reviewer assignment.</p>",
                    image_refs=[],
                ),
                ParsedSection(
                    title="Rollout",
                    section_path="Rollout",
                    content="Rollout will happen in three phases with developer handoff notes.",
                    html_content="<p>Rollout will happen in three phases with developer handoff notes.</p>",
                    image_refs=[],
                ),
            ],
            version_number="5",
            media_dict={
                "MEDIA_ID_1": {
                    "type": "table",
                    "content": "<table><tr><th>Field</th></tr><tr><td>Status</td></tr></table>",
                }
            },
            presentation_source_text="## Section 1: Overview\nThis PRD introduces approval workflow.\n[MEDIA_ID_1]",
        )
        self.service = PRDBriefingService(
            store=self.store,
            confluence=FakeConnector(page),
            text_client=self.text_client,
            voice_service=self.voice_service,
            walkthrough_prewarm_enabled=False,
        )

    def test_parse_presentation_chunks_extracts_json_array(self):
        chunks = parse_presentation_chunks(
            'Here is the JSON:\n```json\n[{"id":"chunk-1","title":"开场","content":"研发先看主流程。"}]\n```'
        )

        self.assertEqual(chunks[0]["id"], "chunk-1")
        self.assertEqual(chunks[0]["title"], "开场")

    def test_parse_presentation_chunks_rejects_invalid_json(self):
        with self.assertRaises(ValueError):
            parse_presentation_chunks('{"id":"chunk-1"}')

    def test_confluence_parser_keeps_media_inside_content_layout_with_toc(self):
        connector = ConfluenceConnector(
            base_url="https://confluence.example",
            email=None,
            api_token=None,
            bearer_token=None,
            store=self.store,
        )
        media_dict = {}

        sections = connector._parse_sections(
            html="""
            <div class="contentLayout2">
              <div class="toc-macro">Table of contents</div>
              <h2>Feature Details</h2>
              <p>Main implementation flow.</p>
              <p><img class="confluence-embedded-image" src="/download/attachments/123/flow.png" width="720"></p>
              <table>
                <tr><th>Status</th><th>Meaning</th></tr>
                <tr><td>NEW</td><td>Create a new outsourced case.</td></tr>
              </table>
            </div>
            """,
            base_url="https://confluence.example",
            source_url="https://confluence.example/pages/viewpage.action?pageId=123",
            session_id="session-1",
            media_dict=media_dict,
        )

        self.assertEqual(len(sections), 1)
        self.assertIn("Main implementation flow.", sections[0].content)
        self.assertGreaterEqual(len(sections[0].image_refs), 1)
        self.assertGreaterEqual(len(sections[0].media_refs), 2)
        self.assertTrue(any(item["type"] == "image" for item in media_dict.values()))
        self.assertTrue(any(item["type"] == "table" for item in media_dict.values()))
        source_text = connector._build_source_text_with_media(sections)
        self.assertIn("[MEDIA_ID_", source_text)
        self.assertIn("[IMAGE] https://confluence.example/download/attachments/123/flow.png", source_text)

    def test_confluence_parser_drops_unresolved_macro_image_placeholders(self):
        connector = ConfluenceConnector(
            base_url="https://confluence.example",
            email=None,
            api_token=None,
            bearer_token=None,
            store=self.store,
        )
        media_dict = {}

        sections = connector._parse_sections(
            html="""
            <table>
              <tr>
                <th>JIRA</th>
                <td>
                  <img src="$iconUrl">SPSK-264073
                  (<img src="/$statusIcon">)
                  <img src="/download/attachments/123/real.png" width="720">
                </td>
              </tr>
            </table>
            """,
            base_url="https://confluence.example",
            source_url="https://confluence.example/pages/viewpage.action?pageId=123",
            session_id="session-1",
            media_dict=media_dict,
        )

        html = sections[0].html_content
        self.assertNotIn("$iconUrl", html)
        self.assertNotIn("$statusIcon", html)
        self.assertIn("SPSK-264073", html)
        self.assertIn("/prd-briefing/image-proxy?src=", html)
        self.assertIn("table", [item["type"] for item in media_dict.values()])
        self.assertIn("image", [item["type"] for item in media_dict.values()])
        self.assertEqual(len(sections[0].image_refs), 1)
        self.assertIn("/download/attachments/123/real.png", sections[0].image_refs[0])

    def test_confluence_parser_registers_storage_attachment_images(self):
        connector = ConfluenceConnector(
            base_url="https://confluence.example",
            email=None,
            api_token=None,
            bearer_token=None,
            store=self.store,
        )
        media_dict = {}

        sections = connector._parse_sections(
            html="""
            <h2>3.1 Scenario Rules</h2>
            <p>Refer to the configuration screenshot.</p>
            <p>
              <ac:image ac:width="960">
                <ri:attachment ri:filename="AF Scenario Sheet.png" />
              </ac:image>
            </p>
            """,
            base_url="https://confluence.example",
            source_url="https://confluence.example/pages/viewpage.action?pageId=123",
            page_id="123",
            session_id="session-1",
            media_dict=media_dict,
        )

        self.assertEqual(len(sections), 1)
        self.assertEqual(len(sections[0].image_refs), 1)
        self.assertEqual(len(sections[0].media_refs), 1)
        image_url = sections[0].image_refs[0]
        self.assertIn("/download/attachments/123/AF%20Scenario%20Sheet.png", image_url)
        media = media_dict[sections[0].media_refs[0]]
        self.assertEqual(media["type"], "image")
        self.assertEqual(media["source_url"], image_url)
        self.assertEqual(media["filename"], "AF Scenario Sheet.png")

    def test_presentation_prompts_include_language_specific_script_rules(self):
        english_prompt = build_presentation_system_prompt("en")
        chinese_prompt = build_presentation_system_prompt("zh")

        self.assertIn("Singapore-neutral English", english_prompt)
        self.assertIn("subject-verb-object", english_prompt)
        self.assertIn("Avoid US or UK slang", english_prompt)
        self.assertIn("中文汉字和英文单词", chinese_prompt)
        self.assertIn("QPS", chinese_prompt)
        self.assertIn("如果...那么", chinese_prompt)

    def test_presentation_cache_hit_reuses_outline_without_second_llm_call(self):
        self.text_client.is_configured = lambda: True
        self.text_client.answer_response = json.dumps([
            {
                "id": "chunk-1",
                "title": "主流程",
                "content": "这一段说明审批主流程。",
                "media_ref": "MEDIA_ID_1",
            }
        ])

        first = self.service.process_prd_for_presentation(
            owner_key="anon:presentation",
            page_ref="https://example.atlassian.net/wiki/pages/123",
        )
        calls_after_first = self.text_client.answer_calls
        second = self.service.process_prd_for_presentation(
            owner_key="anon:presentation",
            page_ref="https://example.atlassian.net/wiki/pages/123",
        )

        self.assertFalse(first["cached"])
        self.assertTrue(second["cached"])
        self.assertEqual(calls_after_first, 1)
        self.assertEqual(self.text_client.answer_calls, 1)
        self.assertEqual(second["chunks"][0]["media"]["type"], "table")
        self.assertEqual(second["session"]["version_number"], "5")

    def test_presentation_cache_misses_when_page_version_changes(self):
        self.text_client.is_configured = lambda: True
        self.text_client.answer_response = json.dumps([
            {"id": "chunk-1", "title": "主流程", "content": "这一段说明审批主流程。"}
        ])

        self.service.process_prd_for_presentation(
            owner_key="anon:presentation",
            page_ref="https://example.atlassian.net/wiki/pages/123",
        )
        self.service.confluence.page.version_number = "6"
        second = self.service.process_prd_for_presentation(
            owner_key="anon:presentation",
            page_ref="https://example.atlassian.net/wiki/pages/123",
        )

        self.assertFalse(second["cached"])
        self.assertEqual(self.text_client.answer_calls, 2)

    def test_manual_presentation_path_repairs_invalid_model_json(self):
        self.text_client.is_configured = lambda: True
        responses = iter(
            [
                "not json",
                json.dumps(
                    [
                        {
                            "id": "chunk-manual",
                            "title": "手工输入",
                            "content": "这一段说明手工输入的 PRD 文本。",
                            "imageUrls": ["https://example.atlassian.net/download/attachments/123/flow.png"],
                        }
                    ]
                ),
            ]
        )

        def create_answer(system_prompt, user_prompt):
            self.text_client.last_system_prompt = system_prompt
            self.text_client.last_user_prompt = user_prompt
            self.text_client.answer_calls += 1
            return next(responses)

        self.text_client.create_answer = create_answer

        payload = self.service.process_prd_for_presentation(
            owner_key="anon:manual-presentation",
            text="Manual PRD text with a submit flow and a fallback image.",
            language="en",
        )

        self.assertEqual(payload["status"], "ok")
        self.assertFalse(payload["cached"])
        self.assertTrue(payload["session"]["page_id"].startswith("manual:"))
        self.assertEqual(payload["session"]["language"], "en")
        self.assertEqual(self.text_client.answer_calls, 2)
        self.assertEqual(payload["chunks"][0]["id"], "chunk-manual")
        self.assertEqual(
            payload["chunks"][0]["imageUrls"],
            ["/prd-briefing/image-proxy?src=https%3A%2F%2Fexample.atlassian.net%2Fdownload%2Fattachments%2F123%2Fflow.png"],
        )

    def test_presentation_generation_validates_source_and_provider(self):
        with self.assertRaisesRegex(ValueError, "PRD text or Confluence page URL is required"):
            self.service.process_prd_for_presentation(owner_key="anon:missing")

        with self.assertRaisesRegex(RuntimeError, "requires Codex to be configured"):
            self.service.process_prd_for_presentation(owner_key="anon:not-configured", text="Manual PRD text")

    def test_presentation_generation_reports_repair_failure(self):
        self.text_client.is_configured = lambda: True
        self.text_client.create_answer = lambda **_kwargs: "still not json"

        with self.assertRaisesRegex(RuntimeError, "valid presentation JSON"):
            self.service.process_prd_for_presentation(owner_key="anon:repair-fail", text="Manual PRD text")

    def test_media_ref_restores_image_table_or_none(self):
        chunks = attach_presentation_media(
            [
                {"id": "chunk-1", "title": "图", "content": "看图。", "media_ref": "MEDIA_ID_1", "imageUrls": []},
                {"id": "chunk-2", "title": "表", "content": "看表。", "media_ref": "MEDIA_ID_2", "imageUrls": []},
                {"id": "chunk-3", "title": "无", "content": "无媒体。", "media_ref": "MEDIA_ID_99", "imageUrls": []},
            ],
            {
                "MEDIA_ID_1": {"type": "image", "content": "/prd-briefing/image-proxy?src=x"},
                "MEDIA_ID_2": {"type": "table", "content": "<table><tr><th>A</th></tr></table>"},
            },
        )

        self.assertEqual(chunks[0]["media"]["type"], "image")
        self.assertEqual(chunks[0]["imageUrls"], ["/prd-briefing/image-proxy?src=x"])
        self.assertEqual(chunks[1]["media"]["type"], "table")
        self.assertEqual(chunks[2]["media"], {"type": "none", "content": ""})

    def test_presentation_chunk_normalizers_filter_unsafe_or_empty_values(self):
        self.assertEqual(
            normalize_image_urls(
                [
                    "",
                    "javascript:alert(1)",
                    "https://example.test/a.png",
                    "/prd-briefing/assets/audio/a.mp3",
                    "/prd-briefing/image-proxy?src=x",
                    "https://example.test/a.png",
                ]
            ),
            [
                "https://example.test/a.png",
                "/prd-briefing/assets/audio/a.mp3",
                "/prd-briefing/image-proxy?src=x",
            ],
        )
        self.assertEqual(normalize_presentation_media({"type": "video", "content": "x"}), {"type": "none", "content": ""})
        self.assertEqual(normalize_presentation_media({"type": "table", "content": "<table></table>"}), {"type": "table", "content": "<table></table>"})
        self.assertEqual(proxy_confluence_image_url("https://example.test/image.png"), "/prd-briefing/image-proxy?src=https%3A%2F%2Fexample.test%2Fimage.png")
        self.assertEqual(proxy_confluence_image_url("/prd-briefing/assets/local.png"), "/prd-briefing/assets/local.png")

        with self.assertRaisesRegex(ValueError, "usable chunks"):
            normalize_presentation_chunks([{"id": "empty", "content": ""}])

    def test_presentation_prompt_and_source_helpers_cover_edge_cases(self):
        long_section = ParsedSection(
            title="Long",
            section_path="3.1 Long Section",
            content=("word " * 1200).strip(),
            html_content="",
            image_refs=["https://example.test/one.png", "https://example.test/two.png"],
        )
        source_text = build_presentation_source_text([long_section])
        self.assertLessEqual(len(source_text), 12_000)
        self.assertIn("## Section 1: 3.1 Long Section", source_text)
        self.assertIn("[IMAGE] https://example.test/one.png", source_text)

        manual_sections = build_sections_from_text(" ".join(f"token{i}" for i in range(80)), chunk_size=120)
        self.assertGreaterEqual(len(manual_sections), 2)
        self.assertEqual(manual_sections[0].section_path, "KB Chunk 1")

        english_prompt = build_presentation_user_prompt(
            source_text="A PRD with an API flow.",
            image_urls=[f"https://example.test/{index}.png" for index in range(30)],
            language="en",
        )
        self.assertIn("Create 4 to 10 English presentation chunks", english_prompt)
        self.assertIn("https://example.test/23.png", english_prompt)
        self.assertNotIn("https://example.test/24.png", english_prompt)
        self.assertEqual(extract_json_array_text("prefix [{\"id\":\"chunk-1\"}] tail"), "[{\"id\":\"chunk-1\"}]")

    def test_generate_audio_preserves_chunk_media_and_uses_version_cache_key(self):
        self.text_client.is_configured = lambda: True
        self.text_client.answer_response = json.dumps([
            {
                "id": "chunk-1",
                "title": "主流程",
                "content": "这一段说明审批主流程。",
                "media_ref": "MEDIA_ID_1",
            }
        ])
        payload = self.service.process_prd_for_presentation(
            owner_key="anon:presentation",
            page_ref="https://example.atlassian.net/wiki/pages/123",
        )
        edited_chunk = dict(payload["chunks"][0])
        edited_chunk["content"] = "这一段编辑后只需要重新生成本段音频。"

        result = self.service.generate_presentation_audio(
            owner_key="anon:presentation",
            session_id=payload["session"]["session_id"],
            chunk=edited_chunk,
        )

        self.assertEqual(result["chunk"]["media"]["type"], "table")
        self.assertEqual(self.voice_service.synthesize_calls[-1]["presentation_cache_key"], "123_5")
        self.assertEqual(self.voice_service.synthesize_calls[-1]["language_code"], "zh")

    def test_presentation_audio_uses_english_session_language(self):
        self.text_client.is_configured = lambda: True
        self.text_client.answer_response = json.dumps([
            {
                "id": "chunk-1",
                "title": "Main Flow",
                "content": "We will build a new button.",
            }
        ])
        payload = self.service.process_prd_for_presentation(
            owner_key="anon:presentation-en",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            language="en",
        )

        self.service.generate_presentation_audio(
            owner_key="anon:presentation-en",
            session_id=payload["session"]["session_id"],
            chunk=dict(payload["chunks"][0]),
        )

        self.assertEqual(self.voice_service.synthesize_calls[-1]["language_code"], "en")

    def test_presentation_audio_cache_is_scoped_by_page_version_and_chunk(self):
        voice = VoiceService(
            store=self.store,
            tts_provider="edge",
            edge_mandarin_voice="zh-CN-XiaozhenNeural",
            edge_english_voice="en-US-JennyNeural",
            edge_rate="-12%",
            edge_mandarin_rate="+0%",
            edge_english_rate="-5%",
        )
        calls = []
        voice._synthesize_with_edge_tts_with_boundaries = lambda **kwargs: (calls.append(kwargs) or (b"mp3", []))
        chunk = {
            "id": "chunk-1",
            "title": "主流程",
            "content": "同一段文本用于验证缓存隔离。",
            "media": {"type": "none", "content": ""},
        }

        first = voice.synthesize_presentation_chunk(
            session_id="session-1",
            owner_key="anon:presentation-audio",
            chunk=chunk,
            language_code="zh",
            presentation_cache_key="123_5",
        )
        second = voice.synthesize_presentation_chunk(
            session_id="session-2",
            owner_key="anon:presentation-audio",
            chunk=chunk,
            language_code="zh",
            presentation_cache_key="123_6",
        )
        third = voice.synthesize_presentation_chunk(
            session_id="session-3",
            owner_key="anon:presentation-audio",
            chunk=chunk,
            language_code="zh",
            presentation_cache_key="123_5",
        )

        self.assertEqual(len(calls), 2)
        self.assertEqual(calls[0]["rate"], "+0%")
        self.assertIn("audio/123_5/chunk-1-", first["audioUrl"])
        self.assertIn("audio/123_6/chunk-1-", second["audioUrl"])
        self.assertEqual(third["audioUrl"], first["audioUrl"])

    def test_presentation_audio_cache_is_scoped_by_language_rate(self):
        voice = VoiceService(
            store=self.store,
            tts_provider="edge",
            edge_mandarin_voice="zh-CN-XiaoxiaoNeural",
            edge_english_voice="en-SG-LunaNeural",
            edge_rate="-12%",
            edge_mandarin_rate="+0%",
            edge_english_rate="-5%",
        )
        calls = []
        voice._synthesize_with_edge_tts_with_boundaries = lambda **kwargs: (calls.append(kwargs) or (b"mp3", []))
        chunk = {
            "id": "chunk-1",
            "title": "Main Flow",
            "content": "We will build a new button.",
            "media": {"type": "none", "content": ""},
        }

        voice.synthesize_presentation_chunk(
            session_id="session-1",
            owner_key="anon:presentation-audio-language",
            chunk=chunk,
            language_code="en",
            presentation_cache_key="123_5",
        )

        self.assertEqual(calls[0]["voice_id"], "en-SG-LunaNeural")
        self.assertEqual(calls[0]["rate"], "-5%")

    def _settings(self) -> Settings:
        return Settings(
            flask_secret_key="secret",
            google_oauth_client_secret_file=Path(self.temp_dir.name) / "client.json",
            google_oauth_redirect_uri=None,
            team_portal_host="127.0.0.1",
            team_portal_port=5000,
            team_portal_base_url=None,
            team_allowed_emails=(),
            team_allowed_email_domains=(),
            team_portal_data_dir=Path(self.temp_dir.name),
            spreadsheet_id="sheet",
            common_tab_name="Common",
            input_tab_name="Input",
            bpmis_base_url="https://example.com",
            bpmis_api_access_token="token",
        )

    def _add_prd_image(self, page: IngestedConfluencePage, *, section_index: int = 2, url: str = "https://example.atlassian.net/download/attachments/123/sheet.png") -> str:
        media_id = f"MEDIA_ID_TEST_{section_index}"
        section = page.sections[section_index - 1]
        section.media_refs = [media_id]
        section.image_refs = [url]
        section.content = f"{section.content}\n[{media_id}]"
        page.media_dict[media_id] = {
            "type": "image",
            "content": f"/prd-briefing/image-proxy?src={url}",
            "source_url": url,
            "section_path": section.section_path,
        }
        return media_id

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_create_session_persists_sections(self):
        payload = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )

        self.assertEqual(payload["session"]["title"], "Payments PRD")
        self.assertEqual(len(payload["sections"]), 2)
        self.assertTrue(payload["session_overview"]["overview"])
        self.assertIn("unclear_rules", payload["session_overview"])
        self.assertIn("missing_edge_cases", payload["session_overview"])
        self.assertIn("unclear_ownership", payload["session_overview"])
        self.assertEqual(payload["sections"][0]["section_path"], "Overview")
        self.assertIn("<p>", payload["sections"][0]["html_content"])
        self.assertTrue(payload["sections"][0]["briefing_notes"])
        self.assertTrue(payload["sections"][0]["briefing_summary"])
        self.assertEqual(self.service.confluence.calls, 1)
        self.assertIn("briefing_blocks", payload)
        self.assertTrue(payload["briefing_blocks"])
        self.assertIn("section_indexes", payload["briefing_blocks"][0])

    def test_recreating_same_prd_keeps_previous_session_chunks(self):
        first = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )
        second = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )

        first_payload = self.service.get_session_payload(
            session_id=first["session"]["session_id"],
            owner_key="anon:test",
        )
        second_payload = self.service.get_session_payload(
            session_id=second["session"]["session_id"],
            owner_key="anon:test",
        )

        self.assertEqual(len(first_payload["sections"]), 2)
        self.assertEqual(len(second_payload["sections"]), 2)
        self.assertNotEqual(first["session"]["session_id"], second["session"]["session_id"])

    def test_answer_question_uses_prd_context(self):
        payload = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )

        answer = self.service.answer_question(
            session_id=payload["session"]["session_id"],
            owner_key="anon:test",
            question="What is the approval workflow?",
        )

        self.assertIn("根据当前检索到的 PRD 内容", answer["answer_text"])
        self.assertIn("可优先查看", answer["answer_text"])
        self.assertEqual(answer["groundedness"], "grounded")
        self.assertGreaterEqual(len(answer["citations"]), 1)

    def test_session_and_audio_methods_validate_missing_records(self):
        with self.assertRaisesRegex(ValueError, "session_id is required"):
            self.service.generate_presentation_audio(owner_key="anon:test", session_id="", chunk={})
        with self.assertRaisesRegex(ValueError, "Briefing session was not found"):
            self.service.generate_presentation_audio(owner_key="anon:test", session_id="missing", chunk={})
        with self.assertRaisesRegex(ValueError, "Briefing session was not found"):
            self.service.get_session_payload(session_id="missing", owner_key="anon:test")
        with self.assertRaisesRegex(ValueError, "Briefing session was not found"):
            self.service.answer_question(session_id="missing", owner_key="anon:test", question="What changed?")
        with self.assertRaisesRegex(ValueError, "Briefing session was not found"):
            self.service.narrate_section(session_id="missing", owner_key="anon:test")

    def test_narrate_section_validates_indexes_and_can_generate_block_audio(self):
        self.text_client.is_configured = lambda: True
        self.voice_service.synthesize = lambda **kwargs: (self.voice_service.synthesize_calls.append(kwargs) or "audio/block.mp3")
        payload = self.service.create_session(
            owner_key="anon:narrate-edges",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )
        session_id = payload["session"]["session_id"]
        block_id = payload["briefing_blocks"][0]["block_id"]

        with self.assertRaisesRegex(ValueError, "Section index is out of range"):
            self.service.narrate_section(
                session_id=session_id,
                owner_key="anon:narrate-edges",
                section_index=99,
                include_audio=False,
            )
        with self.assertRaisesRegex(ValueError, "Briefing block is out of range"):
            self.service.narrate_section(
                session_id=session_id,
                owner_key="anon:narrate-edges",
                briefing_block_id="missing-block",
                include_audio=False,
            )

        result = self.service.narrate_section(
            session_id=session_id,
            owner_key="anon:narrate-edges",
            briefing_block_id=block_id,
            include_audio=True,
        )

        self.assertEqual(result["audio_url"], "/prd-briefing/assets/audio/block.mp3")
        self.assertEqual(result["briefing_block_id"], block_id)
        self.assertEqual(self.voice_service.synthesize_calls[-1]["language_code"], "zh")

    def test_answer_question_falls_back_for_unsupported_and_model_errors(self):
        payload = self.service.create_session(
            owner_key="anon:qa-edges",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )
        unsupported = self.service.answer_question(
            session_id=payload["session"]["session_id"],
            owner_key="anon:qa-edges",
            question="Completely unrelated blockchain settlement wording",
        )
        self.assertEqual(unsupported["groundedness"], "unsupported")
        self.assertIn("无法在当前选择的 PRD", unsupported["answer_text"])

        self.text_client.is_configured = lambda: True
        self.text_client.create_answer = lambda **_kwargs: (_ for _ in ()).throw(RuntimeError("model down"))
        inferred = self.service.answer_question(
            session_id=payload["session"]["session_id"],
            owner_key="anon:qa-edges",
            question="developer handoff",
        )
        self.assertIn("根据当前检索到的 PRD 内容", inferred["answer_text"])

    def test_prd_review_skill_artifact_contains_section_guidance(self):
        skill_path = Path("/Users/NPTSG0388/.codex/skills/prd-review/SKILL.md")
        content = skill_path.read_text(encoding="utf-8")

        self.assertIn("name: prd-review", content)
        self.assertIn("description:", content)
        self.assertIn("Section-by-Section Improvement Suggestions", content)
        self.assertIn("Suggested PRD update", content)
        self.assertIn("Acceptance check", content)
        self.assertIn("Source not found in selected sections", content)

    def test_prd_review_prompt_contains_section_improvement_guidance_and_prd_sections(self):
        page = self.service.confluence.page

        prompt = build_prd_review_prompt(
            jira_id="AF-123",
            jira_link="https://jira/browse/AF-123",
            prd_url=page.source_url,
            page=page,
        )

        self.assertEqual(PRD_REVIEW_PROMPT_VERSION, "v12_prd_review_token_optimized_table_packing")
        self.assertIn("prd-review", prompt)
        self.assertIn("Executive Verdict", prompt)
        self.assertIn("Top Must-Fix Delivery Blockers", prompt)
        self.assertIn("Section Patch Suggestions", prompt)
        self.assertIn("Evidence Coverage", prompt)
        self.assertNotIn("Report Generation Feasibility", prompt)
        self.assertNotIn("Report Template Risks", prompt)
        self.assertNotIn("PRD-to-Template Mapping Gaps", prompt)
        self.assertIn("优先级", prompt)
        self.assertIn("Suggested PRD patch", prompt)
        self.assertIn("Evidence basis", prompt)
        self.assertIn("建议补写", prompt)
        self.assertIn("验收检查", prompt)
        self.assertIn("PM Decision Checklist", prompt)
        self.assertIn("Source not found in selected sections", prompt)
        self.assertIn("不能把 `[MEDIA_ID_x]`", prompt)
        self.assertIn("identifier 必须按 PRD 原文保留", prompt)
        self.assertIn("exact typo", prompt)
        self.assertIn("最终得分", prompt)
        self.assertNotIn("Previous " + "PRD Evidence", prompt)
        self.assertNotIn("previous_" + "prd_url", prompt)
        self.assertNotIn("毒舌", prompt)
        self.assertNotIn("逻辑严密度评估", prompt)
        self.assertIn("AF-123", prompt)
        self.assertIn("This PRD introduces approval workflow", prompt)

    def test_prd_review_prompt_includes_report_sections_only_for_report_prd(self):
        page = self.service.confluence.page
        page.sections[1].section_path = "3.11.1 Report Layout and Field Name & Format"

        prompt = build_prd_review_prompt(
            jira_id="AF-123",
            jira_link="https://jira/browse/AF-123",
            prd_url=page.source_url,
            page=page,
            language="en",
        )

        self.assertIn("Report Generation Feasibility", prompt)
        self.assertIn("Report Template Risks", prompt)
        self.assertIn("PRD-to-Template Mapping Gaps", prompt)
        self.assertIn("Generation feasibility", prompt)
        self.assertIn("Technical generation risk", prompt)
        self.assertIn("PRD mapping gap", prompt)
        self.assertIn("Used in findings", prompt)

    def test_prd_review_prompt_expands_table_media_placeholders(self):
        page = IngestedConfluencePage(
            page_id="table-media",
            title="Anti-Fraud Table PRD",
            source_url="https://example.atlassian.net/wiki/pages/table-media",
            updated_at="2026-05-13T00:00:00Z",
            language="en",
            sections=[
                ParsedSection(
                    title="Scenario",
                    section_path="3.1.1 New Scenario: ApplyCCard",
                    content="CMS calls AF system.\n[MEDIA_ID_1]",
                    media_refs=["MEDIA_ID_1"],
                )
            ],
            media_dict={
                "MEDIA_ID_1": {
                    "type": "table",
                    "content": "<table><tr><th>Entry points</th><td>CMS to call AF system</td></tr><tr><th>Scenario</th><td>ApplyCCard</td></tr></table>",
                }
            },
        )

        prompt = build_prd_review_prompt(
            jira_id="AF-123",
            jira_link="https://jira/browse/AF-123",
            prd_url=page.source_url,
            page=page,
            language="en",
        )

        self.assertIn("[MEDIA_ID_1 table content]", prompt)
        self.assertIn("Entry points | CMS to call AF system", prompt)
        self.assertIn("Scenario | ApplyCCard", prompt)
        self.assertNotIn("\n[MEDIA_ID_1]\n", prompt)

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_review_coverage_reports_confluence_tables(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = IngestedConfluencePage(
            page_id="table-media",
            title="Anti-Fraud Table PRD",
            source_url="https://example.atlassian.net/wiki/pages/table-media",
            updated_at="2026-05-13T00:00:00Z",
            language="en",
            sections=[
                ParsedSection(title="Overview", section_path="Overview", content="Intro"),
                ParsedSection(
                    title="Scenario",
                    section_path="3.1.1 New Scenario: ApplyCCard",
                    content="CMS calls AF system.\n[MEDIA_ID_1]",
                    media_refs=["MEDIA_ID_1"],
                ),
            ],
            media_dict={
                "MEDIA_ID_1": {
                    "type": "table",
                    "content": "<table><tr><th>Entry points</th><td>CMS to call AF system</td></tr><tr><th>Scenario</th><td>ApplyCCard</td></tr></table>",
                }
            },
        )
        service = PRDReviewService(
            store=self.store,
            confluence=FakeConnector(page),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/table-media",
                language="en",
                selected_section_indexes=[2],
            )
        )

        coverage = result["coverage"]
        self.assertEqual(coverage["confluence_tables_total"], 1)
        self.assertEqual(coverage["confluence_tables_reviewed"], 1)
        self.assertEqual(coverage["confluence_tables"][0]["media_id"], "MEDIA_ID_1")
        self.assertEqual(coverage["confluence_tables"][0]["source_section_title"], "3.1.1 New Scenario: ApplyCCard")
        self.assertIn("Entry points | CMS to call AF system", mock_generate.call_args.kwargs["prompt"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_review_compacts_large_confluence_table_by_identifier_rows(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        rows = ["<tr><th>Scenario</th><th>Action</th><th>Rule</th></tr>"]
        for index in range(1, 180):
            scenario = "ApplyCCard" if index == 120 else f"UnrelatedScenario{index}"
            rows.append(
                f"<tr><td>{scenario}</td><td>Action{index}</td><td>{'long rule text ' * 12}</td></tr>"
            )
        page = IngestedConfluencePage(
            page_id="large-table-media",
            title="Anti-Fraud Table PRD",
            source_url="https://example.atlassian.net/wiki/pages/large-table-media",
            updated_at="2026-05-13T00:00:00Z",
            language="en",
            sections=[
                ParsedSection(
                    title="Scenario",
                    section_path="3.1.1 New Scenario: ApplyCCard",
                    content="CMS calls AF system for ApplyCCard.\n[MEDIA_ID_1]",
                    media_refs=["MEDIA_ID_1"],
                )
            ],
            media_dict={"MEDIA_ID_1": {"type": "table", "content": f"<table>{''.join(rows)}</table>"}},
        )
        service = PRDReviewService(
            store=self.store,
            confluence=FakeConnector(page),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/large-table-media",
                language="en",
                selected_section_indexes=[1],
            )
        )

        prompt = mock_generate.call_args.kwargs["prompt"]
        self.assertIn("Scenario | Action | Rule", prompt)
        self.assertIn("ApplyCCard | Action120", prompt)
        self.assertNotIn("UnrelatedScenario179", prompt)
        self.assertIn("Table compacted", prompt)
        self.assertTrue(result["coverage"]["table_truncated"])
        self.assertGreater(result["coverage"]["table_rows_omitted"], 0)

    def test_prd_review_result_cache_uses_prd_updated_at_and_prompt_version(self):
        saved = self.store.save_prd_review_result(
            owner_key="anon:test",
            jira_id="AF-123",
            jira_link="https://jira/browse/AF-123",
            prd_url="https://example/prd",
            prd_updated_at="2026-04-28T00:00:00Z",
            prompt_version=PRD_REVIEW_PROMPT_VERSION,
            status="completed",
            result_markdown="### Review",
            model_id="codex-cli",
            trace={"session_id": "s1"},
        )

        cached = self.store.get_prd_review_result(
            owner_key="anon:test",
            jira_id="AF-123",
            prd_url="https://example/prd",
            prd_updated_at="2026-04-28T00:00:00Z",
            prompt_version=PRD_REVIEW_PROMPT_VERSION,
        )
        stale = self.store.get_prd_review_result(
            owner_key="anon:test",
            jira_id="AF-123",
            prd_url="https://example/prd",
            prd_updated_at="2026-04-29T00:00:00Z",
            prompt_version=PRD_REVIEW_PROMPT_VERSION,
        )
        old_prompt_cached = self.store.get_prd_review_result(
            owner_key="anon:test",
            jira_id="AF-123",
            prd_url="https://example/prd",
            prd_updated_at="2026-04-28T00:00:00Z",
            prompt_version="v3_strict_delivery_logic_review_codex",
        )

        self.assertEqual(saved["result_markdown"], "### Review")
        self.assertEqual(cached["trace"]["session_id"], "s1")
        self.assertIsNone(stale)
        self.assertIsNone(old_prompt_cached)

    def test_prd_briefing_review_prompt_can_be_url_only(self):
        page = self.service.confluence.page

        prompt = build_prd_review_prompt(
            jira_id="",
            jira_link="",
            prd_url=page.source_url,
            page=page,
        )

        self.assertNotIn("Jira ID: -", prompt)
        self.assertNotIn("Jira Link: -", prompt)
        self.assertIn(f"PRD Link: {page.source_url}", prompt)
        self.assertIn("Executive Verdict", prompt)
        self.assertIn("Top Must-Fix Delivery Blockers", prompt)
        self.assertIn("Section Patch Suggestions", prompt)
        self.assertIn("This PRD introduces approval workflow", prompt)

    def test_prd_briefing_review_english_prompt_requests_english_output(self):
        page = self.service.confluence.page

        prompt = build_prd_review_prompt(
            jira_id="",
            jira_link="",
            prd_url=page.source_url,
            page=page,
            language="en",
        )

        self.assertIn("Return only concise Markdown in English", prompt)
        self.assertIn("Executive Verdict", prompt)
        self.assertIn("Top Must-Fix Delivery Blockers", prompt)
        self.assertIn("Section Patch Suggestions", prompt)
        self.assertIn("Suggested PRD patch", prompt)
        self.assertIn("Acceptance check", prompt)
        self.assertIn("Evidence basis", prompt)
        self.assertIn("Preserve scenario names", prompt)
        self.assertIn("exact typo", prompt)
        self.assertNotIn("Previous " + "PRD Evidence", prompt)
        self.assertNotIn("逻辑严密度评估", prompt)

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_review_filters_spurious_identifier_typo_clarification(self, mock_generate):
        page = IngestedConfluencePage(
            page_id="af-card",
            title="Antifraud V3.47 - ID Credit Card",
            source_url="https://confluence.example/display/SPDB/Antifraud+V3.47_0702+-+ID+Credit+Card",
            updated_at="2026-05-12T17:56:27.000+08:00",
            language="en",
            sections=[
                ParsedSection(
                    title="Admin spend",
                    section_path="3.1.16 New Scenario: AdminPortalUpdateCCardSpend",
                    content="CMS calls `AdminPortalUpdateCCardSpend` for the admin portal credit card spend scenario.",
                )
            ],
            ancestor_titles=["Productization PRD - Authentication & Antifraud"],
        )
        service = PRDReviewService(
            store=self.store,
            confluence=FakeConnector(page),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )
        mock_generate.return_value = {
            "result_markdown": """### Executive Verdict
- **Final Score:** 5 / 10

### Secondary Clarifications
- **Priority:** P2
  - **Section:** Section 3.1.16 New Scenario: A dminPortalUpdateCCardSpend
  - **Clarification:** Fix the section title typo after scope is confirmed, so the scenario name is stable for QA traceability.
- **Priority:** P2
  - **Section:** Section 3.1.16 New Scenario: AdminPortalUpdateCCardSpend
  - **Clarification:** Confirm whether maker-checker is required.

### PM Decision Checklist
- Confirm owner.
""",
            "model_id": "codex-cli",
            "trace": {},
        }

        result = service.review_url(
            PRDBriefingReviewRequest(owner_key="anon:test", prd_url=page.source_url, language="en")
        )

        markdown = result["review"]["result_markdown"]
        self.assertNotIn("A dminPortalUpdateCCardSpend", markdown)
        self.assertNotIn("Fix the section title typo", markdown)
        self.assertIn("Confirm whether maker-checker is required", markdown)

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_review_keeps_identifier_typo_when_exact_typo_exists_in_prd(self, mock_generate):
        page = IngestedConfluencePage(
            page_id="af-card-typo",
            title="Antifraud V3.47 - ID Credit Card",
            source_url="https://confluence.example/display/SPDB/Antifraud+V3.47_0702+-+ID+Credit+Card",
            updated_at="2026-05-12T17:56:27.000+08:00",
            language="en",
            sections=[
                ParsedSection(
                    title="Admin spend",
                    section_path="3.1.16 New Scenario: A dminPortalUpdateCCardSpend",
                    content="CMS calls `A dminPortalUpdateCCardSpend` for the admin portal credit card spend scenario.",
                )
            ],
            ancestor_titles=["Productization PRD - Authentication & Antifraud"],
        )
        service = PRDReviewService(
            store=self.store,
            confluence=FakeConnector(page),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )
        mock_generate.return_value = {
            "result_markdown": """### Executive Verdict
- **Final Score:** 5 / 10

### Secondary Clarifications
- **Priority:** P2
  - **Section:** Section 3.1.16 New Scenario: A dminPortalUpdateCCardSpend
  - **Clarification:** Fix the section title typo because the reviewed PRD text contains the exact typo `A dminPortalUpdateCCardSpend`.
""",
            "model_id": "codex-cli",
            "trace": {},
        }

        result = service.review_url(
            PRDBriefingReviewRequest(owner_key="anon:test", prd_url=page.source_url, language="en")
        )

        self.assertIn("A dminPortalUpdateCCardSpend", result["review"]["result_markdown"])

    def test_prd_summary_english_prompt_requests_english_output(self):
        page = self.service.confluence.page

        prompt = build_prd_summary_prompt(
            jira_id="",
            jira_link="",
            prd_url=page.source_url,
            page=page,
            language="en",
        )

        self.assertIn("concise but complete English summary", prompt)
        self.assertIn("Jira ID: -", prompt)
        self.assertIn("PRD Summary", prompt)
        self.assertNotIn("中文摘要", prompt)

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_briefing_review_cache_varies_by_language_and_updated_at(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        service = PRDReviewService(
            store=self.store,
            confluence=self.service.confluence,
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        first = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="zh",
            )
        )
        cached = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="zh",
            )
        )
        english = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
            )
        )

        self.assertFalse(first["cached"])
        self.assertTrue(cached["cached"])
        self.assertFalse(english["cached"])
        self.assertEqual(mock_generate.call_count, 2)
        self.assertEqual(first["review"]["jira_id"], PRD_BRIEFING_REVIEW_CACHE_KEY)
        self.assertEqual(first["review"]["prompt_version"], prd_briefing_review_prompt_version("zh"))
        self.assertEqual(english["review"]["prompt_version"], prd_briefing_review_prompt_version("en"))

        stale = self.store.get_prd_review_result(
            owner_key="anon:test",
            jira_id=PRD_BRIEFING_REVIEW_CACHE_KEY,
            prd_url=self.service.confluence.page.source_url,
            prd_updated_at="2026-04-16T10:00:00Z",
            prompt_version=prd_briefing_review_prompt_version("zh"),
        )
        self.assertIsNone(stale)

    def test_prd_self_assessment_lists_sections_in_source_order(self):
        service = PRDReviewService(
            store=self.store,
            confluence=self.service.confluence,
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        payload = service.list_url_sections(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
            )
        )

        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["prd"]["title"], "Payments PRD")
        self.assertEqual([section["index"] for section in payload["sections"]], [1, 2])
        self.assertEqual(payload["sections"][0]["title"], "Overview")
        self.assertGreater(payload["sections"][0]["char_count"], 0)
        self.assertFalse(payload["sections"][0]["long"])

    def test_prd_self_assessment_lists_spreadsheet_link_counts(self):
        page = self.service.confluence.page
        page.sections[1].spreadsheet_links = [
            SpreadsheetLink(
                title="MAS Outsourcing Register.xlsx",
                url="https://example.atlassian.net/download/attachments/123/MAS%20Outsourcing%20Register.xlsx",
                source_section="Rollout",
                filename="MAS Outsourcing Register.xlsx",
            )
        ]
        service = PRDReviewService(
            store=self.store,
            confluence=self.service.confluence,
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        payload = service.list_url_sections(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
            )
        )

        self.assertEqual(payload["sections"][0]["linked_spreadsheet_count"], 0)
        self.assertEqual(payload["sections"][1]["linked_spreadsheet_count"], 1)

    def test_prd_reviewer_normalization_and_empty_section_boundaries(self):
        empty_page = IngestedConfluencePage(
            page_id="empty",
            title="Empty PRD",
            source_url="https://example.atlassian.net/wiki/pages/empty",
            updated_at="2026-05-20T00:00:00Z",
            language="en",
            sections=[],
        )
        service = PRDReviewService(
            store=self.store,
            confluence=FakeConnector(empty_page),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        with self.assertRaisesRegex(ToolError, "Owner identity"):
            PRDReviewService._normalize_request(
                PRDReviewRequest(owner_key="", jira_id="AF-1", jira_link="", prd_url="https://example/prd")
            )
        with self.assertRaisesRegex(ToolError, "Jira ID"):
            PRDReviewService._normalize_request(
                PRDReviewRequest(owner_key="anon:test", jira_id="", jira_link="", prd_url="https://example/prd")
            )
        with self.assertRaisesRegex(ToolError, "PRD link is required"):
            PRDReviewService._normalize_request(
                PRDReviewRequest(owner_key="anon:test", jira_id="AF-1", jira_link="", prd_url="")
            )
        with self.assertRaisesRegex(ToolError, "HTTP or HTTPS"):
            PRDReviewService._normalize_request(
                PRDReviewRequest(owner_key="anon:test", jira_id="AF-1", jira_link="", prd_url="file:///tmp/prd")
            )
        normalized = PRDReviewService._normalize_request(
            PRDReviewRequest(
                owner_key=" anon:test ",
                jira_id=" AF-1 ",
                jira_link=" https://jira/browse/AF-1 ",
                prd_url=" https://example/prd ",
                force_refresh=1,
                google_credentials={"token": "x"},
            )
        )
        with self.assertRaisesRegex(ToolError, "Owner identity"):
            PRDReviewService._normalize_briefing_review_request(
                PRDBriefingReviewRequest(owner_key="", prd_url="https://example/prd")
            )
        with self.assertRaisesRegex(ToolError, "PRD link is required"):
            PRDReviewService._normalize_briefing_review_request(
                PRDBriefingReviewRequest(owner_key="anon:test", prd_url="")
            )
        with self.assertRaisesRegex(ToolError, "HTTP or HTTPS"):
            PRDReviewService._normalize_briefing_review_request(
                PRDBriefingReviewRequest(owner_key="anon:test", prd_url="ftp://example/prd")
            )
        with self.assertRaisesRegex(ToolError, "readable sections"):
            service.list_url_sections(
                PRDBriefingReviewRequest(owner_key="anon:test", prd_url="https://example.atlassian.net/wiki/pages/empty")
            )

        self.assertEqual(normalized.owner_key, "anon:test")
        self.assertEqual(normalized.jira_id, "AF-1")
        self.assertTrue(normalized.force_refresh)
        self.assertEqual(normalized.google_credentials, {"token": "x"})

    def test_prd_reviewer_source_table_and_section_helper_boundaries(self):
        empty_media_page = IngestedConfluencePage(
            page_id="media",
            title="Media PRD",
            source_url="https://example.atlassian.net/wiki/pages/media",
            updated_at="2026-05-20T00:00:00Z",
            language="en",
            sections=[
                ParsedSection(
                    title="Table",
                    section_path="3.1 Scenario ApplyCard",
                    content="See [MEDIA_ID_1]\nDetached: [MEDIA_ID_2]",
                    media_refs=["", "MEDIA_ID_1", "MEDIA_ID_2", "MEDIA_ID_3"],
                )
            ],
            media_dict={
                "MEDIA_ID_1": {"type": "table", "content": "<table><tr><td></td></tr></table>"},
                "MEDIA_ID_2": {"type": "table", "content": "<table><tr><td>Fallback only</td></tr></table>"},
                "MEDIA_ID_3": {"type": "image", "url": "https://example/image.png"},
            },
        )
        plain_page = IngestedConfluencePage(
            page_id="plain",
            title="Plain PRD",
            source_url="https://example.atlassian.net/wiki/pages/plain",
            updated_at="2026-05-20T00:00:00Z",
            language="en",
            sections=[ParsedSection(title="Blank", section_path="Blank", content="")],
        )
        long_page = IngestedConfluencePage(
            page_id="long",
            title="Long PRD",
            source_url="https://example.atlassian.net/wiki/pages/long",
            updated_at="2026-05-20T00:00:00Z",
            language="en",
            sections=[
                ParsedSection(title="One", section_path="One", content="a" * 20),
                ParsedSection(title="Two", section_path="Two", content="b" * 20),
            ],
        )

        content = _build_review_section_content(page=empty_media_page, section=empty_media_page.sections[0])
        coverage_tables = _collect_table_media_coverage(empty_media_page)
        with self.assertRaisesRegex(ToolError, "readable text"):
            _build_prd_source_payload(plain_page)
        truncated_payload = _build_prd_source_payload(long_page, max_chars=25)
        section_coverage = _build_section_coverage(
            total_sections=2,
            selected_indexes=[1, 2],
            selected_sections=long_page.sections,
            selection_hash="hash",
        )

        self.assertIn("Fallback only", content)
        self.assertEqual(coverage_tables[0]["media_id"], "MEDIA_ID_2")
        self.assertTrue(truncated_payload["truncated"])
        self.assertEqual(_generation_mode_for_page(long_page), "single")
        self.assertEqual(_build_generation_coverage(long_page, mode="hybrid")["mode"], "hybrid")
        self.assertEqual(section_coverage["status"], "full")
        self.assertEqual(_section_selection_hash([2, 1]), _section_selection_hash([2, 1]))
        cache_key = _cache_key_for_section_selection(
            "prefix",
            page=long_page,
            selected_section_indexes=[1],
            linked_artifact_fingerprint="linked",
            google_sheet_screenshot_fingerprint="shot",
        )
        self.assertIn(":sections:", cache_key)
        self.assertIn(":linked:", cache_key)
        self.assertIn(":sheetshots:", cache_key)

    def test_prd_reviewer_table_postprocess_and_batch_helper_tail_boundaries(self):
        page = IngestedConfluencePage(
            page_id="table-tail",
            title="AF Delivery PRD",
            source_url="https://example.atlassian.net/wiki/pages/table-tail",
            updated_at="2026-05-20T00:00:00Z",
            language="en",
            sections=[
                ParsedSection(
                    title="Scenario ApplyCard",
                    section_path="3.1 Scenario ApplyCard",
                    content="ApplyCard owner handles state ApplyCardStatus1.",
                    media_refs=["MEDIA_TABLE", "MEDIA_TABLE", "MEDIA_IMAGE"],
                    image_refs=["https://example.com/sheet.png", "https://example.com/sheet.png", ""],
                )
            ],
            media_dict={
                "MEDIA_TABLE": {
                    "type": "table",
                    "content": "<table><tr><th>Scenario</th><th>Action</th></tr><tr><td>ApplyCard</td><td>Approve</td></tr></table>",
                },
                "MEDIA_IMAGE": {"type": "image", "source_url": "https://example.com/media.png"},
                "MEDIA_OTHER": {"type": "panel", "content": "not table"},
            },
        )
        context = reviewer._new_table_pack_context()
        first_table = reviewer._format_table_media_for_review("MEDIA_TABLE", page.media_dict["MEDIA_TABLE"], section=page.sections[0], table_context=context)
        cached_table = reviewer._format_table_media_for_review("MEDIA_TABLE", page.media_dict["MEDIA_TABLE"], section=page.sections[0], table_context=context)
        empty_rows = reviewer._extract_table_media_rows({"content": ""})
        empty_fallback_rows = reviewer._extract_table_media_rows({"content": "<div>   </div>"})
        budget_omitted = reviewer._pack_table_rows_for_review(media_id="T1", rows=["Field | Rule"], section=None, remaining_chars=0)
        tiny_budget = reviewer._pack_table_rows_for_review(
            media_id="T2",
            rows=["x" * 120, "Scenario ApplyCard | Approve"],
            section=page.sections[0],
            remaining_chars=160,
        )
        no_identifier_terms = reviewer._section_identifier_terms(None)
        typo_without_priority = reviewer._is_spurious_identifier_typo_finding(
            "Fix typo for ApplyCardStatus1",
            source_text="ApplyCardStatus1",
            source_identifier_text=reviewer._normalize_identifier_for_comparison("ApplyCardStatus1"),
        )
        typo_from_normalized_identifier = reviewer._is_spurious_identifier_typo_finding(
            "- **Priority:** P2\n  - **Clarification:** Fix typo for Apply Card Status1",
            source_text="ApplyCardStatus1",
            source_identifier_text=reviewer._normalize_identifier_for_comparison("ApplyCardStatus1"),
        )
        empty_block = reviewer._build_prd_source_block(ParsedSection(title="Blank", section_path="Blank", content=""), 1)
        coverage = reviewer._table_stats_to_coverage(
            [
                {"media_id": "", "rows_total": 1},
                {"media_id": "MEDIA_OTHER", "rows_total": 1},
                {"media_id": "MEDIA_TABLE", "rows_total": 2, "rows_included": 1, "rows_omitted": 1, "truncated": True, "char_count": 99},
            ],
            page=page,
        )
        with self.assertRaisesRegex(ToolError, "readable text"):
            reviewer._split_prd_page_batches(
                IngestedConfluencePage(
                    page_id="empty-batch",
                    title="Empty",
                    source_url="https://example.atlassian.net/wiki/pages/empty-batch",
                    updated_at="2026-05-20T00:00:00Z",
                    language="en",
                    sections=[],
                )
            )
        artifacts = reviewer._collect_section_image_artifacts(page=page, selected_section_indexes=[3])
        postprocessed_without_source = reviewer._postprocess_prd_review_markdown(
            "### Secondary Clarifications\n- **Priority:** P2\n  - **Clarification:** typo\n",
            page=IngestedConfluencePage(
                page_id="blank",
                title="",
                source_url="https://example.atlassian.net/wiki/pages/blank",
                updated_at="",
                language="en",
                sections=[],
            ),
        )

        self.assertEqual(first_table, cached_table)
        self.assertEqual(empty_rows, [])
        self.assertEqual(empty_fallback_rows, [])
        self.assertEqual(budget_omitted["rows_omitted"], 1)
        self.assertIn("row truncated", tiny_budget["text"])
        self.assertEqual(no_identifier_terms, set())
        self.assertFalse(typo_without_priority)
        self.assertTrue(typo_from_normalized_identifier)
        self.assertEqual(empty_block, "")
        self.assertEqual(coverage[0]["media_id"], "MEDIA_TABLE")
        self.assertEqual(len(artifacts), 2)
        self.assertEqual(artifacts[0]["image_id"], "MEDIA_IMAGE")
        self.assertEqual(artifacts[1]["image_id"], "IMAGE_3_1")
        self.assertTrue(reviewer._contains_anti_fraud_marker(" AF "))
        self.assertEqual(postprocessed_without_source.splitlines()[0], "### Secondary Clarifications")

    def test_prd_reviewer_misc_parsing_and_linked_evidence_boundaries(self):
        report_page = IngestedConfluencePage(
            page_id="report",
            title="[AF] Report PRD",
            source_url="https://example.atlassian.net/wiki/pages/report",
            updated_at="2026-05-20T00:00:00Z",
            language="en",
            ancestor_titles=["2.2 Authentication & Antifraud"],
            sections=[
                ParsedSection(
                    title="Report Layout",
                    section_path="3.11.1 Report Layout and Field Name & Format",
                    content="Report format section.",
                    spreadsheet_links=[
                        SpreadsheetLink(
                            title="Unsupported",
                            url="https://example.atlassian.net/download/attachments/123/Legacy.xls",
                            source_section="Report Layout",
                            filename="Legacy.xls",
                        ),
                        SpreadsheetLink(
                            title="Unsupported duplicate",
                            url="https://example.atlassian.net/download/attachments/123/Legacy.xls",
                            source_section="Report Layout",
                            filename="Legacy.xls",
                        ),
                        SpreadsheetLink(
                            title="Other",
                            url="https://example.com/template.csv",
                            source_section="Report Layout",
                            filename="template.csv",
                        ),
                    ],
                )
            ],
        )
        evidence_disabled = _resolve_linked_spreadsheet_evidence(
            confluence=FakeConnector(report_page),
            page=report_page,
            selected_section_indexes=[1],
            google_credentials=None,
            include_linked_spreadsheets=False,
        )
        evidence = _resolve_linked_spreadsheet_evidence(
            confluence=FakeConnector(report_page),
            page=report_page,
            selected_section_indexes=[1],
            google_credentials=None,
            include_linked_spreadsheets=True,
        )
        progress_calls = []

        def old_progress(stage, message, current, total):
            progress_calls.append((stage, message, current, total))

        _emit_prd_progress(old_progress, "stage", "message", 1, 2, estimated_prompt_tokens=10, token_risk="normal")

        self.assertEqual(evidence_disabled, {"artifacts": [], "cache_fingerprint": ""})
        self.assertEqual(len(evidence["artifacts"]), 2)
        self.assertEqual(evidence["artifacts"][0]["reason"], "unsupported_xls_format")
        self.assertEqual(evidence["artifacts"][1]["reason"], "unsupported_link")
        self.assertIn("Linked artifact not reviewed", _build_linked_spreadsheet_prompt_section(evidence))
        self.assertIn("No report-template", _build_linked_spreadsheet_prompt_section({"artifacts": []}))
        self.assertEqual(progress_calls, [("stage", "message", 1, 2)])
        self.assertTrue(_is_anti_fraud_prd(report_page))
        self.assertTrue(_is_report_or_template_prd(report_page))
        self.assertTrue(_is_google_spreadsheet_url("https://docs.google.com/spreadsheets/d/sheet123/edit"))
        self.assertTrue(_is_google_spreadsheet_url("https://drive.google.com/open?id=sheet123"))
        self.assertFalse(_is_google_spreadsheet_url("https://example.com/spreadsheets/d/sheet123"))
        self.assertTrue(_is_confluence_spreadsheet_url("https://example.atlassian.net/download/attachments/123/Template.xlsx"))
        self.assertFalse(_is_confluence_spreadsheet_url("https://example.atlassian.net/download/attachments/123/Template.csv"))
        self.assertEqual(_google_drive_file_id_from_url("https://docs.google.com/spreadsheets/d/sheet123/edit"), "sheet123")
        self.assertEqual(_google_drive_file_id_from_url("https://drive.google.com/open?id=sheet456"), "sheet456")
        self.assertEqual(_google_drive_file_id_from_url("https://drive.google.com/open"), "")
        self.assertEqual(_image_suffix_from_url("https://example.com/a%20b.JPG?x=1"), ".jpg")
        self.assertEqual(_image_suffix_from_url("https://example.com/image"), ".png")
        self.assertEqual(_parse_json_object("```json\n{\"ok\": true}\n```"), {"ok": True})
        self.assertEqual(_parse_json_object("prefix {\"ok\": 1} suffix"), {"ok": 1})
        self.assertIsNone(_parse_json_object("[1,2]"))
        self.assertEqual(_prd_token_risk(10), "normal")
        self.assertEqual(_prd_token_risk(45_000), "elevated")
        self.assertEqual(_prd_token_risk(80_000), "high")
        self.assertEqual(_normalize_selected_section_indexes([1, "2", 2]), [1, 2])
        with self.assertRaisesRegex(ToolError, "integer"):
            _normalize_selected_section_indexes([True])
        with self.assertRaisesRegex(ToolError, "integer"):
            _normalize_selected_section_indexes(["bad"])
        self.assertEqual(_limit_extracted_workbook_text({"text": "abcdef"}, max_chars=3)["text"], "abc")
        self.assertIn("omitted", _limit_extracted_workbook_text({"text": "abcdef"}, max_chars=0)["text"])
        self.assertEqual(_coerce_download_bytes("hello"), b"hello")
        self.assertEqual(_coerce_download_bytes(bytearray(b"hello")), b"hello")
        self.assertEqual(_linked_artifact_public_payload(evidence["artifacts"][0])["status"], "failed")
        self.assertFalse(
            _google_sheet_screenshot_public_payload({"status": "ok", "used_in_findings": True})["cache_hit"]
        )

    def test_prd_reviewer_google_sheet_screenshot_helper_boundaries(self):
        page = IngestedConfluencePage(
            page_id="image-prd",
            title="AF Sheet Screenshot PRD",
            source_url="https://example.atlassian.net/wiki/pages/image-prd",
            updated_at="2026-05-20T00:00:00Z",
            language="en",
            sections=[
                ParsedSection(
                    title="Report Image",
                    section_path="3.11 Report Image",
                    content="See screenshots.",
                    media_refs=["MEDIA_ID_1", "MEDIA_ID_2", "MEDIA_ID_3"],
                    image_refs=["https://example.atlassian.net/download/attachments/123/direct.png", ""],
                )
            ],
            media_dict={
                "MEDIA_ID_1": {"type": "image", "source_url": "https://example.atlassian.net/download/attachments/123/sheet.png"},
                "MEDIA_ID_2": {"type": "image", "source_url": "https://example.atlassian.net/download/attachments/123/sheet.png"},
                "MEDIA_ID_3": {"type": "table", "content": "<table><tr><td>Not image</td></tr></table>"},
            },
        )
        artifacts = _collect_section_image_artifacts(page=page, selected_section_indexes=[1])
        settings = self._settings()
        cache_dir = Path(self.temp_dir.name) / "sheetshot-cache"
        ok_artifact = dict(artifacts[0])
        skipped_artifact = dict(artifacts[0])
        empty_artifact = dict(artifacts[0])
        tool_error_artifact = dict(artifacts[0])
        unexpected_error_artifact = dict(artifacts[0])

        with patch(
            "prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image",
            return_value={
                "is_google_sheet_screenshot": True,
                "reason": "",
                "classification": "visible spreadsheet grid",
                "evidence_text": "Sheet headers: Field, Rule",
            },
        ) as extract:
            _resolve_one_google_sheet_screenshot(
                confluence=FakeAttachmentConnector(page, b"image-bytes"),
                artifact=ok_artifact,
                cache_dir=cache_dir,
                workspace_root=Path(self.temp_dir.name),
                settings=settings,
            )
        cached_artifact = dict(artifacts[0])
        cached_bytes = b"cached-image"
        cached_hash = hashlib.sha256(cached_bytes).hexdigest()[:16]
        _store_cached_google_sheet_screenshot_evidence(
            url=cached_artifact["url"],
            content_hash=cached_hash,
            cache_dir=cache_dir,
            extracted={
                "status": "ok",
                "reason": "",
                "is_google_sheet_screenshot": True,
                "evidence_text": "Cached sheet evidence",
                "evidence_hash": "cached-hash",
            },
        )
        cached_resolve_artifact = dict(artifacts[0])
        with patch(
            "prd_briefing.reviewer._load_cached_google_sheet_screenshot_evidence",
            return_value={"status": "ok", "evidence_text": "Cached sheet evidence"},
        ):
            _resolve_one_google_sheet_screenshot(
                confluence=FakeAttachmentConnector(page, cached_bytes),
                artifact=cached_resolve_artifact,
                cache_dir=cache_dir,
                workspace_root=Path(self.temp_dir.name),
                settings=settings,
            )
        cached_payload = _load_cached_google_sheet_screenshot_evidence(
            url=cached_artifact["url"],
            content_hash=cached_hash,
            cache_dir=cache_dir,
        )

        with patch(
            "prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image",
            return_value={
                "is_google_sheet_screenshot": False,
                "reason": "normal_product_screenshot",
                "classification": "not spreadsheet",
                "evidence_text": "ignored",
            },
        ):
            _resolve_one_google_sheet_screenshot(
                confluence=FakeAttachmentConnector(page, b"other-image"),
                artifact=skipped_artifact,
                cache_dir=None,
                workspace_root=Path(self.temp_dir.name),
                settings=settings,
            )
        _resolve_one_google_sheet_screenshot(
            confluence=FakeAttachmentConnector(page, b""),
            artifact=empty_artifact,
            cache_dir=None,
            workspace_root=Path(self.temp_dir.name),
            settings=settings,
        )
        with patch(
            "prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image",
            side_effect=ToolError("ocr_failed"),
        ):
            _resolve_one_google_sheet_screenshot(
                confluence=FakeAttachmentConnector(page, b"tool-error"),
                artifact=tool_error_artifact,
                cache_dir=None,
                workspace_root=Path(self.temp_dir.name),
                settings=settings,
            )
        with patch(
            "prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image",
            side_effect=RuntimeError("boom"),
        ):
            _resolve_one_google_sheet_screenshot(
                confluence=FakeAttachmentConnector(page, b"unexpected-error"),
                artifact=unexpected_error_artifact,
                cache_dir=None,
                workspace_root=Path(self.temp_dir.name),
                settings=settings,
            )

        key = _google_sheet_screenshot_cache_key(url=ok_artifact["url"], content_hash=ok_artifact["content_hash"])
        bad_cache_path = cache_dir / f"{key}.json"
        bad_cache_path.write_text("{bad-json", encoding="utf-8")
        invalid_cache = _load_cached_google_sheet_screenshot_evidence(
            url=ok_artifact["url"],
            content_hash=ok_artifact["content_hash"],
            cache_dir=cache_dir,
        )
        _store_cached_google_sheet_screenshot_evidence(
            url="https://example/none.png",
            content_hash="hash",
            cache_dir=None,
            extracted={"status": "ok"},
        )
        prompt_section = _build_google_sheet_screenshot_prompt_section(
            {
                "artifacts": [
                    ok_artifact,
                    skipped_artifact,
                    {"status": "failed", "image_id": "IMG_BAD", "source_section_title": "Report", "url": "https://example/bad.png", "reason": ""},
                ]
            }
        )
        metadata_text = _format_template_metadata_for_prompt(
            {
                "workbook_sheet_count": 2,
                "hidden_sheets": ["Hidden"],
                "sheets": [
                    {
                        "name": "Register",
                        "max_row": 10,
                        "max_column": 4,
                        "header_row": ["Field", ""],
                        "merged_range_count": 1,
                        "formula_cell_count": 2,
                        "hidden_rows_count": 1,
                        "hidden_columns_count": 1,
                        "empty_header_count": 1,
                        "duplicate_headers": ["Field"],
                        "multi_level_header_risk": True,
                        "wide_template_risk": False,
                    },
                    "bad-sheet",
                ],
            }
        )

        self.assertEqual(len(artifacts), 2)
        self.assertEqual(ok_artifact["status"], "ok")
        self.assertTrue(cached_resolve_artifact["cache_hit"])
        self.assertEqual(cached_payload["evidence_text"], "Cached sheet evidence")
        self.assertEqual(extract.call_count, 1)
        self.assertEqual(skipped_artifact["status"], "skipped")
        self.assertEqual(empty_artifact["reason"], "empty_download")
        self.assertEqual(tool_error_artifact["reason"], "ocr_failed")
        self.assertIn("ocr_failed: boom", unexpected_error_artifact["reason"])
        self.assertIsNone(invalid_cache)
        self.assertIn("Google Sheet screenshot reviewed", prompt_section)
        self.assertIn("Google Sheet screenshot not reviewed", prompt_section)
        self.assertEqual(
            _build_google_sheet_screenshot_prompt_section(
                {"artifacts": [{"status": "skipped", "reason": "not_google_sheet_screenshot"}]}
            ),
            "",
        )
        self.assertIn("formula_cells=2", metadata_text)
        self.assertEqual(_format_template_metadata_for_prompt({}), "- Metadata unavailable.")

    def test_prd_reviewer_screenshot_cache_and_linked_spreadsheet_failure_tail_boundaries(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            cache_dir = Path(temp_dir) / "sheetshot-cache"
            url = "https://example.com/sheet.png"
            content_hash = "hash-1"
            key = _google_sheet_screenshot_cache_key(url=url, content_hash=content_hash)
            cache_dir.mkdir()
            cache_path = cache_dir / f"{key}.json"
            cache_path.write_text(
                json.dumps(
                    {
                        "version": "old",
                        "url": url,
                        "content_hash": content_hash,
                        "extracted": {"status": "ok"},
                    }
                ),
                encoding="utf-8",
            )
            stale_cache = _load_cached_google_sheet_screenshot_evidence(
                url=url,
                content_hash=content_hash,
                cache_dir=cache_dir,
            )
            cache_path.write_text(
                json.dumps(
                    {
                        "version": reviewer.PRD_GOOGLE_SHEET_SCREENSHOT_EVIDENCE_CACHE_VERSION,
                        "url": "https://example.com/other.png",
                        "content_hash": content_hash,
                        "extracted": {"status": "ok"},
                    }
                ),
                encoding="utf-8",
            )
            mismatched_cache = _load_cached_google_sheet_screenshot_evidence(
                url=url,
                content_hash=content_hash,
                cache_dir=cache_dir,
            )
            with patch.object(Path, "write_text", side_effect=OSError("disk full")):
                _store_cached_google_sheet_screenshot_evidence(
                    url=url,
                    content_hash=content_hash,
                    cache_dir=cache_dir,
                    extracted={"status": "ok"},
                )

            artifact_empty = {
                "url": "https://example.atlassian.net/download/attachments/123/Template.xlsx",
                "filename": "Template.xlsx",
            }
            reviewer._resolve_one_linked_spreadsheet(
                confluence=FakeAttachmentConnector(self.service.confluence.page, b""),
                artifact=artifact_empty,
                google_credentials=None,
                remaining_text_chars=100,
            )
            artifact_tool_error = {
                "url": "https://example.atlassian.net/download/attachments/123/Template.xlsx",
                "filename": "Template.xlsx",
            }
            with patch("prd_briefing.reviewer._extract_workbook_text", side_effect=ToolError("empty_workbook")):
                reviewer._resolve_one_linked_spreadsheet(
                    confluence=FakeAttachmentConnector(self.service.confluence.page, b"xlsx"),
                    artifact=artifact_tool_error,
                    google_credentials=None,
                    remaining_text_chars=100,
                )
            artifact_generic_error = {
                "url": "https://example.atlassian.net/download/attachments/123/Template.xlsx",
                "filename": "Template.xlsx",
            }
            with patch("prd_briefing.reviewer._extract_workbook_text", side_effect=ValueError("bad workbook")):
                reviewer._resolve_one_linked_spreadsheet(
                    confluence=FakeAttachmentConnector(self.service.confluence.page, b"xlsx"),
                    artifact=artifact_generic_error,
                    google_credentials=None,
                    remaining_text_chars=100,
                )

        self.assertIsNone(stale_cache)
        self.assertIsNone(mismatched_cache)
        self.assertEqual(artifact_empty["reason"], "empty_download")
        self.assertEqual(artifact_tool_error["reason"], "empty_workbook")
        self.assertIn("download_or_parse_failed", artifact_generic_error["reason"])

    def test_prd_reviewer_cache_workbook_and_wrapper_tail_boundaries(self):
        import builtins

        url = "https://docs.google.com/spreadsheets/d/sheet123/edit#gid=0"
        credentials = {"token": "x", "scopes": ["https://www.googleapis.com/auth/drive.readonly"]}
        metadata = reviewer._normalize_drive_metadata(
            {
                "id": "sheet123",
                "name": "MAS Outsourcing Register",
                "mimeType": "application/vnd.google-apps.spreadsheet",
                "modifiedTime": "2026-05-12T10:00:00.000Z",
            }
        )
        cache_dir = Path(self.temp_dir.name) / "google-sheet-cache"
        cache_dir.mkdir()
        cache_path = cache_dir / f"{reviewer._google_sheet_artifact_cache_key(url=url, drive_metadata=metadata)}.json"
        drive_service = FakeDriveService(_xlsx_bytes(marker="Cached"), modified_time=metadata["modifiedTime"])

        no_credentials = reviewer._load_cached_google_spreadsheet_artifact(
            url=url,
            google_credentials=None,
            cache_dir=cache_dir,
            max_chars=10,
        )
        no_cache_dir = reviewer._load_cached_google_spreadsheet_artifact(
            url=url,
            google_credentials=credentials,
            cache_dir=None,
            max_chars=10,
        )
        with self.assertRaisesRegex(ToolError, "missing_google_credentials"):
            reviewer._download_google_spreadsheet(url=url, google_credentials=None)
        with self.assertRaisesRegex(ToolError, "google_file_id_not_found"):
            reviewer._download_google_spreadsheet(
                url="https://drive.google.com/open",
                google_credentials=credentials,
            )
        with self.assertRaisesRegex(ToolError, "google_file_id_not_found"):
            reviewer._load_cached_google_spreadsheet_artifact(
                url="https://drive.google.com/open",
                google_credentials=credentials,
                cache_dir=cache_dir,
                max_chars=100,
            )

        real_import = builtins.__import__

        def missing_google_import(name, *args, **kwargs):
            if name == "google.oauth2.credentials":
                raise ImportError("missing google")
            return real_import(name, *args, **kwargs)

        with patch("builtins.__import__", side_effect=missing_google_import):
            with self.assertRaisesRegex(ToolError, "google_drive_reader_unavailable"):
                reviewer._download_google_spreadsheet(url=url, google_credentials=credentials)
        with patch("builtins.__import__", side_effect=missing_google_import):
            with self.assertRaisesRegex(ToolError, "google_drive_reader_unavailable"):
                reviewer._load_cached_google_spreadsheet_artifact(
                    url=url,
                    google_credentials=credentials,
                    cache_dir=cache_dir,
                    max_chars=100,
                )

        native_drive_service = FakeDriveService(_xlsx_bytes(marker="Native XLSX"), modified_time=metadata["modifiedTime"])
        native_drive_service.files_resource.get = lambda **_kwargs: FakeDriveExecute(
            {
                "id": "sheet123",
                "name": "MAS Outsourcing Register.xlsx",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "modifiedTime": metadata["modifiedTime"],
            }
        )
        with patch("bpmis_jira_tool.gmail_dashboard.build_drive_api_service", return_value=native_drive_service):
            native_content, native_metadata = reviewer._download_google_spreadsheet(
                url=url,
                google_credentials=credentials,
            )
        unsupported_drive_service = FakeDriveService(_xlsx_bytes(marker="Unsupported"), modified_time=metadata["modifiedTime"])
        unsupported_drive_service.files_resource.get = lambda **_kwargs: FakeDriveExecute(
            {
                "id": "sheet123",
                "name": "Unsupported",
                "mimeType": "application/pdf",
                "modifiedTime": metadata["modifiedTime"],
            }
        )
        with patch("bpmis_jira_tool.gmail_dashboard.build_drive_api_service", return_value=unsupported_drive_service):
            with self.assertRaisesRegex(ToolError, "unsupported_google_drive_mime_type"):
                reviewer._download_google_spreadsheet(url=url, google_credentials=credentials)

        from googleapiclient.errors import HttpError

        class FakeHttpResponse:
            def __init__(self, status):
                self.status = status
                self.reason = "failed"

        class RaisingExecute:
            def __init__(self, status):
                self.status = status

            def execute(self):
                raise HttpError(resp=FakeHttpResponse(self.status), content=b"failed")

        class RaisingDriveFiles:
            def __init__(self, status):
                self.status = status

            def get(self, **_kwargs):
                return RaisingExecute(self.status)

        class RaisingDriveService:
            def __init__(self, status):
                self.files_resource = RaisingDriveFiles(status)

            def files(self):
                return self.files_resource

        with patch("bpmis_jira_tool.gmail_dashboard.build_drive_api_service", return_value=RaisingDriveService(403)):
            with self.assertRaisesRegex(ToolError, "permission_denied"):
                reviewer._download_google_spreadsheet(url=url, google_credentials=credentials)
        with patch("bpmis_jira_tool.gmail_dashboard.build_drive_api_service", return_value=RaisingDriveService(500)):
            with self.assertRaisesRegex(ToolError, "google_drive_download_failed:500"):
                reviewer._download_google_spreadsheet(url=url, google_credentials=credentials)
        with patch("bpmis_jira_tool.gmail_dashboard.build_drive_api_service", return_value=RaisingDriveService(404)):
            with self.assertRaisesRegex(ToolError, "permission_denied"):
                reviewer._load_cached_google_spreadsheet_artifact(
                    url=url,
                    google_credentials=credentials,
                    cache_dir=cache_dir,
                    max_chars=100,
                )
        with patch("bpmis_jira_tool.gmail_dashboard.build_drive_api_service", return_value=RaisingDriveService(503)):
            with self.assertRaisesRegex(ToolError, "google_drive_metadata_failed:503"):
                reviewer._load_cached_google_spreadsheet_artifact(
                    url=url,
                    google_credentials=credentials,
                    cache_dir=cache_dir,
                    max_chars=100,
                )
        with patch("bpmis_jira_tool.gmail_dashboard.build_drive_api_service", return_value=drive_service):
            missing_cache = reviewer._load_cached_google_spreadsheet_artifact(
                url=url,
                google_credentials=credentials,
                cache_dir=cache_dir,
                max_chars=100,
            )
            cache_path.write_text("{bad-json", encoding="utf-8")
            invalid_json_cache = reviewer._load_cached_google_spreadsheet_artifact(
                url=url,
                google_credentials=credentials,
                cache_dir=cache_dir,
                max_chars=100,
            )
            cache_path.write_text(
                json.dumps({"version": "old", "drive_metadata": metadata, "extracted": {"text": "cached"}}),
                encoding="utf-8",
            )
            stale_version_cache = reviewer._load_cached_google_spreadsheet_artifact(
                url=url,
                google_credentials=credentials,
                cache_dir=cache_dir,
                max_chars=100,
            )
            cache_path.write_text(
                json.dumps(
                    {
                        "version": reviewer.PRD_GOOGLE_SHEET_ARTIFACT_CACHE_VERSION,
                        "drive_metadata": {**metadata, "modifiedTime": "older"},
                        "extracted": {"text": "cached"},
                    }
                ),
                encoding="utf-8",
            )
            mismatched_metadata_cache = reviewer._load_cached_google_spreadsheet_artifact(
                url=url,
                google_credentials=credentials,
                cache_dir=cache_dir,
                max_chars=100,
            )
            cache_path.write_text(
                json.dumps(
                    {
                        "version": reviewer.PRD_GOOGLE_SHEET_ARTIFACT_CACHE_VERSION,
                        "drive_metadata": metadata,
                        "extracted": "bad",
                    }
                ),
                encoding="utf-8",
            )
            non_dict_cache = reviewer._load_cached_google_spreadsheet_artifact(
                url=url,
                google_credentials=credentials,
                cache_dir=cache_dir,
                max_chars=100,
            )
            cache_path.write_text(
                json.dumps(
                    {
                        "version": reviewer.PRD_GOOGLE_SHEET_ARTIFACT_CACHE_VERSION,
                        "drive_metadata": metadata,
                        "extracted": {"text": "cached spreadsheet text"},
                    }
                ),
                encoding="utf-8",
            )
            ok_cache = reviewer._load_cached_google_spreadsheet_artifact(
                url=url,
                google_credentials=credentials,
                cache_dir=cache_dir,
                max_chars=6,
            )
        reviewer._store_cached_google_spreadsheet_artifact(
            url=url,
            drive_metadata={"id": "", "modifiedTime": ""},
            cache_dir=cache_dir,
            extracted={"text": "skip"},
        )
        reviewer._store_cached_google_spreadsheet_artifact(
            url=url,
            drive_metadata=metadata,
            cache_dir=None,
            extracted={"text": "skip"},
        )
        with patch.object(Path, "write_text", side_effect=OSError("disk full")):
            reviewer._store_cached_google_spreadsheet_artifact(
                url=url,
                drive_metadata=metadata,
                cache_dir=cache_dir,
                extracted={"text": "skip"},
            )

        omitted_workbook = reviewer._extract_workbook_text(_xlsx_bytes(marker="Omitted"), max_chars=0)
        truncated_workbook = reviewer._extract_workbook_text(_xlsx_bytes(marker="Truncated"), max_chars=20)

        class EmptyWorkbook:
            worksheets = []

        with patch("openpyxl.load_workbook", return_value=EmptyWorkbook()):
            with self.assertRaisesRegex(ToolError, "empty_workbook"):
                reviewer._extract_workbook_text(b"not-read", max_chars=100)
        with patch("builtins.__import__", side_effect=lambda name, *args, **kwargs: (_ for _ in ()).throw(ImportError("missing openpyxl")) if name == "openpyxl" else real_import(name, *args, **kwargs)):
            with self.assertRaisesRegex(ToolError, "openpyxl_unavailable"):
                reviewer._extract_workbook_text(b"not-read", max_chars=100)

        section = ParsedSection(
            title="Detached",
            section_path="Detached",
            content="No placeholder here.",
            media_refs=["MEDIA_TABLE"],
        )
        detached_page = IngestedConfluencePage(
            page_id="detached",
            title="AF Detached",
            source_url="https://example.atlassian.net/wiki/pages/detached",
            updated_at="2026-05-20T00:00:00Z",
            language="en",
            sections=[section],
            media_dict={"MEDIA_TABLE": {"type": "table", "content": "<table><tr><td>Detached row</td></tr></table>"}},
        )
        detached_content = reviewer._build_review_section_content(page=detached_page, section=section)
        fallback_rows = reviewer._extract_table_media_rows({"content": "<div>Fallback<br>Rows</div>"})
        too_tiny_table = reviewer._pack_table_rows_for_review(media_id="T3", rows=["x" * 80], section=None, remaining_chars=1)
        short_identifier_typo = reviewer._is_spurious_identifier_typo_finding(
            "- **Priority:** P2\n  - **Clarification:** Fix typo ID",
            source_text="",
            source_identifier_text="id",
        )
        short_only_identifier_typo = reviewer._is_spurious_identifier_typo_finding(
            "p2 typo id",
            source_text="",
            source_identifier_text="id",
        )
        with patch("prd_briefing.reviewer._build_prd_source_payload", side_effect=ToolError("empty")):
            batches = reviewer._split_prd_page_batches(detached_page)
        blank_coverage = reviewer._build_section_coverage(
            total_sections=1,
            selected_indexes=[1],
            selected_sections=[ParsedSection(title="Blank", section_path="Blank", content="")],
            selection_hash="blank",
        )
        settings_without_data_dir = replace(self._settings(), team_portal_data_dir=None)
        zh_prompt = build_prd_review_prompt(
            jira_id="AF-ZH",
            jira_link="",
            prd_url=detached_page.source_url,
            page=detached_page,
            language="zh",
            linked_spreadsheet_evidence={"artifacts": [{"status": "ok", "title": "Template.xlsx", "text": "Field\tRule"}]},
        )
        with patch(
            "prd_briefing.reviewer._generate_with_codex",
            return_value={"result_markdown": "### Generated", "model_id": "codex-cli", "trace": {}},
        ) as generate:
            review_generated = reviewer.generate_prd_review_with_codex(
                prompt="Review",
                settings=self._settings(),
                workspace_root=Path(self.temp_dir.name),
                language="en",
            )
            summary_generated = reviewer.generate_prd_summary_with_codex(
                prompt="Summary",
                settings=self._settings(),
                workspace_root=Path(self.temp_dir.name),
                language="en",
            )

        class FakeCodexProvider:
            generated_payload = None

            def __init__(self, **kwargs):
                self.kwargs = kwargs

            def generate(self, *, payload, primary_model, fallback_model):
                FakeCodexProvider.generated_payload = payload
                return type(
                    "FakeCodexResult",
                    (),
                    {"payload": {"text": "provider text", "codex_cli_trace": {"session_id": "provider"}}, "model": primary_model},
                )()

            def extract_text(self, payload):
                return payload["text"]

        with patch("prd_briefing.reviewer.CodexCliBridgeSourceCodeQALLMProvider", FakeCodexProvider), patch(
            "prd_briefing.reviewer.resolve_codex_model", return_value="codex-test"
        ), patch("prd_briefing.reviewer.resolve_codex_reasoning_effort", return_value="low"):
            provider_generated = reviewer._generate_with_codex(
                prompt="Provider prompt",
                settings=self._settings(),
                workspace_root=Path(self.temp_dir.name),
                system_text="System",
                prompt_mode="mode",
                image_paths=["/tmp/sheet.png"],
            )

        self.assertIsNone(no_credentials)
        self.assertIsNone(no_cache_dir)
        self.assertGreater(len(native_content), 0)
        self.assertEqual(native_metadata["mimeType"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIsNone(missing_cache)
        self.assertIsNone(invalid_json_cache)
        self.assertIsNone(stale_version_cache)
        self.assertIsNone(mismatched_metadata_cache)
        self.assertIsNone(non_dict_cache)
        self.assertEqual(ok_cache["text"], "cached")
        self.assertIn("omitted", omitted_workbook["text"])
        self.assertLessEqual(len(truncated_workbook["text"]), 20)
        self.assertIn("Detached row", detached_content)
        self.assertEqual(fallback_rows, ["Fallback\nRows"])
        self.assertEqual(too_tiny_table["rows_included"], 0)
        self.assertFalse(short_identifier_typo)
        self.assertFalse(short_only_identifier_typo)
        self.assertEqual(len(batches), 1)
        self.assertEqual(blank_coverage["sections_assessed"], 0)
        self.assertTrue(reviewer._is_anti_fraud_prd(detached_page))
        self.assertIsNone(reviewer._google_sheet_artifact_cache_dir(settings_without_data_dir))
        self.assertIsNone(reviewer._google_sheet_screenshot_cache_dir(settings_without_data_dir))
        self.assertTrue(reviewer._google_sheet_screenshot_cache_dir(self._settings()).name.endswith("google_sheet_screenshot_evidence"))
        self.assertIsNone(reviewer._parse_json_object(""))
        self.assertEqual(reviewer._likely_header_row([]), [])
        self.assertTrue(reviewer._has_multi_level_header_risk([], ["A1:B2"]))
        self.assertIn("Google Sheet Screenshot Template Assessment", reviewer._google_sheet_screenshot_output_structure_en())
        self.assertIn("Google Sheet 截图模板评估", reviewer._google_sheet_screenshot_output_structure_zh())
        self.assertIn("你是一位资深 PRD", zh_prompt)
        self.assertEqual(review_generated["result_markdown"], "### Generated")
        self.assertEqual(summary_generated["result_markdown"], "### Generated")
        self.assertEqual(generate.call_count, 2)
        self.assertEqual(provider_generated["result_markdown"], "provider text")
        self.assertEqual(provider_generated["trace"], {"session_id": "provider"})
        self.assertEqual(FakeCodexProvider.generated_payload["_codex_image_paths"], ["/tmp/sheet.png"])

    def test_prd_reviewer_image_extraction_and_prompt_builder_boundaries(self):
        page = IngestedConfluencePage(
            page_id="prompt-prd",
            title="Prompt PRD",
            source_url="https://example.atlassian.net/wiki/pages/prompt-prd",
            updated_at="2026-05-20T00:00:00Z",
            language="en",
            sections=[ParsedSection(title="Overview", section_path="Overview", content="Scope and status rules.")],
        )
        batch_outputs = [{"batch_index": 1, "section_titles": ["Overview"], "result_markdown": "### Batch Review"}]
        screenshot_evidence = {
            "artifacts": [
                {
                    "status": "ok",
                    "is_google_sheet_screenshot": True,
                    "image_id": "IMG_1",
                    "source_section_title": "Overview",
                    "url": "https://example/sheet.png",
                    "evidence_text": "Header: Field",
                }
            ]
        }
        linked_evidence = {"artifacts": [{"status": "ok", "title": "Template.xlsx", "text": "Field\tRule"}]}

        with patch(
            "prd_briefing.reviewer._generate_with_codex",
            return_value={"result_markdown": "```json\n{\"is_google_sheet_screenshot\":true,\"evidence_text\":\"Visible grid\"}\n```"},
        ) as generate:
            extracted = _extract_google_sheet_screenshot_evidence_from_image(
                image_bytes=b"png",
                image_url="https://example/sheet.png",
                image_id="IMG_1",
                source_section="Overview",
                settings=self._settings(),
                workspace_root=Path(self.temp_dir.name),
            )
        with patch(
            "prd_briefing.reviewer._generate_with_codex",
            return_value={"result_markdown": "not json"},
        ):
            with self.assertRaisesRegex(ToolError, "ocr_failed"):
                _extract_google_sheet_screenshot_evidence_from_image(
                    image_bytes=b"png",
                    image_url="https://example/sheet.png",
                    image_id="IMG_1",
                    source_section="Overview",
                    settings=self._settings(),
                    workspace_root=Path(self.temp_dir.name),
                )

        summary_batch_en = build_prd_summary_batch_prompt(
            jira_id="AF-1",
            jira_link="https://jira/AF-1",
            prd_url=page.source_url,
            page=page,
            language="en",
            batch_index=1,
            batch_total=2,
        )
        summary_batch_zh = build_prd_summary_batch_prompt(
            jira_id="AF-1",
            jira_link="",
            prd_url=page.source_url,
            page=page,
            language="zh",
            batch_index=1,
            batch_total=2,
        )
        summary_synthesis = build_prd_summary_synthesis_prompt(
            jira_id="AF-1",
            jira_link="",
            prd_url=page.source_url,
            page=page,
            language="en",
            batch_outputs=batch_outputs,
        )
        review_batch_en = build_prd_review_batch_prompt(
            jira_id="AF-1",
            jira_link="",
            prd_url=page.source_url,
            page=page,
            language="en",
            batch_index=1,
            batch_total=2,
        )
        review_batch_zh = build_prd_review_batch_prompt(
            jira_id="AF-1",
            jira_link="",
            prd_url=page.source_url,
            page=page,
            language="zh",
            batch_index=1,
            batch_total=2,
        )
        review_synthesis = build_prd_review_synthesis_prompt(
            jira_id="AF-1",
            jira_link="",
            prd_url=page.source_url,
            page=page,
            language="en",
            batch_outputs=batch_outputs,
            linked_spreadsheet_evidence=linked_evidence,
            google_sheet_screenshot_evidence=screenshot_evidence,
        )
        review_synthesis_zh = build_prd_review_synthesis_prompt(
            jira_id="AF-1",
            jira_link="",
            prd_url=page.source_url,
            page=page,
            language="zh",
            batch_outputs=batch_outputs,
            linked_spreadsheet_evidence={},
            google_sheet_screenshot_evidence={},
        )

        self.assertTrue(extracted["is_google_sheet_screenshot"])
        self.assertIn("image_paths", generate.call_args.kwargs)
        self.assertIn("Summarize this PRD section batch", summary_batch_en)
        self.assertIn("请总结这个 PRD section batch", summary_batch_zh)
        self.assertIn("Batch 1", summary_synthesis)
        self.assertIn("Review this PRD section batch", review_batch_en)
        self.assertIn("请评审这个 PRD section batch", review_batch_zh)
        self.assertIn("Google Sheet Screenshot Template Assessment", review_synthesis)
        self.assertIn("Report Generation Feasibility", review_synthesis)
        self.assertIn("PM Decision Checklist", review_synthesis_zh)
        self.assertIn("Return only the requested Markdown review in English", build_prd_review_system_text("en"))
        self.assertIn("Chinese", build_prd_review_system_text("zh"))
        self.assertTrue(_use_compact_review_prompt(page, {}, {}))
        self.assertFalse(_use_compact_review_prompt(page, linked_evidence, {}))
        self.assertFalse(_use_compact_review_prompt(page, {}, screenshot_evidence))

    def test_prd_self_assessment_section_char_count_includes_table_media(self):
        page = self.service.confluence.page
        page.sections.append(
            ParsedSection(
                title="Admin Portal Changes",
                section_path="3.1.2 Admin Portal Changes",
                content="[MEDIA_ID_1]",
                html_content="<table><tr><th>Field</th></tr><tr><td>Status</td></tr></table>",
                media_refs=["MEDIA_ID_1"],
            )
        )
        service = PRDReviewService(
            store=self.store,
            confluence=self.service.confluence,
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        payload = service.list_url_sections(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
            )
        )

        table_section = payload["sections"][2]
        self.assertEqual(table_section["title"], "3.1.2 Admin Portal Changes")
        self.assertGreater(table_section["char_count"], len("[MEDIA_ID_1]"))
        self.assertGreaterEqual(table_section["char_count"], len("[MEDIA_ID_1 table content]\nField\nStatus"))

    def test_prd_review_prompt_dry_run_omits_empty_jira_context_without_dropping_prd_source(self):
        page = self.service.confluence.page
        source = _build_prd_source_payload(page, max_chars=10_000_000)["source"]

        prompt = build_prd_review_prompt(
            jira_id="",
            jira_link="",
            prd_url=page.source_url,
            page=page,
            language="en",
            linked_spreadsheet_evidence={},
            google_sheet_screenshot_evidence={},
        )
        legacy_prompt = prompt.replace("# Context\n- PRD Title:", "# Context\n- Jira ID: -\n- Jira Link: -\n- PRD Title:", 1)
        team_dashboard_prompt = build_prd_review_prompt(
            jira_id="AF-1",
            jira_link="https://jira.example/browse/AF-1",
            prd_url=page.source_url,
            page=page,
            language="en",
            linked_spreadsheet_evidence={},
            google_sheet_screenshot_evidence={},
        )

        self.assertNotIn("- Jira ID: -", prompt)
        self.assertNotIn("- Jira Link: -", prompt)
        self.assertIn(source, prompt)
        self.assertLess(len(prompt), len(legacy_prompt))
        self.assertIn("- Jira ID: AF-1", team_dashboard_prompt)
        self.assertIn("- Jira Link: https://jira.example/browse/AF-1", team_dashboard_prompt)

    def test_prd_review_table_prompt_dry_run_dedupes_repeated_rows_without_dropping_unique_rows(self):
        page = IngestedConfluencePage(
            page_id="table-dedup",
            title="Table Dedup PRD",
            source_url="https://example.atlassian.net/wiki/pages/table-dedup",
            updated_at="2026-05-24T00:00:00Z",
            language="en",
            sections=[
                ParsedSection(
                    title="Rules",
                    section_path="3. Rules",
                    content="[MEDIA_ID_1]",
                    media_refs=["MEDIA_ID_1"],
                )
            ],
            media_dict={
                "MEDIA_ID_1": {
                    "type": "table",
                    "content": (
                        "<table>"
                        "<tr><th>Field</th><th>Value</th></tr>"
                        "<tr><td>DUPLICATE_TOKEN</td><td>same rule</td></tr>"
                        "<tr><td>DUPLICATE_TOKEN</td><td>same rule</td></tr>"
                        "<tr><td>UNIQUE_TOKEN</td><td>kept rule</td></tr>"
                        "</table>"
                    ),
                }
            },
        )

        source = _build_prd_source_payload(page, max_chars=10_000_000)["source"]
        legacy_source = source.replace("DUPLICATE_TOKEN | same rule\n", "DUPLICATE_TOKEN | same rule\nDUPLICATE_TOKEN | same rule\n", 1)

        self.assertEqual(source.count("DUPLICATE_TOKEN | same rule"), 1)
        self.assertIn("UNIQUE_TOKEN | kept rule", source)
        self.assertLess(len(source), len(legacy_source))

    def test_confluence_parser_extracts_spreadsheet_links_by_section(self):
        connector = ConfluenceConnector(base_url="", email="", api_token="", bearer_token="", store=self.store)

        sections = connector._parse_sections(
            html="""
            <h2>Register</h2>
            <p><a href="/download/attachments/123/MAS%20Outsourcing%20Register.xlsx">MAS Register format</a></p>
            <h2>Appendix</h2>
            <ac:link><ri:attachment ri:filename="Fallback Format.xlsm" /></ac:link>
            """,
            base_url="https://confluence.shopee.io",
            source_url="https://confluence.shopee.io/pages/viewpage.action?pageId=123",
            page_id="123",
            session_id="s1",
            media_dict={},
        )

        self.assertEqual(sections[0].spreadsheet_links[0].title, "MAS Register format")
        self.assertIn("/download/attachments/123/MAS%20Outsourcing%20Register.xlsx", sections[0].spreadsheet_links[0].url)
        self.assertEqual(sections[1].spreadsheet_links[0].filename, "Fallback Format.xlsm")

    def test_confluence_payload_extracts_ancestor_titles(self):
        payload = {
            "ancestors": [
                {"title": "2. Digital Banking PRD"},
                {"title": "2.2 Authentication & Antifraud"},
                {"title": "[AF] Admin Portal"},
                {"title": "[AF] Admin Portal"},
                {"title": ""},
            ]
        }

        self.assertEqual(
            ConfluenceConnector._extract_ancestor_titles(payload),
            ["2. Digital Banking PRD", "2.2 Authentication & Antifraud", "[AF] Admin Portal"],
        )

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_self_assessment_selected_sections_limit_prompt_and_cache(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        service = PRDReviewService(
            store=self.store,
            confluence=self.service.confluence,
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        first = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )
        second = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )
        different_selection = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[1],
            )
        )

        first_prompt = mock_generate.call_args_list[0].kwargs["prompt"]
        self.assertIn("Rollout will happen", first_prompt)
        self.assertNotIn("approval workflow", first_prompt)
        self.assertFalse(first["cached"])
        self.assertTrue(second["cached"])
        self.assertFalse(different_selection["cached"])
        self.assertEqual(mock_generate.call_count, 2)
        self.assertTrue(first["review"]["jira_id"].startswith(f"{PRD_BRIEFING_REVIEW_CACHE_KEY}:sections:"))
        self.assertEqual(first["coverage"]["sections_total"], 2)
        self.assertEqual(first["coverage"]["selected_section_indexes"], [2])
        self.assertEqual(first["coverage"]["selected_section_titles"], ["Rollout"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_self_assessment_all_selected_reuses_full_review_cache_key(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        service = PRDReviewService(
            store=self.store,
            confluence=self.service.confluence,
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        first = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="zh",
                selected_section_indexes=[1, 2],
            )
        )
        cached = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="zh",
            )
        )

        self.assertFalse(first["cached"])
        self.assertTrue(cached["cached"])
        self.assertEqual(first["review"]["jira_id"], PRD_BRIEFING_REVIEW_CACHE_KEY)
        self.assertEqual(mock_generate.call_count, 1)

    def test_prd_self_assessment_selected_sections_validate_indexes(self):
        service = PRDReviewService(
            store=self.store,
            confluence=self.service.confluence,
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        with self.assertRaisesRegex(ToolError, "at least one"):
            service.review_url(
                PRDBriefingReviewRequest(
                    owner_key="anon:test",
                    prd_url="https://example.atlassian.net/wiki/pages/123",
                    selected_section_indexes=[],
                )
            )
        with self.assertRaisesRegex(ToolError, "out of range"):
            service.review_url(
                PRDBriefingReviewRequest(
                    owner_key="anon:test",
                    prd_url="https://example.atlassian.net/wiki/pages/123",
                    selected_section_indexes=[3],
                )
            )
        with self.assertRaisesRegex(ToolError, "must be a list"):
            service.review_url(
                PRDBriefingReviewRequest(
                    owner_key="anon:test",
                    prd_url="https://example.atlassian.net/wiki/pages/123",
                    selected_section_indexes="2",
                )
            )

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_self_assessment_includes_selected_linked_spreadsheet_evidence(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.sections[1].section_path = "3.11.1 Report Layout and Field Name & Format"
        page.sections[1].spreadsheet_links = [
            SpreadsheetLink(
                title="MAS Outsourcing Register.xlsx",
                url="https://example.atlassian.net/download/attachments/123/MAS%20Outsourcing%20Register.xlsx",
                source_section="Rollout",
                filename="MAS Outsourcing Register.xlsx",
            )
        ]
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, _xlsx_bytes(marker="MAS field format")),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        prompt = mock_generate.call_args.kwargs["prompt"]
        self.assertIn("# Linked Spreadsheet Evidence", prompt)
        self.assertIn("MAS Outsourcing Register.xlsx", prompt)
        self.assertIn("Outsourcing Arrangement ID", prompt)
        self.assertIn("MAS field format", prompt)
        self.assertNotIn("approval workflow", prompt)
        self.assertEqual(result["coverage"]["linked_artifacts_total"], 1)
        self.assertEqual(result["coverage"]["linked_artifacts_reviewed"], 1)
        self.assertEqual(result["coverage"]["linked_artifacts"][0]["sheets_extracted"], ["Register Format"])
        self.assertEqual(result["coverage"]["report_templates_reviewed"], 1)

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_self_assessment_skips_linked_spreadsheets_in_non_report_sections(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.sections[1].section_path = "Rollout Plan"
        page.sections[1].spreadsheet_links = [
            SpreadsheetLink(
                title="Random Tracker.xlsx",
                url="https://example.atlassian.net/download/attachments/123/Random%20Tracker.xlsx",
                source_section="Rollout",
                filename="Random Tracker.xlsx",
            )
        ]
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, _xlsx_bytes(marker="should not enter prompt")),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        self.assertEqual(result["coverage"]["report_templates_total"], 0)
        prompt = mock_generate.call_args.kwargs["prompt"]
        self.assertNotIn("should not enter prompt", prompt)
        self.assertIn("No report-template spreadsheet links", prompt)

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_self_assessment_reads_all_report_section_spreadsheets_without_file_limit(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.sections[1].section_path = "3.11.1 Report Layout and Field Name & Format"
        page.sections[1].spreadsheet_links = [
            SpreadsheetLink(
                title=f"Report Template {index}.xlsx",
                url=f"https://example.atlassian.net/download/attachments/123/Report%20Template%20{index}.xlsx",
                source_section="Report Layout",
                filename=f"Report Template {index}.xlsx",
            )
            for index in range(1, 8)
        ]
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, _xlsx_bytes(marker="all templates")),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        self.assertEqual(result["coverage"]["report_templates_total"], 7)
        self.assertEqual(result["coverage"]["report_templates_reviewed"], 7)
        self.assertNotIn("max_linked_spreadsheet_limit", mock_generate.call_args.kwargs["prompt"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_self_assessment_linked_workbook_extracts_first_ten_sheets(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.sections[1].section_path = "3.11.1 Report Layout and Field Name & Format"
        page.sections[1].spreadsheet_links = [
            SpreadsheetLink(
                title="Large Register.xlsx",
                url="https://example.atlassian.net/download/attachments/123/Large%20Register.xlsx",
                source_section="Rollout",
                filename="Large Register.xlsx",
            )
        ]
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, _xlsx_bytes(sheet_count=12)),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        artifact = result["coverage"]["linked_artifacts"][0]
        self.assertEqual(len(artifact["sheets_extracted"]), 10)
        self.assertEqual(artifact["skipped_sheet_count"], 2)
        prompt = mock_generate.call_args.kwargs["prompt"]
        self.assertIn("[Sheet: Sheet 10]", prompt)
        self.assertNotIn("[Sheet: Sheet 11]", prompt)

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_self_assessment_linked_workbook_extracts_template_metadata(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.sections[1].section_path = "3.11.1 Report Layout and Field Name & Format"
        page.sections[1].spreadsheet_links = [
            SpreadsheetLink(
                title="Risky Register.xlsx",
                url="https://example.atlassian.net/download/attachments/123/Risky%20Register.xlsx",
                source_section="Report Layout",
                filename="Risky Register.xlsx",
            )
        ]
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, _xlsx_template_risk_bytes()),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        artifact = result["coverage"]["report_templates"][0]
        metadata = artifact["template_metadata"]["sheets"][0]
        self.assertGreater(metadata["merged_range_count"], 0)
        self.assertGreater(metadata["formula_cell_count"], 0)
        self.assertGreater(metadata["empty_header_count"], 0)
        self.assertIn("Field", metadata["duplicate_headers"])
        self.assertGreater(metadata["hidden_columns_count"], 0)
        prompt = mock_generate.call_args.kwargs["prompt"]
        self.assertIn("Report template metadata", prompt)
        self.assertIn("formula_cells=", prompt)

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_self_assessment_linked_spreadsheet_hash_splits_cache(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.sections[1].section_path = "3.11.1 Report Layout and Field Name & Format"
        page.sections[1].spreadsheet_links = [
            SpreadsheetLink(
                title="MAS Outsourcing Register.xlsx",
                url="https://example.atlassian.net/download/attachments/123/MAS%20Outsourcing%20Register.xlsx",
                source_section="Rollout",
                filename="MAS Outsourcing Register.xlsx",
            )
        ]
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, _xlsx_bytes(marker="initial format")),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        first = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )
        service.confluence = FakeAttachmentConnector(page, _xlsx_bytes(marker="updated format"))
        second = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        self.assertFalse(first["cached"])
        self.assertFalse(second["cached"])
        self.assertEqual(mock_generate.call_count, 2)
        self.assertNotEqual(first["review"]["jira_id"], second["review"]["jira_id"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_self_assessment_linked_spreadsheet_failure_is_coverage_gap(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.sections[1].section_path = "3.11.1 Report Layout and Field Name & Format"
        page.sections[1].spreadsheet_links = [
            SpreadsheetLink(
                title="Legacy Register.xls",
                url="https://example.atlassian.net/download/attachments/123/Legacy%20Register.xls",
                source_section="Rollout",
                filename="Legacy Register.xls",
            )
        ]
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, b"not-used"),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        self.assertEqual(result["coverage"]["linked_artifacts_reviewed"], 0)
        self.assertEqual(result["coverage"]["linked_artifacts_failed"], 1)
        self.assertEqual(result["coverage"]["linked_artifacts"][0]["reason"], "unsupported_xls_format")
        self.assertIn("Linked artifact not reviewed", mock_generate.call_args.kwargs["prompt"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_self_assessment_google_sheet_link_uses_google_credentials(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.sections[1].section_path = "3.11.1 Report Layout and Field Name & Format"
        page.sections[1].spreadsheet_links = [
            SpreadsheetLink(
                title="MAS Outsourcing Register Google Sheet",
                url="https://docs.google.com/spreadsheets/d/sheet123/edit#gid=0",
                source_section="Rollout",
                filename="",
            )
        ]
        service = PRDReviewService(
            store=self.store,
            confluence=FakeConnector(page),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        with patch(
            "bpmis_jira_tool.gmail_dashboard.build_drive_api_service",
            return_value=FakeDriveService(_xlsx_bytes(marker="Google Sheet format")),
        ):
            result = service.review_url(
                PRDBriefingReviewRequest(
                    owner_key="anon:test",
                    prd_url="https://example.atlassian.net/wiki/pages/123",
                    language="en",
                    selected_section_indexes=[2],
                    google_credentials={
                        "token": "x",
                        "scopes": ["https://www.googleapis.com/auth/drive.readonly"],
                    },
                )
            )

        self.assertEqual(result["coverage"]["linked_artifacts_reviewed"], 1)
        self.assertIn("Google Sheet format", mock_generate.call_args.kwargs["prompt"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_self_assessment_google_sheet_artifact_cache_skips_second_export_when_unchanged(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.sections[1].section_path = "3.11.1 Report Layout and Field Name & Format"
        page.sections[1].spreadsheet_links = [
            SpreadsheetLink(
                title="MAS Outsourcing Register Google Sheet",
                url="https://docs.google.com/spreadsheets/d/sheet123/edit#gid=0",
                source_section="Rollout",
                filename="",
            )
        ]
        drive_service = FakeDriveService(_xlsx_bytes(marker="Cached Google Sheet format"))
        service = PRDReviewService(
            store=self.store,
            confluence=FakeConnector(page),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )
        request = PRDBriefingReviewRequest(
            owner_key="anon:test",
            prd_url="https://example.atlassian.net/wiki/pages/123",
            language="en",
            selected_section_indexes=[2],
            force_refresh=True,
            google_credentials={
                "token": "x",
                "scopes": ["https://www.googleapis.com/auth/drive.readonly"],
            },
        )

        with patch("bpmis_jira_tool.gmail_dashboard.build_drive_api_service", return_value=drive_service):
            first = service.review_url(request)
            second = service.review_url(request)

        self.assertEqual(drive_service.files_resource.export_calls, 1)
        self.assertFalse(first["coverage"]["linked_artifacts"][0]["cache_hit"])
        self.assertTrue(second["coverage"]["linked_artifacts"][0]["cache_hit"])
        self.assertIn("Cached Google Sheet format", mock_generate.call_args.kwargs["prompt"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_self_assessment_google_sheet_modified_time_invalidates_artifact_cache(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.sections[1].section_path = "3.11.1 Report Layout and Field Name & Format"
        page.sections[1].spreadsheet_links = [
            SpreadsheetLink(
                title="MAS Outsourcing Register Google Sheet",
                url="https://docs.google.com/spreadsheets/d/sheet123/edit#gid=0",
                source_section="Rollout",
                filename="",
            )
        ]
        drive_service = FakeDriveService(_xlsx_bytes(marker="Google Sheet v1"), modified_time="2026-05-12T10:00:00.000Z")
        service = PRDReviewService(
            store=self.store,
            confluence=FakeConnector(page),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )
        request = PRDBriefingReviewRequest(
            owner_key="anon:test",
            prd_url="https://example.atlassian.net/wiki/pages/123",
            language="en",
            selected_section_indexes=[2],
            force_refresh=True,
            google_credentials={
                "token": "x",
                "scopes": ["https://www.googleapis.com/auth/drive.readonly"],
            },
        )

        with patch("bpmis_jira_tool.gmail_dashboard.build_drive_api_service", return_value=drive_service):
            service.review_url(request)
            drive_service.files_resource.modified_time = "2026-05-13T10:00:00.000Z"
            drive_service.files_resource.workbook_content = _xlsx_bytes(marker="Google Sheet v2")
            result = service.review_url(request)

        self.assertEqual(drive_service.files_resource.export_calls, 2)
        self.assertFalse(result["coverage"]["linked_artifacts"][0]["cache_hit"])
        self.assertIn("Google Sheet v2", mock_generate.call_args.kwargs["prompt"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_review_progress_reports_prd_template_metadata_and_generation_stages(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.sections[1].section_path = "3.11.1 Report Layout and Field Name & Format"
        page.sections[1].spreadsheet_links = [
            SpreadsheetLink(
                title="MAS Outsourcing Register.xlsx",
                url="https://example.atlassian.net/download/attachments/123/MAS%20Outsourcing%20Register.xlsx",
                source_section="Rollout",
                filename="MAS Outsourcing Register.xlsx",
            )
        ]
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, _xlsx_bytes(marker="progress format")),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )
        progress_events = []

        service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            ),
            progress_callback=lambda stage, message, current, total, **kwargs: progress_events.append((stage, message, current, total, kwargs)),
        )

        stages = [event[0] for event in progress_events]
        messages = [event[1] for event in progress_events]
        self.assertIn("reading_prd", stages)
        self.assertIn("reading_report_templates", stages)
        self.assertIn("analyzing_template_metadata", stages)
        self.assertIn("generating_review", stages)
        self.assertTrue(any("Reading 1 report templates" in message for message in messages))
        self.assertTrue(any((event[4].get("estimated_prompt_tokens") or 0) > 0 for event in progress_events))

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    @patch("prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image")
    def test_anti_fraud_image_evidence_is_disabled_and_tables_remain_in_prompt(self, mock_extract, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.title = "Anti-Fraud Risk Decision PRD"
        page.sections[1].content = "Risk Decision scenario configuration.\n[MEDIA_ID_1]"
        self._add_prd_image(page, section_index=2)
        page.sections[1].media_refs = ["MEDIA_ID_1", *page.sections[1].media_refs]
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, b"sheet-image-v1"),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        prompt = mock_generate.call_args.kwargs["prompt"]
        mock_extract.assert_not_called()
        self.assertNotIn("# Google Sheet Screenshot Evidence", prompt)
        self.assertIn("[MEDIA_ID_1 table content]", prompt)
        self.assertFalse(result["coverage"]["google_sheet_screenshot_evidence_enabled"])
        self.assertEqual(result["coverage"]["google_sheet_screenshots_total"], 0)

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    @patch("prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image")
    @unittest.skip("Image evidence scanning is disabled; PRD review uses table media text only.")
    def test_anti_fraud_prd_google_sheet_screenshot_enters_review_prompt(self, mock_extract, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        mock_extract.return_value = {
            "is_google_sheet_screenshot": True,
            "reason": "",
            "classification": "Google Sheets grid with sheet tabs",
            "evidence_text": "Visible fields: Scenario ID, Rule Name, Action. Empty duplicate header in column C.",
        }
        page = self.service.confluence.page
        page.title = "Anti-Fraud Risk Decision PRD"
        page.sections[1].content = "Risk Decision scenario configuration uses a Google Sheet screenshot."
        media_id = self._add_prd_image(page, section_index=2)
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, b"sheet-image-v1"),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        prompt = mock_generate.call_args.kwargs["prompt"]
        self.assertIn("# Google Sheet Screenshot Evidence", prompt)
        self.assertIn("Visible fields: Scenario ID, Rule Name, Action", prompt)
        self.assertIn("Google Sheet Screenshot Template Assessment", prompt)
        self.assertIn("Do not treat `[MEDIA_ID_x]`", prompt)
        self.assertIn(media_id, prompt)
        self.assertEqual(result["coverage"]["google_sheet_screenshots_total"], 1)
        self.assertEqual(result["coverage"]["google_sheet_screenshots_reviewed"], 1)
        self.assertEqual(result["coverage"]["google_sheet_screenshot_images"][0]["status"], "ok")

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    @patch("prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image")
    @unittest.skip("Image evidence scanning is disabled; PRD review uses table media text only.")
    def test_anti_fraud_storage_attachment_image_enters_review_prompt(self, mock_extract, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        mock_extract.return_value = {
            "is_google_sheet_screenshot": True,
            "reason": "",
            "classification": "Google Sheets screenshot from Confluence attachment",
            "evidence_text": "Visible fields: Scenario Code, Trigger, Action.",
        }
        connector = ConfluenceConnector(
            base_url="https://example.atlassian.net",
            email=None,
            api_token=None,
            bearer_token=None,
            store=self.store,
        )
        media_dict = {}
        sections = connector._parse_sections(
            html="""
            <h2>3.1 Scenario Rules</h2>
            <p>CMS scenario setup.</p>
            <p><ac:image ac:width="960"><ri:attachment ri:filename="AF Scenario Sheet.png" /></ac:image></p>
            """,
            base_url="https://example.atlassian.net",
            source_url="https://example.atlassian.net/pages/viewpage.action?pageId=123",
            page_id="123",
            session_id="session-1",
            media_dict=media_dict,
        )
        page = IngestedConfluencePage(
            page_id="123",
            title="Anti-Fraud PRD",
            source_url="https://example.atlassian.net/pages/viewpage.action?pageId=123",
            updated_at="2026-05-12T17:56:27.000+08:00",
            language="en",
            sections=sections,
            media_dict=media_dict,
            ancestor_titles=["Productization PRD - Authentication & Antifraud"],
        )
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, b"sheet-image-v1"),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url=page.source_url,
                language="en",
                selected_section_indexes=[1],
            )
        )

        prompt = mock_generate.call_args.kwargs["prompt"]
        self.assertIn("# Google Sheet Screenshot Evidence", prompt)
        self.assertIn("Visible fields: Scenario Code, Trigger, Action.", prompt)
        self.assertEqual(result["coverage"]["google_sheet_screenshots_reviewed"], 1)
        self.assertIn("/download/attachments/123/AF%20Scenario%20Sheet.png", service.confluence.requested_urls[0][0])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    @patch("prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image")
    @unittest.skip("Image evidence scanning is disabled; PRD review uses table media text only.")
    def test_anti_fraud_table_embedded_image_enters_review_prompt(self, mock_extract, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        mock_extract.return_value = {
            "is_google_sheet_screenshot": True,
            "reason": "",
            "classification": "Google Sheets grid inside Confluence table",
            "evidence_text": "Visible fields: L1 scenario, L2 scenario, Action, Report V1.",
        }
        connector = ConfluenceConnector(
            base_url="https://example.atlassian.net",
            email=None,
            api_token=None,
            bearer_token=None,
            store=self.store,
        )
        media_dict = {}
        sections = connector._parse_sections(
            html="""
            <h2>3.1 Scenario Rules</h2>
            <table>
              <tr><th>Details</th></tr>
              <tr>
                <td>
                  Please refer to gsheet: Row 17
                  <img class="confluence-embedded-image" src="/download/attachments/123/sheet-row-17.png" width="1200">
                </td>
              </tr>
            </table>
            """,
            base_url="https://example.atlassian.net",
            source_url="https://example.atlassian.net/pages/viewpage.action?pageId=123",
            page_id="123",
            session_id="session-1",
            media_dict=media_dict,
        )
        page = IngestedConfluencePage(
            page_id="123",
            title="Anti-Fraud PRD",
            source_url="https://example.atlassian.net/pages/viewpage.action?pageId=123",
            updated_at="2026-05-12T17:56:27.000+08:00",
            language="en",
            sections=sections,
            media_dict=media_dict,
            ancestor_titles=["2.2 Authentication & Antifraud"],
        )
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, b"sheet-image-v1"),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url=page.source_url,
                language="en",
                selected_section_indexes=[1],
            )
        )

        prompt = mock_generate.call_args.kwargs["prompt"]
        self.assertIn("# Google Sheet Screenshot Evidence", prompt)
        self.assertIn("Visible fields: L1 scenario, L2 scenario, Action, Report V1.", prompt)
        self.assertEqual(result["coverage"]["google_sheet_screenshots_reviewed"], 1)
        self.assertTrue(any(item["type"] == "table" for item in media_dict.values()))
        self.assertTrue(any(item["type"] == "image" for item in media_dict.values()))
        self.assertIn("/download/attachments/123/sheet-row-17.png", service.confluence.requested_urls[0][0])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    @patch("prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image")
    def test_non_anti_fraud_prd_does_not_read_image_evidence(self, mock_extract, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        self._add_prd_image(page, section_index=2)
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, b"sheet-image-v1"),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        mock_extract.assert_not_called()
        self.assertEqual(result["coverage"]["google_sheet_screenshots_total"], 0)
        self.assertNotIn("# Google Sheet Screenshot Evidence", mock_generate.call_args.kwargs["prompt"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    @patch("prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image")
    @unittest.skip("Image evidence scanning is disabled; PRD review uses table media text only.")
    def test_anti_fraud_detection_uses_confluence_ancestor_path(self, mock_extract, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        mock_extract.return_value = {
            "is_google_sheet_screenshot": True,
            "reason": "",
            "classification": "Google Sheets grid",
            "evidence_text": "Visible fields: Scenario, Action, Threshold",
        }
        page = self.service.confluence.page
        page.title = "Configuration Tool for Operators"
        page.ancestor_titles = ["2. Digital Banking PRD", "2.2 Authentication & Antifraud", "[AF] Admin Portal"]
        self._add_prd_image(page, section_index=2)
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, b"sheet-image-v1"),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        self.assertEqual(result["coverage"]["google_sheet_screenshots_reviewed"], 1)
        self.assertIn("# Google Sheet Screenshot Evidence", mock_generate.call_args.kwargs["prompt"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    @patch("prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image")
    def test_anti_fraud_detection_fallback_does_not_scan_section_body(self, mock_extract, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.title = "Generic Admin Portal PRD"
        page.ancestor_titles = []
        page.sections[1].content = "This section mentions fraud scenario text, but the title and ancestors do not identify Anti-Fraud."
        self._add_prd_image(page, section_index=2)
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, b"sheet-image-v1"),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        mock_extract.assert_not_called()
        self.assertEqual(result["coverage"]["google_sheet_screenshots_total"], 0)
        self.assertNotIn("# Google Sheet Screenshot Evidence", mock_generate.call_args.kwargs["prompt"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    @patch("prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image")
    @unittest.skip("Image evidence scanning is disabled; PRD review uses table media text only.")
    def test_anti_fraud_detection_fallback_uses_page_title_only(self, mock_extract, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        mock_extract.return_value = {
            "is_google_sheet_screenshot": True,
            "reason": "",
            "classification": "Google Sheets grid",
            "evidence_text": "Visible fields: Rule ID, Action",
        }
        page = self.service.confluence.page
        page.title = "[AF] Admin Portal Configuration PRD"
        page.ancestor_titles = []
        self._add_prd_image(page, section_index=2)
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, b"sheet-image-v1"),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        self.assertEqual(result["coverage"]["google_sheet_screenshots_reviewed"], 1)
        self.assertIn("# Google Sheet Screenshot Evidence", mock_generate.call_args.kwargs["prompt"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    @patch("prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image")
    @unittest.skip("Image evidence scanning is disabled; PRD review uses table media text only.")
    def test_anti_fraud_non_google_sheet_image_is_skipped_and_not_prompted(self, mock_extract, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        mock_extract.return_value = {
            "is_google_sheet_screenshot": False,
            "reason": "not_google_sheet_screenshot",
            "classification": "normal product screenshot",
            "evidence_text": "Button labels should not enter prompt",
        }
        page = self.service.confluence.page
        page.title = "Anti-Fraud PRD"
        self._add_prd_image(page, section_index=2)
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, b"normal-ui-image"),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        prompt = mock_generate.call_args.kwargs["prompt"]
        self.assertNotIn("# Google Sheet Screenshot Evidence", prompt)
        self.assertNotIn("Button labels should not enter prompt", prompt)
        self.assertEqual(result["coverage"]["google_sheet_screenshots_total"], 0)
        self.assertEqual(result["coverage"]["google_sheet_screenshot_images"][0]["reason"], "not_google_sheet_screenshot")

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    @patch("prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image")
    def test_unselected_section_image_is_not_read_for_self_assessment(self, mock_extract, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.title = "Anti-Fraud PRD"
        self._add_prd_image(page, section_index=1)
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, b"sheet-image-v1"),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        mock_extract.assert_not_called()
        self.assertEqual(result["coverage"]["google_sheet_screenshots_total"], 0)
        self.assertNotIn("# Google Sheet Screenshot Evidence", mock_generate.call_args.kwargs["prompt"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    @unittest.skip("Image evidence scanning is disabled; PRD review uses table media text only.")
    def test_google_sheet_screenshot_download_failure_is_coverage_gap(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        page = self.service.confluence.page
        page.title = "Anti-Fraud PRD"
        self._add_prd_image(page, section_index=2)
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, b""),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        result = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        self.assertEqual(result["coverage"]["google_sheet_screenshots_total"], 1)
        self.assertEqual(result["coverage"]["google_sheet_screenshots_failed"], 1)
        self.assertEqual(result["coverage"]["google_sheet_screenshot_images"][0]["reason"], "empty_download")
        self.assertIn("Google Sheet screenshot not reviewed", mock_generate.call_args.kwargs["prompt"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    @patch("prd_briefing.reviewer._extract_google_sheet_screenshot_evidence_from_image")
    @unittest.skip("Image evidence scanning is disabled; PRD review uses table media text only.")
    def test_google_sheet_screenshot_hash_splits_review_cache(self, mock_extract, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        mock_extract.side_effect = lambda *, image_bytes, **_kwargs: {
            "is_google_sheet_screenshot": True,
            "reason": "",
            "classification": "Google Sheet screenshot",
            "evidence_text": f"Visible marker: {image_bytes.decode('utf-8')}",
        }
        page = self.service.confluence.page
        page.title = "Anti-Fraud Risk Decision PRD"
        page.sections[1].content = "Risk Decision scenario configuration."
        self._add_prd_image(page, section_index=2)
        service = PRDReviewService(
            store=self.store,
            confluence=FakeAttachmentConnector(page, b"initial"),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        first = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )
        service.confluence = FakeAttachmentConnector(page, b"updated")
        second = service.review_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
                selected_section_indexes=[2],
            )
        )

        self.assertFalse(first["cached"])
        self.assertFalse(second["cached"])
        self.assertEqual(mock_generate.call_count, 2)
        self.assertNotEqual(first["review"]["jira_id"], second["review"]["jira_id"])

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_prd_review_persists_generation_failure_and_validates_inputs(self, mock_generate):
        mock_generate.side_effect = RuntimeError("model unavailable")
        service = PRDReviewService(
            store=self.store,
            confluence=self.service.confluence,
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        with self.assertRaisesRegex(ToolError, "model unavailable"):
            service.review(
                PRDReviewRequest(
                    owner_key=" anon:test ",
                    jira_id=" AF-123 ",
                    jira_link=" https://jira.example/browse/AF-123 ",
                    prd_url=" https://example.atlassian.net/wiki/pages/123 ",
                )
            )

        failed = self.store.get_prd_review_result(
            owner_key="anon:test",
            jira_id="AF-123",
            prd_url=self.service.confluence.page.source_url,
            prd_updated_at=self.service.confluence.page.updated_at,
            prompt_version=PRD_REVIEW_PROMPT_VERSION,
        )
        self.assertEqual(failed["status"], "failed")
        self.assertIn("model unavailable", failed["error"])

        invalid_requests = [
            PRDReviewRequest(owner_key="", jira_id="AF-123", jira_link="", prd_url="https://example.com"),
            PRDReviewRequest(owner_key="anon:test", jira_id="", jira_link="", prd_url="https://example.com"),
            PRDReviewRequest(owner_key="anon:test", jira_id="AF-123", jira_link="", prd_url="not-a-url"),
            PRDBriefingReviewRequest(owner_key="", prd_url="https://example.com"),
            PRDBriefingReviewRequest(owner_key="anon:test", prd_url="file:///tmp/prd.html"),
        ]
        for request_model in invalid_requests:
            with self.assertRaises(ToolError):
                if isinstance(request_model, PRDBriefingReviewRequest):
                    service.review_url(request_model)
                else:
                    service.review(request_model)

    def test_prd_review_rejects_pages_without_readable_sections(self):
        empty_page = IngestedConfluencePage(
            page_id="empty",
            title="Empty PRD",
            source_url="https://example.atlassian.net/wiki/pages/empty",
            updated_at="2026-05-01T00:00:00Z",
            language="en",
            sections=[],
        )
        service = PRDReviewService(
            store=self.store,
            confluence=FakeConnector(empty_page),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        with self.assertRaisesRegex(ToolError, "readable sections"):
            service.review_url(PRDBriefingReviewRequest(owner_key="anon:test", prd_url=empty_page.source_url))
        with self.assertRaisesRegex(ToolError, "readable sections"):
            service.summarize_url(PRDBriefingReviewRequest(owner_key="anon:test", prd_url=empty_page.source_url))
        with self.assertRaisesRegex(ToolError, "readable sections"):
            service.review(
                PRDReviewRequest(owner_key="anon:test", jira_id="AF-1", jira_link="", prd_url=empty_page.source_url)
            )
        with self.assertRaisesRegex(ToolError, "readable sections"):
            service.summarize(
                PRDReviewRequest(owner_key="anon:test", jira_id="AF-1", jira_link="", prd_url=empty_page.source_url)
            )

    @patch("prd_briefing.reviewer.generate_prd_summary_with_codex")
    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_classic_prd_review_and_summary_cache_success_and_failure_paths(self, mock_review, mock_summary):
        service = PRDReviewService(
            store=self.store,
            confluence=self.service.confluence,
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )
        mock_review.return_value = {
            "result_markdown": "### Classic Review",
            "model_id": "codex-cli",
            "trace": {"session_id": "review-1"},
        }
        mock_summary.return_value = {
            "result_markdown": "### Classic Summary",
            "model_id": "codex-cli",
            "trace": {"session_id": "summary-1"},
        }

        review_request = PRDReviewRequest(
            owner_key="anon:test",
            jira_id="AF-CLASSIC",
            jira_link="https://jira/browse/AF-CLASSIC",
            prd_url=self.service.confluence.page.source_url,
        )
        first_review = service.review(review_request)
        cached_review = service.review(review_request)
        summary_request = PRDReviewRequest(
            owner_key="anon:test",
            jira_id="AF-SUMMARY",
            jira_link="https://jira/browse/AF-SUMMARY",
            prd_url=self.service.confluence.page.source_url,
        )
        first_summary = service.summarize(summary_request)
        cached_summary = service.summarize(summary_request)

        mock_review.side_effect = RuntimeError("review failed")
        with self.assertRaisesRegex(ToolError, "review failed"):
            service.review(
                PRDReviewRequest(
                    owner_key="anon:test",
                    jira_id="AF-FAIL",
                    jira_link="",
                    prd_url=self.service.confluence.page.source_url,
                    force_refresh=True,
                )
            )
        failed_review = self.store.get_prd_review_result(
            owner_key="anon:test",
            jira_id="AF-FAIL",
            prd_url=self.service.confluence.page.source_url,
            prd_updated_at=self.service.confluence.page.updated_at,
            prompt_version=PRD_REVIEW_PROMPT_VERSION,
        )

        self.assertFalse(first_review["cached"])
        self.assertTrue(cached_review["cached"])
        self.assertEqual(first_review["review"]["result_markdown"], "### Classic Review")
        self.assertFalse(first_summary["cached"])
        self.assertTrue(cached_summary["cached"])
        self.assertEqual(first_summary["summary"]["result_markdown"], "### Classic Summary")
        self.assertEqual(mock_review.call_count, 2)
        self.assertEqual(mock_summary.call_count, 1)
        self.assertEqual(failed_review["status"], "failed")
        self.assertIn("review failed", failed_review["error"])

    @patch("prd_briefing.reviewer.generate_prd_summary_with_codex")
    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_url_prd_review_and_summary_failures_are_saved_for_retry_context(self, mock_review, mock_summary):
        service = PRDReviewService(
            store=self.store,
            confluence=self.service.confluence,
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )
        mock_review.side_effect = RuntimeError("url review failed")
        mock_summary.side_effect = RuntimeError("url summary failed")

        with self.assertRaisesRegex(ToolError, "url review failed"):
            service.review_url(
                PRDBriefingReviewRequest(
                    owner_key="anon:test",
                    prd_url=self.service.confluence.page.source_url,
                    language="en",
                    force_refresh=True,
                )
            )
        with self.assertRaisesRegex(ToolError, "url summary failed"):
            service.summarize_url(
                PRDBriefingReviewRequest(
                    owner_key="anon:test",
                    prd_url=self.service.confluence.page.source_url,
                    language="en",
                    force_refresh=True,
                )
            )

        failed_review = self.store.get_prd_review_result(
            owner_key="anon:test",
            jira_id=PRD_BRIEFING_REVIEW_CACHE_KEY,
            prd_url=self.service.confluence.page.source_url,
            prd_updated_at=self.service.confluence.page.updated_at,
            prompt_version=prd_briefing_review_prompt_version("en"),
        )
        failed_summary = self.store.get_prd_review_result(
            owner_key="anon:test",
            jira_id=PRD_URL_SUMMARY_CACHE_KEY,
            prd_url=self.service.confluence.page.source_url,
            prd_updated_at=self.service.confluence.page.updated_at,
            prompt_version=prd_summary_prompt_version("en"),
        )
        self.assertEqual(failed_review["status"], "failed")
        self.assertIn("url review failed", failed_review["error"])
        self.assertEqual(failed_summary["status"], "failed")
        self.assertIn("url summary failed", failed_summary["error"])

    @patch("prd_briefing.reviewer.generate_prd_summary_with_codex")
    def test_prd_url_summary_cache_varies_by_language_and_updated_at(self, mock_generate):
        mock_generate.return_value = {
            "result_markdown": "### Summary",
            "model_id": "codex-cli",
            "trace": {"session_id": "s1"},
        }
        service = PRDReviewService(
            store=self.store,
            confluence=self.service.confluence,
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        first = service.summarize_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="zh",
            )
        )
        cached = service.summarize_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="zh",
            )
        )
        english = service.summarize_url(
            PRDBriefingReviewRequest(
                owner_key="anon:test",
                prd_url="https://example.atlassian.net/wiki/pages/123",
                language="en",
            )
        )

        self.assertFalse(first["cached"])
        self.assertTrue(cached["cached"])
        self.assertFalse(english["cached"])
        self.assertEqual(mock_generate.call_count, 2)
        self.assertEqual(first["summary"]["jira_id"], PRD_URL_SUMMARY_CACHE_KEY)
        self.assertEqual(first["summary"]["prompt_version"], prd_summary_prompt_version("zh"))
        self.assertEqual(english["summary"]["prompt_version"], prd_summary_prompt_version("en"))

        stale = self.store.get_prd_review_result(
            owner_key="anon:test",
            jira_id=PRD_URL_SUMMARY_CACHE_KEY,
            prd_url=self.service.confluence.page.source_url,
            prd_updated_at="2026-04-16T10:00:00Z",
            prompt_version=prd_summary_prompt_version("zh"),
        )
        self.assertIsNone(stale)

    @patch("prd_briefing.reviewer.generate_prd_summary_with_codex")
    def test_prd_summary_failure_is_saved_for_retry_context(self, mock_generate):
        mock_generate.side_effect = RuntimeError("summary failed")
        service = PRDReviewService(
            store=self.store,
            confluence=self.service.confluence,
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        with self.assertRaisesRegex(ToolError, "summary failed"):
            service.summarize(
                PRDReviewRequest(
                    owner_key="anon:test",
                    jira_id="AF-456",
                    jira_link="",
                    prd_url="https://example.atlassian.net/wiki/pages/123",
                    force_refresh=True,
                )
            )

        failed = self.store.get_prd_review_result(
            owner_key="anon:test",
            jira_id="AF-456",
            prd_url=self.service.confluence.page.source_url,
            prd_updated_at=self.service.confluence.page.updated_at,
            prompt_version=PRD_SUMMARY_PROMPT_VERSION,
        )
        self.assertEqual(failed["status"], "failed")
        self.assertIn("summary failed", failed["error"])

    @patch("prd_briefing.reviewer.generate_prd_summary_with_codex")
    def test_long_prd_summary_uses_hybrid_batches_and_covers_late_sections(self, mock_generate):
        long_page = IngestedConfluencePage(
            page_id="long",
            title="Long PRD",
            source_url="https://example.atlassian.net/wiki/pages/long",
            updated_at="2026-05-10T00:00:00Z",
            language="en",
            sections=[
                ParsedSection(title="Early", section_path="1 Early", content="A" * 95_000),
                ParsedSection(title="Late", section_path="2 Late", content="LATE_SECTION_MARKER"),
            ],
        )
        service = PRDReviewService(
            store=self.store,
            confluence=FakeConnector(long_page),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        def generate_side_effect(*, prompt, **kwargs):
            if "# Batch Summaries" in prompt:
                self.assertIn("late batch summary", prompt)
                return {"result_markdown": "### Final Summary", "model_id": "codex-cli", "trace": {"mode": "final"}}
            if "LATE_SECTION_MARKER" in prompt:
                return {"result_markdown": "late batch summary", "model_id": "codex-cli", "trace": {"mode": "batch"}}
            return {"result_markdown": "early batch summary", "model_id": "codex-cli", "trace": {"mode": "batch"}}

        mock_generate.side_effect = generate_side_effect
        result = service.summarize(
            PRDReviewRequest(owner_key="anon:test", jira_id="AF-999", jira_link="", prd_url=long_page.source_url)
        )

        self.assertEqual(result["coverage"]["mode"], "hybrid")
        self.assertEqual(result["coverage"]["sections_covered"], 2)
        self.assertFalse(result["coverage"]["truncated"])
        self.assertEqual(result["summary"]["result_markdown"], "### Final Summary")
        self.assertGreaterEqual(mock_generate.call_count, 3)

    @patch("prd_briefing.reviewer.generate_prd_review_with_codex")
    def test_long_prd_review_uses_hybrid_batches_and_synthesizes_priorities(self, mock_generate):
        long_page = IngestedConfluencePage(
            page_id="long-review",
            title="Long PRD Review",
            source_url="https://example.atlassian.net/wiki/pages/long-review",
            updated_at="2026-05-10T00:00:00Z",
            language="en",
            sections=[
                ParsedSection(title="Early", section_path="1 Early", content="A" * 95_000),
                ParsedSection(title="Late", section_path="2 Late", content="LATE_REVIEW_MARKER"),
            ],
        )
        service = PRDReviewService(
            store=self.store,
            confluence=FakeConnector(long_page),
            settings=self._settings(),
            workspace_root=Path(self.temp_dir.name),
        )

        def generate_side_effect(*, prompt, **kwargs):
            if "# Batch Reviews" in prompt:
                self.assertIn("late review gap", prompt)
                return {"result_markdown": "### Executive Verdict", "model_id": "codex-cli", "trace": {"mode": "final"}}
            if "LATE_REVIEW_MARKER" in prompt:
                return {"result_markdown": "late review gap", "model_id": "codex-cli", "trace": {"mode": "batch"}}
            return {"result_markdown": "early review gap", "model_id": "codex-cli", "trace": {"mode": "batch"}}

        mock_generate.side_effect = generate_side_effect
        result = service.review_url(
            PRDBriefingReviewRequest(owner_key="anon:test", prd_url=long_page.source_url, language="en")
        )

        self.assertEqual(result["coverage"]["mode"], "hybrid")
        self.assertEqual(result["coverage"]["sections_covered"], 2)
        self.assertEqual(result["review"]["result_markdown"], "### Executive Verdict")
        self.assertGreaterEqual(mock_generate.call_count, 3)

    def test_unsupported_question_is_declined(self):
        payload = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )

        answer = self.service.answer_question(
            session_id=payload["session"]["session_id"],
            owner_key="anon:test",
            question="What is the pricing model in Brazil?",
        )

        self.assertEqual(answer["groundedness"], "unsupported")
        self.assertIn("找到这个答案", answer["answer_text"])

    def test_developer_walkthrough_prompt_is_engineering_focused(self):
        self.text_client.is_configured = lambda: True

        payload = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )
        self.text_client.answer_calls = 0
        self.text_client.last_system_prompt = None
        self.text_client.last_user_prompt = None

        result, cached = self.service._compose_walkthrough_section(  # noqa: SLF001
            owner_key="anon:prompt-test",
            section=payload["sections"][0],
        )

        self.assertEqual(result, "LLM answer")
        self.assertFalse(cached)
        self.assertIn("software engineers", self.text_client.last_system_prompt)
        self.assertIn("validation rules", self.text_client.last_system_prompt)
        self.assertIn("implementation", self.text_client.last_user_prompt)
        self.assertIn("这一块主要是", self.text_client.last_user_prompt)

    def test_walkthrough_script_is_cached_after_first_text_model_call(self):
        self.text_client.is_configured = lambda: True
        payload = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )
        self.text_client.answer_calls = 0
        self.text_client.last_system_prompt = None
        self.text_client.last_user_prompt = None

        first = self.service.narrate_section(
            session_id=payload["session"]["session_id"],
            owner_key="anon:test",
            section_index=0,
            include_audio=False,
        )
        second = self.service.narrate_section(
            session_id=payload["session"]["session_id"],
            owner_key="anon:test",
            section_index=0,
            include_audio=False,
        )

        self.assertEqual(first["script"], "LLM answer")
        self.assertEqual(second["script"], "LLM answer")
        self.assertFalse(first["cached"])
        self.assertTrue(second["cached"])
        self.assertEqual(self.text_client.answer_calls, 1)

    def test_create_session_supports_english_walkthrough_prompt_and_cache(self):
        self.text_client.is_configured = lambda: True
        payload = self.service.create_session(
            owner_key="anon:english",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
            language="en",
        )
        self.text_client.answer_calls = 0
        self.text_client.last_system_prompt = None
        self.text_client.last_user_prompt = None

        first = self.service.narrate_section(
            session_id=payload["session"]["session_id"],
            owner_key="anon:english",
            section_index=0,
            include_audio=False,
        )
        second = self.service.narrate_section(
            session_id=payload["session"]["session_id"],
            owner_key="anon:english",
            section_index=0,
            include_audio=False,
        )

        self.assertEqual(payload["session"]["audience"], "developer_en")
        self.assertEqual(first["language"], "en")
        self.assertFalse(first["cached"])
        self.assertTrue(second["cached"])
        self.assertEqual(self.text_client.answer_calls, 1)
        self.assertIn("software engineers in English", self.text_client.last_system_prompt)
        self.assertIn("around 5 to 9 sentences in English", self.text_client.last_user_prompt)

    def test_english_walkthrough_audio_uses_english_language_code(self):
        self.text_client.is_configured = lambda: True
        payload = self.service.create_session(
            owner_key="anon:english-audio",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
            language="en",
        )

        result = self.service.narrate_section(
            session_id=payload["session"]["session_id"],
            owner_key="anon:english-audio",
            section_index=0,
            include_audio=True,
        )

        self.assertEqual(result["language"], "en")
        self.assertEqual(self.voice_service.synthesize_calls[-1]["language_code"], "en")

    def test_tts_text_is_not_truncated_for_edge_voice(self):
        long_text = " ".join([f"Sentence {index} keeps implementation detail." for index in range(80)])

        optimized = optimize_tts_text(long_text, language_code="en")

        self.assertEqual(optimized, long_text)
        self.assertGreater(len(optimized), 520)

    def test_pm_briefing_blocks_filter_metadata_and_merge_related_sections(self):
        sections = [
            {"section_path": "1.1 Version Control", "content": "Version v0.1 PIC someone@example.com"},
            {"section_path": "3.5.1 Search Layout", "content": "The page shows search criteria and default fields. User can click Search button."},
            {"section_path": "3.5.2 Detail Fields", "content": "The detail page displays readonly fields and required form values."},
            {"section_path": "3.6.1 Submit Review", "content": "User can submit for review and status changes from Draft to Pending Review."},
            {"section_path": "3.6.2 Reopen", "content": "User can click Reopen and status changes back to Draft."},
        ]

        blocks = build_pm_briefing_blocks(sections)

        self.assertFalse(any(0 in block["section_indexes"] for block in blocks))
        ui_block = next(block for block in blocks if block["title"] == "页面布局和字段规则")
        state_block = next(block for block in blocks if block["title"] == "状态流转和操作动作")
        self.assertEqual(ui_block["section_indexes"], [1, 2])
        self.assertEqual(state_block["section_indexes"], [3, 4])
        self.assertTrue(ui_block["source_refs"])

    def test_pm_briefing_blocks_use_english_labels_for_english_walkthrough(self):
        sections = [
            {"section_path": "3.5.1 Search Layout", "content": "The page shows search criteria and default fields. User can click Search button."},
            {"section_path": "3.6.1 Submit Review", "content": "User can submit for review and status changes from Draft to Pending Review."},
        ]

        blocks = build_pm_briefing_blocks(sections, language="en")

        self.assertTrue(any(block["title"] == "Page Layout and Field Rules" for block in blocks))
        self.assertTrue(any(block["title"] == "Status Transitions and Actions" for block in blocks))
        self.assertTrue(all("：" not in block["merged_summary"] for block in blocks))
        self.assertTrue(all("说明" not in block["briefing_goal"] for block in blocks))

    def test_pm_briefing_blocks_split_large_related_groups(self):
        sections = [
            {
                "section_path": f"3.8.{index} Report Download",
                "content": "User can download report history and audit export.",
                "html_content": "<table>" + ("<tr><td>Report history</td></tr>" * 20) + "</table>",
            }
            for index in range(1, 7)
        ]

        blocks = build_pm_briefing_blocks(sections)
        reporting_blocks = [block for block in blocks if block["title"].startswith("报表、下载和历史记录")]

        self.assertGreaterEqual(len(reporting_blocks), 2)
        self.assertTrue(all(len(block["section_indexes"]) <= 4 for block in reporting_blocks))
        self.assertEqual(
            [index for block in reporting_blocks for index in block["section_indexes"]],
            list(range(6)),
        )

    def test_pm_briefing_blocks_fall_back_when_sections_have_low_signal(self):
        sections = [
            {"section_path": "Appendix A", "content": "tiny"},
            {"section_path": "Appendix B", "content": "also tiny"},
        ]

        blocks = build_pm_briefing_blocks(sections, language="en")

        self.assertEqual(len(blocks), 1)
        self.assertEqual(blocks[0]["title"], "Core Feature Requirements")
        self.assertEqual(blocks[0]["section_indexes"], [0, 1])
        self.assertIn("implementation capability", blocks[0]["merged_summary"])

    def test_briefing_category_summary_and_overview_helpers_cover_edge_cases(self):
        self.assertEqual(classify_briefing_category("3.1 Audit Report", "download export history"), "reporting")
        self.assertEqual(classify_briefing_category("3.2 Submit Review", "status changes to approved"), "state_actions")
        self.assertEqual(classify_briefing_category("3.3 Field Layout", "readonly tab button"), "ui_rules")
        self.assertEqual(classify_briefing_category("3.4 Permission", "role must not fail validation"), "permission_edge")
        self.assertEqual(classify_briefing_category("3.5 Journey", "click navigation entry"), "workflow")
        self.assertEqual(classify_briefing_category("3.6 Feature", "core capability"), "feature")

        self.assertIn(
            "Review Empty Section",
            build_block_summary("Block", [{"section_path": "Empty Section", "content": ""}], language="en"),
        )
        self.assertIn(
            "system grouped",
            build_block_summary("Block", [{"section_path": "", "content": ""}], language="en"),
        )
        self.assertIn(
            "建议开发优先从主流程",
            build_detail_grounded_overview("Sparse PRD", [{"section_path": "1.1 Version Control", "content": "v1 PIC"}]),
        )
        self.assertEqual(build_scope_items([{"section_path": "1.1 Version Control", "content": "PIC"}]), ["主流程", "页面交互", "状态流转", "规则校验"])

    def test_detail_point_helpers_cover_all_fallback_categories(self):
        sections = [{"section_path": "1.1 Version Control", "content": "PIC someone@example.com"}]

        self.assertIn("页面动作", extract_detail_points(sections, category="developer")[0])
        self.assertIn("页面展示", extract_detail_points(sections, category="frontend")[0])
        self.assertIn("状态流转", extract_detail_points(sections, category="backend")[0])
        self.assertIn("理解偏差", extract_detail_points(sections, category="risks")[0])
        self.assertIn("强校验", extract_detail_points(sections, category="unclear_rules")[0])
        self.assertIn("异常路径", extract_detail_points(sections, category="missing_edge_cases")[0])
        self.assertIn("职责边界", extract_detail_points(sections, category="unclear_ownership")[0])
        self.assertIn("默认值", extract_detail_points(sections, category="open_questions")[0])

        localized = localize_detail_point_zh('User can click on "Submit" button to submit assessment')
        self.assertIn("用户可点击", localized)
        self.assertTrue(localized.endswith("。"))

    def test_prompt_and_retrieval_helpers_cover_empty_and_english_paths(self):
        self.assertEqual(RetrievalService().rank("anything", []), [])
        self.assertIn("natural spoken English", build_walkthrough_section_system_prompt("en"))
        section_prompt = build_walkthrough_section_user_prompt(
            section={"section_path": "3.1 Flow", "content": "User can submit review.", "briefing_summary": "Submit flow"},
            notes=["Default status is Draft"],
            language="en",
        )
        self.assertIn("Write a natural spoken script", section_prompt)
        block_prompt = build_walkthrough_block_user_prompt(
            block={
                "title": "Submit Review",
                "briefing_goal": "Explain submit flow",
                "merged_summary": "Submit changes status.",
                "developer_focus": ["Status changes"],
            },
            source_lines=["[1] 3.1 Submit\nUser can submit review."],
            language="en",
        )
        self.assertIn("around 7 to 12 sentences in English", block_prompt)

    def test_file_text_and_summary_helpers_cover_fallbacks(self):
        self.assertEqual(safe_filename(" ../Bad Name?.txt "), "..-Bad-Name-.txt")
        self.assertEqual(safe_filename("???"), "upload.bin")
        self.assertEqual(truncate_for_prompt("short text", 20), "short text")
        self.assertEqual(truncate_for_prompt("one two three", 8), "one two…")
        self.assertEqual(format_source_text("Short.\nSecond sentence; third sentence."), "Short.\nSecond sentence\nthird sentence.")
        self.assertEqual(build_briefing_summary("Fallback Title", "Too short."), "Fallback Title")
        self.assertEqual(build_presenter_notes("Only Title", ""), ["Only Title"])
        self.assertEqual(summarize_chunk_for_fallback("", limit=10), "")
        self.assertEqual(summarize_chunk_for_fallback("First sentence is far too long for this limit.", limit=18), "First sentence is...")

        md_path = Path(self.temp_dir.name) / "notes.md"
        html_path = Path(self.temp_dir.name) / "notes.html"
        json_path = Path(self.temp_dir.name) / "notes.json"
        unsupported_path = Path(self.temp_dir.name) / "notes.pdf"
        md_path.write_text("# Title\nBody", encoding="utf-8")
        html_path.write_text("<html><body><h1>Title</h1><p>Body</p></body></html>", encoding="utf-8")
        json_path.write_text('{"title":"Body"}', encoding="utf-8")

        self.assertIn("Body", extract_text(md_path))
        self.assertIn("Title", extract_text(html_path))
        self.assertIn("title", extract_text(json_path))
        with self.assertRaisesRegex(ValueError, "Supported knowledge-base files"):
            extract_text(unsupported_path)

    def test_overview_and_script_helpers_cover_low_signal_branches(self):
        self.assertEqual(normalize_overview_list("not-list"), [])
        self.assertEqual(normalize_overview_list(["PIC: user@example.com", "Submit button"]), ["Submit button"])
        self.assertEqual(normalize_detail_sentence("v1.0"), "")
        self.assertEqual(
            normalize_detail_sentence("  valid product sentence with enough detail  "),
            "valid product sentence with enough detail",
        )

        parsed = parse_developer_overview_payload("```json\n{\"background_goal\":\"Goal\",\"implementation_overview\":\"Build flow\"}\n```")
        self.assertEqual(parsed["overview"], "Goal Build flow")
        fallback = parse_developer_overview_payload("not json")
        self.assertEqual(fallback["overview"], "not json")

        self.assertTrue(overview_is_low_signal({"overview": "", "developer_focus": ["Submit button"]}))
        self.assertTrue(overview_is_low_signal({"overview": "Version v1.0", "developer_focus": ["Submit button"]}))
        self.assertTrue(overview_is_low_signal({"overview": "Overview", "developer_focus": [], "scope": []}))

        self.assertEqual(describe_section_topic_zh("Navigation Menu", ""), "页面入口、导航方式和页面切换")
        self.assertEqual(describe_section_topic_zh("Layout Field", ""), "页面布局、字段规则和展示方式")
        self.assertEqual(describe_section_topic_zh("Requirement Rule", ""), "这部分规则要求和校验逻辑")
        self.assertEqual(describe_section_topic_zh("Assessment Review", ""), "业务处理过程里的关键动作和状态变化")
        self.assertIn("页面布局", infer_engineering_focus_zh("Field Layout", "default visible field")[0])
        self.assertIn(
            "主流程",
            build_developer_zh_fallback_script(
                "Generic",
                "No concrete product tokens here.",
                [],
            ),
        )

    def test_module_inference_and_tts_timing_helpers_cover_boundaries(self):
        self.assertEqual(
            infer_impacted_modules(["Workflow Navigation", "Field Detail", "Submit Review", "Report Download"]),
            ["流程和页面跳转", "页面布局和字段交互", "状态流转和操作动作", "报表、下载或历史记录"],
        )
        self.assertEqual(
            infer_impacted_modules_from_sections(
                [
                    {"section_path": "Version Control", "content": "v1"},
                    {"section_path": "3.2 Search Detail Tab", "content": "field layout"},
                    {"section_path": "3.3 Submit Status", "content": "status change"},
                ]
            ),
            ["3.2 Search Detail Tab", "3.3 Submit Status"],
        )

        self.assertEqual(split_presentation_sentences("第一句。第二句；Third?"), ["第一句。", "第二句；", "Third?"])
        self.assertEqual(estimate_tts_duration_seconds("", language_code="zh"), 1.0)
        self.assertGreater(estimate_tts_duration_seconds("a" * 1000, language_code="en"), 1.5)
        self.assertEqual(
            duration_from_edge_boundaries(
                [
                    {"offset": 10_000_000, "duration": 20_000_000},
                    {"Offset": "bad", "Duration": 10_000_000},
                ]
            ),
            3.0,
        )
        self.assertIsNone(duration_from_edge_boundaries([{"offset": "bad"}]))
        timestamps = build_sentence_timestamps("短句。更长的第二句。", duration_seconds=10)
        self.assertEqual(timestamps[0]["start"], 0.0)
        self.assertEqual(timestamps[-1]["end"], 10.0)

    def test_narrate_briefing_block_uses_block_payload_cache(self):
        self.text_client.is_configured = lambda: True
        page = IngestedConfluencePage(
            page_id="456",
            title="Assessment PRD",
            source_url="https://example.atlassian.net/wiki/pages/456",
            updated_at="2026-04-16T10:00:00Z",
            language="en",
            sections=[
                ParsedSection(
                    title="Search Layout",
                    section_path="3.5.1 Search Layout",
                    content="The page shows search criteria and default fields. User can click Search button.",
                    html_content="<p>The page shows search criteria and default fields.</p>",
                    image_refs=[],
                ),
                ParsedSection(
                    title="Detail Fields",
                    section_path="3.5.2 Detail Fields",
                    content="The detail page displays readonly fields and required form values.",
                    html_content="<p>The detail page displays readonly fields and required form values.</p>",
                    image_refs=[],
                ),
            ],
        )
        self.service.confluence = FakeConnector(page)
        payload = self.service.create_session(
            owner_key="anon:block-test",
            page_ref="https://example.atlassian.net/wiki/pages/456",
            mode="walkthrough",
        )
        block_id = payload["briefing_blocks"][0]["block_id"]
        self.text_client.answer_calls = 0

        first = self.service.narrate_section(
            session_id=payload["session"]["session_id"],
            owner_key="anon:block-test",
            briefing_block_id=block_id,
            include_audio=False,
        )
        second = self.service.narrate_section(
            session_id=payload["session"]["session_id"],
            owner_key="anon:block-test",
            briefing_block_id=block_id,
            include_audio=False,
        )

        self.assertEqual(first["script"], "LLM answer")
        self.assertEqual(first["briefing_block_id"], block_id)
        self.assertEqual(first["section_indexes"], [0, 1])
        self.assertFalse(first["cached"])
        self.assertTrue(second["cached"])
        self.assertEqual(self.text_client.answer_calls, 1)

    def test_walkthrough_script_reuses_legacy_cached_entry_from_old_model_id(self):
        self.text_client.is_configured = lambda: True
        payload = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )
        sections = payload["sections"]
        notes = sections[0].get("briefing_notes") or []
        prompt = (
            "You are a product manager briefing a PRD section to software engineers in Mandarin Chinese. "
            "Speak the way PMs normally align requirements with developers during grooming or walkthrough sessions. "
            "Be direct, practical, and structured. First explain the purpose of this section, then the main flow, "
            "then what developers need to build or pay attention to. Call out scope, user actions, system behavior, "
            "validation rules, dependencies, edge cases, and any implementation-sensitive details when present. "
            "Do not sound like a keynote presenter. Do not read the PRD word for word. "
            "Do not mechanically read every field, bullet, or table row. Summarize dense tables into what engineering should understand. "
            "Use spoken PM phrasing that feels normal in a dev sync, for example framing the goal first, then saying what the flow is, "
            "what changes on the page, what gets triggered, what should be validated, and what cases developers need to pay attention to."
        )
        body = build_walkthrough_section_user_prompt(section=sections[0], notes=notes, language="zh")
        section_payload = json.dumps(
            {
                "section_path": sections[0]["section_path"],
                "briefing_summary": sections[0].get("briefing_summary", ""),
                "briefing_notes": notes,
                "content": sections[0]["content"],
                "prompt": prompt,
                "body": body,
            },
            ensure_ascii=False,
            sort_keys=True,
        )
        with self.store.connect() as conn:
            conn.execute(
                """
                delete from briefing_script_cache
                where owner_key = ? and audience = ? and prompt_version = ?
                """,
                ("anon:test", "developer_zh", WALKTHROUGH_SCRIPT_PROMPT_VERSION),
            )
        self.store.cache_script(
            owner_key="anon:test",
            audience="developer_zh",
            model_id="legacy:text-model",
            prompt_version=WALKTHROUGH_SCRIPT_PROMPT_VERSION,
            section_payload=section_payload,
            script="legacy cached script",
        )
        self.text_client.answer_calls = 0

        result = self.service.narrate_section(
            session_id=payload["session"]["session_id"],
            owner_key="anon:test",
            section_index=0,
            include_audio=False,
        )

        self.assertEqual(result["script"], "legacy cached script")
        self.assertTrue(result["cached"])
        self.assertFalse(result["audio_cached"])
        self.assertEqual(self.text_client.answer_calls, 0)

    def test_get_session_payload_marks_cached_sections(self):
        self.text_client.is_configured = lambda: True
        payload = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )

        self.service.narrate_section(
            session_id=payload["session"]["session_id"],
            owner_key="anon:test",
            section_index=0,
            include_audio=False,
        )

        refreshed = self.service.get_session_payload(
            session_id=payload["session"]["session_id"],
            owner_key="anon:test",
        )

        self.assertTrue(refreshed["sections"][0]["walkthrough_cached"])
        self.assertFalse(refreshed["sections"][0]["walkthrough_audio_cached"])
        self.assertFalse(refreshed["sections"][1]["walkthrough_cached"])
        self.assertFalse(refreshed["sections"][1]["walkthrough_audio_cached"])

    def test_get_session_payload_marks_cached_audio_sections(self):
        self.text_client.is_configured = lambda: True
        payload = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )

        result = self.service.narrate_section(
            session_id=payload["session"]["session_id"],
            owner_key="anon:test",
            section_index=0,
            include_audio=False,
        )
        self.voice_service.cached_texts.add(result["script"])

        refreshed = self.service.get_session_payload(
            session_id=payload["session"]["session_id"],
            owner_key="anon:test",
        )

        self.assertTrue(refreshed["sections"][0]["walkthrough_cached"])
        self.assertTrue(refreshed["sections"][0]["walkthrough_audio_cached"])

    def test_walkthrough_script_requires_text_provider(self):
        payload = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )

        with self.assertRaisesRegex(RuntimeError, "Codex to be configured"):
            self.service.narrate_section(
                session_id=payload["session"]["session_id"],
                owner_key="anon:test",
                section_index=0,
                include_audio=False,
            )

    def test_session_overview_filters_metadata_noise(self):
        parsed = parse_session_overview(
            """
            {
              "overview": "这是一个总览",
              "scope": ["Version: v0.1", "Add New Assessment 页面和提交流程"],
              "impacted_modules": ["Date: 24 Nov 2025", "搜索区和详情区"],
              "developer_focus": ["PIC: test@example.com", "状态切换和自动回填"],
              "frontend_focus": [],
              "backend_focus": [],
              "risks": [],
              "unclear_rules": [],
              "missing_edge_cases": [],
              "unclear_ownership": [],
              "open_questions": []
            }
            """
        )
        self.assertEqual(parsed["scope"], ["Add New Assessment 页面和提交流程"])
        self.assertEqual(parsed["impacted_modules"], ["搜索区和详情区"])
        self.assertEqual(parsed["developer_focus"], ["状态切换和自动回填"])

    def test_session_overview_prioritizes_feature_sections(self):
        ranked = select_sections_for_overview([
            {"section_path": "1.2 People Involved", "content": "Role Regional Requester PIC someone@example.com"},
            {"section_path": "2.1 Background", "content": "Background and context"},
            {"section_path": "3.5.2 Add New Assessment", "content": "User can click Add New Assessment button and system auto populate Event"},
            {"section_path": "3.6.8 Reopen SSA", "content": "User can click Reopen and status changes to Draft"},
        ])
        self.assertEqual(ranked[0]["section_path"], "3.5.2 Add New Assessment")
        self.assertIn(ranked[1]["section_path"], {"3.6.8 Reopen SSA", "3.5.2 Add New Assessment"})

    def test_heuristic_overview_extracts_implementation_details(self):
        overview = build_heuristic_session_overview(
            "Payments PRD",
            [
                {"section_path": "1.2 People Involved", "content": "Version v0.1 PIC someone@example.com"},
                {"section_path": "3.5.2 Add New Assessment", "content": "User can click Add New Assessment button and system auto populate Event. By default only seven search criteria are shown."},
                {"section_path": "3.6.8 Reopen SSA", "content": "User can click Reopen and status changes to Draft. Reopen Date is mandatory when Reopen Date is filled."},
            ],
        )
        self.assertNotIn("Version", "".join(overview["developer_focus"]))
        self.assertTrue(any("Add New Assessment" in item or "auto populate" in item or "seven search criteria" in item for item in overview["developer_focus"]))

    def test_developer_overview_uses_single_summary_shape(self):
        payload = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )
        self.assertTrue(payload["session_overview"]["overview"])
        self.assertIn("background_goal", payload["session_overview"])
        self.assertIn("implementation_overview", payload["session_overview"])
        self.assertEqual(payload["session_overview"]["scope"], [])
        self.assertEqual(payload["session_overview"]["developer_focus"], [])

    def test_developer_overview_does_not_block_on_text_model(self):
        self.text_client.is_configured = lambda: True
        payload = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )
        self.text_client.answer_calls = 0

        refreshed = self.service.get_session_payload(
            session_id=payload["session"]["session_id"],
            owner_key="anon:test",
        )

        self.assertTrue(refreshed["session_overview"]["background_goal"])
        self.assertTrue(refreshed["session_overview"]["implementation_overview"])
        self.assertEqual(self.text_client.answer_calls, 0)

    def test_low_signal_overview_is_detected(self):
        overview = {
            "overview": "这个 PRD 主要围绕核心业务流程、页面操作和规则约束展开。",
            "scope": ["1.1 Version Control"],
            "impacted_modules": ["流程和页面跳转"],
            "developer_focus": ["Version: v0.1", "Date: 24 Nov 2025"],
            "frontend_focus": ["页面展示、字段显隐、按钮状态和交互顺序要先对齐。"],
            "backend_focus": ["重点确认状态流转、默认值、系统自动回填和接口出参约束。"],
            "risks": ["部分规则和边界情况可能分散在多个 section 里，开发实现前需要先对齐。"],
            "unclear_rules": [],
            "missing_edge_cases": [],
            "unclear_ownership": [],
            "open_questions": [],
        }
        self.assertTrue(overview_is_low_signal(overview))

    def test_create_session_uses_heuristic_when_overview_is_low_signal(self):
        self.text_client.is_configured = lambda: True
        self.text_client.create_answer = lambda *args, **kwargs: """
        {
          "overview": "这个 PRD 主要围绕核心业务流程、页面操作和规则约束展开。",
          "scope": ["1.1 Version Control", "1.2 People Involved"],
          "impacted_modules": ["流程和页面跳转"],
          "developer_focus": ["Version: v0.1", "Date: 24 Nov 2025"],
          "frontend_focus": ["页面展示、字段显隐、按钮状态和交互顺序要先对齐。"],
          "backend_focus": ["重点确认状态流转、默认值、系统自动回填和接口出参约束。"],
          "risks": ["部分规则和边界情况可能分散在多个 section 里，开发实现前需要先对齐。"],
          "unclear_rules": [],
          "missing_edge_cases": [],
          "unclear_ownership": [],
          "open_questions": []
        }
        """

        payload = self.service.create_session(
            owner_key="anon:test",
            page_ref="https://example.atlassian.net/wiki/pages/123",
            mode="walkthrough",
        )

        self.assertNotIn("Version", "".join(payload["session_overview"]["developer_focus"]))
        self.assertNotIn("Version Control", "".join(payload["session_overview"]["scope"]))
        self.assertTrue(payload["session_overview"]["overview"])


class PRDBriefingServiceCoverageEdgeTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.store = BriefingStore(Path(self.temp_dir.name))

    def tearDown(self):
        self.temp_dir.cleanup()

    def _voice(self) -> VoiceService:
        return VoiceService(
            store=self.store,
            tts_provider="edge",
            edge_mandarin_voice="zh-Test",
            edge_english_voice="en-Test",
            edge_rate="-10%",
            edge_mandarin_rate="-8%",
            edge_english_rate="-4%",
        )

    def _service(self, *, configured: bool = True) -> PRDBriefingService:
        text_client = FakeTextClient()
        text_client.is_configured = lambda: configured
        page = IngestedConfluencePage(
            page_id="123",
            title="Edge PRD",
            source_url="https://example.test/prd",
            updated_at="2026-05-01T00:00:00Z",
            language="en",
            sections=[
                ParsedSection(
                    title="Flow",
                    section_path="3. Flow",
                    content="User can click Submit button. Status changes to Pending Review.",
                    html_content="",
                )
            ],
            version_number="1",
        )
        return PRDBriefingService(
            store=self.store,
            confluence=FakeConnector(page),
            text_client=text_client,
            voice_service=FakeVoiceService(),
            walkthrough_prewarm_enabled=True,
        )

    def test_voice_service_synthesize_cache_failure_and_cache_target_edges(self):
        voice = self._voice()
        with patch.object(voice, "_synthesize_with_edge_tts", return_value=b"mp3-bytes") as synthesize:
            first = voice.synthesize(session_id="s1", text="你好 API", language_code="zh", owner_key="owner")
            second = voice.synthesize(session_id="s1", text="你好 API", language_code="zh", owner_key="owner")

        self.assertTrue(first.endswith(".mp3"))
        self.assertEqual(second, first)
        self.assertEqual(synthesize.call_count, 1)
        self.assertEqual(voice._edge_voice_for_language("en"), "en-Test")
        self.assertEqual(voice._edge_rate_for_language("en"), "-4%")
        self.assertIsNotNone(voice.get_cached_audio_for_text(owner_key="owner", text="你好 API", language_code="zh"))

        with patch.object(voice, "_synthesize_with_edge_tts", side_effect=RuntimeError("tts down")):
            self.assertIsNone(voice.synthesize(session_id="s1", text="new text", language_code="zh", owner_key="owner"))

        voice.tts_provider = "disabled"
        self.assertIsNone(voice.get_cached_audio_for_text(owner_key="owner", text="new text", language_code="zh"))

    def test_voice_service_presentation_audio_boundaries_and_error_edges(self):
        voice = self._voice()
        with self.assertRaisesRegex(ValueError, "Chunk content"):
            voice.synthesize_presentation_chunk(session_id="s1", owner_key="owner", chunk={}, language_code="zh")

        chunk = {"id": "chunk-1", "title": "标题", "content": "第一句。第二句。", "imageUrls": ["https://example.test/a.png"]}
        with patch.object(voice, "_synthesize_with_edge_tts_with_boundaries", return_value=(b"mp3", [{"offset": 0, "duration": 20_000_000}])):
            result = voice.synthesize_presentation_chunk(
                session_id="s1",
                owner_key="owner",
                chunk=chunk,
                language_code="zh",
            )
        self.assertEqual(result["duration"], 2.0)
        self.assertEqual(result["cacheKey"], "")
        self.assertTrue(result["audioUrl"].startswith("/prd-briefing/assets/"))

        with patch.object(voice, "_synthesize_with_edge_tts_with_boundaries", return_value=(b"", [])):
            with self.assertRaisesRegex(RuntimeError, "did not return audio"):
                voice.synthesize_presentation_chunk(session_id="s1", owner_key="owner", chunk={"content": "No audio"}, language_code="en")

    def test_voice_service_edge_tts_stream_and_running_loop_paths(self):
        class FakeCommunicate:
            def __init__(self, text, voice, rate):
                self.text = text
                self.voice = voice
                self.rate = rate

            async def stream(self):
                yield {"type": "audio", "data": b"a"}
                yield {"type": "WordBoundary", "offset": 10_000_000, "duration": 5_000_000}
                yield {"type": "audio", "data": b"b"}

        previous = sys.modules.get("edge_tts")
        sys.modules["edge_tts"] = types.SimpleNamespace(Communicate=FakeCommunicate)
        voice = self._voice()
        try:
            self.assertEqual(voice._synthesize_with_edge_tts(text="hello", voice_id="en", rate="-4%"), b"ab")
            sync_audio, sync_boundaries = voice._synthesize_with_edge_tts_with_boundaries(text="hello", voice_id="en", rate="-4%")
            self.assertEqual(sync_audio, b"ab")
            self.assertEqual(sync_boundaries[0]["type"], "WordBoundary")
            self.assertEqual(asyncio.run(voice._edge_tts_bytes(text="hello", voice_id="en", rate="-4%")), b"ab")
            audio, boundaries = asyncio.run(voice._edge_tts_bytes_and_boundaries(text="hello", voice_id="en", rate="-4%"))
            self.assertEqual(audio, b"ab")
            self.assertEqual(boundaries[0]["type"], "WordBoundary")

            async def run_inside_loop():
                return voice._run_edge_tts_async(voice._edge_tts_bytes(text="hello", voice_id="en", rate="-4%"))

            self.assertEqual(asyncio.run(run_inside_loop()), b"ab")

            async def fail():
                raise RuntimeError("thread failure")

            async def run_failure_inside_loop():
                return voice._run_edge_tts_async(fail())

            with self.assertRaisesRegex(RuntimeError, "thread failure"):
                asyncio.run(run_failure_inside_loop())
        finally:
            if previous is None:
                sys.modules.pop("edge_tts", None)
            else:
                sys.modules["edge_tts"] = previous

    def test_create_session_spawns_prewarm_when_enabled(self):
        service = self._service(configured=True)
        with patch.object(service, "_spawn_prewarm_walkthrough_scripts") as spawn:
            payload = service.create_session(owner_key="owner", page_ref="https://example.test/prd", mode="walkthrough", language="en")

        spawn.assert_called_once()
        self.assertEqual(payload["session"]["audience"], briefing_service.DEVELOPER_AUDIENCE_EN)

    def test_walkthrough_cache_legacy_error_and_prewarm_edges(self):
        service = self._service(configured=True)
        section = {
            "section_path": "3. Flow",
            "briefing_summary": "Summary",
            "briefing_notes": ["Note"],
            "content": "User can click Submit button.",
        }
        lookup = service._build_walkthrough_cache_lookup(owner_key="owner", section=section)
        self.store.cache_script(
            owner_key="owner",
            audience=briefing_service.walkthrough_audience("zh"),
            model_id="legacy-model",
            prompt_version=WALKTHROUGH_SCRIPT_PROMPT_VERSION,
            section_payload=lookup["section_payload"],
            script="legacy script",
        )
        script, cached = service._compose_walkthrough_section(owner_key="owner", section=section)
        self.assertEqual((script, cached), ("legacy script", True))

        block = build_pm_briefing_blocks([section])[0]
        with patch.object(service.store, "get_cached_script", return_value=None), \
            patch.object(service.store, "get_cached_script_any_model", return_value="legacy block script"):
            block_script, block_cached = service._compose_walkthrough_block(owner_key="owner", block=block)
        self.assertEqual((block_script, block_cached), ("legacy block script", True))

        service.text_client.create_answer = lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("llm down"))
        with self.assertRaisesRegex(RuntimeError, "walkthrough script"):
            service._compose_walkthrough_section(owner_key="owner", section={**section, "content": "new content"})
        with self.assertRaisesRegex(RuntimeError, "walkthrough script"):
            service._compose_walkthrough_block(owner_key="owner", block={**block, "block_id": "new-block"})

        service.text_client.is_configured = lambda: False
        self.assertIsNone(service._prewarm_walkthrough_scripts(owner_key="owner", sections=[section]))
        self.assertIsNone(service._spawn_prewarm_walkthrough_scripts(owner_key="owner", sections=[section]))

        service.text_client.is_configured = lambda: True
        with patch.object(service, "_compose_walkthrough_block", side_effect=RuntimeError("ignore")):
            self.assertIsNone(service._prewarm_walkthrough_scripts(owner_key="owner", sections=[section]))

        service.text_client.is_configured = lambda: False
        with self.assertRaisesRegex(RuntimeError, "requires Codex"):
            service._compose_walkthrough_block(owner_key="owner", block={**block, "block_id": "no-config"})

    def test_spawn_prewarm_starts_daemon_thread_when_configured(self):
        service = self._service(configured=True)
        started = []

        class FakeThread:
            def __init__(self, **kwargs):
                self.kwargs = kwargs
                self.daemon = kwargs.get("daemon")
                self.name = kwargs.get("name")

            def start(self):
                started.append(self.kwargs)

        with patch("prd_briefing.service.threading.Thread", FakeThread):
            service._spawn_prewarm_walkthrough_scripts(
                owner_key="owner",
                sections=[{"section_path": "S", "briefing_notes": ["n"], "image_refs": ["i"], "content": "body"}],
                language="en",
            )

        self.assertEqual(started[0]["name"], "prd-briefing-prewarm")
        self.assertTrue(started[0]["daemon"])

    def test_annotation_error_and_answer_fallback_edges(self):
        service = self._service(configured=False)
        section = {"section_path": "3. Flow", "briefing_summary": "", "briefing_notes": [], "content": "Content"}
        with patch.object(service, "_build_walkthrough_cache_lookup", side_effect=RuntimeError("cache down")):
            annotated = service._annotate_section_cache(owner_key="owner", section=section)
        self.assertFalse(annotated["walkthrough_cached"])

        block = {"block_id": "b1", "title": "Block", "source_refs": [], "section_indexes": []}
        with patch.object(service, "_build_walkthrough_block_cache_lookup", side_effect=RuntimeError("cache down")):
            annotated_block = service._annotate_block_cache(owner_key="owner", block=block)
        self.assertFalse(annotated_block["walkthrough_cached"])

        self.assertIn("没有检索到", service._build_fallback_answer(chunks=[], groundedness="unsupported"))
        empty_chunk = briefing_service.ChunkRecord(
            source_id=1,
            owner_key="owner",
            session_id="s1",
            source_type="prd",
            title="PRD",
            section_path="Empty",
            content="",
            html_content="",
            image_refs=[],
            source_url="",
            updated_at="",
        )
        self.assertIn("偏保守的推断", service._build_fallback_answer(chunks=[empty_chunk], groundedness="inference"))

        scored = briefing_service.ChunkRecord(
            source_id=1,
            owner_key="owner",
            session_id="s1",
            source_type="prd",
            title="PRD",
            section_path="Flow",
            content="alpha beta",
            html_content="",
            image_refs=[],
            source_url="",
            updated_at="",
        )
        self.assertEqual(RetrievalService().rank("", [scored]), [])
        self.assertEqual(briefing_service.keyword_score("alpha gamma", "alpha beta"), 0.5)
        scored.score = 0.0
        self.assertFalse(briefing_service.chunk_has_signal(scored))

        inference_service = self._service(configured=False)
        inference_service.confluence.page.sections[0].content = "alpha"
        payload = inference_service.create_session(owner_key="owner", page_ref="https://example.test/prd", mode="walkthrough")
        answer = inference_service.answer_question(
            session_id=payload["session"]["session_id"],
            owner_key="owner",
            question="alpha beta gamma delta epsilon zeta eta theta iota kappa",
        )
        self.assertEqual(answer["groundedness"], "inference")

    def test_overview_and_briefing_helper_edge_branches(self):
        sections = [
            {"section_path": "1.1 Version Control", "content": "v1 owner date"},
            {"section_path": "3. Detail Layout", "content": "User can click Search button. Field display default status must be reviewed carefully."},
            {"section_path": "4. Report Download", "content": "System download history report for audit."},
        ]
        overview = build_heuristic_session_overview("Title", sections)
        self.assertTrue(overview["scope"])
        self.assertIn("页面布局和字段交互", infer_impacted_modules(["Detail layout field"]))
        self.assertTrue(infer_impacted_modules_from_sections(sections))
        self.assertIn("主流程", build_scope_items([])[0])
        self.assertIn("PRD《Empty》", build_detail_grounded_overview("Empty", []))
        self.assertEqual(briefing_service.dedupe_non_empty(["", "A", "a", "B"]), ["A", "B"])
        self.assertFalse(briefing_service.has_feature_level_signal("metadata only"))
        self.assertEqual(briefing_service.normalize_overview_list("bad"), [])
        self.assertTrue(briefing_service.looks_like_metadata_noise("PIC owner@example.com v1.2 24 Nov 2025"))

        self.assertEqual(build_presenter_notes("Title", "Title. - . Valid long presenter note."), ["Valid long presenter note."])
        self.assertEqual(build_presenter_notes("Title", "--------. Valid long note."), ["Valid long note."])
        many_notes = build_presenter_notes(
            "Title",
            "First valid long note. first valid long note. - . Second valid long note. Third valid long note. Fourth valid long note. Fifth valid long note. Sixth valid long note.",
        )
        self.assertEqual(len(many_notes), 5)
        self.assertNotIn("Title", many_notes)
        self.assertEqual(build_presenter_notes("Fallback", ""), ["Fallback"])
        self.assertEqual(build_briefing_summary("Short title", "tiny"), "Short title")
        self.assertEqual(briefing_service.tokenize("the API and status"), {"api", "status"})
        self.assertEqual(briefing_service.keyword_score("", "content"), 0.0)

        metadata_only = [{"section_path": "1.1 Version Control", "content": "v1"}]
        self.assertTrue(build_pm_briefing_blocks(metadata_only))
        many = [
            {"section_path": f"3.{idx} Field Rules", "content": "Field display default status must be reviewed carefully.", "html_content": "x" * 10}
            for idx in range(6)
        ]
        blocks = build_pm_briefing_blocks(many, language="en")
        self.assertTrue(any(block["title"].endswith("1") for block in blocks))
        self.assertEqual(len(briefing_service.split_briefing_entries([(0, {"content": "x", "html_content": "x" * (briefing_service.MAX_BRIEFING_BLOCK_HTML_CHARS + 1)}, 1), (1, {"content": "y"}, 1)])), 2)

        self.assertIn("Review", build_block_summary("Feature", [{"section_path": "Only Title", "content": ""}], language="en"))
        self.assertIn("系统已", build_block_summary("功能", [], language="zh"))
        self.assertEqual(briefing_service.classify_briefing_category("Entry", "User click navigation entry"), "workflow")

    def test_text_normalization_and_script_helper_edge_branches(self):
        self.assertEqual(briefing_service.extract_candidate_sentences([{"section_path": "1.1 Version", "content": "Owner Date"}], limit=3), [])
        self.assertEqual(normalize_detail_sentence("abc 123"), "")
        self.assertEqual(normalize_detail_sentence(""), "")
        self.assertEqual(briefing_service.to_brief_point("x" * 100), "x" * 89 + "…")
        self.assertEqual(localize_detail_point_zh(""), "")
        self.assertIn("用户可点击", localize_detail_point_zh('User can click on "Submit" button to submit assessment.'))
        self.assertIn("这一段需求说明", describe_section_topic_zh("Title", "plain content"))
        self.assertIn("主流程", infer_engineering_focus_zh("Title", "plain content")[0])
        self.assertEqual(briefing_service.derive_source_signals_zh("plain"), [])

        self.assertIn("字段比较多", build_developer_zh_fallback_script("Layout", "field layout search criteria", ["note"]))
        self.assertIn("交互链路", build_developer_zh_fallback_script("Flow", "click submit expand collapse workflow", []))
        self.assertIn("核心规则", build_developer_zh_fallback_script("Plain", "plain", []))

        self.assertEqual(briefing_service.split_text_fragments("short. long enough fragment."), ["long enough fragment."])
        self.assertEqual(build_sections_from_text("", chunk_size=5), [])
        self.assertTrue(build_sections_from_text("one two three four five", chunk_size=8))
        long_sections = [ParsedSection(title="T", section_path="S", content="x" * (briefing_service.PRESENTATION_MAX_SOURCE_CHARS + 100), image_refs=["https://x/img.png"])]
        self.assertLessEqual(len(build_presentation_source_text(long_sections)), briefing_service.PRESENTATION_MAX_SOURCE_CHARS)

        self.assertEqual(briefing_service.infer_focus_items(["Detail table"], ["search field tab"], target="frontend")[-1], "如果页面字段或表格很多，建议先拆清搜索区、结果区和详情区。")
        self.assertEqual(briefing_service.infer_focus_items(["Report"], ["download approval submit"], target="backend")[-1], "涉及下载、提交、审批或审计记录的逻辑，后端规则要先定清楚。")
        self.assertIn("校验规则", infer_engineering_focus_zh("Validation", "required must cannot only")[0])
        self.assertTrue(any("默认值" in item for item in briefing_service.derive_source_signals_zh("By default the field is hidden.")))
        self.assertIn("哪些字段显示", briefing_service.derive_source_signals_zh("show display visible hidden")[0])

    def test_presentation_parsing_media_and_timing_edges(self):
        self.assertEqual(extract_json_array_text("prefix [{\"id\":\"1\"}] suffix"), '[{"id":"1"}]')
        self.assertEqual(extract_json_array_text("```json\n[{\"id\":\"1\"}]\n```"), '[{"id":"1"}]')
        self.assertEqual(extract_json_array_text("no array"), "no array")
        self.assertEqual(normalize_image_urls("bad"), [])
        self.assertEqual(normalize_image_urls(["", "ftp://bad", "https://x/a.png", "https://x/a.png"]), ["https://x/a.png"])
        self.assertEqual(normalize_presentation_media({"type": "video", "content": "x"}), {"type": "none", "content": ""})
        self.assertEqual(briefing_service.normalize_media_ref("see MEDIA_ID_42 now"), "MEDIA_ID_42")
        self.assertEqual(briefing_service.normalize_media_ref("none"), "")
        self.assertEqual(briefing_service.build_presentation_cache_key("page", ""), "page")

        chunks = attach_presentation_media(
            [{"content": "body", "mediaRef": "MEDIA_ID_1", "imageUrls": ["https://x/old.png"]}],
            {"MEDIA_ID_1": {"type": "image", "content": "/prd-briefing/image-proxy?src=x"}},
        )
        self.assertEqual(chunks[0]["media_ref"], "MEDIA_ID_1")
        self.assertEqual(chunks[0]["imageUrls"][0], "/prd-briefing/image-proxy?src=x")

        with self.assertRaisesRegex(ValueError, "usable chunks"):
            normalize_presentation_chunks([{"content": ""}])
        normalized = normalize_presentation_chunks([{"content": "中文api测试", "mediaRef": "MEDIA_ID_2"}])
        self.assertIn("API", normalized[0]["content"])
        self.assertEqual(normalized[0]["media_ref"], "MEDIA_ID_2")

        self.assertEqual(split_presentation_sentences(""), [])
        self.assertEqual(estimate_tts_duration_seconds("", language_code="zh"), 1.0)
        self.assertIsNone(duration_from_edge_boundaries([{"offset": "bad"}]))
        self.assertEqual(duration_from_edge_boundaries([{"Offset": 10_000_000, "Duration": 5_000_000}]), 1.5)
        timestamps = build_sentence_timestamps("第一句。第二句。", duration_seconds=4.0)
        self.assertEqual(timestamps[-1]["end"], 4.0)
        self.assertEqual(optimize_tts_text("short english", language_code="en"), "short english")
        long_text = "这是一句很长的话。" * 80
        self.assertLessEqual(len(optimize_tts_text(long_text, language_code="zh")), 421)
        no_sentence = "长" * 500
        self.assertTrue(optimize_tts_text(no_sentence, language_code="zh").endswith("。"))

    def test_prompt_and_json_helpers_cover_english_and_fence_paths(self):
        section_prompt = build_walkthrough_section_user_prompt(section={"section_path": "S", "briefing_summary": "", "content": "Body"}, notes=[], language="en")
        self.assertIn("sentences in English", section_prompt)
        self.assertIn("Source:", section_prompt)
        self.assertIn("Body", section_prompt)
        self.assertNotIn("Presenter notes:", section_prompt)
        self.assertNotIn("Presenter summary:", section_prompt)
        deduped_notes_prompt = build_walkthrough_section_user_prompt(
            section={"section_path": "S", "briefing_summary": "", "content": "Body"},
            notes=["", "Validate maker approval.", "validate maker approval.", None],
            language="zh",
        )
        self.assertEqual(deduped_notes_prompt.count("Validate maker approval."), 1)
        block_prompt = build_walkthrough_block_user_prompt(block={"title": "B", "briefing_goal": "", "merged_summary": "", "developer_focus": []}, source_lines=[], language="en")
        self.assertIn("natural spoken script", block_prompt)
        self.assertNotIn("Developer focus:", block_prompt)
        self.assertNotIn("Related PRD source sections:", block_prompt)
        self.assertIn("English briefing outline", build_presentation_system_prompt("en"))
        self.assertIn("English presentation chunks", build_presentation_user_prompt(source_text="Source", image_urls=[], language="en"))
        self.assertEqual(briefing_service.strip_code_fences("```json\n{\"a\":1}\n```"), '{"a":1}')
        parsed = parse_session_overview(
            """```json
            {"overview":"O","scope":["S"],"impacted_modules":["M"],"developer_focus":["D"],"frontend_focus":["F"],"backend_focus":["B"],"risks":["R"],"unclear_rules":["U"],"missing_edge_cases":["E"],"unclear_ownership":["O"],"open_questions":["Q"]}
            ```"""
        )
        self.assertEqual(parsed["scope"], ["S"])
        self.assertEqual(parse_developer_overview_payload("not json")["overview"], "not json")
        self.assertEqual(parse_developer_overview_payload('{"background_goal":"BG","implementation_overview":"IO"}')["overview"], "BG IO")

    def test_overview_low_signal_and_fallback_detail_branches(self):
        self.assertTrue(overview_is_low_signal({"overview": "", "scope": [], "developer_focus": []}))
        self.assertTrue(overview_is_low_signal({"overview": "ok", "scope": [], "developer_focus": ["Date: 24 Nov 2025"]}))
        self.assertTrue(overview_is_low_signal({"overview": "ok", "scope": ["Version: v1"], "developer_focus": ["field display rule"]}))
        self.assertTrue(
            overview_is_low_signal(
                {
                    "overview": "ok",
                    "scope": ["核心业务流程"],
                    "developer_focus": ["页面动作", "状态变化", "核心业务流程", "规则约束展开"],
                    "frontend_focus": [],
                    "backend_focus": [],
                    "risks": [],
                    "unclear_rules": [],
                    "missing_edge_cases": [],
                    "unclear_ownership": [],
                    "open_questions": [],
                }
            )
        )
        self.assertTrue(
            overview_is_low_signal(
                {
                    "overview": "ok",
                    "scope": [],
                    "developer_focus": [],
                    "impacted_modules": ["Version"],
                    "frontend_focus": ["Date"],
                    "backend_focus": ["PIC"],
                    "risks": ["Owner"],
                }
            )
        )
        self.assertEqual(normalize_overview_list(["", "Version: v1", "Field rule"]), ["Field rule"])
        self.assertTrue(briefing_service.looks_like_metadata_noise("hello@example.com"))
        self.assertTrue(briefing_service.looks_like_metadata_noise("release v1.2"))
        self.assertTrue(briefing_service.looks_like_metadata_noise("24 November 2025"))

        overview = briefing_service.build_chinese_fallback_overview(
            "Plain",
            [{"section_path": "Plain", "content": "No special product keywords here."}],
        )
        self.assertIn("主要围绕页面操作流程", overview)
        detailed = briefing_service.build_chinese_fallback_overview(
            "Detail",
            [
                {
                    "section_path": "Detail",
                    "content": "assessment id view the assessment detail overview tab withdraw comment review comment submit comment details tab closed verified draft pending review auto populate default required readonly sg regional",
                }
            ],
        )
        self.assertIn("评估详情", detailed)
        self.assertIn("comment 区域", detailed)
        self.assertIn("状态流转", detailed)
        two_part = briefing_service.build_two_part_fallback_overview(
            "Report",
            [{"section_path": "Report", "content": "report download history register closed verified default show display"}],
        )
        self.assertIn("报表", two_part["background_goal"])
        self.assertIn("按钮", two_part["implementation_overview"])

        candidates = briefing_service.extract_candidate_sentences(
            [{"section_path": "3. Field", "content": "Short. This sentence is long enough for extraction."}],
            limit=3,
        )
        self.assertEqual(candidates, ["This sentence is long enough for extraction."])
        self.assertEqual(
            briefing_service.extract_candidate_sentences(
                [{"section_path": "3. Field", "content": "这是一个中文短句字段. Another sentence is long enough for extraction."}],
                limit=3,
            ),
            ["Another sentence is long enough for extraction."],
        )


if __name__ == "__main__":
    unittest.main()
