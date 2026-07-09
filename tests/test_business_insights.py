import io
import json
import os
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest.mock import patch
from zoneinfo import ZoneInfo

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from bpmis_jira_tool.business_insights import (
    AF_FEATURE_CONFIG_TABLE,
    AF_ACTION_LOG_TABLE,
    AF_DETECTION_EFFECTIVENESS_REPORT_ID,
    AF_CARD_3DS_REPORT_ID,
    AF_DEVICE_RISK_REPORT_ID,
    AF_BLACK_WHITE_LIST_TABLE,
    AF_GREY_LIST_TABLE,
    AF_LIST_USAGE_REPORT_ID,
    AF_FACIAL_VERIFICATION_REPORT_ID,
    AF_FACIAL_VERIFICATION_TABLE,
    AF_FRAUD_LOSS_REPORT_ID,
    AF_IDENTIFY_REJECT_TABLE,
    AF_RULE_CHANGE_LOG_REPORT_ID,
    AF_PUNISH_LIST_TABLE,
    AF_REQUEST_STATISTIC_TABLE,
    AF_REVIEW_CASE_TABLE,
    AF_REVIEW_RECORD_TABLE,
    AF_RULE_CONFIG_TABLE,
    AF_RULE_HIT_LOG_TABLE,
    AF_RULE_EFFECTIVENESS_REPORT_ID,
    AF_RULES_FEATURES_REPORT_ID,
    AF_SCENARIOS_ACTIONS_REPORT_ID,
    AF_SCENARIO_GROUP_RELATION_TABLE,
    AF_SCENARIO_GROUP_TABLE,
    AF_TWO_WAY_TEMPLATE_CONFIG_TABLE,
    AF_TWO_WAY_RELATION_CONFIG_TABLE,
    AF_TWO_WAY_COMMUNICATION_TABLE,
    AF_SCENE_TABLE,
    APPLICATION_DISBURSEMENT_FUNNEL_REPORT_ID,
    BusinessInsightsStore,
    LIMIT_UTILIZATION_REPORT_ID,
    PORTFOLIO_REPAYMENT_REPORT_ID,
    UNDERWRITING_FUNNEL_REPORT_ID,
    UNDERWRITING_FUNNEL_TABLE,
    build_af_detection_effectiveness_sql,
    build_af_card_3ds_sql,
    build_af_device_risk_sql,
    build_af_list_usage_sql,
    build_af_facial_verification_sql,
    build_af_fraud_loss_sql,
    build_af_rule_change_log_sql,
    build_af_rule_effectiveness_sql,
    build_af_rules_features_sql,
    build_af_scenarios_actions_sql,
    build_application_disbursement_funnel_sql,
    build_limit_utilization_sql,
    build_portfolio_repayment_sql,
    build_underwriting_funnel_mis_sql,
    build_underwriting_funnel_sql,
    build_underwriting_funnel_workbook,
    product_label,
)
from bpmis_jira_tool.errors import ToolError
from bpmis_jira_tool.web import create_app
from scripts.generate_business_insights_live_reports import (
    REPORT_BUILDERS,
    extract_sql_sections,
    write_visualization,
)


FIXED_NOW = datetime(2026, 5, 26, 10, 30, tzinfo=ZoneInfo("Asia/Singapore"))


def _synthetic_underwriting_rows():
    return [
        {
            "underwriting_id": "UW-1",
            "product_code": "CASH_LOAN",
            "sub_product_code": "NEW",
            "application_submission_time": "2026-04-05 09:00:00",
            "borrower_id": "B1",
            "apply_loan_amount": "10000",
            "apply_loan_tenor": "6",
            "credit_score_partner": "720",
            "underwriting_status": "APPROVED",
            "current_stage": "DONE",
            "step": "FINAL",
            "reject_reason": "",
            "create_date": "2026-04-05 09:00:00",
            "modify_date": "2026-04-05 10:00:00",
        },
        {
            "underwriting_id": "UW-2",
            "product_code": "CASH_LOAN",
            "sub_product_code": "REPEAT",
            "application_submission_time": "2026-05-10 11:00:00",
            "borrower_id": "B2",
            "apply_loan_amount": "5000",
            "apply_loan_tenor": "3",
            "credit_score_partner": "510",
            "underwriting_status": "REJECTED",
            "current_stage": "POLICY",
            "step": "RULE",
            "reject_reason": "LOW_SCORE",
            "create_date": "2026-05-10 11:00:00",
            "modify_date": "2026-05-10 11:30:00",
        },
        {
            "underwriting_id": "UW-3",
            "product_code": "PAY_LATER",
            "sub_product_code": "",
            "application_submission_time": "2026-05-12 13:00:00",
            "borrower_id": "B3",
            "apply_loan_amount": "1500",
            "apply_loan_tenor": "1",
            "credit_score_partner": "650",
            "underwriting_status": "PENDING",
            "current_stage": "MANUAL_REVIEW",
            "step": "REVIEW",
            "reject_reason": "",
            "create_date": "2026-05-12 13:00:00",
            "modify_date": "2026-05-12 13:05:00",
        },
    ]


def _csv_export_bytes(rows=None):
    rows = rows or _synthetic_underwriting_rows()
    headers = list(rows[0].keys())
    lines = [",".join(headers)]
    for row in rows:
        lines.append(",".join(str(row.get(header, "")) for header in headers))
    return ("\n".join(lines) + "\n").encode("utf-8")


class BusinessInsightsTests(unittest.TestCase):
    def test_underwriting_sql_uses_previous_month_and_current_mtd(self):
        sql = build_underwriting_funnel_sql(FIXED_NOW)

        self.assertIn("Credit Risk PH - Underwriting Funnel", sql)
        self.assertIn("Duration: Apr 2026 + May 2026 MTD", sql)
        self.assertIn(UNDERWRITING_FUNNEL_TABLE, sql)
        self.assertIn("application_submission_time >= 1774972800000", sql)
        self.assertIn("application_submission_time < 1779811200000", sql)
        self.assertIn("product_code", sql)
        self.assertIn("sub_product_code", sql)

    def test_underwriting_mis_sql_uses_snapshot_and_aggregation_queries(self):
        sql = build_underwriting_funnel_mis_sql(snapshot_pt_date="2026-05-25", now=FIXED_NOW)

        self.assertIn("Credit Risk PH - Underwriting Funnel MIS", sql)
        self.assertIn("Snapshot: 2026-05-25", sql)
        self.assertIn("pt_date = '2026-05-25'", sql)
        self.assertIn("Summary by Product", sql)
        self.assertIn("Product Funnel", sql)
        self.assertIn("Product Reject Reasons", sql)
        self.assertIn("Product Stage Backlog", sql)
        self.assertIn("Sub-product Funnel", sql)

    def test_new_credit_risk_report_sql_builders_use_accessible_tables(self):
        portfolio_sql = build_portfolio_repayment_sql(snapshot_pt_date="2026-05-25", now=FIXED_NOW)
        limit_sql = build_limit_utilization_sql(snapshot_pt_date="2026-05-25", now=FIXED_NOW)
        funnel_sql = build_application_disbursement_funnel_sql(snapshot_pt_date="2026-05-25", now=FIXED_NOW)

        self.assertIn("Credit Risk PH - Portfolio Repayment", portfolio_sql)
        self.assertIn("cbs_ph_bke_loan_core_db_repay_plan_tab_ss", portfolio_sql)
        self.assertIn("Repay Flow Status", portfolio_sql)
        self.assertIn("Credit Risk PH - Limit Utilization", limit_sql)
        self.assertIn("cbs_ph_bke_loan_core_db_credit_limit_tab_ss_d", limit_sql)
        self.assertIn("Utilization Buckets", limit_sql)
        self.assertIn("Credit Risk PH - Application to Disbursement Funnel", funnel_sql)
        self.assertIn("cbs_ph_bke_loan_txn_db_loan_application_tab_ss", funnel_sql)
        self.assertIn("cbs_ph_bke_loan_core_db_disburse_flow_tab_ss", funnel_sql)

    def test_underwriting_workbook_contains_expected_sheets_and_summary_values(self):
        workbook_bytes = build_underwriting_funnel_workbook(_synthetic_underwriting_rows(), now=FIXED_NOW)
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)

        self.assertEqual(
            workbook.sheetnames,
            [
                "Summary by Product",
                "Product Funnel",
                "Product Reject Reasons",
                "Product Stage Backlog",
                "Sub-product Funnel",
                "Raw Export",
            ],
        )
        summary_rows = list(workbook["Summary by Product"].iter_rows(values_only=True))
        self.assertIn(("Apr 2026", "CASH_LOAN", 1, 1, 0, 0, 1, 10000), summary_rows)
        self.assertIn(("May 2026 MTD", "CASH_LOAN", 1, 0, 1, 0, 0, 5000), summary_rows)
        self.assertIn(("May 2026 MTD", "PAY_LATER", 1, 0, 0, 1, 0, 1500), summary_rows)

    def test_product_codes_are_displayed_as_apollo_product_names_when_known(self):
        rows = [
            {
                **_synthetic_underwriting_rows()[0],
                "underwriting_id": "UW-SPL",
                "product_code": "801",
                "sub_product_code": "101",
            }
        ]
        workbook_bytes = build_underwriting_funnel_workbook(rows, now=FIXED_NOW)
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)

        summary_rows = list(workbook["Summary by Product"].iter_rows(values_only=True))
        subproduct_rows = list(workbook["Sub-product Funnel"].iter_rows(values_only=True))
        raw_rows = list(workbook["Raw Export"].iter_rows(values_only=True))
        self.assertEqual(product_label("812"), "Credit Card")
        self.assertEqual(product_label("812F"), "Credit Card")
        self.assertEqual(product_label("807"), "Employee Loan")
        self.assertEqual(product_label("108"), "Employee Loan")
        self.assertIn(("Apr 2026", "801", 1, 1, 0, 0, 1, 10000), summary_rows)
        self.assertIn(("Apr 2026", "801", "101", "APPROVED", 1, 1), subproduct_rows)
        self.assertIn("801", raw_rows[1])

    def test_visualization_includes_product_filter_for_product_level_data(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "visualization.html"
            write_visualization(
                output_path,
                report_title="Credit Risk PH - Application to Disbursement Funnel",
                snapshot_pt_date="2026-05-25",
                sheets=[
                    (
                        "Funnel Summary by Product",
                        ["period", "product", "applications", "disbursed_loans", "disbursed_principal", "application_to_disbursement_rate"],
                        [
                            ["Apr 2026", "807", 10, 7, 7000, 0.7],
                            ["Apr 2026", "812F", 20, 10, 10000, 0.5],
                        ],
                    )
                ],
            )
            html = output_path.read_text(encoding="utf-8")

        self.assertIn("data-product-filter", html)
        self.assertIn("807 - Employee Loan", html)
        self.assertIn("812F - Credit Card", html)
        self.assertIn('data-product="807 - Employee Loan"', html)
        self.assertIn("Filters product-level charts and tables", html)
        self.assertIn('data-product-visual="1"', html)
        self.assertIn('document.querySelectorAll("[data-global-visual]")', html)

    def test_visualization_tables_render_all_rows_with_pagination(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "visualization.html"
            rows = [["Apr 2026", "807", index, index - 1, index * 100, 0.5] for index in range(1, 56)]
            write_visualization(
                output_path,
                report_title="Credit Risk PH - Application to Disbursement Funnel",
                snapshot_pt_date="2026-05-25",
                sheets=[
                    (
                        "Funnel Summary by Product",
                        ["period", "product", "applications", "disbursed_loans", "disbursed_principal", "application_to_disbursement_rate"],
                        rows,
                    )
                ],
            )
            html = output_path.read_text(encoding="utf-8")

        self.assertIn("data-table-pagination", html)
        self.assertIn('data-page-size="50"', html)
        self.assertIn("<td class=\"num\">55</td>", html)
        self.assertNotIn("Full data is in Excel", html)
        self.assertNotIn("Showing top", html)

    def test_store_persists_metadata_handles_corrupt_metadata_and_missing_artifact(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = BusinessInsightsStore(Path(temp_dir))
            artifact = store.save_underwriting_export(content=_csv_export_bytes(), filename="export.csv", now=FIXED_NOW)
            reloaded = BusinessInsightsStore(Path(temp_dir))
            report = reloaded.report(UNDERWRITING_FUNNEL_REPORT_ID)

            self.assertEqual(report["artifact"]["id"], artifact["id"])
            metadata, artifact_path = reloaded.artifact_path(artifact["id"])
            self.assertEqual(metadata["row_count"], 3)
            self.assertTrue(artifact_path.exists())
            visualization_path = artifact_path.with_suffix(".html")
            visualization_path.write_text("<html><body>Visualization</body></html>", encoding="utf-8")
            payload = json.loads((Path(temp_dir) / "reports.json").read_text(encoding="utf-8"))
            payload["artifacts"][UNDERWRITING_FUNNEL_REPORT_ID]["visualization_filename"] = visualization_path.name
            (Path(temp_dir) / "reports.json").write_text(json.dumps(payload), encoding="utf-8")
            visualization_metadata, resolved_visualization_path = reloaded.visualization_path(artifact["id"])
            self.assertEqual(visualization_metadata["id"], artifact["id"])
            self.assertEqual(resolved_visualization_path, visualization_path.resolve())

            artifact_path.unlink()
            with self.assertRaises(ToolError):
                reloaded.artifact_path(artifact["id"])

            (Path(temp_dir) / "reports.json").write_text("{broken", encoding="utf-8")
            fallback_report = reloaded.report(UNDERWRITING_FUNNEL_REPORT_ID)
            self.assertIsNone(fallback_report["artifact"])

    def test_business_insights_page_nav_tabs_and_headers_render_for_admin(self):
        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(
            os.environ,
            {
                "ENV_FILE": os.devnull,
                "FLASK_SECRET_KEY": "test-secret",
                "TEAM_PORTAL_DATA_DIR": temp_dir,
                "TEAM_PORTAL_BASE_URL": "",
                "TEAM_ALLOWED_EMAILS": "",
                "TEAM_ALLOWED_EMAIL_DOMAINS": "",
                "TEAM_PORTAL_CONFIG_ENCRYPTION_KEY": "",
            },
            clear=True,
        ):
            app = create_app()
            app.testing = True
            with app.test_client() as client:
                with client.session_transaction() as session:
                    session["google_profile"] = {"email": "xiaodong.zheng@npt.sg", "name": "Admin"}
                    session["google_credentials"] = {"token": "x"}
                response = client.get("/business-insights")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        labels = [node.get_text(strip=True) for node in soup.select(".site-switcher-tab")]
        self.assertLess(labels.index("Projects"), labels.index("Business Insights"))
        self.assertLess(labels.index("Business Insights"), labels.index("Others"))
        subtabs = [node.get_text(strip=True) for node in soup.select(".business-insights-tabs .workspace-tab")]
        self.assertEqual(subtabs, ["Anti-fraud", "Credit Risk", "Ops Risk"])
        for table in soup.select(".business-insights-table"):
            headers = [node.get_text(strip=True) for node in table.select("thead th")]
            self.assertEqual(headers, ["Report Name", "Link"])
        self.assertIn("Credit Risk PH - Underwriting Funnel", response.get_data(as_text=True))
        self.assertIn("Credit Risk PH - Portfolio Repayment", response.get_data(as_text=True))
        self.assertIn("Credit Risk PH - Limit Utilization", response.get_data(as_text=True))
        self.assertIn("Credit Risk PH - Application to Disbursement Funnel", response.get_data(as_text=True))
        self.assertIn("Anti-fraud PH - L1+L2 Scenarios, Actions &amp; Auth Steps", response.get_data(as_text=True))
        self.assertIn("Anti-fraud PH - Rules &amp; Features", response.get_data(as_text=True))
        self.assertIn("Anti-fraud PH - Rule Effectiveness / Hit-Rate", response.get_data(as_text=True))
        self.assertIn("Anti-fraud PH - Fraud Loss &amp; Case Outcomes", response.get_data(as_text=True))
        self.assertIn("Anti-fraud PH - Facial Verification / Liveness &amp; Deepfake", response.get_data(as_text=True))
        self.assertIn("Anti-fraud PH - Device &amp; Identity Risk", response.get_data(as_text=True))
        self.assertIn("Anti-fraud PH - Card Fraud &amp; 3DS Authentication", response.get_data(as_text=True))
        self.assertIn("Anti-fraud PH - Blacklist, Whitelist &amp; Greylist", response.get_data(as_text=True))
        # Detection-effectiveness and rule-change-log are folded into other reports, not standalone.
        self.assertNotIn("Anti-fraud PH - Detection Effectiveness &amp; Loss Prevented", response.get_data(as_text=True))
        self.assertNotIn("Anti-fraud PH - Rule Change Log &amp; Governance", response.get_data(as_text=True))
        self.assertEqual(response.get_data(as_text=True).count("Download SQL"), 12)
        self.assertNotIn("Upload Export", response.get_data(as_text=True))
        self.assertIsNone(soup.select_one("[data-business-insights-upload]"))

    def test_business_insights_access_route_sql_ingest_and_download(self):
        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(
            os.environ,
            {
                "ENV_FILE": os.devnull,
                "FLASK_SECRET_KEY": "test-secret",
                "TEAM_PORTAL_DATA_DIR": temp_dir,
                "TEAM_PORTAL_BASE_URL": "https://app.bankpmtool.uk",
                "TEAM_ALLOWED_EMAILS": "",
                "TEAM_ALLOWED_EMAIL_DOMAINS": "npt.sg",
                "TEAM_PORTAL_CONFIG_ENCRYPTION_KEY": "",
            },
            clear=True,
        ):
            app = create_app()
            app.testing = True
            with app.test_client() as client:
                with client.session_transaction() as session:
                    session["google_profile"] = {"email": "teammate@npt.sg", "name": "Teammate"}
                    session["google_credentials"] = {"token": "x"}
                non_admin_page = client.get("/business-insights", follow_redirects=False)
                non_admin_reports = client.get("/api/business-insights/reports?domain=credit-risk")

                with client.session_transaction() as session:
                    session["google_profile"] = {"email": "xiaodong.zheng@npt.sg", "name": "Admin"}
                    session["google_credentials"] = {"token": "x"}
                reports_response = client.get("/api/business-insights/reports?domain=credit-risk")
                sql_response = client.get(f"/api/business-insights/reports/{UNDERWRITING_FUNNEL_REPORT_ID}/sql")
                sql_download_response = client.get(f"/api/business-insights/reports/{UNDERWRITING_FUNNEL_REPORT_ID}/sql?format=raw&download=1")
                portfolio_sql_response = client.get(f"/api/business-insights/reports/{PORTFOLIO_REPAYMENT_REPORT_ID}/sql?format=raw&download=1")
                limit_sql_response = client.get(f"/api/business-insights/reports/{LIMIT_UTILIZATION_REPORT_ID}/sql?format=raw&download=1")
                funnel_sql_response = client.get(f"/api/business-insights/reports/{APPLICATION_DISBURSEMENT_FUNNEL_REPORT_ID}/sql?format=raw&download=1")
                ingest_response = client.post(
                    f"/api/business-insights/reports/{UNDERWRITING_FUNNEL_REPORT_ID}/ingest",
                    data={"file": (io.BytesIO(_csv_export_bytes()), "underwriting_export.csv")},
                    content_type="multipart/form-data",
                )
                artifact = ingest_response.get_json()["artifact"]
                artifact_url = artifact["url"]
                visualization_filename = artifact["filename"].replace(".xlsx", ".html")
                (Path(temp_dir) / "business_insights" / "artifacts" / visualization_filename).write_text(
                    "<html><body>Credit Risk Visualization</body></html>",
                    encoding="utf-8",
                )
                metadata_path = Path(temp_dir) / "business_insights" / "reports.json"
                payload = json.loads(metadata_path.read_text(encoding="utf-8"))
                payload["artifacts"][UNDERWRITING_FUNNEL_REPORT_ID]["visualization_filename"] = visualization_filename
                metadata_path.write_text(json.dumps(payload), encoding="utf-8")
                page_with_visualization = client.get("/business-insights")
                visualized_reports_response = client.get("/api/business-insights/reports?domain=credit-risk")
                visualization_url = visualized_reports_response.get_json()["reports"][0]["artifact"]["visualization_url"]
                # Downloads/visualizations no longer require a password; login is the gate.
                visualization_response = client.get(visualization_url)
                visualization_body = visualization_response.get_data(as_text=True)
                visualization_response.close()
                download_response = client.get(artifact_url)
                download_status = download_response.status_code
                download_mimetype = download_response.mimetype
                download_response.get_data()
                download_response.close()

        self.assertEqual(non_admin_page.status_code, 200)
        self.assertIn("Business Insights", non_admin_page.get_data(as_text=True))
        self.assertEqual(non_admin_reports.status_code, 200)
        # Non-admins only see the public Anti-fraud domain; credit-risk is empty.
        self.assertEqual(non_admin_reports.get_json()["reports"], [])
        self.assertEqual(reports_response.status_code, 200)
        self.assertEqual(len(reports_response.get_json()["reports"]), 4)
        self.assertEqual(sql_response.status_code, 200)
        self.assertIn(UNDERWRITING_FUNNEL_TABLE, sql_response.get_json()["sql"])
        self.assertEqual(sql_download_response.status_code, 200)
        self.assertEqual(sql_download_response.mimetype, "text/plain")
        self.assertIn("attachment; filename=credit-risk-ph-underwriting-funnel.sql", sql_download_response.headers["Content-Disposition"])
        self.assertIn("Summary by Product", sql_download_response.get_data(as_text=True))
        self.assertEqual(portfolio_sql_response.status_code, 200)
        self.assertIn("Portfolio Repayment", portfolio_sql_response.get_data(as_text=True))
        self.assertEqual(limit_sql_response.status_code, 200)
        self.assertIn("Limit Utilization", limit_sql_response.get_data(as_text=True))
        self.assertEqual(funnel_sql_response.status_code, 200)
        self.assertIn("Application to Disbursement Funnel", funnel_sql_response.get_data(as_text=True))
        self.assertEqual(ingest_response.status_code, 200)
        self.assertEqual(ingest_response.get_json()["artifact"]["row_count"], 3)
        self.assertEqual(page_with_visualization.status_code, 200)
        self.assertIn("Open Visualization", page_with_visualization.get_data(as_text=True))
        # No password gate: downloads work directly for logged-in users.
        self.assertEqual(download_status, 200)
        self.assertEqual(visualization_response.status_code, 200)
        self.assertEqual(visualization_response.mimetype, "text/html")
        self.assertIn("Credit Risk Visualization", visualization_body)
        self.assertEqual(download_status, 200)
        self.assertEqual(download_mimetype, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


class BusinessInsightsLoginRequiredTests(unittest.TestCase):
    """Business Insights, Version Plan, and Source Code QA all require login.
    Only admin, allowlisted emails, and allowlisted domains (@npt.sg, @monee.com,
    @seamoney.com) may access.  Non-admin allowed users see Anti-fraud only.
    Anonymous visitors are redirected to the login page."""

    def _probe(self, email=None):
        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(
            os.environ,
            {
                "ENV_FILE": os.devnull,
                "FLASK_SECRET_KEY": "test-secret",
                "TEAM_PORTAL_DATA_DIR": temp_dir,
                "TEAM_PORTAL_BASE_URL": "https://app.bankpmtool.uk",
                "TEAM_ALLOWED_EMAILS": "xiaodong.zheng1991@gmail.com",
                "TEAM_ALLOWED_EMAIL_DOMAINS": "npt.sg,monee.com,seamoney.com",
                "TEAM_PORTAL_CONFIG_ENCRYPTION_KEY": "",
            },
            clear=True,
        ):
            app = create_app()
            app.testing = True
            with app.test_client() as client:
                if email is not None:
                    with client.session_transaction() as session:
                        session["google_profile"] = {"email": email, "name": "Guest"}
                        session["google_credentials"] = {"token": "x"}
                return {
                    "bi_page": client.get("/business-insights?domain=anti-fraud", follow_redirects=False),
                    "af_reports": client.get("/api/business-insights/reports?domain=anti-fraud"),
                    "cr_reports": client.get("/api/business-insights/reports?domain=credit-risk"),
                    "cr_sql": client.get(f"/api/business-insights/reports/{UNDERWRITING_FUNNEL_REPORT_ID}/sql"),
                    "source_code": client.get("/source-code-qa", follow_redirects=False),
                    "version_plan": client.get("/version-plan", follow_redirects=False),
                    "landing": client.get("/portal-home", follow_redirects=False),
                }

    def test_anonymous_visitor_is_redirected_to_login(self):
        r = self._probe(email=None)  # not signed in
        # All three pages require login.
        self.assertEqual(r["bi_page"].status_code, 302)
        self.assertEqual(r["source_code"].status_code, 302)
        self.assertEqual(r["version_plan"].status_code, 302)
        self.assertEqual(r["landing"].status_code, 302)
        # API calls return 401.
        self.assertEqual(r["af_reports"].status_code, 401)

    def test_gmail_test_user_sees_anti_fraud_only(self):
        r = self._probe("xiaodong.zheng1991@gmail.com")
        self.assertEqual(r["bi_page"].status_code, 200)
        html = r["bi_page"].get_data(as_text=True)
        self.assertIn('data-business-insights-tab="anti-fraud"', html)
        self.assertNotIn('data-business-insights-tab="credit-risk"', html)
        self.assertNotIn("data-download-password-form", html)
        self.assertNotIn("Refresh data", html)
        self.assertEqual(r["cr_reports"].status_code, 200)
        self.assertEqual(r["cr_reports"].get_json()["reports"], [])
        self.assertEqual(r["cr_sql"].status_code, 404)
        self.assertEqual(r["source_code"].status_code, 200)
        self.assertEqual(r["version_plan"].status_code, 200)

    def test_monee_and_seamoney_users_see_anti_fraud_only(self):
        for email in ("analyst@monee.com", "analyst@seamoney.com"):
            with self.subTest(email=email):
                r = self._probe(email)
                self.assertEqual(r["bi_page"].status_code, 200)
                html = r["bi_page"].get_data(as_text=True)
                self.assertIn('data-business-insights-tab="anti-fraud"', html)
                self.assertNotIn('data-business-insights-tab="credit-risk"', html)
                self.assertNotIn("data-download-password-form", html)
                self.assertEqual(r["cr_reports"].get_json()["reports"], [])
                self.assertEqual(r["cr_sql"].status_code, 404)
                self.assertEqual(r["source_code"].status_code, 200)
                self.assertEqual(r["version_plan"].status_code, 200)

    def test_external_domain_user_is_blocked(self):
        r = self._probe("stranger@external.com")
        # External domains are not allowlisted: blocked from all surfaces.
        self.assertEqual(r["bi_page"].status_code, 302)
        self.assertEqual(r["bi_page"].headers["Location"], "/access-denied")
        self.assertEqual(r["source_code"].status_code, 302)
        self.assertEqual(r["source_code"].headers["Location"], "/access-denied")
        self.assertEqual(r["version_plan"].status_code, 302)
        self.assertEqual(r["version_plan"].headers["Location"], "/access-denied")
        self.assertEqual(r["landing"].status_code, 302)
        self.assertEqual(r["landing"].headers["Location"], "/access-denied")


class AntiFraudBusinessInsightsTests(unittest.TestCase):
    def test_seeded_reports_include_scenarios_actions_report(self):
        with tempfile.TemporaryDirectory() as tmp:
            store = BusinessInsightsStore(Path(tmp))
            reports = store.reports("anti-fraud")
        ids = {report["id"] for report in reports}
        self.assertIn(AF_SCENARIOS_ACTIONS_REPORT_ID, ids)
        report = next(r for r in reports if r["id"] == AF_SCENARIOS_ACTIONS_REPORT_ID)
        self.assertEqual(report["domain"], "anti-fraud")
        self.assertEqual(report["status"], "generator_ready")
        self.assertIsNone(report["artifact"])

    def test_store_returns_sql_for_scenarios_actions_report(self):
        with tempfile.TemporaryDirectory() as tmp:
            store = BusinessInsightsStore(Path(tmp))
            sql = store.sql_for_report(AF_SCENARIOS_ACTIONS_REPORT_ID, now=FIXED_NOW)
            self.assertTrue(sql.strip())
            self.assertEqual(
                store.sql_filename_for_report(AF_SCENARIOS_ACTIONS_REPORT_ID),
                f"{AF_SCENARIOS_ACTIONS_REPORT_ID}.sql",
            )

    def test_scenarios_actions_sql_uses_granted_tables_and_snapshot(self):
        sql = build_af_scenarios_actions_sql(snapshot_pt_date="2026-05-25")
        self.assertIn(AF_SCENE_TABLE, sql)
        self.assertEqual(AF_SCENE_TABLE, "ods.mbs_ph_seabank_anti_fraud_db_scene_tab_ss")
        self.assertIn("pt_date = '2026-05-25'", sql)
        # Auth steps are surfaced via the action type label and the flow-config challenge steps.
        self.assertIn("Authentication (Auth Step)", sql)
        self.assertIn(AF_SCENARIO_GROUP_RELATION_TABLE, sql)
        self.assertIn(AF_SCENARIO_GROUP_TABLE, sql)
        self.assertIn("default_step", sql)
        self.assertIn("challenge1_step", sql)

    def test_scenarios_actions_sql_has_numbered_sections_for_generator(self):
        sql = build_af_scenarios_actions_sql(snapshot_pt_date="2026-05-25", now=FIXED_NOW)
        sections = extract_sql_sections(sql)
        self.assertTrue(sections, "scenarios/actions SQL produced no generator sections")
        for section in sections:
            self.assertTrue(section.query.strip())

    def test_scenario_flow_section_exposes_enum_name_columns(self):
        sql = build_af_scenarios_actions_sql(snapshot_pt_date="2026-05-25", now=FIXED_NOW)
        flow_section = next(
            section for section in extract_sql_sections(sql) if section.sheet_name == "Scenario Action Auth Flow"
        )
        for column in (
            "l1_scene_name",
            "l1_enum_name",
            "l2_sub_scene_name",
            "l2_enum_name",
            "action_name",
            "action_enum_name",
            "scenario_group_id",
            "scenario_group_name",
            "scenario_group_description",
            "default_step",
            "challenge5_step",
        ):
            self.assertIn(column, flow_section.query)
        # The flow-config table stores names (not codes), so dims join on name.
        self.assertIn("s.name = f.scene", flow_section.query)
        self.assertIn("ss.name = f.sub_scene", flow_section.query)
        self.assertIn("a.name = f.action", flow_section.query)
        self.assertIn("gr.biz_action = f.action", flow_section.query)
        self.assertIn("sg.group_id = gr.scenario_group_id", flow_section.query)
        self.assertNotIn("s.code = f.scene", flow_section.query)

    def test_scenarios_actions_sql_has_authentication_funnel_sections(self):
        sql = build_af_scenarios_actions_sql(snapshot_pt_date="2026-05-25", now=FIXED_NOW)
        names = [s.sheet_name for s in extract_sql_sections(sql)]
        for name in (
            "Authentication Outcome Summary",
            "Auth Outcome by Scene & Type",
            "Challenge Friction by Auth Type",
            "Auth Drop-off by Scene",
        ):
            self.assertIn(name, names)
        # Funnel reads the DWD action log, scoped by pt_date over the window; risk_result drives pass/reject.
        self.assertIn(AF_ACTION_LOG_TABLE, sql)
        self.assertIn("al.pt_date between '2026-03-01' and '2026-05-26'", sql)
        self.assertIn("al.risk_result = '1'", sql)
        self.assertIn("al.risk_result = '0'", sql)
        self.assertIn("is_final_action_in_flow_of_the_day", sql)
        self.assertIn("'May 2026 MTD'", sql)

    def test_scenarios_visualization_renders_auth_funnel_and_flow_table(self):
        flow_headers = [
            "l1_scene_name",
            "l1_enum_name",
            "l2_sub_scene_name",
            "l2_enum_name",
            "action_name",
            "action_enum_name",
            "scenario_group_id",
            "scenario_group_name",
            "scenario_group_description",
            "default_step",
            "challenge1_step",
            "challenge2_step",
            "challenge3_step",
            "challenge4_step",
            "challenge5_step",
        ]
        flow_rows = [
            ["Login", "LOGIN", "Password Login", "PWD_LOGIN", "OTP", "SEND_OTP", "A01", "Login/ Registration/ Onboarding", "All login flows", "step_a", "step_b", "", "", "", ""],
            ["Transfer", "TRANSFER", "P2P", "P2P_TRANSFER", "Face", "FACE_CHECK", "D01", "Transfers", "Money movement", "step_c", "", "", "", "", ""],
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "visualization.html"
            write_visualization(
                output_path,
                report_title="Anti-fraud PH - L1+L2 Scenarios, Actions & Auth Steps",
                snapshot_pt_date="2026-05-25",
                sheets=[
                    ("L1 Scenarios", ["l1_scene_code", "l1_scene_name", "mode", "source"], [["100", "Login", "2", "4"]]),
                    ("Scenario Action Auth Flow", flow_headers, flow_rows),
                    ("Authentication Outcome Summary",
                     ["period", "actions", "distinct_users", "flows", "pass_actions", "reject_actions", "not_evaluated_actions", "reject_rate_pct"],
                     [["Mar 2026", "200000000", "1", "1", "199000000", "400000", "600000", "0.2"],
                      ["Apr 2026", "210000000", "1", "1", "209000000", "420000", "580000", "0.2"],
                      ["May 2026 MTD", "60000000", "1", "1", "59800000", "110000", "90000", "0.18"]]),
                    ("Auth Outcome by Scene & Type",
                     ["scene_name", "authentication_type", "actions", "distinct_users", "pass_actions", "reject_actions", "reject_rate_pct"],
                     [["Transfer", "CHALLENGE_2", "5000000", "1", "4900000", "100000", "2.0"]]),
                    ("Challenge Friction by Auth Type",
                     ["authentication_type", "actions", "distinct_users", "flows", "action_share_pct", "pass_actions", "reject_actions", "reject_rate_pct"],
                     [["DEFAULT", "180000000", "1", "1", "90", "179000000", "1000000", "0.55"],
                      ["CHALLENGE_2", "15000000", "1", "1", "7.5", "14800000", "200000", "1.33"]]),
                    ("Auth Drop-off by Scene",
                     ["scene_name", "flows_started", "flows_completed", "drop_off_rate_pct"],
                     [["Transfer", "1000000", "850000", "15.0"]]),
                ],
                report_id=AF_SCENARIOS_ACTIONS_REPORT_ID,
            )
            html = output_path.read_text(encoding="utf-8")

        # Authentication-outcome KPIs are period-switchable (month selector) and coverage KPIs follow.
        self.assertIn("Authentication Outcomes", html)
        self.assertIn("data-period-kpi", html)
        self.assertIn('<script type="application/json" data-period-json>', html)
        self.assertIn("Scenario Coverage", html)
        self.assertIn("Flow mappings", html)
        # The new authentication funnel panels render.
        self.assertIn("<h2>Auth Outcome by Scene &amp; Type</h2>", html)
        self.assertIn("<h2>Challenge Friction by Auth Type</h2>", html)
        self.assertIn("<h2>Reject Rate by Auth Type (%)</h2>", html)  # bar chart
        self.assertIn("<h2>Auth Drop-off by Scene</h2>", html)
        self.assertIn("CHALLENGE_2", html)
        # The scenario flow table keeps the 50ch step-column rendering.
        self.assertIn('<span class="th-label">scenario_group_id</span>', html)
        self.assertIn('<span class="th-label">scenario_group_name</span>', html)
        self.assertIn('<span class="th-label">scenario_group_description</span>', html)
        self.assertIn("Login/ Registration/ Onboarding", html)
        self.assertIn('<th class="step"><span class="th-label">default_step</span>', html)
        self.assertIn("step_a", html)
        self.assertIn("th.step,td.step{width:50ch;min-width:50ch;max-width:50ch;white-space:normal;", html)
        self.assertIn("Password Login", html)
        # Interactive shell + CSV export are shared across every panel.
        self.assertIn("enhanceTable", html)
        self.assertIn('class="col-filter"', html)
        # No generic credit-risk dashboard chrome.
        self.assertNotIn("data-product-filter", html)
        self.assertNotIn("Data Quality Notes", html)
        # Click-to-view info notes on source / mode / drop-off, plus the auth-step glossary.
        self.assertIn('aria-label="About source"', html)
        self.assertIn('aria-label="About mode"', html)
        self.assertIn('aria-label="About drop_off_rate_pct"', html)
        self.assertIn("col-note-pop", html)
        self.assertIn("<h2>Auth Step Glossary</h2>", html)
        self.assertIn("ALC dynamic-light (Aurora) liveness", html)

    def test_latest_snapshot_resolves_to_anchor_table_max_pt_date(self):
        from scripts import generate_business_insights_live_reports as gen

        self.assertEqual(
            gen.REPORT_SNAPSHOT_ANCHOR_TABLE[AF_SCENARIOS_ACTIONS_REPORT_ID],
            "ods.mbs_ph_seabank_anti_fraud_db_biz_scenario_flow_config_tab_df",
        )
        captured = {}

        def fake_run(session, section, *, poll_seconds, max_polls):
            captured["query"] = section.query
            return ["pt_date"], [["2026-06-08"]], "exec-1"

        with patch.object(gen, "run_workbench_query", fake_run):
            resolved = gen.resolve_snapshot_pt_date(
                object(), AF_SCENARIOS_ACTIONS_REPORT_ID, poll_seconds=1, max_polls=1
            )
        self.assertEqual(resolved, "2026-06-08")
        self.assertIn("max(pt_date)", captured["query"])
        self.assertIn("biz_scenario_flow_config_tab_df", captured["query"])

    def test_unknown_report_has_no_snapshot_anchor(self):
        from scripts import generate_business_insights_live_reports as gen

        self.assertIsNone(
            gen.resolve_snapshot_pt_date(object(), "does-not-exist", poll_seconds=1, max_polls=1)
        )

    def test_refresh_visualizations_publishes_public_artifacts(self):
        from scripts import generate_business_insights_live_reports as gen

        with tempfile.TemporaryDirectory() as temp_dir, patch.object(
            gen, "refresh_existing_visualizations"
        ) as refresh, patch.object(gen, "_publish_to_public_gcs") as publish, patch(
            "sys.argv",
            [
                "generate_business_insights_live_reports.py",
                "--refresh-visualizations",
                "--report-id",
                AF_RULE_EFFECTIVENESS_REPORT_ID,
                "--portal-data-dir",
                temp_dir,
            ],
        ):
            status = gen.main()

        self.assertEqual(status, 0)
        refresh.assert_called_once_with(Path(temp_dir).resolve(), report_ids=[AF_RULE_EFFECTIVENESS_REPORT_ID])
        publish.assert_called_once_with(Path(temp_dir).resolve())

    def test_cloud_run_hydrates_existing_business_insights_artifact_with_ttl(self):
        from bpmis_jira_tool import public_artifacts_gcs as public_gcs

        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(
            os.environ,
            {"TEAM_PORTAL_PUBLIC_GCS_BUCKET": "public-bucket"},
            clear=False,
        ), patch.object(public_gcs, "gcs_fetch_to_file", return_value=True) as fetch:
            ok = public_gcs.hydrate_business_insights_artifact(Path(temp_dir), "viz.html")

        self.assertTrue(ok)
        fetch.assert_called_once_with(
            "public-bucket",
            "business_insights/artifacts/viz.html",
            Path(temp_dir) / "artifacts" / "viz.html",
            max_age_seconds=public_gcs.PUBLIC_GCS_METADATA_TTL_SECONDS,
        )

    def test_generator_report_builders_include_scenarios_actions(self):
        self.assertIn(AF_SCENARIOS_ACTIONS_REPORT_ID, REPORT_BUILDERS)
        title, builder = REPORT_BUILDERS[AF_SCENARIOS_ACTIONS_REPORT_ID]
        self.assertTrue(title.startswith("Anti-fraud PH -"))
        self.assertTrue(callable(builder))

    def test_anti_fraud_report_renders_on_page(self):
        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(
            os.environ,
            {
                "ENV_FILE": os.devnull,
                "FLASK_SECRET_KEY": "test-secret",
                "TEAM_PORTAL_DATA_DIR": temp_dir,
                "TEAM_PORTAL_BASE_URL": "",
                "TEAM_ALLOWED_EMAILS": "",
                "TEAM_ALLOWED_EMAIL_DOMAINS": "",
                "TEAM_PORTAL_CONFIG_ENCRYPTION_KEY": "",
            },
            clear=True,
        ):
            app = create_app()
            app.testing = True
            with app.test_client() as client:
                with client.session_transaction() as session:
                    session["google_profile"] = {"email": "xiaodong.zheng@npt.sg", "name": "Admin"}
                    session["google_credentials"] = {"token": "x"}
                reports_response = client.get("/api/business-insights/reports?domain=anti-fraud")
                page_response = client.get("/business-insights?domain=anti-fraud")

        self.assertEqual(reports_response.status_code, 200)
        returned_ids = {report["id"] for report in reports_response.get_json()["reports"]}
        self.assertEqual(
            returned_ids,
            {
                AF_SCENARIOS_ACTIONS_REPORT_ID,
                AF_RULES_FEATURES_REPORT_ID,
                AF_RULE_EFFECTIVENESS_REPORT_ID,
                AF_FRAUD_LOSS_REPORT_ID,
                AF_FACIAL_VERIFICATION_REPORT_ID,
                AF_DEVICE_RISK_REPORT_ID,
                AF_CARD_3DS_REPORT_ID,
                AF_LIST_USAGE_REPORT_ID,
            },
        )
        # Detection-effectiveness and rule-change-log are folded into the reports above, not standalone.
        self.assertNotIn(AF_DETECTION_EFFECTIVENESS_REPORT_ID, returned_ids)
        self.assertNotIn(AF_RULE_CHANGE_LOG_REPORT_ID, returned_ids)
        self.assertEqual(page_response.status_code, 200)
        self.assertIn(
            "Anti-fraud PH - L1+L2 Scenarios, Actions &amp; Auth Steps",
            page_response.get_data(as_text=True),
        )


class RulesFeaturesBusinessInsightsTests(unittest.TestCase):
    def test_seeded_reports_include_rules_features(self):
        with tempfile.TemporaryDirectory() as tmp:
            store = BusinessInsightsStore(Path(tmp))
            ids = {report["id"] for report in store.reports("anti-fraud")}
        self.assertIn(AF_RULES_FEATURES_REPORT_ID, ids)

    def test_store_returns_sql_for_rules_features(self):
        with tempfile.TemporaryDirectory() as tmp:
            store = BusinessInsightsStore(Path(tmp))
            sql = store.sql_for_report(AF_RULES_FEATURES_REPORT_ID, now=FIXED_NOW)
        self.assertTrue(sql.strip())

    def test_rules_features_sql_has_catalog_and_governance_sections(self):
        sql = build_af_rules_features_sql(snapshot_pt_date="2026-06-08", now=FIXED_NOW)
        sections = extract_sql_sections(sql)
        # Catalogs (1-2), Function-usage (3), Two-Way config (4) + activity (5), then the folded-in rule
        # change-log governance sections (6-8).
        self.assertEqual(
            [s.sheet_name for s in sections],
            ["Rules", "Features", "Function Usage", "Two-Way Communication Config",
             "Two-Way Communication Activity",
             "Change Summary", "Rule Change Detail", "Current Rule Inventory"],
        )
        # Two-Way config pulls the template + relation config tables and exposes the treatment/whitelist knobs.
        self.assertIn(AF_TWO_WAY_TEMPLATE_CONFIG_TABLE, sql)
        self.assertIn(AF_TWO_WAY_RELATION_CONFIG_TABLE, sql)
        for column in ("treatment_tag", "confirmation_window_sec", "temp_whitelist_exempted_rules", "crc_alert"):
            self.assertIn(column, sql)
        # Two-Way activity counts triggers + outcomes from the communication table over the window.
        self.assertIn(AF_TWO_WAY_COMMUNICATION_TABLE, sql)
        for column in ("triggered", "approved", "rejected", "expired_no_response", "approval_rate_pct"):
            self.assertIn(column, sql)
        self.assertIn(AF_RULE_CONFIG_TABLE, sql)
        self.assertIn(AF_FEATURE_CONFIG_TABLE, sql)
        self.assertEqual(AF_RULE_CONFIG_TABLE, "ods.mbs_anti_fraud_rule_config_tab_ss")
        self.assertEqual(AF_FEATURE_CONFIG_TABLE, "ods.mbs_anti_fraud_feature_config_tab_ss")
        self.assertIn("pt_date = '2026-06-08'", sql)
        # Rules catalog surfaces the punish duration and the reject transify key.
        for column in ("punish_length_sec", "punish_duration", "transify_key"):
            self.assertIn(column, sql)
        for column in ("rule_id", "rule_name", "feature_id", "feature_name"):
            self.assertIn(column, sql)
        # Function-usage dimension: per function_id, the feature count + active + example features.
        func_section = next(s for s in sections if s.sheet_name == "Function Usage")
        self.assertIn("fc.function_id", func_section.query)
        self.assertIn("count(1) as features", func_section.query)
        self.assertIn("active_features", func_section.query)
        self.assertIn("collect_set(fc.feature_name)", func_section.query)
        # Governance diff present (added/deactivated/logic-changed classification).
        self.assertIn("full outer join", sql)
        self.assertIn("'Added'", sql)
        self.assertIn("Logic changed", sql)

    def test_generator_report_builders_include_rules_features(self):
        self.assertIn(AF_RULES_FEATURES_REPORT_ID, REPORT_BUILDERS)
        title, builder = REPORT_BUILDERS[AF_RULES_FEATURES_REPORT_ID]
        self.assertEqual(title, "Anti-fraud PH - Rules & Features")
        self.assertTrue(callable(builder))

    def test_visualization_renders_two_searchable_catalogs(self):
        rule_headers = ["rule_id", "rule_name", "outcome_type", "rule_status"]
        rule_rows = [["R1", "High Velocity Login", "Challenge", "Active"], ["R2", "Device Spoof", "Reject", "Active"]]
        feature_headers = ["feature_id", "feature_name", "threshold", "feature_status"]
        feature_rows = [["F1", "Login count 1h", "5", "Active"], ["F2", "Distinct device 24h", "3", "Inactive"]]
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "viz.html"
            write_visualization(
                output_path,
                report_title="Anti-fraud PH - Rules & Features",
                snapshot_pt_date="2026-06-08",
                sheets=[
                    ("Rules", rule_headers, rule_rows),
                    ("Features", feature_headers, feature_rows),
                ],
                report_id=AF_RULES_FEATURES_REPORT_ID,
            )
            html = output_path.read_text(encoding="utf-8")

        # Two searchable catalog panels with distinct placeholders.
        self.assertEqual(html.count('class="search-table"'), 2)
        self.assertIn("<h2>Rules</h2>", html)
        self.assertIn("<h2>Features</h2>", html)
        self.assertIn("Search rule id or name", html)
        self.assertIn("Search feature id or name", html)
        self.assertIn("High Velocity Login", html)
        self.assertIn("Distinct device 24h", html)
        # Generic dashboard chrome is not used.
        self.assertNotIn("data-product-filter", html)
        self.assertNotIn("Data Quality Notes", html)
        # Each column has its own filter input (4 rule cols + 4 feature cols).
        self.assertEqual(html.count('class="col-filter"'), 8)
        self.assertIn('data-col="0"', html)
        self.assertIn('data-col="3"', html)
        # Search and per-column filters are combined (AND) in the shared dashboard script.
        self.assertIn("colFilters", html)
        self.assertIn("._cells", html)
        # Tables are now sortable + value-formatted + threshold-highlighted.
        self.assertIn("enhanceTable", html)
        self.assertIn("flag-bad", html)
        self.assertIn("sortable", html)
        # Enriched with a KPI summary + outcome donut + status charts (still exactly two tables).
        self.assertIn("Catalog Summary", html)
        self.assertIn("Total rules", html)
        self.assertIn("Rules by Outcome Type", html)
        self.assertIn("Features by Status", html)

    def test_visualization_renders_function_usage_dimension(self):
        sheets = [
            ("Rules", ["rule_id", "rule_name", "outcome_type", "rule_status"],
             [["R1", "High Velocity Login", "Challenge", "Active"]]),
            ("Features", ["feature_id", "feature_name", "function_id", "feature_status"],
             [["F1c", "Login count 1h", "F1", "Active"]]),
            ("Function Usage",
             ["function_id", "features", "active_features", "distinct_scenes", "distinct_actions",
              "distinct_time_windows", "max_window_seconds", "example_features"],
             [["F1", "181", "90", "57", "65", "13", "2592000", "Login count 1h | BIN Attack | OTP retries"],
              ["F12", "187", "9", "6", "6", "1", "0", "PN/AR Validation | CRC flow | Soft Token"]]),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "viz.html"
            write_visualization(
                output_path,
                report_title="Anti-fraud PH - Rules & Features",
                snapshot_pt_date="2026-06-08",
                sheets=sheets,
                report_id=AF_RULES_FEATURES_REPORT_ID,
            )
            html = output_path.read_text(encoding="utf-8")
        # Function-usage KPI, ranking bar chart, and searchable table all render.
        self.assertIn("Functions in use", html)
        self.assertIn("<h2>Top Functions by Feature Count</h2>", html)
        self.assertIn("<h2>Function Usage</h2>", html)
        self.assertIn("Search function id", html)
        self.assertIn("example_features", html)
        self.assertIn("BIN Attack", html)


class RuleEffectivenessBusinessInsightsTests(unittest.TestCase):
    def test_seeded_and_generator_registration(self):
        with tempfile.TemporaryDirectory() as tmp:
            ids = {r["id"] for r in BusinessInsightsStore(Path(tmp)).reports("anti-fraud")}
        self.assertIn(AF_RULE_EFFECTIVENESS_REPORT_ID, ids)
        self.assertIn(AF_RULE_EFFECTIVENESS_REPORT_ID, REPORT_BUILDERS)
        self.assertEqual(REPORT_BUILDERS[AF_RULE_EFFECTIVENESS_REPORT_ID][0], "Anti-fraud PH - Rule Effectiveness / Hit-Rate")

    def test_sql_scopes_to_previous_full_month_and_fixes_action_rate(self):
        # FIXED_NOW is May 2026, so the previous full month is April 2026.
        sql = build_af_rule_effectiveness_sql(snapshot_pt_date="2026-06-08", now=FIXED_NOW)
        sections = extract_sql_sections(sql)
        self.assertEqual(
            [s.sheet_name for s in sections],
            [
                "Request Outcome Summary",
                "Reject Rule Hit Summary",
                "Reject Rule Scene Breakdown",
                "Punishment Rule Hit Summary",
                "Punishment Rule Scene Breakdown",
                "Challenge Rule Hit Summary",
                "Challenge Rule Scene Breakdown",
                "Daily Challenge/Reject/Punish",
                "Daily Rule Trigger Trend",
                "Scene/Sub-scene/Action Usage",
                "Rule Scorecard",
                # Folded-in detection effectiveness sections (12-16).
                "Detection Coverage Summary",
                "Detection by Fraud MO Type",
                "Top Detecting Rules",
                "Missed Fraud (Blind Spots)",
                "Daily Detection Trend",
            ],
        )
        self.assertIn(AF_REQUEST_STATISTIC_TABLE, sql)
        self.assertIn(AF_IDENTIFY_REJECT_TABLE, sql)
        self.assertIn(AF_PUNISH_LIST_TABLE, sql)
        self.assertIn("punish_rule_id", sql)
        # Rule precision joins review records to fraud cases (confirmed fraud = fraud_mo_type not Not Fraud).
        self.assertIn(AF_REVIEW_CASE_TABLE, sql)
        self.assertIn(AF_REVIEW_RECORD_TABLE, sql)
        self.assertIn("precision_pct", sql)
        self.assertIn("'not fraud'", sql)
        # Challenge hit-rate from the DWD hit log + action log, classified by outcome_type=2.
        self.assertIn(AF_RULE_HIT_LOG_TABLE, sql)
        self.assertIn(AF_ACTION_LOG_TABLE, sql)
        self.assertIn("outcome_type = 2", sql)
        self.assertIn("trigger_rate_pct", sql)
        self.assertIn("is_rule_triggered = 'Y'", sql)
        # identify_record is empty in ODS, so that table is not queried.
        self.assertNotIn("from ods.mbs_anti_fraud_identify_record", sql)
        self.assertNotIn("Identify Result by Scene", sql)
        # transaction amount is surfaced as PHP, and the breakdown resolves scene names.
        self.assertIn("rejected_amount_php", sql)
        self.assertIn("scene_name", sql)
        # Reject and Punish now carry the fmart benchmark + trigger-rate metric (summary and scene level).
        self.assertIn("benchmark_trxn", sql)
        self.assertIn("trigger_rate_pct", sql)
        self.assertIn("normalised_user_impact_pct", sql)
        # The benchmark denominator comes from the fmart action log scene traffic.
        self.assertIn("scene_traffic", sql)
        self.assertGreaterEqual(sql.count(AF_ACTION_LOG_TABLE), 3)  # reject, punish, challenge sections
        # Spans the last two full months + current MTD (Mar+Apr full, May MTD under the fixed clock).
        self.assertIn("Scope: Mar 2026 – May 2026 MTD", sql)
        self.assertIn("rq.date >= '20260301'", sql)
        self.assertIn("rq.date < '20260527'", sql)
        # Summary is per-period (Mar / Apr / May MTD).
        self.assertIn("'May 2026 MTD'", sql)
        self.assertIn("then 'Mar 2026'", sql)
        # Action rate derived from pass+challenge+reject, not the unpopulated total_req_num column.
        self.assertIn("total_outcomes", sql)
        self.assertNotIn("rq.total_req_num", sql)
        self.assertIn("action_rate_pct", sql)

    def test_visualization_renders_kpis_and_expandable_reject_scene_breakdown(self):
        summary_headers = ["period", "total_outcomes", "pass_num", "challenge_num", "reject_num", "action_rate_pct"]
        summary_rows = [["May 2026", "446218000", "427348342", "18070027", "849571", "4.23"]]
        reject_headers = ["period", "reject_rule", "rule_name", "reject_type", "reject_count", "distinct_users", "distinct_scenes", "rejected_amount_php", "benchmark_trxn", "trigger_rate_pct", "normalised_user_impact_pct"]
        reject_rows = [
            ["May 2026", "U0059", "PIN login velocity", "1", "33968", "20000", "2", "150000.5", "5000000", "0.679", "0.4"],
            ["May 2026", "D0191", "QR scam block", "2", "12479", "9000", "1", "0", "2000000", "0.624", "0.45"],
        ]
        breakdown_headers = ["reject_rule", "reject_type", "scene_name", "reject_count", "distinct_users", "rejected_amount_php", "benchmark_trxn", "trigger_rate_pct", "normalised_user_impact_pct"]
        breakdown_rows = [
            ["U0059", "1", "Login", "18088", "12000", "100000.5", "3000000", "0.603", "0.4"],
            ["U0059", "1", "ShopeepaySeabankInactiveLogin", "15880", "8000", "50000", "2000000", "0.794", "0.4"],
            ["D0191", "2", "QRScan", "12479", "9000", "0", "2000000", "0.624", "0.45"],
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "viz.html"
            write_visualization(
                output_path,
                report_title="Anti-fraud PH - Rule Effectiveness / Hit-Rate",
                snapshot_pt_date="2026-06-08",
                sheets=[
                    ("Request Outcome Summary", summary_headers, summary_rows),
                    ("Reject Rule Hit Summary", reject_headers, reject_rows),
                    ("Reject Rule Scene Breakdown", breakdown_headers, breakdown_rows),
                    ("Punishment Rule Hit Summary", ["period", "punish_rule_id", "rule_name", "punish_count", "distinct_targets", "distinct_scenes", "benchmark_trxn", "trigger_rate_pct", "normalised_user_impact_pct"], [["May 2026", "U0021", "Device block", "18492", "12000", "1", "4000000", "0.462", "0.3"]]),
                    ("Punishment Rule Scene Breakdown", ["punish_rule_id", "scene_name", "punish_count", "distinct_targets", "benchmark_trxn", "trigger_rate_pct", "normalised_user_impact_pct"], [["U0021", "Login", "18492", "12000", "4000000", "0.462", "0.3"]]),
                    ("Challenge Rule Hit Summary", ["rule_id", "rule_name", "rule_status", "review_priority", "challenge_trxn", "challenge_users", "distinct_scenes", "benchmark_trxn", "trigger_rate_pct", "normalised_user_impact_pct"], [["C0116v2", "Device velocity", "active", "1", "513407", "257072", "4", "1744983", "29.422", "14.73"]]),
                    ("Challenge Rule Scene Breakdown", ["rule_id", "scene_name", "challenge_trxn", "challenge_users", "benchmark_trxn", "trigger_rate_pct", "normalised_user_impact_pct"], [["C0116v2", "ApplyVirDCard", "300000", "150000", "1000000", "30.0", "15.0"]]),
                    ("Daily Challenge/Reject/Punish", ["trigger_date", "challenge_num", "reject_num", "punish_num"], [["20260501", "1", "2", "3"]]),
                ],
                report_id=AF_RULE_EFFECTIVENESS_REPORT_ID,
            )
            html = output_path.read_text(encoding="utf-8")

        # KPI cards derived from the summary row.
        self.assertIn("Hit-Rate Summary", html)
        self.assertIn("Action rate", html)
        self.assertIn("4.23%", html)
        self.assertIn('class="kpi"', html)
        # Period-switchable KPI panel + month selector data are embedded for the summary.
        self.assertIn("data-period-kpi", html)
        self.assertIn("data-period-json", html)
        # Expandable rule table with nested scene-breakdown rows.
        self.assertIn("<h2>Reject Rule Hit Summary</h2>", html)
        self.assertIn('class="rule-table"', html)
        self.assertIn('class="expander"', html)
        self.assertIn('class="detail-table"', html)
        self.assertIn("ShopeepaySeabankInactiveLogin", html)  # child scene name
        self.assertIn("rejected_amount_php", html)  # PHP-labelled column
        # Reject now also carries the benchmark + trigger-rate metric (summary and scene level).
        self.assertIn("benchmark_trxn", html)
        self.assertIn("normalised_user_impact_pct", html)
        self.assertIn("0.794", html)  # a scene-level reject trigger rate
        # The breakdown is nested, not rendered as its own standalone panel.
        self.assertNotIn("<h2>Reject Rule Scene Breakdown</h2>", html)
        # The other sections remain plain searchable tables.
        self.assertIn('class="search-table"', html)
        # Punishment Rule Hit Summary is also an expandable panel (nested breakdown, no amount).
        self.assertIn("<h2>Punishment Rule Hit Summary</h2>", html)
        self.assertNotIn("<h2>Punishment Rule Scene Breakdown</h2>", html)
        self.assertIn("punish_rule_id", html)
        self.assertIn("U0021", html)
        # Reject and Punish now carry rule_name alongside the rule id.
        self.assertIn("PIN login velocity", html)
        self.assertIn("Device block", html)
        self.assertEqual(html.count("<th>rule_name</th>"), 3)  # reject, punish, challenge summaries
        # Challenge Rule Hit Summary with the trigger-rate metric, also expandable.
        self.assertIn("<h2>Challenge Rule Hit Summary</h2>", html)
        self.assertNotIn("<h2>Challenge Rule Scene Breakdown</h2>", html)
        self.assertIn("trigger_rate_pct", html)
        self.assertIn("29.422", html)  # trigger rate value
        self.assertIn("ApplyVirDCard", html)  # nested challenge scene
        # Three expandable rule panels now (reject, punishment, challenge).
        self.assertEqual(html.count('class="rule-table"'), 3)


class RuleEffectivenessExtraSectionsTests(unittest.TestCase):
    def _viz(self, extra_sheets):
        base = [("Request Outcome Summary", ["period", "total_outcomes", "pass_num", "challenge_num", "reject_num", "action_rate_pct"], [["May 2026", "1", "1", "0", "0", "0"]])]
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "viz.html"
            write_visualization(
                output_path,
                report_title="Anti-fraud PH - Rule Effectiveness / Hit-Rate",
                snapshot_pt_date="2026-06-08",
                sheets=base + extra_sheets,
                report_id=AF_RULE_EFFECTIVENESS_REPORT_ID,
            )
            return output_path.read_text(encoding="utf-8")

    def test_precision_and_usage_render_as_searchable_tables(self):
        html = self._viz([
            ("Top Detecting Rules", ["rule_id", "rule_name", "flagged_cases", "fraud_cases_caught", "precision_pct", "loss_caught_php"],
             [["U0022v2", "New device transfer", "101", "10", "9.9", "0"]]),
            ("Scene/Sub-scene/Action Usage", ["scene_name", "sub_scene_name", "action_name", "action_type", "transactions", "action_events", "distinct_users"],
             [["Login", "PasswordLogin", "EnterLoginState", "Business", "6089791", "6100000", "1200000"]]),
        ])
        self.assertIn("<h2>Top Detecting Rules</h2>", html)
        self.assertIn("precision_pct", html)
        self.assertIn("New device transfer", html)
        self.assertIn("<h2>Scene/Sub-scene/Action Usage</h2>", html)
        self.assertIn("EnterLoginState", html)
        # New charts: outcome donut + top-scenes bar.
        self.assertIn("<h2>Outcome Mix</h2>", html)
        self.assertIn("<h2>Top Scenes by Transactions</h2>", html)
        # In-page TOC is built client-side for multi-panel reports.
        self.assertIn('nav.className', html)
        self.assertIn("'toc'", html)
        # Searchable, per-column-filterable tables (Request Outcome Summary + precision + usage = 3).
        self.assertEqual(html.count('class="search-table"'), 3)

    def test_daily_trend_renders_filterable_chart(self):
        html = self._viz([
            ("Daily Rule Trigger Trend", ["rule_id", "trigger_date", "trigger_trxn"],
             [["C0024v2", "2026-05-01", "1000"], ["C0024v2", "2026-05-02", "1500"], ["U0059", "2026-05-01", "500"]]),
        ])
        self.assertIn("<h2>Daily Rule Trigger Trend</h2>", html)
        # ECharts line chart: a container div + an inline echarts.init script with embedded data.
        self.assertIn('id="ec-daily-rule-trigger-trend" class="echart"', html)
        self.assertIn("echarts.init", html)
        self.assertIn("C0024v2", html)
        # Not rendered as a flat table.
        self.assertNotIn('<th>trigger_trxn</th>', html)

    def test_rule_scorecard_renders_scatter_and_table(self):
        html = self._viz([
            ("Rule Scorecard",
             ["rule_id", "rule_name", "trigger_trxn", "benchmark_trxn", "trigger_rate_pct", "reviewed_cases", "fraud_cases", "precision_pct"],
             [
                 ["U0022v2", "New device transfer", "50000", "1000000", "5.0", "101", "10", "9.9"],
                 ["D0123v2", "Recipient query limit", "144425", "19265700", "0.75", "1330", "3", "0.23"],
             ]),
        ])
        # Quadrant scatter (precision x trigger-rate) plus a searchable detail table.
        self.assertIn("<h2>Rule Scorecard</h2>", html)
        self.assertIn('id="ec-rule-scorecard" class="echart"', html)
        self.assertIn("echarts.init", html)
        self.assertIn("Rule Scorecard — Detail", html)
        # Cross-filter: scatter click highlights the rule across tables.
        self.assertIn("__afHighlight", html)
        self.assertIn("cross-hit", html)
        self.assertIn("precision_pct", html)
        self.assertIn("New device transfer", html)


class FraudLossBusinessInsightsTests(unittest.TestCase):
    def test_seeded_and_generator_registration(self):
        with tempfile.TemporaryDirectory() as tmp:
            ids = {r["id"] for r in BusinessInsightsStore(Path(tmp)).reports("anti-fraud")}
        self.assertIn(AF_FRAUD_LOSS_REPORT_ID, ids)
        self.assertIn(AF_FRAUD_LOSS_REPORT_ID, REPORT_BUILDERS)
        self.assertEqual(REPORT_BUILDERS[AF_FRAUD_LOSS_REPORT_ID][0], "Anti-fraud PH - Fraud Loss & Case Outcomes")

    def test_sql_sections_and_scope(self):
        sql = build_af_fraud_loss_sql(snapshot_pt_date="2026-06-08", now=FIXED_NOW)
        sections = extract_sql_sections(sql)
        self.assertEqual(
            [s.sheet_name for s in sections],
            [
                "Case & Loss Summary",
                "Loss by Fraud MO Type",
                "Fraud MO Subtype Breakdown",
                "Case Status & SLA",
                "Daily Fraud Loss Trend",
                "Review Pool / Backlog (current)",
            ],
        )
        self.assertIn(AF_REVIEW_CASE_TABLE, sql)
        # Backlog = current pending review-pool records (not yet a case).
        self.assertIn(AF_REVIEW_RECORD_TABLE, sql)
        self.assertIn("review_status = 'PENDING'", sql)
        self.assertIn("pending_records", sql)
        # Spans the last two full months + current MTD (Mar 1 -> May 27 under the fixed clock).
        self.assertIn("case_open_datetime >= '2026-03-01'", sql)
        self.assertIn("case_open_datetime < '2026-05-27'", sql)
        # Summary is per-period.
        self.assertIn("'May 2026 MTD'", sql)
        self.assertIn("not in ('not fraud', 'pending', '')", sql)
        self.assertIn("loss_amt_borne_by_customer", sql)
        self.assertIn("total_loss_php", sql)

    def test_visualization_renders_kpis_mo_expand_and_trend(self):
        sheets = [
            ("Case & Loss Summary",
             ["period", "cases_opened", "fraud_cases", "fraud_rate_pct", "total_loss_php", "loss_customer_php", "loss_bank_php", "loss_third_party_php", "recovered_php", "closed_cases", "open_cases", "avg_review_hours"],
             [["May 2026", "7842", "1330", "16.96", "18169639.46", "17993073.06", "139", "176427.4", "52947.11", "7562", "280", "25.6"]]),
            ("Loss by Fraud MO Type",
             ["fraud_mo_type", "cases", "distinct_subtypes", "total_loss_php", "avg_loss_php", "recovered_php"],
             [["Scam", "881", "5", "15000000", "17026.1", "40000"], ["Not Fraud", "6156", "0", "0", "0", "0"]]),
            ("Fraud MO Subtype Breakdown",
             ["fraud_mo_type", "fraud_mo_subtype", "cases", "total_loss_php", "recovered_php"],
             [["Scam", "Investment Scam", "400", "9000000", "20000"], ["Scam", "Romance Scam", "200", "6000000", "20000"]]),
            ("Case Status & SLA", ["case_status", "cases", "total_loss_php", "avg_review_hours"],
             [["CLOSED", "7562", "16124651.25", "24.1"], ["OPEN", "280", "2044988.21", ""]]),
            ("Daily Fraud Loss Trend", ["fraud_mo_type", "case_open_date", "daily_loss_php", "daily_cases"],
             [["All", "2026-05-01", "500000", "260"], ["Scam", "2026-05-01", "480000", "30"]]),
            ("Review Pool / Backlog (current)", ["source", "pending_records", "distinct_rules", "distinct_users", "oldest_pending", "avg_age_days"],
             [["IHCS", "7698", "1", "5366", "2025-02-21 21:58:02", "186"]]),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "viz.html"
            write_visualization(
                output_path,
                report_title="Anti-fraud PH - Fraud Loss & Case Outcomes",
                snapshot_pt_date="2026-06-08",
                sheets=sheets,
                report_id=AF_FRAUD_LOSS_REPORT_ID,
            )
            html = output_path.read_text(encoding="utf-8")

        # KPI cards (loss formatted with peso sign), MO expandable, status table, trend chart.
        self.assertIn("Fraud Loss Summary", html)
        self.assertIn("Total loss", html)
        self.assertIn("₱", html)  # peso sign
        self.assertIn("<h2>Loss by Fraud MO Type</h2>", html)
        self.assertIn("<h2>Loss by Fraud Type (PHP)</h2>", html)  # MO loss bar chart
        self.assertIn('class="rule-table"', html)
        self.assertIn("Investment Scam", html)  # nested subtype
        self.assertNotIn("<h2>Fraud MO Subtype Breakdown</h2>", html)  # nested, not standalone
        self.assertIn("<h2>Case Status &amp; SLA</h2>", html)
        self.assertIn("<h2>Daily Fraud Loss Trend</h2>", html)
        self.assertIn('id="ec-daily-fraud-loss-trend" class="echart"', html)
        self.assertIn("echarts.init", html)
        # ECharts is loaded from the vendored local asset (no external CDN).
        self.assertIn('/static/vendor/echarts.min.js', html)
        # Current review-pool backlog renders as a searchable table.
        self.assertIn("<h2>Review Pool / Backlog (current)</h2>", html)
        self.assertIn("IHCS", html)
        # The single-row summary is shown as KPIs, not a table panel.
        self.assertNotIn("<h2>Case &amp; Loss Summary</h2>", html)


class DetectionEffectivenessBusinessInsightsTests(unittest.TestCase):
    def test_folded_into_rule_effectiveness_not_standalone(self):
        with tempfile.TemporaryDirectory() as tmp:
            ids = {r["id"] for r in BusinessInsightsStore(Path(tmp)).reports("anti-fraud")}
        # Detection effectiveness is no longer a standalone report; it is folded into Rule Effectiveness.
        self.assertNotIn(AF_DETECTION_EFFECTIVENESS_REPORT_ID, ids)
        self.assertNotIn(AF_DETECTION_EFFECTIVENESS_REPORT_ID, REPORT_BUILDERS)
        names = [s.sheet_name for s in extract_sql_sections(build_af_rule_effectiveness_sql(now=FIXED_NOW))]
        for sheet in ("Detection Coverage Summary", "Detection by Fraud MO Type", "Top Detecting Rules",
                      "Missed Fraud (Blind Spots)", "Daily Detection Trend"):
            self.assertIn(sheet, names)

    def test_sql_sections_and_scope(self):
        sql = build_af_detection_effectiveness_sql(snapshot_pt_date="2026-06-08", now=FIXED_NOW)
        sections = extract_sql_sections(sql)
        self.assertEqual(
            [s.sheet_name for s in sections],
            [
                "Detection Coverage Summary",
                "Detection by Fraud MO Type",
                "Top Detecting Rules",
                "Missed Fraud (Blind Spots)",
                "Daily Detection Trend",
            ],
        )
        # Reuses the proven review_record -> review_case join (rule that flagged the case for review).
        self.assertIn(AF_REVIEW_CASE_TABLE, sql)
        self.assertIn(AF_REVIEW_RECORD_TABLE, sql)
        self.assertIn("rule_detected", sql)
        self.assertIn("loss_leaked_php", sql)
        self.assertIn("not in ('not fraud', 'pending', '')", sql)
        # Spans the last two full months + current MTD (Mar 1 -> May 27 under the fixed clock).
        self.assertIn("case_open_datetime >= '2026-03-01'", sql)
        self.assertIn("case_open_datetime < '2026-05-27'", sql)
        self.assertIn("'May 2026 MTD'", sql)

    def test_detection_panels_render_within_rule_effectiveness(self):
        sheets = [
            ("Detection Coverage Summary",
             ["period", "fraud_cases", "rule_detected_cases", "detection_rate_pct", "fraud_loss_php",
              "loss_rule_detected_php", "loss_leaked_php", "loss_detected_share_pct"],
             [["Mar 2026", "120", "90", "75.0", "5000000", "3800000", "1200000", "76.0"],
              ["Apr 2026", "140", "98", "70.0", "6200000", "4300000", "1900000", "69.4"],
              ["May 2026 MTD", "40", "30", "75.0", "1800000", "1300000", "500000", "72.2"]]),
            ("Detection by Fraud MO Type",
             ["fraud_mo_type", "fraud_cases", "rule_detected_cases", "detection_rate_pct",
              "total_loss_php", "loss_detected_php", "loss_leaked_php"],
             [["Account Takeover", "60", "40", "66.7", "3000000", "2100000", "900000"]]),
            ("Top Detecting Rules",
             ["rule_id", "rule_name", "flagged_cases", "fraud_cases_caught", "precision_pct", "loss_caught_php"],
             [["R001", "ATO velocity", "120", "90", "75.0", "2100000"]]),
            ("Missed Fraud (Blind Spots)",
             ["fraud_mo_type", "fraud_mo_subtype", "missed_cases", "loss_leaked_php", "recovered_php"],
             [["Account Takeover", "SIM swap", "20", "900000", "100000"]]),
            ("Daily Detection Trend", ["series", "case_open_date", "daily_loss_php", "fraud_cases"],
             [["All fraud loss", "2026-05-01", "150000", "5"]]),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "viz.html"
            write_visualization(
                output_path,
                report_title="Anti-fraud PH - Rule Effectiveness / Hit-Rate",
                snapshot_pt_date="2026-06-08",
                sheets=sheets,
                report_id=AF_RULE_EFFECTIVENESS_REPORT_ID,
            )
            html = output_path.read_text(encoding="utf-8")
        # Detection coverage KPIs (period-switchable) + highlights fold into the Rule Effectiveness page.
        self.assertIn("Detection Coverage", html)
        self.assertIn("Detection Highlights", html)
        self.assertIn("data-period-kpi", html)
        self.assertIn("May 2026 MTD", html)
        self.assertIn("<h2>Leaked Loss by Fraud Type (PHP)</h2>", html)
        self.assertIn("<h2>Missed Fraud (Blind Spots)</h2>", html)
        self.assertIn("SIM swap", html)
        self.assertIn('id="ec-daily-detection-trend" class="echart"', html)


class RuleChangeLogBusinessInsightsTests(unittest.TestCase):
    def test_folded_into_rules_features_not_standalone(self):
        with tempfile.TemporaryDirectory() as tmp:
            ids = {r["id"] for r in BusinessInsightsStore(Path(tmp)).reports("anti-fraud")}
        # Rule change log is no longer a standalone report; it is folded into Rules & Features.
        self.assertNotIn(AF_RULE_CHANGE_LOG_REPORT_ID, ids)
        self.assertNotIn(AF_RULE_CHANGE_LOG_REPORT_ID, REPORT_BUILDERS)
        names = [s.sheet_name for s in extract_sql_sections(build_af_rules_features_sql(now=FIXED_NOW))]
        for sheet in ("Change Summary", "Rule Change Detail", "Current Rule Inventory"):
            self.assertIn(sheet, names)

    def test_sql_sections_and_baseline(self):
        sql = build_af_rule_change_log_sql(snapshot_pt_date="2026-06-08", now=FIXED_NOW)
        sections = extract_sql_sections(sql)
        self.assertEqual(
            [s.sheet_name for s in sections],
            ["Change Summary", "Rule Change Detail", "Current Rule Inventory"],
        )
        self.assertIn(AF_RULE_CONFIG_TABLE, sql)
        # Diffs the current snapshot against the snapshot on/before the window start (Mar 1),
        # falling back to the earliest retained snapshot.
        self.assertIn("pt_date <= '2026-03-01'", sql)
        self.assertIn("select min(pt_date)", sql)
        self.assertIn("full outer join", sql)
        self.assertIn("'Added'", sql)
        self.assertIn("'Deactivated'", sql)
        self.assertIn("Logic changed", sql)

    def test_change_log_panels_render_within_rules_features(self):
        sheets = [
            ("Rules", ["rule_id", "rule_name", "outcome_type", "rule_status", "risk_level"],
             [["R1", "Velocity", "Reject", "Active", "high"]]),
            ("Features", ["feature_id", "feature_name", "feature_status"], [["F1", "cnt 1h", "Active"]]),
            ("Change Summary", ["change_type", "rules"],
             [["Added", "5"], ["Deactivated", "3"], ["Logic changed", "4"], ["Unchanged", "900"]]),
            ("Rule Change Detail",
             ["change_type", "rule_id", "rule_name", "status_before", "status_after", "outcome_before",
              "outcome_after", "risk_before", "risk_after", "review_priority_before", "review_priority_after", "logic_changed"],
             [["Added", "R900", "New scam rule", "Inactive/Draft", "Active", "Reject", "Reject", "high", "high", "1", "1", "N"]]),
            ("Current Rule Inventory", ["outcome_type", "rule_status", "risk_level", "rules"],
             [["Reject", "Active", "high", "120"], ["Challenge", "Active", "medium", "80"]]),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "viz.html"
            write_visualization(
                output_path,
                report_title="Anti-fraud PH - Rules & Features",
                snapshot_pt_date="2026-06-08",
                sheets=sheets,
                report_id=AF_RULES_FEATURES_REPORT_ID,
            )
            html = output_path.read_text(encoding="utf-8")
        # Change-log governance folds into the Rules & Features page.
        self.assertIn("Rule Change Summary", html)
        self.assertIn("Rule Change Highlights", html)
        self.assertIn("<h2>Current Rules by Outcome Type</h2>", html)
        self.assertIn("<h2>Rule Change Detail</h2>", html)
        self.assertIn("New scam rule", html)


class FacialVerificationBusinessInsightsTests(unittest.TestCase):
    def test_seeded_and_generator_registration(self):
        with tempfile.TemporaryDirectory() as tmp:
            ids = {r["id"] for r in BusinessInsightsStore(Path(tmp)).reports("anti-fraud")}
        self.assertIn(AF_FACIAL_VERIFICATION_REPORT_ID, ids)
        self.assertIn(AF_FACIAL_VERIFICATION_REPORT_ID, REPORT_BUILDERS)
        self.assertEqual(
            REPORT_BUILDERS[AF_FACIAL_VERIFICATION_REPORT_ID][0],
            "Anti-fraud PH - Facial Verification / Liveness & Deepfake",
        )

    def test_sql_sections_and_scope(self):
        sql = build_af_facial_verification_sql(snapshot_pt_date="2026-06-08", now=FIXED_NOW)
        sections = extract_sql_sections(sql)
        self.assertEqual(
            [s.sheet_name for s in sections],
            [
                "Verification Outcome Summary",
                "Liveness Result Breakdown",
                "Anti-Spoofing QC Breakdown",
                "Facial Match Result Breakdown",
                "Deepfake Score Distribution",
                "Pass Rates by Scene",
                "Human Review (AMR) Outcomes",
                "Review Status & Verdict Detail",
                "Reviewer Workload & Turnaround",
                "CS Review Track",
                "Daily Verification Trend",
            ],
        )
        self.assertIn(AF_FACIAL_VERIFICATION_TABLE, sql)
        # 3-step funnel result columns + spoof / deepfake signals.
        self.assertIn("liveness_check_result = 'LC_SUCCESS'", sql)
        self.assertIn("selfie_qc_anti_spoofing_result = 'SQA_SUCCESS'", sql)
        self.assertIn("facial_matching_result = 'FM_SUCCESS'", sql)
        self.assertIn("SQA_REJECT_FACE_DEEPFAKE", sql)
        self.assertIn("deepfake_spoof_score", sql)
        # Scoped by create_datetime over the window (Mar 1 -> May 27 under the fixed clock).
        self.assertIn("create_datetime >= '2026-03-01'", sql)
        self.assertIn("create_datetime < '2026-05-27'", sql)
        self.assertIn("'May 2026 MTD'", sql)

    def test_visualization_renders_kpis_deepfake_bar_and_period_selector(self):
        sheets = [
            ("Verification Outcome Summary",
             ["period", "checks", "distinct_users", "liveness_pass_rate_pct", "antispoof_pass_rate_pct",
              "match_pass_rate_pct", "overall_pass_rate_pct", "spoof_attack_checks", "deepfake_reject_checks"],
             [["Mar 2026", "1577166", "890554", "96.62", "96.81", "98.22", "91.87", "25823", "8509"],
              ["Apr 2026", "1600729", "935814", "96.83", "97.46", "98.19", "92.67", "20854", "8633"],
              ["May 2026 MTD", "1656245", "981678", "96.75", "97.55", "98.13", "92.61", "18898", "8710"]]),
            ("Liveness Result Breakdown", ["liveness_check_result", "checks", "distinct_users", "share_pct"],
             [["LC_SUCCESS", "21846498", "1", "88.0"], ["LC_AURORA_SPOOF", "182935", "1", "0.7"]]),
            ("Anti-Spoofing QC Breakdown", ["selfie_qc_anti_spoofing_result", "checks", "distinct_users", "share_pct"],
             [["SQA_SUCCESS", "21000000", "1", "95"], ["SQA_REJECT_FACE_DEEPFAKE", "54409", "1", "0.2"]]),
            ("Facial Match Result Breakdown", ["facial_matching_result", "checks", "distinct_users", "avg_match_score", "share_pct"],
             [["FM_SUCCESS", "20000000", "1", "0.82", "98"]]),
            ("Deepfake Score Distribution", ["deepfake_score_band", "checks", "spoof_rejects", "avg_score"],
             [["1. 0.0-0.2", "100", "0", "0.1"], ["5. 0.8-1.0", "5000", "4800", "0.95"]]),
            ("Pass Rates by Scene",
             ["scene_name", "checks", "distinct_users", "liveness_pass_rate_pct", "antispoof_pass_rate_pct", "overall_pass_rate_pct", "spoof_attack_checks"],
             [["UnlockCard", "20000", "1", "70", "95", "62", "500"]]),
            ("Fraud Review Outcomes", ["fraud_review_status", "fraud_review_result", "checks", "distinct_users"],
             [["REVIEWED", "3", "26482", "1"]]),
            ("Daily Verification Trend", ["series", "check_date", "checks"],
             [["Checks", "2026-05-01", "50000"], ["Spoof attacks", "2026-05-01", "800"]]),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "viz.html"
            write_visualization(
                output_path,
                report_title="Anti-fraud PH - Facial Verification / Liveness & Deepfake",
                snapshot_pt_date="2026-06-08",
                sheets=sheets,
                report_id=AF_FACIAL_VERIFICATION_REPORT_ID,
            )
            html = output_path.read_text(encoding="utf-8")
        # Period-switchable verification KPIs drive the month selector.
        self.assertIn("Verification Outcomes", html)
        self.assertIn("data-period-kpi", html)
        self.assertIn('<script type="application/json" data-period-json>', html)
        self.assertIn("May 2026 MTD", html)
        # Highlights + deepfake distribution bar + scene/QC/match tables + trend.
        self.assertIn("<h2>Highlights</h2>", html)
        self.assertIn("<h2>Deepfake Score Distribution</h2>", html)
        self.assertIn("<h2>Anti-Spoofing QC Breakdown</h2>", html)
        self.assertIn("SQA_REJECT_FACE_DEEPFAKE", html)
        self.assertIn('id="ec-daily-verification-trend" class="echart"', html)


class BusinessInsightsGenerationJobTests(unittest.TestCase):
    def _root(self, temp_dir):
        return Path(temp_dir)

    def test_status_is_idle_without_state(self):
        from bpmis_jira_tool import business_insights_jobs as jobs

        with tempfile.TemporaryDirectory() as temp_dir:
            status = jobs.generation_job_status(root_dir=self._root(temp_dir), report_id=AF_RULE_EFFECTIVENESS_REPORT_ID)
        self.assertEqual(status["status"], "idle")

    def test_running_then_completed_via_exit_marker(self):
        from bpmis_jira_tool import business_insights_jobs as jobs

        report_id = AF_RULE_EFFECTIVENESS_REPORT_ID
        with tempfile.TemporaryDirectory() as temp_dir:
            root = self._root(temp_dir)
            with patch("bpmis_jira_tool.business_insights_jobs.subprocess.Popen") as popen:
                popen.return_value = type("Proc", (), {"pid": os.getpid()})()
                started = jobs.start_generation_job(root_dir=root, report_id=report_id, script_path="/tmp/fake.py")
            self.assertEqual(started["status"], "running")
            self.assertTrue(popen.called)

            log_path = root / jobs.JOBS_DIRNAME / f"{report_id}.log"
            log_path.write_text(f"{report_id}: rows=42\n{jobs.EXIT_MARKER}0\n", encoding="utf-8")
            status = jobs.generation_job_status(root_dir=root, report_id=report_id)
        self.assertEqual(status["status"], "completed")
        self.assertEqual(status["exit_code"], 0)

    def test_nonzero_exit_marks_failed_with_session_hint(self):
        from bpmis_jira_tool import business_insights_jobs as jobs

        report_id = AF_RULES_FEATURES_REPORT_ID
        with tempfile.TemporaryDirectory() as temp_dir:
            root = self._root(temp_dir)
            (root / jobs.JOBS_DIRNAME).mkdir(parents=True)
            (root / jobs.JOBS_DIRNAME / f"{report_id}.json").write_text(
                json.dumps({"report_id": report_id, "pid": os.getpid(), "status": "running"}), encoding="utf-8"
            )
            (root / jobs.JOBS_DIRNAME / f"{report_id}.log").write_text(
                f"Data Admin session is not valid: HTTP 401\n{jobs.EXIT_MARKER}1\n", encoding="utf-8"
            )
            status = jobs.generation_job_status(root_dir=root, report_id=report_id)
        self.assertEqual(status["status"], "failed")
        self.assertIn("data-admin.ph.seabank.io", status["error"])

    def test_dead_process_without_marker_is_failed(self):
        from bpmis_jira_tool import business_insights_jobs as jobs

        report_id = PORTFOLIO_REPAYMENT_REPORT_ID
        with tempfile.TemporaryDirectory() as temp_dir:
            root = self._root(temp_dir)
            (root / jobs.JOBS_DIRNAME).mkdir(parents=True)
            (root / jobs.JOBS_DIRNAME / f"{report_id}.json").write_text(
                json.dumps({"report_id": report_id, "pid": 999_999_999, "status": "running"}), encoding="utf-8"
            )
            (root / jobs.JOBS_DIRNAME / f"{report_id}.log").write_text("partial work...\n", encoding="utf-8")
            status = jobs.generation_job_status(root_dir=root, report_id=report_id)
        self.assertEqual(status["status"], "failed")

    def test_start_is_idempotent_while_running(self):
        from bpmis_jira_tool import business_insights_jobs as jobs

        report_id = LIMIT_UTILIZATION_REPORT_ID
        with tempfile.TemporaryDirectory() as temp_dir:
            root = self._root(temp_dir)
            (root / jobs.JOBS_DIRNAME).mkdir(parents=True)
            (root / jobs.JOBS_DIRNAME / f"{report_id}.json").write_text(
                json.dumps({"report_id": report_id, "pid": os.getpid(), "status": "running"}), encoding="utf-8"
            )
            with patch("bpmis_jira_tool.business_insights_jobs.subprocess.Popen") as popen:
                result = jobs.start_generation_job(root_dir=root, report_id=report_id, script_path="/tmp/fake.py")
            self.assertFalse(popen.called)
            self.assertTrue(result.get("already_running"))


class BusinessInsightsGenerationRouteTests(unittest.TestCase):
    def _client_env(self, temp_dir):
        return {
            "ENV_FILE": os.devnull,
            "FLASK_SECRET_KEY": "test-secret",
            "TEAM_PORTAL_DATA_DIR": temp_dir,
            "TEAM_PORTAL_BASE_URL": "",
            "TEAM_ALLOWED_EMAILS": "",
            "TEAM_ALLOWED_EMAIL_DOMAINS": "",
            "TEAM_PORTAL_CONFIG_ENCRYPTION_KEY": "",
        }

    def _admin_session(self, client):
        with client.session_transaction() as session:
            session["google_profile"] = {"email": "xiaodong.zheng@npt.sg", "name": "Admin"}
            session["google_credentials"] = {"token": "x"}

    def _portal_user_session(self, client):
        # A non-admin portal user (any @npt.sg address that is not the admin).
        with client.session_transaction() as session:
            session["google_profile"] = {"email": "teammate@npt.sg", "name": "Teammate"}
            session["google_credentials"] = {"token": "x"}

    def test_refresh_button_renders_for_generator_reports(self):
        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(os.environ, self._client_env(temp_dir), clear=True):
            app = create_app()
            app.testing = True
            with app.test_client() as client:
                self._admin_session(client)
                page = client.get("/business-insights?domain=anti-fraud")
                reports = client.get("/api/business-insights/reports?domain=anti-fraud").get_json()
        soup = BeautifulSoup(page.get_data(as_text=True), "html.parser")
        self.assertTrue(soup.select("[data-business-insights-generate]"))
        rule_eff = next(r for r in reports["reports"] if r["id"] == AF_RULE_EFFECTIVENESS_REPORT_ID)
        self.assertTrue(rule_eff["can_generate"])
        self.assertIn("/generate", rule_eff["generate_url"])
        funnel = next((r for r in reports["reports"] if r["id"] == UNDERWRITING_FUNNEL_REPORT_ID), None)
        if funnel is not None:
            self.assertFalse(funnel["can_generate"])

    def test_generate_rejects_non_generator_report(self):
        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(os.environ, self._client_env(temp_dir), clear=True):
            app = create_app()
            app.testing = True
            with app.test_client() as client:
                self._admin_session(client)
                response = client.post(f"/api/business-insights/reports/{UNDERWRITING_FUNNEL_REPORT_ID}/generate")
        self.assertEqual(response.status_code, 404)

    def test_generate_starts_job_and_status_reports_completion(self):
        report_id = AF_RULE_EFFECTIVENESS_REPORT_ID
        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(os.environ, self._client_env(temp_dir), clear=True):
            app = create_app()
            app.testing = True
            with app.test_client() as client:
                self._admin_session(client)
                with patch(
                    "bpmis_jira_tool.web_business_insights_routes.start_generation_job",
                    return_value={"report_id": report_id, "status": "running", "pid": 4321},
                ) as start:
                    generate = client.post(f"/api/business-insights/reports/{report_id}/generate")
                with patch(
                    "bpmis_jira_tool.web_business_insights_routes.generation_job_status",
                    return_value={"report_id": report_id, "status": "completed", "exit_code": 0},
                ):
                    status = client.get(f"/api/business-insights/reports/{report_id}/generate/status")
        self.assertTrue(start.called)
        self.assertEqual(generate.status_code, 200)
        self.assertEqual(generate.get_json()["job"]["status"], "running")
        self.assertEqual(status.status_code, 200)
        self.assertEqual(status.get_json()["job"]["status"], "completed")

    def test_generate_denies_non_portal_user(self):
        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(os.environ, self._client_env(temp_dir), clear=True):
            app = create_app()
            app.testing = True
            with app.test_client() as client:
                with client.session_transaction() as session:
                    session["google_profile"] = {"email": "outsider@example.com", "name": "Outsider"}
                    session["google_credentials"] = {"token": "x"}
                response = client.post(f"/api/business-insights/reports/{AF_RULE_EFFECTIVENESS_REPORT_ID}/generate")
        self.assertEqual(response.status_code, 403)

    def test_generate_denies_non_admin_portal_user(self):
        report_id = AF_RULE_EFFECTIVENESS_REPORT_ID
        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(os.environ, self._client_env(temp_dir), clear=True):
            app = create_app()
            app.testing = True
            with app.test_client() as client:
                self._portal_user_session(client)
                generate = client.post(f"/api/business-insights/reports/{report_id}/generate")
                status = client.get(f"/api/business-insights/reports/{report_id}/generate/status")
        self.assertEqual(generate.status_code, 403)
        self.assertEqual(status.status_code, 403)

    def test_refresh_button_hidden_for_non_admin_portal_user(self):
        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(os.environ, self._client_env(temp_dir), clear=True):
            app = create_app()
            app.testing = True
            with app.test_client() as client:
                self._portal_user_session(client)
                page = client.get("/business-insights?domain=anti-fraud")
                reports = client.get("/api/business-insights/reports?domain=anti-fraud").get_json()
        # A non-admin can view the reports but must not see the on-demand refresh control.
        self.assertEqual(page.status_code, 200)
        soup = BeautifulSoup(page.get_data(as_text=True), "html.parser")
        self.assertFalse(soup.select("[data-business-insights-generate]"))
        rule_eff = next(r for r in reports["reports"] if r["id"] == AF_RULE_EFFECTIVENESS_REPORT_ID)
        self.assertFalse(rule_eff["can_generate"])
        self.assertNotIn("generate_url", rule_eff)


class DeviceRiskBusinessInsightsTests(unittest.TestCase):
    def test_seeded_and_generator_registration(self):
        with tempfile.TemporaryDirectory() as tmp:
            ids = {r["id"] for r in BusinessInsightsStore(Path(tmp)).reports("anti-fraud")}
        self.assertIn(AF_DEVICE_RISK_REPORT_ID, ids)
        self.assertIn(AF_DEVICE_RISK_REPORT_ID, REPORT_BUILDERS)
        self.assertEqual(REPORT_BUILDERS[AF_DEVICE_RISK_REPORT_ID][0], "Anti-fraud PH - Device & Identity Risk")

    def test_sql_sections_recompute_farming_and_window_by_pt_date(self):
        sql = build_af_device_risk_sql(now=FIXED_NOW)
        sections = extract_sql_sections(sql)
        self.assertEqual(
            [s.sheet_name for s in sections],
            ["Account Farming Summary", "Top Multi-Account Devices", "Multi-Device Accounts",
             "Risk Signal Prevalence", "Risk Signals by Scene", "Multi-Account Device Trend"],
        )
        self.assertIn("dwd_antifraud_action_log_ext_di", sql)
        # Account farming is recomputed (precomputed link columns are encrypted), windowed by pt_date range.
        self.assertIn("count(distinct uid)", sql)
        self.assertIn("count(distinct deviceuuid)", sql)
        self.assertIn("pt_date >=", sql)
        self.assertNotIn("max(pt_date)", sql)  # never scan max() on this 64k-partition table
        # Flag prevalence treats encrypted/clear positives alike via the known-negative exclusion set.
        self.assertIn("cfcd208495d565ef66e7dff9f98764da", sql)

    def test_visualization_renders_farming_kpis_and_signal_bar(self):
        sheets = [
            ("Account Farming Summary",
             ["devices_seen", "devices_5plus_accounts", "devices_10plus_accounts",
              "max_accounts_on_one_device", "pct_devices_5plus_accounts"],
             [["1714805", "320", "48", "66", "0.0187"]]),
            ("Top Multi-Account Devices", ["deviceuuid", "distinct_accounts", "distinct_scenes", "events"],
             [["d25760e5", "66", "4", "210"]]),
            ("Multi-Device Accounts", ["uid", "distinct_devices", "distinct_scenes", "events"], [["uidA", "12", "3", "40"]]),
            ("Risk Signal Prevalence",
             ["total_events", "rooted", "emulator", "vpn", "new_deviceid", "autoclicker"],
             [["23419734", "2", "15", "0", "120000", "6"]]),
            ("Risk Signals by Scene", ["scene_name", "events", "distinct_accounts", "risky_events"],
             [["Login", "900000", "800000", "120"]]),
            ("Multi-Account Device Trend", ["series", "event_date", "value"],
             [["Multi-account devices (>=5)", "2026-06-05", "40"]]),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "viz.html"
            write_visualization(output_path, report_title="Anti-fraud PH - Device & Identity Risk",
                                snapshot_pt_date="2026-06-10", sheets=sheets, report_id=AF_DEVICE_RISK_REPORT_ID)
            html = output_path.read_text(encoding="utf-8")
        self.assertIn("Account Farming", html)
        self.assertIn("Most-shared device", html)
        self.assertIn("<h2>Top Device Risk Signals (events)</h2>", html)
        self.assertIn("<h2>Risk Signal Prevalence</h2>", html)
        self.assertIn('id="ec-multi-account-device-trend" class="echart"', html)
        self.assertIn("d25760e5", html)


class Card3dsBusinessInsightsTests(unittest.TestCase):
    def test_seeded_and_generator_registration(self):
        with tempfile.TemporaryDirectory() as tmp:
            ids = {r["id"] for r in BusinessInsightsStore(Path(tmp)).reports("anti-fraud")}
        self.assertIn(AF_CARD_3DS_REPORT_ID, ids)
        self.assertIn(AF_CARD_3DS_REPORT_ID, REPORT_BUILDERS)
        self.assertEqual(REPORT_BUILDERS[AF_CARD_3DS_REPORT_ID][0], "Anti-fraud PH - Card Fraud & 3DS Authentication")

    def test_sql_sections_and_sources(self):
        sql = build_af_card_3ds_sql(now=FIXED_NOW)
        sections = extract_sql_sections(sql)
        self.assertEqual(
            [s.sheet_name for s in sections],
            ["3DS Authentication Summary", "Outcome by Auth Status", "Frictionless vs Challenge",
             "3DS by Merchant Category (MCC)", "Card Fraud Cases by MO", "Daily 3DS Trend"],
        )
        self.assertIn("pmt_threeds_acs_t_threeds_trans_ss_d", sql)
        self.assertIn("mbs_card_fraud_case_ss_d", sql)
        self.assertIn("trans_status = 'Y'", sql)
        self.assertIn("auth_rate_pct", sql)
        # Challenge = EMV transStatus 'C' (the ODS table has no cnp_decision column; verified on portal).
        self.assertIn("trans_status = 'C'", sql)
        # Spans the last two full months + current MTD (Mar 1 -> May 27 under the fixed clock; period bucketed).
        self.assertIn("pt_date >= '2026-03-01'", sql)

    def test_visualization_renders_auth_kpis_and_outcome_donut(self):
        sheets = [
            ("3DS Authentication Summary",
             ["period", "threeds_txns", "authenticated", "not_authenticated", "challenged", "rejected",
              "unavailable", "auth_rate_pct", "challenge_rate_pct"],
             [["Apr 2026", "5000000", "4000000", "800000", "120000", "50000", "30000", "80.0", "2.4"],
              ["May 2026", "5200000", "4300000", "700000", "130000", "45000", "25000", "82.7", "2.5"],
              ["Jun 2026 MTD", "1800000", "1500000", "220000", "50000", "18000", "12000", "83.3", "2.8"]]),
            ("Outcome by Auth Status", ["auth_status", "threeds_txns", "share_pct"],
             [["Authenticated", "4808715", "79.8"], ["Not authenticated", "1207508", "20.0"]]),
            ("Frictionless vs Challenge",
             ["challenge_indicator", "threeds_txns", "challenged", "authenticated", "challenge_rate_pct"],
             [["04 Challenge requested (mandate)", "95574786", "562405", "73159766", "0.59"]]),
            ("3DS by Merchant Category (MCC)", ["mcc", "threeds_txns", "authenticated", "auth_rate_pct", "purchase_amount_php"],
             [["5399", "900000", "800000", "88.9", "12500000.00"]]),
            ("Card Fraud Cases by MO", ["mo_reason", "sub_mo_reason", "cases"], [["Cards", "Card Not Present Fraud", "42"]]),
            ("Daily 3DS Trend", ["series", "txn_date", "txns"],
             [["Authenticated", "2026-06-01", "150000"], ["Challenge", "2026-06-01", "5000"]]),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "viz.html"
            write_visualization(output_path, report_title="Anti-fraud PH - Card Fraud & 3DS Authentication",
                                snapshot_pt_date="2026-06-10", sheets=sheets, report_id=AF_CARD_3DS_REPORT_ID)
            html = output_path.read_text(encoding="utf-8")
        self.assertIn("3DS Authentication", html)
        self.assertIn("data-period-kpi", html)
        self.assertIn("<h2>3DS Outcome Mix</h2>", html)
        self.assertIn("<h2>Card Fraud Cases by Sub-MO</h2>", html)
        self.assertIn('id="ec-daily-3ds-trend" class="echart"', html)
        self.assertIn("Card Not Present Fraud", html)


class ListUsageBusinessInsightsTests(unittest.TestCase):
    def test_seeded_and_generator_registration(self):
        with tempfile.TemporaryDirectory() as tmp:
            ids = {r["id"] for r in BusinessInsightsStore(Path(tmp)).reports("anti-fraud")}
        self.assertIn(AF_LIST_USAGE_REPORT_ID, ids)
        self.assertIn(AF_LIST_USAGE_REPORT_ID, REPORT_BUILDERS)
        self.assertEqual(REPORT_BUILDERS[AF_LIST_USAGE_REPORT_ID][0], "Anti-fraud PH - Blacklist, Whitelist & Greylist")

    def test_store_returns_sql_for_list_usage(self):
        with tempfile.TemporaryDirectory() as tmp:
            sql = BusinessInsightsStore(Path(tmp)).sql_for_report(AF_LIST_USAGE_REPORT_ID, now=FIXED_NOW)
        self.assertIn(AF_BLACK_WHITE_LIST_TABLE, sql)

    def test_sql_sections_and_sources(self):
        sql = build_af_list_usage_sql(snapshot_pt_date="2026-06-12", now=FIXED_NOW)
        sections = extract_sql_sections(sql)
        self.assertEqual(
            [s.sheet_name for s in sections],
            ["List Overview", "Black/White by Status", "Black/White by ID Type", "Black/White by Source",
             "Black/White by Listed Reason", "Black/White by Scenario", "Monthly Additions", "Greylist Detail"],
        )
        self.assertIn(AF_BLACK_WHITE_LIST_TABLE, sql)
        self.assertIn(AF_GREY_LIST_TABLE, sql)
        # Black/white pinned to the report snapshot.
        self.assertIn("bw.pt_date = '2026-06-12'", sql)
        # Greylist must self-resolve its own latest snapshot (it lags the black/white table), never the anchor.
        self.assertIn(f"gl.pt_date = (select max(pt_date) from {AF_GREY_LIST_TABLE})", sql)
        self.assertNotIn("gl.pt_date = '2026-06-12'", sql)
        # list_type mapping (1 = blacklist, 0/2 = whitelist).
        self.assertIn("when bw.list_type = 1 then 'Blacklist'", sql)
        self.assertIn("when bw.list_type in (0, 2) then 'Whitelist'", sql)
        # Monthly growth derives the month from create_date epoch-ms.
        self.assertIn("from_unixtime(cast(bw.create_date / 1000 as bigint), 'yyyy-MM')", sql)

    def test_greylist_self_resolves_even_without_snapshot(self):
        sql = build_af_list_usage_sql(now=FIXED_NOW)
        self.assertIn(f"bw.pt_date = (select max(pt_date) from {AF_BLACK_WHITE_LIST_TABLE})", sql)
        self.assertIn(f"gl.pt_date = (select max(pt_date) from {AF_GREY_LIST_TABLE})", sql)

    def test_visualization_renders_kpis_donut_and_trend(self):
        sheets = [
            ("List Overview", ["list_name", "entries", "distinct_targets"],
             [["Blacklist", "193012", "161843"], ["Whitelist", "52", "27"], ["Greylist", "8", "8"]]),
            ("Black/White by Status", ["list_name", "status", "entries", "distinct_targets"],
             [["Blacklist", "1", "133968", "133741"], ["Whitelist", "1", "44", "24"]]),
            ("Black/White by ID Type", ["list_name", "id_type", "entries", "distinct_targets"],
             [["Blacklist", "2", "58211", "45527"]]),
            ("Black/White by Source", ["list_name", "source", "entries", "distinct_targets"],
             [["Blacklist", "Others", "104929", "95202"], ["Blacklist", "Anti-Fraud", "51422", "44298"]]),
            ("Black/White by Listed Reason", ["list_name", "listed_reason", "entries", "distinct_targets"],
             [["Blacklist", "Identity Theft", "14898", "14833"]]),
            ("Black/White by Scenario", ["list_name", "scenario", "entries", "distinct_targets"],
             [["Blacklist", "2", "192838", "161758"]]),
            ("Monthly Additions", ["added_month", "list_name", "entries"],
             [["2026-05", "Blacklist", "1200"], ["2026-05", "Whitelist", "3"]]),
            ("Greylist Detail", ["status", "id_type", "source", "listed_reason", "entries"],
             [["Active", "17", "Anti-Fraud", "Others", "4"]]),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "viz.html"
            write_visualization(output_path, report_title="Anti-fraud PH - Blacklist, Whitelist & Greylist",
                                snapshot_pt_date="2026-06-12", sheets=sheets, report_id=AF_LIST_USAGE_REPORT_ID)
            html = output_path.read_text(encoding="utf-8")
        self.assertIn("List Membership", html)
        self.assertIn("<h2>Entries by List</h2>", html)
        self.assertIn("Blacklist Entries by Source", html)
        self.assertIn('id="ec-monthly-additions" class="echart"', html)
        self.assertIn("Greylist Detail", html)


class _FakeSheetValuesCall:
    def __init__(self, payload):
        self.payload = payload

    def execute(self):
        return self.payload


class _FakeSheetValuesResource:
    def __init__(self, values_by_tab):
        self.values_by_tab = values_by_tab
        self.requested_ranges = []

    def batchGet(self, *, spreadsheetId, ranges, majorDimension):
        self.spreadsheet_id = spreadsheetId
        self.requested_ranges = list(ranges)
        value_ranges = []
        for sheet_range in ranges:
            tab_name = sheet_range.split("!", 1)[0].strip("'").replace("''", "'")
            value_ranges.append({"range": sheet_range, "values": self.values_by_tab[tab_name]})
        return _FakeSheetValuesCall({"spreadsheetId": spreadsheetId, "valueRanges": value_ranges})


class _FakeSheetsResource:
    def __init__(self, values_resource):
        self._values_resource = values_resource

    def values(self):
        return self._values_resource


class _FakeSheetsService:
    def __init__(self, values_by_tab):
        self.values_resource = _FakeSheetValuesResource(values_by_tab)

    def spreadsheets(self):
        return _FakeSheetsResource(self.values_resource)


class BusinessInsightsSheetRefreshTests(unittest.TestCase):
    def test_sheet_url_and_scheduled_tab_names_match_scheduler_convention(self):
        from bpmis_jira_tool.business_insights_sheet_refresh import (
            DEFAULT_BUSINESS_INSIGHTS_SHEET_URL,
            google_scopes_include_sheets,
            scheduled_sheet_name,
            spreadsheet_id_from_url,
        )

        self.assertEqual(
            spreadsheet_id_from_url(DEFAULT_BUSINESS_INSIGHTS_SHEET_URL),
            "1F5MSUwnxg8AbGr3rQN1l8nXYkxrBU680FJYhTGzL9qo",
        )
        self.assertEqual(
            scheduled_sheet_name(AF_SCENARIOS_ACTIONS_REPORT_ID, "Auth Outcome by Scene & Type"),
            "1_auth_outcome_by_scene_type",
        )
        self.assertEqual(
            scheduled_sheet_name(AF_RULE_EFFECTIVENESS_REPORT_ID, "Scene/Sub-scene/Action Usage"),
            "3_scene_sub_scene_action_usage",
        )
        self.assertEqual(
            scheduled_sheet_name(AF_CARD_3DS_REPORT_ID, "3DS by Merchant Category (MCC)"),
            "7_3ds_by_merchant_category_mcc",
        )
        self.assertTrue(google_scopes_include_sheets())

    def test_service_account_sheet_credentials_use_sheets_scope(self):
        from bpmis_jira_tool.business_insights_sheet_refresh import (
            GOOGLE_SHEETS_SCOPE,
            load_service_account_google_sheets_credentials,
        )

        with patch(
            "bpmis_jira_tool.business_insights_sheet_refresh.service_account.Credentials.from_service_account_info"
        ) as from_info:
            expected = object()
            from_info.return_value = expected
            credentials = load_service_account_google_sheets_credentials(
                service_account_json=json.dumps({"client_email": "refresh@example.iam.gserviceaccount.com"})
            )

        self.assertIs(credentials, expected)
        from_info.assert_called_once_with(
            {"client_email": "refresh@example.iam.gserviceaccount.com"},
            scopes=[GOOGLE_SHEETS_SCOPE],
        )

    def test_oauth_sheet_credentials_from_json_require_sheets_scope(self):
        from bpmis_jira_tool.business_insights_sheet_refresh import (
            GOOGLE_SHEETS_SCOPE,
            load_oauth_google_sheets_credentials,
        )

        credentials = load_oauth_google_sheets_credentials(
            json.dumps(
                {
                    "token": "access-token",
                    "refresh_token": "refresh-token",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "client_id": "client-id",
                    "client_secret": "client-secret",
                    "scopes": [GOOGLE_SHEETS_SCOPE],
                }
            )
        )

        self.assertEqual(credentials.refresh_token, "refresh-token")
        self.assertIn(GOOGLE_SHEETS_SCOPE, credentials.scopes)

    def test_adc_sheet_credentials_use_sheets_scope(self):
        from bpmis_jira_tool.business_insights_sheet_refresh import (
            GOOGLE_SHEETS_SCOPE,
            load_application_default_google_sheets_credentials,
        )

        expected = object()
        with patch("bpmis_jira_tool.business_insights_sheet_refresh.google.auth.default") as default:
            default.return_value = (expected, "project-id")
            credentials = load_application_default_google_sheets_credentials()

        self.assertIs(credentials, expected)
        default.assert_called_once_with(scopes=[GOOGLE_SHEETS_SCOPE])

    def test_refresh_from_google_sheet_writes_excel_visualization_and_metadata(self):
        from bpmis_jira_tool.business_insights_sheet_refresh import (
            refresh_anti_fraud_reports_from_google_sheet,
            scheduled_sheet_name,
        )

        title, builder = REPORT_BUILDERS[AF_RULES_FEATURES_REPORT_ID]
        sections = extract_sql_sections(builder(snapshot_pt_date=None, now=FIXED_NOW))
        values_by_tab = {}
        for section in sections:
            tab = scheduled_sheet_name(AF_RULES_FEATURES_REPORT_ID, section.sheet_name)
            if section.sheet_name == "Rules":
                values_by_tab[tab] = [
                    ["rule_id", "rule_name", "outcome_type", "rule_status"],
                    ["R1", "Velocity", "Reject", "Active"],
                ]
            elif section.sheet_name == "Features":
                values_by_tab[tab] = [
                    ["feature_id", "feature_name", "function_id", "feature_status"],
                    ["F1", "Login count", "FUNC_1", "Active"],
                ]
            elif section.sheet_name == "Function Usage":
                values_by_tab[tab] = [
                    ["function_id", "features", "active_features"],
                    ["FUNC_1", "1", "1"],
                ]
            else:
                values_by_tab[tab] = [["metric", "value"], [section.sheet_name, "1"]]

        with tempfile.TemporaryDirectory() as temp_dir:
            service = _FakeSheetsService(values_by_tab)
            result = refresh_anti_fraud_reports_from_google_sheet(
                portal_data_dir=Path(temp_dir),
                sheets_service=service,
                report_ids=[AF_RULES_FEATURES_REPORT_ID],
                now=FIXED_NOW,
            )
            metadata_path = Path(temp_dir) / "business_insights" / "reports.json"
            metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
            artifact = metadata["artifacts"][AF_RULES_FEATURES_REPORT_ID]
            workbook_path = Path(temp_dir) / "business_insights" / "artifacts" / artifact["filename"]
            visualization_path = Path(temp_dir) / "business_insights" / "artifacts" / artifact["visualization_filename"]
            workbook = load_workbook(workbook_path, data_only=True)
            visualization_html = visualization_path.read_text(encoding="utf-8")

        self.assertEqual(result["report_count"], 1)
        self.assertEqual(result["reports"][0]["report_id"], AF_RULES_FEATURES_REPORT_ID)
        self.assertEqual(artifact["source_filename"], "google-sheet-scheduled-output")
        self.assertEqual(artifact["source_google_tabs"][0]["google_tab"], "2_rules")
        self.assertIn("Rules", workbook.sheetnames)
        self.assertIn("Features", workbook.sheetnames)
        self.assertIn("Anti-fraud PH - Rules &amp; Features", visualization_html)


if __name__ == "__main__":
    unittest.main()
