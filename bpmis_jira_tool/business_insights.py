from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
import io
import json
import os
from pathlib import Path
import re
import threading
import time as time_module
from typing import Any
import uuid
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bpmis_jira_tool.errors import ToolError

BUSINESS_INSIGHTS_TIMEZONE = ZoneInfo("Asia/Singapore")
BUSINESS_INSIGHTS_DOMAINS: tuple[dict[str, str], ...] = (
    {"key": "anti-fraud", "label": "Anti-fraud"},
    {"key": "credit-risk", "label": "Credit Risk"},
    {"key": "ops-risk", "label": "Ops Risk"},
)
UNDERWRITING_FUNNEL_REPORT_ID = "credit-risk-ph-underwriting-funnel"
PORTFOLIO_REPAYMENT_REPORT_ID = "credit-risk-ph-portfolio-repayment"
LIMIT_UTILIZATION_REPORT_ID = "credit-risk-ph-limit-utilization"
APPLICATION_DISBURSEMENT_FUNNEL_REPORT_ID = "credit-risk-ph-application-disbursement-funnel"
UNDERWRITING_FUNNEL_TABLE = "ods.crms_ph_seabank_crms_db_underwriting_record_tab_ss_d"
REPAY_PLAN_TABLE = "ods.cbs_ph_bke_loan_core_db_repay_plan_tab_ss"
REPAY_DETAIL_TABLE = "ods.cbs_ph_bke_loan_core_db_repay_detail_tab_ss"
REPAY_FLOW_TABLE = "ods.cbs_ph_bke_loan_core_db_repay_flow_tab_ss"
CREDIT_LIMIT_TABLE = "ods.cbs_ph_bke_loan_core_db_credit_limit_tab_ss_d"
CREDIT_LIMIT_EOD_TABLE = "ods.cbs_ph_bke_loan_core_db_credit_limit_eod_tab_ss_d"
FROZEN_LIMIT_DETAIL_TABLE = "ods.cbs_ph_bke_loan_core_db_frozen_limit_detail_tab_ss_d"
FREEZE_LIMIT_FLOW_TABLE = "ods.cbs_ph_bke_loan_core_db_freeze_limit_flow_tab_ss_d"
LOAN_APPLICATION_TABLE = "ods.cbs_ph_bke_loan_txn_db_loan_application_tab_ss"
LOAN_APPLICATION_EXTRA_INFO_TABLE = "ods.cbs_ph_bke_loan_txn_db_loan_application_extra_info_tab_ss"
DISBURSE_FLOW_TABLE = "ods.cbs_ph_bke_loan_core_db_disburse_flow_tab_ss"
LOAN_ACCOUNT_TABLE = "ods.cbs_ph_bke_loan_core_db_loan_account_tab_ss"
UNDERWRITING_FUNNEL_FIELDS: tuple[str, ...] = (
    "underwriting_id",
    "underwriting_purpose",
    "product_code",
    "sub_product_code",
    "application_submission_time",
    "borrower_id",
    "apply_loan_amount",
    "apply_loan_tenor",
    "credit_score_partner",
    "underwriting_status",
    "current_stage",
    "step",
    "reject_reason",
    "create_date",
    "modify_date",
)
PRODUCT_LABELS: dict[str, str] = {
    "101": "SPL",
    "102": "BCL",
    "103": "SCL",
    "104": "Billease BNPL",
    "105": "Billease Cashloan",
    "106": "Juanhand",
    "107": "UDL",
    "108": "Employee Loan",
    "112": "Mabilis",
    "118": "Credit Card Shopee Checkout",
    "119": "Card Purchase",
    "120": "SPL 0%",
    "801": "SPL",
    "802": "BCL",
    "803": "SCL",
    "804": "Billease",
    "805": "Juanhand",
    "806": "UDL",
    "807": "Employee Loan",
    "809": "Mabilis",
    "812": "Credit Card",
    "812F": "Credit Card",
}
PRODUCT_LABEL_COLUMNS: set[str] = {"product", "product_code", "sub-product", "sub_product_code"}

# ---------------------------------------------------------------------------
# Anti-fraud (PH) report identifiers and lake (ODS) source tables.
# ---------------------------------------------------------------------------
# Table names are live-validated ODS tables available through Data Workbench.
# The AF lake uses two naming styles:
#   ods.mbs_ph_seabank_anti_fraud_db_<table>_ss        (scene / sub_scene / action)
#   ods.mbs_ph_seabank_anti_fraud_db_<table>_df        (scenario flow config)
#   ods.mbs_anti_fraud_<table>_ss                      (rule / feature config)
AF_SCENARIOS_ACTIONS_REPORT_ID = "anti-fraud-ph-scenarios-actions-auth-steps"
AF_RULES_FEATURES_REPORT_ID = "anti-fraud-ph-rules-features"
AF_RULE_EFFECTIVENESS_REPORT_ID = "anti-fraud-ph-rule-effectiveness"
AF_FRAUD_LOSS_REPORT_ID = "anti-fraud-ph-fraud-loss-cases"
AF_DETECTION_EFFECTIVENESS_REPORT_ID = "anti-fraud-ph-detection-effectiveness"
AF_RULE_CHANGE_LOG_REPORT_ID = "anti-fraud-ph-rule-change-log"
AF_FACIAL_VERIFICATION_REPORT_ID = "anti-fraud-ph-facial-verification"
AF_DEVICE_RISK_REPORT_ID = "anti-fraud-ph-device-identity-risk"
AF_CARD_3DS_REPORT_ID = "anti-fraud-ph-card-3ds"

AF_SCENE_TABLE = "ods.mbs_ph_seabank_anti_fraud_db_scene_tab_ss"
AF_SUB_SCENE_TABLE = "ods.mbs_ph_seabank_anti_fraud_db_sub_scene_tab_ss"
AF_ACTION_TABLE = "ods.mbs_ph_seabank_anti_fraud_db_action_tab_ss"
AF_SCENARIO_FLOW_CONFIG_TABLE = "ods.mbs_ph_seabank_anti_fraud_db_biz_scenario_flow_config_tab_df"
AF_RULE_CONFIG_TABLE = "ods.mbs_anti_fraud_rule_config_tab_ss"
AF_FEATURE_CONFIG_TABLE = "ods.mbs_anti_fraud_feature_config_tab_ss"
AF_IDENTIFY_REJECT_TABLE = "ods.mbs_anti_fraud_identify_reject_tab_ss"
AF_REQUEST_STATISTIC_TABLE = "ods.mbs_anti_fraud_request_statistic_tab_ss"
AF_RULE_TRIGGER_STATISTIC_TABLE = "ods.mbs_anti_fraud_rule_trigger_statistic_tab_ss"
AF_PUNISH_LIST_TABLE = "ods.mbs_anti_fraud_punish_list_tab_ss"
# Feature-mart DWD layer: clean daily-incremental hit/action logs with scene names built in.
AF_RULE_HIT_LOG_TABLE = "fmart_antifraud.dwd_antifraud_rule_hit_log_di"
AF_ACTION_LOG_TABLE = "fmart_antifraud.dwd_antifraud_action_log_di"
AF_REVIEW_CASE_TABLE = "fmart_antifraud.dwd_antifraud_review_case_df"
AF_REVIEW_RECORD_TABLE = "fmart_antifraud.dwd_antifraud_review_record_df"
# Facial verification / liveness / deepfake authentication checks (cumulative daily _df snapshot).
AF_FACIAL_VERIFICATION_TABLE = "fmart_antifraud.dwd_authentication_facial_verification_df"
# Extended device/identity action log (daily-incremental, event-level; pt_date = event day).
# is_* flags are clear text ('1'/''); precomputed distinct_uid_num_link_* are encrypted/constant, so
# account-farming is recomputed via count(distinct uid) per device/scene. Huge table -> window by pt_date.
AF_ACTION_LOG_EXT_TABLE = "fmart_antifraud.dwd_antifraud_action_log_ext_di"
# Card 3DS (ACS) authentication + risk-based-auth log, and card fraud cases. pt_date = event day.
AF_THREEDS_TRANS_TABLE = "ods.pmt_threeds_acs_t_threeds_trans_ss_d"
AF_THREEDS_RBA_LOG_TABLE = "ods.pmt_threeds_acs_t_threeds_trans_rba_log_ss_d"
AF_CARD_FRAUD_CASE_TABLE = "dm.mbs_card_fraud_case_ss_d"

SEEDED_REPORTS: tuple[dict[str, str], ...] = (
    {
        "id": UNDERWRITING_FUNNEL_REPORT_ID,
        "domain": "credit-risk",
        "name": "Credit Risk PH - Underwriting Funnel",
        "type": "underwriting_funnel",
        "status": "generator_ready",
    },
    {
        "id": PORTFOLIO_REPAYMENT_REPORT_ID,
        "domain": "credit-risk",
        "name": "Credit Risk PH - Portfolio Repayment",
        "type": "portfolio_repayment",
        "status": "generator_ready",
    },
    {
        "id": LIMIT_UTILIZATION_REPORT_ID,
        "domain": "credit-risk",
        "name": "Credit Risk PH - Limit Utilization",
        "type": "limit_utilization",
        "status": "generator_ready",
    },
    {
        "id": APPLICATION_DISBURSEMENT_FUNNEL_REPORT_ID,
        "domain": "credit-risk",
        "name": "Credit Risk PH - Application to Disbursement Funnel",
        "type": "application_disbursement_funnel",
        "status": "generator_ready",
    },
    {
        "id": AF_SCENARIOS_ACTIONS_REPORT_ID,
        "domain": "anti-fraud",
        "name": "Anti-fraud PH - L1+L2 Scenarios, Actions & Auth Steps",
        "type": "af_scenarios_actions",
        "status": "generator_ready",
    },
    {
        "id": AF_RULES_FEATURES_REPORT_ID,
        "domain": "anti-fraud",
        "name": "Anti-fraud PH - Rules & Features",
        "type": "af_rules_features",
        "status": "generator_ready",
    },
    {
        "id": AF_RULE_EFFECTIVENESS_REPORT_ID,
        "domain": "anti-fraud",
        "name": "Anti-fraud PH - Rule Effectiveness / Hit-Rate",
        "type": "af_rule_effectiveness",
        "status": "generator_ready",
    },
    {
        "id": AF_FRAUD_LOSS_REPORT_ID,
        "domain": "anti-fraud",
        "name": "Anti-fraud PH - Fraud Loss & Case Outcomes",
        "type": "af_fraud_loss",
        "status": "generator_ready",
    },
    {
        "id": AF_FACIAL_VERIFICATION_REPORT_ID,
        "domain": "anti-fraud",
        "name": "Anti-fraud PH - Facial Verification / Liveness & Deepfake",
        "type": "af_facial_verification",
        "status": "generator_ready",
    },
    {
        "id": AF_DEVICE_RISK_REPORT_ID,
        "domain": "anti-fraud",
        "name": "Anti-fraud PH - Device & Identity Risk",
        "type": "af_device_risk",
        "status": "generator_ready",
    },
    {
        "id": AF_CARD_3DS_REPORT_ID,
        "domain": "anti-fraud",
        "name": "Anti-fraud PH - Card Fraud & 3DS Authentication",
        "type": "af_card_3ds",
        "status": "generator_ready",
    },
)

# Reports whose artifacts can be regenerated on demand by running the live
# generation CLI (`scripts/generate_business_insights_live_reports.py`) against
# Data Workbench. Keep in sync with that script's REPORT_BUILDERS. Excludes the
# underwriting funnel (ingested via manual Excel upload) and any seeded report
# that has no SQL generator yet.
GENERATOR_REPORT_IDS: frozenset[str] = frozenset(
    {
        PORTFOLIO_REPAYMENT_REPORT_ID,
        LIMIT_UTILIZATION_REPORT_ID,
        APPLICATION_DISBURSEMENT_FUNNEL_REPORT_ID,
        AF_SCENARIOS_ACTIONS_REPORT_ID,
        AF_RULES_FEATURES_REPORT_ID,
        AF_RULE_EFFECTIVENESS_REPORT_ID,
        AF_FRAUD_LOSS_REPORT_ID,
        AF_FACIAL_VERIFICATION_REPORT_ID,
        AF_DEVICE_RISK_REPORT_ID,
        AF_CARD_3DS_REPORT_ID,
    }
)


@dataclass(frozen=True)
class BusinessInsightsPeriod:
    previous_month_start: date
    current_month_start: date
    end_exclusive: date

    @property
    def start_date(self) -> date:
        return self.previous_month_start

    @property
    def current_date(self) -> date:
        return self.end_exclusive - timedelta(days=1)

    @property
    def previous_month_label(self) -> str:
        return self.previous_month_start.strftime("%b %Y")

    @property
    def current_month_label(self) -> str:
        return f"{self.current_month_start.strftime('%b %Y')} MTD"

    @property
    def title(self) -> str:
        return f"{self.previous_month_label} + {self.current_month_label}"


def business_insights_period(now: datetime | None = None) -> BusinessInsightsPeriod:
    active_now = now.astimezone(BUSINESS_INSIGHTS_TIMEZONE) if now else datetime.now(BUSINESS_INSIGHTS_TIMEZONE)
    current_month_start = active_now.date().replace(day=1)
    previous_month_last_day = current_month_start - timedelta(days=1)
    previous_month_start = previous_month_last_day.replace(day=1)
    return BusinessInsightsPeriod(
        previous_month_start=previous_month_start,
        current_month_start=current_month_start,
        end_exclusive=active_now.date() + timedelta(days=1),
    )


@dataclass(frozen=True)
class ReportWindow:
    """The last two full calendar months plus the current month-to-date."""

    periods: tuple[tuple[str, date, date], ...]  # (label, start, end_exclusive) per period
    span_start: date
    span_end_exclusive: date
    span_label: str


def business_insights_window(now: datetime | None = None) -> ReportWindow:
    period = business_insights_period(now)
    current_month_start = period.current_month_start
    previous_month_start = period.previous_month_start
    prev2_month_start = (previous_month_start - timedelta(days=1)).replace(day=1)
    end_exclusive = period.end_exclusive
    periods = (
        (prev2_month_start.strftime("%b %Y"), prev2_month_start, previous_month_start),
        (previous_month_start.strftime("%b %Y"), previous_month_start, current_month_start),
        (current_month_start.strftime("%b %Y") + " MTD", current_month_start, end_exclusive),
    )
    return ReportWindow(
        periods=periods,
        span_start=prev2_month_start,
        span_end_exclusive=end_exclusive,
        span_label=f"{prev2_month_start.strftime('%b %Y')} – {current_month_start.strftime('%b %Y')} MTD",
    )


def _date_to_epoch_millis(value: date) -> int:
    instant = datetime.combine(value, time.min, tzinfo=BUSINESS_INSIGHTS_TIMEZONE)
    return int(instant.timestamp() * 1000)


def build_underwriting_funnel_sql(now: datetime | None = None) -> str:
    period = business_insights_period(now)
    start_ms = _date_to_epoch_millis(period.start_date)
    current_month_start_ms = _date_to_epoch_millis(period.current_month_start)
    end_ms = _date_to_epoch_millis(period.end_exclusive)
    columns = ",\n    ".join(UNDERWRITING_FUNNEL_FIELDS)
    return f"""-- Credit Risk PH - Underwriting Funnel
-- Duration: {period.title}
-- Run this SQL in Data Workbench, download the result, then upload it back to Business Insights.
select
    case
        when application_submission_time >= {current_month_start_ms} then '{period.current_month_label}'
        else '{period.previous_month_label}'
    end as report_period,
    {columns}
from {UNDERWRITING_FUNNEL_TABLE}
where application_submission_time >= {start_ms}
  and application_submission_time < {end_ms}
;"""


def build_underwriting_funnel_mis_sql(*, snapshot_pt_date: str | None = None, now: datetime | None = None) -> str:
    period = business_insights_period(now)
    start_ms = _date_to_epoch_millis(period.start_date)
    current_month_start_ms = _date_to_epoch_millis(period.current_month_start)
    end_ms = _date_to_epoch_millis(period.end_exclusive)
    active_now = now.astimezone(BUSINESS_INSIGHTS_TIMEZONE) if now else datetime.now(BUSINESS_INSIGHTS_TIMEZONE)
    active_now_ms = int(active_now.timestamp() * 1000)
    snapshot_filter = (
        f"pt_date = '{snapshot_pt_date}'"
        if snapshot_pt_date
        else f"pt_date = (select max(pt_date) from {UNDERWRITING_FUNNEL_TABLE})"
    )
    period_expr = (
        f"case when application_submission_time >= {current_month_start_ms} "
        f"then '{period.current_month_label}' else '{period.previous_month_label}' end"
    )
    product_expr = "coalesce(nullif(product_code, ''), 'UNKNOWN')"
    subproduct_expr = "coalesce(nullif(sub_product_code, ''), '-')"
    status_bucket_expr = """case
  when upper(coalesce(underwriting_status, '')) rlike '(APPROV|PASS|SUCCESS|COMPLETED)' then 'APPROVED'
  when upper(coalesce(underwriting_status, '')) rlike '(REJECT|DECLIN|FAIL|DENY)' then 'REJECTED'
  else 'PENDING'
end"""
    base_where = (
        f"{snapshot_filter}\n"
        f"  and application_submission_time >= {start_ms}\n"
        f"  and application_submission_time < {end_ms}"
    )
    return f"""-- Credit Risk PH - Underwriting Funnel MIS
-- Duration: {period.title}
-- Snapshot: {snapshot_pt_date or "latest available pt_date at run time"}
-- Run each query in Data Workbench. These are the aggregation queries used to build the Portal Excel.

-- 1. Summary by Product
select {period_expr} as period, {product_expr} as product,
  count(1) as applications,
  sum(case when {status_bucket_expr} = 'APPROVED' then 1 else 0 end) as approved,
  sum(case when {status_bucket_expr} = 'REJECTED' then 1 else 0 end) as rejected,
  sum(case when {status_bucket_expr} = 'PENDING' then 1 else 0 end) as pending,
  round(avg(cast(apply_loan_amount as double)), 2) as avg_applied_amount
from {UNDERWRITING_FUNNEL_TABLE}
where {base_where}
group by {period_expr}, {product_expr}
order by period, product;

-- 2. Product Funnel
select {period_expr} as period, {product_expr} as product,
  coalesce(nullif(underwriting_status, ''), 'PENDING') as status,
  count(1) as count
from {UNDERWRITING_FUNNEL_TABLE}
where {base_where}
group by {period_expr}, {product_expr}, coalesce(nullif(underwriting_status, ''), 'PENDING')
order by period, product, status;

-- 3. Product Reject Reasons
select {period_expr} as period, {product_expr} as product,
  coalesce(nullif(reject_reason, ''), 'Unspecified') as reject_reason,
  count(1) as count
from {UNDERWRITING_FUNNEL_TABLE}
where {base_where}
  and {status_bucket_expr} = 'REJECTED'
group by {period_expr}, {product_expr}, coalesce(nullif(reject_reason, ''), 'Unspecified')
order by period, product, count desc;

-- 4. Product Stage Backlog
select {period_expr} as period, {product_expr} as product,
  coalesce(nullif(current_stage, ''), 'Unspecified') as current_stage,
  count(1) as count,
  min(create_date) as oldest_create_date_ms,
  round(avg(case when create_date is not null then ({active_now_ms} - cast(create_date as double)) / 86400000.0 else null end), 2) as avg_age_days
from {UNDERWRITING_FUNNEL_TABLE}
where {base_where}
group by {period_expr}, {product_expr}, coalesce(nullif(current_stage, ''), 'Unspecified')
order by period, product, count desc;

-- 5. Sub-product Funnel
select {period_expr} as period, {product_expr} as product, {subproduct_expr} as sub_product,
  coalesce(nullif(underwriting_status, ''), 'PENDING') as status,
  count(1) as count
from {UNDERWRITING_FUNNEL_TABLE}
where {base_where}
group by {period_expr}, {product_expr}, {subproduct_expr}, coalesce(nullif(underwriting_status, ''), 'PENDING')
order by period, product, sub_product, status;
"""


def _snapshot_filter(table: str, snapshot_pt_date: str | None) -> str:
    if snapshot_pt_date:
        return f"pt_date = '{snapshot_pt_date}'"
    return f"pt_date = (select max(pt_date) from {table})"


def _yyyymmdd(value: date) -> str:
    return value.strftime("%Y%m%d")


def _date_key_expr(column_name: str) -> str:
    return f"regexp_replace(cast({column_name} as string), '-', '')"


def build_portfolio_repayment_sql(*, snapshot_pt_date: str | None = None, now: datetime | None = None) -> str:
    period = business_insights_period(now)
    start_key = _yyyymmdd(period.start_date)
    current_key = _yyyymmdd(period.current_month_start)
    end_key = _yyyymmdd(period.end_exclusive)
    plan_snapshot = _snapshot_filter(REPAY_PLAN_TABLE, snapshot_pt_date)
    flow_snapshot = _snapshot_filter(REPAY_FLOW_TABLE, snapshot_pt_date)
    detail_snapshot = _snapshot_filter(REPAY_DETAIL_TABLE, snapshot_pt_date)
    repay_date_key = _date_key_expr("repay_date")
    period_expr = f"case when {repay_date_key} >= '{current_key}' then '{period.current_month_label}' else '{period.previous_month_label}' end"
    plan_period_where = f"{repay_date_key} >= '{start_key}' and {repay_date_key} < '{end_key}'"
    flow_period_where = f"{repay_date_key} >= '{start_key}' and {repay_date_key} < '{end_key}'"
    detail_period_where = f"{repay_date_key} >= '{start_key}' and {repay_date_key} < '{end_key}'"
    dpd_bucket = """case
  when coalesce(eod_overdue_day, overdue_day, 0) <= 0 then 'Current'
  when coalesce(eod_overdue_day, overdue_day, 0) between 1 and 7 then 'DPD 1-7'
  when coalesce(eod_overdue_day, overdue_day, 0) between 8 and 30 then 'DPD 8-30'
  when coalesce(eod_overdue_day, overdue_day, 0) between 31 and 60 then 'DPD 31-60'
  when coalesce(eod_overdue_day, overdue_day, 0) between 61 and 90 then 'DPD 61-90'
  else 'DPD 90+'
end"""
    return f"""-- Credit Risk PH - Portfolio Repayment
-- Duration: {period.title}
-- Snapshot: {snapshot_pt_date or "latest available pt_date at run time"}

-- 1. Summary by Product
select {period_expr} as period,
  coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
  count(distinct loan_no) as loans_due,
  count(1) as plan_rows,
  sum(cast(total_amt as double)) as due_amount,
  sum(cast(repaid_amt as double)) as repaid_amount,
  sum(cast(total_amt as double) - cast(repaid_amt as double)) as outstanding_amount,
  round(sum(cast(repaid_amt as double)) / nullif(sum(cast(total_amt as double)), 0), 4) as repayment_rate,
  sum(case when coalesce(eod_overdue_day, overdue_day, 0) > 0 then 1 else 0 end) as overdue_plan_rows,
  max(coalesce(eod_overdue_day, overdue_day, 0)) as max_dpd
from {REPAY_PLAN_TABLE}
where {plan_snapshot}
  and {plan_period_where}
group by {period_expr}, coalesce(nullif(prod_type, ''), 'UNKNOWN')
order by period, product;

-- 2. DPD Buckets
select {period_expr} as period,
  coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
  {dpd_bucket} as dpd_bucket,
  count(distinct loan_no) as loans,
  sum(cast(total_amt as double)) as due_amount,
  sum(cast(repaid_amt as double)) as repaid_amount,
  sum(cast(total_amt as double) - cast(repaid_amt as double)) as outstanding_amount
from {REPAY_PLAN_TABLE}
where {plan_snapshot}
  and {plan_period_where}
group by {period_expr}, coalesce(nullif(prod_type, ''), 'UNKNOWN'), {dpd_bucket}
order by period, product, dpd_bucket;

-- 3. Repay Flow Status
select {period_expr} as period,
  coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
  coalesce(nullif(repay_status, ''), 'UNKNOWN') as repay_status,
  count(1) as flow_count,
  sum(cast(repay_amt as double)) as repay_amount
from {REPAY_FLOW_TABLE}
where {flow_snapshot}
  and {flow_period_where}
group by {period_expr},
  coalesce(nullif(prod_type, ''), 'UNKNOWN'), coalesce(nullif(repay_status, ''), 'UNKNOWN')
order by period, product, repay_status;

-- 4. Repay Detail by Type
select {period_expr} as period,
  coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
  coalesce(nullif(repay_type, ''), 'UNKNOWN') as repay_type,
  coalesce(nullif(source_of_repay, ''), 'UNKNOWN') as source_of_repay,
  count(1) as detail_count,
  sum(cast(repay_amt as double)) as repay_amount,
  sum(cast(repay_principal as double)) as repay_principal,
  sum(cast(repay_interest as double)) as repay_interest
from {REPAY_DETAIL_TABLE}
where {detail_snapshot}
  and {detail_period_where}
group by {period_expr},
  coalesce(nullif(prod_type, ''), 'UNKNOWN'), coalesce(nullif(repay_type, ''), 'UNKNOWN'), coalesce(nullif(source_of_repay, ''), 'UNKNOWN')
order by period, product, repay_type, source_of_repay;
"""


def build_limit_utilization_sql(*, snapshot_pt_date: str | None = None, now: datetime | None = None) -> str:
    period = business_insights_period(now)
    limit_snapshot = _snapshot_filter(CREDIT_LIMIT_TABLE, snapshot_pt_date)
    eod_snapshot = _snapshot_filter(CREDIT_LIMIT_EOD_TABLE, snapshot_pt_date)
    frozen_snapshot = _snapshot_filter(FROZEN_LIMIT_DETAIL_TABLE, snapshot_pt_date)
    freeze_flow_snapshot = _snapshot_filter(FREEZE_LIMIT_FLOW_TABLE, snapshot_pt_date)
    utilization_expr = "cast(used_limit as double) / nullif(cast(total_limit as double), 0)"
    utilization_bucket = f"""case
  when total_limit is null or cast(total_limit as double) = 0 then 'No Limit'
  when {utilization_expr} < 0.25 then '0-25%'
  when {utilization_expr} < 0.50 then '25-50%'
  when {utilization_expr} < 0.75 then '50-75%'
  when {utilization_expr} < 0.90 then '75-90%'
  else '90%+'
end"""
    return f"""-- Credit Risk PH - Limit Utilization
-- Duration: point-in-time snapshot for {period.current_date.isoformat()}
-- Snapshot: {snapshot_pt_date or "latest available pt_date at run time"}

-- 1. Summary by Product
select coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
  count(distinct client_no) as customers,
  count(1) as limit_rows,
  sum(cast(total_limit as double)) as total_limit,
  sum(cast(used_limit as double)) as used_limit,
  sum(cast(frozen_limit as double)) as frozen_limit,
  case
    when sum(cast(total_limit as double)) > 0
    then sum(cast(total_limit as double) - cast(used_limit as double) - cast(frozen_limit as double))
    else null
  end as available_limit_estimate,
  round(sum(cast(used_limit as double)) / nullif(sum(cast(total_limit as double)), 0), 4) as utilization_rate
from {CREDIT_LIMIT_TABLE}
where {limit_snapshot}
group by coalesce(nullif(prod_type, ''), 'UNKNOWN')
order by product;

-- 2. Utilization Buckets
select coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
  {utilization_bucket} as utilization_bucket,
  count(distinct client_no) as customers,
  sum(cast(total_limit as double)) as total_limit,
  sum(cast(used_limit as double)) as used_limit,
  sum(cast(frozen_limit as double)) as frozen_limit
from {CREDIT_LIMIT_TABLE}
where {limit_snapshot}
group by coalesce(nullif(prod_type, ''), 'UNKNOWN'), {utilization_bucket}
order by product, utilization_bucket;

-- 3. EOD Available Limit
select coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
  coalesce(nullif(status, ''), 'UNKNOWN') as status,
  count(distinct client_no) as customers,
  sum(cast(available_limit as double)) as available_limit,
  sum(cast(earmark_limit as double)) as earmark_limit
from {CREDIT_LIMIT_EOD_TABLE}
where {eod_snapshot}
group by coalesce(nullif(prod_type, ''), 'UNKNOWN'), coalesce(nullif(status, ''), 'UNKNOWN')
order by product, status;

-- 4. Frozen Limit Detail
select coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
  coalesce(nullif(frozen_type, ''), 'UNKNOWN') as frozen_type,
  coalesce(nullif(status, ''), 'UNKNOWN') as status,
  count(1) as frozen_rows,
  sum(cast(current_limit as double)) as frozen_amount
from {FROZEN_LIMIT_DETAIL_TABLE}
where {frozen_snapshot}
group by coalesce(nullif(prod_type, ''), 'UNKNOWN'), coalesce(nullif(frozen_type, ''), 'UNKNOWN'), coalesce(nullif(status, ''), 'UNKNOWN')
order by product, frozen_type, status;

-- 5. Freeze Limit Flow
select coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
  coalesce(nullif(status, ''), 'UNKNOWN') as status,
  count(1) as flow_count,
  sum(cast(amount as double)) as freeze_amount
from {FREEZE_LIMIT_FLOW_TABLE}
where {freeze_flow_snapshot}
group by coalesce(nullif(prod_type, ''), 'UNKNOWN'), coalesce(nullif(status, ''), 'UNKNOWN')
order by product, status;
"""


def build_application_disbursement_funnel_sql(*, snapshot_pt_date: str | None = None, now: datetime | None = None) -> str:
    period = business_insights_period(now)
    start_key = _yyyymmdd(period.start_date)
    current_key = _yyyymmdd(period.current_month_start)
    end_key = _yyyymmdd(period.end_exclusive)
    start_ms = _date_to_epoch_millis(period.start_date)
    end_ms = _date_to_epoch_millis(period.end_exclusive)
    app_snapshot = _snapshot_filter(LOAN_APPLICATION_TABLE, snapshot_pt_date)
    uw_snapshot = _snapshot_filter(UNDERWRITING_FUNNEL_TABLE, snapshot_pt_date)
    disburse_snapshot = _snapshot_filter(DISBURSE_FLOW_TABLE, snapshot_pt_date)
    account_snapshot = _snapshot_filter(LOAN_ACCOUNT_TABLE, snapshot_pt_date)
    apply_date_key = _date_key_expr("apply_date")
    disburse_date_key = _date_key_expr("disburse_date")
    app_period = f"case when {apply_date_key} >= '{current_key}' then '{period.current_month_label}' else '{period.previous_month_label}' end"
    disburse_period = f"case when {disburse_date_key} >= '{current_key}' then '{period.current_month_label}' else '{period.previous_month_label}' end"
    app_where = f"{app_snapshot}\n  and {apply_date_key} >= '{start_key}'\n  and {apply_date_key} < '{end_key}'"
    uw_where = f"{uw_snapshot}\n  and application_submission_time >= {start_ms}\n  and application_submission_time < {end_ms}"
    disburse_where = f"{disburse_snapshot}\n  and {disburse_date_key} >= '{start_key}'\n  and {disburse_date_key} < '{end_key}'"
    return f"""-- Credit Risk PH - Application to Disbursement Funnel
-- Duration: {period.title}
-- Snapshot: {snapshot_pt_date or "latest available pt_date at run time"}

-- 1. Funnel Summary by Product
with app as (
  select {app_period} as period,
    coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
    count(distinct loan_apply_no) as applications,
    sum(cast(principal as double)) as applied_principal,
    sum(case when account_no is not null and account_no <> '' then 1 else 0 end) as applications_with_account,
    sum(case when loan_no is not null and loan_no <> '' then 1 else 0 end) as applications_with_loan_no
  from {LOAN_APPLICATION_TABLE}
  where {app_where}
  group by {app_period}, coalesce(nullif(prod_type, ''), 'UNKNOWN')
),
uw as (
  select case when application_submission_time >= {_date_to_epoch_millis(period.current_month_start)} then '{period.current_month_label}' else '{period.previous_month_label}' end as period,
    coalesce(nullif(product_code, ''), 'UNKNOWN') as product,
    count(distinct underwriting_id) as underwriting_cases,
    sum(case when upper(coalesce(underwriting_status, '')) rlike '(APPROV|PASS|SUCCESS|COMPLETED)' then 1 else 0 end) as approved_cases,
    sum(case when upper(coalesce(underwriting_status, '')) rlike '(REJECT|DECLIN|FAIL|DENY)' then 1 else 0 end) as rejected_cases
  from {UNDERWRITING_FUNNEL_TABLE}
  where {uw_where}
  group by case when application_submission_time >= {_date_to_epoch_millis(period.current_month_start)} then '{period.current_month_label}' else '{period.previous_month_label}' end,
    coalesce(nullif(product_code, ''), 'UNKNOWN')
),
disb as (
  select {disburse_period} as period,
    coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
    count(distinct loan_no) as disbursed_loans,
    sum(cast(principal as double)) as disbursed_principal
  from {DISBURSE_FLOW_TABLE}
  where {disburse_where}
    and upper(coalesce(disburse_status, '')) not rlike '(FAIL|REJECT|CANCEL|REVERS)'
  group by {disburse_period},
    coalesce(nullif(prod_type, ''), 'UNKNOWN')
)
select coalesce(app.period, uw.period, disb.period) as period,
  coalesce(app.product, uw.product, disb.product) as product,
  coalesce(app.applications, 0) as applications,
  coalesce(uw.underwriting_cases, 0) as underwriting_cases,
  coalesce(uw.approved_cases, 0) as approved_cases,
  coalesce(uw.rejected_cases, 0) as rejected_cases,
  coalesce(disb.disbursed_loans, 0) as disbursed_loans,
  coalesce(app.applied_principal, 0) as applied_principal,
  coalesce(disb.disbursed_principal, 0) as disbursed_principal,
  round(coalesce(disb.disbursed_loans, 0) / nullif(coalesce(app.applications, 0), 0), 4) as application_to_disbursement_rate
from app
full outer join uw on app.period = uw.period and app.product = uw.product
full outer join disb on coalesce(app.period, uw.period) = disb.period and coalesce(app.product, uw.product) = disb.product
order by period, product;

-- 2. Loan Application Status
select {app_period} as period,
  coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
  coalesce(nullif(apply_status, ''), 'UNKNOWN') as apply_status,
  count(distinct loan_apply_no) as applications,
  sum(cast(principal as double)) as applied_principal
from {LOAN_APPLICATION_TABLE}
where {app_where}
group by {app_period}, coalesce(nullif(prod_type, ''), 'UNKNOWN'), coalesce(nullif(apply_status, ''), 'UNKNOWN')
order by period, product, apply_status;

-- 3. Disbursement Status
select {disburse_period} as period,
  coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
  coalesce(nullif(disburse_status, ''), 'UNKNOWN') as disburse_status,
  count(1) as disburse_flows,
  count(distinct loan_no) as loans,
  sum(cast(principal as double)) as principal
from {DISBURSE_FLOW_TABLE}
where {disburse_where}
group by {disburse_period},
  coalesce(nullif(prod_type, ''), 'UNKNOWN'), coalesce(nullif(disburse_status, ''), 'UNKNOWN')
order by period, product, disburse_status;

-- 4. Account Opening Snapshot
select coalesce(nullif(prod_type, ''), 'UNKNOWN') as product,
  coalesce(nullif(status, ''), 'UNKNOWN') as account_status,
  coalesce(nullif(account_credit_quality, ''), 'UNKNOWN') as account_credit_quality,
  count(distinct account_no) as accounts
from {LOAN_ACCOUNT_TABLE}
where {account_snapshot}
group by coalesce(nullif(prod_type, ''), 'UNKNOWN'), coalesce(nullif(status, ''), 'UNKNOWN'), coalesce(nullif(account_credit_quality, ''), 'UNKNOWN')
order by product, account_status, account_credit_quality;
"""


def _aliased_snapshot_filter(alias: str, table: str, snapshot_pt_date: str | None) -> str:
    if snapshot_pt_date:
        return f"{alias}.pt_date = '{snapshot_pt_date}'"
    return f"{alias}.pt_date = (select max(pt_date) from {table})"


def _af_report_header(title: str, snapshot_pt_date: str | None, period: BusinessInsightsPeriod | None = None) -> str:
    lines = [
        f"-- {title}",
        f"-- Snapshot: {snapshot_pt_date or 'latest available pt_date at run time'}",
    ]
    if period is not None:
        lines.append(f"-- Duration: {period.title}")
    lines.append("-- Lake (ODS) tables from the RPMAF01 grant (table_access_20260605).")
    lines.append("-- Run each numbered section in Data Workbench (SparkSQL); each becomes one Excel sheet.")
    return "\n".join(lines)


def build_af_scenarios_actions_sql(*, snapshot_pt_date: str | None = None, now: datetime | None = None) -> str:
    scene_snap = _aliased_snapshot_filter("s", AF_SCENE_TABLE, snapshot_pt_date)
    sub_scene_snap = _aliased_snapshot_filter("ss", AF_SUB_SCENE_TABLE, snapshot_pt_date)
    action_snap = _aliased_snapshot_filter("a", AF_ACTION_TABLE, snapshot_pt_date)
    flow_snap = _aliased_snapshot_filter("f", AF_SCENARIO_FLOW_CONFIG_TABLE, snapshot_pt_date)
    # Authentication funnel sections use the DWD action log (daily-incremental), scoped by pt_date over
    # the report window. risk_result: '1' = pass, '0' = reject, '' = not risk-evaluated.
    win = business_insights_window(now)
    span_label = win.span_label
    al_start = win.span_start.isoformat()
    al_end_incl = (win.span_end_exclusive - timedelta(days=1)).isoformat()
    p2_iso, p3_iso = win.periods[1][1].isoformat(), win.periods[2][1].isoformat()
    p1_label, p2_label, p3_label = (p[0] for p in win.periods)
    al_window = f"al.pt_date between '{al_start}' and '{al_end_incl}'"
    al_period = (
        f"case when al.pt_date < '{p2_iso}' then '{p1_label}' "
        f"when al.pt_date < '{p3_iso}' then '{p2_label}' else '{p3_label}' end"
    )
    al_pass = "al.risk_result = '1'"
    al_reject = "al.risk_result = '0'"
    al_evaluated = "al.risk_result in ('0', '1')"
    header = _af_report_header("Anti-fraud PH - L1+L2 Scenarios, Actions & Auth Steps", snapshot_pt_date)
    return f"""{header}
-- Scenario flow config provides the live scene/sub-scene/action/auth-step mapping. Sections 5-8 add the
-- live authentication funnel / customer-friction view from the DWD action log over {span_label}.

-- 1. L1 Scenarios
select
  s.code as l1_scene_code,
  s.name as l1_scene_name,
  s.enum_name as l1_enum_name,
  case s.scene_type when 0 then 'realScene' when 1 then 'logicScene' else concat('scene_type_', cast(s.scene_type as string)) end as l1_scene_type,
  case s.business_category when 0 then 'Retail' when 1 then 'Corporate' else cast(s.business_category as string) end as business_category,
  s.mode,
  s.source,
  s.real_scene_sub_scene,
  s.description
from {AF_SCENE_TABLE} s
where {scene_snap}
order by s.code;

-- 2. L2 Sub-Scenarios
select
  ss.code as l2_sub_scene_code,
  ss.name as l2_sub_scene_name,
  ss.enum_name as l2_enum_name
from {AF_SUB_SCENE_TABLE} ss
where {sub_scene_snap}
order by ss.code;

-- 3. Actions and Auth Steps
select
  a.code as action_code,
  a.name as action_name,
  a.enum_name as action_enum_name,
  case a.type when 1 then 'Business' when 2 then 'Authentication (Auth Step)' else concat('type_', cast(a.type as string)) end as action_type
from {AF_ACTION_TABLE} a
where {action_snap}
order by a.type, a.code;

-- 4. Scenario Action Auth Flow
select
  s.name as l1_scene_name,
  s.enum_name as l1_enum_name,
  ss.name as l2_sub_scene_name,
  ss.enum_name as l2_enum_name,
  a.name as action_name,
  a.enum_name as action_enum_name,
  f.default_step,
  f.challenge1_step,
  f.challenge2_step,
  f.challenge3_step,
  f.challenge4_step,
  f.challenge5_step
-- The flow-config table stores scene/sub_scene/action by name (e.g. 'ActivateFaceID'),
-- not by numeric code, so the dimension tables are joined on name.
from {AF_SCENARIO_FLOW_CONFIG_TABLE} f
left join {AF_SCENE_TABLE} s on s.name = f.scene and {scene_snap}
left join {AF_SUB_SCENE_TABLE} ss on ss.name = f.sub_scene and {sub_scene_snap}
left join {AF_ACTION_TABLE} a on a.name = f.action and {action_snap}
where {flow_snap}
order by f.scene, f.sub_scene, f.action;

-- 5. Authentication Outcome Summary
-- One row per period (last two full months + current MTD) from the action log. pass/reject use the risk
-- decision (risk_result); not_evaluated = actions with no risk decision. flows = distinct bizflow_instance.
select
  {al_period} as period,
  count(1) as actions,
  count(distinct al.uid) as distinct_users,
  count(distinct al.bizflow_instance_id) as flows,
  sum(case when {al_pass} then 1 else 0 end) as pass_actions,
  sum(case when {al_reject} then 1 else 0 end) as reject_actions,
  sum(case when al.risk_result is null or al.risk_result not in ('0', '1') then 1 else 0 end) as not_evaluated_actions,
  round(sum(case when {al_reject} then 1 else 0 end)
        / nullif(sum(case when {al_evaluated} then 1 else 0 end), 0) * 100, 3) as reject_rate_pct
from {AF_ACTION_LOG_TABLE} al
where {al_window}
group by {al_period}
order by min(al.pt_date);

-- 6. Auth Outcome by Scene & Type
-- Outcome funnel by scene and authentication type (DEFAULT = no step-up; CHALLENGE_1/2/3 = step-up
-- tiers). Limited to risk-evaluated actions so the reject rate is clean.
select
  coalesce(nullif(trim(al.scene_name), ''), 'Unspecified') as scene_name,
  coalesce(nullif(trim(al.authentication_type), ''), 'Unspecified') as authentication_type,
  count(1) as actions,
  count(distinct al.uid) as distinct_users,
  sum(case when {al_pass} then 1 else 0 end) as pass_actions,
  sum(case when {al_reject} then 1 else 0 end) as reject_actions,
  round(sum(case when {al_reject} then 1 else 0 end) / nullif(count(1), 0) * 100, 3) as reject_rate_pct
from {AF_ACTION_LOG_TABLE} al
where {al_window} and {al_evaluated}
group by coalesce(nullif(trim(al.scene_name), ''), 'Unspecified'),
  coalesce(nullif(trim(al.authentication_type), ''), 'Unspecified')
order by actions desc;

-- 7. Challenge Friction by Auth Type
-- How much friction each authentication tier adds: volume share, reject rate, and the users/flows it
-- touched. Step-up (CHALLENGE_*) tiers are the added customer friction.
select
  coalesce(nullif(trim(al.authentication_type), ''), 'Unspecified') as authentication_type,
  count(1) as actions,
  count(distinct al.uid) as distinct_users,
  count(distinct al.bizflow_instance_id) as flows,
  round(count(1) / nullif(sum(count(1)) over (), 0) * 100, 2) as action_share_pct,
  sum(case when {al_pass} then 1 else 0 end) as pass_actions,
  sum(case when {al_reject} then 1 else 0 end) as reject_actions,
  round(sum(case when {al_reject} then 1 else 0 end) / nullif(count(1), 0) * 100, 3) as reject_rate_pct
from {AF_ACTION_LOG_TABLE} al
where {al_window} and {al_evaluated}
group by coalesce(nullif(trim(al.authentication_type), ''), 'Unspecified')
order by actions desc;

-- 8. Auth Drop-off by Scene
-- Flows started vs flows that reached a final action that day (is_final_action_in_flow_of_the_day).
-- drop_off_rate_pct approximates abandonment mid-flow (started but no final action recorded).
select
  coalesce(nullif(trim(al.scene_name), ''), 'Unspecified') as scene_name,
  count(distinct al.bizflow_instance_id) as flows_started,
  count(distinct case when al.is_final_action_in_flow_of_the_day = 'Y' then al.bizflow_instance_id end) as flows_reached_final,
  round((count(distinct al.bizflow_instance_id)
         - count(distinct case when al.is_final_action_in_flow_of_the_day = 'Y' then al.bizflow_instance_id end))
        / nullif(count(distinct al.bizflow_instance_id), 0) * 100, 2) as drop_off_rate_pct
from {AF_ACTION_LOG_TABLE} al
where {al_window}
group by coalesce(nullif(trim(al.scene_name), ''), 'Unspecified')
order by flows_started desc;
"""


def _merge_report_sections(sql: str, *, start_number: int) -> str:
    """Strip a builder's header/preamble (everything before its first ``-- 1.`` marker) and renumber its
    ``-- N. Title`` section markers to continue from ``start_number``, so the sections can be appended
    into another report's combined SQL. Section titles (which become Excel sheet names) are preserved."""
    match = re.search(r"^--\s+1\.\s", sql, flags=re.M)
    body = sql[match.start():] if match else sql
    counter = [start_number - 1]

    def _bump(m: "re.Match[str]") -> str:
        counter[0] += 1
        return f"-- {counter[0]}. {m.group(1)}"

    return re.sub(r"^--\s+\d+\.\s+(.+?)\s*$", _bump, body, flags=re.M)


def build_af_rules_features_sql(*, snapshot_pt_date: str | None = None, now: datetime | None = None) -> str:
    rule_snap = _aliased_snapshot_filter("rc", AF_RULE_CONFIG_TABLE, snapshot_pt_date)
    feature_snap = _aliased_snapshot_filter("fc", AF_FEATURE_CONFIG_TABLE, snapshot_pt_date)
    header = _af_report_header("Anti-fraud PH - Rules & Features", snapshot_pt_date)
    base = f"""{header}
-- Full rule and feature catalogs. outcome_type: 1=Punish, 2=Challenge, 3=Reject.
-- rule status > 0 = Active; feature status 1=Active / -1=Inactive.

-- 1. Rules
select
  rc.rule_id,
  rc.rule_name,
  rc.feature_expr,
  case when rc.status > 0 then 'Active' else 'Inactive/Draft' end as rule_status,
  rc.status as status_code,
  case rc.outcome_type when 1 then 'Punish' when 2 then 'Challenge' when 3 then 'Reject' else cast(rc.outcome_type as string) end as outcome_type,
  case rc.real_time when 1 then 'Real-time' else 'Batch' end as execution_mode,
  rc.risk_level,
  rc.priority,
  rc.review_priority,
  rc.punish_action,
  rc.punish_scene,
  rc.punish_sub_scene,
  rc.notice_template
from {AF_RULE_CONFIG_TABLE} rc
where {rule_snap}
order by case when rc.status > 0 then 0 else 1 end, rc.rule_id;

-- 2. Features
select
  fc.feature_id,
  fc.feature_name,
  fc.function_id,
  case fc.status when 1 then 'Active' when -1 then 'Inactive' else cast(fc.status as string) end as feature_status,
  fc.type as feature_type,
  fc.base_obj,
  fc.count_obj,
  fc.time_range as count_window_seconds,
  fc.operator,
  fc.threshold,
  case fc.consecutive when 1 then 'Y' else 'N' end as consecutive,
  fc.scene,
  fc.sub_scene,
  fc.action,
  fc.event_status,
  fc.scenario_type,
  fc.business_category
from {AF_FEATURE_CONFIG_TABLE} fc
where {feature_snap}
order by case when fc.status = 1 then 0 else 1 end, fc.feature_id;
"""
    # Governance: fold in the rule change-log sections (Change Summary / Detail / Current Inventory)
    # from the same rule_config snapshots, renumbered to follow Rules (1) and Features (2).
    governance = _merge_report_sections(
        build_af_rule_change_log_sql(snapshot_pt_date=snapshot_pt_date, now=now), start_number=3
    )
    return base + "\n" + governance


def build_af_rule_effectiveness_sql(*, snapshot_pt_date: str | None = None, now: datetime | None = None) -> str:
    # Scope: last two full calendar months + current MTD. Detail and trend sections aggregate over the
    # whole span; the summary breaks it out per period.
    window = business_insights_window(now)
    span_label = window.span_label
    span_start, span_end_exclusive = window.span_start, window.span_end_exclusive
    p1_label, p2_label, p3_label = (p[0] for p in window.periods)
    # Span-wide bounds reused by every section.
    start_ms = _date_to_epoch_millis(span_start)
    end_ms = _date_to_epoch_millis(span_end_exclusive)
    start_key = _yyyymmdd(span_start)
    end_key_exclusive = _yyyymmdd(span_end_exclusive)
    p2_key, p3_key = _yyyymmdd(window.periods[1][1]), _yyyymmdd(window.periods[2][1])
    month_start_iso = span_start.isoformat()
    month_end_iso = (span_end_exclusive - timedelta(days=1)).isoformat()
    next_month_iso = span_end_exclusive.isoformat()
    span_start_iso, span_end_iso = month_start_iso, month_end_iso
    request_snap = _aliased_snapshot_filter("rq", AF_REQUEST_STATISTIC_TABLE, snapshot_pt_date)
    reject_snap = _aliased_snapshot_filter("r", AF_IDENTIFY_REJECT_TABLE, snapshot_pt_date)
    punish_snap = _aliased_snapshot_filter("p", AF_PUNISH_LIST_TABLE, snapshot_pt_date)
    stat_snap = _aliased_snapshot_filter("rs", AF_RULE_TRIGGER_STATISTIC_TABLE, snapshot_pt_date)
    header = _af_report_header("Anti-fraud PH - Rule Effectiveness / Hit-Rate", snapshot_pt_date)
    base = f"""{header}
-- Scope: {span_label} ({span_start.isoformat()} to {span_end_exclusive.isoformat()} exclusive).
-- Detail and trend sections aggregate over the whole span; the summary breaks it out per period.
-- rule_trigger_log_tab and identify_record_tab are empty in ODS; this uses request_statistic,
-- identify_reject, and rule_trigger_statistic. Each pt_date is a cumulative snapshot, so a single
-- (latest) snapshot is pinned and rows are scoped by date / operation_time to avoid double counting.
-- Action rate uses pass+challenge+reject as the denominator (total_req_num is not populated).

-- 1. Request Outcome Summary
-- One row per period (last two full months + current MTD) for comparison.
select
  case when rq.date < '{p2_key}' then '{p1_label}' when rq.date < '{p3_key}' then '{p2_label}' else '{p3_label}' end as period,
  sum(coalesce(rq.pass_num, 0)) + sum(coalesce(rq.challenge_num, 0)) + sum(coalesce(rq.reject_num, 0)) as total_outcomes,
  sum(coalesce(rq.success_num, 0)) as successful_requests,
  sum(coalesce(rq.fail_num, 0)) as failed_requests,
  sum(coalesce(rq.pass_num, 0)) as pass_num,
  sum(coalesce(rq.challenge_num, 0)) as challenge_num,
  sum(coalesce(rq.reject_num, 0)) as reject_num,
  round((sum(coalesce(rq.challenge_num, 0)) + sum(coalesce(rq.reject_num, 0)))
        / nullif(sum(coalesce(rq.pass_num, 0)) + sum(coalesce(rq.challenge_num, 0)) + sum(coalesce(rq.reject_num, 0)), 0) * 100, 2) as action_rate_pct
from {AF_REQUEST_STATISTIC_TABLE} rq
where {request_snap}
  and rq.date >= '{start_key}'
  and rq.date < '{end_key_exclusive}'
group by case when rq.date < '{p2_key}' then '{p1_label}' when rq.date < '{p3_key}' then '{p2_label}' else '{p3_label}' end
order by min(rq.date);

-- 2. Reject Rule Hit Summary
-- Per reject_rule (+ reject_type) from ODS identify_reject (counts and PHP amount). benchmark_trxn is
-- the transactions in the scenes the rule fired in (fmart action log). trigger_rate_pct = reject events
-- / benchmark; normalised_user_impact_pct = distinct rejected users / benchmark. Amount is in PHP.
with scene_traffic as (
  select scene_name, count(distinct bizflow_instance_id) as scene_trxn
  from {AF_ACTION_LOG_TABLE}
  where pt_date between '{month_start_iso}' and '{month_end_iso}'
  group by scene_name
),
rej as (
  select r.reject_rule, r.reject_type,
    coalesce(s.name, concat('scene_', cast(r.operation_scene as string))) as scene_name,
    r.uid, cast(r.transaction_amount as double) as amt
  from {AF_IDENTIFY_REJECT_TABLE} r
  left join {AF_SCENE_TABLE} s
    on s.code = cast(r.operation_scene as string)
    and s.pt_date = (select max(pt_date) from {AF_SCENE_TABLE})
  where {reject_snap} and r.operation_time >= {start_ms} and r.operation_time < {end_ms}
),
rule_agg as (
  select reject_rule, reject_type,
    count(1) as reject_count, count(distinct uid) as distinct_users, count(distinct scene_name) as distinct_scenes,
    cast(round(sum(coalesce(amt, 0)), 2) as decimal(20, 2)) as rejected_amount_php
  from rej group by reject_rule, reject_type
),
rule_bench as (
  select rs.reject_rule, rs.reject_type, sum(st.scene_trxn) as benchmark_trxn
  from (select distinct reject_rule, reject_type, scene_name from rej) rs
  join scene_traffic st on st.scene_name = rs.scene_name
  group by rs.reject_rule, rs.reject_type
),
rule_dim as (
  select rule_id, max(rule_name) as rule_name
  from {AF_RULE_CONFIG_TABLE}
  where pt_date = (select max(pt_date) from {AF_RULE_CONFIG_TABLE})
  group by rule_id
)
select '{span_label}' as period, a.reject_rule, coalesce(rn.rule_name, '') as rule_name, a.reject_type,
  a.reject_count, a.distinct_users, a.distinct_scenes, a.rejected_amount_php,
  b.benchmark_trxn,
  round(a.reject_count / nullif(b.benchmark_trxn, 0) * 100, 3) as trigger_rate_pct,
  round(a.distinct_users / nullif(b.benchmark_trxn, 0) * 100, 3) as normalised_user_impact_pct
from rule_agg a
left join rule_bench b on b.reject_rule = a.reject_rule and b.reject_type = a.reject_type
left join rule_dim rn on rn.rule_id = a.reject_rule
order by a.reject_count desc;

-- 3. Reject Rule Scene Breakdown
-- Per reject_rule x scene with scene-level benchmark_trxn (scene transactions from the fmart action log)
-- and the trigger_rate_pct / normalised_user_impact_pct for that scene. Amount is in PHP.
with scene_traffic as (
  select scene_name, count(distinct bizflow_instance_id) as scene_trxn
  from {AF_ACTION_LOG_TABLE}
  where pt_date between '{month_start_iso}' and '{month_end_iso}'
  group by scene_name
),
rej as (
  select r.reject_rule, r.reject_type,
    coalesce(s.name, concat('scene_', cast(r.operation_scene as string))) as scene_name,
    r.uid, cast(r.transaction_amount as double) as amt
  from {AF_IDENTIFY_REJECT_TABLE} r
  left join {AF_SCENE_TABLE} s
    on s.code = cast(r.operation_scene as string)
    and s.pt_date = (select max(pt_date) from {AF_SCENE_TABLE})
  where {reject_snap} and r.operation_time >= {start_ms} and r.operation_time < {end_ms}
),
rule_scene as (
  select reject_rule, reject_type, scene_name,
    count(1) as reject_count, count(distinct uid) as distinct_users,
    cast(round(sum(coalesce(amt, 0)), 2) as decimal(20, 2)) as rejected_amount_php
  from rej group by reject_rule, reject_type, scene_name
)
select rsa.reject_rule, rsa.reject_type, rsa.scene_name,
  rsa.reject_count, rsa.distinct_users, rsa.rejected_amount_php,
  st.scene_trxn as benchmark_trxn,
  round(rsa.reject_count / nullif(st.scene_trxn, 0) * 100, 3) as trigger_rate_pct,
  round(rsa.distinct_users / nullif(st.scene_trxn, 0) * 100, 3) as normalised_user_impact_pct
from rule_scene rsa left join scene_traffic st on st.scene_name = rsa.scene_name
order by rsa.reject_rule, rsa.reject_count desc;

-- 4. Punishment Rule Hit Summary
-- Per punish_rule_id from the punish list; start_time scopes the month. Punishments carry no transaction
-- amount; targets are distinct id_value (uid / device / phone). benchmark_trxn / trigger_rate_pct /
-- normalised_user_impact_pct use scene transactions from the fmart action log.
with scene_traffic as (
  select scene_name, count(distinct bizflow_instance_id) as scene_trxn
  from {AF_ACTION_LOG_TABLE}
  where pt_date between '{month_start_iso}' and '{month_end_iso}'
  group by scene_name
),
pun as (
  select p.punish_rule_id,
    coalesce(s.name, concat('scene_', cast(p.scene as string))) as scene_name,
    p.id_value
  from {AF_PUNISH_LIST_TABLE} p
  left join {AF_SCENE_TABLE} s
    on s.code = cast(p.scene as string)
    and s.pt_date = (select max(pt_date) from {AF_SCENE_TABLE})
  where {punish_snap} and p.start_time >= {start_ms} and p.start_time < {end_ms}
),
rule_agg as (
  select punish_rule_id,
    count(1) as punish_count, count(distinct id_value) as distinct_targets, count(distinct scene_name) as distinct_scenes
  from pun group by punish_rule_id
),
rule_bench as (
  select ps.punish_rule_id, sum(st.scene_trxn) as benchmark_trxn
  from (select distinct punish_rule_id, scene_name from pun) ps
  join scene_traffic st on st.scene_name = ps.scene_name
  group by ps.punish_rule_id
),
rule_dim as (
  select rule_id, max(rule_name) as rule_name
  from {AF_RULE_CONFIG_TABLE}
  where pt_date = (select max(pt_date) from {AF_RULE_CONFIG_TABLE})
  group by rule_id
)
select '{span_label}' as period, a.punish_rule_id, coalesce(rn.rule_name, '') as rule_name,
  a.punish_count, a.distinct_targets, a.distinct_scenes,
  b.benchmark_trxn,
  round(a.punish_count / nullif(b.benchmark_trxn, 0) * 100, 3) as trigger_rate_pct,
  round(a.distinct_targets / nullif(b.benchmark_trxn, 0) * 100, 3) as normalised_user_impact_pct
from rule_agg a
left join rule_bench b on b.punish_rule_id = a.punish_rule_id
left join rule_dim rn on rn.rule_id = a.punish_rule_id
order by a.punish_count desc;

-- 5. Punishment Rule Scene Breakdown
-- Per punish_rule_id x scene with scene-level benchmark_trxn and rates (scene transactions from the
-- fmart action log).
with scene_traffic as (
  select scene_name, count(distinct bizflow_instance_id) as scene_trxn
  from {AF_ACTION_LOG_TABLE}
  where pt_date between '{month_start_iso}' and '{month_end_iso}'
  group by scene_name
),
pun as (
  select p.punish_rule_id,
    coalesce(s.name, concat('scene_', cast(p.scene as string))) as scene_name,
    p.id_value
  from {AF_PUNISH_LIST_TABLE} p
  left join {AF_SCENE_TABLE} s
    on s.code = cast(p.scene as string)
    and s.pt_date = (select max(pt_date) from {AF_SCENE_TABLE})
  where {punish_snap} and p.start_time >= {start_ms} and p.start_time < {end_ms}
),
rule_scene as (
  select punish_rule_id, scene_name,
    count(1) as punish_count, count(distinct id_value) as distinct_targets
  from pun group by punish_rule_id, scene_name
)
select rsa.punish_rule_id, rsa.scene_name,
  rsa.punish_count, rsa.distinct_targets,
  st.scene_trxn as benchmark_trxn,
  round(rsa.punish_count / nullif(st.scene_trxn, 0) * 100, 3) as trigger_rate_pct,
  round(rsa.distinct_targets / nullif(st.scene_trxn, 0) * 100, 3) as normalised_user_impact_pct
from rule_scene rsa left join scene_traffic st on st.scene_name = rsa.scene_name
order by rsa.punish_rule_id, rsa.punish_count desc;

-- 6. Challenge Rule Hit Summary
-- Challenge rules (rule_config.outcome_type = 2) sourced from the DWD rule hit log, counting only
-- effective hits (is_rule_triggered = 'Y'). trigger_rate_pct = challenge transactions / transactions
-- in the scenes the rule operates in (from the DWD action log) - an exposure-adjusted hit rate.
-- normalised_user_impact_pct applies the same denominator to distinct challenged users.
with challenge_rules as (
  select
    rule_id,
    rule_name,
    case status when 1 then 'active' when 2 then 'collect data' when -1 then 'inactive' else 'other' end as rule_status,
    review_priority
  from {AF_RULE_CONFIG_TABLE}
  where pt_date = (select max(pt_date) from {AF_RULE_CONFIG_TABLE})
    and outcome_type = 2
),
challenge_hits as (
  select h.rule_id, h.scene_name, h.bizflow_instance_id, h.uid
  from {AF_RULE_HIT_LOG_TABLE} h
  join challenge_rules c on c.rule_id = h.rule_id
  where h.pt_date between '{month_start_iso}' and '{month_end_iso}'
    and h.is_rule_triggered = 'Y'
),
triggers as (
  select rule_id,
    count(distinct bizflow_instance_id) as challenge_trxn,
    count(distinct uid) as challenge_users,
    count(distinct scene_name) as distinct_scenes
  from challenge_hits group by rule_id
),
rule_scenes as (select distinct rule_id, scene_name from challenge_hits),
scene_traffic as (
  select scene_name, count(distinct bizflow_instance_id) as scene_trxn
  from {AF_ACTION_LOG_TABLE}
  where pt_date between '{month_start_iso}' and '{month_end_iso}'
  group by scene_name
),
benchmark as (
  select rs.rule_id, sum(st.scene_trxn) as benchmark_trxn
  from rule_scenes rs join scene_traffic st on st.scene_name = rs.scene_name
  group by rs.rule_id
)
select
  c.rule_id,
  c.rule_name,
  c.rule_status,
  c.review_priority,
  t.challenge_trxn,
  t.challenge_users,
  t.distinct_scenes,
  b.benchmark_trxn,
  round(t.challenge_trxn / nullif(b.benchmark_trxn, 0) * 100, 3) as trigger_rate_pct,
  round(t.challenge_users / nullif(b.benchmark_trxn, 0) * 100, 3) as normalised_user_impact_pct
from challenge_rules c
join triggers t on t.rule_id = c.rule_id
left join benchmark b on b.rule_id = c.rule_id
order by t.challenge_trxn desc;

-- 7. Challenge Rule Scene Breakdown
-- Per challenge rule x scene (scene name from the hit log), with scene-level benchmark_trxn (scene
-- transactions from the action log) and the trigger_rate_pct / normalised_user_impact_pct for that scene.
with scene_traffic as (
  select scene_name, count(distinct bizflow_instance_id) as scene_trxn
  from {AF_ACTION_LOG_TABLE}
  where pt_date between '{month_start_iso}' and '{month_end_iso}'
  group by scene_name
),
rule_scene as (
  select h.rule_id, h.scene_name,
    count(distinct h.bizflow_instance_id) as challenge_trxn,
    count(distinct h.uid) as challenge_users
  from {AF_RULE_HIT_LOG_TABLE} h
  join {AF_RULE_CONFIG_TABLE} rc
    on rc.rule_id = h.rule_id
    and rc.pt_date = (select max(pt_date) from {AF_RULE_CONFIG_TABLE})
  where h.pt_date between '{month_start_iso}' and '{month_end_iso}'
    and h.is_rule_triggered = 'Y'
    and rc.outcome_type = 2
  group by h.rule_id, h.scene_name
)
select rsa.rule_id, rsa.scene_name,
  rsa.challenge_trxn, rsa.challenge_users,
  st.scene_trxn as benchmark_trxn,
  round(rsa.challenge_trxn / nullif(st.scene_trxn, 0) * 100, 3) as trigger_rate_pct,
  round(rsa.challenge_users / nullif(st.scene_trxn, 0) * 100, 3) as normalised_user_impact_pct
from rule_scene rsa left join scene_traffic st on st.scene_name = rsa.scene_name
order by rsa.rule_id, rsa.challenge_trxn desc;

-- 8. Daily Challenge/Reject/Punish
select
  concat_ws('-', substr(rs.date, 1, 4), substr(rs.date, 5, 2), substr(rs.date, 7, 2)) as trigger_date,
  sum(coalesce(rs.challenge_num, 0)) as challenge_num,
  sum(coalesce(rs.reject_num, 0)) as reject_num,
  sum(coalesce(rs.punish_num, 0)) as punish_num
from {AF_RULE_TRIGGER_STATISTIC_TABLE} rs
where {stat_snap}
  and rs.date >= '{start_key}'
  and rs.date < '{end_key_exclusive}'
group by rs.date
order by trigger_date;

-- 9. Rule Precision / Catch Rate
-- For each rule that flagged cases for review (review_record), the share of those cases later confirmed
-- as fraud. Confirmed fraud = review_case.fraud_mo_type not in ('Not Fraud','Pending',''). Scoped to
-- cases opened in the month. precision_pct = fraud_cases / reviewed_cases; loss is in PHP.
with cases as (
  select case_id,
    case when lower(trim(coalesce(fraud_mo_type, ''))) not in ('not fraud', 'pending', '') then 1 else 0 end as is_fraud,
    coalesce(loss_total_amt, 0) as loss
  from {AF_REVIEW_CASE_TABLE}
  where pt_date = (select max(pt_date) from {AF_REVIEW_CASE_TABLE})
    and case_open_datetime >= '{month_start_iso}' and case_open_datetime < '{next_month_iso}'
),
rr as (
  select rule_id, max(rule_name) as rule_name, case_id
  from {AF_REVIEW_RECORD_TABLE}
  where pt_date = (select max(pt_date) from {AF_REVIEW_RECORD_TABLE})
    and review_status = 'REVIEWED' and rule_id is not null and trim(rule_id) <> ''
  group by rule_id, case_id
)
select
  rr.rule_id,
  max(rr.rule_name) as rule_name,
  count(distinct rr.case_id) as reviewed_cases,
  count(distinct case when c.is_fraud = 1 then rr.case_id end) as fraud_cases,
  round(count(distinct case when c.is_fraud = 1 then rr.case_id end) / nullif(count(distinct rr.case_id), 0) * 100, 2) as precision_pct,
  cast(round(sum(case when c.is_fraud = 1 then c.loss else 0 end), 2) as decimal(20, 2)) as fraud_loss_php
from rr join cases c on c.case_id = rr.case_id
group by rr.rule_id
order by fraud_cases desc, reviewed_cases desc;

-- 10. Daily Rule Trigger Trend
-- Daily effective-hit transactions per rule for the top 50 rules by monthly volume (the result API caps
-- at 2000 rows). Powers the filterable daily trend chart.
with daily as (
  select rule_id, pt_date as trigger_date, count(distinct bizflow_instance_id) as trigger_trxn
  from {AF_RULE_HIT_LOG_TABLE}
  where pt_date between '{span_start_iso}' and '{span_end_iso}' and is_rule_triggered = 'Y'
  group by rule_id, pt_date
),
top_rules as (
  select rule_id from daily group by rule_id order by sum(trigger_trxn) desc limit 50
)
select d.rule_id, d.trigger_date, d.trigger_trxn
from daily d join top_rules t on t.rule_id = d.rule_id
order by d.rule_id, d.trigger_date;

-- 11. Scene/Sub-scene/Action Usage
-- Actual transaction volume per scene / sub-scene / action from the action log (all traffic, not just
-- rule hits). Shows which configured scenes and actions are actually exercised.
select
  scene_name,
  sub_scene_name,
  action_name,
  case action_type when '1' then 'Business' when '2' then 'Authentication' else action_type end as action_type,
  count(distinct bizflow_instance_id) as transactions,
  count(1) as action_events,
  count(distinct uid) as distinct_users
from {AF_ACTION_LOG_TABLE}
where pt_date between '{month_start_iso}' and '{month_end_iso}'
group by scene_name, sub_scene_name, action_name, action_type
order by transactions desc;

-- 12. Rule Scorecard
-- Precision x trigger-rate quadrant. Combines, per rule, the trigger rate (hit log: effective
-- triggers / scene traffic) with precision
-- (review records -> confirmed-fraud cases). Limited to rules that flagged cases for review, so both
-- axes are present - the quadrant for tuning: high trigger rate + low precision = noisy / retire.
with review_cases as (
  select case_id,
    case when lower(trim(coalesce(fraud_mo_type, ''))) not in ('not fraud', 'pending', '') then 1 else 0 end as is_fraud
  from {AF_REVIEW_CASE_TABLE}
  where pt_date = (select max(pt_date) from {AF_REVIEW_CASE_TABLE})
    and case_open_datetime >= '{month_start_iso}' and case_open_datetime < '{next_month_iso}'
),
precision_rules as (
  select rr.rule_id, max(rr.rule_name) as rule_name,
    count(distinct rr.case_id) as reviewed_cases,
    count(distinct case when c.is_fraud = 1 then rr.case_id end) as fraud_cases
  from {AF_REVIEW_RECORD_TABLE} rr
  join review_cases c on c.case_id = rr.case_id
  where rr.pt_date = (select max(pt_date) from {AF_REVIEW_RECORD_TABLE})
    and rr.review_status = 'REVIEWED' and rr.rule_id is not null and trim(rr.rule_id) <> ''
  group by rr.rule_id
),
triggers as (
  select rule_id, count(distinct bizflow_instance_id) as trigger_trxn, count(distinct uid) as trigger_users
  from {AF_RULE_HIT_LOG_TABLE}
  where pt_date between '{month_start_iso}' and '{month_end_iso}' and is_rule_triggered = 'Y'
  group by rule_id
),
rule_scenes as (
  select distinct rule_id, scene_name from {AF_RULE_HIT_LOG_TABLE}
  where pt_date between '{month_start_iso}' and '{month_end_iso}' and is_rule_triggered = 'Y'
),
scene_traffic as (
  select scene_name, count(distinct bizflow_instance_id) as scene_trxn
  from {AF_ACTION_LOG_TABLE}
  where pt_date between '{month_start_iso}' and '{month_end_iso}'
  group by scene_name
),
benchmark as (
  select rs.rule_id, sum(st.scene_trxn) as benchmark_trxn
  from rule_scenes rs join scene_traffic st on st.scene_name = rs.scene_name
  group by rs.rule_id
)
select
  p.rule_id,
  p.rule_name,
  t.trigger_trxn,
  b.benchmark_trxn,
  round(t.trigger_trxn / nullif(b.benchmark_trxn, 0) * 100, 3) as trigger_rate_pct,
  p.reviewed_cases,
  p.fraud_cases,
  round(p.fraud_cases / nullif(p.reviewed_cases, 0) * 100, 2) as precision_pct
from precision_rules p
left join triggers t on t.rule_id = p.rule_id
left join benchmark b on b.rule_id = p.rule_id
order by p.fraud_cases desc, trigger_rate_pct desc;
"""
    # Detection effectiveness: fold the rule-hits -> confirmed-fraud-loss sections in as 13-17. Pass
    # snapshot_pt_date=None so the review_case/review_record sections self-resolve their own latest
    # partition (this report's snapshot anchor is request_statistic, a different table).
    detection = _merge_report_sections(
        build_af_detection_effectiveness_sql(snapshot_pt_date=None, now=now), start_number=13
    )
    return base + "\n" + detection


def build_af_fraud_loss_sql(*, snapshot_pt_date: str | None = None, now: datetime | None = None) -> str:
    # Scope: cases opened in the last two full calendar months + the current month-to-date.
    win = business_insights_window(now)
    span_label = win.span_label
    span_start_iso = win.span_start.isoformat()
    span_end_iso = win.span_end_exclusive.isoformat()
    p2_iso, p3_iso = win.periods[1][1].isoformat(), win.periods[2][1].isoformat()
    p1_label, p2_label, p3_label = (p[0] for p in win.periods)
    case_snap = _aliased_snapshot_filter("c", AF_REVIEW_CASE_TABLE, snapshot_pt_date)
    window = f"c.case_open_datetime >= '{span_start_iso}' and c.case_open_datetime < '{span_end_iso}'"
    period_case = (
        f"case when c.case_open_datetime < '{p2_iso}' then '{p1_label}' "
        f"when c.case_open_datetime < '{p3_iso}' then '{p2_label}' else '{p3_label}' end"
    )
    # Used by per-period sections that need an ISO span for the DWD action/review tables.
    month_start_iso = span_start_iso
    next_month_iso = span_end_iso
    # Confirmed fraud = a fraud modus operandi was assigned (not the 'Not Fraud'/'Pending' verdicts).
    fraud_expr = "lower(trim(coalesce(c.fraud_mo_type, ''))) not in ('not fraud', 'pending', '')"
    mo_expr = "case when trim(coalesce(c.fraud_mo_type, '')) = '' then 'Unspecified' else c.fraud_mo_type end"
    # fraud_mo_subtype is stored as a JSON array string (e.g. ["Order (E-commerce)"]); strip the brackets
    # and quotes so it reads as plain text.
    subtype_expr = "case when trim(coalesce(c.fraud_mo_subtype, '')) = '' then 'Unspecified' else regexp_replace(translate(c.fraud_mo_subtype, '[]\"', ''), ',', ', ') end"
    review_hours = (
        "round(avg(case when c.case_status = 'CLOSED' and c.case_closed_timestamp is not null "
        "and c.case_open_timestamp is not null then (c.case_closed_timestamp - c.case_open_timestamp) / 3600000.0 end), 1)"
    )
    header = _af_report_header("Anti-fraud PH - Fraud Loss & Case Outcomes", snapshot_pt_date)
    return f"""{header}
-- Source: DWD review_case (fraud case management). Scope: cases opened in {span_label}.
-- Confirmed fraud = fraud_mo_type not in ('Not Fraud','Pending',''). All amounts are in PHP; the loss
-- split uses the loss_amt_borne_by_* columns (loss_borne_by text is not populated). Review hours use
-- (case_closed_timestamp - case_open_timestamp) for closed cases.

-- 1. Case & Loss Summary
-- One row per period (last two full months + current MTD) for comparison.
select
  {period_case} as period,
  count(1) as cases_opened,
  sum(case when {fraud_expr} then 1 else 0 end) as fraud_cases,
  round(sum(case when {fraud_expr} then 1 else 0 end) / nullif(count(1), 0) * 100, 2) as fraud_rate_pct,
  cast(round(sum(coalesce(c.loss_total_amt, 0)), 2) as decimal(20, 2)) as total_loss_php,
  cast(round(sum(coalesce(c.loss_amt_borne_by_customer, 0)), 2) as decimal(20, 2)) as loss_customer_php,
  cast(round(sum(coalesce(c.loss_amt_borne_by_bank, 0)), 2) as decimal(20, 2)) as loss_bank_php,
  cast(round(sum(coalesce(c.loss_amt_borne_by_third_party, 0)), 2) as decimal(20, 2)) as loss_third_party_php,
  cast(round(sum(coalesce(c.loss_recovered_amt, 0)), 2) as decimal(20, 2)) as recovered_php,
  sum(case when c.case_status = 'CLOSED' then 1 else 0 end) as closed_cases,
  sum(case when c.case_status = 'OPEN' then 1 else 0 end) as open_cases,
  {review_hours} as avg_review_hours
from {AF_REVIEW_CASE_TABLE} c
where {case_snap} and {window}
group by {period_case}
order by min(c.case_open_datetime);

-- 2. Loss by Fraud MO Type
select
  {mo_expr} as fraud_mo_type,
  count(1) as cases,
  count(distinct case when trim(coalesce(c.fraud_mo_subtype, '')) <> '' then c.fraud_mo_subtype end) as distinct_subtypes,
  cast(round(sum(coalesce(c.loss_total_amt, 0)), 2) as decimal(20, 2)) as total_loss_php,
  cast(round(avg(coalesce(c.loss_total_amt, 0)), 2) as decimal(20, 2)) as avg_loss_php,
  cast(round(sum(coalesce(c.loss_recovered_amt, 0)), 2) as decimal(20, 2)) as recovered_php
from {AF_REVIEW_CASE_TABLE} c
where {case_snap} and {window}
group by {mo_expr}
order by total_loss_php desc, cases desc;

-- 3. Fraud MO Subtype Breakdown
select
  {mo_expr} as fraud_mo_type,
  {subtype_expr} as fraud_mo_subtype,
  count(1) as cases,
  cast(round(sum(coalesce(c.loss_total_amt, 0)), 2) as decimal(20, 2)) as total_loss_php,
  cast(round(sum(coalesce(c.loss_recovered_amt, 0)), 2) as decimal(20, 2)) as recovered_php
from {AF_REVIEW_CASE_TABLE} c
where {case_snap} and {window}
group by {mo_expr}, {subtype_expr}
order by {mo_expr}, total_loss_php desc;

-- 4. Case Status & SLA
select
  coalesce(nullif(trim(c.case_status), ''), 'Unspecified') as case_status,
  count(1) as cases,
  cast(round(sum(coalesce(c.loss_total_amt, 0)), 2) as decimal(20, 2)) as total_loss_php,
  {review_hours} as avg_review_hours
from {AF_REVIEW_CASE_TABLE} c
where {case_snap} and {window}
group by coalesce(nullif(trim(c.case_status), ''), 'Unspecified')
order by cases desc;

-- 5. Daily Fraud Loss Trend
-- Daily total loss per fraud MO type plus an 'All' series; powers the filterable trend chart.
with base as (
  select {mo_expr} as fraud_mo_type, substr(c.case_open_datetime, 1, 10) as case_open_date,
    coalesce(c.loss_total_amt, 0) as loss
  from {AF_REVIEW_CASE_TABLE} c
  where {case_snap} and {window}
)
select fraud_mo_type, case_open_date,
  cast(round(sum(loss), 2) as decimal(20, 2)) as daily_loss_php,
  count(1) as daily_cases
from base group by fraud_mo_type, case_open_date
union all
select 'All' as fraud_mo_type, case_open_date,
  cast(round(sum(loss), 2) as decimal(20, 2)) as daily_loss_php,
  count(1) as daily_cases
from base group by case_open_date
order by fraud_mo_type, case_open_date;

-- 6. Review Pool / Backlog (current)
-- Pending review records not yet linked to a case, as of the latest snapshot. This is the CURRENT
-- backlog (not month-scoped). avg_age_days = days since the upstream event.
select
  coalesce(nullif(trim(rr.source), ''), 'Unspecified') as source,
  count(1) as pending_records,
  count(distinct rr.rule_id) as distinct_rules,
  count(distinct rr.uid) as distinct_users,
  min(rr.upstream_event_datetime) as oldest_pending,
  round(avg((unix_timestamp(current_timestamp()) - rr.upstream_event_timestamp / 1000) / 86400.0), 1) as avg_age_days
from {AF_REVIEW_RECORD_TABLE} rr
where rr.pt_date = (select max(pt_date) from {AF_REVIEW_RECORD_TABLE})
  and rr.review_status = 'PENDING'
group by coalesce(nullif(trim(rr.source), ''), 'Unspecified')
order by pending_records desc;
"""


def build_af_detection_effectiveness_sql(*, snapshot_pt_date: str | None = None, now: datetime | None = None) -> str:
    # Detection effectiveness ties the rule engine to confirmed-fraud outcomes: of the fraud cases (and
    # the PHP loss they carry), how much did a rule flag for review vs surface some other way (customer
    # report / chargeback / manual). Reuses the proven review_record -> review_case -> rule join that the
    # Rule Effectiveness scorecard already relies on. Scope: cases opened in the last two full months + MTD.
    win = business_insights_window(now)
    span_label = win.span_label
    span_start_iso = win.span_start.isoformat()
    span_end_iso = win.span_end_exclusive.isoformat()
    p2_iso, p3_iso = win.periods[1][1].isoformat(), win.periods[2][1].isoformat()
    p1_label, p2_label, p3_label = (p[0] for p in win.periods)
    case_snap = _aliased_snapshot_filter("c", AF_REVIEW_CASE_TABLE, snapshot_pt_date)
    window = f"c.case_open_datetime >= '{span_start_iso}' and c.case_open_datetime < '{span_end_iso}'"
    # NB: period_case is evaluated in the final SELECT over the `flagged` CTE (which has no `c` alias),
    # so it references the bare column name rather than c.case_open_datetime.
    period_case = (
        f"case when case_open_datetime < '{p2_iso}' then '{p1_label}' "
        f"when case_open_datetime < '{p3_iso}' then '{p2_label}' else '{p3_label}' end"
    )
    fraud_expr = "lower(trim(coalesce(c.fraud_mo_type, ''))) not in ('not fraud', 'pending', '')"
    mo_expr = "case when trim(coalesce(c.fraud_mo_type, '')) = '' then 'Unspecified' else c.fraud_mo_type end"
    # fraud_mo_subtype is stored as a JSON array string (e.g. ["Order (E-commerce)"]); strip the brackets
    # and quotes so it reads as plain text.
    subtype_expr = "case when trim(coalesce(c.fraud_mo_subtype, '')) = '' then 'Unspecified' else regexp_replace(translate(c.fraud_mo_subtype, '[]\"', ''), ',', ', ') end"
    # A case is "rule-detected" when at least one review_record links it to a non-empty rule_id.
    cases_cte = f"""cases as (
  select c.case_id,
    c.case_open_datetime,
    case when {fraud_expr} then 1 else 0 end as is_fraud,
    {mo_expr} as fraud_mo_type,
    {subtype_expr} as fraud_mo_subtype,
    coalesce(c.loss_total_amt, 0) as loss,
    coalesce(c.loss_recovered_amt, 0) as recovered
  from {AF_REVIEW_CASE_TABLE} c
  where {case_snap} and {window}
),
case_rules as (
  select rr.case_id,
    count(distinct case when rr.rule_id is not null and trim(rr.rule_id) <> '' then rr.rule_id end) as rule_count
  from {AF_REVIEW_RECORD_TABLE} rr
  where rr.pt_date = (select max(pt_date) from {AF_REVIEW_RECORD_TABLE})
  group by rr.case_id
),
flagged as (
  select cs.*, case when coalesce(cr.rule_count, 0) > 0 then 1 else 0 end as rule_detected
  from cases cs left join case_rules cr on cr.case_id = cs.case_id
)"""
    header = _af_report_header("Anti-fraud PH - Detection Effectiveness & Loss Prevented", snapshot_pt_date)
    return f"""{header}
-- Source: DWD review_case (fraud verdict + PHP loss) joined to review_record (which rule flagged the
-- case for review). Scope: cases opened in {span_label}. Confirmed fraud = fraud_mo_type not in
-- ('Not Fraud','Pending',''). rule-detected = >=1 review_record with a non-empty rule_id; the remainder
-- is "leaked" (surfaced outside the rule engine). Loss is in PHP. Per-rule loss is attributed to every
-- rule that flagged the case, so it can exceed the total when multiple rules cover one case.

-- 1. Detection Coverage Summary
-- One row per period (last two full months + current MTD). detection_rate_pct = rule-detected fraud
-- cases / fraud cases; loss_detected_share_pct = rule-detected fraud loss / total fraud loss.
with {cases_cte}
select
  {period_case} as period,
  sum(is_fraud) as fraud_cases,
  sum(case when is_fraud = 1 and rule_detected = 1 then 1 else 0 end) as rule_detected_cases,
  round(sum(case when is_fraud = 1 and rule_detected = 1 then 1 else 0 end) / nullif(sum(is_fraud), 0) * 100, 2) as detection_rate_pct,
  cast(round(sum(case when is_fraud = 1 then loss else 0 end), 2) as decimal(20, 2)) as fraud_loss_php,
  cast(round(sum(case when is_fraud = 1 and rule_detected = 1 then loss else 0 end), 2) as decimal(20, 2)) as loss_rule_detected_php,
  cast(round(sum(case when is_fraud = 1 and rule_detected = 0 then loss else 0 end), 2) as decimal(20, 2)) as loss_leaked_php,
  round(sum(case when is_fraud = 1 and rule_detected = 1 then loss else 0 end)
        / nullif(sum(case when is_fraud = 1 then loss else 0 end), 0) * 100, 2) as loss_detected_share_pct
from flagged
group by {period_case}
order by min(case_open_datetime);

-- 2. Detection by Fraud MO Type
-- Confirmed fraud only. loss_leaked_php highlights the blind spots (fraud the rule engine missed).
with {cases_cte}
select
  fraud_mo_type,
  sum(is_fraud) as fraud_cases,
  sum(case when is_fraud = 1 and rule_detected = 1 then 1 else 0 end) as rule_detected_cases,
  round(sum(case when is_fraud = 1 and rule_detected = 1 then 1 else 0 end) / nullif(sum(is_fraud), 0) * 100, 2) as detection_rate_pct,
  cast(round(sum(case when is_fraud = 1 then loss else 0 end), 2) as decimal(20, 2)) as total_loss_php,
  cast(round(sum(case when is_fraud = 1 and rule_detected = 1 then loss else 0 end), 2) as decimal(20, 2)) as loss_detected_php,
  cast(round(sum(case when is_fraud = 1 and rule_detected = 0 then loss else 0 end), 2) as decimal(20, 2)) as loss_leaked_php
from flagged
where is_fraud = 1
group by fraud_mo_type
order by loss_leaked_php desc, total_loss_php desc;

-- 3. Top Detecting Rules
-- Per rule: the confirmed-fraud cases it flagged and the PHP loss it helped catch, plus its precision
-- (fraud cases / all cases it flagged). The MVP rules - what is actually carrying detection.
with {cases_cte},
links as (
  select rr.case_id, rr.rule_id, max(rr.rule_name) as rule_name
  from {AF_REVIEW_RECORD_TABLE} rr
  where rr.pt_date = (select max(pt_date) from {AF_REVIEW_RECORD_TABLE})
    and rr.rule_id is not null and trim(rr.rule_id) <> ''
  group by rr.case_id, rr.rule_id
)
select
  l.rule_id,
  max(l.rule_name) as rule_name,
  count(distinct l.case_id) as flagged_cases,
  count(distinct case when f.is_fraud = 1 then l.case_id end) as fraud_cases_caught,
  round(count(distinct case when f.is_fraud = 1 then l.case_id end) / nullif(count(distinct l.case_id), 0) * 100, 2) as precision_pct,
  cast(round(sum(case when f.is_fraud = 1 then f.loss else 0 end), 2) as decimal(20, 2)) as loss_caught_php
from links l
join flagged f on f.case_id = l.case_id
group by l.rule_id
order by loss_caught_php desc, fraud_cases_caught desc;

-- 4. Missed Fraud (Blind Spots)
-- Confirmed-fraud cases no rule flagged, by MO type/subtype. The detection backlog to build rules for.
with {cases_cte}
select
  fraud_mo_type,
  fraud_mo_subtype,
  count(1) as missed_cases,
  cast(round(sum(loss), 2) as decimal(20, 2)) as loss_leaked_php,
  cast(round(sum(recovered), 2) as decimal(20, 2)) as recovered_php
from flagged
where is_fraud = 1 and rule_detected = 0
group by fraud_mo_type, fraud_mo_subtype
order by loss_leaked_php desc, missed_cases desc;

-- 5. Daily Detection Trend
-- Daily confirmed-fraud loss split into rule-detected vs leaked, plus an 'All fraud loss' series;
-- powers the filterable trend chart.
with {cases_cte},
daily as (
  select substr(case_open_datetime, 1, 10) as case_open_date,
    sum(case when is_fraud = 1 and rule_detected = 1 then loss else 0 end) as detected_loss,
    sum(case when is_fraud = 1 and rule_detected = 0 then loss else 0 end) as leaked_loss,
    sum(case when is_fraud = 1 then loss else 0 end) as all_loss,
    sum(case when is_fraud = 1 then 1 else 0 end) as fraud_cases
  from flagged group by substr(case_open_datetime, 1, 10)
)
select 'Rule-detected loss' as series, case_open_date,
  cast(round(detected_loss, 2) as decimal(20, 2)) as daily_loss_php, fraud_cases from daily
union all
select 'Leaked loss' as series, case_open_date,
  cast(round(leaked_loss, 2) as decimal(20, 2)) as daily_loss_php, fraud_cases from daily
union all
select 'All fraud loss' as series, case_open_date,
  cast(round(all_loss, 2) as decimal(20, 2)) as daily_loss_php, fraud_cases from daily
order by series, case_open_date;
"""


def build_af_rule_change_log_sql(*, snapshot_pt_date: str | None = None, now: datetime | None = None) -> str:
    # Governance / audit: diff the rule_config snapshot against an earlier one to surface what changed -
    # rules added, retired, activated, deactivated, re-targeted (outcome_type), re-scored (risk_level), or
    # re-logicked (feature_expr). The _ss config table keeps a daily snapshot per pt_date; we compare the
    # current snapshot to the snapshot on/before the report-window start, falling back to the earliest
    # retained snapshot when history does not reach that far back.
    win = business_insights_window(now)
    baseline_iso = win.span_start.isoformat()
    curr_snap = _aliased_snapshot_filter("rc", AF_RULE_CONFIG_TABLE, snapshot_pt_date)
    base_pt = (
        f"coalesce((select max(pt_date) from {AF_RULE_CONFIG_TABLE} where pt_date <= '{baseline_iso}'), "
        f"(select min(pt_date) from {AF_RULE_CONFIG_TABLE}))"
    )
    status_label = "case when {a}.status > 0 then 'Active' else 'Inactive/Draft' end"
    outcome_label = (
        "case {a}.outcome_type when 1 then 'Punish' when 2 then 'Challenge' when 3 then 'Reject' "
        "else cast({a}.outcome_type as string) end"
    )
    # The current and baseline projections, reused by every section. Baseline is pinned to one snapshot.
    snaps_cte = f"""curr as (
  select rc.rule_id, rc.rule_name, rc.status, rc.outcome_type, rc.risk_level,
    rc.priority, rc.review_priority, rc.feature_expr
  from {AF_RULE_CONFIG_TABLE} rc
  where {curr_snap}
),
base as (
  select rc.rule_id, rc.rule_name, rc.status, rc.outcome_type, rc.risk_level,
    rc.priority, rc.review_priority, rc.feature_expr
  from {AF_RULE_CONFIG_TABLE} rc
  where rc.pt_date = {base_pt}
),
joined as (
  select
    coalesce(c.rule_id, b.rule_id) as rule_id,
    coalesce(c.rule_name, b.rule_name) as rule_name,
    b.rule_id as base_rule_id, c.rule_id as curr_rule_id,
    b.status as base_status, c.status as curr_status,
    b.outcome_type as base_outcome, c.outcome_type as curr_outcome,
    b.risk_level as base_risk, c.risk_level as curr_risk,
    b.priority as base_priority, c.priority as curr_priority,
    b.review_priority as base_review_priority, c.review_priority as curr_review_priority,
    b.feature_expr as base_expr, c.feature_expr as curr_expr
  from curr c full outer join base b on c.rule_id = b.rule_id
),
classified as (
  select joined.*,
    case
      when base_rule_id is null then 'Added'
      when curr_rule_id is null then 'Removed'
      when coalesce(base_status, 0) <= 0 and coalesce(curr_status, 0) > 0 then 'Activated'
      when coalesce(base_status, 0) > 0 and coalesce(curr_status, 0) <= 0 then 'Deactivated'
      when coalesce(base_outcome, -999) <> coalesce(curr_outcome, -999) then 'Outcome changed'
      when coalesce(base_risk, '') <> coalesce(curr_risk, '') then 'Risk level changed'
      when coalesce(base_expr, '') <> coalesce(curr_expr, '') then 'Logic changed'
      when coalesce(base_priority, -999) <> coalesce(curr_priority, -999)
        or coalesce(base_review_priority, -999) <> coalesce(curr_review_priority, -999) then 'Priority changed'
      else 'Unchanged'
    end as change_type
  from joined
)"""
    curr_status_l = status_label.format(a="rc")
    curr_outcome_l = outcome_label.format(a="rc")
    header = _af_report_header("Anti-fraud PH - Rule Change Log & Governance", snapshot_pt_date)
    return f"""{header}
-- Source: rule_config snapshots. Current snapshot vs the snapshot on/before {baseline_iso} (the report
-- window start), falling back to the earliest retained snapshot if history is shorter. outcome_type:
-- 1=Punish, 2=Challenge, 3=Reject. status > 0 = Active. If the table retains only recent snapshots the
-- baseline collapses to the earliest available and rules may appear as 'Added' - read change counts with
-- the snapshot span in mind.

-- 1. Change Summary
-- Count of rules per change type since the baseline snapshot.
with {snaps_cte}
select change_type, count(1) as rules
from classified
group by change_type
order by case change_type
  when 'Added' then 1 when 'Activated' then 2 when 'Deactivated' then 3 when 'Removed' then 4
  when 'Outcome changed' then 5 when 'Risk level changed' then 6 when 'Logic changed' then 7
  when 'Priority changed' then 8 else 9 end;

-- 2. Rule Change Detail
-- Every changed rule with its before/after. logic_changed flags a feature_expr edit (the rule's logic).
with {snaps_cte}
select
  change_type,
  rule_id,
  rule_name,
  case when coalesce(base_status, 0) > 0 then 'Active' else 'Inactive/Draft' end as status_before,
  case when coalesce(curr_status, 0) > 0 then 'Active' else 'Inactive/Draft' end as status_after,
  case base_outcome when 1 then 'Punish' when 2 then 'Challenge' when 3 then 'Reject' else cast(base_outcome as string) end as outcome_before,
  case curr_outcome when 1 then 'Punish' when 2 then 'Challenge' when 3 then 'Reject' else cast(curr_outcome as string) end as outcome_after,
  base_risk as risk_before,
  curr_risk as risk_after,
  base_review_priority as review_priority_before,
  curr_review_priority as review_priority_after,
  case when coalesce(base_expr, '') <> coalesce(curr_expr, '') then 'Y' else 'N' end as logic_changed
from classified
where change_type <> 'Unchanged'
order by case change_type
  when 'Added' then 1 when 'Activated' then 2 when 'Deactivated' then 3 when 'Removed' then 4
  when 'Outcome changed' then 5 when 'Risk level changed' then 6 when 'Logic changed' then 7
  when 'Priority changed' then 8 else 9 end, rule_id;

-- 3. Current Rule Inventory
-- The current control surface: active/inactive rule counts by outcome type and risk level.
select
  {curr_outcome_l} as outcome_type,
  {curr_status_l} as rule_status,
  coalesce(nullif(trim(rc.risk_level), ''), 'Unspecified') as risk_level,
  count(1) as rules
from {AF_RULE_CONFIG_TABLE} rc
where {curr_snap}
group by {curr_outcome_l}, {curr_status_l}, coalesce(nullif(trim(rc.risk_level), ''), 'Unspecified')
order by outcome_type, rule_status, risk_level;
"""


def build_af_facial_verification_sql(*, snapshot_pt_date: str | None = None, now: datetime | None = None) -> str:
    # Facial verification is a 3-step funnel per check: liveness (liveness_check_result LC_*) -> selfie
    # anti-spoofing QC (selfie_qc_anti_spoofing_result SQA_*) -> facial matching (facial_matching_result
    # FM_*). Each step only runs if the prior one passed, so the QC/match columns are blank when liveness
    # failed. Scope: checks created in the last two full months + current MTD (create_datetime). The _df
    # table is a cumulative daily snapshot, so the latest pt_date is pinned and rows are date-scoped.
    win = business_insights_window(now)
    span_label = win.span_label
    span_start_iso = win.span_start.isoformat()
    span_end_iso = win.span_end_exclusive.isoformat()
    p2_iso, p3_iso = win.periods[1][1].isoformat(), win.periods[2][1].isoformat()
    p1_label, p2_label, p3_label = (p[0] for p in win.periods)
    fv_snap = _aliased_snapshot_filter("fv", AF_FACIAL_VERIFICATION_TABLE, snapshot_pt_date)
    window = f"fv.create_datetime >= '{span_start_iso}' and fv.create_datetime < '{span_end_iso}'"
    period_case = (
        f"case when fv.create_datetime < '{p2_iso}' then '{p1_label}' "
        f"when fv.create_datetime < '{p3_iso}' then '{p2_label}' else '{p3_label}' end"
    )
    # Step outcomes.
    liveness_pass = "fv.liveness_check_result = 'LC_SUCCESS'"
    antispoof_pass = "fv.selfie_qc_anti_spoofing_result = 'SQA_SUCCESS'"
    match_pass = "fv.facial_matching_result = 'FM_SUCCESS'"
    spoof_attack = (
        "(fv.liveness_check_result = 'LC_AURORA_SPOOF' "
        "or fv.selfie_qc_anti_spoofing_result in ('SQA_REJECT_FACE_SPOOFING', 'SQA_REJECT_FACE_DEEPFAKE'))"
    )
    deepfake_reject = "fv.selfie_qc_anti_spoofing_result = 'SQA_REJECT_FACE_DEEPFAKE'"
    header = _af_report_header("Anti-fraud PH - Facial Verification / Liveness & Deepfake", snapshot_pt_date)
    return f"""{header}
-- Source: DWD authentication_facial_verification. Scope: checks created in {span_label}. 3-step funnel:
-- liveness (LC_SUCCESS) -> anti-spoofing QC (SQA_SUCCESS) -> facial match (FM_SUCCESS); QC/match are
-- blank when the prior step failed, so pass rates use the prior-step-passed count as the denominator.
-- Spoofing attack = LC_AURORA_SPOOF or an SQA face-spoofing / deepfake reject. deepfake_spoof_score in
-- [-1, 1] (<0 = not scored). All rates are %.

-- 1. Verification Outcome Summary
-- One row per period (last two full months + current MTD). Pass rates are conditional on reaching the
-- step: anti-spoof over liveness-passed, match over anti-spoof-passed; overall = full pass / all checks.
select
  {period_case} as period,
  count(1) as checks,
  count(distinct fv.uid) as distinct_users,
  round(sum(case when {liveness_pass} then 1 else 0 end) / nullif(count(1), 0) * 100, 2) as liveness_pass_rate_pct,
  round(sum(case when {antispoof_pass} then 1 else 0 end) / nullif(sum(case when {liveness_pass} then 1 else 0 end), 0) * 100, 2) as antispoof_pass_rate_pct,
  round(sum(case when {match_pass} then 1 else 0 end) / nullif(sum(case when {antispoof_pass} then 1 else 0 end), 0) * 100, 2) as match_pass_rate_pct,
  round(sum(case when {match_pass} then 1 else 0 end) / nullif(count(1), 0) * 100, 2) as overall_pass_rate_pct,
  sum(case when {spoof_attack} then 1 else 0 end) as spoof_attack_checks,
  sum(case when {deepfake_reject} then 1 else 0 end) as deepfake_reject_checks
from {AF_FACIAL_VERIFICATION_TABLE} fv
where {fv_snap} and {window}
group by {period_case}
order by min(fv.create_datetime);

-- 2. Liveness Result Breakdown
-- Why liveness checks pass or fail (customer friction + liveness spoof attempts).
select
  fv.liveness_check_result,
  count(1) as checks,
  count(distinct fv.uid) as distinct_users,
  round(count(1) / nullif(sum(count(1)) over (), 0) * 100, 2) as share_pct
from {AF_FACIAL_VERIFICATION_TABLE} fv
where {fv_snap} and {window}
group by fv.liveness_check_result
order by checks desc;

-- 3. Anti-Spoofing QC Breakdown
-- Outcomes of the selfie anti-spoofing QC (only runs after liveness passed). Surfaces deepfake /
-- face-spoofing rejects and QC friction (eye-close, blur, dark).
select
  fv.selfie_qc_anti_spoofing_result,
  count(1) as checks,
  count(distinct fv.uid) as distinct_users,
  round(count(1) / nullif(sum(count(1)) over (), 0) * 100, 2) as share_pct
from {AF_FACIAL_VERIFICATION_TABLE} fv
where {fv_snap} and {window} and {liveness_pass}
group by fv.selfie_qc_anti_spoofing_result
order by checks desc;

-- 4. Facial Match Result Breakdown
-- Outcomes of the facial matching step (only runs after anti-spoofing passed), with the average score.
select
  fv.facial_matching_result,
  count(1) as checks,
  count(distinct fv.uid) as distinct_users,
  cast(round(avg(fv.facial_matching_score), 4) as decimal(10, 4)) as avg_match_score,
  round(count(1) / nullif(sum(count(1)) over (), 0) * 100, 2) as share_pct
from {AF_FACIAL_VERIFICATION_TABLE} fv
where {fv_snap} and {window} and {antispoof_pass}
group by fv.facial_matching_result
order by checks desc;

-- 5. Deepfake Score Distribution
-- deepfake_spoof_score banded; spoof_rejects = how many in each band the anti-spoofing QC rejected as
-- deepfake / spoofing. High bands with low reject counts are the monitoring blind spots.
select
  case
    when fv.deepfake_spoof_score < 0 then '0. not scored (<0)'
    when fv.deepfake_spoof_score < 0.2 then '1. 0.0-0.2'
    when fv.deepfake_spoof_score < 0.4 then '2. 0.2-0.4'
    when fv.deepfake_spoof_score < 0.6 then '3. 0.4-0.6'
    when fv.deepfake_spoof_score < 0.8 then '4. 0.6-0.8'
    else '5. 0.8-1.0'
  end as deepfake_score_band,
  count(1) as checks,
  sum(case when {spoof_attack} then 1 else 0 end) as spoof_rejects,
  cast(round(avg(fv.deepfake_spoof_score), 4) as decimal(10, 4)) as avg_score
from {AF_FACIAL_VERIFICATION_TABLE} fv
where {fv_snap} and {window}
group by 1
order by deepfake_score_band;

-- 6. Pass Rates by Scene
-- Where friction and spoofing attacks concentrate. Conditional pass rates as in the summary.
select
  coalesce(nullif(trim(fv.scene_name), ''), 'Unspecified') as scene_name,
  count(1) as checks,
  count(distinct fv.uid) as distinct_users,
  round(sum(case when {liveness_pass} then 1 else 0 end) / nullif(count(1), 0) * 100, 2) as liveness_pass_rate_pct,
  round(sum(case when {antispoof_pass} then 1 else 0 end) / nullif(sum(case when {liveness_pass} then 1 else 0 end), 0) * 100, 2) as antispoof_pass_rate_pct,
  round(sum(case when {match_pass} then 1 else 0 end) / nullif(count(1), 0) * 100, 2) as overall_pass_rate_pct,
  sum(case when {spoof_attack} then 1 else 0 end) as spoof_attack_checks
from {AF_FACIAL_VERIFICATION_TABLE} fv
where {fv_snap} and {window}
group by coalesce(nullif(trim(fv.scene_name), ''), 'Unspecified')
order by checks desc;

-- 7. Fraud Review Outcomes
-- Post-hoc fraud review of facial-verification checks (fraud_review_result is a numeric verdict code;
-- non-empty fraud_review_status = the check was pulled for review).
select
  coalesce(nullif(trim(fv.fraud_review_status), ''), 'Not reviewed') as fraud_review_status,
  coalesce(nullif(trim(fv.fraud_review_result), ''), '-') as fraud_review_result,
  count(1) as checks,
  count(distinct fv.uid) as distinct_users
from {AF_FACIAL_VERIFICATION_TABLE} fv
where {fv_snap} and {window}
group by coalesce(nullif(trim(fv.fraud_review_status), ''), 'Not reviewed'),
  coalesce(nullif(trim(fv.fraud_review_result), ''), '-')
order by checks desc;

-- 8. Daily Verification Trend
-- Daily check volume, spoof-attack volume, and liveness failures; powers the filterable trend chart.
with base as (
  select substr(fv.create_datetime, 1, 10) as check_date,
    case when {spoof_attack} then 1 else 0 end as is_spoof,
    case when {liveness_pass} then 0 else 1 end as is_liveness_fail
  from {AF_FACIAL_VERIFICATION_TABLE} fv
  where {fv_snap} and {window}
)
select 'Checks' as series, check_date, count(1) as checks from base group by check_date
union all
select 'Spoof attacks' as series, check_date, sum(is_spoof) as checks from base group by check_date
union all
select 'Liveness failures' as series, check_date, sum(is_liveness_fail) as checks from base group by check_date
order by series, check_date;
"""


def build_af_device_risk_sql(*, snapshot_pt_date: str | None = None, now: datetime | None = None) -> str:
    # Device & identity risk from the extended action log (event-level). Account farming is recomputed
    # as count(distinct uid) per device/account because the precomputed distinct_uid_num_link_* columns
    # are encrypted to a constant. Flag prevalence treats any value that is not a known-negative
    # (''/'0'/their md5 forms) as a positive, so it counts both clear ('1') and encrypted flags.
    # Scope: rolling last 7 days (the table has 64k+ daily partitions; a snapshot max(pt_date) scan is
    # prohibitively slow, and count(distinct) over months would risk the session TTL).
    win = business_insights_window(now)
    end = win.span_end_exclusive
    start = end - timedelta(days=7)
    pt = f"pt_date >= '{start.isoformat()}' and pt_date < '{end.isoformat()}'"
    neg = "('', '0', 'cfcd208495d565ef66e7dff9f98764da', 'd41d8cd98f00b204e9800998ecf8427e')"
    # Exclude sentinel/anonymous keys ('0', '', their md5 forms) from account-farming so the pre-login /
    # unattributed bucket (uid='0') does not masquerade as one device tied to hundreds of thousands of accounts.
    real_uid = f"uid not in {neg}"
    real_dev = f"deviceuuid not in {neg}"
    ext = AF_ACTION_LOG_EXT_TABLE
    header = _af_report_header("Anti-fraud PH - Device & Identity Risk", snapshot_pt_date)
    return f"""{header}
-- Source: {ext} (event-level device/identity signals). Scope: {start.isoformat()} to {end.isoformat()}
-- (rolling last 7 days). Account farming = one device (deviceuuid) or account (uid) tied to many
-- distinct counterparties. Risk-flag prevalence counts events where the flag is set (clear or encrypted).

-- 1. Account Farming Summary
with dev as (
  select deviceuuid, count(distinct uid) as accounts
  from {ext}
  where {pt} and {real_dev} and {real_uid}
  group by deviceuuid
)
select
  count(1) as devices_seen,
  sum(case when accounts >= 5 then 1 else 0 end) as devices_5plus_accounts,
  sum(case when accounts >= 10 then 1 else 0 end) as devices_10plus_accounts,
  max(accounts) as max_accounts_on_one_device,
  cast(round(sum(case when accounts >= 5 then 1 else 0 end) / nullif(count(1), 0) * 100, 4) as decimal(20, 4)) as pct_devices_5plus_accounts
from dev;

-- 2. Top Multi-Account Devices
-- Devices used by 5+ distinct accounts in the window - synthetic-identity / account-farming candidates.
select
  deviceuuid,
  count(distinct uid) as distinct_accounts,
  count(distinct scene_name) as distinct_scenes,
  count(1) as events
from {ext}
where {pt} and {real_dev} and {real_uid}
group by deviceuuid
having count(distinct uid) >= 5
order by distinct_accounts desc
limit 300;

-- 3. Multi-Device Accounts
-- Accounts that hop across 5+ distinct devices in the window - account-takeover / device-churn signal.
select
  uid,
  count(distinct deviceuuid) as distinct_devices,
  count(distinct scene_name) as distinct_scenes,
  count(1) as events
from {ext}
where {pt} and {real_uid} and {real_dev}
group by uid
having count(distinct deviceuuid) >= 5
order by distinct_devices desc
limit 300;

-- 4. Risk Signal Prevalence
-- One wide row: events flagged for each device-risk signal. Flag set = value not in known-negatives.
select
  count(1) as total_events,
  sum(rooted) as rooted, sum(emulator) as emulator, sum(vpn) as vpn, sum(http_proxy) as http_proxy,
  sum(gps_modified) as gps_modified, sum(fake_identity) as fake_identity, sum(fake_deviceinfo) as fake_deviceinfo,
  sum(illegal_imei) as illegal_imei, sum(new_deviceid) as new_deviceid, sum(magisk) as magisk,
  sum(system_debuggable) as system_debuggable, sum(risk_app_root) as risk_app_root,
  sum(risk_app_vpn) as risk_app_vpn, sum(risk_app_fake_gps) as risk_app_fake_gps,
  sum(risk_app_hook) as risk_app_hook, sum(remote_control) as remote_control, sum(autoclicker) as autoclicker
from (
  select
    case when is_root not in {neg} then 1 else 0 end as rooted,
    case when is_emulator not in {neg} then 1 else 0 end as emulator,
    case when is_vpn not in {neg} then 1 else 0 end as vpn,
    case when is_http_proxy not in {neg} then 1 else 0 end as http_proxy,
    case when is_gps_modified not in {neg} then 1 else 0 end as gps_modified,
    case when is_fake_identity not in {neg} then 1 else 0 end as fake_identity,
    case when is_fake_deviceinfo not in {neg} then 1 else 0 end as fake_deviceinfo,
    case when is_illegal_imei not in {neg} then 1 else 0 end as illegal_imei,
    case when is_new_deviceid not in {neg} then 1 else 0 end as new_deviceid,
    case when is_running_magisk not in {neg} then 1 else 0 end as magisk,
    case when is_system_debuggable not in {neg} then 1 else 0 end as system_debuggable,
    case when is_risk_app_install_root not in {neg} then 1 else 0 end as risk_app_root,
    case when is_risk_app_install_vpn not in {neg} then 1 else 0 end as risk_app_vpn,
    case when is_risk_app_install_fake_gps not in {neg} then 1 else 0 end as risk_app_fake_gps,
    case when is_risk_app_install_hook not in {neg} then 1 else 0 end as risk_app_hook,
    case when is_remote_control_acc_enable not in {neg} then 1 else 0 end as remote_control,
    case when is_autoclicker_enable not in {neg} then 1 else 0 end as autoclicker
  from {ext} where {pt}
) f;

-- 5. Risk Signals by Scene
select
  scene_name,
  count(1) as events,
  count(distinct case when {real_uid} then uid end) as distinct_accounts,
  sum(case when (is_root not in {neg} or is_emulator not in {neg} or is_vpn not in {neg}
    or is_http_proxy not in {neg} or is_gps_modified not in {neg} or is_fake_identity not in {neg}
    or is_illegal_imei not in {neg} or is_risk_app_install_root not in {neg}) then 1 else 0 end) as risky_events
from {ext}
where {pt} and scene_name is not null and trim(scene_name) <> ''
group by scene_name
order by risky_events desc, events desc
limit 100;

-- 6. Multi-Account Device Trend
select 'Multi-account devices (>=5)' as series, event_date, multi_account_devices as value
from (
  select pt_date as event_date, count(1) as multi_account_devices from (
    select pt_date, deviceuuid, count(distinct uid) as accounts
    from {ext}
    where {pt} and {real_dev} and {real_uid}
    group by pt_date, deviceuuid
    having count(distinct uid) >= 5
  ) d group by pt_date
) t
order by event_date;
"""


def build_af_card_3ds_sql(*, snapshot_pt_date: str | None = None, now: datetime | None = None) -> str:
    # Card-not-present fraud controls: 3DS (ACS) authentication outcomes + risk-based-auth scoring, plus
    # card fraud cases by modus operandi. trans_status: Y=authenticated, N=not authenticated,
    # C=challenge, R=rejected, U=unavailable, I=info. Amounts are minor units (/100 = PHP).
    # Scope: last two full months + current MTD (pt_date = event day, partition-pruned).
    win = business_insights_window(now)
    span_start = win.span_start.isoformat()
    span_end = win.span_end_exclusive.isoformat()
    (l1, _s1, _e1), (l2, p2, _e2), (l3, p3, _e3) = win.periods
    p2_iso, p3_iso = p2.isoformat(), p3.isoformat()
    trans, cf = AF_THREEDS_TRANS_TABLE, AF_CARD_FRAUD_CASE_TABLE
    pt = f"pt_date >= '{span_start}' and pt_date < '{span_end}'"
    period_expr = (
        f"case when pt_date < '{p2_iso}' then '{l1}' when pt_date < '{p3_iso}' then '{l2}' else '{l3}' end"
    )
    header = _af_report_header("Anti-fraud PH - Card Fraud & 3DS Authentication", snapshot_pt_date)
    return f"""{header}
-- Source: {trans} (3DS ACS auth), {cf} (card fraud cases).
-- Scope: {win.span_label}. trans_status Y=authenticated / N=not auth / C=challenge / R=rejected / U=unavailable.
-- purchase_amount is in minor units; /100 approximates PHP.

-- 1. 3DS Authentication Summary
-- One row per period (last two full months + current MTD).
select
  {period_expr} as period,
  count(1) as threeds_txns,
  sum(case when trans_status = 'Y' then 1 else 0 end) as authenticated,
  sum(case when trans_status = 'N' then 1 else 0 end) as not_authenticated,
  sum(case when trans_status = 'C' then 1 else 0 end) as challenged,
  sum(case when trans_status = 'R' then 1 else 0 end) as rejected,
  sum(case when trans_status = 'U' then 1 else 0 end) as unavailable,
  cast(round(sum(case when trans_status = 'Y' then 1 else 0 end) / nullif(count(1), 0) * 100, 2) as decimal(20, 2)) as auth_rate_pct,
  cast(round(sum(case when trans_status = 'C' then 1 else 0 end) / nullif(count(1), 0) * 100, 2) as decimal(20, 2)) as challenge_rate_pct
from {trans}
where {pt}
group by {period_expr}
order by min(pt_date);

-- 2. Outcome by Auth Status
select
  case trans_status when 'Y' then 'Authenticated' when 'N' then 'Not authenticated' when 'C' then 'Challenge'
    when 'R' then 'Rejected' when 'U' then 'Unavailable' when 'I' then 'Info only' when 'A' then 'Attempted'
    else concat('status_', coalesce(nullif(trim(trans_status), ''), '(blank)')) end as auth_status,
  count(1) as threeds_txns,
  cast(round(count(1) * 100.0 / sum(count(1)) over (), 2) as decimal(20, 2)) as share_pct
from {trans}
where {pt}
group by 1
order by threeds_txns desc;

-- 3. Frictionless vs Challenge
-- By the 3DS Requestor challenge indicator: how often a challenge was requested vs frictionless flow.
select
  case threeds_requestor_chlg_ind
    when '01' then '01 No preference' when '02' then '02 No challenge requested'
    when '03' then '03 Challenge requested' when '04' then '04 Challenge mandated'
    when '05' then '05 No challenge (data share)' when '06' then '06 No challenge (risk analysis)'
    else concat('ind_', coalesce(nullif(trim(threeds_requestor_chlg_ind), ''), '(blank)')) end as challenge_indicator,
  count(1) as threeds_txns,
  sum(case when trans_status = 'C' then 1 else 0 end) as resulted_in_challenge,
  sum(case when trans_status = 'Y' then 1 else 0 end) as authenticated
from {trans}
where {pt}
group by 1
order by threeds_txns desc;

-- 4. 3DS by Merchant Category (MCC)
select
  coalesce(nullif(trim(mcc), ''), '(none)') as mcc,
  count(1) as threeds_txns,
  sum(case when trans_status = 'Y' then 1 else 0 end) as authenticated,
  cast(round(sum(case when trans_status = 'Y' then 1 else 0 end) / nullif(count(1), 0) * 100, 2) as decimal(20, 2)) as auth_rate_pct,
  cast(round(sum(cast(purchase_amount as double)) / 100, 2) as decimal(20, 2)) as purchase_amount_php
from {trans}
where {pt}
group by 1
order by threeds_txns desc
limit 100;

-- 5. Card Fraud Cases by MO
select
  coalesce(nullif(trim(mo_reason), ''), 'Unspecified') as mo_reason,
  coalesce(nullif(trim(regexp_replace(translate(sub_mo_reason, '[]"', ''), ',', ', ')), ''), 'Unspecified') as sub_mo_reason,
  count(distinct case_id) as cases
from {cf}
where {pt}
group by 1, 2
order by cases desc
limit 100;

-- 6. Daily 3DS Trend
with d as (
  select pt_date,
    sum(case when trans_status = 'Y' then 1 else 0 end) as authenticated,
    sum(case when trans_status = 'C' then 1 else 0 end) as challenged,
    sum(case when trans_status = 'R' then 1 else 0 end) as rejected
  from {trans} where {pt} group by pt_date
)
select 'Authenticated' as series, pt_date as txn_date, authenticated as txns from d
union all
select 'Challenge' as series, pt_date as txn_date, challenged as txns from d
union all
select 'Rejected' as series, pt_date as txn_date, rejected as txns from d
order by series, txn_date;
"""


def _normalize_header(value: Any) -> str:
    normalized = re.sub(r"[^a-zA-Z0-9]+", "_", str(value or "").strip().lower()).strip("_")
    return normalized


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def product_label(value: Any) -> str:
    raw = _display_value(value)
    if not raw:
        return "UNKNOWN"
    return PRODUCT_LABELS.get(raw, raw)


def _read_csv_export(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    rows: list[dict[str, Any]] = []
    for row in reader:
        rows.append({_normalize_header(key): value for key, value in (row or {}).items() if key is not None})
    return rows


def _read_xlsx_export(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if not workbook.worksheets:
        return []
    sheet = workbook.worksheets[0]
    raw_rows = list(sheet.iter_rows(values_only=True))
    if not raw_rows:
        return []
    headers = [_normalize_header(value) for value in raw_rows[0]]
    rows: list[dict[str, Any]] = []
    for raw_row in raw_rows[1:]:
        item = {
            headers[index]: value
            for index, value in enumerate(raw_row)
            if index < len(headers) and headers[index]
        }
        if any(_display_value(value) for value in item.values()):
            rows.append(item)
    return rows


def read_underwriting_export(content: bytes, filename: str) -> list[dict[str, Any]]:
    suffix = Path(filename or "").suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx_export(content)
    if suffix in {".csv", ".txt"}:
        return _read_csv_export(content)
    raise ToolError("Upload a Data Workbench export as .csv or .xlsx.")


def _parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.astimezone(BUSINESS_INSIGHTS_TIMEZONE) if value.tzinfo else value.replace(tzinfo=BUSINESS_INSIGHTS_TIMEZONE)
    if isinstance(value, date):
        return datetime.combine(value, time.min, tzinfo=BUSINESS_INSIGHTS_TIMEZONE)
    numeric = _parse_number(value)
    if numeric is not None:
        if numeric > 10_000_000_000:
            return datetime.fromtimestamp(numeric / 1000, tz=BUSINESS_INSIGHTS_TIMEZONE)
        if numeric > 100_000_000:
            return datetime.fromtimestamp(numeric, tz=BUSINESS_INSIGHTS_TIMEZONE)
    text = str(value).strip()
    for candidate in (text, text.replace("Z", "+00:00")):
        try:
            parsed = datetime.fromisoformat(candidate)
            return parsed.astimezone(BUSINESS_INSIGHTS_TIMEZONE) if parsed.tzinfo else parsed.replace(tzinfo=BUSINESS_INSIGHTS_TIMEZONE)
        except ValueError:
            continue
    for pattern in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            parsed = datetime.strptime(text, pattern)
            return parsed.replace(tzinfo=BUSINESS_INSIGHTS_TIMEZONE)
        except ValueError:
            continue
    return None


def _row_text(row: dict[str, Any], *keys: str, default: str = "") -> str:
    for key in keys:
        value = row.get(key)
        text = _display_value(value)
        if text:
            return text
    return default


def _row_datetime(row: dict[str, Any]) -> datetime | None:
    for key in ("application_submission_time", "create_date", "modify_date"):
        parsed = _parse_datetime(row.get(key))
        if parsed is not None:
            return parsed
    return None


def _status_bucket(status: str) -> str:
    normalized = status.strip().upper()
    if not normalized:
        return "PENDING"
    if any(token in normalized for token in ("APPROV", "PASS", "SUCCESS", "COMPLETED")):
        return "APPROVED"
    if any(token in normalized for token in ("REJECT", "DECLIN", "FAIL", "DENY")):
        return "REJECTED"
    return "PENDING"


def _period_label(row: dict[str, Any], period: BusinessInsightsPeriod) -> str:
    explicit = _row_text(row, "report_period")
    if explicit:
        return explicit
    row_date = _row_datetime(row)
    if row_date and row_date.date() >= period.current_month_start:
        return period.current_month_label
    return period.previous_month_label


def _pct(part: int | float, total: int | float) -> float:
    if not total:
        return 0.0
    return round(float(part) / float(total), 4)


def _append_rows(sheet: Any, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    _style_sheet(sheet)


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    header_font = Font(bold=True, color="1F2937")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, column_cells in enumerate(sheet.columns, start=1):
        width = 12
        for cell in column_cells:
            width = max(width, min(len(_display_value(cell.value)) + 2, 42))
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def build_underwriting_funnel_workbook(
    rows: list[dict[str, Any]],
    *,
    now: datetime | None = None,
) -> bytes:
    if not rows:
        raise ToolError("The uploaded export has no data rows.")
    period = business_insights_period(now)
    active_now = now.astimezone(BUSINESS_INSIGHTS_TIMEZONE) if now else datetime.now(BUSINESS_INSIGHTS_TIMEZONE)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    summary: dict[tuple[str, str], dict[str, Any]] = {}
    funnel: dict[tuple[str, str, str], int] = {}
    reject_reasons: dict[tuple[str, str, str], int] = {}
    stage_backlog: dict[tuple[str, str, str], dict[str, Any]] = {}
    subproduct_funnel: dict[tuple[str, str, str, str], int] = {}

    normalized_rows: list[dict[str, Any]] = []
    for raw_row in rows:
        row = {_normalize_header(key): value for key, value in raw_row.items()}
        product = _row_text(row, "product_code", default="UNKNOWN")
        sub_product = _row_text(row, "sub_product_code", default="-") or "-"
        status = _row_text(row, "underwriting_status", default="PENDING").upper()
        status_bucket = _status_bucket(status)
        report_period = _period_label(row, period)
        amount = _parse_number(row.get("apply_loan_amount"))
        row_dt = _row_datetime(row)
        row_date = row_dt.date().isoformat() if row_dt else ""
        reason = _row_text(row, "reject_reason", default="Unspecified")
        stage = _row_text(row, "current_stage", "step", default="Unspecified")

        summary_item = summary.setdefault(
            (report_period, product),
            {"applications": 0, "approved": 0, "rejected": 0, "pending": 0, "amount_sum": 0.0, "amount_count": 0},
        )
        summary_item["applications"] += 1
        summary_item[status_bucket.lower()] += 1
        if amount is not None:
            summary_item["amount_sum"] += amount
            summary_item["amount_count"] += 1

        funnel[(report_period, product, status or "PENDING")] = funnel.get((report_period, product, status or "PENDING"), 0) + 1
        if status_bucket == "REJECTED":
            reject_reasons[(report_period, product, reason)] = reject_reasons.get((report_period, product, reason), 0) + 1
        stage_item = stage_backlog.setdefault(
            (report_period, product, stage),
            {"count": 0, "oldest": None, "age_days_sum": 0.0, "age_count": 0},
        )
        stage_item["count"] += 1
        if row_dt is not None:
            oldest = stage_item["oldest"]
            stage_item["oldest"] = row_dt if oldest is None or row_dt < oldest else oldest
            stage_item["age_days_sum"] += max(0.0, (active_now - row_dt).total_seconds() / 86400)
            stage_item["age_count"] += 1
        subproduct_funnel[(report_period, product, sub_product, status or "PENDING")] = (
            subproduct_funnel.get((report_period, product, sub_product, status or "PENDING"), 0) + 1
        )

        normalized_row = {"report_period": report_period, "application_date": row_date}
        normalized_row.update(row)
        normalized_rows.append(normalized_row)

    summary_rows = []
    for (report_period, product), item in sorted(summary.items()):
        applications = item["applications"]
        avg_amount = round(item["amount_sum"] / item["amount_count"], 2) if item["amount_count"] else ""
        summary_rows.append(
            [
                report_period,
                product,
                applications,
                item["approved"],
                item["rejected"],
                item["pending"],
                _pct(item["approved"], applications),
                avg_amount,
            ]
        )
    _append_rows(
        workbook.create_sheet("Summary by Product"),
        ["Period", "Product", "Applications", "Approved", "Rejected", "Pending", "Approval Rate", "Avg Applied Amount"],
        summary_rows,
    )

    product_totals = {(period_label, product): sum(count for (p, prod, _status), count in funnel.items() if p == period_label and prod == product) for period_label, product, _status in funnel}
    funnel_rows = [
        [report_period, product, status, count, _pct(count, product_totals.get((report_period, product), 0))]
        for (report_period, product, status), count in sorted(funnel.items())
    ]
    _append_rows(workbook.create_sheet("Product Funnel"), ["Period", "Product", "Status", "Count", "% Within Product"], funnel_rows)

    rejected_totals = {
        (period_label, product): sum(count for (p, prod, _reason), count in reject_reasons.items() if p == period_label and prod == product)
        for period_label, product, _reason in reject_reasons
    }
    reason_rows = [
        [report_period, product, reason, count, _pct(count, rejected_totals.get((report_period, product), 0))]
        for (report_period, product, reason), count in sorted(reject_reasons.items())
    ]
    _append_rows(workbook.create_sheet("Product Reject Reasons"), ["Period", "Product", "Reject Reason", "Count", "% of Product Rejections"], reason_rows)

    stage_rows = []
    for (report_period, product, stage), item in sorted(stage_backlog.items()):
        oldest_dt = item["oldest"]
        avg_age = round(item["age_days_sum"] / item["age_count"], 2) if item["age_count"] else ""
        stage_rows.append([report_period, product, stage, item["count"], oldest_dt.date().isoformat() if oldest_dt else "", avg_age])
    _append_rows(workbook.create_sheet("Product Stage Backlog"), ["Period", "Product", "Current Stage", "Count", "Oldest Create Date", "Avg Age Days"], stage_rows)

    subproduct_totals = {
        (period_label, product, sub_product): sum(
            count
            for (p, prod, sub, _status), count in subproduct_funnel.items()
            if p == period_label and prod == product and sub == sub_product
        )
        for period_label, product, sub_product, _status in subproduct_funnel
    }
    subproduct_rows = [
        [report_period, product, sub_product, status, count, _pct(count, subproduct_totals.get((report_period, product, sub_product), 0))]
        for (report_period, product, sub_product, status), count in sorted(subproduct_funnel.items())
    ]
    _append_rows(
        workbook.create_sheet("Sub-product Funnel"),
        ["Period", "Product", "Sub-product", "Status", "Count", "% Within Sub-product"],
        subproduct_rows,
    )

    raw_headers = ["report_period", "application_date", *UNDERWRITING_FUNNEL_FIELDS]
    extra_headers = sorted({key for row in normalized_rows for key in row.keys()} - set(raw_headers))
    raw_sheet = workbook.create_sheet("Raw Export")
    _append_rows(raw_sheet, [*raw_headers, *extra_headers], [[_display_value(row.get(header)) for header in [*raw_headers, *extra_headers]] for row in normalized_rows])

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


class BusinessInsightsStore:
    def __init__(self, root_dir: Path) -> None:
        self.root_dir = Path(root_dir)
        self.metadata_path = self.root_dir / "reports.json"
        self.artifacts_dir = self.root_dir / "artifacts"
        self._lock = threading.Lock()

    def _load(self) -> dict[str, Any]:
        if not self.metadata_path.exists():
            return {"artifacts": {}}
        try:
            payload = json.loads(self.metadata_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {"artifacts": {}}
        return payload if isinstance(payload, dict) else {"artifacts": {}}

    def _persist_locked(self, payload: dict[str, Any]) -> None:
        self.root_dir.mkdir(parents=True, exist_ok=True)
        temp_path = self.metadata_path.with_name(f".{self.metadata_path.name}.{os.getpid()}.tmp")
        temp_path.write_text(json.dumps(payload, ensure_ascii=False, sort_keys=True, indent=2), encoding="utf-8")
        os.replace(temp_path, self.metadata_path)

    def domains(self) -> list[dict[str, str]]:
        return [dict(item) for item in BUSINESS_INSIGHTS_DOMAINS]

    def reports(self, domain: str = "") -> list[dict[str, Any]]:
        domain_key = str(domain or "").strip().lower()
        payload = self._load()
        artifacts = payload.get("artifacts") if isinstance(payload.get("artifacts"), dict) else {}
        reports: list[dict[str, Any]] = []
        for report in SEEDED_REPORTS:
            if domain_key and report["domain"] != domain_key:
                continue
            item = dict(report)
            artifact = artifacts.get(report["id"]) if isinstance(artifacts.get(report["id"]), dict) else None
            item["artifact"] = dict(artifact) if artifact else None
            reports.append(item)
        return reports

    def report(self, report_id: str) -> dict[str, Any] | None:
        normalized_id = str(report_id or "").strip()
        return next((report for report in self.reports() if report["id"] == normalized_id), None)

    def sql_for_report(self, report_id: str, *, now: datetime | None = None) -> str:
        report = self.report(report_id) or {}
        if not report:
            raise ToolError("SQL is not configured for this report yet.")
        artifact = report.get("artifact") if isinstance(report.get("artifact"), dict) else {}
        saved_sql = str(artifact.get("sql") or "").strip() if isinstance(artifact, dict) else ""
        if saved_sql:
            return saved_sql
        builders = {
            UNDERWRITING_FUNNEL_REPORT_ID: lambda: build_underwriting_funnel_mis_sql(now=now),
            PORTFOLIO_REPAYMENT_REPORT_ID: lambda: build_portfolio_repayment_sql(now=now),
            LIMIT_UTILIZATION_REPORT_ID: lambda: build_limit_utilization_sql(now=now),
            APPLICATION_DISBURSEMENT_FUNNEL_REPORT_ID: lambda: build_application_disbursement_funnel_sql(now=now),
            AF_SCENARIOS_ACTIONS_REPORT_ID: lambda: build_af_scenarios_actions_sql(now=now),
            AF_RULES_FEATURES_REPORT_ID: lambda: build_af_rules_features_sql(now=now),
            AF_RULE_EFFECTIVENESS_REPORT_ID: lambda: build_af_rule_effectiveness_sql(now=now),
            AF_FRAUD_LOSS_REPORT_ID: lambda: build_af_fraud_loss_sql(now=now),
            AF_FACIAL_VERIFICATION_REPORT_ID: lambda: build_af_facial_verification_sql(now=now),
            AF_DEVICE_RISK_REPORT_ID: lambda: build_af_device_risk_sql(now=now),
            AF_CARD_3DS_REPORT_ID: lambda: build_af_card_3ds_sql(now=now),
        }
        builder = builders.get(report_id)
        if builder is None:
            raise ToolError("SQL is not configured for this report yet.")
        return builder()

    def sql_filename_for_report(self, report_id: str) -> str:
        if self.report(report_id) is None:
            raise ToolError("SQL is not configured for this report yet.")
        return f"{report_id}.sql"

    def save_underwriting_export(self, *, content: bytes, filename: str, now: datetime | None = None) -> dict[str, Any]:
        rows = read_underwriting_export(content, filename)
        workbook_bytes = build_underwriting_funnel_workbook(rows, now=now)
        artifact_id = uuid.uuid4().hex
        safe_filename = f"{UNDERWRITING_FUNNEL_REPORT_ID}-{artifact_id[:8]}.xlsx"
        self.artifacts_dir.mkdir(parents=True, exist_ok=True)
        artifact_path = self.artifacts_dir / safe_filename
        artifact_path.write_bytes(workbook_bytes)
        metadata = {
            "id": artifact_id,
            "report_id": UNDERWRITING_FUNNEL_REPORT_ID,
            "filename": safe_filename,
            "source_filename": Path(filename or "export").name,
            "row_count": len(rows),
            "created_at": time_module.strftime("%Y-%m-%dT%H:%M:%SZ", time_module.gmtime()),
            "sql": build_underwriting_funnel_sql(now),
        }
        with self._lock:
            payload = self._load()
            artifacts = payload.get("artifacts") if isinstance(payload.get("artifacts"), dict) else {}
            artifacts[UNDERWRITING_FUNNEL_REPORT_ID] = metadata
            payload["artifacts"] = artifacts
            self._persist_locked(payload)
        return dict(metadata)

    def artifact_path(self, artifact_id: str) -> tuple[dict[str, Any], Path]:
        payload = self._load()
        artifacts = payload.get("artifacts") if isinstance(payload.get("artifacts"), dict) else {}
        for metadata in artifacts.values():
            if not isinstance(metadata, dict):
                continue
            if str(metadata.get("id") or "") == str(artifact_id or ""):
                path = (self.artifacts_dir / str(metadata.get("filename") or "")).resolve()
                root = self.artifacts_dir.resolve()
                if root not in path.parents or not path.exists():
                    raise ToolError("Business Insights artifact was not found.")
                return dict(metadata), path
        raise ToolError("Business Insights artifact was not found.")

    def visualization_path(self, artifact_id: str) -> tuple[dict[str, Any], Path]:
        payload = self._load()
        artifacts = payload.get("artifacts") if isinstance(payload.get("artifacts"), dict) else {}
        for metadata in artifacts.values():
            if not isinstance(metadata, dict):
                continue
            if str(metadata.get("id") or "") == str(artifact_id or ""):
                path = (self.artifacts_dir / str(metadata.get("visualization_filename") or "")).resolve()
                root = self.artifacts_dir.resolve()
                if root not in path.parents or not path.exists():
                    raise ToolError("Business Insights visualization was not found.")
                return dict(metadata), path
        raise ToolError("Business Insights visualization was not found.")
