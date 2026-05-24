from __future__ import annotations

import hashlib
import io
import json
import mimetypes
import os
from pathlib import Path
import re
import threading
import time
from typing import Any
import uuid
import zipfile

from bpmis_jira_tool.errors import ToolError
from bpmis_jira_tool.source_code_qa import ALL_COUNTRY, SourceCodeQAService
from bpmis_jira_tool.source_code_qa_sql_artifacts import format_source_code_qa_sql_text as _format_source_code_qa_sql_text


class SourceCodeQASessionStore:
    def __init__(self, storage_path: Path | None = None) -> None:
        self.storage_path = storage_path
        self._sessions: dict[str, dict[str, Any]] = self._load()
        self._lock = threading.Lock()

    def _load(self) -> dict[str, dict[str, Any]]:
        if self.storage_path is None or not self.storage_path.exists():
            return {}
        try:
            payload = json.loads(self.storage_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
        sessions: dict[str, dict[str, Any]] = {}
        for session_id, raw_session in (payload.get("sessions") or {}).items():
            if isinstance(raw_session, dict):
                sessions[str(session_id)] = raw_session
        return sessions

    def _persist_locked(self) -> None:
        if self.storage_path is None:
            return
        try:
            self.storage_path.parent.mkdir(parents=True, exist_ok=True)
            payload = {
                "updated_at": time.time(),
                "sessions": self._sessions,
            }
            temp_path = self.storage_path.with_name(f".{self.storage_path.name}.{os.getpid()}.tmp")
            temp_path.write_text(json.dumps(payload, ensure_ascii=False, sort_keys=True), encoding="utf-8")
            os.replace(temp_path, self.storage_path)
        except OSError:
            return

    @staticmethod
    def _now() -> str:
        return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

    @staticmethod
    def _title_from_question(question: str) -> str:
        normalized = re.sub(r"\s+", " ", str(question or "").strip())
        if not normalized:
            return "New Source Code Chat"
        return normalized[:72] + ("..." if len(normalized) > 72 else "")

    def create(
        self,
        *,
        owner_email: str,
        pm_team: str,
        country: str,
        llm_provider: str,
        title: str = "",
    ) -> dict[str, Any]:
        now = self._now()
        session_payload = {
            "id": uuid.uuid4().hex,
            "owner_email": str(owner_email or "").strip().lower(),
            "title": str(title or "").strip() or "New Source Code Chat",
            "pm_team": str(pm_team or "").strip() or "AF",
            "country": str(country or "").strip() or ALL_COUNTRY,
            "llm_provider": SourceCodeQAService.normalize_query_llm_provider(llm_provider),
            "created_at": now,
            "updated_at": now,
            "messages": [],
            "last_context": None,
            "last_trace_id": "",
            "archived_at": "",
            "archived_by": "",
        }
        with self._lock:
            self._sessions[session_payload["id"]] = session_payload
            self._persist_locked()
            return dict(session_payload)

    def list(self, *, owner_email: str, limit: int = 30) -> list[dict[str, Any]]:
        owner = str(owner_email or "").strip().lower()
        with self._lock:
            sessions = [
                self._public_session(session_payload, include_messages=False)
                for session_payload in self._sessions.values()
                if str(session_payload.get("owner_email") or "").strip().lower() == owner
                and not str(session_payload.get("archived_at") or "").strip()
            ]
        sessions.sort(key=lambda item: str(item.get("updated_at") or ""), reverse=True)
        return sessions[: max(1, min(int(limit or 30), 100))]

    def get(self, session_id: str, *, owner_email: str) -> dict[str, Any] | None:
        owner = str(owner_email or "").strip().lower()
        with self._lock:
            session_payload = self._sessions.get(str(session_id or ""))
            if not session_payload:
                return None
            if str(session_payload.get("owner_email") or "").strip().lower() != owner:
                return None
            return self._public_session(session_payload, include_messages=True)

    def archive(self, session_id: str, *, owner_email: str) -> dict[str, Any] | None:
        owner = str(owner_email or "").strip().lower()
        now = self._now()
        with self._lock:
            session_payload = self._sessions.get(str(session_id or ""))
            if not session_payload:
                return None
            if str(session_payload.get("owner_email") or "").strip().lower() != owner:
                return None
            session_payload["archived_at"] = now
            session_payload["archived_by"] = owner
            session_payload["updated_at"] = now
            self._persist_locked()
            return {
                "status": "ok",
                "session_id": session_payload.get("id") or "",
                "archived_at": now,
            }

    def get_context(self, session_id: str, *, owner_email: str) -> dict[str, Any] | None:
        session_payload = self.get(session_id, owner_email=owner_email)
        context = session_payload.get("last_context") if session_payload else None
        if not isinstance(context, dict):
            return None
        enriched = dict(context)
        enriched.setdefault("session_title", session_payload.get("title") or "")
        return enriched

    @staticmethod
    def _recent_turn_from_context(context: dict[str, Any] | None) -> dict[str, Any] | None:
        if not isinstance(context, dict):
            return None
        question = str(context.get("question") or "").strip()
        answer = str(context.get("answer") or context.get("rendered_answer") or "").strip()
        if not question and not answer:
            return None
        llm_route = context.get("llm_route") if isinstance(context.get("llm_route"), dict) else {}
        return {
            "question": question[:500],
            "answer": answer[:1200],
            "summary": str(context.get("summary") or "")[:500],
            "trace_id": str(context.get("trace_id") or "")[:80],
            "attachments": [
                item for item in (context.get("attachments") or [])[:5]
                if isinstance(item, dict)
            ],
            "llm_provider": str(context.get("llm_provider") or "")[:80],
            "llm_model": str(context.get("llm_model") or "")[:120],
            "matches_snapshot": [
                item for item in (context.get("matches_snapshot") or context.get("matches") or [])[:8]
                if isinstance(item, dict)
            ],
            "codex_candidate_paths": [
                item for item in (
                    context.get("codex_candidate_paths")
                    or (llm_route.get("candidate_paths") if isinstance(llm_route, dict) else [])
                    or []
                )[:12]
                if isinstance(item, dict)
            ],
            "evidence_pack": (
                context.get("evidence_pack")
                if isinstance(context.get("evidence_pack"), dict)
                else {}
            ),
        }

    @classmethod
    def _extend_recent_turns(cls, context: dict[str, Any], previous_context: dict[str, Any] | None) -> dict[str, Any]:
        enriched = dict(context or {})
        recent_turns = [
            item for item in (
                previous_context.get("recent_turns", []) if isinstance(previous_context, dict) else []
            )
            if isinstance(item, dict)
        ]
        previous_turn = cls._recent_turn_from_context(previous_context)
        if previous_turn:
            recent_turns.append(previous_turn)
        deduped: list[dict[str, Any]] = []
        seen: set[tuple[str, str]] = set()
        for item in recent_turns:
            key = (str(item.get("question") or ""), str(item.get("trace_id") or ""))
            if key in seen:
                continue
            seen.add(key)
            deduped.append(item)
        try:
            max_turns = int(enriched.get("codex_session_max_turns") or 8)
        except (TypeError, ValueError):
            max_turns = 8
        enriched["recent_turns"] = deduped[-max(1, min(max_turns, 30)):]
        return enriched

    def append_exchange(
        self,
        session_id: str,
        *,
        owner_email: str,
        pm_team: str,
        country: str,
        llm_provider: str,
        question: str,
        result: dict[str, Any],
        context: dict[str, Any],
        attachments: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any] | None:
        owner = str(owner_email or "").strip().lower()
        now = self._now()
        with self._lock:
            session_payload = self._sessions.get(str(session_id or ""))
            if not session_payload:
                return None
            if str(session_payload.get("owner_email") or "").strip().lower() != owner:
                return None
            title = str(session_payload.get("title") or "").strip()
            if not title or title == "New Source Code Chat":
                session_payload["title"] = self._title_from_question(question)
            previous_context = session_payload.get("last_context") if isinstance(session_payload.get("last_context"), dict) else None
            previous_scope = (
                str(session_payload.get("pm_team") or ""),
                str(session_payload.get("country") or ""),
                str(session_payload.get("llm_provider") or ""),
            )
            current_scope = (
                str(pm_team or "").strip() or str(session_payload.get("pm_team") or ""),
                str(country or "").strip() or str(session_payload.get("country") or ALL_COUNTRY),
                SourceCodeQAService.normalize_query_llm_provider(llm_provider),
            )
            session_payload["pm_team"] = current_scope[0] or "AF"
            session_payload["country"] = current_scope[1] or ALL_COUNTRY
            session_payload["llm_provider"] = current_scope[2]
            session_payload["updated_at"] = now
            next_context = self._extend_recent_turns(context, previous_context)
            if previous_scope != current_scope:
                next_context.pop("codex_cli_session", None)
            session_payload["last_context"] = next_context
            session_payload["last_trace_id"] = str(result.get("trace_id") or "")
            messages = list(session_payload.get("messages") or [])
            normalized_question = str(question or "").strip()
            messages = [
                message for message in messages
                if not (
                    isinstance(message, dict)
                    and message.get("role") == "user"
                    and message.get("pending")
                    and str(message.get("text") or "").strip() == normalized_question
                )
            ]
            messages.extend(
                [
                    {
                        "role": "user",
                        "text": normalized_question,
                        "created_at": now,
                        "attachments": [
                            SourceCodeQAAttachmentStore.public_metadata(item)
                            for item in (attachments or [])
                            if isinstance(item, dict)
                        ],
                    },
                    {
                        "role": "assistant",
                        "text": str(result.get("llm_answer") or result.get("summary") or ""),
                        "created_at": now,
                        "payload": _compact_source_code_qa_session_payload(result),
                    },
                ]
            )
            session_payload["messages"] = messages[-80:]
            self._persist_locked()
            return self._public_session(session_payload, include_messages=True)

    def append_pending_question(
        self,
        session_id: str,
        *,
        owner_email: str,
        pm_team: str,
        country: str,
        llm_provider: str,
        question: str,
        job_id: str,
        attachments: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any] | None:
        owner = str(owner_email or "").strip().lower()
        now = self._now()
        normalized_question = str(question or "").strip()
        normalized_job_id = str(job_id or "").strip()
        if not normalized_question or not normalized_job_id:
            return None
        with self._lock:
            session_payload = self._sessions.get(str(session_id or ""))
            if not session_payload:
                return None
            if str(session_payload.get("owner_email") or "").strip().lower() != owner:
                return None
            title = str(session_payload.get("title") or "").strip()
            if not title or title == "New Source Code Chat":
                session_payload["title"] = self._title_from_question(normalized_question)
            session_payload["pm_team"] = str(pm_team or "").strip() or str(session_payload.get("pm_team") or "AF")
            session_payload["country"] = str(country or "").strip() or str(session_payload.get("country") or ALL_COUNTRY)
            session_payload["llm_provider"] = SourceCodeQAService.normalize_query_llm_provider(llm_provider)
            session_payload["updated_at"] = now
            messages = [
                message for message in list(session_payload.get("messages") or [])
                if not (
                    isinstance(message, dict)
                    and message.get("role") == "user"
                    and message.get("pending")
                    and str(message.get("pending_job_id") or "") == normalized_job_id
                )
            ]
            messages.append(
                {
                    "role": "user",
                    "text": normalized_question,
                    "created_at": now,
                    "attachments": [
                        SourceCodeQAAttachmentStore.public_metadata(item)
                        for item in (attachments or [])
                        if isinstance(item, dict)
                    ],
                    "pending": True,
                    "pending_job_id": normalized_job_id,
                }
            )
            session_payload["messages"] = messages[-80:]
            self._persist_locked()
            return self._public_session(session_payload, include_messages=True)

    def _public_session(self, session_payload: dict[str, Any], *, include_messages: bool) -> dict[str, Any]:
        public_payload = {
            "id": session_payload.get("id") or "",
            "title": session_payload.get("title") or "New Source Code Chat",
            "pm_team": session_payload.get("pm_team") or "",
            "country": session_payload.get("country") or ALL_COUNTRY,
            "llm_provider": session_payload.get("llm_provider") or "codex_cli_bridge",
            "created_at": session_payload.get("created_at") or "",
            "updated_at": session_payload.get("updated_at") or "",
            "archived_at": session_payload.get("archived_at") or "",
            "last_context": session_payload.get("last_context") if include_messages else None,
            "last_trace_id": session_payload.get("last_trace_id") or "",
            "message_count": len(session_payload.get("messages") or []),
        }
        if include_messages:
            if isinstance(public_payload["last_context"], dict):
                public_payload["last_context"] = {
                    **public_payload["last_context"],
                    "session_title": public_payload["title"],
                }
            public_payload["messages"] = list(session_payload.get("messages") or [])
        return public_payload


class SourceCodeQAAttachmentStore:
    MAX_FILE_BYTES = 10 * 1024 * 1024
    MAX_ATTACHMENTS = 5
    MAX_IMAGES = 3
    IMAGE_MIME_TYPES = {"image/png", "image/jpeg", "image/webp", "image/gif"}
    TEXT_EXTENSIONS = {
        ".txt",
        ".md",
        ".csv",
        ".json",
        ".xml",
        ".yaml",
        ".yml",
        ".log",
        ".java",
        ".py",
        ".js",
        ".ts",
        ".tsx",
        ".sql",
        ".properties",
        ".kt",
        ".go",
        ".rb",
        ".php",
        ".html",
        ".css",
        ".sh",
    }
    DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".xlsx"}
    BLOCKED_EXTENSIONS = {
        ".app",
        ".bat",
        ".bin",
        ".cmd",
        ".com",
        ".dmg",
        ".exe",
        ".gz",
        ".jar",
        ".pkg",
        ".rar",
        ".tar",
        ".tgz",
        ".zip",
        ".7z",
    }

    def __init__(self, root_dir: Path | None = None) -> None:
        self.root_dir = root_dir
        self._lock = threading.Lock()

    @staticmethod
    def _now() -> str:
        return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

    @staticmethod
    def _owner_key(owner_email: str) -> str:
        owner = str(owner_email or "").strip().lower() or "local"
        return hashlib.sha256(owner.encode("utf-8")).hexdigest()[:24]

    @staticmethod
    def _safe_session_id(session_id: str) -> str:
        normalized = str(session_id or "").strip()
        if not normalized or not re.fullmatch(r"[A-Za-z0-9_-]{8,80}", normalized):
            raise ToolError("A valid Source Code Q&A session is required before uploading attachments.")
        return normalized

    @staticmethod
    def _safe_filename(filename: str) -> str:
        name = Path(str(filename or "attachment")).name.strip().replace("\x00", "")
        name = re.sub(r"[^A-Za-z0-9._ -]+", "_", name)[:180].strip(" .")
        return name or "attachment"

    def _session_dir(self, *, owner_email: str, session_id: str) -> Path:
        if self.root_dir is None:
            raise ToolError("Source Code Q&A attachments are not configured.")
        return self.root_dir / self._owner_key(owner_email) / self._safe_session_id(session_id)

    def _metadata_path(self, *, owner_email: str, session_id: str) -> Path:
        return self._session_dir(owner_email=owner_email, session_id=session_id) / "metadata.json"

    def _load_metadata_locked(self, *, owner_email: str, session_id: str) -> dict[str, dict[str, Any]]:
        path = self._metadata_path(owner_email=owner_email, session_id=session_id)
        if not path.exists():
            return {}
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
        attachments = payload.get("attachments") if isinstance(payload, dict) else {}
        return {str(key): value for key, value in attachments.items() if isinstance(value, dict)} if isinstance(attachments, dict) else {}

    def _persist_metadata_locked(self, *, owner_email: str, session_id: str, metadata: dict[str, dict[str, Any]]) -> None:
        path = self._metadata_path(owner_email=owner_email, session_id=session_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = path.with_name(f".{path.name}.{os.getpid()}.tmp")
        temp_path.write_text(
            json.dumps({"updated_at": self._now(), "attachments": metadata}, ensure_ascii=False, sort_keys=True),
            encoding="utf-8",
        )
        os.replace(temp_path, path)

    def save_bytes(
        self,
        *,
        owner_email: str,
        session_id: str,
        filename: str,
        content: bytes,
        mime_type: str = "",
    ) -> dict[str, Any]:
        if len(content or b"") <= 0:
            raise ToolError("Attachment file is empty.")
        if len(content) > self.MAX_FILE_BYTES:
            raise ToolError("Attachment is too large. Maximum size is 10MB per file.")
        safe_name = self._safe_filename(filename)
        suffix = Path(safe_name).suffix.lower()
        if suffix in self.BLOCKED_EXTENSIONS:
            raise ToolError("Executable, archive, and unknown binary attachments are not supported.")
        guessed_mime = str(mime_type or mimetypes.guess_type(safe_name)[0] or "").lower()
        kind = self._attachment_kind(safe_name, guessed_mime, content)
        digest = hashlib.sha256(content).hexdigest()
        attachment_id = uuid.uuid4().hex
        stored_name = f"{attachment_id}{suffix or '.bin'}"
        session_dir = self._session_dir(owner_email=owner_email, session_id=session_id)
        metadata = {
            "id": attachment_id,
            "filename": safe_name,
            "stored_name": stored_name,
            "mime_type": guessed_mime or "application/octet-stream",
            "kind": kind,
            "size": len(content),
            "sha256": digest,
            "created_at": self._now(),
            "summary": "",
            "text_char_count": 0,
        }
        if kind in {"text", "document"}:
            extracted = self._extract_attachment_text(safe_name, guessed_mime, content)
            metadata["text_char_count"] = len(extracted)
            metadata["summary"] = extracted[:2000]
        with self._lock:
            existing = self._load_metadata_locked(owner_email=owner_email, session_id=session_id)
            session_dir.mkdir(parents=True, exist_ok=True)
            (session_dir / stored_name).write_bytes(content)
            existing[attachment_id] = metadata
            self._persist_metadata_locked(owner_email=owner_email, session_id=session_id, metadata=existing)
        return self.public_metadata(metadata)

    def resolve_many(self, *, owner_email: str, session_id: str, attachment_ids: list[str]) -> list[dict[str, Any]]:
        requested_ids = [str(item or "").strip() for item in attachment_ids if str(item or "").strip()]
        if len(requested_ids) > self.MAX_ATTACHMENTS:
            raise ToolError(f"At most {self.MAX_ATTACHMENTS} Source Code Q&A attachments are supported per question.")
        with self._lock:
            metadata = self._load_metadata_locked(owner_email=owner_email, session_id=session_id)
        resolved: list[dict[str, Any]] = []
        image_count = 0
        session_dir = self._session_dir(owner_email=owner_email, session_id=session_id)
        for attachment_id in requested_ids:
            item = metadata.get(attachment_id)
            if not item:
                raise ToolError("One or more Source Code Q&A attachments were not found for this session.")
            path = session_dir / str(item.get("stored_name") or "")
            if not path.exists() or not path.is_file():
                raise ToolError(f"Attachment file is missing: {item.get('filename') or attachment_id}")
            enriched = dict(item)
            enriched["path"] = str(path)
            if enriched.get("kind") == "image":
                image_count += 1
                if image_count > self.MAX_IMAGES:
                    raise ToolError(f"At most {self.MAX_IMAGES} image attachments are supported per question.")
            elif enriched.get("kind") in {"text", "document"}:
                try:
                    content = path.read_bytes()
                except OSError as error:
                    raise ToolError(f"Attachment file is unreadable: {item.get('filename') or attachment_id}") from error
                enriched["text"] = self._extract_attachment_text(str(item.get("filename") or ""), str(item.get("mime_type") or ""), content)
            resolved.append(enriched)
        return resolved

    def get_bytes(self, *, owner_email: str, session_id: str, attachment_id: str) -> tuple[dict[str, Any], bytes]:
        resolved = self.resolve_many(owner_email=owner_email, session_id=session_id, attachment_ids=[attachment_id])
        if not resolved:
            raise ToolError("Source Code Q&A attachment was not found.")
        item = resolved[0]
        try:
            content = Path(str(item.get("path") or "")).read_bytes()
        except OSError as error:
            raise ToolError("Source Code Q&A attachment file is unreadable.") from error
        return self.public_metadata(item), content

    @classmethod
    def public_metadata(cls, metadata: dict[str, Any]) -> dict[str, Any]:
        return {
            "id": str(metadata.get("id") or ""),
            "filename": str(metadata.get("filename") or ""),
            "mime_type": str(metadata.get("mime_type") or ""),
            "kind": str(metadata.get("kind") or ""),
            "size": int(metadata.get("size") or 0),
            "sha256": str(metadata.get("sha256") or ""),
            "created_at": str(metadata.get("created_at") or ""),
            "summary": str(metadata.get("summary") or "")[:400],
            "text_char_count": int(metadata.get("text_char_count") or 0),
        }

    def _attachment_kind(self, filename: str, mime_type: str, content: bytes) -> str:
        suffix = Path(filename).suffix.lower()
        mime = str(mime_type or "").lower()
        if mime in self.IMAGE_MIME_TYPES:
            return "image"
        if suffix in self.TEXT_EXTENSIONS or mime.startswith("text/") or mime in {"application/json", "application/xml"}:
            return "text"
        if suffix in self.DOCUMENT_EXTENSIONS:
            return "document"
        if b"\x00" in content[:2048]:
            raise ToolError("Unknown binary attachments are not supported.")
        if suffix:
            return "text"
        raise ToolError("Unsupported attachment type.")

    def _extract_attachment_text(self, filename: str, mime_type: str, content: bytes) -> str:
        suffix = Path(filename).suffix.lower()
        if suffix == ".pdf":
            return self._extract_pdf_text(content)
        if suffix == ".docx":
            return self._extract_docx_text(content)
        if suffix == ".xlsx":
            return self._extract_xlsx_text(content)
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                text = ""
        if not text:
            raise ToolError(f"Unable to parse text from attachment {filename or mime_type}.")
        return re.sub(r"\r\n?", "\n", text).strip()[:16000]

    @staticmethod
    def _extract_pdf_text(content: bytes) -> str:
        try:
            from pypdf import PdfReader  # type: ignore
        except ImportError as error:
            raise ToolError("PDF attachments are supported only when pypdf is installed on the server.") from error
        reader = PdfReader(io.BytesIO(content))
        lines: list[str] = []
        for page in reader.pages[:10]:
            lines.append(str(page.extract_text() or ""))
        text = "\n".join(lines).strip()
        if not text:
            raise ToolError("Unable to extract readable text from this PDF attachment.")
        return text[:16000]

    @staticmethod
    def _extract_docx_text(content: bytes) -> str:
        try:
            from docx import Document  # type: ignore
        except ImportError as error:
            raise ToolError("DOCX attachments are supported only when python-docx is installed on the server.") from error
        document = Document(io.BytesIO(content))
        text = "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text).strip()
        if not text:
            raise ToolError("Unable to extract readable text from this DOCX attachment.")
        return text[:16000]

    @staticmethod
    def _extract_xlsx_text(content: bytes) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise ToolError("XLSX attachments are supported only when openpyxl is installed on the server.") from error
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows: list[str] = []
        for worksheet in workbook.worksheets[:3]:
            rows.append(f"[Sheet: {worksheet.title}]")
            for row in worksheet.iter_rows(max_row=40, max_col=12, values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if any(value.strip() for value in values):
                    rows.append("\t".join(values).rstrip())
        text = "\n".join(rows).strip()
        if not text:
            raise ToolError("Unable to extract readable text from this XLSX attachment.")
        return text[:16000]


class SourceCodeQARuntimeEvidenceStore(SourceCodeQAAttachmentStore):
    ALLOWED_PM_TEAMS = {"AF", "CRMS", "GRC"}
    ALLOWED_COUNTRIES = {"ID", "SG", "PH"}
    SHARED_EVIDENCE_PM_TEAMS = {"AF", "GRC"}
    ALLOWED_SOURCE_TYPES = {"apollo", "db", "data_dictionary", "other"}
    MAX_FILES_PER_SCOPE = 20
    MAX_QUERY_FILES_PER_SCOPE = 8
    MAX_ZIP_MEMBERS = 500
    MAX_ZIP_UNCOMPRESSED_BYTES = 8 * 1024 * 1024
    MAX_ZIP_TEXT_CHARS = 120000
    MAX_DATA_DICTIONARY_XLSX_SHEETS = 120
    MAX_DATA_DICTIONARY_XLSX_ROWS_PER_SHEET = 300
    MAX_DATA_DICTIONARY_XLSX_COLS = 16
    MAX_DATA_DICTIONARY_XLSX_TEXT_CHARS = 120000
    ZIP_TEXT_EXTENSIONS = SourceCodeQAAttachmentStore.TEXT_EXTENSIONS | {
        ".conf",
        ".cfg",
        ".ini",
        ".toml",
        ".env",
    }

    @classmethod
    def _safe_scope(cls, *, pm_team: str, country: str) -> tuple[str, str]:
        normalized_team = str(pm_team or "").strip().upper()
        normalized_country = str(country or "").strip().upper()
        if normalized_team not in cls.ALLOWED_PM_TEAMS:
            raise ToolError("Runtime evidence PM Team must be one of AF, CRMS, or GRC.")
        if normalized_country == ALL_COUNTRY.upper():
            if normalized_team not in cls.SHARED_EVIDENCE_PM_TEAMS:
                raise ToolError("Shared All-country runtime evidence is supported only for AF and GRC.")
            return normalized_team, ALL_COUNTRY
        if normalized_country not in cls.ALLOWED_COUNTRIES:
            raise ToolError("Runtime evidence country must be one of All, ID, SG, or PH.")
        return normalized_team, normalized_country

    @classmethod
    def _safe_source_type(cls, source_type: str) -> str:
        normalized = str(source_type or "").strip().lower() or "other"
        if normalized not in cls.ALLOWED_SOURCE_TYPES:
            raise ToolError("Runtime evidence source type must be apollo, db, data_dictionary, or other.")
        return normalized

    def _scope_dir(self, *, pm_team: str, country: str) -> Path:
        if self.root_dir is None:
            raise ToolError("Source Code Q&A runtime evidence is not configured.")
        safe_team, safe_country = self._safe_scope(pm_team=pm_team, country=country)
        return self.root_dir / safe_team / safe_country

    def _metadata_path(self, *, pm_team: str, country: str) -> Path:
        return self._scope_dir(pm_team=pm_team, country=country) / "metadata.json"

    def _load_metadata_locked(self, *, pm_team: str, country: str) -> dict[str, dict[str, Any]]:
        path = self._metadata_path(pm_team=pm_team, country=country)
        if not path.exists():
            return {}
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
        items = payload.get("evidence") if isinstance(payload, dict) else {}
        return {str(key): value for key, value in items.items() if isinstance(value, dict)} if isinstance(items, dict) else {}

    def _persist_metadata_locked(self, *, pm_team: str, country: str, metadata: dict[str, dict[str, Any]]) -> None:
        path = self._metadata_path(pm_team=pm_team, country=country)
        path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = path.with_name(f".{path.name}.{os.getpid()}.tmp")
        temp_path.write_text(
            json.dumps({"updated_at": self._now(), "evidence": metadata}, ensure_ascii=False, sort_keys=True),
            encoding="utf-8",
        )
        os.replace(temp_path, path)

    def save_bytes(
        self,
        *,
        pm_team: str,
        country: str,
        source_type: str,
        uploaded_by: str,
        filename: str,
        content: bytes,
        mime_type: str = "",
    ) -> dict[str, Any]:
        if len(content or b"") <= 0:
            raise ToolError("Runtime evidence file is empty.")
        if len(content) > self.MAX_FILE_BYTES:
            raise ToolError("Runtime evidence is too large. Maximum size is 10MB per file.")
        safe_team, safe_country = self._safe_scope(pm_team=pm_team, country=country)
        safe_source_type = self._safe_source_type(source_type)
        safe_name = self._safe_filename(filename)
        suffix = Path(safe_name).suffix.lower()
        if suffix in self.BLOCKED_EXTENSIONS and suffix != ".zip":
            raise ToolError("Executable, archive, and unknown binary runtime evidence files are not supported.")
        guessed_mime = str(mime_type or mimetypes.guess_type(safe_name)[0] or "").lower()
        kind = "archive" if suffix == ".zip" else self._attachment_kind(safe_name, guessed_mime, content)
        if kind == "image":
            raise ToolError("Runtime evidence must be a parseable text, spreadsheet, PDF, or document file, not an image.")
        extracted = self._extract_runtime_evidence_text(safe_name, guessed_mime, content, source_type=safe_source_type)
        digest = hashlib.sha256(content).hexdigest()
        evidence_id = uuid.uuid4().hex
        stored_name = f"{evidence_id}{suffix or '.txt'}"
        scope_dir = self._scope_dir(pm_team=safe_team, country=safe_country)
        metadata = {
            "id": evidence_id,
            "filename": safe_name,
            "stored_name": stored_name,
            "mime_type": guessed_mime or "application/octet-stream",
            "kind": kind,
            "source_type": safe_source_type,
            "pm_team": safe_team,
            "country": safe_country,
            "size": len(content),
            "sha256": digest,
            "uploaded_by": str(uploaded_by or "").strip().lower(),
            "created_at": self._now(),
            "summary": extracted[:2000],
            "text_char_count": len(extracted),
        }
        with self._lock:
            existing = self._load_metadata_locked(pm_team=safe_team, country=safe_country)
            scope_dir.mkdir(parents=True, exist_ok=True)
            (scope_dir / stored_name).write_bytes(content)
            existing[evidence_id] = metadata
            if len(existing) > self.MAX_FILES_PER_SCOPE:
                ordered = sorted(existing.values(), key=lambda item: str(item.get("created_at") or ""))
                for stale in ordered[: len(existing) - self.MAX_FILES_PER_SCOPE]:
                    stale_id = str(stale.get("id") or "")
                    stale_name = str(stale.get("stored_name") or "")
                    if stale_name:
                        try:
                            (scope_dir / stale_name).unlink(missing_ok=True)
                        except OSError:
                            pass
                    existing.pop(stale_id, None)
            self._persist_metadata_locked(pm_team=safe_team, country=safe_country, metadata=existing)
        return self.public_metadata(metadata)

    def _extract_attachment_text(self, filename: str, mime_type: str, content: bytes) -> str:
        if Path(filename).suffix.lower() == ".zip":
            return self._extract_zip_text(content)
        return super()._extract_attachment_text(filename, mime_type, content)

    def _extract_runtime_evidence_text(self, filename: str, mime_type: str, content: bytes, *, source_type: str) -> str:
        if str(source_type or "").strip().lower() == "data_dictionary" and Path(filename).suffix.lower() == ".xlsx":
            return self._extract_data_dictionary_xlsx_text(content)
        return self._extract_attachment_text(filename, mime_type, content)

    def _extract_data_dictionary_xlsx_text(self, content: bytes) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise ToolError("XLSX data dictionaries are supported only when openpyxl is installed on the server.") from error
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines: list[str] = []
        sheet_count = 0
        for worksheet in workbook.worksheets[: self.MAX_DATA_DICTIONARY_XLSX_SHEETS]:
            sheet_count += 1
            lines.append(f"[Data dictionary sheet: {worksheet.title}]")
            non_empty_rows = 0
            for row in worksheet.iter_rows(
                max_row=self.MAX_DATA_DICTIONARY_XLSX_ROWS_PER_SHEET,
                max_col=self.MAX_DATA_DICTIONARY_XLSX_COLS,
                values_only=True,
            ):
                values = ["" if value is None else str(value).strip() for value in row]
                values = [value.replace("\r\n", "\n").replace("\r", "\n") for value in values]
                if not any(values):
                    continue
                non_empty_rows += 1
                lines.append("\t".join(values).rstrip())
                if sum(len(line) for line in lines) >= self.MAX_DATA_DICTIONARY_XLSX_TEXT_CHARS:
                    lines.append("...[data dictionary text truncated]")
                    text = "\n".join(lines).strip()
                    if not text:  # pragma: no cover - sheet header is always appended before this guard.
                        raise ToolError("Unable to extract readable text from this XLSX data dictionary.")
                    return text[: self.MAX_DATA_DICTIONARY_XLSX_TEXT_CHARS]
            if non_empty_rows == 0:
                lines.append("(empty sheet)")
        if len(workbook.worksheets) > sheet_count:
            lines.append(f"[Data dictionary skipped sheets: {len(workbook.worksheets) - sheet_count}]")
        text = "\n".join(lines).strip()
        if not text:
            raise ToolError("Unable to extract readable text from this XLSX data dictionary.")
        return text[: self.MAX_DATA_DICTIONARY_XLSX_TEXT_CHARS]

    def _extract_zip_text(self, content: bytes) -> str:
        try:
            archive = zipfile.ZipFile(io.BytesIO(content))
        except zipfile.BadZipFile as error:
            raise ToolError("Unable to read this ZIP runtime evidence file.") from error
        lines: list[str] = []
        total_uncompressed = 0
        readable_members = 0
        skipped_members = 0
        with archive:
            members = [member for member in archive.infolist() if not member.is_dir()]
            if len(members) > self.MAX_ZIP_MEMBERS:
                raise ToolError(f"ZIP runtime evidence contains too many files. Maximum is {self.MAX_ZIP_MEMBERS}.")
            for member in members:
                member_name = str(member.filename or "").replace("\\", "/")
                clean_parts = [part for part in member_name.split("/") if part and part not in {".", ".."}]
                if not clean_parts or clean_parts[0] == "__MACOSX" or len(clean_parts) != len([part for part in member_name.split("/") if part]):
                    skipped_members += 1
                    continue
                member_basename = clean_parts[-1].lower()
                suffix = Path(member_basename).suffix.lower() or (member_basename if member_basename.startswith(".") else "")
                if suffix in self.BLOCKED_EXTENSIONS or suffix not in self.ZIP_TEXT_EXTENSIONS:
                    skipped_members += 1
                    continue
                total_uncompressed += int(member.file_size or 0)
                if total_uncompressed > self.MAX_ZIP_UNCOMPRESSED_BYTES:
                    max_mb = self.MAX_ZIP_UNCOMPRESSED_BYTES // (1024 * 1024)
                    raise ToolError(f"ZIP runtime evidence is too large after extraction. Keep text config files under {max_mb}MB total.")
                try:
                    raw = archive.read(member)
                except (OSError, RuntimeError, zipfile.BadZipFile) as error:
                    raise ToolError(f"Unable to read {member_name} inside this ZIP runtime evidence file.") from error
                if b"\x00" in raw[:2048]:
                    skipped_members += 1
                    continue
                text = ""
                for encoding in ("utf-8-sig", "utf-8", "latin-1"):
                    try:
                        text = raw.decode(encoding)
                        break
                    except UnicodeDecodeError:
                        text = ""
                text = re.sub(r"\r\n?", "\n", text).strip()
                if not text:
                    skipped_members += 1
                    continue
                readable_members += 1
                lines.append(f"[ZIP file: {'/'.join(clean_parts)}]\n{text[:12000]}")
                if sum(len(line) for line in lines) >= self.MAX_ZIP_TEXT_CHARS:
                    lines.append("...[zip text truncated]")
                    break
        if not readable_members:
            raise ToolError("ZIP runtime evidence did not contain readable config/text files.")
        if skipped_members:
            lines.append(f"[ZIP skipped files: {skipped_members}]")
        return "\n\n".join(lines).strip()[: self.MAX_ZIP_TEXT_CHARS]

    @classmethod
    def public_metadata(cls, metadata: dict[str, Any]) -> dict[str, Any]:
        payload = SourceCodeQAAttachmentStore.public_metadata(metadata)
        payload.update(
            {
                "source_type": str(metadata.get("source_type") or ""),
                "pm_team": str(metadata.get("pm_team") or ""),
                "country": str(metadata.get("country") or ""),
                "uploaded_by": str(metadata.get("uploaded_by") or ""),
            }
        )
        return payload

    def list(self, *, pm_team: str, country: str) -> list[dict[str, Any]]:
        safe_team, safe_country = self._safe_scope(pm_team=pm_team, country=country)
        with self._lock:
            metadata = self._load_metadata_locked(pm_team=safe_team, country=safe_country)
        return [
            self.public_metadata(item)
            for item in sorted(metadata.values(), key=lambda value: str(value.get("created_at") or ""), reverse=True)
        ]

    def resolve_scope(self, *, pm_team: str, country: str) -> list[dict[str, Any]]:
        safe_team = str(pm_team or "").strip().upper()
        normalized_country = str(country or "").strip().upper()
        if normalized_country in {"", ALL_COUNTRY.upper()}:
            countries = [ALL_COUNTRY, *sorted(self.ALLOWED_COUNTRIES)] if safe_team in self.SHARED_EVIDENCE_PM_TEAMS else sorted(self.ALLOWED_COUNTRIES)
        else:
            countries = [ALL_COUNTRY, normalized_country] if safe_team in self.SHARED_EVIDENCE_PM_TEAMS else [normalized_country]
        resolved: list[dict[str, Any]] = []
        seen: set[str] = set()
        for scoped_country in countries:
            safe_team, safe_country = self._safe_scope(pm_team=pm_team, country=scoped_country)
            with self._lock:
                metadata = self._load_metadata_locked(pm_team=safe_team, country=safe_country)
            scope_dir = self._scope_dir(pm_team=safe_team, country=safe_country)
            for item in sorted(metadata.values(), key=lambda value: str(value.get("created_at") or ""), reverse=True)[: self.MAX_QUERY_FILES_PER_SCOPE]:
                item_id = str(item.get("id") or "")
                if item_id and item_id in seen:
                    continue
                if item_id:
                    seen.add(item_id)
                path = scope_dir / str(item.get("stored_name") or "")
                if not path.exists() or not path.is_file():
                    continue
                enriched = dict(item)
                enriched["path"] = str(path)
                try:
                    content = path.read_bytes()
                except OSError:
                    continue
                enriched["text"] = self._extract_runtime_evidence_text(
                    str(item.get("filename") or ""),
                    str(item.get("mime_type") or ""),
                    content,
                    source_type=str(item.get("source_type") or ""),
                )
                resolved.append(enriched)
        return resolved[: self.MAX_QUERY_FILES_PER_SCOPE * max(1, len(countries))]

    def delete(self, *, pm_team: str, country: str, evidence_id: str) -> bool:
        safe_team, safe_country = self._safe_scope(pm_team=pm_team, country=country)
        normalized_id = str(evidence_id or "").strip()
        if not re.fullmatch(r"[a-fA-F0-9]{32}", normalized_id):
            raise ToolError("Runtime evidence id is invalid.")
        scope_dir = self._scope_dir(pm_team=safe_team, country=safe_country)
        with self._lock:
            metadata = self._load_metadata_locked(pm_team=safe_team, country=safe_country)
            item = metadata.pop(normalized_id, None)
            if item is None:
                return False
            stored_name = str(item.get("stored_name") or "")
            if stored_name:
                try:
                    (scope_dir / stored_name).unlink(missing_ok=True)
                except OSError:
                    pass
            self._persist_metadata_locked(pm_team=safe_team, country=safe_country, metadata=metadata)
        return True


class SourceCodeQAGeneratedArtifactStore:
    MAX_SQL_BYTES = 2 * 1024 * 1024

    def __init__(self, root_dir: Path | None = None) -> None:
        self.root_dir = root_dir
        self._lock = threading.Lock()

    @staticmethod
    def _now() -> str:
        return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

    @staticmethod
    def _owner_key(owner_email: str) -> str:
        return SourceCodeQAAttachmentStore._owner_key(owner_email)

    @staticmethod
    def _safe_session_id(session_id: str) -> str:
        return SourceCodeQAAttachmentStore._safe_session_id(session_id)

    def _session_dir(self, *, owner_email: str, session_id: str) -> Path:
        if self.root_dir is None:
            raise ToolError("Source Code Q&A generated artifacts are not configured.")
        return self.root_dir / self._owner_key(owner_email) / self._safe_session_id(session_id)

    def _metadata_path(self, *, owner_email: str, session_id: str) -> Path:
        return self._session_dir(owner_email=owner_email, session_id=session_id) / "metadata.json"

    def _load_metadata_locked(self, *, owner_email: str, session_id: str) -> dict[str, dict[str, Any]]:
        path = self._metadata_path(owner_email=owner_email, session_id=session_id)
        if not path.exists():
            return {}
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
        artifacts = payload.get("artifacts") if isinstance(payload, dict) else {}
        return {str(key): value for key, value in artifacts.items() if isinstance(value, dict)} if isinstance(artifacts, dict) else {}

    def _persist_metadata_locked(self, *, owner_email: str, session_id: str, metadata: dict[str, dict[str, Any]]) -> None:
        path = self._metadata_path(owner_email=owner_email, session_id=session_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = path.with_name(f".{path.name}.{os.getpid()}.tmp")
        temp_path.write_text(
            json.dumps({"updated_at": self._now(), "artifacts": metadata}, ensure_ascii=False, sort_keys=True),
            encoding="utf-8",
        )
        os.replace(temp_path, path)

    @classmethod
    def public_metadata(cls, metadata: dict[str, Any]) -> dict[str, Any]:
        return {
            "id": str(metadata.get("id") or ""),
            "filename": str(metadata.get("filename") or "source-code-qa-sql-package.zip"),
            "mime_type": str(metadata.get("mime_type") or "application/zip"),
            "kind": str(metadata.get("kind") or "sql_package"),
            "size": int(metadata.get("size") or 0),
            "sha256": str(metadata.get("sha256") or ""),
            "created_at": str(metadata.get("created_at") or ""),
            "question": str(metadata.get("question") or "")[:240],
            "pm_team": str(metadata.get("pm_team") or ""),
            "country": str(metadata.get("country") or ""),
        }

    def save_sql_package(
        self,
        *,
        owner_email: str,
        session_id: str,
        pm_team: str,
        country: str,
        question: str,
        sql: str,
        readme: str,
    ) -> dict[str, Any]:
        normalized_sql = str(sql or "").strip()
        if not normalized_sql:
            raise ToolError("Generated SQL content is empty.")
        normalized_sql = _format_source_code_qa_sql_text(normalized_sql)
        sql_bytes = normalized_sql.encode("utf-8")
        if len(sql_bytes) > self.MAX_SQL_BYTES:
            raise ToolError("Generated SQL content is too large to package.")
        artifact_id = uuid.uuid4().hex
        filename = "source-code-qa-sql-package.zip"
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("query.sql", normalized_sql + "\n")
            archive.writestr("README.md", str(readme or "").strip() + "\n")
        content = buffer.getvalue()
        digest = hashlib.sha256(content).hexdigest()
        stored_name = f"{artifact_id}.zip"
        metadata = {
            "id": artifact_id,
            "filename": filename,
            "stored_name": stored_name,
            "mime_type": "application/zip",
            "kind": "sql_package",
            "size": len(content),
            "sha256": digest,
            "created_at": self._now(),
            "question": str(question or "").strip(),
            "pm_team": str(pm_team or "").strip().upper(),
            "country": str(country or "").strip().upper() or ALL_COUNTRY,
        }
        with self._lock:
            session_dir = self._session_dir(owner_email=owner_email, session_id=session_id)
            existing = self._load_metadata_locked(owner_email=owner_email, session_id=session_id)
            session_dir.mkdir(parents=True, exist_ok=True)
            (session_dir / stored_name).write_bytes(content)
            existing[artifact_id] = metadata
            self._persist_metadata_locked(owner_email=owner_email, session_id=session_id, metadata=existing)
        return self.public_metadata(metadata)

    def get_bytes(self, *, owner_email: str, session_id: str, artifact_id: str) -> tuple[dict[str, Any], bytes]:
        normalized_id = str(artifact_id or "").strip()
        if not re.fullmatch(r"[a-fA-F0-9]{32}", normalized_id):
            raise ToolError("Generated artifact id is invalid.")
        with self._lock:
            metadata = self._load_metadata_locked(owner_email=owner_email, session_id=session_id)
            item = metadata.get(normalized_id)
            if item is None:
                raise ToolError("Generated artifact was not found.")
            path = self._session_dir(owner_email=owner_email, session_id=session_id) / str(item.get("stored_name") or "")
            try:
                content = path.read_bytes()
            except OSError as error:
                raise ToolError("Generated artifact file is unreadable.") from error
        return self.public_metadata(item), content


def _compact_source_code_qa_session_payload(result: dict[str, Any]) -> dict[str, Any]:
    structured = result.get("structured_answer") if isinstance(result.get("structured_answer"), dict) else {}
    answer_claim_check = result.get("answer_claim_check") if isinstance(result.get("answer_claim_check"), dict) else {}
    llm_route = result.get("llm_route") if isinstance(result.get("llm_route"), dict) else {}
    codex_trace = result.get("codex_cli_trace") if isinstance(result.get("codex_cli_trace"), dict) else {}
    return {
        "status": result.get("status") or "",
        "trace_id": result.get("trace_id") or "",
        "query_mode": result.get("query_mode") or "",
        "deadline_seconds": result.get("deadline_seconds") or 0,
        "deadline_hit": bool(result.get("deadline_hit")),
        "fallback_used": bool(result.get("fallback_used")),
        "fallback_answer_quality": result.get("fallback_answer_quality") or "",
        "fallback_evidence_count": result.get("fallback_evidence_count") or 0,
        "fallback_claim_count": result.get("fallback_claim_count") or 0,
        "deadline_fallback_reason": result.get("deadline_fallback_reason") or "",
        "summary": result.get("summary") or "",
        "llm_answer": result.get("llm_answer") or "",
        "llm_provider": result.get("llm_provider") or "",
        "llm_model": result.get("llm_model") or "",
        "llm_route": {
            "mode": llm_route.get("mode") or "",
            "query_mode": llm_route.get("query_mode") or "",
            "provider": llm_route.get("provider") or "",
            "prompt_mode": llm_route.get("prompt_mode") or "",
            "candidate_paths": (llm_route.get("candidate_paths") or [])[:30],
            "candidate_path_layers": llm_route.get("candidate_path_layers") or {},
            "codex_session_max_turns": llm_route.get("codex_session_max_turns") or 8,
        },
        "structured_answer": {
            "direct_answer": structured.get("direct_answer") or "",
            "claims": (structured.get("claims") or [])[:8],
            "confirmed_points": (structured.get("confirmed_points") or [])[:8],
            "missing_points": (structured.get("missing_points") or [])[:8],
            "evidence_cards": (structured.get("evidence_cards") or [])[:8],
            "citations": (structured.get("citations") or [])[:12],
            "missing_evidence": (structured.get("missing_evidence") or [])[:8],
            "confidence": structured.get("confidence") or "",
        },
        "answer_contract": result.get("answer_contract") or {},
        "answer_quality": result.get("answer_quality") or {},
        "codex_cli_summary": result.get("codex_cli_summary") or {},
        "codex_cli_trace": {
            "session_mode": codex_trace.get("session_mode") or "",
            "command_mode": codex_trace.get("command_mode") or "",
            "session_id": codex_trace.get("session_id") or "",
            "exit_code": codex_trace.get("exit_code"),
            "latency_ms": codex_trace.get("latency_ms"),
            "timeout": bool(codex_trace.get("timeout")),
            "stream_messages": (codex_trace.get("stream_messages") or [])[-20:],
            "command_summaries": (codex_trace.get("command_summaries") or [])[-12:],
            "probable_inspected_files": (codex_trace.get("probable_inspected_files") or [])[-20:],
        },
        "codex_citation_validation": answer_claim_check.get("codex_citation_validation") or {},
        "attachments": [
            SourceCodeQAAttachmentStore.public_metadata(item)
            for item in (result.get("attachments") if isinstance(result.get("attachments"), list) else [])
            if isinstance(item, dict)
        ],
        "runtime_evidence": [
            SourceCodeQARuntimeEvidenceStore.public_metadata(item)
            for item in (result.get("runtime_evidence") if isinstance(result.get("runtime_evidence"), list) else [])
            if isinstance(item, dict)
        ],
        "generated_artifacts": [
            SourceCodeQAGeneratedArtifactStore.public_metadata(item)
            for item in (result.get("generated_artifacts") if isinstance(result.get("generated_artifacts"), list) else [])
            if isinstance(item, dict)
        ],
        "matches": [
            {
                "repo": match.get("repo"),
                "path": match.get("path"),
                "line_start": match.get("line_start"),
                "line_end": match.get("line_end"),
                "retrieval": match.get("retrieval"),
                "trace_stage": match.get("trace_stage"),
                "reason": match.get("reason"),
                "score": match.get("score"),
            }
            for match in (result.get("matches") or [])[:10]
            if isinstance(match, dict)
        ],
    }
