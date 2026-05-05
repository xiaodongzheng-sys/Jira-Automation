from __future__ import annotations

import base64
from collections import deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
import difflib
from email.utils import getaddresses
from functools import lru_cache
import hashlib
import html
import inspect
import io
from http import HTTPStatus
import json
import logging
import mimetypes
import os
from pathlib import Path
import re
import secrets
import sqlite3
import subprocess
import threading
import time
from typing import Any
from urllib.parse import parse_qs, urlparse
import uuid
import zipfile

from flask import Flask, Response, current_app, flash, g, has_app_context, jsonify, redirect, render_template, request, send_file, session, stream_with_context, url_for
from dotenv import load_dotenv
import google_auth_httplib2
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build as build_google_api
from googleapiclient.errors import HttpError
import httplib2
from openpyxl import Workbook
from openpyxl.styles import Font
import requests
from werkzeug.middleware.proxy_fix import ProxyFix

from bpmis_jira_tool.config import Settings
from bpmis_jira_tool.errors import BPMISError, ConfigError, ToolError
from bpmis_jira_tool.google_auth import (
    create_google_authorization_url,
    finish_google_oauth,
    get_google_credentials,
)
from bpmis_jira_tool.local_agent_client import (
    LocalAgentClient,
    RemoteBPMISProjectStore,
    RemoteSeaTalkDashboardService,
    RemoteSeaTalkNameMappingStore,
    RemoteSeaTalkTodoStore,
    RemoteTeamDashboardConfigStore,
    RemoteSourceCodeQAAttachmentStore,
    RemoteSourceCodeQAModelAvailabilityStore,
    RemoteSourceCodeQARuntimeEvidenceStore,
    RemoteSourceCodeQASessionStore,
    RemoteSourceCodeQAService,
    _LOCAL_AGENT_SESSION,
)
from bpmis_jira_tool.gmail_dashboard import GMAIL_READONLY_SCOPE, GmailDashboardService
from bpmis_jira_tool.gmail_sender import StoredGoogleCredentials
from bpmis_jira_tool.meeting_recorder import (
    CALENDAR_READONLY_SCOPE,
    GoogleCalendarMeetingService,
    MeetingProcessingService,
    MeetingRecorderConfig,
    MeetingRecorderRuntime,
    normalize_meeting_transcript_language,
    MeetingRecordStore,
    _utc_now,
    meeting_platform_from_link,
    reminder_eligible_meetings,
)
from bpmis_jira_tool.monthly_report import (
    DEFAULT_MONTHLY_REPORT_RECIPIENT,
    DEFAULT_MONTHLY_REPORT_TEMPLATE,
    MONTHLY_REPORT_PRODUCT_SCOPE,
    MonthlyReportService,
    monthly_report_subject,
    normalize_monthly_report_template,
    resolve_monthly_report_period,
    send_monthly_report_email,
)
from bpmis_jira_tool.report_intelligence import (
    normalize_report_intelligence_config,
)
from bpmis_jira_tool.seatalk_dashboard import SeaTalkDashboardService
from bpmis_jira_tool.google_sheets import GoogleSheetsService
from bpmis_jira_tool.bpmis_projects import BPMISProjectStore, PortalJiraCreationService, PortalProjectSyncService
from bpmis_jira_tool.project_sync import BPMISProjectSyncService
from bpmis_jira_tool.service import JiraCreationService, build_bpmis_client
from bpmis_jira_tool.source_code_qa import CRMS_COUNTRIES, ALL_COUNTRY, CodexCliBridgeSourceCodeQALLMProvider, SourceCodeQAService
from bpmis_jira_tool.user_config import (
    CONFIGURED_FIELDS,
    DEFAULT_DIRECT_VALUES,
    DEFAULT_NEED_UAT_BY_MARKET,
    DEFAULT_SHEET_HEADERS,
    TEAM_DEFAULT_EMAIL_PLACEHOLDER,
    TEAM_PROFILE_DEFAULTS,
    WebConfigStore,
)
from bpmis_jira_tool.work_memory import (
    WorkMemoryStore,
    gmail_attachment_memory_item,
    gmail_drive_link_memory_item,
    gmail_message_memory_item,
    meeting_record_memory_items,
    sent_monthly_report_memory_item_from_gmail_record,
    source_code_qa_memory_item,
    team_dashboard_memory_items,
)
from prd_briefing import create_prd_briefing_blueprint
from prd_briefing.confluence import ConfluenceConnector
from prd_briefing.reviewer import PRDBriefingReviewRequest, PRDReviewRequest, PRDReviewService
from prd_briefing.storage import BriefingStore
from prd_briefing.text_generation import CodexTextGenerationClient


PROJECT_ROOT = Path(__file__).resolve().parent.parent
SOURCE_CODE_QA_EFFORT_DICTIONARY_PATH = PROJECT_ROOT / "config" / "source_code_qa_effort_dictionaries.json"
SOURCE_CODE_QA_DOMAIN_PROFILES_PATH = PROJECT_ROOT / "config" / "source_code_qa_domain_profiles.json"
SOURCE_CODE_QA_DOMAIN_KNOWLEDGE_PATH = PROJECT_ROOT / "config" / "source_code_qa_domain_knowledge_packs.json"
load_dotenv(PROJECT_ROOT / ".env")
MARKET_KEYS = ["ID", "SG", "PH", "Regional"]
PORTAL_ADMIN_EMAIL = "xiaodong.zheng@npt.sg"
PORTAL_TEST_USER_EMAIL = "xiaodong.zheng1991@gmail.com"
TEAM_PROFILE_ADMIN_EMAIL = PORTAL_ADMIN_EMAIL
SYNC_EMAIL_EDIT_ALLOWLIST = {PORTAL_ADMIN_EMAIL}
SOURCE_CODE_QA_BUILTIN_ADMIN_EMAILS = {PORTAL_ADMIN_EMAIL}
GMAIL_SEATALK_BUILTIN_OWNER_EMAILS = {PORTAL_ADMIN_EMAIL}
TEAM_DASHBOARD_ACCESS_EMAILS = {PORTAL_ADMIN_EMAIL}
TEAM_DASHBOARD_LEGACY_DEFAULT_MEMBER_EMAILS = (
    "huixian.nah@npt.sg",
    "jireh.tanyx@npt.sg",
    "keryin.lim@npt.sg",
    "liye.ng@npt.sg",
    "mingming.yeo@npt.sg",
    "chongzj@npt.sg",
    "sabrina.chan@npt.sg",
    "sophia.wangzj@npt.sg",
    "chang.wang@npt.sg",
    "zoey.luxy@npt.sg",
)
TEAM_DASHBOARD_DEFAULT_MEMBER_EMAILS_BY_TEAM = {
    "AF": (
        "jireh.tanyx@npt.sg",
        "keryin.lim@npt.sg",
        "chongzj@npt.sg",
        "chang.wang@npt.sg",
        "zoey.luxy@npt.sg",
        "xiaodong.zheng@npt.sg",
    ),
    "CRMS": (
        "huixian.nah@npt.sg",
        "liye.ng@npt.sg",
        "mingming.yeo@npt.sg",
        "sophia.wangzj@npt.sg",
    ),
    "GRC": (
        "sabrina.chan@npt.sg",
    ),
}
TEAM_DASHBOARD_TEAMS = {
    "AF": "Anti-fraud",
    "CRMS": "Credit Risk",
    "GRC": "Ops Risk",
}
TEAM_DASHBOARD_UNDER_PRD_STATUSES = {"waiting", "prd in progress", "prd reviewed"}
TEAM_DASHBOARD_EXCLUDED_PENDING_STATUSES = {"icebox", "closed", "done"}
TEAM_DASHBOARD_UNDER_PRD_BIZ_PROJECT_STATUSES = {"pending review", "confirmed"}
TEAM_DASHBOARD_PENDING_LIVE_BIZ_PROJECT_STATUSES = {"developing", "testing", "uat"}
TEAM_DASHBOARD_TASK_CACHE_VERSION = 2
MEETING_RECORDER_PROCESS_ACTION = "meeting-recorder-process"
WORK_MEMORY_GMAIL_BACKFILL_ACTION = "work-memory-gmail-backfill"
GOOGLE_DRIVE_READONLY_SCOPE = "https://www.googleapis.com/auth/drive.readonly"
GMAIL_WORK_MEMORY_MAX_ATTACHMENT_BYTES = 12 * 1024 * 1024
GMAIL_WORK_MEMORY_MAX_VIP_ATTACHMENTS_PER_MESSAGE = 5
GMAIL_WORK_MEMORY_CONTENT_CHARS = 16000
GMAIL_WORK_MEMORY_FETCH_BATCH_SIZE = 25
GMAIL_WORK_MEMORY_FETCH_WORKERS = 4
GOOGLE_DRIVE_HTTP_TIMEOUT_SECONDS = 20
TEAM_DASHBOARD_LINK_BIZ_EXCLUDED_TITLE_PHRASES = (
    "sync af productization",
    "productisation upgrade",
    "deployment of productization",
)
_gmail_export_active_users: set[str] = set()
_gmail_export_active_users_lock = threading.Lock()
_source_code_qa_codex_session_locks: dict[str, threading.Lock] = {}
_source_code_qa_codex_session_locks_guard = threading.Lock()


def _bool_env_from_env(name: str, default: bool = False) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


@lru_cache(maxsize=1)
def _current_release_revision() -> str:
    pinned_revision = str(os.environ.get("TEAM_PORTAL_RELEASE_REVISION") or "").strip()
    if pinned_revision:
        return pinned_revision
    try:
        completed = subprocess.run(
            ["git", "-C", str(PROJECT_ROOT), "rev-parse", "HEAD"],
            capture_output=True,
            text=True,
            check=True,
        )
    except (FileNotFoundError, subprocess.CalledProcessError):
        return "unknown"
    revision = completed.stdout.strip()
    return revision or "unknown"


@dataclass
class JobState:
    job_id: str
    action: str
    state: str = "queued"
    title: str = ""
    message: str = ""
    stage: str = "queued"
    current: int = 0
    total: int = 0
    estimated_prompt_tokens: int = 0
    token_risk: str = ""
    error_category: str = ""
    error_code: str = ""
    error_retryable: bool = False
    owner_email: str = ""
    record_id: str = ""
    query_mode: str = ""
    queued_position: int = 0
    eta_seconds_range: list[int] = field(default_factory=list)
    running_user_count: int = 0
    last_progress_at: float = 0
    stalled_retryable: bool = False
    started_at: float = 0
    completed_at: float = 0
    results: list[dict[str, Any]] = field(default_factory=list)
    notice: dict[str, Any] | None = None
    error: str | None = None
    created_at: float = field(default_factory=time.time)
    updated_at: float = field(default_factory=time.time)


class JobStore:
    def __init__(self, storage_path: Path | None = None) -> None:
        self.storage_path = storage_path
        self._loaded_jobs_interrupted = False
        self._jobs: dict[str, JobState] = self._load(mark_interrupted=True)
        self._lock = threading.Lock()
        if self._loaded_jobs_interrupted:
            with self._lock:
                self._persist_locked()

    def _load(self, *, mark_interrupted: bool = False) -> dict[str, JobState]:
        if self.storage_path is None or not self.storage_path.exists():
            return {}
        try:
            payload = json.loads(self.storage_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
        jobs: dict[str, JobState] = {}
        for job_id, raw_job in (payload.get("jobs") or {}).items():
            if not isinstance(raw_job, dict):
                continue
            try:
                job = JobState(**raw_job)
            except TypeError:
                continue
            if mark_interrupted and job.state in {"queued", "running"}:
                job.state = "failed"
                job.stage = "failed"
                job.message = "This job was interrupted by a server restart. Please start it again."
                job.error = job.message
                job.updated_at = time.time()
                self._loaded_jobs_interrupted = True
            jobs[str(job_id)] = job
        return jobs

    def _refresh_locked(self) -> None:
        if self.storage_path is None:
            return
        loaded_jobs = self._load(mark_interrupted=False)
        for job_id, loaded_job in loaded_jobs.items():
            current_job = self._jobs.get(job_id)
            if current_job is None or loaded_job.updated_at >= current_job.updated_at:
                self._jobs[job_id] = loaded_job

    def _persist_locked(self) -> None:
        if self.storage_path is None:
            return
        try:
            self.storage_path.parent.mkdir(parents=True, exist_ok=True)
            payload = {
                "updated_at": time.time(),
                "jobs": {job_id: asdict(job) for job_id, job in self._jobs.items()},
            }
            temp_path = self.storage_path.with_name(f".{self.storage_path.name}.{os.getpid()}.tmp")
            temp_path.write_text(json.dumps(payload, ensure_ascii=False, sort_keys=True), encoding="utf-8")
            os.replace(temp_path, self.storage_path)
        except OSError:
            return

    def create(self, action: str, title: str, *, owner_email: str = "", record_id: str = "") -> JobState:
        with self._lock:
            job = JobState(
                job_id=uuid.uuid4().hex,
                action=action,
                title=title,
                message="Queued and waiting to start.",
                owner_email=str(owner_email or "").strip().lower(),
                record_id=str(record_id or "").strip(),
            )
            self._jobs[job.job_id] = job
            self._persist_locked()
            return job

    def set_owner(self, job_id: str, owner_email: str) -> None:
        with self._lock:
            if job_id not in self._jobs:
                self._refresh_locked()
            job = self._jobs[job_id]
            job.owner_email = str(owner_email or "").strip().lower()
            job.updated_at = time.time()
            self._persist_locked()

    def set_query_mode(self, job_id: str, query_mode: str) -> None:
        with self._lock:
            if job_id not in self._jobs:
                self._refresh_locked()
            job = self._jobs[job_id]
            job.query_mode = str(query_mode or "").strip().lower()
            job.updated_at = time.time()
            self._persist_locked()

    def set_record_id(self, job_id: str, record_id: str) -> None:
        with self._lock:
            if job_id not in self._jobs:
                self._refresh_locked()
            job = self._jobs[job_id]
            job.record_id = str(record_id or "").strip()
            job.updated_at = time.time()
            self._persist_locked()

    def update_queue_metadata(
        self,
        job_id: str,
        *,
        queued_position: int = 0,
        eta_seconds_range: list[int] | None = None,
        running_user_count: int = 0,
        message: str | None = None,
    ) -> None:
        with self._lock:
            if job_id not in self._jobs:
                self._refresh_locked()
            job = self._jobs[job_id]
            job.queued_position = max(0, int(queued_position or 0))
            job.eta_seconds_range = [max(0, int(value or 0)) for value in (eta_seconds_range or [])[:2]]
            job.running_user_count = max(0, int(running_user_count or 0))
            if message is not None:
                job.message = message
            job.updated_at = time.time()
            self._persist_locked()

    def update(
        self,
        job_id: str,
        *,
        state: str | None = None,
        stage: str | None = None,
        message: str | None = None,
        current: int | None = None,
        total: int | None = None,
        estimated_prompt_tokens: int | None = None,
        token_risk: str | None = None,
    ) -> None:
        with self._lock:
            if job_id not in self._jobs:
                self._refresh_locked()
            job = self._jobs[job_id]
            if state is not None:
                if state == "running" and job.state != "running":
                    job.started_at = job.started_at or time.time()
                    job.queued_position = 0
                    job.eta_seconds_range = []
                job.state = state
            if stage is not None:
                job.stage = stage
            if message is not None:
                job.message = message
            if current is not None:
                job.current = current
            if total is not None:
                job.total = total
            if estimated_prompt_tokens is not None:
                job.estimated_prompt_tokens = estimated_prompt_tokens
            if token_risk is not None:
                job.token_risk = token_risk
            job.last_progress_at = time.time()
            job.stalled_retryable = False
            job.updated_at = time.time()
            self._persist_locked()

    def complete(self, job_id: str, *, results: list[dict[str, Any]], notice: dict[str, Any]) -> None:
        with self._lock:
            if job_id not in self._jobs:
                self._refresh_locked()
            job = self._jobs[job_id]
            job.state = "completed"
            job.stage = "completed"
            job.message = "Finished."
            job.results = results
            job.notice = notice
            job.completed_at = time.time()
            job.queued_position = 0
            job.eta_seconds_range = []
            job.last_progress_at = job.completed_at
            job.stalled_retryable = False
            job.updated_at = time.time()
            self._persist_locked()

    def fail(
        self,
        job_id: str,
        error: str,
        *,
        error_category: str = "",
        error_code: str = "",
        error_retryable: bool = True,
    ) -> None:
        with self._lock:
            if job_id not in self._jobs:
                self._refresh_locked()
            job = self._jobs[job_id]
            job.state = "failed"
            job.stage = "failed"
            job.message = error
            job.error = error
            job.error_category = error_category
            job.error_code = error_code
            job.error_retryable = error_retryable
            job.completed_at = time.time()
            job.queued_position = 0
            job.eta_seconds_range = []
            job.last_progress_at = job.completed_at
            job.stalled_retryable = False
            job.updated_at = time.time()
            self._persist_locked()

    def get(self, job_id: str) -> JobState | None:
        with self._lock:
            self._refresh_locked()
            job = self._jobs.get(job_id)
            if job is None:
                return None
            return JobState(**asdict(job))

    def snapshot(self, job_id: str) -> dict[str, Any] | None:
        job = self.get(job_id)
        if not job:
            return None
        payload = asdict(job)
        if payload.get("state") == "running":
            last_progress_at = float(payload.get("last_progress_at") or payload.get("updated_at") or 0)
            stalled = bool(last_progress_at and time.time() - last_progress_at > 180)
            payload["stalled_retryable"] = stalled
            if stalled:
                payload["error_retryable"] = True
        payload["progress"] = {
            "stage": job.stage,
            "current": job.current,
            "total": job.total,
            "message": job.message,
            "estimated_prompt_tokens": job.estimated_prompt_tokens,
            "token_risk": job.token_risk,
        }
        return payload

    def p95_duration_seconds(self, action: str, *, default_seconds: int = 212) -> int:
        with self._lock:
            self._refresh_locked()
            durations = []
            for job in self._jobs.values():
                if job.action != action or job.state != "completed":
                    continue
                started_at = float(job.started_at or job.created_at or 0)
                completed_at = float(job.completed_at or job.updated_at or 0)
                if started_at and completed_at and completed_at >= started_at:
                    durations.append(completed_at - started_at)
            if not durations:
                return default_seconds
            durations.sort()
            index = min(len(durations) - 1, int(round(0.95 * (len(durations) - 1))))
            return max(30, int(durations[index]))

    def latest_completed_result(self, action: str) -> dict[str, Any] | None:
        with self._lock:
            self._refresh_locked()
            candidates = [
                job
                for job in self._jobs.values()
                if job.action == action and job.state == "completed" and job.results
            ]
            if not candidates:
                return None
            latest = max(candidates, key=lambda item: item.updated_at)
            result = latest.results[0] if latest.results else {}
            if not isinstance(result, dict):
                return None
            return {
                **result,
                "job_id": latest.job_id,
                "generated_at": latest.updated_at,
            }

    def list_snapshots(self, *, action: str = "", owner_email: str = "", limit: int = 20) -> list[dict[str, Any]]:
        normalized_action = str(action or "").strip()
        normalized_owner = str(owner_email or "").strip().lower()
        with self._lock:
            self._refresh_locked()
            jobs = [
                job
                for job in self._jobs.values()
                if (not normalized_action or job.action == normalized_action)
                and (not normalized_owner or str(job.owner_email or "").strip().lower() == normalized_owner)
            ]
            jobs.sort(key=lambda item: item.updated_at, reverse=True)
            job_ids = [job.job_id for job in jobs[: max(1, min(int(limit or 20), 100))]]
        return [snapshot for job_id in job_ids if (snapshot := self.snapshot(job_id)) is not None]

    def active_for_record(self, action: str, *, owner_email: str, record_id: str) -> dict[str, Any] | None:
        normalized_owner = str(owner_email or "").strip().lower()
        normalized_record_id = str(record_id or "").strip()
        with self._lock:
            self._refresh_locked()
            candidates = [
                job
                for job in self._jobs.values()
                if job.action == action
                and job.state in {"queued", "running"}
                and str(job.owner_email or "").strip().lower() == normalized_owner
                and str(job.record_id or "").strip() == normalized_record_id
            ]
            if not candidates:
                return None
            latest = max(candidates, key=lambda item: item.updated_at)
            return asdict(latest)


class SourceCodeQAQueryScheduler:
    def __init__(self, *, job_store: JobStore, max_running: int = 2) -> None:
        self.job_store = job_store
        self.max_running = max(1, int(max_running or 2))
        self._lock = threading.Lock()
        self._user_queues: dict[str, deque[tuple[str, Flask, dict[str, Any], Any]]] = {}
        self._user_order: deque[str] = deque()
        self._running: set[str] = set()
        self._running_users: dict[str, int] = {}

    def submit(self, *, app: Flask, job_id: str, payload: dict[str, Any], owner_email: str, runner: Any | None = None) -> None:
        user_key = str(owner_email or "local").strip().lower() or "local"
        with self._lock:
            if user_key not in self._user_queues:
                self._user_queues[user_key] = deque()
                self._user_order.append(user_key)
            self._user_queues[user_key].append((job_id, app, dict(payload), runner))
            self.job_store.set_owner(job_id, user_key)
            self.job_store.set_query_mode(job_id, _source_code_qa_query_mode(payload.get("query_mode")))
            self._refresh_queue_metadata_locked()
            self._start_available_locked()

    def finish(self, job_id: str, owner_email: str) -> None:
        user_key = str(owner_email or "local").strip().lower() or "local"
        with self._lock:
            self._running.discard(job_id)
            if user_key in self._running_users:
                self._running_users[user_key] = max(0, self._running_users[user_key] - 1)
                if self._running_users[user_key] <= 0:
                    self._running_users.pop(user_key, None)
            self._refresh_queue_metadata_locked()
            self._start_available_locked()

    def _start_available_locked(self) -> None:
        while len(self._running) < self.max_running:
            next_item = self._pop_next_locked()
            if next_item is None:
                break
            user_key, job_id, app, payload, runner = next_item
            self._running.add(job_id)
            self._running_users[user_key] = self._running_users.get(user_key, 0) + 1
            self.job_store.update_queue_metadata(
                job_id,
                queued_position=0,
                eta_seconds_range=[],
                running_user_count=len(self._running_users),
                message="Starting Source Code Q&A job.",
            )
            thread = threading.Thread(
                target=self._run_job,
                args=(app, job_id, payload, user_key, runner),
                daemon=True,
            )
            thread.start()
        self._refresh_queue_metadata_locked()

    def _pop_next_locked(self) -> tuple[str, str, Flask, dict[str, Any], Any] | None:
        self._user_order = deque(user_key for user_key in self._user_order if self._user_queues.get(user_key))
        if not self._user_order:
            return None
        ordered = list(self._user_order)
        selected_user = min(ordered, key=lambda user_key: (self._running_users.get(user_key, 0), ordered.index(user_key)))
        self._user_order.remove(selected_user)
        queue = self._user_queues.get(selected_user)
        if not queue:
            self._user_queues.pop(selected_user, None)
            return None
        job_id, app, payload, runner = queue.popleft()
        if queue:
            self._user_order.append(selected_user)
        else:
            self._user_queues.pop(selected_user, None)
        return selected_user, job_id, app, payload, runner

    def _refresh_queue_metadata_locked(self) -> None:
        ordered_jobs = self._simulated_round_robin_locked()
        p95_seconds = self.job_store.p95_duration_seconds("source-code-qa-query", default_seconds=212)
        for index, (job_id, _user_key) in enumerate(ordered_jobs, start=1):
            waves_ahead = max(0, (index - 1) // self.max_running)
            lower = waves_ahead * p95_seconds
            upper = max(lower + 60, (waves_ahead + 1) * p95_seconds)
            self.job_store.update_queue_metadata(
                job_id,
                queued_position=index,
                eta_seconds_range=[lower, upper],
                running_user_count=len(self._running_users),
                message=f"Queued behind {max(0, index - 1)} Source Code Q&A job(s).",
            )

    def _simulated_round_robin_locked(self) -> list[tuple[str, str]]:
        queues = {
            user_key: deque((job_id, user_key) for job_id, _app, _payload, _runner in queue)
            for user_key, queue in self._user_queues.items()
            if queue
        }
        order = deque(user_key for user_key in self._user_order if user_key in queues)
        result: list[tuple[str, str]] = []
        while order:
            user_key = order.popleft()
            queue = queues.get(user_key)
            if not queue:
                continue
            result.append(queue.popleft())
            if queue:
                order.append(user_key)
        return result

    def _run_job(self, app: Flask, job_id: str, payload: dict[str, Any], owner_email: str, runner: Any | None = None) -> None:
        try:
            (runner or _run_source_code_qa_query_job)(app, job_id, payload)
        finally:
            self.finish(job_id, owner_email)


class TeamDashboardConfigStore:
    CONFIG_KEY = "team_dashboard"

    def __init__(self, db_path: Path) -> None:
        self.db_path = db_path
        self._ensure_db()

    def load(self) -> dict[str, Any]:
        with sqlite3.connect(self.db_path) as connection:
            row = connection.execute(
                "SELECT config_json FROM team_dashboard_configs WHERE config_key = ?",
                (self.CONFIG_KEY,),
            ).fetchone()
        if not row:
            return self.default_config()
        try:
            payload = json.loads(row[0])
        except (TypeError, json.JSONDecodeError):
            return self.default_config()
        return self.normalize_config(payload)

    def save(self, config: dict[str, Any]) -> dict[str, Any]:
        normalized = self.normalize_config(config)
        with sqlite3.connect(self.db_path) as connection:
            connection.execute(
                """
                INSERT INTO team_dashboard_configs (config_key, config_json, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(config_key) DO UPDATE SET
                    config_json = excluded.config_json,
                    updated_at = CURRENT_TIMESTAMP
                """,
                (self.CONFIG_KEY, json.dumps(normalized, ensure_ascii=False)),
            )
            connection.commit()
        return normalized

    def default_config(self) -> dict[str, Any]:
        return {
            "teams": {
                team_key: {
                    "label": label,
                    "member_emails": list(TEAM_DASHBOARD_DEFAULT_MEMBER_EMAILS_BY_TEAM.get(team_key, ())),
                }
                for team_key, label in TEAM_DASHBOARD_TEAMS.items()
            },
            "key_project_overrides": {},
            "monthly_report_template": DEFAULT_MONTHLY_REPORT_TEMPLATE,
            "report_intelligence_config": normalize_report_intelligence_config({}),
        }

    def normalize_config(self, config: dict[str, Any]) -> dict[str, Any]:
        raw_teams = config.get("teams") if isinstance(config, dict) else {}
        raw_teams = raw_teams if isinstance(raw_teams, dict) else {}
        raw_key_project_overrides = config.get("key_project_overrides") if isinstance(config, dict) else {}
        raw_key_project_overrides = raw_key_project_overrides if isinstance(raw_key_project_overrides, dict) else {}
        raw_monthly_report_template = config.get("monthly_report_template") if isinstance(config, dict) else ""
        raw_report_intelligence_config = config.get("report_intelligence_config") if isinstance(config, dict) else {}
        raw_task_cache = config.get("task_cache") if isinstance(config, dict) else {}
        raw_task_cache = raw_task_cache if isinstance(raw_task_cache, dict) else {}
        default = self.default_config()
        normalized_teams: dict[str, dict[str, Any]] = {}
        for team_key, label in TEAM_DASHBOARD_TEAMS.items():
            raw_team = raw_teams.get(team_key) if isinstance(raw_teams.get(team_key), dict) else {}
            raw_emails = raw_team.get("member_emails") if isinstance(raw_team, dict) else None
            if raw_emails is None:
                raw_emails = default["teams"][team_key]["member_emails"]
            normalized_emails = _normalize_team_dashboard_emails(raw_emails)
            if set(normalized_emails) == set(TEAM_DASHBOARD_LEGACY_DEFAULT_MEMBER_EMAILS):
                normalized_emails = list(default["teams"][team_key]["member_emails"])
            normalized_teams[team_key] = {
                "label": label,
                "member_emails": normalized_emails,
            }
        normalized_key_project_overrides: dict[str, dict[str, Any]] = {}
        for raw_bpmis_id, raw_override in raw_key_project_overrides.items():
            bpmis_id = str(raw_bpmis_id or "").strip()
            if not bpmis_id or not isinstance(raw_override, dict) or "is_key_project" not in raw_override:
                continue
            normalized_key_project_overrides[bpmis_id] = {
                "is_key_project": bool(raw_override.get("is_key_project")),
                "updated_by": str(raw_override.get("updated_by") or "").strip().lower(),
                "updated_at": str(raw_override.get("updated_at") or "").strip(),
            }
        return {
            "teams": normalized_teams,
            "key_project_overrides": normalized_key_project_overrides,
            "monthly_report_template": normalize_monthly_report_template(raw_monthly_report_template),
            "report_intelligence_config": normalize_report_intelligence_config(raw_report_intelligence_config),
            "task_cache": self._normalize_task_cache(raw_task_cache),
        }

    def _normalize_task_cache(self, task_cache: dict[str, Any]) -> dict[str, Any]:
        version = int(task_cache.get("version") or 1)
        if version != TEAM_DASHBOARD_TASK_CACHE_VERSION:
            return {
                "version": TEAM_DASHBOARD_TASK_CACHE_VERSION,
                "updated_at": "",
                "teams": {},
            }
        raw_teams = task_cache.get("teams") if isinstance(task_cache.get("teams"), dict) else {}
        teams: dict[str, dict[str, Any]] = {}
        for team_key in TEAM_DASHBOARD_TEAMS:
            raw_team = raw_teams.get(team_key)
            if not isinstance(raw_team, dict):
                continue
            teams[team_key] = {
                **raw_team,
                "team_key": team_key,
                "email_signature": str(raw_team.get("email_signature") or "").strip(),
                "cached_at": str(raw_team.get("cached_at") or "").strip(),
                "loaded": bool(raw_team.get("loaded", True)),
                "loading": False,
                "error": "",
                "progress_text": "",
                "under_prd": raw_team.get("under_prd") if isinstance(raw_team.get("under_prd"), list) else [],
                "pending_live": raw_team.get("pending_live") if isinstance(raw_team.get("pending_live"), list) else [],
            }
        return {
            "version": TEAM_DASHBOARD_TASK_CACHE_VERSION,
            "updated_at": str(task_cache.get("updated_at") or "").strip(),
            "teams": teams,
        }

    def _ensure_db(self) -> None:
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(self.db_path) as connection:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS team_dashboard_configs (
                    config_key TEXT PRIMARY KEY,
                    config_json TEXT NOT NULL,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            connection.commit()


class SeaTalkTodoStore:
    def __init__(self, storage_path: Path | None = None) -> None:
        self.storage_path = storage_path
        self._payload = self._load()
        self._lock = threading.Lock()

    def _load(self) -> dict[str, Any]:
        if self.storage_path is None or not self.storage_path.exists():
            return {"owners": {}}
        try:
            payload = json.loads(self.storage_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {"owners": {}}
        return payload if isinstance(payload, dict) else {"owners": {}}

    def _persist_locked(self) -> None:
        if self.storage_path is None:
            return
        try:
            self.storage_path.parent.mkdir(parents=True, exist_ok=True)
            payload = {**self._payload, "updated_at": time.time()}
            temp_path = self.storage_path.with_name(f".{self.storage_path.name}.{os.getpid()}.tmp")
            temp_path.write_text(json.dumps(payload, ensure_ascii=False, sort_keys=True), encoding="utf-8")
            os.replace(temp_path, self.storage_path)
        except OSError:
            return

    @staticmethod
    def _now() -> str:
        return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

    @staticmethod
    def todo_id(todo: dict[str, Any]) -> str:
        explicit = str(todo.get("id") or "").strip()
        if explicit:
            return explicit
        stable = "|".join(
            (
                re.sub(r"[^a-z0-9]+", " ", str(todo.get("domain") or "").lower()).strip(),
                re.sub(r"[^a-z0-9]+", " ", str(todo.get("task") or "").lower()).strip(),
            )
        )
        return hashlib.sha256(stable.encode("utf-8")).hexdigest()[:16]

    def completed_ids(self, *, owner_email: str) -> set[str]:
        owner = str(owner_email or "").strip().lower()
        with self._lock:
            owners = self._payload.get("owners") if isinstance(self._payload.get("owners"), dict) else {}
            owner_payload = owners.get(owner) if isinstance(owners.get(owner), dict) else {}
            completed = owner_payload.get("completed") if isinstance(owner_payload.get("completed"), dict) else {}
            return {str(todo_id) for todo_id in completed if str(todo_id).strip()}

    def processed_until(self, *, owner_email: str) -> str:
        owner = str(owner_email or "").strip().lower()
        with self._lock:
            owners = self._payload.get("owners") if isinstance(self._payload.get("owners"), dict) else {}
            owner_payload = owners.get(owner) if isinstance(owners.get(owner), dict) else {}
            return str(owner_payload.get("todo_processed_until") or "").strip()

    def mark_processed_until(self, *, owner_email: str, processed_until: str) -> None:
        owner = str(owner_email or "").strip().lower()
        value = str(processed_until or "").strip()
        if not owner or not value:
            return
        with self._lock:
            owners = self._payload.setdefault("owners", {})
            owner_payload = owners.setdefault(owner, {})
            current_value = str(owner_payload.get("todo_processed_until") or "").strip()
            if current_value and current_value >= value:
                return
            owner_payload["todo_processed_until"] = value
            self._persist_locked()

    def open_todos(self, *, owner_email: str) -> list[dict[str, Any]]:
        owner = str(owner_email or "").strip().lower()
        with self._lock:
            owners = self._payload.get("owners") if isinstance(self._payload.get("owners"), dict) else {}
            owner_payload = owners.get(owner) if isinstance(owners.get(owner), dict) else {}
            open_items = owner_payload.get("open") if isinstance(owner_payload.get("open"), dict) else {}
            return [dict(todo) for todo in open_items.values() if isinstance(todo, dict)]

    def merge_open_todos(self, *, owner_email: str, todos: list[dict[str, Any]]) -> list[dict[str, Any]]:
        owner = str(owner_email or "").strip().lower()
        if not owner:
            return []
        with self._lock:
            owners = self._payload.setdefault("owners", {})
            owner_payload = owners.setdefault(owner, {})
            open_items = owner_payload.setdefault("open", {})
            completed = owner_payload.get("completed") if isinstance(owner_payload.get("completed"), dict) else {}
            for todo in todos:
                if not isinstance(todo, dict):
                    continue
                todo_id = self.todo_id(todo)
                if not todo_id or todo_id in completed:
                    continue
                similar_id = self._find_similar_open_todo_id(open_items=open_items, todo=todo)
                if similar_id:
                    existing = open_items.get(similar_id) if isinstance(open_items.get(similar_id), dict) else {}
                    open_items[similar_id] = self._merge_similar_open_todo(existing=existing, incoming=todo, todo_id=similar_id)
                    continue
                open_items[todo_id] = {**todo, "id": todo_id, "last_seen_at": self._now()}
            self._persist_locked()
            return [dict(todo) for todo in open_items.values() if isinstance(todo, dict)]

    @classmethod
    def _find_similar_open_todo_id(cls, *, open_items: dict[str, Any], todo: dict[str, Any]) -> str | None:
        for existing_id, existing in open_items.items():
            if not isinstance(existing, dict):
                continue
            if cls._todos_are_similar(existing, todo):
                return str(existing.get("id") or existing_id)
        return None

    @classmethod
    def _todos_are_similar(cls, left: dict[str, Any], right: dict[str, Any]) -> bool:
        left_task = cls._similarity_text(left.get("task"))
        right_task = cls._similarity_text(right.get("task"))
        if not left_task or not right_task:
            return False
        left_domain = SeaTalkDashboardService._normalize_insight_domain(left.get("domain"))
        right_domain = SeaTalkDashboardService._normalize_insight_domain(right.get("domain"))
        same_domain = left_domain == right_domain
        sequence_score = difflib.SequenceMatcher(None, left_task, right_task).ratio()
        token_score = cls._token_overlap_score(left_task, right_task)
        score = max(sequence_score, token_score)
        if score >= (0.78 if same_domain else 0.9):
            return True
        if not same_domain:
            return False
        if not cls._todo_due_compatible(left.get("due"), right.get("due")):
            return False
        left_tokens = cls._informative_todo_tokens(left)
        right_tokens = cls._informative_todo_tokens(right)
        if not left_tokens or not right_tokens:
            return False
        overlap_count = len(left_tokens & right_tokens)
        overlap_ratio = overlap_count / max(1, min(len(left_tokens), len(right_tokens)))
        return overlap_count >= 4 and overlap_ratio >= 0.42

    @staticmethod
    def _similarity_text(value: Any) -> str:
        return re.sub(r"[^\w\u4e00-\u9fff]+", " ", str(value or "").lower()).strip()

    @staticmethod
    def _token_overlap_score(left: str, right: str) -> float:
        left_tokens = {token for token in left.split() if len(token) > 1}
        right_tokens = {token for token in right.split() if len(token) > 1}
        if not left_tokens or not right_tokens:
            return 0.0
        return len(left_tokens & right_tokens) / len(left_tokens | right_tokens)

    @classmethod
    def _informative_todo_tokens(cls, todo: dict[str, Any]) -> set[str]:
        text = cls._similarity_text(f"{todo.get('task') or ''} {todo.get('evidence') or ''}")
        stopwords = {
            "about", "accepted", "after", "aligned", "also", "and", "another", "any", "are", "arrange",
            "asks", "attend", "complete", "discussion", "follow", "for", "from", "help", "if", "invited",
            "join", "keep", "meeting", "needed", "on", "or", "plan", "prepare", "remaining", "says", "session",
            "support", "task", "the", "to", "tool", "up", "with", "xiaodong",
        }
        tokens = {token for token in text.split() if len(token) > 1 and token not in stopwords}
        normalized: set[str] = set()
        for token in tokens:
            normalized.add(token)
            if "-" in token:
                normalized.update(part for part in token.split("-") if len(part) > 1 and part not in stopwords)
        return normalized

    @classmethod
    def _todo_due_compatible(cls, left: Any, right: Any) -> bool:
        left_text = str(left or "").strip().lower()
        right_text = str(right or "").strip().lower()
        if not left_text or left_text == "unknown" or not right_text or right_text == "unknown":
            return True
        left_date = cls._todo_due_date(left_text)
        right_date = cls._todo_due_date(right_text)
        if left_date and right_date:
            return left_date == right_date
        return left_text == right_text

    @staticmethod
    def _todo_due_date(value: str) -> str:
        match = re.search(r"\b(20\d{2}-\d{2}-\d{2})\b", value)
        return match.group(1) if match else ""

    @classmethod
    def _merge_similar_open_todo(cls, *, existing: dict[str, Any], incoming: dict[str, Any], todo_id: str) -> dict[str, Any]:
        merged = {**existing, "id": todo_id, "last_seen_at": cls._now()}
        priority_rank = {"high": 0, "medium": 1, "low": 2, "unknown": 3}
        incoming_priority = str(incoming.get("priority") or "unknown").strip().lower()
        existing_priority = str(existing.get("priority") or "unknown").strip().lower()
        if priority_rank.get(incoming_priority, 3) < priority_rank.get(existing_priority, 3):
            merged["priority"] = incoming_priority
        existing_due = str(existing.get("due") or "").strip()
        incoming_due = str(incoming.get("due") or "").strip()
        if (not existing_due or existing_due.lower() == "unknown") and incoming_due:
            merged["due"] = incoming_due
        if not str(existing.get("evidence") or "").strip() and str(incoming.get("evidence") or "").strip():
            merged["evidence"] = str(incoming.get("evidence") or "").strip()
        return merged

    def mark_completed(self, *, owner_email: str, todo: dict[str, Any]) -> dict[str, Any]:
        owner = str(owner_email or "").strip().lower()
        todo_id = self.todo_id(todo)
        if not owner or not todo_id:
            raise ToolError("SeaTalk to-do completion requires a signed-in owner and a valid task.")
        with self._lock:
            owners = self._payload.setdefault("owners", {})
            owner_payload = owners.setdefault(owner, {})
            completed = owner_payload.setdefault("completed", {})
            completed[todo_id] = {
                "id": todo_id,
                "task": str(todo.get("task") or "").strip(),
                "domain": str(todo.get("domain") or "").strip(),
                "due": str(todo.get("due") or "").strip(),
                "completed_at": self._now(),
            }
            open_items = owner_payload.get("open") if isinstance(owner_payload.get("open"), dict) else {}
            open_items.pop(todo_id, None)
            self._persist_locked()
            return {"status": "ok", "todo_id": todo_id, "completed_at": completed[todo_id]["completed_at"]}


class SeaTalkNameMappingStore:
    def __init__(self, storage_path: Path | None = None) -> None:
        self.storage_path = storage_path
        self._payload = self._load()
        self._lock = threading.Lock()

    @staticmethod
    def normalize_key(value: Any) -> str:
        key = str(value or "").strip()
        if key.startswith("group-") or key.startswith("buddy-"):
            return key
        uid_match = re.match(r"^UID\s+(.+)$", key, re.IGNORECASE)
        if uid_match and uid_match.group(1).strip():
            return f"UID {uid_match.group(1).strip()}"
        return ""

    @staticmethod
    def person_aliases(key: str) -> set[str]:
        if key.startswith("buddy-"):
            suffix = key.removeprefix("buddy-").strip()
            return {f"UID {suffix}"} if suffix else set()
        uid_match = re.match(r"^UID\s+(.+)$", key, re.IGNORECASE)
        if uid_match and uid_match.group(1).strip():
            return {f"buddy-{uid_match.group(1).strip()}"}
        return set()

    @classmethod
    def equivalent_keys(cls, value: Any) -> set[str]:
        key = cls.normalize_key(value)
        return {key, *cls.person_aliases(key)} if key else set()

    @classmethod
    def canonical_display_key(cls, value: Any) -> str:
        key = cls.normalize_key(value)
        if key.startswith("buddy-"):
            suffix = key.removeprefix("buddy-").strip()
            return f"UID {suffix}" if suffix else key
        return key

    @classmethod
    def normalize_mappings(cls, value: Any) -> dict[str, str]:
        if not isinstance(value, dict):
            return {}
        mappings: dict[str, str] = {}
        for raw_key, raw_name in value.items():
            key = cls.normalize_key(raw_key)
            name = " ".join(str(raw_name or "").split())
            if key and name:
                mappings[key] = name[:180]
                for alias in cls.person_aliases(key):
                    mappings[alias] = name[:180]
        return mappings

    def _load(self) -> dict[str, Any]:
        if self.storage_path is None or not self.storage_path.exists():
            return {"mappings": {}}
        try:
            payload = json.loads(self.storage_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {"mappings": {}}
        if not isinstance(payload, dict):
            return {"mappings": {}}
        payload["mappings"] = self.normalize_mappings(payload.get("mappings") if "mappings" in payload else payload)
        return payload

    def _persist_locked(self) -> None:
        if self.storage_path is None:
            return
        try:
            self.storage_path.parent.mkdir(parents=True, exist_ok=True)
            payload = {**self._payload, "updated_at": time.time()}
            temp_path = self.storage_path.with_name(f".{self.storage_path.name}.{os.getpid()}.tmp")
            temp_path.write_text(json.dumps(payload, ensure_ascii=False, sort_keys=True), encoding="utf-8")
            os.replace(temp_path, self.storage_path)
        except OSError:
            return

    def mappings(self) -> dict[str, str]:
        with self._lock:
            return dict(self._payload.get("mappings") or {})

    def replace_mappings(self, mappings: dict[str, str]) -> dict[str, str]:
        normalized = self.normalize_mappings(mappings)
        with self._lock:
            self._payload["mappings"] = normalized
            self._persist_locked()
            return dict(normalized)

    def merge_mappings(self, mappings: dict[str, str]) -> dict[str, str]:
        normalized = self.normalize_mappings(mappings)
        with self._lock:
            current = dict(self._payload.get("mappings") or {})
            current.update(normalized)
            self._payload["mappings"] = current
            self._persist_locked()
            return dict(current)


def _dedupe_seatalk_name_mapping_candidates(rows: Any) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for row in rows if isinstance(rows, list) else []:
        if not isinstance(row, dict):
            continue
        canonical_id = SeaTalkNameMappingStore.canonical_display_key(row.get("id"))
        if not canonical_id:
            continue
        if canonical_id not in merged:
            order.append(canonical_id)
            merged[canonical_id] = {**row, "id": canonical_id}
        else:
            current = merged[canonical_id]
            current["count"] = _safe_int(current.get("count")) + _safe_int(row.get("count"))
            if not current.get("example") and row.get("example"):
                current["example"] = row.get("example")
            if _seatalk_mapping_reason_rank(row.get("priority_reason")) > _seatalk_mapping_reason_rank(current.get("priority_reason")):
                current["priority_reason"] = row.get("priority_reason")
            if current.get("type") != "group" and row.get("type") == "group":
                current["type"] = "group"
    return [merged[key] for key in order]


def _seatalk_mapping_reason_rank(value: Any) -> int:
    reason = str(value or "").strip().lower()
    if "@mentioned" in reason:
        return 3
    if "direct" in reason or "private" in reason:
        return 2
    return 1 if reason else 0


def _safe_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _safe_float(value: Any) -> float | None:
    try:
        if value in (None, ""):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


class SourceCodeQASessionStore:
    def __init__(self, storage_path: Path | None = None) -> None:
        self.storage_path = storage_path
        self._sessions: dict[str, dict[str, Any]] = self._load()
        self._lock = threading.Lock()

    def _load(self) -> dict[str, dict[str, Any]]:
        if self.storage_path is None or not self.storage_path.exists():
            return {}
        try:
            payload = json.loads(self.storage_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
        sessions: dict[str, dict[str, Any]] = {}
        for session_id, raw_session in (payload.get("sessions") or {}).items():
            if isinstance(raw_session, dict):
                sessions[str(session_id)] = raw_session
        return sessions

    def _persist_locked(self) -> None:
        if self.storage_path is None:
            return
        try:
            self.storage_path.parent.mkdir(parents=True, exist_ok=True)
            payload = {
                "updated_at": time.time(),
                "sessions": self._sessions,
            }
            temp_path = self.storage_path.with_name(f".{self.storage_path.name}.{os.getpid()}.tmp")
            temp_path.write_text(json.dumps(payload, ensure_ascii=False, sort_keys=True), encoding="utf-8")
            os.replace(temp_path, self.storage_path)
        except OSError:
            return

    @staticmethod
    def _now() -> str:
        return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

    @staticmethod
    def _title_from_question(question: str) -> str:
        normalized = re.sub(r"\s+", " ", str(question or "").strip())
        if not normalized:
            return "New Source Code Chat"
        return normalized[:72] + ("..." if len(normalized) > 72 else "")

    def create(
        self,
        *,
        owner_email: str,
        pm_team: str,
        country: str,
        llm_provider: str,
        title: str = "",
    ) -> dict[str, Any]:
        now = self._now()
        session_payload = {
            "id": uuid.uuid4().hex,
            "owner_email": str(owner_email or "").strip().lower(),
            "title": str(title or "").strip() or "New Source Code Chat",
            "pm_team": str(pm_team or "").strip() or "AF",
            "country": str(country or "").strip() or ALL_COUNTRY,
            "llm_provider": SourceCodeQAService.normalize_query_llm_provider(llm_provider),
            "created_at": now,
            "updated_at": now,
            "messages": [],
            "last_context": None,
            "last_trace_id": "",
            "archived_at": "",
            "archived_by": "",
        }
        with self._lock:
            self._sessions[session_payload["id"]] = session_payload
            self._persist_locked()
            return dict(session_payload)

    def list(self, *, owner_email: str, limit: int = 30) -> list[dict[str, Any]]:
        owner = str(owner_email or "").strip().lower()
        with self._lock:
            sessions = [
                self._public_session(session_payload, include_messages=False)
                for session_payload in self._sessions.values()
                if str(session_payload.get("owner_email") or "").strip().lower() == owner
                and not str(session_payload.get("archived_at") or "").strip()
            ]
        sessions.sort(key=lambda item: str(item.get("updated_at") or ""), reverse=True)
        return sessions[: max(1, min(int(limit or 30), 100))]

    def get(self, session_id: str, *, owner_email: str) -> dict[str, Any] | None:
        owner = str(owner_email or "").strip().lower()
        with self._lock:
            session_payload = self._sessions.get(str(session_id or ""))
            if not session_payload:
                return None
            if str(session_payload.get("owner_email") or "").strip().lower() != owner:
                return None
            return self._public_session(session_payload, include_messages=True)

    def archive(self, session_id: str, *, owner_email: str) -> dict[str, Any] | None:
        owner = str(owner_email or "").strip().lower()
        now = self._now()
        with self._lock:
            session_payload = self._sessions.get(str(session_id or ""))
            if not session_payload:
                return None
            if str(session_payload.get("owner_email") or "").strip().lower() != owner:
                return None
            session_payload["archived_at"] = now
            session_payload["archived_by"] = owner
            session_payload["updated_at"] = now
            self._persist_locked()
            return {
                "status": "ok",
                "session_id": session_payload.get("id") or "",
                "archived_at": now,
            }

    def get_context(self, session_id: str, *, owner_email: str) -> dict[str, Any] | None:
        session_payload = self.get(session_id, owner_email=owner_email)
        context = session_payload.get("last_context") if session_payload else None
        if not isinstance(context, dict):
            return None
        enriched = dict(context)
        enriched.setdefault("session_title", session_payload.get("title") or "")
        return enriched

    @staticmethod
    def _recent_turn_from_context(context: dict[str, Any] | None) -> dict[str, Any] | None:
        if not isinstance(context, dict):
            return None
        question = str(context.get("question") or "").strip()
        answer = str(context.get("answer") or context.get("rendered_answer") or "").strip()
        if not question and not answer:
            return None
        llm_route = context.get("llm_route") if isinstance(context.get("llm_route"), dict) else {}
        return {
            "question": question[:500],
            "answer": answer[:1200],
            "summary": str(context.get("summary") or "")[:500],
            "trace_id": str(context.get("trace_id") or "")[:80],
            "attachments": [
                item for item in (context.get("attachments") or [])[:5]
                if isinstance(item, dict)
            ],
            "llm_provider": str(context.get("llm_provider") or "")[:80],
            "llm_model": str(context.get("llm_model") or "")[:120],
            "matches_snapshot": [
                item for item in (context.get("matches_snapshot") or context.get("matches") or [])[:8]
                if isinstance(item, dict)
            ],
            "codex_candidate_paths": [
                item for item in (
                    context.get("codex_candidate_paths")
                    or (llm_route.get("candidate_paths") if isinstance(llm_route, dict) else [])
                    or []
                )[:12]
                if isinstance(item, dict)
            ],
            "evidence_pack": (
                context.get("evidence_pack")
                if isinstance(context.get("evidence_pack"), dict)
                else {}
            ),
        }

    @classmethod
    def _extend_recent_turns(cls, context: dict[str, Any], previous_context: dict[str, Any] | None) -> dict[str, Any]:
        enriched = dict(context or {})
        recent_turns = [
            item for item in (
                previous_context.get("recent_turns", []) if isinstance(previous_context, dict) else []
            )
            if isinstance(item, dict)
        ]
        previous_turn = cls._recent_turn_from_context(previous_context)
        if previous_turn:
            recent_turns.append(previous_turn)
        deduped: list[dict[str, Any]] = []
        seen: set[tuple[str, str]] = set()
        for item in recent_turns:
            key = (str(item.get("question") or ""), str(item.get("trace_id") or ""))
            if key in seen:
                continue
            seen.add(key)
            deduped.append(item)
        try:
            max_turns = int(enriched.get("codex_session_max_turns") or 8)
        except (TypeError, ValueError):
            max_turns = 8
        enriched["recent_turns"] = deduped[-max(1, min(max_turns, 30)):]
        return enriched

    def append_exchange(
        self,
        session_id: str,
        *,
        owner_email: str,
        pm_team: str,
        country: str,
        llm_provider: str,
        question: str,
        result: dict[str, Any],
        context: dict[str, Any],
        attachments: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any] | None:
        owner = str(owner_email or "").strip().lower()
        now = self._now()
        with self._lock:
            session_payload = self._sessions.get(str(session_id or ""))
            if not session_payload:
                return None
            if str(session_payload.get("owner_email") or "").strip().lower() != owner:
                return None
            title = str(session_payload.get("title") or "").strip()
            if not title or title == "New Source Code Chat":
                session_payload["title"] = self._title_from_question(question)
            previous_context = session_payload.get("last_context") if isinstance(session_payload.get("last_context"), dict) else None
            previous_scope = (
                str(session_payload.get("pm_team") or ""),
                str(session_payload.get("country") or ""),
                str(session_payload.get("llm_provider") or ""),
            )
            current_scope = (
                str(pm_team or "").strip() or str(session_payload.get("pm_team") or ""),
                str(country or "").strip() or str(session_payload.get("country") or ALL_COUNTRY),
                SourceCodeQAService.normalize_query_llm_provider(llm_provider),
            )
            session_payload["pm_team"] = current_scope[0] or "AF"
            session_payload["country"] = current_scope[1] or ALL_COUNTRY
            session_payload["llm_provider"] = current_scope[2]
            session_payload["updated_at"] = now
            next_context = self._extend_recent_turns(context, previous_context)
            if previous_scope != current_scope:
                next_context.pop("codex_cli_session", None)
            session_payload["last_context"] = next_context
            session_payload["last_trace_id"] = str(result.get("trace_id") or "")
            messages = list(session_payload.get("messages") or [])
            normalized_question = str(question or "").strip()
            messages = [
                message for message in messages
                if not (
                    isinstance(message, dict)
                    and message.get("role") == "user"
                    and message.get("pending")
                    and str(message.get("text") or "").strip() == normalized_question
                )
            ]
            messages.extend(
                [
                    {
                        "role": "user",
                        "text": normalized_question,
                        "created_at": now,
                        "attachments": [
                            SourceCodeQAAttachmentStore.public_metadata(item)
                            for item in (attachments or [])
                            if isinstance(item, dict)
                        ],
                    },
                    {
                        "role": "assistant",
                        "text": str(result.get("llm_answer") or result.get("summary") or ""),
                        "created_at": now,
                        "payload": _compact_source_code_qa_session_payload(result),
                    },
                ]
            )
            session_payload["messages"] = messages[-80:]
            self._persist_locked()
            return self._public_session(session_payload, include_messages=True)

    def append_pending_question(
        self,
        session_id: str,
        *,
        owner_email: str,
        pm_team: str,
        country: str,
        llm_provider: str,
        question: str,
        job_id: str,
        attachments: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any] | None:
        owner = str(owner_email or "").strip().lower()
        now = self._now()
        normalized_question = str(question or "").strip()
        normalized_job_id = str(job_id or "").strip()
        if not normalized_question or not normalized_job_id:
            return None
        with self._lock:
            session_payload = self._sessions.get(str(session_id or ""))
            if not session_payload:
                return None
            if str(session_payload.get("owner_email") or "").strip().lower() != owner:
                return None
            title = str(session_payload.get("title") or "").strip()
            if not title or title == "New Source Code Chat":
                session_payload["title"] = self._title_from_question(normalized_question)
            session_payload["pm_team"] = str(pm_team or "").strip() or str(session_payload.get("pm_team") or "AF")
            session_payload["country"] = str(country or "").strip() or str(session_payload.get("country") or ALL_COUNTRY)
            session_payload["llm_provider"] = SourceCodeQAService.normalize_query_llm_provider(llm_provider)
            session_payload["updated_at"] = now
            messages = [
                message for message in list(session_payload.get("messages") or [])
                if not (
                    isinstance(message, dict)
                    and message.get("role") == "user"
                    and message.get("pending")
                    and str(message.get("pending_job_id") or "") == normalized_job_id
                )
            ]
            messages.append(
                {
                    "role": "user",
                    "text": normalized_question,
                    "created_at": now,
                    "attachments": [
                        SourceCodeQAAttachmentStore.public_metadata(item)
                        for item in (attachments or [])
                        if isinstance(item, dict)
                    ],
                    "pending": True,
                    "pending_job_id": normalized_job_id,
                }
            )
            session_payload["messages"] = messages[-80:]
            self._persist_locked()
            return self._public_session(session_payload, include_messages=True)

    def _public_session(self, session_payload: dict[str, Any], *, include_messages: bool) -> dict[str, Any]:
        public_payload = {
            "id": session_payload.get("id") or "",
            "title": session_payload.get("title") or "New Source Code Chat",
            "pm_team": session_payload.get("pm_team") or "",
            "country": session_payload.get("country") or ALL_COUNTRY,
            "llm_provider": session_payload.get("llm_provider") or "codex_cli_bridge",
            "created_at": session_payload.get("created_at") or "",
            "updated_at": session_payload.get("updated_at") or "",
            "archived_at": session_payload.get("archived_at") or "",
            "last_context": session_payload.get("last_context") if include_messages else None,
            "last_trace_id": session_payload.get("last_trace_id") or "",
            "message_count": len(session_payload.get("messages") or []),
        }
        if include_messages:
            if isinstance(public_payload["last_context"], dict):
                public_payload["last_context"] = {
                    **public_payload["last_context"],
                    "session_title": public_payload["title"],
                }
            public_payload["messages"] = list(session_payload.get("messages") or [])
        return public_payload


class SourceCodeQAAttachmentStore:
    MAX_FILE_BYTES = 10 * 1024 * 1024
    MAX_ATTACHMENTS = 5
    MAX_IMAGES = 3
    IMAGE_MIME_TYPES = {"image/png", "image/jpeg", "image/webp", "image/gif"}
    TEXT_EXTENSIONS = {
        ".txt",
        ".md",
        ".csv",
        ".json",
        ".xml",
        ".yaml",
        ".yml",
        ".log",
        ".java",
        ".py",
        ".js",
        ".ts",
        ".tsx",
        ".sql",
        ".properties",
        ".kt",
        ".go",
        ".rb",
        ".php",
        ".html",
        ".css",
        ".sh",
    }
    DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".xlsx"}
    BLOCKED_EXTENSIONS = {
        ".app",
        ".bat",
        ".bin",
        ".cmd",
        ".com",
        ".dmg",
        ".exe",
        ".gz",
        ".jar",
        ".pkg",
        ".rar",
        ".tar",
        ".tgz",
        ".zip",
        ".7z",
    }

    def __init__(self, root_dir: Path | None = None) -> None:
        self.root_dir = root_dir
        self._lock = threading.Lock()

    @staticmethod
    def _now() -> str:
        return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

    @staticmethod
    def _owner_key(owner_email: str) -> str:
        owner = str(owner_email or "").strip().lower() or "local"
        return hashlib.sha256(owner.encode("utf-8")).hexdigest()[:24]

    @staticmethod
    def _safe_session_id(session_id: str) -> str:
        normalized = str(session_id or "").strip()
        if not normalized or not re.fullmatch(r"[A-Za-z0-9_-]{8,80}", normalized):
            raise ToolError("A valid Source Code Q&A session is required before uploading attachments.")
        return normalized

    @staticmethod
    def _safe_filename(filename: str) -> str:
        name = Path(str(filename or "attachment")).name.strip().replace("\x00", "")
        name = re.sub(r"[^A-Za-z0-9._ -]+", "_", name)[:180].strip(" .")
        return name or "attachment"

    def _session_dir(self, *, owner_email: str, session_id: str) -> Path:
        if self.root_dir is None:
            raise ToolError("Source Code Q&A attachments are not configured.")
        return self.root_dir / self._owner_key(owner_email) / self._safe_session_id(session_id)

    def _metadata_path(self, *, owner_email: str, session_id: str) -> Path:
        return self._session_dir(owner_email=owner_email, session_id=session_id) / "metadata.json"

    def _load_metadata_locked(self, *, owner_email: str, session_id: str) -> dict[str, dict[str, Any]]:
        path = self._metadata_path(owner_email=owner_email, session_id=session_id)
        if not path.exists():
            return {}
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
        attachments = payload.get("attachments") if isinstance(payload, dict) else {}
        return {str(key): value for key, value in attachments.items() if isinstance(value, dict)} if isinstance(attachments, dict) else {}

    def _persist_metadata_locked(self, *, owner_email: str, session_id: str, metadata: dict[str, dict[str, Any]]) -> None:
        path = self._metadata_path(owner_email=owner_email, session_id=session_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = path.with_name(f".{path.name}.{os.getpid()}.tmp")
        temp_path.write_text(
            json.dumps({"updated_at": self._now(), "attachments": metadata}, ensure_ascii=False, sort_keys=True),
            encoding="utf-8",
        )
        os.replace(temp_path, path)

    def save_bytes(
        self,
        *,
        owner_email: str,
        session_id: str,
        filename: str,
        content: bytes,
        mime_type: str = "",
    ) -> dict[str, Any]:
        if len(content or b"") <= 0:
            raise ToolError("Attachment file is empty.")
        if len(content) > self.MAX_FILE_BYTES:
            raise ToolError("Attachment is too large. Maximum size is 10MB per file.")
        safe_name = self._safe_filename(filename)
        suffix = Path(safe_name).suffix.lower()
        if suffix in self.BLOCKED_EXTENSIONS:
            raise ToolError("Executable, archive, and unknown binary attachments are not supported.")
        guessed_mime = str(mime_type or mimetypes.guess_type(safe_name)[0] or "").lower()
        kind = self._attachment_kind(safe_name, guessed_mime, content)
        digest = hashlib.sha256(content).hexdigest()
        attachment_id = uuid.uuid4().hex
        stored_name = f"{attachment_id}{suffix or '.bin'}"
        session_dir = self._session_dir(owner_email=owner_email, session_id=session_id)
        metadata = {
            "id": attachment_id,
            "filename": safe_name,
            "stored_name": stored_name,
            "mime_type": guessed_mime or "application/octet-stream",
            "kind": kind,
            "size": len(content),
            "sha256": digest,
            "created_at": self._now(),
            "summary": "",
            "text_char_count": 0,
        }
        if kind in {"text", "document"}:
            extracted = self._extract_attachment_text(safe_name, guessed_mime, content)
            metadata["text_char_count"] = len(extracted)
            metadata["summary"] = extracted[:2000]
        with self._lock:
            existing = self._load_metadata_locked(owner_email=owner_email, session_id=session_id)
            session_dir.mkdir(parents=True, exist_ok=True)
            (session_dir / stored_name).write_bytes(content)
            existing[attachment_id] = metadata
            self._persist_metadata_locked(owner_email=owner_email, session_id=session_id, metadata=existing)
        return self.public_metadata(metadata)

    def resolve_many(self, *, owner_email: str, session_id: str, attachment_ids: list[str]) -> list[dict[str, Any]]:
        requested_ids = [str(item or "").strip() for item in attachment_ids if str(item or "").strip()]
        if len(requested_ids) > self.MAX_ATTACHMENTS:
            raise ToolError(f"At most {self.MAX_ATTACHMENTS} Source Code Q&A attachments are supported per question.")
        with self._lock:
            metadata = self._load_metadata_locked(owner_email=owner_email, session_id=session_id)
        resolved: list[dict[str, Any]] = []
        image_count = 0
        session_dir = self._session_dir(owner_email=owner_email, session_id=session_id)
        for attachment_id in requested_ids:
            item = metadata.get(attachment_id)
            if not item:
                raise ToolError("One or more Source Code Q&A attachments were not found for this session.")
            path = session_dir / str(item.get("stored_name") or "")
            if not path.exists() or not path.is_file():
                raise ToolError(f"Attachment file is missing: {item.get('filename') or attachment_id}")
            enriched = dict(item)
            enriched["path"] = str(path)
            if enriched.get("kind") == "image":
                image_count += 1
                if image_count > self.MAX_IMAGES:
                    raise ToolError(f"At most {self.MAX_IMAGES} image attachments are supported per question.")
            elif enriched.get("kind") in {"text", "document"}:
                try:
                    content = path.read_bytes()
                except OSError as error:
                    raise ToolError(f"Attachment file is unreadable: {item.get('filename') or attachment_id}") from error
                enriched["text"] = self._extract_attachment_text(str(item.get("filename") or ""), str(item.get("mime_type") or ""), content)
            resolved.append(enriched)
        return resolved

    def get_bytes(self, *, owner_email: str, session_id: str, attachment_id: str) -> tuple[dict[str, Any], bytes]:
        resolved = self.resolve_many(owner_email=owner_email, session_id=session_id, attachment_ids=[attachment_id])
        if not resolved:
            raise ToolError("Source Code Q&A attachment was not found.")
        item = resolved[0]
        try:
            content = Path(str(item.get("path") or "")).read_bytes()
        except OSError as error:
            raise ToolError("Source Code Q&A attachment file is unreadable.") from error
        return self.public_metadata(item), content

    @classmethod
    def public_metadata(cls, metadata: dict[str, Any]) -> dict[str, Any]:
        return {
            "id": str(metadata.get("id") or ""),
            "filename": str(metadata.get("filename") or ""),
            "mime_type": str(metadata.get("mime_type") or ""),
            "kind": str(metadata.get("kind") or ""),
            "size": int(metadata.get("size") or 0),
            "sha256": str(metadata.get("sha256") or ""),
            "created_at": str(metadata.get("created_at") or ""),
            "summary": str(metadata.get("summary") or "")[:400],
            "text_char_count": int(metadata.get("text_char_count") or 0),
        }

    def _attachment_kind(self, filename: str, mime_type: str, content: bytes) -> str:
        suffix = Path(filename).suffix.lower()
        mime = str(mime_type or "").lower()
        if mime in self.IMAGE_MIME_TYPES:
            return "image"
        if suffix in self.TEXT_EXTENSIONS or mime.startswith("text/") or mime in {"application/json", "application/xml"}:
            return "text"
        if suffix in self.DOCUMENT_EXTENSIONS:
            return "document"
        if b"\x00" in content[:2048]:
            raise ToolError("Unknown binary attachments are not supported.")
        if suffix:
            return "text"
        raise ToolError("Unsupported attachment type.")

    def _extract_attachment_text(self, filename: str, mime_type: str, content: bytes) -> str:
        suffix = Path(filename).suffix.lower()
        if suffix == ".pdf":
            return self._extract_pdf_text(content)
        if suffix == ".docx":
            return self._extract_docx_text(content)
        if suffix == ".xlsx":
            return self._extract_xlsx_text(content)
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                text = ""
        if not text:
            raise ToolError(f"Unable to parse text from attachment {filename or mime_type}.")
        return re.sub(r"\r\n?", "\n", text).strip()[:16000]

    @staticmethod
    def _extract_pdf_text(content: bytes) -> str:
        try:
            from pypdf import PdfReader  # type: ignore
        except ImportError as error:
            raise ToolError("PDF attachments are supported only when pypdf is installed on the server.") from error
        reader = PdfReader(io.BytesIO(content))
        lines: list[str] = []
        for page in reader.pages[:10]:
            lines.append(str(page.extract_text() or ""))
        text = "\n".join(lines).strip()
        if not text:
            raise ToolError("Unable to extract readable text from this PDF attachment.")
        return text[:16000]

    @staticmethod
    def _extract_docx_text(content: bytes) -> str:
        try:
            from docx import Document  # type: ignore
        except ImportError as error:
            raise ToolError("DOCX attachments are supported only when python-docx is installed on the server.") from error
        document = Document(io.BytesIO(content))
        text = "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text).strip()
        if not text:
            raise ToolError("Unable to extract readable text from this DOCX attachment.")
        return text[:16000]

    @staticmethod
    def _extract_xlsx_text(content: bytes) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise ToolError("XLSX attachments are supported only when openpyxl is installed on the server.") from error
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows: list[str] = []
        for worksheet in workbook.worksheets[:3]:
            rows.append(f"[Sheet: {worksheet.title}]")
            for row in worksheet.iter_rows(max_row=40, max_col=12, values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if any(value.strip() for value in values):
                    rows.append("\t".join(values).rstrip())
        text = "\n".join(rows).strip()
        if not text:
            raise ToolError("Unable to extract readable text from this XLSX attachment.")
        return text[:16000]


class SourceCodeQARuntimeEvidenceStore(SourceCodeQAAttachmentStore):
    ALLOWED_PM_TEAMS = {"AF", "CRMS", "GRC"}
    ALLOWED_COUNTRIES = {"ID", "SG", "PH"}
    ALLOWED_SOURCE_TYPES = {"apollo", "db", "other"}
    MAX_FILES_PER_SCOPE = 20
    MAX_QUERY_FILES_PER_SCOPE = 8
    MAX_ZIP_MEMBERS = 500
    MAX_ZIP_UNCOMPRESSED_BYTES = 8 * 1024 * 1024
    MAX_ZIP_TEXT_CHARS = 120000
    ZIP_TEXT_EXTENSIONS = SourceCodeQAAttachmentStore.TEXT_EXTENSIONS | {
        ".conf",
        ".cfg",
        ".ini",
        ".toml",
        ".env",
    }

    @classmethod
    def _safe_scope(cls, *, pm_team: str, country: str) -> tuple[str, str]:
        normalized_team = str(pm_team or "").strip().upper()
        normalized_country = str(country or "").strip().upper()
        if normalized_team not in cls.ALLOWED_PM_TEAMS:
            raise ToolError("Runtime evidence PM Team must be one of AF, CRMS, or GRC.")
        if normalized_country not in cls.ALLOWED_COUNTRIES:
            raise ToolError("Runtime evidence country must be one of ID, SG, or PH.")
        return normalized_team, normalized_country

    @classmethod
    def _safe_source_type(cls, source_type: str) -> str:
        normalized = str(source_type or "").strip().lower() or "other"
        if normalized not in cls.ALLOWED_SOURCE_TYPES:
            raise ToolError("Runtime evidence source type must be apollo, db, or other.")
        return normalized

    def _scope_dir(self, *, pm_team: str, country: str) -> Path:
        if self.root_dir is None:
            raise ToolError("Source Code Q&A runtime evidence is not configured.")
        safe_team, safe_country = self._safe_scope(pm_team=pm_team, country=country)
        return self.root_dir / safe_team / safe_country

    def _metadata_path(self, *, pm_team: str, country: str) -> Path:
        return self._scope_dir(pm_team=pm_team, country=country) / "metadata.json"

    def _load_metadata_locked(self, *, pm_team: str, country: str) -> dict[str, dict[str, Any]]:
        path = self._metadata_path(pm_team=pm_team, country=country)
        if not path.exists():
            return {}
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
        items = payload.get("evidence") if isinstance(payload, dict) else {}
        return {str(key): value for key, value in items.items() if isinstance(value, dict)} if isinstance(items, dict) else {}

    def _persist_metadata_locked(self, *, pm_team: str, country: str, metadata: dict[str, dict[str, Any]]) -> None:
        path = self._metadata_path(pm_team=pm_team, country=country)
        path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = path.with_name(f".{path.name}.{os.getpid()}.tmp")
        temp_path.write_text(
            json.dumps({"updated_at": self._now(), "evidence": metadata}, ensure_ascii=False, sort_keys=True),
            encoding="utf-8",
        )
        os.replace(temp_path, path)

    def save_bytes(
        self,
        *,
        pm_team: str,
        country: str,
        source_type: str,
        uploaded_by: str,
        filename: str,
        content: bytes,
        mime_type: str = "",
    ) -> dict[str, Any]:
        if len(content or b"") <= 0:
            raise ToolError("Runtime evidence file is empty.")
        if len(content) > self.MAX_FILE_BYTES:
            raise ToolError("Runtime evidence is too large. Maximum size is 10MB per file.")
        safe_team, safe_country = self._safe_scope(pm_team=pm_team, country=country)
        safe_source_type = self._safe_source_type(source_type)
        safe_name = self._safe_filename(filename)
        suffix = Path(safe_name).suffix.lower()
        if suffix in self.BLOCKED_EXTENSIONS and suffix != ".zip":
            raise ToolError("Executable, archive, and unknown binary runtime evidence files are not supported.")
        guessed_mime = str(mime_type or mimetypes.guess_type(safe_name)[0] or "").lower()
        kind = "archive" if suffix == ".zip" else self._attachment_kind(safe_name, guessed_mime, content)
        if kind == "image":
            raise ToolError("Runtime evidence must be a parseable text, spreadsheet, PDF, or document file, not an image.")
        extracted = self._extract_attachment_text(safe_name, guessed_mime, content)
        digest = hashlib.sha256(content).hexdigest()
        evidence_id = uuid.uuid4().hex
        stored_name = f"{evidence_id}{suffix or '.txt'}"
        scope_dir = self._scope_dir(pm_team=safe_team, country=safe_country)
        metadata = {
            "id": evidence_id,
            "filename": safe_name,
            "stored_name": stored_name,
            "mime_type": guessed_mime or "application/octet-stream",
            "kind": kind,
            "source_type": safe_source_type,
            "pm_team": safe_team,
            "country": safe_country,
            "size": len(content),
            "sha256": digest,
            "uploaded_by": str(uploaded_by or "").strip().lower(),
            "created_at": self._now(),
            "summary": extracted[:2000],
            "text_char_count": len(extracted),
        }
        with self._lock:
            existing = self._load_metadata_locked(pm_team=safe_team, country=safe_country)
            scope_dir.mkdir(parents=True, exist_ok=True)
            (scope_dir / stored_name).write_bytes(content)
            existing[evidence_id] = metadata
            if len(existing) > self.MAX_FILES_PER_SCOPE:
                ordered = sorted(existing.values(), key=lambda item: str(item.get("created_at") or ""))
                for stale in ordered[: len(existing) - self.MAX_FILES_PER_SCOPE]:
                    stale_id = str(stale.get("id") or "")
                    stale_name = str(stale.get("stored_name") or "")
                    if stale_name:
                        try:
                            (scope_dir / stale_name).unlink(missing_ok=True)
                        except OSError:
                            pass
                    existing.pop(stale_id, None)
            self._persist_metadata_locked(pm_team=safe_team, country=safe_country, metadata=existing)
        return self.public_metadata(metadata)

    def _extract_attachment_text(self, filename: str, mime_type: str, content: bytes) -> str:
        if Path(filename).suffix.lower() == ".zip":
            return self._extract_zip_text(content)
        return super()._extract_attachment_text(filename, mime_type, content)

    def _extract_zip_text(self, content: bytes) -> str:
        try:
            archive = zipfile.ZipFile(io.BytesIO(content))
        except zipfile.BadZipFile as error:
            raise ToolError("Unable to read this ZIP runtime evidence file.") from error
        lines: list[str] = []
        total_uncompressed = 0
        readable_members = 0
        skipped_members = 0
        with archive:
            members = [member for member in archive.infolist() if not member.is_dir()]
            if len(members) > self.MAX_ZIP_MEMBERS:
                raise ToolError(f"ZIP runtime evidence contains too many files. Maximum is {self.MAX_ZIP_MEMBERS}.")
            for member in members:
                member_name = str(member.filename or "").replace("\\", "/")
                clean_parts = [part for part in member_name.split("/") if part and part not in {".", ".."}]
                if not clean_parts or clean_parts[0] == "__MACOSX" or len(clean_parts) != len([part for part in member_name.split("/") if part]):
                    skipped_members += 1
                    continue
                member_basename = clean_parts[-1].lower()
                suffix = Path(member_basename).suffix.lower() or (member_basename if member_basename.startswith(".") else "")
                if suffix in self.BLOCKED_EXTENSIONS or suffix not in self.ZIP_TEXT_EXTENSIONS:
                    skipped_members += 1
                    continue
                total_uncompressed += int(member.file_size or 0)
                if total_uncompressed > self.MAX_ZIP_UNCOMPRESSED_BYTES:
                    max_mb = self.MAX_ZIP_UNCOMPRESSED_BYTES // (1024 * 1024)
                    raise ToolError(f"ZIP runtime evidence is too large after extraction. Keep text config files under {max_mb}MB total.")
                try:
                    raw = archive.read(member)
                except (OSError, RuntimeError, zipfile.BadZipFile) as error:
                    raise ToolError(f"Unable to read {member_name} inside this ZIP runtime evidence file.") from error
                if b"\x00" in raw[:2048]:
                    skipped_members += 1
                    continue
                text = ""
                for encoding in ("utf-8-sig", "utf-8", "latin-1"):
                    try:
                        text = raw.decode(encoding)
                        break
                    except UnicodeDecodeError:
                        text = ""
                text = re.sub(r"\r\n?", "\n", text).strip()
                if not text:
                    skipped_members += 1
                    continue
                readable_members += 1
                lines.append(f"[ZIP file: {'/'.join(clean_parts)}]\n{text[:12000]}")
                if sum(len(line) for line in lines) >= self.MAX_ZIP_TEXT_CHARS:
                    lines.append("...[zip text truncated]")
                    break
        if not readable_members:
            raise ToolError("ZIP runtime evidence did not contain readable config/text files.")
        if skipped_members:
            lines.append(f"[ZIP skipped files: {skipped_members}]")
        return "\n\n".join(lines).strip()[: self.MAX_ZIP_TEXT_CHARS]

    @classmethod
    def public_metadata(cls, metadata: dict[str, Any]) -> dict[str, Any]:
        payload = SourceCodeQAAttachmentStore.public_metadata(metadata)
        payload.update(
            {
                "source_type": str(metadata.get("source_type") or ""),
                "pm_team": str(metadata.get("pm_team") or ""),
                "country": str(metadata.get("country") or ""),
                "uploaded_by": str(metadata.get("uploaded_by") or ""),
            }
        )
        return payload

    def list(self, *, pm_team: str, country: str) -> list[dict[str, Any]]:
        safe_team, safe_country = self._safe_scope(pm_team=pm_team, country=country)
        with self._lock:
            metadata = self._load_metadata_locked(pm_team=safe_team, country=safe_country)
        return [
            self.public_metadata(item)
            for item in sorted(metadata.values(), key=lambda value: str(value.get("created_at") or ""), reverse=True)
        ]

    def resolve_scope(self, *, pm_team: str, country: str) -> list[dict[str, Any]]:
        normalized_country = str(country or "").strip().upper()
        countries = sorted(self.ALLOWED_COUNTRIES) if normalized_country in {"", ALL_COUNTRY.upper()} else [normalized_country]
        resolved: list[dict[str, Any]] = []
        for scoped_country in countries:
            safe_team, safe_country = self._safe_scope(pm_team=pm_team, country=scoped_country)
            with self._lock:
                metadata = self._load_metadata_locked(pm_team=safe_team, country=safe_country)
            scope_dir = self._scope_dir(pm_team=safe_team, country=safe_country)
            for item in sorted(metadata.values(), key=lambda value: str(value.get("created_at") or ""), reverse=True)[: self.MAX_QUERY_FILES_PER_SCOPE]:
                path = scope_dir / str(item.get("stored_name") or "")
                if not path.exists() or not path.is_file():
                    continue
                enriched = dict(item)
                enriched["path"] = str(path)
                try:
                    content = path.read_bytes()
                except OSError:
                    continue
                enriched["text"] = self._extract_attachment_text(str(item.get("filename") or ""), str(item.get("mime_type") or ""), content)
                resolved.append(enriched)
        return resolved[: self.MAX_QUERY_FILES_PER_SCOPE * max(1, len(countries))]

    def delete(self, *, pm_team: str, country: str, evidence_id: str) -> bool:
        safe_team, safe_country = self._safe_scope(pm_team=pm_team, country=country)
        normalized_id = str(evidence_id or "").strip()
        if not re.fullmatch(r"[a-fA-F0-9]{32}", normalized_id):
            raise ToolError("Runtime evidence id is invalid.")
        scope_dir = self._scope_dir(pm_team=safe_team, country=safe_country)
        with self._lock:
            metadata = self._load_metadata_locked(pm_team=safe_team, country=safe_country)
            item = metadata.pop(normalized_id, None)
            if item is None:
                return False
            stored_name = str(item.get("stored_name") or "")
            if stored_name:
                try:
                    (scope_dir / stored_name).unlink(missing_ok=True)
                except OSError:
                    pass
            self._persist_metadata_locked(pm_team=safe_team, country=safe_country, metadata=metadata)
        return True


class SourceCodeQAModelAvailabilityStore:
    DEFAULT_AVAILABILITY = {
        "codex_cli_bridge": True,
        "gemini": False,
        "vertex_ai": True,
    }

    def __init__(self, storage_path: Path | None = None) -> None:
        self.storage_path = storage_path
        self._lock = threading.Lock()
        self._availability = self._load()

    def _load(self) -> dict[str, bool]:
        availability = dict(self.DEFAULT_AVAILABILITY)
        if self.storage_path is None or not self.storage_path.exists():
            return availability
        try:
            payload = json.loads(self.storage_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return availability
        raw_availability = payload.get("availability") if isinstance(payload, dict) else {}
        if isinstance(raw_availability, dict):
            for provider in availability:
                if provider in raw_availability:
                    availability[provider] = bool(raw_availability[provider])
        return availability

    def _persist_locked(self) -> None:
        if self.storage_path is None:
            return
        try:
            self.storage_path.parent.mkdir(parents=True, exist_ok=True)
            payload = {
                "updated_at": time.time(),
                "availability": self._availability,
            }
            temp_path = self.storage_path.with_name(f".{self.storage_path.name}.{os.getpid()}.tmp")
            temp_path.write_text(json.dumps(payload, ensure_ascii=False, sort_keys=True), encoding="utf-8")
            os.replace(temp_path, self.storage_path)
        except OSError:
            return

    def get(self) -> dict[str, bool]:
        with self._lock:
            return dict(self._availability)

    def save(self, availability: dict[str, Any]) -> dict[str, bool]:
        with self._lock:
            next_availability = dict(self.DEFAULT_AVAILABILITY)
            for provider in next_availability:
                if provider in availability:
                    next_availability[provider] = bool(availability[provider])
            self._availability = next_availability
            self._persist_locked()
            return dict(self._availability)


def _safe_email_identity(user_identity: dict[str, str | None] | None = None) -> str:
    if not user_identity:
        return ""
    return str(user_identity.get("email") or user_identity.get("config_key") or "").strip().lower()


def _meeting_record_summary(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "record_id": record.get("record_id"),
        "title": record.get("title"),
        "platform": record.get("platform"),
        "meeting_link": record.get("meeting_link"),
        "calendar_event_id": record.get("calendar_event_id"),
        "scheduled_start": record.get("scheduled_start"),
        "scheduled_end": record.get("scheduled_end"),
        "status": record.get("status"),
        "transcript_language": normalize_meeting_transcript_language(record.get("transcript_language")),
        "transcript_language_label": record.get("transcript_language_label") or "",
        "recording_started_at": record.get("recording_started_at"),
        "recording_stopped_at": record.get("recording_stopped_at"),
        "created_at": record.get("created_at"),
        "updated_at": record.get("updated_at"),
        "media": record.get("media") or {},
        "diagnostics_snapshot": record.get("diagnostics_snapshot") or {},
        "audio_preflight": record.get("audio_preflight") or {},
        "recording_health": record.get("recording_health") or {},
        "transcript_status": (record.get("transcript") or {}).get("status"),
        "transcript_quality": (record.get("transcript") or {}).get("quality") or {},
        "minutes_status": (record.get("minutes") or {}).get("status"),
        "email_status": (record.get("email") or {}).get("status"),
        "error": record.get("error") or "",
    }


def _meeting_recorder_diagnostics_payload(settings: Settings) -> dict[str, Any]:
    if _local_agent_meeting_recorder_enabled(settings):
        return _build_local_agent_client(settings).meeting_recorder_diagnostics()
    return _get_meeting_recorder_runtime().diagnostics()


def _meeting_recorder_record_summaries_for_current_user(settings: Settings) -> list[dict[str, Any]]:
    if _local_agent_meeting_recorder_enabled(settings):
        return _build_local_agent_client(settings).meeting_recorder_records(owner_email=_current_google_email())
    return [_meeting_record_summary(record) for record in _get_meeting_record_store().list_records(owner_email=_current_google_email())]


def _meeting_recorder_reminder_debug(
    *,
    reason: str,
    calendar_connected: bool,
    meeting_count: int = 0,
    active_recording: dict[str, Any] | None = None,
    diagnostics: dict[str, Any] | None = None,
    error: Exception | None = None,
) -> dict[str, Any]:
    diagnostics = diagnostics or {}
    return {
        "reason": reason,
        "checked_at": datetime.now(timezone.utc).isoformat(),
        "poll_seconds": 60,
        "timezone": "Asia/Singapore",
        "eligible_window": {"lead_seconds": 120, "grace_seconds": 600, "workday_start_hour": 9, "workday_end_hour": 20},
        "calendar_connected": calendar_connected,
        "eligible_meeting_count": meeting_count,
        "active_recording": bool(active_recording),
        "diagnostics": {
            "audio_capture_label": diagnostics.get("audio_capture_label", ""),
            "audio_capture_mode": diagnostics.get("audio_capture_mode", ""),
            "audio_signal_verified": bool(diagnostics.get("audio_signal_verified")),
            "system_audio_configured": bool(diagnostics.get("system_audio_configured")),
            "ffmpeg_configured": bool(diagnostics.get("ffmpeg_configured")),
            "whisper_cpp_configured": bool(diagnostics.get("whisper_cpp_configured")),
            "whisper_model_exists": bool(diagnostics.get("whisper_model_exists")),
        },
        "error_category": _classify_portal_error(error).get("error_category") if error else "",
        "error_code": _classify_portal_error(error).get("error_code") if error else "",
    }


def _try_acquire_gmail_export_lock(email: str) -> bool:
    normalized = str(email or "").strip().lower()
    if not normalized:
        return False
    with _gmail_export_active_users_lock:
        if normalized in _gmail_export_active_users:
            return False
        _gmail_export_active_users.add(normalized)
        return True


def _release_gmail_export_lock(email: str) -> None:
    normalized = str(email or "").strip().lower()
    if not normalized:
        return
    with _gmail_export_active_users_lock:
        _gmail_export_active_users.discard(normalized)


def _source_code_qa_codex_session_lock(session_id: str) -> threading.Lock:
    normalized = str(session_id or "").strip()
    if not normalized:
        normalized = "_no_session"
    with _source_code_qa_codex_session_locks_guard:
        lock = _source_code_qa_codex_session_locks.get(normalized)
        if lock is None:
            lock = threading.Lock()
            _source_code_qa_codex_session_locks[normalized] = lock
        return lock


def _count_configured_lines(value: Any) -> int:
    return sum(1 for line in str(value or "").splitlines() if line.strip() and not line.strip().startswith("#"))


def _count_routed_components(value: Any) -> int:
    components = {
        parts[2].strip().lower()
        for raw_line in str(value or "").splitlines()
        if (line := raw_line.strip()) and not line.startswith("#")
        if len(parts := [part.strip() for part in line.split("|")]) == 3 and parts[2].strip()
    }
    return len(components)


def _build_mapping_log_summary(config_data: dict[str, Any], *, save_mode: str = "") -> dict[str, Any]:
    return {
        "save_mode": save_mode or "full",
        "pm_team": str(config_data.get("pm_team", "") or "").strip().upper(),
        "system_header": str(config_data.get("system_header", "") or "").strip(),
        "market_header": str(config_data.get("market_header", "") or "").strip(),
        "input_tab_name": str(config_data.get("input_tab_name", "") or "").strip(),
        "has_spreadsheet_link": bool(str(config_data.get("spreadsheet_link", "") or "").strip()),
        "has_bpmis_token": bool(str(config_data.get("bpmis_api_access_token", "") or "").strip()),
        "route_rule_count": _count_configured_lines(config_data.get("component_route_rules_text", "")),
        "route_component_count": _count_routed_components(config_data.get("component_route_rules_text", "")),
        "default_rule_count": _count_configured_lines(config_data.get("component_default_rules_text", "")),
    }


def _build_request_log_context(
    settings: Settings | None = None,
    *,
    user_identity: dict[str, str | None] | None = None,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    context: dict[str, Any] = {
        "request_id": getattr(g, "request_id", None),
        "method": request.method,
        "path": request.path,
        "endpoint": request.endpoint,
        "request_origin": "api" if request.path.startswith("/api/") else "page",
        "remote_addr": request.headers.get("X-Forwarded-For", request.remote_addr),
        "user_agent": request.headers.get("User-Agent", ""),
        "user": _safe_email_identity(user_identity or (settings and _get_user_identity(settings))),
        "google_connected": _google_session_is_connected(),
    }
    if extra:
        context.update(extra)
    return context


def _classify_portal_error(error: Exception | str | None) -> dict[str, Any]:
    message = str(error or "").strip()
    normalized = message.lower()
    exception_type = type(error).__name__ if isinstance(error, BaseException) else ""

    details = {
        "error_message": message,
        "error_type": exception_type or "UnknownError",
        "error_category": "unexpected_internal",
        "error_code": "unexpected_error",
        "error_retryable": False,
    }

    if not message and not exception_type:
        return details

    if isinstance(error, ConfigError):
        details.update(
            {
                "error_category": "host_configuration",
                "error_code": "oauth_host_misconfigured",
                "error_retryable": False,
            }
        )
        return details

    if isinstance(error, ToolError):
        details["error_category"] = "tool_error"
        details["error_code"] = "tool_error"
        if "sign in with your npt google account" in normalized:
            details.update({"error_category": "authentication", "error_code": "auth_required", "error_retryable": True})
        elif "not authorized for the team portal" in normalized:
            details.update({"error_category": "authorization", "error_code": "account_not_allowed", "error_retryable": False})
        elif "team_portal_config_encryption_key" in normalized or "could not decrypt the saved bpmis token" in normalized:
            details.update({"error_category": "host_configuration", "error_code": "portal_encryption_config", "error_retryable": False})
        elif "duplicate system + market -> component rule" in normalized:
            details.update({"error_category": "config_validation", "error_code": "duplicate_route_rule", "error_retryable": False})
        elif "invalid system + market -> component rule" in normalized:
            details.update({"error_category": "config_validation", "error_code": "invalid_route_rule_format", "error_retryable": False})
        elif "invalid component default rule" in normalized:
            details.update({"error_category": "config_validation", "error_code": "invalid_component_default_rule", "error_retryable": False})
        elif "duplicate component default rule" in normalized:
            details.update({"error_category": "config_validation", "error_code": "duplicate_component_default_rule", "error_retryable": False})
        elif "component defaults are missing these routed components" in normalized:
            details.update({"error_category": "config_validation", "error_code": "missing_component_defaults", "error_retryable": False})
        elif "pm team is required" in normalized or "unsupported pm team" in normalized:
            details.update({"error_category": "config_validation", "error_code": "invalid_pm_team", "error_retryable": False})
        elif "version keyword is required" in normalized or "version_id is required" in normalized:
            details.update({"error_category": "request_validation", "error_code": "missing_required_parameter", "error_retryable": False})
        elif "google sheets" in normalized or "spreadsheet" in normalized or "sheet" in normalized:
            details.update({"error_category": "google_sheets", "error_code": "google_sheet_access", "error_retryable": True})
        elif "bpmis" in normalized:
            details.update({"error_category": "bpmis_upstream", "error_code": "bpmis_request_failed", "error_retryable": True})
        return details

    if "connection" in normalized or "timeout" in normalized or "temporar" in normalized:
        details["error_retryable"] = True
    return details


def _classify_http_status(status_code: int) -> dict[str, str]:
    if status_code == HTTPStatus.BAD_REQUEST:
        return {"error_category": "request_validation", "error_code": "bad_request"}
    if status_code == HTTPStatus.UNAUTHORIZED:
        return {"error_category": "authentication", "error_code": "auth_required"}
    if status_code == HTTPStatus.FORBIDDEN:
        return {"error_category": "authorization", "error_code": "forbidden"}
    if status_code == HTTPStatus.NOT_FOUND:
        return {"error_category": "routing", "error_code": "not_found"}
    if status_code == HTTPStatus.SERVICE_UNAVAILABLE:
        return {"error_category": "upstream_unavailable", "error_code": "service_unavailable"}
    if status_code >= HTTPStatus.INTERNAL_SERVER_ERROR:
        return {"error_category": "unexpected_internal", "error_code": "server_error"}
    return {"error_category": "http_error", "error_code": f"http_{status_code}"}


def _log_portal_event(
    event: str,
    *,
    level: int = logging.INFO,
    logger: logging.Logger | None = None,
    **fields: Any,
) -> None:
    active_logger = logger or current_app.logger
    payload = {"event": event, **fields}
    active_logger.log(level, "portal_event %s", json.dumps(payload, ensure_ascii=False, sort_keys=True))


def create_app() -> Flask:
    settings = Settings.from_env()
    package_dir = Path(__file__).resolve().parent
    project_root = package_dir.parent
    data_root = settings.team_portal_data_dir
    if not data_root.is_absolute():
        data_root = (project_root / data_root).resolve()
    config_store = WebConfigStore(
        data_root,
        legacy_root=project_root,
        encryption_key=settings.team_portal_config_encryption_key,
    )
    app = Flask(
        __name__,
        template_folder=str(project_root / "templates"),
        static_folder=str(project_root / "static"),
    )
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)
    app.config["SECRET_KEY"] = settings.flask_secret_key
    app.config["SETTINGS"] = settings
    app.config["CONFIG_STORE"] = config_store
    app.config["BPMIS_PROJECT_STORE"] = BPMISProjectStore(config_store.db_path)
    app.config["TEAM_DASHBOARD_CONFIG_STORE"] = TeamDashboardConfigStore(config_store.db_path)
    app.config["WORK_MEMORY_STORE"] = WorkMemoryStore(data_root / "work_memory" / "memory.db")
    app.config["JOB_STORE"] = JobStore(data_root / "run" / "jobs.json")
    app.config["SOURCE_CODE_QA_QUERY_SCHEDULER"] = SourceCodeQAQueryScheduler(
        job_store=app.config["JOB_STORE"],
        max_running=settings.source_code_qa_codex_concurrency,
    )
    app.config["SEATALK_TODO_STORE"] = SeaTalkTodoStore(data_root / "seatalk" / "completed_todos.json")
    app.config["SEATALK_NAME_MAPPING_STORE"] = SeaTalkNameMappingStore(data_root / "seatalk" / "name_overrides.json")
    app.config["SEATALK_DAILY_CACHE_DIR"] = data_root / "seatalk" / "cache"
    meeting_store = MeetingRecordStore(data_root / "meeting_records")
    app.config["MEETING_RECORD_STORE"] = meeting_store
    app.config["MEETING_RECORDER_RUNTIME"] = MeetingRecorderRuntime(
        store=meeting_store,
        config=_meeting_recorder_config(settings),
    )
    app.config["GOOGLE_CREDENTIAL_STORE"] = StoredGoogleCredentials(
        data_root / "google" / "credentials.json",
        encryption_key=settings.team_portal_config_encryption_key,
    )
    app.config["SOURCE_CODE_QA_SESSION_STORE"] = SourceCodeQASessionStore(data_root / "source_code_qa" / "sessions.json")
    app.config["SOURCE_CODE_QA_ATTACHMENT_STORE"] = SourceCodeQAAttachmentStore(data_root / "source_code_qa" / "attachments")
    app.config["SOURCE_CODE_QA_RUNTIME_EVIDENCE_STORE"] = SourceCodeQARuntimeEvidenceStore(
        data_root / "source_code_qa" / "runtime_evidence"
    )
    app.config["SOURCE_CODE_QA_MODEL_AVAILABILITY_STORE"] = SourceCodeQAModelAvailabilityStore(
        data_root / "source_code_qa" / "model_availability.json"
    )
    app.config["PRD_BRIEFING_STORE"] = BriefingStore(data_root / "prd_briefing")
    app.config["SOURCE_CODE_QA_SERVICE"] = SourceCodeQAService(
        data_root=data_root,
        team_profiles=TEAM_PROFILE_DEFAULTS,
        gitlab_token=settings.source_code_qa_gitlab_token,
        gitlab_username=settings.source_code_qa_gitlab_username,
        llm_provider=settings.source_code_qa_llm_provider,
        gemini_api_key=settings.source_code_qa_gemini_api_key,
        gemini_api_base_url=settings.source_code_qa_gemini_api_base_url,
        openai_api_key=settings.source_code_qa_openai_api_key,
        openai_api_base_url=settings.source_code_qa_openai_api_base_url,
        openai_model=settings.source_code_qa_openai_model,
        openai_cheap_model=settings.source_code_qa_openai_fast_model,
        openai_deep_model=settings.source_code_qa_openai_deep_model,
        openai_fallback_model=settings.source_code_qa_openai_fallback_model,
        gemini_model=settings.source_code_qa_gemini_model,
        gemini_cheap_model=settings.source_code_qa_gemini_fast_model,
        gemini_deep_model=settings.source_code_qa_gemini_deep_model,
        gemini_fallback_model=settings.source_code_qa_gemini_fallback_model,
        vertex_credentials_file=settings.source_code_qa_vertex_credentials_file,
        vertex_project_id=settings.source_code_qa_vertex_project_id,
        vertex_location=settings.source_code_qa_vertex_location,
        vertex_model=settings.source_code_qa_vertex_model,
        vertex_cheap_model=settings.source_code_qa_vertex_fast_model,
        vertex_deep_model=settings.source_code_qa_vertex_deep_model,
        vertex_fallback_model=settings.source_code_qa_vertex_fallback_model,
        query_rewrite_model=settings.source_code_qa_query_rewrite_model,
        planner_model=settings.source_code_qa_planner_model,
        answer_model=settings.source_code_qa_answer_model,
        judge_model=settings.source_code_qa_judge_model,
        repair_model=settings.source_code_qa_repair_model,
        llm_judge_enabled=settings.source_code_qa_llm_judge_enabled,
        semantic_index_model=settings.source_code_qa_embedding_model,
        semantic_index_enabled=settings.source_code_qa_semantic_index_enabled,
        embedding_provider=settings.source_code_qa_embedding_provider,
        embedding_api_key=settings.source_code_qa_embedding_api_key,
        embedding_api_base_url=settings.source_code_qa_embedding_api_base_url,
        llm_cache_ttl_seconds=settings.source_code_qa_llm_cache_ttl_seconds,
        llm_timeout_seconds=settings.source_code_qa_llm_timeout_seconds,
        codex_timeout_seconds=settings.source_code_qa_codex_timeout_seconds,
        codex_concurrency=settings.source_code_qa_codex_concurrency,
        codex_top_path_limit=settings.source_code_qa_codex_top_path_limit,
        codex_repair_enabled=settings.source_code_qa_codex_repair_enabled,
        codex_session_mode=settings.source_code_qa_codex_session_mode,
        codex_session_max_turns=settings.source_code_qa_codex_session_max_turns,
        codex_cache_followups=settings.source_code_qa_codex_cache_followups,
        llm_max_retries=settings.source_code_qa_llm_max_retries,
        llm_backoff_seconds=settings.source_code_qa_llm_backoff_seconds,
        llm_max_backoff_seconds=settings.source_code_qa_llm_max_backoff_seconds,
        git_timeout_seconds=settings.source_code_qa_git_timeout_seconds,
        max_file_bytes=settings.source_code_qa_max_file_bytes,
    )
    app.config["GET_USER_IDENTITY"] = lambda: _get_user_identity(settings)
    app.config["CAN_ACCESS_PRD_BRIEFING"] = lambda: _can_access_prd_briefing(settings)
    app.config["CAN_ACCESS_PRD_SELF_ASSESSMENT"] = lambda: _can_access_prd_self_assessment(settings)
    app.config["CAN_ACCESS_GMAIL_SEATALK_DEMO"] = lambda: _can_access_gmail_seatalk_demo(settings)
    app.register_blueprint(create_prd_briefing_blueprint())

    @app.context_processor
    def inject_primary_navigation():
        current_endpoint = request.endpoint or ""
        if _site_requires_google_login(settings) and not _google_session_is_connected():
            return {
                "site_tabs": [],
                "site_requires_google_login": True,
                "can_access_prd_briefing": False,
                "can_access_prd_self_assessment": False,
                "can_access_gmail_seatalk_demo": False,
                "can_access_meeting_recorder": False,
                "meeting_recorder_reminder_enabled": False,
                "can_access_source_code_qa": False,
                "can_manage_source_code_qa": False,
                "asset_revision": _current_release_revision(),
                "portal_stage": str(settings.team_portal_stage or "").strip().lower(),
            }
        user_identity = _get_user_identity(settings)
        can_access_team_dashboard = _can_access_team_dashboard(user_identity)
        site_tabs = []
        if _can_access_source_code_qa(settings):
            site_tabs.append(
                {
                    "label": "Source Code Q&A",
                    "href": url_for("source_code_qa"),
                    "active": request.path.startswith("/source-code-qa"),
                }
            )
        prd_tab = None
        prd_self_assessment_tab = None
        if _can_access_prd_self_assessment(settings):
            prd_self_assessment_tab = {
                "label": "PRD Self-Assessment",
                "href": url_for("prd_self_assessment_page"),
                "active": current_endpoint == "prd_self_assessment_page",
            }
        if _can_access_prd_briefing(settings):
            prd_tab = {
                "label": "PRD Briefing Tool",
                "href": url_for("prd_briefing.portal"),
                "active": current_endpoint.startswith("prd_briefing"),
            }
        if _can_access_meeting_recorder(settings):
            site_tabs.append(
                {
                    "label": "Meeting Recorder",
                    "href": url_for("meeting_recorder_page"),
                    "active": request.path.startswith("/meeting-recorder"),
                }
            )
        if prd_self_assessment_tab:
            site_tabs.append(prd_self_assessment_tab)
        if prd_tab:
            site_tabs.append(prd_tab)
        if can_access_team_dashboard:
            site_tabs.append(
                {
                    "label": "Team Dashboard",
                    "href": url_for("team_dashboard_page"),
                    "active": current_endpoint == "team_dashboard_page",
                }
            )
        site_tabs.append(
            {
                "label": "BPMIS Automation Tool",
                "href": url_for("index", workspace="run"),
                "active": current_endpoint == "index",
            }
        )
        if _can_access_work_memory(settings):
            site_tabs.append(
                {
                    "label": "AI Memory",
                    "href": url_for("work_memory_page"),
                    "active": request.path.startswith("/work-memory"),
                }
            )
        return {
            "site_tabs": site_tabs,
            "site_requires_google_login": _site_requires_google_login(settings),
            "can_access_prd_briefing": _can_access_prd_briefing(settings),
            "can_access_prd_self_assessment": _can_access_prd_self_assessment(settings),
            "can_access_gmail_seatalk_demo": _can_access_gmail_seatalk_demo(settings),
            "can_access_meeting_recorder": _can_access_meeting_recorder(settings),
            "meeting_recorder_reminder_enabled": _can_access_meeting_recorder(settings),
            "can_access_source_code_qa": _can_access_source_code_qa(settings),
            "can_manage_source_code_qa": _can_manage_source_code_qa(settings),
            "asset_revision": _current_release_revision(),
            "portal_stage": str(settings.team_portal_stage or "").strip().lower(),
        }

    @app.before_request
    def enforce_team_access():
        g.request_id = uuid.uuid4().hex[:12]
        if request.path.rstrip("/") == "/healthz":
            return jsonify({"status": "ok", "revision": _current_release_revision()}), HTTPStatus.OK
        if request.path.startswith("/api/local-agent/"):
            return None
        if request.endpoint in {
            None,
            "static",
            "index",
            "healthz",
            "google_login",
            "google_callback",
            "google_logout",
            "access_denied",
            "prd_briefing.image_proxy",
        }:
            return None
        login_gate = _require_google_login(
            settings,
            api=(request.path.startswith("/api/") or "/api/" in request.path),
        )
        if login_gate is not None:
            return login_gate
        if _current_google_user_is_blocked(settings):
            session.pop("google_credentials", None)
            session.pop("google_profile", None)
            message = "This Google account is not authorized for the team portal. Please contact the maintainer."
            if request.path.startswith("/api/") or "/api/" in request.path:
                return jsonify({"status": "error", "message": message}), HTTPStatus.FORBIDDEN
            flash(message, "error")
            return redirect(url_for("access_denied"))
        return None

    @app.after_request
    def prevent_authenticated_html_caching(response):
        request_id = getattr(g, "request_id", None)
        if request_id:
            response.headers["X-Request-ID"] = request_id
        if (
            _google_session_is_connected()
            and response.mimetype == "text/html"
            and request.method == "GET"
        ):
            response.headers["Cache-Control"] = "no-store, private, max-age=0"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
        if response.status_code >= HTTPStatus.BAD_REQUEST:
            status_details = _classify_http_status(response.status_code)
            _log_portal_event(
                "http_response_error",
                level=(
                    logging.WARNING
                    if response.status_code < HTTPStatus.INTERNAL_SERVER_ERROR
                    or status_details.get("error_category") == "upstream_unavailable"
                    else logging.ERROR
                ),
                **_build_request_log_context(
                    settings,
                    extra={
                        "status_code": response.status_code,
                        **status_details,
                    },
                ),
            )
        return response

    @app.get("/")
    def index():
        if _current_google_user_is_blocked(settings):
            session.pop("google_credentials", None)
            session.pop("google_profile", None)
            flash("This Google account is not authorized for the team portal. Please contact the maintainer.", "error")
            return redirect(url_for("access_denied"))

        if _site_requires_google_login(settings) and not _google_session_is_connected():
            return render_template("login_gate.html", page_title="Sign In")

        if str(request.args.get("workspace") or "").strip() == "team-dashboard":
            return redirect(url_for("team_dashboard_page"))
        requested_workspace_tab = str(request.args.get("workspace") or "").strip()
        has_bpmis_return_state = any(key in session for key in ("default_workspace_tab", "last_results", "run_notice"))
        if (
            not requested_workspace_tab
            and not has_bpmis_return_state
            and _can_access_source_code_qa(settings)
        ):
            return redirect(url_for("source_code_qa"))

        results = _results_for_display(session.pop("last_results", []))
        run_notice = session.pop("run_notice", None)
        user_identity = _get_user_identity(settings)
        config_key = user_identity.get("config_key")
        raw_config_data = None
        effective_team_profiles = {team_key: dict(profile) for team_key, profile in TEAM_PROFILE_DEFAULTS.items()}
        try:
            raw_config_data = _load_user_config_for_identity(settings, user_identity) if config_key else None
            effective_team_profiles = _load_effective_team_profiles(config_store)
        except ToolError as error:
            if not _is_local_agent_unavailable_error(error):
                raise
            current_app.logger.warning(
                "Mac local-agent is unavailable while rendering index; falling back to empty setup/default profiles.",
                exc_info=True,
            )
            flash("Mac local-agent is not reachable right now. The portal is open, but local-agent backed data and actions are temporarily unavailable.", "warning")
        config_data = raw_config_data or config_store._normalize({})
        config_data = _hydrate_setup_defaults(config_data, user_identity, team_profiles=effective_team_profiles)
        _apply_sync_email_policy(config_data, user_identity)
        input_headers: list[str] = []
        has_saved_config = bool(config_key and raw_config_data)
        default_workspace_tab = session.pop("default_workspace_tab", "run" if has_saved_config else "setup")
        allowed_workspace_tabs = {"setup", "run", "productization-upgrade-summary"}
        if _is_team_profile_admin(user_identity):
            allowed_workspace_tabs.add("team-default-admin")
        if requested_workspace_tab in allowed_workspace_tabs:
            default_workspace_tab = requested_workspace_tab

        if (
            "google_credentials" in session
            and _has_explicit_spreadsheet_link(raw_config_data)
            and _google_session_can_call_live_google_apis()
        ):
            try:
                sheets = _build_sheets_service(settings, config_data)
                snapshot = sheets.read_snapshot()
                input_headers = snapshot.headers
            except ToolError as error:
                error_details = _classify_portal_error(error)
                _log_portal_event(
                    "index_sheet_snapshot_tool_error",
                    level=logging.WARNING,
                    **_build_request_log_context(
                        settings,
                        user_identity=user_identity,
                        extra=error_details,
                    ),
                )
            except Exception:
                _log_portal_event(
                    "index_sheet_snapshot_unexpected_error",
                    level=logging.WARNING,
                    **_build_request_log_context(settings, user_identity=user_identity),
                )
                current_app.logger.warning("Google Sheets snapshot read failed on index.", exc_info=True)

        return render_template(
            "index.html",
            settings=settings,
            shared_portal_enabled=_shared_portal_enabled(settings),
            google_connected="google_credentials" in session,
            user_identity=user_identity,
            sync_email_editable=_can_edit_sync_email(user_identity),
            results=results,
            run_notice=run_notice,
            mapping_fields=CONFIGURED_FIELDS,
            mapping_config=config_data,
            team_profiles=_build_team_profiles_for_display(
                config_data,
                user_identity,
                team_profiles=effective_team_profiles,
            ),
            team_profile_admin_configs=effective_team_profiles,
            team_profile_admin_enabled=_is_team_profile_admin(user_identity),
            default_workspace_tab=default_workspace_tab,
            input_headers=input_headers,
            google_authorized=True,
        )

    @app.get("/team-dashboard")
    def team_dashboard_page():
        access_gate = _require_team_dashboard_access(settings)
        if access_gate is not None:
            return access_gate
        return render_template(
            "team_dashboard.html",
            page_title="Team Dashboard",
            user_identity=_get_user_identity(settings),
            team_dashboard_config=_get_team_dashboard_config_store().load(),
            can_manage_team_dashboard=_can_manage_team_dashboard(_get_user_identity(settings)),
            can_view_team_dashboard_monthly_report=_can_access_team_dashboard_monthly_report(
                _get_user_identity(settings)
            ),
            seatalk_configured=_seatalk_dashboard_is_configured(settings),
        )

    @app.get("/prd-self-assessment")
    @app.get("/prd-self-assessment/")
    def prd_self_assessment_page():
        access_gate = _require_prd_self_assessment_access(settings)
        if access_gate is not None:
            return access_gate
        return render_template(
            "prd_self_assessment.html",
            page_title="PRD Self-Assessment",
            user_identity=_get_user_identity(settings),
            review_url=url_for("prd_self_assessment_review_api"),
            summary_url=url_for("prd_self_assessment_summary_api"),
            asset_revision=_current_release_revision(),
        )

    @app.get("/healthz")
    def healthz():
        return jsonify({"status": "ok", "revision": _current_release_revision()}), HTTPStatus.OK

    @app.get("/access-denied")
    def access_denied():
        return render_template("access_denied.html", page_title="Access Restricted"), HTTPStatus.FORBIDDEN

    @app.get("/source-code-qa")
    def source_code_qa():
        access_gate = _require_source_code_qa_access(settings)
        if access_gate is not None:
            return access_gate
        user_identity = _get_user_identity(settings)
        service = _build_source_code_qa_service()
        return render_template(
            "source_code_qa.html",
            page_title="Source Code Q&A",
            user_identity=user_identity,
            options=_source_code_qa_options_payload(service),
            team_profiles=TEAM_PROFILE_DEFAULTS,
            country_options=list(CRMS_COUNTRIES),
            all_country=ALL_COUNTRY,
            can_manage_source_code_qa=_can_manage_source_code_qa(settings),
            asset_revision=_current_release_revision(),
        )

    @app.get("/api/source-code-qa/config")
    def source_code_qa_config_api():
        access_gate = _require_source_code_qa_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            service = _build_source_code_qa_service()
            gemini_service = _build_source_code_qa_service("gemini")
            codex_service = _build_source_code_qa_service("codex_cli_bridge")
            vertex_service = _build_source_code_qa_service("vertex_ai")
            model_availability = _source_code_qa_model_availability()
            return jsonify(
                {
                    "status": "ok",
                    "answer_mode": "auto",
                    "query_mode": "deep",
                    "can_manage": _can_manage_source_code_qa(settings),
                    "auth": _source_code_qa_auth_payload(settings),
                    "git_auth_ready": _source_code_qa_git_auth_ready(service, settings),
                    "llm_ready": service.llm_ready(),
                    "llm_provider": settings.source_code_qa_llm_provider,
                    "llm_providers": {
                        "gemini": {"ready": gemini_service.llm_ready(), "label": "Gemini", "available": model_availability["gemini"]},
                        "codex_cli_bridge": {"ready": codex_service.llm_ready(), "label": "Codex", "available": model_availability["codex_cli_bridge"]},
                        "vertex_ai": {"ready": vertex_service.llm_ready(), "label": "Vertex AI", "available": model_availability["vertex_ai"]},
                    },
                    "llm_model": service.llm_budgets["balanced"]["model"],
                    "llm_cheap_model": service.llm_budgets["cheap"]["model"],
                    "llm_deep_model": service.llm_budgets["deep"]["model"],
                    "llm_fallback_model": service._llm_fallback_model(),
                    "llm_policy": service.llm_policy_payload(),
                    "index_health": service.index_health_payload(),
                    "release_gate": _source_code_qa_release_gate_payload(settings),
                    "domain_knowledge": service.domain_knowledge_payload(),
                    "model_availability": model_availability,
                    "options": _source_code_qa_options_payload(service),
                    "config": service.load_config(),
                }
            )
        except ToolError as error:
            status_code = HTTPStatus.SERVICE_UNAVAILABLE if _is_local_agent_unavailable_error(error) else HTTPStatus.BAD_REQUEST
            return jsonify({"status": "error", "message": str(error), "error_category": _tool_error_category(error)}), status_code
        except Exception as error:  # noqa: BLE001 - keep API clients on JSON even for unexpected failures.
            request_id = getattr(g, "request_id", "")
            current_app.logger.exception("Source Code Q&A config failed unexpectedly")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Source Code Q&A config failed unexpectedly. Please refresh; if it repeats, share the request ID.",
                        "request_id": request_id,
                        "error_category": "source_code_qa_internal",
                        "error_retryable": True,
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.post("/api/source-code-qa/config")
    def source_code_qa_save_config_api():
        access_gate = _require_source_code_qa_manage_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = {}
        try:
            result = _build_source_code_qa_service().save_mapping(
                pm_team=str(payload.get("pm_team") or ""),
                country=str(payload.get("country") or ""),
                repositories=payload.get("repositories") or [],
            )
            return jsonify({"status": "ok", **result})
        except ToolError as error:
            status_code = HTTPStatus.SERVICE_UNAVAILABLE if _is_local_agent_unavailable_error(error) else HTTPStatus.BAD_REQUEST
            return jsonify({"status": "error", "message": str(error), "error_category": _tool_error_category(error)}), status_code

    @app.post("/api/source-code-qa/model-availability")
    def source_code_qa_model_availability_api():
        access_gate = _require_source_code_qa_manage_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        raw_availability = payload.get("availability") if isinstance(payload.get("availability"), dict) else {}
        store = _get_source_code_qa_model_availability_store()
        availability = store.save(raw_availability)
        return jsonify(
            {
                "status": "ok",
                "model_availability": availability,
                "options": _source_code_qa_options_payload(_build_source_code_qa_service()),
            }
        )

    @app.post("/api/source-code-qa/sync")
    def source_code_qa_sync_api():
        access_gate = _require_source_code_qa_manage_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        pm_team = str(payload.get("pm_team") or "")
        country = str(payload.get("country") or "")
        job_store: JobStore = current_app.config["JOB_STORE"]
        job = job_store.create("source-code-qa-sync", title="Sync Source Code Repositories")
        app_obj = current_app._get_current_object()
        thread = threading.Thread(
            target=_run_source_code_qa_sync_job,
            args=(app_obj, job.job_id, settings, pm_team, country),
            daemon=True,
        )
        thread.start()
        return jsonify({"status": "queued", "job_id": job.job_id})

    @app.route("/api/source-code-qa/sessions", methods=["GET", "POST"])
    def source_code_qa_sessions_api():
        access_gate = _require_source_code_qa_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        store = _get_source_code_qa_session_store()
        owner_email = _current_google_email() or "local"
        if request.method == "GET":
            limit = request.args.get("limit", "30")
            try:
                limit_value = int(limit)
            except ValueError:
                limit_value = 30
            return jsonify({"status": "ok", "sessions": store.list(owner_email=owner_email, limit=limit_value)})

        payload = request.get_json(silent=True) or {}
        session_payload = store.create(
            owner_email=owner_email,
            pm_team=str(payload.get("pm_team") or ""),
            country=str(payload.get("country") or ""),
            llm_provider=str(payload.get("llm_provider") or ""),
            title=str(payload.get("title") or ""),
        )
        return jsonify({"status": "ok", "session": session_payload})

    @app.get("/api/source-code-qa/sessions/<session_id>")
    def source_code_qa_session_api(session_id: str):
        access_gate = _require_source_code_qa_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        store = _get_source_code_qa_session_store()
        session_payload = store.get(session_id, owner_email=_current_google_email() or "local")
        if session_payload is None:
            return jsonify({"status": "error", "message": "Source Code Q&A session was not found."}), HTTPStatus.NOT_FOUND
        return jsonify({"status": "ok", "session": session_payload})

    @app.post("/api/source-code-qa/sessions/<session_id>/archive")
    def source_code_qa_session_archive_api(session_id: str):
        access_gate = _require_source_code_qa_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        store = _get_source_code_qa_session_store()
        archived = store.archive(session_id, owner_email=_current_google_email() or "local")
        if archived is None:
            return jsonify({"status": "error", "message": "Source Code Q&A session was not found."}), HTTPStatus.NOT_FOUND
        return jsonify(archived)

    @app.post("/api/source-code-qa/attachments")
    def source_code_qa_attachments_api():
        access_gate = _require_source_code_qa_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        owner_email = _current_google_email() or "local"
        session_id = str(request.form.get("session_id") or "").strip()
        if not session_id:
            return jsonify({"status": "error", "message": "A Source Code Q&A session is required before uploading attachments."}), HTTPStatus.BAD_REQUEST
        if _get_source_code_qa_session_store().get(session_id, owner_email=owner_email) is None:
            return jsonify({"status": "error", "message": "Source Code Q&A session was not found."}), HTTPStatus.NOT_FOUND
        uploaded = request.files.get("file")
        if uploaded is None:
            return jsonify({"status": "error", "message": "Upload a file field named file."}), HTTPStatus.BAD_REQUEST
        try:
            content = uploaded.read()
            attachment = _get_source_code_qa_attachment_store().save_bytes(
                owner_email=owner_email,
                session_id=session_id,
                filename=uploaded.filename or "attachment",
                mime_type=uploaded.mimetype or "",
                content=content,
            )
            return jsonify({"status": "ok", "attachment": attachment})
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.get("/api/source-code-qa/attachments/<attachment_id>")
    def source_code_qa_attachment_api(attachment_id: str):
        access_gate = _require_source_code_qa_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        owner_email = _current_google_email() or "local"
        session_id = str(request.args.get("session_id") or "").strip()
        if not session_id:
            return jsonify({"status": "error", "message": "session_id is required."}), HTTPStatus.BAD_REQUEST
        if _get_source_code_qa_session_store().get(session_id, owner_email=owner_email) is None:
            return jsonify({"status": "error", "message": "Source Code Q&A session was not found."}), HTTPStatus.NOT_FOUND
        try:
            metadata, content = _get_source_code_qa_attachment_store().get_bytes(
                owner_email=owner_email,
                session_id=session_id,
                attachment_id=attachment_id,
            )
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.NOT_FOUND
        return send_file(
            io.BytesIO(content),
            mimetype=metadata.get("mime_type") or "application/octet-stream",
            download_name=metadata.get("filename") or "attachment",
            as_attachment=False,
        )

    @app.route("/api/source-code-qa/runtime-evidence", methods=["GET", "POST"])
    def source_code_qa_runtime_evidence_api():
        access_gate = _require_source_code_qa_manage_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        store = _get_source_code_qa_runtime_evidence_store()
        if request.method == "GET":
            try:
                evidence = store.list(
                    pm_team=str(request.args.get("pm_team") or ""),
                    country=str(request.args.get("country") or ""),
                )
                return jsonify({"status": "ok", "evidence": evidence})
            except ToolError as error:
                return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

        uploaded = request.files.get("file")
        if uploaded is None:
            return jsonify({"status": "error", "message": "Upload a file field named file."}), HTTPStatus.BAD_REQUEST
        try:
            evidence = store.save_bytes(
                pm_team=str(request.form.get("pm_team") or ""),
                country=str(request.form.get("country") or ""),
                source_type=str(request.form.get("source_type") or "other"),
                uploaded_by=_current_google_email() or "local",
                filename=uploaded.filename or "runtime-evidence",
                mime_type=uploaded.mimetype or "",
                content=uploaded.read(),
            )
            return jsonify({"status": "ok", "evidence": evidence, "items": store.list(pm_team=evidence["pm_team"], country=evidence["country"])})
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.delete("/api/source-code-qa/runtime-evidence/<evidence_id>")
    def source_code_qa_runtime_evidence_delete_api(evidence_id: str):
        access_gate = _require_source_code_qa_manage_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            pm_team = str(request.args.get("pm_team") or "")
            country = str(request.args.get("country") or "")
            deleted = _get_source_code_qa_runtime_evidence_store().delete(
                pm_team=pm_team,
                country=country,
                evidence_id=evidence_id,
            )
            return jsonify(
                {
                    "status": "ok",
                    "deleted": deleted,
                    "evidence": _get_source_code_qa_runtime_evidence_store().list(pm_team=pm_team, country=country),
                }
            )
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.post("/api/source-code-qa/query")
    def source_code_qa_query_api():
        access_gate = _require_source_code_qa_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        if payload.get("async"):
            if not _source_code_qa_provider_available(payload.get("llm_provider")):
                return jsonify({"status": "error", "message": "Selected Source Code Q&A model is unavailable."}), HTTPStatus.BAD_REQUEST
            job_store: JobStore = current_app.config["JOB_STORE"]
            job = job_store.create("source-code-qa-query", title="Source Code Q&A Query")
            app_obj = current_app._get_current_object()
            async_payload = dict(payload)
            async_payload["query_mode"] = _source_code_qa_query_mode(payload.get("query_mode"))
            owner_email = _current_google_email() or "local"
            async_payload["_session_owner_email"] = owner_email
            session_id = str(payload.get("session_id") or "").strip()
            if session_id:
                try:
                    attachments = _resolve_source_code_qa_query_attachments(payload, owner_email=owner_email, session_id=session_id)
                    session_payload = _get_source_code_qa_session_store().append_pending_question(
                        session_id,
                        owner_email=owner_email,
                        pm_team=str(payload.get("pm_team") or ""),
                        country=str(payload.get("country") or ""),
                        llm_provider=str(payload.get("llm_provider") or ""),
                        question=str(payload.get("question") or ""),
                        job_id=job.job_id,
                        attachments=attachments,
                    )
                    if session_payload is not None:
                        async_payload["_resolved_attachments"] = attachments
                except ToolError:
                    pass
            scheduler: SourceCodeQAQueryScheduler = current_app.config["SOURCE_CODE_QA_QUERY_SCHEDULER"]
            scheduler.submit(app=app_obj, job_id=job.job_id, payload=async_payload, owner_email=owner_email)
            snapshot = _public_source_code_qa_job_snapshot(job_store.snapshot(job.job_id) or {})
            return jsonify({**snapshot, "status": "queued", "job_id": job.job_id, "session_id": session_id})
        try:
            if not _source_code_qa_provider_available(payload.get("llm_provider")):
                raise ToolError("Selected Source Code Q&A model is unavailable.")
            service = _build_source_code_qa_service(payload.get("llm_provider"))
            session_store = _get_source_code_qa_session_store()
            session_id = str(payload.get("session_id") or "").strip()
            owner_email = _current_google_email() or "local"
            conversation_context = payload.get("conversation_context") if isinstance(payload.get("conversation_context"), dict) else None
            if conversation_context is None and session_id:
                conversation_context = session_store.get_context(session_id, owner_email=owner_email)
            if isinstance(payload.get("_resolved_attachments"), list):
                attachments = payload.get("_resolved_attachments") or []
            else:
                attachments = _resolve_source_code_qa_query_attachments(payload, owner_email=owner_email, session_id=session_id)
            pm_team = str(payload.get("pm_team") or "")
            country = str(payload.get("country") or "")
            query_mode = _source_code_qa_query_mode(payload.get("query_mode"))
            runtime_evidence = _resolve_source_code_qa_runtime_evidence(pm_team=pm_team, country=country)
            auto_sync = _prepare_source_code_qa_auto_sync(service, pm_team=pm_team, country=country)
            def run_query() -> dict[str, Any]:
                return service.query(
                    pm_team=pm_team,
                    country=country,
                    question=str(payload.get("question") or ""),
                    answer_mode=_source_code_qa_public_answer_mode(payload.get("answer_mode")),
                    llm_budget_mode="auto",
                    query_mode=query_mode,
                    conversation_context=conversation_context,
                    attachments=attachments,
                    runtime_evidence=runtime_evidence,
                )

            if service.llm_provider_name == "codex_cli_bridge" and session_id:
                with _source_code_qa_codex_session_lock(session_id):
                    result = run_query()
            else:
                result = run_query()
            result["auto_sync"] = auto_sync
            result["attachments"] = _source_code_qa_public_attachments(attachments)
            result["runtime_evidence"] = _source_code_qa_public_runtime_evidence(runtime_evidence)
            if session_id:
                session_payload = session_store.append_exchange(
                    session_id,
                    owner_email=owner_email,
                    pm_team=str(payload.get("pm_team") or ""),
                    country=str(payload.get("country") or ""),
                    llm_provider=str(payload.get("llm_provider") or ""),
                    question=str(payload.get("question") or ""),
                    result=result,
                    context=_build_source_code_qa_session_context(result, payload),
                    attachments=attachments,
                )
                if session_payload is not None:
                    result["session"] = session_payload
                    result["session_id"] = session_id
            _record_source_code_qa_work_memory(
                owner_email=owner_email,
                pm_team=pm_team,
                country=country,
                question=str(payload.get("question") or ""),
                result=result,
                session_id=session_id,
            )
            return jsonify(result)
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception as error:  # noqa: BLE001 - keep API clients on JSON even for unexpected failures.
            request_id = getattr(g, "request_id", "")
            current_app.logger.exception("Source Code Q&A query failed unexpectedly")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Source Code Q&A failed unexpectedly. Please retry; if it repeats, share the request ID.",
                        "request_id": request_id,
                        "error_category": "source_code_qa_internal",
                        "error_retryable": True,
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.post("/api/source-code-qa/effort-assessment")
    def source_code_qa_effort_assessment_api():
        access_gate = _require_source_code_qa_manage_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        requirement = str(payload.get("requirement") or "").strip()
        if not requirement:
            return jsonify({"status": "error", "message": "Business requirement is empty."}), HTTPStatus.BAD_REQUEST
        if not _source_code_qa_provider_available(payload.get("llm_provider")):
            return jsonify({"status": "error", "message": "Selected Source Code Q&A model is unavailable."}), HTTPStatus.BAD_REQUEST

        job_store: JobStore = current_app.config["JOB_STORE"]
        job = job_store.create("source-code-qa-effort-assessment", title="Source Code Q&A Effort Assessment")
        app_obj = current_app._get_current_object()
        assessment_payload = {
            "pm_team": str(payload.get("pm_team") or ""),
            "country": str(payload.get("country") or ""),
            "language": _source_code_qa_effort_assessment_language(payload.get("language")),
            "requirement": requirement,
            "llm_provider": str(payload.get("llm_provider") or ""),
            "answer_mode": "auto",
            "query_mode": "deep",
            "_session_owner_email": _current_google_email() or "local",
        }
        scheduler: SourceCodeQAQueryScheduler = current_app.config["SOURCE_CODE_QA_QUERY_SCHEDULER"]
        scheduler.submit(
            app=app_obj,
            job_id=job.job_id,
            payload=assessment_payload,
            owner_email=assessment_payload["_session_owner_email"],
            runner=_run_source_code_qa_effort_assessment_job,
        )
        snapshot = _public_source_code_qa_job_snapshot(job_store.snapshot(job.job_id) or {})
        return jsonify({**snapshot, "status": "queued", "job_id": job.job_id})

    @app.get("/api/source-code-qa/effort-assessment/latest")
    def source_code_qa_effort_assessment_latest_api():
        access_gate = _require_source_code_qa_manage_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        result = current_app.config["JOB_STORE"].latest_completed_result("source-code-qa-effort-assessment")
        if not result:
            return jsonify({"status": "empty", "result": {}})
        return jsonify({"status": "ok", "result": result})

    @app.post("/api/source-code-qa/feedback")
    def source_code_qa_feedback_api():
        access_gate = _require_source_code_qa_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        try:
            result = _build_source_code_qa_service().save_feedback(
                user_email=_current_google_email() or "",
                payload=payload,
            )
            return jsonify(result)
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.get("/work-memory")
    def work_memory_page():
        access_gate = _require_work_memory_access(settings)
        if access_gate is not None:
            return access_gate
        return render_template("work_memory.html", page_title="AI Memory")

    @app.get("/api/work-memory/health")
    def work_memory_health_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        return jsonify(_get_work_memory_store().health())

    @app.get("/api/work-memory/recent")
    def work_memory_recent_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        items = _get_work_memory_store().query_work_memory(
            owner_email=_current_google_email(),
            visibility_scope=str(request.args.get("scope") or "owner").strip().lower() or "owner",
            query=str(request.args.get("q") or ""),
            filters={
                "source_type": str(request.args.get("source_type") or "").strip(),
                "item_type": str(request.args.get("item_type") or "").strip(),
            },
            limit=int(request.args.get("limit") or 50),
        )
        return jsonify({"status": "ok", "items": items})

    @app.get("/api/work-memory/review-candidates")
    def work_memory_review_candidates_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        items = _get_work_memory_store().review_candidates(owner_email=_current_google_email(), limit=int(request.args.get("limit") or 50))
        return jsonify({"status": "ok", "items": items})

    @app.get("/api/work-memory/project-timeline")
    def work_memory_project_timeline_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        project_ref = str(request.args.get("project_ref") or request.args.get("q") or "").strip()
        items = _get_work_memory_store().project_timeline(
            project_ref=project_ref,
            owner_email=_current_google_email(),
            visibility_scope=str(request.args.get("scope") or "owner").strip().lower() or "owner",
            limit=int(request.args.get("limit") or 100),
        )
        return jsonify({"status": "ok", "items": items})

    @app.get("/api/work-memory/entity-resolution")
    def work_memory_entity_resolution_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        result = _get_work_memory_store().resolve_work_entity(
            query=str(request.args.get("q") or request.args.get("query") or "").strip(),
            owner_email=_current_google_email(),
            entity_type=str(request.args.get("entity_type") or "").strip(),
        )
        return jsonify(result)

    @app.post("/api/work-memory/feedback")
    def work_memory_feedback_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        try:
            result = _get_work_memory_store().record_memory_feedback(
                item_id=str(payload.get("item_id") or "").strip(),
                action=str(payload.get("action") or "").strip(),
                owner_email=_current_google_email(),
                correction_text=str(payload.get("correction_text") or "").strip(),
                visibility_override=str(payload.get("visibility_override") or "").strip(),
                reason=str(payload.get("reason") or "").strip(),
            )
            return jsonify(result)
        except (KeyError, ValueError) as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.post("/api/work-memory/ingest-sent-monthly-reports")
    def work_memory_ingest_sent_monthly_reports_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            result = _ingest_sent_monthly_reports_from_gmail(settings)
            return jsonify({"status": "ok", **result})
        except (ConfigError, ToolError) as error:
            return jsonify({"status": "error", "message": str(error), **_classify_portal_error(error)}), HTTPStatus.BAD_REQUEST

    @app.post("/api/work-memory/backfill-existing")
    def work_memory_backfill_existing_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        result = _ingest_existing_work_memory_sources(
            settings,
            date_range=str(payload.get("date_range") or "90d").strip() or "90d",
            sources=[str(item or "").strip() for item in payload.get("sources") or [] if str(item or "").strip()] if isinstance(payload.get("sources"), list) else [],
        )
        return jsonify({"status": "ok", **result})

    @app.post("/api/work-memory/distill")
    def work_memory_distill_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        result = _get_work_memory_store().distill_work_memory(
            owner_email=_current_google_email(),
            date_range=str(payload.get("date_range") or "90d").strip() or "90d",
            sources=[str(item or "").strip() for item in payload.get("sources") or [] if str(item or "").strip()] if isinstance(payload.get("sources"), list) else [],
            project_refs=[str(item or "").strip() for item in payload.get("project_refs") or [] if str(item or "").strip()] if isinstance(payload.get("project_refs"), list) else [],
        )
        return jsonify({"status": "ok", **result})

    @app.post("/api/work-memory/ingest-incremental")
    def work_memory_ingest_incremental_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        try:
            result = _run_incremental_memory_ingestion(
                settings,
                window=str(payload.get("window") or "7d").strip() or "7d",
                reconciliation=bool(payload.get("reconciliation")),
            )
            return jsonify({"status": "ok", **result})
        except (ConfigError, ToolError) as error:
            return jsonify({"status": "error", "message": str(error), **_classify_portal_error(error)}), HTTPStatus.BAD_REQUEST

    @app.post("/api/work-memory/backfill-gmail")
    def work_memory_backfill_gmail_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        if not _google_credentials_have_scopes(GMAIL_READONLY_SCOPE):
            return jsonify(
                {
                    "status": "error",
                    "message": "Gmail read permission is missing. Reconnect Google once to grant gmail.readonly.",
                }
            ), HTTPStatus.BAD_REQUEST
        payload = request.get_json(silent=True) or {}
        owner_email = _current_google_email()
        job_store: JobStore = current_app.config["JOB_STORE"]
        active = job_store.active_for_record(WORK_MEMORY_GMAIL_BACKFILL_ACTION, owner_email=owner_email, record_id=owner_email)
        if active:
            return jsonify({**active, "status": "queued" if active.get("state") == "queued" else "running"}), HTTPStatus.ACCEPTED
        try:
            days = max(1, min(int(payload.get("days") or 90), 365))
            max_messages_raw = payload.get("max_messages")
            max_messages = None
            if max_messages_raw is not None and str(max_messages_raw).strip():
                max_messages = max(1, min(int(max_messages_raw), 10000))
        except (TypeError, ValueError):
            return jsonify({"status": "error", "message": "days and max_messages must be numbers."}), HTTPStatus.BAD_REQUEST
        job = job_store.create(WORK_MEMORY_GMAIL_BACKFILL_ACTION, "Gmail Work Memory Backfill", owner_email=owner_email, record_id=owner_email)
        app_obj = current_app._get_current_object()
        runner_payload = {
            "owner_email": owner_email,
            "days": days,
            "max_messages": max_messages,
            "credentials": dict(session.get("google_credentials") or {}),
            "report_intelligence_config": _get_team_dashboard_config_store().load().get("report_intelligence_config") or {},
            "drive_read_enabled": _google_credentials_have_scopes(GOOGLE_DRIVE_READONLY_SCOPE),
        }
        threading.Thread(target=_run_work_memory_gmail_backfill_job, args=(app_obj, job.job_id, runner_payload), daemon=True).start()
        snapshot = job_store.snapshot(job.job_id) or asdict(job)
        return jsonify({**snapshot, "status": "queued", "job_id": job.job_id}), HTTPStatus.ACCEPTED

    @app.get("/api/work-memory/ingestion-jobs")
    def work_memory_ingestion_jobs_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            limit = max(1, min(int(request.args.get("limit") or 20), 100))
        except (TypeError, ValueError):
            return jsonify({"status": "error", "message": "limit must be a number."}), HTTPStatus.BAD_REQUEST
        snapshots = current_app.config["JOB_STORE"].list_snapshots(
            action=WORK_MEMORY_GMAIL_BACKFILL_ACTION,
            owner_email=_current_google_email(),
            limit=limit,
        )
        return jsonify({"status": "ok", "items": snapshots})

    @app.get("/api/superagent/health")
    def superagent_health_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        return jsonify(_get_work_memory_store().superagent_health(owner_email=_current_google_email()))

    @app.post("/api/superagent/query")
    def superagent_query_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        task_type = str(payload.get("task_type") or "general").strip() or "general"
        query_text = str(payload.get("query") or "").strip()
        context = _get_work_memory_store().query_superagent_context(
            query=query_text,
            owner_email=_current_google_email(),
            visibility_scope=str(payload.get("visibility_scope") or "owner").strip().lower() or "owner",
            task_type=task_type,
            limit=int(payload.get("limit") or 12),
        )
        result = _get_work_memory_store().generate_llm_superagent_answer(task_type=task_type, query=query_text, context=context)
        audit = _get_work_memory_store().record_superagent_audit_log(
            owner_email=_current_google_email(),
            user_email=_current_google_email(),
            query=query_text,
            task_type=task_type,
            visibility_scope=str(payload.get("visibility_scope") or "owner").strip().lower() or "owner",
            context=context,
            answer=result,
            metadata={"route": "/api/superagent/query"},
        )
        return jsonify({"status": "ok", "context": context, "audit": audit, **result})

    @app.post("/api/superagent/explain")
    def superagent_explain_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        result = _get_work_memory_store().explain_superagent_answer(
            owner_email=_current_google_email(),
            query=str(payload.get("query") or "").strip(),
            task_type=str(payload.get("task_type") or "general").strip() or "general",
            visibility_scope=str(payload.get("visibility_scope") or "owner").strip().lower() or "owner",
        )
        return jsonify(result)

    @app.post("/api/superagent/eval")
    def superagent_eval_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        result = _get_work_memory_store().run_superagent_eval_cases(
            owner_email=_current_google_email(),
            cases=payload.get("cases") if isinstance(payload.get("cases"), list) else None,
            limit=int(payload.get("limit") or 30),
            suite_id=str(request.args.get("suite_id") or payload.get("suite_id") or "").strip(),
        )
        return jsonify(result)

    @app.get("/api/superagent/audit")
    def superagent_audit_api():
        access_gate = _require_work_memory_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        return jsonify(
            {
                "status": "ok",
                "items": _get_work_memory_store().superagent_audit_log(
                    owner_email=_current_google_email(),
                    limit=int(request.args.get("limit") or 50),
                ),
            }
        )

    @app.get("/auth/google/login")
    def google_login():
        try:
            authorization_url = create_google_authorization_url(settings)
            return redirect(authorization_url)
        except ConfigError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "google_login_config_error",
                level=logging.WARNING,
                **_build_request_log_context(settings, extra=error_details),
            )
            flash(str(error), "error")
            return redirect(url_for("index"))

    @app.get("/auth/google/callback")
    def google_callback():
        try:
            previous_identity = _get_user_identity(settings)
            finish_google_oauth(settings, request.url)
            if _current_google_user_is_blocked(settings):
                session.pop("google_credentials", None)
                session.pop("google_profile", None)
                flash("This Google account is not authorized for the team portal. Please contact the maintainer.", "error")
                return redirect(url_for("access_denied"))
            current_identity = _get_user_identity(settings)
            if previous_identity.get("config_key") and current_identity.get("config_key"):
                _migrate_user_config(settings, previous_identity["config_key"], current_identity["config_key"])
            _persist_owner_google_credentials(settings)
            _log_portal_event(
                "google_callback_success",
                **_build_request_log_context(settings, user_identity=current_identity),
            )
            flash("Google account connected successfully.", "success")
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "google_callback_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(settings, extra=error_details),
            )
            flash(str(error), "error")
        return redirect(url_for("index"))

    @app.get("/gmail-sea-talk-demo")
    def gmail_seatalk_demo():
        access_gate = _require_gmail_seatalk_demo_access(settings)
        if access_gate is not None:
            return access_gate
        return redirect(url_for("team_dashboard_page", tab="seatalk-name-mapping"))

    @app.get("/api/gmail-sea-talk-demo/dashboard")
    def gmail_seatalk_demo_dashboard_api():
        access_gate = _require_gmail_seatalk_demo_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        if not _google_credentials_have_scopes(GMAIL_READONLY_SCOPE):
            return jsonify(
                {
                    "status": "error",
                    "message": "Gmail access is not available for this Google session yet. Please sign in with Google again to grant Gmail read access.",
                }
            ), HTTPStatus.BAD_REQUEST
        try:
            dashboard = _build_gmail_dashboard_service().build_overview()
            return jsonify({"status": "ok", **dashboard})
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "gmail_seatalk_dashboard_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra=error_details,
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "gmail_seatalk_dashboard_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings)),
            )
            current_app.logger.exception("Gmail dashboard load failed.")
            return jsonify({"status": "error", "message": "Gmail data could not be loaded right now. Please try again shortly."}), HTTPStatus.INTERNAL_SERVER_ERROR

    @app.get("/api/gmail-sea-talk-demo/network")
    def gmail_seatalk_demo_network_api():
        access_gate = _require_gmail_seatalk_demo_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        if not _google_credentials_have_scopes(GMAIL_READONLY_SCOPE):
            return jsonify(
                {
                    "status": "error",
                    "message": "Gmail access is not available for this Google session yet. Please sign in with Google again to grant Gmail read access.",
                }
            ), HTTPStatus.BAD_REQUEST
        try:
            network = _build_gmail_dashboard_service().build_network()
            return jsonify({"status": "ok", **network})
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "gmail_seatalk_network_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra=error_details,
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "gmail_seatalk_network_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings)),
            )
            current_app.logger.exception("Gmail network load failed.")
            return jsonify({"status": "error", "message": "Gmail network rankings could not be loaded right now. Please try again shortly."}), HTTPStatus.INTERNAL_SERVER_ERROR

    @app.get("/api/gmail-sea-talk-demo/gmail/export-manifest")
    def gmail_seatalk_demo_gmail_export_manifest():
        access_gate = _require_gmail_seatalk_demo_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        if not _google_credentials_have_scopes(GMAIL_READONLY_SCOPE):
            return jsonify(
                {
                    "status": "error",
                    "message": "Gmail access is not available for this Google session yet. Please sign in with Google again to grant Gmail read access.",
                }
            ), HTTPStatus.BAD_REQUEST
        try:
            manifest = _build_gmail_dashboard_service().build_export_manifest()
            return jsonify({"status": "ok", **manifest})
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "gmail_seatalk_gmail_export_manifest_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra=error_details,
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "gmail_seatalk_gmail_export_manifest_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings)),
            )
            current_app.logger.exception("Gmail export manifest failed.")
            return (
                jsonify({"status": "error", "message": "Gmail export batches could not be prepared right now. Please try again shortly."}),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.get("/api/gmail-sea-talk-demo/gmail/export")
    def gmail_seatalk_demo_gmail_export():
        access_gate = _require_gmail_seatalk_demo_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        if not _google_credentials_have_scopes(GMAIL_READONLY_SCOPE):
            return jsonify(
                {
                    "status": "error",
                    "message": "Gmail access is not available for this Google session yet. Please sign in with Google again to grant Gmail read access.",
                }
            ), HTTPStatus.BAD_REQUEST
        try:
            batch = max(int(request.args.get("batch", "1")), 1)
        except ValueError:
            return jsonify({"status": "error", "message": "Invalid Gmail export batch. Please refresh and try again."}), HTTPStatus.BAD_REQUEST
        user_email = _safe_email_identity(_get_user_identity(settings))
        service = _build_gmail_dashboard_service()
        if not _try_acquire_gmail_export_lock(user_email):
            cached_payload = service.get_cached_export_history_text(batch=batch)
            if cached_payload is not None:
                content, filename = cached_payload
                return send_file(
                    io.BytesIO(content.encode("utf-8")),
                    mimetype="text/plain; charset=utf-8",
                    as_attachment=True,
                    download_name=filename,
                )
            return (
                jsonify({"status": "error", "message": "A Gmail export is already running for this account. Please wait a few seconds and try again."}),
                HTTPStatus.TOO_MANY_REQUESTS,
            )
        try:
            content, filename = service.export_history_text(batch=batch)
            return send_file(
                io.BytesIO(content.encode("utf-8")),
                mimetype="text/plain; charset=utf-8",
                as_attachment=True,
                download_name=filename,
            )
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "gmail_seatalk_gmail_export_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra=error_details,
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "gmail_seatalk_gmail_export_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings)),
            )
            current_app.logger.exception("Gmail history export failed.")
            return (
                jsonify({"status": "error", "message": "Gmail mail history could not be exported right now. Please try again shortly."}),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )
        finally:
            _release_gmail_export_lock(user_email)

    @app.post("/api/gmail-sea-talk-demo/gmail/export-prewarm")
    def gmail_seatalk_demo_gmail_export_prewarm():
        access_gate = _require_gmail_seatalk_demo_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        if not _google_credentials_have_scopes(GMAIL_READONLY_SCOPE):
            return jsonify(
                {
                    "status": "error",
                    "message": "Gmail access is not available for this Google session yet. Please sign in with Google again to grant Gmail read access.",
                }
            ), HTTPStatus.BAD_REQUEST
        try:
            batch = max(int(request.args.get("batch", "1")), 1)
        except ValueError:
            return jsonify({"status": "error", "message": "Invalid Gmail export batch. Please refresh and try again."}), HTTPStatus.BAD_REQUEST
        user_email = _safe_email_identity(_get_user_identity(settings))
        service = _build_gmail_dashboard_service()
        cached_payload = service.get_cached_export_history_text(batch=batch)
        if cached_payload is not None:
            return jsonify({"status": "ok", "cached": True, "batch": batch}), HTTPStatus.OK
        if not _try_acquire_gmail_export_lock(user_email):
            return jsonify({"status": "ok", "cached": False, "in_progress": True, "batch": batch}), HTTPStatus.ACCEPTED
        try:
            service.prewarm_export_history_text(batch=batch)
            return jsonify({"status": "ok", "cached": True, "batch": batch}), HTTPStatus.OK
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "gmail_seatalk_gmail_export_prewarm_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra=error_details,
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "gmail_seatalk_gmail_export_prewarm_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings)),
            )
            current_app.logger.exception("Gmail export prewarm failed.")
            return (
                jsonify({"status": "error", "message": "Gmail export prewarm could not be completed right now. Please try again shortly."}),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )
        finally:
            _release_gmail_export_lock(user_email)

    @app.get("/api/gmail-sea-talk-demo/seatalk")
    def gmail_seatalk_demo_seatalk_api():
        access_gate = _require_gmail_seatalk_demo_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            payload = _build_seatalk_dashboard_service(settings).build_overview()
            return jsonify({"status": "ok", **payload})
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "gmail_seatalk_seatalk_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra=error_details,
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "gmail_seatalk_seatalk_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings)),
            )
            current_app.logger.exception("SeaTalk dashboard load failed.")
            return (
                jsonify({"status": "error", "message": "SeaTalk data could not be loaded right now. Please try again shortly."}),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.get("/api/gmail-sea-talk-demo/seatalk/insights")
    def gmail_seatalk_demo_seatalk_insights_api():
        access_gate = _require_gmail_seatalk_demo_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            owner_email = _current_google_email() or settings.gmail_seatalk_demo_owner_email
            todo_store = _get_seatalk_todo_store(settings)
            service = _build_seatalk_dashboard_service(settings)
            todo_since = todo_store.processed_until(owner_email=owner_email)
            if _callable_accepts_keyword(service.build_insights, "todo_since"):
                payload = service.build_insights(todo_since=todo_since)
            else:
                payload = service.build_insights()
            completed_ids = todo_store.completed_ids(owner_email=owner_email)
            payload = dict(payload)
            open_todos = todo_store.merge_open_todos(
                owner_email=owner_email,
                todos=[todo for todo in (payload.get("my_todos") or []) if isinstance(todo, dict)],
            )
            todo_store.mark_processed_until(
                owner_email=owner_email,
                processed_until=str(payload.get("todo_processed_until") or ""),
            )
            payload["my_todos"] = [
                todo for todo in SeaTalkDashboardService._sort_todos(open_todos)
                if SeaTalkTodoStore.todo_id(todo) not in completed_ids
            ]
            payload["team_todos"] = []
            return jsonify({"status": "ok", **payload})
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "gmail_seatalk_seatalk_insights_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra=error_details,
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "gmail_seatalk_seatalk_insights_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings)),
            )
            current_app.logger.exception("SeaTalk insights load failed.")
            return (
                jsonify({"status": "error", "message": "SeaTalk insights could not be loaded right now. Please try again shortly."}),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.get("/api/gmail-sea-talk-demo/seatalk/project-updates")
    def gmail_seatalk_demo_seatalk_project_updates_api():
        access_gate = _require_gmail_seatalk_demo_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            service = _build_seatalk_dashboard_service(settings)
            if hasattr(service, "build_project_updates"):
                payload = service.build_project_updates()
            else:
                payload = service.build_insights()
            payload = dict(payload)
            payload["my_todos"] = []
            payload["team_todos"] = []
            return jsonify({"status": "ok", **payload})
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "gmail_seatalk_seatalk_project_updates_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra=error_details,
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "gmail_seatalk_seatalk_project_updates_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings)),
            )
            current_app.logger.exception("SeaTalk project updates load failed.")
            return (
                jsonify({"status": "error", "message": "SeaTalk project updates could not be loaded right now. Please try again shortly."}),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.get("/api/gmail-sea-talk-demo/seatalk/todos/open")
    def gmail_seatalk_demo_seatalk_open_todos_api():
        access_gate = _require_gmail_seatalk_demo_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            owner_email = _current_google_email() or settings.gmail_seatalk_demo_owner_email
            todo_store = _get_seatalk_todo_store(settings)
            completed_ids = todo_store.completed_ids(owner_email=owner_email)
            open_todos = [
                todo for todo in SeaTalkDashboardService._sort_todos(todo_store.open_todos(owner_email=owner_email))
                if SeaTalkTodoStore.todo_id(todo) not in completed_ids
            ]
            return jsonify({"status": "ok", "my_todos": open_todos, "team_todos": [], "project_updates": []})
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "gmail_seatalk_open_todos_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra=error_details,
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "gmail_seatalk_open_todos_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings)),
            )
            current_app.logger.exception("SeaTalk open to-dos load failed.")
            return (
                jsonify({"status": "error", "message": "Saved SeaTalk to-dos could not be loaded right now. Please try again shortly."}),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.get("/api/gmail-sea-talk-demo/seatalk/todos")
    def gmail_seatalk_demo_seatalk_todos_api():
        access_gate = _require_gmail_seatalk_demo_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            owner_email = _current_google_email() or settings.gmail_seatalk_demo_owner_email
            todo_store = _get_seatalk_todo_store(settings)
            service = _build_seatalk_dashboard_service(settings)
            todo_since = todo_store.processed_until(owner_email=owner_email)
            if hasattr(service, "build_todos") and _callable_accepts_keyword(service.build_todos, "todo_since"):
                payload = service.build_todos(todo_since=todo_since)
            elif _callable_accepts_keyword(service.build_insights, "todo_since"):
                payload = service.build_insights(todo_since=todo_since)
            else:
                payload = service.build_insights()
            completed_ids = todo_store.completed_ids(owner_email=owner_email)
            payload = dict(payload)
            open_todos = todo_store.merge_open_todos(
                owner_email=owner_email,
                todos=[todo for todo in (payload.get("my_todos") or []) if isinstance(todo, dict)],
            )
            todo_store.mark_processed_until(
                owner_email=owner_email,
                processed_until=str(payload.get("todo_processed_until") or ""),
            )
            payload["project_updates"] = []
            payload["my_todos"] = [
                todo for todo in SeaTalkDashboardService._sort_todos(open_todos)
                if SeaTalkTodoStore.todo_id(todo) not in completed_ids
            ]
            payload["team_todos"] = []
            return jsonify({"status": "ok", **payload})
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "gmail_seatalk_seatalk_todos_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra=error_details,
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "gmail_seatalk_seatalk_todos_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings)),
            )
            current_app.logger.exception("SeaTalk to-do load failed.")
            return (
                jsonify({"status": "error", "message": "SeaTalk to-dos could not be loaded right now. Please try again shortly."}),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.post("/api/gmail-sea-talk-demo/seatalk/todos/complete")
    def gmail_seatalk_demo_seatalk_todo_complete():
        access_gate = _require_gmail_seatalk_demo_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        todo = payload.get("todo") if isinstance(payload.get("todo"), dict) else payload
        try:
            todo_store = _get_seatalk_todo_store(settings)
            result = todo_store.mark_completed(
                owner_email=_current_google_email() or settings.gmail_seatalk_demo_owner_email,
                todo=todo,
            )
            return jsonify(result)
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.route("/api/gmail-sea-talk-demo/seatalk/name-mappings", methods=["GET", "POST"])
    def gmail_seatalk_demo_seatalk_name_mappings():
        access_gate = _require_gmail_seatalk_demo_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        mapping_store = _get_seatalk_name_mapping_store(settings)
        if request.method == "POST":
            payload = request.get_json(silent=True) or {}
            mappings = mapping_store.merge_mappings(payload.get("mappings") if isinstance(payload, dict) else {})
            SeaTalkDashboardService.clear_cache()
            return jsonify({"status": "ok", "mappings": mappings})
        try:
            force_refresh = str(request.args.get("refresh") or "").strip().lower() in {"1", "true", "yes", "on"}
            candidates = _build_seatalk_dashboard_service(settings).build_name_mappings(force_refresh=force_refresh)
            mappings = mapping_store.mappings()
            mapped_keys = {alias for key in mappings for alias in SeaTalkNameMappingStore.equivalent_keys(key)}
            candidates = dict(candidates)
            candidates["unknown_ids"] = _dedupe_seatalk_name_mapping_candidates([
                row for row in (candidates.get("unknown_ids") or [])
                if isinstance(row, dict) and not (SeaTalkNameMappingStore.equivalent_keys(row.get("id")) & mapped_keys)
            ])
            return jsonify({"status": "ok", "mappings": mappings, **candidates})
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "gmail_seatalk_seatalk_name_mappings_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra=error_details,
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "gmail_seatalk_seatalk_name_mappings_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings)),
            )
            current_app.logger.exception("SeaTalk name mapping load failed.")
            return (
                jsonify({"status": "error", "message": "SeaTalk name mappings could not be loaded right now. Please try again shortly."}),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.get("/api/gmail-sea-talk-demo/seatalk/export")
    def gmail_seatalk_demo_seatalk_export():
        access_gate = _require_gmail_seatalk_demo_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            content, filename = _build_seatalk_dashboard_service(settings).export_history_text()
            return send_file(
                io.BytesIO(content.encode("utf-8")),
                mimetype="text/plain; charset=utf-8",
                as_attachment=True,
                download_name=filename,
            )
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "gmail_seatalk_seatalk_export_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra=error_details,
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "gmail_seatalk_seatalk_export_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings)),
            )
            current_app.logger.exception("SeaTalk history export failed.")
            return (
                jsonify({"status": "error", "message": "SeaTalk chat history could not be exported right now. Please try again shortly."}),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.get("/meeting-recorder")
    def meeting_recorder_page():
        access_gate = _require_meeting_recorder_access(settings)
        if access_gate is not None:
            return access_gate
        user_identity = _get_user_identity(settings)
        return render_template(
            "meeting_recorder.html",
            page_title="Meeting Recorder",
            user_identity=user_identity,
            calendar_connected=_google_credentials_have_scopes(CALENDAR_READONLY_SCOPE),
            gmail_send_connected=_google_credentials_have_scopes("https://www.googleapis.com/auth/gmail.send"),
            selected_record_id=str(request.args.get("record") or "").strip(),
            asset_revision=_current_release_revision(),
        )

    @app.get("/api/meeting-recorder/diagnostics")
    def meeting_recorder_diagnostics_api():
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        if _local_agent_meeting_recorder_enabled(settings):
            try:
                return jsonify({"status": "ok", **_build_local_agent_client(settings).meeting_recorder_diagnostics()})
            except ToolError as error:
                return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_GATEWAY
        return jsonify({"status": "ok", **_get_meeting_recorder_runtime().diagnostics()})

    @app.get("/api/meeting-recorder/calendar/upcoming")
    def meeting_recorder_upcoming_api():
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        if not _google_credentials_have_scopes(CALENDAR_READONLY_SCOPE):
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Google Calendar access is not available yet. Sign in with Google again to grant calendar read access.",
                    }
                ),
                HTTPStatus.BAD_REQUEST,
            )
        try:
            meetings = _build_calendar_meeting_service().upcoming_meetings()
            return jsonify({"status": "ok", "meetings": meetings})
        except Exception as error:  # noqa: BLE001
            _log_portal_event(
                "meeting_recorder_calendar_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings), extra=_classify_portal_error(error)),
            )
            current_app.logger.exception("Meeting Recorder calendar load failed.")
            return jsonify({"status": "error", "message": "Upcoming meetings could not be loaded right now."}), HTTPStatus.INTERNAL_SERVER_ERROR

    @app.get("/api/meeting-recorder/reminders")
    def meeting_recorder_reminders_api():
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        if not _google_credentials_have_scopes(CALENDAR_READONLY_SCOPE):
            debug = _meeting_recorder_reminder_debug(reason="calendar_not_connected", calendar_connected=False)
            return jsonify(
                {
                    "status": "ok",
                    "calendar_connected": False,
                    "meetings": [],
                    "active_recording": None,
                    "poll_seconds": 60,
                    "timezone": "Asia/Singapore",
                    "debug": debug,
                }
            )
        try:
            meetings = _build_calendar_meeting_service().upcoming_meetings(days=1, max_results=20)
            eligible = reminder_eligible_meetings(meetings, timezone_name="Asia/Singapore")
            records = _meeting_recorder_record_summaries_for_current_user(settings)
            active_recording = next((record for record in records if str(record.get("status") or "") == "recording"), None)
            diagnostics: dict[str, Any] = {}
            diagnostics_error: Exception | None = None
            try:
                diagnostics = _meeting_recorder_diagnostics_payload(settings)
            except Exception as error:  # noqa: BLE001
                diagnostics_error = error
                diagnostics = {}
            reason = "diagnostics_failed" if diagnostics_error else "eligible_meetings_found" if eligible else "no_eligible_meetings"
            if active_recording:
                reason = "active_recording"
            debug = _meeting_recorder_reminder_debug(
                reason=reason,
                calendar_connected=True,
                meeting_count=len(eligible),
                active_recording=active_recording,
                diagnostics=diagnostics,
                error=diagnostics_error,
            )
            return jsonify(
                {
                    "status": "ok",
                    "calendar_connected": True,
                    "meetings": eligible,
                    "active_recording": active_recording,
                    "diagnostics": diagnostics,
                    "poll_seconds": 60,
                    "timezone": "Asia/Singapore",
                    "debug": debug,
                }
            )
        except Exception as error:  # noqa: BLE001
            _log_portal_event(
                "meeting_recorder_reminders_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings), extra=_classify_portal_error(error)),
            )
            current_app.logger.exception("Meeting Recorder reminders failed.")
            debug = _meeting_recorder_reminder_debug(reason="api_error", calendar_connected=True, error=error)
            return (
                jsonify({"status": "error", "message": "Meeting reminders could not be loaded right now.", "debug": debug}),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.post("/api/meeting-recorder/reminder-telemetry")
    def meeting_recorder_reminder_telemetry_api():
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        allowed_event = str(payload.get("event") or "unknown").strip()[:80]
        extra = {
            "telemetry_event": allowed_event,
            "reason": str(payload.get("reason") or "").strip()[:80],
            "outcome": str(payload.get("outcome") or "").strip()[:80],
            "page_path": str(payload.get("page_path") or "").strip()[:240],
            "capture_path": str(payload.get("capture_path") or "").strip()[:120],
            "meeting_link_present": bool(payload.get("meeting_link_present")),
            "get_user_media_supported": bool(payload.get("get_user_media_supported")),
            "media_recorder_supported": bool(payload.get("media_recorder_supported")),
            "selected_mime_type": str(payload.get("selected_mime_type") or "").strip()[:120],
            "recorder_mime_type": str(payload.get("recorder_mime_type") or "").strip()[:120],
            "recorder_state": str(payload.get("recorder_state") or "").strip()[:80],
            "stop_outcome": str(payload.get("stop_outcome") or "").strip()[:80],
            "active_track_label": str(payload.get("active_track_label") or "").strip()[:160],
            "preferred_device_label": str(payload.get("preferred_device_label") or "").strip()[:160],
            "input_device_count": int(payload.get("input_device_count") or 0),
            "audio_input_labels": str(payload.get("audio_input_labels") or "").strip()[:800],
            "preflight_rms_db": _safe_float(payload.get("preflight_rms_db")),
            "preflight_peak_db": _safe_float(payload.get("preflight_peak_db")),
            "blob_size": int(payload.get("blob_size") or 0),
            "chunk_count": int(payload.get("chunk_count") or 0),
            "elapsed_ms": int(payload.get("elapsed_ms") or 0),
            "record_id": str(payload.get("record_id") or "").strip()[:120],
            "protocol": str(payload.get("protocol") or "").strip()[:40],
            "host": str(payload.get("host") or "").strip()[:200],
            "error_name": str(payload.get("error_name") or "").strip()[:120],
            "meeting_count": int(payload.get("meeting_count") or 0),
            "suppressed_count": int(payload.get("suppressed_count") or 0),
            "active_recording": bool(payload.get("active_recording")),
            "error_category": str(payload.get("error_category") or "").strip()[:120],
            "error_message": str(payload.get("error_message") or "").strip()[:500],
        }
        _log_portal_event(
            "meeting_recorder_reminder_telemetry",
            **_build_request_log_context(settings, user_identity=_get_user_identity(settings), extra=extra),
        )
        return jsonify({"status": "ok"})

    @app.get("/api/meeting-recorder/records")
    def meeting_recorder_records_api():
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        if _local_agent_meeting_recorder_enabled(settings):
            try:
                records = _build_local_agent_client(settings).meeting_recorder_records(owner_email=_current_google_email())
                return jsonify({"status": "ok", "records": records})
            except ToolError as error:
                return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_GATEWAY
        records = _get_meeting_record_store().list_records(owner_email=_current_google_email())
        return jsonify({"status": "ok", "records": [_meeting_record_summary(record) for record in records]})

    @app.get("/api/meeting-recorder/records/<record_id>")
    def meeting_recorder_record_api(record_id: str):
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            if _local_agent_meeting_recorder_enabled(settings):
                record_payload = _build_local_agent_client(settings).meeting_recorder_record(
                    record_id=record_id,
                    owner_email=_current_google_email(),
                )
                return jsonify({"status": "ok", "record": record_payload.get("record") or {}})
            record = _get_meeting_record_store().get_record(record_id)
            if str(record.get("owner_email") or "").strip().lower() != _current_google_email():
                return jsonify({"status": "error", "message": "Meeting record is not available for this Google account."}), HTTPStatus.FORBIDDEN
            return jsonify({"status": "ok", "record": record})
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.post("/api/meeting-recorder/start")
    def meeting_recorder_start_api():
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        try:
            meeting_link = str(payload.get("meeting_link") or payload.get("meetingLink") or "").strip()
            recording_mode = str(payload.get("recording_mode") or payload.get("recordingMode") or "").strip()
            transcript_language = normalize_meeting_transcript_language(payload.get("transcript_language") or payload.get("transcriptLanguage"))
            if not recording_mode:
                recording_mode = "audio_only"
            if _local_agent_meeting_recorder_enabled(settings):
                remote_payload = dict(payload)
                remote_payload.update(
                    {
                        "owner_email": _current_google_email(),
                        "meeting_link": meeting_link,
                        "recording_mode": recording_mode,
                        "transcript_language": transcript_language,
                        "platform": str(payload.get("platform") or meeting_platform_from_link(meeting_link) or "unknown").strip(),
                    }
                )
                result = _build_local_agent_client(settings).meeting_recorder_start(remote_payload)
                return jsonify({"status": "ok", "record": result.get("record") or {}})
            record = _get_meeting_recorder_runtime().start_recording(
                owner_email=_current_google_email(),
                title=str(payload.get("title") or "Untitled meeting").strip(),
                platform=str(payload.get("platform") or meeting_platform_from_link(meeting_link) or "unknown").strip(),
                meeting_link=meeting_link,
                recording_mode=recording_mode,
                calendar_event_id=str(payload.get("calendar_event_id") or payload.get("calendarEventId") or "").strip(),
                scheduled_start=str(payload.get("scheduled_start") or payload.get("scheduledStart") or "").strip(),
                scheduled_end=str(payload.get("scheduled_end") or payload.get("scheduledEnd") or "").strip(),
                attendees=payload.get("attendees") if isinstance(payload.get("attendees"), list) else [],
                transcript_language=transcript_language,
            )
            return jsonify({"status": "ok", "record": _meeting_record_summary(record)})
        except (ConfigError, ToolError) as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.post("/api/meeting-recorder/records/<record_id>/stop")
    def meeting_recorder_stop_api(record_id: str):
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        owner_email = _current_google_email()
        try:
            if _local_agent_meeting_recorder_enabled(settings):
                result = _build_local_agent_client(settings).meeting_recorder_stop(
                    record_id=record_id,
                    owner_email=owner_email,
                )
                process_payload = _meeting_recorder_auto_process_payload(
                    settings=settings,
                    record_id=record_id,
                    owner_email=owner_email,
                )
                return jsonify({"status": "ok", "record": result.get("record") or {}, **process_payload})
            record = _get_meeting_recorder_runtime().stop_recording(record_id=record_id, owner_email=owner_email)
            process_payload = _meeting_recorder_auto_process_payload(
                settings=settings,
                record_id=record_id,
                owner_email=owner_email,
            )
            return jsonify({"status": "ok", "record": _meeting_record_summary(record), **process_payload})
        except (ToolError, requests.RequestException) as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.post("/api/meeting-recorder/records/<record_id>/signal-check")
    def meeting_recorder_signal_check_api(record_id: str):
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            if _local_agent_meeting_recorder_enabled(settings):
                result = _build_local_agent_client(settings).meeting_recorder_signal_check(
                    record_id=record_id,
                    owner_email=_current_google_email(),
                )
                return jsonify({"status": "ok", "record": result.get("record") or {}})
            record = _get_meeting_recorder_runtime().check_recording_signal(record_id=record_id, owner_email=_current_google_email())
            return jsonify({"status": "ok", "record": _meeting_record_summary(record)})
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.post("/api/meeting-recorder/browser-audio")
    def meeting_recorder_browser_audio_api():
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        return jsonify({
            "status": "error",
            "message": "Browser recording fallback is disabled. Grant Screen & System Audio Recording and Microphone permissions, then start recording again.",
        }), HTTPStatus.GONE
        payload = request.get_json(silent=True) or {}
        try:
            audio_base64 = str(payload.get("audio_base64") or "")
            if not audio_base64:
                raise ToolError("Browser audio upload was empty.")
            meeting_link = str(payload.get("meeting_link") or payload.get("meetingLink") or "").strip()
            capture_source = str(payload.get("browser_audio_capture_source") or payload.get("capture_source") or "").strip()
            capture_path = capture_source or ("browser_tab_audio_linked" if meeting_link else "browser_audio_f2f")
            _log_portal_event(
                "meeting_recorder_browser_audio_received",
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra={
                        "capture_path": capture_path,
                        "audio_base64_length": len(audio_base64),
                        "mime_type": str(payload.get("mime_type") or "").strip()[:120],
                        "browser_audio_device_label": str(payload.get("browser_audio_device_label") or "").strip()[:160],
                        "started_at_present": bool(str(payload.get("recording_started_at") or payload.get("started_at") or "").strip()),
                        "stopped_at_present": bool(str(payload.get("recording_stopped_at") or payload.get("stopped_at") or "").strip()),
                    },
                ),
            )
            recording_mode = "audio_only"
            remote_payload = {
                "owner_email": _current_google_email(),
                "title": str(payload.get("title") or "Untitled meeting").strip(),
                "meeting_link": meeting_link,
                "recording_mode": recording_mode,
                "platform": str(payload.get("platform") or meeting_platform_from_link(meeting_link)).strip(),
                "recording_started_at": str(payload.get("recording_started_at") or payload.get("started_at") or "").strip(),
                "recording_stopped_at": str(payload.get("recording_stopped_at") or payload.get("stopped_at") or "").strip(),
                "mime_type": str(payload.get("mime_type") or "").strip(),
                "audio_base64": audio_base64,
                "browser_audio_device_label": str(payload.get("browser_audio_device_label") or "").strip(),
                "browser_audio_capture_source": capture_path,
                "transcript_language": normalize_meeting_transcript_language(payload.get("transcript_language") or payload.get("transcriptLanguage")),
                "browser_audio_preflight": payload.get("browser_audio_preflight") if isinstance(payload.get("browser_audio_preflight"), dict) else {},
            }
            if _local_agent_meeting_recorder_enabled(settings):
                result = _build_local_agent_client(settings).meeting_recorder_browser_audio(remote_payload)
                record = result.get("record") or {}
                record_id = str(record.get("record_id") or "").strip()
                process_payload = _meeting_recorder_auto_process_payload(
                    settings=settings,
                    record_id=record_id,
                    owner_email=_current_google_email(),
                ) if record_id else {"auto_process_error": "Meeting processing could not be queued because the recording id was missing."}
                _log_portal_event(
                    "meeting_recorder_browser_audio_forwarded",
                    **_build_request_log_context(
                        settings,
                        user_identity=_get_user_identity(settings),
                        extra={
                            "capture_path": capture_path,
                            "record_id": str(record.get("record_id") or "").strip()[:120],
                            "record_status": str(record.get("status") or "").strip()[:80],
                            "max_volume_db": ((record.get("recording_health") or {}) if isinstance(record.get("recording_health"), dict) else {}).get("max_volume_db"),
                            "audio_capture_profile": str(((record.get("media") or {}) if isinstance(record.get("media"), dict) else {}).get("audio_capture_profile") or "").strip()[:120],
                        },
                    ),
                )
                return jsonify({"status": "ok", "record": record, **process_payload})
            try:
                audio_bytes = base64.b64decode(audio_base64, validate=True)
            except (ValueError, TypeError) as error:
                raise ToolError("Browser audio upload was not valid base64.") from error
            record = _get_meeting_recorder_runtime().import_browser_audio_recording(
                owner_email=_current_google_email(),
                title=remote_payload["title"],
                platform=remote_payload["platform"],
                meeting_link=meeting_link,
                started_at=remote_payload["recording_started_at"],
                stopped_at=remote_payload["recording_stopped_at"],
                audio_bytes=audio_bytes,
                mime_type=remote_payload["mime_type"],
                device_label=remote_payload["browser_audio_device_label"],
                capture_source=remote_payload["browser_audio_capture_source"],
                preflight_metrics=remote_payload["browser_audio_preflight"],
                transcript_language=normalize_meeting_transcript_language(payload.get("transcript_language") or payload.get("transcriptLanguage")),
            )
            process_payload = _meeting_recorder_auto_process_payload(
                settings=settings,
                record_id=str(record.get("record_id") or ""),
                owner_email=_current_google_email(),
            )
            _log_portal_event(
                "meeting_recorder_browser_audio_saved",
                **_build_request_log_context(
                    settings,
                    user_identity=_get_user_identity(settings),
                    extra={
                        "capture_path": capture_path,
                        "record_id": str(record.get("record_id") or "").strip()[:120],
                        "record_status": str(record.get("status") or "").strip()[:80],
                        "duration_seconds": float(((record.get("recording_health") or {}) if isinstance(record.get("recording_health"), dict) else {}).get("duration_seconds") or 0),
                        "max_volume_db": ((record.get("recording_health") or {}) if isinstance(record.get("recording_health"), dict) else {}).get("max_volume_db"),
                        "audio_capture_profile": str(((record.get("media") or {}) if isinstance(record.get("media"), dict) else {}).get("audio_capture_profile") or "").strip()[:120],
                    },
                ),
            )
            return jsonify({"status": "ok", "record": _meeting_record_summary(record), **process_payload})
        except (ConfigError, ToolError) as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.post("/api/meeting-recorder/records/<record_id>/process")
    def meeting_recorder_process_api(record_id: str):
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            if _local_agent_meeting_recorder_enabled(settings):
                result = _build_local_agent_client(settings).meeting_recorder_process_start(
                    record_id=record_id,
                    owner_email=_current_google_email(),
                )
                return jsonify({
                    "status": "queued",
                    "state": result.get("state") or "queued",
                    "job_id": result.get("job_id") or "",
                    "record": result.get("record") or {},
                })
            payload = _queue_meeting_recorder_process_job(
                app=current_app._get_current_object(),
                settings=settings,
                record_id=record_id,
                owner_email=_current_google_email(),
            )
            return jsonify(payload)
        except (ConfigError, ToolError, requests.RequestException) as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.get("/api/meeting-recorder/process-jobs/<job_id>")
    def meeting_recorder_process_job_api(job_id: str):
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            if _local_agent_meeting_recorder_enabled(settings):
                snapshot = _build_local_agent_client(settings).meeting_recorder_process_job(
                    job_id=job_id,
                    owner_email=_current_google_email(),
                )
                return jsonify(_public_meeting_recorder_process_job_snapshot(snapshot))
            snapshot = _meeting_recorder_process_job_snapshot_for_current_user(job_id)
            if snapshot is None:
                return jsonify({
                    "status": "error",
                    "message": "Meeting Recorder process job was not found.",
                    "error_category": "job_not_found",
                    "error_code": "job_not_found",
                    "error_retryable": False,
                }), HTTPStatus.NOT_FOUND
            return jsonify(_public_meeting_recorder_process_job_snapshot(snapshot))
        except (ConfigError, ToolError, requests.RequestException) as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.post("/api/meeting-recorder/records/<record_id>/repair-video")
    def meeting_recorder_repair_video_api(record_id: str):
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        return jsonify({
            "status": "error",
            "message": "Video recording and playback repair are no longer supported. Meeting Recorder is audio-only.",
        }), HTTPStatus.BAD_REQUEST

    @app.post("/api/meeting-recorder/records/<record_id>/send-email")
    def meeting_recorder_send_email_api(record_id: str):
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        try:
            if _local_agent_meeting_recorder_enabled(settings):
                result = _build_local_agent_client(settings).meeting_recorder_send_email(
                    record_id=record_id,
                    owner_email=_current_google_email(),
                    recipient=str(payload.get("recipient") or "").strip() or _current_google_email(),
                )
                return jsonify({"status": "ok", "email": result.get("email") or {}})
            email_payload = _build_meeting_processing_service(settings).send_minutes_email(
                record_id=record_id,
                owner_email=_current_google_email(),
                recipient=str(payload.get("recipient") or "").strip() or _current_google_email(),
            )
            return jsonify({"status": "ok", "email": email_payload})
        except (ConfigError, ToolError) as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.delete("/api/meeting-recorder/records/<record_id>")
    def meeting_recorder_delete_api(record_id: str):
        access_gate = _require_meeting_recorder_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            if _local_agent_meeting_recorder_enabled(settings):
                _build_local_agent_client(settings).meeting_recorder_delete(
                    record_id=record_id,
                    owner_email=_current_google_email(),
                )
                return jsonify({"status": "ok"})
            _get_meeting_record_store().delete_record(record_id=record_id, owner_email=_current_google_email())
            return jsonify({"status": "ok"})
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.route("/meeting-recorder/assets/<record_id>/<path:relative_path>", methods=["GET", "HEAD"])
    def meeting_recorder_asset(record_id: str, relative_path: str):
        access_gate = _require_meeting_recorder_access(settings)
        if access_gate is not None:
            return access_gate
        as_download = str(request.args.get("download") or "").strip().lower() in {"1", "true", "yes"}
        try:
            if _local_agent_meeting_recorder_enabled(settings):
                upstream = _build_local_agent_client(settings).meeting_recorder_asset_response(
                    record_id=record_id,
                    owner_email=_current_google_email(),
                    relative_path=relative_path,
                    range_header=str(request.headers.get("Range") or ""),
                    method=request.method,
                    download=as_download,
                )
                content_type = str(upstream.headers.get("Content-Type") or "")
                if as_download and "text/html" in content_type.lower():
                    upstream.close()
                    return jsonify({
                        "status": "error",
                        "message": "Meeting audio download returned an HTML response instead of the requested file. Refresh the page and sign in again, then retry.",
                    }), HTTPStatus.BAD_GATEWAY
                excluded_headers = {"content-encoding", "connection", "transfer-encoding"}
                headers = [
                    (key, value)
                    for key, value in upstream.headers.items()
                    if key.lower() not in excluded_headers and (not as_download or key.lower() != "content-disposition")
                ]
                if as_download:
                    filename = Path(upstream.headers.get("X-Meeting-Recorder-Filename") or relative_path).name or "meeting-recording.mp4"
                    headers.append(("Content-Disposition", f'attachment; filename="{filename}"'))
                if request.method == "HEAD":
                    upstream.close()
                    return Response(status=upstream.status_code, headers=headers)

                def stream_upstream():
                    try:
                        for chunk in upstream.iter_content(chunk_size=1024 * 256):
                            if chunk:
                                yield chunk
                    finally:
                        upstream.close()

                return Response(stream_upstream(), status=upstream.status_code, headers=headers, direct_passthrough=True)
            record = _get_meeting_record_store().get_record(record_id)
            if str(record.get("owner_email") or "").strip().lower() != _current_google_email():
                return jsonify({"status": "error", "message": "Meeting record is not available for this Google account."}), HTTPStatus.FORBIDDEN
            root_dir = _get_meeting_record_store().record_dir(record_id).resolve()
            asset_path = (root_dir / relative_path).resolve()
            if root_dir not in asset_path.parents and asset_path != root_dir:
                return jsonify({"status": "error", "message": "Invalid meeting asset path."}), HTTPStatus.BAD_REQUEST
            if not asset_path.exists():
                return jsonify({"status": "error", "message": "Meeting asset not found."}), HTTPStatus.NOT_FOUND
            media = record.get("media") if isinstance(record.get("media"), dict) else {}
            active_media_paths = [str(media.get("audio_path") or "").strip(), str(media.get("video_path") or "").strip()]
            active_asset_paths = {
                (_get_meeting_record_store().root_dir / media_path).resolve()
                for media_path in active_media_paths
                if media_path
            }
            active_asset_paths.update(
                (root_dir / Path(media_path).name).resolve()
                for media_path in active_media_paths
                if media_path
            )
            if str(record.get("status") or "") == "recording" and asset_path in active_asset_paths:
                return jsonify({
                    "status": "error",
                    "message": "Stop the recording before downloading the meeting media file.",
                }), HTTPStatus.CONFLICT
            return send_file(asset_path, conditional=True, as_attachment=as_download, download_name=asset_path.name)
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.post("/auth/google/logout")
    def google_logout():
        session.pop("google_credentials", None)
        session.pop("google_profile", None)
        flash("Google session cleared.", "success")
        return redirect(url_for("index"))

    @app.post("/config/save")
    def save_mapping_config():
        login_gate = _require_google_login(settings)
        if login_gate is not None:
            return login_gate
        try:
            user_identity = _get_user_identity(settings)
            existing_config = _load_user_config_for_identity(settings, user_identity) or config_store._normalize({})
            save_mode = str(request.form.get("save_mode", "") or "").strip()
            config = {
                "spreadsheet_link": request.form.get("spreadsheet_link", ""),
                "input_tab_name": request.form.get("input_tab_name", ""),
                "bpmis_api_access_token": request.form.get("bpmis_api_access_token", ""),
                "pm_team": request.form.get("pm_team", ""),
                "issue_id_header": request.form.get("issue_id_header", ""),
                "jira_ticket_link_header": request.form.get("jira_ticket_link_header", ""),
                "sync_pm_email": request.form.get("sync_pm_email", ""),
                "sync_project_name_header": request.form.get("sync_project_name_header", ""),
                "sync_market_header": request.form.get("sync_market_header", ""),
                "sync_brd_link_header": request.form.get("sync_brd_link_header", ""),
                "component_route_rules_text": request.form.get("component_route_rules_text", ""),
                "component_default_rules_text": request.form.get("component_default_rules_text", ""),
                "market_header": request.form.get("market_header", ""),
                "system_header": request.form.get("system_header", ""),
                "summary_header": request.form.get("summary_header", ""),
                "prd_links_header": request.form.get("prd_links_header", ""),
                "description_header": request.form.get("description_header", ""),
                "task_type_value": request.form.get("task_type_value", ""),
                "priority_value": request.form.get("priority_value", ""),
                "product_manager_value": request.form.get("product_manager_value", ""),
                "reporter_value": request.form.get("reporter_value", ""),
                "biz_pic_value": request.form.get("biz_pic_value", ""),
                "component_by_market": {
                    market: request.form.get(
                        f"component_{market}",
                        str((existing_config.get("component_by_market") or {}).get(market, "")),
                    )
                    for market in MARKET_KEYS
                },
                "need_uat_by_market": {
                    market: request.form.get(f"need_uat_{market}", "")
                    for market in MARKET_KEYS
                },
            }
            _apply_sync_email_policy(config, user_identity)
            config = config_store._normalize(
                _hydrate_setup_defaults(
                    config,
                    user_identity,
                    team_profiles=_load_effective_team_profiles(config_store),
                )
            )
            if save_mode == "route_only":
                _validate_config_security(settings, config)
                _save_route_only_config(existing_config, config, user_identity)
                _log_portal_event(
                    "config_save_success",
                    **_build_request_log_context(
                        settings,
                        user_identity=user_identity,
                        extra=_build_mapping_log_summary(config, save_mode=save_mode),
                    ),
                )
                flash("System + Market to Component was saved. Component owner table refreshed from the latest saved Components.", "success")
                return redirect(url_for("index"))

            _validate_config_security(settings, config)
            _validate_team_profile_setup(config, team_profiles=_load_effective_team_profiles(config_store))

            config_store.build_field_mappings(config)
            _save_user_config_for_identity(settings, user_identity, config)
            _log_portal_event(
                "config_save_success",
                **_build_request_log_context(
                    settings,
                    user_identity=user_identity,
                    extra=_build_mapping_log_summary(config, save_mode=save_mode),
                ),
            )
            flash("Your web Jira config was saved for this user and will be used for preview/run.", "success")
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "config_save_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    extra={
                        **error_details,
                        **_build_mapping_log_summary(config, save_mode=save_mode),
                    },
                ),
            )
            flash(str(error), "error")
        return redirect(url_for("index"))

    @app.post("/config/save-route")
    def save_mapping_route_only():
        login_gate = _require_google_login(settings)
        if login_gate is not None:
            return login_gate
        try:
            user_identity = _get_user_identity(settings)
            existing_config = _load_user_config_for_identity(settings, user_identity) or config_store._normalize({})
            payload = request.get_json(silent=True) or {}
            config = config_store._normalize(
                _hydrate_setup_defaults(
                    {
                        "pm_team": payload.get("pm_team", ""),
                        "system_header": payload.get("system_header", ""),
                        "market_header": payload.get("market_header", ""),
                        "component_route_rules_text": payload.get("component_route_rules_text", ""),
                        "component_default_rules_text": payload.get("component_default_rules_text", ""),
                    },
                    user_identity,
                    team_profiles=_load_effective_team_profiles(config_store),
                )
            )
            _validate_config_security(settings, config)
            saved = _save_route_only_config(
                existing_config,
                config,
                user_identity,
                default_text_override=str(payload.get("component_default_rules_text", "") or ""),
            )
            _log_portal_event(
                "config_save_route_success",
                **_build_request_log_context(
                    settings,
                    user_identity=user_identity,
                    extra=_build_mapping_log_summary(config, save_mode="route_only"),
                ),
            )
            return jsonify(
                {
                    "status": "ok",
                    "message": "System + Market to Component was saved. Component owner table refreshed from the latest saved Components.",
                    "component_route_rules_text": str(saved.get("component_route_rules_text", "") or ""),
                    "component_default_rules_text": str(saved.get("component_default_rules_text", "") or ""),
                }
            )
        except ToolError as error:
            summary_source = config if "config" in locals() else (payload if "payload" in locals() and isinstance(payload, dict) else {})
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "config_save_route_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    extra={
                        **error_details,
                        **_build_mapping_log_summary(summary_source, save_mode="route_only"),
                    },
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    def _save_route_only_config(
        existing_config: dict[str, Any],
        config: dict[str, Any],
        user_identity: dict[str, str | None],
        *,
        default_text_override: str = "",
    ) -> dict[str, Any]:
        route_only_config = dict(existing_config)
        route_only_config["pm_team"] = config.get("pm_team", "")
        route_only_config["system_header"] = config.get("system_header", "")
        route_only_config["market_header"] = config.get("market_header", "")
        route_only_config["component_route_rules_text"] = config.get("component_route_rules_text", "")
        config_store._parse_component_route_rules(str(route_only_config["component_route_rules_text"]))
        default_seed_text = str(default_text_override or "").strip() or str(existing_config.get("component_default_rules_text", "") or "")
        route_only_config["component_default_rules_text"] = config_store.align_component_defaults_to_routes(
            str(route_only_config["component_route_rules_text"]),
            default_seed_text,
        )
        normalized = config_store._normalize(route_only_config)
        _save_user_config_for_identity(settings, user_identity, normalized)
        return normalized

    @app.post("/admin/team-profiles/save")
    def save_team_profile_admin():
        login_gate = _require_google_login(settings)
        if login_gate is not None:
            return login_gate
        user_identity = _get_user_identity(settings)
        if not _is_team_profile_admin(user_identity):
            flash("Only the portal admin can update team default routing.", "error")
            return redirect(url_for("access_denied"))
        try:
            team_key = str(request.form.get("team_key", "") or "").strip().upper()
            team_profiles = _load_effective_team_profiles(config_store)
            if team_key not in team_profiles:
                raise ToolError(f"Unsupported PM Team: {team_key}.")
            saved_profile = _save_team_profile(settings, config_store,
                team_key,
                {
                    "label": str(team_profiles[team_key].get("label", "") or ""),
                    "ready": True,
                    "component_route_rules_text": request.form.get("component_route_rules_text", ""),
                },
            )
            session["default_workspace_tab"] = "team-default-admin"
            _log_portal_event(
                "team_profile_admin_save_success",
                **_build_request_log_context(
                    settings,
                    user_identity=user_identity,
                    extra={
                        "team_key": team_key,
                        "route_rule_count": _count_configured_lines(str(saved_profile.get("component_route_rules_text", "") or "")),
                        "default_rule_count": _count_configured_lines(str(saved_profile.get("component_default_rules_text", "") or "")),
                    },
                ),
            )
            flash(f"{team_profiles[team_key]['label']} team defaults were saved.", "success")
        except ToolError as error:
            error_details = _classify_portal_error(error)
            session["default_workspace_tab"] = "team-default-admin"
            _log_portal_event(
                "team_profile_admin_save_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=user_identity,
                    extra={**error_details, "team_key": str(request.form.get("team_key", "") or "").strip().upper()},
                ),
            )
            flash(str(error), "error")
        return redirect(url_for("index"))

    @app.get("/api/team-dashboard/config")
    def team_dashboard_config():
        access_gate = _require_team_dashboard_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        return jsonify({"status": "ok", "config": _get_team_dashboard_config_store().load()})

    @app.post("/admin/team-dashboard/members")
    def save_team_dashboard_members():
        access_gate = _require_team_dashboard_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        if not _can_manage_team_dashboard(_get_user_identity(settings)):
            return jsonify({"status": "error", "message": "Team Dashboard admin access is restricted."}), HTTPStatus.FORBIDDEN
        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = {
                "teams": {
                    team_key: {"member_emails": request.form.get(f"team_dashboard_members_{team_key}", "")}
                    for team_key in TEAM_DASHBOARD_TEAMS
                }
            }
        store = _get_team_dashboard_config_store()
        existing_config = store.load()
        if isinstance(existing_config.get("key_project_overrides"), dict):
            payload["key_project_overrides"] = existing_config["key_project_overrides"]
        if isinstance(existing_config.get("task_cache"), dict):
            payload["task_cache"] = existing_config["task_cache"]
        payload["monthly_report_template"] = existing_config.get("monthly_report_template") or DEFAULT_MONTHLY_REPORT_TEMPLATE
        payload["report_intelligence_config"] = existing_config.get("report_intelligence_config") or normalize_report_intelligence_config({})
        saved = store.save(payload)
        _log_portal_event(
            "team_dashboard_members_save_success",
            **_build_request_log_context(
                settings,
                user_identity=_get_user_identity(settings),
                extra={
                    "team_counts": {
                        team_key: len(team.get("member_emails") or [])
                        for team_key, team in (saved.get("teams") or {}).items()
                        if isinstance(team, dict)
                    }
                },
            ),
        )
        return jsonify({"status": "ok", "config": saved})

    @app.get("/api/team-dashboard/monthly-report/template")
    def team_dashboard_monthly_report_template():
        access_gate = _require_team_dashboard_monthly_report_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        config = _get_team_dashboard_config_store().load()
        return jsonify(
            {
                "status": "ok",
                "template": normalize_monthly_report_template(config.get("monthly_report_template")),
                "subject": monthly_report_subject(),
                "recipient": DEFAULT_MONTHLY_REPORT_RECIPIENT,
            }
        )

    @app.get("/api/team-dashboard/monthly-report/latest-draft")
    def team_dashboard_monthly_report_latest_draft():
        access_gate = _require_team_dashboard_monthly_report_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        if _remote_bpmis_config_enabled(settings):
            return jsonify(_build_local_agent_client(settings).team_dashboard_monthly_report_latest_draft())
        result = current_app.config["JOB_STORE"].latest_completed_result("team-dashboard-monthly-report-draft")
        draft_markdown = str((result or {}).get("draft_markdown") or "").strip()
        if not draft_markdown:
            return jsonify({"status": "empty", "draft_markdown": ""})
        generation_summary = (result or {}).get("generation_summary") if isinstance((result or {}).get("generation_summary"), dict) else {}
        generation_version = str((result or {}).get("generation_version") or generation_summary.get("generation_version") or "").strip()
        if not generation_version and not generation_summary.get("period_start"):
            return jsonify({"status": "empty", "draft_markdown": "", "message": "Latest Monthly Report draft was generated by an older format."})
        subject = str((result or {}).get("subject") or "").strip() or monthly_report_subject()
        return jsonify(
            {
                "status": "ok",
                "draft_markdown": draft_markdown,
                "subject": subject,
                "job_id": (result or {}).get("job_id") or "",
                "generated_at": (result or {}).get("generated_at") or 0,
                "generation_version": generation_version,
                "period_start": generation_summary.get("period_start") or "",
                "period_end": generation_summary.get("period_end") or "",
                "period_end_exclusive": generation_summary.get("period_end_exclusive") or "",
            }
        )

    @app.post("/admin/team-dashboard/monthly-report-template")
    def save_team_dashboard_monthly_report_template():
        access_gate = _require_team_dashboard_monthly_report_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        user_identity = _get_user_identity(settings)
        if not _can_manage_team_dashboard(user_identity):
            return jsonify({"status": "error", "message": "Team Dashboard admin access is restricted."}), HTTPStatus.FORBIDDEN
        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = {"template": request.form.get("monthly_report_template", "")}
        store = _get_team_dashboard_config_store()
        config = store.load()
        config["monthly_report_template"] = normalize_monthly_report_template(payload.get("template"))
        saved = store.save(config)
        _log_portal_event(
            "team_dashboard_monthly_report_template_save_success",
            **_build_request_log_context(
                settings,
                user_identity=user_identity,
                extra={"template_chars": len(str(saved.get("monthly_report_template") or ""))},
            ),
        )
        return jsonify({"status": "ok", "template": saved.get("monthly_report_template") or DEFAULT_MONTHLY_REPORT_TEMPLATE})

    @app.post("/admin/team-dashboard/report-intelligence")
    def save_team_dashboard_report_intelligence():
        access_gate = _require_team_dashboard_monthly_report_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        user_identity = _get_user_identity(settings)
        if not _can_manage_team_dashboard(user_identity):
            return jsonify({"status": "error", "message": "Team Dashboard admin access is restricted."}), HTTPStatus.FORBIDDEN
        payload = request.get_json(silent=True) or {}
        store = _get_team_dashboard_config_store()
        config = store.load()
        config["report_intelligence_config"] = normalize_report_intelligence_config(payload.get("report_intelligence_config") or payload)
        saved = store.save(config)
        _log_portal_event(
            "team_dashboard_report_intelligence_save_success",
            **_build_request_log_context(
                settings,
                user_identity=user_identity,
                extra={
                    "vip_count": len(saved.get("report_intelligence_config", {}).get("vip_people") or []),
                    "keyword_count": len(saved.get("report_intelligence_config", {}).get("priority_keywords") or []),
                },
            ),
        )
        return jsonify({"status": "ok", "report_intelligence_config": saved.get("report_intelligence_config") or normalize_report_intelligence_config({})})

    @app.route("/api/team-dashboard/report-intelligence/seatalk/name-mappings", methods=["GET", "POST"])
    def team_dashboard_report_intelligence_seatalk_name_mappings():
        access_gate = _require_team_dashboard_monthly_report_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        mapping_store = _get_seatalk_name_mapping_store(settings)
        if request.method == "POST":
            payload = request.get_json(silent=True) or {}
            mappings = mapping_store.merge_mappings(payload.get("mappings") or {})
            SeaTalkDashboardService.clear_cache()
            return jsonify({"status": "ok", "mappings": mappings})
        force_refresh = str(request.args.get("refresh") or "").strip().lower() in {"1", "true", "yes"}
        try:
            candidates = _build_seatalk_dashboard_service(settings).build_name_mappings(force_refresh=force_refresh)
            mappings = mapping_store.mappings()
            mapped_keys = {alias for key in mappings for alias in SeaTalkNameMappingStore.equivalent_keys(key)}
            candidates = dict(candidates)
            candidates["unknown_ids"] = _dedupe_seatalk_name_mapping_candidates([
                row for row in (candidates.get("unknown_ids") or [])
                if isinstance(row, dict) and not (SeaTalkNameMappingStore.equivalent_keys(row.get("id")) & mapped_keys)
            ])
            return jsonify({"status": "ok", "mappings": mappings, **candidates})
        except (ConfigError, ToolError) as error:
            _log_portal_event(
                "team_dashboard_report_intelligence_name_mapping_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings), extra=_classify_portal_error(error)),
            )
            current_app.logger.warning("Team Dashboard Report Intelligence name mapping failed: %s", error)
            return jsonify({"status": "error", "message": str(error), **_classify_portal_error(error)}), HTTPStatus.BAD_REQUEST
        except Exception as error:  # noqa: BLE001
            _log_portal_event(
                "team_dashboard_report_intelligence_name_mapping_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings), extra=_classify_portal_error(error)),
            )
            current_app.logger.exception("Team Dashboard Report Intelligence name mapping failed.")
            return jsonify({"status": "error", "message": "Could not load SeaTalk name mappings.", **_classify_portal_error(error)}), HTTPStatus.INTERNAL_SERVER_ERROR

    @app.get("/api/team-dashboard/tasks")
    def team_dashboard_tasks():
        access_gate = _require_team_dashboard_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        route_started_at = time.monotonic()
        config_started_at = time.monotonic()
        store = _get_team_dashboard_config_store()
        config = store.load()
        config_elapsed = round(time.monotonic() - config_started_at, 3)
        key_project_overrides = config.get("key_project_overrides") if isinstance(config.get("key_project_overrides"), dict) else {}
        requested_team_key = str(request.args.get("team") or request.args.get("team_key") or "").strip().upper()
        force_reload = str(request.args.get("reload") or "").strip().lower() in {"1", "true", "yes"}
        if requested_team_key and requested_team_key not in TEAM_DASHBOARD_TEAMS:
            return (
                jsonify({"status": "error", "message": f"Unknown team: {requested_team_key}."}),
                HTTPStatus.BAD_REQUEST,
            )
        team_items = (
            [(requested_team_key, TEAM_DASHBOARD_TEAMS[requested_team_key])]
            if requested_team_key
            else list(TEAM_DASHBOARD_TEAMS.items())
        )
        if force_reload and not requested_team_key:
            try:
                team_payloads = _load_team_dashboard_tasks_for_all_teams_merged(
                    settings,
                    store,
                    config,
                    config_elapsed=config_elapsed,
                    route_started_at=route_started_at,
                    key_project_overrides=key_project_overrides,
                )
                _record_team_dashboard_work_memory(team_payloads, owner_email=_current_google_email())
                response = jsonify(
                    {
                        "status": "ok",
                        "teams": team_payloads,
                        "team": None,
                        "team_key": "",
                        "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                    }
                )
                response.headers["Cache-Control"] = "no-store, private, max-age=0"
                response.headers["Pragma"] = "no-cache"
                response.headers["Expires"] = "0"
                return response
            except Exception as error:  # noqa: BLE001 - keep the API shape stable on upstream failure.
                error_details = _classify_portal_error(error)
                _log_portal_event(
                    "team_dashboard_tasks_all_team_reload_error",
                    level=logging.WARNING,
                    **_build_request_log_context(
                        settings,
                        user_identity=_get_user_identity(settings),
                        extra=error_details,
                    ),
                )
        team_payloads: list[dict[str, Any]] = []
        has_error = False
        for team_key, label in team_items:
            timing_stats = _team_dashboard_new_timing()
            timing_stats["config_load"] = config_elapsed
            team_config = (config.get("teams") or {}).get(team_key) or {}
            emails = _normalize_team_dashboard_emails(team_config.get("member_emails") or [])
            cache_started_at = time.monotonic()
            cached_team = None if force_reload else _cached_team_dashboard_task_payload(config, team_key, emails)
            _team_dashboard_add_timing(timing_stats, "cache_check", cache_started_at)
            if cached_team is not None:
                timing_stats["total"] = round(time.monotonic() - route_started_at, 3)
                cached_team["timing_stats"] = timing_stats
                cached_team["elapsed_seconds"] = timing_stats["total"]
                team_payloads.append(cached_team)
                continue
            started_at = time.monotonic()
            try:
                bpmis_client = _build_bpmis_client_for_current_user(settings)
                biz_bpmis_client = _build_bpmis_client_for_current_user(settings)
                tasks, biz_projects = _team_dashboard_load_jira_and_biz_projects(
                    bpmis_client,
                    biz_bpmis_client,
                    emails,
                    timing_stats,
                )
                step_started_at = time.monotonic()
                team_payload = _build_team_dashboard_task_group(
                    team_key,
                    label,
                    emails,
                    tasks,
                    biz_projects,
                    key_project_overrides=key_project_overrides,
                )
                _team_dashboard_add_timing(timing_stats, "group_projects", step_started_at)
                step_started_at = time.monotonic()
                _backfill_team_dashboard_empty_project_jira_tasks(bpmis_client, team_payload)
                _remove_team_dashboard_zero_jira_pending_live_projects(team_payload)
                _team_dashboard_add_timing(timing_stats, "backfill_zero_jira_projects", step_started_at)
                timing_stats.update(_team_dashboard_combined_request_timings(bpmis_client, biz_bpmis_client))
                team_payload["elapsed_seconds"] = round(time.monotonic() - started_at, 2)
                team_payload["fetch_stats"] = _team_dashboard_combined_fetch_stats(bpmis_client, biz_bpmis_client)
                timing_stats["total"] = team_payload["elapsed_seconds"]
                team_payload["timing_stats"] = timing_stats
                team_payloads.append(team_payload)
                step_started_at = time.monotonic()
                _store_team_dashboard_task_payload(store, team_key, emails, team_payload)
                _team_dashboard_add_timing(timing_stats, "cache_store", step_started_at)
                timing_stats["total"] = round(time.monotonic() - started_at, 2)
                team_payload["elapsed_seconds"] = timing_stats["total"]
                team_payload["timing_stats"] = timing_stats
                _log_portal_event(
                    "team_dashboard_tasks_team_loaded",
                    **_build_request_log_context(
                        settings,
                        user_identity=_get_user_identity(settings),
                        extra={
                            "team_key": team_key,
                            "email_count": len(emails),
                            "raw_task_count": len(tasks or []),
                            "raw_biz_project_count": len(biz_projects or []),
                            "elapsed_seconds": team_payload["elapsed_seconds"],
                            "fetch_stats": team_payload["fetch_stats"],
                            "timing_stats": team_payload["timing_stats"],
                        },
                    ),
                )
            except Exception as error:  # noqa: BLE001 - keep other team groups renderable.
                has_error = True
                timing_stats["total"] = round(time.monotonic() - started_at, 2)
                error_details = _classify_portal_error(error)
                _log_portal_event(
                    "team_dashboard_tasks_team_error",
                    level=logging.WARNING,
                    **_build_request_log_context(
                        settings,
                        user_identity=_get_user_identity(settings),
                        extra={**error_details, "team_key": team_key},
                    ),
                )
                team_payloads.append(
                    {
                        "team_key": team_key,
                        "label": label,
                        "member_emails": emails,
                        "under_prd": [],
                        "pending_live": [],
                        "error": str(error),
                        "elapsed_seconds": timing_stats["total"],
                        "fetch_stats": {},
                        "timing_stats": timing_stats,
                    }
                )
        _record_team_dashboard_work_memory(team_payloads, owner_email=_current_google_email())
        response = jsonify(
            {
                "status": "partial" if has_error else "ok",
                "teams": team_payloads,
                "team": team_payloads[0] if requested_team_key and team_payloads else None,
                "team_key": requested_team_key,
                "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            }
        )
        response.headers["Cache-Control"] = "no-store, private, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response

    @app.post("/api/team-dashboard/key-projects")
    def save_team_dashboard_key_project():
        access_gate = _require_team_dashboard_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        user_identity = _get_user_identity(settings)
        if not _can_manage_team_dashboard(user_identity):
            return jsonify({"status": "error", "message": "Team Dashboard admin access is restricted."}), HTTPStatus.FORBIDDEN
        payload = request.get_json(silent=True) or {}
        bpmis_id = str(payload.get("bpmis_id") or "").strip()
        if not bpmis_id:
            return jsonify({"status": "error", "message": "BPMIS ID is required."}), HTTPStatus.BAD_REQUEST
        if "is_key_project" not in payload:
            return jsonify({"status": "error", "message": "Key Project value is required."}), HTTPStatus.BAD_REQUEST
        is_key_project = bool(payload.get("is_key_project"))
        store = _get_team_dashboard_config_store()
        config = store.load()
        overrides = config.get("key_project_overrides") if isinstance(config.get("key_project_overrides"), dict) else {}
        overrides[bpmis_id] = {
            "is_key_project": is_key_project,
            "updated_by": str(user_identity.get("email") or "").strip().lower(),
            "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        }
        config["key_project_overrides"] = overrides
        saved = store.save(config)
        effective = _apply_team_dashboard_key_project_state(
            {"bpmis_id": bpmis_id, "priority": str(payload.get("priority") or "").strip()},
            saved.get("key_project_overrides") if isinstance(saved.get("key_project_overrides"), dict) else {},
        )
        _log_portal_event(
            "team_dashboard_key_project_save_success",
            **_build_request_log_context(
                settings,
                user_identity=user_identity,
                extra={
                    "bpmis_id": bpmis_id,
                    "is_key_project": is_key_project,
                    "key_project_source": effective.get("key_project_source"),
                },
            ),
        )
        return jsonify(
            {
                "status": "ok",
                "bpmis_id": bpmis_id,
                "override": (saved.get("key_project_overrides") or {}).get(bpmis_id) or {},
                "is_key_project": effective.get("is_key_project"),
                "key_project_source": effective.get("key_project_source"),
            }
        )

    @app.get("/api/team-dashboard/link-biz-projects")
    def team_dashboard_link_biz_projects():
        access_gate = _require_team_dashboard_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        try:
            config = _get_team_dashboard_config_store().load()
            started_at = time.monotonic()
            rows = _load_team_dashboard_link_biz_jira_rows(settings, config)
            return jsonify({"status": "ok", "rows": rows, "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())})
        except Exception as error:  # noqa: BLE001
            _log_portal_event(
                "team_dashboard_link_biz_project_load_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=_get_user_identity(settings), extra=_classify_portal_error(error)),
            )
            current_app.logger.exception("Team Dashboard Link Biz Project load failed.")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Could not load unlinked Jira tickets. Please retry or share the request ID.",
                        **_classify_portal_error(error),
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.get("/api/team-dashboard/link-biz-projects/jira")
    def team_dashboard_link_biz_project_jira():
        access_gate = _require_team_dashboard_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        user_identity = _get_user_identity(settings)
        try:
            config = _get_team_dashboard_config_store().load()
            started_at = time.monotonic()
            rows = _load_team_dashboard_link_biz_jira_rows(settings, config)
            elapsed_seconds = round(time.monotonic() - started_at, 2)
            _log_portal_event(
                "team_dashboard_link_biz_project_jira_loaded",
                **_build_request_log_context(settings, user_identity=user_identity, extra={"row_count": len(rows), "elapsed_seconds": elapsed_seconds}),
            )
            return jsonify(
                {
                    "status": "ok",
                    "rows": rows,
                    "elapsed_seconds": elapsed_seconds,
                    "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                }
            )
        except Exception as error:  # noqa: BLE001
            _log_portal_event(
                "team_dashboard_link_biz_project_jira_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=user_identity, extra=_classify_portal_error(error)),
            )
            current_app.logger.exception("Team Dashboard Link Biz Project Jira load failed.")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Could not load unlinked Jira tickets. Please retry or share the request ID.",
                        **_classify_portal_error(error),
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.post("/api/team-dashboard/link-biz-projects/suggestions")
    def team_dashboard_link_biz_project_suggestions():
        access_gate = _require_team_dashboard_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        rows = payload.get("rows") if isinstance(payload.get("rows"), list) else []
        team_payloads = payload.get("team_payloads") if isinstance(payload.get("team_payloads"), list) else None
        user_identity = _get_user_identity(settings)
        try:
            config = _get_team_dashboard_config_store().load()
            started_at = time.monotonic()
            result = _suggest_team_dashboard_link_biz_project_rows(settings, config, rows, team_payloads=team_payloads)
            elapsed_seconds = round(time.monotonic() - started_at, 2)
            _log_portal_event(
                "team_dashboard_link_biz_project_suggestions_loaded",
                **_build_request_log_context(
                    settings,
                    user_identity=user_identity,
                    extra={
                        "row_count": len(result["rows"]),
                        "matched_count": result["matched_count"],
                        "team_candidate_count": result["team_candidate_count"],
                        "keyword_candidate_count": result["keyword_candidate_count"],
                        "elapsed_seconds": elapsed_seconds,
                    },
                ),
            )
            return jsonify({"status": "ok", "elapsed_seconds": elapsed_seconds, **result})
        except Exception as error:  # noqa: BLE001
            _log_portal_event(
                "team_dashboard_link_biz_project_suggestions_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=user_identity, extra=_classify_portal_error(error)),
            )
            current_app.logger.exception("Team Dashboard Link Biz Project suggestions failed.")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Could not suggest BPMIS Biz Projects. Please retry or share the request ID.",
                        **_classify_portal_error(error),
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.post("/api/team-dashboard/link-biz-projects")
    def link_team_dashboard_biz_project():
        access_gate = _require_team_dashboard_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        jira_id = _extract_issue_key_from_text(str(payload.get("jira_id") or payload.get("jira_link") or ""))
        jira_link = str(payload.get("jira_link") or "").strip()
        reporter_email = str(payload.get("reporter_email") or payload.get("pm_email") or "").strip().lower()
        bpmis_id = str(
            payload.get("selected_bpmis_id")
            or payload.get("suggested_bpmis_id")
            or payload.get("bpmis_id")
            or ""
        ).strip()
        if not jira_id:
            return jsonify({"status": "error", "message": "Jira ID is required."}), HTTPStatus.BAD_REQUEST
        if not bpmis_id:
            return jsonify({"status": "error", "message": "Suggested BPMIS ID is required."}), HTTPStatus.BAD_REQUEST
        if not reporter_email:
            return jsonify({"status": "error", "message": "Reporter email is required to validate the BPMIS Biz Project owner."}), HTTPStatus.BAD_REQUEST

        user_identity = _get_user_identity(settings)
        try:
            bpmis_client = _build_bpmis_client_for_current_user(settings)
            allowed_candidates = _team_dashboard_link_biz_candidate_projects_by_pm(
                bpmis_client,
                [reporter_email],
                team_payloads=None,
            ).get(reporter_email, [])
            allowed_bpmis_ids = {str(project.get("bpmis_id") or "").strip() for project in allowed_candidates}
            if bpmis_id not in allowed_bpmis_ids:
                raise ToolError("Selected BPMIS Biz Project must belong to the Jira PM and be in an allowed status.")
            linked_detail = bpmis_client.link_jira_ticket_to_project(jira_id, bpmis_id)
            if bpmis_id not in _extract_parent_issue_ids_from_any(linked_detail):
                raise BPMISError("BPMIS link verification failed because the Jira detail does not include this Biz Project parent.")

            project_detail = {}
            try:
                project_detail = bpmis_client.get_issue_detail(bpmis_id)
            except Exception:  # noqa: BLE001 - the verified link is the source of truth; project cache can be sparse.
                project_detail = {}
            project = _normalize_team_dashboard_project(
                {
                    **(project_detail if isinstance(project_detail, dict) else {}),
                    "bpmis_id": bpmis_id,
                    "issue_id": bpmis_id,
                }
            )
            if not project.get("project_name"):
                project["project_name"] = str(
                    (project_detail if isinstance(project_detail, dict) else {}).get("project_name")
                    or (project_detail if isinstance(project_detail, dict) else {}).get("summary")
                    or payload.get("selected_project_title")
                    or payload.get("suggested_project_title")
                    or ""
                ).strip()
            if not jira_link:
                jira_link = f"{_jira_browse_base_url()}{jira_id}"

            _log_portal_event(
                "team_dashboard_link_biz_project_success",
                **_build_request_log_context(
                    settings,
                    user_identity=user_identity,
                    extra={"jira_id": jira_id, "bpmis_id": bpmis_id},
                ),
            )
            return jsonify(
                {
                    "status": "ok",
                    "jira_id": jira_id,
                    "jira_link": jira_link,
                    "bpmis_id": bpmis_id,
                    "project": project,
                    "ticket": {},
                }
            )
        except (BPMISError, ToolError) as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "team_dashboard_link_biz_project_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(settings, user_identity=user_identity, extra={**error_details, "jira_id": jira_id, "bpmis_id": bpmis_id}),
            )
            return jsonify({"status": "error", "message": str(error), **error_details}), HTTPStatus.BAD_REQUEST
        except Exception as error:  # noqa: BLE001
            _log_portal_event(
                "team_dashboard_link_biz_project_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=user_identity, extra={"jira_id": jira_id, "bpmis_id": bpmis_id}),
            )
            current_app.logger.exception("Team Dashboard Link Biz Project failed.")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Link Biz Project failed unexpectedly. Please retry or share the request ID.",
                        **_classify_portal_error(error),
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.post("/api/team-dashboard/monthly-report/draft")
    def team_dashboard_monthly_report_draft():
        access_gate = _require_team_dashboard_monthly_report_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        user_identity = _get_user_identity(settings)
        try:
            config = _get_team_dashboard_config_store().load()
            team_payloads = _load_all_team_dashboard_task_payloads(settings, config)
            report_period = resolve_monthly_report_period()
            request_payload = {
                "template": normalize_monthly_report_template(config.get("monthly_report_template")),
                "team_payloads": team_payloads,
                "report_intelligence_config": normalize_report_intelligence_config(config.get("report_intelligence_config")),
                "period_start": report_period.start.isoformat(),
                "period_end": report_period.end_date,
                "period_end_exclusive": report_period.end_exclusive.isoformat(),
                "product_scope": list(MONTHLY_REPORT_PRODUCT_SCOPE),
            }
            if _remote_bpmis_config_enabled(settings):
                data = _build_local_agent_client(settings).team_dashboard_monthly_report_draft_start(request_payload)
                job_id = str(data.get("job_id") or "").strip()
                if not job_id:
                    raise ToolError("Mac local-agent did not return a Monthly Report job id.")
                _log_portal_event(
                    "team_dashboard_monthly_report_draft_queued",
                    **_build_request_log_context(settings, user_identity=user_identity, extra={"job_id": job_id, "job_backend": "local_agent"}),
                )
                return jsonify({"status": "queued", "job_id": job_id, "job_backend": "local_agent"})
            job_store: JobStore = current_app.config["JOB_STORE"]
            job = job_store.create("team-dashboard-monthly-report-draft", title="Generate Monthly Report Draft")
            app_obj = current_app._get_current_object()
            thread = threading.Thread(
                target=_run_team_dashboard_monthly_report_draft_job,
                args=(app_obj, job.job_id, settings, request_payload, user_identity),
                daemon=True,
            )
            thread.start()
            _log_portal_event(
                "team_dashboard_monthly_report_draft_queued",
                **_build_request_log_context(
                    settings,
                    user_identity=user_identity,
                    extra={"job_id": job.job_id},
                ),
            )
            return jsonify({"status": "queued", "job_id": job.job_id})
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "team_dashboard_monthly_report_draft_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(settings, user_identity=user_identity, extra=error_details),
            )
            return jsonify({"status": "error", "message": str(error), **error_details}), HTTPStatus.BAD_REQUEST
        except Exception as error:  # noqa: BLE001
            _log_portal_event(
                "team_dashboard_monthly_report_draft_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=user_identity, extra=_classify_portal_error(error)),
            )
            current_app.logger.exception("Team Dashboard Monthly Report draft failed.")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Monthly Report draft generation failed unexpectedly. Please retry or share the request ID.",
                        **_classify_portal_error(error),
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.post("/api/team-dashboard/monthly-report/send")
    def team_dashboard_monthly_report_send():
        access_gate = _require_team_dashboard_monthly_report_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        draft_markdown = str(payload.get("draft_markdown") or "").strip()
        subject = str(payload.get("subject") or "").strip() or monthly_report_subject()
        recipient = str(payload.get("recipient") or "").strip() or DEFAULT_MONTHLY_REPORT_RECIPIENT
        user_identity = _get_user_identity(settings)
        try:
            send_payload = {
                "draft_markdown": draft_markdown,
                "subject": subject,
                "recipient": recipient,
            }
            if _local_agent_seatalk_enabled(settings):
                data = _build_local_agent_client(settings).team_dashboard_monthly_report_send(send_payload)
            else:
                result = send_monthly_report_email(
                    credential_store=current_app.config["GOOGLE_CREDENTIAL_STORE"],
                    owner_email=str(settings.gmail_seatalk_demo_owner_email or settings.seatalk_owner_email or "").strip().lower(),
                    recipient=recipient,
                    subject=subject,
                    draft_markdown=draft_markdown,
                )
                data = asdict(result)
            _log_portal_event(
                "team_dashboard_monthly_report_send_success",
                **_build_request_log_context(
                    settings,
                    user_identity=user_identity,
                    extra={"recipient": recipient, "subject": subject, "message_id": str(data.get("message_id") or "")},
                ),
            )
            memory_result = {"recorded": 0, "failed": 0}
            if _google_credentials_have_scopes(GMAIL_READONLY_SCOPE):
                try:
                    memory_result = _ingest_sent_monthly_reports_from_gmail(settings)
                except Exception:  # noqa: BLE001 - sent-mail memory ingestion must not block email sending.
                    current_app.logger.exception("Monthly Report sent-mail Work Memory ingestion failed.")
            else:
                current_app.logger.info("Skipping Monthly Report sent-mail Work Memory ingestion because gmail.readonly scope is absent.")
            data["work_memory"] = memory_result
            return jsonify({"status": "ok", **data})
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "team_dashboard_monthly_report_send_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(settings, user_identity=user_identity, extra=error_details),
            )
            return jsonify({"status": "error", "message": str(error), **error_details}), HTTPStatus.BAD_REQUEST
        except Exception as error:  # noqa: BLE001
            _log_portal_event(
                "team_dashboard_monthly_report_send_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=user_identity, extra=_classify_portal_error(error)),
            )
            current_app.logger.exception("Team Dashboard Monthly Report send failed.")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Monthly Report email failed unexpectedly. Please retry or share the request ID.",
                        **_classify_portal_error(error),
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.post("/api/team-dashboard/prd-review")
    def team_dashboard_prd_review():
        access_gate = _require_team_dashboard_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        user_identity = _get_user_identity(settings)
        review_payload = {
            "owner_key": str(user_identity.get("config_key") or ""),
            "jira_id": str(payload.get("jira_id") or ""),
            "jira_link": str(payload.get("jira_link") or ""),
            "prd_url": str(payload.get("prd_url") or ""),
            "force_refresh": bool(payload.get("force_refresh")),
        }
        try:
            if _local_agent_source_code_qa_enabled(settings):
                data = _build_local_agent_client(settings).prd_review(review_payload)
            else:
                data = _build_prd_review_service(settings).review(PRDReviewRequest(**review_payload))
            _log_portal_event(
                "team_dashboard_prd_review_success",
                **_build_request_log_context(
                    settings,
                    user_identity=user_identity,
                    extra={
                        "jira_id": review_payload["jira_id"],
                        "prd_url_hash": hashlib.sha256(review_payload["prd_url"].encode("utf-8")).hexdigest()[:12],
                        "cached": bool(data.get("cached")),
                    },
                ),
            )
            return jsonify(data)
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "team_dashboard_prd_review_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=user_identity,
                    extra={**error_details, "jira_id": review_payload["jira_id"]},
                ),
            )
            return jsonify({"status": "error", "message": str(error), **error_details}), HTTPStatus.BAD_REQUEST
        except Exception as error:  # noqa: BLE001
            _log_portal_event(
                "team_dashboard_prd_review_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=user_identity, extra={"jira_id": review_payload["jira_id"]}),
            )
            current_app.logger.exception("Team Dashboard PRD review failed.")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "PRD review failed unexpectedly. Please retry or share the request ID.",
                        **_classify_portal_error(error),
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.post("/api/team-dashboard/prd-summary")
    def team_dashboard_prd_summary():
        access_gate = _require_team_dashboard_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        payload = request.get_json(silent=True) or {}
        user_identity = _get_user_identity(settings)
        summary_payload = {
            "owner_key": str(user_identity.get("config_key") or ""),
            "jira_id": str(payload.get("jira_id") or ""),
            "jira_link": str(payload.get("jira_link") or ""),
            "prd_url": str(payload.get("prd_url") or ""),
            "force_refresh": bool(payload.get("force_refresh")),
        }
        try:
            if _local_agent_source_code_qa_enabled(settings):
                data = _build_local_agent_client(settings).prd_summary(summary_payload)
            else:
                data = _build_prd_review_service(settings).summarize(PRDReviewRequest(**summary_payload))
            _log_portal_event(
                "team_dashboard_prd_summary_success",
                **_build_request_log_context(
                    settings,
                    user_identity=user_identity,
                    extra={
                        "jira_id": summary_payload["jira_id"],
                        "prd_url_hash": hashlib.sha256(summary_payload["prd_url"].encode("utf-8")).hexdigest()[:12],
                        "cached": bool(data.get("cached")),
                    },
                ),
            )
            return jsonify(data)
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "team_dashboard_prd_summary_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    user_identity=user_identity,
                    extra={**error_details, "jira_id": summary_payload["jira_id"]},
                ),
            )
            return jsonify({"status": "error", "message": str(error), **error_details}), HTTPStatus.BAD_REQUEST
        except Exception as error:  # noqa: BLE001
            _log_portal_event(
                "team_dashboard_prd_summary_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, user_identity=user_identity, extra={"jira_id": summary_payload["jira_id"]}),
            )
            current_app.logger.exception("Team Dashboard PRD summary failed.")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "PRD summary failed unexpectedly. Please retry or share the request ID.",
                        **_classify_portal_error(error),
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.post("/api/prd-self-assessment/review")
    def prd_self_assessment_review_api():
        return _run_prd_self_assessment_action(settings, action="review")

    @app.post("/api/prd-self-assessment/summary")
    def prd_self_assessment_summary_api():
        return _run_prd_self_assessment_action(settings, action="summary")

    @app.get("/download/default-sheet-template.csv")
    def download_default_sheet_template():
        sample_row = [
            "225159",
            "Standalone Cash Loan",
            "https://docs.google.com/document/d/example",
            "SG",
            "AF",
            "Fraud rule improvement",
            "https://confluence/example-prd",
            "Detailed Jira description goes here.",
            "",
        ]
        csv_lines = [
            ",".join(_csv_escape(header) for header in DEFAULT_SHEET_HEADERS),
            ",".join(_csv_escape(value) for value in sample_row),
        ]
        payload = io.BytesIO("\n".join(csv_lines).encode("utf-8"))
        return send_file(
            payload,
            mimetype="text/csv",
            as_attachment=True,
            download_name="bpmis_jira_default_sheet_template.csv",
        )

    @app.get("/download/default-sheet-template.xlsx")
    def download_default_sheet_template_xlsx():
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"

        sample_row = [
            "225159",
            "Standalone Cash Loan",
            "https://docs.google.com/document/d/example",
            "SG",
            "AF",
            "Fraud rule improvement",
            "https://confluence/example-prd",
            "Detailed Jira description goes here.",
            "",
        ]
        worksheet.append(DEFAULT_SHEET_HEADERS)
        worksheet.append(sample_row)

        header_font = Font(bold=True, color="1F2937")
        for cell in worksheet[1]:
            cell.font = header_font

        worksheet.freeze_panes = "A2"
        for column_letter, width in {
            "A": 16,
            "B": 28,
            "C": 12,
            "D": 30,
            "E": 14,
            "F": 28,
            "G": 34,
            "H": 42,
            "I": 28,
        }.items():
            worksheet.column_dimensions[column_letter].width = width

        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)
        return send_file(
            payload,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="bpmis_jira_default_sheet_template.xlsx",
        )

    @app.post("/preview")
    def preview():
        login_gate = _require_google_login(settings)
        if login_gate is not None:
            return login_gate
        try:
            service = _build_service(settings)
            results, _headers = service.preview()
            session["last_results"] = _serialize_results(results, include_skipped=False)
            _log_portal_event("preview_success", **_build_request_log_context(settings))
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "preview_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(settings, extra=error_details),
            )
            flash(str(error), "error")
        return redirect(url_for("index"))

    @app.post("/run")
    def run():
        login_gate = _require_google_login(settings)
        if login_gate is not None:
            return login_gate
        dry_run = request.form.get("dry_run") == "on"
        try:
            service = _build_service(settings)
            results = service.run(dry_run=dry_run)
            session["last_results"] = _serialize_results(results, include_skipped=False)
            session["run_notice"] = _build_run_notice(results, dry_run=dry_run)
            if dry_run:
                flash("Dry run completed without updating the sheet.", "success")
            else:
                flash("Run completed. Check the results table below.", "success")
            _log_portal_event(
                "run_success",
                **_build_request_log_context(settings, extra={"dry_run": dry_run}),
            )
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "run_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(settings, extra={**error_details, "dry_run": dry_run}),
            )
            flash(str(error), "error")
        return redirect(url_for("index"))

    @app.post("/api/jobs/preview")
    def create_preview_job():
        return _start_job("preview", dry_run=True)

    @app.post("/api/jobs/run")
    def create_run_job():
        return _start_job("run", dry_run=False)

    @app.post("/api/jobs/sync-bpmis-projects")
    def create_sync_bpmis_projects_job():
        return _start_job("sync-bpmis-projects", dry_run=False)

    @app.route("/api/local-agent/<path:agent_path>", methods=["GET", "POST", "PATCH", "DELETE"])
    def local_agent_public_proxy(agent_path: str):
        return _proxy_local_agent_request(agent_path)

    @app.get("/api/bpmis-projects")
    def bpmis_projects():
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate
        user_identity = _get_user_identity(settings)
        store = _get_bpmis_project_store()
        return jsonify({"status": "ok", "projects": store.list_projects(user_key=user_identity["config_key"])})

    @app.delete("/api/bpmis-projects/<bpmis_id>")
    def delete_bpmis_project(bpmis_id: str):
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate
        user_identity = _get_user_identity(settings)
        deleted = _get_bpmis_project_store().soft_delete_project(user_key=user_identity["config_key"], bpmis_id=bpmis_id)
        return jsonify({"status": "ok", "deleted": deleted, "scope": "portal_only"})

    @app.patch("/api/bpmis-projects/order")
    def reorder_bpmis_projects():
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate
        payload = request.get_json(silent=True) or {}
        bpmis_ids = payload.get("bpmis_ids") if isinstance(payload.get("bpmis_ids"), list) else []
        user_identity = _get_user_identity(settings)
        projects = _get_bpmis_project_store().reorder_projects(
            user_key=user_identity["config_key"],
            bpmis_ids=[str(item or "") for item in bpmis_ids],
        )
        return jsonify({"status": "ok", "projects": projects, "scope": "portal_only"})

    @app.patch("/api/bpmis-projects/<bpmis_id>/comment")
    def update_bpmis_project_comment(bpmis_id: str):
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate
        payload = request.get_json(silent=True) or {}
        user_identity = _get_user_identity(settings)
        updated = _get_bpmis_project_store().update_project_comment(
            user_key=user_identity["config_key"],
            bpmis_id=bpmis_id,
            pm_comment=str(payload.get("pm_comment") or ""),
        )
        return jsonify({"status": "ok", "updated": updated, "scope": "portal_only"})

    @app.get("/api/bpmis-projects/<bpmis_id>/jira-options")
    def bpmis_project_jira_options(bpmis_id: str):
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate
        try:
            user_identity = _get_user_identity(settings)
            service = _build_portal_jira_creation_service(settings)
            options = service.jira_options(user_key=user_identity["config_key"], bpmis_id=bpmis_id)
            return jsonify({"status": "ok", **options})
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.get("/api/bpmis-projects/<bpmis_id>/jira-tickets")
    def bpmis_project_jira_tickets(bpmis_id: str):
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate
        try:
            user_identity = _get_user_identity(settings)
            service = _build_portal_jira_creation_service(settings)
            include_live = str(request.args.get("live") or "").strip().lower() in {"1", "true", "yes"}
            tickets = service.list_tickets(
                user_key=user_identity["config_key"],
                bpmis_id=bpmis_id,
                include_live=include_live,
            )
            return jsonify({"status": "ok", "tickets": tickets})
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.delete("/api/bpmis-projects/<bpmis_id>/jira-tickets/<ticket_id>")
    def delete_bpmis_project_jira_ticket(bpmis_id: str, ticket_id: str):
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate
        try:
            user_identity = _get_user_identity(settings)
            service = _build_portal_jira_creation_service(settings)
            deleted = service.delete_ticket(
                user_key=user_identity["config_key"],
                bpmis_id=bpmis_id,
                ticket_id=ticket_id,
            )
            return jsonify({"status": "ok", "deleted": deleted, "scope": "bpmis_and_portal"})
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.patch("/api/bpmis-projects/<bpmis_id>/jira-tickets/<ticket_id>/status")
    def update_bpmis_project_jira_ticket_status(bpmis_id: str, ticket_id: str):
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate
        payload = request.get_json(silent=True) or {}
        status_value = str(payload.get("status") or "").strip() if isinstance(payload, dict) else ""
        if not status_value:
            return jsonify({"status": "error", "message": "Jira status is required."}), HTTPStatus.BAD_REQUEST
        try:
            user_identity = _get_user_identity(settings)
            service = _build_portal_jira_creation_service(settings)
            ticket = service.update_ticket_status(
                user_key=user_identity["config_key"],
                bpmis_id=bpmis_id,
                ticket_id=ticket_id,
                status=status_value,
            )
            return jsonify({"status": "ok", "ticket": ticket})
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.patch("/api/bpmis-projects/<bpmis_id>/jira-tickets/<ticket_id>/version")
    def update_bpmis_project_jira_ticket_version(bpmis_id: str, ticket_id: str):
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate
        payload = request.get_json(silent=True) or {}
        version_name = str(payload.get("version_name") or "").strip() if isinstance(payload, dict) else ""
        version_id = str(payload.get("version_id") or "").strip() if isinstance(payload, dict) else ""
        if not version_name and not version_id:
            return jsonify({"status": "error", "message": "Jira fix version is required."}), HTTPStatus.BAD_REQUEST
        try:
            user_identity = _get_user_identity(settings)
            service = _build_portal_jira_creation_service(settings)
            ticket = service.update_ticket_version(
                user_key=user_identity["config_key"],
                bpmis_id=bpmis_id,
                ticket_id=ticket_id,
                version_name=version_name,
                version_id=version_id,
            )
            return jsonify({"status": "ok", "ticket": ticket})
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.post("/api/bpmis-projects/<bpmis_id>/jira-tickets")
    def create_bpmis_project_jira_tickets(bpmis_id: str):
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate
        payload = request.get_json(silent=True) or {}
        items = payload.get("items") if isinstance(payload, dict) else []
        if not isinstance(items, list):
            return jsonify({"status": "error", "message": "items must be a list."}), HTTPStatus.BAD_REQUEST
        try:
            user_identity = _get_user_identity(settings)
            service = _build_portal_jira_creation_service(settings)
            results = service.create_tickets(user_key=user_identity["config_key"], bpmis_id=bpmis_id, items=items)
            status_code = HTTPStatus.OK if any(result.get("status") == "created" for result in results) else HTTPStatus.BAD_REQUEST
            return jsonify({"status": "ok" if status_code == HTTPStatus.OK else "error", "results": results}), status_code
        except ToolError as error:
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST

    @app.get("/api/productization-upgrade-summary/versions")
    def productization_upgrade_summary_versions():
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate

        started_at = time.monotonic()
        query = str(request.args.get("q") or "").strip()
        if not query:
            return jsonify({"status": "error", "message": "Version keyword is required."}), HTTPStatus.BAD_REQUEST

        try:
            bpmis_client = _build_bpmis_client_for_current_user(settings)
            versions = bpmis_client.search_versions(query)
            return jsonify(
                {
                    "status": "ok",
                    "items": [_serialize_productization_version_candidate(item) for item in versions],
                    "elapsed_seconds": round(time.monotonic() - started_at, 3),
                }
            )
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "productization_version_search_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(settings, extra={**error_details, "query": query}),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "productization_version_search_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, extra={"query": query}),
            )
            current_app.logger.exception("Productization version search failed.")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Unable to search versions right now. Please try again shortly.",
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.get("/api/productization-upgrade-summary/issues")
    def productization_upgrade_summary_issues():
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate

        started_at = time.monotonic()
        version_id = str(request.args.get("version_id") or "").strip()
        show_all_before_team_filtering = str(request.args.get("show_all_before_team_filtering") or "").strip().lower() in {
            "1",
            "true",
            "yes",
            "on",
        }
        if not version_id:
            return jsonify({"status": "error", "message": "version_id is required."}), HTTPStatus.BAD_REQUEST

        try:
            bpmis_client = _build_bpmis_client_for_current_user(settings)
            rows = bpmis_client.list_issues_for_version(version_id)
            config_data = _load_current_user_config(settings)
            raw_count = len(rows)
            rows, filter_metadata = _filter_productization_issue_rows_for_pm_team(
                rows,
                config_data,
                show_all_before_team_filtering=show_all_before_team_filtering,
            )
            normalized_items = [_normalize_productization_issue_row(item) for item in rows]
            return jsonify(
                {
                    "status": "ok",
                    "items": normalized_items,
                    "raw_count": raw_count,
                    "filtered_count": len(rows),
                    "llm_description_generated": False,
                    "llm_generated_count": 0,
                    "codex_detailed_feature": False,
                    "codex_generated_count": 0,
                    "elapsed_seconds": round(time.monotonic() - started_at, 3),
                    **filter_metadata,
                }
            )
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "productization_issue_lookup_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(settings, extra={**error_details, "version_id": version_id}),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "productization_issue_lookup_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, extra={"version_id": version_id}),
            )
            current_app.logger.exception("Productization issue lookup failed.")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Unable to load upgrade tickets right now. Please try again shortly.",
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.get("/api/productization-upgrade-summary/llm-descriptions")
    def productization_upgrade_summary_llm_descriptions():
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate

        started_at = time.monotonic()
        version_id = str(request.args.get("version_id") or "").strip()
        show_all_before_team_filtering = str(request.args.get("show_all_before_team_filtering") or "").strip().lower() in {
            "1",
            "true",
            "yes",
            "on",
        }
        if not version_id:
            return jsonify({"status": "error", "message": "version_id is required."}), HTTPStatus.BAD_REQUEST

        try:
            bpmis_client = _build_bpmis_client_for_current_user(settings)
            rows = bpmis_client.list_issues_for_version(version_id)
            config_data = _load_current_user_config(settings)
            raw_count = len(rows)
            rows, filter_metadata = _filter_productization_issue_rows_for_pm_team(
                rows,
                config_data,
                show_all_before_team_filtering=show_all_before_team_filtering,
            )
            normalized_items = [_normalize_productization_issue_row(item) for item in rows]
            codex_metadata = _apply_codex_productization_detailed_features(
                normalized_items,
                rows,
                settings=settings,
            )
            return jsonify(
                {
                    "status": "ok",
                    "items": normalized_items,
                    "raw_count": raw_count,
                    "filtered_count": len(rows),
                    "elapsed_seconds": round(time.monotonic() - started_at, 3),
                    **codex_metadata,
                    **filter_metadata,
                }
            )
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "productization_llm_description_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(settings, extra=error_details),
            )
            return jsonify({"status": "error", "message": str(error), **error_details}), HTTPStatus.BAD_REQUEST
        except Exception:
            request_id = getattr(g, "request_id", "")
            _log_portal_event(
                "productization_llm_description_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, extra={"version_id": version_id}),
            )
            current_app.logger.exception("Productization LLM Description generation failed.")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Productization LLM Description generation failed unexpectedly. Please retry or share the request ID.",
                        "request_id": request_id,
                        "error_category": "unexpected_internal",
                        "error_code": "server_error",
                        "error_retryable": True,
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    @app.get("/api/jobs/<job_id>")
    def get_job(job_id: str):
        snapshot = current_app.config["JOB_STORE"].snapshot(job_id)
        if snapshot is None:
            if _remote_bpmis_config_enabled(settings):
                try:
                    remote_snapshot = _build_local_agent_client(settings).team_dashboard_monthly_report_job(job_id)
                    if isinstance(remote_snapshot, dict) and str(remote_snapshot.get("state") or ""):
                        return jsonify(remote_snapshot)
                except ToolError:
                    pass
            return jsonify({"status": "error", "message": "Job not found."}), 404
        action = str(snapshot.get("action") or "")
        if action == "source-code-qa-query":
            access_gate = _require_source_code_qa_access(settings, api=True)
            if access_gate is not None:
                return access_gate
            scoped_snapshot = _source_code_qa_job_snapshot_for_current_user(job_id)
            if scoped_snapshot is None:
                return jsonify({"status": "error", "message": "Job not found."}), 404
            return jsonify(_public_source_code_qa_job_snapshot(scoped_snapshot))
        if action == "source-code-qa-effort-assessment":
            access_gate = _require_source_code_qa_manage_access(settings, api=True)
            if access_gate is not None:
                return access_gate
            return jsonify(_public_source_code_qa_job_snapshot(snapshot))
        return jsonify(snapshot)

    @app.get("/api/source-code-qa/query-jobs/<job_id>")
    def source_code_qa_query_job_api(job_id: str):
        access_gate = _require_source_code_qa_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        snapshot = _source_code_qa_job_snapshot_for_current_user(job_id)
        if snapshot is None:
            return jsonify(
                {
                    "status": "error",
                    "message": "Source Code Q&A job was not found.",
                    "error_category": "job_not_found",
                    "error_code": "job_not_found",
                    "error_retryable": False,
                }
            ), HTTPStatus.NOT_FOUND
        return jsonify(_public_source_code_qa_job_snapshot(snapshot))

    @app.get("/api/source-code-qa/query-jobs/<job_id>/events")
    def source_code_qa_query_job_events_api(job_id: str):
        access_gate = _require_source_code_qa_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        if _source_code_qa_job_snapshot_for_current_user(job_id) is None:
            return jsonify(
                {
                    "status": "error",
                    "message": "Source Code Q&A job was not found.",
                    "error_category": "job_not_found",
                    "error_code": "job_not_found",
                    "error_retryable": False,
                }
            ), HTTPStatus.NOT_FOUND

        def event_stream():
            last_payload = ""
            deadline = time.time() + 900
            while time.time() < deadline:
                snapshot = _source_code_qa_job_snapshot_for_current_user(job_id)
                if snapshot is None:
                    payload = {
                        "status": "error",
                        "state": "failed",
                        "message": "Source Code Q&A job was not found.",
                        "error": "Source Code Q&A job was not found.",
                        "error_category": "job_not_found",
                        "error_code": "job_not_found",
                        "error_retryable": False,
                    }
                    yield f"event: failed\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"
                    return
                payload = _public_source_code_qa_job_snapshot(snapshot)
                payload_text = json.dumps(payload, ensure_ascii=False)
                if payload_text != last_payload:
                    event_name = "message"
                    if payload.get("state") == "completed":
                        event_name = "completed"
                    elif payload.get("state") == "failed":
                        event_name = "failed"
                    yield f"event: {event_name}\ndata: {payload_text}\n\n"
                    last_payload = payload_text
                    if payload.get("state") in {"completed", "failed"}:
                        return
                else:
                    yield ": keepalive\n\n"
                time.sleep(0.9)

        return Response(
            stream_with_context(event_stream()),
            mimetype="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "X-Accel-Buffering": "no",
            },
        )

    @app.get("/api/source-code-qa/effort-assessment-jobs/<job_id>")
    def source_code_qa_effort_assessment_job_api(job_id: str):
        access_gate = _require_source_code_qa_manage_access(settings, api=True)
        if access_gate is not None:
            return access_gate
        snapshot = current_app.config["JOB_STORE"].snapshot(job_id)
        if snapshot is None or snapshot.get("action") != "source-code-qa-effort-assessment":
            return jsonify(
                {
                    "status": "error",
                    "message": "Source Code Q&A effort assessment job was not found.",
                    "error_category": "job_not_found",
                    "error_code": "job_not_found",
                    "error_retryable": False,
                }
            ), HTTPStatus.NOT_FOUND
        return jsonify(_public_source_code_qa_job_snapshot(snapshot))

    @app.get("/api/source-code-qa/effort-assessment-jobs/<job_id>/events")
    def source_code_qa_effort_assessment_job_events_api(job_id: str):
        access_gate = _require_source_code_qa_manage_access(settings, api=True)
        if access_gate is not None:
            return access_gate

        def event_stream():
            last_payload = ""
            deadline = time.time() + 900
            while time.time() < deadline:
                snapshot = current_app.config["JOB_STORE"].snapshot(job_id)
                if snapshot is None or snapshot.get("action") != "source-code-qa-effort-assessment":
                    payload = {
                        "status": "error",
                        "state": "failed",
                        "message": "Source Code Q&A effort assessment job was not found.",
                        "error": "Source Code Q&A effort assessment job was not found.",
                        "error_category": "job_not_found",
                        "error_code": "job_not_found",
                        "error_retryable": False,
                    }
                    yield f"event: failed\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"
                    return
                payload = _public_source_code_qa_job_snapshot(snapshot)
                payload_text = json.dumps(payload, ensure_ascii=False)
                if payload_text != last_payload:
                    event_name = "message"
                    if payload.get("state") == "completed":
                        event_name = "completed"
                    elif payload.get("state") == "failed":
                        event_name = "failed"
                    yield f"event: {event_name}\ndata: {payload_text}\n\n"
                    last_payload = payload_text
                    if payload.get("state") in {"completed", "failed"}:
                        return
                else:
                    yield ": keepalive\n\n"
                time.sleep(0.9)

        return Response(
            stream_with_context(event_stream()),
            mimetype="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "X-Accel-Buffering": "no",
            },
        )

    @app.post("/api/spreadsheets/create-template")
    def create_template_spreadsheet():
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate
        if "google_credentials" not in session:
            return jsonify({"status": "error", "message": "Please connect Google Sheets first."}), HTTPStatus.BAD_REQUEST

        config_data = _load_current_user_config(settings)
        input_tab_name = str(config_data.get("input_tab_name", "") or "").strip() or settings.input_tab_name or "Sheet1"
        spreadsheet_title = "BPMIS Automation Tool"
        try:
            created = GoogleSheetsService.create_template_spreadsheet(
                get_google_credentials(),
                spreadsheet_title=spreadsheet_title,
                input_tab=input_tab_name,
                headers=DEFAULT_SHEET_HEADERS,
            )
            _log_portal_event(
                "spreadsheet_template_create_success",
                **_build_request_log_context(
                    settings,
                    extra={"spreadsheet_id": created["spreadsheet_id"], "input_tab_name": input_tab_name},
                ),
            )
            return jsonify(
                {
                    "status": "ok",
                    "message": "A new Google Sheet was created and prefilled with the template header row.",
                    **created,
                }
            )
        except ToolError as error:
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "spreadsheet_template_create_tool_error",
                level=logging.WARNING,
                **_build_request_log_context(
                    settings,
                    extra={**error_details, "input_tab_name": input_tab_name},
                ),
            )
            return jsonify({"status": "error", "message": str(error)}), HTTPStatus.BAD_REQUEST
        except Exception:
            _log_portal_event(
                "spreadsheet_template_create_unexpected_error",
                level=logging.ERROR,
                **_build_request_log_context(settings, extra={"input_tab_name": input_tab_name}),
            )
            current_app.logger.exception("Google Sheet template creation failed.")
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Unable to create a new Google Sheet right now. Please try again shortly.",
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    def _start_job(action: str, *, dry_run: bool):
        login_gate = _require_google_login(settings, api=True)
        if login_gate is not None:
            return login_gate
        if action != "sync-bpmis-projects" and "google_credentials" not in session:
            return jsonify({"status": "error", "message": "Please connect Google Sheets first."}), 400

        user_identity = _get_user_identity(settings)
        config_data = _load_user_config_for_identity(settings, user_identity) or config_store._normalize({})
        config_data = _hydrate_setup_defaults(config_data, user_identity)
        _apply_sync_email_policy(config_data, user_identity)
        config_data["_user_key"] = user_identity["config_key"]
        job_store: JobStore = current_app.config["JOB_STORE"]
        title = {
            "preview": "Preview Eligible Rows",
            "run": "Run Ticket Creation",
            "sync-bpmis-projects": "Sync BPMIS Projects",
        }.get(action, "Background Job")
        job = job_store.create(action, title=title)
        credentials_payload = dict(session.get("google_credentials") or {})

        thread = threading.Thread(
            target=_run_background_job,
            args=(app, job.job_id, action, settings, config_data, credentials_payload, dry_run),
            daemon=True,
        )
        thread.start()
        _log_portal_event(
            "job_queued",
            **_build_request_log_context(
                settings,
                user_identity=user_identity,
                extra={"job_id": job.job_id, "action": action, "dry_run": dry_run},
            ),
        )
        return jsonify({"status": "queued", "job_id": job.job_id})

    return app


def _build_service(settings: Settings) -> JiraCreationService:
    user_identity = _get_user_identity(settings)
    config_store = _get_config_store()
    config_data = _load_user_config_for_identity(settings, user_identity) or config_store._normalize({})
    sheets = _build_sheets_service(settings, config_data)
    field_mappings_override = config_store.build_field_mappings(config_data)
    access_token = _resolve_bpmis_access_token(config_data, settings)
    return JiraCreationService(
        settings,
        sheets,
        access_token=access_token,
        field_mappings_override=field_mappings_override,
    )


def _build_bpmis_client_for_current_user(settings: Settings):
    config_data = _load_current_user_config(settings)
    access_token = _resolve_bpmis_access_token(config_data, settings)
    return build_bpmis_client(settings, access_token=access_token)


def _load_current_user_config(settings: Settings) -> dict[str, Any]:
    user_identity = _get_user_identity(settings)
    config_store = _get_config_store()
    config_data = _load_user_config_for_identity(settings, user_identity) or config_store._normalize({})
    hydrated = _hydrate_setup_defaults(config_data, user_identity)
    _apply_sync_email_policy(hydrated, user_identity)
    return hydrated


def _build_service_from_config(
    settings: Settings,
    config_data: dict[str, Any],
    credentials_payload: dict[str, Any],
) -> JiraCreationService:
    credentials = Credentials(**credentials_payload)
    sheets = _build_sheets_service_with_credentials(settings, config_data, credentials)
    field_mappings_override = _get_config_store().build_field_mappings(config_data)
    access_token = _resolve_bpmis_access_token(config_data, settings)
    return JiraCreationService(
        settings,
        sheets,
        access_token=access_token,
        field_mappings_override=field_mappings_override,
    )


def _build_project_sync_service_from_config(
    settings: Settings,
    config_data: dict[str, Any],
    credentials_payload: dict[str, Any],
) -> BPMISProjectSyncService:
    credentials = Credentials(**credentials_payload)
    sheets = _build_sheets_service_with_credentials(settings, config_data, credentials)
    bpmis_client = build_bpmis_client(settings, access_token=_resolve_bpmis_access_token(config_data, settings))
    return BPMISProjectSyncService(sheets, bpmis_client)


def _build_portal_project_sync_service(settings: Settings, config_data: dict[str, Any]) -> PortalProjectSyncService:
    bpmis_client = build_bpmis_client(settings, access_token=_resolve_bpmis_access_token(config_data, settings))
    return PortalProjectSyncService(_get_bpmis_project_store(), bpmis_client)


def _build_portal_jira_creation_service(settings: Settings) -> PortalJiraCreationService:
    config_data = _load_current_user_config(settings)
    return PortalJiraCreationService(
        store=_get_bpmis_project_store(),
        bpmis_client=build_bpmis_client(settings, access_token=_resolve_bpmis_access_token(config_data, settings)),
        config_store=_get_config_store(),
        config_data=config_data,
    )


def _build_sheets_service(settings: Settings, config_data: dict[str, object] | None = None) -> GoogleSheetsService:
    credentials = get_google_credentials()
    return _build_sheets_service_with_credentials(settings, config_data or {}, credentials)


def _build_gmail_dashboard_service() -> GmailDashboardService:
    credentials = get_google_credentials()
    report_config = {}
    try:
        report_config = _get_team_dashboard_config_store().load().get("report_intelligence_config") or {}
    except Exception:
        report_config = {}
    return GmailDashboardService(credentials=credentials, cache_key=_current_google_email(), report_intelligence_config=report_config)


def _meeting_recorder_config(settings: Settings) -> MeetingRecorderConfig:
    return MeetingRecorderConfig(
        ffmpeg_bin=settings.meeting_recorder_ffmpeg_bin,
        video_input=settings.meeting_recorder_video_input,
        audio_input=settings.meeting_recorder_audio_input,
        video_fps=settings.meeting_recorder_video_fps,
        video_max_width=settings.meeting_recorder_video_max_width,
        video_max_height=settings.meeting_recorder_video_max_height,
        avfoundation_pixel_format=settings.meeting_recorder_avfoundation_pixel_format,
        screen_preflight_timeout_seconds=settings.meeting_recorder_screen_preflight_timeout_seconds,
        audio_only_fallback_on_screen_failure=settings.meeting_recorder_audio_only_fallback_on_screen_failure,
        frame_interval_seconds=settings.meeting_recorder_frame_interval_seconds,
        vision_model=settings.meeting_recorder_vision_model,
        transcribe_provider=settings.meeting_recorder_transcribe_provider,
        whisper_cpp_bin=settings.meeting_recorder_whisper_cpp_bin,
        whisper_model=settings.meeting_recorder_whisper_model,
        whisper_language=settings.meeting_recorder_whisper_language,
        transcript_segment_workers=settings.meeting_recorder_transcript_segment_workers,
        whisper_threads=settings.meeting_recorder_whisper_threads,
        background_nice=settings.meeting_recorder_background_nice,
        capture_status_every_buffers=settings.meeting_recorder_capture_status_every_buffers,
        startup_silence_grace_seconds=settings.meeting_recorder_startup_silence_grace_seconds,
    )


def _get_meeting_record_store() -> MeetingRecordStore:
    return current_app.config["MEETING_RECORD_STORE"]


def _get_meeting_recorder_runtime() -> MeetingRecorderRuntime:
    return current_app.config["MEETING_RECORDER_RUNTIME"]


def _get_work_memory_store() -> WorkMemoryStore:
    return current_app.config["WORK_MEMORY_STORE"]


def _record_work_memory_items(items: list[dict[str, Any]], *, event: str, owner_email: str = "", write_ledger: bool = True) -> dict[str, int]:
    recorded = 0
    failed = 0
    duplicate = 0
    if not items:
        if write_ledger:
            try:
                _get_work_memory_store().record_ingestion_run(
                    source_type=event,
                    owner_email=owner_email or _current_google_email(),
                    status="ok",
                    recorded_count=0,
                    failed_count=0,
                )
            except Exception:
                current_app.logger.debug("Work Memory empty ingestion ledger write failed for %s.", event, exc_info=True)
        return {"recorded": 0, "failed": 0, "duplicate": 0}
    store = _get_work_memory_store()
    for item in items:
        try:
            source_type = str(item.get("source_type") or "").strip()
            source_id = str(item.get("source_id") or "").strip()
            item_type = str(item.get("item_type") or "").strip()
            owner_email = str(item.get("owner_email") or "").strip().lower()
            existing_id = None
            if source_type and source_id and item_type and owner_email:
                existing_id = hashlib.sha256("\x1f".join([source_type, source_id, item_type, owner_email]).encode("utf-8")).hexdigest()[:32]
                if store.get_item(existing_id):
                    duplicate += 1
            store.record_memory_item(**item)
            recorded += 1
        except Exception:  # noqa: BLE001 - memory ingestion must not break the primary tool flow.
            failed += 1
            current_app.logger.exception("Work Memory ingestion failed for %s.", event)
    if write_ledger:
        try:
            store.record_ingestion_run(
                source_type=event,
                owner_email=owner_email or _current_google_email(),
                status="ok" if failed == 0 else "partial_error",
                scanned_count=len(items),
                matched_count=len(items),
                recorded_count=recorded,
                duplicate_count=duplicate,
                failed_count=failed,
            )
        except Exception:
            current_app.logger.debug("Work Memory ingestion ledger write failed for %s.", event, exc_info=True)
    return {"recorded": recorded, "failed": failed, "duplicate": duplicate}


def _record_team_dashboard_work_memory(team_payloads: list[dict[str, Any]], *, owner_email: str) -> dict[str, int]:
    items: list[dict[str, Any]] = []
    for team_payload in team_payloads or []:
        if isinstance(team_payload, dict) and not team_payload.get("error"):
            items.extend(team_dashboard_memory_items(team_payload, owner_email=owner_email))
    return _record_work_memory_items(items, event="team_dashboard")


def _record_meeting_work_memory(record: dict[str, Any]) -> dict[str, int]:
    return _record_work_memory_items(meeting_record_memory_items(record), event="meeting_recorder")


def _record_source_code_qa_work_memory(
    *,
    owner_email: str,
    pm_team: str,
    country: str,
    question: str,
    result: dict[str, Any],
    session_id: str = "",
    job_id: str = "",
) -> dict[str, int]:
    item = source_code_qa_memory_item(
        owner_email=owner_email,
        pm_team=pm_team,
        country=country,
        question=question,
        result=result,
        session_id=session_id,
        job_id=job_id,
    )
    return _record_work_memory_items([item], event="source_code_qa")


SENT_MONTHLY_REPORT_SUBJECT_PATTERN = re.compile(
    r"^\[Banking\]\s+Product\s+Update\s+\("
    r"\d{1,2}\s+[A-Za-z]{3}\s*-\s*\d{1,2}\s+[A-Za-z]{3}"
    r"\)\s+-\s+Anti-Fraud,\s+Credit\s+Risk\s+&\s+Ops\s+Risk$",
    re.IGNORECASE,
)


def _is_sent_monthly_report_subject(subject: str) -> bool:
    return bool(SENT_MONTHLY_REPORT_SUBJECT_PATTERN.match(str(subject or "").strip()))


def _ingest_existing_work_memory_sources(settings: Settings, *, date_range: str = "90d", sources: list[str] | None = None) -> dict[str, Any]:
    owner_email = _current_google_email()
    items: list[dict[str, Any]] = []
    source_counts = {"meeting_records": 0, "team_dashboard_cache": 0}
    source_filter = {str(source or "").strip() for source in sources or [] if str(source or "").strip()}

    if not source_filter or "meeting_recorder" in source_filter:
        try:
            if _local_agent_meeting_recorder_enabled(settings):
                records = _build_local_agent_client(settings).meeting_recorder_records(owner_email=owner_email)
            else:
                records = _get_meeting_record_store().list_records(owner_email=owner_email)
            for record in records:
                if str(record.get("status") or "").strip().lower() != "completed":
                    continue
                record_items = meeting_record_memory_items(record)
                if record_items:
                    source_counts["meeting_records"] += 1
                    items.extend(record_items)
        except Exception:
            current_app.logger.exception("Work Memory meeting backfill failed.")

    if not source_filter or "team_dashboard" in source_filter:
        try:
            config = _get_team_dashboard_config_store().load()
            task_cache = config.get("task_cache") if isinstance(config.get("task_cache"), dict) else {}
            cached_teams = task_cache.get("teams") if isinstance(task_cache.get("teams"), dict) else {}
            for team_key, cached_team in cached_teams.items():
                if not isinstance(cached_team, dict):
                    continue
                team_payload = {
                    **cached_team,
                    "team_key": str(cached_team.get("team_key") or team_key),
                    "member_emails": _normalize_team_dashboard_emails(cached_team.get("member_emails") or []),
                    "loading": False,
                    "loaded": True,
                    "error": "",
                    "progress_text": "",
                    "cache_source": "backfill",
                }
                team_items = team_dashboard_memory_items(team_payload, owner_email=owner_email)
                if team_items:
                    source_counts["team_dashboard_cache"] += 1
                    items.extend(team_items)
        except Exception:
            current_app.logger.exception("Work Memory Team Dashboard cache backfill failed.")

    result = _record_work_memory_items(items, event="work_memory_existing_backfill")
    return {**result, **source_counts, "date_range": date_range, "sources": sorted(source_filter)}


def _ingest_sent_monthly_reports_from_gmail(settings: Settings) -> dict[str, Any]:
    del settings
    if not _google_credentials_have_scopes(GMAIL_READONLY_SCOPE):
        raise ConfigError("Gmail read permission is missing. Reconnect Google once to grant gmail.readonly.")
    service = _build_gmail_dashboard_service()
    owner_email = _current_google_email()
    queries = [
        'in:sent newer_than:180d from:me subject:"[Banking] Product Update"',
        'in:sent newer_than:180d from:me subject:"Anti-Fraud, Credit Risk & Ops Risk"',
        'in:sent newer_than:180d from:me "[Banking] Product Update" "Anti-Fraud, Credit Risk & Ops Risk"',
    ]
    seen_ids: set[str] = set()
    records = []
    for query in queries:
        for message_id in service._list_message_ids(query=query, max_messages=20):
            if message_id in seen_ids:
                continue
            seen_ids.add(message_id)
            records.append(service._fetch_message_full(message_id))
    items = []
    for record in records:
        headers = getattr(record, "headers", {}) or {}
        sender_text = str(headers.get("from") or "").casefold()
        if owner_email and owner_email not in sender_text:
            continue
        if not _is_sent_monthly_report_subject(str(headers.get("subject") or "")):
            continue
        items.append(sent_monthly_report_memory_item_from_gmail_record(owner_email=owner_email, record=record))
    result = _record_work_memory_items(items, event="gmail_sent_monthly_report_scan")
    return {"scanned": len(records), "matched": len(items), **result}


def _run_incremental_memory_ingestion(settings: Settings, *, window: str = "7d", reconciliation: bool = False) -> dict[str, Any]:
    effective_window = "90d" if reconciliation else (str(window or "7d").strip() or "7d")
    sources = ["meeting_recorder", "team_dashboard"]
    result: dict[str, Any] = {
        "window": effective_window,
        "reconciliation": bool(reconciliation),
        "sources": {},
        "recorded": 0,
        "failed": 0,
        "duplicate": 0,
    }
    existing = _ingest_existing_work_memory_sources(settings, date_range=effective_window, sources=sources)
    result["sources"]["existing"] = existing
    for key in ("recorded", "failed", "duplicate"):
        result[key] += int(existing.get(key) or 0)
    if _google_credentials_have_scopes(GMAIL_READONLY_SCOPE):
        try:
            sent_reports = _ingest_sent_monthly_reports_from_gmail(settings)
            result["sources"]["gmail_sent_monthly_report"] = sent_reports
            for key in ("recorded", "failed", "duplicate"):
                result[key] += int(sent_reports.get(key) or 0)
        except Exception as error:  # noqa: BLE001 - incremental ingestion should keep partial source results.
            current_app.logger.exception("Work Memory incremental Gmail sent report ingestion failed.")
            result["sources"]["gmail_sent_monthly_report"] = {"status": "error", "message": str(error)}
            result["failed"] += 1
            _get_work_memory_store().record_ingestion_run(
                source_type="gmail_sent_monthly_report_incremental",
                owner_email=_current_google_email(),
                cursor=effective_window,
                status="error",
                failed_count=1,
                error=str(error),
            )
    else:
        result["sources"]["gmail_sent_monthly_report"] = {
            "status": "skipped",
            "message": "Gmail readonly scope is missing.",
        }
    distilled = _get_work_memory_store().distill_work_memory(owner_email=_current_google_email(), date_range=effective_window)
    result["sources"]["distill"] = distilled
    _get_work_memory_store().record_ingestion_run(
        source_type="work_memory_incremental",
        owner_email=_current_google_email(),
        cursor=effective_window,
        status="ok" if result["failed"] == 0 else "partial_error",
        scanned_count=sum(int((source_result or {}).get("scanned") or (source_result or {}).get("meeting_records") or 0) for source_result in result["sources"].values() if isinstance(source_result, dict)),
        matched_count=result["recorded"],
        recorded_count=result["recorded"],
        duplicate_count=result["duplicate"],
        failed_count=result["failed"],
        metadata={"sources": sources, "reconciliation": bool(reconciliation)},
    )
    return result


def _run_work_memory_gmail_backfill_job(app: Flask, job_id: str, payload: dict[str, Any]) -> None:
    with app.app_context():
        job_store: JobStore = current_app.config["JOB_STORE"]
        owner_email = str(payload.get("owner_email") or "").strip().lower()
        try:
            job_store.update(job_id, state="running", stage="starting", message="Starting Gmail Work Memory backfill.")
            result = _backfill_gmail_work_memory(
                owner_email=owner_email,
                credentials_payload=payload.get("credentials") if isinstance(payload.get("credentials"), dict) else {},
                report_intelligence_config=payload.get("report_intelligence_config") if isinstance(payload.get("report_intelligence_config"), dict) else {},
                days=int(payload.get("days") or 90),
                max_messages=payload.get("max_messages") if payload.get("max_messages") is not None else None,
                drive_read_enabled=bool(payload.get("drive_read_enabled")),
                job_id=job_id,
            )
            job_store.complete(
                job_id,
                results=[result],
                notice={"level": "success" if int(result.get("failed") or 0) == 0 else "warning", "message": "Gmail Work Memory backfill completed."},
            )
        except Exception as error:  # noqa: BLE001 - keep background job failures visible in the job store.
            current_app.logger.exception("Gmail Work Memory backfill failed.")
            _get_work_memory_store().record_ingestion_run(
                source_type="gmail_backfill",
                owner_email=owner_email,
                status="error",
                failed_count=1,
                error=str(error),
            )
            error_details = _classify_portal_error(error)
            job_store.fail(
                job_id,
                str(error),
                error_category=str(error_details.get("error_category") or ""),
                error_code=str(error_details.get("error_code") or ""),
                error_retryable=bool(error_details.get("error_retryable")),
            )


def _backfill_gmail_work_memory(
    *,
    owner_email: str,
    credentials_payload: dict[str, Any],
    report_intelligence_config: dict[str, Any],
    days: int,
    max_messages: int | None,
    drive_read_enabled: bool,
    job_id: str,
) -> dict[str, Any]:
    if not credentials_payload:
        raise ConfigError("Google credentials are missing for Gmail backfill. Reconnect Google and retry.")
    credentials = Credentials(**credentials_payload)
    service = GmailDashboardService(
        credentials=credentials,
        cache_key=owner_email,
        report_intelligence_config=report_intelligence_config,
    )
    job_store: JobStore = current_app.config["JOB_STORE"]
    started_at = datetime.now(timezone.utc).isoformat()
    refs = service.list_work_memory_message_refs(days=days, max_messages=max_messages)
    unique_message_ids: list[str] = []
    seen_message_ids: set[str] = set()
    for ref in refs:
        message_id = str(ref.get("id") or "").strip()
        if message_id and message_id not in seen_message_ids:
            seen_message_ids.add(message_id)
            unique_message_ids.append(message_id)
    original_total = len(unique_message_ids)
    existing_message_ids = _get_work_memory_store().existing_source_ids(
        source_type="gmail",
        owner_email=owner_email,
        source_ids=unique_message_ids,
    )
    pending_message_ids = [message_id for message_id in unique_message_ids if message_id not in existing_message_ids]
    skipped_existing = original_total - len(pending_message_ids)
    total = len(pending_message_ids)
    skip_suffix = f" Skipping {skipped_existing} already recorded." if skipped_existing else ""
    job_store.update(job_id, stage="scanning", message=f"Scanning {total}/{original_total} Gmail message(s).{skip_suffix}", current=0, total=total)
    vip_people = _gmail_work_memory_vip_people(report_intelligence_config)
    key_projects = _gmail_work_memory_key_projects()
    scanned = 0
    matched = 0
    recorded = 0
    duplicate = 0
    failed = 0
    attachment_processed = 0
    drive_links_processed = 0
    last_error = ""
    for batch_start in range(0, total, GMAIL_WORK_MEMORY_FETCH_BATCH_SIZE):
        batch_ids = pending_message_ids[batch_start:batch_start + GMAIL_WORK_MEMORY_FETCH_BATCH_SIZE]
        batch_items: list[dict[str, Any]] = []
        scanned += len(batch_ids)
        if scanned == len(batch_ids) or scanned % 100 == 0 or scanned == total:
            job_store.update(job_id, stage="fetching", message=f"Fetching {scanned}/{total} Gmail message(s); skipped {skipped_existing}.", current=scanned, total=total)
        records, fetch_failures, fetch_error = _fetch_gmail_work_memory_records(
            credentials_payload=credentials_payload,
            owner_email=owner_email,
            report_intelligence_config=report_intelligence_config,
            message_ids=batch_ids,
        )
        failed += fetch_failures
        if fetch_error:
            last_error = fetch_error
        for record in records:
            try:
                message_id = str(getattr(record, "message_id", "") or "").strip()
                if not message_id:
                    continue
                headers = getattr(record, "headers", {}) or {}
                if service.is_export_noise(headers):
                    continue
                matched_vips, vip_roles = _gmail_work_memory_matched_vips(headers, vip_people)
                body_text = str(getattr(record, "body_text", "") or "")
                match_text = "\n".join(
                    [
                        str(headers.get("subject") or ""),
                        str(headers.get("from") or ""),
                        str(headers.get("to") or ""),
                        str(headers.get("cc") or ""),
                        body_text,
                    ]
                )
                report_matches = _gmail_work_memory_report_matches(match_text, report_intelligence_config, key_projects)
                batch_items.append(
                    gmail_message_memory_item(
                        owner_email=owner_email,
                        record=record,
                        matched_vips=matched_vips,
                        vip_email_roles=vip_roles,
                        report_matches=report_matches,
                    )
                )
                matched += 1
                if matched_vips:
                    attachment_items, attachment_failures = _gmail_work_memory_vip_attachment_items(
                        service=service,
                        owner_email=owner_email,
                        record=record,
                        matched_vips=matched_vips,
                    )
                    batch_items.extend(attachment_items)
                    attachment_processed += len(attachment_items)
                    failed += attachment_failures
                    drive_items = _gmail_work_memory_vip_drive_items(
                        credentials=credentials,
                        owner_email=owner_email,
                        record=record,
                        matched_vips=matched_vips,
                        drive_read_enabled=drive_read_enabled,
                    )
                    batch_items.extend(drive_items)
                    drive_links_processed += len(drive_items)
            except Exception as error:  # noqa: BLE001 - one bad message must not stop the backfill.
                failed += 1
                last_error = str(error)
                current_app.logger.warning("Gmail Work Memory message backfill skipped message_id=%s: %s", message_id, error)
        result = _record_work_memory_items(batch_items, event="gmail_backfill", owner_email=owner_email, write_ledger=False)
        recorded += int(result.get("recorded") or 0)
        duplicate += int(result.get("duplicate") or 0)
        failed += int(result.get("failed") or 0)
    status = "ok" if failed == 0 else "partial_error"
    _get_work_memory_store().record_ingestion_run(
        source_type="gmail_backfill",
        owner_email=owner_email,
        cursor=f"days={days};messages={scanned};skipped={skipped_existing}",
        status=status,
        scanned_count=scanned,
        matched_count=matched,
        recorded_count=recorded,
        duplicate_count=duplicate,
        failed_count=failed,
        error=last_error,
        metadata={
            "days": days,
            "max_messages": max_messages,
            "original_message_count": original_total,
            "skipped_existing_message_count": skipped_existing,
            "vip_people_count": len(vip_people),
            "attachment_processed": attachment_processed,
            "drive_links_processed": drive_links_processed,
            "drive_read_enabled": drive_read_enabled,
        },
        started_at=started_at,
        completed_at=datetime.now(timezone.utc).isoformat(),
    )
    distilled = _get_work_memory_store().distill_work_memory(owner_email=owner_email, date_range=f"{days}d", sources=["gmail", "gmail_attachment", "gmail_drive_link"])
    return {
        "status": status,
        "days": days,
        "scanned": scanned,
        "skipped_existing": skipped_existing,
        "original_message_count": original_total,
        "matched": matched,
        "recorded": recorded,
        "duplicate": duplicate,
        "failed": failed,
        "attachment_processed": attachment_processed,
        "drive_links_processed": drive_links_processed,
        "distill": distilled,
    }


def _gmail_work_memory_fetch_workers() -> int:
    raw_value = str(os.getenv("GMAIL_WORK_MEMORY_FETCH_WORKERS") or "").strip()
    if not raw_value:
        return GMAIL_WORK_MEMORY_FETCH_WORKERS
    try:
        return max(1, min(int(raw_value), 8))
    except ValueError:
        return GMAIL_WORK_MEMORY_FETCH_WORKERS


def _fetch_gmail_work_memory_records(
    *,
    credentials_payload: dict[str, Any],
    owner_email: str,
    report_intelligence_config: dict[str, Any],
    message_ids: list[str],
) -> tuple[list[Any], int, str]:
    if not message_ids:
        return [], 0, ""
    workers = min(_gmail_work_memory_fetch_workers(), len(message_ids))
    records: list[Any] = []
    failed = 0
    last_error = ""

    def fetch_one(message_id: str) -> Any:
        local_service = GmailDashboardService(
            credentials=Credentials(**credentials_payload),
            cache_key=owner_email,
            report_intelligence_config=report_intelligence_config,
        )
        return local_service.fetch_work_memory_message(message_id)

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(fetch_one, message_id): message_id for message_id in message_ids}
        for future in as_completed(futures):
            message_id = futures[future]
            try:
                records.append(future.result())
            except Exception as error:  # noqa: BLE001 - one message failure must not stop the backfill batch.
                failed += 1
                last_error = str(error)
                current_app.logger.warning("Gmail Work Memory full fetch failed message_id=%s: %s", message_id, error)
    records.sort(key=lambda item: getattr(item, "internal_date", datetime.min.replace(tzinfo=timezone.utc)), reverse=True)
    return records, failed, last_error


def _gmail_work_memory_vip_people(report_intelligence_config: dict[str, Any]) -> list[dict[str, Any]]:
    normalized = normalize_report_intelligence_config(report_intelligence_config)
    return [vip for vip in normalized.get("vip_people") or [] if isinstance(vip, dict)]


def _gmail_work_memory_key_projects() -> list[dict[str, Any]]:
    try:
        config = _get_team_dashboard_config_store().load()
        task_cache = config.get("task_cache") if isinstance(config.get("task_cache"), dict) else {}
        teams = task_cache.get("teams") if isinstance(task_cache.get("teams"), dict) else {}
        projects: list[dict[str, Any]] = []
        for team in teams.values():
            if not isinstance(team, dict):
                continue
            for section_key in ("under_prd", "pending_live"):
                projects.extend(project for project in (team.get(section_key) or []) if isinstance(project, dict))
        return projects
    except Exception:
        return []


def _gmail_work_memory_matched_vips(headers: dict[str, str], vip_people: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, list[str]]]:
    role_addresses = {
        "from": _gmail_header_addresses(headers.get("from", "")),
        "to": _gmail_header_addresses(headers.get("to", "")),
        "cc": _gmail_header_addresses(headers.get("cc", "")),
    }
    matched: list[dict[str, Any]] = []
    roles: dict[str, list[str]] = {}
    for vip in vip_people:
        emails = {str(email or "").strip().lower() for email in vip.get("emails") or [] if str(email or "").strip()}
        if not emails:
            continue
        vip_roles = sorted(role for role, addresses in role_addresses.items() if emails.intersection(addresses))
        if not vip_roles:
            continue
        label = str(vip.get("display_name") or sorted(emails)[0]).strip()
        matched.append({"display_name": label, "role_tags": vip.get("role_tags") or [], "emails": sorted(emails)})
        roles[label] = vip_roles
    return matched, roles


def _gmail_header_addresses(value: str) -> set[str]:
    return {str(address or "").strip().lower() for _name, address in getaddresses([value or ""]) if str(address or "").strip()}


def _gmail_work_memory_report_matches(text: str, report_intelligence_config: dict[str, Any], key_projects: list[dict[str, Any]]) -> dict[str, Any]:
    try:
        from bpmis_jira_tool.report_intelligence import match_report_intelligence

        return match_report_intelligence(text, config=report_intelligence_config, key_projects=key_projects)
    except Exception:
        return {"matched_vips": [], "matched_keywords": [], "matched_key_projects": []}


def _gmail_work_memory_vip_attachment_items(
    *,
    service: GmailDashboardService,
    owner_email: str,
    record: Any,
    matched_vips: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], int]:
    items: list[dict[str, Any]] = []
    failed = 0
    message_id = str(getattr(record, "message_id", "") or "").strip()
    attachments = [
        attachment
        for attachment in (getattr(record, "attachments", []) or [])
        if _gmail_attachment_is_pdf(attachment)
    ][:GMAIL_WORK_MEMORY_MAX_VIP_ATTACHMENTS_PER_MESSAGE]
    for attachment in attachments:
        try:
            if int(getattr(attachment, "size", 0) or 0) > GMAIL_WORK_MEMORY_MAX_ATTACHMENT_BYTES:
                continue
            content = service.download_attachment(message_id=message_id, attachment_id=str(getattr(attachment, "attachment_id") or ""))
            if len(content) > GMAIL_WORK_MEMORY_MAX_ATTACHMENT_BYTES:
                continue
            text = _extract_pdf_text_for_work_memory(content)
            sha256 = hashlib.sha256(content).hexdigest()
            items.append(
                gmail_attachment_memory_item(
                    owner_email=owner_email,
                    record=record,
                    attachment=attachment,
                    text=text[:GMAIL_WORK_MEMORY_CONTENT_CHARS],
                    sha256=sha256,
                    matched_vips=matched_vips,
                )
            )
        except Exception as error:  # noqa: BLE001 - attachment failures are counted but isolated.
            failed += 1
            current_app.logger.info("VIP Gmail PDF attachment skipped: %s", error)
    return items, failed


def _gmail_attachment_is_pdf(attachment: Any) -> bool:
    filename = str(getattr(attachment, "filename", "") or "").strip().lower()
    mime_type = str(getattr(attachment, "mime_type", "") or "").strip().lower()
    return filename.endswith(".pdf") or mime_type == "application/pdf"


def _extract_pdf_text_for_work_memory(content: bytes) -> str:
    try:
        from pypdf import PdfReader  # type: ignore
    except ImportError as error:
        raise ToolError("PDF attachments are supported only when pypdf is installed on the server.") from error
    reader = PdfReader(io.BytesIO(content))
    lines: list[str] = []
    for page in reader.pages[:12]:
        lines.append(str(page.extract_text() or ""))
    text = "\n".join(lines).strip()
    if not text:
        raise ToolError("Unable to extract readable text from this PDF attachment.")
    return text[:GMAIL_WORK_MEMORY_CONTENT_CHARS]


def _gmail_work_memory_vip_drive_items(
    *,
    credentials: Credentials,
    owner_email: str,
    record: Any,
    matched_vips: list[dict[str, Any]],
    drive_read_enabled: bool,
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    links = list(dict.fromkeys(str(link or "").strip() for link in (getattr(record, "drive_links", []) or []) if str(link or "").strip()))
    for link in links[:GMAIL_WORK_MEMORY_MAX_VIP_ATTACHMENTS_PER_MESSAGE]:
        title = ""
        text = ""
        access_status = "unavailable"
        if drive_read_enabled:
            try:
                title, text = _read_google_drive_link_text(credentials=credentials, url=link)
                access_status = "ok" if text else "unavailable"
            except HttpError as error:
                status = getattr(getattr(error, "resp", None), "status", 0)
                access_status = "permission_denied" if int(status or 0) in {401, 403, 404} else "unavailable"
            except Exception:
                access_status = "unavailable"
        else:
            access_status = "missing_drive_scope"
        items.append(
            gmail_drive_link_memory_item(
                owner_email=owner_email,
                record=record,
                url=link,
                title=title,
                text=text[:GMAIL_WORK_MEMORY_CONTENT_CHARS],
                access_status=access_status,
                matched_vips=matched_vips,
            )
        )
    return items


def _read_google_drive_link_text(*, credentials: Credentials, url: str) -> tuple[str, str]:
    file_id = _google_drive_file_id_from_url(url)
    if not file_id:
        return "", ""
    service = _build_google_drive_service(credentials)
    metadata = service.files().get(fileId=file_id, fields="id,name,mimeType,size").execute()
    title = str(metadata.get("name") or file_id)
    mime_type = str(metadata.get("mimeType") or "")
    if mime_type.startswith("application/vnd.google-apps."):
        for export_mime in ("text/plain", "text/csv"):
            try:
                content = service.files().export_media(fileId=file_id, mimeType=export_mime).execute()
                return title, _bytes_to_text(content)
            except HttpError:
                continue
        return title, ""
    if mime_type == "application/pdf":
        content = service.files().get_media(fileId=file_id).execute()
        return title, _extract_pdf_text_for_work_memory(content)
    return title, ""


def _google_drive_http_timeout_seconds() -> int:
    raw_value = str(os.getenv("GOOGLE_DRIVE_HTTP_TIMEOUT_SECONDS") or "").strip()
    if not raw_value:
        return GOOGLE_DRIVE_HTTP_TIMEOUT_SECONDS
    try:
        return max(5, min(int(raw_value), 120))
    except ValueError:
        return GOOGLE_DRIVE_HTTP_TIMEOUT_SECONDS


def _build_google_drive_service(credentials: Credentials) -> Any:
    http = httplib2.Http(timeout=_google_drive_http_timeout_seconds())
    authed_http = google_auth_httplib2.AuthorizedHttp(credentials, http=http)
    return build_google_api("drive", "v3", http=authed_http, cache_discovery=False)


def _google_drive_file_id_from_url(url: str) -> str:
    parsed = urlparse(str(url or ""))
    path_parts = [part for part in parsed.path.split("/") if part]
    if "d" in path_parts:
        index = path_parts.index("d")
        if index + 1 < len(path_parts):
            return path_parts[index + 1]
    query = parse_qs(parsed.query)
    if query.get("id"):
        return str(query["id"][0] or "").strip()
    if path_parts and path_parts[0] == "open" and query.get("id"):
        return str(query["id"][0] or "").strip()
    return ""


def _bytes_to_text(value: bytes | str) -> str:
    if isinstance(value, str):
        return value.strip()
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return value.decode(encoding).strip()
        except UnicodeDecodeError:
            continue
    return ""


def _build_calendar_meeting_service() -> GoogleCalendarMeetingService:
    return GoogleCalendarMeetingService(get_google_credentials())


def _build_meeting_processing_service(settings: Settings) -> MeetingProcessingService:
    text_client = CodexTextGenerationClient(
        settings=settings,
        workspace_root=PROJECT_ROOT,
        prompt_mode="meeting_recorder_minutes_codex",
        codex_model=settings.prd_briefing_codex_model,
    )
    return MeetingProcessingService(
        store=_get_meeting_record_store(),
        config=_meeting_recorder_config(settings),
        text_client=text_client,
        credential_store=current_app.config.get("GOOGLE_CREDENTIAL_STORE"),
        portal_base_url=settings.team_portal_base_url,
    )


def _persist_owner_google_credentials(settings: Settings) -> None:
    current_email = _current_google_email()
    owner_emails = {
        str(settings.gmail_seatalk_demo_owner_email or settings.seatalk_owner_email or "").strip().lower(),
        str(settings.meeting_recorder_owner_email or "").strip().lower(),
    }
    owner_emails.discard("")
    if not current_email or current_email not in owner_emails:
        return
    store = current_app.config.get("GOOGLE_CREDENTIAL_STORE") if current_app else None
    credentials_payload = dict(session.get("google_credentials") or {})
    if store is None or not credentials_payload:
        return
    store.save(owner_email=current_email, credentials_payload=credentials_payload)


def _build_seatalk_dashboard_service(settings: Settings) -> SeaTalkDashboardService:
    if _local_agent_seatalk_enabled(settings):
        name_mapping_store = _get_seatalk_name_mapping_store(settings) if current_app else None
        return RemoteSeaTalkDashboardService(
            _build_local_agent_client(settings),
            name_mappings_provider=lambda: name_mapping_store.mappings() if name_mapping_store else {},
        )
    name_mapping_store = current_app.config.get("SEATALK_NAME_MAPPING_STORE") if current_app else None
    name_overrides_path = getattr(name_mapping_store, "storage_path", None)
    daily_cache_dir = current_app.config.get("SEATALK_DAILY_CACHE_DIR") if current_app else None
    return SeaTalkDashboardService(
        owner_email=settings.seatalk_owner_email,
        seatalk_app_path=settings.seatalk_local_app_path,
        seatalk_data_dir=settings.seatalk_local_data_dir,
        codex_workspace_root=PROJECT_ROOT,
        codex_model=os.getenv("SOURCE_CODE_QA_CODEX_MODEL", "codex-cli"),
        codex_timeout_seconds=settings.source_code_qa_codex_timeout_seconds,
        codex_concurrency=settings.source_code_qa_codex_concurrency,
        name_overrides_path=name_overrides_path,
        daily_cache_dir=daily_cache_dir,
    )


def _seatalk_dashboard_is_configured(settings: Settings) -> bool:
    if _local_agent_seatalk_enabled(settings):
        try:
            capabilities = _build_local_agent_client(settings).get_health().get("capabilities") or {}
            return bool(capabilities.get("seatalk_configured"))
        except ToolError:
            return False
    app_path = Path(str(settings.seatalk_local_app_path or "")).expanduser()
    data_dir = Path(str(settings.seatalk_local_data_dir or "")).expanduser()
    return bool(app_path.exists() and data_dir.exists() and (data_dir / "config.json").exists())


def _has_explicit_spreadsheet_link(config_data: dict[str, object] | None) -> bool:
    if not config_data:
        return False
    return bool(str(config_data.get("spreadsheet_link", "")).strip())


def _build_sheets_service_with_credentials(
    settings: Settings,
    config_data: dict[str, object] | None,
    credentials,
) -> GoogleSheetsService:
    config_data = config_data or {}
    spreadsheet_id = _resolve_spreadsheet_id(str(config_data.get("spreadsheet_link", "")).strip()) or settings.spreadsheet_id
    input_tab_name = str(config_data.get("input_tab_name", "")).strip() or settings.input_tab_name
    issue_id_header = str(config_data.get("issue_id_header", "")).strip() or "Issue ID"
    jira_ticket_link_header = str(config_data.get("jira_ticket_link_header", "")).strip() or "Jira Ticket Link"
    return GoogleSheetsService(
        credentials=credentials,
        spreadsheet_id=spreadsheet_id,
        common_tab=settings.common_tab_name,
        input_tab=input_tab_name,
        issue_id_header=issue_id_header,
        jira_ticket_link_header=jira_ticket_link_header,
    )


def _get_config_store() -> WebConfigStore:
    return current_app.config["CONFIG_STORE"]


def _get_bpmis_project_store() -> BPMISProjectStore:
    settings: Settings = current_app.config["SETTINGS"]
    if _remote_bpmis_config_enabled(settings):
        return RemoteBPMISProjectStore(_build_local_agent_client(settings))
    return current_app.config["BPMIS_PROJECT_STORE"]


def _get_team_dashboard_config_store() -> TeamDashboardConfigStore:
    settings: Settings = current_app.config["SETTINGS"]
    if _remote_bpmis_config_enabled(settings):
        return RemoteTeamDashboardConfigStore(_build_local_agent_client(settings))
    return current_app.config["TEAM_DASHBOARD_CONFIG_STORE"]


def _build_prd_review_service(settings: Settings) -> PRDReviewService:
    store: BriefingStore = current_app.config["PRD_BRIEFING_STORE"]
    confluence = ConfluenceConnector(
        base_url=settings.confluence_base_url,
        email=settings.confluence_email,
        api_token=settings.confluence_api_token,
        bearer_token=settings.confluence_bearer_token,
        store=store,
    )
    return PRDReviewService(
        store=store,
        confluence=confluence,
        settings=settings,
        workspace_root=PROJECT_ROOT,
    )


def _build_monthly_report_service(settings: Settings) -> MonthlyReportService:
    store: BriefingStore = current_app.config["PRD_BRIEFING_STORE"]
    confluence = ConfluenceConnector(
        base_url=settings.confluence_base_url,
        email=settings.confluence_email,
        api_token=settings.confluence_api_token,
        bearer_token=settings.confluence_bearer_token,
        store=store,
    )
    return MonthlyReportService(
        settings=settings,
        workspace_root=PROJECT_ROOT,
        seatalk_service=_build_seatalk_dashboard_service(settings),
        confluence=confluence,
    )


def _get_source_code_qa_session_store():
    settings: Settings = current_app.config["SETTINGS"]
    if _local_agent_source_code_qa_enabled(settings):
        return RemoteSourceCodeQASessionStore(_build_local_agent_client(settings))
    return current_app.config["SOURCE_CODE_QA_SESSION_STORE"]


def _get_source_code_qa_attachment_store():
    settings: Settings = current_app.config["SETTINGS"]
    if _local_agent_source_code_qa_enabled(settings):
        return RemoteSourceCodeQAAttachmentStore(_build_local_agent_client(settings))
    return current_app.config["SOURCE_CODE_QA_ATTACHMENT_STORE"]


def _get_source_code_qa_runtime_evidence_store():
    settings: Settings = current_app.config["SETTINGS"]
    if _local_agent_source_code_qa_enabled(settings):
        return RemoteSourceCodeQARuntimeEvidenceStore(_build_local_agent_client(settings))
    return current_app.config["SOURCE_CODE_QA_RUNTIME_EVIDENCE_STORE"]


def _get_source_code_qa_model_availability_store():
    settings: Settings = current_app.config["SETTINGS"]
    if _local_agent_source_code_qa_enabled(settings):
        return RemoteSourceCodeQAModelAvailabilityStore(_build_local_agent_client(settings))
    return current_app.config["SOURCE_CODE_QA_MODEL_AVAILABILITY_STORE"]


def _get_seatalk_todo_store(settings: Settings):
    if _local_agent_seatalk_enabled(settings):
        return RemoteSeaTalkTodoStore(_build_local_agent_client(settings))
    return current_app.config["SEATALK_TODO_STORE"]


def _get_seatalk_name_mapping_store(settings: Settings):
    if _local_agent_seatalk_enabled(settings):
        return RemoteSeaTalkNameMappingStore(_build_local_agent_client(settings))
    return current_app.config["SEATALK_NAME_MAPPING_STORE"]


def _remote_bpmis_config_enabled(settings: Settings) -> bool:
    return bool(
        _local_agent_mode_enabled(settings)
        and settings.local_agent_base_url
        and settings.local_agent_hmac_secret
        and settings.local_agent_bpmis_enabled
    )


def _is_local_agent_unavailable_error(error: ToolError) -> bool:
    message = str(error or "").strip().lower()
    if "local-agent" not in message and "local agent" not in message:
        return False
    unavailable_markers = (
        "unavailable",
        "offline",
        "connection refused",
        "connecttimeout",
        "read timed out",
        "timed out",
        "max retries exceeded",
        "err_ngrok_3200",
        "endpoint",
    )
    return any(marker in message for marker in unavailable_markers)


def _tool_error_category(error: ToolError) -> str:
    return "local_agent_unavailable" if _is_local_agent_unavailable_error(error) else "tool_error"


def _load_user_config_for_identity(settings: Settings, user_identity: dict[str, str | None]) -> dict[str, Any] | None:
    user_key = str(user_identity.get("config_key") or "").strip()
    if not user_key:
        return None
    if _remote_bpmis_config_enabled(settings):
        return _build_local_agent_client(settings).bpmis_config_load(user_key=user_key)
    return _get_config_store().load(user_key)


def _save_user_config_for_identity(settings: Settings, user_identity: dict[str, str | None], config: dict[str, Any]) -> dict[str, Any]:
    user_key = str(user_identity.get("config_key") or "").strip()
    if not user_key:
        raise ToolError("Sign in before saving Setup.")
    if _remote_bpmis_config_enabled(settings):
        return _build_local_agent_client(settings).bpmis_config_save(user_key=user_key, config=config)
    return _get_config_store().save(config, user_key)


def _migrate_user_config(settings: Settings, from_user_key: str, to_user_key: str) -> None:
    if _remote_bpmis_config_enabled(settings):
        _build_local_agent_client(settings).bpmis_config_migrate(from_user_key=from_user_key, to_user_key=to_user_key)
        return
    _get_config_store().migrate(from_user_key, to_user_key)


def _save_team_profile(settings: Settings, config_store: WebConfigStore, team_key: str, profile: dict[str, Any]) -> dict[str, Any]:
    if _remote_bpmis_config_enabled(settings):
        return _build_local_agent_client(settings).bpmis_team_profile_save(team_key=team_key, profile=profile)
    return config_store.save_team_profile(team_key, profile)


def _current_google_profile() -> dict[str, Any]:
    return session.get("google_profile") or {}


def _current_google_email() -> str:
    return str(_current_google_profile().get("email") or "").strip().lower()


def _is_portal_admin(email: str | None = None) -> bool:
    current_email = str(email or _current_google_email() or "").strip().lower()
    return current_email == PORTAL_ADMIN_EMAIL


def _is_portal_user(email: str | None = None) -> bool:
    current_email = str(email or _current_google_email() or "").strip().lower()
    return bool(
        current_email
        and (
            current_email.endswith("@npt.sg")
            or current_email == PORTAL_TEST_USER_EMAIL
        )
    )


def _current_google_user_is_blocked(settings: Settings) -> bool:
    if not _shared_portal_enabled(settings):
        return False
    email = _current_google_email()
    if not email:
        return False
    return not _is_portal_user(email)


def _shared_portal_enabled(settings: Settings) -> bool:
    return bool(
        settings.team_portal_base_url
        or settings.team_allowed_email_domains
    )


def _site_requires_google_login(settings: Settings) -> bool:
    return _shared_portal_enabled(settings)


def _google_session_is_connected() -> bool:
    return "google_credentials" in session and bool(_current_google_email())


def _can_access_prd_briefing(settings: Settings) -> bool:
    return _is_portal_user()


def _can_access_prd_self_assessment(settings: Settings) -> bool:
    return _is_portal_user()


def _can_access_gmail_seatalk_demo(settings: Settings) -> bool:
    return _is_portal_admin()


def _can_access_meeting_recorder(settings: Settings) -> bool:
    return _is_portal_admin()


def _can_access_source_code_qa(settings: Settings) -> bool:
    return _is_portal_user()


def _can_manage_source_code_qa(settings: Settings) -> bool:
    return _is_portal_admin()


def _can_access_work_memory(settings: Settings) -> bool:
    return _is_portal_admin()


def _source_code_qa_auth_payload(settings: Settings) -> dict[str, Any]:
    email = _current_google_email()
    owner_email = settings.source_code_qa_owner_email.strip().lower()
    normalized_admins = {PORTAL_ADMIN_EMAIL}
    if _is_portal_admin(email):
        match_source = "portal_admin"
    else:
        match_source = ""
    return {
        "signed_in_email": email,
        "can_manage": _is_portal_admin(email),
        "owner_email": owner_email,
        "admin_email_count": len(normalized_admins),
        "admin_match_source": match_source,
    }


def _source_code_qa_git_auth_ready(service: Any, settings: Settings) -> bool:
    if hasattr(service, "git_auth_ready"):
        return bool(service.git_auth_ready())
    return bool(settings.source_code_qa_gitlab_token)


def _build_source_code_qa_service(llm_provider: str | None = None) -> SourceCodeQAService:
    service: SourceCodeQAService = current_app.config["SOURCE_CODE_QA_SERVICE"]
    resolved = service if llm_provider is None else service.with_llm_provider(str(llm_provider or ""))
    if _local_agent_source_code_qa_enabled(current_app.config["SETTINGS"]):
        return RemoteSourceCodeQAService(_build_local_agent_client(current_app.config["SETTINGS"]), service, llm_provider=llm_provider or resolved.llm_provider_name)
    return resolved


def _source_code_qa_query_sync_mode(settings: Settings) -> str:
    mode = str(os.getenv("SOURCE_CODE_QA_QUERY_SYNC_MODE") or "").strip().lower()
    if mode in {"blocking", "background", "disabled"}:
        return mode
    return "background" if _local_agent_source_code_qa_enabled(settings) else "blocking"


def _prepare_source_code_qa_auto_sync(
    service: Any,
    *,
    pm_team: str,
    country: str,
    progress_callback: Any | None = None,
) -> dict[str, Any]:
    settings: Settings = current_app.config["SETTINGS"]
    mode = _source_code_qa_query_sync_mode(settings)
    key = service.mapping_key(pm_team, country) if hasattr(service, "mapping_key") else f"{pm_team}:{country}"
    if mode == "disabled":
        return {
            "attempted": False,
            "status": "skipped",
            "reason": "query-time repository auto-sync is disabled",
            "key": key,
        }
    if mode == "background":
        if progress_callback:
            progress_callback("auto_sync_queued", "Repository freshness check is running in the background.", 0, 1)
        if hasattr(service, "ensure_synced_today_background"):
            return service.ensure_synced_today_background(pm_team=pm_team, country=country)

        app_obj = current_app._get_current_object()
        logger = current_app.logger

        def run_background_sync() -> None:
            with app_obj.app_context():
                try:
                    service.ensure_synced_today(pm_team=pm_team, country=country)
                except Exception:
                    logger.exception("Source Code Q&A background auto-sync failed for %s.", key)

        threading.Thread(target=run_background_sync, daemon=True).start()
        return {
            "attempted": False,
            "status": "background_queued",
            "reason": "repository freshness check queued in the background",
            "key": key,
        }
    if progress_callback:
        progress_callback("auto_sync_check", "Checking repository sync schedule.", 0, 1)
    result = service.ensure_synced_today(pm_team=pm_team, country=country)
    if progress_callback:
        if result.get("attempted"):
            progress_callback("auto_sync_completed", "Repository auto-sync completed; starting code search.", 1, 1)
        else:
            progress_callback("auto_sync_completed", "Repository indexes do not need scheduled sync; starting code search.", 1, 1)
    return result


def _build_local_agent_client(settings: Settings) -> LocalAgentClient:
    return LocalAgentClient(
        base_url=settings.local_agent_base_url or "",
        hmac_secret=settings.local_agent_hmac_secret or "",
        timeout_seconds=settings.local_agent_timeout_seconds,
        connect_timeout_seconds=settings.local_agent_connect_timeout_seconds,
    )


def _proxy_local_agent_request(agent_path: str):
    settings: Settings = current_app.config["SETTINGS"]
    configured_base_url = (settings.local_agent_base_url or "").strip().rstrip("/")
    local_base_url = configured_base_url or _local_agent_loopback_base_url()
    normalized_agent_path = agent_path.lstrip("/")
    if configured_base_url:
        target_path = f"/api/local-agent/{normalized_agent_path}"
    else:
        target_path = "/healthz" if normalized_agent_path == "healthz" else f"/api/local-agent/{normalized_agent_path}"
    target_url = f"{local_base_url}{target_path}"
    headers = {
        key: value
        for key, value in request.headers.items()
        if key.lower()
        not in {
            "host",
            "content-length",
            "connection",
            "accept-encoding",
        }
    }
    try:
        connect_timeout_seconds = max(
            1,
            min(
                int(settings.local_agent_connect_timeout_seconds or 10),
                int(settings.local_agent_timeout_seconds or 300),
            ),
        )
        response = _LOCAL_AGENT_SESSION.request(
            request.method,
            target_url,
            params=request.args,
            data=request.get_data() if request.method != "GET" else None,
            headers=headers,
            timeout=(connect_timeout_seconds, settings.local_agent_timeout_seconds),
        )
    except requests.RequestException as error:
        return jsonify({"status": "error", "message": f"Mac local-agent is unavailable: {error}"}), HTTPStatus.BAD_GATEWAY

    excluded_headers = {"content-encoding", "content-length", "connection", "transfer-encoding"}
    proxy_headers = [(key, value) for key, value in response.headers.items() if key.lower() not in excluded_headers]
    return current_app.response_class(response.content, status=response.status_code, headers=proxy_headers)


def _local_agent_loopback_base_url() -> str:
    host = str(os.environ.get("LOCAL_AGENT_HOST") or "127.0.0.1").strip() or "127.0.0.1"
    port = str(os.environ.get("LOCAL_AGENT_PORT") or "7007").strip() or "7007"
    return f"http://{host}:{port}".rstrip("/")


def _local_agent_mode_enabled(settings: Settings) -> bool:
    mode = (settings.local_agent_mode or "").strip().lower()
    return mode in {"sync", "remote", "cloud_run", "enabled"}


def _local_agent_source_code_qa_enabled(settings: Settings) -> bool:
    return bool(
        _local_agent_mode_enabled(settings)
        and settings.local_agent_base_url
        and settings.local_agent_hmac_secret
        and settings.local_agent_source_code_qa_enabled
    )


def _local_agent_seatalk_enabled(settings: Settings) -> bool:
    return bool(
        _local_agent_mode_enabled(settings)
        and settings.local_agent_base_url
        and settings.local_agent_hmac_secret
        and settings.local_agent_seatalk_enabled
    )


def _local_agent_meeting_recorder_enabled(settings: Settings) -> bool:
    return bool(
        _local_agent_mode_enabled(settings)
        and settings.local_agent_base_url
        and settings.local_agent_hmac_secret
    )


def _source_code_qa_model_availability() -> dict[str, bool]:
    return _get_source_code_qa_model_availability_store().get()


def _source_code_qa_options_payload(service: SourceCodeQAService) -> dict[str, Any]:
    options = service.options_payload()
    availability = _source_code_qa_model_availability()
    providers = []
    for provider in options.get("llm_providers") or []:
        provider_payload = dict(provider)
        value = str(provider_payload.get("value") or "")
        available = bool(availability.get(value, False))
        provider_payload["available"] = available
        provider_payload["disabled"] = not available
        base_label = str(provider_payload.get("label") or value).replace(" (Unavailable)", "")
        provider_payload["label"] = base_label if available else f"{base_label} (Unavailable)"
        providers.append(provider_payload)
    options["llm_providers"] = providers
    options["runtime_capabilities"] = _source_code_qa_runtime_capabilities_payload()
    return options


def _source_code_qa_runtime_capabilities_payload() -> dict[str, dict[str, dict[str, bool]]]:
    teams = ("AF", "GRC", "CRMS")
    countries = tuple(CRMS_COUNTRIES)
    capabilities = {
        team: {
            country: {"hasConfig": False, "hasDB": False}
            for country in countries
        }
        for team in teams
    }
    try:
        store = _get_source_code_qa_runtime_evidence_store()
        for team in teams:
            for country in countries:
                try:
                    evidence_items = store.list(pm_team=team, country=country)
                except ToolError:
                    evidence_items = []
                for item in evidence_items:
                    source_type = str(item.get("source_type") or "").strip().lower()
                    if source_type == "apollo":
                        capabilities[team][country]["hasConfig"] = True
                    elif source_type == "db":
                        capabilities[team][country]["hasDB"] = True
    except Exception:  # noqa: BLE001 - capability badges are advisory only.
        return capabilities
    return capabilities


def _source_code_qa_provider_available(llm_provider: str | None) -> bool:
    provider = SourceCodeQAService.normalize_query_llm_provider(llm_provider)
    return bool(_source_code_qa_model_availability().get(provider, False))


def _source_code_qa_public_answer_mode(answer_mode: str | None) -> str:
    mode = str(answer_mode or "auto").strip()
    return mode if mode in {"auto", "gemini_flash"} else "auto"


def _source_code_qa_query_mode(query_mode: str | None) -> str:
    return "deep"


def _source_code_qa_attachment_ids(payload: dict[str, Any]) -> list[str]:
    raw_ids = payload.get("attachment_ids") if isinstance(payload, dict) else []
    if raw_ids is None:
        return []
    if not isinstance(raw_ids, list):
        raise ToolError("attachment_ids must be a list.")
    attachment_ids = [str(item or "").strip() for item in raw_ids if str(item or "").strip()]
    if len(attachment_ids) > SourceCodeQAAttachmentStore.MAX_ATTACHMENTS:
        raise ToolError(f"At most {SourceCodeQAAttachmentStore.MAX_ATTACHMENTS} attachments are supported per Source Code Q&A question.")
    return attachment_ids


def _resolve_source_code_qa_query_attachments(
    payload: dict[str, Any],
    *,
    owner_email: str,
    session_id: str,
) -> list[dict[str, Any]]:
    attachment_ids = _source_code_qa_attachment_ids(payload)
    if not attachment_ids:
        return []
    if not session_id:
        raise ToolError("A Source Code Q&A session is required before sending attachments.")
    session_payload = _get_source_code_qa_session_store().get(session_id, owner_email=owner_email)
    if session_payload is None:
        raise ToolError("Source Code Q&A session was not found for these attachments.")
    return _get_source_code_qa_attachment_store().resolve_many(
        owner_email=owner_email,
        session_id=session_id,
        attachment_ids=attachment_ids,
    )


def _source_code_qa_public_attachments(attachments: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        SourceCodeQAAttachmentStore.public_metadata(item)
        for item in attachments
        if isinstance(item, dict)
    ]


def _resolve_source_code_qa_runtime_evidence(*, pm_team: str, country: str) -> list[dict[str, Any]]:
    try:
        return _get_source_code_qa_runtime_evidence_store().resolve_scope(pm_team=pm_team, country=country)
    except ToolError:
        raise
    except Exception as error:  # noqa: BLE001 - runtime evidence must not break code Q&A.
        current_app.logger.warning("Source Code Q&A runtime evidence could not be loaded: %s", error)
        return []


def _source_code_qa_public_runtime_evidence(evidence: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        SourceCodeQARuntimeEvidenceStore.public_metadata(item)
        for item in evidence
        if isinstance(item, dict)
    ]


def _source_code_qa_effort_assessment_language(value: Any) -> str:
    normalized = str(value or "").strip().lower()
    if normalized in {"en", "english"}:
        return "en"
    return "zh"


def _source_code_qa_effort_sentences(text: str) -> list[str]:
    raw = str(text or "").strip()
    if not raw:
        return []
    parts = re.split(r"(?<=[。！？!?])\s+|\n+", raw)
    return [part.strip() for part in parts if part.strip()]


def _source_code_qa_effort_matches(text: str, patterns: list[str]) -> bool:
    lowered = str(text or "").lower()
    return any(re.search(pattern, lowered, flags=re.IGNORECASE) for pattern in patterns)


def _source_code_qa_effort_unique(items: list[Any]) -> list[Any]:
    seen: set[str] = set()
    output: list[Any] = []
    for item in items:
        if item is None:
            continue
        key = json.dumps(item, ensure_ascii=False, sort_keys=True) if isinstance(item, dict) else str(item)
        if key in seen:
            continue
        seen.add(key)
        output.append(item)
    return output


def _source_code_qa_load_json_file(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


@lru_cache(maxsize=1)
def _load_source_code_qa_effort_dictionaries() -> dict[str, Any]:
    return _source_code_qa_load_json_file(SOURCE_CODE_QA_EFFORT_DICTIONARY_PATH)


@lru_cache(maxsize=1)
def _load_source_code_qa_domain_profile_config() -> dict[str, Any]:
    return _source_code_qa_load_json_file(SOURCE_CODE_QA_DOMAIN_PROFILES_PATH)


@lru_cache(maxsize=1)
def _load_source_code_qa_domain_knowledge_config() -> dict[str, Any]:
    return _source_code_qa_load_json_file(SOURCE_CODE_QA_DOMAIN_KNOWLEDGE_PATH)


def _source_code_qa_effort_domain_entries(pm_team: str) -> list[dict[str, Any]]:
    dictionaries = _load_source_code_qa_effort_dictionaries()
    domain = ((dictionaries.get("domains") or {}).get(str(pm_team or "").upper()) or {})
    entries = domain.get("entries") if isinstance(domain, dict) else []
    entries = entries if isinstance(entries, list) else []
    return [entry for entry in entries if isinstance(entry, dict)]


def _source_code_qa_effort_seed_terms(pm_team: str) -> list[str]:
    team = str(pm_team or "").upper()
    terms: list[str] = []
    profiles = _load_source_code_qa_domain_profile_config()
    profile = profiles.get(team) if isinstance(profiles, dict) else {}
    if isinstance(profile, dict):
        for key in ("data_carriers", "source_terms", "api_terms", "config_terms", "logic_terms", "field_population_terms"):
            terms.extend(str(item) for item in (profile.get(key) or []) if item)
    knowledge = _load_source_code_qa_domain_knowledge_config()
    domain = ((knowledge.get("domains") or {}).get(team) or {}) if isinstance(knowledge, dict) else {}
    if isinstance(domain, dict):
        for module in domain.get("module_map") or []:
            if not isinstance(module, dict):
                continue
            terms.append(str(module.get("name") or ""))
            terms.extend(str(item) for item in (module.get("aliases") or []) if item)
            terms.extend(str(item) for item in (module.get("code_hints") or []) if item)
        for term in domain.get("terminology") or []:
            if not isinstance(term, dict):
                continue
            terms.append(str(term.get("term") or ""))
            terms.extend(str(item) for item in (term.get("aliases") or []) if item)
            terms.extend(str(item) for item in (term.get("code_terms") or []) if item)
        retrieval_terms = domain.get("retrieval_terms") if isinstance(domain.get("retrieval_terms"), dict) else {}
        for values in retrieval_terms.values():
            terms.extend(str(item) for item in (values or []) if item)
    return [str(item) for item in _source_code_qa_effort_unique([item for item in terms if str(item or "").strip()])]


def _source_code_qa_effort_entry_applies(entry: dict[str, Any], *, country: str, requirement: str) -> bool:
    countries = [str(item).upper() for item in (entry.get("country_terms") or []) if item]
    if countries and str(country or "").upper() not in countries:
        return False
    aliases = [str(item) for item in (entry.get("business_aliases") or []) if str(item or "").strip()]
    technical_terms = [str(item) for item in (entry.get("technical_terms") or []) if str(item or "").strip()]
    haystack = str(requirement or "").lower()
    for value in [*aliases, *technical_terms]:
        normalized = value.lower().strip()
        if not normalized:
            continue
        if normalized in haystack:
            return True
    return False


def _source_code_qa_effort_group_typed_candidates(entries: list[dict[str, Any]], *, seed_terms: list[str]) -> dict[str, list[str]]:
    grouped: dict[str, list[str]] = {
        "backend_service": [],
        "frontend_surface": [],
        "table_or_config": [],
        "workflow_rule": [],
        "downstream_reporting": [],
    }
    for entry in entries:
        terms = [str(item) for item in (entry.get("technical_terms") or []) if item]
        for surface in entry.get("surfaces") or []:
            surface_key = str(surface or "").strip()
            if surface_key in grouped:
                grouped[surface_key].extend(terms)
    if seed_terms:
        grouped["backend_service"].extend(seed_terms[:30])
    return {key: _source_code_qa_effort_unique(values)[:60] for key, values in grouped.items()}


def _build_source_code_qa_effort_business_plan(
    *,
    pm_team: str,
    country: str,
    language: str,
    requirement: str,
) -> dict[str, Any]:
    raw_requirement = str(requirement or "").strip()
    sentences = _source_code_qa_effort_sentences(raw_requirement)
    option_matches = list(
        re.finditer(
            r"(方案\s*[一二12]|option\s*[12])\s*[:：]?\s*(.*?)(?=(?:方案\s*[一二12]|option\s*[12])\s*[:：]|$)",
            raw_requirement,
            flags=re.IGNORECASE | re.DOTALL,
        )
    )
    options = []
    for index, match in enumerate(option_matches, start=1):
        body = re.sub(r"\s+", " ", match.group(2).strip())
        if body:
            options.append({"id": f"option_{index}", "label": match.group(1).strip(), "summary": body[:1200]})
    has_explicit_options = bool(options)
    if not options:
        options.append({"id": "option_1", "label": "single proposed change", "summary": raw_requirement[:1200]})

    user_segments = []
    if _source_code_qa_effort_matches(raw_requirement, [r"高收入", r"annual\s*income", r"income\s*[><=]", r"120k"]):
        user_segments.append("high income customers")
    if _source_code_qa_effort_matches(raw_requirement, [r"好用户", r"good\s+customer", r"premium"]):
        user_segments.append("qualified or good customers")

    products = []
    if _source_code_qa_effort_matches(raw_requirement, [r"信用卡", r"credit\s*card"]):
        products.append("credit card")
    if _source_code_qa_effort_matches(raw_requirement, [r"现金分期", r"cash\s*installment", r"cash\s*instalment"]):
        products.append("cash installment")
    if _source_code_qa_effort_matches(raw_requirement, [r"cashline", r"cash\s*line"]):
        products.append("cashline")

    limit_types = []
    if _source_code_qa_effort_matches(raw_requirement, [r"额度", r"limit"]):
        limit_types.append("limit amount")
    if _source_code_qa_effort_matches(raw_requirement, [r"信用卡.*额度", r"credit\s*card.*limit"]):
        limit_types.append("credit card limit")
    if _source_code_qa_effort_matches(raw_requirement, [r"现金分期.*(专项)?额度", r"cash\s*installment.*limit"]):
        limit_types.append("cash installment dedicated limit")
    if _source_code_qa_effort_matches(raw_requirement, [r"103", r"104", r"sub\s*product", r"子产品"]):
        limit_types.append("sub-product limit 103/104")

    flow_changes = []
    if _source_code_qa_effort_matches(raw_requirement, [r"申请", r"apply", r"application"]):
        flow_changes.append("application flow")
    if _source_code_qa_effort_matches(raw_requirement, [r"报送", r"submission", r"reporting"]):
        flow_changes.append("reporting or downstream submission")
    if _source_code_qa_effort_matches(raw_requirement, [r"用户教育", r"感知", r"引导", r"education", r"guide"]):
        flow_changes.append("user education or guidance")

    decision_points = [
        sentence[:800]
        for sentence in sentences
        if _source_code_qa_effort_matches(sentence, [r"核心问题", r"是否", r"是不是", r"可以讨论", r"确认", r"how", r"whether"])
    ]
    goals = []
    if _source_code_qa_effort_matches(raw_requirement, [r"策略区分", r"区分", r"separate", r"differentiat"]):
        goals.append("differentiate limit strategy by customer/product context")
    if _source_code_qa_effort_matches(raw_requirement, [r"感知", r"转化", r"教育", r"conversion"]):
        goals.append("make the limit or product path understandable to customers")
    if not goals:
        goals.append("assess technical impact for the requested business change")

    return {
        "raw_requirement": raw_requirement,
        "pm_team": pm_team,
        "country": country,
        "language": language,
        "business_goals": _source_code_qa_effort_unique(goals),
        "options": options,
        "has_explicit_options": has_explicit_options,
        "user_segments": _source_code_qa_effort_unique(user_segments),
        "products": _source_code_qa_effort_unique(products),
        "limit_types": _source_code_qa_effort_unique(limit_types),
        "flow_changes": _source_code_qa_effort_unique(flow_changes),
        "decision_points": _source_code_qa_effort_unique(decision_points)[:8],
    }


def _build_source_code_qa_effort_technical_candidates(
    *,
    pm_team: str,
    country: str,
    business_plan: dict[str, Any],
    requirement: str,
) -> dict[str, Any]:
    raw_requirement = str(requirement or "")
    seed_terms = _source_code_qa_effort_seed_terms(pm_team)
    domain_entries = _source_code_qa_effort_domain_entries(pm_team)
    matched_entries = [
        entry for entry in domain_entries
        if _source_code_qa_effort_entry_applies(entry, country=country, requirement=raw_requirement)
    ]
    terms = [
        "limit",
        "credit limit",
        "product limit",
        "sub product limit",
        "API",
        "config",
        "strategy",
        "workflow",
        "front end screen",
        "application flow",
    ]
    backend_surfaces = ["API validation", "service strategy", "workflow decision rule", "config lookup"]
    frontend_surfaces = ["limit display", "application entry", "customer guidance copy"]
    configs_or_tables = ["limitAmount", "productCode", "productType", "subProductCode"]
    product_terms = []
    limit_terms = []
    evidence_hints = []
    domain_notes = []
    for entry in matched_entries:
        terms.extend(str(item) for item in (entry.get("technical_terms") or []) if item)
        product_terms.extend(str(item) for item in (entry.get("product_terms") or []) if item)
        limit_terms.extend(str(item) for item in (entry.get("limit_terms") or []) if item)
        evidence_hints.extend(str(item) for item in (entry.get("evidence_hints") or []) if item)
        for surface in entry.get("surfaces") or []:
            surface_value = str(surface or "")
            if surface_value == "backend_service":
                backend_surfaces.append(str(entry.get("id") or "backend service impact"))
            elif surface_value == "frontend_surface":
                frontend_surfaces.append(str(entry.get("id") or "frontend impact"))
            elif surface_value in {"table_or_config", "downstream_reporting"}:
                configs_or_tables.extend(str(item) for item in (entry.get("technical_terms") or []) if item)
    terms.extend(seed_terms[:80])

    if str(pm_team or "").upper() == "CRMS":
        backend_surfaces.extend(
            [
                "CRMS underwriting and eligibility decision",
                "borrower/product/sub-product limit calculation",
                "cash installment limit strategy",
                "credit card daily consumption limit strategy",
                "cashline application or redirect flow",
                "downstream reporting payload",
            ]
        )
        frontend_surfaces.extend(
            [
                "credit card and cash installment limit display",
                "cashline application entry",
                "limit explanation and customer education",
            ]
        )
        domain_notes.append("CRMS dictionary v1: income-based credit, cash installment, cashline, and product/sub-product limits.")

    for key in ("products", "limit_types", "flow_changes"):
        for value in business_plan.get(key) or []:
            terms.append(str(value))
    if _source_code_qa_effort_matches(raw_requirement, [r"报送", r"submission", r"reporting"]):
        terms.extend(["reporting", "submission", "report payload"])
    if _source_code_qa_effort_matches(raw_requirement, [r"前端", r"展示", r"入口", r"引导", r"screen", r"display", r"entry"]):
        terms.extend(["screen", "display", "entry point", "guide copy"])

    return {
        "pm_team": pm_team,
        "country": country,
        "search_terms": _source_code_qa_effort_unique(terms)[:80],
        "backend_surfaces": _source_code_qa_effort_unique(backend_surfaces)[:40],
        "frontend_surfaces": _source_code_qa_effort_unique(frontend_surfaces)[:30],
        "configs_or_tables": _source_code_qa_effort_unique(configs_or_tables)[:50],
        "product_terms": _source_code_qa_effort_unique(product_terms)[:40],
        "limit_terms": _source_code_qa_effort_unique(limit_terms)[:40],
        "evidence_hints": _source_code_qa_effort_unique(evidence_hints)[:40],
        "matched_dictionary_entries": [str(entry.get("id") or "") for entry in matched_entries if entry.get("id")],
        "typed_candidates": _source_code_qa_effort_group_typed_candidates(matched_entries, seed_terms=seed_terms),
        "domain_notes": domain_notes,
    }


def _build_source_code_qa_effort_estimation_rubric(
    *,
    business_plan: dict[str, Any],
    technical_candidates: dict[str, Any],
) -> dict[str, Any]:
    text = " ".join(
        [
            " ".join(str(item) for item in business_plan.get("products") or []),
            " ".join(str(item) for item in business_plan.get("limit_types") or []),
            " ".join(str(item) for item in business_plan.get("flow_changes") or []),
            " ".join(str(item) for item in technical_candidates.get("backend_surfaces") or []),
            " ".join(str(item) for item in technical_candidates.get("frontend_surfaces") or []),
        ]
    )
    high_complexity = _source_code_qa_effort_matches(
        text,
        [r"underwriting", r"borrower", r"sub[-\s]?product", r"reporting", r"submission", r"授信", r"额度模型"],
    )
    medium_complexity = _source_code_qa_effort_matches(text, [r"api", r"service", r"strategy", r"workflow", r"limit"])
    frontend_required = bool(technical_candidates.get("frontend_surfaces"))
    option_estimates = []
    for index, option in enumerate(business_plan.get("options") or [], start=1):
        option_text = str(option.get("summary") or "")
        option_high = high_complexity or _source_code_qa_effort_matches(option_text, [r"cashline", r"独立", r"报送", r"模型", r"多产品"])
        option_medium = medium_complexity or _source_code_qa_effort_matches(option_text, [r"额度", r"limit", r"策略", r"rule"])
        be_range = "8-15 PD" if option_high else ("3-6 PD" if option_medium else "1-3 PD")
        fe_range = "3-6 PD" if _source_code_qa_effort_matches(option_text, [r"用户教育", r"感知", r"入口", r"展示", r"guide", r"display"]) else ("1-3 PD" if frontend_required else "0-1 PD")
        option_estimates.append(
            {
                "id": option.get("id") or f"option_{index}",
                "label": option.get("label") or f"Option {index}",
                "be_person_days": be_range,
                "fe_person_days": fe_range,
                "basis": "planning-grade estimate before Dev final sizing",
            }
        )
    return {
        "rules": [
            "Config or rule parameter only: low complexity.",
            "BE API plus service, strategy, or limit-flow change: medium complexity.",
            "Underwriting engine, limit model, reporting, or multi-product limit linkage: high complexity.",
            "FE display, guidance, application entry, and customer education are estimated separately.",
            "Final answer must separate confirmed evidence, inferred impact, and missing evidence.",
        ],
        "complexity_drivers": {
            "backend": "high" if high_complexity else ("medium" if medium_complexity else "low"),
            "frontend": "medium" if frontend_required else "low",
        },
        "option_estimates": option_estimates,
    }


def _source_code_qa_effort_json_block(value: dict[str, Any]) -> str:
    return json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True)[:12000]


def _build_source_code_qa_effort_assessment_prompt(
    *,
    pm_team: str,
    country: str,
    language: str,
    requirement: str,
    llm_provider: str,
    runtime_evidence: list[dict[str, Any]],
    business_plan: dict[str, Any],
    technical_candidates: dict[str, Any],
    estimation_rubric: dict[str, Any],
) -> str:
    team_label = str((TEAM_PROFILE_DEFAULTS.get(pm_team) or {}).get("label") or pm_team or "Selected PM Team").strip()
    output_language = "Chinese" if language == "zh" else "English"
    runtime_items = [
        item for item in runtime_evidence
        if isinstance(item, dict)
    ]
    runtime_summary = ", ".join(
        sorted(
            {
                f"{item.get('pm_team') or pm_team}:{item.get('country') or country}:{item.get('source_type') or 'runtime'}"
                for item in runtime_items
            }
        )
    ) or "none"
    raw_requirement = str(requirement or "").strip()[:8000]
    has_explicit_options = bool(business_plan.get("has_explicit_options"))
    technical_change_section = "方案 1/2 技术改造点" if has_explicit_options else "技术改造点"
    option_rule = (
        "- The requirement contains explicit alternatives; keep the original option labels and compare each option separately."
        if has_explicit_options
        else "- The requirement does not contain explicit Option 1/Option 2 alternatives; do not invent option labels. Use 'proposed change' instead."
    )
    return "\n".join(
        [
            "You are performing a Source Code Q&A Effort Assessment for a new business requirement.",
            "",
            "Context:",
            f"- PM Team: {pm_team} ({team_label})",
            f"- Country: {country}",
            f"- Answer language: {output_language}",
            f"- Selected model provider: {llm_provider or 'default'}",
            f"- Runtime evidence available: {len(runtime_items)} item(s): {runtime_summary}",
            "",
            "Original business requirement, verbatim:",
            raw_requirement,
            "",
            "Business plan extracted from the requirement:",
            _source_code_qa_effort_json_block(business_plan),
            "",
            "Technical candidates for repository evidence search:",
            _source_code_qa_effort_json_block(technical_candidates),
            "",
            "Estimation rubric:",
            _source_code_qa_effort_json_block(estimation_rubric),
            "",
            "Optimized assessment task:",
            "- Use the business plan and technical candidates as the focused search map. Do not rely only on the original business wording.",
            "- Map the requirement to likely impacted repositories, modules, files, APIs, tables, configs, scheduled jobs, front-end screens and components, and tests.",
            "- Use current source-code evidence as the primary basis for implementation impact and person-day estimates.",
            "- If exact table or path lookup misses, record it as a warning and continue with focused technical-candidate search.",
            "- Use runtime evidence only as supporting context. Treat uploaded DB, Apollo, and log evidence separately from source-code proof and cite the evidence type distinctly.",
            "- Separate confirmed code evidence from assumptions, inferred impact, and missing evidence. If a required evidence link is missing, say so explicitly instead of guessing.",
            "- Estimate BE and FE work as ranges in person-days. Use 0 person-days if no FE or BE change is found, but explain why.",
            "- Include QA/test and integration impact in the relevant BE/FE estimate notes instead of creating a third estimate bucket.",
            "",
            "Required output sections:",
            "1. 业务理解",
            f"2. {technical_change_section}",
            "3. BE 人天",
            "4. FE 人天",
            "5. Confirmed / Inferred / Missing Evidence",
            "6. Assumptions / Risks",
            "7. Confirmation Questions",
            "8. Source / Runtime Evidence",
            "",
            "Output rules:",
            f"- Write the final answer in {output_language}.",
            option_rule,
            "- Keep the answer concise but specific enough for PM and engineering planning.",
            "- Cite concrete file paths, classes, functions, APIs, tables, configs, tests, or runtime evidence filenames when available.",
            "- Person-day estimates must be ranges such as 1-2 PD or 3-5 PD, with one sentence explaining the driver for each range.",
            "- If source evidence is weak, still estimate with low confidence and explain missing evidence.",
        ]
    )


def _source_code_qa_effort_missing_evidence(
    *,
    result: dict[str, Any],
    technical_candidates: dict[str, Any],
) -> list[str]:
    missing: list[str] = []
    matches = result.get("matches") if isinstance(result, dict) else []
    if not isinstance(matches, list) or not matches:
        missing.append("No confirmed source-code references were found for the technical candidates.")
    exact_lookup = result.get("exact_lookup") if isinstance(result, dict) else None
    if isinstance(exact_lookup, dict) and exact_lookup.get("terms") and not exact_lookup.get("matched_terms"):
        missing.append("Exact table/path lookup did not match; assessment continued with focused candidate search.")
    if technical_candidates.get("backend_surfaces") and (not isinstance(matches, list) or not matches):
        missing.append("Backend impact surfaces need Dev confirmation against current repositories.")
    return _source_code_qa_effort_unique(missing)


def _source_code_qa_effort_confidence(result: dict[str, Any], missing_evidence: list[str]) -> str:
    matches = result.get("matches") if isinstance(result, dict) else []
    if str(result.get("status") or "").lower() == "no_match":
        return "low"
    if isinstance(matches, list) and len(matches) >= 3 and not missing_evidence:
        return "high"
    if isinstance(matches, list) and matches:
        return "medium"
    return "low"


def _source_code_qa_effort_fallback_answer(
    *,
    language: str,
    business_plan: dict[str, Any],
    technical_candidates: dict[str, Any],
    estimation_rubric: dict[str, Any],
    missing_evidence: list[str],
) -> str:
    options = estimation_rubric.get("option_estimates") or []
    has_explicit_options = bool(business_plan.get("has_explicit_options"))
    option_lines = "\n".join(
        f"- {item.get('label')}: BE {item.get('be_person_days')}, FE {item.get('fe_person_days')} ({item.get('basis')})"
        for item in options
        if isinstance(item, dict)
    )
    candidate_terms = ", ".join(str(item) for item in (technical_candidates.get("search_terms") or [])[:12])
    missing_lines = "\n".join(f"- {item}" for item in missing_evidence) or "- More code evidence is required."
    if language == "zh":
        goals = ", ".join(str(item) for item in business_plan.get("business_goals") or [])
        technical_title = "方案 1/2 技术改造点" if has_explicit_options else "技术改造点"
        confirmation_questions = [
            "- 额度策略是否只改参数，还是需要新增产品/子产品额度模型?",
            "- 是否需要前端新增 cashline 申请入口或额度解释文案?",
        ]
        if has_explicit_options:
            confirmation_questions.insert(0, "- 方案 1 和方案 2 是否二选一，还是都需要落地?")
        return "\n".join(
            [
                "业务理解",
                f"- 目标: {goals or '评估业务需求对应的技术改造范围'}",
                "",
                technical_title,
                "- Confirmed: 当前没有足够 source-code 引用能确认具体文件。",
                f"- Inferred: 需要围绕这些技术候选词继续确认影响面: {candidate_terms}",
                "- Missing: 需要开发确认额度策略、申请流程、前端展示、报送或下游接口是否在当前 repo 覆盖。",
                "",
                "BE 人天 / FE 人天",
                option_lines or "- 单方案: BE 3-6 PD, FE 1-3 PD，低置信度。",
                "",
                "Confirmed / Inferred / Missing Evidence",
                missing_lines,
                "",
                "Assumptions / Risks",
                "- 这是 planning-grade 低置信度估算，不替代 Dev final sizing。",
                "- 如果涉及授信引擎、额度模型、报送或多产品额度联动，BE 复杂度应按高复杂度处理。",
                "",
                "Confirmation Questions",
                *confirmation_questions,
            ]
        )
    technical_title = "Option Technical Changes" if has_explicit_options else "Technical Changes"
    confirmation_questions = [
        "- Is the requested limit change a config-only rule update or a new limit model?",
        "- Does the change require FE display, application entry, or customer education copy?",
    ]
    if has_explicit_options:
        confirmation_questions.insert(0, "- Are the listed options alternatives, or should more than one be implemented?")
    return "\n".join(
        [
            "Business Understanding",
            f"- Goals: {', '.join(str(item) for item in business_plan.get('business_goals') or [])}",
            "",
            technical_title,
            "- Confirmed: No concrete source-code references were found.",
            f"- Inferred: Continue validation around these technical candidates: {candidate_terms}",
            "- Missing: Dev confirmation is required for limit strategy, application flow, FE display, and downstream reporting impact.",
            "",
            "BE / FE Person-days",
            option_lines or "- Single option: BE 3-6 PD, FE 1-3 PD, low confidence.",
            "",
            "Confirmed / Inferred / Missing Evidence",
            missing_lines,
            "",
            "Assumptions / Risks",
            "- This is a planning-grade low-confidence estimate and does not replace Dev final sizing.",
            "",
            "Confirmation Questions",
            *confirmation_questions,
        ]
    )


def _build_source_code_qa_effort_structured_assessment(
    *,
    result: dict[str, Any],
    language: str,
    business_plan: dict[str, Any],
    technical_candidates: dict[str, Any],
    estimation_rubric: dict[str, Any],
    missing_evidence: list[str],
    confidence: str,
) -> dict[str, Any]:
    matches = result.get("matches") if isinstance(result.get("matches"), list) else []
    confirmed_evidence = [
        {
            "repo": str(match.get("repo") or ""),
            "path": str(match.get("path") or ""),
            "line_start": match.get("line_start") or 0,
            "line_end": match.get("line_end") or 0,
        }
        for match in matches[:8]
        if isinstance(match, dict)
    ]
    typed_candidates = technical_candidates.get("typed_candidates") if isinstance(technical_candidates.get("typed_candidates"), dict) else {}
    inferred_impact = [
        {
            "surface": surface,
            "terms": [str(item) for item in (terms or [])[:12]],
        }
        for surface, terms in typed_candidates.items()
        if terms
    ]
    return {
        "version": 1,
        "language": language,
        "confidence": confidence,
        "business_understanding": {
            "goals": business_plan.get("business_goals") or [],
            "user_segments": business_plan.get("user_segments") or [],
            "products": business_plan.get("products") or [],
            "limit_types": business_plan.get("limit_types") or [],
            "flow_changes": business_plan.get("flow_changes") or [],
            "decision_points": business_plan.get("decision_points") or [],
        },
        "option_impacts": [
            {
                "id": item.get("id") or "",
                "label": item.get("label") or "",
                "summary": item.get("summary") or "",
            }
            for item in (business_plan.get("options") or [])
            if isinstance(item, dict)
        ],
        "be_estimate": [
            {
                "option_id": item.get("id") or "",
                "person_days": item.get("be_person_days") or "",
                "basis": item.get("basis") or "",
            }
            for item in (estimation_rubric.get("option_estimates") or [])
            if isinstance(item, dict)
        ],
        "fe_estimate": [
            {
                "option_id": item.get("id") or "",
                "person_days": item.get("fe_person_days") or "",
                "basis": item.get("basis") or "",
            }
            for item in (estimation_rubric.get("option_estimates") or [])
            if isinstance(item, dict)
        ],
        "confirmed_evidence": confirmed_evidence,
        "inferred_impact": inferred_impact,
        "missing_evidence": missing_evidence,
        "questions": (
            ["Are the listed options alternatives, or should more than one be implemented?"]
            if business_plan.get("has_explicit_options")
            else []
        ) + [
            "Is the requested limit change a config-only rule update or a new limit model?",
            "Does the change require FE display, application entry, or customer education copy?",
        ],
        "dictionary_entries": technical_candidates.get("matched_dictionary_entries") or [],
    }


def _normalize_source_code_qa_effort_assessment_result(
    *,
    result: dict[str, Any],
    language: str,
    business_plan: dict[str, Any],
    technical_candidates: dict[str, Any],
    estimation_rubric: dict[str, Any],
) -> dict[str, Any]:
    normalized = dict(result or {})
    missing_evidence = _source_code_qa_effort_missing_evidence(
        result=normalized,
        technical_candidates=technical_candidates,
    )
    confidence = _source_code_qa_effort_confidence(normalized, missing_evidence)
    if str(normalized.get("status") or "").lower() == "no_match":
        normalized["status"] = "ok"
        normalized["effort_evidence_status"] = "warning"
        normalized["summary"] = "Effort assessment completed with low confidence because source-code evidence is missing."
        normalized["llm_answer"] = _source_code_qa_effort_fallback_answer(
            language=language,
            business_plan=business_plan,
            technical_candidates=technical_candidates,
            estimation_rubric=estimation_rubric,
            missing_evidence=missing_evidence,
        )
    else:
        normalized["effort_evidence_status"] = "warning" if missing_evidence else "confirmed"
        if not normalized.get("summary"):
            normalized["summary"] = "Effort assessment completed."
    normalized["assessment_confidence"] = confidence
    normalized["missing_evidence"] = missing_evidence
    normalized["structured_assessment"] = _build_source_code_qa_effort_structured_assessment(
        result=normalized,
        language=language,
        business_plan=business_plan,
        technical_candidates=technical_candidates,
        estimation_rubric=estimation_rubric,
        missing_evidence=missing_evidence,
        confidence=confidence,
    )
    return normalized


def _compact_source_code_qa_session_payload(result: dict[str, Any]) -> dict[str, Any]:
    structured = result.get("structured_answer") if isinstance(result.get("structured_answer"), dict) else {}
    answer_claim_check = result.get("answer_claim_check") if isinstance(result.get("answer_claim_check"), dict) else {}
    llm_route = result.get("llm_route") if isinstance(result.get("llm_route"), dict) else {}
    codex_trace = result.get("codex_cli_trace") if isinstance(result.get("codex_cli_trace"), dict) else {}
    return {
        "status": result.get("status") or "",
        "trace_id": result.get("trace_id") or "",
        "query_mode": result.get("query_mode") or "",
        "deadline_seconds": result.get("deadline_seconds") or 0,
        "deadline_hit": bool(result.get("deadline_hit")),
        "fallback_used": bool(result.get("fallback_used")),
        "fallback_answer_quality": result.get("fallback_answer_quality") or "",
        "fallback_evidence_count": result.get("fallback_evidence_count") or 0,
        "fallback_claim_count": result.get("fallback_claim_count") or 0,
        "deadline_fallback_reason": result.get("deadline_fallback_reason") or "",
        "summary": result.get("summary") or "",
        "llm_answer": result.get("llm_answer") or "",
        "llm_provider": result.get("llm_provider") or "",
        "llm_model": result.get("llm_model") or "",
        "llm_route": {
            "mode": llm_route.get("mode") or "",
            "query_mode": llm_route.get("query_mode") or "",
            "provider": llm_route.get("provider") or "",
            "prompt_mode": llm_route.get("prompt_mode") or "",
            "candidate_paths": (llm_route.get("candidate_paths") or [])[:30],
            "candidate_path_layers": llm_route.get("candidate_path_layers") or {},
            "codex_session_max_turns": llm_route.get("codex_session_max_turns") or 8,
        },
        "structured_answer": {
            "direct_answer": structured.get("direct_answer") or "",
            "claims": (structured.get("claims") or [])[:8],
            "confirmed_points": (structured.get("confirmed_points") or [])[:8],
            "missing_points": (structured.get("missing_points") or [])[:8],
            "evidence_cards": (structured.get("evidence_cards") or [])[:8],
            "citations": (structured.get("citations") or [])[:12],
            "missing_evidence": (structured.get("missing_evidence") or [])[:8],
            "confidence": structured.get("confidence") or "",
        },
        "answer_contract": result.get("answer_contract") or {},
        "answer_quality": result.get("answer_quality") or {},
        "codex_cli_summary": result.get("codex_cli_summary") or {},
        "codex_cli_trace": {
            "session_mode": codex_trace.get("session_mode") or "",
            "command_mode": codex_trace.get("command_mode") or "",
            "session_id": codex_trace.get("session_id") or "",
            "exit_code": codex_trace.get("exit_code"),
            "latency_ms": codex_trace.get("latency_ms"),
            "timeout": bool(codex_trace.get("timeout")),
            "stream_messages": (codex_trace.get("stream_messages") or [])[-20:],
            "command_summaries": (codex_trace.get("command_summaries") or [])[-12:],
            "probable_inspected_files": (codex_trace.get("probable_inspected_files") or [])[-20:],
        },
        "codex_citation_validation": answer_claim_check.get("codex_citation_validation") or {},
        "attachments": _source_code_qa_public_attachments(result.get("attachments") if isinstance(result.get("attachments"), list) else []),
        "runtime_evidence": _source_code_qa_public_runtime_evidence(
            result.get("runtime_evidence") if isinstance(result.get("runtime_evidence"), list) else []
        ),
        "matches": [
            {
                "repo": match.get("repo"),
                "path": match.get("path"),
                "line_start": match.get("line_start"),
                "line_end": match.get("line_end"),
                "retrieval": match.get("retrieval"),
                "trace_stage": match.get("trace_stage"),
                "reason": match.get("reason"),
                "score": match.get("score"),
            }
            for match in (result.get("matches") or [])[:10]
            if isinstance(match, dict)
        ],
    }


def _build_source_code_qa_session_context(result: dict[str, Any], request_payload: dict[str, Any]) -> dict[str, Any]:
    compact = _compact_source_code_qa_session_payload(result)
    matches = compact.get("matches") or []
    codex_trace = compact.get("codex_cli_trace") if isinstance(compact.get("codex_cli_trace"), dict) else {}
    candidate_paths = (compact.get("llm_route") or {}).get("candidate_paths") or []
    inspected_paths: list[dict[str, Any]] = []
    for raw_path in codex_trace.get("probable_inspected_files") or []:
        raw_text = str(raw_path or "")
        matched = None
        for candidate in candidate_paths:
            if isinstance(candidate, dict) and str(candidate.get("path") or "") and str(candidate.get("path") or "") in raw_text:
                matched = candidate
                break
        if matched:
            inspected_paths.append({**matched, "source": "codex_cli_trace"})
    if not inspected_paths:
        inspected_paths = [
            item for item in candidate_paths[:5]
            if isinstance(item, dict) and str(item.get("trace_stage") or "") == "followup_memory"
        ]
    session_id = str(codex_trace.get("session_id") or "").strip()
    return {
        "key": f"{request_payload.get('pm_team') or ''}:{request_payload.get('country') or ALL_COUNTRY}",
        "pm_team": request_payload.get("pm_team") or "",
        "country": request_payload.get("country") or ALL_COUNTRY,
        "question": request_payload.get("question") or "",
        "trace_id": compact.get("trace_id") or "",
        "summary": compact.get("summary") or "",
        "answer": compact.get("llm_answer") or "",
        "rendered_answer": compact.get("llm_answer") or "",
        "attachments": compact.get("attachments") or [],
        "llm_provider": compact.get("llm_provider") or "",
        "llm_model": compact.get("llm_model") or "",
        "llm_route": compact.get("llm_route") or {},
        "codex_session_max_turns": (compact.get("llm_route") or {}).get("codex_session_max_turns") or 8,
        "codex_cli_summary": compact.get("codex_cli_summary") or {},
        "codex_cli_trace": codex_trace,
        "codex_cli_session": {
            "session_id": session_id,
            "mode": codex_trace.get("session_mode") or "",
            "last_used_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        } if session_id else {},
        "codex_inspected_paths": inspected_paths[:12],
        "codex_citation_validation": compact.get("codex_citation_validation") or {},
        "codex_candidate_paths": candidate_paths,
        "repo_scope": list(dict.fromkeys([match.get("repo") for match in matches if match.get("repo")]))[:8],
        "matches": matches[:8],
        "matches_snapshot": matches[:10],
        "trace_paths": (result.get("trace_paths") or [])[:5],
        "query_mode": compact.get("query_mode") or "",
        "deadline_seconds": compact.get("deadline_seconds") or 0,
        "deadline_hit": bool(compact.get("deadline_hit")),
        "fallback_used": bool(compact.get("fallback_used")),
        "fallback_answer_quality": compact.get("fallback_answer_quality") or "",
        "fallback_evidence_count": compact.get("fallback_evidence_count") or 0,
        "fallback_claim_count": compact.get("fallback_claim_count") or 0,
        "deadline_fallback_reason": compact.get("deadline_fallback_reason") or "",
        "structured_answer": compact.get("structured_answer") or {},
        "answer_contract": compact.get("answer_contract") or {},
        "evidence_pack": result.get("evidence_pack") or {},
        "answer_quality": compact.get("answer_quality") or {},
    }


def _read_json_file(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _callable_accepts_keyword(func: Any, keyword: str) -> bool:
    try:
        signature = inspect.signature(func)
    except (TypeError, ValueError):
        return True
    return any(
        parameter.kind == inspect.Parameter.VAR_KEYWORD or name == keyword
        for name, parameter in signature.parameters.items()
    )


def _source_code_qa_release_gate_payload(settings: Settings) -> dict[str, Any]:
    data_root = settings.team_portal_data_dir
    if not data_root.is_absolute():
        data_root = (PROJECT_ROOT / data_root).resolve()
    gate = _read_json_file(data_root / "run" / "source_code_qa_release_gate.json")
    eval_status = _read_json_file(data_root / "run" / "source_code_qa_eval_status.json")
    latest_eval = _read_json_file(data_root / "source_code_qa" / "eval_runs" / "latest.json")
    status = str(gate.get("status") or eval_status.get("state") or latest_eval.get("status") or "missing")
    updated_at = gate.get("timestamp") or latest_eval.get("timestamp") or eval_status.get("updated_at")
    return {
        "status": status,
        "updated_at": updated_at,
        "summary": gate.get("summary") or eval_status.get("message") or "",
        "thresholds": gate.get("thresholds") or {},
        "checks": gate.get("checks") or {},
        "latest_eval": {
            "status": latest_eval.get("status"),
            "eval": latest_eval.get("eval") or {},
            "llm_smoke": latest_eval.get("llm_smoke") or {},
            "report_path": latest_eval.get("report_path"),
        },
    }


def _google_credentials_have_scopes(*required_scopes: str) -> bool:
    payload = dict(session.get("google_credentials") or {})
    scopes = payload.get("scopes") or []
    normalized_scopes = {str(scope).strip() for scope in scopes if str(scope).strip()}
    return all(scope in normalized_scopes for scope in required_scopes)


def _google_session_can_call_live_google_apis() -> bool:
    payload = dict(session.get("google_credentials") or {})
    token = str(payload.get("token") or "").strip()
    if not token:
        return False
    refresh_fields = (
        "refresh_token",
        "token_uri",
        "client_id",
        "client_secret",
    )
    return all(str(payload.get(field) or "").strip() for field in refresh_fields)


def _require_google_login(settings: Settings, *, api: bool = False):
    if _site_requires_google_login(settings):
        if not _google_session_is_connected():
            message = "Sign in with your NPT Google account before using the shared portal."
            if api:
                return jsonify({"status": "error", "message": message}), HTTPStatus.UNAUTHORIZED
            return redirect(url_for("index"))
        if _current_google_user_is_blocked(settings):
            message = "This Google account is not authorized for the team portal."
            if api:
                return jsonify({"status": "error", "message": message}), HTTPStatus.FORBIDDEN
            flash(message, "error")
            return redirect(url_for("access_denied"))
    return None


def _require_gmail_seatalk_demo_access(settings: Settings, *, api: bool = False):
    login_gate = _require_google_login(settings, api=api)
    if login_gate is not None:
        return login_gate
    message = f"SeaTalk Management is restricted to {PORTAL_ADMIN_EMAIL}."
    if not _can_access_gmail_seatalk_demo(settings):
        if api:
            return jsonify({"status": "error", "message": message}), HTTPStatus.FORBIDDEN
        flash(message, "error")
        return redirect(url_for("index"))
    return None


def _require_meeting_recorder_access(settings: Settings, *, api: bool = False):
    login_gate = _require_google_login(settings, api=api)
    if login_gate is not None:
        return login_gate
    message = f"Meeting Recorder is restricted to {PORTAL_ADMIN_EMAIL}."
    if not _can_access_meeting_recorder(settings):
        if api:
            return jsonify({"status": "error", "message": message}), HTTPStatus.FORBIDDEN
        flash(message, "error")
        return redirect(url_for("index"))
    return None


def _require_source_code_qa_access(settings: Settings, *, api: bool = False):
    login_gate = _require_google_login(settings, api=api)
    if login_gate is not None:
        return login_gate
    message = "Source Code Q&A is available to signed-in @npt.sg users and the configured test account."
    if not _can_access_source_code_qa(settings):
        if api:
            return jsonify({"status": "error", "message": message}), HTTPStatus.FORBIDDEN
        flash(message, "error")
        return redirect(url_for("index"))
    return None


def _require_source_code_qa_manage_access(settings: Settings, *, api: bool = False):
    access_gate = _require_source_code_qa_access(settings, api=api)
    if access_gate is not None:
        return access_gate
    auth_payload = _source_code_qa_auth_payload(settings)
    message = (
        f"Source Code Q&A repository admin is restricted to {PORTAL_ADMIN_EMAIL}. "
        f"Signed in as {auth_payload['signed_in_email'] or 'unknown'}."
    )
    if not _can_manage_source_code_qa(settings):
        if api:
            return jsonify({"status": "error", "message": message, "auth": auth_payload}), HTTPStatus.FORBIDDEN
        flash(message, "error")
        return redirect(url_for("source_code_qa"))
    return None


def _require_work_memory_access(settings: Settings, *, api: bool = False):
    login_gate = _require_google_login(settings, api=api)
    if login_gate is not None:
        return login_gate
    message = f"AI Memory is restricted to {PORTAL_ADMIN_EMAIL}."
    if not _can_access_work_memory(settings):
        if api:
            return jsonify({"status": "error", "message": message}), HTTPStatus.FORBIDDEN
        flash(message, "error")
        return redirect(url_for("index"))
    return None


def _require_team_dashboard_access(settings: Settings, *, api: bool = False):
    login_gate = _require_google_login(settings, api=api)
    if login_gate is not None:
        return login_gate
    user_identity = _get_user_identity(settings)
    message = f"Team Dashboard is restricted to {PORTAL_ADMIN_EMAIL}."
    if not _can_access_team_dashboard(user_identity):
        if api:
            return jsonify({"status": "error", "message": message}), HTTPStatus.FORBIDDEN
        flash(message, "error")
        return redirect(url_for("access_denied"))
    return None


def _require_prd_self_assessment_access(settings: Settings, *, api: bool = False):
    login_gate = _require_google_login(settings, api=api)
    if login_gate is not None:
        return login_gate
    message = "PRD Self-Assessment is available to signed-in npt.sg users and the configured test account."
    if not _can_access_prd_self_assessment(settings):
        if api:
            return jsonify({"status": "error", "message": message}), HTTPStatus.FORBIDDEN
        flash(message, "error")
        return redirect(url_for("access_denied"))
    return None


def _run_prd_self_assessment_action(settings: Settings, *, action: str):
    access_gate = _require_prd_self_assessment_access(settings, api=True)
    if access_gate is not None:
        return access_gate
    payload = request.get_json(silent=True) or {}
    user_identity = _get_user_identity(settings)
    request_payload = {
        "owner_key": str(user_identity.get("config_key") or ""),
        "prd_url": str(payload.get("prd_url") or payload.get("page_ref") or ""),
        "language": str(payload.get("language") or "zh"),
        "force_refresh": bool(payload.get("force_refresh")),
    }
    event_prefix = f"prd_self_assessment_{action}"
    try:
        if _local_agent_source_code_qa_enabled(settings):
            client = _build_local_agent_client(settings)
            data = (
                client.prd_self_assessment_summary(request_payload)
                if action == "summary"
                else client.prd_self_assessment_review(request_payload)
            )
        else:
            service = _build_prd_review_service(settings)
            request_model = PRDBriefingReviewRequest(**request_payload)
            data = service.summarize_url(request_model) if action == "summary" else service.review_url(request_model)
        _log_portal_event(
            f"{event_prefix}_success",
            **_build_request_log_context(
                settings,
                user_identity=user_identity,
                extra={
                    "prd_url_hash": hashlib.sha256(request_payload["prd_url"].encode("utf-8")).hexdigest()[:12],
                    "language": request_payload["language"],
                    "cached": bool(data.get("cached")),
                },
            ),
        )
        return jsonify(data)
    except ToolError as error:
        error_details = _classify_portal_error(error)
        _log_portal_event(
            f"{event_prefix}_tool_error",
            level=logging.WARNING,
            **_build_request_log_context(
                settings,
                user_identity=user_identity,
                extra={**error_details, "language": request_payload["language"]},
            ),
        )
        return jsonify({"status": "error", "message": str(error), **error_details}), HTTPStatus.BAD_REQUEST
    except Exception as error:  # noqa: BLE001
        _log_portal_event(
            f"{event_prefix}_unexpected_error",
            level=logging.ERROR,
            **_build_request_log_context(settings, user_identity=user_identity, extra=_classify_portal_error(error)),
        )
        current_app.logger.exception("PRD Self-Assessment %s failed.", action)
        return (
            jsonify(
                {
                    "status": "error",
                    "message": f"PRD {action} failed unexpectedly. Please retry or share the request ID.",
                    **_classify_portal_error(error),
                }
            ),
            HTTPStatus.INTERNAL_SERVER_ERROR,
        )


def _require_team_dashboard_monthly_report_access(settings: Settings, *, api: bool = False):
    access_gate = _require_team_dashboard_access(settings, api=api)
    if access_gate is not None:
        return access_gate
    user_identity = _get_user_identity(settings)
    if _can_access_team_dashboard_monthly_report(user_identity):
        return None
    message = "Monthly Report is restricted to xiaodong.zheng@npt.sg."
    if api:
        return jsonify({"status": "error", "message": message}), HTTPStatus.FORBIDDEN
    flash(message, "error")
    return redirect(url_for("access_denied"))


def _validate_config_security(settings: Settings, config_data: dict[str, Any]) -> None:
    portal_token = str(config_data.get("bpmis_api_access_token", "") or "").strip()
    if _shared_portal_enabled(settings) and portal_token and not settings.team_portal_config_encryption_key:
        raise ToolError(
            "TEAM_PORTAL_CONFIG_ENCRYPTION_KEY must be configured on the host before saving BPMIS tokens in shared mode."
        )


def _load_effective_team_profiles(config_store: WebConfigStore) -> dict[str, dict[str, Any]]:
    profiles = {team_key: dict(profile) for team_key, profile in TEAM_PROFILE_DEFAULTS.items()}
    stored_profiles = config_store.load_team_profiles()
    if has_app_context():
        settings = current_app.config.get("SETTINGS")
        if isinstance(settings, Settings) and _remote_bpmis_config_enabled(settings):
            stored_profiles = _build_local_agent_client(settings).bpmis_team_profiles_load()
    for team_key, stored_profile in stored_profiles.items():
        if team_key not in profiles:
            continue
        profiles[team_key].update(stored_profile)
    return profiles


def _is_team_profile_admin(user_identity: dict[str, str | None]) -> bool:
    return _is_portal_admin(str(user_identity.get("email") or ""))


def _can_access_team_dashboard(user_identity: dict[str, str | None]) -> bool:
    return _is_portal_admin(str(user_identity.get("email") or ""))


def _can_manage_team_dashboard(user_identity: dict[str, str | None]) -> bool:
    return _is_portal_admin(str(user_identity.get("email") or ""))


def _can_access_team_dashboard_monthly_report(user_identity: dict[str, str | None]) -> bool:
    return _is_portal_admin(str(user_identity.get("email") or ""))


def _can_edit_sync_email(user_identity: dict[str, str | None]) -> bool:
    return _is_portal_admin(str(user_identity.get("email") or ""))


def _apply_sync_email_policy(config_data: dict[str, Any], user_identity: dict[str, str | None]) -> None:
    email = str(user_identity.get("email") or "").strip().lower()
    if email and not _can_edit_sync_email(user_identity):
        config_data["sync_pm_email"] = email


def _hydrate_setup_defaults(
    config_data: dict[str, Any],
    user_identity: dict[str, str | None],
    *,
    team_profiles: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    hydrated = dict(config_data)
    email = str(user_identity.get("email") or "").strip().lower()
    pm_team = str(hydrated.get("pm_team", "") or "").strip().upper()
    effective_team_profiles = team_profiles or TEAM_PROFILE_DEFAULTS
    if pm_team:
        hydrated["pm_team"] = pm_team
    for field, default_value in DEFAULT_DIRECT_VALUES.items():
        if not str(hydrated.get(field, "") or "").strip():
            hydrated[field] = default_value
    profile = effective_team_profiles.get(pm_team)
    if profile:
        if not str(hydrated.get("component_route_rules_text", "") or "").strip():
            hydrated["component_route_rules_text"] = str(profile.get("component_route_rules_text", "") or "")
        if not str(hydrated.get("component_default_rules_text", "") or "").strip():
            hydrated["component_default_rules_text"] = str(profile.get("component_default_rules_text", "") or "")
    existing_need_uat = hydrated.get("need_uat_by_market", {})
    hydrated["need_uat_by_market"] = {
        market: str((existing_need_uat or {}).get(market, "") or DEFAULT_NEED_UAT_BY_MARKET.get(market, "")).strip()
        for market in MARKET_KEYS
    }
    if email:
        for field in ("component_route_rules_text", "component_default_rules_text"):
            value = str(hydrated.get(field, "") or "")
            if TEAM_DEFAULT_EMAIL_PLACEHOLDER in value:
                hydrated[field] = value.replace(TEAM_DEFAULT_EMAIL_PLACEHOLDER, email)
        for field in ("sync_pm_email", "product_manager_value", "reporter_value", "biz_pic_value"):
            if not str(hydrated.get(field, "") or "").strip():
                hydrated[field] = email
    return hydrated


def _build_team_profiles_for_display(
    config_data: dict[str, Any],
    user_identity: dict[str, str | None],
    *,
    team_profiles: dict[str, dict[str, Any]] | None = None,
) -> dict[str, dict[str, Any]]:
    display_email = (
        str(user_identity.get("email") or "").strip().lower()
        or str(config_data.get("sync_pm_email", "") or "").strip().lower()
        or str(config_data.get("product_manager_value", "") or "").strip().lower()
        or str(config_data.get("reporter_value", "") or "").strip().lower()
        or str(config_data.get("biz_pic_value", "") or "").strip().lower()
    )

    profiles: dict[str, dict[str, Any]] = {}
    for team_key, profile in (team_profiles or TEAM_PROFILE_DEFAULTS).items():
        rendered_profile = dict(profile)
        if display_email:
            for field in ("component_route_rules_text", "component_default_rules_text"):
                value = str(rendered_profile.get(field, "") or "")
                if TEAM_DEFAULT_EMAIL_PLACEHOLDER in value:
                    rendered_profile[field] = value.replace(TEAM_DEFAULT_EMAIL_PLACEHOLDER, display_email)
        profiles[team_key] = rendered_profile
    return profiles


def _validate_team_profile_setup(
    config_data: dict[str, Any],
    *,
    team_profiles: dict[str, dict[str, Any]] | None = None,
) -> None:
    effective_team_profiles = team_profiles or TEAM_PROFILE_DEFAULTS
    pm_team = str(config_data.get("pm_team", "") or "").strip().upper()
    valid_team_labels = ", ".join(profile["label"] for profile in effective_team_profiles.values())
    if not pm_team:
        raise ToolError(f"PM Team is required. Choose {valid_team_labels} before saving setup.")
    if pm_team not in effective_team_profiles:
        raise ToolError(f"Unsupported PM Team: {pm_team}. Choose {valid_team_labels}.")
    profile = effective_team_profiles[pm_team]
    has_routing = bool(str(config_data.get("component_route_rules_text", "") or "").strip())
    has_defaults = bool(str(config_data.get("component_default_rules_text", "") or "").strip())
    if not profile.get("ready") and not (has_routing and has_defaults):
        raise ToolError(
            f"{profile['label']} default setup is not available yet. Open Advanced mapping overrides and enter Component Routing and Component Defaults manually for now."
        )


def _resolve_spreadsheet_id(value: str) -> str | None:
    if not value:
        return None
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", value)
    if match:
        return match.group(1)
    if re.fullmatch(r"[a-zA-Z0-9-_]{20,}", value):
        return value
    return None


def _build_run_notice(results: list[object], dry_run: bool) -> dict[str, object]:
    created = [result for result in results if getattr(result, "status", "") == "created"]
    errors = [result for result in results if getattr(result, "status", "") == "error"]
    skipped = [result for result in results if getattr(result, "status", "") == "skipped"]
    previews = [result for result in results if getattr(result, "status", "") == "preview"]

    title = "Dry Run Ready" if dry_run else ("Run Completed" if not errors else "Run Completed With Issues")
    tone = "success" if not errors else "warning"
    summary = (
        f"{len(previews)} ready, {len(errors)} error, {len(skipped)} skipped."
        if dry_run
        else f"{len(created)} created, {len(errors)} error, {len(skipped)} skipped."
    )

    details: list[str] = []
    for result in created[:3]:
        ticket = getattr(result, "ticket_key", None) or getattr(result, "ticket_link", None) or "-"
        details.append(f"Created row {result.row_number}: {ticket}")
    for result in errors[:3]:
        details.append(f"Row {result.row_number}: {result.message}")
    if not details:
        for result in skipped[:3]:
            details.append(f"Row {result.row_number}: {result.message}")

    return {
        "title": title,
        "tone": tone,
        "summary": summary,
        "details": details,
    }


def _build_sync_notice(results: list[object]) -> dict[str, object]:
    created = [result for result in results if getattr(result, "status", "") == "created"]
    updated = [result for result in results if getattr(result, "status", "") == "updated"]
    errors = [result for result in results if getattr(result, "status", "") == "error"]
    skipped = [result for result in results if getattr(result, "status", "") == "skipped"]

    details: list[str] = []
    for result in created[:3]:
        details.append(f"Added BPMIS Issue ID {result.issue_id}.")
    for result in updated[:3]:
        details.append(f"Updated BPMIS Issue ID {result.issue_id}.")
    for result in errors[:3]:
        details.append(f"{result.issue_id or 'Unknown Issue'}: {result.message}")
    if not details:
        for result in skipped[:3]:
            details.append(f"{result.issue_id}: {result.message}")

    return {
        "title": "BPMIS Sync Completed" if not errors else "BPMIS Sync Completed With Issues",
        "tone": "success" if not errors else "warning",
        "summary": f"{len(created)} added, {len(updated)} updated, {len(skipped)} skipped, {len(errors)} error.",
        "details": details,
    }


def _results_for_display(results: list[dict[str, Any]] | list[object], *, include_skipped: bool = False) -> list[dict[str, Any]]:
    filtered: list[dict[str, Any]] = []
    for result in results:
        payload = result if isinstance(result, dict) else result.__dict__
        if not include_skipped and str(payload.get("status", "") or "").lower() == "skipped":
            continue
        filtered.append(dict(payload))
    return filtered


def _classify_source_code_qa_job_error(message: str) -> dict[str, Any]:
    normalized = str(message or "").lower()
    if "local-agent" in normalized or "local agent" in normalized or "connection refused" in normalized:
        return {"error_category": "local_agent_offline", "error_code": "local_agent_unavailable", "error_retryable": True}
    if "ngrok" in normalized or "err_ngrok_3200" in normalized or "gateway" in normalized or "html error" in normalized:
        return {"error_category": "gateway_disconnected", "error_code": "gateway_disconnected", "error_retryable": True}
    if "rate limit" in normalized or "quota" in normalized:
        return {"error_category": "codex_timeout_or_rate_limit", "error_code": "llm_rate_limited", "error_retryable": True}
    if "timeout" in normalized or "timed out" in normalized:
        return {"error_category": "codex_timeout_or_rate_limit", "error_code": "llm_timeout", "error_retryable": True}
    return {"error_category": "job_failed", "error_code": "source_code_qa_job_failed", "error_retryable": True}


def _public_source_code_qa_job_snapshot(snapshot: dict[str, Any]) -> dict[str, Any]:
    payload = dict(snapshot)
    payload.pop("owner_email", None)
    payload.setdefault("status", "ok")
    payload["progress"] = snapshot.get("progress") if isinstance(snapshot.get("progress"), dict) else {
        "stage": snapshot.get("stage") or "",
        "current": snapshot.get("current") or 0,
        "total": snapshot.get("total") or 0,
        "message": snapshot.get("message") or "",
    }
    state = str(payload.get("state") or "")
    payload["query_mode"] = _source_code_qa_query_mode(payload.get("query_mode"))
    payload["queued_position"] = int(payload.get("queued_position") or 0)
    payload["eta_seconds_range"] = [
        max(0, int(value or 0))
        for value in (payload.get("eta_seconds_range") if isinstance(payload.get("eta_seconds_range"), list) else [])
    ][:2]
    payload["running_user_count"] = int(payload.get("running_user_count") or 0)
    payload["last_progress_at"] = float(payload.get("last_progress_at") or payload.get("updated_at") or 0)
    if payload.get("stalled_retryable"):
        payload["error_category"] = payload.get("error_category") or "job_stalled"
        payload["error_code"] = payload.get("error_code") or "job_stalled_retryable"
        payload["error_retryable"] = True
    if state == "running":
        payload.setdefault("error_category", "job_running")
        payload.setdefault("error_code", "")
        payload.setdefault("error_retryable", True)
    if state == "queued":
        payload.setdefault("error_category", "job_queued")
        payload.setdefault("error_code", "")
        payload.setdefault("error_retryable", True)
    if state == "failed":
        classification = _classify_source_code_qa_job_error(str(payload.get("error") or payload.get("message") or ""))
        for key, value in classification.items():
            if not payload.get(key):
                payload[key] = value
    return payload


def _source_code_qa_job_snapshot_for_current_user(job_id: str) -> dict[str, Any] | None:
    snapshot = current_app.config["JOB_STORE"].snapshot(job_id)
    if snapshot is None or snapshot.get("action") != "source-code-qa-query":
        return None
    owner_email = str(snapshot.get("owner_email") or "").strip().lower()
    current_email = _current_google_email()
    if owner_email and owner_email != current_email and not _can_manage_source_code_qa(current_app.config["SETTINGS"]):
        return None
    return snapshot


def _sanitize_meeting_recorder_job_error(error: object, *, unexpected: bool = False) -> str:
    if unexpected:
        return "Meeting processing failed unexpectedly. Check server logs for details."
    message = " ".join(str(error or "").split()).strip() or "Meeting processing failed."
    if re.search(r"traceback|token|secret|authorization|api[_ -]?key", message, re.IGNORECASE):
        return "Meeting processing failed. Check server logs for details."
    return message[:500]


def _meeting_recorder_process_job_snapshot_for_current_user(job_id: str) -> dict[str, Any] | None:
    snapshot = current_app.config["JOB_STORE"].snapshot(job_id)
    if snapshot is None or snapshot.get("action") != MEETING_RECORDER_PROCESS_ACTION:
        return None
    owner_email = str(snapshot.get("owner_email") or "").strip().lower()
    if owner_email and owner_email != _current_google_email():
        return None
    return snapshot


def _public_meeting_recorder_process_job_snapshot(snapshot: dict[str, Any]) -> dict[str, Any]:
    payload = dict(snapshot)
    payload.pop("owner_email", None)
    payload.setdefault("status", "ok")
    payload["progress"] = snapshot.get("progress") if isinstance(snapshot.get("progress"), dict) else {
        "stage": snapshot.get("stage") or "",
        "current": snapshot.get("current") or 0,
        "total": snapshot.get("total") or 0,
        "message": snapshot.get("message") or "",
    }
    state = str(payload.get("state") or "")
    if state == "queued":
        payload.setdefault("message", "Meeting processing is queued.")
        payload.setdefault("error_category", "job_queued")
        payload.setdefault("error_code", "")
        payload.setdefault("error_retryable", True)
    elif state == "running":
        payload.setdefault("message", "Meeting processing is running.")
        payload.setdefault("error_category", "job_running")
        payload.setdefault("error_code", "")
        payload.setdefault("error_retryable", True)
    elif state == "failed":
        payload["error_category"] = payload.get("error_category") or "meeting_processing_failed"
        payload["error_code"] = payload.get("error_code") or "meeting_processing_failed"
        payload["error_retryable"] = bool(payload.get("error_retryable", True))
    return payload


def _meeting_record_for_processing_job(record_id: str, owner_email: str) -> dict[str, Any]:
    record = _get_meeting_record_store().get_record(record_id)
    if str(record.get("owner_email") or "").strip().lower() != str(owner_email or "").strip().lower():
        raise ToolError("Meeting record is not available for this Google account.")
    status = str(record.get("status") or "").strip().lower()
    if status in {"recorded", "failed", "completed", "processing"}:
        return record
    raise ToolError("Stop the recording before processing this meeting.")


def _meeting_recorder_auto_process_payload(*, settings: Settings, record_id: str, owner_email: str) -> dict[str, Any]:
    try:
        if _local_agent_meeting_recorder_enabled(settings):
            result = _build_local_agent_client(settings).meeting_recorder_process_start(
                record_id=record_id,
                owner_email=owner_email,
                send_email_on_complete=True,
            )
            return {
                "state": result.get("state") or "queued",
                "job_id": result.get("job_id") or "",
            }
        payload = _queue_meeting_recorder_process_job(
            app=current_app._get_current_object(),
            settings=settings,
            record_id=record_id,
            owner_email=owner_email,
            send_email_on_complete=True,
        )
        return {
            "state": payload.get("state") or "queued",
            "job_id": payload.get("job_id") or "",
        }
    except (ConfigError, ToolError, requests.RequestException) as error:
        current_app.logger.warning(
            "Meeting Recorder auto process queue failed. record_id=%s",
            record_id,
            exc_info=True,
        )
        return {"auto_process_error": _sanitize_meeting_recorder_job_error(error)}


def _queue_meeting_recorder_process_job(
    *,
    app: Flask,
    settings: Settings,
    record_id: str,
    owner_email: str,
    send_email_on_complete: bool = False,
) -> dict[str, Any]:
    del settings
    normalized_owner = str(owner_email or "").strip().lower()
    record = _meeting_record_for_processing_job(record_id, normalized_owner)
    job_store: JobStore = current_app.config["JOB_STORE"]
    active_job = job_store.active_for_record(
        MEETING_RECORDER_PROCESS_ACTION,
        owner_email=normalized_owner,
        record_id=str(record.get("record_id") or record_id),
    )
    if active_job is not None:
        return {
            "status": "queued",
            "state": active_job.get("state") or "queued",
            "job_id": active_job.get("job_id") or "",
            "record": _meeting_record_summary(record),
        }
    if str(record.get("status") or "").strip().lower() == "processing":
        record["status"] = "recorded"
        record["error"] = ""
        _get_meeting_record_store().save_record(record)
    job = job_store.create(
        MEETING_RECORDER_PROCESS_ACTION,
        title="Process Meeting Recording",
        owner_email=normalized_owner,
        record_id=str(record.get("record_id") or record_id),
    )
    job_store.update(
        job.job_id,
        state="queued",
        stage="queued",
        message="Meeting processing is queued.",
        current=0,
        total=1,
    )
    thread = threading.Thread(
        target=_run_meeting_recorder_process_job,
        args=(app, job.job_id, str(record.get("record_id") or record_id), normalized_owner, bool(send_email_on_complete)),
        daemon=True,
    )
    thread.start()
    return {
        "status": "queued",
        "state": "queued",
        "job_id": job.job_id,
        "record": _meeting_record_summary(record),
    }


def _mark_meeting_record_process_failed(record_id: str, owner_email: str, message: str) -> None:
    try:
        record = _get_meeting_record_store().get_record(record_id)
        if str(record.get("owner_email") or "").strip().lower() != str(owner_email or "").strip().lower():
            return
        record["status"] = "failed"
        record["error"] = message
        _get_meeting_record_store().save_record(record)
    except Exception:  # pragma: no cover - best effort cleanup after a worker failure.
        current_app.logger.exception("Failed to mark meeting recorder record as failed.")


def _mark_meeting_record_email_failed(record_id: str, owner_email: str, message: str) -> None:
    try:
        record = _get_meeting_record_store().get_record(record_id)
        if str(record.get("owner_email") or "").strip().lower() != str(owner_email or "").strip().lower():
            return
        record["email"] = {
            "status": "failed",
            "error": message,
            "failed_at": _utc_now(),
            "recipient": str(owner_email or "").strip().lower(),
        }
        _get_meeting_record_store().save_record(record)
    except Exception:  # pragma: no cover - best effort annotation after processing succeeds.
        current_app.logger.exception("Failed to mark meeting recorder email as failed.")


def _send_meeting_recorder_minutes_email_after_process(
    *,
    settings: Settings,
    record_id: str,
    owner_email: str,
) -> dict[str, Any] | None:
    try:
        return _build_meeting_processing_service(settings).send_minutes_email(
            record_id=record_id,
            owner_email=owner_email,
            recipient=owner_email,
        )
    except (ConfigError, ToolError) as error:
        message = _sanitize_meeting_recorder_job_error(error)
        _mark_meeting_record_email_failed(record_id, owner_email, message)
        current_app.logger.warning(
            "Meeting Recorder auto email failed. record_id=%s",
            record_id,
            exc_info=True,
        )
        return {"status": "failed", "error": message, "recipient": str(owner_email or "").strip().lower()}


def _run_meeting_recorder_process_job(
    app: Flask,
    job_id: str,
    record_id: str,
    owner_email: str,
    send_email_on_complete: bool = False,
) -> None:
    with app.app_context():
        job_store: JobStore = app.config["JOB_STORE"]
        job_store.update(
            job_id,
            state="running",
            stage="processing",
            message="Transcribing audio and generating meeting minutes.",
            current=0,
            total=1,
        )
        try:
            record = _build_meeting_processing_service(app.config["SETTINGS"]).process_recording(
                record_id=record_id,
                owner_email=owner_email,
            )
            email_payload = (
                _send_meeting_recorder_minutes_email_after_process(
                    settings=app.config["SETTINGS"],
                    record_id=record_id,
                    owner_email=owner_email,
                )
                if send_email_on_complete
                else None
            )
            completed_record = _get_meeting_record_store().get_record(record_id)
            _record_meeting_work_memory(completed_record)
            details = [f"Record: {record_id}"]
            if email_payload:
                if email_payload.get("status") == "sent":
                    details.append(f"Email sent to {email_payload.get('recipient') or owner_email}")
                elif email_payload.get("status") == "failed":
                    details.append("Email was not sent automatically.")
            job_store.complete(
                job_id,
                results=[{"record": _meeting_record_summary(completed_record), "email": email_payload or {}}],
                notice={
                    "title": "Meeting Processing Completed",
                    "tone": "success",
                    "summary": "Transcript and minutes are ready.",
                    "details": details,
                },
            )
        except ToolError as error:
            message = _sanitize_meeting_recorder_job_error(error)
            _mark_meeting_record_process_failed(record_id, owner_email, message)
            job_store.fail(
                job_id,
                message,
                error_category="meeting_processing_failed",
                error_code="meeting_processing_failed",
                error_retryable=True,
            )
        except Exception as error:  # pragma: no cover - defensive guard for background worker failures.
            message = _sanitize_meeting_recorder_job_error(error, unexpected=True)
            current_app.logger.error(
                "Meeting Recorder process job failed unexpectedly. job_id=%s record_id=%s",
                job_id,
                record_id,
            )
            _mark_meeting_record_process_failed(record_id, owner_email, message)
            job_store.fail(
                job_id,
                message,
                error_category="meeting_processing_failed",
                error_code="meeting_processing_unexpected_error",
                error_retryable=True,
            )


def _serialize_results(results: list[object], *, include_skipped: bool = True) -> list[dict[str, Any]]:
    return _results_for_display(results, include_skipped=include_skipped)


def _run_source_code_qa_sync_job(
    app: Flask,
    job_id: str,
    settings: Settings,
    pm_team: str,
    country: str,
) -> None:
    del settings
    with app.app_context():
        job_store: JobStore = app.config["JOB_STORE"]
        job_store.update(
            job_id,
            state="running",
            stage="syncing",
            message="Syncing repositories and rebuilding the source-code index.",
            current=0,
            total=1,
        )
        try:
            result = _build_source_code_qa_service().sync(pm_team=pm_team, country=country)
            status = str(result.get("status") or "ok")
            summary = "Source repositories are synced." if status == "ok" else "Source repository sync completed with issues."
            job_store.complete(
                job_id,
                results=[result],
                notice={
                    "title": "Source Code Sync",
                    "tone": "success" if status == "ok" else "warning",
                    "summary": summary,
                    "details": [f"Status: {status}", f"Repositories: {len(result.get('results') or [])}"],
                },
            )
        except ToolError as error:
            job_store.fail(job_id, str(error))
        except Exception as error:  # pragma: no cover - defensive guard for background worker failures.
            app.logger.exception("Source code QA sync job failed unexpectedly.")
            job_store.fail(job_id, f"Unexpected error: {error}")


def _run_source_code_qa_query_job(app: Flask, job_id: str, payload: dict[str, Any]) -> None:
    with app.app_context():
        job_store: JobStore = app.config["JOB_STORE"]

        def progress_callback(stage: str, message: str, current: int, total: int) -> None:
            job_store.update(
                job_id,
                state="running",
                stage=stage,
                message=message,
                current=current,
                total=total,
            )

        try:
            if not _source_code_qa_provider_available(payload.get("llm_provider")):
                raise ToolError("Selected Source Code Q&A model is unavailable.")
            service = _build_source_code_qa_service(payload.get("llm_provider"))
            pm_team = str(payload.get("pm_team") or "")
            country = str(payload.get("country") or "")
            query_mode = _source_code_qa_query_mode(payload.get("query_mode"))
            session_store = _get_source_code_qa_session_store()
            session_id = str(payload.get("session_id") or "").strip()
            owner_email = str(payload.get("_session_owner_email") or "").strip().lower() or "local"
            conversation_context = payload.get("conversation_context") if isinstance(payload.get("conversation_context"), dict) else None
            if conversation_context is None and session_id:
                conversation_context = session_store.get_context(session_id, owner_email=owner_email)
            if isinstance(payload.get("_resolved_attachments"), list):
                attachments = payload.get("_resolved_attachments") or []
            else:
                attachments = _resolve_source_code_qa_query_attachments(payload, owner_email=owner_email, session_id=session_id)
            runtime_evidence = _resolve_source_code_qa_runtime_evidence(pm_team=pm_team, country=country)
            auto_sync = _prepare_source_code_qa_auto_sync(
                service,
                pm_team=pm_team,
                country=country,
                progress_callback=progress_callback,
            )
            def run_query() -> dict[str, Any]:
                return service.query(
                    pm_team=pm_team,
                    country=country,
                    question=str(payload.get("question") or ""),
                    answer_mode=_source_code_qa_public_answer_mode(payload.get("answer_mode")),
                    llm_budget_mode="auto",
                    query_mode=query_mode,
                    conversation_context=conversation_context,
                    attachments=attachments,
                    runtime_evidence=runtime_evidence,
                    progress_callback=progress_callback,
                )

            if service.llm_provider_name == "codex_cli_bridge" and session_id:
                progress_callback("codex_session_lock", "Waiting for this chat's Codex session slot.", 0, 1)
                with _source_code_qa_codex_session_lock(session_id):
                    result = run_query()
            else:
                result = run_query()
            result["auto_sync"] = auto_sync
            result["attachments"] = _source_code_qa_public_attachments(attachments)
            result["runtime_evidence"] = _source_code_qa_public_runtime_evidence(runtime_evidence)
            if session_id:
                session_write_started = time.perf_counter()
                session_payload = session_store.append_exchange(
                    session_id,
                    owner_email=owner_email,
                    pm_team=pm_team,
                    country=country,
                    llm_provider=str(payload.get("llm_provider") or ""),
                    question=str(payload.get("question") or ""),
                    result=result,
                    context=_build_source_code_qa_session_context(result, payload),
                    attachments=attachments,
                )
                current_app.logger.warning(
                    "source_code_qa_timing %s",
                    json.dumps(
                        {
                            "event": "source_code_qa_timing",
                            "component": "session_write",
                            "elapsed_ms": int((time.perf_counter() - session_write_started) * 1000),
                            "job_id": job_id,
                            "trace_id": str(result.get("trace_id") or ""),
                            "session_id": session_id,
                            "owner_email": owner_email,
                            "message_count": len(session_payload.get("messages") or []) if isinstance(session_payload, dict) else 0,
                            "status": "ok" if session_payload is not None else "missing_session",
                        },
                        ensure_ascii=False,
                        sort_keys=True,
                    ),
                )
                if session_payload is not None:
                    result["session"] = session_payload
                    result["session_id"] = session_id
            status = str(result.get("status") or "ok")
            started_at = float((job_store.snapshot(job_id) or {}).get("started_at") or time.time())
            elapsed_seconds = max(0.0, time.time() - started_at)
            if elapsed_seconds > 120:
                current_app.logger.warning(
                    "source_code_qa_slow_query job_id=%s owner=%s elapsed_seconds=%.1f stage=%s",
                    job_id,
                    owner_email,
                    elapsed_seconds,
                    str((job_store.snapshot(job_id) or {}).get("stage") or ""),
                )
            _record_source_code_qa_work_memory(
                owner_email=owner_email,
                pm_team=pm_team,
                country=country,
                question=str(payload.get("question") or ""),
                result=result,
                session_id=session_id,
                job_id=job_id,
            )
            job_store.complete(
                job_id,
                results=[result],
                notice={
                    "title": "Source Code Q&A",
                    "tone": "success" if status == "ok" else "warning",
                    "summary": result.get("summary") or "Source Code Q&A completed.",
                    "details": [f"Status: {status}", f"Trace: {result.get('trace_id') or 'n/a'}"],
                },
            )
        except ToolError as error:
            job_store.fail(job_id, str(error), **_classify_source_code_qa_job_error(str(error)))
        except Exception as error:  # pragma: no cover - defensive guard for background worker failures.
            app.logger.exception("Source code QA query job failed unexpectedly.")
            message = f"Unexpected error: {error}"
            job_store.fail(job_id, message, **_classify_source_code_qa_job_error(message))


def _run_source_code_qa_effort_assessment_job(app: Flask, job_id: str, payload: dict[str, Any]) -> None:
    with app.app_context():
        job_store: JobStore = app.config["JOB_STORE"]

        def progress_callback(stage: str, message: str, current: int, total: int) -> None:
            job_store.update(
                job_id,
                state="running",
                stage=stage,
                message=message,
                current=current,
                total=total,
            )

        try:
            if not _source_code_qa_provider_available(payload.get("llm_provider")):
                raise ToolError("Selected Source Code Q&A model is unavailable.")
            service = _build_source_code_qa_service(payload.get("llm_provider"))
            pm_team = str(payload.get("pm_team") or "")
            country = str(payload.get("country") or "")
            language = _source_code_qa_effort_assessment_language(payload.get("language"))
            requirement = str(payload.get("requirement") or "").strip()
            if not requirement:
                raise ToolError("Business requirement is empty.")
            progress_callback("assessment_prompt", "Building optimized effort assessment prompt.", 0, 1)
            runtime_evidence = _resolve_source_code_qa_runtime_evidence(pm_team=pm_team, country=country)
            business_plan = _build_source_code_qa_effort_business_plan(
                pm_team=pm_team,
                country=country,
                language=language,
                requirement=requirement,
            )
            technical_candidates = _build_source_code_qa_effort_technical_candidates(
                pm_team=pm_team,
                country=country,
                business_plan=business_plan,
                requirement=requirement,
            )
            estimation_rubric = _build_source_code_qa_effort_estimation_rubric(
                business_plan=business_plan,
                technical_candidates=technical_candidates,
            )
            optimized_prompt = _build_source_code_qa_effort_assessment_prompt(
                pm_team=pm_team,
                country=country,
                language=language,
                requirement=requirement,
                llm_provider=str(payload.get("llm_provider") or ""),
                runtime_evidence=runtime_evidence,
                business_plan=business_plan,
                technical_candidates=technical_candidates,
                estimation_rubric=estimation_rubric,
            )
            result = service.query(
                pm_team=pm_team,
                country=country,
                question=optimized_prompt,
                answer_mode="auto",
                llm_budget_mode="auto",
                query_mode="deep",
                conversation_context=None,
                attachments=[],
                runtime_evidence=runtime_evidence,
                progress_callback=progress_callback,
            )
            result = _normalize_source_code_qa_effort_assessment_result(
                result=result,
                language=language,
                business_plan=business_plan,
                technical_candidates=technical_candidates,
                estimation_rubric=estimation_rubric,
            )
            result["assessment"] = {
                "type": "effort_assessment",
                "pm_team": pm_team,
                "country": country,
                "language": language,
                "requirement": requirement,
                "business_plan": business_plan,
                "technical_candidates": technical_candidates,
                "estimation_rubric": estimation_rubric,
                "structured_assessment": result.get("structured_assessment") or {},
                "confidence": result.get("assessment_confidence") or "low",
                "missing_evidence": result.get("missing_evidence") or [],
                "evidence_status": result.get("effort_evidence_status") or "warning",
            }
            result["runtime_evidence"] = _source_code_qa_public_runtime_evidence(runtime_evidence)
            status = str(result.get("status") or "ok")
            notice_tone = "warning" if result.get("effort_evidence_status") == "warning" else ("success" if status == "ok" else "warning")
            job_store.complete(
                job_id,
                results=[result],
                notice={
                    "title": "Effort Assessment",
                    "tone": notice_tone,
                    "summary": result.get("summary") or "Effort assessment completed.",
                    "details": [
                        f"Status: {status}",
                        f"Evidence: {result.get('effort_evidence_status') or 'n/a'}",
                        f"Trace: {result.get('trace_id') or 'n/a'}",
                    ],
                },
            )
        except ToolError as error:
            job_store.fail(job_id, str(error), **_classify_source_code_qa_job_error(str(error)))
        except Exception as error:  # pragma: no cover - defensive guard for background worker failures.
            app.logger.exception("Source code QA effort assessment job failed unexpectedly.")
            message = f"Unexpected error: {error}"
            job_store.fail(job_id, message, **_classify_source_code_qa_job_error(message))


def _run_team_dashboard_monthly_report_draft_job(
    app: Flask,
    job_id: str,
    settings: Settings,
    request_payload: dict[str, Any],
    user_identity: dict[str, str | None],
) -> None:
    with app.app_context():
        job_store: JobStore = app.config["JOB_STORE"]

        def progress_callback(
            stage: str,
            message: str,
            current: int,
            total: int,
            *,
            estimated_prompt_tokens: int = 0,
            token_risk: str = "",
        ) -> None:
            job_store.update(
                job_id,
                state="running",
                stage=stage,
                message=message,
                current=current,
                total=total,
                estimated_prompt_tokens=estimated_prompt_tokens,
                token_risk=token_risk,
            )

        try:
            progress_callback("preparing_sources", "Preparing Monthly Report sources.", 0, 0)
            if _local_agent_seatalk_enabled(settings):
                data = _build_local_agent_client(settings).team_dashboard_monthly_report_draft(
                    request_payload,
                    progress_callback=progress_callback,
                )
            else:
                data = _build_monthly_report_service(settings).generate_draft(
                    **request_payload,
                    progress_callback=progress_callback,
                )
            evidence = data.get("evidence_summary") if isinstance(data.get("evidence_summary"), dict) else {}
            generation = data.get("generation_summary") if isinstance(data.get("generation_summary"), dict) else {}
            _log_portal_event(
                "team_dashboard_monthly_report_draft_success",
                user=_safe_email_identity(user_identity),
                job_id=job_id,
                **evidence,
                **{f"generation_{key}": value for key, value in generation.items() if key in {"total_batches", "max_batch_estimated_tokens", "final_estimated_tokens", "elapsed_seconds"}},
            )
            job_store.complete(
                job_id,
                results=[data],
                notice={
                    "title": "Monthly Report",
                    "tone": "success",
                    "summary": "Monthly Report draft generated.",
                    "details": [
                        f"Total batches: {generation.get('total_batches', 0)}",
                        f"Final estimated tokens: {generation.get('final_estimated_tokens', generation.get('estimated_prompt_tokens', 0))}",
                    ],
                },
            )
        except ToolError as error:
            _log_portal_event(
                "team_dashboard_monthly_report_draft_tool_error",
                level=logging.WARNING,
                user=_safe_email_identity(user_identity),
                job_id=job_id,
                **_classify_portal_error(error),
            )
            job_store.fail(job_id, str(error))
        except Exception as error:  # pragma: no cover - defensive guard for background worker failures.
            _log_portal_event(
                "team_dashboard_monthly_report_draft_unexpected_error",
                level=logging.ERROR,
                user=_safe_email_identity(user_identity),
                job_id=job_id,
                **_classify_portal_error(error),
            )
            app.logger.exception("Team Dashboard Monthly Report draft job failed unexpectedly.")
            job_store.fail(job_id, f"Unexpected error: {error}")


def _run_background_job(
    app: Flask,
    job_id: str,
    action: str,
    settings: Settings,
    config_data: dict[str, Any],
    credentials_payload: dict[str, Any],
    dry_run: bool,
) -> None:
    with app.app_context():
        job_store: JobStore = app.config["JOB_STORE"]
        logger = app.logger
        job_store.update(
            job_id,
            state="running",
            stage="starting",
            message="Preparing your request.",
            current=0,
            total=0,
        )
        _log_portal_event(
            "job_started",
            logger=logger,
            job_id=job_id,
            action=action,
            dry_run=dry_run,
            pm_team=str(config_data.get("pm_team", "") or "").strip().upper(),
            user=str(config_data.get("sync_pm_email", "") or "").strip().lower(),
        )

        def progress_callback(stage: str, message: str, current: int, total: int) -> None:
            job_store.update(
                job_id,
                state="running",
                stage=stage,
                message=message,
                current=current,
                total=total,
            )

        try:
            if action == "sync-bpmis-projects":
                service = _build_portal_project_sync_service(settings, config_data)
                results = service.sync_projects(
                    user_key=str(config_data.get("_user_key", "")).strip(),
                    pm_email=str(config_data.get("sync_pm_email", "")).strip(),
                    progress_callback=progress_callback,
                )
                notice = _build_sync_notice(results)
            else:
                service = _build_service_from_config(settings, config_data, credentials_payload)
                if dry_run:
                    results, _headers = service.preview(progress_callback=progress_callback)
                else:
                    results = service.run(dry_run=False, progress_callback=progress_callback)
                notice = _build_run_notice(results, dry_run=dry_run)
            include_skipped = action == "sync-bpmis-projects"
            job_store.complete(
                job_id,
                results=_serialize_results(results, include_skipped=include_skipped),
                notice=notice,
            )
            _log_portal_event(
                "job_completed",
                logger=logger,
                job_id=job_id,
                action=action,
                dry_run=dry_run,
                result_count=len(results),
            )
        except ToolError as error:
            job_store.fail(job_id, str(error))
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "job_tool_error",
                level=logging.WARNING,
                logger=logger,
                job_id=job_id,
                action=action,
                dry_run=dry_run,
                **error_details,
            )
        except Exception as error:  # noqa: BLE001
            job_store.fail(job_id, f"Unexpected error: {error}")
            error_details = _classify_portal_error(error)
            _log_portal_event(
                "job_unexpected_error",
                level=logging.ERROR,
                logger=logger,
                job_id=job_id,
                action=action,
                dry_run=dry_run,
                **error_details,
            )
            logger.exception("Background job failed unexpectedly.")


def _csv_escape(value: str) -> str:
    text = str(value)
    if any(character in text for character in [",", "\"", "\n"]):
        return "\"" + text.replace("\"", "\"\"") + "\""
    return text


def _resolve_bpmis_access_token(config_data: dict[str, Any], settings: Settings) -> str | None:
    configured_token = str(config_data.get("bpmis_api_access_token", "") or "").strip()
    return configured_token or settings.bpmis_api_access_token


def _normalize_team_dashboard_emails(value: Any) -> list[str]:
    if isinstance(value, str):
        raw_values = re.split(r"[\s,;]+", value)
    elif isinstance(value, list):
        raw_values = value
    else:
        raw_values = []
    normalized: list[str] = []
    for raw_email in raw_values:
        email = str(raw_email or "").strip().lower()
        if email and email not in normalized:
            normalized.append(email)
    return normalized


def _build_team_dashboard_task_group(
    team_key: str,
    label: str,
    member_emails: list[str],
    tasks: list[dict[str, Any]],
    biz_projects: list[dict[str, Any]] | None = None,
    *,
    key_project_overrides: dict[str, Any] | None = None,
) -> dict[str, Any]:
    under_prd: list[dict[str, Any]] = []
    pending_live: list[dict[str, Any]] = []
    seen: set[str] = set()
    for task in tasks:
        normalized_task = _normalize_team_dashboard_task(task)
        dedupe_key = normalized_task["jira_id"] or normalized_task["issue_id"]
        if dedupe_key and dedupe_key in seen:
            continue
        if dedupe_key:
            seen.add(dedupe_key)
        status_key = normalized_task["jira_status"].strip().casefold()
        if status_key in TEAM_DASHBOARD_UNDER_PRD_STATUSES:
            under_prd.append(normalized_task)
        elif status_key not in TEAM_DASHBOARD_EXCLUDED_PENDING_STATUSES:
            pending_live.append(normalized_task)
    under_prd_biz_projects, pending_live_biz_projects = _split_team_dashboard_biz_projects_by_status(biz_projects or [])
    under_prd_projects = _group_team_dashboard_tasks_by_project(under_prd)
    under_prd_projects = _merge_team_dashboard_biz_projects(under_prd_projects, under_prd_biz_projects)
    _sort_team_dashboard_under_prd_projects(under_prd_projects)
    pending_live_projects = _group_team_dashboard_tasks_by_project(pending_live, sort_by_release=True)
    pending_live_projects = _merge_team_dashboard_biz_projects(
        pending_live_projects,
        pending_live_biz_projects,
        sort_by_release=True,
    )
    _apply_team_dashboard_key_project_states(under_prd_projects, key_project_overrides or {})
    _apply_team_dashboard_key_project_states(pending_live_projects, key_project_overrides or {})
    return {
        "team_key": team_key,
        "label": label,
        "member_emails": member_emails,
        "under_prd": under_prd_projects,
        "pending_live": pending_live_projects,
        "error": "",
    }


def _team_dashboard_biz_projects_for_emails(bpmis_client: Any, emails: list[str]) -> list[dict[str, Any]]:
    projects: dict[str, dict[str, Any]] = {}
    normalized_emails = _normalize_team_dashboard_emails(emails)
    if hasattr(bpmis_client, "list_biz_projects_for_pm_emails"):
        try:
            rows = bpmis_client.list_biz_projects_for_pm_emails(normalized_emails)
        except Exception as error:  # noqa: BLE001 - fall back to the stable single-email path.
            current_app.logger.warning("Team Dashboard bulk Biz Project lookup failed: %s", error)
        else:
            for row in rows or []:
                project = _normalize_team_dashboard_project(row if isinstance(row, dict) else {})
                matched_emails = _normalize_team_dashboard_emails(
                    row.get("matched_pm_emails") if isinstance(row, dict) else []
                )
                if not matched_emails:
                    matched_emails = _normalize_team_dashboard_emails(project.get("matched_pm_emails") or [])
                if not matched_emails:
                    regional_pm = str(project.get("regional_pm_pic") or "").strip().lower()
                    if regional_pm in normalized_emails:
                        matched_emails = [regional_pm]
                _merge_team_dashboard_biz_project_lookup(projects, project, matched_emails)
            return list(projects.values())
    for email in normalized_emails:
        normalized_email = str(email or "").strip().lower()
        if not normalized_email:
            continue
        try:
            rows = bpmis_client.list_biz_projects_for_pm_email(normalized_email)
        except Exception as error:  # noqa: BLE001 - Jira data should still render when project coverage fails.
            current_app.logger.warning("Team Dashboard Biz Project lookup failed for %s: %s", normalized_email, error)
            continue
        for row in rows or []:
            project = _normalize_team_dashboard_project(row if isinstance(row, dict) else {})
            _merge_team_dashboard_biz_project_lookup(projects, project, [normalized_email])
    return list(projects.values())


def _team_dashboard_load_jira_and_biz_projects(
    jira_bpmis_client: Any,
    biz_bpmis_client: Any,
    emails: list[str],
    timing_stats: dict[str, float],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    app_obj = current_app._get_current_object()

    def load_jira_tasks() -> tuple[list[dict[str, Any]], float]:
        started_at = time.monotonic()
        rows = jira_bpmis_client.list_jira_tasks_created_by_emails(
            emails,
            max_pages=_team_dashboard_jira_max_pages(),
            enrich_missing_parent=False,
            release_after=_team_dashboard_jira_release_after(),
        )
        return rows, round(time.monotonic() - started_at, 3)

    def load_biz_projects() -> tuple[list[dict[str, Any]], float]:
        started_at = time.monotonic()
        with app_obj.app_context():
            rows = _team_dashboard_biz_projects_for_emails(biz_bpmis_client, emails)
        return rows, round(time.monotonic() - started_at, 3)

    started_at = time.monotonic()
    with ThreadPoolExecutor(max_workers=2) as executor:
        jira_future = executor.submit(load_jira_tasks)
        biz_future = executor.submit(load_biz_projects)
        tasks, jira_elapsed = jira_future.result()
        biz_projects, biz_elapsed = biz_future.result()
    timing_stats["list_jira_tasks"] = jira_elapsed
    timing_stats["list_biz_projects"] = biz_elapsed
    timing_stats["list_jira_and_biz_projects"] = round(time.monotonic() - started_at, 3)
    timing_stats.update(_team_dashboard_combined_request_timings(jira_bpmis_client, biz_bpmis_client))
    return tasks, biz_projects


def _load_team_dashboard_tasks_for_all_teams_merged(
    settings: Settings,
    store: TeamDashboardConfigStore | RemoteTeamDashboardConfigStore,
    config: dict[str, Any],
    *,
    config_elapsed: float,
    route_started_at: float,
    key_project_overrides: dict[str, Any],
) -> list[dict[str, Any]]:
    started_at = time.monotonic()
    shared_timing = _team_dashboard_new_timing()
    shared_timing["config_load"] = config_elapsed
    team_items: list[tuple[str, str, list[str]]] = []
    all_emails: list[str] = []
    for team_key, label in TEAM_DASHBOARD_TEAMS.items():
        team_config = (config.get("teams") or {}).get(team_key) or {}
        emails = _normalize_team_dashboard_emails(team_config.get("member_emails") or [])
        team_items.append((team_key, label, emails))
        all_emails.extend(emails)
    all_emails = _normalize_team_dashboard_emails(all_emails)
    bpmis_client = _build_bpmis_client_for_current_user(settings)
    biz_bpmis_client = _build_bpmis_client_for_current_user(settings)
    tasks, biz_projects = _team_dashboard_load_jira_and_biz_projects(
        bpmis_client,
        biz_bpmis_client,
        all_emails,
        shared_timing,
    )

    team_payloads: list[dict[str, Any]] = []
    group_started_at = time.monotonic()
    for team_key, label, emails in team_items:
        team_tasks = _filter_team_dashboard_tasks_for_emails(tasks, emails)
        team_biz_projects = _filter_team_dashboard_biz_projects_for_emails(biz_projects, emails)
        team_payloads.append(
            _build_team_dashboard_task_group(
                team_key,
                label,
                emails,
                team_tasks,
                team_biz_projects,
                key_project_overrides=key_project_overrides,
            )
        )
    _team_dashboard_add_timing(shared_timing, "group_projects", group_started_at)

    backfill_started_at = time.monotonic()
    _backfill_all_team_dashboard_empty_project_jira_tasks(bpmis_client, team_payloads)
    for team_payload in team_payloads:
        _remove_team_dashboard_zero_jira_pending_live_projects(team_payload)
    _team_dashboard_add_timing(shared_timing, "backfill_zero_jira_projects", backfill_started_at)
    shared_timing.update(_team_dashboard_combined_request_timings(bpmis_client, biz_bpmis_client))

    fetch_stats = _team_dashboard_combined_fetch_stats(bpmis_client, biz_bpmis_client)
    elapsed = round(time.monotonic() - started_at, 2)
    for team_payload in team_payloads:
        timing_stats = dict(shared_timing)
        timing_stats["total"] = elapsed
        team_payload["elapsed_seconds"] = elapsed
        team_payload["fetch_stats"] = fetch_stats
        team_payload["timing_stats"] = timing_stats

    cache_started_at = time.monotonic()
    for team_key, _label, emails in team_items:
        team_payload = next((payload for payload in team_payloads if payload.get("team_key") == team_key), None)
        if team_payload is not None:
            _store_team_dashboard_task_payload(store, team_key, emails, team_payload)
    _team_dashboard_add_timing(shared_timing, "cache_store", cache_started_at)
    elapsed = round(time.monotonic() - started_at, 2)
    for team_payload in team_payloads:
        timing_stats = dict(shared_timing)
        timing_stats["total"] = elapsed
        team_payload["elapsed_seconds"] = elapsed
        team_payload["fetch_stats"] = fetch_stats
        team_payload["timing_stats"] = timing_stats

    for team_key, _label, emails in team_items:
        team_payload = next((payload for payload in team_payloads if payload.get("team_key") == team_key), {})
        _log_portal_event(
            "team_dashboard_tasks_team_loaded",
            **_build_request_log_context(
                settings,
                user_identity=_get_user_identity(settings),
                extra={
                    "team_key": team_key,
                    "email_count": len(emails),
                    "raw_task_count": len(_filter_team_dashboard_tasks_for_emails(tasks, emails)),
                    "raw_biz_project_count": len(_filter_team_dashboard_biz_projects_for_emails(biz_projects, emails)),
                    "elapsed_seconds": elapsed,
                    "fetch_stats": fetch_stats,
                    "timing_stats": team_payload.get("timing_stats") or {},
                    "all_team_merged_reload": True,
                    "route_elapsed_seconds": round(time.monotonic() - route_started_at, 2),
                },
            ),
        )
    return team_payloads


def _filter_team_dashboard_tasks_for_emails(tasks: list[dict[str, Any]], emails: list[str]) -> list[dict[str, Any]]:
    allowed_emails = set(_normalize_team_dashboard_emails(emails))
    if not allowed_emails:
        return []
    filtered: list[dict[str, Any]] = []
    for task in tasks or []:
        if not isinstance(task, dict):
            continue
        normalized_task = _normalize_team_dashboard_task(task)
        pm_email = str(normalized_task.get("pm_email") or "").strip().lower()
        if pm_email in allowed_emails:
            filtered.append(task)
    return filtered


def _filter_team_dashboard_biz_projects_for_emails(
    biz_projects: list[dict[str, Any]],
    emails: list[str],
) -> list[dict[str, Any]]:
    allowed_emails = set(_normalize_team_dashboard_emails(emails))
    if not allowed_emails:
        return []
    filtered: list[dict[str, Any]] = []
    for raw_project in biz_projects or []:
        if not isinstance(raw_project, dict):
            continue
        project = _normalize_team_dashboard_project(raw_project)
        matched_emails = _normalize_team_dashboard_emails(project.get("matched_pm_emails") or [])
        if not matched_emails:
            regional_pm = str(project.get("regional_pm_pic") or "").strip().lower()
            if regional_pm in allowed_emails:
                matched_emails = [regional_pm]
        team_matches = [email for email in matched_emails if email in allowed_emails]
        if not team_matches:
            continue
        project["matched_pm_emails"] = team_matches
        filtered.append(project)
    return filtered


def _merge_team_dashboard_biz_project_lookup(
    projects: dict[str, dict[str, Any]],
    project: dict[str, Any],
    matched_emails: list[str],
) -> None:
    bpmis_id = str(project.get("bpmis_id") or "").strip()
    if not bpmis_id:
        return
    existing = projects.setdefault(
        bpmis_id,
        {
            **project,
            "matched_pm_emails": [],
        },
    )
    for key, value in project.items():
        if value and not existing.get(key):
            existing[key] = value
    matched = existing.setdefault("matched_pm_emails", [])
    for normalized_email in _normalize_team_dashboard_emails(matched_emails):
        if normalized_email not in matched:
            matched.append(normalized_email)


def _team_dashboard_fetch_stats(bpmis_client: Any) -> dict[str, int]:
    stats = getattr(bpmis_client, "request_stats", None)
    if not isinstance(stats, dict):
        return {}
    return {
        key: int(stats.get(key) or 0)
        for key in (
            "api_call_count",
            "issue_created_before_cutoff_count",
            "issue_release_before_cutoff_count",
            "issue_release_missing_included_count",
            "issue_detail_lookup_count",
            "issue_detail_bulk_lookup_count",
            "issue_detail_bulk_issue_count",
            "issue_detail_single_fallback_count",
            "jira_live_detail_lookup_count",
            "jira_live_bulk_lookup_count",
            "jira_live_bulk_issue_count",
            "jira_live_status_override_count",
            "bpmis_release_query_filter_probe_count",
            "bpmis_release_query_filter_enabled_count",
            "bpmis_release_query_filter_disabled_count",
            "bpmis_release_query_filter_probe_failed_count",
            "bpmis_release_query_filter_used_count",
            "issue_detail_enrichment_skipped_count",
            "issue_list_created_cutoff_hit",
            "issue_list_page_cap_hit",
            "issue_list_page_count",
            "issue_rows_scanned",
            "issue_tree_page_count",
            "issue_tree_rows_scanned",
            "issue_tree_fallback_count",
            "release_version_lookup_count",
            "release_version_count",
            "release_version_lookup_failed_count",
            "team_dashboard_zero_jira_fallback_candidate_count",
            "team_dashboard_zero_jira_bulk_project_count",
            "team_dashboard_zero_jira_bulk_hit_count",
            "team_dashboard_zero_jira_bulk_failed_count",
            "team_dashboard_zero_jira_per_project_fallback_count",
            "team_dashboard_zero_jira_per_project_fallback_skipped_count",
            "user_lookup_count",
        )
    }


def _team_dashboard_combined_fetch_stats(*bpmis_clients: Any) -> dict[str, int]:
    combined: dict[str, int] = {}
    seen: set[int] = set()
    for bpmis_client in bpmis_clients:
        if bpmis_client is None:
            continue
        identity = id(bpmis_client)
        if identity in seen:
            continue
        seen.add(identity)
        for key, value in _team_dashboard_fetch_stats(bpmis_client).items():
            combined[key] = int(combined.get(key) or 0) + int(value or 0)
    return combined


def _team_dashboard_combined_request_timings(*bpmis_clients: Any) -> dict[str, float]:
    combined: dict[str, float] = {}
    seen: set[int] = set()
    for bpmis_client in bpmis_clients:
        if bpmis_client is None:
            continue
        identity = id(bpmis_client)
        if identity in seen:
            continue
        seen.add(identity)
        timings = getattr(bpmis_client, "request_timings", None)
        if not isinstance(timings, dict):
            continue
        for key, value in timings.items():
            try:
                numeric = float(value or 0.0)
            except (TypeError, ValueError):
                continue
            combined[key] = round(float(combined.get(key) or 0.0) + numeric, 3)
    return combined


def _team_dashboard_new_timing() -> dict[str, float]:
    return {
        "config_load": 0.0,
        "cache_check": 0.0,
        "list_jira_tasks": 0.0,
        "list_biz_projects": 0.0,
        "list_jira_and_biz_projects": 0.0,
        "bpmis_user_lookup": 0.0,
        "release_versions": 0.0,
        "issue_tree_reporter": 0.0,
        "issue_tree_jiraRegionalPmPicId": 0.0,
        "parent_detail_bulk": 0.0,
        "jira_live_bulk": 0.0,
        "zero_jira_bulk": 0.0,
        "group_projects": 0.0,
        "backfill_zero_jira_projects": 0.0,
        "cache_store": 0.0,
        "total": 0.0,
    }


def _team_dashboard_add_timing(timing_stats: dict[str, float], key: str, started_at: float) -> None:
    timing_stats[key] = round(float(timing_stats.get(key) or 0.0) + (time.monotonic() - started_at), 3)


def _team_dashboard_increment_request_stat(bpmis_client: Any, key: str, amount: int = 1) -> None:
    stats = getattr(bpmis_client, "request_stats", None)
    if isinstance(stats, dict):
        stats[key] = int(stats.get(key) or 0) + amount


def _load_all_team_dashboard_task_payloads(settings: Settings, config: dict[str, Any]) -> list[dict[str, Any]]:
    key_project_overrides = config.get("key_project_overrides") if isinstance(config.get("key_project_overrides"), dict) else {}
    bpmis_client = _build_bpmis_client_for_current_user(settings)
    team_payloads: list[dict[str, Any]] = []
    for team_key, label in TEAM_DASHBOARD_TEAMS.items():
        team_config = (config.get("teams") or {}).get(team_key) or {}
        emails = _normalize_team_dashboard_emails(team_config.get("member_emails") or [])
        tasks = bpmis_client.list_jira_tasks_created_by_emails(
            emails,
            max_pages=_team_dashboard_jira_max_pages(),
            enrich_missing_parent=False,
            release_after=_team_dashboard_monthly_report_jira_release_after(),
        )
        biz_projects = _team_dashboard_biz_projects_for_emails(bpmis_client, emails)
        team_payload = _build_team_dashboard_task_group(
            team_key,
            label,
            emails,
            tasks,
            biz_projects,
            key_project_overrides=key_project_overrides,
        )
        _backfill_team_dashboard_empty_project_jira_tasks(bpmis_client, team_payload)
        team_payloads.append(team_payload)
    return team_payloads


def _load_team_dashboard_link_biz_project_payloads(settings: Settings, config: dict[str, Any], bpmis_client: Any | None = None) -> list[dict[str, Any]]:
    key_project_overrides = config.get("key_project_overrides") if isinstance(config.get("key_project_overrides"), dict) else {}
    bpmis_client = bpmis_client or _build_bpmis_client_for_current_user(settings)
    team_payloads: list[dict[str, Any]] = []
    for team_key, label in TEAM_DASHBOARD_TEAMS.items():
        team_config = (config.get("teams") or {}).get(team_key) or {}
        emails = _normalize_team_dashboard_emails(team_config.get("member_emails") or [])
        tasks = bpmis_client.list_jira_tasks_created_by_emails(
            emails,
            max_pages=_team_dashboard_jira_max_pages(),
            enrich_missing_parent=False,
            release_after=_team_dashboard_jira_release_after(),
        )
        biz_projects = _team_dashboard_biz_projects_for_emails(bpmis_client, emails)
        team_payload = _build_team_dashboard_task_group(
            team_key,
            label,
            emails,
            tasks,
            biz_projects,
            key_project_overrides=key_project_overrides,
        )
        _backfill_team_dashboard_empty_project_jira_tasks(bpmis_client, team_payload)
        _remove_team_dashboard_zero_jira_pending_live_projects(team_payload)
        team_payloads.append(team_payload)
    return team_payloads


def _team_dashboard_task_cache_signature(emails: list[str]) -> str:
    return "|".join(_normalize_team_dashboard_emails(emails))


def _cached_team_dashboard_task_payload(config: dict[str, Any], team_key: str, emails: list[str]) -> dict[str, Any] | None:
    task_cache = config.get("task_cache") if isinstance(config.get("task_cache"), dict) else {}
    if int(task_cache.get("version") or 1) != TEAM_DASHBOARD_TASK_CACHE_VERSION:
        return None
    cached_teams = task_cache.get("teams") if isinstance(task_cache.get("teams"), dict) else {}
    cached_team = cached_teams.get(team_key)
    if not isinstance(cached_team, dict):
        return None
    if str(cached_team.get("email_signature") or "") != _team_dashboard_task_cache_signature(emails):
        return None
    payload = {
        **cached_team,
        "team_key": team_key,
        "member_emails": emails,
        "loading": False,
        "loaded": True,
        "error": "",
        "progress_text": "",
        "cache_source": "server",
    }
    key_project_overrides = config.get("key_project_overrides") if isinstance(config.get("key_project_overrides"), dict) else {}
    for section_key in ("under_prd", "pending_live"):
        projects = payload.get(section_key)
        if isinstance(projects, list):
            _apply_team_dashboard_key_project_states(projects, key_project_overrides)
    return payload


def _store_team_dashboard_task_payload(
    store: TeamDashboardConfigStore | RemoteTeamDashboardConfigStore,
    team_key: str,
    emails: list[str],
    team_payload: dict[str, Any],
) -> None:
    if not team_key or team_payload.get("error"):
        return
    config = store.load()
    task_cache = config.get("task_cache") if isinstance(config.get("task_cache"), dict) else {}
    cached_teams = task_cache.get("teams") if isinstance(task_cache.get("teams"), dict) else {}
    cached_team = {
        **team_payload,
        "team_key": team_key,
        "member_emails": emails,
        "email_signature": _team_dashboard_task_cache_signature(emails),
        "cached_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "loading": False,
        "loaded": True,
        "error": "",
        "progress_text": "",
    }
    cached_teams[team_key] = cached_team
    config["task_cache"] = {
        "version": TEAM_DASHBOARD_TASK_CACHE_VERSION,
        "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "teams": cached_teams,
    }
    store.save(config)


def _load_team_dashboard_link_biz_jira_rows(settings: Settings, config: dict[str, Any]) -> list[dict[str, Any]]:
    bpmis_client = _build_bpmis_client_for_current_user(settings)
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for team_key, _label in TEAM_DASHBOARD_TEAMS.items():
        team_config = (config.get("teams") or {}).get(team_key) or {}
        emails = _normalize_team_dashboard_emails(team_config.get("member_emails") or [])
        tasks = bpmis_client.list_jira_tasks_created_by_emails(
            emails,
            max_pages=_team_dashboard_jira_max_pages(),
            enrich_missing_parent=False,
            release_after=_team_dashboard_jira_release_after(),
        )
        for raw_task in tasks or []:
            if not isinstance(raw_task, dict):
                continue
            task = _normalize_team_dashboard_task(raw_task)
            if str((task.get("parent_project") or {}).get("bpmis_id") or "").strip():
                continue
            status_key = str(task.get("jira_status") or "").strip().casefold()
            if status_key not in TEAM_DASHBOARD_UNDER_PRD_STATUSES and status_key in TEAM_DASHBOARD_EXCLUDED_PENDING_STATUSES:
                continue
            jira_id = str(task.get("jira_id") or task.get("issue_id") or "").strip()
            if not jira_id or jira_id in seen:
                continue
            seen.add(jira_id)
            rows.append(_team_dashboard_link_biz_row_from_ticket(team_key, task, {}))
    rows.sort(key=lambda row: (str(row.get("team_key") or ""), str(row.get("jira_id") or "")))
    return rows


def _build_team_dashboard_link_biz_project_rows(team_payloads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidate_projects: list[dict[str, Any]] = []
    unlinked_items: list[tuple[str, dict[str, Any]]] = []
    for team in team_payloads or []:
        team_key = str(team.get("team_key") or "").strip()
        for section_key in ("under_prd", "pending_live"):
            for project in team.get(section_key) or []:
                if not isinstance(project, dict):
                    continue
                bpmis_id = str(project.get("bpmis_id") or "").strip()
                tickets = project.get("jira_tickets") if isinstance(project.get("jira_tickets"), list) else []
                if bpmis_id:
                    candidate_projects.append(project)
                    continue
                for ticket in tickets:
                    if isinstance(ticket, dict):
                        unlinked_items.append((team_key, ticket))

    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for team_key, ticket in unlinked_items:
        jira_id = str(ticket.get("jira_id") or ticket.get("issue_id") or "").strip()
        if not jira_id or jira_id in seen:
            continue
        seen.add(jira_id)
        pm_email = str(ticket.get("pm_email") or ticket.get("reporter_email") or "").strip().lower()
        pm_candidates = _team_dashboard_link_biz_filter_candidates_for_pm(candidate_projects, pm_email)
        suggestion = _suggest_team_dashboard_biz_project(ticket, _tag_team_dashboard_candidate_source(pm_candidates, "pm"))
        suggestion["select_biz_project_options"] = _team_dashboard_link_biz_project_options(pm_candidates)
        rows.append(_team_dashboard_link_biz_row_from_ticket(team_key, ticket, suggestion))
    rows.sort(key=lambda row: (str(row.get("team_key") or ""), str(row.get("jira_id") or "")))
    return rows


def _team_dashboard_link_biz_row_from_ticket(
    team_key: str,
    ticket: dict[str, Any],
    suggestion: dict[str, Any],
    *,
    select_options: list[dict[str, str]] | None = None,
) -> dict[str, Any]:
    jira_id = str(ticket.get("jira_id") or ticket.get("issue_id") or "").strip()
    jira_title = str(ticket.get("jira_title") or "").strip()
    special_version = _team_dashboard_link_biz_special_version(jira_title)
    row = {
        "team_key": team_key,
        "jira_id": jira_id,
        "jira_link": str(ticket.get("jira_link") or (f"{_jira_browse_base_url()}{jira_id}" if jira_id else "")).strip(),
        "jira_title": jira_title,
        "reporter_email": str(ticket.get("pm_email") or ticket.get("reporter_email") or "").strip().lower(),
        "suggested_bpmis_id": str(suggestion.get("bpmis_id") or ""),
        "suggested_project_title": str(suggestion.get("project_name") or ""),
        "match_score": float(suggestion.get("match_score") or 0.0),
        "match_source": str(suggestion.get("match_source") or ""),
        "select_biz_project_options": list(suggestion.get("select_biz_project_options") or []),
    }
    if special_version:
        row["special_version"] = special_version
        row["match_source"] = row["match_source"] or "version"
    if select_options is not None:
        row["select_biz_project_options"] = select_options
    return row


def _suggest_team_dashboard_link_biz_project_rows(
    settings: Settings,
    config: dict[str, Any],
    rows: list[Any],
    *,
    team_payloads: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    bpmis_client = _build_bpmis_client_for_current_user(settings)
    parsed_tickets: list[tuple[str, dict[str, str]]] = []
    pm_emails: list[str] = []
    for raw_row in rows:
        if not isinstance(raw_row, dict):
            continue
        ticket = {
            "jira_id": str(raw_row.get("jira_id") or "").strip(),
            "jira_link": str(raw_row.get("jira_link") or "").strip(),
            "jira_title": str(raw_row.get("jira_title") or "").strip(),
            "pm_email": str(raw_row.get("reporter_email") or raw_row.get("pm_email") or "").strip().lower(),
        }
        if not ticket["jira_id"]:
            continue
        parsed_tickets.append((str(raw_row.get("team_key") or ""), ticket))
        if ticket["pm_email"]:
            pm_emails.append(ticket["pm_email"])
    candidates_by_pm = _team_dashboard_link_biz_candidate_projects_by_pm(
        bpmis_client,
        _normalize_team_dashboard_emails(pm_emails),
        team_payloads=team_payloads,
    )
    suggested_rows: list[dict[str, Any]] = []
    all_options: dict[str, dict[str, str]] = {}
    version_candidate_count = 0
    version_search_count = 0
    version_option_cache: dict[str, list[dict[str, str]]] = {}
    for team_key, ticket in parsed_tickets:
        special_version = _team_dashboard_link_biz_special_version(ticket["jira_title"])
        if special_version:
            version_search_count += 1
            if special_version not in version_option_cache:
                version_option_cache[special_version] = _team_dashboard_link_biz_version_project_options(
                    bpmis_client,
                    special_version,
                )
            version_options = version_option_cache[special_version]
            version_candidate_count += len(version_options)
            suggestion = version_options[0] if version_options else {}
            suggested_rows.append(
                _team_dashboard_link_biz_row_from_ticket(
                    team_key,
                    ticket,
                    suggestion,
                    select_options=version_options,
                )
            )
            continue
        pm_candidates = candidates_by_pm.get(ticket["pm_email"], [])
        row_options = _team_dashboard_link_biz_project_options(pm_candidates)
        for option in row_options:
            all_options[str(option.get("bpmis_id") or "")] = option
        suggestion = _suggest_team_dashboard_biz_project(ticket, _tag_team_dashboard_candidate_source(pm_candidates, "pm"))
        suggestion["select_biz_project_options"] = row_options
        suggested_rows.append(_team_dashboard_link_biz_row_from_ticket(team_key, ticket, suggestion))

    matched_count = len([row for row in suggested_rows if str(row.get("suggested_bpmis_id") or "").strip()])
    unique_candidate_ids = {
        str(project.get("bpmis_id") or "").strip()
        for candidates in candidates_by_pm.values()
        for project in candidates
        if str(project.get("bpmis_id") or "").strip()
    }
    return {
        "rows": suggested_rows,
        "matched_count": matched_count,
        "team_candidate_count": len(unique_candidate_ids),
        "keyword_candidate_count": 0,
        "keyword_search_count": 0,
        "version_candidate_count": version_candidate_count,
        "version_search_count": version_search_count,
        "select_biz_project_options": sorted(
            all_options.values(),
            key=lambda item: (str(item.get("project_name") or "").casefold(), str(item.get("bpmis_id") or "").casefold()),
        ),
    }


def _team_dashboard_link_biz_candidate_projects_from_payloads(team_payloads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for team in team_payloads or []:
        if not isinstance(team, dict):
            continue
        for section_key in ("under_prd", "pending_live"):
            for project in team.get(section_key) or []:
                if isinstance(project, dict) and str(project.get("bpmis_id") or "").strip():
                    candidates.append({**project, "team_key": str(team.get("team_key") or project.get("team_key") or "").strip()})
    return _tag_team_dashboard_candidate_source(_dedupe_team_dashboard_candidate_projects(candidates), "team")


def _team_dashboard_link_biz_allowed_project_statuses() -> set[str]:
    return set(TEAM_DASHBOARD_UNDER_PRD_BIZ_PROJECT_STATUSES) | set(TEAM_DASHBOARD_PENDING_LIVE_BIZ_PROJECT_STATUSES)


def _team_dashboard_link_biz_project_status_allowed(project: dict[str, Any]) -> bool:
    status_key = str(project.get("status") or "").strip().casefold()
    return status_key in _team_dashboard_link_biz_allowed_project_statuses()


def _team_dashboard_link_biz_filter_candidates_for_pm(candidates: list[dict[str, Any]], pm_email: str) -> list[dict[str, Any]]:
    normalized_pm = str(pm_email or "").strip().lower()
    if not normalized_pm:
        return []
    filtered: list[dict[str, Any]] = []
    for raw_project in candidates or []:
        if not isinstance(raw_project, dict):
            continue
        project = _normalize_team_dashboard_project(raw_project)
        if not _team_dashboard_link_biz_project_status_allowed(project):
            continue
        matched_emails = _normalize_team_dashboard_emails(project.get("matched_pm_emails") or [])
        if not matched_emails:
            regional_pm = str(project.get("regional_pm_pic") or "").strip().lower()
            if regional_pm:
                matched_emails = _normalize_team_dashboard_emails([regional_pm])
        if normalized_pm not in matched_emails:
            continue
        filtered.append(
            {
                **project,
                "matched_pm_emails": matched_emails,
                "team_key": str(raw_project.get("team_key") or project.get("team_key") or "").strip(),
            }
        )
    return _dedupe_team_dashboard_candidate_projects(filtered)


def _team_dashboard_link_biz_candidate_projects_by_pm(
    bpmis_client: Any,
    pm_emails: list[str],
    *,
    team_payloads: list[dict[str, Any]] | None = None,
) -> dict[str, list[dict[str, Any]]]:
    normalized_emails = _normalize_team_dashboard_emails(pm_emails)
    if not normalized_emails:
        return {}
    candidates_by_pm: dict[str, list[dict[str, Any]]] = {email: [] for email in normalized_emails}
    payload_candidates = _team_dashboard_link_biz_candidate_projects_from_payloads(team_payloads or []) if team_payloads else []
    for email in normalized_emails:
        candidates_by_pm[email] = _team_dashboard_link_biz_filter_candidates_for_pm(payload_candidates, email)

    missing_emails = [email for email in normalized_emails if not candidates_by_pm.get(email)]
    if missing_emails:
        loaded_projects = _team_dashboard_biz_projects_for_emails(bpmis_client, missing_emails)
        for email in missing_emails:
            candidates_by_pm[email] = _team_dashboard_link_biz_filter_candidates_for_pm(loaded_projects, email)
    return candidates_by_pm


def _team_dashboard_link_biz_project_options(candidates: list[dict[str, Any]]) -> list[dict[str, str]]:
    options: dict[str, dict[str, str]] = {}
    for raw_project in candidates or []:
        project = _normalize_team_dashboard_project(raw_project if isinstance(raw_project, dict) else {})
        if not _team_dashboard_link_biz_project_status_allowed(project):
            continue
        bpmis_id = str(project.get("bpmis_id") or "").strip()
        project_name = str(project.get("project_name") or "").strip()
        if not bpmis_id or not project_name:
            continue
        options[bpmis_id] = {
            "bpmis_id": bpmis_id,
            "project_name": project_name,
            "team_key": str(raw_project.get("team_key") or "").strip() if isinstance(raw_project, dict) else "",
            "market": str(project.get("market") or "").strip(),
        }
    return sorted(options.values(), key=lambda item: (item["project_name"].casefold(), item["bpmis_id"].casefold()))


def _team_dashboard_zero_jira_biz_project_options(team_payloads: list[dict[str, Any]]) -> list[dict[str, str]]:
    options: dict[str, dict[str, str]] = {}
    for team in team_payloads or []:
        if not isinstance(team, dict):
            continue
        for section_key in ("under_prd", "pending_live"):
            for raw_project in team.get(section_key) or []:
                if not isinstance(raw_project, dict):
                    continue
                project = _normalize_team_dashboard_project(raw_project)
                bpmis_id = str(project.get("bpmis_id") or "").strip()
                project_name = str(project.get("project_name") or "").strip()
                if not bpmis_id or not project_name or len(raw_project.get("jira_tickets") or []) > 0:
                    continue
                options[bpmis_id] = {
                    "bpmis_id": bpmis_id,
                    "project_name": project_name,
                    "team_key": str(team.get("team_key") or "").strip(),
                    "market": str(project.get("market") or "").strip(),
                }
    return sorted(options.values(), key=lambda item: (item["project_name"].casefold(), item["bpmis_id"].casefold()))


def _dedupe_team_dashboard_candidate_projects(candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduped: dict[str, dict[str, Any]] = {}
    for raw_project in candidates:
        project = _normalize_team_dashboard_project(raw_project if isinstance(raw_project, dict) else {})
        if isinstance(raw_project, dict) and raw_project.get("team_key"):
            project["team_key"] = str(raw_project.get("team_key") or "").strip()
        bpmis_id = str(project.get("bpmis_id") or "").strip()
        if not bpmis_id or bpmis_id in deduped:
            continue
        deduped[bpmis_id] = project
    return list(deduped.values())


def _tag_team_dashboard_candidate_source(candidates: list[dict[str, Any]], source: str) -> list[dict[str, Any]]:
    return [{**(candidate if isinstance(candidate, dict) else {}), "match_source": source} for candidate in candidates]


def _team_dashboard_link_biz_keyword_fallback_threshold() -> float:
    raw_value = str(os.getenv("TEAM_DASHBOARD_LINK_BIZ_KEYWORD_FALLBACK_SCORE") or "0.78").strip()
    try:
        return max(0.0, min(1.0, float(raw_value)))
    except ValueError:
        return 0.78


def _team_dashboard_link_biz_keywords(title: str) -> str:
    normalized = _normalize_team_dashboard_link_match_text(title)
    stop_words = {"the", "and", "for", "with", "from", "into", "jira", "feature", "support", "tech"}
    tokens = [token for token in normalized.split() if len(token) >= 3 and token not in stop_words]
    return " ".join(tokens[:8])


def _team_dashboard_link_biz_title_excluded(title: str) -> bool:
    normalized = str(title or "").casefold()
    return any(phrase in normalized for phrase in TEAM_DASHBOARD_LINK_BIZ_EXCLUDED_TITLE_PHRASES)


def _team_dashboard_link_biz_special_version(title: str) -> str:
    if not _team_dashboard_link_biz_title_excluded(title):
        return ""
    match = re.search(r"(?<!\d)(\d+\.\d+\.\d{2})(?!\d)", str(title or ""))
    return match.group(1) if match else ""


def _team_dashboard_link_biz_version_project_options(bpmis_client: Any, version: str) -> list[dict[str, str]]:
    prefix = f"AF_v{version}"
    if not version or not hasattr(bpmis_client, "search_versions") or not hasattr(bpmis_client, "list_issues_for_version"):
        return []
    try:
        version_rows = bpmis_client.search_versions(prefix) or []
    except Exception:  # noqa: BLE001 - version hints are best-effort; normal linking must still work.
        return []

    matching_versions: list[dict[str, str]] = []
    for raw_version in version_rows:
        if not isinstance(raw_version, dict):
            continue
        version_item = _serialize_productization_version_candidate(raw_version)
        version_name = str(version_item.get("version_name") or "").strip()
        version_id = str(version_item.get("version_id") or "").strip()
        if version_id and version_name.casefold().startswith(prefix.casefold()):
            matching_versions.append(version_item)

    options: dict[str, dict[str, str]] = {}
    for version_item in matching_versions:
        version_id = str(version_item.get("version_id") or "").strip()
        if not version_id:
            continue
        try:
            issue_rows = bpmis_client.list_issues_for_version(version_id) or []
        except Exception:  # noqa: BLE001 - ignore one bad version and keep other candidate versions usable.
            continue
        for issue in issue_rows:
            if not isinstance(issue, dict):
                continue
            for parent_id in sorted(_extract_parent_issue_ids_from_any(issue)):
                if parent_id in options:
                    continue
                project = _team_dashboard_link_biz_project_option_from_parent(bpmis_client, parent_id)
                if project:
                    options[parent_id] = project
    return sorted(options.values(), key=lambda item: (item["project_name"].casefold(), item["bpmis_id"].casefold()))


def _team_dashboard_link_biz_project_option_from_parent(bpmis_client: Any, parent_id: str) -> dict[str, str]:
    bpmis_id = str(parent_id or "").strip()
    if not bpmis_id:
        return {}
    detail: dict[str, Any] = {}
    if hasattr(bpmis_client, "get_issue_detail"):
        try:
            raw_detail = bpmis_client.get_issue_detail(bpmis_id)
            detail = raw_detail if isinstance(raw_detail, dict) else {}
        except Exception:  # noqa: BLE001 - a parent ID without detail should not break all suggestions.
            detail = {}
    project = _normalize_team_dashboard_project({**detail, "bpmis_id": bpmis_id, "issue_id": bpmis_id})
    project_name = str(project.get("project_name") or "").strip() or _extract_first_text(
        detail,
        "project_name",
        "summary",
        "title",
        "name",
        "issueName",
    )
    if not project_name:
        project_name = f"BPMIS {bpmis_id}"
    return {
        "bpmis_id": bpmis_id,
        "project_name": project_name,
        "team_key": "AF",
        "market": str(project.get("market") or "").strip(),
        "match_score": 1.0,
        "match_source": "version",
    }


def _normalize_team_dashboard_link_match_text(value: str) -> str:
    text = str(value or "").strip()
    while True:
        updated = re.sub(r"^\s*\[[^\]]+\]\s*", "", text).strip()
        if updated == text:
            break
        text = updated
    text = re.sub(r"[_\-|:/\\]+", " ", text.casefold())
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _suggest_team_dashboard_biz_project(ticket: dict[str, Any], candidates: list[dict[str, Any]]) -> dict[str, Any]:
    source = _normalize_team_dashboard_link_match_text(str(ticket.get("jira_title") or ""))
    best: dict[str, Any] = {}
    best_score = -1.0
    best_sort_key: tuple[str, str] = ("", "")
    for project in candidates:
        bpmis_id = str(project.get("bpmis_id") or "").strip()
        project_name = str(project.get("project_name") or "").strip()
        target = _normalize_team_dashboard_link_match_text(project_name)
        if not bpmis_id or not target:
            continue
        score = difflib.SequenceMatcher(None, source, target).ratio() if source else 0.0
        sort_key = (project_name.casefold(), bpmis_id.casefold())
        if score > best_score or (score == best_score and (not best or sort_key < best_sort_key)):
            best_score = score
            best_sort_key = sort_key
            best = {
                "bpmis_id": bpmis_id,
                "project_name": project_name,
                "match_score": round(score, 4),
                "match_source": str(project.get("match_source") or "team"),
            }
    return best


def _extract_parent_issue_ids_from_any(value: Any) -> set[str]:
    if not isinstance(value, dict):
        return set()
    parent_values = value.get("parentIds")
    if parent_values is None:
        parent_values = value.get("parentIssueId")
    if parent_values is None and isinstance(value.get("raw_jira"), dict):
        return _extract_parent_issue_ids_from_any(value["raw_jira"])
    if not isinstance(parent_values, list):
        parent_values = [parent_values] if parent_values not in (None, "") else []
    ids: set[str] = set()
    for parent in parent_values:
        if isinstance(parent, dict):
            candidate = parent.get("id") or parent.get("issueId") or parent.get("value")
        else:
            candidate = parent
        text = str(candidate or "").strip()
        if text:
            ids.add(text)
    return ids


def _backfill_team_dashboard_empty_project_jira_tasks(bpmis_client: Any, team_payload: dict[str, Any]) -> None:
    candidates: list[dict[str, Any]] = []
    for section_key in ("under_prd", "pending_live"):
        projects = team_payload.get(section_key)
        if not isinstance(projects, list):
            continue
        for project in projects:
            if not isinstance(project, dict) or project.get("jira_tickets"):
                continue
            bpmis_id = str(project.get("bpmis_id") or "").strip()
            if not bpmis_id:
                continue
            emails = _team_dashboard_project_fallback_emails(project)
            if not emails:
                continue
            candidates.append({"project": project, "bpmis_id": bpmis_id, "emails": emails})
    _team_dashboard_increment_request_stat(
        bpmis_client,
        "team_dashboard_zero_jira_fallback_candidate_count",
        len(candidates),
    )
    bulk_lookup_handled = False
    if candidates and hasattr(bpmis_client, "list_jira_tasks_for_projects_created_by_emails"):
        project_ids = [candidate["bpmis_id"] for candidate in candidates]
        all_emails: list[str] = []
        for candidate in candidates:
            all_emails.extend(candidate["emails"])
        _team_dashboard_increment_request_stat(
            bpmis_client,
            "team_dashboard_zero_jira_bulk_project_count",
            len(project_ids),
        )
        try:
            grouped_rows = bpmis_client.list_jira_tasks_for_projects_created_by_emails(
                project_ids,
                _normalize_team_dashboard_emails(all_emails),
            )
        except Exception as error:  # noqa: BLE001 - keep the older per-project fallback available.
            _team_dashboard_increment_request_stat(bpmis_client, "team_dashboard_zero_jira_bulk_failed_count")
            current_app.logger.warning("Team Dashboard bulk Jira fallback lookup failed: %s", error)
            grouped_rows = None
        if isinstance(grouped_rows, dict):
            bulk_lookup_handled = True
            for candidate in candidates:
                rows = grouped_rows.get(candidate["bpmis_id"]) or []
                tickets = _normalize_team_dashboard_project_fallback_rows(
                    rows,
                    candidate["project"],
                    candidate["emails"],
                )
                if tickets:
                    _team_dashboard_increment_request_stat(bpmis_client, "team_dashboard_zero_jira_bulk_hit_count")
                    candidate["project"]["jira_tickets"] = tickets
                    candidate["project"]["task_count"] = len(tickets)
                    _apply_team_dashboard_project_release_date(candidate["project"])
    if bulk_lookup_handled:
        skipped_count = sum(
            1
            for candidate in candidates
            if isinstance(candidate.get("project"), dict) and not candidate["project"].get("jira_tickets")
        )
        _team_dashboard_increment_request_stat(
            bpmis_client,
            "team_dashboard_zero_jira_per_project_fallback_skipped_count",
            skipped_count,
        )
    if bulk_lookup_handled:
        for section_key in ("under_prd", "pending_live"):
            projects = team_payload.get(section_key)
            if section_key == "under_prd" and isinstance(projects, list):
                _sort_team_dashboard_under_prd_projects(projects)
        return
    for section_key in ("under_prd", "pending_live"):
        projects = team_payload.get(section_key)
        if not isinstance(projects, list):
            continue
        for project in projects:
            if not isinstance(project, dict) or project.get("jira_tickets"):
                continue
            bpmis_id = str(project.get("bpmis_id") or "").strip()
            if not bpmis_id:
                continue
            _team_dashboard_increment_request_stat(bpmis_client, "team_dashboard_zero_jira_per_project_fallback_count")
            tickets = _team_dashboard_project_fallback_jira_tasks(bpmis_client, project)
            if not tickets:
                continue
            project["jira_tickets"] = tickets
            project["task_count"] = len(tickets)
            _apply_team_dashboard_project_release_date(project)
        if section_key == "under_prd":
            _sort_team_dashboard_under_prd_projects(projects)


def _backfill_all_team_dashboard_empty_project_jira_tasks(
    bpmis_client: Any,
    team_payloads: list[dict[str, Any]],
) -> None:
    candidates: list[dict[str, Any]] = []
    for team_payload in team_payloads or []:
        if not isinstance(team_payload, dict):
            continue
        for section_key in ("under_prd", "pending_live"):
            projects = team_payload.get(section_key)
            if not isinstance(projects, list):
                continue
            for project in projects:
                if not isinstance(project, dict) or project.get("jira_tickets"):
                    continue
                bpmis_id = str(project.get("bpmis_id") or "").strip()
                if not bpmis_id:
                    continue
                emails = _team_dashboard_project_fallback_emails(project)
                if not emails:
                    continue
                candidates.append(
                    {
                        "team_key": str(team_payload.get("team_key") or ""),
                        "section_key": section_key,
                        "project": project,
                        "bpmis_id": bpmis_id,
                        "emails": emails,
                    }
                )
    _team_dashboard_increment_request_stat(
        bpmis_client,
        "team_dashboard_zero_jira_fallback_candidate_count",
        len(candidates),
    )
    if not candidates:
        return
    if not hasattr(bpmis_client, "list_jira_tasks_for_projects_created_by_emails"):
        for team_payload in team_payloads or []:
            if isinstance(team_payload, dict):
                _backfill_team_dashboard_empty_project_jira_tasks(bpmis_client, team_payload)
        return

    project_ids = list(dict.fromkeys(candidate["bpmis_id"] for candidate in candidates))
    all_emails: list[str] = []
    for candidate in candidates:
        all_emails.extend(candidate["emails"])
    all_emails = _normalize_team_dashboard_emails(all_emails)
    _team_dashboard_increment_request_stat(
        bpmis_client,
        "team_dashboard_zero_jira_bulk_project_count",
        len(project_ids),
    )
    try:
        grouped_rows = bpmis_client.list_jira_tasks_for_projects_created_by_emails(project_ids, all_emails)
    except Exception as error:  # noqa: BLE001 - preserve old per-team fallback behavior when bulk fails.
        _team_dashboard_increment_request_stat(bpmis_client, "team_dashboard_zero_jira_bulk_failed_count")
        current_app.logger.warning("Team Dashboard all-team bulk Jira fallback lookup failed: %s", error)
        for team_payload in team_payloads or []:
            if isinstance(team_payload, dict):
                _backfill_team_dashboard_empty_project_jira_tasks(bpmis_client, team_payload)
        return
    if not isinstance(grouped_rows, dict):
        for team_payload in team_payloads or []:
            if isinstance(team_payload, dict):
                _backfill_team_dashboard_empty_project_jira_tasks(bpmis_client, team_payload)
        return

    for candidate in candidates:
        rows = grouped_rows.get(candidate["bpmis_id"]) or []
        tickets = _normalize_team_dashboard_project_fallback_rows(
            rows,
            candidate["project"],
            candidate["emails"],
        )
        if tickets:
            _team_dashboard_increment_request_stat(bpmis_client, "team_dashboard_zero_jira_bulk_hit_count")
            candidate["project"]["jira_tickets"] = tickets
            candidate["project"]["task_count"] = len(tickets)
            _apply_team_dashboard_project_release_date(candidate["project"])
    skipped_count = sum(
        1
        for candidate in candidates
        if isinstance(candidate.get("project"), dict) and not candidate["project"].get("jira_tickets")
    )
    _team_dashboard_increment_request_stat(
        bpmis_client,
        "team_dashboard_zero_jira_per_project_fallback_skipped_count",
        skipped_count,
    )
    for team_payload in team_payloads or []:
        if not isinstance(team_payload, dict):
            continue
        projects = team_payload.get("under_prd")
        if isinstance(projects, list):
            _sort_team_dashboard_under_prd_projects(projects)


def _remove_team_dashboard_zero_jira_pending_live_projects(team_payload: dict[str, Any]) -> None:
    pending_live = team_payload.get("pending_live")
    if not isinstance(pending_live, list):
        return
    team_payload["pending_live"] = [
        project
        for project in pending_live
        if isinstance(project, dict) and len(project.get("jira_tickets") or []) > 0
    ]


def _team_dashboard_project_fallback_emails(project: dict[str, Any]) -> list[str]:
    emails = _normalize_team_dashboard_emails(project.get("matched_pm_emails") or [])
    if not emails:
        regional_pm = str(project.get("regional_pm_pic") or "").strip()
        if "@" in regional_pm:
            emails = _normalize_team_dashboard_emails([regional_pm])
    return emails


def _team_dashboard_project_parent_payload(project: dict[str, Any]) -> dict[str, str]:
    return {
        "bpmis_id": str(project.get("bpmis_id") or "").strip(),
        "project_name": str(project.get("project_name") or "").strip(),
        "market": str(project.get("market") or "").strip(),
        "priority": str(project.get("priority") or "").strip(),
        "regional_pm_pic": str(project.get("regional_pm_pic") or "").strip(),
        "status": str(project.get("status") or "").strip(),
    }


def _normalize_team_dashboard_project_fallback_rows(
    rows: list[dict[str, Any]],
    project: dict[str, Any],
    emails: list[str],
) -> list[dict[str, Any]]:
    parent_project = _team_dashboard_project_parent_payload(project)
    allowed_emails = set(_normalize_team_dashboard_emails(emails))
    tickets: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        row_email = str(row.get("pm_email") or "").strip().lower()
        if allowed_emails and row_email and row_email not in allowed_emails:
            continue
        normalized = _normalize_team_dashboard_task(
            {
                **row,
                "pm_email": row.get("pm_email") or (next(iter(allowed_emails)) if allowed_emails else ""),
                "parent_project": row.get("parent_project") if isinstance(row.get("parent_project"), dict) else parent_project,
            }
        )
        status_key = str(normalized.get("jira_status") or "").strip().casefold()
        if status_key in TEAM_DASHBOARD_EXCLUDED_PENDING_STATUSES:
            continue
        dedupe_key = normalized.get("jira_id") or normalized.get("issue_id")
        if dedupe_key and dedupe_key in seen:
            continue
        if dedupe_key:
            seen.add(dedupe_key)
        tickets.append(normalized)
    tickets.sort(key=_team_dashboard_sort_key)
    return tickets


def _team_dashboard_project_fallback_jira_tasks(bpmis_client: Any, project: dict[str, Any]) -> list[dict[str, Any]]:
    bpmis_id = str(project.get("bpmis_id") or "").strip()
    if not bpmis_id or not hasattr(bpmis_client, "list_jira_tasks_for_project_created_by_email"):
        return []
    emails = _team_dashboard_project_fallback_emails(project)
    if not emails:
        return []
    tickets: list[dict[str, Any]] = []
    seen: set[str] = set()
    parent_project = _team_dashboard_project_parent_payload(project)
    for email in emails:
        try:
            rows = bpmis_client.list_jira_tasks_for_project_created_by_email(bpmis_id, email)
        except Exception as error:  # noqa: BLE001 - keep the project card renderable when fallback lookup fails.
            current_app.logger.warning("Team Dashboard Jira fallback lookup failed for %s/%s: %s", bpmis_id, email, error)
            continue
        for row in rows or []:
            if not isinstance(row, dict):
                continue
            normalized = _normalize_team_dashboard_task(
                {
                    **row,
                    "pm_email": row.get("pm_email") or email,
                    "parent_project": row.get("parent_project") if isinstance(row.get("parent_project"), dict) else parent_project,
                }
            )
            status_key = str(normalized.get("jira_status") or "").strip().casefold()
            if status_key in TEAM_DASHBOARD_EXCLUDED_PENDING_STATUSES:
                continue
            dedupe_key = normalized.get("jira_id") or normalized.get("issue_id")
            if dedupe_key and dedupe_key in seen:
                continue
            if dedupe_key:
                seen.add(dedupe_key)
            tickets.append(normalized)
    tickets.sort(key=_team_dashboard_sort_key)
    return tickets


def _team_dashboard_jira_max_pages() -> int:
    raw_value = str(os.getenv("TEAM_DASHBOARD_JIRA_MAX_PAGES") or "5").strip()
    try:
        return max(1, int(raw_value))
    except ValueError:
        return 5


def _team_dashboard_jira_release_after() -> str:
    configured = str(os.getenv("TEAM_DASHBOARD_JIRA_RELEASE_AFTER") or "").strip()
    if configured:
        return configured
    return time.strftime("%Y-%m-%d", time.localtime())


def _team_dashboard_monthly_report_jira_release_after() -> str:
    configured = str(os.getenv("TEAM_DASHBOARD_MONTHLY_REPORT_JIRA_RELEASE_AFTER") or "").strip()
    if configured:
        return configured
    return time.strftime("%Y-%m-%d", time.localtime(time.time() - 60 * 60 * 24 * 45))


def _normalize_team_dashboard_task(task: dict[str, Any]) -> dict[str, Any]:
    jira_id = str(task.get("jira_id") or task.get("ticket_key") or "").strip()
    issue_id = str(task.get("issue_id") or "").strip()
    jira_link = str(task.get("jira_link") or task.get("ticket_link") or "").strip()
    if not jira_link and jira_id:
        jira_link = f"{_jira_browse_base_url()}{jira_id}"
    raw_prd_links = task.get("prd_links")
    if not raw_prd_links and task.get("prd_link"):
        raw_prd_links = str(task.get("prd_link") or "").splitlines()
    prd_links = _team_dashboard_link_items(raw_prd_links)
    return {
        "issue_id": issue_id,
        "jira_id": jira_id or issue_id,
        "jira_link": jira_link,
        "jira_title": str(task.get("jira_title") or "").strip(),
        "pm_email": str(task.get("pm_email") or "").strip().lower(),
        "jira_status": str(task.get("jira_status") or task.get("status") or "").strip(),
        "created_at": str(task.get("created_at") or task.get("created") or "").strip(),
        "release_date": _format_team_dashboard_release_date(task.get("release_date") or task.get("release")),
        "version": str(task.get("version") or task.get("fix_version_name") or "").strip(),
        "description": str(task.get("description") or task.get("desc") or task.get("jiraDescription") or "").strip(),
        "prd_links": prd_links,
        "parent_project": _normalize_team_dashboard_project(task.get("parent_project") if isinstance(task.get("parent_project"), dict) else {}),
    }


def _normalize_team_dashboard_project(project: dict[str, Any]) -> dict[str, str]:
    bpmis_id = str(project.get("bpmis_id") or project.get("issue_id") or "").strip()
    matched_pm_emails = _normalize_team_dashboard_emails(project.get("matched_pm_emails") or [])
    normalized: dict[str, Any] = {
        "bpmis_id": bpmis_id,
        "project_name": str(project.get("project_name") or "").strip(),
        "market": str(project.get("market") or "").strip(),
        "priority": str(project.get("priority") or "").strip(),
        "regional_pm_pic": str(project.get("regional_pm_pic") or "").strip(),
        "status": str(project.get("status") or project.get("biz_project_status") or "").strip(),
    }
    if matched_pm_emails:
        normalized["matched_pm_emails"] = matched_pm_emails
    return normalized


def _split_team_dashboard_biz_projects_by_status(
    biz_projects: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    under_prd: list[dict[str, Any]] = []
    pending_live: list[dict[str, Any]] = []
    for raw_project in biz_projects:
        project = _normalize_team_dashboard_project(raw_project if isinstance(raw_project, dict) else {})
        status_key = str(project.get("status") or "").strip().casefold()
        if status_key in TEAM_DASHBOARD_UNDER_PRD_BIZ_PROJECT_STATUSES:
            under_prd.append(project)
        elif status_key in TEAM_DASHBOARD_PENDING_LIVE_BIZ_PROJECT_STATUSES:
            pending_live.append(project)
    return under_prd, pending_live


def _group_team_dashboard_tasks_by_project(
    tasks: list[dict[str, Any]],
    *,
    sort_by_release: bool = False,
) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for task in tasks:
        project = task.get("parent_project") if isinstance(task.get("parent_project"), dict) else {}
        project = _normalize_team_dashboard_project(project)
        key = project.get("bpmis_id") or "unknown"
        if key not in grouped:
            if key == "unknown":
                project = {
                    "bpmis_id": "",
                    "project_name": "BPMIS unavailable",
                    "market": "",
                    "priority": "",
                    "regional_pm_pic": "",
                }
            grouped[key] = {
                **project,
                "jira_tickets": [],
                "task_count": 0,
                "release_date": "-",
                "release_date_sort": "",
            }
        grouped[key]["jira_tickets"].append(task)
        grouped[key]["task_count"] = len(grouped[key]["jira_tickets"])

    projects = list(grouped.values())
    for project in projects:
        project["jira_tickets"].sort(key=_team_dashboard_sort_key)
        _apply_team_dashboard_project_release_date(project)
    if sort_by_release:
        projects.sort(key=_team_dashboard_project_release_sort_key)
    else:
        projects.sort(key=_team_dashboard_project_name_sort_key)
    for project in projects:
        project.pop("release_date_sort", None)
    return projects


def _merge_team_dashboard_biz_projects(
    projects: list[dict[str, Any]],
    biz_projects: list[dict[str, Any]],
    *,
    sort_by_release: bool = False,
) -> list[dict[str, Any]]:
    by_id: dict[str, dict[str, Any]] = {
        str(project.get("bpmis_id") or "").strip(): project
        for project in projects
        if str(project.get("bpmis_id") or "").strip()
    }
    merged = list(projects)
    for raw_project in biz_projects:
        project = _normalize_team_dashboard_project(raw_project if isinstance(raw_project, dict) else {})
        bpmis_id = project.get("bpmis_id")
        if not bpmis_id:
            continue
        existing = by_id.get(bpmis_id)
        if existing:
            for key in ("project_name", "market", "priority", "regional_pm_pic", "status"):
                if project.get(key) and not existing.get(key):
                    existing[key] = project[key]
            _merge_team_dashboard_project_pm_emails(existing, project.get("matched_pm_emails") or [])
            continue
        project.update(
            {
                "jira_tickets": [],
                "task_count": 0,
                "release_date": "-",
                "release_date_sort": "",
            }
        )
        merged.append(project)
        by_id[bpmis_id] = project
    if sort_by_release:
        merged.sort(key=_team_dashboard_project_release_sort_key)
    else:
        merged.sort(key=_team_dashboard_project_name_sort_key)
    for project in merged:
        project.pop("release_date_sort", None)
    return merged


def _merge_team_dashboard_project_pm_emails(project: dict[str, Any], emails: list[str]) -> None:
    existing = _normalize_team_dashboard_emails(project.get("matched_pm_emails") or [])
    for email in _normalize_team_dashboard_emails(emails):
        if email not in existing:
            existing.append(email)
    if existing:
        project["matched_pm_emails"] = existing


def _apply_team_dashboard_key_project_states(projects: list[dict[str, Any]], overrides: dict[str, Any]) -> None:
    for project in projects:
        _apply_team_dashboard_key_project_state(project, overrides)


def _apply_team_dashboard_key_project_state(project: dict[str, Any], overrides: dict[str, Any]) -> dict[str, Any]:
    bpmis_id = str(project.get("bpmis_id") or "").strip()
    override = overrides.get(bpmis_id) if bpmis_id and isinstance(overrides, dict) else None
    if isinstance(override, dict) and "is_key_project" in override:
        is_key_project = bool(override.get("is_key_project"))
        project["is_key_project"] = is_key_project
        project["key_project_source"] = "manual_on" if is_key_project else "manual_off"
        project["key_project_override"] = {
            "is_key_project": is_key_project,
            "updated_by": str(override.get("updated_by") or "").strip().lower(),
            "updated_at": str(override.get("updated_at") or "").strip(),
        }
        return project
    priority = str(project.get("priority") or "").strip().casefold()
    is_priority_default = priority in {"sp", "p0"}
    project["is_key_project"] = is_priority_default
    project["key_project_source"] = "priority_default" if is_priority_default else "none"
    project.pop("key_project_override", None)
    return project


def _apply_team_dashboard_project_release_date(project: dict[str, Any]) -> None:
    latest = None
    for task in project.get("jira_tickets") or []:
        parsed, _text = _parse_team_dashboard_release_date(task.get("release_date"))
        if parsed and (latest is None or parsed > latest):
            latest = parsed
    if latest:
        project["release_date"] = time.strftime("%Y-%m-%d", latest)
        project["release_date_sort"] = time.strftime("%Y-%m-%d", latest)
    else:
        project["release_date"] = "-"
        project["release_date_sort"] = ""


def _format_team_dashboard_release_date(value: Any) -> str:
    parsed, text = _parse_team_dashboard_release_date(value)
    if parsed:
        return time.strftime("%Y-%m-%d", parsed)
    return text


def _parse_team_dashboard_release_date(value: Any) -> tuple[time.struct_time | None, str]:
    text = str(value or "").strip()
    if not text:
        return None, ""
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return time.strptime(text[:10], pattern), text
        except ValueError:
            continue
    return None, text


def _team_dashboard_project_name_sort_key(project: dict[str, Any]) -> tuple[str, str]:
    return (
        str(project.get("project_name") or "").casefold(),
        str(project.get("bpmis_id") or "").casefold(),
    )


def _sort_team_dashboard_under_prd_projects(projects: list[dict[str, Any]]) -> None:
    projects.sort(key=_team_dashboard_under_prd_project_sort_key)


def _team_dashboard_under_prd_project_sort_key(project: dict[str, Any]) -> tuple[int, str, str, str]:
    release_sort = str(project.get("release_date_sort") or "").strip()
    if not release_sort:
        parsed, _text = _parse_team_dashboard_release_date(project.get("release_date"))
        if parsed:
            release_sort = time.strftime("%Y-%m-%d", parsed)
    jira_count = len(project.get("jira_tickets") or [])
    if release_sort:
        bucket = 0
    elif jira_count > 0:
        bucket = 1
    else:
        bucket = 2
    return (
        bucket,
        release_sort,
        str(project.get("project_name") or "").casefold(),
        str(project.get("bpmis_id") or "").casefold(),
    )


def _team_dashboard_project_release_sort_key(project: dict[str, Any]) -> tuple[int, str, str, str]:
    release_sort = str(project.get("release_date_sort") or "").strip()
    if not release_sort:
        parsed, _text = _parse_team_dashboard_release_date(project.get("release_date"))
        if parsed:
            release_sort = time.strftime("%Y-%m-%d", parsed)
    return (
        0 if release_sort else 1,
        release_sort,
        str(project.get("project_name") or "").casefold(),
        str(project.get("bpmis_id") or "").casefold(),
    )


def _team_dashboard_link_items(value: Any) -> list[dict[str, str]]:
    raw_links: list[str] = []
    if isinstance(value, list):
        for item in value:
            if isinstance(item, dict):
                raw_links.append(str(item.get("url") or item.get("label") or "").strip())
            else:
                raw_links.append(str(item or "").strip())
    elif isinstance(value, str):
        raw_links.extend(item.strip() for item in re.split(r"[\n,]+", value) if item.strip())
    deduped: list[dict[str, str]] = []
    seen: set[str] = set()
    for link in raw_links:
        if not link or link in seen:
            continue
        seen.add(link)
        deduped.append({"label": link, "url": link})
    return deduped


def _team_dashboard_sort_key(task: dict[str, Any]) -> tuple[str, str, str]:
    return (
        str(task.get("pm_email") or "").casefold(),
        str(task.get("version") or "").casefold(),
        str(task.get("jira_id") or "").casefold(),
    )


def _serialize_productization_version_candidate(row: dict[str, Any]) -> dict[str, str]:
    version_id = str(row.get("id") or row.get("versionId") or "").strip()
    version_name = (
        str(row.get("fullName") or row.get("name") or row.get("versionName") or row.get("label") or "").strip()
    )
    market = _coerce_display_text(row.get("marketId") or row.get("market") or row.get("country"))
    label = version_name
    if market:
        label = f"{version_name} · {market}"
    return {
        "version_id": version_id,
        "version_name": version_name,
        "market": market,
        "label": label,
    }


def _normalize_productization_issue_row(row: dict[str, Any]) -> dict[str, Any]:
    ticket_key = _extract_first_text(
        row,
        "jiraKey",
        "ticketKey",
        "jiraIssueKey",
        "issueKey",
        "key",
    )
    ticket_link = _normalize_productization_ticket_url(
        _extract_first_text(row, "jiraLink", "ticketLink", "jiraUrl", "url", "link")
    )
    if not ticket_key:
        ticket_key = _extract_issue_key_from_text(ticket_link)
    if not ticket_link and ticket_key:
        ticket_link = f"{_jira_browse_base_url()}{ticket_key}"

    return {
        "jira_ticket_number": ticket_key or "-",
        "jira_ticket_url": ticket_link or "",
        "feature_summary": _extract_first_text(row, "summary", "title", "jiraSummary") or "-",
        "detailed_feature": _format_productization_description_text(
            _extract_first_text(row, "desc", "description", "jiraDescription")
        ),
        "detailed_feature_source": "jira_description",
        "pm": _extract_person_display(
            _extract_first_value(row, "jiraRegionalPmPicId", "regionalPmPic", "productManager", "pm", "regionalPm")
        )
        or "-",
        "prd_links": _extract_link_values(
            _extract_first_value(row, "jiraPrdLink", "prdLink", "prdLinks", "prd", "brdLink")
        ),
    }


def _apply_codex_productization_detailed_features(
    normalized_items: list[dict[str, Any]],
    raw_rows: list[dict[str, Any]],
    *,
    settings: Settings,
) -> dict[str, Any]:
    if not normalized_items:
        return {
            "llm_description_generated": True,
            "llm_generated_count": 0,
            "codex_detailed_feature": True,
            "codex_generated_count": 0,
        }

    prompt_items = []
    for normalized, raw in zip(normalized_items, raw_rows):
        prompt_items.append(
            {
                "jira_ticket_number": str(normalized.get("jira_ticket_number") or "-"),
                "feature_summary": str(normalized.get("feature_summary") or "-"),
                "jira_description": _format_productization_description_text(
                    _extract_first_text(raw, "desc", "description", "jiraDescription")
                )[:6000],
            }
        )

    generated = _generate_productization_detailed_features_with_codex(prompt_items, settings=settings)
    generated_by_ticket = {
        str(item.get("jira_ticket_number") or "").strip(): str(item.get("detailed_feature") or "").strip()
        for item in generated
        if isinstance(item, dict)
    }

    generated_count = 0
    for item in normalized_items:
        ticket_number = str(item.get("jira_ticket_number") or "").strip()
        detailed_feature = generated_by_ticket.get(ticket_number, "")
        if detailed_feature:
            item["detailed_feature"] = detailed_feature
            item["detailed_feature_source"] = "codex"
            generated_count += 1
    return {
        "llm_description_generated": True,
        "llm_generated_count": generated_count,
        "codex_detailed_feature": True,
        "codex_generated_count": generated_count,
    }


def _generate_productization_detailed_features_with_codex(
    prompt_items: list[dict[str, str]],
    *,
    settings: Settings,
) -> list[dict[str, str]]:
    if _local_agent_source_code_qa_enabled(settings):
        return [
            {
                "jira_ticket_number": str(item.get("jira_ticket_number") or "").strip(),
                "detailed_feature": _clean_codex_productization_detailed_feature(str(item.get("detailed_feature") or "")),
            }
            for item in _build_local_agent_client(settings).productization_llm_descriptions(items=prompt_items)
            if isinstance(item, dict)
        ]
    return _generate_productization_detailed_features_with_local_codex(prompt_items, settings=settings)


def _generate_productization_detailed_features_with_local_codex(
    prompt_items: list[dict[str, str]],
    *,
    settings: Settings,
) -> list[dict[str, str]]:
    provider = CodexCliBridgeSourceCodeQALLMProvider(
        workspace_root=PROJECT_ROOT,
        timeout_seconds=settings.source_code_qa_codex_timeout_seconds,
        concurrency_limit=settings.source_code_qa_codex_concurrency,
        session_mode="ephemeral",
        codex_binary=os.getenv("SOURCE_CODE_QA_CODEX_BINARY") or None,
    )
    prompt = (
        "Generate English Detailed Feature text for each Jira ticket from its Jira Description.\n"
        "Rules:\n"
        "- Output strict JSON only, with shape: {\"items\":[{\"jira_ticket_number\":\"...\",\"detailed_feature\":\"...\"}]}.\n"
        "- Keep one item per input Jira ticket and preserve the jira_ticket_number exactly.\n"
        "- Write in clear product/engineering English.\n"
        "- Summarize the functional change and expected behavior, not implementation chatter.\n"
        "- If the description is empty or not meaningful, use \"-\".\n"
        "- Do not include Markdown fences, citations, explanations, or Chinese text.\n\n"
        f"Input JSON:\n{json.dumps({'items': prompt_items}, ensure_ascii=False)}"
    )
    result = provider.generate(
        payload={
            "systemInstruction": {"parts": [{"text": "You are a concise product feature summarizer."}]},
            "contents": [{"parts": [{"text": prompt}]}],
            "codex_prompt_mode": "productization_detailed_feature_v1",
        },
        primary_model=os.getenv("SOURCE_CODE_QA_CODEX_MODEL", "codex-cli"),
        fallback_model=os.getenv("SOURCE_CODE_QA_CODEX_MODEL", "codex-cli"),
    )
    text = provider.extract_text(result.payload)
    payload = _parse_codex_json_object(text)
    items = payload.get("items") if isinstance(payload, dict) else None
    if not isinstance(items, list):
        raise ToolError("Codex returned an invalid Detailed Feature payload.")
    return [
        {
            "jira_ticket_number": str(item.get("jira_ticket_number") or "").strip(),
            "detailed_feature": _clean_codex_productization_detailed_feature(str(item.get("detailed_feature") or "")),
        }
        for item in items
        if isinstance(item, dict)
    ]


def _parse_codex_json_object(text: str) -> dict[str, Any]:
    raw = str(text or "").strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.I)
        raw = re.sub(r"\s*```$", "", raw)
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError as error:
        start = raw.find("{")
        end = raw.rfind("}")
        if start < 0 or end <= start:
            raise ToolError("Codex returned unreadable Detailed Feature JSON.") from error
        try:
            payload = json.loads(raw[start : end + 1])
        except json.JSONDecodeError as nested_error:
            raise ToolError("Codex returned unreadable Detailed Feature JSON.") from nested_error
    if not isinstance(payload, dict):
        raise ToolError("Codex returned an invalid Detailed Feature payload.")
    return payload


def _clean_codex_productization_detailed_feature(value: str) -> str:
    text = _format_productization_description_text(value)
    if not text:
        return "-"
    text = re.sub(r"```(?:json)?|```", "", text, flags=re.I).strip()
    return text or "-"


def _filter_productization_issue_rows_for_pm_team(
    rows: list[dict[str, Any]],
    config_data: dict[str, Any],
    *,
    show_all_before_team_filtering: bool = False,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    allowed_components = _productization_allowed_components_for_pm_team(config_data)
    if show_all_before_team_filtering:
        return rows, {"team_filter_applied": False, "show_all_before_team_filtering": True}
    if not allowed_components:
        return rows, {"team_filter_applied": False, "show_all_before_team_filtering": False}
    filtered_rows = [row for row in rows if _productization_issue_matches_components(row, allowed_components)]
    return filtered_rows, {"team_filter_applied": True, "show_all_before_team_filtering": False}


def _productization_allowed_components_for_pm_team(config_data: dict[str, Any]) -> set[str]:
    pm_team = str(config_data.get("pm_team", "") or "").strip().upper()
    if pm_team == "AF":
        return {"dbp-anti-fraud", "anti-fraud"}
    return set()


def _productization_issue_matches_components(row: dict[str, Any], allowed_components: set[str]) -> bool:
    issue_components = _extract_productization_issue_components(row)
    return bool(issue_components and issue_components.intersection(allowed_components))


def _extract_productization_issue_components(row: dict[str, Any]) -> set[str]:
    raw_value = _extract_first_value(
        row,
        "componentId",
        "component",
        "components",
        "jiraComponent",
        "jiraComponentId",
    )
    flattened = _flatten_productization_component_values(raw_value)
    return {component.lower() for component in flattened if component}


def _flatten_productization_component_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        parts = [part.strip() for part in re.split(r"[;,/|]", text) if part.strip()]
        return parts or [text]
    if isinstance(value, dict):
        parts: list[str] = []
        for key in ("displayName", "name", "label", "value", "fullName", "id"):
            text = str(value.get(key) or "").strip()
            if text:
                parts.append(text)
        return parts
    if isinstance(value, list):
        parts: list[str] = []
        for item in value:
            parts.extend(_flatten_productization_component_values(item))
        return parts
    return [str(value).strip()]


def _normalize_productization_ticket_url(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text.startswith(("http://", "https://")):
        return text
    issue_key = _extract_issue_key_from_text(text)
    if issue_key:
        return f"{_jira_browse_base_url()}{issue_key}"
    return text


def _format_productization_description_text(value: str) -> str:
    if not str(value or "").strip():
        return "-"

    text = html.unescape(value)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"</p\s*>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    text = text.replace("\r", "\n")
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.split("\n")]
    chunks = [line for line in lines if line]
    return "\n".join(chunks).strip() if chunks else "-"


def _extract_first_value(row: dict[str, Any], *keys: str) -> Any:
    containers = [row]
    for nested_key in ("fields", "mapping", "data", "detail", "row"):
        nested = row.get(nested_key)
        if isinstance(nested, dict):
            containers.append(nested)

    for key in keys:
        lowered_key = key.lower()
        for container in containers:
            for candidate_key, value in container.items():
                if str(candidate_key).lower() == lowered_key:
                    return value
    return None


def _extract_first_text(row: dict[str, Any], *keys: str) -> str:
    value = _extract_first_value(row, *keys)
    return _coerce_display_text(value)


def _coerce_display_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, dict):
        for key in ("displayName", "name", "emailAddress", "label", "value", "fullName", "id"):
            text = str(value.get(key) or "").strip()
            if text:
                return text
        return ""
    if isinstance(value, list):
        parts = [_coerce_display_text(item) for item in value]
        parts = [part for part in parts if part]
        return ", ".join(parts)
    return str(value).strip()


def _extract_person_display(value: Any) -> str:
    if isinstance(value, list):
        people = [_extract_person_display(item) for item in value]
        people = [person for person in people if person]
        return ", ".join(people)
    if isinstance(value, dict):
        for key in ("displayName", "name", "emailAddress", "label", "username", "value"):
            text = str(value.get(key) or "").strip()
            if text:
                return text
        return ""
    return _coerce_display_text(value)


def _extract_link_values(value: Any) -> list[dict[str, str]]:
    links = _flatten_links(value)
    deduped: list[dict[str, str]] = []
    seen: set[str] = set()
    for link in links:
        if link in seen:
            continue
        seen.add(link)
        deduped.append({"label": link, "url": link})
    return deduped


def _flatten_links(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        links: list[str] = []
        for item in value:
            links.extend(_flatten_links(item))
        return links
    if isinstance(value, dict):
        links: list[str] = []
        for key in ("url", "link", "href", "value"):
            links.extend(_flatten_links(value.get(key)))
        return links
    text = str(value).strip()
    if not text:
        return []
    matches = re.findall(r"https?://[^\s,]+", text)
    if matches:
        return matches
    return [text] if text.startswith("http://") or text.startswith("https://") else []


def _extract_issue_key_from_text(value: str) -> str:
    match = re.search(r"\b([A-Z][A-Z0-9]+-\d+)\b", value or "")
    return match.group(1) if match else ""


def _jira_browse_base_url() -> str:
    return "https://jira.shopee.io/browse/"


def _get_user_identity(settings: Settings | None = None) -> dict[str, str | None]:
    profile = _current_google_profile()
    email = _current_google_email()
    name = str(profile.get("name") or "").strip()

    if email:
        return {
            "config_key": f"google:{email}",
            "display_name": name or email,
            "email": email,
            "mode": "google",
        }

    if settings and _shared_portal_enabled(settings):
        return {
            "config_key": None,
            "display_name": "Sign in with your NPT Google account",
            "email": None,
            "mode": "guest",
        }

    anonymous_key = session.get("anonymous_user_key")
    if not anonymous_key:
        anonymous_key = secrets.token_hex(8)
        session["anonymous_user_key"] = anonymous_key

    return {
        "config_key": f"anon:{anonymous_key}",
        "display_name": f"Anonymous Session {anonymous_key[:6]}",
        "email": None,
        "mode": "anonymous",
    }
