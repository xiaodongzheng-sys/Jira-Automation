#!/usr/bin/env python3
"""Generate Business Insights live report artifacts from Data Workbench.

This is an operational helper for the Mac-hosted live portal. It uses the
current Chrome Data Admin live session, runs aggregate SparkSQL sections in
Data Workbench, and writes Excel plus self-contained HTML visualization
artifacts into TEAM_PORTAL_DATA_DIR/business_insights.
"""
from __future__ import annotations

import argparse
import base64
from dataclasses import dataclass
from datetime import UTC, datetime
import html
import json
import os
from pathlib import Path
import re
import sqlite3
import subprocess
import sys
import time
from typing import Any, Callable
import uuid
from zoneinfo import ZoneInfo

import requests
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from bpmis_jira_tool.timefmt import format_gmt8  # noqa: E402
from bpmis_jira_tool.business_insights import (  # noqa: E402
    AF_BLACK_WHITE_LIST_TABLE,
    AF_CARD_3DS_REPORT_ID,
    AF_DEVICE_RISK_REPORT_ID,
    AF_FACIAL_VERIFICATION_REPORT_ID,
    AF_FACIAL_VERIFICATION_TABLE,
    AF_FRAUD_LOSS_REPORT_ID,
    AF_LIST_USAGE_REPORT_ID,
    AF_REQUEST_STATISTIC_TABLE,
    AF_REVIEW_CASE_TABLE,
    AF_RULE_CONFIG_TABLE,
    AF_RULE_EFFECTIVENESS_REPORT_ID,
    AF_RULES_FEATURES_REPORT_ID,
    AF_SCENARIO_FLOW_CONFIG_TABLE,
    AF_SCENARIOS_ACTIONS_REPORT_ID,
    APPLICATION_DISBURSEMENT_FUNNEL_REPORT_ID,
    CREDIT_LIMIT_TABLE,
    LIMIT_UTILIZATION_REPORT_ID,
    LOAN_APPLICATION_TABLE,
    PRODUCT_LABELS,
    PRODUCT_LABEL_COLUMNS,
    PORTFOLIO_REPAYMENT_REPORT_ID,
    REPAY_PLAN_TABLE,
    UNDERWRITING_FUNNEL_REPORT_ID,
    build_af_card_3ds_sql,
    build_af_device_risk_sql,
    build_af_list_usage_sql,
    build_af_facial_verification_sql,
    build_af_fraud_loss_sql,
    build_af_rule_effectiveness_sql,
    build_af_rules_features_sql,
    build_af_scenarios_actions_sql,
    build_application_disbursement_funnel_sql,
    build_limit_utilization_sql,
    build_portfolio_repayment_sql,
    product_label,
)

DATA_ADMIN_BASE_URL = "https://data-admin.ph.seabank.io"
DATA_ADMIN_TEAM_ID = "T0300036"
LIVE_HOST_ROOT = Path("/Users/NPTSG0388/Workspace/jira-creation-stack-host")
DEFAULT_PORTAL_DATA_DIR = LIVE_HOST_ROOT / ".team-portal"
REPORT_BUILDERS: dict[str, tuple[str, Callable[..., str]]] = {
    PORTFOLIO_REPAYMENT_REPORT_ID: ("Credit Risk PH - Portfolio Repayment", build_portfolio_repayment_sql),
    LIMIT_UTILIZATION_REPORT_ID: ("Credit Risk PH - Limit Utilization", build_limit_utilization_sql),
    APPLICATION_DISBURSEMENT_FUNNEL_REPORT_ID: (
        "Credit Risk PH - Application to Disbursement Funnel",
        build_application_disbursement_funnel_sql,
    ),
    AF_SCENARIOS_ACTIONS_REPORT_ID: (
        "Anti-fraud PH - L1+L2 Scenarios, Actions & Auth Steps",
        build_af_scenarios_actions_sql,
    ),
    AF_RULES_FEATURES_REPORT_ID: (
        "Anti-fraud PH - Rules & Features",
        build_af_rules_features_sql,
    ),
    AF_RULE_EFFECTIVENESS_REPORT_ID: (
        "Anti-fraud PH - Rule Effectiveness / Hit-Rate",
        build_af_rule_effectiveness_sql,
    ),
    AF_FRAUD_LOSS_REPORT_ID: (
        "Anti-fraud PH - Fraud Loss & Case Outcomes",
        build_af_fraud_loss_sql,
    ),
    AF_FACIAL_VERIFICATION_REPORT_ID: (
        "Anti-fraud PH - Facial Verification / Liveness & Deepfake",
        build_af_facial_verification_sql,
    ),
    AF_DEVICE_RISK_REPORT_ID: (
        "Anti-fraud PH - Device & Identity Risk",
        build_af_device_risk_sql,
    ),
    AF_CARD_3DS_REPORT_ID: (
        "Anti-fraud PH - Card Fraud & 3DS Authentication",
        build_af_card_3ds_sql,
    ),
    AF_LIST_USAGE_REPORT_ID: (
        "Anti-fraud PH - Blacklist, Whitelist & Greylist",
        build_af_list_usage_sql,
    ),
}

# Anchor table per report used to resolve the latest available pt_date when the
# snapshot is "latest". Tables in a report share a daily snapshot, so the anchor's
# max(pt_date) is used to pin every section to the same snapshot.
REPORT_SNAPSHOT_ANCHOR_TABLE: dict[str, str] = {
    PORTFOLIO_REPAYMENT_REPORT_ID: REPAY_PLAN_TABLE,
    LIMIT_UTILIZATION_REPORT_ID: CREDIT_LIMIT_TABLE,
    APPLICATION_DISBURSEMENT_FUNNEL_REPORT_ID: LOAN_APPLICATION_TABLE,
    AF_SCENARIOS_ACTIONS_REPORT_ID: AF_SCENARIO_FLOW_CONFIG_TABLE,
    AF_RULES_FEATURES_REPORT_ID: AF_RULE_CONFIG_TABLE,
    AF_RULE_EFFECTIVENESS_REPORT_ID: AF_REQUEST_STATISTIC_TABLE,
    AF_FRAUD_LOSS_REPORT_ID: AF_REVIEW_CASE_TABLE,
    AF_FACIAL_VERIFICATION_REPORT_ID: AF_FACIAL_VERIFICATION_TABLE,
    AF_LIST_USAGE_REPORT_ID: AF_BLACK_WHITE_LIST_TABLE,
}

LATEST_SNAPSHOT = "latest"

PREFERRED_PRODUCT_CODES: dict[str, str] = {
    "Credit Card": "812F",
}
PREFERRED_SUB_PRODUCT_CODES: dict[str, str] = {
    "Employee Loan": "108",
}

# ISO 18245 Merchant Category Codes - common ones, used to render "code - name" in the 3DS MCC table.
# Unknown codes fall back to the bare code (see _mcc_label); 3000-3999 are carrier/rental/lodging ranges.
MCC_LABELS: dict[str, str] = {
    "4111": "Local / commuter transport", "4121": "Taxicabs & rideshare", "4131": "Bus lines",
    "4214": "Freight & courier", "4411": "Cruise lines", "4511": "Airlines", "4722": "Travel agencies",
    "4784": "Tolls & bridge fees", "4789": "Transportation services", "4812": "Telecom equipment",
    "4814": "Telecom services", "4816": "Computer network / information services", "4829": "Money transfer",
    "4899": "Cable & streaming services", "4900": "Utilities (electric, gas, water)",
    "5111": "Office supplies & printing", "5172": "Petroleum products", "5200": "Home supply stores",
    "5300": "Wholesale clubs", "5310": "Discount stores", "5311": "Department stores",
    "5331": "Variety stores", "5399": "General merchandise", "5411": "Grocery & supermarkets",
    "5412": "Convenience stores", "5422": "Meat / freezer provisioners", "5441": "Candy & confectionery",
    "5451": "Dairy products", "5462": "Bakeries", "5499": "Specialty food stores",
    "5511": "Car dealers (new & used)", "5541": "Service stations (fuel)", "5542": "Automated fuel dispensers",
    "5611": "Men's clothing", "5621": "Women's clothing", "5631": "Women's accessories",
    "5641": "Children's clothing", "5651": "Family clothing", "5661": "Shoe stores",
    "5691": "Clothing stores", "5712": "Furniture & furnishings", "5722": "Household appliances",
    "5732": "Electronics stores", "5733": "Music & instrument stores", "5734": "Computer software stores",
    "5735": "Record / media stores", "5811": "Caterers", "5812": "Restaurants & eating places",
    "5813": "Bars, lounges & nightclubs", "5814": "Fast food", "5815": "Digital goods - media/books",
    "5816": "Digital goods - games", "5817": "Digital goods - apps", "5818": "Digital goods - large merchant",
    "5912": "Pharmacies & drug stores", "5921": "Liquor stores", "5941": "Sporting goods",
    "5942": "Book stores", "5944": "Jewelry & watches", "5945": "Toy & game stores",
    "5946": "Camera & photo supply", "5947": "Gift & novelty shops", "5964": "Direct marketing - catalog",
    "5965": "Direct marketing - retail", "5967": "Direct marketing - inbound telemarketing",
    "5968": "Direct marketing - subscriptions", "5969": "Direct marketing - other",
    "5999": "Miscellaneous retail", "6010": "Manual cash disbursement", "6011": "ATM cash withdrawal",
    "6012": "Financial institution merchandise", "6051": "Quasi-cash / crypto / wallets",
    "6211": "Securities & brokers", "6300": "Insurance", "6513": "Real estate agents & rentals",
    "6540": "Prepaid / wallet top-up (non-financial)", "7011": "Lodging & hotels",
    "7032": "Recreational camps", "7210": "Laundry & cleaning", "7230": "Beauty & barber shops",
    "7273": "Dating & escort services", "7298": "Health & beauty spas", "7299": "Personal services",
    "7311": "Advertising services", "7349": "Cleaning & maintenance", "7372": "Computer programming / IT",
    "7392": "Management & consulting", "7399": "Business services", "7512": "Car rental agencies",
    "7523": "Parking lots & garages", "7832": "Cinemas", "7841": "Video rental",
    "7922": "Theatrical & event tickets", "7929": "Bands & entertainers", "7941": "Sports clubs & promoters",
    "7991": "Tourist attractions", "7994": "Video game arcades", "7995": "Betting & gambling",
    "7997": "Membership clubs (sports/golf)", "7999": "Recreation services", "8011": "Doctors & physicians",
    "8021": "Dentists", "8062": "Hospitals", "8071": "Medical & dental labs", "8099": "Health services",
    "8211": "Schools (elementary/secondary)", "8220": "Colleges & universities", "8241": "Correspondence schools",
    "8299": "Educational services", "8398": "Charitable & social orgs", "8651": "Political organizations",
    "8999": "Professional services", "9211": "Court costs & fines", "9222": "Government fines",
    "9311": "Tax payments", "9399": "Government services", "9402": "Postal services",
    "9405": "Government - intra-government",
}
@dataclass(frozen=True)
class WorkbenchSection:
    sheet_name: str
    query: str


def _decode_jwt_payload(token: str) -> dict[str, Any]:
    try:
        payload = token.split(".")[1]
        payload += "=" * ((4 - len(payload) % 4) % 4)
        decoded = base64.urlsafe_b64decode(payload.encode("ascii"))
        data = json.loads(decoded)
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def _chrome_safe_storage_secret() -> bytes:
    return subprocess.check_output(
        ["security", "find-generic-password", "-w", "-s", "Chrome Safe Storage"],
        stderr=subprocess.DEVNULL,
    ).strip()


def _chrome_cookie_key() -> bytes:
    return PBKDF2HMAC(
        algorithm=hashes.SHA1(),
        length=16,
        salt=b"saltysalt",
        iterations=1003,
        backend=default_backend(),
    ).derive(_chrome_safe_storage_secret())


def _decrypt_chrome_cookie(encrypted_value: bytes, key: bytes) -> str:
    if not encrypted_value.startswith(b"v10"):
        raise RuntimeError("Unsupported Chrome cookie encryption format.")
    decryptor = Cipher(algorithms.AES(key), modes.CBC(b" " * 16), backend=default_backend()).decryptor()
    plaintext = decryptor.update(encrypted_value[3:])
    pad = plaintext[-1]
    if 1 <= pad <= 16:
        plaintext = plaintext[:-pad]
    # Chrome prefixes v10 cookie plaintext with a 32-byte host digest.
    return plaintext[32:].decode("utf-8")


def load_data_admin_token(*, chrome_profile: str = "Default") -> str:
    cookies_path = Path.home() / "Library/Application Support/Google/Chrome" / chrome_profile / "Cookies"
    if not cookies_path.exists():
        raise RuntimeError(f"Chrome cookie DB not found: {cookies_path}")
    key = _chrome_cookie_key()
    connection = sqlite3.connect(cookies_path)
    try:
        row = connection.execute(
            """
            select encrypted_value
            from cookies
            where host_key = 'data-admin.ph.seabank.io'
              and name = 'bank-admin-token'
            order by expires_utc desc
            limit 1
            """
        ).fetchone()
    finally:
        connection.close()
    if not row:
        raise RuntimeError("Data Admin live bank-admin-token cookie was not found.")
    token = _decrypt_chrome_cookie(bytes(row[0]), key)
    if not token.startswith("eyJ"):
        raise RuntimeError("Decrypted Data Admin token is not a JWT.")
    return token


def build_data_admin_session(token: str) -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "portal-request": "true",
            "cache": "no-store",
            "bank-admin-token": token,
            "dual-active-mark": "live",
            "content-type": "application/json",
            "portal-permission-code": "workbench_process",
        }
    )
    session.cookies.update({"bank-admin-token": token, "dual-active-mark": "live"})
    return session


def validate_data_admin_session(session: requests.Session) -> str:
    response = session.get(f"{DATA_ADMIN_BASE_URL}/api/session/session/v1/userInfo", timeout=10)
    if response.status_code != 200:
        raise RuntimeError(f"Data Admin session is not valid: HTTP {response.status_code} {response.text[:160]}")
    payload = response.json()
    if payload.get("code") != 0:
        raise RuntimeError(f"Data Admin userInfo failed: {payload}")
    data = payload.get("data") if isinstance(payload.get("data"), dict) else {}
    return str(data.get("email") or data.get("userId") or "")


def extract_sql_sections(sql: str) -> list[WorkbenchSection]:
    matches = list(re.finditer(r"^--\s+\d+\.\s+(.+?)\s*$", sql, flags=re.M))
    sections: list[WorkbenchSection] = []
    for index, match in enumerate(matches):
        sheet_name = match.group(1).strip()[:31]
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(sql)
        query = sql[start:end].strip()
        query = re.sub(r"^--.*?$", "", query, flags=re.M).strip().rstrip(";").strip()
        if query:
            sections.append(WorkbenchSection(sheet_name=sheet_name, query=query))
    return sections


def resolve_snapshot_pt_date(
    session: requests.Session,
    report_id: str,
    *,
    poll_seconds: int,
    max_polls: int,
) -> str | None:
    """Return the latest available pt_date for the report's anchor table.

    Returns None when no anchor is configured; callers then let each SQL section
    fall back to its own ``max(pt_date)`` subquery.
    """
    table = REPORT_SNAPSHOT_ANCHOR_TABLE.get(report_id)
    if not table:
        return None
    _schema, rows, _execution_id = run_workbench_query(
        session,
        WorkbenchSection(sheet_name="resolve snapshot", query=f"select max(pt_date) as pt_date from {table}"),
        poll_seconds=poll_seconds,
        max_polls=max_polls,
    )
    if rows and rows[0] and rows[0][0]:
        return str(rows[0][0]).strip()
    return None


def run_workbench_query(
    session: requests.Session,
    section: WorkbenchSection,
    *,
    poll_seconds: int,
    max_polls: int,
) -> tuple[list[str], list[list[Any]], str]:
    run_response = session.post(
        f"{DATA_ADMIN_BASE_URL}/api/workbench-service/adhoc/run",
        json={
            "engine": "SparkSQL",
            "script": section.query,
            "teamId": DATA_ADMIN_TEAM_ID,
            "parameterList": [],
            "sparkParam": "",
        },
        timeout=30,
    )
    run_response.raise_for_status()
    run_payload = run_response.json()
    if run_payload.get("code") != 0:
        raise RuntimeError(f"{section.sheet_name}: Data Workbench run failed: {run_payload}")
    data = run_payload.get("data") if isinstance(run_payload.get("data"), dict) else {}
    task = (data.get("taskList") or [{}])[0]
    execution_id = str(task.get("executionId") or data.get("runId") or "")
    if not execution_id:
        raise RuntimeError(f"{section.sheet_name}: Data Workbench did not return executionId.")
    print(f"{section.sheet_name}: execution_id={execution_id}", flush=True)

    final_status = ""
    last_log = ""
    for _attempt in range(max_polls):
        log_response = session.post(
            f"{DATA_ADMIN_BASE_URL}/api/workbench-service/adhoc/log",
            json={"executionId": execution_id, "startOffset": 0},
            timeout=60,
        )
        if log_response.status_code == 401:
            raise RuntimeError(f"{section.sheet_name}: Data Admin session became invalid while polling.")
        log_response.raise_for_status()
        log_payload = log_response.json()
        log_data = log_payload.get("data") if isinstance(log_payload.get("data"), dict) else {}
        final_status = str(log_data.get("status") or "")
        last_log = str(log_data.get("log") or log_response.text)
        if final_status in {"SUCCESS", "FAILED", "CANCELED"}:
            break
        time.sleep(poll_seconds)
    if final_status != "SUCCESS":
        raise RuntimeError(f"{section.sheet_name}: Data Workbench ended with {final_status}: {last_log[-1200:]}")

    result_response = session.post(
        f"{DATA_ADMIN_BASE_URL}/api/workbench-service/adhoc/result",
        json={"executionId": execution_id},
        timeout=90,
    )
    if result_response.status_code == 401:
        raise RuntimeError(f"{section.sheet_name}: Data Admin session became invalid while fetching result.")
    result_response.raise_for_status()
    result_payload = result_response.json()
    if result_payload.get("code") != 0:
        raise RuntimeError(f"{section.sheet_name}: Data Workbench result failed: {result_payload}")
    table_data = (result_payload.get("data") or {}).get("tableData") or {}
    schema = [str(item) for item in (table_data.get("schema") or [])]
    rows = table_data.get("data") or []
    print(f"{section.sheet_name}: rows={len(rows)}", flush=True)
    return schema, rows, execution_id


def style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    header_font = Font(bold=True, color="1F2937")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, column_cells in enumerate(sheet.columns, start=1):
        width = 12
        for cell in column_cells:
            width = max(width, min(len(str(cell.value or "")) + 2, 44))
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _excel_sheet_title(sheet_name: str) -> str:
    # openpyxl rejects : \ / ? * [ ] in sheet titles, which are capped at 31 chars.
    return re.sub(r"[:\\/?*\[\]]", " ", str(sheet_name)).strip()[:31] or "Sheet"


# Canonical sheet names that _excel_sheet_title mangles (slashes become spaces). Used to restore the
# original names when sheets are read back from Excel, so write_visualization's sheet lookups keep
# matching on --refresh-visualizations runs.
_CANONICAL_SHEET_NAMES = (
    "Scene/Sub-scene/Action Usage",
    "Daily Challenge/Reject/Punish",
    "Review Pool / Backlog (current)",
)
_SHEET_TITLE_TO_CANONICAL = {_excel_sheet_title(name): name for name in _CANONICAL_SHEET_NAMES}


def _canonical_sheet_name(excel_title: str) -> str:
    return _SHEET_TITLE_TO_CANONICAL.get(str(excel_title), str(excel_title))


def _typed_cell(value: Any) -> Any:
    """Coerce Data Workbench string output to a real number when the round-trip is lossless, so the
    Excel artifact gets numeric cells (sortable/summable) instead of text. Leading-zero codes
    ('041…', '01') and long identifiers (>15 digits, float precision) stay strings."""
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not re.fullmatch(r"-?\d+(\.\d+)?", text):
        return value
    if "." not in text:
        if len(text.lstrip("-")) > 15:
            return value
        coerced = int(text)
        return coerced if str(coerced) == text else value
    coerced = float(text)
    return coerced if repr(coerced) == text else value


def write_workbook(path: Path, sheets: list[tuple[str, list[str], list[list[Any]]]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, schema, rows in sheets:
        sheet = workbook.create_sheet(_excel_sheet_title(sheet_name))
        sheet.append(schema)
        for row in rows:
            sheet.append([_typed_cell(value) for value in row])
        style_sheet(sheet)
    workbook.save(path)


def normalize_product_labels(
    sheets: list[tuple[str, list[str], list[list[Any]]]],
    *,
    preserve_raw_export: bool = True,
) -> list[tuple[str, list[str], list[list[Any]]]]:
    normalized_sheets: list[tuple[str, list[str], list[list[Any]]]] = []
    for sheet_name, headers, rows in sheets:
        if preserve_raw_export and sheet_name == "Raw Export":
            normalized_sheets.append((sheet_name, headers, rows))
            continue
        product_offsets = {
            offset
            for offset, header in enumerate(headers)
            if str(header).strip().lower() in PRODUCT_LABEL_COLUMNS
        }
        if not product_offsets:
            normalized_sheets.append((sheet_name, headers, rows))
            continue
        normalized_rows = []
        for row in rows:
            normalized_row = list(row)
            for offset in product_offsets:
                if offset < len(normalized_row):
                    normalized_row[offset] = product_label(normalized_row[offset])
            normalized_rows.append(normalized_row)
        normalized_sheets.append((sheet_name, headers, normalized_rows))
    return normalized_sheets


def _number(value: Any) -> float:
    try:
        if value in (None, ""):
            return 0.0
        return float(str(value).replace(",", ""))
    except Exception:
        return 0.0


def _is_number(value: Any) -> bool:
    if isinstance(value, bool) or value in (None, ""):
        return False
    if isinstance(value, (int, float)):
        return True
    try:
        float(str(value).replace(",", ""))
    except Exception:
        return False
    return True


def _format_number(value: Any, *, suffix: str = "") -> str:
    if not _is_number(value):
        return str(value or "")
    number = _number(value)
    sign = "-" if number < 0 else ""
    absolute = abs(number)
    if absolute >= 1_000_000_000_000:
        text = f"{sign}{absolute / 1_000_000_000_000:.2f}T"
    elif absolute >= 1_000_000_000:
        text = f"{sign}{absolute / 1_000_000_000:.2f}B"
    elif absolute >= 1_000_000:
        text = f"{sign}{absolute / 1_000_000:.2f}M"
    elif absolute >= 10_000:
        text = f"{sign}{absolute:,.0f}"
    elif absolute == int(absolute):
        text = f"{sign}{absolute:,.0f}"
    else:
        text = f"{sign}{absolute:,.2f}"
    return f"{text}{suffix}"


def _preferred_code_for_label(label: str, *, header: str = "") -> str:
    normalized_header = header.strip().lower()
    if normalized_header in {"sub-product", "sub_product_code"} and label in PREFERRED_SUB_PRODUCT_CODES:
        return PREFERRED_SUB_PRODUCT_CODES[label]
    if normalized_header in {"product", "product_code"} and label in PREFERRED_PRODUCT_CODES:
        return PREFERRED_PRODUCT_CODES[label]
    candidates = [code for code, mapped in PRODUCT_LABELS.items() if mapped == label]
    if not candidates:
        return ""
    if normalized_header in {"product", "product_code"}:
        product_candidates = [code for code in candidates if code.upper().startswith("8")]
        if product_candidates:
            return sorted(product_candidates)[0]
    if normalized_header in {"sub-product", "sub_product_code"}:
        subproduct_candidates = [code for code in candidates if not code.upper().startswith("8")]
        if subproduct_candidates:
            return sorted(subproduct_candidates)[0]
    return sorted(candidates)[0]


def _product_display_label(product: Any, *, header: str = "product") -> str:
    raw = str(product or "").strip()
    if not raw:
        return ""
    if " - " in raw:
        return raw
    label = product_label(raw)
    if label != raw:
        return f"{raw} - {label}"
    preferred_code = _preferred_code_for_label(raw, header=header)
    return f"{preferred_code} - {raw}" if preferred_code else raw


def _format_cell(header: str, value: Any) -> str:
    lowered = header.lower()
    if lowered in PRODUCT_LABEL_COLUMNS:
        return _product_display_label(value, header=lowered)
    if lowered.endswith("rate") or lowered.startswith("%") or "% " in lowered:
        number = _number(value)
        if value in (None, ""):
            return ""
        if abs(number) <= 1:
            number *= 100
        return f"{number:.1f}%"
    return _format_number(value) if _is_number(value) else str(value or "")


def _mcc_label(code: str) -> str:
    """Human-readable name for an ISO 18245 MCC, or '' when unknown."""
    name = MCC_LABELS.get(code)
    if name:
        return name
    try:
        n = int(code)
    except (TypeError, ValueError):
        return ""
    if 3000 <= n <= 3299:
        return "Airline"
    if 3300 <= n <= 3499:
        return "Car rental"
    if 3500 <= n <= 3999:
        return "Lodging / hotel"
    return ""


def _format_mcc_cell(value: Any) -> str:
    """Render an MCC as 'code - name', stripping any thousands separator / trailing .0 from the code."""
    raw = str(value if value is not None else "").strip()
    if not raw or raw == "(none)":
        return raw or "(none)"
    code = raw.replace(",", "")
    if code.endswith(".0") and code[:-2].isdigit():
        code = code[:-2]
    name = _mcc_label(code)
    return f"{code} - {name}" if name else code


def _product_data_attr(product: Any, *, header: str = "product") -> str:
    label = _product_display_label(product, header=header)
    return f' data-product="{html.escape(label, quote=True)}"' if label else ""


def _product_filter_options(sheets: list[tuple[str, list[str], list[list[Any]]]]) -> list[str]:
    products: set[str] = set()
    for sheet_name, headers, rows in sheets:
        if sheet_name == "Raw Export":
            continue
        product_offsets = [
            offset
            for offset, header in enumerate(headers)
            if str(header).strip().lower() in {"product", "product_code"}
        ]
        for row in rows:
            for offset in product_offsets:
                if offset < len(row):
                    products.add(_product_display_label(row[offset], header="product"))
    return sorted(product for product in products if product and product != "UNKNOWN")


def _product_filter_html(products: list[str]) -> str:
    if len(products) < 2:
        return ""
    options = ['<option value="">All products</option>']
    options.extend(
        f'<option value="{html.escape(product, quote=True)}">{html.escape(product)}</option>'
        for product in products
    )
    return (
        '<section class="filter-card wide">'
        '<div><p class="eyebrow">View Controls</p><h2>Product Filter</h2>'
        '<p>Filters product-level charts and tables. All-product visuals are hidden when a single product is selected.</p></div>'
        f'<label><span>Product</span><select data-product-filter>{"".join(options)}</select></label>'
        "</section>"
    )


def _table_html(headers: list[str], rows: list[list[Any]], *, page_size: int = 50) -> str:
    header_html = "".join(f"<th>{html.escape(str(header))}</th>" for header in headers)
    product_offset = next(
        (
            offset
            for offset, header in enumerate(headers)
            if str(header).strip().lower() in {"product", "product_code"}
        ),
        None,
    )
    body_rows = []
    for row in rows:
        cells = []
        for index, header in enumerate(headers):
            value = row[index] if index < len(row) else ""
            class_name = "num" if _is_number(value) else ""
            cells.append(f'<td class="{class_name}">{html.escape(_format_cell(str(header), value))}</td>')
        product_attr = _product_data_attr(row[product_offset], header=str(headers[product_offset]).strip().lower()) if product_offset is not None and product_offset < len(row) else ""
        body_rows.append(f"<tr{product_attr}>" + "".join(cells) + "</tr>")
    controls = (
        '<div class="table-pagination" data-table-pagination>'
        '<button type="button" data-page-prev>Previous</button>'
        '<span data-page-info></span>'
        '<button type="button" data-page-next>Next</button>'
        '</div>'
        if len(rows) > page_size
        else ""
    )
    return (
        f'<div class="table-wrap"><table class="bi-table" data-page-size="{page_size}">'
        f'<thead><tr>{header_html}</tr></thead><tbody>{"".join(body_rows)}</tbody></table></div>{controls}'
    )


def _bar_chart(
    title: str,
    labels: list[Any],
    values: list[Any],
    *,
    value_suffix: str = "",
    labels_are_products: bool = False,
) -> str:
    pairs = [
        (_product_display_label(label, header="product") if labels_are_products else str(label or "UNKNOWN"), _number(value))
        for label, value in zip(labels, values, strict=False)
    ]
    pairs = [item for item in pairs if item[1] != 0][:12]
    if not pairs:
        return ""
    maximum = max(value for _label, value in pairs) or 1
    rows = []
    for label, value in pairs:
        width = max(2.0, value / maximum * 100.0)
        product_attr = _product_data_attr(label) if labels_are_products else ""
        rows.append(
            f'<div class="bar-row"{product_attr}>'
            f"<span>{html.escape(label)}</span>"
            f'<div class="bar-track"><div class="bar" style="width:{width:.1f}%"></div></div>'
            f"<b>{html.escape(_format_number(value, suffix=value_suffix))}</b>"
            "</div>"
        )
    scope = ' data-product-visual="1"' if labels_are_products else ' data-global-visual="1"'
    return f'<section class="panel"{scope}><h2>{html.escape(title)}</h2>{"".join(rows)}</section>'


def _group_sum(headers: list[str], rows: list[list[Any]], label_column: str, value_column: str) -> list[tuple[str, float]]:
    index = _sheet_index(headers)
    if label_column not in index or value_column not in index:
        return []
    grouped: dict[str, float] = {}
    label_offset = index[label_column]
    value_offset = index[value_column]
    for row in rows:
        label = str(row[label_offset] if label_offset < len(row) and row[label_offset] not in (None, "") else "UNKNOWN")
        if label_column.lower() in {"product", "product_code", "sub-product", "sub_product_code"}:
            label = _product_display_label(label, header=label_column.lower())
        value = _number(row[value_offset] if value_offset < len(row) else None)
        grouped[label] = grouped.get(label, 0.0) + value
    return sorted(grouped.items(), key=lambda item: item[1], reverse=True)


def _donut_chart(title: str, values: list[tuple[str, float, str]]) -> str:
    filtered = [(label, value, color) for label, value, color in values if value > 0]
    total = sum(value for _label, value, _color in filtered)
    if total <= 0:
        return ""
    start = 0.0
    segments = []
    legend = []
    for label, value, color in filtered:
        end = start + value / total * 100.0
        segments.append(f"{color} {start:.2f}% {end:.2f}%")
        legend.append(
            '<div class="legend-row">'
            f'<span class="legend-dot" style="background:{html.escape(color)}"></span>'
            f"<span>{html.escape(label)}</span><b>{html.escape(_format_number(value))}</b>"
            "</div>"
        )
        start = end
    return (
        f'<section class="panel" data-global-visual="1"><h2>{html.escape(title)}</h2>'
        '<div class="donut-layout">'
        f'<div class="donut" style="background:conic-gradient({",".join(segments)})"><span>{html.escape(_format_number(total))}</span></div>'
        f'<div class="legend">{"".join(legend)}</div>'
        "</div></section>"
    )


def _stacked_bar_chart(
    title: str,
    labels: list[str],
    series: list[tuple[str, list[float], str]],
    *,
    value_suffix: str = "",
    labels_are_products: bool = False,
) -> str:
    if not labels or not series:
        return ""
    rows = []
    for row_index, label in enumerate(labels[:12]):
        display_label = _product_display_label(label, header="product") if labels_are_products else label
        total = sum(values[row_index] for _name, values, _color in series if row_index < len(values))
        if total <= 0:
            continue
        segments = []
        for name, values, color in series:
            value = values[row_index] if row_index < len(values) else 0.0
            if value <= 0:
                continue
            width = max(2.0, value / total * 100.0)
            segments.append(
                f'<span class="stack-segment" title="{html.escape(name)}: {html.escape(_format_number(value, suffix=value_suffix))}" '
                f'style="width:{width:.2f}%;background:{html.escape(color)}"></span>'
            )
        product_attr = _product_data_attr(display_label, header="product") if labels_are_products else ""
        rows.append(
            f'<div class="stack-row"{product_attr}>'
            f"<span>{html.escape(display_label)}</span>"
            f'<div class="stack-track">{"".join(segments)}</div>'
            f"<b>{html.escape(_format_number(total, suffix=value_suffix))}</b>"
            "</div>"
        )
    legend = "".join(
        f'<span class="stack-legend"><i style="background:{html.escape(color)}"></i>{html.escape(name)}</span>'
        for name, _values, color in series
    )
    scope = ' data-product-visual="1"' if labels_are_products else ' data-global-visual="1"'
    return f'<section class="panel"{scope}><h2>{html.escape(title)}</h2><div class="stack-legend-wrap">{legend}</div>{"".join(rows)}</section>' if rows else ""


def _heatmap_table(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    row_column: str,
    column_column: str,
    value_column: str,
    max_columns: int = 8,
    value_suffix: str = "",
) -> str:
    index = _sheet_index(headers)
    if row_column not in index or column_column not in index or value_column not in index:
        return ""
    matrix: dict[str, dict[str, float]] = {}
    column_totals: dict[str, float] = {}
    row_totals: dict[str, float] = {}
    for row in rows:
        row_label = str(row[index[row_column]] if index[row_column] < len(row) and row[index[row_column]] not in (None, "") else "UNKNOWN")
        col_label = str(row[index[column_column]] if index[column_column] < len(row) and row[index[column_column]] not in (None, "") else "UNKNOWN")
        if row_column.lower() in {"product", "product_code", "sub-product", "sub_product_code"}:
            row_label = _product_display_label(row_label, header=row_column.lower())
        if column_column.lower() in {"product", "product_code", "sub-product", "sub_product_code"}:
            col_label = _product_display_label(col_label, header=column_column.lower())
        value = _number(row[index[value_column]] if index[value_column] < len(row) else None)
        matrix.setdefault(row_label, {})[col_label] = matrix.setdefault(row_label, {}).get(col_label, 0.0) + value
        column_totals[col_label] = column_totals.get(col_label, 0.0) + value
        row_totals[row_label] = row_totals.get(row_label, 0.0) + value
    row_labels = [label for label, _total in sorted(row_totals.items(), key=lambda item: item[1], reverse=True)[:10]]
    col_labels = [label for label, _total in sorted(column_totals.items(), key=lambda item: item[1], reverse=True)[:max_columns]]
    if not row_labels or not col_labels:
        return ""
    max_value = max((matrix.get(row_label, {}).get(col_label, 0.0) for row_label in row_labels for col_label in col_labels), default=0.0)
    if max_value <= 0:
        return ""
    header_html = "".join(f"<th>{html.escape(label)}</th>" for label in col_labels)
    body = []
    for row_label in row_labels:
        product_attr = _product_data_attr(row_label, header=row_column.lower()) if row_column.lower() in PRODUCT_LABEL_COLUMNS else ""
        cells = []
        for col_label in col_labels:
            value = matrix.get(row_label, {}).get(col_label, 0.0)
            alpha = 0.08 + (value / max_value * 0.54 if value > 0 else 0)
            cells.append(
                f'<td class="num heat" style="background:rgba(23,105,224,{alpha:.2f})">{html.escape(_format_number(value, suffix=value_suffix))}</td>'
            )
        body.append(f"<tr{product_attr}><th>{html.escape(row_label)}</th>{''.join(cells)}</tr>")
    scope = ' data-product-visual="1"' if row_column.lower() in PRODUCT_LABEL_COLUMNS else ' data-global-visual="1"'
    return (
        f'<section class="panel wide"{scope}><h2>{html.escape(title)}</h2>'
        f'<div class="table-wrap heatmap"><table><thead><tr><th>{html.escape(row_column)}</th>{header_html}</tr></thead>'
        f"<tbody>{''.join(body)}</tbody></table></div></section>"
    )


def _comparison_cards(title: str, metrics: list[tuple[str, float, float, str]]) -> str:
    cards = []
    for label, previous, current, suffix in metrics:
        delta = current - previous
        favorable_delta = delta <= 0 if "outstanding" in label.lower() else delta >= 0
        tone = "good" if favorable_delta else "watch"
        cards.append(
            f'<div class="comparison-card {tone}">'
            f"<span>{html.escape(label)}</span>"
            f"<strong>{html.escape(_format_number(current, suffix=suffix))}</strong>"
            f"<small>Apr: {html.escape(_format_number(previous, suffix=suffix))} | Change: {html.escape(_format_number(delta, suffix=suffix))}</small>"
            "</div>"
        )
    return f'<section class="panel wide" data-global-visual="1"><h2>{html.escape(title)}</h2><div class="comparison-grid">{"".join(cards)}</div></section>' if cards else ""


def _insights_panel(insights: list[tuple[str, str, str]]) -> str:
    if not insights:
        return ""
    cards = []
    for title, value, detail in insights[:4]:
        cards.append(
            '<div class="insight-card">'
            f"<span>{html.escape(title)}</span>"
            f"<strong>{html.escape(value)}</strong>"
            f"<small>{html.escape(detail)}</small>"
            "</div>"
        )
    return f'<section class="insights-card"><p class="eyebrow">Business Insights</p><h2>What to Watch</h2><div class="insight-grid">{"".join(cards)}</div></section>'


def _sheet_index(headers: list[str]) -> dict[str, int]:
    return {str(header): offset for offset, header in enumerate(headers)}


def _column_values(headers: list[str], rows: list[list[Any]], column_name: str) -> list[Any]:
    index = _sheet_index(headers)
    if column_name not in index:
        return []
    offset = index[column_name]
    return [row[offset] if offset < len(row) else None for row in rows]


def _sum_column(headers: list[str], rows: list[list[Any]], column_name: str) -> float:
    return sum(_number(value) for value in _column_values(headers, rows, column_name))


def _kpi_card(label: str, value: Any, *, tone: str = "neutral", suffix: str = "") -> str:
    return (
        f'<div class="kpi {html.escape(tone)}">'
        f"<span>{html.escape(label)}</span>"
        f"<strong>{html.escape(_format_number(value, suffix=suffix))}</strong>"
        "</div>"
    )


def _analyze_sheets(sheets: list[tuple[str, list[str], list[list[Any]]]]) -> list[str]:
    notes: list[str] = []
    for sheet_name, headers, rows in sheets:
        header_index = {str(header).lower(): offset for offset, header in enumerate(headers)}
        for column_index, header in enumerate(headers):
            numeric_values = []
            for row in rows:
                value = row[column_index] if column_index < len(row) else None
                if _is_number(value):
                    numeric_values.append(_number(value))
            if not numeric_values:
                continue
            gt_int64 = sum(1 for value in numeric_values if abs(value) > 9_223_372_036_854_775_807)
            negatives = sum(1 for value in numeric_values if value < 0)
            lowered = str(header).lower()
            if gt_int64:
                notes.append(f"{sheet_name}.{header}: {gt_int64} values exceed signed 64-bit integer range.")
            if negatives and any(token in lowered for token in ("available", "outstanding", "amount", "limit")):
                notes.append(f"{sheet_name}.{header}: {negatives} negative aggregate values detected; validate business definition before using as available capacity.")
            if lowered.endswith("rate"):
                outside = sum(1 for value in numeric_values if value < 0 or value > 1)
                if outside:
                    notes.append(f"{sheet_name}.{header}: {outside} rate values are outside 0-100%.")
    return notes


def _overview_cards(report_title: str, sheets: list[tuple[str, list[str], list[list[Any]]]]) -> str:
    summary = next(((headers, rows) for sheet, headers, rows in sheets if sheet in {"Summary by Product", "Funnel Summary by Product"}), None)
    if not summary:
        return ""
    headers, rows = summary
    cards: list[str] = []
    if "Applications" in headers:
        cards.extend(
            [
                _kpi_card("Applications", _sum_column(headers, rows, "Applications")),
                _kpi_card("Approved", _sum_column(headers, rows, "Approved"), tone="good"),
                _kpi_card("Rejected", _sum_column(headers, rows, "Rejected"), tone="watch"),
            ]
        )
    elif "applications" in headers:
        cards.extend(
            [
                _kpi_card("Applications", _sum_column(headers, rows, "applications")),
                _kpi_card("Approved Cases", _sum_column(headers, rows, "approved_cases"), tone="good"),
                _kpi_card("Disbursed Loans", _sum_column(headers, rows, "disbursed_loans"), tone="good"),
            ]
        )
    elif "due_amount" in headers:
        cards.extend(
            [
                _kpi_card("Due Amount", _sum_column(headers, rows, "due_amount")),
                _kpi_card("Repaid Amount", _sum_column(headers, rows, "repaid_amount"), tone="good"),
                _kpi_card("Outstanding", _sum_column(headers, rows, "outstanding_amount"), tone="watch"),
            ]
        )
    elif "total_limit" in headers:
        cards.extend(
            [
                _kpi_card("Customers", _sum_column(headers, rows, "customers")),
                _kpi_card("Total Limit", _sum_column(headers, rows, "total_limit")),
                _kpi_card("Used Limit", _sum_column(headers, rows, "used_limit"), tone="watch"),
            ]
        )
    if not cards:
        return ""
    return f'<section class="hero-card"><p class="eyebrow">Executive View</p><h2>{html.escape(report_title)}</h2><div class="kpi-grid">{"".join(cards)}</div></section>'


def _period_metric(headers: list[str], rows: list[list[Any]], period: str, metric: str) -> float:
    index = _sheet_index(headers)
    if "period" not in index or metric not in index:
        return 0.0
    return sum(_number(row[index[metric]] if index[metric] < len(row) else None) for row in rows if str(row[index["period"]] if index["period"] < len(row) else "") == period)


def _underwriting_sections(lookup: dict[str, tuple[list[str], list[list[Any]]]]) -> list[str]:
    sections: list[str] = []
    summary = lookup.get("Summary by Product")
    if summary:
        headers, rows = summary
        sections.append(
            _donut_chart(
                "Application Decision Mix",
                [
                    ("Approved", _sum_column(headers, rows, "Approved"), "#087443"),
                    ("Rejected", _sum_column(headers, rows, "Rejected"), "#b42318"),
                    ("Pending", _sum_column(headers, rows, "Pending"), "#b54708"),
                ],
            )
        )
        product_col = "Product" if "Product" in headers else "product"
        approved_col = "Approved" if "Approved" in headers else "approved"
        applications_col = "Applications" if "Applications" in headers else "applications"
        if product_col in headers and approved_col in headers and applications_col in headers:
            product_index = headers.index(product_col)
            approved_index = headers.index(approved_col)
            applications_index = headers.index(applications_col)
            aggregates: dict[str, list[float]] = {}
            for row in rows:
                product = _product_display_label(row[product_index] if product_index < len(row) else "UNKNOWN")
                bucket = aggregates.setdefault(product, [0.0, 0.0])
                bucket[0] += _number(row[approved_index] if approved_index < len(row) else None)
                bucket[1] += _number(row[applications_index] if applications_index < len(row) else None)
            pairs = [
                (product, approved / applications * 100.0)
                for product, (approved, applications) in aggregates.items()
                if applications > 0
            ]
            pairs = sorted(pairs, key=lambda item: item[1], reverse=True)[:12]
            sections.append(
                _bar_chart(
                    "Approval Rate by Product",
                    [item[0] for item in pairs],
                    [item[1] for item in pairs],
                    value_suffix="%",
                    labels_are_products=True,
                )
            )
    funnel = lookup.get("Product Funnel")
    if funnel:
        headers, rows = funnel
        index = _sheet_index(headers)
        if {"Product", "Status", "Count"}.issubset(index):
            products = [label for label, _total in _group_sum(headers, rows, "Product", "Count")[:10]]
            statuses = ["APPROVED", "REJECTED", "PENDING"]
            colors = {"APPROVED": "#087443", "REJECTED": "#b42318", "PENDING": "#b54708"}
            series = []
            for status in statuses:
                values = []
                for product in products:
                    values.append(
                        sum(
                            _number(row[index["Count"]] if index["Count"] < len(row) else None)
                            for row in rows
                            if _product_display_label(row[index["Product"]] if index["Product"] < len(row) else "") == product
                            and str(row[index["Status"]] if index["Status"] < len(row) else "") == status
                        )
                    )
                series.append((status.title(), values, colors[status]))
            sections.append(_stacked_bar_chart("Funnel Mix by Product", products, series, labels_are_products=True))
    rejects = lookup.get("Product Reject Reasons")
    if rejects:
        headers, rows = rejects
        top = _group_sum(headers, rows, "Reject Reason", "Count")[:12]
        sections.append(_bar_chart("Top Reject Reasons", [item[0] for item in top], [item[1] for item in top]))
    backlog = lookup.get("Product Stage Backlog")
    if backlog:
        headers, rows = backlog
        sections.append(
            _heatmap_table(
                "Backlog Heatmap by Product and Stage",
                headers,
                rows,
                row_column="Product",
                column_column="Current Stage",
                value_column="Count",
                max_columns=6,
            )
        )
    return [section for section in sections if section]


def _portfolio_sections(lookup: dict[str, tuple[list[str], list[list[Any]]]]) -> list[str]:
    sections: list[str] = []
    summary = lookup.get("Summary by Product")
    if summary:
        headers, rows = summary
        sections.append(
            _comparison_cards(
                "Apr 2026 vs May 2026 MTD",
                [
                    ("Due Amount", _period_metric(headers, rows, "Apr 2026", "due_amount"), _period_metric(headers, rows, "May 2026 MTD", "due_amount"), ""),
                    ("Repaid Amount", _period_metric(headers, rows, "Apr 2026", "repaid_amount"), _period_metric(headers, rows, "May 2026 MTD", "repaid_amount"), ""),
                    ("Outstanding", _period_metric(headers, rows, "Apr 2026", "outstanding_amount"), _period_metric(headers, rows, "May 2026 MTD", "outstanding_amount"), ""),
                ],
            )
        )
        index = _sheet_index(headers)
        if {"product", "repayment_rate"}.issubset(index):
            pairs = [
                (
                    str(row[index["period"]] if index["period"] < len(row) else ""),
                    _product_display_label(row[index["product"]] if index["product"] < len(row) else "UNKNOWN"),
                    _number(row[index["repayment_rate"]] if index["repayment_rate"] < len(row) else None),
                )
                for row in rows
            ]
            mtd_pairs = [(product, rate * 100 if rate <= 1 else rate) for period, product, rate in pairs if period == "May 2026 MTD"]
            mtd_pairs = sorted(mtd_pairs, key=lambda item: item[1], reverse=True)[:12]
            sections.append(
                _bar_chart(
                    "May MTD Repayment Rate by Product",
                    [item[0] for item in mtd_pairs],
                    [item[1] for item in mtd_pairs],
                    value_suffix="%",
                    labels_are_products=True,
                )
            )
    dpd = lookup.get("DPD Buckets")
    if dpd:
        headers, rows = dpd
        sections.append(
            _heatmap_table(
                "Outstanding Amount Heatmap by DPD Bucket",
                headers,
                rows,
                row_column="product",
                column_column="dpd_bucket",
                value_column="outstanding_amount",
                max_columns=8,
            )
        )
    flow = lookup.get("Repay Flow Status")
    if flow:
        headers, rows = flow
        top = _group_sum(headers, rows, "repay_status", "repay_amount")[:10]
        sections.append(_bar_chart("Repay Flow Amount by Status", [item[0] for item in top], [item[1] for item in top]))
    return [section for section in sections if section]


def _limit_sections(lookup: dict[str, tuple[list[str], list[list[Any]]]]) -> list[str]:
    sections: list[str] = []
    summary = lookup.get("Summary by Product")
    if summary:
        headers, rows = summary
        index = _sheet_index(headers)
        valid_rows = [row for row in rows if "total_limit" in index and _number(row[index["total_limit"]] if index["total_limit"] < len(row) else None) > 0]
        if valid_rows and {"product", "used_limit", "available_limit_estimate"}.issubset(index):
            labels = [_product_display_label(row[index["product"]] if index["product"] < len(row) else "UNKNOWN") for row in valid_rows[:10]]
            used = [_number(row[index["used_limit"]] if index["used_limit"] < len(row) else None) for row in valid_rows[:10]]
            available = [_number(row[index["available_limit_estimate"]] if index["available_limit_estimate"] < len(row) else None) for row in valid_rows[:10]]
            sections.append(
                _stacked_bar_chart(
                    "Used vs Available Limit by Product",
                    labels,
                    [("Used", used, "#b54708"), ("Available", available, "#087443")],
                    labels_are_products=True,
                )
            )
        if {"product", "utilization_rate"}.issubset(index):
            pairs = []
            for row in rows:
                rate = _number(row[index["utilization_rate"]] if index["utilization_rate"] < len(row) else None)
                if rate > 0:
                    pairs.append((_product_display_label(row[index["product"]] if index["product"] < len(row) else "UNKNOWN"), rate * 100 if rate <= 1 else rate))
            pairs = sorted(pairs, key=lambda item: item[1], reverse=True)[:12]
            sections.append(
                _bar_chart(
                    "Utilization Rate by Product",
                    [item[0] for item in pairs],
                    [item[1] for item in pairs],
                    value_suffix="%",
                    labels_are_products=True,
                )
            )
    buckets = lookup.get("Utilization Buckets")
    if buckets:
        headers, rows = buckets
        sections.append(
            _heatmap_table(
                "Customer Count Heatmap by Utilization Bucket",
                headers,
                rows,
                row_column="product",
                column_column="utilization_bucket",
                value_column="customers",
                max_columns=7,
            )
        )
    available = lookup.get("EOD Available Limit")
    if available:
        headers, rows = available
        top = _group_sum(headers, rows, "status", "available_limit")[:8]
        sections.append(_bar_chart("Available Limit by Status", [item[0] for item in top], [item[1] for item in top]))
    return [section for section in sections if section]


def _specialized_sections(report_title: str, lookup: dict[str, tuple[list[str], list[list[Any]]]]) -> list[str]:
    normalized = report_title.lower()
    if "underwriting" in normalized:
        return _underwriting_sections(lookup)
    if "portfolio repayment" in normalized:
        return _portfolio_sections(lookup)
    if "limit utilization" in normalized:
        return _limit_sections(lookup)
    return []


def _business_insights(report_title: str, lookup: dict[str, tuple[list[str], list[list[Any]]]]) -> str:
    normalized = report_title.lower()
    insights: list[tuple[str, str, str]] = []
    summary = lookup.get("Summary by Product")
    funnel_summary = lookup.get("Funnel Summary by Product")
    if "application to disbursement" in normalized and funnel_summary:
        headers, rows = funnel_summary
        index = _sheet_index(headers)
        if {"applications", "disbursed_loans", "disbursed_principal", "application_to_disbursement_rate", "product"}.issubset(index):
            applications = _sum_column(headers, rows, "applications")
            disbursed_loans = _sum_column(headers, rows, "disbursed_loans")
            disbursed_principal = _sum_column(headers, rows, "disbursed_principal")
            conversion_rate = disbursed_loans / applications * 100 if applications else 0
            product_rates = []
            product_dropoffs = []
            for row in rows:
                product = _product_display_label(row[index["product"]] if index["product"] < len(row) else "UNKNOWN")
                product_apps = _number(row[index["applications"]] if index["applications"] < len(row) else None)
                product_disbursed = _number(row[index["disbursed_loans"]] if index["disbursed_loans"] < len(row) else None)
                if product_apps > 0:
                    product_rates.append((product, product_disbursed / product_apps * 100))
                    product_dropoffs.append((product, product_apps - product_disbursed))
            lowest_conversion = min(product_rates, key=lambda item: item[1], default=("UNKNOWN", 0))
            largest_dropoff = max(product_dropoffs, key=lambda item: item[1], default=("UNKNOWN", 0))
            insights.append(("Application to disbursement", f"{conversion_rate:.1f}%", f"{_format_number(disbursed_loans)} disbursed loans from {_format_number(applications)} applications."))
            insights.append(("Disbursed principal", _format_number(disbursed_principal), "Total principal disbursed in the covered period."))
            insights.append(("Lowest conversion product", lowest_conversion[0], f"{lowest_conversion[1]:.1f}% application-to-disbursement rate."))
            insights.append(("Largest funnel drop-off", largest_dropoff[0], f"{_format_number(largest_dropoff[1])} applications did not become disbursed loans."))
    elif "underwriting" in normalized and summary:
        headers, rows = summary
        index = _sheet_index(headers)
        product_col = "Product" if "Product" in index else "product"
        applications_col = "Applications" if "Applications" in index else "applications"
        approved_col = "Approved" if "Approved" in index else "approved"
        rejected_col = "Rejected" if "Rejected" in index else "rejected"
        if {product_col, applications_col, approved_col, rejected_col}.issubset(index):
            total_apps = _sum_column(headers, rows, applications_col)
            total_approved = _sum_column(headers, rows, approved_col)
            approval_rate = total_approved / total_apps * 100 if total_apps else 0
            product_totals = []
            for row in rows:
                product_totals.append((
                    _product_display_label(row[index[product_col]] if index[product_col] < len(row) else "UNKNOWN"),
                    _number(row[index[applications_col]] if index[applications_col] < len(row) else None),
                    _number(row[index[approved_col]] if index[approved_col] < len(row) else None),
                ))
            top_product = max(product_totals, key=lambda item: item[1], default=("UNKNOWN", 0, 0))
            low_approval = min(
                ((product, approved / apps * 100) for product, apps, approved in product_totals if apps > 0),
                key=lambda item: item[1],
                default=("UNKNOWN", 0),
            )
            insights.append(("Overall approval", f"{approval_rate:.1f}%", f"{_format_number(total_approved)} approvals from {_format_number(total_apps)} applications."))
            insights.append(("Largest application source", top_product[0], f"{_format_number(top_product[1])} applications in the covered period."))
            insights.append(("Lowest approval product", low_approval[0], f"{low_approval[1]:.1f}% approval rate, worth checking policy/rejection mix."))
        rejects = lookup.get("Product Reject Reasons")
        if rejects:
            top = _group_sum(rejects[0], rejects[1], "Reject Reason", "Count")[:1]
            if top:
                insights.append(("Top reject driver", top[0][0], f"{_format_number(top[0][1])} rejects across products."))
    elif "portfolio repayment" in normalized and summary:
        headers, rows = summary
        due = _sum_column(headers, rows, "due_amount")
        repaid = _sum_column(headers, rows, "repaid_amount")
        outstanding = _sum_column(headers, rows, "outstanding_amount")
        rate = repaid / due * 100 if due else 0
        by_product = _group_sum(headers, rows, "product", "outstanding_amount")
        insights.append(("Portfolio repayment", f"{rate:.1f}%", f"{_format_number(repaid)} repaid against {_format_number(due)} due."))
        insights.append(("Outstanding exposure", _format_number(outstanding), "Use DPD heatmap below to locate aging concentration."))
        if by_product:
            insights.append(("Largest outstanding product", by_product[0][0], f"{_format_number(by_product[0][1])} outstanding amount."))
        dpd = lookup.get("DPD Buckets")
        if dpd:
            top = _group_sum(dpd[0], dpd[1], "dpd_bucket", "outstanding_amount")[:1]
            if top:
                insights.append(("Highest DPD exposure bucket", top[0][0], f"{_format_number(top[0][1])} outstanding amount."))
    elif "limit utilization" in normalized and summary:
        headers, rows = summary
        by_used = _group_sum(headers, rows, "product", "used_limit")
        index = _sheet_index(headers)
        inconsistent = 0
        valid_total_limit = 0.0
        valid_used_limit = 0.0
        if {"total_limit", "used_limit"}.issubset(index):
            for row in rows:
                row_total = _number(row[index["total_limit"]] if index["total_limit"] < len(row) else None)
                row_used = _number(row[index["used_limit"]] if index["used_limit"] < len(row) else None)
                if row_total == 0 and row_used != 0:
                    inconsistent += 1
                if row_total > 0:
                    valid_total_limit += row_total
                    valid_used_limit += row_used
        utilization = valid_used_limit / valid_total_limit * 100 if valid_total_limit else 0
        insights.append(("Booked utilization", f"{utilization:.1f}%", "Only products with positive total_limit are included."))
        if by_used:
            insights.append(("Largest used limit product", by_used[0][0], f"{_format_number(by_used[0][1])} used limit."))
        if inconsistent:
            insights.append(("Data definition caveat", f"{inconsistent} products", "total_limit is zero while used limit is non-zero; availability is undefined."))
        available = lookup.get("EOD Available Limit")
        if available:
            top = _group_sum(available[0], available[1], "status", "available_limit")[:1]
            if top:
                insights.append(("Largest available status", top[0][0], f"{_format_number(top[0][1])} available limit."))
    return _insights_panel(insights)


def _searchable_table_panel(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    placeholder: str,
    step_columns: set[int] | None = None,
    column_notes: dict[str, str] | None = None,
    note: str = "",
) -> str:
    step_columns = step_columns or set()
    column_notes = column_notes or {}

    def _th(index: int, header: str) -> str:
        cls = ' class="step"' if index in step_columns else ""
        note = column_notes.get(str(header))
        info = (
            f'<button type="button" class="col-info" data-note="{html.escape(str(note), quote=True)}" '
            f'aria-label="About {html.escape(str(header), quote=True)}">&#9432;</button>'
            if note
            else ""
        )
        return (
            f"<th{cls}><span class=\"th-label\">{html.escape(str(header))}</span>{info}"
            f'<input class="col-filter" type="text" data-col="{index}" '
            f'placeholder="Filter" aria-label="Filter {html.escape(str(header))}"></th>'
        )

    header_html = "".join(_th(index, header) for index, header in enumerate(headers))

    def _cell(value: Any) -> str:
        return "" if value is None else str(value)

    body_rows = []
    for row in rows:
        cells = "".join(
            f'<td{" class=\"step\"" if index in step_columns else ""}>'
            f"{html.escape(_cell(row[index] if index < len(row) else ''))}</td>"
            for index in range(len(headers))
        )
        body_rows.append(f"<tr>{cells}</tr>")
    table_html = (
        f'<div class="table-wrap"><table class="search-table"><thead><tr>{header_html}</tr></thead>'
        f'<tbody>{"".join(body_rows)}</tbody></table></div>'
        if headers
        else '<p class="empty">No rows were returned.</p>'
    )
    total = len(rows)
    note_html = f'<p class="note">{html.escape(note)}</p>' if note else ""
    return (
        f'<section class="panel"><h2>{html.escape(title)}</h2>{note_html}'
        '<div class="search-bar">'
        f'<input type="search" data-search placeholder="{html.escape(placeholder)}" aria-label="{html.escape(placeholder)}">'
        f'<span class="count" data-count>{total} of {total} rows</span>'
        f"</div>{table_html}</section>"
    )


def _period_label(summary, row) -> str:
    headers, _ = summary
    cols = {str(c): i for i, c in enumerate(headers)}
    i = cols.get("period")
    return str(row[i]) if row is not None and i is not None and i < len(row) else ""


def _period_rows(summary):
    """For a per-period summary (rows ordered Apr, May, Jun-MTD), return
    (cur_full, prev_full, cur_label, prev_label, mtd_row, scope_label)."""
    headers, rows = summary
    labels = [_period_label(summary, r) for r in rows]
    fulls = [(label, r) for label, r in zip(labels, rows) if "MTD" not in label]
    mtds = [r for label, r in zip(labels, rows) if "MTD" in label]
    cur_label, cur = fulls[-1] if fulls else ((labels[-1] if labels else ""), (rows[-1] if rows else None))
    prev_label, prev = fulls[-2] if len(fulls) >= 2 else ("", None)
    scope_label = f"{labels[0]} – {labels[-1]}" if labels else ""
    return cur, prev, cur_label, prev_label, (mtds[-1] if mtds else None), scope_label


def _delta_vs(cur_val, prev_val, prev_label: str) -> tuple[str, str]:
    if prev_val in (None, "") or cur_val in (None, ""):
        return ("", "")
    prev_n = _number(prev_val)
    if prev_n == 0:
        return ("", "")
    pct = round((_number(cur_val) - prev_n) / prev_n * 100, 1)
    direction = "up" if pct > 0 else ("down" if pct < 0 else "")
    return (f"{abs(pct)}% vs {prev_label}", direction)


def _kpi_cards_panel(title: str, pairs: list[tuple]) -> str:
    """Each card is (label, value) or (label, value, delta_text, delta_dir) where delta_dir in up/down/''."""
    if not pairs:
        return ""
    parts = []
    for item in pairs:
        label, value = item[0], item[1]
        delta_text = item[2] if len(item) > 2 else ""
        delta_dir = item[3] if len(item) > 3 else ""
        arrow = "&#9650; " if delta_dir == "up" else ("&#9660; " if delta_dir == "down" else "")
        delta_html = f'<small class="delta {html.escape(delta_dir)}">{arrow}{html.escape(delta_text)}</small>' if delta_text else ""
        parts.append(
            f'<div class="kpi"><span>{html.escape(label)}</span><strong>{html.escape(value)}</strong>{delta_html}</div>'
        )
    return f'<section class="panel"><h2>{html.escape(title)}</h2><div class="kpi-grid">{"".join(parts)}</div></section>'


def _kpi_fmt(value: Any, fmt: str) -> str:
    if value in (None, ""):
        return ""
    if fmt == "money":
        return f"₱{_format_number(value)}"
    if fmt == "pct":
        return f"{value}%"
    if fmt == "hours":
        return f"{value}h"
    return _format_number(value)


def _period_kpi_panel(title_base: str, summary, specs: list[tuple[str, str, str]]) -> str:
    """Period-switchable KPI panel. summary=(headers, rows) with a 'period' column; specs=list of
    (label, column, fmt) where fmt in {'pct','money','hours','num'}. Renders cards for the latest full
    month (delta vs the prior full month) and embeds per-period data so the month selector can recompute
    each card client-side. Returns '' when there is no period summary to drive it."""
    headers, rows = summary if summary else ([], [])
    if not rows:
        return ""
    cols = {str(c): i for i, c in enumerate(headers)}
    if "period" not in cols:
        return ""
    cur, prev, cur_label, prev_label, _mtd, scope_label = _period_rows(summary)

    def cell(row, col):
        i = cols.get(col)
        return row[i] if row is not None and i is not None and i < len(row) else None

    periods = [str(cell(r, "period")) for r in rows]
    data: dict[str, dict[str, float]] = {}
    for r in rows:
        period = str(cell(r, "period"))
        data[period] = {col: _number(cell(r, col)) for (_l, col, _f) in specs if _is_number(cell(r, col))}

    cards_html: list[str] = []
    spec_meta: list[dict[str, str]] = []
    for (label, col, fmt) in specs:
        value = cell(cur, col)
        if value in (None, ""):
            continue
        dtext, ddir = _delta_vs(value, cell(prev, col), prev_label)
        arrow = "&#9650; " if ddir == "up" else ("&#9660; " if ddir == "down" else "")
        delta_html = f'<small class="delta {html.escape(ddir)}">{arrow}{html.escape(dtext)}</small>' if dtext else ""
        cards_html.append(
            f'<div class="kpi" data-kpi-col="{html.escape(col)}" data-kpi-fmt="{html.escape(fmt)}" '
            f'data-kpi-label="{html.escape(label)}"><span>{html.escape(label)} ({html.escape(cur_label)})</span>'
            f'<strong>{html.escape(_kpi_fmt(value, fmt))}</strong>{delta_html}</div>'
        )
        spec_meta.append({"label": label, "col": col, "fmt": fmt})
    if not cards_html:
        return ""
    payload = {"base": title_base, "periods": periods, "defaultPeriod": cur_label, "specs": spec_meta, "data": data}
    title = f"{title_base} — {scope_label}" if scope_label else title_base
    return (
        f'<section class="panel" data-period-kpi>'
        f'<script type="application/json" data-period-json>{_js_data(payload)}</script>'
        f'<h2>{html.escape(title)}</h2><div class="kpi-grid">{"".join(cards_html)}</div></section>'
    )


def _insight_panel(title: str, cards: list[tuple[str, str, str]]) -> str:
    """Auto-insight strip. Each card is (label, value, level) where level in '', bad, warn, good."""
    cards = [(label, value, level) for label, value, level in cards if value not in (None, "")]
    if not cards:
        return ""
    html_cards = "".join(
        f'<div class="insight-card {html.escape(level)}"><span class="label">{html.escape(label)}</span>'
        f'<span class="value">{html.escape(value)}</span></div>'
        for label, value, level in cards
    )
    return f'<section class="panel"><h2>{html.escape(title)}</h2><div class="insights">{html_cards}</div></section>'


def _echart_id(title: str) -> str:
    return "ec-" + (re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-") or "chart")


def _js_data(value: Any) -> str:
    # JSON safe to embed inside an inline <script> (avoid closing the tag).
    return json.dumps(value).replace("</", "<\\/")


# ECharts init templates. Placeholders (__ID__/__DATA__/__XLABEL__/__YLABEL__) are filled via
# str.replace so the JS braces don't collide with Python f-strings.
_TREND_JS = r"""
(function(){
  if(!window.echarts){return;}
  var el=document.getElementById('__ID__'); if(!el){return;}
  var series=__DATA__; var keys=Object.keys(series);
  if(!keys.length){el.innerHTML='<p class="empty">No data.</p>';return;}
  var dset={}; keys.forEach(function(k){series[k].forEach(function(pt){dset[pt[0]]=1;});});
  var dates=Object.keys(dset).sort();
  var totals=keys.map(function(k){return [k,series[k].reduce(function(a,p){return a+p[1];},0)];})
                 .sort(function(a,b){return b[1]-a[1];});
  var sel={}; totals.forEach(function(t,i){sel[t[0]]=i<5;});
  var chart=echarts.init(el);
  chart.setOption({
    tooltip:{trigger:'axis'},
    legend:{type:'scroll',top:0,selected:sel,data:totals.map(function(t){return t[0];})},
    grid:{left:70,right:24,top:38,bottom:64},
    xAxis:{type:'category',data:dates},
    yAxis:{type:'value'},
    dataZoom:[{type:'slider',start:0,end:100},{type:'inside'}],
    series:keys.map(function(k){var m={}; series[k].forEach(function(p){m[p[0]]=p[1];});
      return {name:k,type:'line',showSymbol:false,connectNulls:true,
        data:dates.map(function(d){return m[d]!=null?m[d]:null;})};})
  });
  window.addEventListener('resize',function(){chart.resize();});
})();
"""

_SCATTER_JS = r"""
(function(){
  if(!window.echarts){return;}
  var el=document.getElementById('__ID__'); if(!el){return;}
  var pts=__DATA__; if(!pts.length){el.innerHTML='<p class="empty">No data.</p>';return;}
  var med=function(a){var s=a.slice().sort(function(x,y){return x-y;});var m=Math.floor(s.length/2);
    return s.length%2?s[m]:(s[m-1]+s[m])/2;};
  var xmed=med(pts.map(function(p){return p.x;})), ymed=med(pts.map(function(p){return p.y;}));
  var smax=Math.max.apply(null,pts.map(function(p){return p.size||0;}).concat([1]));
  var chart=echarts.init(el);
  chart.setOption({
    tooltip:{formatter:function(pa){var d=pa.data;return '<b>'+d.label+'</b><br/>'+(d.name2?d.name2+'<br/>':'')
      +'__XLABEL__: '+d.value[0]+'<br/>__YLABEL__: '+d.value[1]+'<br/>volume: '+(d.size||0).toLocaleString();}},
    grid:{left:66,right:28,top:24,bottom:58},
    xAxis:{name:'__XLABEL__',nameLocation:'middle',nameGap:32,type:'value',scale:true},
    yAxis:{name:'__YLABEL__',nameLocation:'middle',nameGap:48,type:'value',scale:true},
    dataZoom:[{type:'inside',xAxisIndex:0},{type:'inside',yAxisIndex:0}],
    series:[{type:'scatter',
      symbolSize:function(v,p){return 8+Math.sqrt((p.data.size||0)/smax)*34;},
      itemStyle:{opacity:0.65,color:function(p){return (p.data.value[0]>=xmed&&p.data.value[1]<ymed)?'#b42318'
        :(p.data.value[1]>=ymed?'#087443':'#1769e0');}},
      data:pts.map(function(p){return {value:[p.x,p.y],label:p.label,name2:p.name,size:p.size};}),
      markLine:{silent:true,symbol:'none',lineStyle:{color:'#e4a11b',type:'dashed'},
        data:[{xAxis:xmed},{yAxis:ymed}]}
    }]
  });
  chart.on('click',function(pa){ if(window.__afHighlight && pa.data && pa.data.label){ window.__afHighlight(pa.data.label); } });
  window.addEventListener('resize',function(){chart.resize();});
})();
"""


def _daily_trend_panel(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    rule_column: str,
    date_column: str,
    value_column: str,
    note: str = "",
) -> str:
    """Interactive multi-line daily trend (ECharts): legend toggle + time brushing (dataZoom)."""
    idx = {str(col): offset for offset, col in enumerate(headers)}
    r_i, d_i, v_i = idx.get(rule_column), idx.get(date_column), idx.get(value_column)
    series: dict[str, list[list[Any]]] = {}
    if None not in (r_i, d_i, v_i):
        for row in rows:
            rid = "" if row[r_i] is None else str(row[r_i])
            day = "" if row[d_i] is None else str(row[d_i])
            val = _number(row[v_i]) if v_i < len(row) else 0.0
            series.setdefault(rid, []).append([day, val])
    if not series:
        return f'<section class="panel"><h2>{html.escape(title)}</h2><p class="empty">No trend data was returned.</p></section>'
    chart_id = _echart_id(title)
    data_json = _js_data({rid: sorted(pts) for rid, pts in series.items()})
    script = _TREND_JS.replace("__ID__", chart_id).replace("__DATA__", data_json)
    full_note = "Toggle series in the legend; drag the slider to zoom the date range."
    if note:
        full_note = f"{note} {full_note}"
    return (
        f'<section class="panel"><h2>{html.escape(title)}</h2>'
        f'<p class="note">{html.escape(full_note)}</p>'
        f'<div id="{chart_id}" class="echart" style="height:360px"></div>'
        f"<script>{script}</script></section>"
    )


def _scatter_quadrant_panel(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    x_column: str,
    y_column: str,
    label_column: str,
    size_column: str,
    x_label: str,
    y_label: str,
) -> str:
    """Interactive precision x trigger-rate quadrant scatter (ECharts): zoom, tooltips, median lines."""
    idx = {str(col): offset for offset, col in enumerate(headers)}

    def cell(row: list[Any], col: str) -> Any:
        return row[idx[col]] if col in idx and idx[col] < len(row) else None

    pts = []
    for row in rows:
        xv, yv = cell(row, x_column), cell(row, y_column)
        if xv in (None, "") or yv in (None, ""):
            continue
        label = cell(row, label_column)
        pts.append(
            {
                "label": "" if label is None else str(label),
                "x": _number(xv),
                "y": _number(yv),
                "size": _number(cell(row, size_column)),
            }
        )
    if not pts:
        return f'<section class="panel"><h2>{html.escape(title)}</h2><p class="empty">No scorecard data was returned.</p></section>'
    chart_id = _echart_id(title)
    script = (
        _SCATTER_JS.replace("__ID__", chart_id)
        .replace("__DATA__", _js_data(pts))
        .replace("__XLABEL__", x_label.replace("'", ""))
        .replace("__YLABEL__", y_label.replace("'", ""))
    )
    return (
        f'<section class="panel"><h2>{html.escape(title)}</h2>'
        f'<p class="note">Bubble size = volume; dashed lines = medians. '
        f'Red = high {html.escape(x_label.lower())} / low {html.escape(y_label.lower())} (tune); green = high {html.escape(y_label.lower())}.</p>'
        f'<div id="{chart_id}" class="echart" style="height:440px"></div>'
        f"<script>{script}</script></section>"
    )


_DONUT_JS = r"""
(function(){ if(!window.echarts){return;} var el=document.getElementById('__ID__'); if(!el){return;}
  var data=__DATA__; if(!data.length){el.innerHTML='<p class="empty">No data.</p>';return;}
  var chart=echarts.init(el);
  chart.setOption({tooltip:{trigger:'item',formatter:function(p){return p.name+': '+(p.value||0).toLocaleString()+' ('+p.percent+'%)';}},
    legend:{bottom:0},
    series:[{type:'pie',radius:['45%','72%'],avoidLabelOverlap:true,label:{formatter:'{b}\n{d}%'},
      data:data}]});
  window.addEventListener('resize',function(){chart.resize();});
})();
"""

_BAR_JS = r"""
(function(){ if(!window.echarts){return;} var el=document.getElementById('__ID__'); if(!el){return;}
  var data=__DATA__; if(!data.length){el.innerHTML='<p class="empty">No data.</p>';return;}
  var cats=data.map(function(d){return d.name;}).reverse();
  var vals=data.map(function(d){return d.value;}).reverse();
  var chart=echarts.init(el);
  chart.setOption({tooltip:{trigger:'axis',axisPointer:{type:'shadow'},
      valueFormatter:function(v){return '__PREFIX__'+(v||0).toLocaleString();}},
    grid:{left:8,right:60,top:12,bottom:24,containLabel:true},
    xAxis:{type:'value'}, yAxis:{type:'category',data:cats,axisLabel:{width:210,overflow:'truncate'}},
    series:[{type:'bar',data:vals,itemStyle:{color:'#1769e0'},
      label:{show:true,position:'right',formatter:function(p){return '__PREFIX__'+(p.value||0).toLocaleString();}}}]});
  window.addEventListener('resize',function(){chart.resize();});
})();
"""


def _donut_panel(title: str, pairs: list[tuple[str, Any]], *, note: str = "") -> str:
    data = [{"name": str(name), "value": _number(value)} for name, value in pairs if _number(value) > 0]
    if not data:
        return ""
    chart_id = _echart_id(title)
    script = _DONUT_JS.replace("__ID__", chart_id).replace("__DATA__", _js_data(data))
    note_html = f'<p class="note">{html.escape(note)}</p>' if note else ""
    return (
        f'<section class="panel"><h2>{html.escape(title)}</h2>{note_html}'
        f'<div id="{chart_id}" class="echart" style="height:320px"></div><script>{script}</script></section>'
    )


_FUNNEL_JS = r"""
(function(){ if(!window.echarts){return;} var el=document.getElementById('__ID__'); if(!el){return;}
  var data=__DATA__; if(!data.length){el.innerHTML='<p class="empty">No data.</p>';return;}
  var max=data[0].value||1;
  var chart=echarts.init(el);
  chart.setOption({tooltip:{trigger:'item',formatter:function(p){
      var share=max?Math.round(p.value/max*10000)/100:0;
      return p.name+': '+(p.value||0).toLocaleString()+' ('+share+'% of '+data[0].name.toLowerCase()+')';}},
    series:[{type:'funnel',sort:'none',gap:3,top:8,bottom:8,left:'5%',width:'90%',minSize:'45%',
      label:{position:'inside',color:'#fff',fontWeight:600,
        formatter:function(p){var share=max?Math.round(p.value/max*10000)/100:0;
        return p.name+'  '+(p.value||0).toLocaleString()+' ('+share+'%)';}},
      data:data}]});
  window.addEventListener('resize',function(){chart.resize();});
})();
"""


def _funnel_panel(title: str, stages: list[tuple[str, Any]], *, note: str = "") -> str:
    """Sequential funnel (ECharts). stages = ordered (label, count); rendered in given order with
    share-of-first labels."""
    data = [{"name": str(name), "value": _number(value)} for name, value in stages if _number(value) > 0]
    if len(data) < 2:
        return ""
    chart_id = _echart_id(title)
    script = _FUNNEL_JS.replace("__ID__", chart_id).replace("__DATA__", _js_data(data))
    note_html = f'<p class="note">{html.escape(note)}</p>' if note else ""
    return (
        f'<section class="panel"><h2>{html.escape(title)}</h2>{note_html}'
        f'<div id="{chart_id}" class="echart" style="height:300px"></div><script>{script}</script></section>'
    )


def _bar_panel(title: str, pairs: list[tuple[str, Any]], *, note: str = "", prefix: str = "") -> str:
    data = [{"name": str(name), "value": _number(value)} for name, value in pairs if _number(value) > 0]
    if not data:
        return ""
    chart_id = _echart_id(title)
    script = _BAR_JS.replace("__ID__", chart_id).replace("__DATA__", _js_data(data)).replace("__PREFIX__", prefix)
    note_html = f'<p class="note">{html.escape(note)}</p>' if note else ""
    height = max(220, 30 * len(data) + 60)
    return (
        f'<section class="panel"><h2>{html.escape(title)}</h2>{note_html}'
        f'<div id="{chart_id}" class="echart" style="height:{height}px"></div><script>{script}</script></section>'
    )


def _expandable_rule_panel(
    title: str,
    placeholder: str,
    summary: tuple[list[str], list[list[Any]]],
    detail: tuple[list[str], list[list[Any]]],
    *,
    key_columns: tuple[str, ...],
    main_columns: tuple[str, ...],
    detail_columns: tuple[str, ...],
    name_column: str,
    note: str = "",
) -> str:
    """Render a per-rule table whose rows expand to a nested scene breakdown.

    `summary` is one row per rule; `detail` is one row per rule x scene. Rows are
    linked by `key_columns`. A single search box filters rules by their own text
    plus their child scene names.
    """
    s_head, s_rows = summary
    d_head, d_rows = detail
    si = {str(col): offset for offset, col in enumerate(s_head)}
    di = {str(col): offset for offset, col in enumerate(d_head)}

    def cell(value: Any) -> str:
        return "" if value is None else str(value)

    def sval(row: list[Any], col: str) -> str:
        return cell(row[si[col]]) if col in si and si[col] < len(row) else ""

    def dval(row: list[Any], col: str) -> str:
        return cell(row[di[col]]) if col in di and di[col] < len(row) else ""

    grouped: dict[tuple[str, ...], list[list[Any]]] = {}
    for row in d_rows:
        grouped.setdefault(tuple(dval(row, col) for col in key_columns), []).append(row)

    head_html = "<th></th>" + "".join(f"<th>{html.escape(col)}</th>" for col in main_columns)
    body: list[str] = []
    for srow in s_rows:
        key = tuple(sval(srow, col) for col in key_columns)
        key_attr = html.escape("|".join(key), quote=True)
        kids = grouped.get(key, [])
        scene_text = " ".join(dval(k, name_column) for k in kids)
        data_text = html.escape(" ".join([*key, scene_text]).lower(), quote=True)
        cells = "".join(f"<td>{html.escape(sval(srow, col))}</td>" for col in main_columns)
        expander = (
            '<button class="expander" type="button" aria-expanded="false" aria-label="Toggle scene breakdown">&#9654;</button>'
            if kids
            else ""
        )
        body.append(
            f'<tr class="rule-row" data-key="{key_attr}" data-text="{data_text}">'
            f'<td class="exp-cell">{expander}</td>{cells}</tr>'
        )
        if kids:
            d_head_html = "".join(f"<th>{html.escape(col)}</th>" for col in detail_columns)
            d_body = "".join(
                "<tr>" + "".join(f"<td>{html.escape(dval(k, col))}</td>" for col in detail_columns) + "</tr>"
                for k in kids
            )
            inner = f'<table class="detail-table"><thead><tr>{d_head_html}</tr></thead><tbody>{d_body}</tbody></table>'
            body.append(
                f'<tr class="detail-row" data-key="{key_attr}" hidden>'
                f'<td class="detail-cell" colspan="{len(main_columns) + 1}">{inner}</td></tr>'
            )
    total = len(s_rows)
    table_html = (
        f'<div class="table-wrap"><table class="rule-table"><thead><tr>{head_html}</tr></thead>'
        f'<tbody>{"".join(body)}</tbody></table></div>'
        if s_rows
        else '<p class="empty">No rows were returned.</p>'
    )
    note_html = f'<p class="note">{html.escape(note)}</p>' if note else ""
    return (
        f'<section class="panel rule-panel"><h2>{html.escape(title)}</h2>{note_html}'
        '<div class="search-bar">'
        f'<input type="search" data-search placeholder="{html.escape(placeholder)}" aria-label="{html.escape(placeholder)}">'
        f'<span class="count" data-count>{total} of {total} rules</span>'
        f"</div>{table_html}</section>"
    )


# Shared dashboard interaction JS (plain string -> no f-string brace escaping). Handles table search,
# per-column filters, expand/collapse, click-to-sort, number formatting, and threshold highlighting.
_DASHBOARD_JS = r"""
(function () {
  function parseNum(s){ if(s==null) return null; var t=String(s).replace(/[,₱%\s]/g,''); if(t===''||isNaN(Number(t))) return null; return Number(t); }
  function fmt(n, dec){ if(dec==null){ dec = (n%1!==0)?2:0; } return n.toLocaleString(undefined,{minimumFractionDigits:0,maximumFractionDigits:dec}); }
  function idLike(h){ return /(^id$|_id$|^uid$|^mcc$|uuid|_code$|date|status_code|trigger_date)/.test(h); }
  function colKind(h){ if(h.slice(-4)==='_pct') return 'pct'; if(h.slice(-4)==='_php') return 'php'; return 'num'; }
  function flagClass(h,v){ if(v==null) return '';
    if(h.indexOf('trigger_rate_pct')>=0){ if(v>=20) return 'flag-bad'; if(v>=5) return 'flag-warn'; }
    else if(h.indexOf('precision_pct')>=0){ if(v<1) return 'flag-bad'; if(v<5) return 'flag-warn'; }
    else if(h.indexOf('avg_age_days')>=0){ if(v>=90) return 'flag-bad'; if(v>=30) return 'flag-warn'; }
    return ''; }
  function headerLabels(table){ var hr = table.tHead ? table.tHead.rows[table.tHead.rows.length-1] : null; if(!hr) return [];
    return Array.prototype.map.call(hr.cells, function(th){ var l=th.querySelector('.th-label'); return (l?l.textContent:th.textContent).trim().toLowerCase(); }); }
  function enhanceTable(table, sortable){
    var headers = headerLabels(table); if(!headers.length) return;
    var tb = table.tBodies[0]; if(!tb) return;
    Array.prototype.forEach.call(tb.rows, function(tr){
      var raws=[];
      Array.prototype.forEach.call(tr.cells, function(td,i){
        var header = headers[i]||'';
        var raw = td.getAttribute('data-raw'); if(raw==null){ raw=td.textContent.trim(); td.setAttribute('data-raw',raw); }
        raws.push(raw.toLowerCase());
        var num = (header && !idLike(header)) ? parseNum(raw) : null;
        if(num!=null){
          td.setAttribute('data-sort', String(num));
          var k = colKind(header);
          td.textContent = k==='pct' ? (fmt(num,3)+'%') : (k==='php' ? ('₱'+fmt(num,2)) : fmt(num));
          var fc = flagClass(header,num); if(fc) td.classList.add(fc);
        } else { td.setAttribute('data-sort', raw.toLowerCase()); }
      });
      tr._cells = raws;
      if(!tr.dataset.text){ tr.dataset.text = raws.join(' '); }
    });
    if(sortable){
      var hr = table.tHead.rows[table.tHead.rows.length-1];
      Array.prototype.forEach.call(hr.cells, function(th,i){
        th.classList.add('sortable');
        th.addEventListener('click', function(e){
          if(e.target && ((e.target.classList && e.target.classList.contains('col-filter')) || (e.target.closest && e.target.closest('.col-info')))) return;
          var asc = th.getAttribute('data-dir')!=='asc';
          Array.prototype.forEach.call(hr.cells, function(o){ o.removeAttribute('data-dir'); o.classList.remove('sort-asc','sort-desc'); });
          th.setAttribute('data-dir', asc?'asc':'desc'); th.classList.add(asc?'sort-asc':'sort-desc');
          var rows = Array.prototype.slice.call(tb.rows);
          rows.sort(function(a,b){
            var av=a.cells[i]?a.cells[i].getAttribute('data-sort'):''; var bv=b.cells[i]?b.cells[i].getAttribute('data-sort'):'';
            var an=parseFloat(av), bn=parseFloat(bv); var both=!isNaN(an)&&!isNaN(bn);
            var c = both ? an-bn : String(av).localeCompare(String(bv)); return asc?c:-c;
          });
          rows.forEach(function(r){ tb.appendChild(r); });
        });
      });
    }
  }
  document.querySelectorAll('.panel').forEach(function(panel){
    var table = panel.querySelector('table.search-table'); if(!table) return;
    enhanceTable(table, true);
    var input = panel.querySelector('[data-search]'); var counter = panel.querySelector('[data-count]');
    var colFilters = Array.prototype.slice.call(panel.querySelectorAll('.col-filter'));
    var tb = table.tBodies[0]; var rows = Array.prototype.slice.call(tb.rows); var total = rows.length;
    var apply = function(){
      var q = ((input&&input.value)||'').trim().toLowerCase();
      var active = colFilters.map(function(f){ return [Number(f.dataset.col),(f.value||'').trim().toLowerCase()]; }).filter(function(p){ return p[1]; });
      var vis=0;
      rows.forEach(function(r){ var m = !q || (r.dataset.text||'').indexOf(q)>=0;
        if(m){ for(var k=0;k<active.length;k++){ if(((r._cells||[])[active[k][0]]||'').indexOf(active[k][1])<0){ m=false; break; } } }
        r.classList.toggle('no-match', !m); if(m) vis++; });
      if(counter) counter.textContent = vis+' of '+total+' rows';
    };
    if(input) input.addEventListener('input', apply); colFilters.forEach(function(f){ f.addEventListener('input', apply); }); apply();
  });
  document.querySelectorAll('.rule-panel').forEach(function(panel){
    var table = panel.querySelector('table.rule-table'); if(!table) return;
    enhanceTable(table, false);
    Array.prototype.forEach.call(table.querySelectorAll('.detail-table'), function(dt){ enhanceTable(dt, false); });
    var input = panel.querySelector('[data-search]'); var counter = panel.querySelector('[data-count]');
    var mainRows = Array.prototype.slice.call(table.querySelectorAll('tr.rule-row'));
    var details = {}; table.querySelectorAll('tr.detail-row').forEach(function(d){ details[d.dataset.key]=d; });
    var total = mainRows.length;
    table.querySelectorAll('.expander').forEach(function(btn){
      btn.addEventListener('click', function(){ var tr=btn.closest('tr'); var d=details[tr.dataset.key];
        var open = btn.getAttribute('aria-expanded')==='true'; btn.setAttribute('aria-expanded', String(!open)); btn.innerHTML = open?'&#9654;':'&#9660;'; if(d) d.hidden=open; });
    });
    var apply = function(){ var q=((input&&input.value)||'').trim().toLowerCase(); var vis=0;
      mainRows.forEach(function(tr){ var m = !q || (tr.dataset.text||'').indexOf(q)>=0; tr.hidden=!m;
        var d=details[tr.dataset.key]; if(d&&!m){ d.hidden=true; var b=tr.querySelector('.expander'); if(b){ b.setAttribute('aria-expanded','false'); b.innerHTML='&#9654;'; } } if(m) vis++; });
      if(counter) counter.textContent = vis+' of '+total+' rules'; };
    if(input) input.addEventListener('input', apply); apply();
    // Cross-filter: clicking a rule row (not the expander) highlights that rule everywhere.
    mainRows.forEach(function(tr){ tr.addEventListener('click', function(e){
      if(e.target && e.target.closest && e.target.closest('.expander')) return;
      var firstData = tr.querySelector('td:not(.exp-cell)'); if(firstData && window.__afHighlight){ window.__afHighlight(firstData.getAttribute('data-raw')||firstData.textContent.trim()); }
    }); });
  });
  // ---- cross-filter / linked highlight across panels ----
  var current=null, chip=null;
  function ruleCells(){ return document.querySelectorAll('table.search-table tbody td:first-child, table.rule-table tbody tr.rule-row td:nth-child(2)'); }
  function clearHighlight(){ current=null;
    document.querySelectorAll('.cross-hit').forEach(function(el){ el.classList.remove('cross-hit'); });
    if(chip){ chip.style.display='none'; } }
  window.__afHighlight = function(ruleId){
    ruleId = (ruleId||'').trim(); if(!ruleId){ return; }
    if(current===ruleId){ clearHighlight(); return; }
    document.querySelectorAll('.cross-hit').forEach(function(el){ el.classList.remove('cross-hit'); });
    current=ruleId; var key=ruleId.toLowerCase(); var first=null; var hits=0;
    document.querySelectorAll('table.rule-table tbody tr.rule-row').forEach(function(tr){
      var k=(tr.dataset.key||'').toLowerCase();
      if(k===key || k.indexOf(key+'|')===0){ tr.classList.add('cross-hit'); hits++; if(!first) first=tr; }
    });
    document.querySelectorAll('table.search-table tbody tr').forEach(function(tr){
      var c=tr.cells[0]; var v=c?((c.getAttribute('data-raw')||c.textContent).trim().toLowerCase()):'';
      if(v===key){ tr.classList.add('cross-hit'); hits++; if(!first) first=tr; }
    });
    if(!chip){ chip=document.createElement('div'); chip.className='xfilter-chip'; document.body.appendChild(chip);
      chip.addEventListener('click', clearHighlight); }
    chip.innerHTML='Highlighting <b>'+ruleId+'</b> ('+hits+') &times;'; chip.style.display='block';
    if(first){ first.scrollIntoView({behavior:'smooth', block:'center'}); }
  };
})();
(function(){
  // In-page table of contents for multi-panel reports.
  var main=document.querySelector('main'); if(!main){return;}
  var panels=Array.prototype.filter.call(main.children, function(el){ return el.classList && el.classList.contains('panel'); });
  var heads=[];
  panels.forEach(function(p){ var h=p.querySelector(':scope > h2'); if(h){ heads.push({panel:p, h:h}); } });
  if(heads.length<3){return;}
  var nav=document.createElement('nav'); nav.className='toc';
  heads.forEach(function(o,i){ if(!o.panel.id){ o.panel.id='panel-'+i; }
    var a=document.createElement('a'); a.href='#'+o.panel.id; a.textContent=o.h.textContent; nav.appendChild(a); });
  main.insertBefore(nav, main.firstChild);
})();
(function(){
  // Per-table CSV export. Adds a "Download CSV" button to each table panel; exports the rows currently
  // visible (search / column filters applied) using the unformatted data-raw values.
  function slug(s){ return (String(s||'table').toLowerCase().replace(/[^a-z0-9]+/g,'-').replace(/^-+|-+$/g,'').slice(0,60))||'table'; }
  function esc(s){ s=String(s==null?'':s); return /[",\n\r]/.test(s) ? '"'+s.replace(/"/g,'""')+'"' : s; }
  function labelOf(th){ var l=th.querySelector('.th-label'); return (l?l.textContent:th.textContent).trim(); }
  function exportTable(table, name){
    var hr = table.tHead ? table.tHead.rows[table.tHead.rows.length-1] : null; if(!hr){ return; }
    var hcells = Array.prototype.slice.call(hr.cells);
    var keep = hcells.map(function(th){ return !th.classList.contains('exp-cell'); });
    var labels = hcells.filter(function(th,i){ return keep[i]; }).map(labelOf);
    var lines = [labels.map(esc).join(',')];
    var bodyRows;
    if(table.classList.contains('rule-table')){
      bodyRows = Array.prototype.slice.call(table.querySelectorAll('tbody tr.rule-row')).filter(function(tr){ return !tr.hidden; });
    } else {
      bodyRows = Array.prototype.slice.call(table.tBodies[0].rows).filter(function(tr){ return !tr.classList.contains('no-match'); });
    }
    bodyRows.forEach(function(tr){
      var vals=[];
      Array.prototype.forEach.call(tr.cells, function(td,i){ if(!keep[i]){ return; }
        var raw=td.getAttribute('data-raw'); vals.push(raw!=null?raw:td.textContent.trim()); });
      lines.push(vals.map(esc).join(','));
    });
    var blob = new Blob(['﻿'+lines.join('\r\n')], {type:'text/csv;charset=utf-8;'});
    var a = document.createElement('a'); a.href = URL.createObjectURL(blob); a.download = slug(name)+'.csv';
    document.body.appendChild(a); a.click();
    setTimeout(function(){ URL.revokeObjectURL(a.href); a.remove(); }, 0);
  }
  document.querySelectorAll('.panel').forEach(function(panel){
    var table = panel.querySelector('table.search-table, table.rule-table'); if(!table){ return; }
    var h2 = panel.querySelector(':scope > h2'); var name = h2 ? h2.textContent : 'table';
    var btn = document.createElement('button'); btn.type='button'; btn.className='csv-btn'; btn.textContent='Download CSV';
    btn.addEventListener('click', function(){ exportTable(table, name); });
    var bar = panel.querySelector('.search-bar');
    if(bar){ bar.appendChild(btn); }
    else { var d=document.createElement('div'); d.className='search-bar'; d.appendChild(btn); panel.insertBefore(d, table.closest('.table-wrap')||table); }
  });
})();
(function(){
  // Month selector: focus the report on one period. Recomputes [data-period-kpi] cards (with delta vs
  // the prior period) and filters tables that carry a 'period' column. "All months" restores the
  // overview (KPIs = latest full month, all table rows shown). Absent when there is no period data.
  var kpiPanels = Array.prototype.slice.call(document.querySelectorAll('[data-period-kpi]'));
  var meta = kpiPanels.map(function(p){ var s=p.querySelector('script[data-period-json]'); if(!s){ return null; }
    try{ return JSON.parse(s.textContent); }catch(e){ return null; } });
  function periodCol(table){ var hr=table.tHead?table.tHead.rows[table.tHead.rows.length-1]:null; if(!hr){ return -1; }
    for(var i=0;i<hr.cells.length;i++){ var l=hr.cells[i].querySelector('.th-label');
      var t=(l?l.textContent:hr.cells[i].textContent).trim().toLowerCase(); if(t==='period'){ return i; } } return -1; }
  var periodTables=[]; document.querySelectorAll('table.search-table').forEach(function(t){ var ci=periodCol(t); if(ci>=0){ periodTables.push([t,ci]); } });
  var periods=[];
  meta.forEach(function(m){ if(m){ (m.periods||[]).forEach(function(p){ if(periods.indexOf(p)<0){ periods.push(p); } }); } });
  if(!periods.length){ periodTables.forEach(function(pt){ Array.prototype.forEach.call(pt[0].tBodies[0].rows, function(r){
    var c=r.cells[pt[1]]; var v=(c.getAttribute('data-raw')||c.textContent).trim(); if(v && periods.indexOf(v)<0){ periods.push(v); } }); }); }
  if(periods.length < 2){ return; }
  var bar=document.createElement('div'); bar.className='period-bar';
  var label=document.createElement('span'); label.className='period-label'; label.textContent='Month';
  var sel=document.createElement('select'); sel.className='period-select';
  var optAll=document.createElement('option'); optAll.value='__all__'; optAll.textContent='All months'; sel.appendChild(optAll);
  periods.forEach(function(p){ var o=document.createElement('option'); o.value=p; o.textContent=p; sel.appendChild(o); });
  bar.appendChild(label); bar.appendChild(sel);
  var main=document.querySelector('main'); if(!main){ return; }
  var toc=main.querySelector('.toc');
  if(toc && toc.nextSibling){ main.insertBefore(bar, toc.nextSibling); } else { main.insertBefore(bar, main.firstChild); }
  function fmtNum(n){ var dec=(n%1!==0)?2:0; return n.toLocaleString(undefined,{minimumFractionDigits:0,maximumFractionDigits:dec}); }
  function fmtVal(v,fmt){ if(v==null||isNaN(v)){ return ''; } if(fmt==='money'){ return '₱'+fmtNum(v); }
    if(fmt==='pct'){ return v+'%'; } if(fmt==='hours'){ return v+'h'; } return fmtNum(v); }
  function renderKpis(period){
    kpiPanels.forEach(function(panel,pi){ var m=meta[pi]; if(!m){ return; }
      var shown = (period==='__all__') ? m.defaultPeriod : period;
      var idx=m.periods.indexOf(shown); var prevP = idx>0 ? m.periods[idx-1] : null;
      var row=m.data[shown]||{}; var prow=prevP?(m.data[prevP]||{}):{};
      Array.prototype.forEach.call(panel.querySelectorAll('.kpi'), function(card){
        var col=card.getAttribute('data-kpi-col'); if(!col){ return; }
        var fmt=card.getAttribute('data-kpi-fmt'); var base=card.getAttribute('data-kpi-label');
        var v=row[col]; var span=card.querySelector('span'); var strong=card.querySelector('strong'); var delta=card.querySelector('.delta');
        if(span){ span.textContent=base+' ('+shown+')'; }
        if(strong){ strong.textContent = (v==null?'':fmtVal(v,fmt)); }
        if(delta){ var pv=prevP?prow[col]:null;
          if(pv!=null && pv!==0 && v!=null){ var pct=Math.round((v-pv)/pv*1000)/10; var dir=pct>0?'up':(pct<0?'down':'');
            delta.className='delta '+dir; delta.innerHTML=(dir==='up'?'&#9650; ':(dir==='down'?'&#9660; ':''))+Math.abs(pct)+'% vs '+prevP; delta.style.display=''; }
          else { delta.style.display='none'; } }
      });
    });
  }
  function filterTables(period){
    periodTables.forEach(function(pt){ var t=pt[0],ci=pt[1];
      Array.prototype.forEach.call(t.tBodies[0].rows, function(r){
        if(period==='__all__'){ r.classList.remove('period-hide'); return; }
        var c=r.cells[ci]; var v=(c.getAttribute('data-raw')||c.textContent).trim();
        r.classList.toggle('period-hide', v!==period); }); });
  }
  sel.addEventListener('change', function(){ renderKpis(sel.value); filterTables(sel.value); });
})();
(function(){
  // Charts depend on the portal-served ECharts bundle; if it failed to load (e.g. the HTML file was
  // downloaded and opened standalone), say so instead of leaving silent blank panels.
  window.addEventListener('load', function(){
    if(window.echarts){ return; }
    document.querySelectorAll('.echart').forEach(function(el){
      el.innerHTML = '<p class="empty">Chart could not load its library - open this report from the portal to see the interactive chart. The table data on this page is unaffected.</p>';
    });
  });
})();
(function(){
  // Click an info icon (next to a column header) to toggle a note popover explaining that column.
  var pop=null;
  function close(){ if(pop){ pop.remove(); pop=null; } }
  document.addEventListener('click', function(e){
    var btn = e.target.closest ? e.target.closest('.col-info') : null;
    if(btn){
      var owned = pop && pop._owner===btn;
      close();
      if(owned){ return; }
      pop=document.createElement('div'); pop.className='col-note-pop'; pop._owner=btn;
      pop.textContent=btn.getAttribute('data-note')||'';
      document.body.appendChild(pop);
      var r=btn.getBoundingClientRect();
      var left=Math.max(8, Math.min(window.scrollX+r.left, window.scrollX+window.innerWidth-pop.offsetWidth-8));
      pop.style.top=(window.scrollY+r.bottom+6)+'px'; pop.style.left=left+'px';
      return;
    }
    if(pop && !(e.target.closest && e.target.closest('.col-note-pop'))){ close(); }
  });
})();
"""


_ISO_DATE_PREFIX = re.compile(r"\d{4}-\d{2}-\d{2}")


def _data_window(sheets: list[tuple[str, list[str], list[list[Any]]]]) -> tuple[str, str]:
    """(earliest, latest) ISO date seen in any date-named column across the sheets - the honest data
    coverage span for the header (daily partitions can lag the query window by a day)."""
    earliest = ""
    latest = ""
    for _sheet_name, headers, rows in sheets:
        date_columns = [offset for offset, header in enumerate(headers) if "date" in str(header).lower()]
        for offset in date_columns:
            for row in rows:
                if offset < len(row) and row[offset] not in (None, ""):
                    text = str(row[offset])[:10]
                    if _ISO_DATE_PREFIX.fullmatch(text):
                        if text > latest:
                            latest = text
                        if not earliest or text < earliest:
                            earliest = text
    return earliest, latest


def _data_through(sheets: list[tuple[str, list[str], list[list[Any]]]]) -> str:
    """Header coverage marker: a 'min → max' window when the data spans multiple dates, else the single
    'data through' date. Empty when no date column is present (pure config snapshots)."""
    earliest, latest = _data_window(sheets)
    if not latest:
        return ""
    return f"{earliest} → {latest}" if earliest and earliest != latest else latest


def _searchable_tables_document(
    report_title: str,
    snapshot_pt_date: str,
    panels: list[str],
    *,
    intro_html: str = "",
    data_through: str = "",
) -> str:
    generated_at = format_gmt8(datetime.now(UTC))
    if not data_through:
        coverage = ""
    elif "→" in data_through:
        coverage = f" Data window {html.escape(data_through)}."
    else:
        coverage = f" Data through {html.escape(data_through)}."
    body = (intro_html + "".join(panels)) or '<section class="panel"><p class="empty">No data was returned.</p></section>'
    body_html = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>{html.escape(report_title)} Visualization</title>
<style>
:root{{--ink:#182230;--muted:#667085;--line:#d9e2ec;--bg:#f5f7fb;--blue:#1769e0;}}
*{{box-sizing:border-box;}}
body{{margin:0;background:var(--bg);color:var(--ink);font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;}}
header{{background:linear-gradient(135deg,#102a43,#173b5f);color:#fff;padding:30px 38px;}}
header h1{{margin:0 0 8px;font-size:28px;overflow-wrap:anywhere;}} header p{{margin:0;color:#dbeafe;}}
main{{padding:24px 34px 38px;}}
.panel{{background:#fff;border:1px solid var(--line);border-radius:8px;padding:18px;box-shadow:0 1px 2px rgba(16,42,67,.06);}}
.panel + .panel{{margin-top:18px;}}
h2{{margin:0 0 14px;font-size:18px;}}
.search-bar{{display:flex;align-items:center;gap:12px;margin-bottom:14px;flex-wrap:wrap;}}
.search-bar input{{flex:1;min-width:240px;height:40px;border:1px solid #cbd5e1;border-radius:6px;padding:0 12px;font-size:14px;}}
.search-bar .count{{color:var(--muted);font-size:13px;white-space:nowrap;}}
.csv-btn{{flex:0 0 auto;height:36px;border:1px solid #cbd5e1;border-radius:6px;background:#fff;color:var(--blue);font-size:13px;font-weight:600;padding:0 12px;cursor:pointer;white-space:nowrap;}}
.csv-btn:hover{{border-color:var(--blue);background:#f1f6ff;}}
.period-bar{{display:flex;align-items:center;gap:10px;margin:0 0 14px;}}
.period-bar .period-label{{font-size:13px;font-weight:700;color:#344054;}}
.period-bar .period-select{{height:36px;border:1px solid #cbd5e1;border-radius:6px;padding:0 10px;font-size:14px;background:#fff;color:var(--ink);min-width:160px;}}
tr.period-hide{{display:none;}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;}}
.kpi{{border:1px solid #e4e7ec;border-radius:8px;padding:14px;background:#fafcff;}}
.kpi span{{display:block;color:var(--muted);font-size:12px;margin-bottom:6px;}}
.kpi strong{{display:block;font-size:22px;}}
.kpi .delta{{display:block;margin-top:4px;font-size:12px;color:var(--muted);}}
.kpi .delta.up{{color:#087443;}} .kpi .delta.down{{color:#b42318;}}
.table-wrap{{overflow:auto;border:1px solid #edf1f7;border-radius:6px;max-height:74vh;}}
table{{width:100%;border-collapse:collapse;font-size:13px;}}
th,td{{border-bottom:1px solid #edf1f7;padding:8px 10px;text-align:left;white-space:nowrap;vertical-align:top;}}
th{{background:#f1f6ff;font-weight:700;color:#344054;position:sticky;top:0;}}
th .th-label{{display:block;margin-bottom:6px;}}
th .col-filter{{display:block;width:100%;min-width:90px;font-weight:400;font-size:12px;padding:3px 6px;border:1px solid #cbd5e1;border-radius:4px;background:#fff;}}
.col-info{{border:none;background:none;color:var(--blue);cursor:pointer;font-size:13px;padding:0 4px;line-height:1;vertical-align:middle;}}
.col-info:hover{{color:#0b4fb0;}}
.col-note-pop{{position:absolute;z-index:60;max-width:340px;background:#102a43;color:#fff;font-size:12px;font-weight:400;line-height:1.45;padding:10px 12px;border-radius:8px;box-shadow:0 4px 14px rgba(16,42,67,.35);white-space:normal;}}
th.step,td.step{{width:50ch;min-width:50ch;max-width:50ch;white-space:normal;overflow-wrap:anywhere;word-break:break-word;}}
.exp-cell{{width:34px;text-align:center;}}
.expander{{border:none;background:none;cursor:pointer;font-size:11px;color:var(--blue);padding:2px 4px;line-height:1;}}
td.detail-cell{{padding:0;background:#f8fbff;}}
.detail-table{{width:100%;border-collapse:collapse;}}
.detail-table th{{position:static;background:#eef4ff;font-size:12px;}}
.detail-table th,.detail-table td{{border-bottom:1px solid #e6eefb;}}
tr.no-match{{display:none;}}
.empty{{padding:18px;color:var(--muted);}}
.echart{{width:100%;}}
.note{{margin:0 0 12px;color:var(--muted);font-size:12px;}}
.insights{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px;}}
.insight-card{{border:1px solid #e4e7ec;border-left:4px solid var(--blue);border-radius:8px;padding:12px 14px;background:#fafcff;}}
.insight-card.bad{{border-left-color:var(--red,#b42318);}} .insight-card.warn{{border-left-color:#b54708;}} .insight-card.good{{border-left-color:#087443;}}
.insight-card .label{{display:block;color:var(--muted);font-size:12px;margin-bottom:4px;}}
.insight-card .value{{display:block;font-size:15px;font-weight:650;overflow-wrap:anywhere;}}
td.num,td[data-sort]{{font-variant-numeric:tabular-nums;}}
td.flag-bad{{background:#fde7e6;color:#912018;font-weight:650;}} td.flag-warn{{background:#fdf0e3;color:#9c4a16;}}
th.sortable{{cursor:pointer;user-select:none;}} th.sortable:hover{{background:#e4eefb;}}
th.sort-asc .th-label::after,th.sort-desc .th-label::after{{font-size:10px;color:var(--blue);margin-left:4px;}}
th.sort-asc .th-label::after{{content:"\\25B2";}} th.sort-desc .th-label::after{{content:"\\25BC";}}
th.sort-asc::after{{content:" \\25B2";font-size:10px;color:var(--blue);}} th.sort-desc::after{{content:" \\25BC";font-size:10px;color:var(--blue);}}
tr.cross-hit > td{{background:#eaf2ff !important;box-shadow:inset 3px 0 0 var(--blue);}}
table.rule-table tbody tr.rule-row{{cursor:pointer;}}
.xfilter-chip{{display:none;position:fixed;right:18px;bottom:18px;z-index:50;background:#102a43;color:#fff;
padding:9px 14px;border-radius:18px;font-size:13px;cursor:pointer;box-shadow:0 2px 8px rgba(16,42,67,.3);}}
.xfilter-chip b{{font-weight:700;}}
.toc{{position:sticky;top:0;z-index:40;display:flex;gap:8px;overflow-x:auto;white-space:nowrap;
padding:10px 4px;margin:-8px 0 6px;background:rgba(245,247,251,.96);backdrop-filter:blur(3px);border-bottom:1px solid var(--line);}}
.toc a{{flex:0 0 auto;font-size:12px;color:#344054;text-decoration:none;padding:5px 10px;border:1px solid var(--line);border-radius:14px;background:#fff;}}
.toc a:hover{{border-color:var(--blue);color:var(--blue);}}
.panel{{scroll-margin-top:52px;}}
</style></head><body>
<script src="/static/vendor/echarts.min.js"></script>
<header><h1>{html.escape(report_title)}</h1><p>Snapshot {html.escape(snapshot_pt_date)}.{coverage} Generated {generated_at} from Data Workbench output.</p></header>
<main>{body}</main>"""
    return body_html + "\n<script>\n" + _DASHBOARD_JS + "\n</script>\n</body></html>"


# --- Scenarios report: column explanations (info-icon popovers) + auth-step glossary ---
_SCENE_COL_NOTES = {
    "source": (
        "Channel the scene originates from (SceneSourceEnum): 1 = Seabank app (BANKING_APP), "
        "2 = Seabank SDK embedded in Shopee / ShopeePay (BANKING_SDK), 3 = Admin / internal, "
        "4 = Other (e.g. loans, money-in). Sources 1 and 2 are authentication scenes."
    ),
    "mode": (
        "Only set on logic (aggregate) scenes; blank for real scenes. 1 = include — the scene equals "
        "this explicit list of real scenes (e.g. 'All Seabank loans'). 2 = exclude — the scene equals "
        "all real scenes except a list (e.g. 'All except HelpCenterFacialVerification')."
    ),
}
_DROPOFF_COL_NOTES = {
    "drop_off_rate_pct": (
        "Cross-day, flow-level. Each flow (bizflow_instance_id) is followed across days AND scenes and "
        "attributed to its entry scene (the scene of its first action); 'completed' = the flow's terminal "
        "action (latest timestamp, in any scene) succeeded (action_status = 1). Drop-off = flows that did "
        "not complete. This replaces the per-day is_final_action_in_flow_of_the_day flag, which only marks "
        "the last action of the day and 'may not be the final action in the design of the bizflow' - so "
        "intermediate scenes (e.g. DPDataTopUp, whose flow continues to DPOrderPaid) no longer show a "
        "false ~100% drop-off. CAVEAT: drop-off only means 'user did not finish authenticating' for genuine "
        "multi-step auth flows (Login, Transfer, etc.). A scene that is itself a terminal-NEGATIVE or "
        "single-action event - e.g. CardInvalidTxn (an invalid-card transaction that is declined by design) "
        "- has no successful terminal action to begin with, so it reads ~100%. That is the scene being "
        "blocked/declined, NOT customers abandoning; read those scenes as 'block rate', not 'abandonment'."
    ),
    "flows_completed": (
        "Distinct flows whose terminal action (the latest action across all days and scenes) succeeded "
        "(action_status = 1)."
    ),
}
# Auth-step codes from the risk-decision engine (StepEnum, dbp-antifraud-common). 'type' is the engine's
# own classification: normal = a real (usually user-facing) step; internal = orchestration step with no
# user interaction. In the flow-config columns (default_step / challenge*_step) a cell may also be 'NA'
# (no step), 'EXP:…' (a runtime rule that selects the step), or comma-separated codes (run in sequence).
_AUTH_STEP_GLOSSARY = [
    ("BE", "Flow start", "normal"),
    ("BD", "Flow complete (final / done)", "normal"),
    ("BSTD", "Flow aborted / terminated", "normal"),
    ("BP", "Verify PIN", "normal"),
    ("BPW", "Verify password", "normal"),
    ("BSO", "Verify SMS OTP", "normal"),
    ("BSON", "Verify SMS OTP for a new phone (user enters the number)", "normal"),
    ("BEO", "Verify Email OTP", "normal"),
    ("BO", "SMS or Email OTP (user chooses)", "normal"),
    ("BOP", "Verify one-time PIN", "normal"),
    ("BOPW", "Verify one-time password", "normal"),
    ("BAOP", "Send one-time PIN", "normal"),
    ("BAOPW", "Send one-time password", "normal"),
    ("BBIO", "TouchID / FaceID (device biometrics)", "normal"),
    ("BBIOST", "TouchID / FaceID + Soft Token", "normal"),
    ("BST", "Activate + verify Soft Token (standalone)", "normal"),
    ("BSTC", "Activate + verify Soft Token, return CA cert to client", "normal"),
    ("BSV", "Verify Soft Token only", "normal"),
    ("BPST", "PIN + Soft Token", "normal"),
    ("BPWST", "Password + Soft Token", "normal"),
    ("BSOST", "SMS OTP + Soft Token", "normal"),
    ("BPPWST", "PIN or password + Soft Token", "normal"),
    ("BPBIOST", "PIN or login TouchID / FaceID + Soft Token", "normal"),
    ("BTST", "Transaction biometric / PIN / password + Soft Token", "normal"),
    ("BLO", "Existing-user login (TouchID/FaceID or password) + step-up (varies by action)", "normal"),
    ("BLOST", "Existing-user login (TouchID/FaceID or password) + Soft Token (varies by action)", "normal"),
    ("BFV", "Facial verification — ALC dynamic-light (Aurora) liveness", "normal"),
    ("BSFV", "Facial verification — SLC static liveness", "normal"),
    ("BLC", "Liveness check (standalone)", "normal"),
    ("BPFV", "PIN or facial verification", "normal"),
    ("BPWFV", "Password or facial verification", "normal"),
    ("BSOFV", "SMS OTP or facial verification", "normal"),
    ("BPSO", "PIN or SMS OTP", "normal"),
    ("BPBIO", "PIN or TouchID / FaceID", "normal"),
    ("BSGP", "SingPass facial recognition", "normal"),
    ("BPWSGP", "Password or SingPass facial recognition", "normal"),
    ("BPSGP", "PIN or SingPass facial recognition", "normal"),
    ("BNSFV", "SingPass-login facial verification", "normal"),
    ("BPNSFV", "SingPass-login facial or PIN", "normal"),
    ("BPWNSFV", "SingPass-login facial or password", "normal"),
    ("BDOB", "Date-of-birth (DOB) verification", "normal"),
    ("BND", "NIK + DOB verification", "normal"),
    ("BPND", "PIN or NIK + DOB verification", "normal"),
    ("BPDOB", "PIN or DOB verification", "normal"),
    ("BPWDOB", "Password or DOB verification", "normal"),
    ("BNRIC", "Verify NRIC", "normal"),
    ("BTBSO", "Transaction biometric or SMS OTP", "normal"),
    ("BKFI", "Shopee x Seabank FaceID login", "normal"),
    ("BKTI", "Shopee x Seabank TouchID login", "normal"),
    ("BH5P", "H5 (web) verify PIN", "normal"),
    ("BH5SO", "H5 (web) verify SMS OTP", "normal"),
    ("BPN", "Send push notification (PN)", "normal"),
    ("BPAN", "Send PN + AR message", "normal"),
    ("BPAV", "Verify card-auth PN + AR token", "normal"),
    ("BNO", "Chain-trigger notification-rule flow", "normal"),
    ("BJP", "Jump SDK -> app, keep SDK page open", "normal"),
    ("BJPD", "Jump SDK -> app, close SDK page", "normal"),
    ("BJA", "Jump SDK -> app, may pass verification", "normal"),
    ("BJSP", "Jump app -> Shopee SDK, may pass verification", "normal"),
    ("BUDL", "Link user-device info", "internal"),
    ("BULUL", "Uplift-unlock AF penalty list", "internal"),
    ("BSKP", "Skip PIN", "internal"),
    ("BRT", "Post rule-treatment operation", "internal"),
    ("NA", "No authentication step configured", "config"),
    ("EXP:…", "Runtime rule that picks the step by app version / PIN-vs-password / login type", "config"),
]


def _auth_step_glossary_panel() -> str:
    rows = [[code, meaning, step_type] for code, meaning, step_type in _AUTH_STEP_GLOSSARY]
    return _searchable_table_panel(
        "Auth Step Glossary",
        ["step_code", "meaning", "step_type"],
        rows,
        placeholder="Search step code or meaning…",
        column_notes={
            "step_code": (
                "Auth-step codes issued by the risk-decision engine (StepEnum). Comma-separated codes = "
                "multiple steps required in sequence; 'EXP:…' = a runtime rule that selects the step by "
                "app version / PIN-vs-password / login type."
            ),
            "step_type": (
                "Engine classification: normal = a real (usually user-facing) auth step; internal = "
                "orchestration step with no user interaction; config = a flow-config placeholder, not a "
                "real step."
            ),
        },
    )


# Facial-verification result codes from the authentication engine (LCResultEnum / SQAResultEnum /
# facial-match result). Used to build a plain-language glossary so the raw LC_/SQA_/FM_ codes in the
# breakdown sheets are readable. (group, code, meaning).
_FACIAL_RESULT_GLOSSARY = [
    ("Liveness (LC_)", "LC_SUCCESS", "Passed liveness"),
    ("Liveness (LC_)", "LC_FRAUD", "Liveness check flagged as fraud"),
    ("Liveness (LC_)", "LC_AURORA_SPOOF", "Aurora liveness detected a spoof (presentation attack)"),
    ("Liveness (LC_)", "LC_AURORA_BRIGHT", "Too bright — flagged for possible deepfake"),
    ("Liveness (LC_)", "LC_BLUR", "Face too blurry"),
    ("Liveness (LC_)", "LC_BLOCKED", "Face blocked / obstructed"),
    ("Liveness (LC_)", "LC_BRIGHT", "Face too bright"),
    ("Liveness (LC_)", "LC_DARK", "Face too dark"),
    ("Liveness (LC_)", "LC_EYE_INVALID", "Eyes not valid (closed / not detected)"),
    ("Liveness (LC_)", "LC_TIMEEXPIRED", "Liveness check timed out"),
    ("Liveness (LC_)", "LC_ALLRETRIESFAILED", "Timed out after all retries"),
    ("Liveness (LC_)", "LC_CAPTUREIMAGEFAILED", "Camera failed to capture an image"),
    ("Liveness (LC_)", "LC_NETWORK_ERROR", "Network issue during the check"),
    ("Liveness (LC_)", "LC_INTERACTION_FAIL", "User did not complete the interactive prompts"),
    ("Liveness (LC_)", "LC_NO_CAMERA_PERMISSION", "Camera permission not granted"),
    ("Liveness (LC_)", "LC_SIGNATURE_ERROR", "Image signature invalid — image was altered"),
    ("Liveness (LC_)", "LC_ERROR", "Unknown liveness error"),
    ("Anti-spoof QC (SQA_)", "SQA_SUCCESS", "Passed selfie anti-spoofing QC"),
    ("Anti-spoof QC (SQA_)", "SQA_REJECT_FACE_SPOOFING", "Detected a spoof face (photo/mask/replay)"),
    ("Anti-spoof QC (SQA_)", "SQA_REJECT_FACE_DEEPFAKE", "Face suspected to be a deepfake"),
    ("Anti-spoof QC (SQA_)", "SQA_REJECT_FACE_SHALLOWFAKE", "Face suspected to be a shallowfake (edited)"),
    ("Anti-spoof QC (SQA_)", "SQA_REJECT_FACE_SCREENSHOT", "Face suspected to be a screenshot"),
    ("Anti-spoof QC (SQA_)", "SQA_REJECT_FACE_DETECTION", "No face detected"),
    ("Anti-spoof QC (SQA_)", "SQA_REJECT_FACE_QC_FAILED_BRIGHT", "QC fail: face too bright"),
    ("Anti-spoof QC (SQA_)", "SQA_REJECT_FACE_QC_FAILED_DARK", "QC fail: face too dark"),
    ("Anti-spoof QC (SQA_)", "SQA_REJECT_FACE_QC_FAILED_BLUR", "QC fail: face too blurry"),
    ("Anti-spoof QC (SQA_)", "SQA_REJECT_FACE_QC_FAILED_BLOCKED", "QC fail: face blocked"),
    ("Anti-spoof QC (SQA_)", "SQA_REJECT_FACE_QC_FAILED_EYECLOSE", "QC fail: eyes closed"),
    ("Anti-spoof QC (SQA_)", "SQA_ERROR_IMAGE_IDENTICAL", "The two images submitted were identical"),
    ("Anti-spoof QC (SQA_)", "SQA_ERROR_FACE_DOWNLOAD", "Could not download the photo"),
    ("Anti-spoof QC (SQA_)", "SQA_ERROR_FACE_SIGNATURE_NONE", "Image had no signature"),
    ("Anti-spoof QC (SQA_)", "SQA_ERROR_HONEYPOT_TRIGGER", "Image signature contained a forbidden (honeypot) key"),
    ("Anti-spoof QC (SQA_)", "SQA_ERROR", "Unknown anti-spoofing error"),
    ("Face match (FM_)", "FM_SUCCESS", "Selfie matched the reference face"),
    ("Face match (FM_)", "FM_ERROR_FACE_MISMATCHED", "Selfie did not match the reference face"),
    ("Face match (FM_)", "FM_ERROR", "Unknown face-matching error"),
]


def _facial_result_glossary_panel() -> str:
    rows = [[group, code, meaning] for group, code, meaning in _FACIAL_RESULT_GLOSSARY]
    return _searchable_table_panel(
        "Result Code Glossary",
        ["step", "result_code", "meaning"],
        rows,
        placeholder="Search result code or meaning…",
        note="Plain-language meaning of the LC_ (liveness), SQA_ (anti-spoofing QC) and FM_ (face match) "
             "result codes used in the breakdown tables. Source: authentication-engine LCResultEnum / "
             "SQAResultEnum.",
        column_notes={
            "result_code": (
                "Engine result code. LC_ = liveness check, SQA_ = selfie anti-spoofing QC, FM_ = facial "
                "matching. SUCCESS = passed; REJECT_/ERROR_ = the failure reason."
            )
        },
    )


# Device-risk signal glossary (fmart_antifraud action-log boolean flags). Each flag = events where the
# device/identity signal fired. (signal, meaning).
_DEVICE_RISK_SIGNAL_GLOSSARY = [
    ("rooted", "Device is rooted (Android superuser access)"),
    ("emulator", "Running on an emulator, not a physical phone"),
    ("vpn", "A VPN is active on the device"),
    ("http proxy", "Traffic is routed through an HTTP proxy"),
    ("gps modified", "GPS location is being spoofed / mocked"),
    ("fake identity", "Identity signals look fabricated"),
    ("fake deviceinfo", "Device fingerprint looks fabricated / tampered"),
    ("illegal imei", "IMEI is invalid or blacklisted"),
    ("new deviceid", "First time this device id has been seen"),
    ("magisk", "Magisk (root-hiding framework) detected"),
    ("system debuggable", "OS build is debuggable (developer / modified build)"),
    ("risk app root", "A known rooting app is installed"),
    ("risk app vpn", "A known VPN app is installed"),
    ("risk app fake gps", "A known GPS-spoofing app is installed"),
    ("risk app hook", "A known hooking / instrumentation app is installed (e.g. Frida/Xposed)"),
    ("remote control", "A remote-control / screen-share accessibility service is enabled"),
    ("autoclicker", "An auto-clicker / automation service is enabled"),
]


def _device_risk_signal_glossary_panel() -> str:
    rows = [[signal, meaning] for signal, meaning in _DEVICE_RISK_SIGNAL_GLOSSARY]
    return _searchable_table_panel(
        "Risk Signal Glossary",
        ["risk_signal", "meaning"],
        rows,
        placeholder="Search signal or meaning…",
        note="What each device / identity risk signal means. A signal 'fires' for an event when the "
             "device telemetry sets that flag (e.g. rooted, emulator, VPN, fake GPS).",
    )


# Shared column info-notes for backend-coded columns, so raw enum codes are explained inline.
_RULES_COL_NOTES = {
    "rule_status": (
        "Engine RuleStatus: Active = live and enforcing its outcome; Collect Data = dry-run / shadow "
        "(the rule evaluates and logs hits but does NOT enforce its action — used to observe a rule before "
        "turning it on); Inactive = off. Only Active rules actually block/challenge/punish in production."
    ),
    "status_code": (
        "Raw rule status integer from rule_config: 1 = Active, 2 = Collect Data (dry-run/shadow), "
        "-1 = Inactive. See the rule_status column for the label."
    ),
    "review_priority": "Manual review priority assigned to cases this rule challenges (higher = reviewed sooner).",
    "punish_length_sec": (
        "Configured punishment duration in seconds (rule_config.punish_length). Negative = no limit "
        "(permanent block); empty/null = not a punish rule. The punish_duration column shows it in hours."
    ),
    "punish_duration": "Human-readable punishment duration derived from punish_length_sec.",
    "transify_key": (
        "Transify (i18n) key configured for the rejection — resolves to the localized error message / "
        "error code shown to the customer when this rule rejects. Empty for rules that don't reject."
    ),
}
_FEATURES_COL_NOTES = {
    "feature_type": "Engine feature type code (rule_config feature taxonomy). 1 = the standard metric-based feature.",
    "event_status": (
        "Engine EventStatus the feature counts: 0 = Fail events only, 1 = Success events only, 2 = All "
        "events (pass + fail)."
    ),
    "scenario_type": (
        "ProtectRule scenario scope: 0 = Exclusively (this scene only), 1 = Inclusively (this scene and "
        "its children)."
    ),
    "business_category": "Customer segment: 0 = Retail, 1 = Corporate.",
    "function_id": (
        "Metric function the feature calls (e.g. F1, F10). The function's calculation logic lives in the "
        "engine's function catalog / design docs; the Function Usage sheet lists how each is used."
    ),
    "consecutive": "Whether the feature requires consecutive occurrences (Y) or any occurrences (N).",
}
_REJECT_TYPE_NOTE = (
    "Engine RejectType family (output class): 1 = hard reject (black-list / punish-list / realtime "
    "reject), 2 = challenge, 0 = pass (notification / white-list / normal)."
)
_AUTH_TYPE_NOTE = (
    "Authentication tier the risk engine assigned to the action (source AuthenTypeEnum). DEFAULT (L0) = "
    "engine required no step-up auth factor; CHALLENGE_1/2/3 (L1-L3) = step-up challenge tiers of "
    "increasing friction. A blank value is shown as 'Non-interactive (engine-scored)': the action "
    "carried no challenge tier - a business action, or an engine decision with no user-facing challenge. "
    "Note 'engine-scored' is this report's label for the blank case, not a value from the source enum."
)


def write_visualization(
    path: Path,
    *,
    report_title: str,
    snapshot_pt_date: str,
    sheets: list[tuple[str, list[str], list[list[Any]]]],
    report_id: str = "",
) -> None:
    if report_id == AF_SCENARIOS_ACTIONS_REPORT_ID:
        sc_lookup = {sheet_name: (headers, rows) for sheet_name, headers, rows in sheets}
        flow = sc_lookup.get("Scenario Action Auth Flow")
        fheaders, frows = flow if flow else ([], [])

        def _distinct(col: str):
            idx = next((i for i, h in enumerate(fheaders) if str(h) == col), None)
            if idx is None:
                return None
            return len({str(r[idx]) for r in frows if idx < len(r) and r[idx] not in (None, "")})

        # Authentication-outcome KPIs (period-switchable) lead; scenario-coverage KPIs follow.
        auth_summary = sc_lookup.get("Authentication Outcome Summary")
        auth_kpi = _period_kpi_panel(
            "Authentication Outcomes",
            auth_summary,
            [
                ("Actions", "actions", "num"), ("Pass", "pass_actions", "num"),
                ("Reject", "reject_actions", "num"), ("Reject rate", "reject_rate_pct", "pct"),
                ("Flows", "flows", "num"), ("Users", "distinct_users", "num"),
            ],
        )
        coverage_kpis: list[tuple] = []
        if frows:
            coverage_kpis.append(("Flow mappings", _format_number(len(frows))))
        for label, col in (("L1 scenes", "l1_scene_name"), ("L2 sub-scenes", "l2_sub_scene_name"), ("Actions", "action_name")):
            distinct = _distinct(col)
            if distinct:
                coverage_kpis.append((label, _format_number(distinct)))
        # Auto-insights: highest reject-rate scene/auth-type, step-up challenge load, worst drop-off.
        insight_cards: list[tuple[str, str, str]] = []
        by_scene = sc_lookup.get("Auth Outcome by Scene & Type")
        if by_scene and by_scene[1]:
            bc = {str(c): i for i, c in enumerate(by_scene[0])}
            if {"scene_name", "authentication_type", "reject_rate_pct", "actions"} <= set(bc):
                sized = [r for r in by_scene[1] if _number(r[bc["actions"]] if bc["actions"] < len(r) else 0) >= 1000]
                if sized:
                    worst = max(sized, key=lambda r: _number(r[bc["reject_rate_pct"]] if bc["reject_rate_pct"] < len(r) else 0))
                    rr = _number(worst[bc["reject_rate_pct"]])
                    insight_cards.append((
                        "Highest reject scene (≥1k actions)",
                        f"{worst[bc['scene_name']]} / {worst[bc['authentication_type']]} — {rr}%",
                        "bad" if rr >= 5 else ("warn" if rr >= 1 else ""),
                    ))
        friction = sc_lookup.get("Challenge Friction by Auth Type")
        if friction and friction[1]:
            fc = {str(c): i for i, c in enumerate(friction[0])}
            if "authentication_type" in fc and "action_share_pct" in fc:
                challenge = sum(
                    _number(r[fc["action_share_pct"]]) for r in friction[1]
                    if fc["action_share_pct"] < len(r) and str(r[fc["authentication_type"]]).upper().startswith("CHALLENGE")
                )
                if challenge > 0:
                    insight_cards.append(("Step-up challenge load", f"{round(challenge, 2)}% of actions", "warn" if challenge >= 10 else ""))
        dropoff = sc_lookup.get("Auth Drop-off by Scene")
        if dropoff and dropoff[1]:
            dc = {str(c): i for i, c in enumerate(dropoff[0])}
            if {"scene_name", "drop_off_rate_pct", "flows_started"} <= set(dc):
                sized = [r for r in dropoff[1] if _number(r[dc["flows_started"]] if dc["flows_started"] < len(r) else 0) >= 1000]
                if sized:
                    worst = max(sized, key=lambda r: _number(r[dc["drop_off_rate_pct"]] if dc["drop_off_rate_pct"] < len(r) else 0))
                    dr = _number(worst[dc["drop_off_rate_pct"]])
                    insight_cards.append(("Worst drop-off scene", f"{worst[dc['scene_name']]} — {dr}%", "warn" if dr >= 20 else ""))
        intro = _insight_panel("Highlights", insight_cards) + auth_kpi + _kpi_cards_panel("Scenario Coverage", coverage_kpis)
        # Funnel panels first (the new authentication view), then the scenario flow + config catalogs.
        panels: list[str] = []
        placeholders = {
            "Auth Outcome by Scene & Type": "Search scene or auth type…",
            "Auth Drop-off by Scene": "Search scene…",
            "Auth Step Success by Method": "Search method…",
            "L1 Scenarios": "Search scene…",
            "L2 Sub-Scenarios": "Search sub-scene…",
            "Actions and Auth Steps": "Search action…",
        }
        ordered = [
            "Auth Outcome by Scene & Type",
            "Challenge Friction by Auth Type",
            "Auth Drop-off by Scene",
            "Auth Step Success by Method",
            "Scenario Action Auth Flow",
            "L1 Scenarios",
            "L2 Sub-Scenarios",
            "Actions and Auth Steps",
        ]
        for sheet_name in ordered:
            table = sc_lookup.get(sheet_name)
            if not table:
                continue
            headers, rows = table
            if sheet_name == "Challenge Friction by Auth Type":
                fc = {str(c): i for i, c in enumerate(headers)}
                if "authentication_type" in fc and "reject_rate_pct" in fc:
                    pairs = [
                        (str(r[fc["authentication_type"]]), r[fc["reject_rate_pct"]])
                        for r in rows
                        if fc["authentication_type"] < len(r) and fc["reject_rate_pct"] < len(r)
                    ]
                    bar = _bar_panel("Reject Rate by Auth Type (%)", pairs, note="Reject rate per authentication tier (DEFAULT vs step-up challenges).")
                    if bar:
                        panels.append(bar)
                panels.append(_searchable_table_panel(
                    sheet_name, headers, rows, placeholder="Search auth type…",
                    column_notes={"authentication_type": _AUTH_TYPE_NOTE}))
                continue
            if sheet_name == "Scenario Action Auth Flow":
                step_cols = {i for i, h in enumerate(headers) if str(h).strip().lower().endswith("_step")}
                panels.append(_searchable_table_panel(sheet_name, headers, rows, placeholder="Search scene name or step…", step_columns=step_cols))
                continue
            notes = _SCENE_COL_NOTES if sheet_name == "L1 Scenarios" else (
                _DROPOFF_COL_NOTES if sheet_name == "Auth Drop-off by Scene" else None
            )
            table_note = {
                "Auth Outcome by Scene & Type": (
                    "Risk-evaluated actions only (risk_result pass/reject) - actions with no risk "
                    "decision are excluded, so totals sit below the Authentication Outcomes summary."
                ),
                "Auth Step Success by Method": (
                    "Per-method pass/fail of each authentication step (action_status: 1 = success, "
                    "0 = failed) over the window. This is whether the step ITSELF passed (correct OTP, "
                    "biometric match, right PIN/password), not the risk-engine decision (risk_result) "
                    "used by the sections above. SoftToken / PIN dominate volume; SMS-OTP and Password "
                    "carry the most friction; Email-OTP and one-time pin/pwd are near-zero (legacy) in PH."
                ),
            }.get(sheet_name, "")
            if notes is None and sheet_name == "Auth Outcome by Scene & Type":
                notes = {"authentication_type": _AUTH_TYPE_NOTE}
            panels.append(_searchable_table_panel(
                sheet_name, headers, rows, placeholder=placeholders.get(sheet_name, "Search…"), column_notes=notes,
                note=table_note,
            ))
        panels.append(_auth_step_glossary_panel())
        path.write_text(
            _searchable_tables_document(report_title, snapshot_pt_date, panels, intro_html=intro, data_through=_data_through(sheets)),
            encoding="utf-8",
        )
        return
    if report_id == AF_RULES_FEATURES_REPORT_ID:
        flow_lookup = {sheet_name: (headers, rows) for sheet_name, headers, rows in sheets}
        rules = flow_lookup.get("Rules")
        features = flow_lookup.get("Features")
        # Raw engine codes that older artifacts (or unmapped values) leave in the data; charts show the
        # OutcomeType/FeatureStatus meaning instead of a bare digit.
        outcome_code_labels = {"4": "Reject+Punish", "5": "Challenge+Punish", "6": "Pass", "3": "Notification"}
        feature_status_labels = {"2": "Retired (status 2)", "1": "Active", "-1": "Inactive"}

        def _relabel(agg: dict[str, float], labels: dict[str, str]) -> dict[str, float]:
            relabelled: dict[str, float] = {}
            for key, count in agg.items():
                label = labels.get(key, key)
                relabelled[label] = relabelled.get(label, 0.0) + count
            return relabelled

        def _count_by(table, col):
            if not table:
                return {}
            hdrs, rws = table
            idx = next((i for i, h in enumerate(hdrs) if str(h) == col), None)
            if idx is None:
                return {}
            agg: dict[str, float] = {}
            for r in rws:
                if idx < len(r) and r[idx] not in (None, ""):
                    key = str(r[idx])
                    agg[key] = agg.get(key, 0.0) + 1
            return agg

        kpis: list[tuple] = []
        if rules and rules[1]:
            status = _count_by(rules, "rule_status")
            kpis.append(("Total rules", _format_number(len(rules[1]))))
            if status.get("Active"):
                kpis.append(("Active rules (enforcing)", _format_number(status["Active"])))
            if status.get("Collect Data"):
                kpis.append(("Collect Data (shadow)", _format_number(status["Collect Data"])))
        if features and features[1]:
            fstatus = _count_by(features, "feature_status")
            kpis.append(("Total features", _format_number(len(features[1]))))
            if fstatus.get("Active"):
                kpis.append(("Active features", _format_number(fstatus["Active"])))
        func_usage = flow_lookup.get("Function Usage")
        if func_usage and func_usage[1]:
            kpis.append(("Functions in use", _format_number(len(func_usage[1]))))
        intro = _kpi_cards_panel("Catalog Summary", kpis)
        # Governance (folded-in rule change log): KPIs + highlights from the rule_config snapshot diff.
        chg_counts: dict[str, float] = {}
        chg_summary = flow_lookup.get("Change Summary")
        if chg_summary and chg_summary[1]:
            sc = {str(c): i for i, c in enumerate(chg_summary[0])}
            if "change_type" in sc and "rules" in sc:
                for r in chg_summary[1]:
                    if sc["change_type"] < len(r) and sc["rules"] < len(r):
                        chg_counts[str(r[sc["change_type"]])] = _number(r[sc["rules"]])
        chg_insights: list[tuple[str, str, str]] = []
        net = (chg_counts.get("Added", 0) + chg_counts.get("Activated", 0)) - (chg_counts.get("Removed", 0) + chg_counts.get("Deactivated", 0))
        if chg_counts:
            sign = "+" if net >= 0 else ""
            chg_insights.append(("Net change in active rules", f"{sign}{_format_number(net)}", "good" if net > 0 else ("warn" if net < 0 else "")))
        if chg_counts.get("Deactivated"):
            chg_insights.append(("Rules deactivated", _format_number(chg_counts["Deactivated"]), "warn"))
        if chg_counts.get("Logic changed"):
            chg_insights.append(("Rules re-logicked", _format_number(chg_counts["Logic changed"]), "warn"))
        chg_kpis = [
            (label, _format_number(chg_counts[label]))
            for label in ("Added", "Activated", "Deactivated", "Removed", "Outcome changed", "Risk level changed", "Logic changed", "Priority changed")
            if chg_counts.get(label)
        ]
        # Annotate the comparison window the change-log diff uses (current snapshot vs the rule_config
        # snapshot at the reporting-window start, ~2 months back). Baseline matches build_af_rule_change_log_sql.
        from bpmis_jira_tool.business_insights import business_insights_window  # local import: window helper
        _chg_baseline = business_insights_window(datetime.now(UTC)).span_start.isoformat()
        chg_window_note = (
            f"Comparison window: the current rule_config snapshot ({snapshot_pt_date}) "
            f"vs the snapshot at the start of the reporting window (≈ {_chg_baseline} — the first day of the "
            f"earliest of the three reporting months, about two months back). Rules present only now show as "
            f"'Added'; only in the baseline as 'Removed'; etc. If snapshot history is shorter than that span, "
            f"the baseline falls back to the earliest retained snapshot, which can inflate 'Added'."
        )
        chg_window_panel = (
            f'<section class="panel"><h2>Rule Change Window</h2><p class="note">{html.escape(chg_window_note)}</p></section>'
            if chg_counts else ""
        )
        intro = intro + chg_window_panel + _insight_panel("Rule Change Highlights", chg_insights) + _kpi_cards_panel("Rule Change Summary", chg_kpis)

        chart_panels: list[str] = []
        outcome = _relabel(_count_by(rules, "outcome_type"), outcome_code_labels)
        if outcome:
            donut = _donut_panel(
                "Rules by Outcome Type",
                sorted(outcome.items(), key=lambda kv: kv[1], reverse=True),
                note="Configured rules split by enforcement outcome (engine OutcomeType enum).",
            )
            if donut:
                chart_panels.append(donut)
        risk = _count_by(rules, "risk_level")
        if risk:
            bar = _bar_panel(
                "Rules by Risk Level",
                sorted(risk.items(), key=lambda kv: kv[1], reverse=True),
                note="Configured rule count per risk level.",
            )
            if bar:
                chart_panels.append(bar)
        fstatus_chart = _relabel(_count_by(features, "feature_status"), feature_status_labels)
        if fstatus_chart:
            fbar = _bar_panel(
                "Features by Status",
                sorted(fstatus_chart.items(), key=lambda kv: kv[1], reverse=True),
                note="Feature config rows per status. Only Active (status 1) is loaded by the engine; Inactive is -1; 'Retired (status 2)' rows are superseded config kept in the snapshot but not live (1=valid / 2=invalid convention).",
            )
            if fbar:
                chart_panels.append(fbar)
        if func_usage and func_usage[1]:
            uc = {str(c): i for i, c in enumerate(func_usage[0])}
            if "function_id" in uc and "features" in uc:
                pairs = [
                    (str(r[uc["function_id"]]), r[uc["features"]])
                    for r in func_usage[1]
                    if uc["function_id"] < len(r) and uc["features"] < len(r)
                ]
                ubar = _bar_panel(
                    "Top Functions by Feature Count",
                    sorted(pairs, key=lambda kv: _number(kv[1]), reverse=True)[:15],
                    note="Metric functions (function_id) ranked by how many configured features use them.",
                )
                if ubar:
                    chart_panels.append(ubar)

        outcome_note = (
            "Engine OutcomeType enum: code 1 = Reject when the rule is real-time, Punish when batch; "
            "2 = Challenge; 3 = Notification; 4 = Reject+Punish; 5 = Challenge+Punish; 6 = Pass. "
            "Artifacts generated before 2026-06-13 labelled code 3 as 'Reject' and real-time code-1 "
            "rules as 'Punish'; bare digits are codes the label map did not cover."
        )
        feature_status_note = (
            "Engine FeatureStatus enum defines only 1 = Active and -1 = Inactive; the rule engine loads "
            "ONLY status = 1. status = 2 is not in that enum — it is a retired/superseded config row "
            "(the codebase's wider convention is 1 = 生效/valid, 2 = 失效/invalid), kept in the snapshot "
            "but not live. So only the Active (1) count reflects features currently in force."
        )
        table_panels: list[str] = []
        if rules:
            table_panels.append(
                _searchable_table_panel(
                    "Rules", rules[0], rules[1], placeholder="Search rule id or name…",
                    column_notes={"outcome_type": outcome_note, **_RULES_COL_NOTES},
                )
            )
        if features:
            table_panels.append(
                _searchable_table_panel(
                    "Features", features[0], features[1], placeholder="Search feature id or name…",
                    column_notes={"feature_status": feature_status_note, **_FEATURES_COL_NOTES},
                )
            )
        if func_usage:
            table_panels.append(
                _searchable_table_panel(
                    "Function Usage", func_usage[0], func_usage[1], placeholder="Search function id…",
                    column_notes={"function_id": _FEATURES_COL_NOTES["function_id"]},
                )
            )
        two_way = flow_lookup.get("Two-Way Communication Config")
        if two_way and two_way[1]:
            table_panels.append(
                _searchable_table_panel(
                    "Two-Way Communication Config", two_way[0], two_way[1],
                    placeholder="Search template, rule or treatment…",
                    note="Online config of the Two-Way Communication treatment: the customer is asked to "
                         "approve/reject a flagged transaction within a confirmation window before it "
                         "settles. One row per template × relation (treatment); *_sec columns are seconds.",
                )
            )
        two_way_act = flow_lookup.get("Two-Way Communication Activity")
        if two_way_act and two_way_act[1]:
            ah, ar = two_way_act
            ac = {str(c): i for i, c in enumerate(ah)}
            if "template_id" in ac and "triggered" in ac:
                trig_pairs = [
                    (str(r[ac["template_id"]]), _number(r[ac["triggered"]]))
                    for r in ar if ac["triggered"] < len(r)
                ]
                tbar = _bar_panel(
                    "Two-Way Triggers by Template",
                    sorted(trig_pairs, key=lambda kv: kv[1], reverse=True)[:15],
                    note="Two-way confirmations triggered per template over the reporting window.",
                )
                if tbar:
                    chart_panels.append(tbar)
            table_panels.append(
                _searchable_table_panel(
                    "Two-Way Communication Activity", two_way_act[0], two_way_act[1],
                    placeholder="Search template…",
                    note="How often each two-way template fired and how customers responded over the "
                         "reporting window. approved=status 1+5, rejected=2+6, expired_no_response=3, "
                         "pending=0 (awaiting customer); response_rate=responded/triggered, "
                         "approval_rate=approved/responded.",
                )
            )
        # Governance panels (folded-in rule change log): current-inventory bar + change-log tables.
        governance_panels: list[str] = []
        inventory = flow_lookup.get("Current Rule Inventory")
        if inventory and inventory[1]:
            ih, ir = inventory
            ic = {str(c): i for i, c in enumerate(ih)}
            if "outcome_type" in ic and "rules" in ic:
                agg: dict[str, float] = {}
                for r in ir:
                    key = str(r[ic["outcome_type"]]) if ic["outcome_type"] < len(r) else ""
                    agg[key] = agg.get(key, 0.0) + _number(r[ic["rules"]] if ic["rules"] < len(r) else 0)
                inv_bar = _bar_panel(
                    "Current Rules by Outcome Type",
                    sorted(_relabel(agg, outcome_code_labels).items(), key=lambda kv: kv[1], reverse=True),
                    note="Active + inactive rule count per outcome type in the current snapshot.",
                )
                if inv_bar:
                    governance_panels.append(inv_bar)
        chg_column_notes = {
            "Rule Change Detail": {"outcome_before": outcome_note, "outcome_after": outcome_note},
            "Current Rule Inventory": {"outcome_type": outcome_note},
        }
        chg_table_notes = {"Change Summary": chg_window_note, "Rule Change Detail": chg_window_note}
        for chg_sheet in ("Change Summary", "Rule Change Detail", "Current Rule Inventory"):
            sec = flow_lookup.get(chg_sheet)
            if sec:
                placeholder = "Search rule id, name or change type…" if chg_sheet == "Rule Change Detail" else (
                    "Search outcome, status or risk…" if chg_sheet == "Current Rule Inventory" else "Filter…"
                )
                governance_panels.append(_searchable_table_panel(
                    chg_sheet, sec[0], sec[1], placeholder=placeholder,
                    column_notes=chg_column_notes.get(chg_sheet),
                    note=chg_table_notes.get(chg_sheet, ""),
                ))
        path.write_text(
            _searchable_tables_document(
                report_title, snapshot_pt_date, chart_panels + table_panels + governance_panels,
                intro_html=intro, data_through=_data_through(sheets),
            ),
            encoding="utf-8",
        )
        return
    if report_id == AF_RULE_EFFECTIVENESS_REPORT_ID:
        eff_lookup = {sheet_name: (headers, rows) for sheet_name, headers, rows in sheets}
        # KPI cards: latest full month headline + delta vs the prior full month, switchable by month.
        cur = None
        cur_action_rate = None
        summary = eff_lookup.get("Request Outcome Summary")
        if summary and summary[1]:
            cur, _prev, _cur_label, _prev_label, _mtd, _scope = _period_rows(summary)
            scols = {str(c): i for i, c in enumerate(summary[0])}
            cur_action_rate = cur[scols["action_rate_pct"]] if cur is not None and "action_rate_pct" in scols and scols["action_rate_pct"] < len(cur) else None
        kpi_panel = _period_kpi_panel(
            "Hit-Rate Summary",
            summary,
            [
                ("Total outcomes", "total_outcomes", "num"), ("Pass", "pass_num", "num"),
                ("Challenge", "challenge_num", "num"), ("Reject", "reject_num", "num"),
                ("Action rate", "action_rate_pct", "pct"),
            ],
        )
        # Auto-insight highlights from the scorecard (most over-firing rule, lowest precision).
        insight_cards: list[tuple[str, str, str]] = []
        scorecard = eff_lookup.get("Rule Scorecard")
        if scorecard and scorecard[1]:
            sc_cols = {str(c): i for i, c in enumerate(scorecard[0])}

            def _scell(row, col):
                return row[sc_cols[col]] if col in sc_cols and sc_cols[col] < len(row) else None

            scored = [r for r in scorecard[1] if _scell(r, "trigger_rate_pct") not in (None, "")]
            if scored:
                top = max(scored, key=lambda r: _number(_scell(r, "trigger_rate_pct")))
                tr = _number(_scell(top, "trigger_rate_pct"))
                insight_cards.append((
                    "Most over-firing rule",
                    f"{_scell(top, 'rule_id')} — {tr}% of scene traffic",
                    "bad" if tr >= 20 else ("warn" if tr >= 5 else ""),
                ))
            reviewed = [r for r in scorecard[1] if _number(_scell(r, "reviewed_cases")) >= 50 and _scell(r, "precision_pct") not in (None, "")]
            if reviewed:
                worst = min(reviewed, key=lambda r: _number(_scell(r, "precision_pct")))
                pp = _number(_scell(worst, "precision_pct"))
                insight_cards.append((
                    "Lowest precision (≥50 reviews)",
                    f"{_scell(worst, 'rule_id')} — {pp}% fraud-confirmed",
                    "bad" if pp < 1 else ("warn" if pp < 5 else "good"),
                ))
        if cur_action_rate not in (None, ""):
            ar = _number(cur_action_rate)
            insight_cards.append(("Action rate (challenge+reject)", f"{cur_action_rate}%", "warn" if ar >= 10 else ""))
        intro = _insight_panel("Highlights", insight_cards) + kpi_panel
        # Detection effectiveness (folded in): coverage KPIs + blind-spot / top-rule highlights.
        det_summary = eff_lookup.get("Detection Coverage Summary")
        det_kpi = _period_kpi_panel(
            "Detection Coverage",
            det_summary,
            [
                ("Detection rate", "detection_rate_pct", "pct"),
                ("Confirmed-fraud loss", "fraud_loss_php", "money"),
                ("Loss rule-detected", "loss_rule_detected_php", "money"),
                ("Loss leaked", "loss_leaked_php", "money"),
                ("Loss detected share", "loss_detected_share_pct", "pct"),
            ],
        ) if det_summary else ""
        det_insights: list[tuple[str, str, str]] = []
        if det_summary and det_summary[1]:
            d_cur, _dp, d_label, _dpl, _dm, _ds = _period_rows(det_summary)
            dcols = {str(c): i for i, c in enumerate(det_summary[0])}
            dr = d_cur[dcols["detection_rate_pct"]] if d_cur is not None and "detection_rate_pct" in dcols and dcols["detection_rate_pct"] < len(d_cur) else None
            if dr not in (None, ""):
                drn = _number(dr)
                det_insights.append((f"Detection rate ({d_label})", f"{dr}%", "good" if drn >= 80 else ("warn" if drn >= 60 else "bad")))
        det_by_mo = eff_lookup.get("Detection by Fraud MO Type")
        if det_by_mo and det_by_mo[1]:
            mc = {str(c): i for i, c in enumerate(det_by_mo[0])}
            if "fraud_mo_type" in mc and "loss_leaked_php" in mc:
                worst = max(det_by_mo[1], key=lambda r: _number(r[mc["loss_leaked_php"]] if mc["loss_leaked_php"] < len(r) else 0))
                lv = _number(worst[mc["loss_leaked_php"]])
                if lv > 0:
                    det_insights.append(("Biggest blind spot", f"{worst[mc['fraud_mo_type']]} — ₱{_format_number(lv)} leaked", "bad"))
        det_top = eff_lookup.get("Top Detecting Rules")
        if det_top and det_top[1]:
            tc = {str(c): i for i, c in enumerate(det_top[0])}
            if "rule_id" in tc and "loss_caught_php" in tc:
                best = det_top[1][0]
                cv = _number(best[tc["loss_caught_php"]] if tc["loss_caught_php"] < len(best) else 0)
                if cv > 0:
                    det_insights.append(("Top detecting rule", f"{best[tc['rule_id']]} — ₱{_format_number(cv)} caught", "good"))
        intro = intro + _insight_panel("Detection Highlights", det_insights) + det_kpi
        placeholders = {
            "Request Outcome Summary": "Filter…",
            "Daily Challenge/Reject/Punish": "Search date…",
            "Scene/Sub-scene/Action Usage": "Search scene, sub-scene or action…",
            "Top Detecting Rules": "Search rule id or name…",
            "Missed Fraud (Blind Spots)": "Search fraud type…",
            "Detection Coverage Summary": "Filter…",
        }
        reject_breakdown = eff_lookup.get("Reject Rule Scene Breakdown")
        punish_breakdown = eff_lookup.get("Punishment Rule Scene Breakdown")
        challenge_breakdown = eff_lookup.get("Challenge Rule Scene Breakdown")
        # Breakdown sheets are nested into their summary panels, not rendered standalone.
        nested_sheets = {"Reject Rule Scene Breakdown", "Punishment Rule Scene Breakdown", "Challenge Rule Scene Breakdown"}
        panels = []
        for sheet_name, headers, rows in sheets:
            if sheet_name in nested_sheets:
                continue
            if sheet_name == "Reject Rule Hit Summary" and reject_breakdown:
                panels.append(
                    _expandable_rule_panel(
                        "Reject Rule Hit Summary",
                        "Search rule, type or scene…",
                        (headers, rows),
                        reject_breakdown,
                        key_columns=("reject_rule", "reject_type"),
                        main_columns=(
                            "reject_rule",
                            "rule_name",
                            "reject_type",
                            "reject_count",
                            "distinct_users",
                            "distinct_scenes",
                            "rejected_amount_php",
                            "benchmark_trxn",
                            "trigger_rate_pct",
                            "normalised_user_impact_pct",
                        ),
                        detail_columns=(
                            "scene_name",
                            "reject_count",
                            "distinct_users",
                            "rejected_amount_php",
                            "benchmark_trxn",
                            "trigger_rate_pct",
                            "normalised_user_impact_pct",
                        ),
                        name_column="scene_name",
                        note=(
                            "reject_type is the engine RejectType family: 1 = hard reject (black-list / "
                            "punish-list / realtime reject). Expand a rule to see its scene breakdown."
                        ),
                    )
                )
                continue
            if sheet_name == "Punishment Rule Hit Summary" and punish_breakdown:
                panels.append(
                    _expandable_rule_panel(
                        "Punishment Rule Hit Summary",
                        "Search rule or scene…",
                        (headers, rows),
                        punish_breakdown,
                        key_columns=("punish_rule_id",),
                        main_columns=(
                            "punish_rule_id",
                            "rule_name",
                            "punish_count",
                            "distinct_targets",
                            "distinct_scenes",
                            "benchmark_trxn",
                            "trigger_rate_pct",
                            "normalised_user_impact_pct",
                        ),
                        detail_columns=(
                            "scene_name",
                            "punish_count",
                            "distinct_targets",
                            "benchmark_trxn",
                            "trigger_rate_pct",
                            "normalised_user_impact_pct",
                        ),
                        name_column="scene_name",
                    )
                )
                continue
            if sheet_name == "Challenge Rule Hit Summary" and challenge_breakdown:
                panels.append(
                    _expandable_rule_panel(
                        "Challenge Rule Hit Summary",
                        "Search rule or scene…",
                        (headers, rows),
                        challenge_breakdown,
                        key_columns=("rule_id",),
                        main_columns=(
                            "rule_id",
                            "rule_name",
                            "rule_status",
                            "review_priority",
                            "challenge_trxn",
                            "challenge_users",
                            "distinct_scenes",
                            "benchmark_trxn",
                            "trigger_rate_pct",
                            "normalised_user_impact_pct",
                        ),
                        detail_columns=(
                            "scene_name",
                            "challenge_trxn",
                            "challenge_users",
                            "benchmark_trxn",
                            "trigger_rate_pct",
                            "normalised_user_impact_pct",
                        ),
                        name_column="scene_name",
                    )
                )
                continue
            if sheet_name == "Daily Rule Trigger Trend":
                panels.append(
                    _daily_trend_panel(
                        "Daily Rule Trigger Trend",
                        headers,
                        rows,
                        rule_column="rule_id",
                        date_column="trigger_date",
                        value_column="trigger_trxn",
                        note="Top 50 rules by total triggers over the window; top 5 pre-selected.",
                    )
                )
                continue
            if sheet_name == "Rule Scorecard":
                panels.append(
                    _scatter_quadrant_panel(
                        "Rule Scorecard",
                        headers,
                        rows,
                        x_column="trigger_rate_pct",
                        y_column="precision_pct",
                        label_column="rule_id",
                        size_column="trigger_trxn",
                        x_label="Trigger rate %",
                        y_label="Precision %",
                    )
                )
                panels.append(
                    _searchable_table_panel(
                        "Rule Scorecard — Detail", headers, rows, placeholder="Search rule id or name…"
                    )
                )
                continue
            if sheet_name == "Request Outcome Summary":
                panels.append(_searchable_table_panel(sheet_name, headers, rows, placeholder="Filter…"))
                if cur is not None:
                    scols = {str(c): i for i, c in enumerate(headers)}

                    def _cg(col):
                        return cur[scols[col]] if col in scols and scols[col] < len(cur) else None

                    donut = _donut_panel(
                        "Outcome Mix",
                        [("Pass", _cg("pass_num")), ("Challenge", _cg("challenge_num")), ("Reject", _cg("reject_num"))],
                        note="Share of antifraud outcomes (pass / challenge / reject) for the latest full month.",
                    )
                    if donut:
                        panels.append(donut)
                continue
            if sheet_name == "Scene/Sub-scene/Action Usage":
                col = {str(c): i for i, c in enumerate(headers)}
                if "scene_name" in col and "transactions" in col:
                    agg: dict[str, float] = {}
                    for r in rows:
                        scene = str(r[col["scene_name"]]) if col["scene_name"] < len(r) else ""
                        agg[scene] = agg.get(scene, 0.0) + _number(r[col["transactions"]] if col["transactions"] < len(r) else 0)
                    top = sorted(agg.items(), key=lambda kv: kv[1], reverse=True)[:15]
                    bar = _bar_panel("Top Scenes by Transactions", top, note="Highest-traffic scenes this month (top 15).")
                    if bar:
                        panels.append(bar)
                panels.append(_searchable_table_panel(sheet_name, headers, rows, placeholder=placeholders.get(sheet_name, "Search…")))
                continue
            if sheet_name == "Detection by Fraud MO Type":
                mo_col = {str(c): i for i, c in enumerate(headers)}
                if "fraud_mo_type" in mo_col and "loss_leaked_php" in mo_col:
                    pairs = [
                        (str(r[mo_col["fraud_mo_type"]]), r[mo_col["loss_leaked_php"]])
                        for r in rows
                        if mo_col["fraud_mo_type"] < len(r) and mo_col["loss_leaked_php"] < len(r)
                    ]
                    bar = _bar_panel(
                        "Leaked Loss by Fraud Type (PHP)",
                        sorted(pairs, key=lambda kv: _number(kv[1]), reverse=True)[:12],
                        note="Confirmed fraud loss the rule engine did not flag, by modus operandi.",
                        prefix="₱",
                    )
                    if bar:
                        panels.append(bar)
                panels.append(_searchable_table_panel(
                    sheet_name, headers, rows, placeholder="Search fraud type…",
                    note="Confirmed-fraud cases only.",
                ))
                continue
            if sheet_name == "Daily Detection Trend":
                panels.append(
                    _daily_trend_panel(
                        "Daily Detection Trend",
                        headers,
                        rows,
                        rule_column="series",
                        date_column="case_open_date",
                        value_column="daily_loss_php",
                    )
                )
                continue
            table_notes = {
                "Detection Coverage Summary": (
                    "Scope: confirmed-fraud cases only (Not Fraud / Pending excluded), so the loss here "
                    "is smaller than the Fraud Loss report's all-cases total."
                ),
            }
            panels.append(
                _searchable_table_panel(
                    sheet_name,
                    headers,
                    rows,
                    placeholder=placeholders.get(sheet_name, "Search…"),
                    note=table_notes.get(sheet_name, ""),
                )
            )
        path.write_text(
            _searchable_tables_document(report_title, snapshot_pt_date, panels, intro_html=intro, data_through=_data_through(sheets)),
            encoding="utf-8",
        )
        return
    if report_id == AF_FRAUD_LOSS_REPORT_ID:
        loss_lookup = {sheet_name: (headers, rows) for sheet_name, headers, rows in sheets}
        summary = loss_lookup.get("Case & Loss Summary")
        intro = _period_kpi_panel(
            "Fraud Loss Summary",
            summary,
            [
                ("Cases opened", "cases_opened", "num"), ("Confirmed fraud", "fraud_cases", "num"),
                ("Fraud rate", "fraud_rate_pct", "pct"), ("Total loss (all cases)", "total_loss_php", "money"),
                ("Borne by customer", "loss_customer_php", "money"), ("Recovered", "recovered_php", "money"),
                ("Avg review time", "avg_review_hours", "hours"),
            ],
        )
        subtype_breakdown = loss_lookup.get("Fraud MO Subtype Breakdown")
        nested_sheets = {"Fraud MO Subtype Breakdown"}
        placeholders = {
            "Case Status & SLA": "Search status…",
            "Review Pool / Backlog (current)": "Search source…",
        }
        panels = []
        for sheet_name, headers, rows in sheets:
            if sheet_name == "Case & Loss Summary" or sheet_name in nested_sheets:
                continue  # summary -> KPI cards; subtype breakdown -> nested under the MO panel
            if sheet_name == "Loss by Fraud MO Type" and subtype_breakdown:
                mo_col = {str(c): i for i, c in enumerate(headers)}
                if "fraud_mo_type" in mo_col and "total_loss_php" in mo_col:
                    mo_pairs = [
                        (str(r[mo_col["fraud_mo_type"]]), r[mo_col["total_loss_php"]])
                        for r in rows
                        if mo_col["fraud_mo_type"] < len(r) and mo_col["total_loss_php"] < len(r)
                    ]
                    bar = _bar_panel(
                        "Loss by Fraud Type (PHP)",
                        sorted(mo_pairs, key=lambda kv: _number(kv[1]), reverse=True)[:12],
                        note="Total confirmed fraud loss by modus operandi.",
                        prefix="₱",
                    )
                    if bar:
                        panels.append(bar)
                panels.append(
                    _expandable_rule_panel(
                        "Loss by Fraud MO Type",
                        "Search fraud type or subtype…",
                        (headers, rows),
                        subtype_breakdown,
                        key_columns=("fraud_mo_type",),
                        main_columns=(
                            "fraud_mo_type",
                            "cases",
                            "distinct_subtypes",
                            "total_loss_php",
                            "avg_loss_php",
                            "recovered_php",
                        ),
                        detail_columns=("fraud_mo_subtype", "cases", "total_loss_php", "recovered_php"),
                        name_column="fraud_mo_subtype",
                    )
                )
                continue
            if sheet_name == "Daily Fraud Loss Trend":
                panels.append(
                    _daily_trend_panel(
                        "Daily Fraud Loss Trend",
                        headers,
                        rows,
                        rule_column="fraud_mo_type",
                        date_column="case_open_date",
                        value_column="daily_loss_php",
                    )
                )
                continue
            panels.append(
                _searchable_table_panel(sheet_name, headers, rows, placeholder=placeholders.get(sheet_name, "Search…"))
            )
        path.write_text(
            _searchable_tables_document(report_title, snapshot_pt_date, panels, intro_html=intro, data_through=_data_through(sheets)),
            encoding="utf-8",
        )
        return
    if report_id == AF_FACIAL_VERIFICATION_REPORT_ID:
        fv_lookup = {sheet_name: (headers, rows) for sheet_name, headers, rows in sheets}
        summary = fv_lookup.get("Verification Outcome Summary")
        cur = cur_label = None
        if summary and summary[1]:
            cur, _p, cur_label, _pl, _m, _s = _period_rows(summary)
        kpi_panel = _period_kpi_panel(
            "Verification Outcomes",
            summary,
            [
                ("Checks", "checks", "num"), ("Liveness pass", "liveness_pass_rate_pct", "pct"),
                ("Anti-spoof pass", "antispoof_pass_rate_pct", "pct"), ("Match pass", "match_pass_rate_pct", "pct"),
                ("Overall pass", "overall_pass_rate_pct", "pct"), ("Spoof attacks", "spoof_attack_checks", "num"),
            ],
        )
        # Auto-insights: overall pass rate, spoof-attack volume, worst-friction scene.
        insight_cards: list[tuple[str, str, str]] = []
        if summary and summary[1]:
            scols = {str(c): i for i, c in enumerate(summary[0])}

            def _cv(row, col):
                return row[scols[col]] if row is not None and col in scols and scols[col] < len(row) else None

            op = _cv(cur, "overall_pass_rate_pct")
            if op not in (None, ""):
                opn = _number(op)
                insight_cards.append((f"Overall pass rate ({cur_label})", f"{op}%", "good" if opn >= 90 else ("warn" if opn >= 80 else "bad")))
            sa = _cv(cur, "spoof_attack_checks")
            if sa not in (None, ""):
                insight_cards.append((f"Spoof attacks ({cur_label})", _format_number(sa), "warn" if _number(sa) > 0 else ""))
        by_scene = fv_lookup.get("Pass Rates by Scene")
        if by_scene and by_scene[1]:
            bc = {str(c): i for i, c in enumerate(by_scene[0])}
            if {"scene_name", "overall_pass_rate_pct", "checks"} <= set(bc):
                sized = [r for r in by_scene[1] if _number(r[bc["checks"]] if bc["checks"] < len(r) else 0) >= 1000]
                if sized:
                    worst = min(sized, key=lambda r: _number(r[bc["overall_pass_rate_pct"]] if bc["overall_pass_rate_pct"] < len(r) else 100))
                    wp = _number(worst[bc["overall_pass_rate_pct"]])
                    insight_cards.append(("Lowest-pass scene (≥1k checks)", f"{worst[bc['scene_name']]} — {wp}%", "bad" if wp < 80 else ("warn" if wp < 90 else "")))
        intro = _insight_panel("Highlights", insight_cards) + kpi_panel
        panels = []

        def _fv_stage(sheet_key: str, result_col: str, success_value: str) -> tuple[float, float]:
            table = fv_lookup.get(sheet_key)
            if not table:
                return 0.0, 0.0
            t_cols = {str(c): i for i, c in enumerate(table[0])}
            if result_col not in t_cols or "checks" not in t_cols:
                return 0.0, 0.0
            total = sum(_number(r[t_cols["checks"]]) for r in table[1] if t_cols["checks"] < len(r))
            passed = sum(
                _number(r[t_cols["checks"]]) for r in table[1]
                if t_cols[result_col] < len(r) and t_cols["checks"] < len(r)
                and str(r[t_cols[result_col]]) == success_value
            )
            return total, passed

        all_checks, liveness_passed = _fv_stage("Liveness Result Breakdown", "liveness_check_result", "LC_SUCCESS")
        _qc_total, antispoof_passed = _fv_stage("Anti-Spoofing QC Breakdown", "selfie_qc_anti_spoofing_result", "SQA_SUCCESS")
        _fm_total, match_passed = _fv_stage("Facial Match Result Breakdown", "facial_matching_result", "FM_SUCCESS")
        funnel = _funnel_panel(
            "Verification Funnel",
            [
                ("All checks", all_checks),
                ("Liveness passed", liveness_passed),
                ("Anti-spoof passed", antispoof_passed),
                ("Face match passed", match_passed),
            ],
            note="Whole-window funnel: each step only runs when the prior step passed. Hover a stage for its share of all checks.",
        )
        if funnel:
            panels.append(funnel)
        for sheet_name, headers, rows in sheets:
            if sheet_name == "Verification Outcome Summary":
                continue  # -> KPI cards
            if sheet_name == "Deepfake Score Distribution":
                dc = {str(c): i for i, c in enumerate(headers)}
                if "deepfake_score_band" in dc and "checks" in dc:
                    pairs = [
                        (str(r[dc["deepfake_score_band"]]), r[dc["checks"]])
                        for r in rows
                        if dc["deepfake_score_band"] < len(r) and dc["checks"] < len(r)
                    ]
                    bar = _bar_panel("Deepfake Score Distribution", pairs, note="Facial checks by deepfake_spoof_score band (higher band = more genuine / live; lower = more spoof / deepfake-like).")
                    if bar:
                        panels.append(bar)
                panels.append(_searchable_table_panel(sheet_name, headers, rows, placeholder="Search score band…"))
                continue
            if sheet_name == "Daily Verification Trend":
                panels.append(
                    _daily_trend_panel(
                        "Daily Verification Trend",
                        headers,
                        rows,
                        rule_column="series",
                        date_column="check_date",
                        value_column="checks",
                    )
                )
                continue
            placeholders = {
                "Liveness Result Breakdown": "Search liveness result…",
                "Anti-Spoofing QC Breakdown": "Search QC result…",
                "Facial Match Result Breakdown": "Search match result…",
                "Pass Rates by Scene": "Search scene…",
                "Fraud Review Outcomes": "Search review status…",
                "Human Review (AMR) Outcomes": "Search review status…",
                "Review Status & Verdict Detail": "Search status or verdict…",
                "CS Review Track": "Search CS status…",
            }
            review_notes = {
                "fraud_review_status": (
                    "Whether the check was pulled into the post-hoc fraud review workflow. 'Not reviewed' "
                    "= never pulled. A non-empty value means it entered review; workflow-status codes map "
                    "to ReviewStatusEnum (1 = Draft, 2 = Pending review, 3 = Review rejected, 4 = Review "
                    "approved)."
                ),
                "fraud_review_result": (
                    "KYC Ops manual review verdict (source kyc_review_result): 1 = Same person (genuine / "
                    "identity matches), 2 = Different person, 3 = Hard to say (inconclusive), 4 = Spoofing "
                    "(fraudulent). '-' = no verdict recorded."
                ),
            }
            result_col_note = {
                "liveness_check_result": (
                    "Engine LCResultEnum code — see the Result Code Glossary panel. LC_SUCCESS = passed; "
                    "LC_AURORA_SPOOF / LC_FRAUD = spoof / fraud; the rest are capture-quality failures."
                ),
                "selfie_qc_anti_spoofing_result": (
                    "Engine SQAResultEnum code — see the Result Code Glossary panel. SQA_SUCCESS = passed; "
                    "SQA_REJECT_FACE_DEEPFAKE / _SPOOFING = attack; SQA_REJECT_FACE_QC_FAILED_* = QC issues."
                ),
                "facial_matching_result": (
                    "Face-match result — see the Result Code Glossary panel. FM_SUCCESS = matched; "
                    "FM_ERROR_FACE_MISMATCHED = did not match."
                ),
            }
            cs_review_note = {
                "cs_review_status": (
                    "Status in the CS (Customer Service) review queue for this facial check — a SEPARATE "
                    "queue from the fraud / AMR review. 0 = not picked up by CS review (the vast majority). "
                    "Non-zero = the check entered a CS-review state. The exact per-code meaning is defined "
                    "in the CS / authentication system's own status scheme (not in the anti-fraud config), "
                    "so codes are shown raw; ask the CS/KYC team for the full legend if a non-zero code "
                    "needs interpreting."
                ),
            }
            col_notes = None
            if sheet_name in ("Fraud Review Outcomes", "Human Review (AMR) Outcomes", "Review Status & Verdict Detail"):
                col_notes = review_notes
            elif sheet_name == "CS Review Track":
                col_notes = cs_review_note
            elif sheet_name in ("Liveness Result Breakdown", "Anti-Spoofing QC Breakdown", "Facial Match Result Breakdown"):
                col_notes = result_col_note
            panels.append(_searchable_table_panel(
                sheet_name, headers, rows,
                placeholder=placeholders.get(sheet_name, "Search…"),
                column_notes=col_notes,
            ))
        panels.append(_facial_result_glossary_panel())
        path.write_text(
            _searchable_tables_document(report_title, snapshot_pt_date, panels, intro_html=intro, data_through=_data_through(sheets)),
            encoding="utf-8",
        )
        return
    if report_id == AF_DEVICE_RISK_REPORT_ID:
        dr_lookup = {sheet_name: (headers, rows) for sheet_name, headers, rows in sheets}
        kpis: list[tuple] = []
        farming = dr_lookup.get("Account Farming Summary")
        if farming and farming[1]:
            fh, fr = farming
            fc = {str(c): i for i, c in enumerate(fh)}
            row0 = fr[0]

            def _fv(col):
                return row0[fc[col]] if col in fc and fc[col] < len(row0) else None

            for label, col, suf in [
                ("Devices w/ 5+ accounts", "devices_5plus_accounts", ""),
                ("Devices w/ 10+ accounts", "devices_10plus_accounts", ""),
                ("Max accounts on one device", "max_accounts_on_one_device", ""),
                ("% devices 5+ accounts", "pct_devices_5plus_accounts", "%"),
            ]:
                v = _fv(col)
                if v not in (None, ""):
                    kpis.append((label, f"{_format_number(v)}{suf}" if _is_number(v) else f"{v}{suf}"))
        insight_cards: list[tuple[str, str, str]] = []
        top_dev = dr_lookup.get("Top Multi-Account Devices")
        if top_dev and top_dev[1]:
            tc = {str(c): i for i, c in enumerate(top_dev[0])}
            if "distinct_accounts" in tc:
                n = _number(top_dev[1][0][tc["distinct_accounts"]])
                if n > 0:
                    insight_cards.append(("Most-shared device", f"{int(n)} accounts on one device", "bad" if n >= 20 else "warn"))
        multi_dev = dr_lookup.get("Multi-Device Accounts")
        if multi_dev and multi_dev[1]:
            mc = {str(c): i for i, c in enumerate(multi_dev[0])}
            if "distinct_devices" in mc:
                n = _number(multi_dev[1][0][mc["distinct_devices"]])
                if n > 0:
                    insight_cards.append(("Most device-hopping account", f"{int(n)} devices for one account", "bad" if n >= 20 else "warn"))
        dr_from, dr_to = _data_window(sheets)
        window_txt = (
            f"Rolling 7-day window — events from {dr_from} to {dr_to} (the trend may stop a day short of the "
            f"query window while the latest daily partition lands)."
            if dr_from and dr_to else
            "Rolling 7-day window of device/identity events (the last 7 daily partitions at run time)."
        )
        window_panel = f'<section class="panel"><h2>Reporting Window</h2><p class="note">{html.escape(window_txt)}</p></section>'
        intro = window_panel + _insight_panel("Highlights", insight_cards) + _kpi_cards_panel("Account Farming — last 7 days", kpis)
        panels = []
        for sheet_name, headers, rows in sheets:
            if sheet_name == "Account Farming Summary":
                continue
            if sheet_name == "Risk Signal Prevalence":
                if rows:
                    rc = {str(c): i for i, c in enumerate(headers)}
                    total = _number(rows[0][rc["total_events"]]) if "total_events" in rc else 0
                    sig_rows = []
                    for col in headers:
                        c = str(col)
                        if c == "total_events":
                            continue
                        v = _number(rows[0][rc[c]] if rc[c] < len(rows[0]) else 0)
                        pct = round(v / total * 100, 4) if total else 0
                        sig_rows.append([c.replace("_", " "), int(v), pct])
                    sig_rows.sort(key=lambda r: r[1], reverse=True)
                    bar = _bar_panel(
                        "Top Device Risk Signals (events)",
                        [(r[0], r[1]) for r in sig_rows[:12] if r[1] > 0],
                        note="Events flagged for each device-risk signal in the window.",
                    )
                    if bar:
                        panels.append(bar)
                    panels.append(_searchable_table_panel(
                        "Risk Signal Prevalence", ["risk_signal", "events", "share_of_events_pct"], sig_rows,
                        placeholder="Search signal…",
                        note="One row per device-risk signal: events where the flag fired and its share of "
                             "all events, across the full window (all 17 signals, no scene filter). One event "
                             "can trip several signals, so these rows count independently and the column does "
                             "NOT sum to total events — it is not comparable to 'risky_events' in Risk Signals "
                             "by Scene. See the Risk Signal Glossary panel for what each signal means.",
                        column_notes={"risk_signal": (
                            "Device / identity risk flag from the action log; see the Risk Signal Glossary "
                            "panel for the plain-language meaning of each."
                        ), "events": (
                            "Events where THIS signal fired. An event can fire multiple signals, so these "
                            "counts overlap and do not add up to total events."
                        )}))
                continue
            if sheet_name == "Multi-Account Device Trend":
                panels.append(_daily_trend_panel(
                    "Multi-Account Device Trend", headers, rows,
                    rule_column="series", date_column="event_date", value_column="value"))
                continue
            placeholder = {
                "Top Multi-Account Devices": "Search device…",
                "Multi-Device Accounts": "Search account…",
                "Risk Signals by Scene": "Search scene…",
            }.get(sheet_name, "Search…")
            table_note = {
                "Top Multi-Account Devices": "Devices tied to 5+ distinct accounts in the window - top 300 by distinct accounts.",
                "Multi-Device Accounts": "Accounts seen on 5+ distinct devices in the window - top 300 by distinct devices.",
                "Risk Signals by Scene": "Top 100 scenes by risky-event count.",
            }.get(sheet_name, "")
            signal_scene_note = {"risky_events": (
                "Events in this scene where AT LEAST ONE of 8 core risk signals fired — is_root, "
                "is_emulator, is_vpn, is_http_proxy, is_gps_modified, is_fake_identity, is_illegal_imei, "
                "is_risk_app_install_root (counted once per event, not per signal). This is a different "
                "metric from Risk Signal Prevalence (which counts all 17 signals separately over every "
                "event), so the two will NOT sum to the same number. Top 100 scenes; null scenes excluded."
            ), "events": "All events recorded in this scene in the window (risky or not)."
            } if sheet_name == "Risk Signals by Scene" else None
            panels.append(_searchable_table_panel(
                sheet_name, headers, rows, placeholder=placeholder, note=table_note, column_notes=signal_scene_note))
        panels.append(_device_risk_signal_glossary_panel())
        path.write_text(
            _searchable_tables_document(report_title, snapshot_pt_date, panels, intro_html=intro, data_through=_data_through(sheets)),
            encoding="utf-8",
        )
        return
    if report_id == AF_CARD_3DS_REPORT_ID:
        c3_lookup = {sheet_name: (headers, rows) for sheet_name, headers, rows in sheets}
        summary = c3_lookup.get("3DS Authentication Summary")
        kpi_panel = _period_kpi_panel(
            "3DS Authentication",
            summary,
            [
                ("3DS txns", "threeds_txns", "num"), ("Authenticated", "authenticated", "num"),
                ("Auth rate", "auth_rate_pct", "pct"), ("Challenged", "challenged", "num"),
                ("Challenge rate", "challenge_rate_pct", "pct"),
            ],
        ) if summary else ""
        insight_cards = []
        if summary and summary[1]:
            cur, _p, cur_label, _pl, _m, _s = _period_rows(summary)
            sc = {str(c): i for i, c in enumerate(summary[0])}

            def _cv3(col):
                return cur[sc[col]] if cur is not None and col in sc and sc[col] < len(cur) else None

            ar = _cv3("auth_rate_pct")
            if ar not in (None, ""):
                arn = _number(ar)
                insight_cards.append((f"Auth rate ({cur_label})", f"{ar}%", "good" if arn >= 90 else ("warn" if arn >= 75 else "bad")))
            cr = _cv3("challenge_rate_pct")
            if cr not in (None, ""):
                insight_cards.append((f"Challenge rate ({cur_label})", f"{cr}%", "warn" if _number(cr) >= 20 else ""))
        intro = _insight_panel("Highlights", insight_cards) + kpi_panel
        # Display the txn-count column as "3DS_txns" (raw SQL alias stays threeds_txns for lookups).
        def _disp(hs: list[Any]) -> list[Any]:
            return ["3DS_txns" if str(h) == "threeds_txns" else h for h in hs]
        auth_status_note = (
            "EMV 3DS transaction status (transStatus) - the issuer ACS's verdict. "
            "Authenticated (Y) = cardholder verified in a frictionless flow; Not authenticated (N) = denied; "
            "Challenge (C) = a step-up (e.g. OTP / biometric) was required before a final Y/N; "
            "Rejected (R) = issuer refused, do not authorise; Unavailable (U) = could not be performed "
            "(technical issue); Info only (I) = informational acknowledgement, no authentication performed "
            "(e.g. whitelist status); Attempted (A) = proof of attempt, not a full verification."
        )
        panels = []
        for sheet_name, headers, rows in sheets:
            if sheet_name == "Outcome by Auth Status":
                col = {str(c): i for i, c in enumerate(headers)}
                if "auth_status" in col and "threeds_txns" in col:
                    donut = _donut_panel(
                        "3DS Outcome Mix",
                        [(str(r[col["auth_status"]]), r[col["threeds_txns"]]) for r in rows
                         if col["auth_status"] < len(r) and col["threeds_txns"] < len(r)],
                        note="Share of 3DS authentication outcomes over the span.")
                    if donut:
                        panels.append(donut)
                panels.append(_searchable_table_panel(
                    sheet_name, _disp(headers), rows, placeholder="Search status…",
                    column_notes={"auth_status": auth_status_note}))
                continue
            if sheet_name == "Card Fraud Cases by MO":
                col = {str(c): i for i, c in enumerate(headers)}
                if "sub_mo_reason" in col and "cases" in col:
                    bar = _bar_panel(
                        "Card Fraud Cases by Sub-MO",
                        sorted([(str(r[col["sub_mo_reason"]]), r[col["cases"]]) for r in rows
                                if col["sub_mo_reason"] < len(r) and col["cases"] < len(r)],
                               key=lambda kv: _number(kv[1]), reverse=True)[:12],
                        note="Confirmed card fraud cases by modus operandi.")
                    if bar:
                        panels.append(bar)
                panels.append(_searchable_table_panel(
                    sheet_name, headers, rows, placeholder="Search MO…",
                    note="Confirmed card fraud cases by modus operandi - top 100 MO / sub-MO combinations by case count.",
                ))
                continue
            if sheet_name == "Daily 3DS Trend":
                panels.append(_daily_trend_panel(
                    "Daily 3DS Trend", headers, rows,
                    rule_column="series", date_column="txn_date", value_column="txns"))
                continue
            placeholder = {
                "3DS Authentication Summary": "Filter…",
                "Frictionless vs Challenge": "Filter…",
                "3DS by Merchant Category (MCC)": "Search MCC or category…",
            }.get(sheet_name, "Search…")
            table_note = {
                "3DS by Merchant Category (MCC)": "Top 100 merchant categories by 3DS transaction volume over the window.",
            }.get(sheet_name, "")
            col_notes = {
                "3DS Authentication Summary": {"challenged": (
                    "3DS authentications the issuer ACS flagged Challenge Required (EMV transStatus = C) - a "
                    "step-up (e.g. OTP / biometric) was demanded. Genuinely rare (~0.2%): 3DS 2.x is "
                    "risk-based-auth driven, so the issuer authenticates most transactions frictionlessly and "
                    "the merchant's challenge request (even '04 Mandate') does not force one."
                )},
                "3DS by Merchant Category (MCC)": {"mcc": (
                    "ISO 18245 Merchant Category Code, shown as 'code - name' for the merchant's line of "
                    "business (e.g. 5399 - General merchandise, 4814 - Telecom services). '(none)' = no MCC "
                    "on the transaction; a bare code = name not in our lookup."
                )},
                "Frictionless vs Challenge": {"challenge_indicator": (
                    "3DS Requestor Challenge Indicator (EMV 3DS field) - the merchant's challenge *request* "
                    "sent before authentication, not the outcome. 01 no preference; 02 no challenge requested; "
                    "03 challenge requested (requestor preference); 04 challenge requested (mandate); "
                    "05 no challenge (risk analysis already done); 06 no challenge (data share only); "
                    "07 no challenge (SCA already done); 08 no challenge (whitelist exemption); "
                    "09 challenge requested (whitelist prompt). The ACS decides via risk-based auth, so even "
                    "'04 Mandate' is mostly frictionless - see the challenged / challenge_rate_pct columns."
                )},
            }.get(sheet_name)
            display_rows = rows
            if sheet_name == "3DS by Merchant Category (MCC)":
                mci = {str(c): i for i, c in enumerate(headers)}.get("mcc")
                if mci is not None:
                    display_rows = [
                        [(_format_mcc_cell(v) if i == mci else v) for i, v in enumerate(r)] for r in rows
                    ]
            panels.append(_searchable_table_panel(
                sheet_name, _disp(headers), display_rows, placeholder=placeholder,
                note=table_note, column_notes=col_notes))
        path.write_text(
            _searchable_tables_document(report_title, snapshot_pt_date, panels, intro_html=intro, data_through=_data_through(sheets)),
            encoding="utf-8",
        )
        return
    if report_id == AF_LIST_USAGE_REPORT_ID:
        lu_lookup = {sheet_name: (headers, rows) for sheet_name, headers, rows in sheets}
        list_type_note = (
            "Blacklist = list_type 1; Whitelist = list_type 0/2 (the column comment says 0/1 but live "
            "data carries 1 and 2). Greylist comes from a separate daily-incremental table whose snapshot "
            "lags the black/white list, so it resolves its own latest pt_date."
        )
        status_note = "Engine lifecycle code (1 / 2 / 10 observed live); the meaning is defined in the AF list config."
        id_type_note = (
            "The kind of identifier this entry lists (NOT a risk label - the risk reason is the "
            "listed_reason column). Black/white/grey share one code set (AF IdTypeEnum): "
            "1 = User ID; 2 = Mobile no.; 3 = KTP no.; 4 = Device ID; 5 = IP; 6 = IPRange; 7 = Country; "
            "8 = City; 9 = Payee Account (deprecated); 10 = Recipient Mobile no.; 11 = Merchant Code; "
            "12 = Payer Account; 13 = Card Identifier (PH); 14 = CNPTxnDevice; 15 = ID no.; 16 = MCC; "
            "17 = Merchant Bank Name; 18 = Terminal Code; 20 = UEN (SG); 21 = Shopee UID; 22 = Shop ID; "
            "23 = Recipient Account No.; 24 = Recipient Bank Name; 25 = Paynow Proxy Mobile no.; "
            "26 = Paynow Proxy NRIC No.; 27 = Paynow Proxy UEN; 28 = Paynow Proxy VPA; 29 = Token Device; "
            "30 = Customer ID; 31 = User ID & Source of Fund; 32 = Card Identifier & Source of Fund; "
            "33 = User ID & Token Device; 34 = Card Identifier & Token Device; 35-38 = GPS Hex11-14; "
            "39 = Recipient Account No. Range; -1 = not set."
        )
        scenario_note = "Applicable scene code; scenario 2 is the global/default scope, the 1xxx codes are specific scenes."
        overview = lu_lookup.get("List Overview")
        kpis: list[tuple] = []
        donut_pairs: list[tuple[str, Any]] = []
        insight_cards: list[tuple[str, str, str]] = []
        if overview and overview[1]:
            oc = {str(c): i for i, c in enumerate(overview[0])}
            ni, ei = oc.get("list_name"), oc.get("entries")
            sizes: dict[str, float] = {}
            for r in overview[1]:
                name = str(r[ni]) if ni is not None and ni < len(r) else ""
                ent = r[ei] if ei is not None and ei < len(r) else 0
                kpis.append((f"{name} entries", _format_number(ent)))
                donut_pairs.append((name, ent))
                sizes[name] = _number(ent)
            bl, wl, gl = sizes.get("Blacklist", 0.0), sizes.get("Whitelist", 0.0), sizes.get("Greylist", 0.0)
            if wl > 0:
                insight_cards.append(("Blacklist : Whitelist", f"{_format_number(round(bl / wl))} : 1", ""))
            insight_cards.append(("Greylist entries", _format_number(gl), "warn" if gl == 0 else ""))
        intro = _insight_panel("Highlights", insight_cards) + _kpi_cards_panel("List Membership", kpis)
        panels = []
        if donut_pairs:
            donut = _donut_panel(
                "Entries by List", donut_pairs,
                note="Current membership per list. Black/white pinned to the report snapshot; greylist to its own latest.")
            if donut:
                panels.append(donut)

        def _blacklist_pairs(headers: list[str], rows: list[list[Any]], key: str) -> list[tuple[str, Any]]:
            col = {str(c): i for i, c in enumerate(headers)}
            if "list_name" not in col or key not in col or "entries" not in col:
                return []
            li, ki, ent = col["list_name"], col[key], col["entries"]
            pairs = [
                (str(r[ki]), r[ent]) for r in rows
                if li < len(r) and ki < len(r) and ent < len(r) and str(r[li]) == "Blacklist"
            ]
            return sorted(pairs, key=lambda kv: _number(kv[1]), reverse=True)[:12]

        placeholders = {
            "List Overview": "Search list…",
            "Black/White by Status": "Search list or status…",
            "Black/White by ID Type": "Search list or id type…",
            "Black/White by Source": "Search list or source…",
            "Black/White by Listed Reason": "Search reason…",
            "Black/White by Scenario": "Search list or scenario…",
            "Greylist Detail": "Search…",
        }
        col_notes_by_sheet = {
            "List Overview": {"list_name": list_type_note},
            "Black/White by Status": {"status": status_note, "list_name": list_type_note},
            "Black/White by ID Type": {"id_type": id_type_note},
            "Black/White by Scenario": {"scenario": scenario_note},
            "Greylist Detail": {"id_type": id_type_note},
        }
        for sheet_name, headers, rows in sheets:
            if sheet_name == "Monthly Additions":
                panels.append(_daily_trend_panel(
                    "Monthly Additions", headers, rows,
                    rule_column="list_name", date_column="added_month", value_column="entries",
                    note="New black/white entries by the month they were created (create_date)."))
                continue
            if sheet_name == "Black/White by Source":
                bar = _bar_panel(
                    "Blacklist Entries by Source", _blacklist_pairs(headers, rows, "source"),
                    note="Top sources feeding the blacklist.")
                if bar:
                    panels.append(bar)
            if sheet_name == "Black/White by Listed Reason":
                bar = _bar_panel(
                    "Blacklist Entries by Reason", _blacklist_pairs(headers, rows, "listed_reason"),
                    note="Top listed reasons on the blacklist.")
                if bar:
                    panels.append(bar)
            panels.append(_searchable_table_panel(
                sheet_name, headers, rows,
                placeholder=placeholders.get(sheet_name, "Search…"),
                column_notes=col_notes_by_sheet.get(sheet_name)))
        path.write_text(
            _searchable_tables_document(report_title, snapshot_pt_date, panels, intro_html=intro, data_through=_data_through(sheets)),
            encoding="utf-8",
        )
        return
    lookup = {sheet_name: (headers, rows) for sheet_name, headers, rows in sheets}
    sections: list[str] = []
    product_options = _product_filter_options(sheets)
    product_filter = _product_filter_html(product_options)
    if product_filter:
        sections.append(product_filter)
    sections.append(_overview_cards(report_title, sheets))
    insight_panel = _business_insights(report_title, lookup)
    if insight_panel:
        sections.append(insight_panel)
    quality_notes = _analyze_sheets(sheets)
    if quality_notes:
        notes = "".join(f"<li>{html.escape(note)}</li>" for note in quality_notes[:10])
        sections.append(f'<section class="quality-card"><h2>Data Quality Notes</h2><ul>{notes}</ul></section>')
    else:
        sections.append('<section class="quality-card good"><h2>Data Quality Notes</h2><p>No signed 64-bit overflow, negative capacity, or rate-bound anomalies detected in aggregate sheets.</p></section>')
    specialized = _specialized_sections(report_title, lookup)
    sections.extend(specialized)
    summary = lookup.get("Summary by Product") or lookup.get("Funnel Summary by Product")
    if summary:
        headers, rows = summary
        index = {header: offset for offset, header in enumerate(headers)}
        product_index = index.get("product", 1 if len(headers) > 1 else 0)
        if not specialized:
            metric = next(
                (candidate for candidate in ("repayment_rate", "utilization_rate", "application_to_disbursement_rate") if candidate in index),
                None,
            )
            if metric:
                multiplier = 100 if rows and _number(rows[0][index[metric]]) <= 1 else 1
                sections.append(
                    _bar_chart(
                        f"{metric.replace('_', ' ').title()} by Product",
                        [row[product_index] for row in rows],
                        [_number(row[index[metric]]) * multiplier for row in rows],
                        value_suffix="%",
                        labels_are_products=True,
                    )
                )
            amount_metric = next(
                (candidate for candidate in ("outstanding_amount", "used_limit", "disbursed_principal", "applications") if candidate in index),
                None,
            )
            if amount_metric:
                sections.append(
                    _bar_chart(
                        f"{amount_metric.replace('_', ' ').title()} by Product",
                        [row[product_index] for row in rows],
                        [row[index[amount_metric]] for row in rows],
                        labels_are_products=True,
                    )
                )
        sections.append(f'<section class="panel wide"><h2>Summary by Product</h2>{_table_html(headers, rows)}</section>')
    for sheet_name, headers, rows in sheets:
        if sheet_name == "Summary by Product":
            continue
        sections.append(f'<section class="panel wide"><h2>{html.escape(sheet_name)}</h2>{_table_html(headers, rows)}</section>')
    generated_at = format_gmt8(datetime.now(UTC))
    document = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>{html.escape(report_title)} Visualization</title>
<style>
:root{{--ink:#182230;--muted:#667085;--line:#d9e2ec;--bg:#f5f7fb;--blue:#1769e0;--green:#087443;--amber:#b54708;--red:#b42318;}}
*{{box-sizing:border-box;}}
body{{margin:0;background:var(--bg);color:var(--ink);font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;}}
header{{background:linear-gradient(135deg,#102a43,#173b5f);color:#fff;padding:30px 38px;}}
header h1{{margin:0 0 8px;font-size:30px;letter-spacing:0;overflow-wrap:anywhere;word-break:break-word;}} header p{{margin:0;color:#dbeafe;overflow-wrap:anywhere;}}
main{{padding:24px 34px 38px;display:grid;grid-template-columns:repeat(12,minmax(0,1fr));gap:18px;}}
.hero-card,.quality-card,.insights-card,.filter-card,.panel{{background:#fff;border:1px solid var(--line);border-radius:8px;padding:18px;box-shadow:0 1px 2px rgba(16,42,67,.06);}}
.hero-card,.quality-card,.insights-card,.filter-card,.panel,main>*{{min-width:0;}}
.hero-card,.quality-card,.insights-card,.filter-card,.wide{{grid-column:1/-1;}} .panel{{grid-column:span 6;}}
.eyebrow{{margin:0 0 6px;color:var(--blue);font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.08em;}}
h2{{margin:0 0 14px;font-size:18px;letter-spacing:0;overflow-wrap:anywhere;word-break:break-word;}}
.hero-card h2{{font-size:22px;margin-bottom:16px;}}
.kpi-grid{{display:grid;grid-template-columns:repeat(3,minmax(160px,1fr));gap:12px;}}
.kpi{{border:1px solid #e4e7ec;border-radius:8px;padding:14px;background:#fafcff;}}
.kpi span{{display:block;color:var(--muted);font-size:12px;margin-bottom:6px;}} .kpi strong{{display:block;font-size:24px;}}
.kpi.good strong{{color:var(--green);}} .kpi.watch strong{{color:var(--amber);}}
.filter-card{{display:flex;align-items:flex-end;justify-content:space-between;gap:18px;}}
.filter-card h2{{margin-bottom:8px;}} .filter-card p{{margin:0;color:#475467;line-height:1.35;}}
.filter-card label{{display:grid;gap:6px;min-width:220px;color:#475467;font-size:12px;font-weight:700;}}
.filter-card select{{height:40px;border:1px solid #cbd5e1;border-radius:6px;background:#fff;color:var(--ink);font-size:14px;padding:0 12px;}}
.insights-card{{border-left:4px solid var(--blue);background:linear-gradient(180deg,#fff,#f8fbff);}}
.insight-grid{{display:grid;grid-template-columns:repeat(4,minmax(160px,1fr));gap:12px;}}
.insight-card{{border:1px solid #dbeafe;border-radius:8px;padding:14px;background:#fff;}}
.insight-card span{{display:block;color:var(--muted);font-size:12px;margin-bottom:6px;}} .insight-card strong{{display:block;font-size:21px;line-height:1.2;overflow-wrap:anywhere;}}
.insight-card small{{display:block;margin-top:8px;color:#475467;line-height:1.35;overflow-wrap:anywhere;}}
.quality-card{{border-left:4px solid var(--amber);}} .quality-card.good{{border-left-color:var(--green);}}
.quality-card ul{{margin:0;padding-left:18px;color:#344054;}} .quality-card li{{margin:6px 0;}}
.bar-row{{display:grid;grid-template-columns:minmax(120px,220px) 1fr minmax(90px,auto);gap:12px;align-items:center;margin:10px 0;}}
.bar-row span{{color:#344054;font-weight:600;min-width:0;overflow-wrap:anywhere;}} .bar-row b{{text-align:right;font-variant-numeric:tabular-nums;}}
.bar-track{{height:18px;background:#e5e7eb;border-radius:4px;overflow:hidden;}} .bar{{height:100%;background:linear-gradient(90deg,#1769e0,#39a0ff);}}
.donut-layout{{display:grid;grid-template-columns:180px 1fr;gap:20px;align-items:center;}}
.donut{{width:156px;height:156px;border-radius:50%;display:grid;place-items:center;position:relative;box-shadow:inset 0 0 0 1px #e4e7ec;}}
.donut:after{{content:"";position:absolute;width:92px;height:92px;border-radius:50%;background:#fff;box-shadow:0 0 0 1px #edf1f7;}}
.donut span{{position:relative;z-index:1;font-weight:800;font-size:18px;}}
.legend-row{{display:grid;grid-template-columns:14px 1fr auto;gap:8px;align-items:center;margin:8px 0;}}
.legend-row b,.comparison-card strong{{font-variant-numeric:tabular-nums;}} .legend-dot{{width:10px;height:10px;border-radius:50%;display:inline-block;}}
.stack-legend-wrap{{display:flex;gap:14px;flex-wrap:wrap;margin:0 0 12px;color:#475467;font-size:12px;}}
.stack-legend i{{display:inline-block;width:10px;height:10px;border-radius:2px;margin-right:5px;vertical-align:-1px;}}
.stack-row{{display:grid;grid-template-columns:minmax(90px,160px) 1fr minmax(90px,auto);gap:12px;align-items:center;margin:10px 0;}}
.stack-row>span{{font-weight:600;color:#344054;min-width:0;overflow-wrap:anywhere;}} .stack-row>b{{text-align:right;font-variant-numeric:tabular-nums;}}
.stack-track{{height:20px;background:#e5e7eb;border-radius:4px;overflow:hidden;display:flex;}} .stack-segment{{display:block;height:100%;min-width:2px;}}
.comparison-grid{{display:grid;grid-template-columns:repeat(3,minmax(180px,1fr));gap:12px;}}
.comparison-card{{border:1px solid #e4e7ec;border-radius:8px;padding:14px;background:#fbfdff;}}
.comparison-card span{{display:block;color:var(--muted);font-size:12px;margin-bottom:6px;}} .comparison-card strong{{display:block;font-size:22px;}}
.comparison-card small{{display:block;margin-top:7px;color:#475467;}} .comparison-card.good strong{{color:var(--green);}} .comparison-card.watch strong{{color:var(--amber);}}
.table-wrap{{overflow:auto;border:1px solid #edf1f7;border-radius:6px;}} table{{width:100%;border-collapse:collapse;font-size:13px;}} th,td{{border-bottom:1px solid #edf1f7;padding:8px 10px;text-align:left;white-space:nowrap;}} th{{background:#f1f6ff;font-weight:700;color:#344054;position:sticky;top:0;}} td.num{{text-align:right;font-variant-numeric:tabular-nums;}}
.table-pagination{{display:flex;align-items:center;justify-content:flex-end;gap:10px;margin-top:10px;color:#475467;font-size:12px;}}
.table-pagination button{{height:32px;border:1px solid #cbd5e1;border-radius:6px;background:#fff;color:#344054;padding:0 10px;font-weight:650;cursor:pointer;}}
.table-pagination button:disabled{{color:#98a2b3;background:#f8fafc;cursor:not-allowed;}}
.heatmap td.heat{{color:#12263f;font-weight:650;}}
.note{{color:var(--muted);font-size:12px;margin:10px 0 0;}}
@media(max-width:900px){{body{{overflow-x:hidden;}}header{{padding:28px 18px;}}header h1{{font-size:28px;}}main{{grid-template-columns:1fr;padding-left:16px;padding-right:16px;}}.panel{{grid-column:1/-1;}}.filter-card{{display:grid;align-items:stretch;}}.filter-card label{{min-width:0;}}.kpi-grid,.comparison-grid,.insight-grid{{grid-template-columns:1fr;}}.bar-row,.stack-row,.donut-layout{{grid-template-columns:1fr;gap:6px;}}.bar-row b,.stack-row>b{{text-align:left;}}}}
</style></head><body><header><h1>{html.escape(report_title)}</h1><p>Snapshot {html.escape(snapshot_pt_date)}. Generated {generated_at} from Data Workbench aggregate output.</p></header><main>{"".join(sections)}</main><script>
(() => {{
  const filter = document.querySelector("[data-product-filter]");
  const tables = Array.from(document.querySelectorAll("table.bi-table"));
  const productMatches = (node, selected) => !selected || !node.hasAttribute("data-product") || node.getAttribute("data-product") === selected;
  const updateTables = (selected) => {{
    tables.forEach((table) => {{
      const pageSize = Number(table.getAttribute("data-page-size") || "50");
      const rows = Array.from(table.tBodies[0]?.rows || []);
      const visibleRows = rows.filter((row) => productMatches(row, selected));
      const pages = Math.max(1, Math.ceil(visibleRows.length / pageSize));
      const currentPage = Math.min(Number(table.dataset.page || "1"), pages);
      table.dataset.page = String(currentPage);
      const start = (currentPage - 1) * pageSize;
      const end = start + pageSize;
      rows.forEach((row) => {{
        const filtered = !productMatches(row, selected);
        const pageIndex = visibleRows.indexOf(row);
        row.hidden = filtered || pageIndex < start || pageIndex >= end;
      }});
      const panel = table.closest(".panel");
      const controls = panel?.querySelector("[data-table-pagination]");
      if (!controls) return;
      const info = controls.querySelector("[data-page-info]");
      const previous = controls.querySelector("[data-page-prev]");
      const next = controls.querySelector("[data-page-next]");
      const first = visibleRows.length ? start + 1 : 0;
      const last = Math.min(end, visibleRows.length);
      if (info) info.textContent = `${{first}}-${{last}} of ${{visibleRows.length}}`;
      if (previous) previous.disabled = currentPage <= 1;
      if (next) next.disabled = currentPage >= pages;
    }});
  }};
  document.querySelectorAll("[data-table-pagination]").forEach((controls) => {{
    const table = controls.closest(".panel")?.querySelector("table.bi-table");
    if (!table) return;
    controls.querySelector("[data-page-prev]")?.addEventListener("click", () => {{
      table.dataset.page = String(Math.max(1, Number(table.dataset.page || "1") - 1));
      apply();
    }});
    controls.querySelector("[data-page-next]")?.addEventListener("click", () => {{
      table.dataset.page = String(Number(table.dataset.page || "1") + 1);
      apply();
    }});
  }});
  const apply = () => {{
    const selected = filter?.value || "";
    document.querySelectorAll("[data-product]").forEach((node) => {{
      if (node.closest("table.bi-table")) return;
      node.hidden = !productMatches(node, selected);
    }});
    document.querySelectorAll("[data-global-visual]").forEach((node) => {{
      node.hidden = Boolean(selected);
    }});
    document.querySelectorAll("[data-product-visual]").forEach((panel) => {{
      const productNodes = Array.from(panel.querySelectorAll("[data-product]"));
      panel.hidden = Boolean(selected) && !productNodes.some((node) => productMatches(node, selected));
    }});
    updateTables(selected);
  }};
  filter?.addEventListener("change", () => {{
    tables.forEach((table) => {{ table.dataset.page = "1"; }});
    apply();
  }});
  apply();
}})();
</script></body></html>"""
    path.write_text(document, encoding="utf-8")


def reports_metadata_path(portal_data_dir: Path) -> Path:
    return portal_data_dir / "business_insights" / "reports.json"


def artifacts_dir(portal_data_dir: Path) -> Path:
    return portal_data_dir / "business_insights" / "artifacts"


def load_metadata(portal_data_dir: Path) -> dict[str, Any]:
    path = reports_metadata_path(portal_data_dir)
    if not path.exists():
        return {"artifacts": {}}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"artifacts": {}}
    return payload if isinstance(payload, dict) else {"artifacts": {}}


def persist_metadata(portal_data_dir: Path, payload: dict[str, Any]) -> None:
    path = reports_metadata_path(portal_data_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, sort_keys=True, indent=2), encoding="utf-8")
    temp_path.replace(path)


def report_has_artifacts(portal_data_dir: Path, report_id: str) -> bool:
    payload = load_metadata(portal_data_dir)
    artifact = (payload.get("artifacts") or {}).get(report_id)
    if not isinstance(artifact, dict):
        return False
    root = artifacts_dir(portal_data_dir)
    return bool(
        artifact.get("filename")
        and artifact.get("visualization_filename")
        and (root / str(artifact["filename"])).exists()
        and (root / str(artifact["visualization_filename"])).exists()
    )


def sheets_from_workbook(
    path: Path,
    *,
    include_raw_export: bool = False,
    normalize_products: bool = True,
) -> list[tuple[str, list[str], list[list[Any]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: list[tuple[str, list[str], list[list[Any]]]] = []
        for sheet in workbook.worksheets:
            raw_rows = list(sheet.iter_rows(values_only=True))
            if not raw_rows:
                continue
            headers = [str(value or "") for value in raw_rows[0]]
            rows = [list(row) for row in raw_rows[1:] if any(value not in (None, "") for value in row)]
            if sheet.title == "Raw Export" and not include_raw_export:
                continue
            sheets.append((_canonical_sheet_name(sheet.title), headers, rows))
        return normalize_product_labels(sheets) if normalize_products else sheets
    finally:
        workbook.close()


def refresh_existing_visualizations(portal_data_dir: Path, *, report_ids: list[str] | None = None) -> None:
    payload = load_metadata(portal_data_dir)
    artifacts = payload.get("artifacts") if isinstance(payload.get("artifacts"), dict) else {}
    selected = set(report_ids or artifacts.keys())
    root = artifacts_dir(portal_data_dir)
    title_by_report = {
        UNDERWRITING_FUNNEL_REPORT_ID: "Credit Risk PH - Underwriting Funnel",
        **{report_id: title for report_id, (title, _builder) in REPORT_BUILDERS.items()},
    }
    for report_id, artifact in artifacts.items():
        if report_id not in selected or not isinstance(artifact, dict):
            continue
        filename = str(artifact.get("filename") or "")
        if not filename:
            continue
        workbook_path = root / filename
        if not workbook_path.exists():
            print(f"{report_id}: workbook missing; skipping visualization refresh.", flush=True)
            continue
        visualization_filename = str(artifact.get("visualization_filename") or filename.replace(".xlsx", ".html"))
        write_visualization(
            root / visualization_filename,
            report_title=title_by_report.get(report_id, report_id),
            snapshot_pt_date=str(artifact.get("snapshot_pt_date") or "latest available pt_date at run time"),
            sheets=sheets_from_workbook(workbook_path, normalize_products=False),
            report_id=report_id,
        )
        artifact["visualization_filename"] = visualization_filename
        print(f"{report_id}: refreshed visualization={visualization_filename}", flush=True)
    payload["artifacts"] = artifacts
    persist_metadata(portal_data_dir, payload)


def normalize_existing_product_labels(portal_data_dir: Path, *, report_ids: list[str] | None = None) -> None:
    payload = load_metadata(portal_data_dir)
    artifacts = payload.get("artifacts") if isinstance(payload.get("artifacts"), dict) else {}
    selected = set(report_ids or artifacts.keys())
    root = artifacts_dir(portal_data_dir)
    title_by_report = {
        UNDERWRITING_FUNNEL_REPORT_ID: "Credit Risk PH - Underwriting Funnel",
        **{report_id: title for report_id, (title, _builder) in REPORT_BUILDERS.items()},
    }
    for report_id, artifact in artifacts.items():
        if report_id not in selected or not isinstance(artifact, dict):
            continue
        filename = str(artifact.get("filename") or "")
        if not filename:
            continue
        workbook_path = root / filename
        if not workbook_path.exists():
            print(f"{report_id}: workbook missing; skipping product label normalization.", flush=True)
            continue
        sheets = sheets_from_workbook(workbook_path, include_raw_export=True, normalize_products=False)
        write_workbook(workbook_path, normalize_product_labels(sheets))
        visualization_filename = str(artifact.get("visualization_filename") or filename.replace(".xlsx", ".html"))
        write_visualization(
            root / visualization_filename,
            report_title=title_by_report.get(report_id, report_id),
            snapshot_pt_date=str(artifact.get("snapshot_pt_date") or "latest available pt_date at run time"),
            sheets=[sheet for sheet in sheets if sheet[0] != "Raw Export"],
            report_id=report_id,
        )
        artifact["visualization_filename"] = visualization_filename
        print(f"{report_id}: normalized product labels in excel={filename} visualization={visualization_filename}", flush=True)
    payload["artifacts"] = artifacts
    persist_metadata(portal_data_dir, payload)


def update_report_artifact(
    portal_data_dir: Path,
    *,
    report_id: str,
    metadata: dict[str, Any],
) -> None:
    payload = load_metadata(portal_data_dir)
    artifacts = payload.get("artifacts") if isinstance(payload.get("artifacts"), dict) else {}
    artifacts[report_id] = metadata
    payload["artifacts"] = artifacts
    persist_metadata(portal_data_dir, payload)


def generate_report(
    *,
    session: requests.Session,
    portal_data_dir: Path,
    report_id: str,
    snapshot_pt_date: str,
    now: datetime,
    skip_existing: bool,
    poll_seconds: int,
    max_polls: int,
) -> None:
    if skip_existing and report_has_artifacts(portal_data_dir, report_id):
        print(f"{report_id}: existing Excel and visualization found; skipping.", flush=True)
        return
    report_config = REPORT_BUILDERS.get(report_id)
    if report_config is None:
        raise RuntimeError(f"Unsupported report id: {report_id}")
    title, builder = report_config
    sql = builder(snapshot_pt_date=snapshot_pt_date, now=now)
    display_snapshot = snapshot_pt_date or "latest available pt_date at run time"
    sections = extract_sql_sections(sql)
    if not sections:
        raise RuntimeError(f"{report_id}: no SQL sections found.")

    print(f"{report_id}: generating {len(sections)} sheets", flush=True)
    sheets: list[tuple[str, list[str], list[list[Any]]]] = []
    executions = []
    for section in sections:
        schema, rows, execution_id = run_workbench_query(session, section, poll_seconds=poll_seconds, max_polls=max_polls)
        sheets.append((section.sheet_name, schema, rows))
        executions.append({"sheet": section.sheet_name, "execution_id": execution_id, "rows": len(rows)})

    artifact_id = uuid.uuid4().hex
    xlsx_filename = f"{report_id}-{artifact_id[:8]}.xlsx"
    html_filename = f"{report_id}-{artifact_id[:8]}.html"
    root = artifacts_dir(portal_data_dir)
    root.mkdir(parents=True, exist_ok=True)
    display_sheets = normalize_product_labels(sheets)
    write_workbook(root / xlsx_filename, display_sheets)
    write_visualization(root / html_filename, report_title=title, snapshot_pt_date=display_snapshot, sheets=sheets, report_id=report_id)
    metadata = {
        "id": artifact_id,
        "report_id": report_id,
        "filename": xlsx_filename,
        "visualization_filename": html_filename,
        "source_filename": f"data-workbench-aggregates-{snapshot_pt_date or 'latest'}.xlsx",
        "row_count": sum(len(rows) for _sheet, _headers, rows in sheets),
        "created_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "snapshot_pt_date": snapshot_pt_date,
        "sql": sql,
        "workbench_executions": executions,
    }
    update_report_artifact(portal_data_dir, report_id=report_id, metadata=metadata)
    print(f"{report_id}: completed rows={metadata['row_count']} excel={xlsx_filename} visualization={html_filename}", flush=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--report-id",
        action="append",
        default=[],
        help="Report id to generate. Repeatable. Defaults to missing generator-backed reports.",
    )
    parser.add_argument("--all", action="store_true", help="Generate all supported generator-backed reports.")
    parser.add_argument("--skip-existing", action="store_true", help="Skip reports that already have Excel and visualization artifacts.")
    parser.add_argument("--check-session", action="store_true", help="Only validate the Data Admin live session and exit.")
    parser.add_argument(
        "--refresh-visualizations",
        action="store_true",
        help="Regenerate HTML visualizations from existing Excel artifacts without calling Data Workbench.",
    )
    parser.add_argument(
        "--normalize-product-labels",
        action="store_true",
        help="Rewrite existing Excel and HTML artifacts so known product codes display Apollo product names.",
    )
    parser.add_argument(
        "--snapshot-pt-date",
        default=LATEST_SNAPSHOT,
        help="Data Workbench pt_date snapshot to use, or 'latest' (default) to resolve the newest available partition at run time.",
    )
    parser.add_argument(
        "--portal-data-dir",
        default=os.environ.get("TEAM_PORTAL_DATA_DIR") or str(DEFAULT_PORTAL_DATA_DIR),
        help="Portal data dir containing business_insights/reports.json.",
    )
    parser.add_argument("--chrome-profile", default="Default", help="Chrome profile directory for Data Admin cookies.")
    parser.add_argument("--poll-seconds", type=int, default=3, help="Seconds between Data Workbench log polls.")
    parser.add_argument("--max-polls", type=int, default=240, help="Maximum Data Workbench log polls per SQL section.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    portal_data_dir = Path(args.portal_data_dir).expanduser().resolve()
    if args.refresh_visualizations:
        selected_report_ids = args.report_id if args.report_id else None
        refresh_existing_visualizations(portal_data_dir, report_ids=selected_report_ids)
        _publish_to_public_gcs(portal_data_dir)
        return 0
    if args.normalize_product_labels:
        selected_report_ids = args.report_id if args.report_id else None
        normalize_existing_product_labels(portal_data_dir, report_ids=selected_report_ids)
        _publish_to_public_gcs(portal_data_dir)
        return 0

    token = load_data_admin_token(chrome_profile=args.chrome_profile)
    payload = _decode_jwt_payload(token)
    user = (payload.get("user") or {}).get("email") if isinstance(payload.get("user"), dict) else ""
    print(f"loaded_data_admin_token_for={user or 'unknown'}", flush=True)
    session = build_data_admin_session(token)
    validated_user = validate_data_admin_session(session)
    print(f"validated_data_admin_session={validated_user}", flush=True)
    if args.check_session:
        return 0

    if args.all:
        report_ids = list(REPORT_BUILDERS)
    elif args.report_id:
        report_ids = args.report_id
    else:
        report_ids = [report_id for report_id in REPORT_BUILDERS if not report_has_artifacts(portal_data_dir, report_id)]
    if not report_ids:
        print("No reports selected.", flush=True)
        return 0
    now = datetime.now(ZoneInfo("Asia/Singapore"))
    for report_id in report_ids:
        snapshot_pt_date = args.snapshot_pt_date
        if snapshot_pt_date == LATEST_SNAPSHOT:
            snapshot_pt_date = resolve_snapshot_pt_date(
                session, report_id, poll_seconds=args.poll_seconds, max_polls=args.max_polls
            )
            print(f"{report_id}: resolved latest snapshot pt_date={snapshot_pt_date}", flush=True)
        generate_report(
            session=session,
            portal_data_dir=portal_data_dir,
            report_id=report_id,
            snapshot_pt_date=snapshot_pt_date,
            now=now,
            skip_existing=args.skip_existing,
            poll_seconds=args.poll_seconds,
            max_polls=args.max_polls,
        )
    _publish_to_public_gcs(portal_data_dir)
    return 0


def _publish_to_public_gcs(portal_data_dir: Path) -> None:
    """Mirror refreshed reports to the public GCS bucket (best-effort)."""
    try:
        from bpmis_jira_tool.public_artifacts_gcs import publish_business_insights_dir, public_gcs_publish_bucket

        if not public_gcs_publish_bucket():
            return
        uploaded = publish_business_insights_dir(portal_data_dir / "business_insights")
        print(f"published_business_insights_files_to_gcs={uploaded}", flush=True)
    except Exception as error:  # noqa: BLE001 - publishing must not fail the refresh
        print(f"publish_business_insights_to_gcs_failed={error}", flush=True)


if __name__ == "__main__":
    raise SystemExit(main())
